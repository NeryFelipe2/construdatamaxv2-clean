#!/usr/bin/env python3
"""
GERAR_CRONOGRAMA_MACRO.PY — Cronograma multinúcleo + Export Primavera P6 + OpenProject
ConstruData - HydroNetwork · CT 11481051 · FCN Construções e Saneamento

WBS:
  Contrato SLNR Santos
    ├── N01 Verde e Teteu (180 NS, 2621m)
    │     ├── Fase 1: Escavação
    │     ├── Fase 2: Assentamento
    │     ├── ...
    │     └── NS_001, NS_002, ...
    ├── N02 Pantanal Baixo
    ├── N03 São Manoel
    └── ...

Exports:
  1. MS Project XML (já existia — agora multinúcleo)
  2. Primavera P6 XER (novo)
  3. OpenProject CSV (novo)
"""

import json, math, os, csv
from datetime import datetime, timedelta
from pathlib import Path
from xml.etree.ElementTree import Element, SubElement, tostring
from xml.dom import minidom

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


# ══════════════════════════════════════════════════════════
# FASES PADRÃO POR NÚCLEO
# ══════════════════════════════════════════════════════════

FASES = [
    {"id": "TOP", "nome": "Topografia + Cadastro", "pct_ext": 1.0, "dias_fixo": 15, "predecessora": None},
    {"id": "PRJ", "nome": "Projeto + Planejamento", "pct_ext": 1.0, "dias_fixo": 10, "predecessora": "TOP"},
    {"id": "ESC", "nome": "Escavação",              "pct_ext": 1.0, "dias_por_100m": 4, "predecessora": "PRJ"},
    {"id": "ASS", "nome": "Assentamento Tubulação",  "pct_ext": 1.0, "dias_por_100m": 3, "predecessora": "ESC"},
    {"id": "PVS", "nome": "Montagem PVs/PIs",       "pct_ext": 0.3, "dias_por_100m": 2, "predecessora": "ASS"},
    {"id": "REA", "nome": "Reaterro + Compactação",  "pct_ext": 1.0, "dias_por_100m": 2, "predecessora": "PVS"},
    {"id": "TST", "nome": "Teste Estanqueidade",     "pct_ext": 1.0, "dias_fixo": 5,  "predecessora": "REA"},
    {"id": "LAV", "nome": "Lavagem + Coliformes",    "pct_ext": 1.0, "dias_fixo": 7,  "predecessora": "TST"},
    {"id": "PAV", "nome": "Pavimentação CBUQ",       "pct_ext": 1.0, "dias_por_100m": 2, "predecessora": "REA"},
    {"id": "CAD", "nome": "Cadastro NTS 292",        "pct_ext": 1.0, "dias_fixo": 10, "predecessora": "TST"},
    {"id": "LIG", "nome": "Ligações Prediais",       "pct_ext": 1.0, "dias_por_100m": 3, "predecessora": "LAV"},
    {"id": "MED", "nome": "Medição + BM",            "pct_ext": 1.0, "dias_fixo": 5,  "predecessora": "CAD"},
]


def _dias_uteis(data_inicio, n_dias):
    """Calcula data fim pulando fins de semana."""
    dt = data_inicio
    dias = 0
    while dias < n_dias:
        dt += timedelta(days=1)
        if dt.weekday() < 5:
            dias += 1
    return dt


def gerar_cronograma_nucleo(nucleo_nome, extensao_m, n_trechos, data_inicio, equipes=2):
    """Gera cronograma de 1 núcleo com todas as fases."""
    tarefas = []
    dt_atual = data_inicio
    fase_datas = {}
    
    for fase in FASES:
        # Duração
        if "dias_fixo" in fase:
            duracao = fase["dias_fixo"]
        else:
            duracao = max(1, int(math.ceil(extensao_m / 100 * fase["dias_por_100m"] / equipes)))
        
        # Início: depende da predecessora
        pred = fase["predecessora"]
        if pred and pred in fase_datas:
            inicio = fase_datas[pred]["fim"]
        else:
            inicio = dt_atual
        
        fim = _dias_uteis(inicio, duracao)
        fase_datas[fase["id"]] = {"inicio": inicio, "fim": fim}
        
        tarefas.append({
            "id": fase["id"],
            "nome": fase["nome"],
            "inicio": inicio.strftime("%Y-%m-%d"),
            "fim": fim.strftime("%Y-%m-%d"),
            "duracao_dias": duracao,
            "predecessora": pred,
            "nucleo": nucleo_nome,
        })
    
    return tarefas


def gerar_cronograma_macro(nucleos, data_inicio_str="2026-04-01"):
    """
    Gera cronograma macro de TODOS os núcleos.
    
    Args:
        nucleos: [{nome, extensao_m, n_trechos, equipes}]
        data_inicio_str: data início do contrato
    
    Returns:
        WBS completo com todas as tarefas
    """
    data_base = datetime.strptime(data_inicio_str, "%Y-%m-%d")
    wbs = {
        "projeto": "SLNR Santos — CT 11481051",
        "empresa": "FCN Construções e Saneamento",
        "data_inicio": data_inicio_str,
        "nucleos": [],
        "total_tarefas": 0,
    }
    
    dt_atual = data_base
    task_id = 1
    
    for nuc in nucleos:
        nome = nuc["nome"]
        ext = nuc.get("extensao_m", 0)
        n_tr = nuc.get("n_trechos", 0)
        eq = nuc.get("equipes", 2)
        
        tarefas = gerar_cronograma_nucleo(nome, ext, n_tr, dt_atual, eq)
        
        # Calcular datas do núcleo
        inicio_nuc = min(t["inicio"] for t in tarefas)
        fim_nuc = max(t["fim"] for t in tarefas)
        duracao_nuc = (datetime.strptime(fim_nuc, "%Y-%m-%d") - datetime.strptime(inicio_nuc, "%Y-%m-%d")).days
        
        nucleo_wbs = {
            "nome": nome,
            "extensao_m": ext,
            "n_trechos": n_tr,
            "equipes": eq,
            "inicio": inicio_nuc,
            "fim": fim_nuc,
            "duracao_dias": duracao_nuc,
            "fases": tarefas,
        }
        
        wbs["nucleos"].append(nucleo_wbs)
        wbs["total_tarefas"] += len(tarefas)
        
        # Próximo núcleo começa 2 semanas depois (overlap parcial)
        dt_atual = _dias_uteis(datetime.strptime(inicio_nuc, "%Y-%m-%d"), 10)
    
    # Datas globais
    if wbs["nucleos"]:
        wbs["data_fim"] = max(n["fim"] for n in wbs["nucleos"])
        wbs["duracao_total_dias"] = (datetime.strptime(wbs["data_fim"], "%Y-%m-%d") - data_base).days
    
    return wbs


# ══════════════════════════════════════════════════════════
# EXPORT MS PROJECT XML
# ══════════════════════════════════════════════════════════

def exportar_project_xml(wbs, path):
    """Exporta cronograma macro para MS Project XML."""
    NS = "http://schemas.microsoft.com/project"
    
    proj = Element("Project", xmlns=NS)
    SubElement(proj, "Name").text = wbs["projeto"]
    SubElement(proj, "StartDate").text = wbs["data_inicio"] + "T08:00:00"
    SubElement(proj, "CalendarUID").text = "1"
    
    # Calendar
    cals = SubElement(proj, "Calendars")
    cal = SubElement(cals, "Calendar")
    SubElement(cal, "UID").text = "1"
    SubElement(cal, "Name").text = "Padrão"
    SubElement(cal, "IsBaseCalendar").text = "1"
    
    tasks = SubElement(proj, "Tasks")
    uid = 0
    
    # Task 0: projeto
    t0 = SubElement(tasks, "Task")
    SubElement(t0, "UID").text = str(uid)
    SubElement(t0, "ID").text = str(uid)
    SubElement(t0, "Name").text = wbs["projeto"]
    SubElement(t0, "OutlineLevel").text = "0"
    SubElement(t0, "Start").text = wbs["data_inicio"] + "T08:00:00"
    SubElement(t0, "Summary").text = "1"
    uid += 1
    
    for nuc in wbs["nucleos"]:
        # Núcleo (nível 1)
        tn = SubElement(tasks, "Task")
        SubElement(tn, "UID").text = str(uid)
        SubElement(tn, "ID").text = str(uid)
        SubElement(tn, "Name").text = f"{nuc['nome']} ({nuc['extensao_m']:.0f}m)"
        SubElement(tn, "OutlineLevel").text = "1"
        SubElement(tn, "Start").text = nuc["inicio"] + "T08:00:00"
        SubElement(tn, "Finish").text = nuc["fim"] + "T17:00:00"
        SubElement(tn, "Summary").text = "1"
        nuc_uid = uid
        uid += 1
        
        pred_map = {}
        for fase in nuc["fases"]:
            tf = SubElement(tasks, "Task")
            SubElement(tf, "UID").text = str(uid)
            SubElement(tf, "ID").text = str(uid)
            SubElement(tf, "Name").text = fase["nome"]
            SubElement(tf, "OutlineLevel").text = "2"
            SubElement(tf, "Start").text = fase["inicio"] + "T08:00:00"
            SubElement(tf, "Finish").text = fase["fim"] + "T17:00:00"
            dur_hours = fase["duracao_dias"] * 8
            SubElement(tf, "Duration").text = f"PT{dur_hours}H0M0S"
            SubElement(tf, "Summary").text = "0"
            
            if fase["predecessora"] and fase["predecessora"] in pred_map:
                pl = SubElement(tf, "PredecessorLink")
                SubElement(pl, "PredecessorUID").text = str(pred_map[fase["predecessora"]])
                SubElement(pl, "Type").text = "1"
            
            pred_map[fase["id"]] = uid
            uid += 1
    
    xml_str = minidom.parseString(tostring(proj, encoding="unicode")).toprettyxml(indent="  ")
    with open(path, "w", encoding="utf-8") as f:
        f.write(xml_str)
    
    return path


# ══════════════════════════════════════════════════════════
# EXPORT PRIMAVERA P6 XER
# ══════════════════════════════════════════════════════════

def exportar_primavera_xer(wbs, path):
    """
    Exporta cronograma para formato Primavera P6 XER.
    XER é um formato tabular (tipo CSV com tabs) usado pelo Oracle Primavera.
    """
    lines = []
    lines.append("ERMHDR\t13.0")
    lines.append(f"--- Generated by ConstruData - HydroNetwork | {datetime.now().isoformat()}")
    
    # PROJECT table
    lines.append("%T\tPROJECT")
    lines.append("%F\tproj_id\tproj_short_name\tplan_start_date\tplan_end_date")
    lines.append(f"%R\t1\tSLNR\t{wbs['data_inicio']}\t{wbs.get('data_fim', wbs['data_inicio'])}")
    
    # CALENDAR table
    lines.append("%T\tCALENDAR")
    lines.append("%F\tclndr_id\tclndr_name\tdefault_flag")
    lines.append("%R\t1\tPadrão 5d/sem\tY")
    
    # WBS table
    lines.append("%T\tPROJWBS")
    lines.append("%F\twbs_id\twbs_short_name\twbs_name\tparent_wbs_id\tproj_id")
    lines.append(f"%R\t1\tSLNR\t{wbs['projeto']}\t\t1")
    
    wbs_id = 2
    wbs_map = {}
    for nuc in wbs["nucleos"]:
        lines.append(f"%R\t{wbs_id}\t{nuc['nome'][:10]}\t{nuc['nome']}\t1\t1")
        wbs_map[nuc["nome"]] = wbs_id
        wbs_id += 1
    
    # TASK (ACTIVITY) table
    lines.append("%T\tTASK")
    lines.append("%F\ttask_id\ttask_code\ttask_name\twbs_id\tproj_id\ttarget_start_date\ttarget_end_date\ttarget_drtn_hr_cnt\ttask_type")
    
    task_id = 1
    for nuc in wbs["nucleos"]:
        nuc_wbs = wbs_map.get(nuc["nome"], 1)
        for fase in nuc["fases"]:
            code = f"{nuc['nome'][:3].upper()}_{fase['id']}"
            dur_h = fase["duracao_dias"] * 8
            lines.append(f"%R\t{task_id}\t{code}\t{fase['nome']}\t{nuc_wbs}\t1\t{fase['inicio']}\t{fase['fim']}\t{dur_h}\tTT_Task")
            task_id += 1
    
    # TASKPRED table
    lines.append("%T\tTASKPRED")
    lines.append("%F\ttask_pred_id\ttask_id\tpred_task_id\tpred_type")
    
    pred_id = 1
    task_id_map = {}
    tid = 1
    for nuc in wbs["nucleos"]:
        for fase in nuc["fases"]:
            code = f"{nuc['nome'][:3].upper()}_{fase['id']}"
            task_id_map[f"{nuc['nome']}_{fase['id']}"] = tid
            
            if fase["predecessora"]:
                pred_key = f"{nuc['nome']}_{fase['predecessora']}"
                if pred_key in task_id_map:
                    lines.append(f"%R\t{pred_id}\t{tid}\t{task_id_map[pred_key]}\tPR_FS")
                    pred_id += 1
            tid += 1
    
    lines.append("%E")
    
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    
    return path


# ══════════════════════════════════════════════════════════
# EXPORT OPENPROJECT CSV
# ══════════════════════════════════════════════════════════

def exportar_openproject_csv(wbs, path):
    """Exporta cronograma para OpenProject (CSV com WBS)."""
    rows = []
    
    # Header OpenProject
    rows.append({
        "Subject": wbs["projeto"],
        "Type": "Phase",
        "Status": "New",
        "Start date": wbs["data_inicio"],
        "Finish date": wbs.get("data_fim", ""),
        "Parent": "",
        "Estimated time": "",
    })
    
    for nuc in wbs["nucleos"]:
        # Núcleo como fase
        rows.append({
            "Subject": nuc["nome"],
            "Type": "Phase",
            "Status": "New",
            "Start date": nuc["inicio"],
            "Finish date": nuc["fim"],
            "Parent": wbs["projeto"],
            "Estimated time": f"{nuc['duracao_dias'] * 8}h",
        })
        
        for fase in nuc["fases"]:
            rows.append({
                "Subject": fase["nome"],
                "Type": "Task",
                "Status": "New",
                "Start date": fase["inicio"],
                "Finish date": fase["fim"],
                "Parent": nuc["nome"],
                "Estimated time": f"{fase['duracao_dias'] * 8}h",
            })
    
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=rows[0].keys())
        w.writeheader()
        w.writerows(rows)
    
    return path


# ══════════════════════════════════════════════════════════
# CRONOGRAMA POR NS — sequência executiva definida pelo usuário
# ══════════════════════════════════════════════════════════

# Paleta de cores por equipe (Chart.js / XLSX)
_CORES_EQUIPE = [
    "#1565c0", "#2e7d32", "#e65100", "#6a1b9a",
    "#00695c", "#c62828", "#0277bd", "#558b2f",
]

_CORES_XLSX = [
    "FF1565C0", "FF2E7D32", "FFE65100", "FF6A1B9A",
    "FF00695C", "FFC62828", "FF0277BD", "FF558B2F",
]


def gerar_cronograma_por_ns(ns_sequencia, data_inicio_str,
                             equipes=4, prod_m_dia=6.0,
                             nucleo="REDE", out_dir=None):
    """
    Gera cronograma detalhado por NS a partir da sequência executiva.

    Args:
        ns_sequencia: lista de dicts na ordem de execução:
            [{"ordem": 1, "trecho_idx": 2, "pv_ini": "PV001",
              "pv_fim": "PI054", "ext_m": 45.3, "rua": "Rua A"}, ...]
        data_inicio_str: "YYYY-MM-DD"
        equipes: N cadeias paralelas (round-robin)
        prod_m_dia: m executados por equipe por dia útil
        nucleo: nome do núcleo
        out_dir: Path | str — se fornecido, exporta arquivos

    Returns:
        dict com tarefas, total_dias, data_fim, extensao_total_m, n_equipes
    """
    if not ns_sequencia:
        return {"tarefas": [], "total_dias": 0, "data_fim": data_inicio_str,
                "extensao_total_m": 0.0, "n_equipes": equipes}

    data_inicio = datetime.strptime(data_inicio_str, "%Y-%m-%d")

    # Cadeia de datas por equipe (cada equipe tem sua própria fila)
    cadeia_fim = {e: data_inicio for e in range(1, equipes + 1)}

    tarefas = []
    for i, ns in enumerate(ns_sequencia):
        equipe = (i % equipes) + 1
        ext_m = float(ns.get("ext_m") or 0)
        duracao = max(1, math.ceil(ext_m / prod_m_dia)) if ext_m > 0 else 1
        inicio = cadeia_fim[equipe]
        fim = _dias_uteis(inicio, duracao)
        cadeia_fim[equipe] = fim

        tarefas.append({
            "ns_id":           ns.get("ordem", i + 1),
            "nome":            f"NS{ns.get('ordem', i+1):03d} {ns.get('pv_ini','?')}→{ns.get('pv_fim','?')}",
            "pv_ini":          ns.get("pv_ini", ""),
            "pv_fim":          ns.get("pv_fim", ""),
            "rua":             ns.get("rua", ""),
            "ext_m":           round(ext_m, 1),
            "equipe":          equipe,
            "inicio":          inicio.strftime("%Y-%m-%d"),
            "fim":             fim.strftime("%Y-%m-%d"),
            "duracao_dias":    duracao,
            "predecessora_ns_id": None,   # round-robin: sem predecessora explícita
        })

    # Totais
    data_fim = max(cadeia_fim[e] for e in cadeia_fim)
    total_dias = (data_fim - data_inicio).days
    ext_total = sum(t["ext_m"] for t in tarefas)

    resultado = {
        "nucleo":           nucleo,
        "tarefas":          tarefas,
        "total_dias":       total_dias,
        "data_inicio":      data_inicio_str,
        "data_fim":         data_fim.strftime("%Y-%m-%d"),
        "extensao_total_m": round(ext_total, 1),
        "n_equipes":        equipes,
        "prod_m_dia":       prod_m_dia,
        "gerado_em":        datetime.now().isoformat(),
    }

    if out_dir:
        _exportar_cronograma_ns(resultado, out_dir)

    return resultado


def _exportar_cronograma_ns(resultado, out_dir):
    """Exporta cronograma por NS para XLSX, XML, HTML e JSON."""
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)

    # ── JSON (sequência executiva) ────────────────────────────────────────
    json_path = out / "SEQUENCIA_EXECUTIVA.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(resultado, f, indent=2, ensure_ascii=False)

    # ── XLSX ──────────────────────────────────────────────────────────────
    if _HAS_OPENPYXL:
        _exportar_ns_xlsx(resultado, out / "CRONOGRAMA_NS_SEQUENCIA.xlsx")

    # ── MS Project XML ────────────────────────────────────────────────────
    _exportar_ns_project_xml(resultado, out / "CRONOGRAMA_NS.xml")

    # ── Gantt HTML (Chart.js) ─────────────────────────────────────────────
    _exportar_ns_gantt_html(resultado, out / "GANTT_NS.html")


def _exportar_ns_xlsx(resultado, path):
    """Planilha XLSX com uma linha por NS, cores por equipe."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cronograma NS"

    # Cabeçalho
    headers = ["NS", "Trecho", "Rua", "Ext (m)", "Equipe",
               "Início", "Fim", "Duração (d)"]
    hdr_fill = PatternFill("solid", fgColor="FF0A2140")
    hdr_font = Font(bold=True, color="FFFFFFFF", size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = hdr_align

    ws.row_dimensions[1].height = 28
    col_widths = [8, 28, 22, 10, 8, 14, 14, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    thin = Side(style="thin", color="FFCCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_i, t in enumerate(resultado["tarefas"], 2):
        eq = t["equipe"]
        fill_color = _CORES_XLSX[(eq - 1) % len(_CORES_XLSX)]
        row_fill = PatternFill("solid", fgColor=fill_color)
        font_row = Font(color="FFFFFFFF", size=9)
        align_c = Alignment(horizontal="center", vertical="center")
        align_l = Alignment(horizontal="left", vertical="center")

        values = [
            f"NS{t['ns_id']:03d}",
            f"{t['pv_ini']} → {t['pv_fim']}",
            t["rua"],
            t["ext_m"],
            f"Equipe {eq}",
            t["inicio"],
            t["fim"],
            t["duracao_dias"],
        ]
        for col_i, val in enumerate(values, 1):
            c = ws.cell(row=row_i, column=col_i, value=val)
            c.fill = row_fill
            c.font = font_row
            c.border = border
            c.alignment = align_c if col_i not in (2, 3) else align_l

    # Rodapé com totais
    last = len(resultado["tarefas"]) + 2
    ws.cell(row=last, column=1, value="TOTAL").font = Font(bold=True, size=9)
    ws.cell(row=last, column=4, value=resultado["extensao_total_m"]).font = Font(bold=True, size=9)
    ws.cell(row=last, column=7, value=f"Fim: {resultado['data_fim']}").font = Font(bold=True, size=9)
    ws.cell(row=last, column=8, value=resultado["total_dias"]).font = Font(bold=True, size=9)

    wb.save(path)


def _exportar_ns_project_xml(resultado, path):
    """MS Project XML — uma tarefa por NS, agrupadas por equipe."""
    NS_URI = "http://schemas.microsoft.com/project"
    proj = Element("Project", xmlns=NS_URI)
    SubElement(proj, "Name").text = f"{resultado['nucleo']} — Cronograma NS"
    SubElement(proj, "StartDate").text = resultado["data_inicio"] + "T08:00:00"
    SubElement(proj, "CalendarUID").text = "1"

    cals = SubElement(proj, "Calendars")
    cal = SubElement(cals, "Calendar")
    SubElement(cal, "UID").text = "1"
    SubElement(cal, "Name").text = "Padrão"
    SubElement(cal, "IsBaseCalendar").text = "1"

    tasks = SubElement(proj, "Tasks")
    uid = 0

    # Sumário raiz
    t0 = SubElement(tasks, "Task")
    SubElement(t0, "UID").text = "0"
    SubElement(t0, "ID").text = "0"
    SubElement(t0, "Name").text = f"{resultado['nucleo']} — Sequência Executiva"
    SubElement(t0, "OutlineLevel").text = "0"
    SubElement(t0, "Start").text = resultado["data_inicio"] + "T08:00:00"
    SubElement(t0, "Summary").text = "1"
    uid = 1

    for t in resultado["tarefas"]:
        tf = SubElement(tasks, "Task")
        SubElement(tf, "UID").text = str(uid)
        SubElement(tf, "ID").text = str(uid)
        SubElement(tf, "Name").text = t["nome"]
        SubElement(tf, "OutlineLevel").text = "1"
        SubElement(tf, "Start").text = t["inicio"] + "T08:00:00"
        SubElement(tf, "Finish").text = t["fim"] + "T17:00:00"
        SubElement(tf, "Duration").text = f"PT{t['duracao_dias'] * 8}H0M0S"
        SubElement(tf, "Summary").text = "0"
        SubElement(tf, "Notes").text = f"Equipe {t['equipe']} | {t['ext_m']}m | {t.get('rua','')}"
        uid += 1

    xml_str = minidom.parseString(tostring(proj, encoding="unicode")).toprettyxml(indent="  ")
    with open(path, "w", encoding="utf-8") as f:
        f.write(xml_str)


def _exportar_ns_gantt_html(resultado, path):
    """Gantt interativo em HTML + Chart.js, colorido por equipe."""
    tarefas = resultado["tarefas"]
    nucleo  = resultado["nucleo"]
    equipes = resultado["n_equipes"]

    # Converter datas para labels legíveis e offsets (dias desde início)
    data_base = datetime.strptime(resultado["data_inicio"], "%Y-%m-%d")

    rows_js = []
    for t in tarefas:
        ini_d = (datetime.strptime(t["inicio"], "%Y-%m-%d") - data_base).days
        eq = t["equipe"]
        cor = _CORES_EQUIPE[(eq - 1) % len(_CORES_EQUIPE)]
        label = f"NS{t['ns_id']:03d} {t['pv_ini']}→{t['pv_fim']}"
        tooltip = f"{label} | Equipe {eq} | {t['ext_m']}m | {t.get('rua','')} | {t['inicio']}→{t['fim']}"
        rows_js.append(
            f"  {{label:{json.dumps(label)}, inicio:{ini_d}, dur:{t['duracao_dias']}, "
            f"equipe:{eq}, cor:{json.dumps(cor)}, tooltip:{json.dumps(tooltip)}}}"
        )

    rows_js_str = "[\n" + ",\n".join(rows_js) + "\n]"

    leg_items = "".join(
        f'<span class="leg-item"><span class="leg-dot" style="background:{_CORES_EQUIPE[i % len(_CORES_EQUIPE)]}"></span>Equipe {i+1}</span>'
        for i in range(equipes)
    )

    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="utf-8"/>
<title>Gantt NS — {nucleo}</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Segoe UI',Arial,sans-serif;background:#0d1117;color:#e6edf3;padding:16px}}
h1{{font-size:18px;margin-bottom:4px;color:#58a6ff}}
.sub{{font-size:12px;color:#8b949e;margin-bottom:12px}}
.legend{{display:flex;flex-wrap:wrap;gap:8px;margin-bottom:12px}}
.leg-item{{display:flex;align-items:center;gap:4px;font-size:11px;color:#c9d1d9}}
.leg-dot{{width:12px;height:12px;border-radius:3px}}
#gantt{{width:100%;overflow-x:auto}}
canvas{{display:block}}
.tooltip-box{{position:fixed;background:#161b22;border:1px solid #30363d;
  border-radius:6px;padding:6px 10px;font-size:11px;color:#e6edf3;
  pointer-events:none;display:none;max-width:320px;z-index:999}}
</style>
</head>
<body>
<h1>Gantt — Sequência Executiva de NS</h1>
<div class="sub">{nucleo} &nbsp;|&nbsp; {len(tarefas)} NS &nbsp;|&nbsp; {equipes} equipes &nbsp;|&nbsp; {resultado['extensao_total_m']:.0f} m &nbsp;|&nbsp; {resultado['data_inicio']} → {resultado['data_fim']}</div>
<div class="legend">{leg_items}</div>
<div id="gantt"><canvas id="c"></canvas></div>
<div class="tooltip-box" id="tip"></div>
<script>
const DATA = {rows_js_str};
const BASE_DATE = new Date("{resultado['data_inicio']}");
const TOTAL_DIAS = {resultado['total_dias']};

const ROW_H = 22, PAD_L = 180, PAD_R = 20, PAD_T = 36, BAR_H = 14;
const canvas = document.getElementById('c');
const ctx = canvas.getContext('2d');
const tip = document.getElementById('tip');

function dateLabel(d) {{
  const dt = new Date(BASE_DATE); dt.setDate(dt.getDate()+d);
  return dt.toLocaleDateString('pt-BR',{{day:'2-digit',month:'2-digit'}});
}}

function draw() {{
  const W = Math.max(document.getElementById('gantt').clientWidth, 900);
  const H = PAD_T + DATA.length * ROW_H + 30;
  canvas.width = W; canvas.height = H;
  const BAR_W = W - PAD_L - PAD_R;

  ctx.fillStyle='#0d1117'; ctx.fillRect(0,0,W,H);

  // Cabeçalho de datas
  ctx.fillStyle='#8b949e'; ctx.font='10px Segoe UI';
  const step = Math.max(1, Math.round(TOTAL_DIAS/20));
  for(let d=0; d<=TOTAL_DIAS; d+=step) {{
    const x = PAD_L + (d/TOTAL_DIAS)*BAR_W;
    ctx.fillStyle='#30363d'; ctx.fillRect(x,PAD_T-20,1,H-PAD_T+16);
    ctx.fillStyle='#8b949e'; ctx.fillText(dateLabel(d), x-14, PAD_T-6);
  }}

  DATA.forEach((t,i) => {{
    const y = PAD_T + i*ROW_H;
    // Zebra
    ctx.fillStyle = i%2===0?'#161b22':'#0d1117';
    ctx.fillRect(0, y, W, ROW_H);
    // Label
    ctx.fillStyle='#c9d1d9'; ctx.font='9px Segoe UI';
    ctx.fillText(t.label, 4, y+ROW_H/2+3.5);
    // Barra
    const bx = PAD_L + (t.inicio/TOTAL_DIAS)*BAR_W;
    const bw = Math.max(4, (t.dur/TOTAL_DIAS)*BAR_W);
    ctx.fillStyle=t.cor;
    ctx.beginPath();
    ctx.roundRect(bx, y+3, bw, BAR_H, 3);
    ctx.fill();
  }});
}}

draw(); window.addEventListener('resize', draw);

canvas.addEventListener('mousemove', e => {{
  const rect = canvas.getBoundingClientRect();
  const mx = e.clientX-rect.left, my = e.clientY-rect.top;
  const i = Math.floor((my-PAD_T)/ROW_H);
  if(i>=0 && i<DATA.length) {{
    const t=DATA[i];
    const bx=PAD_L+(t.inicio/TOTAL_DIAS)*(canvas.width-PAD_L-PAD_R);
    const bw=Math.max(4,(t.dur/TOTAL_DIAS)*(canvas.width-PAD_L-PAD_R));
    if(mx>=bx && mx<=bx+bw) {{
      tip.style.display='block';
      tip.style.left=(e.clientX+12)+'px';
      tip.style.top=(e.clientY-10)+'px';
      tip.textContent=t.tooltip;
      return;
    }}
  }}
  tip.style.display='none';
}});
canvas.addEventListener('mouseleave',()=>tip.style.display='none');
</script>
</body>
</html>"""

    with open(path, "w", encoding="utf-8") as f:
        f.write(html)


# ══════════════════════════════════════════════════════════
# EXPORT XLSX — Cronograma Macro
# ══════════════════════════════════════════════════════════

def exportar_macro_xlsx(wbs, path):
    """Exporta cronograma macro para XLSX com abas RESUMO e TAREFAS."""
    if not _HAS_OPENPYXL:
        return None

    wb = openpyxl.Workbook()

    # ── Aba RESUMO ────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "RESUMO"

    hdr_fill  = PatternFill("solid", fgColor="FF0A2140")
    hdr_font  = Font(bold=True, color="FFFFFFFF", size=10)
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin      = Side(style="thin", color="FFCCCCCC")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Título
    ws.merge_cells("A1:G1")
    tc = ws["A1"]
    tc.value   = f"CRONOGRAMA MACRO — {wbs.get('projeto', 'SLNR Santos')}"
    tc.font    = Font(bold=True, color="FF00C4FF", size=13)
    tc.fill    = PatternFill("solid", fgColor="FF0A2140")
    tc.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    # Cabeçalhos
    headers = ["Núcleo", "Extensão (m)", "Trechos", "Equipes", "Início", "Fim", "Duração (d)"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = hdr_align; c.border = border

    ws.column_dimensions["A"].width = 24
    for col in range(2, 8):
        ws.column_dimensions[get_column_letter(col)].width = 14

    # Dados por núcleo
    for row_i, nuc in enumerate(wbs.get("nucleos", []), 3):
        eq  = (row_i - 3) % len(_CORES_XLSX)
        fill = PatternFill("solid", fgColor=_CORES_XLSX[eq])
        font_r = Font(color="FFFFFFFF", size=9)
        for col, val in enumerate([
            nuc["nome"], nuc["extensao_m"], nuc["n_trechos"], nuc["equipes"],
            nuc["inicio"], nuc["fim"], nuc["duracao_dias"]
        ], 1):
            c = ws.cell(row=row_i, column=col, value=val)
            c.fill = fill; c.font = font_r; c.border = border
            c.alignment = Alignment(horizontal="center" if col > 1 else "left")

    # KPIs finais
    last = len(wbs.get("nucleos", [])) + 3
    ws.cell(row=last, column=1, value="TOTAL CONTRATO").font = Font(bold=True)
    ws.cell(row=last, column=7, value=wbs.get("duracao_total_dias", 0)).font = Font(bold=True)
    ws.cell(row=last + 1, column=1, value=f"Início: {wbs.get('data_inicio','')}").font = Font(italic=True, size=8)
    ws.cell(row=last + 1, column=4, value=f"Fim: {wbs.get('data_fim','')}").font = Font(italic=True, size=8)

    # ── Aba TAREFAS ───────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("TAREFAS")
    t_headers = ["Núcleo", "Fase", "ID", "Início", "Fim", "Duração (d)", "Predecessora"]
    for col, h in enumerate(t_headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = hdr_align; c.border = border

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 28
    for col in ["C", "D", "E", "F", "G"]:
        ws2.column_dimensions[col].width = 14

    row_i = 2
    for ni, nuc in enumerate(wbs.get("nucleos", [])):
        fill = PatternFill("solid", fgColor=_CORES_XLSX[ni % len(_CORES_XLSX)])
        font_r = Font(color="FFFFFFFF", size=9)
        for fase in nuc.get("fases", []):
            for col, val in enumerate([
                nuc["nome"], fase["nome"], fase["id"],
                fase["inicio"], fase["fim"], fase["duracao_dias"],
                fase.get("predecessora", "")
            ], 1):
                c = ws2.cell(row=row_i, column=col, value=val)
                c.fill = fill; c.font = font_r; c.border = border
                c.alignment = Alignment(horizontal="left" if col <= 2 else "center")
            row_i += 1

    wb.save(str(path))
    return str(path)


def gerar_tudo(nucleos, data_inicio="2026-04-01", out_dir="."):
    """Gera cronograma macro + exporta para os 3 formatos."""
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)
    
    wbs = gerar_cronograma_macro(nucleos, data_inicio)
    
    paths = []
    
    # MS Project
    p1 = str(out / "CRONOGRAMA_MACRO_SLNR.xml")
    exportar_project_xml(wbs, p1)
    paths.append(p1)
    
    # Primavera P6
    p2 = str(out / "CRONOGRAMA_MACRO_SLNR.xer")
    exportar_primavera_xer(wbs, p2)
    paths.append(p2)
    
    # OpenProject
    p3 = str(out / "CRONOGRAMA_MACRO_SLNR.csv")
    exportar_openproject_csv(wbs, p3)
    paths.append(p3)
    
    # JSON
    p4 = str(out / "CRONOGRAMA_MACRO_SLNR.json")
    with open(p4, "w", encoding="utf-8") as f:
        json.dump(wbs, f, indent=2, ensure_ascii=False)
    paths.append(p4)

    # XLSX
    p5 = str(out / "CRONOGRAMA_MACRO_SLNR.xlsx")
    try:
        exportar_macro_xlsx(wbs, p5)
        paths.append(p5)
    except Exception:
        pass

    return wbs, paths


if __name__ == "__main__":
    # Dados reais dos núcleos SLNR Santos
    nucleos = [
        {"nome": "Verde e Teteu",   "extensao_m": 2621,  "n_trechos": 180, "equipes": 3},
        {"nome": "São Manoel",      "extensao_m": 1275,  "n_trechos": 16,  "equipes": 2},
        {"nome": "Vila Israel",     "extensao_m": 11509, "n_trechos": 861, "equipes": 3},
        {"nome": "Pantanal Baixo",  "extensao_m": 6720,  "n_trechos": 189, "equipes": 3},
        {"nome": "Vila Criadores",  "extensao_m": 4138,  "n_trechos": 130, "equipes": 2},
        {"nome": "João Carlos",     "extensao_m": 3000,  "n_trechos": 100, "equipes": 2},
    ]
    
    wbs, paths = gerar_tudo(nucleos, "2026-04-01", "/home/claude/CRONO_MACRO")
    
    print("═" * 60)
    print("  CRONOGRAMA MACRO — SLNR Santos")
    print("═" * 60)
    print(f"\n  {len(wbs['nucleos'])} núcleos | {wbs['total_tarefas']} tarefas")
    print(f"  Início: {wbs['data_inicio']} | Fim: {wbs.get('data_fim','?')} | {wbs.get('duracao_total_dias',0)} dias")
    
    for nuc in wbs["nucleos"]:
        print(f"\n  {nuc['nome']:20s} | {nuc['extensao_m']:>6.0f}m | {nuc['duracao_dias']:3d}d | {nuc['inicio']} → {nuc['fim']}")
    
    print(f"\n  Exportado:")
    for p in paths:
        print(f"    {os.path.basename(p)} ({os.path.getsize(p)/1024:.0f}KB)")
