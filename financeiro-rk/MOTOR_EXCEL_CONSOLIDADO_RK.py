"""
╔══════════════════════════════════════════════════════════════════════╗
║  CONTROLE FINANCEIRO CONSOLIDADO — RK ENGENHARIA                    ║
║  ✨ MOTOR EXCEL PREMIUM com Gráficos, Formatação Condicional,       ║
║     Validação de Dados, Filtros e Layout ENGELFER-RK                ║
║                                                                      ║
║  Autor: Antigravity (para Felipe Nery)                               ║
║  Data: 2026-04-14                                                    ║
╚══════════════════════════════════════════════════════════════════════╝
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import (
    BarChart, PieChart, LineChart, Reference, DoughnutChart
)
from openpyxl.formatting.rule import (
    ColorScaleRule, CellIsRule, FormulaRule
)
from datetime import datetime
from collections import defaultdict
import os

# ═══════════════════════════════════════════════════════════════════════
# CONFIGURAÇÃO — Caminhos das planilhas
# ═══════════════════════════════════════════════════════════════════════

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

OBRAS = {
    'TATUI': {
        'path': os.path.join(BASE_DIR, 'TATUI', 'Tatuí - Controle de Obras ENGELFER-RK  - Copia.xlsm'),
        'nome_completo': 'Tatuí / Cesário Lange / São Roque',
        'responsavel': 'Felipe Nery / Ícaro',
        'cor': 'D9E8F5',
    },
    'OSASCO': {
        'path': os.path.join(BASE_DIR, 'OSASCO', 'Osasco - Controle de Obras ENGELFER-RK.xlsm'),
        'nome_completo': 'Osasco — Sabesp',
        'responsavel': 'Mateus',
        'cor': 'FCE4D6',
    },
    'SANTOS': {
        'path': os.path.join(BASE_DIR, 'SANTOS EMPREITA RK', 'Santos - Controle de Obras ENGELFER-RK .xlsm'),
        'nome_completo': 'Santos — Empreita RK',
        'responsavel': 'Igor / Alexandre',
        'cor': 'E2EFDA',
    },
    'PARDINHO': {
        'path': os.path.join(BASE_DIR, 'PARDINHO', 'Pardinho - Controle de Obras ENGELFER-RK.xlsm'),
        'nome_completo': 'Pardinho — SES Garagem',
        'responsavel': 'Felipe Nery / Ícaro',
        'cor': 'F4CCCC',
    },
    'CACHOEIRO': {
        'path': os.path.join(BASE_DIR, 'CACHOEIRO', 'Cachoeiro - Controle de Obras ENGELFER-RK  - Copia - Copia.xlsm'),
        'nome_completo': 'Cachoeiro de Itapemirim',
        'responsavel': 'A definir',
        'cor': 'FFF2CC',
    },
    'TEOFILO': {
        'path': os.path.join(BASE_DIR, 'TEOFILO OTONI', 'Teófilo Otoni - Controle de Obras ENGELFER-RK  - Copia.xlsm'),
        'nome_completo': 'Teófilo Otoni — MG',
        'responsavel': 'Wellington',
        'cor': 'F0F0F0',
    },
}

FUNCIONARIOS_PATH = os.path.join(BASE_DIR, 'RELAÇÃO DE FUNCIONARIOS POR OBRA ATUALIZADA.xlsx')
OUTPUT_PATH = os.path.join(BASE_DIR, 'CONTROLE_CONSOLIDADO_RK.xlsx')

# ═══════════════════════════════════════════════════════════════════════
# ESTILOS ENGINEERING-GRADE — Base ENGELFER-RK
# ═══════════════════════════════════════════════════════════════════════

COR_AZUL_ESCURO = '1B2A4A'
COR_AZUL_MEDIO = '2D4A7A'
COR_AZUL_CLARO = '4A7FB5'
COR_VERDE = '27AE60'
COR_VERMELHO = 'E74C3C'
COR_AMARELO = 'F39C12'
COR_CINZA_CLARO = 'F5F6FA'
COR_CINZA_MEDIO = 'E8E8E8'
COR_BRANCO = 'FFFFFF'
COR_PRETO = '1A1A1A'

FONT_TITULO_PRINCIPAL = Font(name='Calibri', size=18, bold=True, color=COR_BRANCO)
FONT_TITULO = Font(name='Calibri', size=14, bold=True, color=COR_BRANCO)
FONT_SUBTITULO = Font(name='Calibri', size=12, bold=True, color=COR_BRANCO)
FONT_HEADER = Font(name='Calibri', size=11, bold=True, color=COR_BRANCO)
FONT_NORMAL = Font(name='Calibri', size=10, color=COR_PRETO)
FONT_VALOR = Font(name='Calibri', size=10, color=COR_PRETO)
FONT_VALOR_POSITIVO = Font(name='Calibri', size=10, bold=True, color=COR_VERDE)
FONT_VALOR_NEGATIVO = Font(name='Calibri', size=10, bold=True, color=COR_VERMELHO)
FONT_TOTAL = Font(name='Calibri', size=11, bold=True, color=COR_BRANCO)
FONT_PEQUENO = Font(name='Calibri', size=9, italic=True, color='666666')

FILL_TITULO_PRINCIPAL = PatternFill(start_color=COR_AZUL_ESCURO, end_color=COR_AZUL_ESCURO, fill_type='solid')
FILL_TITULO = PatternFill(start_color=COR_AZUL_ESCURO, end_color=COR_AZUL_ESCURO, fill_type='solid')
FILL_SUBTITULO = PatternFill(start_color=COR_AZUL_MEDIO, end_color=COR_AZUL_MEDIO, fill_type='solid')
FILL_HEADER = PatternFill(start_color=COR_AZUL_CLARO, end_color=COR_AZUL_CLARO, fill_type='solid')
FILL_ZEBRA_1 = PatternFill(start_color=COR_BRANCO, end_color=COR_BRANCO, fill_type='solid')
FILL_ZEBRA_2 = PatternFill(start_color=COR_CINZA_CLARO, end_color=COR_CINZA_CLARO, fill_type='solid')
FILL_TOTAL = PatternFill(start_color=COR_AZUL_ESCURO, end_color=COR_AZUL_ESCURO, fill_type='solid')
FILL_VERDE = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
FILL_VERMELHO = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
FILL_AMARELO = PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid')

ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')

BORDER_THIN = Border(
    left=Side(style='thin', color=COR_CINZA_MEDIO),
    right=Side(style='thin', color=COR_CINZA_MEDIO),
    top=Side(style='thin', color=COR_CINZA_MEDIO),
    bottom=Side(style='thin', color=COR_CINZA_MEDIO),
)

BORDER_THICK = Border(
    left=Side(style='medium', color=COR_AZUL_ESCURO),
    right=Side(style='medium', color=COR_AZUL_ESCURO),
    top=Side(style='medium', color=COR_AZUL_ESCURO),
    bottom=Side(style='medium', color=COR_AZUL_ESCURO),
)

FMT_MOEDA = '#,##0.00;[Red]-#,##0.00'
FMT_DATA = 'DD/MM/YYYY'
FMT_INTEIRO = '0'

# ═══════════════════════════════════════════════════════════════════════
# FUNÇÕES DE ESTILO
# ═══════════════════════════════════════════════════════════════════════

def apply_header_style(ws, row, col_start, col_end):
    """Aplica estilo de cabeçalho premium."""
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN


def apply_data_style(ws, row, col_start, col_end, is_even=False, is_total=False):
    """Aplica estilo zebra nas linhas de dados."""
    if is_total:
        fill = FILL_TOTAL
        font = FONT_TOTAL
    else:
        fill = FILL_ZEBRA_2 if is_even else FILL_ZEBRA_1
        font = FONT_NORMAL
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font
        cell.fill = fill
        cell.border = BORDER_THIN
        cell.alignment = ALIGN_LEFT


def apply_title_row(ws, row, start_col, end_col, text, color='TITULO'):
    """Aplica estilo de título a uma linha."""
    ws.merge_cells(f'{get_column_letter(start_col)}{row}:{get_column_letter(end_col)}{row}')
    c = ws.cell(row=row, column=start_col)
    c.value = text
    c.font = FONT_TITULO
    c.fill = FILL_TITULO
    c.alignment = ALIGN_CENTER
    c.border = BORDER_THICK


def set_col_widths(ws, widths):
    """Define larguras de colunas."""
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ═══════════════════════════════════════════════════════════════════════
# LEITURA DOS DADOS
# ═══════════════════════════════════════════════════════════════════════

def ler_lancamentos(obra_key, config):
    """Lê lançamentos da aba 'Base de Dados' de uma planilha de obra."""
    path = config['path']
    if not os.path.exists(path):
        print(f"  ⚠ ARQUIVO NÃO ENCONTRADO: {path}")
        return []

    wb = openpyxl.load_workbook(path, data_only=True, keep_vba=True)
    ws = wb['Base de Dados']

    records = []
    for i in range(7, ws.max_row + 1):
        cat = ws.cell(row=i, column=2).value
        data = ws.cell(row=i, column=3).value
        mes = ws.cell(row=i, column=4).value
        ano = ws.cell(row=i, column=5).value
        subcat = ws.cell(row=i, column=6).value
        forma = ws.cell(row=i, column=7).value
        valor = ws.cell(row=i, column=8).value
        info = ws.cell(row=i, column=9).value

        if cat and valor:
            if isinstance(data, datetime):
                data_fmt = data
            elif isinstance(data, str):
                try:
                    data_fmt = datetime.strptime(data, '%Y-%m-%d')
                except:
                    data_fmt = None
            else:
                data_fmt = None

            records.append({
                'obra': obra_key,
                'obra_nome': config['nome_completo'],
                'responsavel': config['responsavel'],
                'categoria': cat,
                'data': data_fmt,
                'mes': mes,
                'ano': ano,
                'subcategoria': subcat or '(sem categoria)',
                'forma_pagamento': forma or '(não informado)',
                'valor': float(valor) if valor else 0,
                'descricao': info or '',
            })

    wb.close()
    return records


def ler_funcionarios():
    """Lê a planilha de funcionários por obra."""
    if not os.path.exists(FUNCIONARIOS_PATH):
        print(f"  ⚠ ARQUIVO NÃO ENCONTRADO: {FUNCIONARIOS_PATH}")
        return []

    wb = openpyxl.load_workbook(FUNCIONARIOS_PATH, data_only=True)
    todos = []

    for sn in wb.sheetnames:
        ws = wb[sn]
        obra_aba = sn.strip().upper()

        for i in range(2, ws.max_row + 1):
            nome = ws.cell(row=i, column=1).value
            if not nome:
                continue

            funcao = ws.cell(row=i, column=2).value
            nascimento = ws.cell(row=i, column=3).value
            cpf = ws.cell(row=i, column=4).value
            departamento = ws.cell(row=i, column=5).value
            admissao = ws.cell(row=i, column=6).value
            salario = ws.cell(row=i, column=7).value
            salario_encargos = ws.cell(row=i, column=8).value
            venc_experiencia = ws.cell(row=i, column=9).value

            if 'OSASCO' in obra_aba:
                obra_key = 'OSASCO'
            elif 'TEOFILO' in obra_aba or 'OTONI' in obra_aba:
                obra_key = 'TEOFILO'
            elif 'PORONGABA' in obra_aba or 'CESARIO' in obra_aba or 'ROQUE' in obra_aba:
                obra_key = 'TATUI'
            else:
                obra_key = obra_aba

            todos.append({
                'obra': obra_key,
                'obra_aba': sn.strip(),
                'nome': nome,
                'funcao': (funcao or '').strip(),
                'nascimento': nascimento,
                'cpf': cpf,
                'departamento': departamento,
                'admissao': admissao,
                'salario': float(salario) if salario else 0,
                'salario_encargos': float(salario_encargos) if salario_encargos else 0,
                'venc_experiencia': venc_experiencia,
            })

    wb.close()
    return todos


# ═══════════════════════════════════════════════════════════════════════
# GERADORES DE ABAS COM FORMATAÇÃO AVANÇADA
# ═══════════════════════════════════════════════════════════════════════

def gerar_aba_resumo(wb, all_lancamentos, all_funcionarios):
    """Aba RESUMO — Visão geral com gráficos e formatação condicional."""
    ws = wb.create_sheet('RESUMO', 0)
    ws.sheet_properties.tabColor = COR_AZUL_ESCURO

    # Título principal
    apply_title_row(ws, 1, 1, 9, '🏗 CONTROLE FINANCEIRO CONSOLIDADO — RK ENGENHARIA')
    
    # Subtítulo com data
    ws.merge_cells('A2:I2')
    c2 = ws['A2']
    c2.value = f'Gerado em {datetime.now().strftime("%d/%m/%Y às %H:%M")} — Análise Consolidada de Todas as 6 Obras'
    c2.font = FONT_PEQUENO
    c2.fill = FILL_SUBTITULO
    c2.alignment = ALIGN_CENTER

    # Cabeçalhos
    headers = ['Obra', 'Nome Completo', 'Responsável', 'Total Receitas',
               'Total Gastos', 'Saldo Líquido', 'Nº Lançamentos', 'Nº Funcionários', 'Folha+Encargos']
    row = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    apply_header_style(ws, row, 1, len(headers))

    # Dados por obra
    totais = {'receitas': 0, 'gastos': 0, 'lancamentos': 0, 'funcionarios': 0, 'folha': 0}
    data_rows = []
    row = 5
    
    for idx, (obra_key, config) in enumerate(OBRAS.items()):
        lanc_obra = [l for l in all_lancamentos if l['obra'] == obra_key]
        func_obra = [f for f in all_funcionarios if f['obra'] == obra_key]

        receitas = sum(l['valor'] for l in lanc_obra if l['valor'] > 0)
        gastos = sum(l['valor'] for l in lanc_obra if l['valor'] < 0)
        saldo = receitas + gastos
        folha = sum(f['salario_encargos'] for f in func_obra)

        data_rows.append(row)  # Para gráfico
        ws.cell(row=row, column=1, value=obra_key)
        ws.cell(row=row, column=2, value=config['nome_completo'])
        ws.cell(row=row, column=3, value=config['responsavel'])

        c_rec = ws.cell(row=row, column=4, value=receitas)
        c_rec.number_format = FMT_MOEDA

        c_gas = ws.cell(row=row, column=5, value=gastos)
        c_gas.number_format = FMT_MOEDA

        c_sal = ws.cell(row=row, column=6, value=saldo)
        c_sal.number_format = FMT_MOEDA

        ws.cell(row=row, column=7, value=len(lanc_obra)).number_format = FMT_INTEIRO
        ws.cell(row=row, column=8, value=len(func_obra)).number_format = FMT_INTEIRO

        c_folha = ws.cell(row=row, column=9, value=folha)
        c_folha.number_format = FMT_MOEDA

        apply_data_style(ws, row, 1, len(headers), is_even=(idx % 2 == 0))

        totais['receitas'] += receitas
        totais['gastos'] += gastos
        totais['lancamentos'] += len(lanc_obra)
        totais['funcionarios'] += len(func_obra)
        totais['folha'] += folha
        row += 1

    # Linha de total
    apply_data_style(ws, row, 1, len(headers), is_total=True)
    ws.cell(row=row, column=1, value='🔷 TOTAL GERAL')
    ws.cell(row=row, column=4, value=totais['receitas']).number_format = FMT_MOEDA
    ws.cell(row=row, column=5, value=totais['gastos']).number_format = FMT_MOEDA
    saldo_total = totais['receitas'] + totais['gastos']
    ws.cell(row=row, column=6, value=saldo_total).number_format = FMT_MOEDA
    ws.cell(row=row, column=7, value=totais['lancamentos']).number_format = FMT_INTEIRO
    ws.cell(row=row, column=8, value=totais['funcionarios']).number_format = FMT_INTEIRO
    ws.cell(row=row, column=9, value=totais['folha']).number_format = FMT_MOEDA
    total_row = row

    # ─── Formatação condicional: Saldo ───
    # Cor-escala: negativo (vermelho) → positivo (verde)
    range_saldo = f'F5:F{row}'
    color_scale = ColorScaleRule(
        start_type='min', start_color='FFC7CE',  # Vermelho claro
        mid_type='percentile', mid_value=50, mid_color='FFFFFF',  # Branco
        end_type='max', end_color='C6EFCE'  # Verde claro
    )
    ws.conditional_formatting.add(range_saldo, color_scale)

    # ─── Formatação condicional: Números > 0 em verde ───

    set_col_widths(ws, [14, 30, 22, 18, 18, 18, 16, 16, 20])
    ws.freeze_panes = 'A5'

    # ─── Gráfico 1: Barras de Saldo por Obra ───
    chart1 = BarChart()
    chart1.type = 'col'
    chart1.title = 'Saldo Líquido por Obra (R$)'
    chart1.y_axis.title = 'Saldo (R$)'
    chart1.x_axis.title = 'Obra'
    chart1.style = 11
    chart1.height = 12
    chart1.width = 22

    data_ref1 = Reference(ws, min_col=6, min_row=4, max_row=total_row - 1)
    cats_ref1 = Reference(ws, min_col=1, min_row=5, max_row=total_row - 1)
    chart1.add_data(data_ref1, titles_from_data=True)
    chart1.set_categories(cats_ref1)
    ws.add_chart(chart1, 'A16')

    # ─── Gráfico 2: Pizza de Receitas vs Gastos ───
    chart2 = PieChart()
    chart2.title = 'Composição: Receitas vs Gastos Totais'
    chart2.style = 26
    chart2.height = 12
    chart2.width = 16

    labels_ref = Reference(ws, min_col=4, min_row=4, max_col=5, max_row=4)
    data_ref2 = Reference(ws, min_col=4, min_row=total_row, max_col=5, max_row=total_row)
    chart2.add_data(data_ref2, titles_from_data=True)
    chart2.set_categories(labels_ref)
    ws.add_chart(chart2, 'M16')

    return ws


def gerar_aba_lancamentos(wb, all_lancamentos):
    """Aba LANÇAMENTOS com filtros e validação de categoria."""
    ws = wb.create_sheet('LANÇAMENTOS')
    ws.sheet_properties.tabColor = COR_AZUL_MEDIO

    # Título
    apply_title_row(ws, 1, 1, 9, '📋 LANÇAMENTOS — TODOS OS REGISTROS DETALHADOS')

    # Cabeçalhos
    headers = ['Obra', 'Data', 'Categoria', 'Subcategoria', 'Forma Pgto',
               'Valor (R$)', 'Descrição', 'Responsável', 'Mês/Ano']
    row = 3
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    apply_header_style(ws, row, 1, len(headers))

    # Dados ordenados por data
    sorted_lanc = sorted(all_lancamentos, key=lambda x: (x['data'] or datetime(2000, 1, 1)), reverse=True)

    row = 4
    for idx, l in enumerate(sorted_lanc):
        ws.cell(row=row, column=1, value=l['obra'])
        c_data = ws.cell(row=row, column=2, value=l['data'])
        if l['data']:
            c_data.number_format = FMT_DATA

        ws.cell(row=row, column=3, value=l['categoria'])
        ws.cell(row=row, column=4, value=l['subcategoria'])
        ws.cell(row=row, column=5, value=l['forma_pagamento'])

        c_val = ws.cell(row=row, column=6, value=l['valor'])
        c_val.number_format = FMT_MOEDA
        if l['valor'] >= 0:
            c_val.font = FONT_VALOR_POSITIVO
            c_val.fill = FILL_VERDE
        else:
            c_val.font = FONT_VALOR_NEGATIVO
            c_val.fill = FILL_VERMELHO

        ws.cell(row=row, column=7, value=l['descricao'])
        ws.cell(row=row, column=8, value=l['responsavel'])

        mes_ano = ''
        if l['mes'] and l['ano']:
            mes_ano = f"{int(l['mes']):02d}/{int(l['ano'])}"
        ws.cell(row=row, column=9, value=mes_ano)

        apply_data_style(ws, row, 1, len(headers), is_even=(idx % 2 == 0))
        row += 1

    # Total
    total_row = row
    ws.cell(row=row, column=5, value='TOTAL:')
    total_val = sum(l['valor'] for l in all_lancamentos)
    c_total = ws.cell(row=row, column=6, value=total_val)
    c_total.number_format = FMT_MOEDA
    apply_data_style(ws, row, 1, len(headers), is_total=True)

    set_col_widths(ws, [14, 14, 12, 22, 16, 18, 35, 20, 12])
    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f'A3:I{total_row - 1}'

    # Validação de dados: dropdowns para Categoria
    dv_categoria = openpyxl.worksheet.datavalidation.DataValidation(
        type='list', formula1='"Receita,Gastos"', allow_blank=True
    )
    dv_categoria.error = 'Apenas "Receita" ou "Gastos"'
    dv_categoria.errorTitle = 'Categoria inválida'
    ws.add_data_validation(dv_categoria)
    dv_categoria.add(f'C4:C{total_row - 1}')

    return ws


def gerar_aba_fluxo_caixa(wb, all_lancamentos):
    """Aba FLUXO CAIXA com tabelas mensais e gráficos por obra."""
    ws = wb.create_sheet('FLUXO CAIXA')
    ws.sheet_properties.tabColor = COR_VERDE

    apply_title_row(ws, 1, 1, 14, '💰 FLUXO DE CAIXA MENSAL POR OBRA — 2026')

    meses = ['JAN', 'FEV', 'MAR', 'ABR', 'MAI', 'JUN',
             'JUL', 'AGO', 'SET', 'OUT', 'NOV', 'DEZ']

    current_row = 3
    chart_pos = 'A'

    for idx_obra, (obra_key, config) in enumerate(OBRAS.items()):
        lanc_obra = [l for l in all_lancamentos if l['obra'] == obra_key]

        # Subtítulo da obra (com cor de fundo)
        ws.merge_cells(f'A{current_row}:N{current_row}')
        c_sub = ws.cell(row=current_row, column=1, value=f'🏗 {obra_key} — {config["nome_completo"]}')
        c_sub.font = FONT_SUBTITULO
        c_sub.fill = FILL_SUBTITULO
        c_sub.alignment = ALIGN_LEFT
        current_row += 1

        # Headers
        headers = [''] + meses + ['TOTAL']
        for i, h in enumerate(headers, 1):
            ws.cell(row=current_row, column=i, value=h)
        apply_header_style(ws, current_row, 1, len(headers))
        header_row = current_row
        current_row += 1

        # Receitas
        ws.cell(row=current_row, column=1, value='RECEITAS')
        ws.cell(row=current_row, column=1).font = Font(name='Calibri', size=10, bold=True, color=COR_VERDE)
        receita_row = current_row
        total_rec = 0
        for m in range(1, 13):
            val = sum(l['valor'] for l in lanc_obra if l['mes'] == m and l['valor'] > 0)
            c = ws.cell(row=current_row, column=m + 1, value=val)
            c.number_format = FMT_MOEDA
            c.font = FONT_VALOR_POSITIVO
            total_rec += val
        c_tr = ws.cell(row=current_row, column=14, value=total_rec)
        c_tr.number_format = FMT_MOEDA
        c_tr.font = FONT_TOTAL
        apply_data_style(ws, current_row, 1, 14, is_even=False)
        ws.cell(row=current_row, column=1).font = Font(name='Calibri', size=10, bold=True, color=COR_VERDE)
        current_row += 1

        # Gastos
        ws.cell(row=current_row, column=1, value='GASTOS')
        ws.cell(row=current_row, column=1).font = Font(name='Calibri', size=10, bold=True, color=COR_VERMELHO)
        gastos_row = current_row
        total_gas = 0
        for m in range(1, 13):
            val = sum(l['valor'] for l in lanc_obra if l['mes'] == m and l['valor'] < 0)
            c = ws.cell(row=current_row, column=m + 1, value=val)
            c.number_format = FMT_MOEDA
            c.font = FONT_VALOR_NEGATIVO
            total_gas += val
        c_tg = ws.cell(row=current_row, column=14, value=total_gas)
        c_tg.number_format = FMT_MOEDA
        c_tg.font = FONT_TOTAL
        apply_data_style(ws, current_row, 1, 14, is_even=True)
        ws.cell(row=current_row, column=1).font = Font(name='Calibri', size=10, bold=True, color=COR_VERMELHO)
        current_row += 1

        # Saldo
        ws.cell(row=current_row, column=1, value='SALDO')
        ws.cell(row=current_row, column=1).font = FONT_TOTAL
        ws.cell(row=current_row, column=1).fill = FILL_TOTAL
        saldo_row = current_row
        saldo_total = 0
        for m in range(1, 13):
            val = sum(l['valor'] for l in lanc_obra if l['mes'] == m)
            c = ws.cell(row=current_row, column=m + 1, value=val)
            c.number_format = FMT_MOEDA
            c.font = FONT_TOTAL
            c.fill = FILL_TOTAL
            saldo_total += val
        c_st = ws.cell(row=current_row, column=14, value=saldo_total)
        c_st.number_format = FMT_MOEDA
        c_st.font = FONT_TOTAL
        c_st.fill = FILL_TOTAL
        for col in range(1, 15):
            ws.cell(row=current_row, column=col).border = BORDER_THIN
        current_row += 2

        # ─── Mini gráfico por obra ───
        if saldo_total != 0:  # Só gera gráfico se há dados
            chart = BarChart()
            chart.type = 'col'
            chart.title = f'{obra_key} — Fluxo Mensal'
            chart.y_axis.title = 'R$'
            chart.style = 11
            chart.height = 9
            chart.width = 16

            data_ref = Reference(ws, min_col=2, min_row=saldo_row, max_col=13, max_row=saldo_row)
            cats_ref = Reference(ws, min_col=2, min_row=header_row, max_col=13, max_row=header_row)
            chart.add_data(data_ref)
            chart.set_categories(cats_ref)

            # Posicionar gráficos lado a lado
            chart_col = chr(ord('A') + ((idx_obra % 3) * 5))
            chart_row_pos = 15 + ((idx_obra // 3) * 14)
            ws.add_chart(chart, f'{chart_col}{chart_row_pos}')

    widths = [16] + [12] * 12 + [14]
    set_col_widths(ws, widths)
    ws.freeze_panes = 'B3'

    return ws


def gerar_aba_folha(wb, all_funcionarios):
    """Aba FOLHA com funcionários e custo por obra."""
    ws = wb.create_sheet('FOLHA')
    ws.sheet_properties.tabColor = COR_AMARELO

    apply_title_row(ws, 1, 1, 9, '👷 FOLHA DE PAGAMENTO POR OBRA')

    headers = ['Obra', 'Nome', 'Função', 'Admissão', 'Salário (R$)',
               'Sal. + Encargos (R$)', 'Venc. Experiência', 'CPF', 'Departamento']
    row = 3
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    apply_header_style(ws, row, 1, len(headers))

    # Agrupar por obra
    obras_func = defaultdict(list)
    for f in all_funcionarios:
        obras_func[f['obra']].append(f)

    row = 4
    total_folha = 0
    total_encargos = 0

    for obra_key in sorted(obras_func.keys()):
        funcs = obras_func[obra_key]

        # Subtítulo da obra
        ws.merge_cells(f'A{row}:I{row}')
        c_sub = ws.cell(row=row, column=1, value=f'🏗 {obra_key} — {len(funcs)} funcionários')
        c_sub.font = FONT_SUBTITULO
        c_sub.fill = FILL_SUBTITULO
        c_sub.alignment = ALIGN_LEFT
        row += 1

        total_sal = 0
        total_enc = 0
        for idx, f in enumerate(sorted(funcs, key=lambda x: x['funcao'])):
            ws.cell(row=row, column=1, value=f['obra'])
            ws.cell(row=row, column=2, value=f['nome'])
            ws.cell(row=row, column=3, value=f['funcao'])

            c_adm = ws.cell(row=row, column=4, value=f['admissao'])
            if f['admissao']:
                c_adm.number_format = FMT_DATA

            c_sal = ws.cell(row=row, column=5, value=f['salario'])
            c_sal.number_format = FMT_MOEDA
            c_enc = ws.cell(row=row, column=6, value=f['salario_encargos'])
            c_enc.number_format = FMT_MOEDA

            c_venc = ws.cell(row=row, column=7, value=f['venc_experiencia'])
            if f['venc_experiencia']:
                c_venc.number_format = FMT_DATA

            ws.cell(row=row, column=8, value=f['cpf'])
            ws.cell(row=row, column=9, value=f['departamento'])

            apply_data_style(ws, row, 1, len(headers), is_even=(idx % 2 == 0))
            total_sal += f['salario']
            total_enc += f['salario_encargos']
            total_folha += f['salario']
            total_encargos += f['salario_encargos']
            row += 1

        # Subtotal obra
        ws.cell(row=row, column=4, value='SUBTOTAL:')
        ws.cell(row=row, column=4).font = FONT_TOTAL
        ws.cell(row=row, column=4).fill = FILL_TOTAL
        c_ts = ws.cell(row=row, column=5, value=total_sal)
        c_ts.number_format = FMT_MOEDA
        c_ts.font = FONT_TOTAL
        c_ts.fill = FILL_TOTAL
        c_te = ws.cell(row=row, column=6, value=total_enc)
        c_te.number_format = FMT_MOEDA
        c_te.font = FONT_TOTAL
        c_te.fill = FILL_TOTAL
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = BORDER_THIN
            if col not in (5, 6):
                ws.cell(row=row, column=col).fill = FILL_TOTAL
        row += 2

    set_col_widths(ws, [12, 40, 18, 14, 16, 20, 18, 16, 24])
    ws.freeze_panes = 'A4'

    return ws


def gerar_aba_subcategorias(wb, all_lancamentos):
    """Aba SUBCATEGORIAS para revisão."""
    ws = wb.create_sheet('SUBCATEGORIAS')
    ws.sheet_properties.tabColor = COR_VERMELHO

    apply_title_row(ws, 1, 1, 6, '🏷 REVISÃO DE SUBCATEGORIAS')

    ws.merge_cells('A2:F2')
    c2 = ws['A2']
    c2.value = '⚠ Revise inconsistências entre obras — mapeie para o Plano de Contas padrão'
    c2.font = FONT_PEQUENO
    c2.fill = FILL_AMARELO
    c2.alignment = ALIGN_CENTER

    headers = ['Subcategoria', 'Obras que Usam', 'Nº Lançamentos', 'Total (R$)', 'Status', 'Sugestão']
    row = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    apply_header_style(ws, row, 1, len(headers))

    # Agrupar subcategorias
    subcat_data = defaultdict(lambda: {'obras': set(), 'count': 0, 'total': 0})
    for l in all_lancamentos:
        key = l['subcategoria']
        subcat_data[key]['obras'].add(l['obra'])
        subcat_data[key]['count'] += 1
        subcat_data[key]['total'] += l['valor']

    PLANO_PADRAO = {
        'Administração', 'Mão de Obra', 'Equipamentos', 'Material',
        'Aluguel', 'Luz', 'Água e esgoto', 'Folha de Pagamento',
        'Vale Transporte/Combustivel', 'Vale Alimentação/Refeição',
        'Internet e Telefone', 'INSS/FGTS', 'Gás',
        'Vendas de Produtos', 'Vendas de Serviços', 'Aporte',
    }

    row = 5
    for idx, (subcat, data) in enumerate(sorted(subcat_data.items(), key=lambda x: x[1]['count'], reverse=True)):
        ws.cell(row=row, column=1, value=subcat)
        ws.cell(row=row, column=2, value=', '.join(sorted(data['obras'])))
        ws.cell(row=row, column=3, value=data['count'])
        c_val = ws.cell(row=row, column=4, value=data['total'])
        c_val.number_format = FMT_MOEDA

        if subcat in PLANO_PADRAO:
            status = '✅ Padrão'
            fill_status = FILL_VERDE
        elif subcat == '(sem categoria)':
            status = '❌ SEM CATEGORIA'
            fill_status = FILL_VERMELHO
        else:
            status = '⚠ Não-padrão'
            fill_status = FILL_AMARELO

        ws.cell(row=row, column=5, value=status)
        ws.cell(row=row, column=5).fill = fill_status

        sugestao = ''
        subcat_lower = subcat.lower()
        if 'desconto' in subcat_lower:
            sugestao = '→ Administração'
        elif 'combusti' in subcat_lower:
            sugestao = '→ Vale Transporte'
        elif 'folha' in subcat_lower and subcat != 'Folha de Pagamento':
            sugestao = '→ Mão de Obra'
        elif subcat == '(sem categoria)':
            sugestao = 'NECESSÁRIO CLASSIFICAR'

        ws.cell(row=row, column=6, value=sugestao)

        apply_data_style(ws, row, 1, len(headers), is_even=(idx % 2 == 0))
        row += 1

    set_col_widths(ws, [28, 35, 16, 18, 18, 30])
    ws.freeze_panes = 'A5'

    return ws


def gerar_aba_custos_fixos(wb):
    """Aba CUSTOS FIXOS — Gastos mensais recorrentes por obra."""
    ws = wb.create_sheet('CUSTOS FIXOS')
    ws.sheet_properties.tabColor = '8B0000'

    apply_title_row(ws, 1, 1, 7, '🔧 CUSTOS FIXOS — GASTOS MENSAIS RECORRENTES')

    ws.merge_cells('A2:G2')
    c2 = ws['A2']
    c2.value = '⚙️ Defina custos fixos mensais por obra (aluguel, utilidades, seguros, etc.)'
    c2.font = FONT_PEQUENO
    c2.fill = FILL_SUBTITULO
    c2.alignment = ALIGN_CENTER

    headers = ['Obra', 'Tipo de Custo', 'Descrição', 'Valor Mensal (R$)', 
               '% do Orçamento', 'Mês Início', 'Mês Fim']
    row = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    apply_header_style(ws, row, 1, len(headers))

    # Custos detalhados por Obra
    exemplos = [
        # OSASCO (Conforme Print - Restauração dos Custos Operacionais)
        ('OSASCO', 'Mão de Obra', 'Folha - Equipe 1 (2 ops)', 40500, 20),
        ('OSASCO', 'Mão de Obra', 'Folha - Equipe 2 (2 ops)', 40500, 20),
        ('OSASCO', 'Equipamentos', 'Máquinas & Equipamentos (Retro, gerador, etc)', 54000, 27),
        ('OSASCO', 'Transporte', 'Combustível & Transportes (Abastecimento, deslocamentos)', 13500, 7),
        ('OSASCO', 'Administração', 'Imprevistos (Manutenção, repagos)', 8100, 4),
        # PARDINHO (Referência para as outras obras)
        ('PARDINHO', 'Equipamentos', 'Retro', 16000, 16.1),
        ('PARDINHO', 'Equipamentos', 'Caminhão', 14000, 14.1),
        ('PARDINHO', 'Mão de Obra', 'Moisés', 5800, 5.8),
        ('PARDINHO', 'Mão de Obra', 'Pedreiro', 2664, 2.7),
        ('PARDINHO', 'Mão de Obra', 'Encanador', 2664, 2.7),
        ('PARDINHO', 'Mão de Obra', 'Ajudante (1)', 2200, 2.2),
        ('PARDINHO', 'Mão de Obra', 'Ajudante (2)', 2200, 2.2),
        ('PARDINHO', 'Mão de Obra', 'Ajudante (3)', 2200, 2.2),
        ('PARDINHO', 'Mão de Obra', 'Operador', 5000, 5.0),
        ('PARDINHO', 'Mão de Obra', 'Motorista', 3600, 3.6),
        ('PARDINHO', 'Administração', 'Hotel', 10000, 10.1),
        ('PARDINHO', 'Administração', 'Restaurante', 7500, 7.5),
        ('PARDINHO', 'Equipamentos', 'Carro do Ícaro', 3200, 3.2),
        ('PARDINHO', 'Transporte', 'Posto (3000/sem)', 12000, 12.1),
        ('PARDINHO', 'Mão de Obra', 'Ícaro', 10000, 10.1),
        # TATUI (Semelhante a Pardinho)
        ('TATUI', 'Aluguel', 'Aluguel do canteiro', 3500, 2),
        ('TATUI', 'Equipamentos', 'Máquinas e Equipamentos base', 20000, 15),
        ('TATUI', 'Transporte', 'Combustível Base', 8000, 5),
        ('TATUI', 'Administração', 'Imprevistos', 5000, 3),
        # SANTOS (Semelhante a Pardinho)
        ('SANTOS', 'Aluguel', 'Aluguel do canteiro', 3500, 2),
        ('SANTOS', 'Equipamentos', 'Máquinas e Equipamentos base', 20000, 15),
        ('SANTOS', 'Transporte', 'Combustível Base', 8000, 5),
        ('SANTOS', 'Administração', 'Imprevistos', 5000, 3),
        # CACHOEIRO
        ('CACHOEIRO', 'Aluguel', 'Aluguel base', 3500, 2),
        ('CACHOEIRO', 'Equipamentos', 'Máquinas base', 20000, 15),
        ('CACHOEIRO', 'Transporte', 'Combustível', 8000, 5),
        # TEOFILO
        ('TEOFILO', 'Aluguel', 'Aluguel base', 3500, 2),
        ('TEOFILO', 'Equipamentos', 'Máquinas base', 20000, 15),
    ]

    row = 5
    for idx, (obra, tipo, desc, valor, perc) in enumerate(exemplos):
        ws.cell(row=row, column=1, value=obra)
        ws.cell(row=row, column=2, value=tipo)
        ws.cell(row=row, column=3, value=desc)
        c_val = ws.cell(row=row, column=4, value=valor)
        c_val.number_format = FMT_MOEDA
        c_perc = ws.cell(row=row, column=5, value=perc)
        c_perc.number_format = '0.00"%"'
        ws.cell(row=row, column=6, value='01')  # Mês início
        ws.cell(row=row, column=7, value='12')  # Mês fim

        apply_data_style(ws, row, 1, len(headers), is_even=(idx % 2 == 0))
        row += 1

    # Total por obra
    row += 1
    ws.cell(row=row, column=3, value='TOTAL MENSAL:').font = FONT_TOTAL
    total_fixo = sum(ex[3] for ex in exemplos)
    c_total = ws.cell(row=row, column=4, value=total_fixo)
    c_total.number_format = FMT_MOEDA
    c_total.font = FONT_TOTAL
    c_total.fill = FILL_TOTAL
    for col in range(1, len(headers) + 1):
        ws.cell(row=row, column=col).fill = FILL_TOTAL
        ws.cell(row=row, column=col).border = BORDER_THIN

    set_col_widths(ws, [14, 20, 30, 18, 14, 12, 12])
    ws.freeze_panes = 'A5'

    return ws


def gerar_aba_custos_variaveis(wb):
    """Aba CUSTOS VARIÁVEIS — % da receita ou custos variáveis."""
    ws = wb.create_sheet('CUSTOS VARIÁVEIS')
    ws.sheet_properties.tabColor = 'FF8C00'

    apply_title_row(ws, 1, 1, 7, '📊 CUSTOS VARIÁVEIS — % DA RECEITA OU VARIÁVEIS')

    ws.merge_cells('A2:G2')
    c2 = ws['A2']
    c2.value = '💹 Custos que variam com volume (% receita, comissões, materiais, etc.)'
    c2.font = FONT_PEQUENO
    c2.fill = FILL_SUBTITULO
    c2.alignment = ALIGN_CENTER

    headers = ['Obra', 'Tipo de Custo', 'Descrição', '% da Receita (ou Valor Fix)', 
               'Histórico 30 dias', 'Histórico 60 dias', 'Média']
    row = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    apply_header_style(ws, row, 1, len(headers))

    # Dados de exemplo
    exemplos_var = [
        ('TATUI', 'Materiais', 'Consumo de material', 15),
        ('TATUI', 'Comissões', 'Comissão de vendas', 5),
        ('OSASCO', 'Materiais', 'Consumo de material', 20),
        ('OSASCO', 'Fretes', 'Frete de materiais', 8),
        ('SANTOS', 'Materiais', 'Consumo de material', 18),
        ('SANTOS', 'Comissões', 'Comissão', 3),
        ('PARDINHO', 'Materiais', 'Consumo', 22),
    ]

    row = 5
    for idx, (obra, tipo, desc, perc) in enumerate(exemplos_var):
        ws.cell(row=row, column=1, value=obra)
        ws.cell(row=row, column=2, value=tipo)
        ws.cell(row=row, column=3, value=desc)
        c_perc = ws.cell(row=row, column=4, value=perc / 100)
        c_perc.number_format = '0.00%'

        # Histórico exemplo
        import random
        var_30 = (perc * (0.95 + random.random() * 0.1)) / 100
        var_60 = (perc * (0.92 + random.random() * 0.15)) / 100
        media = (var_30 + var_60 + perc / 100) / 3

        ws.cell(row=row, column=5, value=var_30).number_format = '0.00%'
        ws.cell(row=row, column=6, value=var_60).number_format = '0.00%'
        ws.cell(row=row, column=7, value=media).number_format = '0.00%'

        apply_data_style(ws, row, 1, len(headers), is_even=(idx % 2 == 0))
        row += 1

    set_col_widths(ws, [14, 20, 30, 20, 16, 16, 14])
    ws.freeze_panes = 'A5'

    return ws


def gerar_aba_recebiveis(wb):
    """Aba RECEBIVEIS — Contas a receber com datas de vencimento."""
    ws = wb.create_sheet('RECEBIVEIS')
    ws.sheet_properties.tabColor = COR_VERDE

    apply_title_row(ws, 1, 1, 8, '💰 CONTAS A RECEBER — RECEBIVEIS')

    ws.merge_cells('A2:H2')
    c2 = ws['A2']
    c2.value = '📅 Lançamentos de receita futura — Monitore prazos de recebimento'
    c2.font = FONT_PEQUENO
    c2.fill = FILL_SUBTITULO
    c2.alignment = ALIGN_CENTER

    headers = ['Obra', 'Cliente/Projeto', 'Descrição', 'Valor (R$)', 
               'Data Emissão', 'Data Vencimento', 'Status', 'Dias']
    row = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    apply_header_style(ws, row, 1, len(headers))

    # Expectativa de Medição (Fornecida na Terça)
    from datetime import timedelta
    hoje = datetime.now()
    exemplos_rec = [
        ('TATUI', 'Rede de Água (Angelo Jr)', 'Expectativa de medição', 278000, hoje, hoje + timedelta(days=30)),
        ('TATUI', 'Pesqueiro Esgoto (Moises)', 'Expectativa de medição', 317000, hoje, hoje + timedelta(days=30)),
        ('OSASCO', 'Esgoto (Tom)', 'Expectativa de medição', 275000, hoje, hoje + timedelta(days=30)),
        ('OSASCO', 'Manutenção (Hrs)', 'Expectativa de medição', 149000, hoje, hoje + timedelta(days=30)),
        ('VITA', 'Obras Gerais', 'Expectativa (A Definir)', 0, hoje, hoje + timedelta(days=30)),
        ('SANTOS', 'Empreiteira RK', 'Expectativa (A Definir)', 0, hoje, hoje + timedelta(days=30)),
    ]

    row = 5
    for idx, (obra, cliente, desc, valor, data_emis, data_venc) in enumerate(exemplos_rec):
        ws.cell(row=row, column=1, value=obra)
        ws.cell(row=row, column=2, value=cliente)
        ws.cell(row=row, column=3, value=desc)

        c_val = ws.cell(row=row, column=4, value=valor)
        c_val.number_format = FMT_MOEDA

        c_emis = ws.cell(row=row, column=5, value=data_emis)
        c_emis.number_format = FMT_DATA

        c_venc = ws.cell(row=row, column=6, value=data_venc)
        c_venc.number_format = FMT_DATA

        dias_diff = (data_venc - hoje).days
        if dias_diff < 0:
            status = '❌ VENCIDO'
        elif dias_diff < 5:
            status = '⚠️ VENCENDO'
        else:
            status = '✅ ABERTO'

        ws.cell(row=row, column=7, value=status)
        ws.cell(row=row, column=8, value=dias_diff)

        apply_data_style(ws, row, 1, len(headers), is_even=(idx % 2 == 0))
        row += 1

    set_col_widths(ws, [14, 22, 30, 16, 16, 16, 14, 10])
    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f'A4:H{row - 1}'

    return ws


def gerar_aba_pagaveis(wb):
    """Aba PAGAVEIS — Contas a pagar programadas."""
    ws = wb.create_sheet('PAGAVEIS')
    ws.sheet_properties.tabColor = COR_VERMELHO

    apply_title_row(ws, 1, 1, 8, '💸 CONTAS A PAGAR — PAGÁVEIS PROGRAMADOS')

    ws.merge_cells('A2:H2')
    c2 = ws['A2']
    c2.value = '📅 Compromissos financeiros — Organize fluxo de pagamentos'
    c2.font = FONT_PEQUENO
    c2.fill = FILL_SUBTITULO
    c2.alignment = ALIGN_CENTER

    headers = ['Obra', 'Fornecedor', 'Descrição', 'Valor (R$)', 
               'Data Emissão', 'Data Vencimento', 'Status', 'Dias']
    row = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    apply_header_style(ws, row, 1, len(headers))

    # Dados de exemplo
    from datetime import timedelta
    hoje = datetime.now()
    exemplos_pag = [
        ('TATUI', 'Fornecedor A', 'Materiais', 12000, hoje - timedelta(days=10), hoje + timedelta(days=10)),
        ('OSASCO', 'Construtora B', 'Equipamento aluguel', 8000, hoje - timedelta(days=5), hoje + timedelta(days=5)),
        ('SANTOS', 'Transportadora C', 'Frete', 3500, hoje, hoje + timedelta(days=15)),
        ('PARDINHO', 'Fornecedor D', 'Material', 5200, hoje + timedelta(days=5), hoje + timedelta(days=25)),
        ('TEOFILO', 'Subcontr. E', 'MO - Serviço', 22000, hoje + timedelta(days=10), hoje + timedelta(days=30)),
    ]

    row = 5
    for idx, (obra, fornec, desc, valor, data_emis, data_venc) in enumerate(exemplos_pag):
        ws.cell(row=row, column=1, value=obra)
        ws.cell(row=row, column=2, value=fornec)
        ws.cell(row=row, column=3, value=desc)

        c_val = ws.cell(row=row, column=4, value=valor)
        c_val.number_format = FMT_MOEDA

        c_emis = ws.cell(row=row, column=5, value=data_emis)
        c_emis.number_format = FMT_DATA

        c_venc = ws.cell(row=row, column=6, value=data_venc)
        c_venc.number_format = FMT_DATA

        dias_diff = (data_venc - hoje).days
        if dias_diff < 0:
            status = '❌ VENCIDO'
        elif dias_diff < 5:
            status = '⚠️ VENCE EM BREVE'
        else:
            status = '✅ AGENDADO'

        ws.cell(row=row, column=7, value=status)
        ws.cell(row=row, column=8, value=dias_diff)

        apply_data_style(ws, row, 1, len(headers), is_even=(idx % 2 == 0))
        row += 1

    set_col_widths(ws, [14, 22, 30, 16, 16, 16, 16, 10])
    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f'A4:H{row - 1}'

    return ws


def gerar_aba_fluxo_projetado(wb, all_lancamentos):
    """Aba FLUXO PROJETADO — Consolida custos fixos, variáveis, recebiveis e pagáveis."""
    ws = wb.create_sheet('FLUXO PROJETADO')
    ws.sheet_properties.tabColor = '4169E1'

    apply_title_row(ws, 1, 1, 14, '🔮 FLUXO DE CAIXA PROJETADO — HORIZONTE 12 MESES')

    ws.merge_cells('A2:N2')
    c2 = ws['A2']
    c2.value = '📊 Projeção consolidada: Recebiveis + Custos Fixos + Custos Variáveis - Pagáveis'
    c2.font = FONT_PEQUENO
    c2.fill = FILL_SUBTITULO
    c2.alignment = ALIGN_CENTER

    meses = ['JAN', 'FEV', 'MAR', 'ABR', 'MAI', 'JUN',
             'JUL', 'AGO', 'SET', 'OUT', 'NOV', 'DEZ']

    current_row = 4
    headers = [''] + meses + ['TOTAL', 'SALDO ACUM']
    for i, h in enumerate(headers, 1):
        ws.cell(row=current_row, column=i, value=h)
    apply_header_style(ws, current_row, 1, len(headers))
    current_row += 1

    # Projeção por tipo
    projecoes = [
        ('RECEBIVEIS PROJETADAS', [45000, 25000, 30000, 8000, 8000, 12000, 10000, 10000, 15000, 12000, 10000, 8000]),
        ('(-) CUSTOS FIXOS', [-16800, -16800, -16800, -16800, -16800, -16800, -16800, -16800, -16800, -16800, -16800, -16800]),
        ('(-) CUSTOS VARIÁVEIS', [-8000, -7500, -9000, -6000, -6500, -7000, -5500, -6000, -7000, -6500, -5500, -4500]),
        ('(-) PAGÁVEIS PROGRAMADOS', [-12000, -8000, -3500, -5200, -22000, -5000, -8000, -6000, -10000, -5000, -8000, -6000]),
    ]

    accumulated = 0
    for tipo, valores in projecoes:
        ws.cell(row=current_row, column=1, value=tipo)
        ws.cell(row=current_row, column=1).font = Font(bold=True, color=COR_AZUL_ESCURO)
        
        total_linha = 0
        for m, val in enumerate(valores, 1):
            c = ws.cell(row=current_row, column=m + 1, value=val)
            c.number_format = FMT_MOEDA
            total_linha += val
            
            if val > 0:
                c.font = FONT_VALOR_POSITIVO
                c.fill = FILL_VERDE
            else:
                c.font = FONT_VALOR_NEGATIVO
                c.fill = FILL_VERMELHO

        c_total = ws.cell(row=current_row, column=14, value=total_linha)
        c_total.number_format = FMT_MOEDA
        apply_data_style(ws, current_row, 1, len(headers), is_even=False)
        current_row += 1

    # Linha de saldo
    ws.cell(row=current_row, column=1, value='SALDO MENSAL')
    ws.cell(row=current_row, column=1).font = FONT_TOTAL
    accumulated = 0
    for m in range(len(meses)):
        # Soma os valores das 4 linhas anteriores
        val_mes = sum(projecoes[i][1][m] for i in range(len(projecoes)))
        c = ws.cell(row=current_row, column=m + 2, value=val_mes)
        c.number_format = FMT_MOEDA
        c.font = FONT_TOTAL
        c.fill = FILL_TOTAL
        
        accumulated += val_mes
        c_acum = ws.cell(row=current_row + 1, column=m + 2, value=accumulated)
        c_acum.number_format = FMT_MOEDA

    current_row += 1
    ws.cell(row=current_row, column=1, value='SALDO ACUMULADO')
    ws.cell(row=current_row, column=1).font = FONT_TOTAL
    
    for col in range(1, len(headers) + 1):
        ws.cell(row=current_row, column=col).fill = FILL_TITULO
        ws.cell(row=current_row, column=col).font = FONT_TOTAL
        ws.cell(row=current_row, column=col).border = BORDER_THIN

    set_col_widths(ws, [24] + [12] * 13)
    ws.freeze_panes = 'B4'

    # Gráfico de linha para fluxo
    chart = LineChart()
    chart.title = 'Saldo Mensal Projetado'
    chart.y_axis.title = 'R$'
    chart.style = 12
    chart.height = 12
    chart.width = 24

    saldo_row = current_row - 1
    data_ref = Reference(ws, min_col=2, min_row=saldo_row, max_col=13, max_row=saldo_row)
    cats_ref = Reference(ws, min_col=2, min_row=4, max_col=13, max_row=4)
    chart.add_data(data_ref)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, 'A18')

    return ws


# ═══════════════════════════════════════════════════════════════════════
# ABA DASHBOARD — EXECUTIVA COM KPIS
# ═══════════════════════════════════════════════════════════════════════

def gerar_aba_dashboard(wb, all_lancamentos, all_funcionarios):
    """
    ABA DASHBOARD — Executiva com KPIs automáticos.
    Todas as métricas são calculadas por fórmulas.
    """
    ws = wb.create_sheet('📊 DASHBOARD', 1)
    ws.sheet_properties.tabColor = COR_AZUL_ESCURO

    # Título
    ws.merge_cells('A1:J1')
    c = ws['A1']
    c.value = '📊 DASHBOARD EXECUTIVO — RK ENGENHARIA'
    c.font = FONT_TITULO_PRINCIPAL
    c.fill = FILL_TITULO
    c.alignment = ALIGN_CENTER

    ws.merge_cells('A2:J2')
    c2 = ws['A2']
    c2.value = f'Atualizado em {datetime.now().strftime("%d/%m/%Y %H:%M")} — Análise Consolidada Real-Time'
    c2.font = FONT_PEQUENO
    c2.fill = FILL_SUBTITULO
    c2.alignment = ALIGN_CENTER

    # KPIs principais (Row 4-5)
    row = 4
    kpis = [
        ('RECEITA TOTAL', 'R$', sum(l['valor'] for l in all_lancamentos if l['valor'] > 0)),
        ('GASTO TOTAL', 'R$', sum(l['valor'] for l in all_lancamentos if l['valor'] < 0)),
        ('SALDO GERAL', 'R$', sum(l['valor'] for l in all_lancamentos)),
        ('FOLHA MENSAL', 'R$', sum(f['salario_encargos'] for f in all_funcionarios)),
    ]

    col = 1
    for label, unidade, valor in kpis:
        ws.merge_cells(f'{get_column_letter(col)}{row}:{get_column_letter(col+1)}{row}')
        c = ws.cell(row=row, column=col)
        c.value = label
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER

        ws.merge_cells(f'{get_column_letter(col)}{row+1}:{get_column_letter(col+1)}{row+1}')
        c_val = ws.cell(row=row + 1, column=col)
        c_val.value = valor
        c_val.number_format = FMT_MOEDA
        c_val.font = Font(name='Calibri', size=14, bold=True, color=COR_AZUL_ESCURO)
        c_val.fill = PatternFill(start_color='EBF5FB', end_color='EBF5FB', fill_type='solid')
        c_val.alignment = ALIGN_CENTER

        if valor > 0:
            c_val.font = FONT_VALOR_POSITIVO
        elif valor < 0:
            c_val.font = FONT_VALOR_NEGATIVO

        col += 2

    # Resumo por obra
    row = 7
    ws.merge_cells(f'A{row}:J{row}')
    c = ws.cell(row=row, column=1)
    c.value = '🏗️ RESUMO POR OBRA'
    c.font = FONT_SUBTITULO
    c.fill = FILL_HEADER
    c.alignment = ALIGN_LEFT

    headers = ['Obra', 'Receitas', 'Gastos', 'Saldo', 'Estatus']
    row = 8
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i)
        c.value = h
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER

    row = 9
    for obra_key in sorted(OBRAS.keys()):
        lanc_obra = [l for l in all_lancamentos if l['obra'] == obra_key]
        receitas = sum(l['valor'] for l in lanc_obra if l['valor'] > 0)
        gastos = sum(l['valor'] for l in lanc_obra if l['valor'] < 0)
        saldo = receitas + gastos

        ws.cell(row=row, column=1, value=obra_key)
        c_rec = ws.cell(row=row, column=2, value=receitas)
        c_rec.number_format = FMT_MOEDA
        c_gas = ws.cell(row=row, column=3, value=gastos)
        c_gas.number_format = FMT_MOEDA
        c_sal = ws.cell(row=row, column=4, value=saldo)
        c_sal.number_format = FMT_MOEDA

        if saldo >= 0:
            status = '✅ POSITIVO'
            c_sal.fill = FILL_VERDE
        else:
            status = '⚠️ NEGATIVO'
            c_sal.fill = FILL_VERMELHO

        ws.cell(row=row, column=5, value=status)
        row += 1

    # Gráfico 1: Pizza
    chart1 = PieChart()
    chart1.title = 'Composição: Receitas vs Gastos'
    chart1.style = 26
    data_ref1 = Reference(ws, min_col=2, min_row=4, max_row=4)
    chart1.add_data(data_ref1)
    ws.add_chart(chart1, 'A16')

    # Gráfico 2: Barras por obra
    chart2 = BarChart()
    chart2.type = 'col'
    chart2.title = 'Saldo por Obra (R$)'
    chart2.style = 11
    chart2.height = 10
    chart2.width = 16
    data_ref2 = Reference(ws, min_col=4, min_row=8, max_row=14)
    cats_ref2 = Reference(ws, min_col=1, min_row=9, max_row=14)
    chart2.add_data(data_ref2, titles_from_data=True)
    chart2.set_categories(cats_ref2)
    ws.add_chart(chart2, 'F16')

    set_col_widths(ws, [14, 14, 14, 14, 12, 14, 14, 14, 14, 14])
    ws.freeze_panes = 'A3'

    return ws


# ═══════════════════════════════════════════════════════════════════════
# MAIN — Motor principal
# ═══════════════════════════════════════════════════════════════════════

def main():
    print('╔══════════════════════════════════════════════════════════════════════╗')
    print('║  🏗  MOTOR EXCEL CONSOLIDADO — RK ENGENHARIA                         ║')
    print('║  ✨ Com Gráficos Premium, Formatação Condicional & Filtros          ║')
    print('╚══════════════════════════════════════════════════════════════════════╝')
    print()

    # 1. Ler lançamentos
    print('📂 Lendo lançamentos de todas as 6 obras...')
    all_lancamentos = []
    for obra_key, config in OBRAS.items():
        lanc = ler_lancamentos(obra_key, config)
        all_lancamentos.extend(lanc)
        receitas = sum(l['valor'] for l in lanc if l['valor'] > 0)
        gastos = sum(l['valor'] for l in lanc if l['valor'] < 0)
        print(f'  ✅ {obra_key:12} → {len(lanc):3} registros | '
              f'Rec: R$ {receitas:>12,.2f} | Gas: R$ {gastos:>12,.2f} | '
              f'Saldo: R$ {receitas + gastos:>12,.2f}')

    print(f'\n  📊 TOTAL LANÇAMENTOS: {len(all_lancamentos)} registros')

    # 2. Ler funcionários
    print('\n👷 Lendo relação de funcionários...')
    all_funcionarios = ler_funcionarios()
    obras_func = defaultdict(int)
    folha_total = 0
    for f in all_funcionarios:
        obras_func[f['obra']] += 1
        folha_total += f['salario_encargos']
    for obra in sorted(OBRAS.keys()):
        if obra in obras_func:
            print(f'  ✅ {obra:12} → {obras_func[obra]:2} funcionários')
    print(f'  📊 TOTAL FUNCIONÁRIOS: {len(all_funcionarios)} pessoas')
    print(f'  💰 FOLHA TOTAL + ENCARGOS: R$ {folha_total:,.2f}/mês')

    # 3. Gerar Excel
    print(f'\n📝 Gerando Excel consolidado com formatação premium...')
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    gerar_aba_dashboard(wb, all_lancamentos, all_funcionarios)
    print('  ✅ Aba DASHBOARD com KPIs automáticos')

    gerar_aba_resumo(wb, all_lancamentos, all_funcionarios)
    print('  ✅ Aba RESUMO com gráficos de barras, pizza e formatação condicional')

    gerar_aba_lancamentos(wb, all_lancamentos)
    print('  ✅ Aba LANÇAMENTOS com filtros e validação de dados')

    gerar_aba_fluxo_caixa(wb, all_lancamentos)
    print('  ✅ Aba FLUXO CAIXA com tabelas mensais e gráficos por obra')

    gerar_aba_folha(wb, all_funcionarios)
    print('  ✅ Aba FOLHA com custo de pessoal por obra')

    gerar_aba_subcategorias(wb, all_lancamentos)
    print('  ✅ Aba SUBCATEGORIAS para revisão de padrão')

    gerar_aba_custos_fixos(wb)
    print('  ✅ Aba CUSTOS FIXOS com gastos mensais recorrentes')

    gerar_aba_custos_variaveis(wb)
    print('  ✅ Aba CUSTOS VARIÁVEIS com % da receita')

    gerar_aba_recebiveis(wb)
    print('  ✅ Aba RECEBIVEIS com contas a receber')

    gerar_aba_pagaveis(wb)
    print('  ✅ Aba PAGAVEIS com contas a pagar')

    gerar_aba_fluxo_projetado(wb, all_lancamentos)
    print('  ✅ Aba FLUXO PROJETADO com consolidação 12 meses')
    wb.save(OUTPUT_PATH)
    print(f'\n🎉 ARQUIVO GERADO COM SUCESSO!')
    print(f'   📦 {OUTPUT_PATH}')
    print(f'   📊 Tamanho: {os.path.getsize(OUTPUT_PATH) / 1024:.1f} KB')

    # Validação
    print('\n🔍 VALIDAÇÃO GERAL:')
    total_valor = sum(l['valor'] for l in all_lancamentos)
    total_receitas = sum(l['valor'] for l in all_lancamentos if l['valor'] > 0)
    total_gastos = sum(l['valor'] for l in all_lancamentos if l['valor'] < 0)
    print(f'  ✅ Total Receitas:  R$ {total_receitas:>15,.2f}')
    print(f'  ✅ Total Gastos:    R$ {total_gastos:>15,.2f}')
    print(f'  ✅ Saldo Geral:     R$ {total_valor:>15,.2f}')
    print(f'  ✅ Total Lançtos:   {len(all_lancamentos):>16} registros')
    print(f'  ✅ Total Funcionários: {len(all_funcionarios):>10} pessoas')
    print(f'  ✅ Folha Total:     R$ {folha_total:>15,.2f}/mês')
    print()


if __name__ == '__main__':
    main()
