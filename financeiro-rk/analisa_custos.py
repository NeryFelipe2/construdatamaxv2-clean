#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import pandas as pd
from openpyxl import load_workbook
import os

os.chdir('c:\\Users\\felip\\Desktop\\_ORGANIZADO\\01-OBRAS-RK')

# Lista de projetos e seus arquivos
projetos = {
    'PARDINHO': 'PARDINHO/Pardinho - Controle de Obras ENGELFER-RK.xlsm',
    'OSASCO': 'OSASCO/Osasco - Controle de Obras ENGELFER-RK.xlsm',
    'CACHOEIRO': 'CACHOEIRO/Cachoeiro - Controle de Obras ENGELFER-RK  - Copia - Copia.xlsm',
    'SANTOS': 'SANTOS EMPREITA RK/Santos - Controle de Obras ENGELFER-RK .xlsm',
}

print("=" * 80)
print("ANÁLISE DE CUSTOS FIXOS - TODOS OS PROJETOS")
print("=" * 80)

for nome, caminho in projetos.items():
    if not os.path.exists(caminho):
        print(f"\n❌ {nome}: Arquivo não encontrado em {caminho}")
        continue
    
    print(f"\n✓ {nome}: {caminho}")
    try:
        wb = load_workbook(caminho)
        print(f"  Abas disponíveis: {', '.join(wb.sheetnames[:8])}")
        
        # Procurar por "custo fixo" em todas as abas
        for aba in wb.sheetnames:
            ws = wb[aba]
            for row in ws.iter_rows(min_row=1, max_row=100, values_only=True):
                linha_texto = ' '.join([str(x).lower() for x in row if x])
                if 'custo' in linha_texto and 'fixo' in linha_texto:
                    print(f"  ➜ ENCONTRADO EM '{aba}':")
                    print(f"    {row}")
                    
    except Exception as e:
        print(f"  ⚠️ Erro ao ler: {e}")

print("\n" + "=" * 80)
print("CUSTOS FIXOS INFORMADOS DE PARDINHO:")
print("=" * 80)

custos_pardinho = {
    'Retro': 16000,
    'Caminhão': 14000,
    'Moisés': 5800,
    'Pedreiro': 2664,
    'Encanador': 2664,
    'Ajudante 1': 2200,
    'Ajudante 2': 2200,
    'Ajudante 3': 2200,
    'Operador': 5000,
    'Motorista': 3600,
    'Hotel': 10000,
    'Restaurante': 7500,
    'Carro do Ícaro': 3200,
    'Posto': 3000,
    'Ícaro': 10000,
}

total = 0
for item, valor in custos_pardinho.items():
    print(f"  {item:.<30} R$ {valor:>10,.2f}")
    total += valor

print(f"  {'-'*40}")
print(f"  {'TOTAL MENSAL':.<30} R$ {total:>10,.2f}")
print(f"  {'TOTAL ANUAL':.<30} R$ {total * 12:>10,.2f}")
