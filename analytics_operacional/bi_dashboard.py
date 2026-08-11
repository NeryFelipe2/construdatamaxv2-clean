"""BI operacional para Analytics/XGBoost."""
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from analytics_operacional.xgboost_responsaveis import OUT_JSON, RESPONSAVEIS, _dataset, rodar_xgboost_responsaveis

ROOT = Path(__file__).resolve().parents[1]
OUT_BI_PDF = ROOT / "BI_ANALYTICS_OPERACIONAL_20260426.pdf"
OUT_BI_XLSX = ROOT / "BI_ANALYTICS_OPERACIONAL_20260426.xlsx"


def _load_json() -> dict[str, Any] | None:
    if not OUT_JSON.exists():
        return None
    try:
        return json.loads(OUT_JSON.read_text(encoding="utf-8"))
    except Exception:
        return None


def _series() -> dict[str, list[dict[str, Any]]]:
    out: dict[str, list[dict[str, Any]]] = {}
    for key, cfg in RESPONSAVEIS.items():
        df = _dataset(cfg["nucleos"])
        if df.empty:
            out[key] = []
            continue
        out[key] = [
            {
                "data": str(row["data"]),
                "producao_equiv": float(row["producao_equiv"]),
                "producao_m": float(row["producao_m"]),
                "producao_un": float(row["producao_un"]),
                "producao_etapa": float(row.get("producao_etapa") or 0),
                "bloqueio_score": int(row["bloqueio_score"]),
                "atividade_score": int(row["atividade_score"]),
            }
            for _, row in df.iterrows()
        ]
    return out


def _risco(modelo: dict[str, Any]) -> str:
    conf = float(modelo.get("confianca") or 0)
    r2 = modelo.get("r2")
    if conf < 0.5 or (isinstance(r2, (int, float)) and r2 < 0):
        return "ALTO"
    if conf < 0.75:
        return "MEDIO"
    return "BAIXO"


def _resumo_executivo(cards: dict[str, Any], comparativo: list[dict[str, Any]]) -> list[str]:
    total_equiv = cards.get("total_equiv", 0)
    notas = [
        f"Base consolidada com {cards['rdos']} RDOs e {total_equiv} unidades equivalentes de producao.",
        "O BI agora separa metros, unidades e etapas para nao esconder a producao textual dos RDOs.",
        f"Previsao operacional dos proximos 7 dias: {cards['previsao_7_dias']} unidades equivalentes.",
    ]
    for item in comparativo:
        nome = item.get("responsavel") or item.get("chave")
        notas.append(
            f"{nome}: {item.get('rdos', 0)} RDOs, {item.get('total_equiv', 0)} equiv., "
            f"risco {item.get('risco')} e confianca {item.get('confianca')}."
        )
    return notas


def _acoes_operacionais() -> list[str]:
    return [
        "Engenheiro envia texto diario com RDO, mao de obra, custo direto/indireto, desvios e producao planejada.",
        "Operacao cola em Preencher com Texto; o sistema estrutura RDO, planejamento, custos, desvios e fluxo.",
        "Analytics recalcula XGBoost por responsavel e atualiza tendencia, risco, confianca e previsao semanal.",
        "Gestao usa o PDF/Excel para cobrar desvios, medicao prevista, recebimento e caixa projetado.",
    ]


def montar_bi_payload(recalcular: bool = False) -> dict[str, Any]:
    base = rodar_xgboost_responsaveis() if recalcular else (_load_json() or rodar_xgboost_responsaveis())
    resultados = base.get("resultados", {})
    total_rdos = sum(int(r.get("rdos") or 0) for r in resultados.values())
    total_m = sum(float(r.get("total_m") or 0) for r in resultados.values())
    total_un = sum(float(r.get("total_un") or 0) for r in resultados.values())
    total_etapa = sum(float(r.get("total_etapa") or 0) for r in resultados.values())
    total_equiv = sum(float(r.get("total_equiv") or 0) for r in resultados.values())
    previsao_7 = sum(float((r.get("modelo") or {}).get("previsao_7_dias") or 0) for r in resultados.values())
    cards = {
        "rdos": total_rdos,
        "total_m": round(total_m, 2),
        "total_un": round(total_un, 2),
        "total_etapa": round(total_etapa, 2),
        "total_equiv": round(total_equiv, 2),
        "previsao_7_dias": round(previsao_7, 2),
        "responsaveis": len(resultados),
    }
    comparativo = []
    for key, r in resultados.items():
        modelo = r.get("modelo") or {}
        comparativo.append({
            "chave": key,
            "responsavel": r.get("responsavel"),
            "rdos": r.get("rdos"),
            "periodo": r.get("periodo"),
            "total_m": r.get("total_m"),
            "total_un": r.get("total_un"),
            "total_etapa": r.get("total_etapa"),
            "total_equiv": r.get("total_equiv"),
            "media_dia_equiv": r.get("media_dia_equiv"),
            "previsao_proximo_dia": modelo.get("previsao_proximo_dia"),
            "previsao_7_dias": modelo.get("previsao_7_dias"),
            "confianca": modelo.get("confianca"),
            "mae": modelo.get("mae"),
            "r2": modelo.get("r2"),
            "risco": _risco(modelo),
            "observacao": r.get("observacao"),
            "importancias": modelo.get("importancias") or [],
        })
    return {
        "gerado_em": datetime.now().isoformat(timespec="seconds"),
        "fonte": str(OUT_JSON),
        "cards": cards,
        "comparativo": comparativo,
        "series": _series(),
        "resumo_executivo": _resumo_executivo(cards, comparativo),
        "acoes_operacionais": _acoes_operacionais(),
        "modelo_raw": resultados,
    }


def _style_header(row: Any, fill: str = "123A75") -> None:
    for cell in row:
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")


def _apply_sheet_style(wb: Workbook) -> None:
    thin = Side(style="thin", color="D9E2EF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col in range(1, sheet.max_column + 1):
            values = [str(sheet.cell(row, col).value or "") for row in range(1, min(sheet.max_row, 35) + 1)]
            width = min(42, max(12, max(len(v) for v in values) + 3))
            sheet.column_dimensions[get_column_letter(col)].width = width


def _serie_agregada(payload: dict[str, Any]) -> list[list[Any]]:
    datas = sorted({row["data"] for serie in payload["series"].values() for row in serie})
    headers = ["Data"] + list(payload["series"].keys()) + ["Total"]
    rows = [headers]
    for data in datas:
        vals = []
        for key in payload["series"].keys():
            vals.append(round(sum(float(r["producao_equiv"] or 0) for r in payload["series"][key] if r["data"] == data), 2))
        rows.append([data] + vals + [round(sum(vals), 2)])
    return rows


def exportar_bi_excel(payload: dict[str, Any] | None = None) -> Path:
    payload = payload or montar_bi_payload(False)
    wb = Workbook()
    dash = wb.active
    dash.title = "Dashboard"
    fill_h = "123A75"

    dash.append(["ConstruData BI Analytics Operacional", "", "", "Gerado em", payload["gerado_em"]])
    dash.merge_cells("A1:C1")
    dash["A1"].font = Font(size=16, bold=True, color="123A75")
    dash["D1"].font = Font(bold=True)
    dash.append([])
    dash.append(["Indicador", "Valor", "Leitura"])
    _style_header(dash[3], fill_h)
    cards = payload["cards"]
    metricas = [
        ["RDOs", cards["rdos"], "Historico consolidado"],
        ["Producao equiv.", cards["total_equiv"], "m + un + etapas"],
        ["Total m", cards["total_m"], "Producao linear"],
        ["Total un", cards["total_un"], "Unidades executadas"],
        ["Total etapas", cards.get("total_etapa", 0), "RDO textual estruturado"],
        ["Previsao 7 dias", cards["previsao_7_dias"], "Saida do XGBoost"],
    ]
    for row in metricas:
        dash.append(row)
    dash.append([])
    dash.append(["Resumo executivo"])
    dash["A11"].font = Font(bold=True, color="123A75")
    for nota in payload["resumo_executivo"]:
        dash.append([nota])
    dash.append([])
    dash.append(["Rotina operacional"])
    dash[f"A{dash.max_row}"].font = Font(bold=True, color="123A75")
    for acao in payload["acoes_operacionais"]:
        dash.append([acao])

    ws2 = wb.create_sheet("Comparativo")
    headers = [
        "Responsavel", "RDOs", "Periodo", "Total m", "Total un", "Etapas", "Equiv",
        "Media dia", "Prev prox dia", "Prev 7 dias", "Confianca", "MAE", "R2", "Risco", "Observacao",
    ]
    ws2.append(headers)
    _style_header(ws2[1], fill_h)
    for item in payload["comparativo"]:
        ws2.append([
            item["responsavel"], item["rdos"], item["periodo"], item["total_m"], item["total_un"],
            item.get("total_etapa"), item.get("total_equiv"), item["media_dia_equiv"],
            item["previsao_proximo_dia"], item["previsao_7_dias"], item["confianca"],
            item["mae"], item["r2"], item["risco"], item["observacao"],
        ])

    ws3 = wb.create_sheet("Serie Diaria")
    ws3.append(["Responsavel", "Data", "Producao equiv", "Producao m", "Producao un", "Etapas", "Bloqueios", "Atividade"])
    _style_header(ws3[1], fill_h)
    for key, serie in payload["series"].items():
        for row in serie:
            ws3.append([key, row["data"], row["producao_equiv"], row["producao_m"], row["producao_un"], row.get("producao_etapa", 0), row["bloqueio_score"], row["atividade_score"]])

    ws4 = wb.create_sheet("Serie BI")
    for row in _serie_agregada(payload):
        ws4.append(row)
    _style_header(ws4[1], fill_h)

    if ws2.max_row > 1:
        bar = BarChart()
        bar.title = "Producao equivalente por responsavel"
        bar.y_axis.title = "Equivalente"
        bar.x_axis.title = "Responsavel"
        data = Reference(ws2, min_col=7, min_row=1, max_row=ws2.max_row)
        cats = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        bar.height = 7
        bar.width = 14
        dash.add_chart(bar, "E4")
    if ws4.max_row > 2:
        line = LineChart()
        line.title = "Tendencia diaria"
        line.y_axis.title = "Equivalente"
        data = Reference(ws4, min_col=2, max_col=ws4.max_column, min_row=1, max_row=ws4.max_row)
        cats = Reference(ws4, min_col=1, min_row=2, max_row=ws4.max_row)
        line.add_data(data, titles_from_data=True)
        line.set_categories(cats)
        line.height = 7
        line.width = 14
        dash.add_chart(line, "E19")

    _apply_sheet_style(wb)
    wb.save(OUT_BI_XLSX)
    return OUT_BI_XLSX


def _fmt(value: Any) -> str:
    if isinstance(value, (int, float)):
        s = f"{float(value):,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")
        return s[:-3] if s.endswith(",00") else s
    return str(value or "-")


def _table(data: list[list[Any]], widths: list[float], font_size: int = 8) -> Table:
    t = Table(data, colWidths=widths, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#123a75")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), font_size),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#c8d2e0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f7fb")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
    ]))
    return t


def exportar_bi_pdf(payload: dict[str, Any] | None = None) -> Path:
    payload = payload or montar_bi_payload(False)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="Small", parent=styles["Normal"], fontSize=8, leading=10))
    styles.add(ParagraphStyle(name="Section", parent=styles["Heading2"], textColor=colors.HexColor("#123a75"), spaceAfter=8))
    doc = SimpleDocTemplate(
        str(OUT_BI_PDF),
        pagesize=landscape(A4),
        rightMargin=1.0 * cm,
        leftMargin=1.0 * cm,
        topMargin=0.8 * cm,
        bottomMargin=0.8 * cm,
    )
    story: list[Any] = [
        Paragraph("BI Analytics Operacional - ConstruData", styles["Title"]),
        Paragraph(f"Gerado em {payload['gerado_em']} | Fonte: {payload['fonte']}", styles["Small"]),
        Spacer(1, 0.25 * cm),
        Paragraph("Resumo executivo", styles["Section"]),
    ]
    for nota in payload["resumo_executivo"]:
        story.append(Paragraph(f"- {nota}", styles["Normal"]))
    story.append(Spacer(1, 0.25 * cm))

    cards = payload["cards"]
    story.append(_table([
        ["RDOs", "Producao equiv.", "Total m", "Total un", "Etapas", "Prev. 7 dias"],
        [_fmt(cards["rdos"]), _fmt(cards["total_equiv"]), _fmt(cards["total_m"]), _fmt(cards["total_un"]), _fmt(cards.get("total_etapa", 0)), _fmt(cards["previsao_7_dias"])],
    ], [3.3 * cm, 4.0 * cm, 3.4 * cm, 3.4 * cm, 3.4 * cm, 3.8 * cm], 9))
    story.append(Spacer(1, 0.35 * cm))

    rows: list[list[Any]] = [["Responsavel", "RDOs", "Periodo", "Base", "Prev. dia", "Prev. 7d", "Conf.", "Risco", "Leitura"]]
    for item in payload["comparativo"]:
        rows.append([
            item["responsavel"],
            _fmt(item["rdos"]),
            item["periodo"],
            f"{_fmt(item['total_m'])} m / {_fmt(item['total_un'])} un / {_fmt(item.get('total_etapa', 0))} etapas",
            _fmt(item["previsao_proximo_dia"]),
            _fmt(item["previsao_7_dias"]),
            _fmt(item["confianca"]),
            item["risco"],
            Paragraph(str(item["observacao"] or ""), styles["Small"]),
        ])
    story.append(Paragraph("Comparativo por responsavel", styles["Section"]))
    story.append(_table(rows, [4.5 * cm, 1.3 * cm, 3.0 * cm, 4.0 * cm, 2.0 * cm, 2.0 * cm, 1.5 * cm, 1.7 * cm, 7.2 * cm], 7))
    story.append(Spacer(1, 0.35 * cm))

    story.append(Paragraph("Leitura dos modelos XGBoost", styles["Section"]))
    model_rows = [["Responsavel", "Variaveis que mais pesaram", "Uso pratico"]]
    for item in payload["comparativo"]:
        tops = ", ".join(f"{i['feature']} ({_fmt(i['peso'])})" for i in (item.get("importancias") or [])[:5]) or "Sem peso calculado"
        uso = "Cobrar atualizacao diaria de desvios e comparar previsao x executado no fechamento do dia."
        model_rows.append([item["responsavel"], Paragraph(tops, styles["Small"]), Paragraph(uso, styles["Small"])])
    story.append(_table(model_rows, [4.5 * cm, 10.0 * cm, 11.0 * cm], 8))

    story.append(PageBreak())
    story.append(Paragraph("Serie diaria recente", styles["Section"]))
    serie_rows = [["Responsavel", "Data", "Equiv.", "m", "un", "etapas", "Bloqueios", "Atividade"]]
    flat = []
    for key, serie in payload["series"].items():
        for row in serie:
            flat.append([key, row["data"], row["producao_equiv"], row["producao_m"], row["producao_un"], row.get("producao_etapa", 0), row["bloqueio_score"], row["atividade_score"]])
    for row in sorted(flat, key=lambda x: (x[1], x[0]))[-28:]:
        serie_rows.append([row[0], row[1], _fmt(row[2]), _fmt(row[3]), _fmt(row[4]), _fmt(row[5]), _fmt(row[6]), _fmt(row[7])])
    story.append(_table(serie_rows, [4.0 * cm, 2.6 * cm, 2.3 * cm, 2.3 * cm, 2.3 * cm, 2.3 * cm, 2.4 * cm, 2.4 * cm], 8))
    story.append(Spacer(1, 0.35 * cm))

    story.append(Paragraph("Rotina para alimentar o BI", styles["Section"]))
    for acao in payload["acoes_operacionais"]:
        story.append(Paragraph(f"- {acao}", styles["Normal"]))
    doc.build(story)
    return OUT_BI_PDF
