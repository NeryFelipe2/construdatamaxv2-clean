import json
import math
import re
from pathlib import Path
from datetime import datetime
from collections import defaultdict, Counter

import geopandas as gpd
from shapely.geometry import shape
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_RE_NON_PRINTABLE = re.compile(r'[\x00-\x1f\x7f]')
_RE_SPACES = re.compile(r'\s+')

def limpar(rua: str) -> str:
    if not rua or str(rua).strip() in ("","nan","None","Sem Rua","1"):
        return "Sem Rua"
    t = str(rua).replace("\n"," ").replace("\r"," ")
    t = _RE_NON_PRINTABLE.sub('', t)
    return _RE_SPACES.sub(' ', t).strip() or "Sem Rua"


def e_cadastro(trecho: dict, nucleo: str) -> tuple[bool, str]:
    """Determina se um trecho provavelmente e CADASTRO (rede existente) e nao obra nova."""
    dn = trecho.get("dn_mm") or 200
    nome_nuc = str(nucleo).lower()
    
    # Sao Manoel: PVs com "(1)" no nome = cadastro SABESP
    if "sao manoel" in nome_nuc or "manuel" in nome_nuc:
        pv_ini = str(trecho.get("pv_ini",""))
        pv_fim = str(trecho.get("pv_fim",""))
        if pv_ini.endswith("(1)") or pv_fim.endswith("(1)"):
            return True, "CADASTRO - Rede existente"
            
    # Prolongamentos / Redes Maiores: DN >= 300mm = provavel coletor tronco existente
    if dn >= 300 and "agua" not in nome_nuc:
        if any(k in nome_nuc for k in ["prolong","pantanal","israel","teteu"]):
            return True, f"CADASTRO? - DN{dn}mm"
            
    return False, ""


def carregar_shp(shapefile_path: Path) -> list[dict]:
    import os
    os.environ["SHAPE_ENCODING"] = "latin-1"
    try:
        gdf = gpd.read_file(str(shapefile_path))
    except Exception as e:
        import pyogrio
        gdf = pyogrio.read_dataframe(str(shapefile_path), encoding="latin-1")
        
    gdf['length_m'] = gdf.geometry.length
    linhas = []
    for _, row in gdf.iterrows():
        g = row.geometry
        if g is None: continue
        mid = g.interpolate(0.5, normalized=True)
        dn = 200
        layer = str(row.get('GM_LAYER',''))
        if 'DN 150' in layer: dn = 150
        elif 'DN 200' in layer: dn = 200
        elif 'DN 300' in layer: dn = 300
        elif 'DN 63' in layer: dn = 63
        else:
            d = row.get('Diametro')
            if d:
                d = float(d)
                dn = round(d*1000) if d < 1 else round(d) if d < 1000 else 200
        mat = str(row.get('material','PVC'))
        tipo = 'AGUA' if 'agua' in str(row.get('tipo','')).lower() or 'PEAD' in mat else 'ESGOTO'
        
        # Para Polygon / MultiLineString, tratar coordenadas adequadamente (simplificacao)
        try:
            if g.geom_type == 'LineString':
                c = g.coords
            else:
                continue
            linhas.append({
                'mx': mid.x, 'my': mid.y,
                'sx': c[0][0], 'sy': c[0][1],
                'ex': c[-1][0], 'ey': c[-1][1],
                'length': row['length_m'], 'dn': dn, 'material': mat,
                'tipo': tipo, 'data': str(row.get('Data_Inst',''))
            })
        except:
            pass
    return linhas


def match_shp(trecho: dict, pvs: dict, linhas: list[dict], tol=15.0) -> tuple[bool, dict]:
    p0 = pvs.get(trecho.get("pv_ini"), {})
    p1 = pvs.get(trecho.get("pv_fim"), {})
    x0, y0 = p0.get("x",0), p0.get("y",0)
    x1, y1 = p1.get("x",0), p1.get("y",0)
    if x0==0 or x1==0: return False, None
    mx = (x0+x1)/2
    my = (y0+y1)/2
    
    best_d = float('inf')
    best = None
    for l in linhas:
        d = math.hypot(l['mx']-mx, l['my']-my)
        if d < best_d: 
            best_d = d
            best = l
            
    if best_d <= tol: return True, best
    
    for l in linhas:
        d1 = math.hypot(l['sx']-x0, l['sy']-y0)
        d2 = math.hypot(l['ex']-x1, l['ey']-y1)
        if d1<=tol and d2<=tol: return True, l
        d1b = math.hypot(l['sx']-x1, l['sy']-y1)
        d2b = math.hypot(l['ex']-x0, l['ey']-y0)
        if d1b<=tol and d2b<=tol: return True, l
    return False, None


def processar_lote_auditoria(projetos: list[dict], shapefiles_dir: Path, output_dir: Path) -> dict:
    """
    Recebe um lote de projetos e varre os SHPs de execuçao.
    projetos = [{"nucleo": "...", "tipo": "ESGOTO", "pvs": {...}, "trechos": [...]}]
    """
    # 1. Carregar todos os SHPs (multiplos shapefiles na pasta web recebida, ou um arquivo direto)
    linhas_shp = []
    if shapefiles_dir and shapefiles_dir.exists():
        if shapefiles_dir.is_file() and shapefiles_dir.suffix.lower() == '.shp':
            try:
                linhas_shp.extend(carregar_shp(shapefiles_dir))
            except Exception as e:
                print(f"Erro lendo SHP {shapefiles_dir}: {e}")
        else:
            for shp_file in shapefiles_dir.glob("**/*.shp"):
                try:
                    linhas_shp.extend(carregar_shp(shp_file))
                except Exception as e:
                    print(f"Erro lendo SHP {shp_file}: {e}")
                
    # 2. Processar Redes
    todos_trechos_processados = []
    stats = {"exec":0, "pend":0, "cad":0, "ext_e":0, "ext_p":0, "ext_c":0, "n_nucleos": len(projetos)}
    
    for proj in projetos:
        nucleo = proj.get("nucleo", "Rede Desconhecida")
        tipo = proj.get("tipo", "ESGOTO").upper()
        pvs = proj.get("pvs", {})
        trechos = proj.get("trechos", [])
        
        if not trechos: continue
        
        n_exec = n_pend = n_cad = 0
        ext_e = ext_p = ext_c = 0
        por_rua = defaultdict(list)
        
        for t in trechos:
            is_cad, motiv = e_cadastro(t, nucleo)
            t["_cadastro"] = is_cad
            t["_motivo_cad"] = motiv
            if is_cad:
                t["_exec"] = False
                n_cad += 1
                ext_c += t.get("ext_m", 0)
            else:
                is_exec, match_data = match_shp(t, pvs, linhas_shp)
                t["_exec"] = is_exec
                t["_match_data"] = match_data
                if is_exec:
                    n_exec += 1
                    ext_e += t.get("ext_m", 0)
                else:
                    n_pend += 1
                    ext_p += t.get("ext_m", 0)
                    
            rua = limpar(t.get("rua"))
            por_rua[rua].append(t)
            
        stats["exec"] += n_exec; stats["ext_e"] += ext_e
        stats["pend"] += n_pend; stats["ext_p"] += ext_p
        stats["cad"] += n_cad; stats["ext_c"] += ext_c
        
        for rua, lista in por_rua.items():
            todos_trechos_processados.append({
                "nucleo": nucleo, "tipo": tipo, "rua": rua,
                "trechos": lista, "pvs": pvs
            })

    output_dir.mkdir(parents=True, exist_ok=True)
    
    # GERAR CONSOLIDADO V4
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Consolidado v4"
    hf=Font(name="Calibri",bold=True,color="FFFFFF",size=10)
    hfl=PatternFill(start_color="1b5e20",fill_type="solid")
    brd=Border(left=Side("thin","cccccc"),bottom=Side("thin","cccccc"))
    exec_f=PatternFill(start_color="c8e6c9",fill_type="solid")
    pend_f=PatternFill(start_color="ffcdd2",fill_type="solid")
    cad_f=PatternFill(start_color="e0e0e0",fill_type="solid")
    ef=Font(bold=True,color="1b5e20"); pf=Font(bold=True,color="c62828"); cf=Font(italic=True,color="757575")
    
    heads=["Nucleo","Tipo","Rua","NS","PV Mont","PV Jus","DN(mm)","Ext(m)","Mat",
           "DECL","STATUS","Motivo Cad","Layer"]
    ws.append(heads)
    row_idx = 2
    for res in todos_trechos_processados:
        for i, t in enumerate(res["trechos"]):
            status = "EXEC" if t["_exec"] else "CADASTRO" if t["_cadastro"] else "PENDENTE"
            vals = [
                res["nucleo"], res["tipo"], res["rua"], i+1,
                t.get("pv_ini"), t.get("pv_fim"), t.get("dn_mm",200),
                round(t.get("ext_m",0),2), t.get("material","PVC"),
                t.get("decl_pct"), status, t.get("_motivo_cad",""),
                t.get("layer","")
            ]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=row_idx, column=col, value=val)
                if status=="EXEC": c.fill=exec_f; c.font=ef
                elif status=="PENDENTE": c.fill=pend_f; c.font=pf
                else: c.fill=cad_f; c.font=cf
            row_idx += 1
            
    consolidado_path = output_dir / "CONSOLIDADO_V4_WEB.xlsx"
    wb.save(str(consolidado_path))
    
    # GERAR MATERIAIS V4 (NTS)
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active; ws2.title = "Materiais PENDENTES V4 NTS"
    row2 = 1
    
    for res in todos_trechos_processados:
        tipo = res["tipo"]
        t_pend = [t for t in res["trechos"] if not t["_exec"] and not t["_cadastro"]]
        if not t_pend: continue
        
        ws2[f"A{row2}"]=f"NUCLEO: {res['nucleo']} | RUA: {res['rua']} | {tipo}"
        ws2[f"A{row2}"].font=Font(bold=True, size=12, color="b71c1c")
        row2 += 2
        
        # Agrupar trechos pendentes de acordo as bitolas NTS (PVC: 200/300 | PEAD: 63/110)
        ext_por_dn = defaultdict(float)
        for t in t_pend:
            dn_orig = t.get("dn_mm") or (200 if tipo=="ESGOTO" else 63)
            if tipo == "ESGOTO":
                dn = 200 if dn_orig <= 200 else 300
            else:
                dn = 63 if dn_orig <= 63 else 110
            ext_por_dn[dn] += t.get("ext_m", 0)
            
        n_pvs = len(t_pend) + 1  # 1 PV a mais que trechos (linearizacao)
        
        if tipo == "ESGOTO":
            for dn, ext_dn in sorted(ext_por_dn.items()):
                nb = math.ceil(ext_dn / 6.0)
                ws2[f"A{row2}"]=f"Tubo PVC Corrugado/Macico Esgoto JEI DN {dn}mm - NTS 048"; ws2[f"D{row2}"]=math.ceil(ext_dn)
                ws2[f"A{row2}"].font=Font(bold=True)
                row2+=1
                for desc,qt in [
                    (f"  - Anel de Borracha p/ Tubo PVC DN {dn}mm", nb + 1),
                    (f"  - Pasta Lubrificante p/ Junta Elastica", math.ceil(nb * 0.05)),
                    (f"  - Luva de Correr PVC Esgoto DN {dn}mm", max(math.ceil(nb * 0.05), 1))
                ]:
                    ws2[f"A{row2}"]=desc; ws2[f"D{row2}"]=qt; row2+=1
                    
            ws2[f"A{row2}"]="Poco de Visita (PV) Pre-Moldado Concreto D=1,20m - NTS 015"; ws2[f"D{row2}"]=n_pvs; ws2[f"A{row2}"].font=Font(bold=True)
            row2+=1
            for desc,qt in [
                ("  - Tampao FoFo Articulado", n_pvs),
                ("  - Degrau de FF", n_pvs * 5),
                ("  - Anel de Ajuste Concreto", n_pvs * 2),
                ("  - Argamassa de Assentamento Traço 1:3", math.ceil(n_pvs * 1.5))
            ]:
                ws2[f"A{row2}"]=desc; ws2[f"D{row2}"]=qt; row2+=1
        else: # AGUA
            for dn, ext_dn in sorted(ext_por_dn.items()):
                nb = math.ceil(ext_dn / 6.0)
                ws2[f"A{row2}"]=f"Tubo PEAD PE 100 SDR 17 PN 10 DN {dn}mm - NTS 194"; ws2[f"D{row2}"]=math.ceil(ext_dn)
                ws2[f"A{row2}"].font=Font(bold=True)
                row2+=1
                for desc,qt in [
                    (f"  - Luva Eletrofusao PEAD DN {dn}mm", math.ceil(nb * 1.1)),
                    (f"  - Te Eletrofusao PEAD DN {dn}mm", max(math.ceil(nb * 0.05), 1)),
                    (f"  - Cap Eletrofusao PEAD DN {dn}mm", 2)
                ]:
                    ws2[f"A{row2}"]=desc; ws2[f"D{row2}"]=qt; row2+=1
                    
        # Insumos Civis Regra Unificada NTS
        ext_p = sum(ext_por_dn.values())
        if tipo == "ESGOTO":
            itens_civil = [
                ("Escavacao mecanica / manual em vala", math.ceil(ext_p * 0.9)),
                ("Areia Lavada Media (Envoltoria)", math.ceil(ext_p * 0.35)),
                ("Brita N 1 ou 2 (Base/Reaterro)", math.ceil(ext_p * 0.25)),
                ("Reposicao Pavimento Asfaltico (CBUQ)", math.ceil(ext_p * 0.9))
            ]
        else:
            itens_civil = [
                ("Escavacao mecanica / manual em vala (Agua)", math.ceil(ext_p * 0.6)),
                ("Areia Lavada Media (Envoltoria - Agua)", math.ceil(ext_p * 0.2)),
                ("Reposicao Pavimento Asfaltico (CBUQ)", math.ceil(ext_p * 0.6))
            ]
        for desc, qt in itens_civil:
            ws2[f"A{row2}"]=desc; ws2[f"D{row2}"]=qt; row2+=1
        row2 += 1
        
    materiais_path = output_dir / "MATERIAIS_PENDENTES_V4_WEB.xlsx"
    wb2.save(str(materiais_path))
    
    return {
        "status_geral": stats,
        "consolidado_xlsx": str(consolidado_path.name),
        "materiais_xlsx": str(materiais_path.name)
    }
