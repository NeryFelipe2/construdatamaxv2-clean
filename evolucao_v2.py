#!/usr/bin/env python3
"""
EVOLUÇÃO v2 — PLANILHAS COM PRINTS + ANÁLISE "SEM RUA"
=======================================================
Lê os DADOS.json já gerados, regenera:
  1. Planilha por rua COM prints satélite embutidos
  2. Trechos "Sem Rua" com print da rede e explicação do porquê
  3. MATERIAIS_POR_RUA_VALERIA_v2.xlsx
  4. TRECHOS_CONSOLIDADO_v2.xlsx com prints
"""
import sys, os, json, math, re, warnings, traceback
from pathlib import Path
from datetime import datetime
from collections import defaultdict, Counter
import unicodedata

warnings.filterwarnings("ignore")

MOTOR_DIR = Path(r"C:\Users\felip\Downloads\NOVA NS Versao 5")
sys.path.insert(0, str(MOTOR_DIR))

from gerar_ns import calc_manning, CONTRATO, NS_VERSION

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

try:
    import contextily as cx
    HAS_CTX = True
except ImportError:
    HAS_CTX = False

BASE = Path(r"C:\Users\felip\Desktop\GERAR NS COM ITENS POR RUA DE PV A PV\SAIDA_COMPLETA_POR_RUA")
# Incluir também a pasta com ruas corretas do JC+SM
BASE_NOMES = Path(r"C:\Users\felip\Desktop\GERAR NS COM ITENS POR RUA DE PV A PV\SAIDA_POR_RUA_COM_NOMES")
OUTPUT = BASE / "_PLANILHAS_V2"

def slug(t):
    if not t: return "DESCONHECIDO"
    t = str(t).replace("\n"," ").replace("\r"," ")
    t = re.sub(r'[\x00-\x1f\x7f]','',t)
    n = unicodedata.normalize("NFKD",t).encode("ascii","ignore").decode("ascii")
    s = re.sub(r'[^\w\s-]','_',n).strip()
    return re.sub(r'\s+',' ',s)[:80] or "DESCONHECIDO"

def limpar(rua):
    if not rua or str(rua).strip() in ("","nan","None","Sem Rua","1"):
        return "Sem Rua"
    t = str(rua).replace("\n"," ").replace("\r"," ")
    t = re.sub(r'[\x00-\x1f\x7f]','',t)
    return re.sub(r'\s+',' ',t).strip() or "Sem Rua"

# ──────────────────────────────────────────────────────────────────────
# CARREGAR TODOS OS DADOS DOS PROJETOS JÁ PROCESSADOS
# ──────────────────────────────────────────────────────────────────────
def carregar_dados_projeto(pasta_projeto):
    """Lê todos os NS{xxx}_DADOS.json de um projeto."""
    ns_campo = pasta_projeto / "01_NS_CAMPO"
    if not ns_campo.exists():
        return None, None
    pvs = {}
    trechos = []
    for ns_dir in sorted(ns_campo.iterdir()):
        if not ns_dir.is_dir():
            continue
        for f in ns_dir.glob("*_DADOS.json"):
            try:
                with open(f, encoding="utf-8") as fh:
                    d = json.load(fh)
                t = d.get("trecho", {})
                if t.get("pv_ini") and t.get("pv_fim"):
                    t["rua"] = limpar(t.get("rua"))
                    trechos.append(t)
                    if d.get("pv_montante"):
                        pvs[t["pv_ini"]] = d["pv_montante"]
                    if d.get("pv_jusante"):
                        pvs[t["pv_fim"]] = d["pv_jusante"]
            except Exception:
                pass
    return pvs, trechos

# ──────────────────────────────────────────────────────────────────────
# GERAR IMAGEM SATÉLITE
# ──────────────────────────────────────────────────────────────────────
def gerar_img(pvs, trecho, out_path):
    p0 = pvs.get(trecho.get("pv_ini"), {})
    p1 = pvs.get(trecho.get("pv_fim"), {})
    x0,y0 = p0.get("x",0), p0.get("y",0)
    x1,y1 = p1.get("x",0), p1.get("y",0)
    if x0==0 or x1==0 or y0==0 or y1==0:
        return False
    fig,ax = plt.subplots(figsize=(6,4),dpi=100)
    mg = max(abs(x1-x0),abs(y1-y0),30)*0.6
    ax.set_xlim(min(x0,x1)-mg,max(x0,x1)+mg)
    ax.set_ylim(min(y0,y1)-mg,max(y0,y1)+mg)
    sat=False
    if HAS_CTX:
        for z in [18,17,16]:
            try:
                cx.add_basemap(ax,crs="EPSG:31983",source=cx.providers.Esri.WorldImagery,zoom=z)
                sat=True; break
            except: pass
    if not sat:
        ax.set_facecolor("#e8f0e8"); ax.grid(True,color="#ccc",alpha=0.5)
    ax.plot([x0,x1],[y0,y1],'r-',lw=3,zorder=5)
    for nm,pv,c in [(trecho.get("pv_ini"),p0,"red"),(trecho.get("pv_fim"),p1,"blue")]:
        ax.plot(pv["x"],pv["y"],'o',color=c,ms=8,zorder=6)
        ct=pv.get("ct","")
        ctt=f"\nCT={ct:.2f}" if isinstance(ct,(int,float)) and ct>0 else ""
        ax.annotate(f"{nm}{ctt}",(pv["x"],pv["y"]),fontsize=6,color="white" if sat else "black",
                    fontweight="bold",ha="center",va="bottom",xytext=(0,8),textcoords="offset points",
                    bbox=dict(boxstyle="round,pad=0.2",fc="black" if sat else "white",alpha=0.7),zorder=7)
    dn=trecho.get("dn_mm","?"); ext=trecho.get("ext_m",0); rua=trecho.get("rua","")
    ax.set_title(f"DN{dn}mm | {ext:.1f}m | {rua}",fontsize=8,fontweight="bold")
    ax.tick_params(labelsize=5)
    fig.tight_layout()
    fig.savefig(str(out_path),dpi=100,bbox_inches="tight",pad_inches=0.1)
    plt.close(fig)
    return True

def gerar_img_rede(pvs, trechos_lista, titulo, out_path):
    """Gera imagem panorâmica de TODOS os trechos (visão da rede)."""
    xs, ys = [], []
    for t in trechos_lista:
        p0 = pvs.get(t.get("pv_ini"),{})
        p1 = pvs.get(t.get("pv_fim"),{})
        if p0.get("x",0)>0 and p1.get("x",0)>0:
            xs.extend([p0["x"],p1["x"]]); ys.extend([p0["y"],p1["y"]])
    if len(xs)<2:
        return False
    fig,ax = plt.subplots(figsize=(8,6),dpi=100)
    mg = max(max(xs)-min(xs),max(ys)-min(ys))*0.15
    ax.set_xlim(min(xs)-mg,max(xs)+mg); ax.set_ylim(min(ys)-mg,max(ys)+mg)
    sat=False
    if HAS_CTX:
        for z in [17,16,15]:
            try:
                cx.add_basemap(ax,crs="EPSG:31983",source=cx.providers.Esri.WorldImagery,zoom=z)
                sat=True; break
            except: pass
    if not sat:
        ax.set_facecolor("#e8f0e8"); ax.grid(True,color="#ccc",alpha=0.5)
    for t in trechos_lista:
        p0=pvs.get(t.get("pv_ini"),{}); p1=pvs.get(t.get("pv_fim"),{})
        if p0.get("x",0)>0 and p1.get("x",0)>0:
            ax.plot([p0["x"],p1["x"]],[p0["y"],p1["y"]],'r-',lw=2,zorder=5,alpha=0.8)
            ax.plot(p0["x"],p0["y"],'o',color="yellow",ms=4,zorder=6)
            ax.plot(p1["x"],p1["y"],'o',color="yellow",ms=4,zorder=6)
    ax.set_title(titulo,fontsize=9,fontweight="bold")
    ax.tick_params(labelsize=5)
    fig.tight_layout()
    fig.savefig(str(out_path),dpi=100,bbox_inches="tight",pad_inches=0.1)
    plt.close(fig)
    return True

# ──────────────────────────────────────────────────────────────────────
# ANÁLISE "SEM RUA"
# ──────────────────────────────────────────────────────────────────────
def analisar_sem_rua(trecho, pvs):
    """Analisa POR QUE um trecho ficou sem rua e dá explicação."""
    pvi = pvs.get(trecho.get("pv_ini"),{})
    pvf = pvs.get(trecho.get("pv_fim"),{})
    x0,y0 = pvi.get("x",0), pvi.get("y",0)
    x1,y1 = pvf.get("x",0), pvf.get("y",0)
    ext = trecho.get("ext_m",0)
    dn = trecho.get("dn_mm") or 200
    nome_ini = str(trecho.get("pv_ini",""))
    nome_fim = str(trecho.get("pv_fim",""))

    razoes = []

    # 1. Sem coordenadas
    if x0==0 or x1==0:
        razoes.append("PV sem coordenadas UTM (Null/Dummy no XML)")
        return "SEM COORDENADA — Ponto fictício no projeto Civil 3D (StartNull/EndNull). Provavelmente é INTERLIGAÇÃO com rede existente ou PONTO DE LANÇAMENTO.", razoes

    # 2. NullStruct
    if "Null" in nome_ini or "Null" in nome_fim:
        razoes.append("PV é NullStruct (ponto de interligação)")
        return "INTERLIGAÇÃO — Ponto onde a rede nova conecta com rede existente SABESP. Sem logradouro associado no mapa.", razoes

    # 3. Trecho muito curto (< 3m) = pode ser ligação de caixa
    if ext < 3:
        razoes.append(f"Trecho muito curto ({ext:.1f}m)")
        return f"TRECHO CURTO ({ext:.1f}m) — Provavelmente ligação entre PVs adjacentes ou conexão de caixa de inspeção.", razoes

    # 4. DN grande (>= 300) em áreas sem mapa = pode ser coletor tronco/beira córrego
    if dn >= 300:
        razoes.append(f"DN grande ({dn}mm) = possível coletor tronco")
        return f"COLETOR TRONCO DN{dn}mm — Rede de grande diâmetro, geralmente segue curso d'água ou fundo de vale. Sem logradouro no mapa cartográfico.", razoes

    # 5. Genérico - sem mapa na região
    razoes.append("Sem correspondência no mapa cartográfico QGIS")
    return "SEM MAPA — Trecho em área não coberta pelo mapa cartográfico QGIS. Pode ser: extensão da rede, beco sem nome, ou área de difícil acesso.", razoes

# ──────────────────────────────────────────────────────────────────────
# GERAR PLANILHA V2 POR RUA COM PRINTS + ANÁLISE SEM RUA
# ──────────────────────────────────────────────────────────────────────
def gerar_planilha_rua_v2(trechos, pvs, nucleo, tipo, rua, out_path, img_dir):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = rua[:31]  # max 31 chars para nome de aba

    # Estilos premium
    hf = Font(name="Calibri",bold=True,color="FFFFFF",size=10)
    hfill = PatternFill(start_color="0d47a1",end_color="0d47a1",fill_type="solid")
    df = Font(name="Calibri",size=9)
    da = Alignment(horizontal="center",vertical="center",wrap_text=True)
    brd = Border(left=Side("thin","999999"),right=Side("thin","999999"),
                 top=Side("thin","999999"),bottom=Side("thin","999999"))
    alt = PatternFill(start_color="e3f2fd",end_color="e3f2fd",fill_type="solid")
    sem_rua_fill = PatternFill(start_color="fff3e0",end_color="fff3e0",fill_type="solid")
    warn_font = Font(name="Calibri",size=8,color="bf360c",italic=True)

    is_sem_rua = (rua == "Sem Rua")

    # Título
    ws.merge_cells("A1:P1")
    ws["A1"] = f"TRECHOS — {rua} — {nucleo} — {tipo}"
    ws["A1"].font = Font(name="Calibri",bold=True,size=14,color="0d47a1")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:P2")
    ws["A2"] = f"CT {CONTRATO} | SE LIGA NA REDE | {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(name="Calibri",size=9,color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = ["NS","PV Mont","PV Jus","DN(mm)","Ext(m)","Mat",
               "CT Mont","CF Mont","Prof M","CT Jus","CF Jus","Prof J",
               "Decl‰","V(m/s)","Print Satélite"]
    if is_sem_rua:
        headers.append("Análise/Motivo")
    for c,h in enumerate(headers,1):
        cell = ws.cell(row=3,column=c,value=h)
        cell.font=hf; cell.fill=hfill; cell.alignment=da; cell.border=brd

    for i,t in enumerate(trechos):
        r = i+4
        pvi = pvs.get(t.get("pv_ini"),{})
        pvf = pvs.get(t.get("pv_fim"),{})
        hidr = calc_manning(t.get("dn_mm"), t.get("decl_mm"))
        vals = [
            f"NS-{i+1:04d}", t.get("pv_ini",""), t.get("pv_fim",""),
            t.get("dn_mm"), round(t.get("ext_m",0),2), t.get("material","PVC"),
            pvi.get("ct"), pvi.get("cf"), pvi.get("prof"),
            pvf.get("ct"), pvf.get("cf"), pvf.get("prof"),
            round((t.get("decl_mm") or 0)*1000,2) if t.get("decl_mm") else None,
            hidr.get("v_ms"), ""
        ]
        if is_sem_rua:
            motivo, _ = analisar_sem_rua(t, pvs)
            vals.append(motivo)
        for c,v in enumerate(vals,1):
            cell = ws.cell(row=r,column=c,value=v)
            cell.font = df if c < len(vals) else warn_font
            cell.alignment = da
            cell.border = brd
            if is_sem_rua:
                cell.fill = sem_rua_fill
            elif i%2==1:
                cell.fill = alt

        # Imagem satélite do trecho
        img_path = Path(img_dir) / f"trecho_{i+1:03d}.png"
        try:
            if not img_path.exists():
                gerar_img(pvs, t, img_path)
            if img_path.exists():
                img = XlImage(str(img_path))
                img.width = 300; img.height = 180
                ws.add_image(img, f"O{r}")
                ws.row_dimensions[r].height = 140
        except Exception:
            pass

    # Se é "Sem Rua", adicionar imagem panorâmica da rede no final
    if is_sem_rua and trechos:
        r_net = len(trechos) + 6
        net_img = Path(img_dir) / "REDE_PANORAMICA.png"
        try:
            gerar_img_rede(pvs, trechos, f"VISÃO GERAL — Trechos Sem Rua — {nucleo}", net_img)
            if net_img.exists():
                ws.merge_cells(f"A{r_net}:P{r_net}")
                ws[f"A{r_net}"] = "▼ VISÃO PANORÂMICA DA REDE (trechos sem rua)"
                ws[f"A{r_net}"].font = Font(name="Calibri",bold=True,size=12,color="e65100")
                img = XlImage(str(net_img))
                img.width = 800; img.height = 500
                ws.add_image(img, f"A{r_net+1}")
                ws.row_dimensions[r_net+1].height = 400
        except Exception:
            pass

    # Ajustar larguras
    widths = [8,14,14,7,7,5,8,8,6,8,8,6,7,7,45]
    if is_sem_rua:
        widths.append(50)
    for c,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # Rodapé
    r_fin = len(trechos) + 5
    ws.merge_cells(f"A{r_fin}:O{r_fin}")
    ext = sum(t.get("ext_m",0) for t in trechos)
    ws[f"A{r_fin}"] = f"TOTAL: {len(trechos)} trechos | {ext:.1f}m | {tipo}"
    ws[f"A{r_fin}"].font = Font(name="Calibri",bold=True,size=11,color="0d47a1")
    wb.save(str(out_path))
    return len(trechos)

# ──────────────────────────────────────────────────────────────────────
# VALÉRIA V2
# ──────────────────────────────────────────────────────────────────────
def gerar_valeria_v2(todos, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Materiais por Rua v2"
    hf=Font(name="Calibri",bold=True,color="FFFFFF",size=10)
    hfl=PatternFill(start_color="0d47a1",end_color="0d47a1",fill_type="solid")
    sf=Font(name="Calibri",size=9,color="444444")
    itf=Font(name="Calibri",bold=True,size=9)
    nf=Font(name="Calibri",bold=True,size=12,color="1a237e")
    rf=Font(name="Calibri",bold=True,size=10,color="2e7d32")
    brd=Border(left=Side("thin","cccccc"),right=Side("thin","cccccc"),top=Side("thin","cccccc"),bottom=Side("thin","cccccc"))
    ws.merge_cells("A1:I1")
    ws["A1"]="MATERIAIS POR RUA v2 — PROJETOS COMPLETOS"
    ws["A1"].font=Font(name="Calibri",bold=True,size=14,color="0d47a1")
    ws.merge_cells("A2:I2")
    ws["A2"]=f"CT {CONTRATO} | SE LIGA NA REDE | {datetime.now().strftime('%d/%m/%Y')}"
    ws["A2"].font=Font(name="Calibri",size=9,color="666666")
    heads=["MATERIAL","UN","REDE","QTD","METRAGEM","NÚCLEO","RUA","TIPO","OBS"]
    for c,h in enumerate(heads,1):
        cell=ws.cell(row=3,column=c,value=h)
        cell.font=hf;cell.fill=hfl;cell.border=brd
    row=4
    por_nucleo=defaultdict(list)
    for d in todos: por_nucleo[d["nucleo"]].append(d)
    for nucleo,dados_nucleo in sorted(por_nucleo.items()):
        ws.merge_cells(f"A{row}:I{row}")
        ws[f"A{row}"]=f"═══ {nucleo.upper()} ═══"
        ws[f"A{row}"].font=nf; row+=1
        for dados in sorted(dados_nucleo,key=lambda d:d.get("rua","")):
            rua=dados.get("rua","Sem Rua"); tipo=dados.get("tipo","ESGOTO")
            trechos=dados.get("trechos",[]); pvs_d=dados.get("pvs",{})
            if not trechos: continue
            ws[f"A{row}"]=f"  ▸ {rua}"; ws[f"A{row}"].font=rf
            ws[f"C{row}"]=tipo; ws[f"F{row}"]=nucleo; ws[f"G{row}"]=rua
            ws[f"H{row}"]=tipo; row+=1
            ext_total=sum(t.get("ext_m",0) for t in trechos)
            n_pvs_t=len(trechos)+1
            dn_def=200 if tipo=="ESGOTO" else 63
            dns=Counter((t.get("dn_mm") or dn_def) for t in trechos)
            if tipo=="ESGOTO":
                for dn,_ in sorted(dns.items(),key=lambda x:x[0] or 0):
                    ext_dn=sum(t.get("ext_m",0) for t in trechos if (t.get("dn_mm") or dn_def)==dn)
                    nb=math.ceil(ext_dn/6)
                    ws[f"A{row}"]=f"Tubo PVC DN{dn}mm"; ws[f"B{row}"]="barra"
                    ws[f"C{row}"]="ESG"; ws[f"D{row}"]=nb; ws[f"E{row}"]=f"{ext_dn:.1f}m"
                    ws[f"A{row}"].font=itf
                    for c in range(1,10): ws.cell(row=row,column=c).border=brd
                    row+=1
                    for desc,un,qt in [(f"   • Anel Borracha DN{dn}","pc",nb+1),
                                       ("   • Pasta Lubrificante","kg",round(nb*0.04,2)),
                                       (f"   • Luva correr PVC DN{dn}","pc",max(nb-1,0))]:
                        ws[f"A{row}"]=desc;ws[f"B{row}"]=un;ws[f"D{row}"]=qt
                        ws[f"A{row}"].font=sf
                        for c in range(1,10): ws.cell(row=row,column=c).border=brd
                        row+=1
                ws[f"A{row}"]="PV concreto"; ws[f"B{row}"]="un"; ws[f"C{row}"]="ESG"
                ws[f"D{row}"]=n_pvs_t; ws[f"A{row}"].font=itf
                for c in range(1,10): ws.cell(row=row,column=c).border=brd
                row+=1
                for desc,un,qt in [("   • Tampão FF DN600/900","pc",n_pvs_t),
                                   ("   • Laje Cobertura/Cone","pc",n_pvs_t),
                                   ("   • Anel Borracha Junta","pc",n_pvs_t*3),
                                   ("   • Degrau Polipropileno","pc",n_pvs_t*6)]:
                    ws[f"A{row}"]=desc;ws[f"B{row}"]=un;ws[f"D{row}"]=qt
                    ws[f"A{row}"].font=sf
                    for c in range(1,10): ws.cell(row=row,column=c).border=brd
                    row+=1
                for desc,un,qt in [("Escavação mecanizada","m³",round(ext_total*0.6*1.5,2)),
                                   ("Areia (lastro+envolt.)","m³",round(ext_total*0.32,2)),
                                   ("Brita (dreno)","m³",round(ext_total*0.22,2)),
                                   ("Pavimentação CBUQ","m²",round(ext_total*0.88,2))]:
                    ws[f"A{row}"]=desc;ws[f"B{row}"]=un;ws[f"C{row}"]="ESG"
                    ws[f"D{row}"]=qt;ws[f"A{row}"].font=itf
                    for c in range(1,10): ws.cell(row=row,column=c).border=brd
                    row+=1
            else:  # AGUA
                for dn,_ in sorted(dns.items(),key=lambda x:x[0] or 0):
                    ext_dn=sum(t.get("ext_m",0) for t in trechos if (t.get("dn_mm") or dn_def)==dn)
                    nb=math.ceil(ext_dn/6)
                    ws[f"A{row}"]=f"Tubo PEAD DN{dn}mm"; ws[f"B{row}"]="barra"
                    ws[f"C{row}"]="AG"; ws[f"D{row}"]=nb; ws[f"E{row}"]=f"{ext_dn:.1f}m"
                    ws[f"A{row}"].font=itf
                    for c in range(1,10): ws.cell(row=row,column=c).border=brd
                    row+=1
                    for desc,un,qt in [(f"   • Sela Eletrofusão DN{dn}","pc",max(round(nb*0.1),1)),
                                       (f"   • Luva Eletrofusão DN{dn}","pc",max(round(nb*0.05),1))]:
                        ws[f"A{row}"]=desc;ws[f"B{row}"]=un;ws[f"D{row}"]=qt
                        ws[f"A{row}"].font=sf
                        for c in range(1,10): ws.cell(row=row,column=c).border=brd
                        row+=1
            row+=1
    for col,w in [("A",45),("B",8),("C",6),("D",10),("E",12),("F",20),("G",25),("H",8),("I",15)]:
        ws.column_dimensions[col].width=w
    wb.save(str(out_path))
    print(f"  Valéria v2: {row} linhas → {out_path.name}")

# ──────────────────────────────────────────────────────────────────────
# CONSOLIDADO V2
# ──────────────────────────────────────────────────────────────────────
def gerar_consolidado_v2(todos, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Consolidado v2"
    hf=Font(name="Calibri",bold=True,color="FFFFFF",size=10)
    hfl=PatternFill(start_color="1b5e20",end_color="1b5e20",fill_type="solid")
    df=Font(name="Calibri",size=9)
    brd=Border(left=Side("thin","cccccc"),right=Side("thin","cccccc"),top=Side("thin","cccccc"),bottom=Side("thin","cccccc"))
    alt=PatternFill(start_color="e8f5e9",end_color="e8f5e9",fill_type="solid")
    srfill=PatternFill(start_color="fff3e0",end_color="fff3e0",fill_type="solid")
    wf=Font(name="Calibri",size=8,color="bf360c",italic=True)
    ws.merge_cells("A1:R1")
    ws["A1"]="CONSOLIDADO v2 — TODOS OS NÚCLEOS POR RUA — COM ANÁLISE"
    ws["A1"].font=Font(name="Calibri",bold=True,size=14,color="1b5e20")
    heads=["Núcleo","Tipo","Rua","NS","PV Mont","PV Jus","DN(mm)","Ext(m)","Mat",
           "CT Mont","CF Mont","CT Jus","CF Jus","Decl‰","V(m/s)","Q(l/s)","Análise Sem Rua","Print"]
    for c,h in enumerate(heads,1):
        cell=ws.cell(row=3,column=c,value=h)
        cell.font=hf;cell.fill=hfl;cell.border=brd
    row=4
    for res in todos:
        nucleo=res.get("nucleo",""); tipo=res.get("tipo",""); rua=res.get("rua","Sem Rua")
        pvs=res.get("pvs",{})
        for i,t in enumerate(res.get("trechos",[])):
            pvi=pvs.get(t.get("pv_ini"),{}); pvf=pvs.get(t.get("pv_fim"),{})
            hidr=calc_manning(t.get("dn_mm"),t.get("decl_mm"))
            motivo=""
            if rua=="Sem Rua":
                motivo,_=analisar_sem_rua(t,pvs)
            vals=[nucleo,tipo,rua,f"NS-{i+1:04d}",t.get("pv_ini",""),t.get("pv_fim",""),
                  t.get("dn_mm"),round(t.get("ext_m",0),2),t.get("material","PVC"),
                  pvi.get("ct"),pvi.get("cf"),pvf.get("ct"),pvf.get("cf"),
                  round((t.get("decl_mm") or 0)*1000,2) if t.get("decl_mm") else None,
                  hidr.get("v_ms"),hidr.get("q_ls"),motivo,""]
            for c,v in enumerate(vals,1):
                cell=ws.cell(row=row,column=c,value=v)
                cell.font=wf if c==17 and motivo else df
                cell.border=brd
                if rua=="Sem Rua": cell.fill=srfill
                elif row%2==0: cell.fill=alt
            row+=1
    row+=1
    tot_ext=sum(sum(t.get("ext_m",0) for t in r.get("trechos",[])) for r in todos)
    ws[f"A{row}"]="TOTAIS"; ws[f"A{row}"].font=Font(bold=True,size=12,color="1b5e20")
    ws[f"A{row+1}"]=f"Total trechos: {row-5}"
    ws[f"A{row+2}"]=f"Extensão total: {tot_ext:.1f}m"
    for c in range(1,19): ws.column_dimensions[get_column_letter(c)].width=12
    ws.column_dimensions["A"].width=22; ws.column_dimensions["C"].width=30
    ws.column_dimensions["Q"].width=55
    wb.save(str(out_path))
    print(f"  Consolidado v2: {row} linhas → {out_path.name}")

# ──────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────
def main():
    try: sys.stdout.reconfigure(encoding="utf-8",errors="replace")
    except: pass
    print("="*72)
    print("  EVOLUÇÃO v2 — PLANILHAS COM PRINTS + ANÁLISE SEM RUA")
    print("="*72)
    OUTPUT.mkdir(parents=True,exist_ok=True)

    # Projetos a processar (preferir versões com rua correta)
    projetos = []

    # JC e SM com ruas corretas
    if (BASE_NOMES / "Joao Carlos").exists():
        projetos.append(("Joao Carlos","ESGOTO",BASE_NOMES / "Joao Carlos"))
    if (BASE_NOMES / "Sao Manoel").exists():
        projetos.append(("Sao Manoel","ESGOTO",BASE_NOMES / "Sao Manoel"))

    # Todos os outros de SAIDA_COMPLETA_POR_RUA
    for p in sorted(BASE.iterdir()):
        if not p.is_dir() or p.name.startswith("_"):
            continue
        # Pular JC e SM originais (sem rua) já que temos versão corrigida
        if p.name in ("Joao Carlos_ESGOTO","Sao Manoel_ESGOTO"):
            continue
        parts = p.name.rsplit("_",1)
        nucleo = parts[0] if len(parts)>1 else p.name
        tipo = parts[1] if len(parts)>1 else "ESGOTO"
        projetos.append((nucleo,tipo,p))

    print(f"  {len(projetos)} projetos a processar\n")

    todos_valeria = []
    todos_consolidado = []

    for nucleo,tipo,pasta in projetos:
        print(f"\n{'─'*60}")
        print(f"  {nucleo} ({tipo})")
        print(f"{'─'*60}")

        pvs, trechos = carregar_dados_projeto(pasta)
        if not trechos:
            print(f"  SKIP — sem trechos")
            continue
        print(f"  {len(pvs)} PVs, {len(trechos)} trechos")

        # Separar por rua
        por_rua = defaultdict(list)
        for t in trechos:
            por_rua[limpar(t.get("rua"))].append(t)

        nucleo_slug = slug(nucleo)
        pasta_out = OUTPUT / f"{nucleo_slug}_{slug(tipo)}"
        pasta_out.mkdir(parents=True,exist_ok=True)

        for rua,lista in sorted(por_rua.items()):
            rua_slug = slug(rua)
            pasta_rua = pasta_out / rua_slug
            pasta_rua.mkdir(parents=True,exist_ok=True)
            img_dir = pasta_rua / "imgs"
            img_dir.mkdir(exist_ok=True)

            print(f"    {rua}: {len(lista)} trechos", end="")
            try:
                gerar_planilha_rua_v2(lista, pvs, nucleo, tipo, rua,
                                      pasta_rua / f"Trechos_{rua_slug}.xlsx", str(img_dir))
                print(" ✓")
            except Exception as e:
                print(f" ERR: {e}")

            todos_valeria.append({"nucleo":nucleo,"tipo":tipo,"rua":rua,"trechos":lista,"pvs":pvs})
            todos_consolidado.append({"nucleo":nucleo,"tipo":tipo,"rua":rua,"trechos":lista,"pvs":pvs})

    # Valéria v2
    print(f"\n{'='*60}")
    print("  Gerando VALÉRIA v2...")
    gerar_valeria_v2(todos_valeria, OUTPUT / "MATERIAIS_POR_RUA_VALERIA_v2.xlsx")

    # Consolidado v2
    print("  Gerando CONSOLIDADO v2...")
    gerar_consolidado_v2(todos_consolidado, OUTPUT / "TRECHOS_CONSOLIDADO_v2.xlsx")

    print(f"\n{'='*72}")
    print(f"  CONCLUÍDO! Saída: {OUTPUT}")
    print(f"{'='*72}")

if __name__=="__main__":
    main()
