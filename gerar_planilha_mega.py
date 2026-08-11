#!/usr/bin/env python3
"""
GERAR_PLANILHA_MEGA.PY — Planilha MEGA Integrada de todos os trechos PV a PV
com quantitativos completos e nomes das ruas.

Uso:
    python gerar_planilha_mega.py <DXF_OU_DWG> [--saida PASTA] [--nucleo NOME]

Gera na pasta PLANEJAMENTO/:
    MEGA_INTEGRADA_{NUCLEO}.xlsx
"""
import sys, os, math, json
from pathlib import Path
from datetime import datetime

# ── Parâmetros de cálculo ────────────────────────────────────────────────────
LARGURA_VALA = 0.60   # m
BDI = 1.25
PROF_PADRAO = 1.50    # m (quando não tem CT/CF)

SINAPI = {
    "tubo_pvc_100":  28.50, "tubo_pvc_150":  35.80, "tubo_pvc_200":  47.12,
    "tubo_pvc_250":  62.30, "tubo_pvc_300":  89.45, "tubo_pvc_400": 142.60,
    "escavacao":     30.77, "reaterro":   19.97, "lastro":   85.50,
    "envoltorio":    75.30, "brita":      95.20, "pavimentacao": 97.80,
    "pv_concreto_1200": 3078.00,
    "escoramento_m2": 28.50,
}


def calc_quant(ext_m, prof_m, dn_mm, lv=LARGURA_VALA):
    """Quantitativos de 1 trecho."""
    if not ext_m or ext_m <= 0:
        return {}
    H = prof_m or PROF_PADRAO
    d = (dn_mm or 200) / 1000
    esc = lv * H * ext_m
    lastro = lv * 0.10 * ext_m
    h_env = min(d + 0.30, H - 0.10)
    envolt = max(0, lv * h_env * ext_m - math.pi * (d / 2) ** 2 * ext_m)
    brita = lv * 0.15 * ext_m
    tubo_v = math.pi * (d / 2) ** 2 * ext_m
    reat = max(0, esc - lastro - envolt - brita - tubo_v)
    bota = max(0, esc - reat)
    pav = lv * ext_m * 1.20
    escor = 2 * H * ext_m  # 2 faces da vala
    barras = max(1, math.ceil(ext_m / 6))
    return {
        "esc_m3": round(esc, 2),
        "lastro_m3": round(lastro, 2),
        "envolt_m3": round(envolt, 2),
        "brita_m3": round(brita, 2),
        "reat_m3": round(reat, 2),
        "bota_m3": round(bota, 2),
        "pav_m2": round(pav, 2),
        "escor_m2": round(escor, 2),
        "tubo_barras": barras,
    }


def calc_custos(ext_m, dn_mm, q):
    """Custos de 1 trecho."""
    chave = f"tubo_pvc_{dn_mm or 200}"
    pt = SINAPI.get(chave, SINAPI["tubo_pvc_200"])
    ct = pt * (ext_m or 0) * BDI
    ce = SINAPI["escavacao"] * (q.get("esc_m3") or 0) * BDI
    cr = SINAPI["reaterro"] * (q.get("reat_m3") or 0) * BDI
    cl = SINAPI["lastro"] * (q.get("lastro_m3") or 0) * BDI
    cb = SINAPI["brita"] * (q.get("brita_m3") or 0) * BDI
    cp = SINAPI["pavimentacao"] * (q.get("pav_m2") or 0) * BDI
    ces = SINAPI["escoramento_m2"] * (q.get("escor_m2") or 0) * BDI
    cpv = SINAPI["pv_concreto_1200"] * BDI
    total = ct + ce + cr + cl + cb + cp + ces + cpv
    return {
        "custo_tubo": round(ct, 2),
        "custo_escav": round(ce, 2),
        "custo_reat": round(cr, 2),
        "custo_lastro": round(cl, 2),
        "custo_brita": round(cb, 2),
        "custo_pav": round(cp, 2),
        "custo_escor": round(ces, 2),
        "custo_pv": round(cpv, 2),
        "custo_total": round(total, 2),
    }


def prof_media(t, pvs):
    """Profundidade média do trecho."""
    p0 = pvs.get(t.get("pv_ini"), {})
    p1 = pvs.get(t.get("pv_fim"), {})
    pi = t.get("prof_ini") or p0.get("prof")
    pf = t.get("prof_fim") or p1.get("prof")
    if pi and pf:
        return round((pi + pf) / 2, 2)
    if pi:
        return pi
    if pf:
        return pf
    # Tentar calcular de CT - CF
    ct0 = t.get("ct_ini") or p0.get("ct")
    cf0 = t.get("cf_ini") or p0.get("cf")
    ct1 = t.get("ct_fim") or p1.get("ct")
    cf1 = t.get("cf_fim") or p1.get("cf")
    profs = []
    if ct0 and cf0:
        profs.append(ct0 - cf0)
    if ct1 and cf1:
        profs.append(ct1 - cf1)
    if profs:
        return round(sum(profs) / len(profs), 2)
    return PROF_PADRAO


def associar_ruas_gpkg(trechos, pvs, gpkg_path, tol=50.0):
    """Associa ruas aos trechos a partir de textos do GPKG cartográfico."""
    import numpy as np
    try:
        import geopandas as gpd
    except ImportError:
        print("WARN: geopandas não disponível para ler GPKG")
        return
    df = gpd.read_file(gpkg_path, layer='texts')
    mask = df['text'].notna()
    PREF = ("RUA ", "BECO ", "TRAV", "AV ", "ESTRADA", "VIELA", "ALAMEDA", "ACESSO", "R. ", "R ")
    ruas_df = df[mask][df[mask]['text'].str.upper().str.strip().apply(
        lambda x: any(x.startswith(p) for p in PREF))].copy()
    # Corrigir triple-encoding (UTF-8 -> cp1252 -> UTF-8 -> cp1252 -> UTF-8)
    _ENC_MAP = {
        'Ã\x83\xe2\x80\x9c': 'Ó', 'Ã\x83\xc6\x92': 'Ã', 'Ã\x83\xe2\x80\xb0': 'É',
        'Ã\x83\xc2\x81': 'Á', 'Ã\x83\xc2\x8d': 'Í', 'Ã\x83\xc2\x89': 'É',
        'Ã\x83\xc2\x93': 'Ó', 'Ã\x83\xc2\x9a': 'Ú', 'Ã\x83\xc2\xa3': 'ã',
        'Ã\x83\xc2\xa9': 'é', 'Ã\x83\xc2\xad': 'í', 'Ã\x83\xc2\xb3': 'ó',
        'Ã\x83\xc2\xba': 'ú', 'Ã\x83\xc2\xa1': 'á', 'Ã\x83\xc2\xa7': 'ç',
        'Ã\x83\xc2\xaa': 'ê', 'Ã\x83\xc2\xb4': 'ô', 'Ã\x83\xc2\xa2': 'â',
        'Ã\u201c': 'Ó', 'Ã\u201d': 'Ó',
        'Ã‰': 'É', 'Ã\u0093': 'Ó', 'Ãƒ': 'Ã', 'Ã\x81': 'Á', 'Ã\x8d': 'Í',
        'Ã\x9a': 'Ú', 'Ã£': 'ã', 'Ã©': 'é', 'Ã­': 'í', 'Ã³': 'ó',
        'Ãº': 'ú', 'Ã¡': 'á', 'Ã§': 'ç', 'Ãª': 'ê', 'Ã´': 'ô', 'Ã¢': 'â',
    }
    def _fix_enc(t):
        for bad, good in sorted(_ENC_MAP.items(), key=lambda x: -len(x[0])):
            t = t.replace(bad, good)
        t = t.replace('\\P', ' ').replace('\\p', ' ')  # AutoCAD line break
        return t.strip()
    ruas_df['text'] = ruas_df['text'].apply(_fix_enc)
    # Filtrar por coordenadas reais (ignorar carimbo/legenda)
    ruas_df = ruas_df[ruas_df.geometry.x > 10000]
    if len(ruas_df) == 0:
        print("WARN: nenhum texto de rua no GPKG")
        return
    rua_xy = np.array([[g.x, g.y] for g in ruas_df.geometry])
    rua_txts = ruas_df['text'].values
    n = 0
    for t in trechos:
        if t.get("rua") and t["rua"] != "Sem Rua":
            continue
        p0 = pvs.get(t.get("pv_ini"), {})
        p1 = pvs.get(t.get("pv_fim"), {})
        if not p0.get("x") or not p1.get("x"):
            continue
        mx = (p0["x"] + p1["x"]) / 2
        my = (p0["y"] + p1["y"]) / 2
        dists = np.hypot(rua_xy[:, 0] - mx, rua_xy[:, 1] - my)
        idx = int(np.argmin(dists))
        if dists[idx] <= tol:
            t["rua"] = rua_txts[idx].strip()
            n += 1
    print(f"GPKG: {n} ruas associadas de {len(rua_txts)} textos")


def gerar_mega(pvs, trechos, nucleo, saida_dir):
    """Gera planilha MEGA INTEGRADA."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("ERRO: pip install openpyxl")
        return None

    wb = Workbook()

    # ── Estilos ──────────────────────────────────────────────────────────────
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    num_fmt_2 = '0.00'
    num_fmt_1 = '0.0'
    money_fmt = '#,##0.00'
    thin = Side(style="thin")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    sub_fill = PatternFill("solid", fgColor="D6E4F0")
    tot_fill = PatternFill("solid", fgColor="FFF2CC")
    tot_font = Font(bold=True, size=10)

    # ── Calcular dados ───────────────────────────────────────────────────────
    rows = []
    for i, t in enumerate(trechos):
        ext = t.get("ext_m", 0)
        dn = t.get("dn_mm") or 200
        pm = prof_media(t, pvs)
        q = calc_quant(ext, pm, dn)
        c = calc_custos(ext, dn, q)
        rua = t.get("rua", "Sem Rua")
        rows.append({
            "id": f"T-{i+1:03d}",
            "ns": f"NS{i+1:03d}",
            "rua": rua,
            "pv_ini": t.get("pv_ini", ""),
            "pv_fim": t.get("pv_fim", ""),
            "trecho": f"{t.get('pv_ini','')}->{t.get('pv_fim','')}",
            "ext_m": round(ext, 2),
            "dn_mm": dn,
            "material": t.get("material", "PVC"),
            "prof_ini": t.get("prof_ini") or pvs.get(t.get("pv_ini",""),{}).get("prof"),
            "prof_fim": t.get("prof_fim") or pvs.get(t.get("pv_fim",""),{}).get("prof"),
            "prof_med": pm,
            "decl_mm": t.get("decl_mm"),
            **q, **c,
        })

    # ═══════════════════════════════════════════════════════════════════════════
    # ABA 1: TRECHOS (todos PV a PV)
    # ═══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "TRECHOS"

    cols = [
        ("ID", 8), ("NS", 8), ("Rua", 30), ("PV Inicio", 12), ("PV Fim", 12),
        ("Trecho", 22), ("Comp(m)", 10), ("DN(mm)", 8), ("Material", 10),
        ("Prof Ini(m)", 10), ("Prof Fim(m)", 10), ("Prof Med(m)", 10),
        ("Decl(m/m)", 10),
        ("Escav(m3)", 10), ("Lastro(m3)", 10), ("Envolt(m3)", 10),
        ("Brita(m3)", 10), ("Reaterro(m3)", 10), ("Bota-fora(m3)", 10),
        ("Pavim(m2)", 10), ("Escor(m2)", 10), ("Tubos(barras)", 10),
        ("PVs(un)", 8),
        ("Custo Tubo(R$)", 14), ("Custo Escav(R$)", 14),
        ("Custo Total(R$)", 14),
    ]

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    title_cell = ws.cell(1, 1, f"MEGA INTEGRADA - {nucleo.upper()} - {len(rows)} TRECHOS PV a PV")
    title_cell.font = Font(bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center")

    ws.cell(2, 1, f"Gerado: {datetime.now().strftime('%d/%m/%Y %H:%M')} | "
            f"Extensão: {sum(r['ext_m'] for r in rows):,.1f}m | "
            f"PVs: {len(pvs)}")

    # Header
    for j, (name, width) in enumerate(cols, 1):
        cell = ws.cell(4, j, name)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = brd
        ws.column_dimensions[cell.column_letter].width = width

    keys = ["id", "ns", "rua", "pv_ini", "pv_fim", "trecho",
            "ext_m", "dn_mm", "material", "prof_ini", "prof_fim", "prof_med",
            "decl_mm",
            "esc_m3", "lastro_m3", "envolt_m3", "brita_m3", "reat_m3", "bota_m3",
            "pav_m2", "escor_m2", "tubo_barras", None,
            "custo_tubo", "custo_escav", "custo_total"]

    for i, r in enumerate(rows):
        row_num = 5 + i
        for j, k in enumerate(keys, 1):
            if k is None:
                val = 1  # PVs por trecho
            else:
                val = r.get(k, "")
            cell = ws.cell(row_num, j, val if val is not None else "")
            cell.border = brd
            if isinstance(val, float):
                cell.number_format = num_fmt_2
            if k and "custo" in str(k):
                cell.number_format = money_fmt

    # Totais
    tot_row = 5 + len(rows)
    ws.cell(tot_row, 1, "TOTAL").font = tot_font
    for j in range(1, len(cols) + 1):
        ws.cell(tot_row, j).fill = tot_fill
        ws.cell(tot_row, j).font = tot_font
        ws.cell(tot_row, j).border = brd

    # Somas nas colunas numéricas
    sum_cols = {7: "ext_m", 14: "esc_m3", 15: "lastro_m3", 16: "envolt_m3",
                17: "brita_m3", 18: "reat_m3", 19: "bota_m3", 20: "pav_m2",
                21: "escor_m2", 22: "tubo_barras", 23: None,
                24: "custo_tubo", 25: "custo_escav", 26: "custo_total"}
    for col_idx, key in sum_cols.items():
        if key is None:
            val = len(rows)
        else:
            val = sum(r.get(key, 0) or 0 for r in rows)
        cell = ws.cell(tot_row, col_idx, round(val, 2))
        cell.number_format = money_fmt if key and "custo" in str(key) else num_fmt_2

    # Filtros
    ws.auto_filter.ref = f"A4:{chr(64+len(cols))}{tot_row-1}"
    ws.freeze_panes = "A5"

    # ═══════════════════════════════════════════════════════════════════════════
    # ABA 2: QUANTITATIVOS POR RUA
    # ═══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("QUANTITATIVOS POR RUA")

    ruas_dict = {}
    for r in rows:
        rua = r["rua"]
        if rua not in ruas_dict:
            ruas_dict[rua] = {"ext_m": 0, "trechos": 0, "esc_m3": 0,
                              "lastro_m3": 0, "envolt_m3": 0, "brita_m3": 0,
                              "reat_m3": 0, "pav_m2": 0, "escor_m2": 0,
                              "tubo_barras": 0, "custo_total": 0}
        d = ruas_dict[rua]
        d["trechos"] += 1
        for k in ["ext_m", "esc_m3", "lastro_m3", "envolt_m3", "brita_m3",
                   "reat_m3", "pav_m2", "escor_m2", "tubo_barras", "custo_total"]:
            d[k] += r.get(k, 0) or 0

    ws2.merge_cells("A1:L1")
    ws2.cell(1, 1, f"QUANTITATIVOS POR RUA - {nucleo.upper()}").font = Font(bold=True, size=14, color="1F4E79")

    rua_cols = [("Rua", 30), ("Trechos", 8), ("Ext(m)", 10), ("Escav(m3)", 10),
                ("Lastro(m3)", 10), ("Envolt(m3)", 10), ("Brita(m3)", 10),
                ("Reaterro(m3)", 10), ("Pavim(m2)", 10), ("Escor(m2)", 10),
                ("Tubos(barras)", 10), ("Custo Total(R$)", 14)]

    for j, (name, w) in enumerate(rua_cols, 1):
        cell = ws2.cell(3, j, name)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = brd
        ws2.column_dimensions[cell.column_letter].width = w

    rua_keys = ["trechos", "ext_m", "esc_m3", "lastro_m3", "envolt_m3",
                "brita_m3", "reat_m3", "pav_m2", "escor_m2", "tubo_barras", "custo_total"]

    for i, (rua, d) in enumerate(sorted(ruas_dict.items(), key=lambda x: -x[1]["ext_m"])):
        row_num = 4 + i
        ws2.cell(row_num, 1, rua).border = brd
        for j, k in enumerate(rua_keys, 2):
            cell = ws2.cell(row_num, j, round(d[k], 2))
            cell.border = brd
            cell.number_format = money_fmt if "custo" in k else num_fmt_2

    # Total
    rua_tot = 4 + len(ruas_dict)
    ws2.cell(rua_tot, 1, "TOTAL").font = tot_font
    for j in range(1, len(rua_cols) + 1):
        ws2.cell(rua_tot, j).fill = tot_fill
        ws2.cell(rua_tot, j).font = tot_font
        ws2.cell(rua_tot, j).border = brd
    for j, k in enumerate(rua_keys, 2):
        val = sum(d[k] for d in ruas_dict.values())
        ws2.cell(rua_tot, j, round(val, 2)).number_format = money_fmt if "custo" in k else num_fmt_2

    # ═══════════════════════════════════════════════════════════════════════════
    # ABA 3: RESUMO
    # ═══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("RESUMO")
    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 20
    ws3.column_dimensions["C"].width = 12

    ws3.merge_cells("A1:C1")
    ws3.cell(1, 1, f"RESUMO MEGA INTEGRADA - {nucleo.upper()}").font = Font(bold=True, size=14, color="1F4E79")

    resumo = [
        ("Total de Trechos PV a PV", len(rows), "trechos"),
        ("Total de PVs", len(pvs), "unidades"),
        ("Extensão Total da Rede", round(sum(r["ext_m"] for r in rows), 1), "m"),
        ("Ruas/Logradouros", len([r for r in ruas_dict if r != "Sem Rua"]), "ruas"),
        ("", "", ""),
        ("QUANTITATIVOS CONSOLIDADOS", "", ""),
        ("Escavação Total", round(sum(r.get("esc_m3",0) or 0 for r in rows), 2), "m³"),
        ("Lastro", round(sum(r.get("lastro_m3",0) or 0 for r in rows), 2), "m³"),
        ("Envoltória", round(sum(r.get("envolt_m3",0) or 0 for r in rows), 2), "m³"),
        ("Brita", round(sum(r.get("brita_m3",0) or 0 for r in rows), 2), "m³"),
        ("Reaterro", round(sum(r.get("reat_m3",0) or 0 for r in rows), 2), "m³"),
        ("Pavimentação", round(sum(r.get("pav_m2",0) or 0 for r in rows), 2), "m²"),
        ("Escoramento", round(sum(r.get("escor_m2",0) or 0 for r in rows), 2), "m²"),
        ("Tubos (barras 6m)", sum(r.get("tubo_barras",0) or 0 for r in rows), "barras"),
        ("", "", ""),
        ("CUSTOS (c/ BDI 25%)", "", ""),
        ("Custo Total", round(sum(r.get("custo_total",0) or 0 for r in rows), 2), "R$"),
        ("Custo/metro", round(sum(r.get("custo_total",0) or 0 for r in rows) / max(1, sum(r["ext_m"] for r in rows)), 2), "R$/m"),
    ]

    for i, (desc, val, un) in enumerate(resumo, 3):
        ws3.cell(i, 1, desc).border = brd
        c = ws3.cell(i, 2, val)
        c.border = brd
        if isinstance(val, float):
            c.number_format = money_fmt if "R$" in un else num_fmt_2
        ws3.cell(i, 3, un).border = brd
        if desc and desc.isupper():
            ws3.cell(i, 1).font = tot_font
            ws3.cell(i, 1).fill = sub_fill

    # ═══════════════════════════════════════════════════════════════════════════
    # ABA 4: MATERIAIS CONSOLIDADOS
    # ═══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("MATERIAIS")
    ws4.merge_cells("A1:F1")
    ws4.cell(1, 1, f"MATERIAIS CONSOLIDADOS - {nucleo.upper()}").font = Font(bold=True, size=14, color="1F4E79")

    mat_cols = [("Material", 35), ("Un.", 8), ("Qtd Total", 12),
                ("Custo Unit(R$)", 14), ("Total s/BDI(R$)", 14), ("Total c/BDI(R$)", 14)]
    for j, (name, w) in enumerate(mat_cols, 1):
        cell = ws4.cell(3, j, name)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = brd
        ws4.column_dimensions[cell.column_letter].width = w

    # Agregar materiais por DN
    dn_counts = {}
    for r in rows:
        dn = r["dn_mm"]
        if dn not in dn_counts:
            dn_counts[dn] = {"ext_m": 0, "barras": 0}
        dn_counts[dn]["ext_m"] += r["ext_m"]
        dn_counts[dn]["barras"] += r.get("tubo_barras", 0)

    mat_rows = []
    for dn in sorted(dn_counts):
        d = dn_counts[dn]
        chave = f"tubo_pvc_{dn}"
        pu = SINAPI.get(chave, SINAPI["tubo_pvc_200"])
        mat_rows.append((f"Tubo PVC DN{dn}mm (barras 6m)", "barra", d["barras"],
                         pu * 6, d["barras"] * pu * 6, d["barras"] * pu * 6 * BDI))
        mat_rows.append((f"Tubo PVC DN{dn}mm", "m", round(d["ext_m"], 1),
                         pu, round(d["ext_m"] * pu, 2), round(d["ext_m"] * pu * BDI, 2)))

    tot_barras = sum(r.get("tubo_barras", 0) for r in rows)
    mat_rows.append(("PV concreto DN1200mm", "pc", len(pvs),
                     SINAPI["pv_concreto_1200"],
                     len(pvs) * SINAPI["pv_concreto_1200"],
                     len(pvs) * SINAPI["pv_concreto_1200"] * BDI))

    q_esc = sum(r.get("esc_m3", 0) or 0 for r in rows)
    q_reat = sum(r.get("reat_m3", 0) or 0 for r in rows)
    q_last = sum(r.get("lastro_m3", 0) or 0 for r in rows)
    q_brit = sum(r.get("brita_m3", 0) or 0 for r in rows)
    q_pav = sum(r.get("pav_m2", 0) or 0 for r in rows)
    q_escor = sum(r.get("escor_m2", 0) or 0 for r in rows)

    for desc, un, qtd, pu_key in [
        ("Escavação mecânica", "m³", q_esc, "escavacao"),
        ("Reaterro compactado", "m³", q_reat, "reaterro"),
        ("Areia lastro", "m³", q_last, "lastro"),
        ("Brita dreno", "m³", q_brit, "brita"),
        ("Pavimentação", "m²", q_pav, "pavimentacao"),
        ("Escoramento", "m²", q_escor, "escoramento_m2"),
    ]:
        pu = SINAPI.get(pu_key, 0)
        mat_rows.append((desc, un, round(qtd, 2), pu, round(qtd * pu, 2), round(qtd * pu * BDI, 2)))

    for i, (desc, un, qtd, pu, total_s, total_c) in enumerate(mat_rows, 4):
        ws4.cell(i, 1, desc).border = brd
        ws4.cell(i, 2, un).border = brd
        ws4.cell(i, 3, qtd).border = brd
        ws4.cell(i, 3).number_format = num_fmt_2
        ws4.cell(i, 4, pu).border = brd
        ws4.cell(i, 4).number_format = money_fmt
        ws4.cell(i, 5, total_s).border = brd
        ws4.cell(i, 5).number_format = money_fmt
        ws4.cell(i, 6, total_c).border = brd
        ws4.cell(i, 6).number_format = money_fmt

    # Total geral materiais
    mt = 4 + len(mat_rows)
    ws4.cell(mt, 1, "TOTAL GERAL").font = tot_font
    for j in range(1, 7):
        ws4.cell(mt, j).fill = tot_fill
        ws4.cell(mt, j).font = tot_font
        ws4.cell(mt, j).border = brd
    ws4.cell(mt, 5, round(sum(r[4] for r in mat_rows), 2)).number_format = money_fmt
    ws4.cell(mt, 6, round(sum(r[5] for r in mat_rows), 2)).number_format = money_fmt

    # ── Salvar ───────────────────────────────────────────────────────────────
    saida_dir = Path(saida_dir)
    saida_dir.mkdir(parents=True, exist_ok=True)
    nome = f"MEGA_INTEGRADA_{nucleo.upper().replace(' ', '_')}.xlsx"
    out_path = saida_dir / nome
    wb.save(str(out_path))
    print(f"OK: {out_path}")
    print(f"   {len(rows)} trechos | {len(pvs)} PVs | {len(ruas_dict)} ruas")
    return str(out_path)


# ═══════════════════════════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser(description="Gera planilha MEGA INTEGRADA")
    ap.add_argument("arquivo", help="DXF, DWG ou JSON da rede")
    ap.add_argument("--saida", default=None, help="Pasta de saída")
    ap.add_argument("--nucleo", default=None, help="Nome do núcleo")
    ap.add_argument("--gpkg", default=None, help="GPKG cartografia para ruas")
    args = ap.parse_args()

    path = Path(args.arquivo)
    nucleo = args.nucleo or path.stem.replace("_", " ").title()

    ext = path.suffix.lower()
    if ext == ".dxf":
        from ler_dxf_gdal import ler_dxf_gdal
        pvs, trechos, ruas, meta = ler_dxf_gdal(str(path))
    elif ext == ".dwg":
        from ler_dwg_universal import ler_dwg_universal
        pvs, trechos, meta = ler_dwg_universal(str(path))
    elif ext == ".xml":
        from ler_landxml import ler_landxml
        pvs, trechos, ruas, meta = ler_landxml(str(path))
    elif ext == ".json":
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        pvs = data.get("pvs", {})
        trechos = data.get("trechos", [])
    else:
        print(f"Formato não suportado: {ext}")
        sys.exit(1)

    if not trechos:
        print("ERRO: nenhum trecho encontrado!")
        sys.exit(1)

    # Enriquecer trechos
    try:
        from gerar_ns import enriquecer_trechos
        trechos = enriquecer_trechos(trechos, pvs)
    except ImportError:
        pass

    if args.gpkg:
        associar_ruas_gpkg(trechos, pvs, args.gpkg)

    saida = args.saida or str(path.parent / "PLANEJAMENTO")
    gerar_mega(pvs, trechos, nucleo, saida)
