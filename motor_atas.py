#!/usr/bin/env python3
"""
MOTOR_ATAS.PY — Motor Integrado de Atas de Reunião
ConstruData - HydroNetwork · FCN Construções e Saneamento

Integração de Alto Impacto: Supabase / n8n
Objetivo: Evitar hiper-fragmentação, unificando a geração em .xlsx 
e o dispáro via Databricks/Supabase num módulo estrutural único.
"""

import os
import json
import logging
import requests
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Credentials
SUPABASE_URL = os.environ.get("SUPABASE_URL", "https://vblfdikfobsirwpdnybw.supabase.co")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY", "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InZibGZkaWtmb2JzaXJ3cGRueWJ3Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzMzNzAwODIsImV4cCI6MjA4ODk0NjA4Mn0.GOx3HoMh3P2Zzxz8BxNsfQBfXwsNZNQsdVc3nJaqRy4")

# Mock de fones da base (para n8n WhatsApp)
TELEFONES_MAP = {
    "Eng. Ícaro": "5537998268576",  # Real Ícaro (Pardinho/Tatuí)
    "Eng. Mateus": "5561991015639", # Real Mateus (Osasco/CSU)
    "Eng. Luiz": "5537999425397",   # Real Luiz Fernando (Diretoria)
}

class MotorAtas:
    def __init__(self, data_reuniao, titulo, local):
        self.data_reuniao = data_reuniao
        self.titulo = titulo
        self.local = local
        self.headers = {
            "apikey": SUPABASE_KEY,
            "Authorization": f"Bearer {SUPABASE_KEY}",
            "Content-Type": "application/json",
            "Prefer": "return=minimal"
        }
        self.tarefas_para_enviar = []

    def adicionar_tarefas(self, responsavel, descricao, prazo=None, prioridade="normal"):
        self.tarefas_para_enviar.append({
            "delegante": "Diretoria",
            "responsavel": responsavel,
            "responsavel_phone": TELEFONES_MAP.get(responsavel, ""),
            "descricao": descricao,
            "prazo": prazo,
            "prioridade": prioridade,
            "origem": "ata_reuniao",
            "metadata": {"data_ata": self.data_reuniao, "titulo": self.titulo}
        })
        
    def disparar_para_supabase(self):
        """Envia todas as tarefas carregadas da ata diretamente para o Supabase (n8n hook)"""
        logging.info(f"🚀 Iniciando push de {len(self.tarefas_para_enviar)} tarefas para o Supabase DataLake...")
        
        endpoint = f"{SUPABASE_URL}/rest/v1/tarefas"
        
        # Faz uma limpa nos Nones pro formato JSONb SQL do Supabase
        cleaned_payload = []
        for t in self.tarefas_para_enviar:
            t_copy = t.copy()
            if not t_copy.get("prazo"):
                t_copy["prazo"] = None
            cleaned_payload.append(t_copy)

        try:
            r = requests.post(endpoint, headers=self.headers, json=cleaned_payload)
            if r.status_code in (200, 201, 204):
                logging.info(f"✅ Ingestão Suprema Realizada com Sucesso! {len(cleaned_payload)} tarefas gravadas no schema público do Supabase.")
            else:
                logging.error(f"❌ Falha no Supabase: {r.status_code} - {r.text}")
        except Exception as e:
            logging.error(f"❌ Exceção Crítica no HTTP push: {e}")

    def gerar_excel(self, filename="ATA_OUTPUT.xlsx"):
        """Recria o gerador do projeto de forma dinâmica baseada nas tarefas injetadas no fluxo."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ata de Reunião"

        # Colors & Fonts Stylers (Da versão anterior)
        HEADER_COLOR = "002060"
        SUBHEADER_COLOR = "4F81BD"
        WHITE = "FFFFFF"
        LIGHT_BLUE = "DCE6F1"

        font_title = Font(name='Segoe UI', size=16, bold=True, color=WHITE)
        font_subtitle = Font(name='Segoe UI', size=12, bold=True, color=WHITE)
        font_bold = Font(name='Segoe UI', size=11, bold=True)
        font_normal = Font(name='Segoe UI', size=11)
        font_warning = Font(name='Segoe UI', size=11, bold=True, color="9C6500")

        align_center = Alignment(horizontal="center", vertical="center")
        align_left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        def _merge_and_format(ranges, value, font, alignment, fill_color=None):
            ws.merge_cells(ranges)
            cell = ws[ranges.split(':')[0]]
            cell.value = value
            cell.font = font
            cell.alignment = alignment
            if fill_color:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            
            # borders over merged
            for r in ws[ranges]:
                for c in r:
                    c.border = thin_border

        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 5
        ws.column_dimensions['C'].width = 60
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 3

        row = 2
        _merge_and_format(f'B{row}:E{row}', f"ATA DE REUNIÃO - {self.titulo.upper()}", font_title, align_center, HEADER_COLOR)
        ws.row_dimensions[row].height = 30
        row += 2

        # INFOS
        _merge_and_format(f'B{row}:E{row}', "INFORMAÇÕES GERAIS", font_subtitle, align_center, SUBHEADER_COLOR)
        row += 1
        ws[f'B{row}'] = "Data:"; ws[f'B{row}'].font = font_bold; ws[f'C{row}'] = self.data_reuniao
        ws[f'D{row}'] = "Local:"; ws[f'D{row}'].font = font_bold; ws[f'E{row}'] = self.local
        row += 2

        # TAREFAS
        _merge_and_format(f'B{row}:E{row}', "DIRETRIZES E TAREFAS DISCUTIDAS", font_subtitle, align_center, SUBHEADER_COLOR)
        row += 1

        ws[f'B{row}'] = "#"; ws[f'B{row}'].font = font_bold; ws[f'B{row}'].fill = PatternFill(fill_type="solid", start_color=LIGHT_BLUE)
        ws[f'C{row}'] = "Ação / Diretriz"; ws[f'C{row}'].font = font_bold; ws[f'C{row}'].fill = PatternFill(fill_type="solid", start_color=LIGHT_BLUE)
        ws[f'D{row}'] = "Responsável"; ws[f'D{row}'].font = font_bold;  ws[f'D{row}'].fill = PatternFill(fill_type="solid", start_color=LIGHT_BLUE)
        ws[f'E{row}'] = "Prazo/Prioridade"; ws[f'E{row}'].font = font_bold; ws[f'E{row}'].fill = PatternFill(fill_type="solid", start_color=LIGHT_BLUE)
        row += 1

        for i, t in enumerate(self.tarefas_para_enviar, 1):
            ws[f'B{row}'] = i; ws[f'B{row}'].alignment = align_center; ws[f'B{row}'].border = thin_border
            ws[f'C{row}'] = t["descricao"]; ws[f'C{row}'].alignment = align_left_top; ws[f'C{row}'].border = thin_border
            ws[f'D{row}'] = t["responsavel"]; ws[f'D{row}'].alignment = align_center; ws[f'D{row}'].border = thin_border
            
            p_val = t.get("prazo") if t.get("prazo") else t.get("prioridade", "").upper()
            ws[f'E{row}'] = p_val
            ws[f'E{row}'].font = font_warning if t.get("prioridade") in ["alta", "urgente"] else font_normal
            ws[f'E{row}'].alignment = align_center
            ws[f'E{row}'].border = thin_border
            
            ws.row_dimensions[row].height = 45 # wrap text needs height
            row += 1

        wb.save(filename)
        logging.info(f"✅ Arquivo {filename} gerado fisicamente no disco com sucesso!")

if __name__ == "__main__":
    motor = MotorAtas(
        data_reuniao="19/04/2026",
        titulo="Alinhamento Estratégico - CGMT & CSU",
        local="Consórcio CGMT & Consórcio CSU"
    )

    # 1. Pautas Ícaro
    motor.adicionar_tarefas(
        responsavel="Eng. Ícaro",
        descricao="Elaborar e apresentar diários de obra de Tatuí impreterivelmente até amanhã à noite.",
        prazo="2026-04-20",
        prioridade="urgente"
    )
    motor.adicionar_tarefas(
        responsavel="Eng. Ícaro",
        descricao="Iniciar imediatamente as duas obras sob sua responsabilidade com produção total.",
        prioridade="alta"
    )
    motor.adicionar_tarefas(
        responsavel="Eng. Ícaro",
        descricao="Realizar a contratação urgente de efetivo, visto déficit e falta de visibilidade para a diretoria.",
        prioridade="urgente"
    )
    motor.adicionar_tarefas(
        responsavel="Eng. Ícaro",
        descricao="Produzir vídeos das duas obras para fins de criação de conteúdo institucional.",
        prioridade="normal"
    )
    motor.adicionar_tarefas(
        responsavel="Eng. Ícaro",
        descricao="Concluir as duas obras até o dia 30/04, sendo necessário extensão de jornada (período noturno) se necessário.",
        prazo="2026-04-30",
        prioridade="urgente"
    )

    # 2. Pautas Mateus
    motor.adicionar_tarefas(
        responsavel="Eng. Mateus",
        descricao="Providenciar a mobilização/mobiliário da casa e apresentar cotação para validação.",
        prioridade="alta"
    )
    motor.adicionar_tarefas(
        responsavel="Eng. Mateus",
        descricao="Efetuar com urgência a troca da locadora de equipamentos leves.",
        prioridade="urgente"
    )
    motor.adicionar_tarefas(
        responsavel="Eng. Mateus",
        descricao="Realizar o desligamento do colaborador (pedreiro) devido a comportamento inadequado.",
        prioridade="alta"
    )
    motor.adicionar_tarefas(
        responsavel="Eng. Mateus",
        descricao="Contratação imediata de efetivo para segregação de frentes (prazo com 1 semana de atraso).",
        prioridade="urgente"
    )
    motor.adicionar_tarefas(
        responsavel="Eng. Mateus",
        descricao="Garantir produção e entrega da frente de serviço até o dia 30/04, sem exceções.",
        prazo="2026-04-30",
        prioridade="urgente"
    )
    motor.adicionar_tarefas(
        responsavel="Eng. Mateus",
        descricao="Reestruturar e otimizar a logística de materiais na obra (visando produtividade).",
        prioridade="alta"
    )

    # 3. Pautas Luiz (Adicionadas a pedido da Diretoria)
    motor.adicionar_tarefas(
        responsavel="Eng. Luiz",
        descricao="Aprovação final das medições dos fornecedores e envio do report financeiro da quinzena.",
        prazo="2026-04-22",
        prioridade="alta"
    )
    motor.adicionar_tarefas(
        responsavel="Eng. Luiz",
        descricao="Auditoria rigorosa de Qualidade e Segurança (QSMS) nas frentes de serviço segregadas.",
        prioridade="urgente"
    )
    motor.adicionar_tarefas(
        responsavel="Eng. Luiz",
        descricao="Apoio estratégico na validação do orçamento para mobiliário e equipamentos leves.",
        prioridade="normal"
    )

    # Motor em Ação "Antigravity":
    motor.gerar_excel(os.path.join(os.getcwd(), "ATA_REUNIAO_19_04_2026.xlsx"))
    motor.disparar_para_supabase()
