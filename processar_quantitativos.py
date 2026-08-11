#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PROCESSAR QUANTITATIVOS — PROJETADO × EXECUTADO × SALDO
========================================================
Motor: ConstruData HydroNetwork v9 (NOVA NS Versao 5)
CT 11481051 — Consórcio Se Liga Na Rede — Santos

Pipeline:
1. Lê TODOS os projetos (DXF esgoto/água + XML prolongamentos) da pasta PROJETOS
2. Resolve nomes de ruas via GPKG das cartografias
3. Calcula volumes PV a PV: Escavação, Reaterro, Areia, Brita, Pavimentação
4. Carrega SHP executado (linhas executado.shp) de TODA OBRA
5. Match geoespacial trecho × shp (midpoint + endpoints, tol=15m)
6. Gera planilha final: POR TRECHO + POR RUA + POR NÚCLEO + RESUMO GERAL
"""
import sys, os, json, math, re, warnings, unicodedata
from pathlib import Path
from datetime import datetime
from collections import defaultdict, Counter

warnings.filterwarnings("ignore")

# ── Setup motor ──────────────────────────────────────────────────────────────
MOTOR_DIR = Path(r"C:\Users\felip\Downloads\NOVA NS Versao 5")
sys.path.insert(0, str(MOTOR_DIR))

from ler_landxml import ler_landxml
from ler_dxf_gdal import ler_dxf_gdal
from gerar_ns import enriquecer_trechos, calc_manning, calcular_materiais, CONTRATO

try:
    from motor_custo import custo_trecho, FATORES, BDI
    HAS_CUSTO = True
except ImportError:
    HAS_CUSTO = False
    FATORES = {"CBUQ provisório": 0.88, "Areia": 0.25, "Berço": 0.24, "Brita base": 0.22}

import numpy as np
import geopandas as gpd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── PASTAS ───────────────────────────────────────────────────────────────────
INPUT_DIR   = Path(r"C:\Users\felip\Desktop\PROJETOS\Nova pasta")
EXEC_DIR    = Path(r"C:\Users\felip\Desktop\TODA OBRA")
OUTPUT_FILE = INPUT_DIR / "QUANTITATIVO_GLOBAL_EXEC_vs_PROJ.xlsx"

# ── PROJETOS (mesma config do processar_tudo_por_rua.py) ─────────────────────
PROJETOS = [
    # XMLs (LandXML)
    {"arquivo": "ESTUDO - CT JOÃO CARLOS DA SILVA.xml",  "nucleo": "Joao Carlos",             "tipo": "ESGOTO", "gpkg": "MAPA JOÃO CARLOS DA SILVA_RV07 (QGIS).gpkg"},
    {"arquivo": "ESTUDO - SÃO MANOEL.xml",               "nucleo": "Sao Manoel",              "tipo": "ESGOTO", "gpkg": "MAPA SÃO MANOEL_RV05 (QGIS).gpkg"},
    {"arquivo": "PROLONGAMENTO CRIADORES.xml",            "nucleo": "Prolongamento Criadores",  "tipo": "ESGOTO", "gpkg": "CARTOGRAFIA VILA CRIADORES_R06 (QGIS).gpkg"},
    {"arquivo": "PROLONGAMENTO PANTANAL BAIXO.xml",       "nucleo": "Prolongamento Pantanal",   "tipo": "ESGOTO", "gpkg": "MAPA PANTANAL BAIXO_R04 (QGIS).gpkg"},
    {"arquivo": "PROLONGAMENTO SÃO MANOEL.xml",           "nucleo": "Prolongamento Sao Manoel", "tipo": "ESGOTO", "gpkg": "MAPA SÃO MANOEL_RV05 (QGIS).gpkg"},
    {"arquivo": "PROLONGAMENTO TETEU ALT-01.xml",         "nucleo": "Prolongamento Teteu Alt",  "tipo": "ESGOTO", "gpkg": "MAPA TETEU-VALE VERDE_R04 (QGIS).gpkg"},

    # DXFs ESGOTO
    {"arquivo": "CRIADORES_ESGOTO_REV.02.dxf",           "nucleo": "Vila Criadores",           "tipo": "ESGOTO", "gpkg": "CARTOGRAFIA VILA CRIADORES_R06 (QGIS).gpkg"},
    {"arquivo": "ISRAEL_ESGOTO.dxf",                      "nucleo": "Vila Israel",              "tipo": "ESGOTO", "gpkg": "MAPA PANTANAL ALTO (VILA ISRAEL)_RV02 (QGIS).gpkg"},
    {"arquivo": "PANTANAL_ESGOTO.dxf",                    "nucleo": "Pantanal Baixo",           "tipo": "ESGOTO", "gpkg": "MAPA PANTANAL BAIXO_R04 (QGIS).gpkg"},
    {"arquivo": "TETÉU_ESGOTO-GAE019.dxf",               "nucleo": "Morro do Teteu",           "tipo": "ESGOTO", "gpkg": "MAPA TETEU-VALE VERDE_R04 (QGIS).gpkg"},

    # DXFs ÁGUA
    {"arquivo": "Projeto Criadores- AGUA.dxf",            "nucleo": "Vila Criadores",           "tipo": "AGUA",   "gpkg": "CARTOGRAFIA VILA CRIADORES_R06 (QGIS).gpkg"},
    {"arquivo": "ISRAEL_ÁGUA.dxf",                        "nucleo": "Vila Israel",              "tipo": "AGUA",   "gpkg": "MAPA PANTANAL ALTO (VILA ISRAEL)_RV02 (QGIS).gpkg"},
    {"arquivo": "TETEU_ÁGUA.dxf",                         "nucleo": "Morro do Teteu",           "tipo": "AGUA",   "gpkg": "MAPA TETEU-VALE VERDE_R04 (QGIS).gpkg"},
]

# KM reais de ESGOTO por núcleo (para validação)
KM_REAIS_ESG = {
    "Morro do Teteu": 4.0, "Joao Carlos": 1.3, "Vila Criadores": 2.3,
    "Sao Manoel": 2.3, "Pantanal Baixo": 2.3, "Vila Israel": 1.4,
}

# ══════════════════════════════════════════════════════════════════════════════
# FUNÇÕES AUXILIARES (do motor original)
# ══════════════════════════════════════════════════════════════════════════════

def log(msg, nivel="INFO"):
    ts = datetime.now().strftime("%H:%M:%S")
    prefix = {"OK": "[OK]  ", "WARN": "[!]   ", "STEP": ">>> ", "ERR": "[X]  "}.get(nivel, "      ")
    line = f"[{ts}] {prefix}{msg}"
    try:
        print(line)
    except UnicodeEncodeError:
        print(line.encode("ascii", "replace").decode())


def limpar_rua(rua):
    if not rua or str(rua).strip() in ("", "nan", "None", "Sem Rua", "1"):
        return "Sem Rua"
    txt = str(rua).replace("\n", " ").replace("\r", " ")
    txt = re.sub(r'[\x00-\x1f\x7f]', '', txt)
    txt = re.sub(r'\s+', ' ', txt).strip()
    return txt if txt else "Sem Rua"


def carregar_ruas_gpkg(gpkg_path, trechos, pvs):
    """Carrega nomes de ruas do GeoPackage e associa aos trechos (do exportar_completo.py)."""
    if not gpkg_path or not Path(gpkg_path).exists():
        return 0
    try:
        gdf = gpd.read_file(str(gpkg_path))
        rua_col = None
        for col in gdf.columns:
            c = col.upper()
            if any(k in c for k in ["RUA", "LOGRADOURO", "NOME", "NAME", "DESCRICAO", "LABEL", "TEXT"]):
                rua_col = col
                break
        if not rua_col:
            return 0
        ruas_gdf = gdf[gdf[rua_col].notna() & (gdf[rua_col] != "")].copy()
        if len(ruas_gdf) == 0:
            return 0
        centroids = ruas_gdf.geometry.centroid
        rua_xy = np.array([[c.x, c.y] for c in centroids])
        rua_names = ruas_gdf[rua_col].values
        n_associados = 0
        TOL = 100.0
        for t in trechos:
            if t.get("rua") and str(t["rua"]).strip() not in ("", "Sem Rua", "nan", "None"):
                continue
            p0 = pvs.get(t.get("pv_ini"), {})
            p1 = pvs.get(t.get("pv_fim"), {})
            if not p0.get("x") or not p1.get("x"):
                continue
            mx = (p0["x"] + p1["x"]) / 2
            my = (p0["y"] + p1["y"]) / 2
            dists = np.hypot(rua_xy[:, 0] - mx, rua_xy[:, 1] - my)
            idx = int(np.argmin(dists))
            if dists[idx] <= TOL:
                t["rua"] = str(rua_names[idx]).strip()
                n_associados += 1
        return n_associados
    except Exception as e:
        log(f"  Erro GPKG: {e}", "WARN")
        return 0


# ══════════════════════════════════════════════════════════════════════════════
# CARREGAR SHP EXECUTADO (lógica do compilar_executado_v3.py)
# ══════════════════════════════════════════════════════════════════════════════

def carregar_shp_executado():
    """Carrega TODAS as linhas executadas dos shapefiles em TODA OBRA."""
    log("Carregando SHP executado...", "STEP")
    linhas = []

    # Buscar todos os .shp com "linhas" no nome em TODA OBRA (recursive)
    for root, dirs, files in os.walk(str(EXEC_DIR)):
        for f in files:
            if f.lower().endswith('.shp') and 'linhas' in f.lower():
                shp_path = Path(root) / f
                log(f"  SHP: {shp_path.name}", "INFO")
                try:
                    cpg = shp_path.with_suffix(".cpg")
                    cpg.write_text("latin-1", encoding="utf-8")
                    os.environ["SHAPE_ENCODING"] = "latin-1"
                    gdf = gpd.read_file(str(shp_path))
                    gdf['length_m'] = gdf.geometry.length
                    for _, row in gdf.iterrows():
                        g = row.geometry
                        if g is None:
                            continue
                        mid = g.interpolate(0.5, normalized=True)
                        dn = 200
                        layer = str(row.get('GM_LAYER', ''))
                        if 'DN 150' in layer: dn = 150
                        elif 'DN 200' in layer: dn = 200
                        elif 'DN 300' in layer: dn = 300
                        elif 'DN 63' in layer: dn = 63
                        else:
                            d = row.get('Diametro')
                            if d:
                                d = float(d)
                                dn = round(d*1000) if d < 1 else round(d) if d < 1000 else 200
                        mat = str(row.get('material', 'PVC'))
                        tipo = 'AGUA' if 'agua' in str(row.get('tipo', '')).lower() or 'PEAD' in mat else 'ESGOTO'
                        linhas.append({
                            'mx': mid.x, 'my': mid.y,
                            'sx': g.coords[0][0], 'sy': g.coords[0][1],
                            'ex': g.coords[-1][0], 'ey': g.coords[-1][1],
                            'length': row['length_m'], 'dn': dn, 'material': mat,
                            'tipo': tipo, 'data': str(row.get('Data_Inst', ''))
                        })
                except Exception as e:
                    log(f"  Erro SHP {f}: {e}", "WARN")

    log(f"  Total linhas executadas: {len(linhas)}", "OK")
    return linhas


def match_shp(trecho, pvs, linhas, tol=15.0):
    """Match geoespacial trecho × shp (midpoint + endpoints). Do compilar_executado_v3.py."""
    p0 = pvs.get(trecho.get("pv_ini"), {})
    p1 = pvs.get(trecho.get("pv_fim"), {})
    x0, y0 = p0.get("x", 0), p0.get("y", 0)
    x1, y1 = p1.get("x", 0), p1.get("y", 0)
    if x0 == 0 or x1 == 0:
        return False, None
    mx = (x0 + x1) / 2
    my = (y0 + y1) / 2
    best_d = float('inf')
    best = None
    for l in linhas:
        d = math.hypot(l['mx'] - mx, l['my'] - my)
        if d < best_d:
            best_d = d
            best = l
    if best_d <= tol:
        return True, best
    for l in linhas:
        d1 = math.hypot(l['sx'] - x0, l['sy'] - y0)
        d2 = math.hypot(l['ex'] - x1, l['ey'] - y1)
        if d1 <= tol and d2 <= tol:
            return True, l
        d1b = math.hypot(l['sx'] - x1, l['sy'] - y1)
        d2b = math.hypot(l['ex'] - x0, l['ey'] - y0)
        if d1b <= tol and d2b <= tol:
            return True, l
    return False, None


def e_cadastro(trecho, nucleo):
    """Determina se um trecho provavelmente é CADASTRO (do compilar_executado_v3.py)."""
    dn = trecho.get("dn_mm") or 200
    if "Sao Manoel" in nucleo:
        pv_ini = str(trecho.get("pv_ini", ""))
        pv_fim = str(trecho.get("pv_fim", ""))
        if pv_ini.endswith("(1)") or pv_fim.endswith("(1)"):
            return True, "CADASTRO — Rede '1' do XML (cadastro SABESP existente)"
    if dn >= 300 and "AGUA" not in nucleo:
        if any(k in nucleo for k in ["Prolong", "Pantanal", "Israel", "Teteu"]):
            return True, f"CADASTRO? — DN{dn}mm = coletor tronco (provável rede existente)"
    return False, ""


# ══════════════════════════════════════════════════════════════════════════════
# CÁLCULO DE QUANTITATIVOS (do _gerar_planilha_trechos_completa)
# ══════════════════════════════════════════════════════════════════════════════

def calc_quantitativos(trecho, pvs):
    """Calcula volumes de escavação, reaterro, areia, brita, pavimentação para 1 trecho."""
    dn_mm = trecho.get("dn_mm") or 200
    ext_m = float(trecho.get("ext_m") or 0)

    pvi = pvs.get(trecho.get("pv_ini"), {})
    pvf = pvs.get(trecho.get("pv_fim"), {})
    prof_i = pvi.get("prof")
    prof_f = pvf.get("prof")

    # Fallback: calcular prof a partir de CT - CF
    if not prof_i and pvi.get("ct") and pvi.get("cf"):
        prof_i = abs(float(pvi["ct"]) - float(pvi["cf"]))
    if not prof_f and pvf.get("ct") and pvf.get("cf"):
        prof_f = abs(float(pvf["ct"]) - float(pvf["cf"]))

    profs = [p for p in [prof_i, prof_f] if p and p > 0]
    prof_media = sum(profs) / len(profs) if profs else 1.5

    calc_dn = float(dn_mm) / 1000.0
    largura_vala = max(0.60, calc_dn + 0.50)

    vol_escavacao = round(ext_m * largura_vala * prof_media, 3)
    vol_reaterro  = round(vol_escavacao * 0.85, 3)
    area_pav      = round(ext_m * FATORES.get("CBUQ provisório", 0.88), 3)
    vol_areia     = round(ext_m * (FATORES.get("Areia", 0.25) + FATORES.get("Berço", 0.24)), 3)
    vol_brita     = round(ext_m * FATORES.get("Brita base", 0.22), 3)
    n_barras      = math.ceil(ext_m / 6) if ext_m > 0 else 0

    return {
        "ext_m": round(ext_m, 2),
        "prof_media": round(prof_media, 2),
        "largura_vala": round(largura_vala, 2),
        "vol_escavacao": vol_escavacao,
        "vol_reaterro": vol_reaterro,
        "area_pav": area_pav,
        "vol_areia": vol_areia,
        "vol_brita": vol_brita,
        "n_barras": n_barras,
    }


# ══════════════════════════════════════════════════════════════════════════════
# PIPELINE PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════

def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

    print("=" * 72)
    print("  QUANTITATIVOS GLOBAL: PROJETADO × EXECUTADO × SALDO")
    print(f"  Motor ConstruData HydroNetwork v9 | CT {CONTRATO}")
    print(f"  {len(PROJETOS)} projetos | {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print("=" * 72)

    # ── 1. CARREGAR SHP EXECUTADO ────────────────────────────────────────────
    linhas_exec = carregar_shp_executado()

    # ── 2. PROCESSAR PROJETOS ────────────────────────────────────────────────
    all_rows = []
    stats = {"exec": 0, "pend": 0, "cad": 0, "ext_e": 0, "ext_p": 0, "ext_c": 0}

    for pi, proj in enumerate(PROJETOS):
        arquivo = proj["arquivo"]
        nucleo  = proj["nucleo"]
        tipo    = proj["tipo"]
        gpkg_name = proj.get("gpkg")

        arq_path = INPUT_DIR / arquivo
        if not arq_path.exists():
            log(f"[{pi+1}/{len(PROJETOS)}] NÃO ENCONTRADO: {arquivo}", "WARN")
            continue

        log(f"[{pi+1}/{len(PROJETOS)}] {arquivo} → {nucleo} ({tipo})", "STEP")

        # 2a. LER ARQUIVO
        try:
            ext = arq_path.suffix.lower()
            if ext == ".xml":
                pvs, trechos, ruas_xml, meta = ler_landxml(str(arq_path))
            elif ext == ".dxf":
                pvs, trechos, ruas_raw, meta = ler_dxf_gdal(str(arq_path))
            else:
                log(f"  Formato não suportado: {ext}", "ERR")
                continue
        except Exception as e:
            log(f"  ERRO LEITURA: {e}", "ERR")
            continue

        if not trechos:
            log(f"  Sem trechos! Pulando.", "WARN")
            continue

        # 2b. ENRIQUECER
        trechos = enriquecer_trechos(trechos, pvs)
        ext_total = sum(t.get("ext_m", 0) for t in trechos)
        log(f"  Rede: {len(pvs)} PVs, {len(trechos)} trechos, {ext_total:.0f}m", "OK")

        # 2c. CARREGAR RUAS DO GPKG (CARTOGRAFIA)
        if gpkg_name:
            gpkg_path = INPUT_DIR / gpkg_name
            if gpkg_path.exists():
                n_ruas = carregar_ruas_gpkg(str(gpkg_path), trechos, pvs)
                log(f"  Ruas GPKG: {n_ruas} trechos com rua via {gpkg_name}", "OK")
            else:
                log(f"  GPKG não encontrado: {gpkg_name}", "WARN")

        # 2d. MATCH EXECUTADO + CÁLCULO QUANTITATIVOS
        n_exec = 0; n_pend = 0; n_cad = 0
        for t in trechos:
            rua = limpar_rua(t.get("rua"))
            t["rua"] = rua

            is_cad, motivo_cad = e_cadastro(t, nucleo)
            quant = calc_quantitativos(t, pvs)

            if is_cad:
                status = "🔘 CADASTRO"
                n_cad += 1
                stats["cad"] += 1
                stats["ext_c"] += quant["ext_m"]
                is_exec = False
                match = None
            else:
                is_exec, match = match_shp(t, pvs, linhas_exec)
                if is_exec:
                    status = "✅ EXECUTADO"
                    n_exec += 1
                    stats["exec"] += 1
                    stats["ext_e"] += quant["ext_m"]
                else:
                    status = "⏳ PENDENTE"
                    n_pend += 1
                    stats["pend"] += 1
                    stats["ext_p"] += quant["ext_m"]

            pvi = pvs.get(t.get("pv_ini"), {})
            pvf = pvs.get(t.get("pv_fim"), {})
            hidr = calc_manning(t.get("dn_mm"), t.get("decl_mm"))

            row = {
                "Núcleo": nucleo,
                "Tipo": tipo,
                "Rua": rua,
                "PV Montante": t.get("pv_ini", ""),
                "PV Jusante": t.get("pv_fim", ""),
                "DN (mm)": t.get("dn_mm"),
                "Extensão (m)": quant["ext_m"],
                "Material": t.get("material", "PVC"),
                "CT Mont.": pvi.get("ct"),
                "CF Mont.": pvi.get("cf"),
                "CT Jus.": pvf.get("ct"),
                "CF Jus.": pvf.get("cf"),
                "Prof. Média (m)": quant["prof_media"],
                "Larg. Vala (m)": quant["largura_vala"],
                "Decl (‰)": round((t.get("decl_mm") or 0) * 1000, 2) if t.get("decl_mm") else None,
                "V (m/s)": hidr.get("v_ms"),
                "Q (l/s)": hidr.get("q_ls"),
                "STATUS": status,
                "Data Exec": match.get("data", "") if match else "",
                # PROJETADO
                "Escavação PROJ (m³)": quant["vol_escavacao"],
                "Reaterro PROJ (m³)": quant["vol_reaterro"],
                "Areia PROJ (m³)": quant["vol_areia"],
                "Brita PROJ (m³)": quant["vol_brita"],
                "Paviment. PROJ (m²)": quant["area_pav"],
                "Tubos PROJ (barras)": quant["n_barras"],
                # EXECUTADO
                "Escavação EXEC (m³)": quant["vol_escavacao"] if is_exec else 0,
                "Reaterro EXEC (m³)": quant["vol_reaterro"] if is_exec else 0,
                "Areia EXEC (m³)": quant["vol_areia"] if is_exec else 0,
                "Brita EXEC (m³)": quant["vol_brita"] if is_exec else 0,
                "Paviment. EXEC (m²)": quant["area_pav"] if is_exec else 0,
                "Tubos EXEC (barras)": quant["n_barras"] if is_exec else 0,
                # SALDO
                "Escavação SALDO (m³)": 0 if is_exec else quant["vol_escavacao"],
                "Reaterro SALDO (m³)": 0 if is_exec else quant["vol_reaterro"],
                "Areia SALDO (m³)": 0 if is_exec else quant["vol_areia"],
                "Brita SALDO (m³)": 0 if is_exec else quant["vol_brita"],
                "Paviment. SALDO (m²)": 0 if is_exec else quant["area_pav"],
                "Tubos SALDO (barras)": 0 if is_exec else quant["n_barras"],
                # Cadastro flag
                "_is_cad": is_cad,
            }
            all_rows.append(row)

        log(f"  {nucleo} ({tipo}): {n_exec} exec + {n_pend} pend + {n_cad} cad", "OK")

    # ── 3. GERAR PLANILHA ────────────────────────────────────────────────────
    log("Gerando planilha final...", "STEP")

    wb = openpyxl.Workbook()

    # Estilos
    hf = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    hfl = PatternFill(start_color="1b5e20", end_color="1b5e20", fill_type="solid")
    hfl2 = PatternFill(start_color="0d47a1", end_color="0d47a1", fill_type="solid")
    hfl3 = PatternFill(start_color="b71c1c", end_color="b71c1c", fill_type="solid")
    df_font = Font(name="Calibri", size=9)
    brd = Border(left=Side("thin", "cccccc"), right=Side("thin", "cccccc"),
                 top=Side("thin", "cccccc"), bottom=Side("thin", "cccccc"))
    exec_f = PatternFill(start_color="c8e6c9", end_color="c8e6c9", fill_type="solid")
    pend_f = PatternFill(start_color="ffcdd2", end_color="ffcdd2", fill_type="solid")
    cad_f  = PatternFill(start_color="e0e0e0", end_color="e0e0e0", fill_type="solid")

    # ── ABA 1: POR TRECHO ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Por Trecho"

    ws1.merge_cells("A1:AK1")
    ws1["A1"] = "QUANTITATIVOS GLOBAL: PROJETADO × EXECUTADO × SALDO — PV a PV"
    ws1["A1"].font = Font(bold=True, size=14, color="1b5e20")
    ws1.merge_cells("A2:AK2")
    ws1["A2"] = f"CT {CONTRATO} | {datetime.now().strftime('%d/%m/%Y %H:%M')} | Motor ConstruData HydroNetwork v9"
    ws1["A2"].font = Font(size=9, color="666666")

    heads_base = ["Núcleo", "Tipo", "Rua", "PV Montante", "PV Jusante", "DN (mm)",
                  "Extensão (m)", "Material", "CT Mont.", "CF Mont.", "CT Jus.", "CF Jus.",
                  "Prof. Média (m)", "Larg. Vala (m)", "Decl (‰)", "V (m/s)", "Q (l/s)", "STATUS", "Data Exec"]
    heads_proj = ["Escavação PROJ (m³)", "Reaterro PROJ (m³)", "Areia PROJ (m³)",
                  "Brita PROJ (m³)", "Paviment. PROJ (m²)", "Tubos PROJ (barras)"]
    heads_exec = ["Escavação EXEC (m³)", "Reaterro EXEC (m³)", "Areia EXEC (m³)",
                  "Brita EXEC (m³)", "Paviment. EXEC (m²)", "Tubos EXEC (barras)"]
    heads_saldo = ["Escavação SALDO (m³)", "Reaterro SALDO (m³)", "Areia SALDO (m³)",
                   "Brita SALDO (m³)", "Paviment. SALDO (m²)", "Tubos SALDO (barras)"]
    all_heads = heads_base + heads_proj + heads_exec + heads_saldo

    for c, h in enumerate(all_heads, 1):
        cell = ws1.cell(row=3, column=c, value=h)
        cell.font = hf
        cell.border = brd
        # Cor por grupo
        if h in heads_proj:
            cell.fill = hfl   # verde
        elif h in heads_exec:
            cell.fill = hfl2  # azul
        elif h in heads_saldo:
            cell.fill = hfl3  # vermelho
        else:
            cell.fill = PatternFill(start_color="37474f", end_color="37474f", fill_type="solid")

    row_num = 4
    for r in all_rows:
        is_cad = r.pop("_is_cad", False)
        for c, h in enumerate(all_heads, 1):
            cell = ws1.cell(row=row_num, column=c, value=r.get(h))
            cell.font = df_font
            cell.border = brd
            if is_cad:
                cell.fill = cad_f
            elif "EXECUTADO" in str(r.get("STATUS", "")):
                cell.fill = exec_f
            elif "PENDENTE" in str(r.get("STATUS", "")):
                cell.fill = pend_f
        row_num += 1

    for c in range(1, len(all_heads) + 1):
        ws1.column_dimensions[get_column_letter(c)].width = 14
    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["C"].width = 28
    ws1.column_dimensions["R"].width = 16

    # ── ABA 2: POR RUA ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Por Rua")
    rua_agg = defaultdict(lambda: {
        "nucleo": "", "tipo": "", "n_trechos": 0, "ext_proj": 0, "ext_exec": 0,
        "esc_proj": 0, "esc_exec": 0, "reat_proj": 0, "reat_exec": 0,
        "areia_proj": 0, "areia_exec": 0, "brita_proj": 0, "brita_exec": 0,
        "pav_proj": 0, "pav_exec": 0, "tubos_proj": 0, "tubos_exec": 0,
    })
    for r in all_rows:
        key = (r.get("Núcleo", ""), r.get("Rua", "Sem Rua"), r.get("Tipo", ""))
        a = rua_agg[key]
        a["nucleo"] = r.get("Núcleo", "")
        a["tipo"] = r.get("Tipo", "")
        a["n_trechos"] += 1
        a["ext_proj"] += r.get("Extensão (m)", 0)
        a["ext_exec"] += r.get("Extensão (m)", 0) if "EXECUTADO" in str(r.get("STATUS", "")) else 0
        for campo, col_proj, col_exec in [
            ("esc", "Escavação PROJ (m³)", "Escavação EXEC (m³)"),
            ("reat", "Reaterro PROJ (m³)", "Reaterro EXEC (m³)"),
            ("areia", "Areia PROJ (m³)", "Areia EXEC (m³)"),
            ("brita", "Brita PROJ (m³)", "Brita EXEC (m³)"),
            ("pav", "Paviment. PROJ (m²)", "Paviment. EXEC (m²)"),
            ("tubos", "Tubos PROJ (barras)", "Tubos EXEC (barras)"),
        ]:
            a[f"{campo}_proj"] += r.get(col_proj, 0) or 0
            a[f"{campo}_exec"] += r.get(col_exec, 0) or 0

    heads_rua = ["Núcleo", "Tipo", "Rua", "Trechos", "Ext PROJ (m)", "Ext EXEC (m)", "Ext SALDO (m)", "% Exec",
                 "Esc PROJ (m³)", "Esc EXEC (m³)", "Esc SALDO (m³)",
                 "Reat PROJ (m³)", "Reat EXEC (m³)", "Reat SALDO (m³)",
                 "Areia PROJ (m³)", "Areia EXEC (m³)", "Areia SALDO (m³)",
                 "Brita PROJ (m³)", "Brita EXEC (m³)", "Brita SALDO (m³)",
                 "Pav PROJ (m²)", "Pav EXEC (m²)", "Pav SALDO (m²)",
                 "Tubos PROJ", "Tubos EXEC", "Tubos SALDO"]

    for c, h in enumerate(heads_rua, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = hf
        cell.fill = hfl
        cell.border = brd

    row2 = 2
    for (nucleo, rua, tipo), a in sorted(rua_agg.items()):
        ext_saldo = a["ext_proj"] - a["ext_exec"]
        pct = (a["ext_exec"] / a["ext_proj"] * 100) if a["ext_proj"] > 0 else 0
        values = [
            nucleo, tipo, rua, a["n_trechos"],
            round(a["ext_proj"], 1), round(a["ext_exec"], 1), round(ext_saldo, 1), f"{pct:.0f}%",
            round(a["esc_proj"], 1), round(a["esc_exec"], 1), round(a["esc_proj"] - a["esc_exec"], 1),
            round(a["reat_proj"], 1), round(a["reat_exec"], 1), round(a["reat_proj"] - a["reat_exec"], 1),
            round(a["areia_proj"], 1), round(a["areia_exec"], 1), round(a["areia_proj"] - a["areia_exec"], 1),
            round(a["brita_proj"], 1), round(a["brita_exec"], 1), round(a["brita_proj"] - a["brita_exec"], 1),
            round(a["pav_proj"], 1), round(a["pav_exec"], 1), round(a["pav_proj"] - a["pav_exec"], 1),
            a["tubos_proj"], a["tubos_exec"], a["tubos_proj"] - a["tubos_exec"],
        ]
        for c, v in enumerate(values, 1):
            cell = ws2.cell(row=row2, column=c, value=v)
            cell.font = df_font
            cell.border = brd
            if pct >= 100:
                cell.fill = exec_f
            elif pct > 0:
                cell.fill = PatternFill(start_color="fff9c4", end_color="fff9c4", fill_type="solid")
            else:
                cell.fill = pend_f
        row2 += 1

    for c in range(1, len(heads_rua) + 1):
        ws2.column_dimensions[get_column_letter(c)].width = 14
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["C"].width = 28

    # ── ABA 3: POR NÚCLEO ────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Por Núcleo")
    nuc_agg = defaultdict(lambda: {
        "tipo": "", "n_trechos": 0, "n_exec": 0, "n_pend": 0, "n_cad": 0,
        "ext_proj": 0, "ext_exec": 0, "ext_cad": 0,
        "esc_proj": 0, "esc_exec": 0, "reat_proj": 0, "reat_exec": 0,
        "areia_proj": 0, "areia_exec": 0, "brita_proj": 0, "brita_exec": 0,
        "pav_proj": 0, "pav_exec": 0, "tubos_proj": 0, "tubos_exec": 0,
    })
    for r in all_rows:
        key = (r.get("Núcleo", ""), r.get("Tipo", ""))
        a = nuc_agg[key]
        a["tipo"] = r.get("Tipo", "")
        a["n_trechos"] += 1
        st = str(r.get("STATUS", ""))
        if "EXECUTADO" in st:
            a["n_exec"] += 1
            a["ext_exec"] += r.get("Extensão (m)", 0)
        elif "CADASTRO" in st:
            a["n_cad"] += 1
            a["ext_cad"] += r.get("Extensão (m)", 0)
        else:
            a["n_pend"] += 1
        a["ext_proj"] += r.get("Extensão (m)", 0)
        for campo, col_proj, col_exec in [
            ("esc", "Escavação PROJ (m³)", "Escavação EXEC (m³)"),
            ("reat", "Reaterro PROJ (m³)", "Reaterro EXEC (m³)"),
            ("areia", "Areia PROJ (m³)", "Areia EXEC (m³)"),
            ("brita", "Brita PROJ (m³)", "Brita EXEC (m³)"),
            ("pav", "Paviment. PROJ (m²)", "Paviment. EXEC (m²)"),
            ("tubos", "Tubos PROJ (barras)", "Tubos EXEC (barras)"),
        ]:
            a[f"{campo}_proj"] += r.get(col_proj, 0) or 0
            a[f"{campo}_exec"] += r.get(col_exec, 0) or 0

    heads3 = ["Núcleo", "Tipo", "Trechos", "Exec", "Pend", "Cad",
              "km PROJ", "km EXEC", "km SALDO", "km Cad", "% Exec",
              "Esc PROJ (m³)", "Esc EXEC (m³)", "Esc SALDO (m³)",
              "Reat PROJ (m³)", "Reat EXEC (m³)", "Reat SALDO (m³)",
              "Pav PROJ (m²)", "Pav EXEC (m²)", "Pav SALDO (m²)"]

    for c, h in enumerate(heads3, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = hf
        cell.fill = hfl2
        cell.border = brd

    r3 = 2
    for (nucleo, tipo), a in sorted(nuc_agg.items()):
        ext_obra = a["ext_exec"] + (a["ext_proj"] - a["ext_exec"] - a["ext_cad"])
        pct = (a["ext_exec"] / ext_obra * 100) if ext_obra > 0 else 0
        values = [
            nucleo, tipo, a["n_trechos"], a["n_exec"], a["n_pend"], a["n_cad"],
            round(a["ext_proj"]/1000, 2), round(a["ext_exec"]/1000, 2),
            round((a["ext_proj"] - a["ext_exec"] - a["ext_cad"])/1000, 2),
            round(a["ext_cad"]/1000, 2), f"{pct:.0f}%",
            round(a["esc_proj"], 1), round(a["esc_exec"], 1), round(a["esc_proj"] - a["esc_exec"], 1),
            round(a["reat_proj"], 1), round(a["reat_exec"], 1), round(a["reat_proj"] - a["reat_exec"], 1),
            round(a["pav_proj"], 1), round(a["pav_exec"], 1), round(a["pav_proj"] - a["pav_exec"], 1),
        ]
        for c, v in enumerate(values, 1):
            cell = ws3.cell(row=r3, column=c, value=v)
            cell.font = df_font
            cell.border = brd
            if pct >= 80:
                cell.fill = exec_f
            elif pct >= 40:
                cell.fill = PatternFill(start_color="fff9c4", end_color="fff9c4", fill_type="solid")
            else:
                cell.fill = pend_f
        r3 += 1

    for c in range(1, len(heads3) + 1):
        ws3.column_dimensions[get_column_letter(c)].width = 14
    ws3.column_dimensions["A"].width = 25

    # ── ABA 4: RESUMO GERAL ─────────────────────────────────────────────────
    ws4 = wb.create_sheet("Resumo Geral")
    obra_total = stats["ext_e"] + stats["ext_p"]
    pct_geral = stats["ext_e"] / obra_total * 100 if obra_total else 0

    ws4["A1"] = "RESUMO GERAL — QUANTITATIVOS CT 11481051"
    ws4["A1"].font = Font(bold=True, size=16, color="1b5e20")
    ws4["A3"] = f"✅ Executados: {stats['exec']} trechos ({stats['ext_e']:.0f}m = {stats['ext_e']/1000:.1f}km)"
    ws4["A3"].font = Font(bold=True, size=12, color="1b5e20")
    ws4["A4"] = f"⏳ Pendentes: {stats['pend']} trechos ({stats['ext_p']:.0f}m = {stats['ext_p']/1000:.1f}km)"
    ws4["A4"].font = Font(bold=True, size=12, color="c62828")
    ws4["A5"] = f"🔘 Cadastro: {stats['cad']} trechos ({stats['ext_c']:.0f}m) — EXCLUÍDO DO CÔMPUTO"
    ws4["A5"].font = Font(size=11, italic=True, color="757575")
    ws4["A7"] = f"📊 PROGRESSO REAL: {pct_geral:.1f}% ({stats['ext_e']:.0f}m de {obra_total:.0f}m = {obra_total/1000:.1f}km)"
    ws4["A7"].font = Font(bold=True, size=14, color="0d47a1")

    # Totais globais de volumes
    totais = {"esc_p": 0, "esc_e": 0, "reat_p": 0, "reat_e": 0,
              "areia_p": 0, "areia_e": 0, "brita_p": 0, "brita_e": 0,
              "pav_p": 0, "pav_e": 0}
    for r in all_rows:
        totais["esc_p"] += r.get("Escavação PROJ (m³)", 0) or 0
        totais["esc_e"] += r.get("Escavação EXEC (m³)", 0) or 0
        totais["reat_p"] += r.get("Reaterro PROJ (m³)", 0) or 0
        totais["reat_e"] += r.get("Reaterro EXEC (m³)", 0) or 0
        totais["areia_p"] += r.get("Areia PROJ (m³)", 0) or 0
        totais["areia_e"] += r.get("Areia EXEC (m³)", 0) or 0
        totais["brita_p"] += r.get("Brita PROJ (m³)", 0) or 0
        totais["brita_e"] += r.get("Brita EXEC (m³)", 0) or 0
        totais["pav_p"] += r.get("Paviment. PROJ (m²)", 0) or 0
        totais["pav_e"] += r.get("Paviment. EXEC (m²)", 0) or 0

    ws4["A9"] = "VOLUMES GLOBAIS"
    ws4["A9"].font = Font(bold=True, size=13, color="37474f")
    vol_heads = ["Grandeza", "PROJETADO", "EXECUTADO", "SALDO"]
    for c, h in enumerate(vol_heads, 1):
        cell = ws4.cell(row=10, column=c, value=h)
        cell.font = hf
        cell.fill = hfl2
        cell.border = brd

    vol_rows = [
        ("Escavação (m³)", totais["esc_p"], totais["esc_e"]),
        ("Reaterro (m³)", totais["reat_p"], totais["reat_e"]),
        ("Areia lastro+envoltória (m³)", totais["areia_p"], totais["areia_e"]),
        ("Brita dreno (m³)", totais["brita_p"], totais["brita_e"]),
        ("Pavimentação CBUQ (m²)", totais["pav_p"], totais["pav_e"]),
    ]
    for i, (nome, proj, exe) in enumerate(vol_rows):
        r = 11 + i
        saldo = proj - exe
        for c, v in enumerate([nome, round(proj, 1), round(exe, 1), round(saldo, 1)], 1):
            cell = ws4.cell(row=r, column=c, value=v)
            cell.font = df_font
            cell.border = brd

    ws4.column_dimensions["A"].width = 35
    ws4.column_dimensions["B"].width = 16
    ws4.column_dimensions["C"].width = 16
    ws4.column_dimensions["D"].width = 16

    # ── SALVAR ───────────────────────────────────────────────────────────────
    wb.save(str(OUTPUT_FILE))
    log(f"Planilha salva: {OUTPUT_FILE}", "OK")

    print(f"\n{'=' * 72}")
    print(f"  RESULTADO FINAL")
    print(f"{'=' * 72}")
    print(f"  ✅ Executados: {stats['exec']} trechos ({stats['ext_e']:.0f}m = {stats['ext_e']/1000:.1f}km)")
    print(f"  ⏳ Pendentes:  {stats['pend']} trechos ({stats['ext_p']:.0f}m = {stats['ext_p']/1000:.1f}km)")
    print(f"  🔘 Cadastro:   {stats['cad']} trechos ({stats['ext_c']:.0f}m) — EXCLUÍDO")
    print(f"  📊 PROGRESSO: {pct_geral:.1f}% (obra: {obra_total:.0f}m = {obra_total/1000:.1f}km)")
    print(f"  📁 Saída: {OUTPUT_FILE}")
    print(f"{'=' * 72}")


if __name__ == "__main__":
    main()
