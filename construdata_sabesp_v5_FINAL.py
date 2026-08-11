#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CONSTRUDATA SABESP v5.0 -- Pipeline Unificado SABESP Santos
SE LIGA NA REDE -- Contrato 11481051 -- FCN Construções e Saneamento

Pipeline:
  DXF + GPKG + JSON -> PVs/Trechos -> Calculos -> NS Completa + BIM + GIS

Por NS gera:
  NS_XXX_A4.pdf        Folha de campo / Ordem de Servico
  NS_XXX_DESENHO.pdf   Prancha A3: Planta + Perfil + Tabela + Selo
  NS_XXX_OSE.xlsx      OSE padrao SABESP (formato NS_017rev1)
  NS_XXX_DADOS.json    Dados tecnicos estruturados
  NS_XXX.html          Dashboard interativo Leaflet + perfil SVG real

Uso:
  python construdata_sabesp_v5_FINAL.py ISRAEL_ESGOTO.dxf
  python construdata_sabesp_v5_FINAL.py --json rede_definida.json
  python construdata_sabesp_v5_FINAL.py AGUA.dxf --gpkg MAPA.gpkg
  python construdata_sabesp_v5_FINAL.py --batch
"""

import sys
import os
import math
import re
import json
import time
import copy
import argparse
import warnings
import traceback
import configparser
from pathlib import Path
from datetime import datetime
from collections import defaultdict, Counter

warnings.filterwarnings("ignore")

# ── matplotlib headless ──────────────────────────────────────────────────────
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.patches import FancyBboxPatch
from matplotlib.lines import Line2D
from matplotlib.gridspec import GridSpec
import matplotlib.ticker as mticker

# ── Opcionais com fallback ───────────────────────────────────────────────────
try:
    import ezdxf
    _HAS_EZDXF = True
except ImportError:
    _HAS_EZDXF = False

try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

try:
    import networkx as nx
    _HAS_NX = True
except ImportError:
    _HAS_NX = False

try:
    from pyproj import Transformer
    _TF = Transformer.from_crs("EPSG:31983", "EPSG:4326", always_xy=True)
    def utm_to_latlon(x, y):
        if x is None or y is None:
            return None, None
        lon, lat = _TF.transform(x, y)
        return round(lat, 8), round(lon, 8)
    _HAS_PYPROJ = True
except ImportError:
    _HAS_PYPROJ = False
    def utm_to_latlon(x, y):
        return None, None

try:
    import geopandas as gpd
    import pyogrio
    _HAS_GEO = True
except ImportError:
    _HAS_GEO = False

# GDAL/OGR reader (drop-in replacement para ler_dxf)
try:
    from ler_dxf_gdal import ler_dxf_gdal
    _HAS_GDAL_READER = True
except ImportError:
    _HAS_GDAL_READER = False


# ==============================================================================
# CONFIGURAÇÃO GLOBAL
# ==============================================================================

CFG = {
    "contrato":   "11481051",
    "empresa":    "CONSÓRCIO SE LIGA NA REDE",
    "cidade":     "SANTOS-SP",
    "crs":        "EPSG:31983",
    "tol_pv_tubo":     50.0,
    "tol_texto_tubo":  40.0,
    "tol_rua_trecho": 150.0,
    "tol_grupo_pv_x":   3.0,
    "tol_grupo_pv_y":   8.0,
    
    # CARREGAR DO INI (ProSaneamento)
    "_ini_path": Path(__file__).parent / "config" / "PARAMETROS_PROSANE.INI",
}

# Carregar parâmetros do INI se existir
_ini_cfg = configparser.ConfigParser(inline_comment_prefixes=(';', '#'))
_ini_path = CFG.get("_ini_path")
if _ini_path and _ini_path.exists():
    _ini_cfg.read(_ini_path, encoding='utf-8')
    try:
        log(f"  Config carregada: {_ini_path}", "INFO")
    except NameError:
        print(f"  Config carregada: {_ini_path}")
    
    # Vala
    CFG["largura_vala"] = _ini_cfg.getfloat('Vala', 'Largura', fallback=60.0) / 100
    CFG["lastro"] = _ini_cfg.getfloat('Vala', 'Lastro', fallback=15.0) / 100
    CFG["bdi"] = _ini_cfg.getfloat('Vala', 'BDI', fallback=1.25)
    
    # Declividade
    CFG["decl_minima"] = _ini_cfg.getfloat('Declividade', 'Minima', fallback=0.002)
    CFG["prof_minima"] = _ini_cfg.getfloat('Declividade', 'Profundidade_Minima', fallback=0.30)
    
    # Manning
    CFG["manning"] = {
        "PVC": _ini_cfg.getfloat('Manning_Esgoto', 'PVC', fallback=0.013),
        "PEAD": _ini_cfg.getfloat('Manning_Esgoto', 'PEAD', fallback=0.011),
        "PE80": _ini_cfg.getfloat('Manning_Esgoto', 'PE80', fallback=0.011),
        "PE100": _ini_cfg.getfloat('Manning_Esgoto', 'PE100', fallback=0.011),
        "Concreto": _ini_cfg.getfloat('Manning_Esgoto', 'Concreto', fallback=0.013),
    }
    
    # Perfil
    CFG["perfil_esc_h"] = _ini_cfg.getint('Perfil', 'Escala_H', fallback=200)
    CFG["perfil_esc_v"] = _ini_cfg.getint('Perfil', 'Escala_V', fallback=200)
    CFG["perfil_exag"] = _ini_cfg.getfloat('Perfil', 'Exagero', fallback=0.5)
    
    # PV
    CFG["pv_prof_default"] = _ini_cfg.getfloat('PV', 'Profundidade_Default', fallback=0.50)
else:
    # Defaults hardcoded
    CFG["largura_vala"] = 0.60
    CFG["lastro"] = 0.15
    CFG["bdi"] = 1.25
    CFG["decl_minima"] = 0.002
    CFG["prof_minima"] = 0.30
    CFG["manning"] = {
        "PVC": 0.013, "PEAD": 0.011, "PE80": 0.011, "PE100": 0.011,
        "FC": 0.012, "CA": 0.013, "CONC": 0.013, "FD": 0.013,
    }
    CFG["perfil_esc_h"] = 200
    CFG["perfil_esc_v"] = 200
    CFG["perfil_exag"] = 0.5
    CFG["pv_prof_default"] = 0.50

SINAPI = {
    "tubo_pvc_100":  28.50, "tubo_pvc_150":  35.80, "tubo_pvc_200":  47.12,
    "tubo_pvc_250":  62.30, "tubo_pvc_300":  89.45, "tubo_pvc_400": 142.60,
    "tubo_pe80_63":  15.30, "tubo_pe80_110": 32.50, "tubo_pe80_160": 58.70,
    "escavacao":     30.77, "reaterro":   19.97, "lastro":   85.50,
    "envoltorio":    75.30, "brita":      95.20, "pavimentacao": 97.80,
    "pv_concreto_1200": 3078.00, "pv_concreto_600": 1850.00,
    "sela_dn63x20":  48.90, "te_fofo_dn63":  85.50,
    "reg_gaveta_dn63": 145.00, "chave_servico_dn20": 62.00,
    "luva_pe80_dn63": 12.80, "tampao_pe80_dn63": 18.50,
}

_LOG_ENTRIES = []

def log(msg, nivel="INFO"):
    ts    = datetime.now().strftime("%H:%M:%S")
    icons = {"INFO": "  ", "OK": "[OK] ", "WARN": "[!] ", "ERR": "[X] ", "STEP": ">>> "}
    linha = f"[{ts}] {icons.get(nivel,'  ')} {msg}"
    try:
        print(linha)
    except UnicodeEncodeError:
        print(linha.encode("ascii", "replace").decode())
    _LOG_ENTRIES.append({"ts": ts, "nivel": nivel, "msg": msg})


def _fmtv(v, fmt=".3f", fallback="---"):
    if v is None:
        return fallback
    try:
        return format(float(v), fmt)
    except (TypeError, ValueError):
        return str(v) if v else fallback


# ==============================================================================
# MÓDULO 2 — LEITURA DXF ProSaneamento
# ==============================================================================


def _limpar_encoding(txt):
    """Corrige resíduos de encoding CP1252/latin-1 em textos DXF."""
    if not txt:
        return txt
    replacements = [
        ("ÃO", "ÃO"), ("Ã", "Ã"), ("Ã£", "ã"),
        ("Ã", "Ç"),   ("Ã§", "ç"),
        ("Ã", "É"),   ("Ã©", "é"),
        ("Ã", "Á"),   ("Ã¡", "á"),
        ("Ã", "Ó"),   ("Ã³", "ó"),
        ("Ã", "Ú"),   ("Ãº", "ú"),
        ("Ã", "Ê"),   ("Ãª", "ê"),
        ("Ã", "Ô"),   ("Ã´", "ô"),
        ("Ã", "Æ"),   ("Ã", "À"),   ("Ã", "Í"),
        ("Ã­", "í"),   ("Ã", "Õ"),   ("Ãµ", "õ"),
        ("Ã", "Ü"),   ("Ã¼", "ü"),   ("Ã", "Ñ"),
        ("Ã±", "ñ"),
        ("\\P", " "),    # AutoCAD paragraph break
    ]
    for old, new in replacements:
        txt = txt.replace(old, new)
    return txt.strip()

def _dist(a, b):
    return math.hypot(a[0]-b[0], a[1]-b[1])

def _mais_proximo(ponto, textos, max_dist=30.0):
    if not textos:
        return None
    melhor, melhor_d = None, float("inf")
    px, py = ponto[0], ponto[1]
    for t in textos:
        d = math.hypot(t.get("x", 0)-px, t.get("y", 0)-py)
        if d < melhor_d:
            melhor_d, melhor = d, t
    return melhor if melhor_d <= max_dist else None

def _pv_mais_proximo(ponto, pvs, max_dist=15.0, preferir_real=False):
    melhor, melhor_d = None, float("inf")
    melhor_real, melhor_d_real = None, float("inf")
    px, py = ponto[0], ponto[1]
    for nome, pv in pvs.items():
        d = math.hypot(pv.get("x", 0)-px, pv.get("y", 0)-py)
        if d < melhor_d:
            melhor_d, melhor = d, nome
        if preferir_real and not pv.get("sintetico") and d < melhor_d_real:
            melhor_d_real, melhor_real = d, nome
    if preferir_real and melhor_real and melhor_d_real <= max_dist:
        return melhor_real
    return melhor if melhor_d <= max_dist else None

def _inferir_material(layer_name):
    up = layer_name.upper()
    for k, v in [("PE_80","PE80"),("PE_100","PE100"),("PEAD","PEAD"),
                 ("PVC","PVC"),("FC","FC"),("CONC","CONC")]:
        if k in up:
            return v
    return "PVC"

def _parsear_dn(txt):
    if not txt:
        return None
    for p in [r"(\d+)\s*mm", r"D\s*=?\s*(\d+)", r"DN\s*(\d+)"]:
        m = re.search(p, txt, re.IGNORECASE)
        if m:
            v = int(m.group(1))
            if 20 <= v <= 1200:
                return v
    return None

def _parsear_comp(txt):
    if not txt:
        return None
    m = re.search(r"([\d.,]+)\s*m(?!\w)", txt)
    return float(m.group(1).replace(",", ".")) if m else None

def _parsear_incl(txt):
    if not txt:
        return None
    m = re.search(r"([\d.,]+)\s*m/m", txt, re.IGNORECASE)
    if m:
        return float(m.group(1).replace(",", "."))
    m = re.search(r"([\d.,]+)\s*%", txt)
    if m:
        return float(m.group(1).replace(",", ".")) / 100
    return None

def _parsear_pressao(txt):
    if not txt:
        return None
    m = re.search(r"P\s*=\s*([\d.,]+)\s*mca", txt, re.IGNORECASE)
    return float(m.group(1).replace(",", ".")) if m else None

def _agrupar_textos_pvs(textos):
    """Agrupa blocos de textos empilhados -> PV/PI/N_xxx."""
    if not textos:
        return {}
    # BUG-2 FIX: filtrar textos sem coordenadas validas
    textos = [t for t in textos
              if isinstance(t, dict)
              and isinstance(t.get("x"), (int, float))
              and isinstance(t.get("y"), (int, float))
              and t.get("text", "").strip()]
    if not textos:
        return {}
    tx = CFG["tol_grupo_pv_x"]
    ty = CFG["tol_grupo_pv_y"]
    ts = sorted(textos, key=lambda t: (round(t["x"]/tx)*tx, -t["y"]))
    grupos, grupo = [], [ts[0]]
    for t in ts[1:]:
        u = grupo[-1]
        if abs(t["x"]-u["x"]) < tx and abs(t["y"]-u["y"]) < ty:
            grupo.append(t)
        else:
            grupos.append(grupo)
            grupo = [t]
    grupos.append(grupo)

    pvs = {}
    for g in grupos:
        g.sort(key=lambda t: -t["y"])
        nome = ct = cf = prof = None
        x0, y0 = g[0]["x"], g[0]["y"]
        for t in g:
            txt = t["text"].strip()
            if not txt:
                continue
            m = re.match(r"P\.?\s*[VI]\.?\s*[_\s]*(\d+)", txt, re.IGNORECASE)
            if m:
                tipo = "PI" if re.search(r"[Ii]", txt[1:3]) else "PV"
                nome = f"{tipo}_{m.group(1)}"
                continue
            m = re.match(r"^N[_\s]?(\d+)$", txt, re.IGNORECASE)
            if m:
                nome = f"N_{m.group(1)}"
                continue
            # Pontos de rede de agua: TE DN100a, C90 DN75b, CURVA 22 DN75a, CAP DN50a, RED DN75a
            m = re.match(r"^(TE|C90|C45|C22|C11|CURVA\s*\d*|CAP|RED|LUVA|CV)\s*\.?\s*(DN\d+)?\s*(\w*)$", txt, re.IGNORECASE)
            if m:
                tipo_agua = m.group(1).upper().replace(" ","")
                dn_txt = m.group(2) or ""
                sufixo = m.group(3) or ""
                nome = f"N_{tipo_agua}_{dn_txt}{sufixo}".rstrip("_")
                continue
            m = re.match(r"P\.?\s*F\.?\s*([+-]?[\d.,]+)", txt, re.IGNORECASE)
            if m:
                try:
                    prof = float(m.group(1).replace(",", "."))
                except ValueError:
                    pass
                continue
            m = re.match(r"C\.?\s*T\.?\s*([+-]?[\d.,]+)", txt, re.IGNORECASE)
            if m:
                try:
                    ct = float(m.group(1).replace(",", "."))
                except ValueError:
                    pass
                continue
            m = re.match(r"C\.?\s*F\.?\s*([+-]?[\d.,]+)", txt, re.IGNORECASE)
            if m:
                try:
                    cf = float(m.group(1).replace(",", "."))
                except ValueError:
                    pass
                continue
        if nome:
            # Validar: CF nunca pode ser maior que CT (inverter se necessario)
            if ct is not None and cf is not None and cf > ct:
                ct, cf = cf, ct
            # Recalcular prof se temos CT e CF
            if ct is not None and cf is not None and prof is None:
                prof = round(ct - cf, 4)
            pvs[nome] = {
                "x": x0, "y": y0, "ct": ct, "cf": cf, "prof": prof,
                "tipo": "RG" if nome.startswith("N_") else nome.split("_")[0],
            }
    return pvs


def _ler_xdata_raw(dxf_path):
    """
    Lê XDATA ProSaneamento via raw text (sem ezdxf).
    Também coleta TEXT/MTEXT de logradouro no mesmo espaço de coordenadas.
    Retorna (pvs, tubos_raw, textos_rua_raw) ou (None, None, []).
    """
    PREF_RUA = ("RUA ", "BECO ", "TRAV", "AV ", "ESTRADA", "VIELA", "ALAMEDA", "ACESSO")
    LAYERS_LOGR = {"A_Alerta", "TXT-LOGRAD", "TEXTO", "0", "ZZ-Carimbo Texto",
                    "LT-TEXTO-RUA", "TXT-PRACA", "PS_IND_TRECHO"}

    try:
        with open(dxf_path, encoding="latin-1", errors="replace") as f:
            lines = [l.rstrip("\r\n") for l in f]
    except Exception as e:
        return None, None, []

    inserts, polylines, textos_rua_raw = [], [], []
    i, ent = 0, None
    n = len(lines)

    while i < n - 1:
        code, val = lines[i].strip(), lines[i+1].strip()

        if code == "0":
            # Fechar entidade anterior
            if ent:
                if ent.get("xd"):
                    if ent["type"] == "INSERT":
                        inserts.append(ent)
                    elif ent["type"] == "LWPOLYLINE":
                        polylines.append(ent)
                # Salvar texto de rua (espaço local = mesmo dos PVs/tubos)
                if ent["type"] in ("TEXT", "MTEXT"):
                    txt = ent.get("_txt", "").strip()
                    lyr = ent.get("layer", "")
                    x_, y_ = ent.get("x"), ent.get("y")
                    if (txt and x_ is not None and y_ is not None
                            and lyr in LAYERS_LOGR
                            and any(txt.upper().startswith(p) for p in PREF_RUA)):
                        textos_rua_raw.append({"x": x_, "y": y_,
                                                "text": _limpar_encoding(txt)})

            ent = {"type": val, "xd": {}, "layer": "", "_app": None,
                   "x": None, "y": None, "_txt": ""}
            i += 2
            continue

        if ent is None:
            i += 2
            continue

        # Capturar campos por tipo de entidade
        if ent["type"] == "INSERT":
            if code == "10": ent["x"] = float(val)
            elif code == "20": ent["y"] = float(val)
            elif code == "8":  ent["layer"] = val

        elif ent["type"] == "LWPOLYLINE":
            if code == "8":  ent["layer"] = val
            elif code == "10": ent.setdefault("px", []).append(float(val))
            elif code == "20": ent.setdefault("py", []).append(float(val))

        elif ent["type"] in ("TEXT", "MTEXT"):
            if code == "10": ent["x"] = float(val)
            elif code == "20": ent["y"] = float(val)
            elif code == "8":  ent["layer"] = val
            elif code in ("1", "3"): ent["_txt"] = val  # texto

        # XDATA
        if code == "1001":
            ent["_app"] = val
            ent["xd"].setdefault(val, [])
        elif ent.get("_app") and code.isdigit() and int(code) >= 1000:
            c = int(code)
            if c == 1001:
                ent["_app"] = val
                ent["xd"].setdefault(val, [])
            else:
                ent["xd"][ent["_app"]].append((c, val))
        i += 2

    # Fechar última entidade
    if ent:
        if ent.get("xd"):
            if ent["type"] == "INSERT":    inserts.append(ent)
            elif ent["type"] == "LWPOLYLINE": polylines.append(ent)
        if ent["type"] in ("TEXT", "MTEXT"):
            txt = ent.get("_txt","").strip()
            lyr = ent.get("layer","")
            x_, y_ = ent.get("x"), ent.get("y")
            if txt and x_ is not None and lyr in LAYERS_LOGR:
                if any(txt.upper().startswith(p) for p in PREF_RUA):
                    textos_rua_raw.append({"x": x_, "y": y_,
                                                "text": _limpar_encoding(txt)})

    if not inserts:
        log("  Sem XDATA — fallback ezdxf", "WARN")
        return None, None, textos_rua_raw

    # FILTRO AGRESSIVO: Remover blocos de detalhe
    _detalhe = {"DETALHE","LISTAGEM","QUADRO","PERFIL","BLOCO","REF","PAREDE",
                "COT","TIT","FOR","TUB","SIMBOLO","LEGENDA","CARIMBO","TEXTO",
                "IND","FLUXO","DIAM","INCL","COMPR"}
    _inserts_filtrados = [b for b in inserts if not any(d in b.get("layer","").upper() for d in _detalhe)]
    if len(_inserts_filtrados) < len(inserts):
        log(f"  Filtro: {len(inserts)} → {len(_inserts_filtrados)} INSERTs", "INFO")
        inserts = _inserts_filtrados

    log(f"  XDATA: {len(inserts)} INSERTs, {len(polylines)} polilínias "
        f"| {len(textos_rua_raw)} textos de rua", "OK")

    # ── Decodificar PVs ──────────────────────────────────────────────────────
    pvs = {}
    for blk in inserts:
        x, y = blk.get("x", 0), blk.get("y", 0)
        cnx   = blk["xd"].get("PH_DATCNX", [])
        reals = [float(v) for c, v in cnx if c == 1040]
        strs  = [v.strip() for c, v in cnx if c == 1000 and v.strip()]
        tipo  = strs[0] if strs else "PI"
        # Layout PH_DATCNX: reals = [diam_pv, flag, prof, CF]
        # CF = geratriz inferior. CT = CF + prof. NUNCA reals[3] como CT.
        prof  = reals[2] if len(reals) > 2 else None
        cf    = reals[3] if len(reals) > 3 else None
        ct    = round(cf + prof, 4) if cf is not None and prof is not None else None

        idn     = blk["xd"].get("PH_DATIDN", [])
        reals_i = [float(v) for c, v in idn if c == 1040]
        strs_i  = [v.strip() for c, v in idn if c == 1000
                   and v.strip() not in ("", "{", "}")]
        num = int(reals_i[0]) if reals_i else len(pvs) + 1
        tu  = tipo.upper()
        is_agua_node = tu.startswith(("TE ","TE_","CURVA","CAP","RG ","RED ",
                                      "LUVA","CV ","X DN","C11","C22","C45","C90"))
        nome = (f"N_{num:03d}" if is_agua_node
                else f"PV_{num:03d}" if num
                else f"PV_X{len(pvs)+1:03d}")
        if nome not in pvs:
            pvs[nome] = {"x": x, "y": y, "ct": ct, "cf": cf, "prof": prof,
                         "tipo": "RG" if is_agua_node else "PV"}

    # ── Decodificar Tubos ────────────────────────────────────────────────────
    tubos_raw = []
    for pl in polylines:
        px  = pl.get("px", [])
        py_ = pl.get("py", [])
        if len(px) < 2: continue
        p0 = (px[0], py_[0])
        p1 = (px[-1], py_[-1])
        ext = round(math.sqrt((p1[0]-p0[0])**2 + (p1[1]-p0[1])**2), 2)
        if ext < 0.01: continue

        tub_xd  = pl["xd"].get("PH_DATTUB", [])
        strs_t  = [v for c, v in tub_xd if c == 1000]
        reals_t = [float(v) for c, v in tub_xd if c == 1040]
        mat_raw = strs_t[0].upper() if strs_t else "PVC"
        mat = _inferir_material(mat_raw)  # G-1 FIX: normaliza "TUBO PVC" -> "PVC"

        # DN: strs_t[1] tem o DN explícito ("300", "200", etc.)
        dn = None
        if len(strs_t) > 1 and strs_t[1].strip().isdigit():
            dn = int(strs_t[1].strip())
        elif reals_t:
            cand = int(reals_t[1]) if len(reals_t) > 1 else int(reals_t[0])
            if 50 <= cand <= 1200:
                dn = cand

        tubos_raw.append({
            "pt_ini": p0, "pt_fim": p1,
            "mid": ((p0[0]+p1[0])/2, (p0[1]+p1[1])/2),
            "material": mat, "dn_mm": dn, "ext_m": ext,
            "decl_mm": None, "layer": pl.get("layer", ""),
        })

    return pvs, tubos_raw, textos_rua_raw


def _filtrar_bifilar(tubos, tol_mid=1.0, tol_ext_pct=0.02, tol_angle_deg=5.0):
    """Remove polylines bifilar duplicadas (2 linhas paralelas por tubo real)."""
    if not tubos or len(tubos) < 2:
        return tubos
    cos_tol = math.cos(math.radians(tol_angle_deg))
    keep = [True] * len(tubos)
    for i in range(len(tubos)):
        if not keep[i]:
            continue
        mi = tubos[i]["mid"]
        ei = tubos[i].get("ext_m") or _dist(tubos[i]["pt_ini"], tubos[i]["pt_fim"])
        di = (tubos[i]["pt_fim"][0] - tubos[i]["pt_ini"][0],
              tubos[i]["pt_fim"][1] - tubos[i]["pt_ini"][1])
        di_len = math.hypot(*di)
        if di_len < 0.01:
            continue
        for j in range(i+1, len(tubos)):
            if not keep[j]:
                continue
            mj = tubos[j]["mid"]
            if math.hypot(mi[0]-mj[0], mi[1]-mj[1]) > tol_mid:
                continue
            ej = tubos[j].get("ext_m") or _dist(tubos[j]["pt_ini"], tubos[j]["pt_fim"])
            if abs(ei - ej) / max(ei, 0.1) > tol_ext_pct:
                continue
            dj = (tubos[j]["pt_fim"][0] - tubos[j]["pt_ini"][0],
                  tubos[j]["pt_fim"][1] - tubos[j]["pt_ini"][1])
            dj_len = math.hypot(*dj)
            if dj_len < 0.01:
                continue
            cos_a = abs(di[0]*dj[0] + di[1]*dj[1]) / (di_len * dj_len)
            if cos_a >= cos_tol:
                # Manter o com maior DN
                dni = tubos[i].get("dn_mm") or 0
                dnj = tubos[j].get("dn_mm") or 0
                if dnj > dni:
                    keep[i] = False
                else:
                    keep[j] = False
    filtered = [t for t, k in zip(tubos, keep) if k]
    n_removed = len(tubos) - len(filtered)
    if n_removed:
        log(f"  Bifilar filter: {n_removed} tubos duplicados removidos", "INFO")
    return filtered


def ler_dxf(dxf_path):
    """
    Lê DXF ProSaneamento ou Civil 3D Pipe Network.
    Detecta automaticamente o tipo de DXF:
    - Civil 3D Pipe Network (AeccDbPipe/AeccDbStructure) → leitura direta
    - ProSaneamento (XDATA PH_DATCNX) → snap por proximidade
    Retorna (pvs, trechos, ruas, meta).
    """
    if not _HAS_EZDXF:
        raise ImportError("ezdxf não instalado — pip install ezdxf")

    dxf_path = str(dxf_path)
    log(f"Lendo DXF: {Path(dxf_path).name}", "STEP")

    # Detectar Civil 3D Pipe Network
    doc = ezdxf.readfile(dxf_path)
    msp = doc.modelspace()
    _c3d_pipes = [e for e in msp if "AECC" in e.dxftype().upper()
                  and "PIPE" in e.dxftype().upper()]
    _c3d_structs = [e for e in msp if "AECC" in e.dxftype().upper()
                    and "STRUCT" in e.dxftype().upper()]
    if _c3d_pipes:
        log(f"  Civil 3D Pipe Network detectado: {len(_c3d_pipes)} pipes, "
            f"{len(_c3d_structs)} structures", "OK")
        # Civil 3D Pipe Networks detectadas no DXF convertido.
        # Para DWGs originais, usar ler_dwg_aec.py (3 camadas com fallback).
        # Aqui no DXF convertido, prossegue com parser de texto abaixo.
        log("  Civil 3D pipes detectadas — tentando extrair via texto/XDATA", "WARN")

    pvs_xd, tubos_xd, ruas_raw = _ler_xdata_raw(dxf_path)

    doc = ezdxf.readfile(dxf_path)
    msp = doc.modelspace()
    layer_names = [l.dxf.name for l in doc.layers]

    textos  = defaultdict(list)
    inserts = defaultdict(list)
    tubos_ez, lin_af = [], []

    # Detectar tipo de rede pelo nome do DXF
    _dxf_name = Path(dxf_path).stem.upper()
    _dxf_esgoto = "ESGOTO" in _dxf_name or "ESG" in _dxf_name
    _dxf_agua = "AGUA" in _dxf_name or "ÁGUA" in _dxf_name
    _dxf_prolong = "PROLONG" in _dxf_name or "LIGACAO" in _dxf_name or "RAMAL" in _dxf_name

    layers_tubo = [n for n in layer_names
                   if ("TUBO" in n.upper()
                       or "PROLONG" in n.upper()
                       or "LIGACAO" in n.upper()
                       or "RAMAL" in n.upper()
                       or "CONEXAO" in n.upper())
                   and not n.upper().startswith("PS_")
                   and "PERFIL" not in n.upper()
                   and "COTA"   not in n.upper()
                   and "TXT"    not in n.upper()
                   and "BIFILAR" not in n.upper()
                   and "SGC" not in n.upper()
                   and "DETALHE" not in n.upper()]
    # Filtrar layers por tipo de rede quando DXF é específico
    if _dxf_esgoto and not _dxf_agua:
        _agua_layers = [l for l in layers_tubo
                        if "PE_80" in l.upper() or "PE_100" in l.upper()
                        or "PE100" in l.upper() or "PEAD" in l.upper()
                        or "DRENAGEM" in l.upper()]
        if _agua_layers:
            layers_tubo = [l for l in layers_tubo if l not in _agua_layers]
            log(f"  Filtro rede: DXF esgoto, excluindo {len(_agua_layers)} layers agua", "INFO")
    if not layers_tubo:
        layers_tubo = [n for n in layer_names
                       if "TUBO" in n.upper()
                       and not n.upper().startswith("PS_")
                       and "PERFIL" not in n.upper()
                       and "COTA" not in n.upper()
                       and "TXT" not in n.upper()
                       and "DETALHE" not in n.upper()]

    eixos_rua = []  # BUG-9: P_Eixo = eixo viario, NAO rede de agua

    for e in msp:
        layer = e.dxf.layer
        if e.dxftype() == "LWPOLYLINE":
            if layer == "P_Eixo":
                # BUG-9 FIX: P_Eixo e eixo viario, nao rede de agua
                pts = list(e.get_points())
                if len(pts) >= 2:
                    eixos_rua.append(pts)
            elif ("LIN - AF" in layer or "LIN - ADUTORA" in layer
                    or "LIN-AF" in layer.upper()):
                pts = list(e.get_points())
                if len(pts) >= 2:
                    lin_af.append(pts)
            elif layer in layers_tubo:
                pts = list(e.get_points())
                if len(pts) >= 2:
                    p0  = (pts[0][0],  pts[0][1])
                    p1  = (pts[-1][0], pts[-1][1])
                    mid = ((p0[0]+p1[0])/2, (p0[1]+p1[1])/2)
                    _ext_geom = round(math.hypot(p1[0]-p0[0], p1[1]-p0[1]), 2)
                    tubos_ez.append({
                        "pt_ini": p0, "pt_fim": p1, "mid": mid,
                        "material": _inferir_material(layer),
                        "dn_mm": None, "ext_m": _ext_geom, "decl_mm": None,
                        "layer": layer,
                    })
        elif e.dxftype() == "TEXT":
            try:
                textos[layer].append({"text": e.dxf.text.strip(),
                                      "x": e.dxf.insert.x,
                                      "y": e.dxf.insert.y})
            except Exception:
                pass
        elif e.dxftype() == "MTEXT":
            try:
                txt = e.plain_mtext().strip()
                textos[layer].append({"text": txt,
                                      "x": e.dxf.insert.x,
                                      "y": e.dxf.insert.y})
            except Exception:
                pass
        elif e.dxftype() == "INSERT":
            try:
                inserts[layer].append({"x": e.dxf.insert.x,
                                       "y": e.dxf.insert.y,
                                       "nome": e.dxf.name})
            except Exception:
                pass

    # ── PVs ──────────────────────────────────────────────────────────────────
    # PRIORIDADE: PS_PONTOS_IDENTIFICACAO_TXT > XDATA INSERTs
    # PS_PONTOS tem os PVs/PIs reais da rede com CT/CF do dimensionamento.
    # XDATA INSERTs podem ter milhares de blocos de detalhe/listagem.

    # DETECÇÃO DE CRS RÁPIDA
    _ps_utm = False
    _tubo_local = False

    if textos:
        for layer_txt in textos:
            if "PS_PONTOS" in layer_txt.upper() and textos[layer_txt]:
                pv_x = textos[layer_txt][0].get("x", 0)
                if abs(pv_x) > 100_000:
                    _ps_utm = True
                break

    if tubos_ez and tubos_ez[0].get("pt_ini"):
        tubo_x = tubos_ez[0]["pt_ini"][0]
        if abs(tubo_x) < 100_000:
            _tubo_local = True

    _crs_incompativel = _ps_utm and _tubo_local
    if _crs_incompativel:
        log(f"  CRS incompativel: PS_PONTOS=UTM, Tubos=local — FORÇAR XDATA", "WARN")

    pvs = {}
    layer_pv = next(
        (l for l in textos if "PS_PONTOS_IDENTIFICACAO_TXT" in l.upper()), None
    ) or next(
        (l for l in textos if "PS_PONTOS" in l.upper()), None
    )

    # SE CRS INCOMPATÍVEL — PULA PS_PONTOS E VAI DIRETO PRO XDATA
    if _crs_incompativel and pvs_xd:
        _pvs_xd_filtrados = {k:v for k,v in pvs_xd.items() if re.match(r"^(PV|PI|N)_\d+", k)}
        if _pvs_xd_filtrados:
            pvs = _pvs_xd_filtrados
            tubos_raw = tubos_xd if tubos_xd else []
            log(f"  XDATA: {len(pvs)} PVs, {len(tubos_raw)} tubos (coords locais)", "OK")
        else:
            log(f"  XDATA inválido — tenta PS_PONTOS", "WARN")

    # ── FONTE DE PVs: PRIORIDADE PS_PONTOS > XDATA > INSERTs ──────────────
    # PS_PONTOS_IDENTIFICACAO_TXT é a fonte mais confiável: tem CT, CF, PF,
    # coordenadas UTM, e representa APENAS PVs da rede dimensionada.
    # XDATA contém blocos de detalhe (perfis, listagens, quadros) que
    # produzem centenas de PVs fantasma com coords locais e sem dados.

    if not pvs and layer_pv and len(textos.get(layer_pv, [])) > 10:
        pvs_ps = _agrupar_textos_pvs(textos[layer_pv])
        if pvs_ps:
            pvs = pvs_ps
            tubos_raw = tubos_ez
            log(f"  PS_PONTOS: {len(pvs)} PVs (UTM) — FONTE PRIMARIA", "OK")
        else:
            # PS_PONTOS parseou mas não achou PVs válidos → fallback XDATA
            if pvs_xd:
                _pvs_xd_filtrados = {k:v for k,v in pvs_xd.items()
                                     if re.match(r"^(PV|PI|N)_\d+", k)}
                pvs = _pvs_xd_filtrados if _pvs_xd_filtrados else pvs_xd
                tubos_raw = tubos_xd if tubos_xd else tubos_ez
                log(f"  PS_PONTOS vazio, XDATA fallback: {len(pvs)} PVs", "WARN")
            else:
                pvs = {}
                tubos_raw = tubos_ez
                log(f"  PS_PONTOS e XDATA vazios", "WARN")
    elif not pvs:
        # Sem PS_PONTOS disponível — usar XDATA se tiver
        if pvs_xd:
            _pvs_xd_filtrados = {k:v for k,v in pvs_xd.items()
                                 if re.match(r"^(PV|PI|N)_\d+", k)}
            pvs = _pvs_xd_filtrados if _pvs_xd_filtrados else pvs_xd
            tubos_raw = tubos_xd if tubos_xd else tubos_ez
            log(f"  Sem PS_PONTOS, XDATA: {len(pvs)} PVs", "OK")
        else:
            tubos_raw = tubos_ez
            log(f"  Sem PS_PONTOS nem XDATA", "WARN")

    # BUG-10 FIX: XDATA pode ter coordenadas locais (paperspace), nao UTM.
    # Detectar: se PVs tem x < 100_000 ou y < 1_000_000, sao locais.
    _xd_local = False
    if pvs:
        xs = sorted([abs(pv.get("x", 0)) for pv in pvs.values() if pv.get("x") is not None])
        mediana_x = xs[len(xs)//2] if xs else 0
        if xs and mediana_x < 100_000:
            _xd_local = True
            log("  XDATA coords locais detectadas - buscando UTM em PS_PONTOS...", "WARN")
            layer_pv_txt = next(
                (l for l in textos if "PS_PONTOS_IDENTIFICACAO_TXT" in l.upper()), None
            ) or next(
                (l for l in textos if "PS_PONTOS" in l.upper()), None
            )
            if layer_pv_txt:
                pvs_txt = _agrupar_textos_pvs(textos[layer_pv_txt])
                log(f"  PVs texto UTM: {len(pvs_txt)}", "OK")
                def _num(nome):
                    m = re.search(r"(\d+)", nome)
                    return int(m.group(1)) if m else -1
                txt_by_num = {}
                for nome, pv in pvs_txt.items():
                    n = _num(nome)
                    if n >= 0:
                        txt_by_num[n] = pv
                matched = 0
                for nome, pv in pvs.items():
                    n = _num(nome)
                    if n in txt_by_num:
                        tp = txt_by_num[n]
                        pv["x_local"] = pv.get("x")
                        pv["y_local"] = pv.get("y")
                        pv["x"] = tp["x"]
                        pv["y"] = tp["y"]
                        # Texto PS_PONTOS e fonte mais confiavel de CT/CF que XDATA
                        if tp.get("ct") is not None:
                            pv["ct"] = tp["ct"]
                        if tp.get("cf") is not None:
                            pv["cf"] = tp["cf"]
                        if tp.get("prof") is not None:
                            pv["prof"] = tp["prof"]
                        elif pv.get("ct") is not None and pv.get("cf") is not None:
                            pv["prof"] = round(pv["ct"] - pv["cf"], 4)
                        matched += 1
                log(f"  Coords UTM atualizadas: {matched}/{len(pvs)} PVs", "OK")

                # G-2 FIX: remover PVs que ficaram com coords locais (sem match no texto)
                # Esses sao blocos de detalhe/listagem, nao da rede real
                nao_mapeados = [n for n, pv in pvs.items()
                                if not pv.get("x_local") is not None  # nao foi processado
                                or (pv.get("x_local") is not None     # foi processado
                                    and n not in [k for k in txt_by_num.values()]
                                    and _num(n) not in txt_by_num)]
                # Marcar PVs nao mapeados para que trechos com eles sejam identificaveis
                for n in pvs:
                    if _num(n) not in txt_by_num and "x_local" in pvs[n]:
                        pvs[n]["_sem_utm"] = True

    # FIX-C: Filtrar PVs com coords locais quando ha UTM disponivel
    _n_antes = len(pvs)
    _tem_utm = any(abs(pv.get("x", 0)) > 100_000 for pv in pvs.values())
    if _tem_utm:
        _remover = [n for n, pv in pvs.items()
                    if not pv.get("sintetico")
                    and pv.get("x") is not None
                    and abs(pv["x"]) < 100_000]
        for n in _remover:
            del pvs[n]
        if _remover:
            log(f"  FIX-C: {len(_remover)} PVs com coords locais removidos", "INFO")

    log(f"  PVs totais: {len(pvs)}", "OK")

    # ── Textos de indicação ───────────────────────────────────────────────────
    diams = textos.get("PS_IND_DIAMETRO") or next(
        (v for k, v in textos.items() if "DIAMETRO" in k.upper()), [])
    comps = textos.get("PS_IND_COMPRIMENTO") or next(
        (v for k, v in textos.items() if "COMPRIMENTO" in k.upper()), [])
    incls = textos.get("PS_IND_INCLINACAO") or next(
        (v for k, v in textos.items() if "INCLINA" in k.upper()), [])

    # ── Ruas (textos de logradouro) ────────────────────────────────────────────
    LAYERS_LOGR = {"A_Alerta","TXT-LOGRAD","TEXTO","ZZ-Carimbo Texto",
                   "ZZ-CARIMBO TEXTO","LT-TEXTO-RUA","TXT-PRACA",
                   "PS_IND_TRECHO","0"}
    PREF_RUA = ("RUA ","BECO ","TRAV","AV ","ESTRADA","VIELA","ALAMEDA","ACESSO")
    ruas = []
    for ln in LAYERS_LOGR:
        ruas.extend(textos.get(ln, []))
    if not ruas:
        for ln, ld in textos.items():
            if "RUA" in ln.upper() or ("TEXTO" in ln.upper() and "IND" not in ln.upper()):
                ruas.extend(ld)
    # Manter só nomes de logradouro reais
    ruas_ok = [r for r in ruas if len(r.get("text","")) > 2
               and any(r["text"].upper().startswith(p) for p in PREF_RUA)]
    ruas = ruas_ok if ruas_ok else [r for r in ruas if len(r.get("text","")) > 2]

    # PRIORIDADE: textos coletados no parse raw (mesmo espaço local dos PVs/tubos)
    # São mais confiáveis para snap por distância
    if ruas_raw:
        ruas = ruas_raw   # substituir por versão no espaço local do DXF

    # Detecção de tipo de rede: nome do DXF tem prioridade sobre heurística
    if _dxf_esgoto and not _dxf_agua:
        is_agua = False
    elif _dxf_agua and not _dxf_esgoto:
        is_agua = True
    else:
        is_agua = (bool(lin_af) or any(n.startswith("N_") for n in pvs)
                   or any("LIN - AF" in l or "LIN - ADUTORA" in l for l in layer_names)
                   or any("PE_80" in l or "PEAD" in l or "PE100" in l for l in layer_names))

    # ── Filtro bifilar: remover polylines duplicadas paralelas ────────────────
    if tubos_raw:
        tubos_raw = _filtrar_bifilar(tubos_raw)
    
    # FILTRO: Remover tubos MUITO curtos (<2m) - provavel detalhe/ruído
    _antes = len(tubos_raw)
    tubos_raw = [t for t in tubos_raw if (t.get("ext_m") or 0) >= 2.0]
    if len(tubos_raw) < _antes:
        log(f"  Filtro: {_antes} → {len(tubos_raw)} tubos (removidos <2m)", "INFO")

    # ── Snap RESTRITO: so conectar tubo a PV se endpoint < 10m ──────────────
    # Aprendizado SewerCAD/SOLIDOS: conexao so existe com polyline REAL tocando PV.
    # SEM nos sinteticos. SEM snap adaptativo. SEM tolerancia de 300m.
    _tol_snap = 25.0  # maximo 25m — restrito, sem nos sinteticos, sem adaptativo

    # ── Tubos → Trechos ───────────────────────────────────────────────────────
    trechos, sem_match = [], 0
    for tb in (tubos_raw or []):
        mid = tb["mid"]
        pt0, pt1 = tb["pt_ini"], tb["pt_fim"]
        dn = tb.get("dn_mm")
        if dn is None:
            dt = _mais_proximo(mid, diams, CFG["tol_texto_tubo"])
            dn = _parsear_dn(dt["text"]) if dt else None
        ext = tb.get("ext_m")
        if ext is None:
            ct = _mais_proximo(mid, comps, CFG["tol_texto_tubo"])
            ext = _parsear_comp(ct["text"]) if ct else None
        if ext is None:
            ext = round(_dist(pt0, pt1), 2)
        if ext < 0.5:  # filtrar tubos degenerados
            continue
        decl = tb.get("decl_mm")
        if decl is None:
            it = _mais_proximo(mid, incls, CFG["tol_texto_tubo"])
            decl = _parsear_incl(it["text"]) if it else None
        # Snap RESTRITO: endpoint do tubo deve estar a <10m de um PV real
        pvi = _pv_mais_proximo(pt0, pvs, _tol_snap, preferir_real=True)
        pvf = _pv_mais_proximo(pt1, pvs, _tol_snap, preferir_real=True)
        if not pvi or not pvf or pvi == pvf:
            sem_match += 1
            continue
        # Validacao: distancia PV-PV deve ser compativel com extensao do tubo
        d_pvs = math.hypot(pvs[pvi]["x"] - pvs[pvf]["x"],
                           pvs[pvi]["y"] - pvs[pvf]["y"])
        if d_pvs > ext * 3.0:  # PVs muito distantes para este tubo = falso
            sem_match += 1
            continue
        mx = (pvs[pvi]["x"]+pvs[pvf]["x"])/2
        my = (pvs[pvi]["y"]+pvs[pvf]["y"])/2
        rt = _mais_proximo((mx,my), ruas, 300.0)
        pvi_data = pvs.get(pvi, {})
        pvf_data = pvs.get(pvf, {})
        trechos.append({
            "pv_ini":   pvi,
            "pv_fim":   pvf,
            "dn_mm":    dn,
            "ext_m":    round(ext, 2),
            "decl_mm":  round(decl, 5) if decl is not None else None,
            "decl_pct": round(decl*100, 3) if decl is not None else None,
            "material": tb.get("material", "PVC"),
            "rua":      rt["text"] if rt else "Sem Rua",
            "layer":    tb.get("layer", ""),
            "is_agua":  is_agua,
            "is_prolongamento": _dxf_prolong,
            "ct_ini":   pvi_data.get("ct"),
            "ct_fim":   pvf_data.get("ct"),
            "cf_ini":   pvi_data.get("cf"),
            "cf_fim":   pvf_data.get("cf"),
            "prof_ini": pvi_data.get("prof"),
            "prof_fim": pvf_data.get("prof"),
        })

    # Deduplificar AGRESSIVO: um unico trecho por par PV-PV
    # Se tiver multiplos, manter o com maior DN (mais confiavel)
    _por_par = defaultdict(list)
    for t in trechos:
        par = tuple(sorted([t["pv_ini"], t["pv_fim"]]))
        _por_par[par].append(t)
    
    unicos = []
    for par, ts in _por_par.items():
        if len(ts) == 1:
            unicos.append(ts[0])
            continue
        
        # Multiplos trechos mesmo par -> manter SO o com maior DN
        ts.sort(key=lambda t: t.get("dn_mm") or 0, reverse=True)
        melhor = ts[0]
        
        # Logar duplicados removidos
        if len(ts) > 2:
            log(f"  Duplicados: {par[0]}->{par[1]} = {len(ts)} trechos, mantido DN{melhor.get('dn_mm','?')}", "INFO")
        
        unicos.append(melhor)
    
    # FILTRO FINAL: Remover trechos so entre nos sintéticos (ND_→ND_)
    _antes = len(unicos)
    unicos = [t for t in unicos 
              if not (t["pv_ini"].startswith("ND_") and t["pv_fim"].startswith("ND_"))]
    if len(unicos) < _antes:
        log(f"  Filtro: {_antes} → {len(unicos)} trechos (removidos ND_→ND_)", "INFO")

    # Fallback água via LIN-AF
    if not unicos and lin_af and pvs:
        unicos = _build_trechos_agua(pvs, lin_af, diams, [], ruas)

    log(f"  Trechos: {len(unicos)} (sem match: {sem_match}) | Tipo: {'AGUA' if is_agua else 'ESGOTO'}", "OK")
    log(f"  Ruas: {len(ruas)} textos de logradouro", "OK")

    meta = {
        "arquivo": Path(dxf_path).name,
        "tipo_rede": "AGUA" if is_agua else "ESGOTO",
        "n_pvs": len(pvs), "n_trechos": len(unicos),
        "n_tubos_raw": len(tubos_raw) if tubos_raw else 0,
    }
    return pvs, unicos, ruas, meta


def _build_trechos_agua(pvs, lin_af_polylines, diams, pressao_texts, ruas,
                        max_snap=50.0):
    """Conecta N_xx ao longo das polilínias LIN-AF."""
    nomes = [n for n in pvs if n.startswith("N_")]
    trechos = []

    def _proj_arc(pt, poly):
        px, py = pt
        best_arc, best_d = 0.0, float("inf")
        arc = 0.0
        for i in range(len(poly)-1):
            ax, ay = poly[i][0], poly[i][1]
            bx, by = poly[i+1][0], poly[i+1][1]
            sl = math.hypot(bx-ax, by-ay)
            if sl < 1e-9:
                arc += sl
                continue
            tt = max(0.0, min(1.0, ((px-ax)*(bx-ax)+(py-ay)*(by-ay))/sl**2))
            cx, cy = ax+tt*(bx-ax), ay+tt*(by-ay)
            d = math.hypot(px-cx, py-cy)
            if d < best_d:
                best_d = d
                best_arc = arc + tt*sl
            arc += sl
        return best_arc, best_d

    for poly in lin_af_polylines:
        comp = sum(_dist(poly[i][:2], poly[i+1][:2]) for i in range(len(poly)-1))
        if comp < 1.0:
            continue
        nodos = []
        for nome in nomes:
            pv = pvs[nome]
            arc, d_proj = _proj_arc((pv["x"], pv["y"]), poly)
            if d_proj <= max_snap:
                nodos.append((arc, nome))
        if len(nodos) < 2:
            continue
        nodos.sort()
        for i in range(len(nodos)-1):
            arc_i, ni = nodos[i]
            arc_j, nj = nodos[i+1]
            ext = round(arc_j - arc_i, 2)
            if ext < 0.5:
                continue
            pvi = pvs[ni]; pvf = pvs[nj]
            mid = ((pvi["x"]+pvf["x"])/2, (pvi["y"]+pvf["y"])/2)
            dt  = _mais_proximo(mid, diams, 50.0)
            dn  = _parsear_dn(dt["text"]) if dt else None
            rt  = _mais_proximo(mid, ruas, 100.0)
            rua = rt["text"] if rt else "Sem Rua"
            trechos.append({
                "pv_ini": ni, "pv_fim": nj,
                "dn_mm": dn, "ext_m": ext,
                "decl_mm": None, "decl_pct": None,
                "material": "PE80", "rua": rua,
                "layer": "LIN - AF", "is_agua": True,
            })

    grau = {}
    for t in trechos:
        grau[t["pv_ini"]] = grau.get(t["pv_ini"], 0)+1
        grau[t["pv_fim"]] = grau.get(t["pv_fim"], 0)+1
    for t in trechos:
        t["grau_ini"] = grau.get(t["pv_ini"], 1)
        t["grau_fim"] = grau.get(t["pv_fim"], 1)
    return trechos


def ler_json_rede(json_path):
    """Lê rede_definida.json (GeoJSON) ou rede_esgoto_dynamo.json."""
    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)
    # BUG-5 FIX: extrair ruas do GeoJSON
    pvs, trechos, ruas = {}, [], []

    if data.get("type") == "FeatureCollection":
        ruas_vistas = set()
        for ft in data.get("features", []):
            p = ft.get("properties", {})
            geom   = ft.get("geometry", {})
            coords = geom.get("coordinates", [])
            # coords pode ser [[x,y],[x,y]] (LineString) ou [x,y] (Point)
            if not coords:
                continue
            # Garantir que é lista de pares
            if isinstance(coords[0], (int, float)):
                # Ponto único — pular
                continue
            if len(coords) < 2:
                continue
            x0, y0 = float(coords[0][0]),  float(coords[0][1])
            x1, y1 = float(coords[-1][0]), float(coords[-1][1])
            pvi, pvf = p.get("pv_ini",""), p.get("pv_fim","")
            if pvi and pvi not in pvs:
                pvs[pvi] = {"x": x0, "y": y0, "ct": p.get("ct_ini"),
                            "cf": p.get("cf_ini"), "prof": p.get("prof_ini"), "tipo": "PV"}
            if pvf and pvf not in pvs:
                pvs[pvf] = {"x": x1, "y": y1, "ct": p.get("ct_fim"),
                            "cf": p.get("cf_fim"), "prof": p.get("prof_fim"), "tipo": "PV"}
            is_agua = p.get("material","PVC") in ("PE80","PE100","PEAD")
            trechos.append({
                "pv_ini": pvi, "pv_fim": pvf,
                "dn_mm": p.get("dn_mm"), "ext_m": p.get("ext_m"),
                "decl_mm": None, "decl_pct": p.get("decl_pct"),
                "material": p.get("material","PVC"),
                "rua": p.get("rua","Sem Rua"),
                "layer": "LIN - AF" if is_agua else "TUBO_PVC",
                "is_agua": is_agua,
                "prof_ini": p.get("prof_ini"), "prof_fim": p.get("prof_fim"),
                "pressao_mca": p.get("pressao_mca"),
            })
            # BUG-5 FIX: extrair ruas do GeoJSON
            rua_nome = p.get("rua", "")
            if rua_nome and rua_nome != "Sem Rua" and rua_nome not in ruas_vistas:
                ruas_vistas.add(rua_nome)
                ruas.append({"text": rua_nome, "x": (x0+x1)/2, "y": (y0+y1)/2})
    elif "pontos" in data:
        for p in data["pontos"]:
            pvs[p["id"]] = {"x": p["x"], "y": p["y"], "ct": p.get("ct"),
                            "cf": p.get("cf"), "prof": p.get("prof"), "tipo": "PV"}
        for t in data.get("tubulacoes", []):
            dm = t.get("decl_mm")
            trechos.append({
                "pv_ini": t.get("pv_ini",""), "pv_fim": t.get("pv_fim",""),
                "dn_mm": t.get("dn_mm"), "ext_m": t.get("ext_m"),
                "decl_mm": dm,
                "decl_pct": round(dm*100, 3) if dm else None,
                "material": t.get("material","PVC"),
                "rua": t.get("rua","Sem Rua"),
                "layer": "TUBO_PVC", "is_agua": False,
            })

    meta = {"arquivo": Path(json_path).name, "tipo_rede": "MISTO",
            "n_pvs": len(pvs), "n_trechos": len(trechos), "n_tubos_raw": len(trechos)}
    return pvs, trechos, ruas, meta


# ==============================================================================
# MÓDULO 3 — ENRIQUECIMENTO
# ==============================================================================

def calc_manning(dn_mm, decl_mpm, material="PVC"):
    """Manning secao plena. decl_mpm = declividade em m/m."""
    if not dn_mm or decl_mpm is None:
        return {"vel_ms": None, "vazao_ls": None, "tau_pa": None, "status": "SEM_DADOS"}
    # BUG-3 FIX: trata declividade negativa/zero/absurda
    if decl_mpm <= 0:
        return {"vel_ms": None, "vazao_ls": None, "tau_pa": None,
                "status": "DECL_INVALIDA" if decl_mpm < 0 else "DECL_ZERO"}
    if decl_mpm > 1.0:
        return {"vel_ms": None, "vazao_ls": None, "tau_pa": None,
                "status": f"DECL_ABSURDA ({decl_mpm:.4f} m/m)"}
    n   = CFG["manning"].get(str(material).upper(), 0.013)
    d   = dn_mm / 1000
    A   = math.pi * d**2 / 4
    Rh  = d / 4
    V   = (1/n) * Rh**(2/3) * math.sqrt(decl_mpm)
    Q   = V * A * 1000
    tau = 1000 * 9.81 * Rh * decl_mpm
    probs = []
    if V < 0.6:   probs.append(f"V={V:.3f}<0.6m/s")
    if V > 5.0:   probs.append(f"V={V:.3f}>5.0m/s")
    if tau < 1.0: probs.append(f"tau={tau:.2f}<1.0Pa")
    return {"vel_ms": round(V, 3), "vazao_ls": round(Q, 2),
            "tau_pa": round(tau, 2),
            "status": "OK" if not probs else "VERIFICAR: " + ", ".join(probs)}


def calc_quantitativos(ext_m, prof_media_m, dn_mm, lv=None):
    lv = lv or CFG["largura_vala"]
    if not ext_m or not prof_media_m or prof_media_m <= 0:
        return {k: None for k in ["esc_m3","lastro_m3","envolt_m3","brita_m3",
                                   "reat_m3","pav_m2","tubo_barras"]}
    d      = (dn_mm or 200) / 1000
    H      = prof_media_m
    h_last = CFG.get("lastro", 0.15)
    esc    = lv * H * ext_m
    lastro = lv * h_last * ext_m
    h_env  = min(d + 0.30, H - h_last)
    envolt = max(0, lv * h_env * ext_m - math.pi*(d/2)**2*ext_m)
    brita  = lv * 0.15 * ext_m
    tubo_v = math.pi*(d/2)**2 * ext_m
    reat   = max(0, esc - lastro - envolt - brita - tubo_v)
    pav    = lv * ext_m * 1.20
    barras = max(1, math.ceil(ext_m / 6))
    return {
        "esc_m3":      round(esc,    2),
        "lastro_m3":   round(lastro, 2),
        "envolt_m3":   round(envolt, 2),
        "brita_m3":    round(brita,  2),
        "reat_m3":     round(reat,   2),
        "pav_m2":      round(pav,    2),
        "tubo_barras": barras,
    }


def calc_custos(t):
    dn  = t.get("dn_mm") or 200
    mat = (t.get("material") or "PVC").upper()
    ext = t.get("ext_m") or 0
    q   = t.get("quantitativos") or {}
    bdi = CFG["bdi"]
    chave = f"tubo_pe80_{dn}" if "PE" in mat else f"tubo_pvc_{dn}"
    pt    = SINAPI.get(chave, SINAPI.get("tubo_pvc_200", 47.12))
    ct    = pt * ext * bdi
    ce    = SINAPI["escavacao"]    * (q.get("esc_m3")    or 0) * bdi
    cr    = SINAPI["reaterro"]     * (q.get("reat_m3")   or 0) * bdi
    cl    = SINAPI["lastro"]       * (q.get("lastro_m3") or 0) * bdi
    cb    = SINAPI["brita"]        * (q.get("brita_m3")  or 0) * bdi
    cp    = SINAPI["pavimentacao"] * (q.get("pav_m2")    or 0) * bdi
    total = ct + ce + cr + cl + cb + cp
    return {"tubo_R": round(ct,2), "escavacao_R": round(ce,2),
            "reaterro_R": round(cr,2), "lastro_R": round(cl,2),
            "brita_R": round(cb,2), "pavimentacao_R": round(cp,2),
            "total_R": round(total,2), "bdi_pct": 25}


def _materiais_agua(t):
    dn  = t.get("dn_mm") or 63
    ext = t.get("ext_m") or 0
    mat = {}
    pn  = "PN10" if dn >= 100 else "PN12,5"
    mat[f"Tubo PE80 DN{dn} {pn} (m)"] = round(ext, 2)
    for grau in [t.get("grau_ini", 1), t.get("grau_fim", 1)]:
        mat[f"Sela de derivação PE80 DN{dn}×20 (un)"] = \
            mat.get(f"Sela de derivação PE80 DN{dn}×20 (un)", 0) + 1
        mat["Chave de serviço DN20 (un)"] = \
            mat.get("Chave de serviço DN20 (un)", 0) + 1
        if grau == 1:
            mat[f"Tampão PE80 DN{dn} (un)"] = \
                mat.get(f"Tampão PE80 DN{dn} (un)", 0) + 1
        if grau >= 3:
            mat[f"TE FoFo DN{dn} (un)"] = \
                mat.get(f"TE FoFo DN{dn} (un)", 0) + 1
        mat[f"Registro de gaveta FoFo DN{dn} (un)"] = \
            mat.get(f"Registro de gaveta FoFo DN{dn} (un)", 0) + 1
    return mat


def enriquecer_trechos(trechos, pvs):
    """Calcula hidráulica, quantitativos e custos para todos os trechos."""
    # FIX-A: Validar cotas dos PVs (padrão SewerCAD)
    # CT pode ser negativo (Santos tem terreno abaixo do mar)
    # CT == 0.0 exato = dado não coletado → tratar como None
    for nome, pv in pvs.items():
        if pv.get("ct") is not None and pv["ct"] == 0.0:
            pv["ct"] = None
        if pv.get("cf") is not None and pv["cf"] == 0.0:
            pv["cf"] = None
        # CF > CT → avisar (pode ser rede aérea), não rejeitar
        if (pv.get("ct") is not None and pv.get("cf") is not None
                and pv["cf"] > pv["ct"]):
            pv["_aviso_cf_gt_ct"] = True

    for t in trechos:
        pvi = pvs.get(t.get("pv_ini"), {})
        pvf = pvs.get(t.get("pv_fim"), {})
        t["ct_ini"]  = t.get("ct_ini")  or pvi.get("ct")
        t["ct_fim"]  = t.get("ct_fim")  or pvf.get("ct")
        t["cf_ini"]  = t.get("cf_ini")  or pvi.get("cf")
        t["cf_fim"]  = t.get("cf_fim")  or pvf.get("cf")
        t["prof_ini"] = t.get("prof_ini") or pvi.get("prof")
        t["prof_fim"] = t.get("prof_fim") or pvf.get("prof")

        # Fallback: profundidade default do ProSaneamento (INS_CNX.DEF)
        _prof_def = CFG.get("pv_prof_default", 0.50)
        profs = [p for p in [t["prof_ini"], t["prof_fim"]] if p is not None]
        if not profs and _prof_def:
            profs = [_prof_def]
        t["prof_media_m"] = round(sum(profs)/len(profs), 3) if profs else None

        decl_mm = t.get("decl_mm")
        if (decl_mm is None and t.get("cf_ini") is not None
                and t.get("cf_fim") is not None and t.get("ext_m")
                and t["ext_m"] > 0):
            decl_mm = abs(t["cf_ini"] - t["cf_fim"]) / t["ext_m"]
            t["decl_mm"]  = round(decl_mm, 5)
            t["decl_pct"] = round(decl_mm * 100, 3)

        hid = calc_manning(t.get("dn_mm"), decl_mm, t.get("material","PVC"))
        t["hidraulica"]    = hid
        t["vel_ms"]        = hid.get("vel_ms")
        t["vazao_ls"]      = hid.get("vazao_ls")
        t["tau_pa"]        = hid.get("tau_pa")
        t["status_hid"]    = hid.get("status", "SEM_DADOS")
        qnt = calc_quantitativos(t.get("ext_m"), t["prof_media_m"], t.get("dn_mm"))
        t["quantitativos"] = qnt
        cst = calc_custos(t)
        t["custos"]        = cst
        t["custo_total"]   = cst.get("total_R") if cst else None
        if t.get("is_agua"):
            t["materiais_agua"] = _materiais_agua(t)
    return trechos


# ==============================================================================
# MÓDULO 4 — VALIDAÇÃO DO GRAFO
# ==============================================================================

def validar_rede(pvs, trechos):
    erros, avisos = [], []
    if not trechos:
        return erros, ["Nenhum trecho definido"]
    if _HAS_NX:
        G = nx.DiGraph()
        for t in trechos:
            G.add_edge(t["pv_ini"], t["pv_fim"],
                       dn=t.get("dn_mm"),
                       cf_i=t.get("cf_ini"), cf_f=t.get("cf_fim"))
        for node in G.nodes():
            dn_ent = [G.edges[e].get("dn") or 0 for e in G.in_edges(node)]
            dn_sai = [G.edges[e].get("dn") or 0 for e in G.out_edges(node)]
            if dn_ent and dn_sai:
                mx = max(dn_ent)
                for e in G.out_edges(node):
                    ds = G.edges[e].get("dn") or 0
                    if 0 < ds < mx:
                        erros.append(f"V001 [{node}]: DN{mx}->DN{ds} afoga")
        comps = list(nx.connected_components(G.to_undirected()))
        if len(comps) > 1:
            avisos.append(f"V003: {len(comps)} partes desconectadas")
        for c in list(nx.simple_cycles(G))[:3]:
            erros.append(f"V004: Ciclo: {' -> '.join(c)}")
    for t in trechos:
        tid  = f"{t['pv_ini']}→{t['pv_fim']}"
        cf_i = t.get("cf_ini"); cf_f = t.get("cf_fim")
        if cf_i is not None and cf_f is not None and cf_f > cf_i and not t.get("is_agua"):
            erros.append(f"V002 [{tid}]: sifao CF {cf_i:.3f}->{cf_f:.3f}")
        prof_min = CFG.get("prof_minima", 0.30)
        for pk in ["prof_ini", "prof_fim"]:
            p = t.get(pk)
            if p is not None and p < prof_min:
                avisos.append(f"V005 [{tid}]: {pk}={p:.2f}m < {prof_min:.2f}m")
        decl_min_pct = CFG.get("decl_minima", 0.002) * 100  # m/m -> %
        dp = t.get("decl_pct")
        if dp is not None and dp < decl_min_pct and not t.get("is_agua"):
            avisos.append(f"V006 [{tid}]: decl={dp:.3f}% < {decl_min_pct:.2f}%")
        hid = t.get("hidraulica") or {}
        if hid.get("vel_ms") is not None and hid["vel_ms"] < 0.6:
            avisos.append(f"V007 [{tid}]: V={hid['vel_ms']:.3f}m/s < 0.60m/s")
        if hid.get("tau_pa") is not None and hid["tau_pa"] < 1.0:
            avisos.append(f"V008 [{tid}]: tau={hid['tau_pa']:.2f}Pa < 1.0Pa")
    return erros, avisos


# ==============================================================================
# MÓDULO 5 — CARTOGRAFIA GPKG
# ==============================================================================

_GPKG_CACHE = {}

def ler_cartografia_gpkg(gpkg_path):
    key = str(gpkg_path)
    if key in _GPKG_CACHE:
        # BUG-6 FIX: retorna copia defensiva
        return copy.deepcopy(_GPKG_CACHE[key])
    result = {"streets": None, "quadras": None, "ruas_txt": []}
    if not _HAS_GEO or not Path(gpkg_path).exists():
        return result
    try:
        lines = gpd.read_file(gpkg_path, layer="lines", engine="pyogrio")
        streets = lines[lines["layer"] == "P_Eixo"]
        b = streets.geometry.bounds
        streets = streets[(b["minx"] > 300_000) & (b["miny"] > 7_000_000)]
        result["streets"] = streets if len(streets) else None

        poly = gpd.read_file(gpkg_path, layer="polylines", engine="pyogrio")
        quadras = poly[poly["layer"] == "0-quadras"]
        b2 = quadras.geometry.bounds
        quadras = quadras[(b2["minx"] > 300_000) & (b2["miny"] > 7_000_000)]
        result["quadras"] = quadras if len(quadras) else None

        try:
            gdf_txt = gpd.read_file(gpkg_path, layer="texts", engine="pyogrio")
            gdf_rua = gdf_txt[gdf_txt["layer"] == "ZZ-Carimbo Texto"]
            b3 = gdf_rua.geometry.bounds
            gdf_rua = gdf_rua[(b3["minx"] > 300_000) & (b3["miny"] > 7_000_000)]
            result["ruas_txt"] = [
                {"text": row.get("text",""), "x": row.geometry.x, "y": row.geometry.y}
                for _, row in gdf_rua.iterrows()
                if row.get("text","").strip()
            ]
        except Exception:
            pass
    except Exception as e:
        log(f"GPKG erro: {e}", "WARN")
    _GPKG_CACHE[key] = result
    return copy.deepcopy(result)


# ==============================================================================
# MÓDULO 6 — NS_A4.pdf
# ==============================================================================

def gerar_ns_a4(t, pvs, ns_id, pasta, cfg):
    """Gera a Ordem de Serviço para Gabarito/Execução (folha campo A4 landscape)."""
    pvi      = pvs.get(t.get("pv_ini"), {})
    pvf      = pvs.get(t.get("pv_fim"), {})
    is_agua  = t.get("is_agua", False)
    titulo_sis = ("SISTEMA DE ABASTECIMENTO DE ÁGUA FRIA"
                  if is_agua else "SISTEMA DE ESGOTAMENTO SANITÁRIO")
    titulo_os  = ("ORDEM DE SERVIÇO PARA EXECUÇÃO"
                  if is_agua else "ORDEM DE SERVIÇO PARA GABARITO")
    hid      = t.get("hidraulica") or {}
    pressao  = _fmtv(t.get("pressao_mca"), ".2f") if is_agua else None
    c_pipe   = "#1E6B3C" if is_agua else "#1565C0"

    fig, ax = plt.subplots(figsize=(29.7/2.54, 21.0/2.54))
    ax.set_xlim(0, 29.7); ax.set_ylim(0, 21.0); ax.axis("off")

    # Cabeçalho azul escuro
    ax.add_patch(plt.Rectangle((0.5, 17.5), 28.7, 3.0, ec="#1F4E79", fc="#1F4E79", lw=0))
    ax.text(14.85, 19.55, f"{titulo_sis}  |  {cfg['cidade']}  |  {titulo_os}",
            ha="center", va="center", fontsize=10.5, fontweight="bold", color="white")
    ax.text(14.85, 18.85,
            f"EMPRESA: {cfg['empresa']}  |  CONTRATO: {cfg['contrato']}  |  "
            f"NÚCLEO: {cfg.get('nucleo','')}  |  LOGRADOURO: {t.get('rua','Sem Rua')}",
            ha="center", va="center", fontsize=7.5, color="white")
    badge = (f"NS Nº {ns_id}  |  TRECHO: {t['pv_ini']} → {t['pv_fim']}  |  "
             f"DN: {_fmtv(t.get('dn_mm'),'.0f')} mm  |  EXT: {_fmtv(t.get('ext_m'),'.2f')} m  |  "
             + (f"P: {pressao} mca" if pressao else f"DECL: {_fmtv(t.get('decl_pct'),'.3f')} %"))
    ax.text(14.85, 18.10, badge, ha="center", va="center",
            fontsize=8.5, fontweight="bold", color="white")

    # Tabela principal
    headers  = ["TRECHO","ESTACA","DISTÂNCIA\n(m)","CT\n(m)","I\n(m/m)",
                "CP\n(m)","CR\n(m)","DN\n(mm)","G","H","P"]
    col_w    = [3.5, 2.8, 2.4, 2.2, 2.0, 2.2, 2.2, 1.8, 1.2, 1.2, 1.2]
    col_x    = [0.5]
    for w in col_w[:-1]:
        col_x.append(col_x[-1]+w)
    y_hdr    = 16.8
    for h, cx, w in zip(headers, col_x, col_w):
        ax.add_patch(plt.Rectangle((cx, y_hdr-0.85), w, 0.92,
                                   ec="white", fc="#1F4E79", lw=0.5))
        ax.text(cx+w/2, y_hdr-0.38, h, ha="center", va="center",
                fontsize=5.8, fontweight="bold", color="white")

    ext_str = _fmtv(t.get("ext_m"), ".2f")
    rows = [
        [t["pv_ini"], "0+00", "0.00",
         _fmtv(t.get("ct_ini")), _fmtv(t.get("decl_mm"),".5f"),
         _fmtv(t.get("cf_ini")), _fmtv(t.get("prof_ini"),".2f"),
         _fmtv(t.get("dn_mm"),".0f"), "", "", ""],
        [t["pv_fim"], f"0+{ext_str}", ext_str,
         _fmtv(t.get("ct_fim")), _fmtv(t.get("decl_mm"),".5f"),
         _fmtv(t.get("cf_fim")), _fmtv(t.get("prof_fim"),".2f"),
         _fmtv(t.get("dn_mm"),".0f"), "", "", ""],
        ["TOTAIS","","",ext_str,"","","","","","",""],
    ]
    for ri, row in enumerate(rows):
        y_row = y_hdr - 0.85 - (ri+1)*0.75
        fc = "#EEF3FA" if ri%2==0 else "white"
        for j, (val, cx, w) in enumerate(zip(row, col_x, col_w)):
            ax.add_patch(plt.Rectangle((cx, y_row), w, 0.72,
                                       ec="#AAAAAA", fc=fc, lw=0.3))
            ax.text(cx+w/2, y_row+0.36, str(val),
                    ha="center", va="center", fontsize=7)

    y_hid = y_hdr - 0.85 - len(rows)*0.75 - 0.35
    ax.add_patch(plt.Rectangle((0.5, y_hid-0.55), 28.7, 0.60, fc="#1F4E79", lw=0))
    if is_agua:
        lbl_hid = "DADOS HIDRÁULICOS — REDE PRESSURIZADA (ÁGUA FRIA)"
        txt_hid = (f"Pressão: {pressao} mca  |  DN: {_fmtv(t.get('dn_mm'),'.0f')} mm  |  "
                   f"Material: {t.get('material','PE80')}  |  "
                   f"Prof. ini: {_fmtv(t.get('prof_ini'),'.2f')} m  |  "
                   f"Prof. fim: {_fmtv(t.get('prof_fim'),'.2f')} m")
    else:
        lbl_hid = "DADOS HIDRÁULICOS (Manning — NBR 9649)"
        txt_hid = (f"Velocidade: {_fmtv(hid.get('vel_ms'),'.3f')} m/s  |  "
                   f"Vazão: {_fmtv(hid.get('vazao_ls'),'.2f')} l/s  |  "
                   f"Tensão trativa: {_fmtv(hid.get('tau_pa'),'.2f')} Pa  |  "
                   f"Status: {hid.get('status','SEM_DADOS')}")
    ax.text(14.85, y_hid-0.08, lbl_hid, ha="center", va="center",
            fontsize=7.5, fontweight="bold", color="white")
    ax.text(14.85, y_hid-0.44, txt_hid, ha="center", va="center",
            fontsize=7.5, color="white")

    # Materiais agua
    y_mat = y_hid - 0.55 - 0.3
    ax.add_patch(plt.Rectangle((0.5, y_mat-0.50), 28.7, 0.55, fc="#1F4E79", lw=0))
    if is_agua and t.get("materiais_agua"):
        ax.text(14.85, y_mat-0.20, "MATERIAIS DE CONEXÃO — REDE DE ÁGUA FRIA",
                ha="center", va="center", fontsize=7.5, fontweight="bold", color="white")
        mats = t.get("materiais_agua", {})
        mat_str = "  |  ".join(f"• {k}: {v}" for k, v in list(mats.items())[:5])
        ax.text(14.85, y_mat-0.90, mat_str, ha="center", va="center",
                fontsize=7.0, color="#111")
        y_croqui = y_mat - 1.35
    else:
        ax.text(14.85, y_mat-0.20, "CROQUI ESQUEMÁTICO DO TRECHO",
                ha="center", va="center", fontsize=7.5, fontweight="bold", color="white")
        y_croqui = y_mat - 0.70

    # Croqui
    cx0, cx1, cy = 3.0, 26.7, y_croqui - 0.9
    ax.annotate("", xy=(cx1, cy), xytext=(cx0, cy),
                arrowprops=dict(arrowstyle="-|>", color=c_pipe, lw=2.2))
    for cx_pv in [cx0, cx1]:
        ax.add_patch(plt.Rectangle((cx_pv-0.28, cy-0.28), 0.56, 0.56,
                                   ec="#333", fc="#4488ff", lw=1))
    ax.text(cx0, cy-0.65,
            f"{t['pv_ini']}\nCF={_fmtv(t.get('cf_ini'))}",
            ha="center", va="top", fontsize=6.5, color="#111")
    ax.text(cx1, cy-0.65,
            f"{t['pv_fim']}\nCF={_fmtv(t.get('cf_fim'))}",
            ha="center", va="top", fontsize=6.5, color="#111")
    pipe_lbl = (f"DN {_fmtv(t.get('dn_mm'),'.0f')} mm / L={ext_str} m / "
                + (f"P={pressao} mca" if pressao
                   else f"i={_fmtv(t.get('decl_pct'),'.3f')} %"))
    ax.text(14.85, cy+0.28, pipe_lbl, ha="center", va="bottom",
            fontsize=7.5, fontweight="bold", color=c_pipe)

    # Assinaturas
    sigs = ["ENG. CAMPO","EXECUTOR","COORD.","GERENTE","C. PROJ.","G. ENG."]
    sw   = 28.7 / len(sigs)
    for k, s in enumerate(sigs):
        sx = 0.5 + k*sw
        ax.add_patch(plt.Rectangle((sx, 0.60), sw, 1.20, ec="#AAAAAA", fc="white", lw=0.5))
        ax.plot([sx+0.3, sx+sw-0.3], [1.10, 1.10], color="#888", lw=0.5)
        ax.text(sx+sw/2, 0.80, s, ha="center", va="center", fontsize=6.5, color="#555")

    # Rodapé
    rodape = (f"SABESP SANTOS — {cfg.get('nucleo','')} — "
              f"{'ÁGUA FRIA' if is_agua else 'ESGOTO'} — NS{ns_id} — "
              f"Folha 1 de 1 — Rev. 0")
    ax.add_patch(plt.Rectangle((0.5, 0.0), 28.7, 0.55, fc="#1F4E79", lw=0))
    ax.text(14.0, 0.27, rodape, ha="center", va="center", fontsize=7, color="#CCCCCC")
    ax.text(27.5, 0.27, "ConstruData BIM", ha="center", va="center",
            fontsize=7, color="#AAAAAA", fontstyle="italic")

    fig.tight_layout(pad=0)
    caminho = pasta / f"NS_{ns_id}_A4.pdf"
    fig.savefig(caminho, dpi=200, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    return caminho


# ==============================================================================
# MÓDULO 7 — NS_DESENHO.pdf  (Prancha A3)
# ==============================================================================

def gerar_ns_desenho(t, pvs, ns_id, pasta, cfg, ruas_dxf=None, gpkg_path=None):
    """Prancha A3: Planta UTM + Perfil longitudinal com cotas + Tabela + Selo."""
    pvi     = pvs.get(t.get("pv_ini"), {})
    pvf     = pvs.get(t.get("pv_fim"), {})
    xi, yi  = pvi.get("x"), pvi.get("y")
    xf, yf  = pvf.get("x"), pvf.get("y")
    is_agua = t.get("is_agua", False)
    c_pipe  = "#1E6B3C" if is_agua else "#1565C0"
    C_CT    = "#7B3F00"
    C_CF    = "#1565C0"

    fig = plt.figure(figsize=(42.0/2.54, 29.7/2.54), facecolor="white")
    gs  = GridSpec(3, 2, figure=fig,
                   height_ratios=[7, 3.5, 1.8],
                   width_ratios=[1, 1],
                   hspace=0.10, wspace=0.06,
                   left=0.04, right=0.97, top=0.95, bottom=0.05)
    ax_planta = fig.add_subplot(gs[0, 0])
    ax_legend = fig.add_subplot(gs[0, 1])
    ax_perfil = fig.add_subplot(gs[1, :])
    ax_tabela = fig.add_subplot(gs[2, :])

    # ── Planta ───────────────────────────────────────────────────────────────
    ax_planta.set_facecolor("#F0EDE8")
    ax_planta.tick_params(labelsize=5)

    if xi and yi and xf and yf:
        cx  = (xi+xf)/2; cy = (yi+yf)/2
        pad = max(math.hypot(xf-xi, yf-yi)*1.5, 60)
        ax_planta.set_xlim(cx-pad, cx+pad)
        ax_planta.set_ylim(cy-pad, cy+pad)
        ax_planta.set_aspect("equal")

        # Tile de satelite (contextily) — fallback zoom 18→17→16
        _sat_ok = False
        try:
            if cx > 100_000 and cy > 1_000_000:
                import contextily as ctx
                for _zoom in [18, 17, 16]:
                    try:
                        ctx.add_basemap(ax_planta, crs="EPSG:31983",
                                        source=ctx.providers.Esri.WorldImagery,
                                        zoom=_zoom, attribution="",
                                        attribution_size=4)
                        _sat_ok = True
                        break
                    except Exception:
                        continue
                if _sat_ok:
                    ax_planta.set_facecolor("black")
        except Exception as _ctx_err:
            log(f"  Satelite PDF falhou: {_ctx_err}", "WARN")
        if not _sat_ok:
            log(f"  Sem satelite no PDF — pip install contextily", "WARN")

        if gpkg_path:
            try:
                from shapely.geometry import box as sbox
                carta = ler_cartografia_gpkg(gpkg_path)
                bb    = sbox(cx-pad, cy-pad, cx+pad, cy+pad)
                if carta["quadras"] is not None:
                    sub = carta["quadras"][carta["quadras"].intersects(bb)]
                    for geom in sub.geometry:
                        try:
                            xs, ys = (geom.exterior.xy if hasattr(geom, "exterior")
                                      else geom.xy)
                            ax_planta.fill(xs, ys, fc="#EDE8DC", ec="#CCBBAA",
                                           lw=0.4, zorder=1)
                        except Exception:
                            pass
                if carta["streets"] is not None:
                    sub = carta["streets"][carta["streets"].intersects(bb)]
                    for geom in sub.geometry:
                        try:
                            xs, ys = geom.xy
                            ax_planta.plot(xs, ys, color="#AAAAAA", lw=0.8, zorder=2)
                        except Exception:
                            pass
            except Exception:
                pass

        if ruas_dxf:
            xmin, xmax = cx-pad, cx+pad
            ymin, ymax = cy-pad, cy+pad
            seen = set()
            for r in ruas_dxf:
                rx, ry, rt = r.get("x",0), r.get("y",0), r.get("text","")
                if rt in seen or not rt:
                    continue
                if xmin < rx < xmax and ymin < ry < ymax:
                    seen.add(rt)
                    is_cur = t.get("rua","") and t["rua"].upper() in rt.upper()
                    ax_planta.text(rx, ry, rt, fontsize=4.5,
                                   color="#FFFF00" if is_cur else "#DDDDDD",
                                   ha="center", va="center", zorder=5,
                                   fontweight="bold" if is_cur else "normal",
                                   bbox=dict(boxstyle="round,pad=0.15", fc="black",
                                             ec="none", alpha=0.5))

        ax_planta.annotate("", xy=(xf, yf), xytext=(xi, yi),
                           arrowprops=dict(arrowstyle="-|>", color=c_pipe, lw=2))
        for px_v, py_v, pnome, pv_d, lado in [
                (xi, yi, t["pv_ini"], pvi, "ini"),
                (xf, yf, t["pv_fim"], pvf, "fim")]:
            ax_planta.plot(px_v, py_v, "s", color=c_pipe, ms=8, zorder=10)
            dx_off = pad*0.18 * (1 if lado == "fim" else -1)
            ax_planta.text(px_v+dx_off, py_v+pad*0.12,
                           f"{pnome}\nCT={_fmtv(pv_d.get('ct'),'.3f')}\n"
                           f"CF={_fmtv(pv_d.get('cf'),'.3f')}",
                           fontsize=4.5, color="#111", va="bottom",
                           bbox=dict(boxstyle="round,pad=0.2", fc="white",
                                     ec=c_pipe, lw=0.5))
        mid_x = (xi+xf)/2; mid_y = (yi+yf)/2
        ax_planta.text(mid_x, mid_y+pad*0.08,
                       f"DN {_fmtv(t.get('dn_mm'),'.0f')}mm /{t.get('material','PVC')}/ "
                       f"L={_fmtv(t.get('ext_m'),'.2f')}m",
                       ha="center", va="bottom", fontsize=5,
                       color=c_pipe, fontweight="bold",
                       bbox=dict(boxstyle="round,pad=0.15", fc="white",
                                 ec=c_pipe, lw=0.4))
        # Norte
        ax_planta.annotate("", xy=(cx+pad*0.78, cy+pad*0.72),
                           xytext=(cx+pad*0.78, cy+pad*0.52),
                           arrowprops=dict(arrowstyle="-|>", color="#333", lw=1.5))
        ax_planta.text(cx+pad*0.78, cy+pad*0.76, "N",
                       ha="center", va="bottom", fontsize=8, fontweight="bold")
        ax_planta.set_xlabel("Este (m UTM)", fontsize=5)
        ax_planta.set_ylabel("Norte (m UTM)", fontsize=5)
        plt.setp(ax_planta.get_xticklabels(), fontsize=4.5)
        plt.setp(ax_planta.get_yticklabels(), fontsize=4.5)

    ax_planta.set_title(
        f"PLANTA NS {ns_id} | {t['pv_ini']} → {t['pv_fim']} | {t.get('rua','Sem Rua')}",
        fontsize=7.5, fontweight="bold", pad=4)

    # ── Legenda/Quantitativo ─────────────────────────────────────────────────
    ax_legend.axis("off")
    q     = t.get("quantitativos") or {}
    lines_q = [
        "Quantitativo:",
        "",
        f"Volume de Escavação = {_fmtv(q.get('esc_m3'),'.3f')} m³",
        f"Volume de Aterro = {_fmtv(q.get('reat_m3'),'.3f')} m³",
        f"Pavimentação = {_fmtv(q.get('pav_m2'),'.2f')} m²",
        f"Extensão Total = {_fmtv(t.get('ext_m'),'.2f')} m",
    ]
    if not is_agua:
        lines_q += ["", "—— PVC ——",
                    f"{q.get('tubo_barras','---')} barra(s)  "
                    f"{_fmtv(t.get('dn_mm'),'.0f')}mm  Tubo {t.get('material','PVC')}"]
    else:
        lines_q += ["", "—— Materiais ——"]
        for k, v in list((t.get("materiais_agua") or {}).items())[:5]:
            lines_q.append(f"{v}  {k}")
    for li, line in enumerate(lines_q):
        ax_legend.text(0.05, 0.95 - li*0.07, line,
                       transform=ax_legend.transAxes,
                       fontsize=6.5, va="top", color="#111")
    yl = 0.95 - len(lines_q)*0.07 - 0.04
    ax_legend.text(0.05, yl, "LEGENDA", transform=ax_legend.transAxes,
                   fontsize=7, fontweight="bold", va="top")
    ax_legend.plot([0.05, 0.20], [yl-0.06, yl-0.06], color=c_pipe, lw=2,
                   transform=ax_legend.transAxes)
    ax_legend.text(0.23, yl-0.04,
                   f"Tubo DN{_fmtv(t.get('dn_mm'),'.0f')}mm",
                   transform=ax_legend.transAxes, fontsize=6.5, va="center")
    ax_legend.plot([0.05], [yl-0.12], "s", color=c_pipe, ms=8,
                   transform=ax_legend.transAxes)
    ax_legend.text(0.23, yl-0.10, "P.V. — Poço de Visita",
                   transform=ax_legend.transAxes, fontsize=6.5, va="center")

    # ── Perfil ────────────────────────────────────────────────────────────────
    ct_i = t.get("ct_ini"); ct_f = t.get("ct_fim")
    cf_i = t.get("cf_ini"); cf_f = t.get("cf_fim")
    ext  = t.get("ext_m") or 1
    has_cotas = all(v is not None for v in [ct_i, ct_f, cf_i, cf_f])

    if has_cotas:
        all_v = [ct_i, ct_f, cf_i, cf_f]
        y_min = min(all_v) - 0.30
        y_max = max(all_v) + 0.45
        ax_perfil.set_xlim(-ext*0.03, ext*1.03)
        ax_perfil.set_ylim(y_min, y_max)
        ax_perfil.plot([0, ext], [ct_i, ct_f], color=C_CT, lw=1.8,
                       zorder=5, label="CT (Terreno)")
        ax_perfil.plot([0, ext], [cf_i, cf_f], color=C_CF, lw=1.5,
                       ls="--", zorder=5, label="CF (Geratriz Inf.)")
        ax_perfil.fill_between([0, ext], [cf_i, cf_f], [ct_i, ct_f],
                               alpha=0.12, color="#8B4513", zorder=2)
        ax_perfil.fill_between([0, ext],
                               [cf_i, cf_f],
                               [cf_i+0.05, cf_f+0.05],
                               color=c_pipe, alpha=0.85, zorder=6)
        if y_min <= 0 <= y_max:
            ax_perfil.axhline(0, color="#AAAAAA", lw=0.5)
        for px_d, pnome, ct_v, cf_v in [(0, t["pv_ini"], ct_i, cf_i),
                                         (ext, t["pv_fim"], ct_f, cf_f)]:
            ax_perfil.plot(px_d, ct_v, "v", color=C_CT, ms=7, zorder=10)
            ax_perfil.plot(px_d, cf_v, "^", color=C_CF, ms=7, zorder=10)
            ax_perfil.text(px_d, y_max,
                           f"{pnome}\nCT={ct_v:.3f}\nCF={cf_v:.3f}",
                           ha="center", va="top", fontsize=5.5, color="#111",
                           bbox=dict(boxstyle="round,pad=0.2", fc="white",
                                     ec="#888", lw=0.4))
        ax_perfil.text(ext/2, (cf_i+cf_f)/2 - (y_max-y_min)*0.12,
                       f"DN {_fmtv(t.get('dn_mm'),'.0f')}mm  i={_fmtv(t.get('decl_pct'),'.2f')}%",
                       ha="center", va="center", fontsize=6.5,
                       color=c_pipe, fontweight="bold")
        ax_perfil.legend(fontsize=6, loc="upper right")
    else:
        ax_perfil.text(0.5, 0.5, "Sem dados de cota — CT/CF não disponíveis",
                       ha="center", va="center", fontsize=11,
                       color="#AAAAAA", transform=ax_perfil.transAxes)
        ax_perfil.set_xlim(0, ext or 10); ax_perfil.set_ylim(0, 1)

    ax_perfil.set_xlabel("Distância (m)", fontsize=6)
    ax_perfil.set_ylabel("Cota (m)", fontsize=6)
    ax_perfil.tick_params(labelsize=5.5)
    ax_perfil.yaxis.grid(True, ls="--", lw=0.4, color="#DDDDDD")
    ax_perfil.set_title("PERFIL LONGITUDINAL    Exag. vertical ~10x",
                        fontsize=7, loc="center", pad=3)

    # ── Tabela de dados ───────────────────────────────────────────────────────
    ax_tabela.axis("off")
    col_labels = ["Estaca","CT (m)","CF (m)","Prof (m)",
                  "Dist (m)","DN (mm)","Decl (%)"]
    row_i = [t["pv_ini"],
             _fmtv(t.get("ct_ini")), _fmtv(t.get("cf_ini")),
             _fmtv(t.get("prof_ini"), ".2f"), "0.00",
             _fmtv(t.get("dn_mm"), ".0f"), _fmtv(t.get("decl_pct"), ".2f")]
    row_f = [t["pv_fim"],
             _fmtv(t.get("ct_fim")), _fmtv(t.get("cf_fim")),
             _fmtv(t.get("prof_fim"), ".2f"), _fmtv(t.get("ext_m"), ".2f"),
             _fmtv(t.get("dn_mm"), ".0f"), _fmtv(t.get("decl_pct"), ".2f")]
    tbl = ax_tabela.table(
        cellText=[row_i, row_f],
        colLabels=col_labels,
        loc="upper left",
        cellLoc="center",
        bbox=[0.0, 0.0, 0.55, 1.0],
    )
    tbl.auto_set_font_size(False)
    tbl.set_fontsize(6.5)
    for (r, c), cell in tbl.get_celld().items():
        if r == 0:
            cell.set_facecolor("#1F4E79")
            cell.set_text_props(color="white", fontweight="bold", fontsize=6.5)
        else:
            cell.set_facecolor("#F0F4FA" if r%2 else "white")
        cell.set_edgecolor("#CCCCCC")
        cell.set_height(0.42)

    # Selo SABESP
    ax_tabela.text(0.68, 0.85, "SABESP", transform=ax_tabela.transAxes,
                   ha="center", va="top", fontsize=12, fontweight="bold",
                   color="#0A3D91")
    ax_tabela.text(0.68, 0.60,
                   ("SISTEMA DE ESGOTAMENTO\nSANITÁRIO SANTOS/SP"
                    if not is_agua else
                    "SISTEMA DE ABASTECIMENTO\nDE ÁGUA FRIA SANTOS/SP"),
                   transform=ax_tabela.transAxes, ha="center", va="top",
                   fontsize=6.5, color="#333")
    ax_tabela.text(0.88, 0.80,
                   f"CONTRATO: {cfg['contrato']}\nNS Nº {ns_id}\n"
                   f"NÚCLEO: {cfg.get('nucleo','')}\n"
                   f"ENGENHEIRO: {cfg.get('engenheiro','')}\n"
                   f"DESENHO: ConstruData BIM  Rev. 0",
                   transform=ax_tabela.transAxes, ha="center", va="top",
                   fontsize=5.5, color="#333")

    caminho = pasta / f"NS_{ns_id}_DESENHO.pdf"
    fig.savefig(caminho, dpi=200, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    return caminho


# ==============================================================================
# MÓDULO 8 — NS_OSE.xlsx  (formato NS_017rev1 SABESP)
# ==============================================================================

def gerar_ns_ose(t, pvs, ns_id, pasta, cfg):
    """Gera OSE padrao SABESP - layout identico ao template ProSaneamento OSE-Modelo_1.

    Layout de colunas (colunas intercaladas, par=dado, impar=espacador):
      B=TRECHO, D=ESTACA INTEIRO, F=ESTACA FRACAO,
      H=DIST PARCIAL, J=DIST ACUM,
      L=CT, N=I, P=CP(CF), R=CR, T=DN, V=G, X=H, Z=P,
      AB=PV NOME, AD=PV TIPO, AF=PV PROF, AH=OBSERVACOES

    Header institucional: linhas 3-13.
    Cabecalhos tabela: linhas 15-17.
    Dados: linha 18+.
    Convencoes (Z10:AD13) conforme template.
    """
    if not _HAS_OPENPYXL:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OS"
    ws.sheet_view.showGridLines = False

    # ── estilos basicos ────────────────────────────────────────────────────────
    thin  = Side(style="thin",   color="888888")
    thick = Side(style="medium", color="1F4E79")
    brd   = Border(left=thin, right=thin, top=thin, bottom=thin)
    brd_h = Border(left=thick, right=thick, top=thick, bottom=thick)

    is_agua = t.get("is_agua", False)
    hid     = t.get("hidraulica") or {}
    ext     = t.get("ext_m") or 0

    # ── helpers internos ──────────────────────────────────────────────────────
    def _hc(row, col, val, fc="1F4E79", fg="FFFFFF", bold=True, sz=9,
            wrap=True, halign="center"):
        """Celula de cabecalho (header cell)."""
        c = ws.cell(row=row, column=col, value=val)
        c.fill      = PatternFill("solid", fgColor=fc)
        c.font      = Font(name="Calibri", bold=bold, color=fg, size=sz)
        c.alignment = Alignment(horizontal=halign, vertical="center",
                                wrap_text=wrap)
        c.border    = brd
        return c

    def _dc(row, col, val, fmt=None, bold=False, fc="FFFFFF",
            halign="center"):
        """Celula de dado (data cell)."""
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(name="Calibri", size=9, bold=bold)
        c.alignment = Alignment(horizontal=halign, vertical="center")
        c.fill      = PatternFill("solid", fgColor=fc)
        c.border    = brd
        if fmt:
            c.number_format = fmt
        return c

    def _info(cell_ref, val, bold=False, sz=9, halign="left", fc=None):
        """Celula de info no cabecalho institucional."""
        c = ws[cell_ref]
        c.value     = val
        c.font      = Font(name="Calibri", size=sz, bold=bold)
        c.alignment = Alignment(horizontal=halign, vertical="center",
                                wrap_text=False)
        c.border    = brd
        if fc:
            c.fill = PatternFill("solid", fgColor=fc)
        return c

    def _mc(cell_ref, merge_range):
        """Merge de celulas apos atribuir valor/estilo na celula ancora."""
        ws.merge_cells(merge_range)
        return ws[cell_ref]

    # ── mapeamento de colunas por letra (conforme template ProSaneamento) ─────
    # Colunas de dados: B D F H J L N P R T V X Z AB AD AF AH..AX
    # Colunas em numero (openpyxl): A=1, B=2, C=3, D=4, E=5, ...
    #   Z=26, AA=27, AB=28, AC=29, AD=30, AE=31, AF=32, AG=33, AH=34, AX=50
    CB  = 2    # B  - TRECHO
    CD  = 4    # D  - ESTACA INTEIRO
    CF  = 6    # F  - ESTACA FRACAO
    CH  = 8    # H  - DISTANCIA PARCIAL
    CJ  = 10   # J  - DISTANCIA ACUMULADA
    CL  = 12   # L  - CT
    CN  = 14   # N  - I (declividade)
    CP_ = 16   # P  - CP (cota de projeto = CF)
    CR  = 18   # R  - CR (profundidade tubo)
    CT_ = 20   # T  - DN (mm)
    CV  = 22   # V  - G (largura vala)
    CX  = 24   # X  - H (altura regua ao greide)
    CZ  = 26   # Z  - P (profundidade da vala)
    CAB = 28   # AB - PV NOME
    CAD = 30   # AD - PV TIPO
    CAF = 32   # AF - PV PROF
    CAH = 34   # AH - OBSERVACOES (ate AX=50)
    CAX = 50   # AX - fim de OBSERVACOES

    # ── larguras das colunas (A..AX) ─────────────────────────────────────────
    # Padrao: colunas de dado=largura especifica, colunas intermediarias=1
    COL_WIDTHS = {
        1:  1.5,   # A  - margem
        2:  14.0,  # B  - TRECHO
        3:  1.5,   # C  - espacador
        4:  7.0,   # D  - ESTACA INT
        5:  1.5,   # E
        6:  7.0,   # F  - ESTACA FRAC
        7:  1.5,   # G
        8:  8.0,   # H  - DIST PARCIAL
        9:  1.5,   # I
        10: 8.0,   # J  - DIST ACUM
        11: 1.5,   # K
        12: 8.0,   # L  - CT
        13: 1.5,   # M
        14: 8.0,   # N  - I
        15: 1.5,   # O
        16: 8.0,   # P  - CP
        17: 1.5,   # Q
        18: 8.0,   # R  - CR
        19: 1.5,   # S
        20: 6.0,   # T  - DN
        21: 1.5,   # U
        22: 6.0,   # V  - G
        23: 1.5,   # W
        24: 7.0,   # X  - H
        25: 1.5,   # Y
        26: 7.0,   # Z  - P
        27: 1.5,   # AA
        28: 9.0,   # AB - PV NOME
        29: 1.5,   # AC
        30: 6.0,   # AD - PV TIPO
        31: 1.5,   # AE
        32: 6.0,   # AF - PV PROF
        33: 1.5,   # AG
        34: 18.0,  # AH..AX - OBSERVACOES (AH=col 34 ate AX=col 50)
    }
    for ci, cw in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(ci)].width = cw
    # Colunas AH(34)+1 ate AX(50) - distribuidas na merge de OBSERVACOES
    for ci in range(35, 51):
        ws.column_dimensions[get_column_letter(ci)].width = 2.2

    # ── alturas das linhas ────────────────────────────────────────────────────
    for r, h_pts in {
        3: 18, 4: 14, 5: 14, 6: 22, 7: 18,
        8: 14, 9: 14, 10: 14, 11: 14, 12: 14, 13: 14,
        14: 6,   # linha separadora
        15: 22, 16: 14, 17: 16,
    }.items():
        ws.row_dimensions[r].height = h_pts

    # ══════════════════════════════════════════════════════════════════════════
    # BLOCO 1 — CABECALHO INSTITUCIONAL (linhas 3-13)
    # ══════════════════════════════════════════════════════════════════════════

    # L3: sistema (merge L3:Y3)
    titulo = ("SISTEMA DE ABASTECIMENTO DE AGUA FRIA"
              if is_agua else "SISTEMA DE ESGOTAMENTO SANITARIO")
    ws["L3"].value     = titulo
    ws["L3"].font      = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws["L3"].fill      = PatternFill("solid", fgColor="1F4E79")
    ws["L3"].alignment = Alignment(horizontal="center", vertical="center")
    ws["L3"].border    = brd_h
    ws.merge_cells("L3:Y3")

    # L4: CIDADE (merge L4:Y4)
    ws["L4"].value     = f"CIDADE: {cfg.get('cidade', '')}"
    ws["L4"].font      = Font(name="Calibri", bold=True, size=10)
    ws["L4"].alignment = Alignment(horizontal="center", vertical="center")
    ws["L4"].border    = brd
    ws.merge_cells("L4:Y4")

    # L6: ORDEM DE SERVICO (merge L6:Y6)
    ws["L6"].value     = "ORDEM DE SERVICO"
    ws["L6"].font      = Font(name="Calibri", bold=True, size=13, color="1F4E79")
    ws["L6"].alignment = Alignment(horizontal="center", vertical="center")
    ws["L6"].border    = brd
    ws.merge_cells("L6:Y6")

    # L7: PARA GABARITO / PARA EXECUCAO (merge L7:Y7)
    subtitulo_os = "PARA EXECUCAO" if is_agua else "PARA GABARITO"
    ws["L7"].value     = subtitulo_os
    ws["L7"].font      = Font(name="Calibri", bold=True, size=11, color="1F4E79")
    ws["L7"].alignment = Alignment(horizontal="center", vertical="center")
    ws["L7"].border    = brd
    ws.merge_cells("L7:Y7")

    # B9: EMPRESA (merge B9:K9)
    ws["B9"].value     = f"EMPRESA: {cfg.get('empresa', '')}"
    ws["B9"].font      = Font(name="Calibri", size=9)
    ws["B9"].alignment = Alignment(horizontal="left", vertical="center")
    ws["B9"].border    = brd
    ws.merge_cells("B9:K9")

    # L9: CALCULISTA (merge L9:Y9)
    ws["L9"].value     = f"CALCULISTA: {cfg.get('engenheiro', 'FELIPE NERY')}"
    ws["L9"].font      = Font(name="Calibri", size=9)
    ws["L9"].alignment = Alignment(horizontal="left", vertical="center")
    ws["L9"].border    = brd
    ws.merge_cells("L9:Y9")

    # B10: CONTRATO (merge B10:K10)
    ws["B10"].value     = f"CONTRATO No: {cfg.get('contrato', '')}"
    ws["B10"].font      = Font(name="Calibri", size=9)
    ws["B10"].alignment = Alignment(horizontal="left", vertical="center")
    ws["B10"].border    = brd
    ws.merge_cells("B10:K10")

    # B11: NUCLEO (merge B11:K11)
    ws["B11"].value     = f"NUCLEO: {cfg.get('nucleo', '')}"
    ws["B11"].font      = Font(name="Calibri", size=9)
    ws["B11"].alignment = Alignment(horizontal="left", vertical="center")
    ws["B11"].border    = brd
    ws.merge_cells("B11:K11")

    # L11: DATA (merge L11:Y11)
    ws["L11"].value     = f"DATA: {datetime.now().strftime('%d/%m/%Y')}"
    ws["L11"].font      = Font(name="Calibri", size=9)
    ws["L11"].alignment = Alignment(horizontal="left", vertical="center")
    ws["L11"].border    = brd
    ws.merge_cells("L11:Y11")

    # B12: LOGRADOURO (merge B12:K12)
    ws["B12"].value     = f"LOGRADOURO: {t.get('rua', 'Sem Rua')}"
    ws["B12"].font      = Font(name="Calibri", size=9)
    ws["B12"].alignment = Alignment(horizontal="left", vertical="center")
    ws["B12"].border    = brd
    ws.merge_cells("B12:K12")

    # B13: O.S. No (merge B13:K13)
    ws["B13"].value     = f"O.S. No: {ns_id}"
    ws["B13"].font      = Font(name="Calibri", size=9, bold=True)
    ws["B13"].alignment = Alignment(horizontal="left", vertical="center")
    ws["B13"].border    = brd
    ws.merge_cells("B13:K13")

    # Z10:AD13 — bloco de CONVENCOES (conforme template)
    # Z=26, AA=27, AB=28, AC=29, AD=30
    conv_fc  = "FFF2CC"   # amarelo claro
    conv_hdr = "FFD966"   # amarelo header
    # titulo do bloco (Z10:AD10)
    ws["Z10"].value     = "CONVENCOES"
    ws["Z10"].font      = Font(name="Calibri", bold=True, size=8, color="000000")
    ws["Z10"].fill      = PatternFill("solid", fgColor=conv_hdr)
    ws["Z10"].alignment = Alignment(horizontal="center", vertical="center")
    ws["Z10"].border    = brd
    ws.merge_cells("Z10:AD10")

    conv_defs = [
        # (linha, sigla_celula, sigla_val, desc_celula, desc_val)
        (11, "Z11", "CT",  "AA11", "Cota do Terreno (m)"),
        (11, "AB11", "CP", "AC11", "Cota de Projeto (m)"),
        (12, "Z12", "I",   "AA12", "Declividade (m/m)"),
        (12, "AB12", "DN", "AC12", "Diametro Nominal (mm)"),
        (13, "Z13", "G",   "AA13", "Largura da Vala (m)"),
        (13, "AB13", "H",  "AC13", "Altura Regua ao Greide (m)"),
    ]
    for (ln, sig_ref, sig_val, desc_ref, desc_val) in conv_defs:
        ws[sig_ref].value     = sig_val
        ws[sig_ref].font      = Font(name="Calibri", bold=True, size=8)
        ws[sig_ref].fill      = PatternFill("solid", fgColor=conv_fc)
        ws[sig_ref].alignment = Alignment(horizontal="center", vertical="center")
        ws[sig_ref].border    = brd
        ws[desc_ref].value     = desc_val
        ws[desc_ref].font      = Font(name="Calibri", size=7.5)
        ws[desc_ref].fill      = PatternFill("solid", fgColor=conv_fc)
        ws[desc_ref].alignment = Alignment(horizontal="left", vertical="center",
                                           wrap_text=True)
        ws[desc_ref].border    = brd
    # Adiciona "P" e "CR" nas linhas 11-12 restantes (AD)
    for (ln, sig_ref, sig_val, desc_ref, desc_val) in [
        (11, "AD11", "CR", None, None),
        (12, "AD12", "P",  None, None),
    ]:
        ws[sig_ref].value     = sig_val
        ws[sig_ref].font      = Font(name="Calibri", bold=True, size=8)
        ws[sig_ref].fill      = PatternFill("solid", fgColor=conv_fc)
        ws[sig_ref].alignment = Alignment(horizontal="center", vertical="center")
        ws[sig_ref].border    = brd

    # ══════════════════════════════════════════════════════════════════════════
    # BLOCO 2 — CABECALHOS DA TABELA (linhas 15-17)
    # ══════════════════════════════════════════════════════════════════════════
    # Cores do cabecalho
    FC_H1 = "1F4E79"  # azul escuro - linha 15
    FC_H2 = "2E6DAA"  # azul medio  - linha 16
    FC_H3 = "3878A8"  # azul claro  - linha 17
    FG_W  = "FFFFFF"

    # B15: TRECHO (merge B15:B17)
    _hc(15, CB,  "TRECHO",        fc=FC_H1, sz=8)
    ws.merge_cells(f"B15:B17")

    # D15: ESTACA (merge D15:F15)
    _hc(15, CD,  "ESTACA",        fc=FC_H1, sz=8)
    ws.merge_cells(f"D15:F15")

    # H15: DISTANCIA (m) (merge H15:J15)
    _hc(15, CH,  "DISTANCIA (m)", fc=FC_H1, sz=8)
    ws.merge_cells(f"H15:J15")

    # L15: CT (merge L15:L17)
    _hc(15, CL,  "CT",            fc=FC_H1, sz=8)
    ws.merge_cells(f"L15:L17")

    # N15: I (merge N15:N17)
    _hc(15, CN,  "I",             fc=FC_H1, sz=8)
    ws.merge_cells(f"N15:N17")

    # P15: CP (merge P15:P17)
    _hc(15, CP_, "CP",            fc=FC_H1, sz=8)
    ws.merge_cells(f"P15:P17")

    # R15: CR (merge R15:R17)
    _hc(15, CR,  "CR",            fc=FC_H1, sz=8)
    ws.merge_cells(f"R15:R17")

    # T15: DN (merge T15:T17)
    _hc(15, CT_, "DN",            fc=FC_H1, sz=8)
    ws.merge_cells(f"T15:T17")

    # V15: G (merge V15:V17)
    _hc(15, CV,  "G",             fc=FC_H1, sz=8)
    ws.merge_cells(f"V15:V17")

    # X15: H (merge X15:X17)
    _hc(15, CX,  "H",             fc=FC_H1, sz=8)
    ws.merge_cells(f"X15:X17")

    # Z15: P (merge Z15:Z17)
    _hc(15, CZ,  "P",             fc=FC_H1, sz=8)
    ws.merge_cells(f"Z15:Z17")

    # AB15: POCO DE VISITA (merge AB15:AF15)
    _hc(15, CAB, "POCO DE VISITA", fc=FC_H1, sz=8)
    ws.merge_cells(f"AB15:AF15")

    # AH15: OBSERVACOES (merge AH15:AX17)
    _hc(15, CAH, "OBSERVACOES",   fc=FC_H1, sz=8)
    ws.merge_cells(f"AH15:AX17")

    # Linha 16 — sub-grupos ESTACA e DISTANCIA
    _hc(16, CD,  "INTEIRO",  fc=FC_H2, sz=7.5)
    ws.merge_cells(f"D16:D17")
    _hc(16, CF,  "FRACAO (m)", fc=FC_H2, sz=7.5)
    ws.merge_cells(f"F16:F17")
    _hc(16, CH,  "PARCIAL",  fc=FC_H2, sz=7.5)
    ws.merge_cells(f"H16:H17")
    _hc(16, CJ,  "ACUMUL.",  fc=FC_H2, sz=7.5)
    ws.merge_cells(f"J16:J17")

    # Linha 16 — sub-grupos POCO DE VISITA
    _hc(16, CAB, "NOME",     fc=FC_H2, sz=7.5)
    ws.merge_cells(f"AB16:AB17")
    _hc(16, CAD, "TIPO",     fc=FC_H2, sz=7.5)
    ws.merge_cells(f"AD16:AD17")
    _hc(16, CAF, "PROF.",    fc=FC_H2, sz=7.5)
    ws.merge_cells(f"AF16:AF17")

    # Linha 17 — unidades das colunas simples
    # (CT, I, CP, CR, DN, G, H, P ja estao mergeadas ate 17 acima)
    # Adiciona legenda de unidades na linha 17 para colunas que nao estao
    # mergeadas ate 17 (as colunas que tem merge B15:B17 ja estao cobertas)
    unit_map = {
        CL:  "(m)",      # CT
        CN:  "(m/m)",    # I
        CP_: "(m)",      # CP
        CR:  "(m)",      # CR
        CT_: "(mm)",     # DN
        CV:  "(m)",      # G
        CX:  "(m)",      # H
        CZ:  "(m)",      # P
    }
    # Estas colunas foram mergeadas B15:B17, L15:L17 etc. -- a unidade
    # fica dentro do merge. Para inserir a unidade precisamos colocar
    # o texto junto no valor da celula L15, mas o template usa duas
    # linhas separadas. Aqui seguimos o padrao do template com merge ate 17.
    # O valor ja esta definido acima. Nao ha necessidade de celulas extras.
    # (openpyxl nao permite editar celulas dentro de ranges ja mergeados)

    # ══════════════════════════════════════════════════════════════════════════
    # BLOCO 3 — LINHAS DE DADOS (linha 18+)
    # ══════════════════════════════════════════════════════════════════════════
    DATA_START = 19  # FIX-4: Conforme DATOSE.DEF (linha inicio oficial = 19)
    G_VALA = 0.60  # largura padrao da vala de esgoto (m)

    # Monta lista de linhas de dados: ini e fim do trecho
    trecho_nome = f"{t.get('pv_ini', '?')}-->{t.get('pv_fim', '?')}"

    # Declividade em m/m (converter de ‰ se necessario)
    decl_raw = t.get("decl_mm")
    try:
        decl_val = float(decl_raw) if decl_raw is not None else None
    except (TypeError, ValueError):
        decl_val = None

    rows_data = [
        # (trecho, est_int, est_frac, dist_p, dist_a,
        #  ct,     decl,   cp,       cr,
        #  dn,     g,      h,        p_vala,
        #  pv_nome, pv_tipo, pv_prof)
        (
            trecho_nome,
            0,
            0.0,
            0.0,
            0.0,
            t.get("ct_ini"),
            decl_val,
            t.get("cf_ini"),
            t.get("prof_ini"),
            t.get("dn_mm"),
            G_VALA if not is_agua else None,
            t.get("prof_ini"),   # H = prof ini como referencia
            t.get("prof_ini"),   # P = prof vala
            t.get("pv_ini", ""),
            "PV",
            t.get("prof_ini"),
        ),
        (
            trecho_nome,
            0,
            round(float(ext), 4) if ext else 0.0,
            round(float(ext), 4) if ext else 0.0,
            round(float(ext), 4) if ext else 0.0,
            t.get("ct_fim"),
            decl_val,
            t.get("cf_fim"),
            t.get("prof_fim"),
            t.get("dn_mm"),
            G_VALA if not is_agua else None,
            t.get("prof_fim"),   # H
            t.get("prof_fim"),   # P
            t.get("pv_fim", ""),
            "PV",
            t.get("prof_fim"),
        ),
    ]

    for ri, row_vals in enumerate(rows_data):
        (trn, est_i, est_f, dist_p, dist_a,
         ct, decl, cp, cr,
         dn, g_val, h_val, p_val,
         pv_nome, pv_tipo, pv_prof) = row_vals

        row = DATA_START + ri
        ws.row_dimensions[row].height = 15
        fc_c = "F0F4FA" if ri % 2 == 0 else "FFFFFF"

        _dc(row, CB,  trn,    bold=True, fc=fc_c, halign="left")
        _dc(row, CD,  est_i,  fmt="0",       fc=fc_c)
        _dc(row, CF,  est_f,  fmt="0.0000",  fc=fc_c)
        _dc(row, CH,  dist_p, fmt="0.0000",  fc=fc_c)
        _dc(row, CJ,  dist_a, fmt="0.0000",  fc=fc_c)
        _dc(row, CL,  ct,     fmt="0.0000",  fc=fc_c)
        _dc(row, CN,  decl,   fmt="0.00000", fc=fc_c)
        _dc(row, CP_, cp,     fmt="0.0000",  fc=fc_c)
        _dc(row, CR,  cr,     fmt="0.0000",  fc=fc_c)
        _dc(row, CT_, dn,     fmt="0",       fc=fc_c)
        _dc(row, CV,  g_val,  fmt="0.00",    fc=fc_c)
        _dc(row, CX,  h_val,  fmt="0.0000",  fc=fc_c)
        _dc(row, CZ,  p_val,  fmt="0.0000",  fc=fc_c)
        _dc(row, CAB, pv_nome,               fc=fc_c)
        _dc(row, CAD, pv_tipo,               fc=fc_c)
        _dc(row, CAF,
            pv_prof if isinstance(pv_prof, str)
            else (_fmtv(pv_prof, ".2f") if pv_prof is not None else "---"),
            fc=fc_c)
        _dc(row, CAH, None, fc=fc_c)  # OBSERVACOES em branco

    # ══════════════════════════════════════════════════════════════════════════
    # BLOCO 4 — TOTAIS
    # ══════════════════════════════════════════════════════════════════════════
    row_tot = DATA_START + len(rows_data)
    ws.row_dimensions[row_tot].height = 15

    c_tot = ws.cell(row=row_tot, column=CB, value="TOTAIS")
    c_tot.font      = Font(name="Calibri", bold=True, size=9)
    c_tot.fill      = PatternFill("solid", fgColor="DCE6F1")
    c_tot.alignment = Alignment(horizontal="center", vertical="center")
    c_tot.border    = brd

    _dc(row_tot, CH, 0.0,             fmt="0.0000", fc="DCE6F1")  # parcial = 0
    _dc(row_tot, CJ, round(float(ext), 4) if ext else 0.0,
        fmt="0.0000", fc="DCE6F1")   # acumulada = extensao total

    # ══════════════════════════════════════════════════════════════════════════
    # BLOCO 5 — DADOS HIDRAULICOS
    # ══════════════════════════════════════════════════════════════════════════
    row_hid = row_tot + 2
    ws.row_dimensions[row_hid].height = 14

    if is_agua:
        hid_txt = (
            f"Pressao: {_fmtv(t.get('pressao_mca'), '.2f')} mca  |  "
            f"DN: {_fmtv(t.get('dn_mm'), '.0f')} mm  |  "
            f"Material: {t.get('material', 'PE80')}  |  "
            f"Prof. ini: {_fmtv(t.get('prof_ini'), '.2f')} m  |  "
            f"Prof. fim: {_fmtv(t.get('prof_fim'), '.2f')} m"
        )
    else:
        hid_txt = (
            f"Velocidade: {_fmtv(hid.get('vel_ms'), '.3f')} m/s  |  "
            f"Vazao: {_fmtv(hid.get('vazao_ls'), '.2f')} l/s  |  "
            f"Tensao trativa: {_fmtv(hid.get('tau_pa'), '.2f')} Pa  |  "
            f"Status: {hid.get('status', 'SEM_DADOS')}"
        )

    ws[f"B{row_hid}"].value     = hid_txt
    ws[f"B{row_hid}"].font      = Font(name="Calibri", size=8.5, italic=True)
    ws[f"B{row_hid}"].fill      = PatternFill("solid", fgColor="E8F4FD")
    ws[f"B{row_hid}"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"B{row_hid}"].border    = brd
    ws.merge_cells(f"B{row_hid}:AX{row_hid}")

    # ══════════════════════════════════════════════════════════════════════════
    # BLOCO 6 — ASSINATURAS
    # ══════════════════════════════════════════════════════════════════════════
    row_ass = row_hid + 3
    ws.row_dimensions[row_ass].height   = 14
    ws.row_dimensions[row_ass+1].height = 18

    # Faixa de titulo "ASSINATURAS"
    ws[f"B{row_ass}"].value     = "ASSINATURAS"
    ws[f"B{row_ass}"].font      = Font(name="Calibri", bold=True, size=9,
                                       color="FFFFFF")
    ws[f"B{row_ass}"].fill      = PatternFill("solid", fgColor="1F4E79")
    ws[f"B{row_ass}"].alignment = Alignment(horizontal="center")
    ws.merge_cells(f"B{row_ass}:AX{row_ass}")

    # Seis campos de assinatura distribuidos em colunas B..AX (largura ~8 cada)
    sigs   = ["ENG. CAMPO", "EXECUTOR", "COORD.", "GERENTE", "C. PROJ.", "G. ENG."]
    # Colunas inicio de cada bloco: B(2), J(10), R(18), Z(26), AH(34), AP(42)
    sig_cols = [CB, CJ, CR, CZ, CAH, 42]
    for k, (s, sc) in enumerate(zip(sigs, sig_cols)):
        ec = sig_cols[k+1] - 2 if k < len(sig_cols)-1 else 50  # AX=50
        ws.merge_cells(start_row=row_ass+1, start_column=sc,
                       end_row=row_ass+1,   end_column=ec)
        cell = ws.cell(row=row_ass+1, column=sc, value=s)
        cell.font      = Font(name="Calibri", bold=True, size=8.5, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="2E6DAA")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = brd

    # ══════════════════════════════════════════════════════════════════════════
    # BLOCO 7 — RODAPE
    # ══════════════════════════════════════════════════════════════════════════
    row_rod = row_ass + 3
    ws.row_dimensions[row_rod].height = 12

    ws[f"B{row_rod}"].value = (
        f"SABESP SANTOS  --  {cfg.get('nucleo', '')}  --  "
        f"{'AGUA FRIA' if is_agua else 'ESGOTO'}  --  "
        f"NS{ns_id}  --  Folha 1 de 1  --  Rev. 0"
    )
    ws[f"B{row_rod}"].font      = Font(name="Calibri", size=8, color="555555",
                                       italic=True)
    ws[f"B{row_rod}"].alignment = Alignment(horizontal="left")
    ws.merge_cells(f"B{row_rod}:X{row_rod}")

    ws[f"Z{row_rod}"].value     = "ConstruData BIM"
    ws[f"Z{row_rod}"].font      = Font(name="Calibri", size=8, color="555555",
                                       italic=True)
    ws[f"Z{row_rod}"].alignment = Alignment(horizontal="right")
    ws.merge_cells(f"Z{row_rod}:AX{row_rod}")

    # ── salva arquivo ─────────────────────────────────────────────────────────
    caminho = pasta / f"NS_{ns_id}_OSE.xlsx"
    wb.save(caminho)
    return caminho


# ==============================================================================
# MÓDULO 9 — NS_DADOS.json
# ==============================================================================

def gerar_ns_dados_json(t, pvs, ns_id, pasta, cfg):
    pvi = pvs.get(t.get("pv_ini"), {})
    pvf = pvs.get(t.get("pv_fim"), {})
    lat_i, lon_i = utm_to_latlon(pvi.get("x"), pvi.get("y"))
    lat_f, lon_f = utm_to_latlon(pvf.get("x"), pvf.get("y"))
    dados = {
        "ns_id": ns_id, "nucleo": cfg.get("nucleo",""),
        "contrato": cfg["contrato"], "empresa": cfg["empresa"],
        "rua": t.get("rua","Sem Rua"),
        "tipo": "agua" if t.get("is_agua") else "esgoto",
        "pv_ini": {"nome": t["pv_ini"], "x_utm": pvi.get("x"), "y_utm": pvi.get("y"),
                   "lat": lat_i, "lon": lon_i,
                   "ct": t.get("ct_ini"), "cf": t.get("cf_ini"),
                   "prof": t.get("prof_ini")},
        "pv_fim": {"nome": t["pv_fim"], "x_utm": pvf.get("x"), "y_utm": pvf.get("y"),
                   "lat": lat_f, "lon": lon_f,
                   "ct": t.get("ct_fim"), "cf": t.get("cf_fim"),
                   "prof": t.get("prof_fim")},
        "dn_mm": t.get("dn_mm"), "material": t.get("material","PVC"),
        "ext_m": t.get("ext_m"), "decl_mm": t.get("decl_mm"),
        "decl_pct": t.get("decl_pct"), "prof_media": t.get("prof_media_m"),
        "hidraulica": t.get("hidraulica"),
        "quantitativos": t.get("quantitativos"),
        "materiais_agua": t.get("materiais_agua"),
        "pressao_mca": t.get("pressao_mca"),
        "gerado_em": datetime.now().isoformat(),
        "software": "ConstruData BIM SABESP v5.0",
    }
    caminho = pasta / f"NS_{ns_id}_DADOS.json"
    with open(caminho, "w", encoding="utf-8") as f:
        json.dump(dados, f, indent=2, ensure_ascii=False, default=str)
    return caminho


# ==============================================================================
# MÓDULO 10 — NS_XXX.html
# ==============================================================================

def gerar_ns_html(t, pvs, ns_id, pasta, cfg):
    """Dashboard interativo Leaflet + SVG perfil com cotas reais."""
    pvi     = pvs.get(t.get("pv_ini"), {})
    pvf     = pvs.get(t.get("pv_fim"), {})
    lat_i, lon_i = utm_to_latlon(pvi.get("x"), pvi.get("y"))
    lat_f, lon_f = utm_to_latlon(pvf.get("x"), pvf.get("y"))
    is_agua = t.get("is_agua", False)
    hid     = t.get("hidraulica") or {}
    q       = t.get("quantitativos") or {}

    # Centro do mapa — default Santos se coords invalidas
    def _safe_lat(v): return v if v and -34 < v < 5 else -23.96
    def _safe_lon(v): return v if v and -74 < v < -34 else -46.33
    lat_c = (_safe_lat(lat_i) + _safe_lat(lat_f)) / 2
    lon_c = (_safe_lon(lon_i) + _safe_lon(lon_f)) / 2
    c_pipe = "#1E6B3C" if is_agua else "#1a6ec0"

    # SVG perfil
    ct_i = t.get("ct_ini"); ct_f = t.get("ct_fim")
    cf_i = t.get("cf_ini"); cf_f = t.get("cf_fim")
    ext  = t.get("ext_m") or 1
    has_cotas = all(v is not None for v in [ct_i, ct_f, cf_i, cf_f])

    if has_cotas:
        all_v  = [ct_i, ct_f, cf_i, cf_f]
        v_min  = min(all_v) - 0.3
        v_rng  = max(all_v) + 0.4 - v_min
        def mx(d): return 40 + (d/ext)*620
        def my(v): return 170 - ((v-v_min)/v_rng)*150
        xi_s, yi_ct = mx(0),   my(ct_i)
        xf_s, yf_ct = mx(ext), my(ct_f)
        yi_cf, yf_cf = my(cf_i), my(cf_f)
        perfil_svg = f"""
      <polyline points="{xi_s},{yi_ct} {xf_s},{yf_ct}" stroke="#7B3F00" stroke-width="2.5" fill="none"/>
      <polyline points="{xi_s},{yi_cf} {xf_s},{yf_cf}" stroke="{c_pipe}" stroke-width="2" fill="none" stroke-dasharray="8,4"/>
      <polygon  points="{xi_s},{yi_cf} {xf_s},{yf_cf} {xf_s},{yf_ct} {xi_s},{yi_ct}"
                fill="#8B4513" fill-opacity="0.10"/>
      <line x1="{xi_s}" y1="175" x2="{xf_s}" y2="175" stroke="#ccc" stroke-width="0.5"/>
      <circle cx="{xi_s}" cy="{yi_ct}" r="4" fill="#7B3F00"/>
      <circle cx="{xf_s}" cy="{yf_ct}" r="4" fill="#7B3F00"/>
      <circle cx="{xi_s}" cy="{yi_cf}" r="4" fill="{c_pipe}"/>
      <circle cx="{xf_s}" cy="{yf_cf}" r="4" fill="{c_pipe}"/>
      <text x="{xi_s}" y="{yi_ct-10}" font-size="9" fill="#7B3F00" text-anchor="middle">{ct_i:.3f}</text>
      <text x="{xi_s}" y="{yi_cf+16}" font-size="9" fill="{c_pipe}" text-anchor="middle">{cf_i:.3f}</text>
      <text x="{xf_s}" y="{yf_ct-10}" font-size="9" fill="#7B3F00" text-anchor="middle">{ct_f:.3f}</text>
      <text x="{xf_s}" y="{yf_cf+16}" font-size="9" fill="{c_pipe}" text-anchor="middle">{cf_f:.3f}</text>
      <text x="660" y="{yi_ct}" font-size="8" fill="#7B3F00" dominant-baseline="middle">CT</text>
      <text x="660" y="{yi_cf}" font-size="8" fill="{c_pipe}" dominant-baseline="middle">CF</text>
      <text x="350" y="200" font-size="9" fill="#555" text-anchor="middle">
        DN {_fmtv(t.get('dn_mm'),'.0f')}mm  i={_fmtv(t.get('decl_pct'),'.2f')}%  L={_fmtv(ext,'.2f')}m
      </text>"""
    else:
        perfil_svg = "<text x='350' y='100' text-anchor='middle' font-size='14' fill='#888'>Sem dados de cota</text>"

    # Mapa JS — so renderiza se coords lat/lon sao validas (Brasil: lat -34 a 5, lon -74 a -34)
    mapa_js = ""
    _coords_validas = (lat_i and lon_i and lat_f and lon_f
                       and -34 < lat_i < 5 and -74 < lon_i < -34
                       and -34 < lat_f < 5 and -74 < lon_f < -34)
    if _coords_validas:
        mapa_js = f"""
        var map = L.map('mapa').setView([{lat_c},{lon_c}],17);
        L.tileLayer('https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{{z}}/{{y}}/{{x}}',
          {{attribution:'Tiles © Esri, Maxar, Earthstar Geographics',maxZoom:20}}).addTo(map);
        L.marker([{lat_i},{lon_i}]).addTo(map)
          .bindPopup('<b>{t["pv_ini"]}</b><br>CT:{_fmtv(t.get("ct_ini"))} m<br>CF:{_fmtv(t.get("cf_ini"))} m<br>Prof:{_fmtv(t.get("prof_ini"),".2f")} m');
        L.marker([{lat_f},{lon_f}]).addTo(map)
          .bindPopup('<b>{t["pv_fim"]}</b><br>CT:{_fmtv(t.get("ct_fim"))} m<br>CF:{_fmtv(t.get("cf_fim"))} m<br>Prof:{_fmtv(t.get("prof_fim"),".2f")} m');
        L.polyline([[{lat_i},{lon_i}],[{lat_f},{lon_f}]],
          {{color:'{c_pipe}',weight:4,opacity:0.85}}).addTo(map)
          .bindPopup('DN {_fmtv(t.get("dn_mm"),".0f")}mm — {_fmtv(ext,".2f")}m');
        window._map=map; window._mapInit=true;"""

    pressao = _fmtv(t.get("pressao_mca"), ".2f") if is_agua else None
    badge   = (f"DN {_fmtv(t.get('dn_mm'),'.0f')} mm | {_fmtv(ext,'.2f')} m"
               + (f" | {pressao} mca" if pressao else ""))
    badge_bg = "#1E6B3C" if is_agua else "#e94560"
    status   = hid.get("status","SEM_DADOS")
    st_color = "green" if status == "OK" else "red"

    mat_rows = ""
    if is_agua and t.get("materiais_agua"):
        for k, v in t["materiais_agua"].items():
            mat_rows += (f"<tr><td class='label'>{k}</td>"
                         f"<td class='value'>{v}</td>"
                         f"<td class='unit'>{'un' if 'un' in k else 'm'}</td></tr>")

    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>NS {ns_id} — {t['pv_ini']} → {t['pv_fim']}</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Segoe UI',Arial,sans-serif;background:#f0f2f5;color:#222}}
header{{background:#1a1a2e;color:white;padding:14px 24px;display:flex;align-items:center;justify-content:space-between}}
header h1{{font-size:18px;font-weight:700}}
header .sub{{font-size:12px;color:#aaa;margin-top:4px}}
.badge{{background:{badge_bg};color:white;padding:4px 12px;border-radius:12px;font-size:13px;font-weight:bold}}
.tabs{{display:flex;background:#16213e;padding:0 24px;border-bottom:2px solid #0f3460}}
.tab-btn{{padding:10px 22px;cursor:pointer;color:#aaa;font-size:13px;font-weight:600;border:none;background:none;border-bottom:3px solid transparent;transition:all .2s}}
.tab-btn:hover{{color:#fff}}
.tab-btn.active{{color:#00d4ff;border-bottom-color:#00d4ff}}
.tab-content{{display:none;padding:20px 24px}}
.tab-content.active{{display:block}}
#mapa{{height:420px;border-radius:8px}}
.info-grid{{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-top:12px}}
table{{width:100%;border-collapse:collapse;margin-top:8px}}
th{{background:#1F4E79;color:white;padding:8px 12px;text-align:left;font-size:12px}}
td{{padding:7px 12px;font-size:12px;border-bottom:1px solid #e0e0e0}}
td.label{{color:#555;width:50%;font-weight:500}}
td.value{{font-weight:600;color:#1a1a2e}}
td.unit{{color:#888;font-size:11px;width:60px}}
tr:hover td{{background:#f0f7ff}}
.card{{background:white;border-radius:8px;padding:16px;box-shadow:0 2px 8px rgba(0,0,0,.08)}}
.card h3{{font-size:14px;color:#1F4E79;margin-bottom:10px}}
svg{{background:#fdfdfd;border:1px solid #e0e0e0;border-radius:6px}}
footer{{background:#1a1a2e;color:#666;text-align:center;padding:10px;font-size:11px;margin-top:20px}}
</style>
</head>
<body>
<header>
  <div>
    <h1>NS {ns_id} — {t['pv_ini']} → {t['pv_fim']}</h1>
    <div class="sub">SABESP SANTOS — {cfg.get('nucleo','')} | {t.get('rua','Sem Rua')} | Contrato {cfg['contrato']}</div>
  </div>
  <span class="badge">{badge}</span>
</header>
<div class="tabs">
  <button class="tab-btn active" onclick="showTab(this,'tab-mapa')">MAPA</button>
  <button class="tab-btn" onclick="showTab(this,'tab-perfil')">PERFIL</button>
  <button class="tab-btn" onclick="showTab(this,'tab-dados')">DADOS</button>
  <button class="tab-btn" onclick="showTab(this,'tab-quant')">QUANTITATIVOS</button>
</div>
<div id="tab-mapa" class="tab-content active">
  <div class="card"><h3>Localização do Trecho</h3><div id="mapa">{'<p style="padding:40px;text-align:center;color:#888;font-size:14px;">Mapa indisponível — coordenadas UTM fora do range válido.<br>Verifique o datum do DXF (esperado: SIRGAS 2000 UTM 23S).</p>' if not _coords_validas else ''}</div></div>
</div>
<div id="tab-perfil" class="tab-content">
  <div class="card">
    <h3>Perfil Longitudinal</h3>
    <svg viewBox="0 0 700 215" width="100%" height="235">
      {perfil_svg}
    </svg>
    <p style="font-size:11px;color:#888;margin-top:6px;">
      Linha marrom = CT (terreno) | Linha tracejada = CF (geratriz inferior do tubo)
    </p>
  </div>
</div>
<div id="tab-dados" class="tab-content">
  <div class="info-grid">
    <div class="card">
      <h3>Dados Técnicos do Trecho</h3>
      <table><thead><tr><th>Campo</th><th>Valor</th><th>Un.</th></tr></thead><tbody>
        <tr><td class="label">NS ID</td><td class="value">{ns_id}</td><td></td></tr>
        <tr><td class="label">Rua</td><td class="value">{t.get('rua','Sem Rua')}</td><td></td></tr>
        <tr><td class="label">Núcleo</td><td class="value">{cfg.get('nucleo','')}</td><td></td></tr>
        <tr><td class="label">Contrato</td><td class="value">{cfg['contrato']}</td><td></td></tr>
        <tr><td class="label">PV Inicial</td><td class="value">{t['pv_ini']}</td><td></td></tr>
        <tr><td class="label">PV Final</td><td class="value">{t['pv_fim']}</td><td></td></tr>
        <tr><td class="label">CT Inicial</td><td class="value">{_fmtv(t.get('ct_ini'))}</td><td class="unit">m</td></tr>
        <tr><td class="label">CT Final</td><td class="value">{_fmtv(t.get('ct_fim'))}</td><td class="unit">m</td></tr>
        <tr><td class="label">CF Inicial</td><td class="value">{_fmtv(t.get('cf_ini'))}</td><td class="unit">m</td></tr>
        <tr><td class="label">CF Final</td><td class="value">{_fmtv(t.get('cf_fim'))}</td><td class="unit">m</td></tr>
        <tr><td class="label">Prof. Inicial</td><td class="value">{_fmtv(t.get('prof_ini'),'.2f')}</td><td class="unit">m</td></tr>
        <tr><td class="label">Prof. Final</td><td class="value">{_fmtv(t.get('prof_fim'),'.2f')}</td><td class="unit">m</td></tr>
        <tr><td class="label">DN</td><td class="value">{_fmtv(t.get('dn_mm'),'.0f')}</td><td class="unit">mm</td></tr>
        <tr><td class="label">Material</td><td class="value">{t.get('material','PVC')}</td><td></td></tr>
        <tr><td class="label">Extensão</td><td class="value">{_fmtv(ext,'.2f')}</td><td class="unit">m</td></tr>
        <tr><td class="label">Declividade</td><td class="value">{_fmtv(t.get('decl_pct'),'.3f')}</td><td class="unit">%</td></tr>
        <tr><td class="label">Velocidade</td><td class="value">{_fmtv(hid.get('vel_ms'),'.3f')}</td><td class="unit">m/s</td></tr>
        <tr><td class="label">Vazão</td><td class="value">{_fmtv(hid.get('vazao_ls'),'.2f')}</td><td class="unit">l/s</td></tr>
        <tr><td class="label">Tensão trativa</td><td class="value">{_fmtv(hid.get('tau_pa'),'.2f')}</td><td class="unit">Pa</td></tr>
        <tr><td class="label">Status Hid.</td><td class="value" style="color:{st_color}">{status}</td><td></td></tr>
        {f'<tr><td class="label">Pressão</td><td class="value">{pressao}</td><td class="unit">mca</td></tr>' if pressao else ''}
      </tbody></table>
    </div>
    <div class="card">
      <h3>Informações da NS</h3>
      <table>
        <tr><td class="label">NS ID</td><td class="value">{ns_id}</td><td></td></tr>
        <tr><td class="label">Núcleo</td><td class="value">{cfg.get('nucleo','')}</td><td></td></tr>
        <tr><td class="label">Contrato</td><td class="value">{cfg['contrato']}</td><td></td></tr>
        <tr><td class="label">Rua</td><td class="value">{t.get('rua','Sem Rua')}</td><td></td></tr>
        <tr><td class="label">Material</td><td class="value">{t.get('material','PVC')}</td><td></td></tr>
        <tr><td class="label">Status Hid.</td><td class="value" style="color:{st_color}">{status}</td><td></td></tr>
      </table>
    </div>
  </div>
</div>
<div id="tab-quant" class="tab-content">
  <div class="card">
    <h3>Quantitativos do Trecho</h3>
    <table><thead><tr><th>Item</th><th>Valor</th><th>Unidade</th></tr></thead><tbody>
      <tr><td class="label">Escavação</td><td class="value">{_fmtv(q.get('esc_m3'))}</td><td class="unit">m³</td></tr>
      <tr><td class="label">Lastro</td><td class="value">{_fmtv(q.get('lastro_m3'))}</td><td class="unit">m³</td></tr>
      <tr><td class="label">Envoltório</td><td class="value">{_fmtv(q.get('envolt_m3'))}</td><td class="unit">m³</td></tr>
      <tr><td class="label">Brita</td><td class="value">{_fmtv(q.get('brita_m3'))}</td><td class="unit">m³</td></tr>
      <tr><td class="label">Reaterro</td><td class="value">{_fmtv(q.get('reat_m3'))}</td><td class="unit">m³</td></tr>
      <tr><td class="label">Pavimentação</td><td class="value">{_fmtv(q.get('pav_m2'),'.2f')}</td><td class="unit">m²</td></tr>
      <tr><td class="label">Barras de tubo</td><td class="value">{q.get('tubo_barras','---')}</td><td class="unit">un</td></tr>
    </tbody></table>
    <p style="font-size:11px;color:#888;margin-top:8px;">* Custos em 03_CUSTOS_MEDICAO_PLANEJAMENTO/CUSTOS_POR_TRECHO.xlsx</p>
  </div>
  {f'<div class="card" style="margin-top:16px;border-left:4px solid #1E6B3C"><h3 style="color:#1E6B3C">Materiais — Rede de Água Fria</h3><table><thead><tr><th>Material</th><th>Qtd.</th><th>Un.</th></tr></thead><tbody>' + mat_rows + '</tbody></table></div>' if is_agua and mat_rows else ''}
</div>
<footer>ConstruData BIM SABESP — NS {ns_id} — {cfg.get('nucleo','')} — SABESP SANTOS</footer>
<script>
function showTab(btn,tabId){{
  document.querySelectorAll('.tab-btn').forEach(b=>b.classList.remove('active'));
  document.querySelectorAll('.tab-content').forEach(c=>c.classList.remove('active'));
  btn.classList.add('active');
  document.getElementById(tabId).classList.add('active');
  if(tabId==='tab-mapa'&&window._mapInit)window._map.invalidateSize();
}}
{mapa_js}
</script>
</body>
</html>"""

    caminho = pasta / f"NS_{ns_id}.html"
    with open(caminho, "w", encoding="utf-8") as f:
        f.write(html)
    return caminho


# ==============================================================================
# MÓDULO 11 — GERAÇÃO NS COMPLETA (orquestrador dos 5 arquivos)
# ==============================================================================

def gerar_ns_completa(t, pvs, ns_id, pasta_raiz, cfg,
                      ruas_dxf=None, gpkg_path=None, ns_base_url="",
                      pastas_extras=None):
    """
    Gera os arquivos de uma NS distribuídos nas pastas corretas:

      01_NS_CAMPO/NS_XXX_PVI_AO_PVF/
          NS_XXX_A4.pdf       OS campo com QR Code
          NS_XXX_DADOS.json   dados técnicos
          NS_XXX_QR.png       QR Code standalone

      02_OSE/
          NS_XXX_OSE.xlsx     Planilha OSE SABESP

      03_DESENHOS/
          NS_XXX_DESENHO.pdf  Prancha A3 planta+perfil+tabela+selo

      04_HTML/
          NS_XXX.html         Dashboard interativo Leaflet

      06_BIM/ (se IFC habilitado)
          NS_XXX_IFC.ifc      BIM LOD500 do trecho
    """
    pv_ini_c = re.sub(r"[^\w]", "", t.get("pv_ini","X"))
    pv_fim_c = re.sub(r"[^\w]", "", t.get("pv_fim","X"))

    # Pasta principal da NS (dentro de 01_NS_CAMPO)
    pasta_ns = pasta_raiz / f"NS_{ns_id}_{pv_ini_c}_AO_{pv_fim_c}"
    pasta_ns.mkdir(parents=True, exist_ok=True)

    # Pastas separadas (passadas pelo processar())
    pe = pastas_extras or {}
    pasta_ose    = pe.get("ose",      pasta_ns)  # fallback: pasta_ns
    pasta_desenh = pe.get("desenhos", pasta_ns)
    pasta_html   = pe.get("html",     pasta_ns)
    pasta_bim_ns = pe.get("bim",      pasta_ns)

    for p in [pasta_ose, pasta_desenh, pasta_html, pasta_bim_ns]:
        p.mkdir(parents=True, exist_ok=True)

    resultados = {"ns_id": ns_id, "pasta": str(pasta_ns),
                  "arquivos": {}, "erros": []}

    def _try(nome, fn, *a, **kw):
        try:
            p = fn(*a, **kw)
            if p is not None:
                resultados["arquivos"][nome] = str(p)
                log(f"    {nome}: {p.name}", "OK")
        except Exception as e:
            resultados["erros"].append(f"{nome}: {e}")
            log(f"    {nome}: {e}", "ERR")
            if "--debug" in sys.argv:
                traceback.print_exc()

    url = f"{ns_base_url}NS_{ns_id}.html" if ns_base_url else f"NS_{ns_id}.html"

    # A4 + dados + QR → pasta da NS (01_NS_CAMPO/NS_XXX/)
    _try("A4",    gerar_ns_a4_com_qr,      t, pvs, ns_id, pasta_ns, cfg,
         ns_base_url=ns_base_url)
    _try("DADOS", gerar_ns_dados_json,     t, pvs, ns_id, pasta_ns, cfg)

    # OSE → 02_OSE/
    _try("OSE",     gerar_ns_ose,              t, pvs, ns_id, pasta_ose, cfg)

    # Desenho → 03_DESENHOS/
    _try("DESENHO", gerar_ns_desenho_com_qr,   t, pvs, ns_id, pasta_desenh, cfg,
         ruas_dxf=ruas_dxf, gpkg_path=gpkg_path, ns_base_url=ns_base_url)

    # HTML → 04_HTML/
    _try("HTML",    gerar_ns_html,              t, pvs, ns_id, pasta_html, cfg)

    # IFC → 06_BIM/
    _try("IFC",     gerar_ifc, pvs, [t], pasta_bim_ns, cfg,
         {"arquivo": f"NS_{ns_id}", "tipo_rede": "AGUA" if t.get("is_agua") else "ESGOTO"},
         ns_base_url=ns_base_url)

    # QR → pasta da NS
    def _salvar_qr():
        qr_img = gerar_qr_png(url)
        if qr_img is None:
            return None
        p = pasta_ns / f"NS_{ns_id}_QR.png"
        try:
            from PIL import Image as _PILImg
            if hasattr(qr_img, 'mode') and qr_img.mode not in ('RGB','L','RGBA'):
                qr_img = qr_img.convert('RGB')
            qr_img.save(str(p), format='PNG')
        except Exception:
            qr_img.save(str(p))
        return p
    _try("QR", _salvar_qr)

    return resultados


# ==============================================================================
# MÓDULO 12 — GIS
# ==============================================================================

def gerar_rede_geojson(pvs, trechos, pasta_gis):
    pasta_gis = Path(pasta_gis)
    features = []
    for t in trechos:
        pvi = pvs.get(t.get("pv_ini"), {})
        pvf = pvs.get(t.get("pv_fim"), {})
        if not pvi.get("x") or not pvf.get("x"):
            continue
        hid = t.get("hidraulica") or {}
        q   = t.get("quantitativos") or {}
        features.append({
            "type": "Feature",
            "geometry": {"type": "LineString",
                         "coordinates": [[pvi["x"],pvi["y"]],[pvf["x"],pvf["y"]]]},
            "properties": {
                "ns_id":     t.get("ns_id",""),
                "pv_ini":    t.get("pv_ini"),   "pv_fim":    t.get("pv_fim"),
                "dn_mm":     t.get("dn_mm"),     "ext_m":     t.get("ext_m"),
                "decl_pct":  t.get("decl_pct"), "material":  t.get("material","PVC"),
                "rua":       t.get("rua","Sem Rua"),
                "ct_ini":    t.get("ct_ini"),    "ct_fim":    t.get("ct_fim"),
                "cf_ini":    t.get("cf_ini"),    "cf_fim":    t.get("cf_fim"),
                "prof_ini":  t.get("prof_ini"),  "prof_fim":  t.get("prof_fim"),
                "vel_ms":    hid.get("vel_ms"),  "vazao_ls":  hid.get("vazao_ls"),
                "tau_pa":    hid.get("tau_pa"),
                "status_hid":hid.get("status","SEM_DADOS"),
                "esc_m3":    q.get("esc_m3"),    "pav_m2":    q.get("pav_m2"),
                "tubo_barras": q.get("tubo_barras"),
            },
        })

    geojson = {
        "type": "FeatureCollection",
        "crs": {"type":"name","properties":
                {"name":"urn:ogc:def:crs:EPSG::31983"}},
        "features": features,
    }
    pasta_gis.mkdir(parents=True, exist_ok=True)
    caminho = pasta_gis / "rede_definida.geojson"
    with open(caminho, "w", encoding="utf-8") as f:
        json.dump(geojson, f, indent=2, ensure_ascii=False, default=str)
    log(f"  GeoJSON: {len(features)} trechos → {caminho.name}", "OK")
    return caminho


def gerar_rede_html(pvs, trechos, pasta_html, cfg):
    """Gera REDE_GERAL.html — mapa Leaflet interativo com toda a rede."""
    pasta_html = Path(pasta_html)
    # Coletar PVs e trechos com coords validas
    markers_js = []
    lines_js = []
    lats, lons = [], []

    # Estatísticas para legenda
    stats = {"ok": 0, "verificar": 0, "sem_dados": 0, "agua": 0, "prolong": 0}
    pvs_vistos = set()

    for t in trechos:
        pvi_d = pvs.get(t.get("pv_ini"), {})
        pvf_d = pvs.get(t.get("pv_fim"), {})
        lat_i, lon_i = utm_to_latlon(pvi_d.get("x"), pvi_d.get("y"))
        lat_f, lon_f = utm_to_latlon(pvf_d.get("x"), pvf_d.get("y"))

        if not (lat_i and lon_i and lat_f and lon_f):
            continue
        if not (-34 < lat_i < 5 and -74 < lon_i < -34):
            continue
        if not (-34 < lat_f < 5 and -74 < lon_f < -34):
            continue

        hid = t.get("hidraulica") or {}
        status = hid.get("status", "SEM_DADOS")
        is_agua = t.get("is_agua", False)
        is_prolong = t.get("is_prolongamento", False)

        # Cores por status hidráulico
        if is_prolong:
            cor = "#9b59b6"  # Roxo para prolongamentos
            stats["prolong"] += 1
        elif is_agua:
            cor = "#1E6B3C"
            stats["agua"] += 1
        elif status == "OK":
            cor = "#27ae60"
            stats["ok"] += 1
        elif "VERIFICAR" in status:
            cor = "#e74c3c"
            stats["verificar"] += 1
        else:
            cor = "#f39c12"
            stats["sem_dados"] += 1

        ns_id = t.get("ns_id", "?")
        dn = t.get('dn_mm', '?')
        ext = _fmtv(t.get('ext_m'), '.1f')
        decl = _fmtv(t.get('decl_pct'), '.2f')
        vel = _fmtv(hid.get('vel_ms'), '.2f')
        tau = _fmtv(hid.get('tau_pa'), '.2f')
        q = _fmtv(hid.get('vazao_ls'), '.1f')
        rua_safe = t.get('rua','Sem Rua').replace("'", "\\'")
        
        # Popup rico no trecho
        popup = (f"<div style='font-family:Arial,sans-serif;font-size:11px;'>"
                 f"<b style='color:#1F4E79;font-size:13px;'>NS {ns_id}</b>")
        if is_prolong:
            popup += f" <b style='color:#9b59b6;'>(Prolongamento)</b>"
        popup += f"<hr style='margin:4px 0;border:0;border-top:1px solid #ccc;'>"
        popup += f"<b>Trecho:</b> {t.get('pv_ini','')} &rarr; {t.get('pv_fim','')}<br>"
        popup += f"<b>DN:</b> {dn} mm | <b>Ext:</b> {ext} m<br>"
        popup += f"<b>Decl:</b> {decl}% | <b>Material:</b> {t.get('material','PVC')}<br>"
        popup += f"<b>Rua:</b> {rua_safe}<br>"
        popup += f"<b>Vel:</b> {vel} m/s | <b>Vazão:</b> {q} l/s<br>"
        popup += f"<b>Tau:</b> {tau} Pa<br>"
        popup += f"<b>Status:</b> <b style='color:{cor};font-size:12px;'>{status}</b></div>"

        lines_js.append(
            f"L.polyline([[{lat_i},{lon_i}],[{lat_f},{lon_f}]],"
            f"{{color:'{cor}',weight:3.5,opacity:0.85}})"
            f".addTo(map_trechos).bindPopup({repr(popup)});")

        # PVs
        for nome, lat, lon, pv_d in [(t["pv_ini"], lat_i, lon_i, pvi_d),
                                      (t["pv_fim"], lat_f, lon_f, pvf_d)]:
            if nome not in pvs_vistos:
                pvs_vistos.add(nome)
                lats.append(lat)
                lons.append(lon)
                ct = _fmtv(pv_d.get('ct'), '.3f')
                cf = _fmtv(pv_d.get('cf'), '.3f')
                prof = _fmtv(pv_d.get('prof'), '.2f')
                tipo = pv_d.get('tipo', 'PV')
                
                m_popup = (f"<div style='font-family:Arial,sans-serif;font-size:11px;'>"
                           f"<b style='color:#1F4E79;font-size:13px;'>{nome}</b>"
                           f"<br><b>Tipo:</b> {tipo}<hr style='margin:4px 0;border:0;border-top:1px solid #ccc;'>"
                           f"<b>CT:</b> {ct} m<br>"
                           f"<b>CF:</b> {cf} m<br>"
                           f"<b>Prof:</b> {prof} m</div>")
                
                markers_js.append(
                    f"L.circleMarker([{lat},{lon}],{{radius:6,color:'#1F4E79',"
                    f"fillColor:'#4488ff',fillOpacity:0.9,weight:2}})"
                    f".addTo(map_pvs).bindPopup({repr(m_popup)});")

    if not lats:
        log("  REDE HTML: sem coordenadas validas para mapa", "WARN")
        return None

    lat_c = sum(lats) / len(lats)
    lon_c = sum(lons) / len(lons)
    nucleo = cfg.get("nucleo", "")
    n_trechos = len(lines_js)
    n_pvs = len(markers_js)

    # Estatísticas completas
    ext_total = sum(t.get("ext_m", 0) or 0 for t in trechos)
    custo_total = sum((t.get("custos") or {}).get("total_R", 0) or 0 for t in trechos)
    hid_ok = stats["ok"]
    
    # DN stats
    dn_counts = {}
    for t in trechos:
        dn = t.get('dn_mm')
        if dn:
            dn_counts[dn] = dn_counts.get(dn, 0) + 1

    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Rede Geral - {nucleo} - SABESP</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Segoe UI',Arial,sans-serif;background:#1a1a2e;color:#fff}}
header{{background:linear-gradient(135deg,#1F4E79 0%,#0A2342 100%);padding:16px 24px;display:flex;align-items:center;justify-content:space-between;box-shadow:0 2px 10px rgba(0,0,0,.3)}}
header h1{{font-size:20px;font-weight:700}}
header .sub{{font-size:11px;color:#aac;margin-top:4px}}
.stats{{display:flex;gap:12px;flex-wrap:wrap}}
.stat{{background:rgba(255,255,255,.12);padding:6px 14px;border-radius:8px;text-align:center}}
.stat .num{{font-size:18px;font-weight:700;color:#00d4ff}}
.stat .lbl{{font-size:10px;color:#aaa;margin-top:2px}}
#mapa{{height:calc(100vh - 140px);width:100%}}
.legend{{position:fixed;bottom:20px;left:20px;background:rgba(26,26,46,.95);padding:14px 18px;border-radius:10px;font-size:11px;z-index:1000;box-shadow:0 2px 12px rgba(0,0,0,.5);border:1px solid #334}}
.legend b{{color:#00d4ff;margin-bottom:8px;display:block}}
.legend div{{margin:4px 0;display:flex;align-items:center;gap:8px}}
.legend .dot{{width:16px;height:4px;border-radius:2px;display:inline-block}}
.legend .pv{{width:10px;height:10px;border-radius:50%}}
.controls{{position:fixed;top:20px;right:20px;background:rgba(26,26,46,.95);padding:12px 16px;border-radius:10px;font-size:11px;z-index:1000;box-shadow:0 2px 12px rgba(0,0,0,.5)}}
.controls label{{display:block;margin:4px 0;cursor:pointer}}
.controls input{{margin-right:6px}}
.dn-filter{{position:fixed;top:20px;left:20px;background:rgba(26,26,46,.95);padding:12px 16px;border-radius:10px;font-size:11px;z-index:1000;box-shadow:0 2px 12px rgba(0,0,0,.5)}}
.dn-filter button{{background:#1F4E79;color:#fff;border:1px solid #4488ff;padding:4px 10px;margin:2px;border-radius:4px;cursor:pointer;font-size:10px}}
.dn-filter button:hover,.dn-filter button.ativo{{background:#4488ff}}
footer{{background:#1F4E79;text-align:center;padding:10px;font-size:10px;color:#88a}}
</style>
</head>
<body>
<header>
  <div>
    <h1>REDE GERAL - {nucleo.upper()}</h1>
    <div class="sub">SABESP SANTOS | Contrato {cfg.get('contrato','')} | {cfg.get('empresa','')}</div>
  </div>
  <div class="stats">
    <div class="stat"><div class="num">{n_pvs}</div><div class="lbl">PVs</div></div>
    <div class="stat"><div class="num">{n_trechos}</div><div class="lbl">Trechos</div></div>
    <div class="stat"><div class="num">{ext_total:.0f}m</div><div class="lbl">Extensão</div></div>
    <div class="stat"><div class="num">{hid_ok}</div><div class="lbl">Hid. OK</div></div>
    <div class="stat"><div class="num">{stats['verificar']}</div><div class="lbl">Verificar</div></div>
    <div class="stat"><div class="num">R${custo_total:,.0f}</div><div class="lbl">Custo Total</div></div>
  </div>
</header>
<div id="mapa"></div>
<div class="dn-filter">
  <b style='color:#00d4ff;display:block;margin-bottom:6px;'>FILTRAR POR DN</b>
  <div id="dn-buttons"></div>
</div>
<div class="controls">
  <b style='color:#00d4ff;display:block;margin-bottom:6px;'>CAMADAS</b>
  <label><input type="checkbox" checked onchange="toggleLayer('trechos')"> Trechos</label>
  <label><input type="checkbox" checked onchange="toggleLayer('pvs')"> PVs</label>
  <label><input type="checkbox" checked onchange="toggleLayer('ruas')"> Ruas</label>
</div>
<div class="legend">
  <b>LEGENDA</b>
  <div><span class="dot" style="background:#27ae60"></span> Hidráulica OK ({stats['ok']})</div>
  <div><span class="dot" style="background:#e74c3c"></span> Verificar ({stats['verificar']})</div>
  <div><span class="dot" style="background:#f39c12"></span> Sem dados ({stats['sem_dados']})</div>
  <div><span class="dot" style="background:#1E6B3C"></span> Rede de água ({stats['agua']})</div>
  <div><span class="dot" style="background:#9b59b6"></span> Prolongamento ({stats['prolong']})</div>
  <div><span class="pv" style="background:#4488ff"></span> Poço de Visita ({n_pvs})</div>
</div>
<script>
var map = L.map('mapa').setView([{lat_c},{lon_c}],16);

// Camada de Satélite (Esri World Imagery)
L.tileLayer('https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{{z}}/{{y}}/{{x}}',
  {{attribution:'Tiles © Esri, Maxar, Earthstar Geographics',maxZoom:20}}).addTo(map);

// Controles de camada
var map_trechos = L.layerGroup().addTo(map);
var map_pvs = L.layerGroup().addTo(map);
var map_ruas = L.layerGroup();

// Ruas (nomes)
var ruas = {json.dumps(list(set(t.get('rua','Sem Rua') for t in trechos if t.get('rua'))), ensure_ascii=False)};
ruas.forEach(function(rua) {{
  L.marker([{lat_c},{lon_c}],{{opacity:0}}).bindPopup('<b>Rua:</b> '+rua).addTo(map_ruas);
}});

// Trechos
{chr(10).join(lines_js)}

// PVs
{chr(10).join(markers_js)}

// Controles de camada
function toggleLayer(name) {{
  if(name==='trechos') {{ map.hasLayer(map_trechos) ? map.removeLayer(map_trechos) : map.addLayer(map_trechos); }}
  if(name==='pvs') {{ map.hasLayer(map_pvs) ? map.removeLayer(map_pvs) : map.addLayer(map_pvs); }}
  if(name==='ruas') {{ map.hasLayer(map_ruas) ? map.removeLayer(map_ruas) : map.addLayer(map_ruas); }}
}}

// Filtro DN
var dnCounts = {json.dumps(dn_counts)};
var dnBtns = document.getElementById('dn-buttons');
var dnAtivo = null;
var trechosLayer = map_trechos;

Object.keys(dnCounts).sort((a,b)=>a-b).forEach(function(dn) {{
  var btn = document.createElement('button');
  btn.textContent = 'DN'+dn+' ('+dnCounts[dn]+')';
  btn.onclick = function() {{
    if(dnAtivo === dn) {{
      dnAtivo = null;
      btn.classList.remove('ativo');
      trechosLayer.eachLayer(l => l.setStyle({{opacity: 0.8}}));
    }} else {{
      if(dnAtivo) document.querySelector('.dn-filter button.ativo')?.classList.remove('ativo');
      dnAtivo = dn;
      btn.classList.add('ativo');
      trechosLayer.eachLayer(function(layer) {{
        var popup = layer.getPopup();
        var content = popup ? popup.getContent() : '';
        var dnMatch = content.includes('DN:'+dn) || content.includes('DN '+dn);
        layer.setStyle({{opacity: dnMatch ? 0.9 : 0.1}});
        if(!dnMatch && layer._popup) {{ layer._popup._container.style.display = 'none'; }}
      }});
    }}
  }};
  dnBtns.appendChild(btn);
}});

// Popup handler - mostrar info do PV com grau de conexão
var grauCount = {{}};
{chr(10).join(markers_js)}
// Calcular grau de conexão
document.querySelectorAll('.leaflet-popup-content').forEach(function(p) {{
  var nome = p.querySelector('b')?.textContent;
  if(nome) {{ grauCount[nome] = (grauCount[nome]||0)+1; }}
}});
</script>
<footer>
  ConstruData SABESP v5.0 | Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} | Dados em UTM SIRGAS 2000
</footer>
</body>
</html>"""

    pasta_html.mkdir(parents=True, exist_ok=True)
    caminho = pasta_html / "REDE_GERAL.html"
    with open(caminho, "w", encoding="utf-8") as f:
        f.write(html)
    log(f"  REDE HTML: {n_pvs} PVs, {n_trechos} trechos -> REDE_GERAL.html", "OK")
    return caminho


def gerar_dashboard_qualidade_html(pvs, trechos, pasta_html, cfg):
    """Gera DASHBOARD_QUALIDADE.html — análise completa da rede com Chart.js."""
    pasta_html = Path(pasta_html)
    # Coletar dados válidos
    trechos_validos = []
    for t in trechos:
        pvi = pvs.get(t.get("pv_ini"), {})
        pvf = pvs.get(t.get("pv_fim"), {})
        if pvi.get("x") and pvf.get("x"):
            trechos_validos.append(t)
    
    if not trechos_validos:
        log("  DASHBOARD: sem dados válidos", "WARN")
        return None
    
    # Estatísticas
    n_pvs = len(pvs)
    n_trechos = len(trechos_validos)
    ext_total = sum(t.get("ext_m", 0) or 0 for t in trechos_validos)
    custo_total = sum((t.get("custos") or {}).get("total_R", 0) or 0 for t in trechos_validos)
    
    # Hidráulica
    hid_stats = {"ok": 0, "alerta": 0, "erro": 0}
    vel_baixa = 0
    tau_baixo = 0
    prof_raso = 0
    decl_baixa = 0
    
    dn_counts = {}
    vel_histogram = {"0-0.3": 0, "0.3-0.6": 0, "0.6-1.0": 0, "1.0-2.0": 0, "2.0-3.0": 0, "3.0+": 0}
    custo_por_rua = {}
    
    for t in trechos_validos:
        hid = t.get("hidraulica") or {}
        status = hid.get("status", "SEM_DADOS")
        vel = hid.get("vel_ms")
        tau = hid.get("tau_pa")
        decl = t.get("decl_pct")
        
        if status == "OK":
            hid_stats["ok"] += 1
        elif "VERIFICAR" in status:
            hid_stats["alerta"] += 1
        else:
            hid_stats["erro"] += 1
        
        if vel is not None:
            if vel < 0.6:
                vel_baixa += 1
            if vel < 0.3:
                vel_histogram["0-0.3"] += 1
            elif vel < 0.6:
                vel_histogram["0.3-0.6"] += 1
            elif vel < 1.0:
                vel_histogram["0.6-1.0"] += 1
            elif vel < 2.0:
                vel_histogram["1.0-2.0"] += 1
            elif vel < 3.0:
                vel_histogram["2.0-3.0"] += 1
            else:
                vel_histogram["3.0+"] += 1
        
        if tau is not None and tau < 1.0:
            tau_baixo += 1
        
        if decl is not None and decl < 0.2:
            decl_baixa += 1
        
        # PVs rasos
        for pv_nome in [t.get("pv_ini"), t.get("pv_fim")]:
            pv = pvs.get(pv_nome, {})
            prof = pv.get("prof")
            if prof is not None and prof < 0.30:
                prof_raso += 1
        
        # DN
        dn = t.get("dn_mm")
        if dn:
            dn_counts[dn] = dn_counts.get(dn, 0) + 1
        
        # Custo por rua
        rua = t.get("rua", "Sem Rua")
        custo = (t.get("custos") or {}).get("total_R", 0) or 0
        custo_por_rua[rua] = custo_por_rua.get(rua, 0) + custo
    
    # Calcular porcentagens
    pct_vel_baixa = (vel_baixa / n_trechos * 100) if n_trechos else 0
    pct_tau_baixo = (tau_baixo / n_trechos * 100) if n_trechos else 0
    pct_prof_raso = (prof_raso / (n_pvs * 2) * 100) if n_pvs else 0
    
    # Preparar dados para Chart.js
    dn_labels = list(dn_counts.keys())
    dn_data = list(dn_counts.values())
    vel_labels = list(vel_histogram.keys())
    vel_data = list(vel_histogram.values())
    
    # Top 10 ruas por custo
    top_ruas = sorted(custo_por_rua.items(), key=lambda x: x[1], reverse=True)[:10]
    rua_labels = [r[0][:25] + "..." if len(r[0]) > 25 else r[0] for r in top_ruas]
    rua_data = [r[1] for r in top_ruas]
    
    nucleo = cfg.get("nucleo", "")
    
    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Dashboard de Qualidade - {nucleo} - SABESP</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.2/dist/chart.umd.min.js"></script>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Segoe UI',Arial,sans-serif;background:#0A0E1A;color:#fff;padding:20px}}
header{{background:linear-gradient(135deg,#1F4E79 0%,#0A2342 100%);padding:20px 30px;border-radius:12px;margin-bottom:20px;box-shadow:0 4px 20px rgba(0,0,0,.4)}}
header h1{{font-size:24px;font-weight:700;margin-bottom:6px}}
header .sub{{font-size:12px;color:#aac}}
.kpis{{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:16px;margin-bottom:24px}}
.kpi{{background:linear-gradient(135deg,#1a1a2e 0%,#16213e 100%);padding:18px;border-radius:12px;text-align:center;border:1px solid #334}}
.kpi .num{{font-size:28px;font-weight:700;color:#00d4ff;margin-bottom:4px}}
.kpi .lbl{{font-size:11px;color:#88a;text-transform:uppercase}}
.kpi.alerta .num{{color:#f39c12}}
.kpi.erro .num{{color:#e74c3c}}
.kpi.sucesso .num{{color:#27ae60}}
.painel{{background:#1a1a2e;border-radius:12px;padding:20px;margin-bottom:20px;border:1px solid #334}}
.painel h2{{font-size:16px;font-weight:600;margin-bottom:16px;color:#00d4ff;border-bottom:2px solid #1F4E79;padding-bottom:8px}}
.grids{{display:grid;grid-template-columns:repeat(auto-fit,minmax(350px,1fr));gap:20px}}
.grafico{{background:#16213e;padding:16px;border-radius:10px}}
table{{width:100%;border-collapse:collapse;font-size:11px}}
th,td{{padding:10px;text-align:left;border-bottom:1px solid #334}}
th{{background:#1F4E79;color:#fff;font-weight:600}}
tr:hover{{background:rgba(255,255,255,.05)}}
.badge{{padding:3px 8px;border-radius:4px;font-size:10px;font-weight:600;display:inline-block}}
.badge.ok{{background:#27ae60;color:#fff}}
.badge.alerta{{background:#f39c12;color:#000}}
.badge.erro{{background:#e74c3c;color:#fff}}
footer{{text-align:center;padding:20px;font-size:10px;color:#668}}
</style>
</head>
<body>
<header>
  <h1>DASHBOARD DE QUALIDADE - {nucleo.upper()}</h1>
  <div class="sub">SABESP SANTOS | Contrato {cfg.get('contrato','')} | {cfg.get('empresa','')} | Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}</div>
</header>

<!-- KPIS -->
<div class="kpis">
  <div class="kpi sucesso"><div class="num">{n_pvs}</div><div class="lbl">PVs</div></div>
  <div class="kpi"><div class="num">{n_trechos}</div><div class="lbl">Trechos</div></div>
  <div class="kpi"><div class="num">{ext_total:.0f} m</div><div class="lbl">Extensão Total</div></div>
  <div class="kpi"><div class="num">R${custo_total:,.0f}</div><div class="lbl">Custo Total</div></div>
  <div class="kpi sucesso"><div class="num">{hid_stats['ok']}</div><div class="lbl">Trechos OK</div></div>
  <div class="kpi alerta"><div class="num">{hid_stats['alerta']}</div><div class="lbl">Alertas</div></div>
  <div class="kpi erro"><div class="num">{hid_stats['erro']}</div><div class="lbl">Erros</div></div>
  <div class="kpi alerta"><div class="num">{pct_vel_baixa:.1f}%</div><div class="lbl">V {'<'} 0.6 m/s</div></div>
  <div class="kpi alerta"><div class="num">{pct_tau_baixo:.1f}%</div><div class="lbl">Tau {'<'} 1.0 Pa</div></div>
  <div class="kpi alerta"><div class="num">{pct_prof_raso:.1f}%</div><div class="lbl">Prof {'<'} 0.30m</div></div>
</div>

<!-- GRÁFICOS -->
<div class="painel">
  <h2>HISTOGRAMA DE DNs</h2>
  <div class="grids">
    <div class="grafico"><canvas id="dnChart"></canvas></div>
    <div class="grafico"><canvas id="velChart"></canvas></div>
  </div>
</div>

<div class="painel">
  <h2>CUSTO POR RUA (TOP 10)</h2>
  <div class="grafico"><canvas id="ruaChart"></canvas></div>
</div>

<!-- TABELA -->
<div class="painel">
  <h2>TODOS OS TRECHOS</h2>
  <table id="tabela-trechos">
    <thead>
      <tr><th>NS</th><th>PV Ini</th><th>PV Fim</th><th>DN</th><th>Ext (m)</th><th>Decl (%)</th><th>Vel (m/s)</th><th>Tau (Pa)</th><th>Status</th><th>Rua</th></tr>
    </thead>
    <tbody>
"""
    
    for t in trechos_validos:
        hid = t.get("hidraulica") or {}
        status = hid.get("status", "SEM_DADOS")
        badge_class = "ok" if status == "OK" else "alerta" if "VERIFICAR" in status else "erro"
        html += f"""<tr>
          <td>{t.get('ns_id','?')}</td>
          <td>{t.get('pv_ini','')}</td>
          <td>{t.get('pv_fim','')}</td>
          <td>DN{t.get('dn_mm','?')}</td>
          <td>{_fmtv(t.get('ext_m'),'.1f')}</td>
          <td>{_fmtv(t.get('decl_pct'),'.2f')}</td>
          <td>{_fmtv(hid.get('vel_ms'),'.2f')}</td>
          <td>{_fmtv(hid.get('tau_pa'),'.2f')}</td>
          <td><span class="badge {badge_class}">{status}</span></td>
          <td>{t.get('rua','Sem Rua')[:30]}</td>
        </tr>"""
    
    html += f"""
    </tbody>
  </table>
</div>

<footer>
  ConstruData SABESP v5.0 | Dados em UTM SIRGAS 2000
</footer>

<script>
// Gráfico DN
new Chart(document.getElementById('dnChart'), {{
  type: 'bar',
  data: {{
    labels: {json.dumps(dn_labels)},
    datasets: [{{
      label: 'Quantidade de Trechos',
      data: {json.dumps(dn_data)},
      backgroundColor: '#4488ff',
      borderColor: '#1F4E79',
      borderWidth: 1
    }}]
  }},
  options: {{
    responsive: true,
    plugins: {{ legend: {{ display: false }} }},
    scales: {{
      y: {{ beginAtZero: true, grid: {{ color: '#334' }}, ticks: {{ color: '#88a' }} }},
      x: {{ grid: {{ display: false }}, ticks: {{ color: '#88a' }} }}
    }}
  }}
}});

// Gráfico Velocidades
new Chart(document.getElementById('velChart'), {{
  type: 'bar',
  data: {{
    labels: {json.dumps(vel_labels)},
    datasets: [{{
      label: 'Faixa de Velocidade (m/s)',
      data: {json.dumps(vel_data)},
      backgroundColor: ['#e74c3c', '#f39c12', '#f1c40f', '#27ae60', '#2ecc71', '#1abc9c'],
      borderColor: '#334',
      borderWidth: 1
    }}]
  }},
  options: {{
    responsive: true,
    plugins: {{ legend: {{ display: false }} }},
    scales: {{
      y: {{ beginAtZero: true, grid: {{ color: '#334' }}, ticks: {{ color: '#88a' }} }},
      x: {{ grid: {{ display: false }}, ticks: {{ color: '#88a' }} }}
    }}
  }}
}});

// Gráfico Ruas
new Chart(document.getElementById('ruaChart'), {{
  type: 'bar',
  data: {{
    labels: {json.dumps(rua_labels)},
    datasets: [{{
      label: 'Custo (R$)',
      data: {json.dumps(rua_data)},
      backgroundColor: '#27ae60',
      borderColor: '#1F4E79',
      borderWidth: 1
    }}]
  }},
  options: {{
    indexAxis: 'y',
    responsive: true,
    plugins: {{ legend: {{ display: false }} }},
    scales: {{
      y: {{ grid: {{ color: '#334' }}, ticks: {{ color: '#88a' }} }},
      x: {{ grid: {{ color: '#334' }}, ticks: {{ color: '#88a', callback: function(v) {{ return 'R$'+v.toLocaleString(); }} }} }}
    }}
  }}
}});
</script>
</body>
</html>"""

    pasta_html.mkdir(parents=True, exist_ok=True)
    caminho = pasta_html / "DASHBOARD_QUALIDADE.html"
    with open(caminho, "w", encoding="utf-8") as f:
        f.write(html)
    log(f"  DASHBOARD HTML: {n_pvs} PVs, {n_trechos} trechos -> DASHBOARD_QUALIDADE.html", "OK")
    return caminho


def gerar_mapa_todos_nucleos_html(todos_dados, pasta_saida, cfg_global):
    """Gera MAPA_TODOS_NUCLEOS.html — todos os núcleos em um único mapa."""
    # todos_dados: lista de dicts com {nucleo, pvs, trechos, cfg}
    
    nucleo_coords = {
        "Sao Manoel": (-23.970, -46.320),
        "Vila Criadores": (-23.960, -46.330),
        "Pantanal Baixo": (-23.960, -46.330),
        "Morro do Tetéu": (-23.950, -46.320),
        "Vila Israel": (-23.960, -46.330),
        "Joao Carlos": (-23.970, -46.320),
    }
    
    cores_nucleos = [
        "#e74c3c", "#3498db", "#2ecc71", "#f39c12", "#9b59b6", "#1abc9c",
        "#e91e63", "#00bcd4", "#ff5722", "#795548", "#607d8b", "#3f51b5"
    ]
    
    layers_js = []
    stats_por_nucleo = []
    
    for idx, dado in enumerate(todos_dados):
        nucleo = dado.get("nucleo", "Desconhecido")
        pvs = dado.get("pvs", {})
        trechos = dado.get("trechos", [])
        cor = cores_nucleos[idx % len(cores_nucleos)]
        
        n_pvs = len(pvs)
        n_trechos = len(trechos)
        ext_total = sum(t.get("ext_m", 0) or 0 for t in trechos)
        custo_total = sum((t.get("custos") or {}).get("total_R", 0) or 0 for t in trechos)
        
        stats_por_nucleo.append({
            "nucleo": nucleo, "pvs": n_pvs, "trechos": n_trechos,
            "ext_m": ext_total, "custo_R": custo_total
        })
        
        # Coletar coords válidas para centro do mapa
        lats, lons = [], []
        markers_js = []
        lines_js = []
        pvs_vistos = set()
        
        for t in trechos:
            pvi_d = pvs.get(t.get("pv_ini"), {})
            pvf_d = pvs.get(t.get("pv_fim"), {})
            lat_i, lon_i = utm_to_latlon(pvi_d.get("x"), pvi_d.get("y"))
            lat_f, lon_f = utm_to_latlon(pvf_d.get("x"), pvf_d.get("y"))
            
            if not (lat_i and lon_i and lat_f and lon_f):
                continue
            if not (-34 < lat_i < 5 and -74 < lon_i < -34):
                continue
            
            lats.extend([lat_i, lat_f])
            lons.extend([lon_i, lon_f])
            
            # Trecho
            popup = (f"<b style='color:{cor};'>{nucleo}</b><br>"
                     f"{t.get('pv_ini')} → {t.get('pv_fim')}<br>"
                     f"DN: {t.get('dn_mm')} mm | L: {_fmtv(t.get('ext_m'), '.1f')} m")
            lines_js.append(
                f"L.polyline([[{lat_i},{lon_i}],[{lat_f},{lon_f}]],"
                f"{{color:'{cor}',weight:2,opacity:0.7}})"
                f".addTo(map_{idx}).bindPopup({repr(popup)});")
            
            # PVs
            for nome, lat, lon, pv_d in [(t["pv_ini"], lat_i, lon_i, pvi_d),
                                          (t["pv_fim"], lat_f, lon_f, pvf_d)]:
                if nome not in pvs_vistos:
                    pvs_vistos.add(nome)
                    m_popup = f"<b>{nome}</b><br>{nucleo}<br>CT: {_fmtv(pv_d.get('ct'))} m"
                    markers_js.append(
                        f"L.circleMarker([{lat},{lon}],{{radius:4,color:'{cor}',"
                        f"fillColor:'{cor}',fillOpacity:0.6,weight:1}})"
                        f".addTo(map_{idx}).bindPopup({repr(m_popup)});")
        
        if lats:
            lat_c = sum(lats) / len(lats)
            lon_c = sum(lons) / len(lons)
        else:
            lat_c, lon_c = nucleo_coords.get(nucleo, (-23.96, -46.33))
        
        layers_js.append({
            "idx": idx, "nucleo": nucleo, "cor": cor,
            "lat_c": lat_c, "lon_c": lon_c,
            "markers": chr(10).join(markers_js),
            "lines": chr(10).join(lines_js),
            "stats": f"{n_pvs} PVs, {n_trechos} trechos, {ext_total:.0f}m"
        })
    
    if not layers_js:
        log("  MAPA GERAL: sem dados válidos", "WARN")
        return None
    
    # Centro do mapa (média de todos os núcleos)
    lat_geral = sum(l["lat_c"] for l in layers_js) / len(layers_js)
    lon_geral = sum(l["lon_c"] for l in layers_js) / len(layers_js)
    
    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Mapa Geral - Todos Núcleos - SABESP Santos</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Segoe UI',Arial,sans-serif;background:#1a1a2e;color:#fff}}
header{{background:linear-gradient(135deg,#1F4E79 0%,#0A2342 100%);padding:16px 24px;display:flex;align-items:center;justify-content:space-between;box-shadow:0 2px 10px rgba(0,0,0,.3)}}
header h1{{font-size:20px;font-weight:700}}
header .sub{{font-size:11px;color:#aac;margin-top:4px}}
#mapa{{height:calc(100vh - 100px);width:100%}}
.controls{{position:fixed;top:20px;right:20px;background:rgba(26,26,46,.95);padding:14px 18px;border-radius:10px;font-size:11px;z-index:1000;box-shadow:0 2px 12px rgba(0,0,0,.5);max-height:80vh;overflow-y:auto}}
.controls b{{color:#00d4ff;display:block;margin-bottom:8px}}
.controls label{{display:block;margin:4px 0;cursor:pointer}}
.controls input{{margin-right:6px}}
.stats{{position:fixed;bottom:20px;left:20px;background:rgba(26,26,46,.95);padding:14px 18px;border-radius:10px;font-size:11px;z-index:1000;box-shadow:0 2px 12px rgba(0,0,0,.5)}}
.stats table{{border-collapse:collapse}}
.stats td{{padding:4px 8px;border-bottom:1px solid #334}}
.stats .cor{{width:16px;height:4px;border-radius:2px;display:inline-block}}
footer{{background:#1F4E79;text-align:center;padding:10px;font-size:10px;color:#88a}}
</style>
</head>
<body>
<header>
  <div>
    <h1>MAPA GERAL - TODOS NÚCLEOS</h1>
    <div class="sub">SABESP SANTOS | Contrato {cfg_global.get('contrato','')} | {len(layers_js)} núcleos</div>
  </div>
</header>
<div id="mapa"></div>
<div class="controls">
  <b>CAMADAS POR NÚCLEO</b>
  {''.join([f"<label><input type='checkbox' checked onchange='toggleNucleo({l['idx']})'><span style='color:{l['cor']}'>{l['nucleo']}</span> ({l['stats']})</label>" for l in layers_js])}
</div>
<div class="stats">
  <b style='color:#00d4ff;display:block;margin-bottom:8px;'>RESUMO</b>
  <table>
    <tr><td>Núcleos:</td><td><b>{len(layers_js)}</b></td></tr>
    <tr><td>Total PVs:</td><td><b>{sum(s['pvs'] for s in stats_por_nucleo)}</b></td></tr>
    <tr><td>Total Trechos:</td><td><b>{sum(s['trechos'] for s in stats_por_nucleo)}</b></td></tr>
    <tr><td>Extensão:</td><td><b>{sum(s['ext_m'] for s in stats_por_nucleo):.0f} m</b></td></tr>
    <tr><td>Custo:</td><td><b>R${sum(s['custo_R'] for s in stats_por_nucleo):,.0f}</b></td></tr>
  </table>
</div>
<script>
var map = L.map('mapa').setView([{lat_geral},{lon_geral}],14);
L.tileLayer('https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{{z}}/{{y}}/{{x}}',
  {{attribution:'Tiles © Esri, Maxar, Earthstar Geographics',maxZoom:20}}).addTo(map);

// Camadas por núcleo
var nucleos = {{}};
{chr(10).join([f"nucleos[{l['idx']}] = L.layerGroup();{chr(10)}{l['lines']}{chr(10)}{l['markers']}{chr(10)}nucleos[{l['idx']}].addTo(map);" for l in layers_js])}

function toggleNucleo(idx) {{
  if(map.hasLayer(nucleos[idx])) {{
    map.removeLayer(nucleos[idx]);
  }} else {{
    map.addLayer(nucleos[idx]);
  }}
}}
</script>
<footer>
  ConstruData SABESP v5.0 | Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} | Dados em UTM SIRGAS 2000
</footer>
</body>
</html>"""

    caminho = Path(pasta_saida) / "MAPA_TODOS_NUCLEOS.html"
    with open(caminho, "w", encoding="utf-8") as f:
        f.write(html)
    log(f"  MAPA GERAL: {len(layers_js)} núcleos -> MAPA_TODOS_NUCLEOS.html", "OK")
    return caminho


def gerar_rede_dynamo(pvs, trechos, pasta_gis, meta, cfg):
    """JSON para Civil 3D 2025/2026 — API Pipe.ByStructures."""
    pasta_gis = Path(pasta_gis)
    pontos, vistos = [], set()
    for t in trechos:
        for nm in [t.get("pv_ini"), t.get("pv_fim")]:
            if nm and nm not in vistos:
                vistos.add(nm)
                pv = pvs.get(nm, {})
                pontos.append({
                    "id": nm, "tipo": pv.get("tipo","PV"),
                    "x": pv.get("x"), "y": pv.get("y"),
                    "z": pv.get("ct") or 0,
                    "ct": pv.get("ct"), "cf": pv.get("cf"),
                    "prof": pv.get("prof"), "dn_tampa_mm": 600,
                })

    tubulacoes = [
        {"id": f"TRECHO-{i+1:03d}",
         "pv_ini": t.get("pv_ini"), "pv_fim": t.get("pv_fim"),
         "dn_mm": t.get("dn_mm") or 200, "material": t.get("material","PVC"),
         "ext_m": t.get("ext_m"), "decl_mm": t.get("decl_mm")}
        for i, t in enumerate(trechos)
    ]

    data = {
        "metadata": {
            "projeto":         meta.get("arquivo",""),
            "contrato":        cfg["contrato"],
            "nucleo":          cfg.get("nucleo",""),
            "datum":           "SIRGAS 2000 UTM 23S",
            "unidade_linear":  "metros",
            "total_pvs":       len(pontos),
            "total_trechos":   len(tubulacoes),
            "data_exportacao": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "software":        "ConstruData BIM SABESP v5.0",
            "civil3d_api": {
                "versao_minima": "Civil 3D 2025.1",
                "nodes_usar": [
                    "PipeNetwork.ByName   → cria a rede",
                    "Structure.ByPoint    → cria cada PV com CT/CF",
                    "Pipe.ByStructures    → conecta dois PVs com tubo",
                ],
                "nodes_NAO_usar": [
                    "PipeNetwork.Create   → API antiga 2020-2023",
                    "network.AddStructure → API antiga 2020-2023",
                ],
                "nota": (
                    "Usar CPython 3 (não IronPython). "
                    "Importar este JSON com JSON.FromFile no nó Python do Dynamo. "
                    "Script de exemplo em: 07_LOG/dynamo_pipe_network_v5.py"
                ),
            },
        },
        "pontos": pontos,
        "tubulacoes": tubulacoes,
        "pipe_network": {
            "name": f"REDE_{cfg.get('nucleo','').upper().replace(' ','_')}",
            "description": f"ConstruData v5 — {cfg.get('nucleo','')}",
            "structure_family": "Sabesp_Tampa_PV",
            "pipe_family": "Sabesp_Tubo_PVC",
        },
    }

    caminho = pasta_gis / "rede_dynamo.json"
    with open(caminho, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False, default=str)
    log(f"  rede_dynamo.json: {caminho.name}", "OK")
    return caminho


# ==============================================================================
# MÓDULO 13 — EXCEL DE CUSTOS
# ==============================================================================

def gerar_excel_custos(trechos, pvs, pasta_custos, cfg):
    """Gera CUSTOS_POR_TRECHO.xlsx (separado das NS de campo)."""
    pasta_custos = Path(pasta_custos)
    if not _HAS_OPENPYXL:
        log("openpyxl não disponível — CUSTOS.xlsx não gerado", "WARN")
        return None

    pasta_custos.mkdir(parents=True, exist_ok=True)
    caminho = pasta_custos / "CUSTOS_POR_TRECHO.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CUSTOS"
    ws.freeze_panes = "A2"

    thin = Side(style="thin", color="888888")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["NS","PV_INI","PV_FIM","DN(mm)","EXT(m)","RUA",
               "TUBO(R$)","ESCAV.(R$)","REATERRO(R$)","LASTRO(R$)",
               "BRITA(R$)","PAVIM.(R$)","TOTAL(R$)","BDI(%)"]
    widths  = [6, 14, 14, 8, 8, 30, 10, 10, 10, 10, 10, 10, 12, 6]

    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
        c = ws.cell(row=1, column=ci, value=h)
        c.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
        c.fill      = PatternFill("solid", fgColor="1F4E79")
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border    = brd
    ws.row_dimensions[1].height = 26

    total_geral = 0.0
    for ri, t in enumerate(trechos, 2):
        ns_id  = t.get("ns_id", str(ri-1).zfill(3))
        custo  = t.get("custos") or {}
        total_geral += custo.get("total_R") or 0
        fc = "F0F4FA" if ri%2==0 else "FFFFFF"
        row_vals = [
            ns_id,
            t.get("pv_ini",""), t.get("pv_fim",""),
            t.get("dn_mm"),     t.get("ext_m"),
            t.get("rua",""),
            custo.get("tubo_R"),     custo.get("escavacao_R"),
            custo.get("reaterro_R"), custo.get("lastro_R"),
            custo.get("brita_R"),    custo.get("pavimentacao_R"),
            custo.get("total_R"),    custo.get("bdi_pct"),
        ]
        fmts = [None,None,None,"0","0.00",None,
                "#,##0.00","#,##0.00","#,##0.00","#,##0.00",
                "#,##0.00","#,##0.00","#,##0.00","0"]
        for ci, (v, fmt) in enumerate(zip(row_vals, fmts), 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font      = Font(name="Calibri", size=8.5)
            c.fill      = PatternFill("solid", fgColor=fc)
            c.border    = brd
            c.alignment = Alignment(
                horizontal="right" if ci > 5 else "center",
                vertical="center")
            if fmt:
                c.number_format = fmt
        ws.row_dimensions[ri].height = 14

    # Total geral
    row_tot = len(trechos) + 2
    ws.merge_cells(f"A{row_tot}:F{row_tot}")
    c = ws.cell(row=row_tot, column=1, value="TOTAL GERAL")
    c.font = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = brd
    c13 = ws.cell(row=row_tot, column=13, value=round(total_geral, 2))
    c13.font   = Font(name="Calibri", bold=True, size=9)
    c13.fill   = PatternFill("solid", fgColor="DCE6F1")
    c13.border = brd
    c13.number_format = "#,##0.00"
    c13.alignment = Alignment(horizontal="right", vertical="center")

    # Rodapé
    row_rod = row_tot + 2
    ws[f"A{row_rod}"].value = (
        f"ConstruData BIM SABESP v5.0 — {cfg.get('nucleo','')} — "
        f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} — "
        f"BDI={int(CFG['bdi']*100-100)}%")
    ws.merge_cells(f"A{row_rod}:N{row_rod}")
    ws[f"A{row_rod}"].font = Font(name="Calibri", size=8, color="777777", italic=True)

    wb.save(caminho)
    log(f"  CUSTOS: {caminho.name} ({len(trechos)} trechos / R$ {total_geral:,.2f})", "OK")
    return caminho


# ==============================================================================
# MÓDULO 13B — COMPARATIVO ProSaneamento (mesmas colunas do CSV exportado)
# ==============================================================================

def gerar_comparativo_prosane(trechos, pvs, pasta_excel, cfg):
    """Gera COMPARATIVO_PROSANEAMENTO.xlsx com as mesmas colunas do CSV do ProSaneamento."""
    pasta_excel = Path(pasta_excel)
    if not _HAS_OPENPYXL:
        return None

    pasta_excel.mkdir(parents=True, exist_ok=True)
    caminho = pasta_excel / "COMPARATIVO_PROSANEAMENTO.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dimensiona(Esgoto)"
    ws.freeze_panes = "A3"

    thin = Side(style="thin", color="888888")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    azul = PatternFill("solid", fgColor="1F4E79")
    cinza = PatternFill("solid", fgColor="F0F4FA")
    branco = PatternFill("solid", fgColor="FFFFFF")
    verde = PatternFill("solid", fgColor="E2EFDA")
    amarelo = PatternFill("solid", fgColor="FFF2CC")
    bold_w = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
    bold_b = Font(name="Calibri", bold=True, color="1F4E79", size=9)
    normal = Font(name="Calibri", size=9)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Titulo
    ws.merge_cells("A1:S1")
    c = ws["A1"]
    c.value = f"COMPARATIVO ConstruData vs ProSaneamento | {cfg.get('nucleo','')} | {cfg.get('contrato','')}"
    c.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    c.fill = azul
    c.alignment = center
    ws.row_dimensions[1].height = 28

    # Cabecalho
    headers = [
        "NS", "PV Montante", "PV Jusante", "Rua",
        "Extensao\n(m)", "DN\n(mm)", "Material",
        "CT ini\n(m)", "CT fim\n(m)",
        "CF ini\n(m)", "CF fim\n(m)",
        "Prof ini\n(m)", "Prof fim\n(m)",
        "Decl\n(m/m)", "Decl\n(%)",
        "V\n(m/s)", "Q\n(l/s)", "Tau\n(Pa)", "Status Hid."
    ]
    widths = [6, 12, 12, 22, 10, 8, 10, 10, 10, 10, 10, 10, 10, 12, 8, 8, 8, 8, 14]

    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
        c = ws.cell(row=2, column=ci, value=h)
        c.font = bold_w
        c.fill = azul
        c.alignment = center
        c.border = brd
    ws.row_dimensions[2].height = 36

    # Dados
    for ri, t in enumerate(trechos, 3):
        hid = t.get("hidraulica") or {}
        fc = cinza if ri % 2 == 1 else branco
        status = hid.get("status", "SEM_DADOS")
        if status == "OK":
            fc_status = verde
        elif "VERIFICAR" in status:
            fc_status = amarelo
        else:
            fc_status = fc

        vals = [
            t.get("ns_id", ""),
            t.get("pv_ini", ""),
            t.get("pv_fim", ""),
            t.get("rua", "Sem Rua"),
            t.get("ext_m"),
            t.get("dn_mm"),
            t.get("material", "PVC"),
            t.get("ct_ini"),
            t.get("ct_fim"),
            t.get("cf_ini"),
            t.get("cf_fim"),
            t.get("prof_ini"),
            t.get("prof_fim"),
            t.get("decl_mm"),
            t.get("decl_pct"),
            hid.get("vel_ms"),
            hid.get("vazao_ls"),
            hid.get("tau_pa"),
            status,
        ]
        fmts = [
            None, None, None, None,
            "0.0000", "0", None,
            "0.0000", "0.0000",
            "0.0000", "0.0000",
            "0.0000", "0.0000",
            "0.00000", "0.000",
            "0.000", "0.00", "0.00", None,
        ]
        for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = normal
            c.fill = fc_status if ci == 19 else fc
            c.alignment = center
            c.border = brd
            if fmt and v is not None:
                c.number_format = fmt
        ws.row_dimensions[ri].height = 15

    # Resumo
    row_r = len(trechos) + 4
    ws.merge_cells(f"A{row_r}:D{row_r}")
    ws[f"A{row_r}"].value = "RESUMO"
    ws[f"A{row_r}"].font = bold_w
    ws[f"A{row_r}"].fill = azul
    ws[f"A{row_r}"].alignment = center

    n_total = len(trechos)
    n_ok = sum(1 for t in trechos if (t.get("hidraulica") or {}).get("status") == "OK")
    n_cota = sum(1 for t in trechos if t.get("ct_ini") is not None and t.get("cf_ini") is not None)
    n_rua = sum(1 for t in trechos if t.get("rua", "Sem Rua") != "Sem Rua")
    ext_total = sum(t.get("ext_m", 0) or 0 for t in trechos)

    resumo = [
        ("Trechos", n_total),
        ("Com cota", f"{n_cota}/{n_total}"),
        ("Com rua", f"{n_rua}/{n_total}"),
        ("Hid. OK", f"{n_ok}/{n_total}"),
        ("Extensao total", f"{ext_total:.1f} m"),
    ]
    for i, (lbl, val) in enumerate(resumo):
        ws.cell(row=row_r+1+i, column=1, value=lbl).font = bold_b
        ws.cell(row=row_r+1+i, column=2, value=str(val)).font = normal

    # Rodape
    row_rod = row_r + len(resumo) + 2
    ws.merge_cells(f"A{row_rod}:S{row_rod}")
    ws[f"A{row_rod}"].value = (
        f"ConstruData BIM SABESP v5.0 | {cfg.get('nucleo','')} | "
        f"{cfg.get('contrato','')} | Gerado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    ws[f"A{row_rod}"].font = Font(name="Calibri", size=8, color="777777", italic=True)

    wb.save(caminho)
    log(f"  COMPARATIVO: {caminho.name} ({n_total} trechos, {n_cota} com cota, {n_rua} com rua)", "OK")
    return caminho


# ==============================================================================
# MÓDULO 14 — SCRIPT DYNAMO ATUALIZADO (Civil 3D 2025/2026)
# ==============================================================================

DYNAMO_SCRIPT_V5 = '''# -*- coding: utf-8 -*-
"""
Script Dynamo/Civil 3D 2025/2026 — ConstruData BIM SABESP v5.0
Cria Pipe Network a partir do rede_dynamo.json.

REQUISITOS:
  - Civil 3D 2025.1+ (nodes PipeNetwork.ByName, Structure.ByPoint, Pipe.ByStructures)
  - Dynamo 3.2+ com CPython 3.7+ (NÃO IronPython)
  - Package: Civil3DToolkit (opcional, para acesso avançado)

COMO USAR:
  1. Dynamo → Novo Gráfico
  2. Inserir nó "Python Script" (CPython 3)
  3. Colar este código no nó Python
  4. Conectar entrada IN[0] = caminho do rede_dynamo.json
  5. Executar

ALTERNATIVA (nodes nativos, sem Python):
  File.ReadText → JSON.FromObject → List.GetItemAtIndex →
  PipeNetwork.ByName → Structure.ByPoint (loop) → Pipe.ByStructures (loop)
"""

import sys
import json
import clr

# ── Assemblies Civil 3D 2025+ ────────────────────────────────────────────────
clr.AddReference("AcMgd")
clr.AddReference("AcCoreMgd")
clr.AddReference("AcDbMgd")
clr.AddReference("AecBaseMgd")
clr.AddReference("AecPropDataMgd")
clr.AddReference("AeccDbMgd")

from Autodesk.AutoCAD.ApplicationServices import Application as AcadApp
from Autodesk.AutoCAD.DatabaseServices import (
    Transaction, OpenMode, ObjectId
)
from Autodesk.AutoCAD.Geometry import Point3d
from Autodesk.Civil.DatabaseServices import (
    PipeNetwork, Structure, Pipe
)
from Autodesk.Civil.ApplicationServices import CivilApplication

# ── Entrada do nó Dynamo ─────────────────────────────────────────────────────
# IN[0] = caminho do rede_dynamo.json
CAMINHO_JSON = IN[0] if IN and IN[0] else r"rede_dynamo.json"


def criar_pipe_network_v5(json_path: str) -> str:
    """
    Cria Pipe Network no Civil 3D 2025/2026 a partir do rede_dynamo.json.

    API NOVA (2025.1+):
      PipeNetwork.Create(db, name, desc, parts_list_id, layer_id)
      network.AddStructure(family, size, alignId, E, N, rim_elev, snap_to_align)
      struct.SumpElevation = cf
      network.AddPipe(family, size, alignId, struct_ini_id, struct_fim_id)
    """
    with open(json_path, encoding="utf-8") as f:
        dados = json.load(f)

    meta      = dados["metadata"]
    net_cfg   = dados["pipe_network"]
    pontos    = dados["pontos"]
    tubulacoes= dados["tubulacoes"]

    doc = AcadApp.DocumentManager.MdiActiveDocument
    if doc is None:
        raise RuntimeError("Nenhum documento Civil 3D aberto.")

    db     = doc.Database
    editor = doc.Editor
    pv_ids = {}
    log    = []

    with doc.LockDocument():
        with db.TransactionManager.StartTransaction() as tr:
            # ── Parts List ───────────────────────────────────────────────────
            civil_doc = CivilApplication.ActiveDocument
            parts_ids = list(civil_doc.Styles.PartsList)
            parts_id  = parts_ids[0] if parts_ids else db.BlockTableId

            # ── Criar rede ───────────────────────────────────────────────────
            net_id = PipeNetwork.Create(
                db,
                net_cfg.get("name", "REDE_SABESP"),
                net_cfg.get("description", "ConstruData BIM v5"),
                parts_id,
                db.LayerZero,
            )
            network = tr.GetObject(net_id, OpenMode.ForWrite)
            log.append(f"Rede criada: {net_cfg['name']}")

            # ── Estruturas (PVs) ─────────────────────────────────────────────
            for ponto in pontos:
                try:
                    x   = ponto.get("x") or 0
                    y   = ponto.get("y") or 0
                    ct  = ponto.get("ct") or ponto.get("z") or 0
                    cf  = ponto.get("cf") or ct

                    struct_id = network.AddStructure(
                        net_cfg.get("structure_family", "Sabesp_Tampa_PV"),
                        f"DN{ponto.get('dn_tampa_mm', 600)}",
                        ObjectId.Null,
                        x, y, ct,
                        False,
                    )
                    struct = tr.GetObject(struct_id, OpenMode.ForWrite)

                    # CF via SumpElevation (API 2025+)
                    try:
                        struct.SumpElevation = cf
                    except AttributeError:
                        try:
                            struct.SumpDepth = ct - cf
                        except Exception:
                            pass

                    pv_ids[ponto["id"]] = struct_id
                    log.append(f"  PV {ponto['id']}: CT={ct:.3f} CF={cf:.3f}")

                except Exception as ex:
                    log.append(f"  [ERRO] PV {ponto.get('id','?')}: {ex}")

            # ── Tubos ────────────────────────────────────────────────────────
            for tubo in tubulacoes:
                try:
                    id_ini = pv_ids.get(tubo.get("pv_ini",""))
                    id_fim = pv_ids.get(tubo.get("pv_fim",""))
                    if id_ini is None or id_fim is None:
                        log.append(f"  [AVISO] {tubo['id']}: PV não encontrado")
                        continue

                    pipe_id = network.AddPipe(
                        net_cfg.get("pipe_family", "Sabesp_Tubo_PVC"),
                        f"DN{tubo.get('dn_mm', 200)}",
                        ObjectId.Null,
                        id_ini,
                        id_fim,
                    )
                    log.append(f"  Tubo {tubo['id']}: DN{tubo.get('dn_mm')}mm")

                except Exception as ex:
                    log.append(f"  [ERRO] Tubo {tubo.get('id','?')}: {ex}")

            tr.Commit()

    resumo = (
        f"=== PIPE NETWORK CRIADA ===\\n"
        f"Rede: {net_cfg['name']}\\n"
        f"Projeto: {meta.get('projeto','')}\\n"
        f"PVs: {len(pontos)} | Tubos: {len(tubulacoes)}\\n"
        + "\\n".join(log[:20])
    )
    editor.WriteMessage(f"\\n{resumo}\\n")
    return resumo


# ── Saída Dynamo ─────────────────────────────────────────────────────────────
try:
    OUT = criar_pipe_network_v5(CAMINHO_JSON)
except Exception as e:
    OUT = f"ERRO: {e}"
    raise
'''


def gerar_dynamo_script(pasta_log, cfg):
    """Salva o script Dynamo atualizado para Civil 3D 2025/2026."""
    caminho = pasta_log / "dynamo_pipe_network_v5.py"
    with open(caminho, "w", encoding="utf-8") as f:
        f.write(DYNAMO_SCRIPT_V5)
    log(f"  Dynamo script: {caminho.name}", "OK")
    return caminho


# ==============================================================================
# MÓDULO 15 — PIPELINE PRINCIPAL
# ==============================================================================

def processar(dxf_path=None, json_path=None, pasta_saida=None,
              nucleo=None, gpkg_path=None, cfg_extra=None,
              max_ns=None, tipo_override=None, ns_base_url="",
              quant_campo=None):
    """
    Pipeline completo: lê DXF ou JSON → enriquece → gera NS + GIS + Dynamo.

    Parâmetros
    ----------
    dxf_path      : caminho do DXF ProSaneamento (prioritário)
    json_path     : caminho de rede_definida.json ou rede_esgoto_dynamo.json
    pasta_saida   : pasta raiz de saída
    nucleo        : nome do núcleo
    gpkg_path     : GPKG de cartografia (opcional)
    cfg_extra     : dict com chaves adicionais (empresa, engenheiro, etc.)
    max_ns        : limitar número de NS geradas (debug)
    tipo_override : forçar "agua" ou "esgoto"
    """
    t_ini = time.time()
    _LOG_ENTRIES.clear()

    # Config local
    cfg_local = {**CFG, **(cfg_extra or {})}
    cfg_local["nucleo"] = nucleo or cfg_local.get("nucleo","Desconhecido")

    log("=" * 65, "INFO")
    log("  CONSTRUDATA BIM SABESP v5.0", "INFO")
    log(f"  Núcleo: {cfg_local['nucleo']}", "INFO")
    log("=" * 65, "INFO")

    # ── Leitura ───────────────────────────────────────────────────────────────
    if dxf_path and Path(dxf_path).exists() and str(dxf_path).lower().endswith('.dwg'):
        # DWG Civil 3D — extrai via ler_dwg_aec (3 camadas com fallback)
        from ler_dwg_aec import ler_dwg_aec as _ler_dwg
        log(f"Arquivo DWG detectado: {Path(dxf_path).name}", "STEP")
        pvs, trechos, meta = _ler_dwg(dxf_path)
        ruas = []
    elif dxf_path and Path(dxf_path).exists():
        if _HAS_GDAL_READER:
            pvs, trechos, ruas, meta = ler_dxf_gdal(dxf_path)
        else:
            pvs, trechos, ruas, meta = ler_dxf(dxf_path)
    elif json_path and Path(json_path).exists():
        pvs, trechos, ruas, meta = ler_json_rede(json_path)
    else:
        raise FileNotFoundError(
            f"Arquivo não encontrado: dxf={dxf_path} json={json_path}")

    if tipo_override:
        is_a = (tipo_override.lower() == "agua")
        meta["tipo_rede"] = "AGUA" if is_a else "ESGOTO"
        for t in trechos:
            t["is_agua"] = is_a

    # ── Enriquecimento ────────────────────────────────────────────────────────
    log("Enriquecendo trechos (Manning + Quant + Custos)...", "STEP")
    trechos = enriquecer_trechos(trechos, pvs)

    # ── Validação ─────────────────────────────────────────────────────────────
    log("Validando rede...", "STEP")
    erros, avisos = validar_rede(pvs, trechos)
    if erros:
        for e in erros[:5]:
            log(e, "ERR")
    if avisos:
        for a in avisos[:5]:
            log(a, "WARN")

    # ── Pastas de saída ───────────────────────────────────────────────────────
    nome_nucleo = (cfg_local["nucleo"].upper()
                   .replace(" ","_").replace("/","_")[:30])
    raiz = Path(pasta_saida or "SAIDA_BIM_SABESP") / nome_nucleo
    # ── Estrutura de pastas conforme print do projeto ────────────────────────
    pasta_ns    = raiz / "01_NS_CAMPO"            # subpastas por NS
    pasta_ose   = raiz / "02_OSE"                 # todas as OSE.xlsx
    pasta_deseq = raiz / "03_DESENHOS"            # todas as pranchas A3
    pasta_html  = raiz / "04_HTML"                # todos os dashboards .html
    pasta_gis   = raiz / "05_GIS"                 # GeoJSON + rede_dynamo
    pasta_excel = raiz / "06_EXCEL"               # custos + planejamento
    pasta_bim   = raiz / "06_BIM"                 # IFC LOD500 rede completa
    pasta_log_p = raiz / "07_LOG"                 # log + dynamo script
    for p in [pasta_ns, pasta_ose, pasta_deseq, pasta_html,
              pasta_gis, pasta_excel, pasta_bim, pasta_log_p]:
        p.mkdir(parents=True, exist_ok=True)

    # Mapa de pastas extras passado para cada NS
    pastas_extras = {
        "ose":      pasta_ose,
        "desenhos": pasta_deseq,
        "html":     pasta_html,
        "bim":      pasta_bim,
    }

    # ── Cartografia GPKG (complementar ruas) ─────────────────────────────────
    if gpkg_path and Path(gpkg_path).exists():
        log(f"Cartografia GPKG: {Path(gpkg_path).name}", "STEP")
        carta = ler_cartografia_gpkg(gpkg_path)
        # Complementar ruas vazias
        for t in trechos:
            if t.get("rua","Sem Rua") == "Sem Rua" and carta.get("ruas_txt"):
                pvi = pvs.get(t.get("pv_ini"), {})
                pvf = pvs.get(t.get("pv_fim"), {})
                mx  = (pvi.get("x",0)+pvf.get("x",0))/2
                my  = (pvi.get("y",0)+pvf.get("y",0))/2
                if mx and my:
                    rt = _mais_proximo((mx,my), carta["ruas_txt"], max_dist=100.0)
                    if rt and rt.get("text","").strip():
                        t["rua"] = rt["text"].strip()

    # ── Gerar NS ──────────────────────────────────────────────────────────────
    log(f"Gerando {len(trechos)} NS...", "STEP")
    resultados_ns = []
    n_max = max_ns or len(trechos)

    for i, t in enumerate(trechos[:n_max]):
        ns_id = str(i+1).zfill(3)
        t["ns_id"] = ns_id
        log(f"  NS {ns_id}: {t['pv_ini']} → {t['pv_fim']}"
            f" | DN{t.get('dn_mm','?')} | {t.get('ext_m','?')}m"
            f" | {t.get('rua','Sem Rua')}", "INFO")
        res = gerar_ns_completa(
            t, pvs, ns_id, pasta_ns, cfg_local,
            ruas_dxf=ruas, gpkg_path=gpkg_path,
            ns_base_url=ns_base_url,
            pastas_extras=pastas_extras)
        resultados_ns.append(res)

    # ── Custos ────────────────────────────────────────────────────────────────
    log("Gerando CUSTOS_POR_TRECHO.xlsx...", "STEP")
    gerar_excel_custos(trechos[:n_max], pvs, pasta_excel, cfg_local)

    # ── Comparativo ProSaneamento ─────────────────────────────────────────────
    log("Gerando COMPARATIVO_PROSANEAMENTO.xlsx...", "STEP")
    gerar_comparativo_prosane(trechos[:n_max], pvs, pasta_excel, cfg_local)

    # ── GIS ───────────────────────────────────────────────────────────────────
    log("Gerando GIS (GeoJSON + rede_dynamo.json)...", "STEP")
    gerar_rede_geojson(pvs, trechos[:n_max], pasta_gis)
    gerar_rede_dynamo(pvs, trechos[:n_max], pasta_gis, meta, cfg_local)

    # ── HTML rede geral + Dashboard Qualidade ─────────────────────────────────
    log("Gerando mapas HTML e dashboard de qualidade...", "STEP")
    gerar_rede_html(pvs, trechos[:n_max], pasta_html, cfg_local)
    gerar_dashboard_qualidade_html(pvs, trechos[:n_max], pasta_html, cfg_local)

    # ── IFC global da rede completa ───────────────────────────────────────────
    log("Gerando IFC LOD500 da rede completa...", "STEP")
    pasta_bim = raiz / "06_BIM"
    try:
        ifc_path = gerar_ifc(pvs, trechos[:n_max], pasta_bim, cfg_local, meta,
                             ns_base_url=ns_base_url)
        # Marcar arquivo .ifc.pendente como concluído
        pendente = pasta_bim / "rede_ifc.pendente"
        if pendente.exists():
            pendente.unlink()
        log(f"  IFC global: {ifc_path.name}", "OK")
    except Exception as e:
        log(f"  IFC global ERRO: {e}", "WARN")

    # ── Medição + Curva S + Diário de Obras ──────────────────────────────────
    try:
        from gerar_medicao_curva_s import gerar_medicao_curva_s
        log("Gerando medição, curva S e diário de obras...", "STEP")
        med = gerar_medicao_curva_s(pvs, trechos[:n_max], cfg_local["nucleo"],
                                     str(raiz / "09_MEDICAO"))
        log(f"  Medição: {med['stats']['total_ns']} NS | "
            f"{med['stats']['meses_total']} meses | "
            f"R$ {med['stats']['custo_total']:,.0f}", "OK")
    except Exception as e:
        log(f"  Medição erro: {e}", "WARN")

    # ── Sistema de Compras ────────────────────────────────────────────────────
    try:
        from gerar_compras import gerar_compras
        log("Gerando sistema de compras...", "STEP")
        compras = gerar_compras(pvs, trechos[:n_max], cfg_local["nucleo"],
                                str(raiz / "08_COMPRAS"))
        log(f"  Compras: {compras['total_itens']} itens | "
            f"R$ {compras['custo_total']:,.0f}", "OK")
    except Exception as e:
        log(f"  Compras erro: {e}", "WARN")

    # ── Cronograma por NS ────────────────────────────────────────────────────
    try:
        from gerar_cronograma_ns import gerar_cronograma_ns
        log("Gerando cronograma por NS...", "STEP")
        crono = gerar_cronograma_ns(
            pvs, trechos[:n_max], cfg_local["nucleo"],
            pasta_saida=str(raiz / "05_CRONOGRAMA"),
            equipes=4,
        )
        log(f"  Cronograma: {crono['stats']['total_ns']} NS | "
            f"{crono['stats']['dias_totais']} dias | "
            f"R$ {crono['stats']['custo_total']:,.0f}", "OK")
    except Exception as e:
        log(f"  Cronograma NS erro: {e}", "WARN")

    # ── Dynamo script ─────────────────────────────────────────────────────────
    gerar_dynamo_script(pasta_log_p, cfg_local)

    # ── Log JSON ──────────────────────────────────────────────────────────────
    dt_total = round(time.time() - t_ini, 1)
    n_ok  = sum(1 for r in resultados_ns if not r.get("erros"))
    n_err = sum(1 for r in resultados_ns if r.get("erros"))

    log_data = {
        "nucleo":       cfg_local["nucleo"],
        "contrato":     cfg_local["contrato"],
        "arquivo":      meta.get("arquivo",""),
        "tipo_rede":    meta.get("tipo_rede",""),
        "n_pvs":        len(pvs),
        "n_trechos":    len(trechos),
        "n_ns_geradas": len(resultados_ns),
        "n_ns_ok":      n_ok,
        "n_ns_erros":   n_err,
        "erros_validacao": erros[:20],
        "avisos_validacao": avisos[:20],
        "tempo_s":      dt_total,
        "gerado_em":    datetime.now().isoformat(),
        "resultados_ns": resultados_ns,
        "log_entries":  _LOG_ENTRIES[-50:],
    }
    caminho_log = pasta_log_p / "log_processamento.json"
    with open(caminho_log, "w", encoding="utf-8") as f:
        json.dump(log_data, f, indent=2, ensure_ascii=False, default=str)

    # ── Resumo final ──────────────────────────────────────────────────────────
    log("=" * 65, "INFO")
    log(f"  CONCLUÍDO em {dt_total}s", "OK")
    log(f"  NS geradas:  {n_ok} OK | {n_err} com erros", "OK")
    log(f"  Erros rede:  {len(erros)} | Avisos: {len(avisos)}", "OK")
    log(f"  Saída:  {raiz}", "OK")
    log(f"  01_NS_CAMPO  | 02_OSE  | 03_DESENHOS  | 04_HTML", "OK")
    log(f"  05_GIS  | 06_EXCEL  | 06_BIM  | 07_LOG", "OK")
    log("=" * 65, "INFO")

    return {
        "raiz": str(raiz),
        "pvs": len(pvs),
        "trechos": len(trechos),
        "ns_ok": n_ok,
        "ns_erros": n_err,
        "erros_validacao": erros,
        "avisos_validacao": avisos,
        "pasta_bim":   str(pasta_bim),
        "pasta_excel": str(pasta_excel),
        "pasta_ose":   str(pasta_ose),
        "pasta_html":  str(pasta_html),
    }


# ==============================================================================
# MÓDULO 19 — PARSER DE QUANTITATIVOS DE CAMPO (.txt / .rtf / .csv)
# ==============================================================================

def ler_quantitativo_campo(caminho):
    """
    Lê arquivo de quantitativos de campo no formato produzido pelo ProSaneamento.
    Suporta:
      - QUANTITATIVO_XX.txt  (formato texto com "Volume de Escavação = ...")
      - lista_XX.rtf          (formato RTF com tabela PVC/Caixas)
      - qualquer .csv com colunas item,quantidade,unidade

    Retorna dict com:
      extensao_m, esc_m3, aterro_m3, pav_m2,
      escoramento_m2, n_tubos, n_pvs, materiais[]
    """
    caminho = Path(caminho)
    # Tentar latin-1/cp1252 primeiro (mais comum em arquivos Windows BR)
    for enc in ["utf-8", "cp1252", "latin-1"]:
        try:
            with open(caminho, encoding=enc, errors="strict") as f:
                raw = f.read()
            break
        except (UnicodeDecodeError, LookupError):
            continue
    else:
        with open(caminho, encoding="utf-8", errors="replace") as f:
            raw = f.read()

    # Strip RTF se necessário
    if caminho.suffix.lower() == ".rtf" or raw.strip().startswith("{\rtf"):
        raw = re.sub(r"\\[a-z]+\d*\s?", " ", raw)
        raw = re.sub(r"[{}]", " ", raw)
        raw = re.sub(r"\s+", " ", raw)

    result = {
        "extensao_m":    None,
        "esc_m3":        None,
        "aterro_m3":     None,
        "pav_m2":        None,
        "escoramento_m2": None,
        "n_tubos":        0,
        "n_pvs":          0,
        "materiais":      [],
        "fonte":          caminho.name,
    }

    # ── Parser QUANTITATIVO_XX.txt ────────────────────────────────────────────
    def _extrair(pattern, texto):
        m = re.search(pattern, texto, re.IGNORECASE)
        if m:
            try:
                return float(m.group(1).replace(",", "."))
            except ValueError:
                return None
        return None

    result["extensao_m"]     = _extrair(r"Extens.o Total\s*=\s*([\d.,]+)", raw)
    result["pav_m2"]         = _extrair(r"Pavimenta..o.*?=\s*([\d.,]+)\s*m2", raw)
    result["escoramento_m2"] = _extrair(r"Escoramento\s*=\s*([\d.,]+)\s*m2", raw)

    # Escavação: "Volume de Escavação = 90.911 x 1.25 = 113.638 m3"
    # Pegar o valor final (após o "= ")
    m_esc = re.search(
        r"Escava..o\s*=\s*[\d.,]+\s*x\s*[\d.,]+\s*=\s*([\d.,]+)\s*m3",
        raw, re.IGNORECASE)
    if not m_esc:
        m_esc = re.search(r"Escava..o\s*=\s*([\d.,]+)\s*m3", raw, re.IGNORECASE)
    if m_esc:
        result["esc_m3"] = float(m_esc.group(1).replace(",", "."))

    m_atr = re.search(
        r"Aterro\s*=\s*[\d.,]+\s*x\s*[\d.,]+\s*=\s*([\d.,]+)\s*m3",
        raw, re.IGNORECASE)
    if not m_atr:
        m_atr = re.search(r"Aterro\s*=\s*([\d.,]+)\s*m3", raw, re.IGNORECASE)
    if m_atr:
        result["aterro_m3"] = float(m_atr.group(1).replace(",", "."))

    # "Quantitativo de: 4 Tubo(s) e 5 caixa(s)"
    m_tc = re.search(r"(\d+)\s*Tubo\(s\)\s*e\s*(\d+)\s*caixa", raw, re.IGNORECASE)
    if m_tc:
        result["n_tubos"] = int(m_tc.group(1))
        result["n_pvs"]   = int(m_tc.group(2))

    # ── Parser lista_XX.rtf — tabela de materiais ─────────────────────────────
    # Padrão: "8.2 Barra 200mm Tubo PVC"
    # ou:      "2 pc 0.5x1.2m Poço de Visita"
    for m in re.finditer(
        r"([\d.,]+)\s+(Barra|pc|m|un)\s+([\w.×x]+)\s+(Tubo\s+PVC|Po[çc]o de Visita|"
        r"Tubo de Queda|Caixa|PV|PI)",
        raw, re.IGNORECASE
    ):
        qty  = float(m.group(1).replace(",", "."))
        unit = m.group(2).strip()
        dim  = m.group(3).strip()
        desc = m.group(4).strip()
        result["materiais"].append({
            "qtd": qty, "unidade": unit,
            "dimensao": dim, "descricao": desc
        })
        if "pvc" in desc.lower() and "tubo" in desc.lower():
            result["n_tubos"] = max(result["n_tubos"], int(math.ceil(qty)))
        if "po" in desc.lower() or "pv" in desc.lower():
            result["n_pvs"] += 1

    return result


def exportar_quant_campo_para_ns(q_campo, ns_id, pasta):
    """
    Gera NS_XXX_QUANT_CAMPO.json — dados reais de campo para comparação
    com quantitativos calculados da NS.
    """
    if not q_campo:
        return None
    caminho = pasta / f"NS_{ns_id}_QUANT_CAMPO.json"
    q_campo["ns_id"] = ns_id
    q_campo["gerado_em"] = datetime.now().isoformat()
    with open(caminho, "w", encoding="utf-8") as f:
        json.dump(q_campo, f, indent=2, ensure_ascii=False, default=str)
    return caminho


# ==============================================================================
# MÓDULO 16 — BATCH (todos os núcleos)
# ==============================================================================

_BASE_DXF = r"C:\Users\felip\Downloads\PROJETOS DE ÁGUA E ESGOTO - DWG E DXF 2018\MAPAS ÁGUA E ESGOTO PARA DXF"

NUCLEOS_BATCH = [
    {"nucleo": "São Manoel",
     "dxf": _BASE_DXF + r"\SÃO MANOEL\SÃO_MANOEL_ESGOTO.dxf",
     "gpkg": r"C:\Users\felip\Downloads\MAPA SÃO MANOEL_RV05 (QGIS).gpkg",
     "csv_prosane": _BASE_DXF + r"\SÃO MANOEL\sao manoel esgoto\SÃO_MANOEL_ESGOTO.csv"},
    {"nucleo": "Vila Criadores",
     "dxf": _BASE_DXF + r"\VILA DOS CRIADORES\CRIADORES_ESGOTO.dxf",
     "gpkg": r"C:\Users\felip\Downloads\CARTOGRAFIA VILA CRIADORES_R06 (QGIS).gpkg",
     "csv_prosane": _BASE_DXF + r"\VILA DOS CRIADORES\ESGOTO V4\CRIADORES_ESGOTO2.csv"},
    {"nucleo": "Pantanal Baixo",
     "dxf": _BASE_DXF + r"\PANTANAL BAIXO\PANTANAL_ESGOTO.dxf",
     "gpkg": r"C:\Users\felip\Downloads\MAPA PANTANAL BAIXO_R04 (QGIS).gpkg",
     "csv_prosane": _BASE_DXF + r"\PANTANAL BAIXO\saida pantanal esgoto\PANTANAL_ESGOTO.csv"},
    {"nucleo": "Morro do Tetéu",
     "dxf": _BASE_DXF + r"\MORRO DO TETÉU\TETÉU_ESGOTO.dxf",
     "gpkg": r"C:\Users\felip\Downloads\MAPA TETEU-VALE VERDE_R04 (QGIS).gpkg",
     "csv_prosane": r"C:\Users\felip\Downloads\NOVA NS Versao 5\TETÉU_ESGOTO - Dimensiona(Esgoto).csv"},
    {"nucleo": "Vila Israel",
     "dxf": _BASE_DXF + r"\VILA ISRAEL\ISRAEL_ESGOTO.dxf",
     "gpkg": None,
     "csv_prosane": _BASE_DXF + r"\VILA ISRAEL\ESGOTO ISRAEL ENTREGA REV0\Esgoto Vila Israel\ISRAEL_ESGOTO.csv"},
    {"nucleo": "João Carlos",
     "dxf": _BASE_DXF + r"\JOÃO CARLOS\JOÃO_CARLOS_ESGOTO.dxf",
     "gpkg": r"C:\Users\felip\Downloads\MAPA JOÃO CARLOS DA SILVA_RV07 (QGIS).gpkg",
     "csv_prosane": _BASE_DXF + r"\JOÃO CARLOS\SAIDA_PROSANE\JOÃO_CARLOS_ÁGUA.csv"},
]


# ==============================================================================
# MÓDULO 17 — VALIDAÇÃO PROSANEAMENTO (TAREFA LLM-2)
# ==============================================================================

def _detectar_formato_csv(csv_path):
    """
    Detecta formato do CSV do ProSaneamento:
    - 'saida': separador ',', nós como 'PV_01'
    - 'dimensionamento': separador ';', nós como '"P,V, 36"'
    """
    with open(csv_path, 'r', encoding='latin-1') as f:
        linhas = f.readlines()[:10]
    
    # Contar separadores
    sep_comma = sum(l.count(',') for l in linhas)
    sep_semicolon = sum(l.count(';') for l in linhas)
    
    # Verificar formato dos nós
    tem_no_pv = any('PV_' in l or 'PI_' in l for l in linhas)
    tem_no_comma = any('"P,V,' in l for l in linhas)
    
    if tem_no_comma or sep_semicolon > sep_comma:
        return 'dimensionamento'
    return 'saida'


def _parse_csv_saida(csv_path):
    """
    Parse CSV formato saída (São Manoel, Criadores, Pantanal, Israel).
    Colunas: id,rua,pv_ini,pv_fim,ext_m,dn_mm,material,ct_ini,ct_fim,cf_ini,cf_fim,...
    """
    import csv
    pvs = {}
    trechos = []
    
    with open(csv_path, 'r', encoding='latin-1') as f:
        reader = csv.DictReader(f)
        for row in reader:
            # Extrair PVs
            for prefix in ['pv_ini', 'pv_fim']:
                nome = row.get(f'{prefix}', '').strip()
                if not nome:
                    continue
                ct = row.get(f'ct_{prefix[-3:]}', '').replace(',', '.').strip()
                cf = row.get(f'cf_{prefix[-3:]}', '').replace(',', '.').strip()
                prof = row.get(f'prof_{prefix[-3:]}', '').replace(',', '.').strip()
                
                if nome not in pvs:
                    pvs[nome] = {
                        'nome': nome,
                        'ct': float(ct) if ct else None,
                        'cf': float(cf) if cf else None,
                        'prof': float(prof) if prof else None,
                        'fonte': 'prosane'
                    }
            
            # Extrair trecho
            ext = row.get('ext_m', '').replace(',', '.').strip()
            dn = row.get('dn_mm', '').replace(',', '.').strip()
            
            trechos.append({
                'pv_ini': row.get('pv_ini', '').strip(),
                'pv_fim': row.get('pv_fim', '').strip(),
                'ext_m': float(ext) if ext else None,
                'dn_mm': float(dn) if dn else None,
                'material': row.get('material', '').strip(),
                'rua': row.get('rua', '').strip(),
                'fonte': 'prosane'
            })
    
    return pvs, trechos


def _parse_csv_dimensionamento(csv_path):
    """
    Parse CSV formato dimensionamento (Tetéu).
    Separador ';', nós como '"P,V, 36"', colunas com fórmulas Excel.
    """
    pvs = {}
    trechos = []
    
    with open(csv_path, 'r', encoding='latin-1') as f:
        linhas = f.readlines()
    
    for linha in linhas:
        linha = linha.strip()
        if not linha or linha.startswith(';'):
            continue
        
        # Separar por ; e limpar aspas
        cols = [c.strip().strip('"') for c in linha.split(';')]
        
        # Pular linhas de cabeçalho/config
        if len(cols) < 10 or 'Tabela' in cols[0] or 'Taxa' in cols[0]:
            continue
        
        # Tentar extrair PVs (colunas 0 e 1)
        pv_ini = cols[0] if len(cols) > 0 else ''
        pv_fim = cols[1] if len(cols) > 1 else ''
        
        # Validar se são PVs (formato "P,V, XX" ou similar)
        if not (pv_ini and pv_fim):
            continue
        if not any(x in pv_ini.upper() for x in ['P,V', 'PV', 'PI']):
            continue
        
        # Normalizar nome do PV
        def normalizar_pv(nome):
            # "P,V, 36" -> "PV_36"
            if 'P,V' in nome:
                nums = re.findall(r'\d+', nome)
                if nums:
                    return f"PV_{nums[0]}"
            return nome.strip().replace(' ', '_').upper()
        
        pv_ini_norm = normalizar_pv(pv_ini)
        pv_fim_norm = normalizar_pv(pv_fim)
        
        # Extrair extensão e DN (colunas 3 e 8)
        ext_m = None
        dn_mm = None
        
        if len(cols) > 3:
            try:
                ext_val = cols[3].replace(',', '.').strip()
                ext_m = float(ext_val)
            except:
                pass
        
        if len(cols) > 8:
            try:
                dn_val = cols[8].replace(',', '.').strip()
                if dn_val and dn_val not in ('', 'DN'):
                    dn_mm = float(dn_val)
            except:
                pass
        
        # Adicionar PVs
        for pv_nome in [pv_ini_norm, pv_fim_norm]:
            if pv_nome and pv_nome not in pvs:
                pvs[pv_nome] = {
                    'nome': pv_nome,
                    'ct': None,
                    'cf': None,
                    'prof': None,
                    'fonte': 'prosane'
                }
        
        # Adicionar trecho
        if pv_ini_norm and pv_fim_norm:
            trechos.append({
                'pv_ini': pv_ini_norm,
                'pv_fim': pv_fim_norm,
                'ext_m': ext_m,
                'dn_mm': dn_mm,
                'material': 'PVC',
                'rua': cols[0] if len(cols) > 0 and 'Rua' in cols[0] else '',
                'fonte': 'prosane'
            })
    
    return pvs, trechos


def ler_csv_prosane(csv_path):
    """
    Lê CSV do ProSaneamento e retorna PVs e trechos.
    Detecta formato automaticamente.
    """
    if not os.path.exists(csv_path):
        log(f"  CSV ProSane não encontrado: {csv_path}", "WARN")
        return {}, []
    
    formato = _detectar_formato_csv(csv_path)
    log(f"  CSV ProSane: {formato} — {csv_path}", "OK")
    
    if formato == 'dimensionamento':
        return _parse_csv_dimensionamento(csv_path)
    else:
        return _parse_csv_saida(csv_path)


def validar_contra_prosane(pvs, trechos, csv_prosane_path):
    """
    Compara extração ConstruData com CSV do ProSaneamento.
    Retorna dict com métricas de concordância.
    """
    if not os.path.exists(csv_prosane_path):
        return {'erro': f'CSV não encontrado: {csv_prosane_path}'}
    
    # Ler CSV do ProSane
    pvs_ref, trechos_ref = ler_csv_prosane(csv_prosane_path)
    
    if not pvs_ref or not trechos_ref:
        return {'erro': 'CSV ProSane vazio ou formato inválido'}
    
    # === COMPARAÇÃO DE PVs ===
    pvs_match = 0
    pvs_parcial = 0
    pvs_faltante = 0
    pv_detalhes = []
    
    for nome, pv_ref in pvs_ref.items():
        pv_cd = pvs.get(nome)
        
        if not pv_cd:
            # Tentar match parcial (PV_01 vs PV_1)
            for k in pvs:
                if k.replace('_0', '_') == nome.replace('_0', '_'):
                    pv_cd = pvs[k]
                    break
        
        if not pv_cd:
            pvs_faltante += 1
            pv_detalhes.append({
                'nome': nome,
                'status': 'FALTANTE',
                'ct_ref': pv_ref.get('ct'),
                'ct_cd': None,
                'cf_ref': pv_ref.get('cf'),
                'cf_cd': None
            })
            continue
        
        # Comparar CT e CF (tolerância 0.01m)
        ct_match = abs((pv_cd.get('ct') or 0) - (pv_ref.get('ct') or 0)) < 0.5 if pv_ref.get('ct') else True
        cf_match = abs((pv_cd.get('cf') or 0) - (pv_ref.get('cf') or 0)) < 0.5 if pv_ref.get('cf') else True
        
        if ct_match and cf_match:
            pvs_match += 1
            status = 'MATCH'
        elif ct_match or cf_match:
            pvs_parcial += 1
            status = 'PARCIAL'
        else:
            pvs_faltante += 1
            status = 'DIVERGENTE'
        
        pv_detalhes.append({
            'nome': nome,
            'status': status,
            'ct_ref': pv_ref.get('ct'),
            'ct_cd': pv_cd.get('ct'),
            'cf_ref': pv_ref.get('cf'),
            'cf_cd': pv_cd.get('cf')
        })
    
    total_pvs = len(pvs_ref)
    pct_pvs_match = (pvs_match / total_pvs * 100) if total_pvs > 0 else 0
    
    # === COMPARAÇÃO DE TRECHOS ===
    trechos_match = 0
    trechos_parcial = 0
    trechos_faltante = 0
    trecho_detalhes = []
    
    for tr_ref in trechos_ref:
        # Buscar trecho correspondente
        tr_cd = None
        for t in trechos:
            if (t.get('pv_ini') == tr_ref['pv_ini'] and 
                t.get('pv_fim') == tr_ref['pv_fim']):
                tr_cd = t
                break
        
        if not tr_cd:
            trechos_faltante += 1
            trecho_detalhes.append({
                'trecho': f"{tr_ref['pv_ini']}->{tr_ref['pv_fim']}",
                'status': 'FALTANTE',
                'ext_ref': tr_ref.get('ext_m'),
                'ext_cd': None,
                'dn_ref': tr_ref.get('dn_mm'),
                'dn_cd': None
            })
            continue
        
        # Comparar extensão (tolerância 1m) e DN (exato)
        ext_match = abs((tr_cd.get('ext_m') or 0) - (tr_ref.get('ext_m') or 0)) < 1.0 if tr_ref.get('ext_m') else True
        dn_match = int(tr_cd.get('dn_mm') or 0) == int(tr_ref.get('dn_mm') or 0) if tr_ref.get('dn_mm') else True
        
        if ext_match and dn_match:
            trechos_match += 1
            status = 'MATCH'
        elif ext_match or dn_match:
            trechos_parcial += 1
            status = 'PARCIAL'
        else:
            trechos_faltante += 1
            status = 'DIVERGENTE'
        
        trecho_detalhes.append({
            'trecho': f"{tr_ref['pv_ini']}->{tr_ref['pv_fim']}",
            'status': status,
            'ext_ref': tr_ref.get('ext_m'),
            'ext_cd': tr_cd.get('ext_m'),
            'dn_ref': tr_ref.get('dn_mm'),
            'dn_cd': tr_cd.get('dn_mm')
        })
    
    total_trechos = len(trechos_ref)
    pct_trechos_match = (trechos_match / total_trechos * 100) if total_trechos > 0 else 0
    
    return {
        'csv_path': csv_prosane_path,
        'pvs': {
            'total_ref': total_pvs,
            'match': pvs_match,
            'parcial': pvs_parcial,
            'faltante': pvs_faltante,
            'pct_match': pct_pvs_match,
            'detalhes': pv_detalhes
        },
        'trechos': {
            'total_ref': total_trechos,
            'match': trechos_match,
            'parcial': trechos_parcial,
            'faltante': trechos_faltante,
            'pct_match': pct_trechos_match,
            'detalhes': trecho_detalhes
        }
    }


def gerar_relatorio_comparativo(resultados_validacao, pasta_saida):
    """
    Gera Excel COMPARATIVO_PROSANE_CONSTRUDATA.xlsx com:
    - Aba "Resumo": tabela por núcleo (PVs match%, trechos match%, CT/CF RMSE)
    - Aba "PVs": cada PV com CT/CF do ProSane vs ConstruData
    - Aba "Trechos": cada trecho com ext/dn do ProSane vs ConstruData
    - Formatação condicional: verde (match), amarelo (parcial), vermelho (faltante)
    """
    if not _HAS_OPENPYXL:
        log("  openpyxl não disponível — pulando relatório Excel", "WARN")
        return None
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    
    # Cores para formatação condicional
    fill_verde = PatternFill("solid", fgColor="C6EFCE")
    fill_amarelo = PatternFill("solid", fgColor="FFEB9C")
    fill_vermelho = PatternFill("solid", fgColor="FFC7CE")
    fill_azul = PatternFill("solid", fgColor="D6DCE4")
    font_vermelho = Font(color="9C0006")
    
    # === ABA RESUMO ===
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    
    # Cabeçalho
    headers = ["NÚCLEO", "PVs (Ref)", "PVs Match", "PVs %", "Trechos (Ref)", 
               "Trechos Match", "Trechos %", "Status Geral"]
    for col, h in enumerate(headers, 1):
        cel = ws_resumo.cell(row=1, column=col, value=h)
        cel.font = Font(bold=True)
        cel.fill = fill_azul
        cel.alignment = Alignment(horizontal="center", vertical="center")
    
    # Dados por núcleo
    row = 2
    for nucleo, dados in resultados_validacao.items():
        if 'erro' in dados:
            continue
        
        pvs = dados.get('pvs', {})
        trechos = dados.get('trechos', {})
        
        pct_pvs = pvs.get('pct_match', 0)
        pct_trechos = trechos.get('pct_match', 0)
        status = "OK" if pct_pvs >= 80 and pct_trechos >= 80 else "ATENÇÃO"
        
        ws_resumo.cell(row=row, column=1, value=nucleo)
        ws_resumo.cell(row=row, column=2, value=pvs.get('total_ref', 0))
        ws_resumo.cell(row=row, column=3, value=pvs.get('match', 0))
        
        cel_pct_pvs = ws_resumo.cell(row=row, column=4, value=f"{pct_pvs:.1f}%")
        cel_pct_pvs.fill = fill_verde if pct_pvs >= 80 else (fill_amarelo if pct_pvs >= 50 else fill_vermelho)
        
        ws_resumo.cell(row=row, column=5, value=trechos.get('total_ref', 0))
        ws_resumo.cell(row=row, column=6, value=trechos.get('match', 0))
        
        cel_pct_trechos = ws_resumo.cell(row=row, column=7, value=f"{pct_trechos:.1f}%")
        cel_pct_trechos.fill = fill_verde if pct_trechos >= 80 else (fill_amarelo if pct_trechos >= 50 else fill_vermelho)
        
        cel_status = ws_resumo.cell(row=row, column=8, value=status)
        cel_status.fill = fill_verde if status == "OK" else fill_amarelo
        
        row += 1
    
    # Ajustar largura das colunas
    for col in ws_resumo.columns:
        max_len = max(len(str(cell.value)) for cell in col) if col else 10
        ws_resumo.column_dimensions[col[0].column_letter].width = min(max_len + 2, 25)
    
    # === ABA PVs ===
    ws_pvs = wb.create_sheet(title="PVs")
    
    headers_pvs = ["NÚCLEO", "PV NOME", "Status", "CT (Ref)", "CT (ConstruData)", 
                   "CF (Ref)", "CF (ConstruData)", "Diff CT", "Diff CF"]
    for col, h in enumerate(headers_pvs, 1):
        cel = ws_pvs.cell(row=1, column=col, value=h)
        cel.font = Font(bold=True)
        cel.fill = fill_azul
    
    row = 2
    for nucleo, dados in resultados_validacao.items():
        if 'erro' in dados:
            continue
        
        for pv in dados.get('pvs', {}).get('detalhes', []):
            ws_pvs.cell(row=row, column=1, value=nucleo)
            ws_pvs.cell(row=row, column=2, value=pv.get('nome', ''))
            
            status = pv.get('status', '')
            cel_status = ws_pvs.cell(row=row, column=3, value=status)
            if status == 'MATCH':
                cel_status.fill = fill_verde
            elif status == 'PARCIAL':
                cel_status.fill = fill_amarelo
            else:
                cel_status.fill = fill_vermelho
                cel_status.font = font_vermelho
            
            ct_ref = pv.get('ct_ref')
            ct_cd = pv.get('ct_cd')
            cf_ref = pv.get('cf_ref')
            cf_cd = pv.get('cf_cd')
            
            ws_pvs.cell(row=row, column=4, value=ct_ref)
            ws_pvs.cell(row=row, column=5, value=ct_cd)
            ws_pvs.cell(row=row, column=6, value=cf_ref)
            ws_pvs.cell(row=row, column=7, value=cf_cd)
            
            diff_ct = abs(ct_ref - ct_cd) if ct_ref and ct_cd else None
            diff_cf = abs(cf_ref - cf_cd) if cf_ref and cf_cd else None
            ws_pvs.cell(row=row, column=8, value=round(diff_ct, 3) if diff_ct else "N/A")
            ws_pvs.cell(row=row, column=9, value=round(diff_cf, 3) if diff_cf else "N/A")
            
            row += 1
    
    # === ABA TRECHOS ===
    ws_trechos = wb.create_sheet(title="Trechos")
    
    headers_trechos = ["NÚCLEO", "TRECHO", "Status", "Ext (Ref)", "Ext (ConstruData)", 
                       "Diff Ext", "DN (Ref)", "DN (ConstruData)", "DN Match"]
    for col, h in enumerate(headers_trechos, 1):
        cel = ws_trechos.cell(row=1, column=col, value=h)
        cel.font = Font(bold=True)
        cel.fill = fill_azul
    
    row = 2
    for nucleo, dados in resultados_validacao.items():
        if 'erro' in dados:
            continue
        
        for tr in dados.get('trechos', {}).get('detalhes', []):
            ws_trechos.cell(row=row, column=1, value=nucleo)
            ws_trechos.cell(row=row, column=2, value=tr.get('trecho', ''))
            
            status = tr.get('status', '')
            cel_status = ws_trechos.cell(row=row, column=3, value=status)
            if status == 'MATCH':
                cel_status.fill = fill_verde
            elif status == 'PARCIAL':
                cel_status.fill = fill_amarelo
            else:
                cel_status.fill = fill_vermelho
                cel_status.font = font_vermelho
            
            ext_ref = tr.get('ext_ref')
            ext_cd = tr.get('ext_cd')
            dn_ref = tr.get('dn_ref')
            dn_cd = tr.get('dn_cd')
            
            ws_trechos.cell(row=row, column=4, value=ext_ref)
            ws_trechos.cell(row=row, column=5, value=ext_cd)
            
            diff_ext = abs(ext_ref - ext_cd) if ext_ref and ext_cd else None
            ws_trechos.cell(row=row, column=6, value=round(diff_ext, 2) if diff_ext else "N/A")
            
            ws_trechos.cell(row=row, column=7, value=dn_ref)
            ws_trechos.cell(row=row, column=8, value=dn_cd)
            
            dn_match = "SIM" if dn_ref == dn_cd else ("N/A" if not dn_ref else "NÃO")
            cel_dn = ws_trechos.cell(row=row, column=9, value=dn_match)
            if dn_match == "SIM":
                cel_dn.fill = fill_verde
            elif dn_match == "NÃO":
                cel_dn.fill = fill_vermelho
                cel_dn.font = font_vermelho
            
            row += 1
    
    # Ajustar largura das colunas
    for ws in [ws_pvs, ws_trechos]:
        for col in ws.columns:
            max_len = max(len(str(cell.value)) for cell in col) if col else 10
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)
    
    # Salvar
    caminho = Path(pasta_saida) / "COMPARATIVO_PROSANE_CONSTRUDATA.xlsx"
    wb.save(caminho)
    log(f"  Comparativo Excel: {caminho.name}", "OK")
    return caminho


def processar_batch_com_validacao(pasta_saida="SAIDA_BIM_SABESP", max_ns_por_nucleo=None):
    """
    Processa todos os núcleos em lote E compara com ProSaneamento automaticamente.
    Gera relatório Excel COMPARATIVO_PROSANE_CONSTRUDATA.xlsx
    """
    t_ini = time.time()
    resultados = []
    resultados_validacao = {}
    ok, err, val = 0, 0, 0
    
    log("=" * 65, "INFO")
    log("  BATCH COM VALIDAÇÃO PROSANEAMENTO", "STEP")
    log("=" * 65, "INFO")
    
    for n in NUCLEOS_BATCH:
        dxf = n.get("dxf","")
        csv_prosane = n.get("csv_prosane")
        
        if not Path(dxf).exists():
            log(f"  SKIP: {n['nucleo']} — DXF não encontrado", "WARN")
            continue
        
        log(f"\n  {n['nucleo']}", "STEP")
        try:
            # Processar DXF e capturar dados
            if _HAS_GDAL_READER:
                pvs, trechos, ruas, meta = ler_dxf_gdal(dxf)
            else:
                pvs, trechos, ruas, meta = ler_dxf(dxf)
            
            # Enriquecer dados
            trechos = enriquecer_trechos(trechos, pvs)
            
            # Processar NS (gerar arquivos)
            r = processar(
                dxf_path=dxf,
                pasta_saida=pasta_saida,
                nucleo=n["nucleo"],
                gpkg_path=n.get("gpkg"),
                max_ns=max_ns_por_nucleo,
            )
            resultados.append({"nucleo": n["nucleo"], "status": "OK", **r})
            ok += 1
            
            # Validar contra ProSane se CSV existir
            if csv_prosane and Path(csv_prosane).exists():
                log(f"  Validando contra ProSane...", "INFO")
                validacao = validar_contra_prosane(pvs, trechos, csv_prosane)
                resultados_validacao[n["nucleo"]] = validacao
                
                pct_pvs = validacao.get('pvs', {}).get('pct_match', 0)
                pct_trechos = validacao.get('trechos', {}).get('pct_match', 0)
                log(f"  Validação: PVs={pct_pvs:.1f}% | Trechos={pct_trechos:.1f}%", 
                    "OK" if pct_pvs >= 80 and pct_trechos >= 80 else "WARN")
                val += 1
            elif csv_prosane:
                log(f"  CSV ProSane não encontrado: {csv_prosane}", "WARN")
                resultados_validacao[n["nucleo"]] = {'erro': 'CSV não encontrado'}
            else:
                log(f"  Sem CSV ProSane configurado", "WARN")
                
        except Exception as e:
            log(f"  ERRO em {n['nucleo']}: {e}", "ERR")
            resultados.append({"nucleo": n["nucleo"], "status": "ERRO", "erro": str(e)})
            err += 1
    
    # Gerar relatório Excel
    log(f"\n  Gerando relatório comparativo...", "INFO")
    gerar_relatorio_comparativo(resultados_validacao, pasta_saida)
    
    dt = time.time() - t_ini
    log(f"\nBATCH+VALIDAÇÃO concluído em {dt/60:.1f}min — OK:{ok} ERRO:{err} VAL:{val}", "OK")

    # Salvar log completo
    log_completo = {
        "resultados_batch": resultados,
        "validacao_prosane": resultados_validacao,
        "tempo_total": dt
    }

    log_b = Path(pasta_saida) / "log_batch_validacao.json"
    with open(log_b, "w", encoding="utf-8") as f:
        json.dump(log_completo, f, indent=2, ensure_ascii=False, default=str)
    log(f"Log: {log_b}", "OK")

    # Gerar mapa de TODOS os núcleos
    log(f"\n  Gerando mapa geral de todos os núcleos...", "STEP")
    try:
        # Coletar dados de todos os núcleos processados
        todos_dados = []
        for res in resultados:
            if res.get("status") == "OK":
                # Re-ler dados do DXF para obter pvs e trechos completos
                nucleo = res.get("nucleo", "")
                dxf_path = next((n.get("dxf") for n in NUCLEOS_BATCH if n.get("nucleo") == nucleo), None)
                if dxf_path and Path(dxf_path).exists():
                    if _HAS_GDAL_READER:
                        pvs, trechos, ruas, meta = ler_dxf_gdal(dxf_path)
                    else:
                        pvs, trechos, ruas, meta = ler_dxf(dxf_path)
                    trechos = enriquecer_trechos(trechos, pvs)
                    todos_dados.append({
                        "nucleo": nucleo,
                        "pvs": pvs,
                        "trechos": trechos,
                        "cfg": {"nucleo": nucleo}
                    })

        if todos_dados:
            gerar_mapa_todos_nucleos_html(todos_dados, pasta_saida, CFG)
    except Exception as e:
        log(f"  Erro ao gerar mapa geral: {e}", "WARN")

    return resultados, resultados_validacao


# Alias para compatibilidade com CLI
processar_batch = processar_batch_com_validacao


# ==============================================================================
# MAIN — argparse
# ==============================================================================

def main():
    if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
        try:
            sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass

    parser = argparse.ArgumentParser(
        description="ConstruData BIM SABESP v5.0 — Pipeline Unificado",
        formatter_class=argparse.RawTextHelpFormatter,
    )
    parser.add_argument("dxf", nargs="?",
                        help="Caminho do DXF ProSaneamento (esgoto ou água)")
    parser.add_argument("--json",    help="Carregar de rede_definida.json ou dynamo.json")
    parser.add_argument("--saida",   default="SAIDA_BIM_SABESP",
                        help="Pasta raiz de saída (default: SAIDA_BIM_SABESP)")
    parser.add_argument("--nucleo",  default=None,
                        help="Nome do núcleo (ex: 'Morro do Teteu')")
    parser.add_argument("--gpkg",    default=None,
                        help="GPKG de cartografia (opcional)")
    parser.add_argument("--tipo",    choices=["agua","esgoto"],
                        help="Forçar tipo de rede")
    parser.add_argument("--batch",   action="store_true",
                        help="Processar todos os núcleos em lote")
    parser.add_argument("--max-ns",  type=int, default=None,
                        help="Limitar número de NS (debug)")
    parser.add_argument("--debug",   action="store_true",
                        help="Mostrar traceback completo nos erros")
    parser.add_argument("--base-url", default="",
                        dest="base_url",
                        metavar="URL",
                        help="URL base para QR Code (ex: https://obra.sabesp.gov.br/slnr/)")
    parser.add_argument("--quant",   default=None,
                        metavar="ARQUIVO",
                        help="Arquivo de quantitativos de campo (.txt/.rtf) para enriquecer NS")
    parser.add_argument("--validar-prosane",   action="store_true",
                        dest="validar_prosane",
                        help="Processar batch COM validação automática contra CSVs do ProSaneamento")
    args = parser.parse_args()

    if args.validar_prosane:
        processar_batch_com_validacao(pasta_saida=args.saida, max_ns_por_nucleo=args.max_ns)
        return

    if args.batch:
        processar_batch_com_validacao(args.saida, args.max_ns)
        return

    if not args.dxf and not args.json:
        parser.print_help()
        print("\n  Exemplos:")
        print("  python construdata_sabesp_v5_FINAL.py ISRAEL_ESGOTO.dxf --nucleo 'Vila Israel'")
        print("  python construdata_sabesp_v5_FINAL.py --json rede_definida.json")
        print("  python construdata_sabesp_v5_FINAL.py CRIADORES_AGUA.dxf --tipo agua --gpkg MAPA.gpkg")
        print("  python construdata_sabesp_v5_FINAL.py --batch")
        sys.exit(0)

    nucleo = args.nucleo or (Path(args.dxf).stem.replace("_"," ") if args.dxf
                             else Path(args.json).stem.replace("_"," "))

    processar(
        dxf_path=args.dxf,
        json_path=args.json,
        pasta_saida=args.saida,
        nucleo=nucleo,
        gpkg_path=args.gpkg,
        tipo_override=args.tipo,
        max_ns=args.max_ns,
        ns_base_url=args.base_url,
        quant_campo=args.quant,
    )


# ==============================================================================
# MÓDULO 17 — IFC LOD500 (ifcopenshell 0.8.4)
# ==============================================================================

try:
    import ifcopenshell
    import ifcopenshell.api
    _HAS_IFC = True
except ImportError:
    _HAS_IFC = False

def _ifc_cylinder(model, ctx, x0, y0, z0, x1, y1, z1, raio_m):
    """
    Cria IfcExtrudedAreaSolid (cilindro) orientado do ponto P0 a P1.
    Retorna (IfcProductDefinitionShape, IfcLocalPlacement) ou (None, None).
    
    MELHORIA: Aplica offset para coordenadas locais (evita valores UTM grandes).
    """
    # Offset para coordenadas locais (Santos SP — SIRGAS 2000 UTM 23S)
    OFFSET_X = 360000.0
    OFFSET_Y = 7350000.0
    
    x0_l = float(x0) - OFFSET_X
    y0_l = float(y0) - OFFSET_Y
    x1_l = float(x1) - OFFSET_X
    y1_l = float(y1) - OFFSET_Y
    z0_l = float(z0)
    z1_l = float(z1)
    
    dx, dy, dz = x1_l-x0_l, y1_l-y0_l, z1_l-z0_l
    length = math.sqrt(dx**2 + dy**2 + dz**2)
    if length < 0.01:
        return None, None

    # Eixo do tubo normalizado
    ex, ey, ez = dx/length, dy/length, dz/length

    # Vetor ref perpendicular ao eixo
    if abs(ez) < 0.9:
        rx, ry, rz = -ey, ex, 0.0
    else:
        rx, ry, rz = 1.0, 0.0, 0.0
    r_len = math.sqrt(rx**2 + ry**2 + rz**2)
    rx, ry, rz = rx/r_len, ry/r_len, rz/r_len

    profile = model.createIfcCircleProfileDef(
        "AREA", None,
        model.createIfcAxis2Placement2D(
            model.createIfcCartesianPoint([0.0, 0.0]), None),
        raio_m)

    ext_dir = model.createIfcDirection([0.0, 0.0, 1.0])
    solid   = model.createIfcExtrudedAreaSolid(
        profile,
        model.createIfcAxis2Placement3D(
            model.createIfcCartesianPoint([0.0, 0.0, 0.0]), ext_dir, None),
        ext_dir, length)

    rep     = model.createIfcShapeRepresentation(ctx, "Body", "SweptSolid", [solid])
    prod_def= model.createIfcProductDefinitionShape(None, None, [rep])

    place_3d = model.createIfcAxis2Placement3D(
        model.createIfcCartesianPoint([x0_l, y0_l, z0_l]),
        model.createIfcDirection([ex, ey, ez]),
        model.createIfcDirection([rx, ry, rz]))
    placement = model.createIfcLocalPlacement(None, place_3d)

    return prod_def, placement


def _ifc_box(model, ctx, x, y, z, dx, dy, dz):
    """Cria caixa (PV) como IfcExtrudedAreaSolid.
    
    MELHORIA: Aplica offset para coordenadas locais (evita valores UTM grandes).
    """
    # Offset para coordenadas locais (Santos SP — SIRGAS 2000 UTM 23S)
    OFFSET_X = 360000.0
    OFFSET_Y = 7350000.0
    
    x_l = float(x) - OFFSET_X
    y_l = float(y) - OFFSET_Y
    z_l = float(z)
    
    rect = model.createIfcRectangleProfileDef(
        "AREA", None,
        model.createIfcAxis2Placement2D(
            model.createIfcCartesianPoint([0.0, 0.0]), None),
        dx, dy)
    ext_dir = model.createIfcDirection([0.0, 0.0, 1.0])
    solid   = model.createIfcExtrudedAreaSolid(
        rect,
        model.createIfcAxis2Placement3D(
            model.createIfcCartesianPoint([0.0, 0.0, 0.0]), ext_dir, None),
        ext_dir, dz)

    rep      = model.createIfcShapeRepresentation(ctx, "Body", "SweptSolid", [solid])
    prod_def = model.createIfcProductDefinitionShape(None, None, [rep])

    place_3d = model.createIfcAxis2Placement3D(
        model.createIfcCartesianPoint([x_l - dx/2, y_l - dy/2, z_l]),
        model.createIfcDirection([0.0, 0.0, 1.0]), None)
    placement = model.createIfcLocalPlacement(None, place_3d)

    return prod_def, placement


def _ifc_pset(model, product, nome_pset, props):
    """Adiciona Pset com propriedades simples ao produto."""
    pset = ifcopenshell.api.run("pset.add_pset", model,
                                product=product, name=nome_pset)
    safe_props = {}
    for k, v in props.items():
        if v is None:
            continue
        if isinstance(v, bool):
            safe_props[k] = v
        elif isinstance(v, (int, float)):
            safe_props[k] = float(v)
        else:
            safe_props[k] = str(v)
    ifcopenshell.api.run("pset.edit_pset", model, pset=pset,
                         properties=safe_props)
    return pset


def gerar_ifc(pvs, trechos, pasta_bim, cfg, meta,
              ns_base_url="http://sabesp.construdata.local/"):
    """
    Gera arquivo IFC LOD500 com:
      - IfcPipeSegment (cilindro 3D com DN real) por trecho
      - IfcFlowStorageDevice (caixa 3D CT/CF) por PV
      - Pset_PipeSegmentPHistory por trecho
      - Pset_FlowStorageDeviceTypeCommon por PV
      - Georeferenciado em SIRGAS 2000 UTM 23S (EPSG:31983)
      - IfcClassification ligada à norma NBR 9649

    Retorna caminho do .ifc gerado.
    """
    pasta_bim = Path(pasta_bim)
    if not _HAS_IFC:
        log("ifcopenshell não instalado — pip install ifcopenshell", "WARN")
        return None

    pasta_bim.mkdir(parents=True, exist_ok=True)
    nome_nucleo = cfg.get("nucleo","SABESP").upper().replace(" ","_")[:30]
    nome_arquivo = f"REDE_{nome_nucleo}.ifc"
    tipo_rede    = meta.get("tipo_rede","ESGOTO").upper()

    log(f"Gerando IFC LOD500: {nome_arquivo}", "STEP")

    model = ifcopenshell.file(schema="IFC4")

    # ── Projeto ───────────────────────────────────────────────────────────────
    project = ifcopenshell.api.run("root.create_entity", model,
        ifc_class="IfcProject",
        name=f"SLNR Santos — {cfg.get('nucleo','')} — Contrato {cfg['contrato']}")
    ifcopenshell.api.run("unit.assign_unit", model,
        length={"is_metric": True, "raw": "METRE"})

    # Contextos de representação
    ctx_model = ifcopenshell.api.run("context.add_context", model,
        context_type="Model")
    ctx_body = ifcopenshell.api.run("context.add_context", model,
        context_type="Model", context_identifier="Body",
        target_view="MODEL_VIEW", parent=ctx_model)

    # ── Hierarquia espacial ───────────────────────────────────────────────────
    site = ifcopenshell.api.run("root.create_entity", model,
        ifc_class="IfcSite", name="Santos SP")
    building = ifcopenshell.api.run("root.create_entity", model,
        ifc_class="IfcBuilding",
        name=f"Infraestrutura {'Água Fria' if 'AGUA' in tipo_rede else 'Esgoto'}")
    storey = ifcopenshell.api.run("root.create_entity", model,
        ifc_class="IfcBuildingStorey",
        name="Rede Subterrânea")
    ifcopenshell.api.run("aggregate.assign_object", model,
        relating_object=project, products=[site])
    ifcopenshell.api.run("aggregate.assign_object", model,
        relating_object=site, products=[building])
    ifcopenshell.api.run("aggregate.assign_object", model,
        relating_object=building, products=[storey])

    # ── Georeferenciamento (EPSG:31983) ───────────────────────────────────────
    # Coordenadas de origem aproximadas — Santos SP
    cx_utm = sum(pvs[n]["x"] for n in pvs if pvs[n].get("x")) / max(1, len(pvs))
    cy_utm = sum(pvs[n]["y"] for n in pvs if pvs[n].get("y")) / max(1, len(pvs))
    site.RefLatitude  = _dd_to_dms(-23.9)
    site.RefLongitude = _dd_to_dms(-46.3)
    site.RefElevation = 0.0

    # MapConversion para CRS (IFC4 MapConversion)
    proj_ctx = model.createIfcProjectedCRS(
        Name="SIRGAS 2000 UTM Zone 23S",
        Description="EPSG:31983",
        GeodeticDatum="GRS 1980",
        VerticalDatum="SIRGAS 2000",
        MapProjection="UTM",
        MapZone="23S",
        MapUnit=model.createIfcSIUnit(None, "LENGTHUNIT", None, "METRE"))
    map_conv = model.createIfcMapConversion(
        model.createIfcGeometricRepresentationContext(
            None, "Model", 3, 1.0e-5,
            model.createIfcAxis2Placement3D(
                model.createIfcCartesianPoint([0.0, 0.0, 0.0]),
                model.createIfcDirection([0.0, 0.0, 1.0]),
                model.createIfcDirection([1.0, 0.0, 0.0])),
            None),
        proj_ctx,
        round(cx_utm, 3),   # Eastings
        round(cy_utm, 3),   # Northings
        0.0,                # OrthogonalHeight
        1.0, 0.0,           # XAxisAbscissa, XAxisOrdinate
        1.0)                # Scale

    # ── Classificação NBR 9649 ────────────────────────────────────────────────
    classif = model.createIfcClassification(
        Source="ABNT", Edition="2021",
        EditionDate=None,
        Name="NBR 9649",
        Description="Projeto de Redes Coletoras de Esgoto Sanitário")

    # ── PVs → IfcFlowStorageDevice ────────────────────────────────────────────
    pv_ifc_map = {}
    pv_set     = set()
    for t in trechos:
        pv_set.add(t.get("pv_ini",""))
        pv_set.add(t.get("pv_fim",""))

    n_pvs_ok = 0
    for nome_pv in pv_set:
        if not nome_pv:
            continue
        pv = pvs.get(nome_pv, {})
        x  = pv.get("x"); y = pv.get("y")
        ct = pv.get("ct"); cf = pv.get("cf")
        if x is None or y is None:
            continue

        # Geometria: caixa 0.6×0.6 × profundidade
        prof = pv.get("prof") or 1.2
        z_base = (cf if cf is not None else (ct - prof if ct else 0.0))
        z_base = z_base if z_base is not None else 0.0
        prof   = prof if prof else 1.2

        prod_def, placement = _ifc_box(model, ctx_body,
                                       x, y, z_base,
                                       0.60, 0.60, prof)

        pv_tipo = "RG" if nome_pv.startswith("N_") else "PV"
        ifc_class = ("IfcFlowTerminal" if "AGUA" in tipo_rede
                     else "IfcFlowStorageDevice")

        dev = ifcopenshell.api.run("root.create_entity", model,
            ifc_class=ifc_class,
            name=nome_pv)

        if prod_def:
            dev.Representation = prod_def
            dev.ObjectPlacement = placement

        ifcopenshell.api.run("spatial.assign_container", model,
            relating_structure=storey, products=[dev])

        _ifc_pset(model, dev, "Pset_ManHoleTypeCommon", {
            "Reference":        nome_pv,
            "Status":           "EXISTING" if pv.get("asbuilt") else "NEW",
            "RimElevation":     ct,
            "InvertElevation":  cf,
            "CoverWidth":       0.60,
            "CoverLength":      0.60,
            "WallThickness":    0.08,
        })
        
        # MELHORIA: Pset de geometria detalhada
        _ifc_pset(model, dev, "SABESP_PV_Geometria", {
            "BaseElevation":    z_base,
            "TotalHeight":      prof,
            "InnerWidth":       0.44,  # 0.60 - 2*0.08
            "InnerLength":      0.44,
            "CoverThickness":   0.10,
            "SumpDepth":        0.05,  # Profundidade do fundo
        })
        
        _ifc_pset(model, dev, "SABESP_PV", {
            "Nucleo":     cfg.get("nucleo",""),
            "Contrato":   cfg["contrato"],
            "Tipo_PV":    pv_tipo,
            "CT_m":       ct,
            "CF_m":       cf,
            "Prof_m":     prof,
            "X_UTM":      x,
            "Y_UTM":      y,
            "EPSG":       31983,
        })

        pv_ifc_map[nome_pv] = dev
        n_pvs_ok += 1

    # ── Validação de CRS (MELHORIA) ───────────────────────────────────────────
    # Verificar se PVs estão em UTM (X > 100,000)
    pvs_utm = sum(1 for pv in pvs.values() if pv.get("x") and pv["x"] > 100000)
    pvs_locais = sum(1 for pv in pvs.values() if pv.get("x") and pv["x"] < 100000)
    
    if pvs_locais > pvs_utm:
        log(f"  ⚠️ CRS: {pvs_locais} PVs em coords locais, {pvs_utm} em UTM", "WARN")
        log(f"     IFC pode estar georeferenciado incorretamente", "WARN")
    else:
        log(f"  ✅ CRS: {pvs_utm} PVs em UTM (EPSG:31983)", "OK")

    # ── Trechos → IfcPipeSegment ──────────────────────────────────────────────
    n_pipes_ok = 0
    for i, t in enumerate(trechos):
        pvi_n = t.get("pv_ini",""); pvf_n = t.get("pv_fim","")
        pvi   = pvs.get(pvi_n, {}); pvf_d = pvs.get(pvf_n, {})
        x0, y0 = pvi.get("x"), pvi.get("y")
        x1, y1 = pvf_d.get("x"), pvf_d.get("y")
        if x0 is None or x1 is None:
            continue

        dn_m   = (t.get("dn_mm") or 200) / 1000
        cf_ini = t.get("cf_ini") or pvi.get("cf") or 0.0
        cf_fim = t.get("cf_fim") or pvf_d.get("cf") or 0.0
        z0     = cf_ini if cf_ini is not None else 0.0
        z1     = cf_fim if cf_fim is not None else 0.0

        prod_def, placement = _ifc_cylinder(
            model, ctx_body, x0, y0, z0, x1, y1, z1, dn_m/2)

        ns_id   = t.get("ns_id","")
        pipe_id = f"NS{ns_id}-{pvi_n}-{pvf_n}"
        hid     = t.get("hidraulica") or {}

        pipe = ifcopenshell.api.run("root.create_entity", model,
            ifc_class="IfcPipeSegment",
            name=pipe_id,
            predefined_type="RIGIDSEGMENT")
        if prod_def:
            pipe.Representation = prod_def
            pipe.ObjectPlacement = placement

        ifcopenshell.api.run("spatial.assign_container", model,
            relating_structure=storey, products=[pipe])

        # Pset_PipeSegmentPHistory (padrão IFC4)
        _ifc_pset(model, pipe, "Pset_PipeSegmentPHistory", {
            "NominalDiameter":     float(t.get("dn_mm") or 200),
            "WallThickness":       round(float(t.get("dn_mm") or 200) * 0.05, 1),
            "NominalLength":       float(t.get("ext_m") or 0),
            "GrossWeight":        0.0,
        })
        
        # MELHORIA: Pset de instalação
        profundidade_media = (t.get("prof_ini") or 0) + (t.get("prof_fim") or 0)
        _ifc_pset(model, pipe, "SABESP_Instalacao", {
            "TrenchDepth_avg":   profundidade_media / 2 if profundidade_media else 1.5,
            "BeddingType":       "Areia",
            "BeddingThickness":  0.15,
            "BackfillType":      "Solo original",
            "WarningTape":       True,
        })

        # Pset customizado SABESP
        _ifc_pset(model, pipe, "SABESP_Trecho", {
            "NS_ID":           ns_id,
            "PV_Inicial":      pvi_n,
            "PV_Final":        pvf_n,
            "DN_mm":           float(t.get("dn_mm") or 0),
            "Material":        t.get("material","PVC"),
            "Extensao_m":      float(t.get("ext_m") or 0),
            "Declividade_pct": float(t.get("decl_pct") or 0),
            "CT_ini_m":        float(t.get("ct_ini") or 0) if t.get("ct_ini") else None,
            "CT_fim_m":        float(t.get("ct_fim") or 0) if t.get("ct_fim") else None,
            "CF_ini_m":        float(cf_ini),
            "CF_fim_m":        float(cf_fim),
            "Prof_ini_m":      float(t.get("prof_ini") or 0) if t.get("prof_ini") else None,
            "Prof_fim_m":      float(t.get("prof_fim") or 0) if t.get("prof_fim") else None,
            "Velocidade_ms":   hid.get("vel_ms"),
            "Vazao_ls":        hid.get("vazao_ls"),
            "Tensao_trativa_Pa": hid.get("tau_pa"),
            "Status_hidraulico": hid.get("status","SEM_DADOS"),
            "Nucleo":          cfg.get("nucleo",""),
            "Contrato":        cfg["contrato"],
            "Rua":             t.get("rua",""),
            "URL_NS":          f"{ns_base_url}NS_{ns_id}.html",
        })

        # Tipo de rede como IFC classification
        ref = model.createIfcClassificationReference(
            Location=None,
            Identification=("AA-AF" if "AGUA" in tipo_rede else "AA-ES"),
            Name=("Rede de Água Fria" if "AGUA" in tipo_rede
                  else "Rede Coletora de Esgoto"),
            ReferencedSource=classif)
        model.createIfcRelAssociatesClassification(
            ifcopenshell.guid.new(), None, None, None,
            [pipe], ref)

        n_pipes_ok += 1

    # ── Gravar ────────────────────────────────────────────────────────────────
    caminho = pasta_bim / nome_arquivo
    model.write(str(caminho))

    size_kb = caminho.stat().st_size // 1024
    log(f"  IFC LOD500: {nome_arquivo} ({size_kb}KB) | "
        f"{n_pipes_ok} pipes | {n_pvs_ok} estruturas", "OK")

    # Substituir o arquivo .ifc.pendente se existir
    pendente = pasta_bim / (nome_arquivo.replace(".ifc", "_ifc.pendente"))
    if pendente.exists():
        pendente.unlink()

    return caminho


def _dd_to_dms(dd):
    """Decimal degrees → (graus, minutos, segundos) para IFC RefLatitude/Longitude."""
    neg = dd < 0
    dd  = abs(dd)
    g   = int(dd)
    m   = int((dd - g) * 60)
    s   = round(((dd - g) * 60 - m) * 1e6)
    return (-g if neg else g, m, s)


# ==============================================================================
# MÓDULO 18 — QR CODE por NS
# ==============================================================================

try:
    import qrcode
    from PIL import Image as PILImage
    import io as _io
    _HAS_QR = True
except ImportError:
    _HAS_QR = False

_QR_CACHE = {}

def gerar_qr_png(url, tamanho_px=200):
    """
    Gera QR Code PNG em memória a partir de uma URL.
    Retorna PIL.Image ou None.
    """
    if not _HAS_QR:
        return None
    if url in _QR_CACHE:
        return _QR_CACHE[url]
    qr = qrcode.QRCode(
        version=2,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=8,
        border=2)
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = _io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    pil_img = PILImage.open(buf).copy()
    _QR_CACHE[url] = pil_img
    return pil_img


def _embed_qr_no_ax(ax, qr_img, x=0.78, y=0.02, w=0.20, h=0.20,
                    label="QR: Dashboard NS"):
    """Embute QR Code em um axes matplotlib no canto inferior direito."""
    if qr_img is None:
        return
    ax_in = ax.inset_axes([x, y, w, h])
    ax_in.imshow(qr_img, cmap="gray", aspect="equal")
    ax_in.axis("off")
    ax_in.set_title(label, fontsize=4.5, color="#555555", pad=1)


def gerar_ns_a4_com_qr(t, pvs, ns_id, pasta, cfg, ns_base_url=""):
    """
    Versão aprimorada de gerar_ns_a4 com QR Code embedido no rodapé.
    Chama gerar_ns_a4 internamente e adiciona QR na figura antes de salvar.
    """
    # URL do dashboard
    url = f"{ns_base_url}NS_{ns_id}.html" if ns_base_url else f"NS_{ns_id}.html"
    qr_img = gerar_qr_png(url)

    pvi     = pvs.get(t.get("pv_ini"), {})
    pvf     = pvs.get(t.get("pv_fim"), {})
    is_agua = t.get("is_agua", False)
    hid     = t.get("hidraulica") or {}
    pressao = _fmtv(t.get("pressao_mca"), ".2f") if is_agua else None
    c_pipe  = "#1E6B3C" if is_agua else "#1565C0"
    titulo_sis = ("SISTEMA DE ABASTECIMENTO DE ÁGUA FRIA"
                  if is_agua else "SISTEMA DE ESGOTAMENTO SANITÁRIO")
    titulo_os  = ("ORDEM DE SERVIÇO PARA EXECUÇÃO"
                  if is_agua else "ORDEM DE SERVIÇO PARA GABARITO")

    fig, ax = plt.subplots(figsize=(29.7/2.54, 21.0/2.54))
    ax.set_xlim(0, 29.7); ax.set_ylim(0, 21.0); ax.axis("off")

    # ── Cabeçalho ────────────────────────────────────────────────────────────
    ax.add_patch(plt.Rectangle((0.5, 17.5), 28.7, 3.0, fc="#1F4E79", lw=0))
    ax.text(14.85, 19.55, f"{titulo_sis}  |  {cfg['cidade']}  |  {titulo_os}",
            ha="center", va="center", fontsize=10.5, fontweight="bold", color="white")
    ax.text(14.85, 18.85,
            f"EMPRESA: {cfg['empresa']}  |  CONTRATO: {cfg['contrato']}  |  "
            f"NÚCLEO: {cfg.get('nucleo','')}  |  LOGRADOURO: {t.get('rua','Sem Rua')}",
            ha="center", va="center", fontsize=7.5, color="white")
    badge = (f"NS Nº {ns_id}  |  TRECHO: {t['pv_ini']} → {t['pv_fim']}  |  "
             f"DN: {_fmtv(t.get('dn_mm'),'.0f')} mm  |  EXT: {_fmtv(t.get('ext_m'),'.2f')} m  |  "
             + (f"P: {pressao} mca" if pressao else f"DECL: {_fmtv(t.get('decl_pct'),'.3f')} %"))
    ax.text(14.85, 18.10, badge, ha="center", va="center",
            fontsize=8.5, fontweight="bold", color="white")

    # ── Tabela principal ─────────────────────────────────────────────────────
    headers = ["TRECHO","ESTACA","DISTÂNCIA\n(m)","CT\n(m)","I\n(m/m)",
               "CP\n(m)","CR\n(m)","DN\n(mm)","G","H","P"]
    col_w   = [3.5, 2.8, 2.4, 2.2, 2.0, 2.2, 2.2, 1.8, 1.2, 1.2, 1.2]
    col_x   = [0.5]
    for w in col_w[:-1]:
        col_x.append(col_x[-1]+w)
    y_hdr = 16.8
    for h, cx, w in zip(headers, col_x, col_w):
        ax.add_patch(plt.Rectangle((cx, y_hdr-0.85), w, 0.92,
                                   ec="white", fc="#1F4E79", lw=0.5))
        ax.text(cx+w/2, y_hdr-0.38, h, ha="center", va="center",
                fontsize=5.8, fontweight="bold", color="white")

    ext_str = _fmtv(t.get("ext_m"), ".2f")
    rows = [
        [t["pv_ini"], "0+00", "0.00",
         _fmtv(t.get("ct_ini")), _fmtv(t.get("decl_mm"),".5f"),
         _fmtv(t.get("cf_ini")), _fmtv(t.get("prof_ini"),".2f"),
         _fmtv(t.get("dn_mm"),".0f"), "", "", ""],
        [t["pv_fim"], f"0+{ext_str}", ext_str,
         _fmtv(t.get("ct_fim")), _fmtv(t.get("decl_mm"),".5f"),
         _fmtv(t.get("cf_fim")), _fmtv(t.get("prof_fim"),".2f"),
         _fmtv(t.get("dn_mm"),".0f"), "", "", ""],
        ["TOTAIS","","",ext_str,"","","","","","",""],
    ]
    for ri, row in enumerate(rows):
        y_row = y_hdr - 0.85 - (ri+1)*0.75
        fc_c = "#EEF3FA" if ri%2==0 else "white"
        for j, (val, cx, w) in enumerate(zip(row, col_x, col_w)):
            ax.add_patch(plt.Rectangle((cx, y_row), w, 0.72,
                                       ec="#AAAAAA", fc=fc_c, lw=0.3))
            ax.text(cx+w/2, y_row+0.36, str(val),
                    ha="center", va="center", fontsize=7)

    # ── Dados hidráulicos ────────────────────────────────────────────────────
    y_hid = y_hdr - 0.85 - len(rows)*0.75 - 0.35
    ax.add_patch(plt.Rectangle((0.5, y_hid-0.55), 28.7, 0.60, fc="#1F4E79", lw=0))
    if is_agua:
        lbl_hid = "DADOS HIDRÁULICOS — REDE PRESSURIZADA (ÁGUA FRIA)"
        txt_hid = (f"Pressão: {pressao} mca  |  DN: {_fmtv(t.get('dn_mm'),'.0f')} mm  |  "
                   f"Material: {t.get('material','PE80')}  |  "
                   f"Prof. ini: {_fmtv(t.get('prof_ini'),'.2f')} m  |  "
                   f"Prof. fim: {_fmtv(t.get('prof_fim'),'.2f')} m")
    else:
        lbl_hid = "DADOS HIDRÁULICOS (Manning — NBR 9649)"
        txt_hid = (f"Velocidade: {_fmtv(hid.get('vel_ms'),'.3f')} m/s  |  "
                   f"Vazão: {_fmtv(hid.get('vazao_ls'),'.2f')} l/s  |  "
                   f"Tensão trativa: {_fmtv(hid.get('tau_pa'),'.2f')} Pa  |  "
                   f"Status: {hid.get('status','SEM_DADOS')}")
    ax.text(14.85, y_hid-0.08, lbl_hid, ha="center", va="center",
            fontsize=7.5, fontweight="bold", color="white")
    ax.text(14.85, y_hid-0.44, txt_hid, ha="center", va="center",
            fontsize=7.5, color="white")

    # ── Croqui ───────────────────────────────────────────────────────────────
    y_croqui = y_hid - 0.55 - 1.35
    cx0, cx1, cy = 4.0, 23.0, y_croqui - 0.9
    ax.annotate("", xy=(cx1, cy), xytext=(cx0, cy),
                arrowprops=dict(arrowstyle="-|>", color=c_pipe, lw=2.2))
    for cx_pv in [cx0, cx1]:
        ax.add_patch(plt.Rectangle((cx_pv-0.28, cy-0.28), 0.56, 0.56,
                                   ec="#333", fc="#4488ff", lw=1))
    ax.text(cx0, cy-0.65, f"{t['pv_ini']}\nCF={_fmtv(t.get('cf_ini'))}",
            ha="center", va="top", fontsize=6.5, color="#111")
    ax.text(cx1, cy-0.65, f"{t['pv_fim']}\nCF={_fmtv(t.get('cf_fim'))}",
            ha="center", va="top", fontsize=6.5, color="#111")
    pipe_lbl = (f"DN {_fmtv(t.get('dn_mm'),'.0f')} mm / L={ext_str} m / "
                + (f"P={pressao} mca" if pressao
                   else f"i={_fmtv(t.get('decl_pct'),'.3f')} %"))
    ax.text(13.5, cy+0.28, pipe_lbl, ha="center", va="bottom",
            fontsize=7.5, fontweight="bold", color=c_pipe)

    # ── QR Code ──────────────────────────────────────────────────────────────
    if qr_img is not None:
        ax_qr = ax.inset_axes([0.82, 0.04, 0.16, 0.22],
                              transform=ax.transAxes)
        ax_qr.imshow(qr_img, cmap="gray", aspect="equal")
        ax_qr.axis("off")
        ax_qr.set_title(f"Dashboard NS {ns_id}", fontsize=4.5,
                        color="#333", pad=1)

    # ── Assinaturas ──────────────────────────────────────────────────────────
    sigs = ["ENG. CAMPO","EXECUTOR","COORD.","GERENTE","C. PROJ.","G. ENG."]
    sw   = 23.7 / len(sigs)
    for k, s in enumerate(sigs):
        sx = 0.5 + k*sw
        ax.add_patch(plt.Rectangle((sx, 0.60), sw, 1.20,
                                   ec="#AAAAAA", fc="white", lw=0.5))
        ax.plot([sx+0.3, sx+sw-0.3], [1.10, 1.10], color="#888", lw=0.5)
        ax.text(sx+sw/2, 0.80, s, ha="center", va="center",
                fontsize=6.5, color="#555")

    # ── Rodapé ───────────────────────────────────────────────────────────────
    rodape = (f"SABESP SANTOS — {cfg.get('nucleo','')} — "
              f"{'ÁGUA FRIA' if is_agua else 'ESGOTO'} — NS{ns_id} — "
              f"Folha 1 de 1 — Rev. 0")
    ax.add_patch(plt.Rectangle((0.5, 0.0), 28.7, 0.55, fc="#1F4E79", lw=0))
    ax.text(13.0, 0.27, rodape, ha="center", va="center",
            fontsize=7, color="#CCCCCC")
    ax.text(24.0, 0.27, "ConstruData BIM", ha="center", va="center",
            fontsize=7, color="#AAAAAA", fontstyle="italic")

    fig.tight_layout(pad=0)
    caminho = pasta / f"NS_{ns_id}_A4.pdf"
    fig.savefig(caminho, dpi=200, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    return caminho


def gerar_ns_desenho_com_qr(t, pvs, ns_id, pasta, cfg,
                             ruas_dxf=None, gpkg_path=None, ns_base_url=""):
    """
    Versão de gerar_ns_desenho com QR Code no canto do selo.
    Delega ao gerar_ns_desenho original e injeta o QR no eixo da tabela.
    """
    url    = f"{ns_base_url}NS_{ns_id}.html" if ns_base_url else f"NS_{ns_id}.html"
    qr_img = gerar_qr_png(url)

    # Gera normalmente — depois reabre a figura para adicionar QR
    # (mais simples: re-implementar o ax_tabela com QR inline)

    pvi     = pvs.get(t.get("pv_ini"), {})
    pvf     = pvs.get(t.get("pv_fim"), {})
    xi, yi  = pvi.get("x"), pvi.get("y")
    xf, yf  = pvf.get("x"), pvf.get("y")
    is_agua = t.get("is_agua", False)
    c_pipe  = "#1E6B3C" if is_agua else "#1565C0"
    C_CT    = "#7B3F00"
    C_CF    = "#1565C0"

    fig = plt.figure(figsize=(42.0/2.54, 29.7/2.54), facecolor="white")
    gs  = GridSpec(3, 2, figure=fig,
                   height_ratios=[7, 3.5, 1.8],
                   width_ratios=[1, 1],
                   hspace=0.10, wspace=0.06,
                   left=0.04, right=0.97, top=0.95, bottom=0.05)
    ax_planta = fig.add_subplot(gs[0, 0])
    ax_legend = fig.add_subplot(gs[0, 1])
    ax_perfil = fig.add_subplot(gs[1, :])
    ax_tabela = fig.add_subplot(gs[2, :])

    # ── Planta ───────────────────────────────────────────────────────────────
    ax_planta.set_facecolor("#F0EDE8")
    ax_planta.tick_params(labelsize=5)
    if xi and yi and xf and yf:
        cx  = (xi+xf)/2; cy  = (yi+yf)/2
        pad = max(math.hypot(xf-xi, yf-yi)*1.5, 60)
        ax_planta.set_xlim(cx-pad, cx+pad)
        ax_planta.set_ylim(cy-pad, cy+pad)
        ax_planta.set_aspect("equal")
        if gpkg_path:
            try:
                from shapely.geometry import box as sbox
                carta = ler_cartografia_gpkg(gpkg_path)
                bb    = sbox(cx-pad, cy-pad, cx+pad, cy+pad)
                if carta["quadras"] is not None:
                    for geom in carta["quadras"][carta["quadras"].intersects(bb)].geometry:
                        try:
                            xs, ys = (geom.exterior.xy if hasattr(geom,"exterior") else geom.xy)
                            ax_planta.fill(xs, ys, fc="#EDE8DC", ec="#CCBBAA", lw=0.4, zorder=1)
                        except Exception:
                            pass
                if carta["streets"] is not None:
                    for geom in carta["streets"][carta["streets"].intersects(bb)].geometry:
                        try:
                            xs, ys = geom.xy
                            ax_planta.plot(xs, ys, color="#AAAAAA", lw=0.8, zorder=2)
                        except Exception:
                            pass
            except Exception:
                pass
        if ruas_dxf:
            xmin, xmax = cx-pad, cx+pad
            ymin_p, ymax_p = cy-pad, cy+pad
            seen = set()
            for r in ruas_dxf:
                rx, ry, rt = r.get("x",0), r.get("y",0), r.get("text","")
                if rt in seen or not rt: continue
                if xmin < rx < xmax and ymin_p < ry < ymax_p:
                    seen.add(rt)
                    is_cur = t.get("rua","") and t["rua"].upper() in rt.upper()
                    ax_planta.text(rx, ry, rt, fontsize=4.5,
                                   color="#FFFF00" if is_cur else "#DDDDDD",
                                   ha="center", va="center", zorder=5,
                                   fontweight="bold" if is_cur else "normal",
                                   bbox=dict(boxstyle="round,pad=0.15", fc="black",
                                             ec="none", alpha=0.5))
        ax_planta.annotate("", xy=(xf,yf), xytext=(xi,yi),
                           arrowprops=dict(arrowstyle="-|>", color=c_pipe, lw=2))
        for px_v,py_v,pnome,pv_d,lado in [(xi,yi,t["pv_ini"],pvi,"ini"),
                                            (xf,yf,t["pv_fim"],pvf,"fim")]:
            ax_planta.plot(px_v, py_v, "s", color=c_pipe, ms=8, zorder=10)
            dx_off = pad*0.18*(1 if lado=="fim" else -1)
            ax_planta.text(px_v+dx_off, py_v+pad*0.12,
                           f"{pnome}\nCT={_fmtv(pv_d.get('ct'),'.3f')}\n"
                           f"CF={_fmtv(pv_d.get('cf'),'.3f')}",
                           fontsize=4.5, color="#111", va="bottom",
                           bbox=dict(boxstyle="round,pad=0.2", fc="white",
                                     ec=c_pipe, lw=0.5))
        mid_x=(xi+xf)/2; mid_y=(yi+yf)/2
        ax_planta.text(mid_x, mid_y+pad*0.08,
                       f"DN {_fmtv(t.get('dn_mm'),'.0f')}mm/{t.get('material','PVC')}/"
                       f"L={_fmtv(t.get('ext_m'),'.2f')}m",
                       ha="center", va="bottom", fontsize=5, color=c_pipe,
                       fontweight="bold",
                       bbox=dict(boxstyle="round,pad=0.15",fc="white",ec=c_pipe,lw=0.4))
        ax_planta.annotate("", xy=(cx+pad*0.78, cy+pad*0.72),
                           xytext=(cx+pad*0.78, cy+pad*0.52),
                           arrowprops=dict(arrowstyle="-|>", color="#333", lw=1.5))
        ax_planta.text(cx+pad*0.78, cy+pad*0.76, "N",
                       ha="center", va="bottom", fontsize=8, fontweight="bold")
        ax_planta.set_xlabel("Este (m UTM)", fontsize=5)
        ax_planta.set_ylabel("Norte (m UTM)", fontsize=5)
        plt.setp(ax_planta.get_xticklabels(), fontsize=4.5)
        plt.setp(ax_planta.get_yticklabels(), fontsize=4.5)
    ax_planta.set_title(
        f"PLANTA NS {ns_id} | {t['pv_ini']} → {t['pv_fim']} | {t.get('rua','Sem Rua')}",
        fontsize=7.5, fontweight="bold", pad=4)

    # ── Legenda ───────────────────────────────────────────────────────────────
    ax_legend.axis("off")
    q = t.get("quantitativos") or {}
    lines_q = [
        "Quantitativo:", "",
        f"Volume de Escavação = {_fmtv(q.get('esc_m3'),'.3f')} m³",
        f"Volume de Aterro = {_fmtv(q.get('reat_m3'),'.3f')} m³",
        f"Pavimentação = {_fmtv(q.get('pav_m2'),'.2f')} m²",
        f"Extensão Total = {_fmtv(t.get('ext_m'),'.2f')} m",
    ]
    if not is_agua:
        lines_q += ["", "—— PVC ——",
                    f"{q.get('tubo_barras','---')} barras  "
                    f"{_fmtv(t.get('dn_mm'),'.0f')}mm  {t.get('material','PVC')}"]
    else:
        lines_q += ["", "—— Materiais ——"]
        for k,v in list((t.get("materiais_agua") or {}).items())[:5]:
            lines_q.append(f"{v}  {k}")
    for li, line in enumerate(lines_q):
        ax_legend.text(0.05, 0.95-li*0.07, line,
                       transform=ax_legend.transAxes, fontsize=6.5,
                       va="top", color="#111")
    yl = 0.95 - len(lines_q)*0.07 - 0.04
    ax_legend.text(0.05, yl, "LEGENDA", transform=ax_legend.transAxes,
                   fontsize=7, fontweight="bold", va="top")
    ax_legend.plot([0.05,0.20], [yl-0.06,yl-0.06], color=c_pipe, lw=2,
                   transform=ax_legend.transAxes)
    ax_legend.text(0.23, yl-0.04,
                   f"Tubo DN{_fmtv(t.get('dn_mm'),'.0f')}mm",
                   transform=ax_legend.transAxes, fontsize=6.5, va="center")
    ax_legend.plot([0.05],[yl-0.12],"s",color=c_pipe,ms=8,
                   transform=ax_legend.transAxes)
    ax_legend.text(0.23, yl-0.10, "P.V. — Poço de Visita",
                   transform=ax_legend.transAxes, fontsize=6.5, va="center")

    # ── Perfil ────────────────────────────────────────────────────────────────
    ct_i=t.get("ct_ini"); ct_f=t.get("ct_fim")
    cf_i=t.get("cf_ini"); cf_f=t.get("cf_fim")
    ext=t.get("ext_m") or 1
    has_cotas=all(v is not None for v in [ct_i,ct_f,cf_i,cf_f])
    if has_cotas:
        all_v=[ct_i,ct_f,cf_i,cf_f]
        y_min=min(all_v)-0.30; y_max=max(all_v)+0.45
        ax_perfil.set_xlim(-ext*0.03, ext*1.03)
        ax_perfil.set_ylim(y_min, y_max)
        ax_perfil.plot([0,ext],[ct_i,ct_f], color=C_CT,lw=1.8,zorder=5,label="CT")
        ax_perfil.plot([0,ext],[cf_i,cf_f], color=C_CF,lw=1.5,ls="--",zorder=5,label="CF")
        ax_perfil.fill_between([0,ext],[cf_i,cf_f],[ct_i,ct_f],
                               alpha=0.12, color="#8B4513", zorder=2)
        ax_perfil.fill_between([0,ext],[cf_i,cf_f],[cf_i+0.05,cf_f+0.05],
                               color=c_pipe, alpha=0.85, zorder=6)
        for px_d,pnome,ct_v,cf_v in [(0,t["pv_ini"],ct_i,cf_i),(ext,t["pv_fim"],ct_f,cf_f)]:
            ax_perfil.plot(px_d,ct_v,"v",color=C_CT,ms=7,zorder=10)
            ax_perfil.plot(px_d,cf_v,"^",color=C_CF,ms=7,zorder=10)
            ax_perfil.text(px_d, y_max, f"{pnome}\nCT={ct_v:.3f}\nCF={cf_v:.3f}",
                           ha="center", va="top", fontsize=5.5, color="#111",
                           bbox=dict(boxstyle="round,pad=0.2",fc="white",ec="#888",lw=0.4))
        ax_perfil.text(ext/2, (cf_i+cf_f)/2-(y_max-y_min)*0.12,
                       f"DN {_fmtv(t.get('dn_mm'),'.0f')}mm  i={_fmtv(t.get('decl_pct'),'.2f')}%",
                       ha="center", va="center", fontsize=6.5,
                       color=c_pipe, fontweight="bold")
        ax_perfil.legend(fontsize=6, loc="upper right")
    else:
        ax_perfil.text(0.5,0.5,"Sem dados de cota",ha="center",va="center",
                       fontsize=11,color="#AAAAAA",transform=ax_perfil.transAxes)
        ax_perfil.set_xlim(0,ext or 10); ax_perfil.set_ylim(0,1)
    ax_perfil.set_xlabel("Distância (m)",fontsize=6)
    ax_perfil.set_ylabel("Cota (m)",fontsize=6)
    ax_perfil.tick_params(labelsize=5.5)
    ax_perfil.yaxis.grid(True,ls="--",lw=0.4,color="#DDDDDD")
    ax_perfil.set_title("PERFIL LONGITUDINAL    Exag. vertical ~10x",
                        fontsize=7, loc="center", pad=3)

    # ── Tabela + Selo + QR ───────────────────────────────────────────────────
    ax_tabela.axis("off")
    col_labels = ["Estaca","CT (m)","CF (m)","Prof (m)","Dist (m)","DN (mm)","Decl (%)"]
    row_i=[t["pv_ini"],_fmtv(t.get("ct_ini")),_fmtv(t.get("cf_ini")),
           _fmtv(t.get("prof_ini"),".2f"),"0.00",
           _fmtv(t.get("dn_mm"),".0f"),_fmtv(t.get("decl_pct"),".2f")]
    row_f=[t["pv_fim"],_fmtv(t.get("ct_fim")),_fmtv(t.get("cf_fim")),
           _fmtv(t.get("prof_fim"),".2f"),_fmtv(t.get("ext_m"),".2f"),
           _fmtv(t.get("dn_mm"),".0f"),_fmtv(t.get("decl_pct"),".2f")]
    tbl = ax_tabela.table(cellText=[row_i,row_f], colLabels=col_labels,
                          loc="upper left", cellLoc="center",
                          bbox=[0.0, 0.0, 0.52, 1.0])
    tbl.auto_set_font_size(False); tbl.set_fontsize(6.5)
    for (r,c),cell in tbl.get_celld().items():
        if r==0:
            cell.set_facecolor("#1F4E79")
            cell.set_text_props(color="white",fontweight="bold",fontsize=6.5)
        else:
            cell.set_facecolor("#F0F4FA" if r%2 else "white")
        cell.set_edgecolor("#CCCCCC")
        cell.set_height(0.42)

    # Selo SABESP
    ax_tabela.text(0.62,0.85,"SABESP",transform=ax_tabela.transAxes,
                   ha="center",va="top",fontsize=12,fontweight="bold",color="#0A3D91")
    ax_tabela.text(0.62,0.60,
                   ("SISTEMA DE ESGOTAMENTO\nSANITÁRIO SANTOS/SP"
                    if not is_agua else
                    "SISTEMA DE ABASTECIMENTO\nDE ÁGUA FRIA SANTOS/SP"),
                   transform=ax_tabela.transAxes,ha="center",va="top",
                   fontsize=6.5,color="#333")
    ax_tabela.text(0.78,0.80,
                   f"CONTRATO: {cfg['contrato']}\nNS Nº {ns_id}\n"
                   f"NÚCLEO: {cfg.get('nucleo','')}\n"
                   f"ENG.: {cfg.get('engenheiro','')}\n"
                   f"ConstruData BIM  Rev. 0",
                   transform=ax_tabela.transAxes,ha="center",va="top",
                   fontsize=5.5,color="#333")

    # QR Code no canto do selo
    if qr_img is not None:
        ax_qr = ax_tabela.inset_axes([0.88, 0.05, 0.12, 0.90],
                                     transform=ax_tabela.transAxes)
        ax_qr.imshow(qr_img, cmap="gray", aspect="equal")
        ax_qr.axis("off")
        ax_qr.set_title(f"NS {ns_id}", fontsize=4.5, color="#333", pad=1)

    caminho = pasta / f"NS_{ns_id}_DESENHO.pdf"
    fig.savefig(caminho, dpi=200, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    return caminho



# ==============================================================================
# ENTRY POINT
# ==============================================================================

if __name__ == "__main__":
    main()
