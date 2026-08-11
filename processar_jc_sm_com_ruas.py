#!/usr/bin/env python3
"""
PROCESSAR JOAO CARLOS + SAO MANOEL — COM RUAS CORRETAS + PRINT SATÉLITE
========================================================================
Usa a estratégia CALIBRADA:
  1. LandXML → network names = ruas
  2. MAPA DXF QGIS → complemento de ruas por proximidade
  3. Imagem satélite por trecho embutida na planilha Excel
"""
import sys, os, json, math, traceback, re, warnings
from pathlib import Path
from datetime import datetime
from collections import defaultdict, Counter
import unicodedata

warnings.filterwarnings("ignore")

MOTOR_DIR = Path(r"C:\Users\felip\Downloads\NOVA NS Versao 5")
sys.path.insert(0, str(MOTOR_DIR))

from ler_landxml import ler_landxml
from gerar_ns import (
    gerar_ns_a4, gerar_ns_desenho, gerar_ns_sat, gerar_html,
    gerar_geojson, enriquecer_trechos, calcular_materiais,
    calc_manning, CONTRATO, NS_VERSION, _ns_folder_name
)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage

import numpy as np
import geopandas as gpd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

try:
    import contextily as cx
    HAS_CTX = True
except ImportError:
    HAS_CTX = False

try:
    from exportar_completo import _gerar_planilha_trechos_completa
    HAS_EXPORT = True
except ImportError:
    HAS_EXPORT = False

INPUT_DIR = Path(r"C:\Users\felip\Desktop\GERAR NS COM ITENS POR RUA DE PV A PV")
OUTPUT_BASE = INPUT_DIR / "SAIDA_POR_RUA_COM_NOMES"


def slug(texto):
    if not texto:
        return "DESCONHECIDO"
    txt = str(texto).replace("\n", " ").replace("\r", " ").replace("\t", " ")
    txt = re.sub(r'[\x00-\x1f\x7f]', '', txt)
    normalizado = unicodedata.normalize("NFKD", txt)
    ascii_txt = normalizado.encode("ascii", "ignore").decode("ascii")
    safe = re.sub(r'[^\w\s-]', '_', ascii_txt).strip()
    safe = re.sub(r'\s+', ' ', safe)
    return safe[:80] or "DESCONHECIDO"


def limpar_rua(rua):
    if not rua or str(rua).strip() in ("", "nan", "None", "Sem Rua", "1"):
        return "Sem Rua"
    txt = str(rua).replace("\n", " ").replace("\r", " ")
    txt = re.sub(r'[\x00-\x1f\x7f]', '', txt)
    txt = re.sub(r'\s+', ' ', txt).strip()
    return txt if txt else "Sem Rua"


def log(msg, nivel="INFO"):
    ts = datetime.now().strftime("%H:%M:%S")
    pfx = {"OK": "[OK]  ", "WARN": "[!]   ", "STEP": ">>> ", "ERR": "[X]  "}.get(nivel, "      ")
    try:
        print(f"[{ts}] {pfx}{msg}")
    except UnicodeEncodeError:
        print(f"[{ts}] {pfx}" + msg.encode("ascii", "replace").decode())


# ══════════════════════════════════════════════════════════════════════
# CARREGAR RUAS DO MAPA DXF (igual ao script calibrado que funcionou!)
# ══════════════════════════════════════════════════════════════════════

def carregar_ruas_mapa_dxf(mapa_dxf_path, trechos, pvs, tol=80.0):
    """
    Carrega textos de rua do MAPA DXF do QGIS e associa aos trechos.
    Este é o método que FUNCIONOU na versão calibrada do São Manoel.
    """
    if not Path(mapa_dxf_path).exists():
        log(f"  Mapa DXF não encontrado: {mapa_dxf_path}", "WARN")
        return 0

    try:
        gdf_mapa = gpd.read_file(str(mapa_dxf_path))
    except Exception as e:
        log(f"  Erro lendo mapa DXF: {e}", "WARN")
        return 0

    # Filtrar textos que são nomes de rua
    PREF_RUA = ('RUA', 'BECO', 'TRAV', 'AV', 'ESTRADA', 'VIELA', 'ALAMEDA',
                'ACESSO', 'CAMINHO', 'LARGO', 'PRAÇA', 'PRACA')

    if 'Text' not in gdf_mapa.columns:
        log(f"  Mapa DXF sem coluna 'Text'", "WARN")
        return 0

    mask = gdf_mapa['Text'].notna()
    gdf_ruas = gdf_mapa[mask & gdf_mapa['Text'].str.upper().str.strip().apply(
        lambda t: any(str(t).startswith(p) for p in PREF_RUA) if t else False
    )]

    if len(gdf_ruas) == 0:
        # Fallback: qualquer texto > 3 chars
        gdf_ruas = gdf_mapa[mask & (gdf_mapa['Text'].str.len() > 3)]

    if len(gdf_ruas) == 0:
        log(f"  Mapa DXF sem textos de rua", "WARN")
        return 0

    # Extrair coordenadas e nomes
    rua_data = []
    for _, row in gdf_ruas.iterrows():
        try:
            g = row.geometry
            if g is not None and g.geom_type == 'Point' and abs(g.x) > 1000:
                rua_data.append((g.x, g.y, str(row['Text']).strip()))
        except Exception:
            pass

    if not rua_data:
        log(f"  Mapa DXF sem pontos válidos de rua", "WARN")
        return 0

    rua_xy = np.array([[r[0], r[1]] for r in rua_data])
    rua_txts = [r[2] for r in rua_data]

    n_rua = 0
    for t in trechos:
        # Pular se já tem rua válida
        rua_atual = str(t.get("rua") or "").strip()
        if rua_atual and rua_atual not in ("Sem Rua", "", "nan", "None", "1"):
            continue

        p0 = pvs.get(t.get("pv_ini"), {})
        p1 = pvs.get(t.get("pv_fim"), {})
        if not p0.get("x") or not p1.get("x"):
            continue
        if p0["x"] == 0 or p1["x"] == 0:
            continue

        mx = (p0["x"] + p1["x"]) / 2
        my = (p0["y"] + p1["y"]) / 2
        dists = np.hypot(rua_xy[:, 0] - mx, rua_xy[:, 1] - my)
        idx = int(np.argmin(dists))
        if dists[idx] <= tol:
            t["rua"] = rua_txts[idx]
            n_rua += 1

    return n_rua


# ══════════════════════════════════════════════════════════════════════
# USAR NETWORK NAMES DO XML COMO RUAS
# ══════════════════════════════════════════════════════════════════════

def associar_ruas_network_xml(xml_path, trechos):
    """
    Relê o XML diretamente para extrair os nomes das PipeNetworks.
    No Civil 3D, cada rede tem um nome que geralmente corresponde à rua.
    Ex: 'REDE RUA LIBERDADE', 'RCE BECO 16', etc.
    """
    import xml.etree.ElementTree as ET

    tree = ET.parse(str(xml_path))
    root = tree.getroot()
    ns = {'lx': 'http://www.landxml.org/schema/LandXML-1.2'}

    # Mapear pipe name -> network name
    pipe_to_network = {}
    for pn in root.findall('.//lx:PipeNetwork', ns):
        net_name = pn.get('name', '')
        for pipe in pn.findall('.//lx:Pipe', ns):
            pipe_name = pipe.get('name', '')
            if pipe_name:
                pipe_to_network[pipe_name] = net_name

    if not pipe_to_network:
        # Tentar sem namespace
        for pn in root.iter():
            if 'PipeNetwork' in pn.tag:
                net_name = pn.get('name', '')
                for pipe in pn.iter():
                    if 'Pipe' in pipe.tag and pipe.get('name'):
                        pipe_to_network[pipe.get('name')] = net_name

    n_assoc = 0
    for t in trechos:
        nome_trecho = t.get("nome_trecho", "")
        if nome_trecho and nome_trecho in pipe_to_network:
            net_name = pipe_to_network[nome_trecho]
            if net_name and net_name != "1":
                t["rua"] = limpar_rua(net_name)
                n_assoc += 1

    return n_assoc


# ══════════════════════════════════════════════════════════════════════
# GERAR IMAGEM SATÉLITE DO TRECHO (PNG)
# ══════════════════════════════════════════════════════════════════════

def gerar_img_trecho(pvs, trecho, out_path):
    p0 = pvs.get(trecho["pv_ini"], {})
    p1 = pvs.get(trecho["pv_fim"], {})
    if not p0.get("x") or not p1.get("x") or p0["x"] == 0 or p1["x"] == 0:
        return False

    fig, ax = plt.subplots(figsize=(6, 4), dpi=100)
    x0, y0 = p0["x"], p0["y"]
    x1, y1 = p1["x"], p1["y"]
    margin = max(abs(x1-x0), abs(y1-y0), 30) * 0.6

    ax.set_xlim(min(x0,x1)-margin, max(x0,x1)+margin)
    ax.set_ylim(min(y0,y1)-margin, max(y0,y1)+margin)

    sat_ok = False
    if HAS_CTX:
        for zoom in [18, 17, 16]:
            try:
                cx.add_basemap(ax, crs="EPSG:31983",
                               source=cx.providers.Esri.WorldImagery, zoom=zoom)
                sat_ok = True
                break
            except Exception:
                continue

    if not sat_ok:
        ax.set_facecolor("#e8f0e8")
        ax.grid(True, color="#ccc", alpha=0.5)

    ax.plot([x0,x1],[y0,y1],'r-',linewidth=3,zorder=5)

    for nome, pv, cor in [(trecho["pv_ini"],p0,"red"),(trecho["pv_fim"],p1,"blue")]:
        ax.plot(pv["x"],pv["y"],'o',color=cor,markersize=8,zorder=6)
        ct = pv.get("ct","")
        ct_txt = f"\nCT={ct:.3f}" if isinstance(ct,(int,float)) and ct > 0 else ""
        ax.annotate(f"{nome}{ct_txt}",(pv["x"],pv["y"]),
                    fontsize=7,color="white" if sat_ok else "black",
                    fontweight="bold",ha="center",va="bottom",
                    xytext=(0,8),textcoords="offset points",
                    bbox=dict(boxstyle="round,pad=0.2",fc="black" if sat_ok else "white",alpha=0.7),
                    zorder=7)

    dn = trecho.get("dn_mm","?")
    ext = trecho.get("ext_m",0)
    rua = trecho.get("rua","")
    ax.set_title(f"DN{dn}mm | {ext:.1f}m | {rua}", fontsize=8, fontweight="bold")
    ax.tick_params(labelsize=5)
    fig.tight_layout()
    fig.savefig(str(out_path), dpi=100, bbox_inches="tight", pad_inches=0.1)
    plt.close(fig)
    return True


# ══════════════════════════════════════════════════════════════════════
# PLANILHA DE TRECHOS POR RUA COM IMAGEM SATÉLITE
# ══════════════════════════════════════════════════════════════════════

def gerar_planilha_rua(trechos, pvs, nucleo, rua, out_path, img_dir):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trechos"

    hf = Font(name="Calibri",bold=True,color="FFFFFF",size=10)
    hfill = PatternFill(start_color="1a237e",end_color="1a237e",fill_type="solid")
    df = Font(name="Calibri",size=9)
    da = Alignment(horizontal="center",vertical="center")
    brd = Border(left=Side("thin","999999"),right=Side("thin","999999"),
                 top=Side("thin","999999"),bottom=Side("thin","999999"))
    alt = PatternFill(start_color="f0f4ff",end_color="f0f4ff",fill_type="solid")

    ws.merge_cells("A1:O1")
    ws["A1"] = f"TRECHOS — {rua} — {nucleo} — ESGOTO"
    ws["A1"].font = Font(name="Calibri",bold=True,size=14,color="1a237e")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:O2")
    ws["A2"] = f"CT {CONTRATO} | SE LIGA NA REDE | {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(name="Calibri",size=9,color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = ["NS","PV Mont","PV Jus","DN(mm)","Ext(m)","Mat",
               "CT Mont","CF Mont","Prof Mont","CT Jus","CF Jus","Prof Jus",
               "Decl‰","V(m/s)","Imagem Satélite"]
    for c,h in enumerate(headers,1):
        cell = ws.cell(row=3,column=c,value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = da; cell.border = brd

    for i,t in enumerate(trechos):
        r = i + 4
        pvi = pvs.get(t["pv_ini"],{})
        pvf = pvs.get(t["pv_fim"],{})
        hidr = calc_manning(t.get("dn_mm"), t.get("decl_mm"))

        vals = [
            f"NS-{i+1:04d}", t.get("pv_ini",""), t.get("pv_fim",""),
            t.get("dn_mm"), round(t.get("ext_m",0),2), t.get("material","PVC"),
            pvi.get("ct"), pvi.get("cf"), pvi.get("prof"),
            pvf.get("ct"), pvf.get("cf"), pvf.get("prof"),
            round((t.get("decl_mm") or 0)*1000,2) if t.get("decl_mm") else None,
            hidr.get("v_ms"), ""
        ]
        for c,v in enumerate(vals,1):
            cell = ws.cell(row=r,column=c,value=v)
            cell.font = df; cell.alignment = da; cell.border = brd
            if i%2==1: cell.fill = alt

        # Imagem satélite
        img_path = Path(img_dir) / f"trecho_{i+1:03d}.png"
        try:
            if gerar_img_trecho(pvs, t, img_path):
                img = XlImage(str(img_path))
                img.width = 300; img.height = 180
                ws.add_image(img, f"O{r}")
                ws.row_dimensions[r].height = 140
        except Exception:
            pass

    widths = [8,14,14,8,8,6,9,9,7,9,9,7,7,7,45]
    for c,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(c)].width = w

    r_fin = len(trechos) + 5
    ws.merge_cells(f"A{r_fin}:O{r_fin}")
    ext = sum(t.get("ext_m",0) for t in trechos)
    ws[f"A{r_fin}"] = f"TOTAL: {len(trechos)} trechos | {ext:.1f}m"
    ws[f"A{r_fin}"].font = Font(name="Calibri",bold=True,size=11,color="1a237e")
    wb.save(str(out_path))


# ══════════════════════════════════════════════════════════════════════
# PROCESSAR UM PROJETO (XML) COM RUAS CORRETAS
# ══════════════════════════════════════════════════════════════════════

def processar_projeto(xml_name, nucleo, mapa_dxf_name):
    """Processa UM projeto XML com ruas corretas do mapa DXF."""
    xml_path = INPUT_DIR / xml_name
    mapa_path = INPUT_DIR / mapa_dxf_name

    if not xml_path.exists():
        log(f"XML não encontrado: {xml_path}", "ERR")
        return

    print("\n" + "=" * 72)
    print(f"  PROJETO: {nucleo}")
    print(f"  XML:     {xml_name}")
    print(f"  MAPA:    {mapa_dxf_name}")
    print("=" * 72)

    # 1. Ler LandXML
    log("Lendo LandXML...", "STEP")
    pvs, trechos, ruas_xml, meta = ler_landxml(str(xml_path))
    log(f"  {len(pvs)} PVs, {len(trechos)} trechos, {sum(t.get('ext_m',0) for t in trechos):.0f}m", "OK")

    # 2. Enriquecer
    trechos = enriquecer_trechos(trechos, pvs)

    # 3. Associar ruas pelo NETWORK NAME do XML (método que funcionou!)
    log("Associando ruas pelo network name do XML...", "STEP")
    n_net = associar_ruas_network_xml(str(xml_path), trechos)
    log(f"  {n_net} trechos com rua via network name", "OK")

    # 4. Complementar com MAPA DXF do QGIS (para ruas sem network name)
    log("Complementando ruas via mapa DXF QGIS...", "STEP")
    n_mapa = carregar_ruas_mapa_dxf(str(mapa_path), trechos, pvs)
    log(f"  {n_mapa} trechos complementados via mapa DXF", "OK")

    # 5. Fallback: rede como rua
    for t in trechos:
        rua = limpar_rua(t.get("rua"))
        t["rua"] = rua

    # Estatísticas
    n_com_rua = sum(1 for t in trechos if t.get("rua") != "Sem Rua")
    log(f"  TOTAL com rua: {n_com_rua}/{len(trechos)}", "OK")

    # 6. Criar pastas de saída
    nucleo_slug = slug(nucleo)
    pasta_projeto = OUTPUT_BASE / nucleo_slug
    pastas = {}
    for nome in ["01_NS_CAMPO", "02_DESENHOS", "03_HTML", "04_GIS",
                  "05_PLANILHAS", "11_POR_RUA", "12_LOG"]:
        p = pasta_projeto / nome
        p.mkdir(parents=True, exist_ok=True)
        pastas[nome] = p

    # 7. Gerar NS (A4 + Desenho + Satélite + HTML)
    log("Gerando NS...", "STEP")
    n_ok = n_err = 0
    for i, t in enumerate(trechos):
        ns_id = i + 1
        pv_i = t.get("pv_ini", "PVX")
        pv_f = t.get("pv_fim", "PVY")
        ns_name = _ns_folder_name(ns_id, pv_i, pv_f)
        ns_dir = pastas["01_NS_CAMPO"] / ns_name
        ns_dir.mkdir(parents=True, exist_ok=True)
        try:
            gerar_ns_a4(ns_id, t, pvs, nucleo, str(ns_dir / f"NS{ns_id:03d}_A4.pdf"))
            try:
                gerar_ns_desenho(ns_id, t, pvs, trechos, nucleo,
                                 str(pastas["02_DESENHOS"] / f"NS{ns_id:03d}_DESENHO.pdf"))
            except Exception: pass
            try:
                gerar_ns_sat(ns_id, t, pvs, nucleo,
                             str(pastas["02_DESENHOS"] / f"NS{ns_id:03d}_SAT.pdf"))
            except Exception: pass
            try:
                gerar_html(ns_id, t, pvs, trechos, nucleo,
                           str(pastas["03_HTML"] / f"NS{ns_id:03d}.html"))
            except Exception: pass

            # JSON de dados
            materiais = calcular_materiais(t, pvs)
            dados = {
                "ns_id": ns_id, "nucleo": nucleo, "contrato": CONTRATO,
                "tipo_rede": "ESGOTO",
                "trecho": {k:v for k,v in t.items()},
                "pv_montante": pvs.get(pv_i, {}),
                "pv_jusante": pvs.get(pv_f, {}),
                "hidraulica": calc_manning(t.get("dn_mm"), t.get("decl_mm")),
                "materiais": materiais,
                "gerado_em": datetime.now().isoformat(),
            }
            with open(ns_dir / f"NS{ns_id:03d}_DADOS.json", "w", encoding="utf-8") as f:
                json.dump(dados, f, indent=2, ensure_ascii=False)

            n_ok += 1
            if ns_id <= 3 or ns_id % 50 == 0:
                rua_txt = t.get("rua", "")
                log(f"  NS{ns_id:03d}: {pv_i}→{pv_f} DN{t.get('dn_mm')} {t.get('ext_m',0):.1f}m [{rua_txt}]", "OK")
        except Exception as e:
            n_err += 1
            if ns_id <= 5:
                log(f"  NS{ns_id:03d}: ERRO - {e}", "ERR")

    log(f"  NS: {n_ok} ok, {n_err} erros", "OK")

    # 8. GeoJSON
    try:
        gerar_geojson(trechos, pvs, str(pastas["04_GIS"] / "rede.geojson"))
    except Exception: pass

    # 9. HTML Rede Geral
    try:
        gerar_html(0, trechos[0], pvs, trechos, nucleo,
                   str(pastas["03_HTML"] / "REDE_GERAL.html"))
    except Exception: pass

    # 10. Planilha MESTRE
    if HAS_EXPORT:
        try:
            _gerar_planilha_trechos_completa(
                trechos, pvs, nucleo, "ESGOTO",
                pastas["05_PLANILHAS"] / f"TRECHOS_MESTRE_{nucleo_slug}.xlsx"
            )
            log(f"  Planilha MESTRE: {len(trechos)} trechos", "OK")
        except Exception as e:
            log(f"  Erro planilha mestre: {e}", "WARN")

    # 11. SEPARAR POR RUA COM IMAGENS SATÉLITE
    log("Separando por rua com imagens satélite...", "STEP")
    trechos_por_rua = defaultdict(list)
    for t in trechos:
        trechos_por_rua[t.get("rua", "Sem Rua")].append(t)

    for rua, lista in sorted(trechos_por_rua.items()):
        rua_slug = slug(rua)
        pasta_rua = pastas["11_POR_RUA"] / rua_slug
        pasta_rua.mkdir(parents=True, exist_ok=True)
        img_dir = pasta_rua / "imgs"
        img_dir.mkdir(exist_ok=True)

        log(f"  {rua} ({len(lista)} trechos)...", "INFO")
        try:
            gerar_planilha_rua(lista, pvs, nucleo, rua,
                               pasta_rua / f"Trechos_{rua_slug}.xlsx",
                               str(img_dir))
        except Exception as e:
            log(f"    ERRO: {e}", "ERR")

    log(f"  TOTAL: {len(trechos_por_rua)} ruas", "OK")

    # 12. Log
    log_data = {
        "projeto": xml_name, "nucleo": nucleo,
        "n_pvs": len(pvs), "n_trechos": len(trechos),
        "ext_total_m": round(sum(t.get("ext_m",0) for t in trechos), 1),
        "n_ruas": len(trechos_por_rua),
        "n_com_rua": n_com_rua,
        "n_ns_ok": n_ok, "n_ns_err": n_err,
        "ruas": {rua: len(lista) for rua, lista in trechos_por_rua.items()},
        "gerado_em": datetime.now().isoformat(),
    }
    with open(pastas["12_LOG"] / "log.json", "w", encoding="utf-8") as f:
        json.dump(log_data, f, indent=2, ensure_ascii=False)

    print(f"\n  ✓ PROJETO {nucleo} CONCLUÍDO!")
    print(f"    {n_ok} NS | {len(trechos_por_rua)} ruas | {n_com_rua}/{len(trechos)} com nome")
    return log_data


# ══════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════

def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception: pass

    OUTPUT_BASE.mkdir(parents=True, exist_ok=True)

    print("=" * 72)
    print("  PROCESSAMENTO FOCADO — JOÃO CARLOS + SÃO MANOEL")
    print("  COM RUAS CORRETAS + PRINT SATÉLITE NA PLANILHA")
    print("=" * 72)

    # PROJETO 1: JOÃO CARLOS
    r1 = processar_projeto(
        xml_name="ESTUDO - CT JOÃO CARLOS DA SILVA.xml",
        nucleo="Joao Carlos",
        mapa_dxf_name="MAPA JOÃO CARLOS DA SILVA_RV07 (QGIS).dxf"
    )

    # PROJETO 2: SÃO MANOEL
    r2 = processar_projeto(
        xml_name="ESTUDO - SÃO MANOEL.xml",
        nucleo="Sao Manoel",
        mapa_dxf_name="MAPA SÃO MANOEL_RV05 (QGIS).dxf"
    )

    print("\n" + "=" * 72)
    print("  CONCLUÍDO!")
    print("=" * 72)
    if r1:
        print(f"  João Carlos: {r1['n_trechos']} trechos, {r1['n_ruas']} ruas, {r1['n_com_rua']} com nome")
    if r2:
        print(f"  São Manoel:  {r2['n_trechos']} trechos, {r2['n_ruas']} ruas, {r2['n_com_rua']} com nome")
    print(f"  Saída: {OUTPUT_BASE}")
    print("=" * 72)


if __name__ == "__main__":
    main()
