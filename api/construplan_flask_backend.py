#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
╔══════════════════════════════════════════════════════════════════════════════╗
║  ConstruPlan v1.0 — Backend Python                                          ║
║  Autor: NeryFelipe · 2026                                                   ║
║                                                                              ║
║  SUPERLOG: Este arquivo contém TODA a lógica do sistema.                    ║
║  Se você perdeu o contexto, leia este arquivo — ele é o DNA do ConstruPlan. ║
║                                                                              ║
║  MÓDULOS:                                                                   ║
║    1. Engines: Manning/NBR9649, Quantitativos, CPM/PERT, EVM, LPS           ║
║    2. Importadores: DXF, IFC(CSV), SHP, GeoJSON, GeoPackage, GMW, XLSX      ║
║    3. Planejador 5D: Físico + Financeiro + BIM LOD500                       ║
║    4. Mapa de Calor: por atributos de rede de saneamento                    ║
║    5. PMBOK + EVM: SPI, CPI, EAC, VAC em tempo real                         ║
║    6. LPS (Last Planner System): PPC, restrições, lookahead                 ║
║    7. Replanejamento automático sob premissas                                ║
║    8. Super Log persistente (sobrevive troca de perfil/sessão)               ║
║                                                                              ║
║  ROUTES:                                                                    ║
║    POST /api/importar         — importa qualquer arquivo suportado          ║
║    POST /api/hidraulica        — Manning + tensão trativa NBR9649            ║
║    POST /api/quantitativos    — volumes vala/berço/PV por trecho             ║
║    POST /api/cronograma       — CPM/PERT + Gantt + LPS                      ║
║    POST /api/evm              — EVM: SPI, CPI, EAC, VAC                     ║
║    POST /api/orcamento        — BDI + SINAPI/SABESP                         ║
║    POST /api/planejamento5d   — físico + financeiro + equipes                ║
║    POST /api/mapa_calor       — heatmap por atributos de rede               ║
║    POST /api/lps              — Last Planner System                         ║
║    POST /api/replanejamento   — replanejamento automático                   ║
║    POST /api/pmbok            — métricas PMBOK                              ║
║    GET  /api/trechos          — listar trechos em memória                   ║
║    POST /api/trechos          — salvar/atualizar trechos                    ║
║    GET  /api/logs             — stream de logs (SSE)                        ║
║    GET  /api/superlog         — SUPERLOG completo do projeto                ║
║    POST /api/export/excel     — exportar Excel 10 abas                      ║
║    POST /api/export/geojson   — exportar GeoJSON enriquecido                ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""

from cronograma_cpm import compute_cpm
import os, sys, json, csv, math, logging, traceback, threading, datetime, re, io, copy, uuid
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple

from flask import Flask, request, jsonify, send_file, Response
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════════════
# CONFIGURAÇÃO FLASK
# ═══════════════════════════════════════════════════════════════════════════════
app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB

PORT     = 5002
BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR.parent / "data"
LOGS_DIR = BASE_DIR.parent / "logs"
SUPERLOG_FILE = LOGS_DIR / "superlog.json"

for d in [DATA_DIR, LOGS_DIR, DATA_DIR / "uploads", DATA_DIR / "output"]:
    d.mkdir(parents=True, exist_ok=True)

# ═══════════════════════════════════════════════════════════════════════════════
# SUPERLOG — sobrevive entre sessões e perfis
# ═══════════════════════════════════════════════════════════════════════════════
_mensagens: List[Dict] = []
_lock = threading.Lock()

def _load_superlog():
    global _mensagens
    if SUPERLOG_FILE.exists():
        try:
            with open(SUPERLOG_FILE, 'r', encoding='utf-8') as f:
                _mensagens = json.load(f)
        except:
            _mensagens = []

def _save_superlog():
    try:
        with open(SUPERLOG_FILE, 'w', encoding='utf-8') as f:
            json.dump(_mensagens[-5000:], f, ensure_ascii=False, indent=2)
    except:
        pass

def log(msg: str, nivel: str = "INFO", modulo: str = "SYSTEM"):
    print(f"[{nivel}][{modulo}] {msg}")
    entry = {
        "id":     str(uuid.uuid4())[:8],
        "ts":     datetime.datetime.now().isoformat(),
        "nivel":  nivel,
        "modulo": modulo,
        "msg":    msg
    }
    with _lock:
        _mensagens.append(entry)
        if len(_mensagens) > 5000:
            _mensagens.pop(0)
    _save_superlog()

_load_superlog()
log("ConstruPlan v1.0 iniciado", "INFO", "BOOT")

# ═══════════════════════════════════════════════════════════════════════════════
# ESTADO EM MEMÓRIA (sincronizado com disco)
# ═══════════════════════════════════════════════════════════════════════════════
_trechos: List[Dict]  = []
_projetos: Dict[str, Dict] = {}
_state_file = DATA_DIR / "output" / "construplan_state.json"

def _save_state():
    try:
        with open(_state_file, 'w', encoding='utf-8') as f:
            json.dump({"trechos": _trechos, "projetos": _projetos}, f, ensure_ascii=False, indent=2)
    except Exception as e:
        log(f"Erro ao salvar estado: {e}", "ERROR", "STATE")

def _load_state():
    global _trechos, _projetos
    if _state_file.exists():
        try:
            with open(_state_file, 'r', encoding='utf-8') as f:
                s = json.load(f)
                _trechos  = s.get("trechos", [])
                _projetos = s.get("projetos", {})
                log(f"Estado carregado: {len(_trechos)} trechos, {len(_projetos)} projetos", "INFO", "STATE")
        except Exception as e:
            log(f"Erro ao carregar estado: {e}", "WARN", "STATE")

_load_state()

# ═══════════════════════════════════════════════════════════════════════════════
# CORS
# ═══════════════════════════════════════════════════════════════════════════════
@app.after_request
def add_cors(response):
    response.headers["Access-Control-Allow-Origin"]  = "*"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type,Authorization"
    response.headers["Access-Control-Allow-Methods"] = "GET,POST,PUT,DELETE,OPTIONS"
    return response

@app.route("/api/hello")
def hello():
    return jsonify(ok=True, version="ConstruPlan v1.0", trechos=len(_trechos))

# ═══════════════════════════════════════════════════════════════════════════════
# ENGINE 1 — TABELAS FÍSICAS (NBR 9649, ABNT)
# ═══════════════════════════════════════════════════════════════════════════════
DIE_TABLE = {
    "PVC":      {75:0.085, 100:0.110, 150:0.160, 200:0.200, 250:0.250, 300:0.315, 400:0.400, 500:0.500, 600:0.630},
    "PEAD":     {75:0.090, 100:0.110, 150:0.160, 200:0.225, 250:0.280, 300:0.335, 400:0.450, 500:0.560},
    "CONCRETO": {200:0.260, 250:0.310, 300:0.370, 400:0.480, 500:0.590, 600:0.700, 800:0.920},
    "CERAMICA": {100:0.120, 150:0.170, 200:0.230},
}
MANNING_N = {"PVC":0.013, "PEAD":0.011, "CONCRETO":0.015, "CERAMICA":0.014}

SINAPI_BASE = [
    {"codigo":"74125/003","descricao":"Tubo PVC rígido esgoto DN 200mm","unidade":"m","preco_ref":64.89},
    {"codigo":"74100/001","descricao":"Escavação vala prof 0-1,5m mec","unidade":"m³","preco_ref":8.12},
    {"codigo":"74090/002","descricao":"Reaterro compactado manual","unidade":"m³","preco_ref":14.37},
    {"codigo":"74091/001","descricao":"Areia fina lavada (lastro/envolt.)","unidade":"m³","preco_ref":118.40},
    {"codigo":"73965/002","descricao":"Poço de visita PV-1200 moldado","unidade":"un","preco_ref":2840.00},
    {"codigo":"74125/001","descricao":"Tubo PVC rígido esgoto DN 150mm","unidade":"m","preco_ref":43.20},
    {"codigo":"74125/005","descricao":"Tubo PVC rígido esgoto DN 300mm","unidade":"m","preco_ref":128.50},
    {"codigo":"74101/003","descricao":"Escavação vala prof 1,5-3m mec","unidade":"m³","preco_ref":12.86},
    {"codigo":"97645/001","descricao":"CBUQ faixa C camada rolamento","unidade":"m²","preco_ref":38.74},
    {"codigo":"94494",    "descricao":"Calçada conc simples e=6cm","unidade":"m²","preco_ref":42.18},
    {"codigo":"97300/001","descricao":"Escav. mec. vala prof 3-4m","unidade":"m³","preco_ref":18.94},
    {"codigo":"74200/002","descricao":"Concreto fck=20MPa para PV","unidade":"m³","preco_ref":418.00},
]

CFG_OBRA_DEFAULT = {
    "largura_minima_m":       0.80,
    "folga_lateral_m":        0.15,
    "talude_hv":              0.00,
    "lastro_areia_m":         0.10,
    "envoltorio_areia_m":     0.30,
    "dreno_ativo":            True,
    "dreno_brita_m":          0.20,
    "fator_empolamento":      1.30,
    "pv_diam_int_m":          1.20,
    "pv_espessura_parede_m":  0.12,
    "pv_folga_exc_m":         0.30,
    "perc_asfalto":           0.55,
    "perc_passeio":           0.30,
    "bgs_espessura_m":        0.15,
    "cbuq_espessura_m":       0.05,
    "prod_escavacao_m3d":     80.0,
    "prod_assentamento_md":   80.0,
    "n_equipes_assentamento": 1,
    "eficiencia_campo":       0.85,
    "data_inicio_obra":       "2025-03-01",
    "data_fim_contratual":    "2026-06-30",
    "custo_escavacao_m3":     50.00,
    "custo_reaterro_m3":      35.00,
    "custo_asfalto_m2":       85.00,
    "custo_passeio_m2":       45.00,
    "custo_pv_un":          3000.00,
    "custo_tubo_150_m":     1648.38,
    "custo_tubo_200_m":     1888.49,
    "custo_tubo_250_m":     2100.00,
    "custo_tubo_300_m":     2400.00,
    "bdi_perc":               25.00,
    "n_equipes_escavacao":    1,
    "feriados": [],
}

def get_die(dn: int, mat: str) -> float:
    mat_up = mat.upper()
    key = "PVC"
    for k in DIE_TABLE:
        if k in mat_up:
            key = k; break
    tb = DIE_TABLE[key]
    if dn in tb: return tb[dn]
    dns = sorted(tb.keys())
    if dn <= dns[0]: return tb[dns[0]]
    if dn >= dns[-1]: return tb[dns[-1]]
    for i in range(len(dns)-1):
        if dns[i] <= dn <= dns[i+1]:
            t = (dn - dns[i]) / (dns[i+1] - dns[i])
            return tb[dns[i]] + t * (tb[dns[i+1]] - tb[dns[i]])
    return dn/1000 + 0.04

def get_manning(mat: str) -> float:
    mat_up = mat.upper()
    for k, v in MANNING_N.items():
        if k in mat_up: return v
    return 0.013

# ═══════════════════════════════════════════════════════════════════════════════
# ENGINE 2 — HIDRÁULICA (Manning + NBR 9649:1986)
# ═══════════════════════════════════════════════════════════════════════════════
def hydro(dn_m: float, decl_mm: float, n: float, q_ls: float) -> dict:
    """Manning + lâmina parcial + tensão trativa — ABNT NBR 9649:1986"""
    I = decl_mm / 1000.0
    Q = q_ls / 1000.0
    D = dn_m
    if I <= 0 or D <= 0:
        return {"vel":0, "yd":0, "tt":0, "qP":0, "ok":False, "alertas":[]}
    A  = math.pi * D * D / 4
    P  = math.pi * D
    Rh = A / P
    qP = (1/n) * A * (Rh**(2/3)) * (I**0.5)
    if Q <= 0:
        tt = round(9810 * Rh * I, 3)
        return {"vel":0, "yd":0, "tt":tt, "qP":round(qP*1000,3), "ok":tt>=1.0, "alertas":[]}
    lo, hi = 0.001, 0.999
    for _ in range(80):
        mid = (lo + hi) / 2
        th  = 2 * math.acos(max(-1, min(1, 1 - 2*mid)))
        a   = D*D/4 * (th - math.sin(th)) / 2
        p_  = D * th / 2
        if p_ <= 0: lo = mid; continue
        rh_ = a / p_
        qc  = (1/n) * a * (rh_**(2/3)) * (I**0.5)
        if qc < Q: lo = mid
        else:       hi = mid
    yd   = mid
    th_f = 2 * math.acos(max(-1, min(1, 1 - 2*yd)))
    a_f  = D*D/4 * (th_f - math.sin(th_f)) / 2
    vel  = Q / a_f if a_f > 0 else 0
    p_f  = D * th_f / 2
    rh_f = a_f / p_f if p_f > 0 else 0
    tt   = 9810 * rh_f * I

    alertas = []
    if vel < 0.6:   alertas.append("⚠️ Velocidade mínima < 0,60 m/s (NBR 9649 §4.2)")
    if vel > 5.0:   alertas.append("⚠️ Velocidade > 5,0 m/s — verificar erosão")
    if yd > 0.75:   alertas.append("⚠️ Lâmina > 75% D — risco de pressurização")
    if tt < 1.0:    alertas.append("⚠️ Tensão trativa < 1,0 Pa — risco de deposição")
    if tt > 8.0:    alertas.append("⚠️ Tensão trativa > 8,0 Pa — verificar ancoragem")

    return {
        "vel":     round(vel, 3),
        "yd":      round(yd, 4),
        "yd_pct":  round(yd * 100, 1),
        "tt":      round(tt, 3),
        "qP":      round(qP * 1000, 3),
        "ok":      len(alertas) == 0,
        "alertas": alertas,
    }

# ═══════════════════════════════════════════════════════════════════════════════
# ENGINE 3 — NORMALIZAÇÃO DE CAMPOS
# ═══════════════════════════════════════════════════════════════════════════════
def normalizar_trecho(t: dict) -> dict:
    def fv(keys, default=0.0):
        for k in keys:
            if k in t and t[k] not in (None, ""):
                try: return float(t[k])
                except: pass
        return default
    def sv(keys, default=""):
        for k in keys:
            if k in t and t[k] not in (None, ""):
                return str(t[k]).strip()
        return default

    ext  = fv(["extensao","Extensao","extensão","length","comprimento"], 0)
    dn   = fv(["dn","DN","diametro","Diametro","diameter"], 200)
    mat  = sv(["material","Material","mat"], "PVC")
    ctm  = fv(["cotaTerrenoMontante","cota_terreno_montante","Cota_Ter_Montante","CTm","ctm"], 0)
    ctj  = fv(["cotaTerrenoJusante","cota_terreno_jusante","Cota_Ter_Jusante","CTj","ctj"], 0)
    gim  = fv(["cotaFundoMontante","cota_fundo_montante","Cota_Col_Montante","GIm","gim","cota_soleira_mont"], 0)
    gij  = fv(["cotaFundoJusante","cota_fundo_jusante","Cota_Col_Jusante","GIj","gij","cota_soleira_jus"], 0)
    profm = fv(["profundidade","prof_montante","profMont"], 0)

    if profm == 0 and ctm > 0 and gim > 0:
        die_val = get_die(int(dn), mat)
        profm = ctm - gim - die_val
        if profm < 0: profm = ctm - gim if ctm > gim else 1.5

    decl = fv(["declividade","Declividade","slope","i_mm"], 0)
    if decl == 0 and ext > 0 and gim > 0 and gij > 0:
        decl = (gim - gij) / ext
    if decl > 1: decl /= 1000

    n_tag   = sv(["montante","Montante","PV_Mont","pv_montante","origem"], "PV-01")
    j_tag   = sv(["jusante","Jusante","PV_Jus","pv_jusante","destino"], "PV-02")
    rua     = sv(["nomeRua","nome_rua","rua","Rua","logradouro","street"], "")
    bairro  = sv(["bairro","Bairro","district"], "")
    area_tag = sv(["area_irregular","area","tag_nucleo","TAG_NUCLEO","projeto"], "")

    # Atributos extras (vêm de IFC/CSV GlobalMapper)
    ifc_name = sv(["IFC_NAME","ifc_name","nome_ifc"], "")
    ifc_desc = sv(["IFC_DESCRIPTION","ifc_description","desc_ifc"], "")
    glob_id  = sv(["IFC_GLOBALID","globalid","IFC_GLOBALID"], "")

    # Custo direto (se vier da tabela SABESP)
    custo_unit = fv(["custo_unitario","custo_unit","preco_unit","valor_unit"], 0)

    return {
        "id":                   sv(["id","ID"], str(uuid.uuid4())[:8]),
        "montante":             n_tag,
        "jusante":              j_tag,
        "extensao":             round(ext, 3),
        "dn":                   int(dn),
        "material":             mat,
        "cotaTerrenoMontante":  round(ctm, 3),
        "cotaTerrenoJusante":   round(ctj, 3),
        "cotaFundoMontante":    round(gim, 3),
        "cotaFundoJusante":     round(gij, 3),
        "profundidade":         round(profm, 3),
        "declividade":          round(decl, 6),
        "nomeRua":              rua,
        "bairro":               bairro,
        "area_irregular":       area_tag,
        "ifc_name":             ifc_name,
        "ifc_descricao":        ifc_desc,
        "ifc_globalid":         glob_id,
        "custo_unitario":       custo_unit,
        "_raw": t,
    }

# ═══════════════════════════════════════════════════════════════════════════════
# ENGINE 4 — QUANTITATIVOS
# ═══════════════════════════════════════════════════════════════════════════════
def vols(B: float, die: float, PGI: float) -> dict:
    L  = max(0.70, die + 0.08)
    v0 = B * L * min(PGI, 1)
    v1 = B * L * max(0, min(PGI-1, 1))
    v2 = B * L * max(0, min(PGI-2, 1))
    v3 = B * L * max(0, min(PGI-3, 1))
    v4 = B * L * max(0, PGI-4)
    return {"v0":round(v0,3),"v1":round(v1,3),"v2":round(v2,3),
            "v3":round(v3,3),"v4":round(v4,4),"tot":round(v0+v1+v2+v3+v4,3),"L":round(L,3)}

def calc_quant_trecho(t: dict, cfg: dict = None) -> dict:
    t  = normalizar_trecho(t)
    c  = {**CFG_OBRA_DEFAULT, **(cfg or {})}
    ext    = float(t["extensao"]) or 1
    dn_mm  = float(t["dn"])
    dn_m   = dn_mm / 1000.0
    ct_m   = float(t["cotaTerrenoMontante"])
    ct_j   = float(t["cotaTerrenoJusante"])
    gi_m   = float(t["cotaFundoMontante"])
    gi_j   = float(t["cotaFundoJusante"])
    mat    = t["material"]
    die    = get_die(int(dn_mm), mat)

    prof_m = (ct_m - gi_m - die) if ct_m > gi_m else float(t.get("profundidade", 1.5))
    prof_j = (ct_j - gi_j - die) if ct_j > gi_j else prof_m
    if prof_m <= 0: prof_m = float(t.get("profundidade", 1.5))
    if prof_j <= 0: prof_j = prof_m
    prof_med = (prof_m + prof_j) / 2

    B  = max(c["largura_minima_m"], die + 2 * c["folga_lateral_m"])
    B += c["talude_hv"] * prof_med * 2

    vM = vols(B, die, prof_m)
    vJ = vols(B, die, prof_j)

    vol_esc_tot = round((vM["tot"] + vJ["tot"]) / 2 * ext, 3)
    vol_esc_0   = round((vM["v0"]  + vJ["v0"]) / 2 * ext, 3)
    vol_esc_1   = round((vM["v1"]  + vJ["v1"]) / 2 * ext, 3)
    vol_esc_2   = round((vM["v2"]  + vJ["v2"]) / 2 * ext, 3)
    vol_esc_3   = round((vM["v3"]  + vJ["v3"]) / 2 * ext, 3)
    vol_esc_4   = round((vM["v4"]  + vJ["v4"]) / 2 * ext, 3)

    vol_lastro   = round(B * c["lastro_areia_m"] * ext, 3)
    vol_envolt   = round(B * c["envoltorio_areia_m"] * ext, 3)
    vol_reaterro = round(max(0, vol_esc_tot - vol_lastro - vol_envolt - (math.pi*(die/2)**2)*ext), 3)
    vol_reaterro_comp = round(vol_reaterro / c["fator_empolamento"], 3)

    # PV
    pv_diam_ext = c["pv_diam_int_m"] + 2 * c["pv_espessura_parede_m"]
    B_pv        = pv_diam_ext + 2 * c["pv_folga_exc_m"]
    vol_esc_pv  = round(B_pv * B_pv * prof_m, 3)

    # Pavimentação
    area_asf  = round(B * ext * c["perc_asfalto"], 2)
    area_pass = round(B * ext * c["perc_passeio"], 2)

    # Hidráulica
    decl  = float(t["declividade"])
    n_    = get_manning(mat)
    h_res = hydro(dn_m, decl * 1000, n_, 0)

    # Custo
    custos = {
        "escavacao":  round(vol_esc_tot * c["custo_escavacao_m3"], 2),
        "reaterro":   round(vol_reaterro_comp * c["custo_reaterro_m3"], 2),
        "asfalto":    round(area_asf * c["custo_asfalto_m2"], 2),
        "passeio":    round(area_pass * c["custo_passeio_m2"], 2),
        "tubo":       round(ext * c.get(f"custo_tubo_{int(dn_mm)}_m", c.get("custo_tubo_200_m", 1888.49)), 2),
        "pv":         round(c["custo_pv_un"], 2),
    }
    custos["total_sem_bdi"] = sum(custos.values())
    custos["bdi"] = round(custos["total_sem_bdi"] * c["bdi_perc"] / 100, 2)
    custos["total_com_bdi"] = round(custos["total_sem_bdi"] + custos["bdi"], 2)

    return {
        **t,
        "vol_escavacao_tot": vol_esc_tot,
        "vol_esc_0_1m":      vol_esc_0,
        "vol_esc_1_2m":      vol_esc_1,
        "vol_esc_2_3m":      vol_esc_2,
        "vol_esc_3_4m":      vol_esc_3,
        "vol_esc_4m":        vol_esc_4,
        "vol_lastro_areia":  vol_lastro,
        "vol_envoltorio":    vol_envolt,
        "vol_reaterro_comp": vol_reaterro_comp,
        "vol_esc_pv":        vol_esc_pv,
        "largura_vala_m":    round(B, 3),
        "area_asfalto_m2":   area_asf,
        "area_passeio_m2":   area_pass,
        "prof_media_m":      round(prof_med, 3),
        "hidraulica":        h_res,
        "custos":            custos,
    }

# ═══════════════════════════════════════════════════════════════════════════════
# ENGINE 5 — DIAS ÚTEIS + CRONOGRAMA CPM
# ═══════════════════════════════════════════════════════════════════════════════
def dias_uteis(inicio: datetime.date, n: int, feriados=None) -> datetime.date:
    feriados = set(feriados or [])
    d = inicio; count = 0
    while count < n:
        d += datetime.timedelta(days=1)
        if d.weekday() < 5 and d not in feriados:
            count += 1
    return d

def calc_cronograma_sequencial(trechos: list, cfg: dict = None) -> dict:
    c       = {**CFG_OBRA_DEFAULT, **(cfg or {})}
    inicio  = datetime.datetime.strptime(c["data_inicio_obra"], "%Y-%m-%d").date()
    prod    = c["prod_assentamento_md"] * c["n_equipes_assentamento"] * c["eficiencia_campo"]
    resultado = []
    dt_atual  = inicio
    custo_acum = 0

    for t in trechos:
        ext  = float(t.get("extensao", 0)) or 1
        dur  = max(1, math.ceil(ext / prod))
        dt_fim = dias_uteis(dt_atual, dur, c.get("feriados", []))
        q    = calc_quant_trecho(t, c)
        custo_acum += q["custos"]["total_com_bdi"]

        resultado.append({
            "montante":    t.get("montante",""),
            "jusante":     t.get("jusante",""),
            "extensao":    ext,
            "duracao_dias": dur,
            "dataInicio":  dt_atual.strftime("%Y-%m-%d"),
            "dataFim":     dt_fim.strftime("%Y-%m-%d"),
            "custo":       q["custos"]["total_com_bdi"],
            "custo_acum":  round(custo_acum, 2),
            "dn":          t.get("dn", 200),
            "rua":         t.get("nomeRua", t.get("rua", "")),
        })
        dt_atual = dt_fim

    total_dias  = (resultado[-1]["dataFim"] if resultado else inicio.strftime("%Y-%m-%d"))
    total_custo = round(sum(r["custo"] for r in resultado), 2)

    return {
        "ok": True,
        "total_trechos": len(resultado),
        "total_extensao": round(sum(r["extensao"] for r in resultado), 2),
        "total_dias": (datetime.datetime.strptime(resultado[-1]["dataFim"],"%Y-%m-%d").date() - inicio).days if resultado else 0,
        "total_custo": total_custo,
        "data_inicio": inicio.strftime("%Y-%m-%d"),
        "data_fim_prevista": resultado[-1]["dataFim"] if resultado else "",
        "trechos": resultado,
    }

# ═══════════════════════════════════════════════════════════════════════════════
# ENGINE 6 — EVM (Earned Value Management) — PMBOK 7ª ed
# ═══════════════════════════════════════════════════════════════════════════════
def calc_evm(bac: float, pv: float, ev: float, ac: float) -> dict:
    """
    BAC = Budget at Completion
    PV  = Planned Value
    EV  = Earned Value
    AC  = Actual Cost
    """
    cv  = round(ev - ac, 2)            # Cost Variance
    sv  = round(ev - pv, 2)            # Schedule Variance
    cpi = round(ev / ac, 4) if ac > 0 else 0   # Cost Performance Index
    spi = round(ev / pv, 4) if pv > 0 else 0   # Schedule Performance Index
    eac_cpi  = round(bac / cpi, 2) if cpi > 0 else 0    # EAC by CPI
    eac_cpis = round(ac + (bac - ev) / (cpi * spi), 2) if (cpi * spi) > 0 else 0
    eac_rem  = round(ac + (bac - ev), 2)         # EAC by remaining
    etc      = round(eac_cpi - ac, 2)            # Estimate to Complete
    vac      = round(bac - eac_cpi, 2)           # Variance at Completion
    tcpi     = round((bac - ev) / (bac - ac), 4) if (bac - ac) > 0 else 0

    status_custo   = "✅ Abaixo do orçamento" if cv >= 0 else "🔴 Acima do orçamento"
    status_prazo   = "✅ Adiantado" if sv >= 0 else "🔴 Atrasado"

    return {
        "ok": True,
        "BAC": bac, "PV": pv, "EV": ev, "AC": ac,
        "CV": cv, "SV": sv, "CPI": cpi, "SPI": spi,
        "EAC_CPI": eac_cpi, "EAC_CPIS": eac_cpis, "EAC_REM": eac_rem,
        "ETC": etc, "VAC": vac, "TCPI": tcpi,
        "perc_concluido": round(ev/bac*100, 1) if bac > 0 else 0,
        "status_custo":   status_custo,
        "status_prazo":   status_prazo,
        "alertas": [status_custo if cv < 0 else "", status_prazo if sv < 0 else ""]
    }

# ═══════════════════════════════════════════════════════════════════════════════
# ENGINE 7 — LPS (Last Planner System)
# ═══════════════════════════════════════════════════════════════════════════════
def calc_lps(semanas: list) -> dict:
    """
    semanas = [{"semana": "S1", "planejadas": 10, "concluidas": 8, "restricoes": [...]}]
    Retorna PPC por semana e acumulado
    """
    resultado = []
    total_p = total_c = 0
    for s in semanas:
        p = int(s.get("planejadas", 0))
        c = int(s.get("concluidas", 0))
        ppc = round(c / p * 100, 1) if p > 0 else 0
        total_p += p; total_c += c
        resultado.append({
            **s,
            "ppc": ppc,
            "status": "✅" if ppc >= 80 else ("⚠️" if ppc >= 60 else "🔴"),
        })
    ppc_acum = round(total_c / total_p * 100, 1) if total_p > 0 else 0
    return {
        "ok": True,
        "semanas": resultado,
        "ppc_acumulado": ppc_acum,
        "meta_ppc": 80,
        "status_geral": "✅ Meta atingida" if ppc_acum >= 80 else "🔴 Abaixo da meta",
        "total_planejadas": total_p,
        "total_concluidas": total_c,
    }

# ═══════════════════════════════════════════════════════════════════════════════
# ENGINE 8 — MAPA DE CALOR por atributos de rede
# ═══════════════════════════════════════════════════════════════════════════════
def calc_mapa_calor(trechos: list, atributo: str = "profundidade") -> dict:
    """
    Gera dados para mapa de calor baseado em qualquer atributo numérico.
    Suporta: profundidade, declividade, dn, custo, tensao_trativa, velocidade
    """
    valores = []
    features = []

    for t in trechos:
        t_norm = normalizar_trecho(t)
        quant  = calc_quant_trecho(t_norm)

        # Resolver atributo
        if atributo == "profundidade":
            val = quant.get("prof_media_m", 0)
        elif atributo == "declividade":
            val = quant.get("declividade", 0) * 1000  # em ‰
        elif atributo == "tensao_trativa":
            val = quant.get("hidraulica", {}).get("tt", 0)
        elif atributo == "velocidade":
            val = quant.get("hidraulica", {}).get("vel", 0)
        elif atributo == "custo_m":
            val = quant["custos"]["total_com_bdi"] / max(quant["extensao"], 1)
        elif atributo == "dn":
            val = quant.get("dn", 200)
        else:
            val = quant.get(atributo, 0) or 0

        valores.append(val)
        features.append({
            "montante": t_norm["montante"],
            "jusante":  t_norm["jusante"],
            "rua":      t_norm["nomeRua"],
            "extensao": t_norm["extensao"],
            "valor":    round(float(val), 4),
            "atributo": atributo,
        })

    if not valores:
        return {"ok": False, "erro": "Nenhum trecho"}

    vmin = min(valores)
    vmax = max(valores)
    vmed = sum(valores) / len(valores)
    vrange = vmax - vmin or 1

    # Classificar em quintis (0-4) para cor
    def cor_quintil(v):
        q = (v - vmin) / vrange
        if q < 0.2: return "#00ff00"   # verde
        if q < 0.4: return "#aaff00"   # amarelo-verde
        if q < 0.6: return "#ffff00"   # amarelo
        if q < 0.8: return "#ff8800"   # laranja
        return "#ff0000"               # vermelho

    for f in features:
        f["cor"]     = cor_quintil(f["valor"])
        f["quintil"] = int((f["valor"] - vmin) / vrange * 4.99)
        f["label"]   = f"{atributo}: {f['valor']:.2f}"

    # Alertas automáticos baseados em thresholds de engenharia
    alertas = []
    atrib_thresh = {
        "profundidade":    {"warn": 3.0, "crit": 4.0, "unit": "m"},
        "declividade":     {"warn": 0.5, "crit": 0.2, "unit": "‰", "inv": True},
        "tensao_trativa":  {"warn": 1.0, "crit": 0.8, "unit": "Pa", "inv": True},
        "velocidade":      {"warn": 0.6, "crit": 0.4, "unit": "m/s", "inv": True},
    }
    if atributo in atrib_thresh:
        thresh = atrib_thresh[atributo]
        inv    = thresh.get("inv", False)
        for f in features:
            v = f["valor"]
            if (not inv and v >= thresh["crit"]) or (inv and v <= thresh["crit"]):
                alertas.append(f"🔴 {f['montante']}→{f['jusante']}: {atributo}={v:.2f}{thresh['unit']} CRÍTICO")
            elif (not inv and v >= thresh["warn"]) or (inv and v <= thresh["warn"]):
                alertas.append(f"⚠️ {f['montante']}→{f['jusante']}: {atributo}={v:.2f}{thresh['unit']}")

    return {
        "ok": True,
        "atributo": atributo,
        "min": round(vmin, 4),
        "max": round(vmax, 4),
        "media": round(vmed, 4),
        "features": features,
        "alertas": alertas[:50],
        "escala": [
            {"label": "Baixo",   "cor": "#00ff00"},
            {"label": "Médio-baixo", "cor": "#aaff00"},
            {"label": "Médio",   "cor": "#ffff00"},
            {"label": "Médio-alto","cor": "#ff8800"},
            {"label": "Alto",    "cor": "#ff0000"},
        ]
    }

# ═══════════════════════════════════════════════════════════════════════════════
# ENGINE 9 — IMPORTADORES (DXF, IFC via CSV, CSV, GeoJSON, XLSX)
# ═══════════════════════════════════════════════════════════════════════════════
def importar_csv_globalmapper(content: str) -> List[Dict]:
    """Importa CSV exportado pelo GlobalMapper de arquivos IFC (PV/PI/tubos)"""
    trechos = []
    try:
        reader = csv.DictReader(io.StringIO(content), delimiter=';')
        pvs = []
        for row in reader:
            nome = row.get("IFC_NAME","").strip()
            desc = row.get("IFC_DESCRIPTION","").strip()
            globid = row.get("IFC_GLOBALID","").strip()
            # Extrair dimensões da descrição: "0.5x1.2m" → diâm x prof
            dims = re.findall(r'[\d\.]+', desc)
            prof = float(dims[-1]) if dims else 1.0
            diam = float(dims[0]) * 1000 if dims else 500.0  # mm
            pvs.append({
                "ifc_name": nome,
                "ifc_desc": desc,
                "ifc_globalid": globid,
                "profundidade": prof,
                "dn_pv": diam,
                "tipo": "PV" if "visita" in nome.lower() or "pv" in nome.lower() else "PI",
            })
        # Gerar trechos sequenciais entre PVs
        for i in range(len(pvs) - 1):
            p1, p2 = pvs[i], pvs[i+1]
            trechos.append({
                "montante":  f"{p1['tipo']}-{i+1:03d}",
                "jusante":   f"{p2['tipo']}-{i+2:03d}",
                "extensao":  20.0,  # padrão quando não há coordenadas
                "dn":        200,
                "material":  "PVC",
                "profundidade": (p1["profundidade"] + p2["profundidade"]) / 2,
                "ifc_name":  p1["ifc_name"],
                "ifc_descricao": p1["ifc_desc"],
                "ifc_globalid": p1["ifc_globalid"],
            })
    except Exception as e:
        log(f"Erro importar CSV GlobalMapper: {e}", "ERROR", "IMPORT")
    return trechos

def importar_csv_sabesp(content: str) -> List[Dict]:
    """Importa tabela de custos SABESP (base.csv)"""
    itens = []
    try:
        reader = csv.DictReader(io.StringIO(content), delimiter=';')
        for row in reader:
            codigo = row.get("codigo","").strip()
            if not codigo or codigo == "Item": continue
            try:
                custo = float(str(row.get("custo_unitario","0")).replace(",","."))
            except:
                custo = 0.0
            itens.append({
                "codigo":       codigo,
                "descricao":    row.get("descricao",""),
                "unidade":      row.get("unidade",""),
                "custo_unitario": custo,
                "categoria":    row.get("categoria",""),
            })
    except Exception as e:
        log(f"Erro importar CSV SABESP: {e}", "ERROR", "IMPORT")
    return itens

def importar_csv_generico(content: str) -> List[Dict]:
    """Importa CSV genérico tentando detectar formato"""
    # Detectar separador
    sep = ';' if content.count(';') > content.count(',') else ','
    try:
        df = pd.read_csv(io.StringIO(content), sep=sep, encoding='utf-8', low_memory=False)
        records = df.where(pd.notnull(df), None).to_dict('records')
        return [normalizar_trecho(r) for r in records]
    except Exception as e:
        log(f"Erro importar CSV genérico: {e}", "ERROR", "IMPORT")
        return []

def importar_geojson(content: str) -> List[Dict]:
    """Importa GeoJSON com features de trechos"""
    trechos = []
    try:
        gj = json.loads(content)
        for feat in gj.get("features", []):
            props = feat.get("properties", {})
            geom  = feat.get("geometry", {})
            if geom.get("type") == "LineString":
                coords = geom.get("coordinates", [])
                if len(coords) >= 2:
                    import math as m
                    dx = coords[-1][0] - coords[0][0]
                    dy = coords[-1][1] - coords[0][1]
                    props["extensao"] = round(m.sqrt(dx**2 + dy**2), 2)
            trechos.append(normalizar_trecho(props))
    except Exception as e:
        log(f"Erro importar GeoJSON: {e}", "ERROR", "IMPORT")
    return trechos

def importar_dxf(content: str) -> List[Dict]:
    """Parser DXF — extrai LWPOLYLINE e MTEXT com info de rede"""
    trechos = []
    lines   = content.splitlines()
    pvs     = {}
    # Reusar parser do construdata_backend
    i = 0
    while i < len(lines):
        if lines[i].strip() == "0" and i+1 < len(lines) and lines[i+1].strip() in ("MTEXT","TEXT"):
            j = i + 2
            layer = handle = texto = ""
            while j < len(lines) and not (lines[j].strip() == "0"):
                code = lines[j].strip()
                val  = lines[j+1].strip() if j+1 < len(lines) else ""
                if code == "8":  layer  = val
                if code == "5":  handle = val
                if code == "1":  texto  = val
                j += 2
            # TAG da área irregular
            tm = re.match(r'(0\d{5})', texto)
            if tm:
                pvs[handle] = {"tag": tm.group(1), "nome": texto, "layer": layer}
            i = j
        else:
            i += 1

    # LWPOLYLINE como trechos
    parts = content.split('\n  0\n')
    idx   = 0
    for part in parts:
        if not part.strip().startswith('LWPOLYLINE'): continue
        gc = {}
        lns = part.strip().split('\n')
        k = 1
        while k < len(lns) - 1:
            c = lns[k].strip(); v = lns[k+1].strip()
            if c not in gc: gc[c] = []
            gc[c].append(v)
            k += 2
        layer = gc.get('8', [''])[0]
        if 'AREA' not in layer.upper() and 'ESGOTO' not in layer.upper() and 'TUBO' not in layer.upper():
            idx += 1; continue
        xs = gc.get('10', [])
        ys = gc.get('20', [])
        if len(xs) >= 2:
            try:
                coords = [(float(x), float(y)) for x, y in zip(xs, ys)]
                ext = sum(math.sqrt((coords[i+1][0]-coords[i][0])**2 + (coords[i+1][1]-coords[i][1])**2)
                         for i in range(len(coords)-1))
                trechos.append({
                    "montante":  f"PV-{idx:03d}",
                    "jusante":   f"PV-{idx+1:03d}",
                    "extensao":  round(ext, 2),
                    "dn":        200,
                    "material":  "PVC",
                    "layer":     layer,
                    "source":    "DXF",
                })
                idx += 1
            except:
                pass

    log(f"DXF: {len(trechos)} trechos importados", "INFO", "IMPORT")
    return trechos

def importar_xlsx(content_bytes: bytes) -> List[Dict]:
    """Importa planilha Excel com trechos ou custos"""
    trechos = []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content_bytes), data_only=True)
        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            headers = [str(c.value or "").strip() for c in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(v is None for v in row): continue
                d = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
                trechos.append(d)
        log(f"XLSX: {len(trechos)} linhas importadas", "INFO", "IMPORT")
    except Exception as e:
        log(f"Erro importar XLSX: {e}", "ERROR", "IMPORT")
    return trechos

# ═══════════════════════════════════════════════════════════════════════════════
# ENGINE 10 — PLANEJADOR 5D (físico + financeiro + BIM)
# ═══════════════════════════════════════════════════════════════════════════════
def planejamento_5d(trechos: list, cfg: dict = None) -> dict:
    """
    5D = 3D (geometria BIM) + 4D (tempo) + 5D (custo)
    Retorna cronograma com curva S, cash flow e EVM projetado
    """
    c      = {**CFG_OBRA_DEFAULT, **(cfg or {})}
    cron   = calc_cronograma_sequencial(trechos, c)
    bac    = cron["total_custo"]

    # Curva S (distribuição de custo ao longo do tempo)
    curva_s = []
    acum    = 0
    for t in cron["trechos"]:
        acum += t["custo"]
        curva_s.append({
            "data":       t["dataFim"],
            "custo_per":  t["custo"],
            "custo_acum": round(acum, 2),
            "pct":        round(acum / bac * 100, 2) if bac > 0 else 0,
        })

    # Resumo 5D por faixa de profundidade
    faixas = {"0-1m": 0, "1-2m": 0, "2-3m": 0, ">3m": 0}
    for t in trechos:
        q = calc_quant_trecho(t, c)
        p = q.get("prof_media_m", 0)
        k = "0-1m" if p < 1 else ("1-2m" if p < 2 else ("2-3m" if p < 3 else ">3m"))
        faixas[k] += q["custos"]["total_com_bdi"]

    return {
        "ok": True,
        "resumo_fisico": {
            "total_extensao_m": cron["total_extensao"],
            "total_trechos":    cron["total_trechos"],
            "total_dias":       cron["total_dias"],
        },
        "resumo_financeiro": {
            "BAC":        round(bac, 2),
            "custo_m":    round(bac / cron["total_extensao"], 2) if cron["total_extensao"] > 0 else 0,
            "por_faixa":  {k: round(v, 2) for k, v in faixas.items()},
        },
        "curva_s": curva_s,
        "cronograma": cron["trechos"],
        "bim_lod": "LOD 300 (geometria paramétrica por dimensões IFC)",
    }

# ═══════════════════════════════════════════════════════════════════════════════
# ENGINE 11 — PMBOK (métricas e status do projeto)
# ═══════════════════════════════════════════════════════════════════════════════
def calc_pmbok(projeto: dict) -> dict:
    """
    Calcula métricas PMBOK 7ª ed: áreas de conhecimento, performance, riscos
    """
    bac    = float(projeto.get("bac", 0))
    pv     = float(projeto.get("pv", 0))
    ev     = float(projeto.get("ev", 0))
    ac     = float(projeto.get("ac", 0))
    riscos = projeto.get("riscos", [])

    evm    = calc_evm(bac, pv, ev, ac)
    lps_d  = projeto.get("lps_semanas", [])
    lps_r  = calc_lps(lps_d) if lps_d else {"ppc_acumulado": 0}

    # Score de saúde do projeto (0-100)
    score = 100
    if evm["CPI"] < 1:  score -= int((1 - evm["CPI"]) * 30)
    if evm["SPI"] < 1:  score -= int((1 - evm["SPI"]) * 25)
    if lps_r["ppc_acumulado"] < 80: score -= int((80 - lps_r["ppc_acumulado"]) / 4)
    score = max(0, score)

    # Classificar riscos
    riscos_altos = [r for r in riscos if r.get("impacto","") == "ALTO"]

    return {
        "ok": True,
        "evm": evm,
        "lps": lps_r,
        "score_saude": score,
        "status": "🟢 Saudável" if score >= 75 else ("🟡 Atenção" if score >= 50 else "🔴 Crítico"),
        "riscos_altos": len(riscos_altos),
        "areas_criticas": [
            "Custo" if evm["CPI"] < 1 else None,
            "Prazo" if evm["SPI"] < 1 else None,
            "Qualidade" if lps_r["ppc_acumulado"] < 70 else None,
        ],
        "recomendacoes": _recomendacoes_pmbok(evm, lps_r, score),
    }

def _recomendacoes_pmbok(evm, lps, score):
    rec = []
    if evm["CPI"] < 0.9:
        rec.append("🔴 CPI crítico — revisar composições de custo e negociar fornecedores")
    elif evm["CPI"] < 1.0:
        rec.append("⚠️ CPI abaixo de 1.0 — monitorar custos semanalmente")
    if evm["SPI"] < 0.9:
        rec.append("🔴 SPI crítico — aumentar equipes ou revisar sequência de atividades")
    elif evm["SPI"] < 1.0:
        rec.append("⚠️ SPI abaixo de 1.0 — replanejamento de curto prazo recomendado")
    if lps["ppc_acumulado"] < 70:
        rec.append("🔴 PPC < 70% — eliminar restrições antes de comprometer atividades")
    if not rec:
        rec.append("✅ Projeto dentro dos parâmetros — manter monitoramento semanal")
    return rec

# ═══════════════════════════════════════════════════════════════════════════════
# ENGINE 12 — REPLANEJAMENTO AUTOMÁTICO
# ═══════════════════════════════════════════════════════════════════════════════
def replanejamento(trechos: list, realizado: list, cfg: dict = None) -> dict:
    """
    Compara planejado vs realizado e gera novo plano.
    realizado = [{"montante":"PV-01","jusante":"PV-02","concluido":True/False,"custo_real":X}]
    """
    c = {**CFG_OBRA_DEFAULT, **(cfg or {})}
    mapa_real = {f"{r['montante']}-{r['jusante']}": r for r in realizado}

    pendentes      = []
    total_ev       = 0
    total_ac       = 0
    total_pv       = 0

    for t in trechos:
        key  = f"{t.get('montante','')}-{t.get('jusante','')}"
        real = mapa_real.get(key, {})
        q    = calc_quant_trecho(t, c)
        pv_t = q["custos"]["total_com_bdi"]
        total_pv += pv_t

        if real.get("concluido"):
            total_ev += pv_t
            total_ac += float(real.get("custo_real", pv_t))
        else:
            pendentes.append(t)

    novo_cron  = calc_cronograma_sequencial(pendentes, c) if pendentes else {"trechos":[], "total_custo":0}
    bac        = sum(calc_quant_trecho(t,c)["custos"]["total_com_bdi"] for t in trechos)
    evm_atual  = calc_evm(bac, total_pv, total_ev, total_ac)

    # Ajustar produtividade baseado em SPI real
    spi = evm_atual["SPI"]
    if spi < 0.85 and pendentes:
        fator = 1 / spi if spi > 0 else 1.2
        c2    = {**c, "n_equipes_assentamento": math.ceil(c["n_equipes_assentamento"] * fator)}
        novo_cron = calc_cronograma_sequencial(pendentes, c2)
        log(f"Replanejamento: SPI={spi:.2f} → equipes ajustadas para {c2['n_equipes_assentamento']}", "WARN", "REPLAN")

    return {
        "ok": True,
        "evm_atual": evm_atual,
        "trechos_concluidos": len(trechos) - len(pendentes),
        "trechos_pendentes":  len(pendentes),
        "novo_cronograma":    novo_cron,
        "recomendacao":       _recomendacoes_pmbok(evm_atual, {"ppc_acumulado": 80}, 75),
    }

# ═══════════════════════════════════════════════════════════════════════════════
# EXPORT — EXCEL 10 ABAS
# ═══════════════════════════════════════════════════════════════════════════════
VERDE  = "1E6B3A"; BRANCO = "FFFFFF"; CINZA  = "F2F2F2"; AZUL   = "1F4E79"
AMARELO= "FFD700"; LARANJA= "FF6600"; VERMELHO="CC0000"

def _fill(hex_c): return PatternFill("solid", fgColor=hex_c)
def _font(bold=False, size=10, color=None):
    kw = {"bold": bold, "size": size}
    if color: kw["color"] = color
    return Font(**kw)
def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def gerar_excel(trechos_raw: list, cfg_in: dict = None) -> str:
    c    = {**CFG_OBRA_DEFAULT, **(cfg_in or {})}
    rows = [calc_quant_trecho(t, c) for t in trechos_raw]
    wb   = openpyxl.Workbook()
    wb.remove(wb.active)

    def cabecalho(ws, titulo, colunas):
        ws.sheet_view.showGridLines = False
        ws.merge_cells(f"A1:{get_column_letter(len(colunas))}1")
        ws["A1"] = titulo
        ws["A1"].font = _font(True, 14, BRANCO)
        ws["A1"].fill = _fill(VERDE)
        ws["A1"].alignment = _align()
        ws.row_dimensions[1].height = 30
        for j, col in enumerate(colunas, 1):
            cell = ws.cell(2, j, col)
            cell.font = _font(True, 10, BRANCO)
            cell.fill = _fill(AZUL)
            cell.alignment = _align()
        ws.row_dimensions[2].height = 20

    # ABA 1 — RESUMO
    ws = wb.create_sheet("RESUMO")
    cabecalho(ws, "ConstruPlan v1.0 — Resumo Executivo", ["Indicador", "Valor", "Unidade"])
    resumo = [
        ("Total de trechos",    len(rows),                                "un"),
        ("Extensão total",      round(sum(r["extensao"] for r in rows),2), "m"),
        ("Vol. escavação",      round(sum(r["vol_escavacao_tot"] for r in rows),2), "m³"),
        ("Vol. reaterro",       round(sum(r["vol_reaterro_comp"] for r in rows),2), "m³"),
        ("Área asfalto",        round(sum(r["area_asfalto_m2"] for r in rows),2), "m²"),
        ("Área passeio",        round(sum(r["area_passeio_m2"] for r in rows),2), "m²"),
        ("Custo s/ BDI",        round(sum(r["custos"]["total_sem_bdi"] for r in rows),2), "R$"),
        ("BDI",                 round(sum(r["custos"]["bdi"] for r in rows),2), "R$"),
        ("Custo c/ BDI",        round(sum(r["custos"]["total_com_bdi"] for r in rows),2), "R$"),
        ("Custo/metro",         round(sum(r["custos"]["total_com_bdi"] for r in rows)/max(sum(r["extensao"] for r in rows),1),2), "R$/m"),
    ]
    for i, (ind, val, un) in enumerate(resumo, 3):
        ws.cell(i, 1, ind)
        ws.cell(i, 2, val)
        ws.cell(i, 3, un)
        if i % 2 == 0: ws.cell(i,1).fill = ws.cell(i,2).fill = ws.cell(i,3).fill = _fill(CINZA)

    # ABA 2 — TRECHOS
    ws2 = wb.create_sheet("TRECHOS")
    cols2 = ["#","Montante","Jusante","Rua","Extensão(m)","DN(mm)","Material",
             "Prof.Med(m)","Decl(‰)","Vol.Esc(m³)","Reaterro(m³)","Custo c/BDI(R$)"]
    cabecalho(ws2, "ConstruPlan v1.0 — Trechos", cols2)
    for i, r in enumerate(rows, 3):
        data = [i-2, r["montante"], r["jusante"], r.get("nomeRua",""),
                r["extensao"], r["dn"], r["material"],
                r["prof_media_m"], round(r["declividade"]*1000,4),
                r["vol_escavacao_tot"], r["vol_reaterro_comp"],
                r["custos"]["total_com_bdi"]]
        for j, v in enumerate(data, 1):
            cell = ws2.cell(i, j, v)
            if i % 2 == 0: cell.fill = _fill(CINZA)

    # ABA 3 — HIDRÁULICA
    ws3 = wb.create_sheet("HIDRÁULICA")
    cols3 = ["Montante","Jusante","DN(mm)","Decl(‰)","Vel(m/s)","Lâmina(%)","T.Trat(Pa)","Q_pleno(L/s)","Status"]
    cabecalho(ws3, "ConstruPlan v1.0 — Verificação Hidráulica NBR 9649", cols3)
    for i, r in enumerate(rows, 3):
        h = r.get("hidraulica", {})
        ok = h.get("ok", False)
        data = [r["montante"], r["jusante"], r["dn"],
                round(r["declividade"]*1000,4),
                h.get("vel",0), h.get("yd_pct",0), h.get("tt",0), h.get("qP",0),
                "✅ OK" if ok else "⚠️ " + "; ".join(h.get("alertas",["Verificar"]))]
        for j, v in enumerate(data, 1):
            cell = ws3.cell(i, j, v)
            if not ok: cell.fill = _fill("FFEEEE")
            elif i % 2 == 0: cell.fill = _fill(CINZA)

    # ABA 4 — CRONOGRAMA
    ws4 = wb.create_sheet("CRONOGRAMA")
    cron = calc_cronograma_sequencial(trechos_raw, c)
    cols4 = ["Montante","Jusante","Extensão(m)","Duração(dias)","Início","Fim","Custo(R$)","Acumulado(R$)"]
    cabecalho(ws4, "ConstruPlan v1.0 — Cronograma Físico-Financeiro", cols4)
    for i, t in enumerate(cron["trechos"], 3):
        for j, v in enumerate([t["montante"],t["jusante"],t["extensao"],t["duracao_dias"],
                                t["dataInicio"],t["dataFim"],t["custo"],t["custo_acum"]], 1):
            ws4.cell(i, j, v)

    # ABA 5 — CUSTOS DETALHADOS
    ws5 = wb.create_sheet("CUSTOS")
    cols5 = ["Montante","Jusante","Escavação","Reaterro","Tubo","PV","Asfalto","Passeio","S/BDI","BDI","C/BDI"]
    cabecalho(ws5, "ConstruPlan v1.0 — Planilha de Custos", cols5)
    for i, r in enumerate(rows, 3):
        cu = r["custos"]
        for j, v in enumerate([r["montante"],r["jusante"],cu["escavacao"],cu["reaterro"],
                                cu["tubo"],cu["pv"],cu["asfalto"],cu["passeio"],
                                cu["total_sem_bdi"],cu["bdi"],cu["total_com_bdi"]], 1):
            ws5.cell(i, j, v)

    # ABA 6 — SINAPI
    ws6 = wb.create_sheet("SINAPI")
    cols6 = ["Código","Descrição","Unid","Preço Ref (R$)"]
    cabecalho(ws6, "SINAPI SP 2025 — Referência ConstruPlan", cols6)
    for i, s in enumerate(SINAPI_BASE, 3):
        for j, v in enumerate([s["codigo"],s["descricao"],s["unidade"],s["preco_ref"]], 1):
            ws6.cell(i, j, v)

    # ABA 7 — MAPA DE CALOR (tabela)
    ws7 = wb.create_sheet("MAPA_CALOR")
    hm  = calc_mapa_calor(trechos_raw, "profundidade")
    cols7 = ["Montante","Jusante","Rua","Extensão(m)","Profundidade(m)","Classificação"]
    cabecalho(ws7, "ConstruPlan v1.0 — Mapa de Calor: Profundidade", cols7)
    cores_mapa = {"0":"00CC00","1":"88CC00","2":"CCCC00","3":"CC6600","4":"CC0000"}
    for i, f in enumerate(hm.get("features",[]), 3):
        data = [f["montante"],f["jusante"],f.get("rua",""),f["extensao"],f["valor"],f["label"]]
        for j, v in enumerate(data, 1):
            cell = ws7.cell(i, j, v)
            cell.fill = _fill(cores_mapa.get(str(f.get("quintil",0)), "FFFFFF"))

    # ABA 8 — LPS
    ws8 = wb.create_sheet("LPS")
    cols8 = ["Semana","Planejadas","Concluídas","PPC(%)","Status"]
    cabecalho(ws8, "ConstruPlan v1.0 — Last Planner System (PPC)", cols8)
    ws8.cell(3, 1, "Sem dados de LPS — inserir via API /api/lps")

    # ABA 9 — EVM
    ws9 = wb.create_sheet("EVM")
    cabecalho(ws9, "ConstruPlan v1.0 — Earned Value Management", ["Indicador","Valor","Status"])
    bac_tot = cron["total_custo"]
    evm_ex  = calc_evm(bac_tot, bac_tot*0.5, bac_tot*0.4, bac_tot*0.45)
    evm_data = [
        ("BAC",  evm_ex["BAC"],  "Budget at Completion"),
        ("PV",   evm_ex["PV"],   "Planned Value"),
        ("EV",   evm_ex["EV"],   "Earned Value"),
        ("AC",   evm_ex["AC"],   "Actual Cost"),
        ("CV",   evm_ex["CV"],   "Cost Variance"),
        ("SV",   evm_ex["SV"],   "Schedule Variance"),
        ("CPI",  evm_ex["CPI"],  "Cost Performance Index"),
        ("SPI",  evm_ex["SPI"],  "Schedule Performance Index"),
        ("EAC",  evm_ex["EAC_CPI"], "Estimate at Completion"),
        ("VAC",  evm_ex["VAC"],  "Variance at Completion"),
    ]
    for i, (k, v, s) in enumerate(evm_data, 3):
        ws9.cell(i, 1, k); ws9.cell(i, 2, v); ws9.cell(i, 3, s)

    # ABA 10 — SUPERLOG
    ws10 = wb.create_sheet("SUPERLOG")
    cabecalho(ws10, "ConstruPlan v1.0 — SuperLog do Projeto", ["Timestamp","Nível","Módulo","Mensagem"])
    for i, entry in enumerate(_mensagens[-500:], 3):
        ws10.cell(i,1,entry.get("ts",""))
        ws10.cell(i,2,entry.get("nivel",""))
        ws10.cell(i,3,entry.get("modulo",""))
        ws10.cell(i,4,entry.get("msg",""))

    out = str(DATA_DIR / "output" / "construplan_export.xlsx")
    wb.save(out)
    log(f"Excel exportado: {out}", "INFO", "EXPORT")
    return out

# ═══════════════════════════════════════════════════════════════════════════════
# ROUTES — API REST
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/")
def root():
    return jsonify({
        "sistema": "ConstruPlan v1.0",
        "autor":   "NeryFelipe · 2026",
        "rotas":   [r.rule for r in app.url_map.iter_rules()],
        "trechos": len(_trechos),
        "projetos": list(_projetos.keys()),
    })

@app.route("/api/trechos", methods=["GET","POST","PUT","DELETE","OPTIONS"])
def api_trechos():
    global _trechos
    if request.method == "OPTIONS": return jsonify(ok=True)
    if request.method == "GET":
        return jsonify({"ok":True,"trechos":_trechos,"total":len(_trechos)})
    if request.method == "POST":
        data = request.get_json(force=True) or {}
        novos = data.get("trechos", [data])
        for t in novos:
            t_n = normalizar_trecho(t)
            # Upsert por ID
            idx = next((i for i,x in enumerate(_trechos) if x["id"]==t_n["id"]), None)
            if idx is not None: _trechos[idx] = t_n
            else: _trechos.append(t_n)
        _save_state()
        log(f"Trechos salvos: {len(novos)} (total: {len(_trechos)})", "INFO", "TRECHOS")
        return jsonify({"ok":True,"total":len(_trechos)})
    if request.method == "DELETE":
        _trechos.clear()
        _save_state()
        return jsonify({"ok":True,"msg":"Trechos removidos"})

@app.route("/api/importar", methods=["POST","OPTIONS"])
def api_importar():
    if request.method == "OPTIONS": return jsonify(ok=True)
    f = request.files.get("file")
    if not f: return jsonify(error="Nenhum arquivo enviado"), 400
    nome = f.filename.lower()
    raw  = f.read()
    log(f"Importando: {f.filename} ({len(raw)//1024}KB)", "INFO", "IMPORT")
    try:
        if nome.endswith(".csv"):
            text = raw.decode('utf-8', errors='replace')
            # Detectar tipo de CSV
            if "IFC_NAME" in text or "IFC_GLOBALID" in text:
                trechos = importar_csv_globalmapper(text)
                tipo = "CSV GlobalMapper/IFC"
            elif "codigo;descricao" in text or "custo_unitario" in text:
                itens = importar_csv_sabesp(text)
                return jsonify({"ok":True,"tipo":"CSV SABESP","itens":itens,"total":len(itens)})
            else:
                trechos = importar_csv_generico(text)
                tipo = "CSV Genérico"
        elif nome.endswith(".geojson") or nome.endswith(".json"):
            trechos = importar_geojson(raw.decode('utf-8', errors='replace'))
            tipo = "GeoJSON"
        elif nome.endswith(".dxf"):
            trechos = importar_dxf(raw.decode('utf-8', errors='replace'))
            tipo = "DXF"
        elif nome.endswith((".xlsx",".xls")):
            trechos = importar_xlsx(raw)
            tipo = "Excel"
        else:
            return jsonify(error=f"Formato não suportado: {nome.split('.')[-1]}"), 400

        # Salvar em memória
        for t in trechos:
            t_n = normalizar_trecho(t)
            _trechos.append(t_n)
        _save_state()
        log(f"Importado: {len(trechos)} trechos de {tipo}", "INFO", "IMPORT")
        return jsonify({"ok":True,"tipo":tipo,"trechos":trechos,"total":len(trechos)})
    except Exception as e:
        log(f"Erro importar {nome}: {e}", "ERROR", "IMPORT")
        return jsonify(error=str(e), tb=traceback.format_exc()), 500

@app.route("/api/hidraulica", methods=["POST","OPTIONS"])
def api_hidraulica():
    if request.method == "OPTIONS": return jsonify(ok=True)
    try:
        data = request.get_json(force=True) or {}
        trechos = data.get("trechos", _trechos)
        cfg     = data.get("cfg", {})
        if not trechos: return jsonify(error="Nenhum trecho"), 400
        resultados = []
        for t in trechos:
            t_n  = normalizar_trecho(t)
            dn_m = t_n["dn"] / 1000
            mat  = t_n["material"]
            decl = t_n["declividade"]
            n_   = get_manning(mat)
            q_ls = float(t_n.get("q_ls", data.get("q_ls", 0)))
            h    = hydro(dn_m, decl * 1000, n_, q_ls)
            resultados.append({**t_n, "hidraulica": h})
        log(f"Hidráulica: {len(resultados)} trechos", "INFO", "HIDRAULICA")
        return jsonify({"ok":True,"resultados":resultados})
    except Exception as e:
        return jsonify(error=str(e)), 500

@app.route("/api/quantitativos", methods=["POST","OPTIONS"])
def api_quantitativos():
    if request.method == "OPTIONS": return jsonify(ok=True)
    try:
        data    = request.get_json(force=True) or {}
        trechos = data.get("trechos", _trechos)
        cfg     = data.get("cfg", {})
        if not trechos: return jsonify(error="Nenhum trecho"), 400
        rows = [calc_quant_trecho(t, cfg) for t in trechos]
        totais = {
            "extensao_m":   round(sum(r["extensao"] for r in rows), 2),
            "vol_esc_tot":  round(sum(r["vol_escavacao_tot"] for r in rows), 2),
            "reaterro_m3":  round(sum(r["vol_reaterro_comp"] for r in rows), 2),
            "asfalto_m2":   round(sum(r["area_asfalto_m2"] for r in rows), 2),
            "passeio_m2":   round(sum(r["area_passeio_m2"] for r in rows), 2),
            "custo_total":  round(sum(r["custos"]["total_com_bdi"] for r in rows), 2),
        }
        log(f"Quantitativos: {len(rows)} trechos, custo total R$ {totais['custo_total']:,.2f}", "INFO", "QUANT")
        return jsonify({"ok":True,"trechos":rows,"totais":totais})
    except Exception as e:
        return jsonify(error=str(e), tb=traceback.format_exc()), 500

@app.route("/api/cronograma", methods=["POST","OPTIONS"])
def api_cronograma():
    if request.method == "OPTIONS": return jsonify(ok=True)
    try:
        data    = request.get_json(force=True) or {}
        trechos = data.get("trechos", _trechos)
        cfg     = data.get("cfg", {})
        modo    = data.get("modo", "sequencial")
        if not trechos: return jsonify(error="Nenhum trecho"), 400
        if modo == "cpm":
            tasks = data.get("tasks", [])
            if not tasks: return jsonify(error="CPM requer lista 'tasks'"), 400
            result = compute_cpm(tasks)
        else:
            result = calc_cronograma_sequencial(trechos, cfg)
        log(f"Cronograma ({modo}): {len(trechos)} trechos", "INFO", "CRON")
        return jsonify({"ok":True, **result})
    except Exception as e:
        return jsonify(error=str(e), tb=traceback.format_exc()), 500

@app.route("/api/evm", methods=["POST","OPTIONS"])
def api_evm():
    if request.method == "OPTIONS": return jsonify(ok=True)
    try:
        data = request.get_json(force=True) or {}
        bac  = float(data.get("BAC",0) or data.get("bac",0))
        pv   = float(data.get("PV",0)  or data.get("pv",0))
        ev   = float(data.get("EV",0)  or data.get("ev",0))
        ac   = float(data.get("AC",0)  or data.get("ac",0))
        if bac <= 0: return jsonify(error="BAC inválido"), 400
        result = calc_evm(bac, pv, ev, ac)
        log(f"EVM: CPI={result['CPI']:.2f} SPI={result['SPI']:.2f}", "INFO", "EVM")
        return jsonify(result)
    except Exception as e:
        return jsonify(error=str(e)), 500

@app.route("/api/mapa_calor", methods=["POST","OPTIONS"])
def api_mapa_calor():
    if request.method == "OPTIONS": return jsonify(ok=True)
    try:
        data      = request.get_json(force=True) or {}
        trechos   = data.get("trechos", _trechos)
        atributo  = data.get("atributo", "profundidade")
        if not trechos: return jsonify(error="Nenhum trecho"), 400
        result = calc_mapa_calor(trechos, atributo)
        log(f"Mapa calor '{atributo}': {len(trechos)} features", "INFO", "MAPA")
        return jsonify(result)
    except Exception as e:
        return jsonify(error=str(e)), 500

@app.route("/api/lps", methods=["POST","OPTIONS"])
def api_lps():
    if request.method == "OPTIONS": return jsonify(ok=True)
    try:
        data   = request.get_json(force=True) or {}
        semanas = data.get("semanas", [])
        if not semanas: return jsonify(error="Nenhuma semana informada"), 400
        result = calc_lps(semanas)
        log(f"LPS: PPC acumulado = {result['ppc_acumulado']}%", "INFO", "LPS")
        return jsonify(result)
    except Exception as e:
        return jsonify(error=str(e)), 500

@app.route("/api/planejamento5d", methods=["POST","OPTIONS"])
def api_planejamento5d():
    if request.method == "OPTIONS": return jsonify(ok=True)
    try:
        data    = request.get_json(force=True) or {}
        trechos = data.get("trechos", _trechos)
        cfg     = data.get("cfg", {})
        if not trechos: return jsonify(error="Nenhum trecho"), 400
        result = planejamento_5d(trechos, cfg)
        log(f"Planejamento 5D: BAC=R${result['resumo_financeiro']['BAC']:,.2f}", "INFO", "5D")
        return jsonify(result)
    except Exception as e:
        return jsonify(error=str(e), tb=traceback.format_exc()), 500

@app.route("/api/pmbok", methods=["POST","OPTIONS"])
def api_pmbok():
    if request.method == "OPTIONS": return jsonify(ok=True)
    try:
        data   = request.get_json(force=True) or {}
        result = calc_pmbok(data)
        log(f"PMBOK: Score={result['score_saude']}, {result['status']}", "INFO", "PMBOK")
        return jsonify(result)
    except Exception as e:
        return jsonify(error=str(e)), 500

@app.route("/api/replanejamento", methods=["POST","OPTIONS"])
def api_replanejamento():
    if request.method == "OPTIONS": return jsonify(ok=True)
    try:
        data      = request.get_json(force=True) or {}
        trechos   = data.get("trechos", _trechos)
        realizado = data.get("realizado", [])
        cfg       = data.get("cfg", {})
        if not trechos: return jsonify(error="Nenhum trecho"), 400
        result = replanejamento(trechos, realizado, cfg)
        log(f"Replanejamento: {result['trechos_concluidos']} concluídos / {result['trechos_pendentes']} pendentes", "INFO", "REPLAN")
        return jsonify(result)
    except Exception as e:
        return jsonify(error=str(e), tb=traceback.format_exc()), 500

@app.route("/api/orcamento", methods=["POST","OPTIONS"])
def api_orcamento():
    if request.method == "OPTIONS": return jsonify(ok=True)
    try:
        data    = request.get_json(force=True) or {}
        trechos = data.get("trechos", _trechos)
        cfg     = data.get("cfg", {})
        if not trechos: return jsonify(error="Nenhum trecho"), 400
        rows    = [calc_quant_trecho(t, cfg) for t in trechos]
        bac     = round(sum(r["custos"]["total_com_bdi"] for r in rows), 2)
        bdi_val = round(sum(r["custos"]["bdi"] for r in rows), 2)
        return jsonify({
            "ok": True,
            "BAC": bac,
            "BDI_total": bdi_val,
            "sinapi": SINAPI_BASE,
            "por_trecho": [{"montante":r["montante"],"jusante":r["jusante"],"custos":r["custos"]} for r in rows],
        })
    except Exception as e:
        return jsonify(error=str(e)), 500

@app.route("/api/export/excel", methods=["POST","OPTIONS"])
def api_export_excel():
    if request.method == "OPTIONS": return jsonify(ok=True)
    try:
        data    = request.get_json(force=True) or {}
        trechos = data.get("trechos", _trechos)
        cfg     = data.get("cfg", {})
        if not trechos: return jsonify(error="Nenhum trecho"), 400
        path = gerar_excel(trechos, cfg)
        return send_file(path, as_attachment=True, download_name="construplan_export.xlsx",
                        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        return jsonify(error=str(e), tb=traceback.format_exc()), 500

@app.route("/api/export/geojson", methods=["POST","OPTIONS"])
def api_export_geojson():
    if request.method == "OPTIONS": return jsonify(ok=True)
    try:
        data    = request.get_json(force=True) or {}
        trechos = data.get("trechos", _trechos)
        atrib   = data.get("atributo", "profundidade")
        hm      = calc_mapa_calor(trechos, atrib)
        features = [{
            "type": "Feature",
            "geometry": {"type": "Point", "coordinates": [0, 0]},  # sem coords reais
            "properties": f,
        } for f in hm.get("features", [])]
        gj = {"type":"FeatureCollection","features":features}
        buf = io.BytesIO(json.dumps(gj, ensure_ascii=False).encode('utf-8'))
        buf.seek(0)
        return send_file(buf, as_attachment=True, download_name=f"construplan_{atrib}.geojson",
                        mimetype="application/geo+json")
    except Exception as e:
        return jsonify(error=str(e)), 500

@app.route("/api/logs", methods=["GET"])
def api_logs():
    limit = int(request.args.get("limit", 200))
    return jsonify({"ok":True,"logs":_mensagens[-limit:],"total":len(_mensagens)})

@app.route("/api/superlog", methods=["GET","DELETE"])
def api_superlog():
    if request.method == "DELETE":
        _mensagens.clear()
        _save_superlog()
        return jsonify({"ok":True,"msg":"Superlog limpo"})
    return jsonify({
        "ok": True,
        "total": len(_mensagens),
        "arquivo": str(SUPERLOG_FILE),
        "logs": _mensagens[-int(request.args.get("limit",1000)):],
        "resumo": {
            "INFO":  sum(1 for m in _mensagens if m["nivel"]=="INFO"),
            "WARN":  sum(1 for m in _mensagens if m["nivel"]=="WARN"),
            "ERROR": sum(1 for m in _mensagens if m["nivel"]=="ERROR"),
        }
    })

@app.route("/api/config", methods=["GET","POST"])
def api_config():
    if request.method == "GET":
        return jsonify(CFG_OBRA_DEFAULT)
    data = request.get_json(force=True) or {}
    CFG_OBRA_DEFAULT.update(data)
    log(f"Config atualizada: {list(data.keys())}", "INFO", "CONFIG")
    return jsonify({"ok":True, "config": CFG_OBRA_DEFAULT})

@app.route("/api/sinapi", methods=["GET"])
def api_sinapi():
    return jsonify({"ok":True,"sinapi":SINAPI_BASE,"total":len(SINAPI_BASE)})

# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    log(f"ConstruPlan backend iniciando na porta {PORT}", "INFO", "BOOT")
    app.run(debug=True, port=PORT, host="0.0.0.0")
