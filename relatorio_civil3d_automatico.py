#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Relatorio automatico Civil 3D -> materiais/quantitativos.

Uso:
  python relatorio_civil3d_automatico.py "C:\\obra\\rede.xml" --nucleo "Sao Manoel"
  python relatorio_civil3d_automatico.py "C:\\obra\\rede.dwg" --saida "C:\\obra\\RELATORIOS"

Regra pratica:
  - XML LandXML do Civil 3D e a fonte preferida.
  - DWG/DXF entram como fallback usando os leitores existentes do motor.
"""

from __future__ import annotations

import argparse
import json
import math
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ler_landxml import ler_landxml

try:
    from ler_dxf_gdal import ler_dxf_gdal
except Exception:  # pragma: no cover - fallback opcional
    ler_dxf_gdal = None

try:
    from ler_dwg_universal import ler_dwg_universal
except Exception:  # pragma: no cover - fallback opcional
    ler_dwg_universal = None


CONTRATO = "11481051"

FATORES_PADRAO = {
    "reaterro_pct": 0.85,
    "pavimento_m2_por_m": 0.88,
    "areia_lastro_m3_por_m": 0.25,
    "areia_berco_m3_por_m": 0.24,
    "brita_m3_por_m": 0.22,
    "comprimento_barra_m": 6.0,
}


def _num(v, default=0.0):
    try:
        if v in (None, ""):
            return default
        x = float(v)
        return x if math.isfinite(x) else default
    except Exception:
        return default


def _safe_name(texto, default="RELATORIO"):
    texto = str(texto or default).strip()
    texto = re.sub(r"[^\w\-]+", "_", texto, flags=re.UNICODE).strip("_")
    return texto or default


def _prof_pv(pv):
    prof = _num(pv.get("prof"), 0)
    if prof > 0:
        return prof
    ct, cf = _num(pv.get("ct"), 0), _num(pv.get("cf"), 0)
    return abs(ct - cf) if ct and cf else 0


def carregar_rede(caminho):
    """Carrega rede Civil 3D a partir de XML, DXF ou DWG."""
    path = Path(caminho)
    if not path.exists():
        raise FileNotFoundError(path)

    suf = path.suffix.lower()
    if suf == ".xml":
        pvs, trechos, ruas, meta = ler_landxml(path)
        return pvs, trechos, ruas, meta

    if suf == ".dxf":
        if not ler_dxf_gdal:
            raise RuntimeError("ler_dxf_gdal indisponivel.")
        pvs, trechos, ruas, meta = ler_dxf_gdal(path)
        return pvs, trechos, ruas, meta

    if suf == ".dwg":
        xml_irmao = path.with_suffix(".xml")
        if xml_irmao.exists():
            pvs, trechos, ruas, meta = ler_landxml(xml_irmao)
            meta["arquivo_dwg"] = path.name
            meta["fonte_usada"] = "LandXML ao lado do DWG"
            return pvs, trechos, ruas, meta
        if not ler_dwg_universal:
            raise RuntimeError("ler_dwg_universal indisponivel e XML ao lado do DWG nao encontrado.")
        dados = ler_dwg_universal(path)
        if len(dados) == 4:
            return dados
        pvs, trechos, meta = dados
        return pvs, trechos, [], meta

    raise ValueError(f"Formato nao suportado: {path.suffix}")


def deduplicar_trechos_geom(trechos, pvs):
    """Remove tubos repetidos do Civil 3D usando geometria, DN e extensao."""
    saida = []
    vistos = set()
    for t in trechos:
        p0 = pvs.get(t.get("pv_ini"), {})
        p1 = pvs.get(t.get("pv_fim"), {})
        x0, y0 = _num(p0.get("x"), None), _num(p0.get("y"), None)
        x1, y1 = _num(p1.get("x"), None), _num(p1.get("y"), None)
        dn = int(_num(t.get("dn_mm"), 0) or 0)
        ext = _num(t.get("ext_m"), 0)
        if x0 is not None and y0 is not None and x1 is not None and y1 is not None:
            if ext <= 0:
                ext = math.hypot(x1 - x0, y1 - y0)
            par = tuple(sorted([(round(x0, 1), round(y0, 1)), (round(x1, 1), round(y1, 1))]))
            key = (par, round(ext, 1), dn)
        else:
            par = tuple(sorted([str(t.get("pv_ini")), str(t.get("pv_fim"))]))
            key = (par, round(ext, 1), dn)
        if key in vistos:
            continue
        vistos.add(key)
        saida.append(t)
    return saida, len(trechos) - len(saida)


def enriquecer_trecho(trecho, pvs):
    t = dict(trecho)
    p0 = pvs.get(t.get("pv_ini"), {})
    p1 = pvs.get(t.get("pv_fim"), {})
    t["ct_ini"] = t.get("ct_ini") or p0.get("ct")
    t["cf_ini"] = t.get("cf_ini") or p0.get("cf")
    t["ct_fim"] = t.get("ct_fim") or p1.get("ct")
    t["cf_fim"] = t.get("cf_fim") or p1.get("cf")
    t["prof_ini"] = t.get("prof_ini") or _prof_pv(p0)
    t["prof_fim"] = t.get("prof_fim") or _prof_pv(p1)
    return t


def calcular_quantitativo(trecho, pvs, fatores=None):
    """Calcula material/volume por trecho PV a PV."""
    fatores = fatores or FATORES_PADRAO
    t = enriquecer_trecho(trecho, pvs)

    dn_mm = int(_num(t.get("dn_mm"), 200) or 200)
    ext_m = _num(t.get("ext_m"), 0)
    profs = [p for p in (_num(t.get("prof_ini"), 0), _num(t.get("prof_fim"), 0)) if p > 0]
    prof_media = sum(profs) / len(profs) if profs else 1.5
    largura_vala = max(0.60, dn_mm / 1000.0 + 0.50)

    escavacao = ext_m * largura_vala * prof_media
    areia = ext_m * (fatores["areia_lastro_m3_por_m"] + fatores["areia_berco_m3_por_m"])
    brita = ext_m * fatores["brita_m3_por_m"]
    pavimento = ext_m * fatores["pavimento_m2_por_m"]
    barras = math.ceil(ext_m / fatores["comprimento_barra_m"]) if ext_m > 0 else 0

    q = {
        "dn_mm": dn_mm,
        "ext_m": round(ext_m, 2),
        "prof_media_m": round(prof_media, 3),
        "largura_vala_m": round(largura_vala, 3),
        "escavacao_m3": round(escavacao, 3),
        "reaterro_m3": round(escavacao * fatores["reaterro_pct"], 3),
        "areia_m3": round(areia, 3),
        "brita_m3": round(brita, 3),
        "pavimento_m2": round(pavimento, 3),
        "tubos_barras": barras,
        "luvas_pc": max(barras - 1, 0),
        "aneis_pc": barras + 1 if barras else 0,
        "pasta_kg": round(barras * 0.04, 3),
    }
    t.update(q)
    return t


def montar_relatorio(caminho, nucleo=None, fatores=None):
    pvs, trechos, _ruas, meta = carregar_rede(caminho)
    bruto = len(trechos)
    trechos, duplicados = deduplicar_trechos_geom(trechos, pvs)
    meta["n_trechos_bruto"] = bruto
    meta["n_trechos"] = len(trechos)
    meta["duplicados_geom_ignorados"] = duplicados
    nucleo = nucleo or meta.get("nucleo") or Path(caminho).stem
    fatores = fatores or FATORES_PADRAO

    linhas = [calcular_quantitativo(t, pvs, fatores) for t in trechos]
    dns = Counter()
    por_dn = defaultdict(lambda: defaultdict(float))
    por_rua = defaultdict(lambda: defaultdict(float))
    materiais = defaultdict(lambda: {"descricao": "", "unidade": "", "grupo": "", "quantidade": 0.0})

    def add_mat(desc, un, qtd, grupo):
        key = (grupo, desc, un)
        materiais[key]["grupo"] = grupo
        materiais[key]["descricao"] = desc
        materiais[key]["unidade"] = un
        materiais[key]["quantidade"] += qtd

    for t in linhas:
        dn = t["dn_mm"]
        rua = t.get("rua") or "Sem Rua"
        dns[dn] += 1

        for bucket in (por_dn[dn], por_rua[rua]):
            bucket["trechos"] += 1
            bucket["ext_m"] += t["ext_m"]
            bucket["escavacao_m3"] += t["escavacao_m3"]
            bucket["reaterro_m3"] += t["reaterro_m3"]
            bucket["areia_m3"] += t["areia_m3"]
            bucket["brita_m3"] += t["brita_m3"]
            bucket["pavimento_m2"] += t["pavimento_m2"]
            bucket["tubos_barras"] += t["tubos_barras"]

        add_mat(f"Tubo PVC DN{dn}mm", "barra", t["tubos_barras"], "Tubos")
        add_mat(f"Luva correr PVC DN{dn}mm", "pc", t["luvas_pc"], "Conexoes")
        add_mat(f"Anel borracha DN{dn}mm", "pc", t["aneis_pc"], "Conexoes")
        add_mat("Pasta lubrificante", "kg", t["pasta_kg"], "Insumos")
        add_mat("Escavacao de vala", "m3", t["escavacao_m3"], "Movimento de terra")
        add_mat("Reaterro compactado", "m3", t["reaterro_m3"], "Movimento de terra")
        add_mat("Areia lastro/envoltoria", "m3", t["areia_m3"], "Insumos")
        add_mat("Brita base/dreno", "m3", t["brita_m3"], "Insumos")
        add_mat("Recomposicao de pavimento", "m2", t["pavimento_m2"], "Pavimento")

    pvs_usados = {t.get("pv_ini") for t in linhas} | {t.get("pv_fim") for t in linhas}
    pvs_validos = [pv for pv in pvs_usados if pv and not str(pv).upper().startswith(("STARTNULL", "ENDNULL"))]
    add_mat("PV/PI concreto", "pc", len(pvs_validos), "Estruturas")

    totais = defaultdict(float)
    for t in linhas:
        for k in ("ext_m", "escavacao_m3", "reaterro_m3", "areia_m3", "brita_m3", "pavimento_m2", "tubos_barras"):
            totais[k] += t[k]

    return {
        "meta": {
            **meta,
            "nucleo": nucleo,
            "arquivo": str(Path(caminho)),
            "gerado_em": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "contrato": CONTRATO,
        },
        "pvs": pvs,
        "trechos": linhas,
        "materiais": sorted(materiais.values(), key=lambda x: (x["grupo"], x["descricao"])),
        "por_dn": dict(sorted(por_dn.items())),
        "por_rua": dict(sorted(por_rua.items())),
        "totais": {k: round(v, 3) for k, v in totais.items()},
        "dns": dict(dns),
        "fatores": fatores,
    }


def _style_table(ws, header_row, max_col):
    fill = PatternFill("solid", fgColor="123B5D")
    font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style="thin", color="D0D7DE"),
        right=Side(style="thin", color="D0D7DE"),
        top=Side(style="thin", color="D0D7DE"),
        bottom=Side(style="thin", color="D0D7DE"),
    )
    for c in range(1, max_col + 1):
        cell = ws.cell(header_row, c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = ws.cell(header_row + 1, 1)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(max_col)}{ws.max_row}"
    for col in range(1, max_col + 1):
        width = max(len(str(ws.cell(r, col).value or "")) for r in range(1, ws.max_row + 1))
        ws.column_dimensions[get_column_letter(col)].width = min(max(width + 2, 10), 42)


def _add_rows(ws, headers, rows, start=1):
    for c, h in enumerate(headers, 1):
        ws.cell(start, c, h)
    for r, row in enumerate(rows, start + 1):
        for c, h in enumerate(headers, 1):
            ws.cell(r, c, row.get(h, ""))
    _style_table(ws, start, len(headers))


def gerar_xlsx(relatorio, out_path):
    wb = Workbook()
    meta = relatorio["meta"]
    totais = relatorio["totais"]

    ws = wb.active
    ws.title = "Resumo"
    ws["A1"] = "RELATORIO CIVIL 3D - MATERIAIS E QUANTITATIVOS"
    ws["A1"].font = Font(size=14, bold=True, color="123B5D")
    ws["A2"] = f"Contrato {meta.get('contrato')} | {meta.get('nucleo')} | {meta.get('gerado_em')}"
    resumo = [
        ("Arquivo fonte", meta.get("arquivo")),
        ("Motor", meta.get("motor", meta.get("fonte_usada", "Civil 3D/LandXML"))),
        ("PVs lidos", meta.get("n_pvs", len(relatorio["pvs"]))),
        ("Trechos", len(relatorio["trechos"])),
        ("Trechos brutos Civil 3D", meta.get("n_trechos_bruto", len(relatorio["trechos"]))),
        ("Duplicados ignorados", meta.get("duplicados_geom_ignorados", 0)),
        ("Extensao total (m)", round(totais.get("ext_m", 0), 2)),
        ("Escavacao (m3)", round(totais.get("escavacao_m3", 0), 2)),
        ("Reaterro (m3)", round(totais.get("reaterro_m3", 0), 2)),
        ("Areia (m3)", round(totais.get("areia_m3", 0), 2)),
        ("Brita (m3)", round(totais.get("brita_m3", 0), 2)),
        ("Pavimento (m2)", round(totais.get("pavimento_m2", 0), 2)),
        ("Tubos (barras)", int(totais.get("tubos_barras", 0))),
    ]
    for i, (k, v) in enumerate(resumo, 4):
        ws.cell(i, 1, k)
        ws.cell(i, 2, v)
    _style_table(ws, 3, 2)

    ws = wb.create_sheet("Por Trecho")
    heads = [
        "Rua", "PV Inicio", "PV Fim", "DN", "Extensao (m)", "Material",
        "CT Ini", "CF Ini", "CT Fim", "CF Fim", "Prof Media (m)",
        "Larg Vala (m)", "Decl m/m", "Escavacao (m3)", "Reaterro (m3)",
        "Areia (m3)", "Brita (m3)", "Pavimento (m2)", "Tubos (barras)",
    ]
    rows = []
    for t in relatorio["trechos"]:
        rows.append({
            "Rua": t.get("rua", "Sem Rua"),
            "PV Inicio": t.get("pv_ini"),
            "PV Fim": t.get("pv_fim"),
            "DN": t.get("dn_mm"),
            "Extensao (m)": t.get("ext_m"),
            "Material": t.get("material", "PVC"),
            "CT Ini": t.get("ct_ini"),
            "CF Ini": t.get("cf_ini"),
            "CT Fim": t.get("ct_fim"),
            "CF Fim": t.get("cf_fim"),
            "Prof Media (m)": t.get("prof_media_m"),
            "Larg Vala (m)": t.get("largura_vala_m"),
            "Decl m/m": t.get("decl_mm"),
            "Escavacao (m3)": t.get("escavacao_m3"),
            "Reaterro (m3)": t.get("reaterro_m3"),
            "Areia (m3)": t.get("areia_m3"),
            "Brita (m3)": t.get("brita_m3"),
            "Pavimento (m2)": t.get("pavimento_m2"),
            "Tubos (barras)": t.get("tubos_barras"),
        })
    _add_rows(ws, heads, rows)

    ws = wb.create_sheet("Materiais")
    _add_rows(ws, ["Grupo", "Descricao", "Unidade", "Quantidade"], [
        {
            "Grupo": m["grupo"],
            "Descricao": m["descricao"],
            "Unidade": m["unidade"],
            "Quantidade": round(m["quantidade"], 3),
        }
        for m in relatorio["materiais"]
    ])

    bucket_heads = ["Chave", "Trechos", "Extensao (m)", "Escavacao (m3)", "Reaterro (m3)", "Areia (m3)", "Brita (m3)", "Pavimento (m2)", "Tubos (barras)"]
    for sheet_name, data in (("Por DN", relatorio["por_dn"]), ("Por Rua", relatorio["por_rua"])):
        ws = wb.create_sheet(sheet_name)
        bucket_rows = []
        for chave, vals in data.items():
            bucket_rows.append({
                "Chave": chave,
                "Trechos": int(vals.get("trechos", 0)),
                "Extensao (m)": round(vals.get("ext_m", 0), 2),
                "Escavacao (m3)": round(vals.get("escavacao_m3", 0), 3),
                "Reaterro (m3)": round(vals.get("reaterro_m3", 0), 3),
                "Areia (m3)": round(vals.get("areia_m3", 0), 3),
                "Brita (m3)": round(vals.get("brita_m3", 0), 3),
                "Pavimento (m2)": round(vals.get("pavimento_m2", 0), 3),
                "Tubos (barras)": int(vals.get("tubos_barras", 0)),
            })
        _add_rows(ws, bucket_heads, bucket_rows)

    ws = wb.create_sheet("Parametros")
    _add_rows(ws, ["Parametro", "Valor"], [{"Parametro": k, "Valor": v} for k, v in relatorio["fatores"].items()])

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def gerar_relatorio(caminho, saida=None, nucleo=None, json_saida=True):
    relatorio = montar_relatorio(caminho, nucleo=nucleo)
    base = Path(caminho)
    out_dir = Path(saida) if saida else base.parent / "RELATORIOS_CIVIL3D"
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    nome = f"RELATORIO_CIVIL3D_{_safe_name(nucleo or base.stem)}_{stamp}"
    xlsx = gerar_xlsx(relatorio, out_dir / f"{nome}.xlsx")
    json_path = None
    if json_saida:
        json_path = out_dir / f"{nome}.json"
        json_path.write_text(json.dumps(relatorio, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return {"xlsx": str(xlsx), "json": str(json_path) if json_path else None, "relatorio": relatorio}


def main():
    parser = argparse.ArgumentParser(description="Gera relatorio automatico Civil 3D de materiais/quantitativos.")
    parser.add_argument("arquivo", help="LandXML, DXF ou DWG do Civil 3D")
    parser.add_argument("--saida", help="Pasta de saida. Padrao: RELATORIOS_CIVIL3D ao lado do arquivo")
    parser.add_argument("--nucleo", help="Nome do nucleo/obra")
    parser.add_argument("--sem-json", action="store_true", help="Gera apenas XLSX")
    args = parser.parse_args()

    result = gerar_relatorio(args.arquivo, saida=args.saida, nucleo=args.nucleo, json_saida=not args.sem_json)
    rel = result["relatorio"]
    print("=" * 72)
    print("RELATORIO CIVIL 3D GERADO")
    print(f"Trechos: {len(rel['trechos'])}")
    print(f"Extensao: {rel['totais'].get('ext_m', 0):.2f} m")
    print(f"XLSX: {result['xlsx']}")
    if result["json"]:
        print(f"JSON: {result['json']}")
    print("=" * 72)


if __name__ == "__main__":
    main()
