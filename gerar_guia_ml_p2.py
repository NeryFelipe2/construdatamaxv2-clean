#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""PARTE 2 — abas 05-15 do GUIA_ML_1000_LIGACOES.xlsx"""
import sys, io, os, datetime, warnings
warnings.filterwarnings("ignore")
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.series import SeriesLabel

BASE = os.path.dirname(os.path.abspath(__file__))
OUT_FILE = os.path.join(BASE,"TRECHOS_NS_COMPLETOS","GUIA_ML","GUIA_ML_1000_LIGACOES.xlsx")

AZUL_ESC="1F4E79";AZUL_MED="2E75B6";AZUL_CLA="D6E4F0"
VERDE_ESC="1E6B3C";VERDE_CLA="E2EFDA"
VERM_ESC="C00000";VERM_CLA="FCE4D6"
AMAR_ESC="7F6000";AMAR_CLA="FFF2CC"
CINZA="F2F2F2";BRANCO="FFFFFF";LARANJA="FF6600"

def pf(c): return PatternFill("solid", fgColor=c)
def fn(b=False,c="000000",s=9): return Font(bold=b,color=c,size=s,name="Calibri")
def al(h="center",v="center",wrap=False):
    return Alignment(horizontal=h,vertical=v,wrap_text=wrap)
THIN=Border(left=Side(style="thin"),right=Side(style="thin"),
            top=Side(style="thin"),bottom=Side(style="thin"))
def cw(ws,col,w): ws.column_dimensions[get_column_letter(col)].width=w
def rh(ws,row,h): ws.row_dimensions[row].height=h
def mc(ws,r1,c1,r2,c2):
    ws.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)
def wr(ws,row,col,val,fill=None,font=None,aln=None,border=None,nfmt=None):
    c=ws.cell(row,col,val)
    if fill:  c.fill=fill
    if font:  c.font=font
    if aln:   c.alignment=aln
    if border:c.border=border
    if nfmt:  c.number_format=nfmt
    return c
def hdr(ws,row,cols_vals,widths=None,fill_col=AZUL_MED):
    for ci,(h,w) in enumerate(zip(cols_vals,widths or [12]*len(cols_vals)),1):
        wr(ws,row,ci,h,pf(fill_col),fn(True,"FFFFFF",9),al("center",wrap=True),THIN)
        if widths: cw(ws,ci,w)
    rh(ws,row,30)
def linha(ws,row,vals,fill_=None,font_=None,aln_=None,border_=THIN):
    f=font_ or fn(False,"000000",9)
    a=aln_  or al("center")
    for ci,v in enumerate(vals,1):
        wr(ws,row,ci,v,fill_,f,a,border_)
def titulo_aba(ws,t1,t2,ncols=16):
    mc(ws,1,1,1,ncols)
    wr(ws,1,1,t1,pf(AZUL_ESC),fn(True,"FFFFFF",14),al("center",wrap=True))
    rh(ws,1,28)
    mc(ws,2,1,2,ncols)
    wr(ws,2,1,t2,pf(AZUL_MED),fn(False,"FFFFFF",10),al("center",wrap=True))
    rh(ws,2,18)
    ws.freeze_panes="A3"
    return 3

print("Carregando workbook existente...")
wb = openpyxl.load_workbook(OUT_FILE)
NCOLS = 16

# ╔══════════════════════════════════════╗
# ║  ABA 05 — CRONOGRAMA MESTRE          ║
# ╚══════════════════════════════════════╝
ws5 = wb.create_sheet("05_CRONOGRAMA_MESTRE")
rn = titulo_aba(ws5,"05 — CRONOGRAMA MESTRE: Abr-Dez 2026",
                "Gantt visual com todos os processos | █=em execução ▲=mobilização ✓=concluído",22)

# meses abr-dez
MESES = ["ABR","MAI","JUN","JUL","AGO","SET","OUT","NOV","DEZ"]
# semanas por mês (4)
SEMS  = [f"S{i}" for i in range(1,5)]

# header duplo
wr(ws5,rn,1,"ATIVIDADE/GRUPO",pf(AZUL_ESC),fn(True,"FFFFFF",9),al("left"))
wr(ws5,rn,2,"RESP",pf(AZUL_ESC),fn(True,"FFFFFF",9),al("center"))
wr(ws5,rn,3,"DIFIC",pf(AZUL_ESC),fn(True,"FFFFFF",9),al("center"))
cw(ws5,1,34); cw(ws5,2,12); cw(ws5,3,7)
col_ini = 4
for mi, mes in enumerate(MESES):
    for si, sem in enumerate(SEMS):
        col = col_ini + mi*4 + si
        wr(ws5,rn,col,f"{mes}\n{sem}",pf(AZUL_MED),fn(True,"FFFFFF",8),
           al("center",wrap=True),THIN)
        cw(ws5,col,5)
rh(ws5,rn,30); rn+=1

GANTT_ROWS = [
    # (atividade, resp, dific, [tuplas (mes_ini_0, sem_ini_0, duracao_sems, cor)])
    # mes_ini: 0=ABR,1=MAI,...  sem_ini: 0-3
    ("═══ NÚCLEOS EM EXECUÇÃO ═══","","",None),
    ("Morro do Tetéu","FCN",4,[(0,0,16,"2E75B6")]),
    ("Pantanal Baixo","FCN",4,[(0,0,16,"2E75B6")]),
    ("Vila dos Criadores","FCN",3,[(0,0,16,"2E75B6")]),
    ("São Manoel","FCN",3,[(0,0,6,"2E75B6")]),
    ("João Carlos","FCN",3,[(0,0,2,"2E75B6")]),
    ("Vila Israel","FCN",4,[(0,0,6,"2E75B6")]),
    ("","","",None),
    ("═══ NOVOS GRUPOS ═══","","",None),
    ("R3 Cruzeiro/Prog II  ▲Mob 07/Abr","FCN",3,[(0,1,1,"FFC000"),(0,2,14,"70AD47")]),
    ("R8 Noroeste/Gilda  ▲Mob 07/Abr","FCN",3,[(0,1,1,"FFC000"),(0,2,14,"70AD47")]),
    ("R7 Piratininga/NW  ▲Mob 22/Abr","FCN",3,[(0,3,1,"FFC000"),(1,0,13,"70AD47")]),
    ("R4 Cruzeiro/Penha  ▲Mob 22/Abr","FCN",3,[(0,3,1,"FFC000"),(1,0,13,"70AD47")]),
    ("R2 Cruzeiro/Prog I  ▲Mob 12/Mai","FCN",2,[(1,1,1,"FFC000"),(1,2,12,"92D050")]),
    ("R5 Cruzeiro/Bento  ▲Mob 12/Mai","FCN",2,[(1,1,1,"FFC000"),(1,2,12,"92D050")]),
    ("R1 Zona Leste  ▲Mob 01/Jun","FCN",3,[(2,0,1,"FFC000"),(2,1,11,"70AD47")]),
    ("R6 Monte Serrat  ▲Mob 01/Jun","FCN",4,[(2,0,1,"FFC000"),(2,1,11,"2E75B6")]),
    ("","","",None),
    ("═══ PROCESSOS DE APOIO (paralelos) ═══","","",None),
    ("Cadastro Social — R3+R8","Assist.Soc.","-",[(0,0,4,"FFD966")]),
    ("Cadastro Social — R7+R4","Assist.Soc.","-",[(0,2,4,"FFD966")]),
    ("Cadastro Social — R2+R5","Assist.Soc.","-",[(1,0,4,"FFD966")]),
    ("Cadastro Social — R1+R6","Assist.Soc.","-",[(2,0,4,"FFD966")]),
    ("Cadastro Comercial — todos","SABESP+FCN","-",[(0,0,18,"FFD966")]),
    ("Cartografia/As-built — contínuo","Topógrafo","-",[(0,0,36,"BDD7EE")]),
    ("","","",None),
    ("═══ LAVAGEM DE REDE (agendar 2d antes) ═══","","",None),
    ("Lavagem R3+R8 (após rede pronta)","SABESP","-",[(1,0,1,"FF0000")]),
    ("Lavagem R7+R4","SABESP","-",[(1,2,1,"FF0000")]),
    ("Lavagem R2+R5","SABESP","-",[(2,0,1,"FF0000")]),
    ("Lavagem R1+R6","SABESP","-",[(2,2,1,"FF0000")]),
    ("","","",None),
    ("═══ MEDIÇÃO E FATURAMENTO ═══","","",None),
    ("Boletim Medição Mensal (dia 20)","FCN","-",[(0,3,1,"C9C9C9"),(1,3,1,"C9C9C9"),
                                                   (2,3,1,"C9C9C9"),(3,3,1,"C9C9C9"),
                                                   (4,3,1,"C9C9C9"),(5,3,1,"C9C9C9"),
                                                   (6,3,1,"C9C9C9"),(7,3,1,"C9C9C9"),
                                                   (8,3,1,"C9C9C9")]),
]

total_cols = len(MESES)*4

for g_row in GANTT_ROWS:
    nome, resp, dific, bars = g_row
    is_header = nome.startswith("═══")
    is_empty  = nome==""

    if is_empty:
        rh(ws5,rn,6); rn+=1; continue

    fill_row = pf(AZUL_CLA) if is_header else None
    font_row = fn(True,"1F4E79",9) if is_header else fn(False,"000000",8)

    mc(ws5,rn,1,rn,1)
    wr(ws5,rn,1,nome,fill_row,font_row,al("left"))
    wr(ws5,rn,2,resp,fill_row,fn(False,"666666",8),al("center"))
    wr(ws5,rn,3,str(dific) if dific else "",fill_row,fn(False,"666666",8),al("center"))

    # pintar colunas gantt
    for col in range(col_ini, col_ini+total_cols):
        wr(ws5,rn,col,"",pf(CINZA) if not is_header else fill_row,None,None,THIN)

    if bars:
        for (m_ini, s_ini, dur, cor) in bars:
            for step in range(dur):
                idx = m_ini*4 + s_ini + step
                if idx < total_cols:
                    col = col_ini + idx
                    wr(ws5,rn,col,"█",pf(cor),fn(False,cor,8),al("center"),THIN)

    rh(ws5,rn,14); rn+=1

# legenda
rn+=1
mc(ws5,rn,1,rn,4)
wr(ws5,rn,1,"LEGENDA:",pf(AZUL_ESC),fn(True,"FFFFFF",9),al("left"))
legenda = [
    ("2E75B6","█ DIFIC 3-4 (azul)"),("70AD47","█ DIFIC 2-3 novo (verde)"),
    ("92D050","█ DIFIC 1-2 (verde claro)"),("FFC000","▲ Mobilização"),
    ("FFD966","Processos apoio"),("FF0000","Lavagem SABESP"),("C9C9C9","Medição"),
]
for i,(cor,desc) in enumerate(legenda):
    wr(ws5,rn,5+i*2,"■",pf(cor),fn(True,cor,9),al("center"))
    wr(ws5,rn,6+i*2,desc,None,fn(False,"000000",8),al("left"))

print("  05_CRONOGRAMA_MESTRE ok")

# ╔══════════════════════════════════════╗
# ║  ABA 06 — CHECKLIST NUCLEO           ║
# ╚══════════════════════════════════════╝
ws6 = wb.create_sheet("06_CHECKLIST_NUCLEO")
rn = titulo_aba(ws6,"06 — CHECKLIST: Abertura de Novo Núcleo",
                "20+ passos obrigatórios | Use como lista de verificação antes de iniciar qualquer grupo",12)

mc(ws6,rn,1,rn,12)
wr(ws6,rn,1,"► IMPRIMIR ESTE CHECKLIST e assinar cada etapa antes de mobilizar equipes",
   pf(VERM_CLA),fn(True,VERM_ESC,10),al("left")); rn+=1

hdr(ws6,rn,["#","Fase","Atividade","Responsável","Prazo\n(dias antes)","Status","Observação/Doc"],
    [5,14,42,16,12,10,40]); rn+=1

CHECKLIST = [
    # (num, fase, atividade, resp, prazo_antes, doc_exigido)
    ("01","PROJETO",    "Projeto executivo aprovado pela SABESP",         "FCN",  "30 dias", "Plantas aprovadas c/ carimbo SABESP"),
    ("02","PROJETO",    "Levantamento topográfico/planta baixa executado", "Topógrafo", "25 dias", "Arquivo DWG/PDF planta"),
    ("03","LICENÇAS",   "ART/RRT assinada e registrada no CREA/CAU",     "Eng. Resp.", "20 dias", "Protocolo ART digital"),
    ("04","LICENÇAS",   "Licença de sinalização (Prefeitura Santos)",      "Adm. FCN",  "15 dias", "Alvará sinalização"),
    ("05","LICENÇAS",   "Comunicação à SABESP (Nota de Início)",          "FCN",  "10 dias", "E-mail protocolo"),
    ("06","AMBIENTAL",  "Verificar necessidade licença ambiental (CETESB)","Eng. Amb.", "20 dias", "Consulta prévia CETESB"),
    ("07","MATERIAL",   "Pedido de compra de tubos, conexões, PVs",       "Compras",   "15 dias", "Ordem de compra emitida"),
    ("08","MATERIAL",   "Entrega confirmada: tubos no canteiro",          "Almoxarifado","5 dias", "Nota fiscal + conferência"),
    ("09","MATERIAL",   "EPI completo para todas as equipes",             "Segurança", "5 dias",  "Checklist EPI assinado"),
    ("10","SOCIAL",     "Cadastro Social iniciado (pesquisa domiciliar)", "Assist.Soc.","15 dias", "Lista de imóveis cadastrados"),
    ("11","SOCIAL",     "Reunião informativa com moradores",              "FCN+SABESP","10 dias", "Ata de reunião"),
    ("12","COMERCIAL",  "Cadastro Comercial SABESP: confirmar dados",     "SABESP+FCN","10 dias", "Planilha CC atualizada"),
    ("13","EQUIPES",    "Equipes contratadas/designadas com CTPS",       "RH",         "10 dias", "Lista de funcionários"),
    ("14","EQUIPES",    "Exames admissionais realizados",                 "RH",         "7 dias",  "ASO (atestado saúde ocup.)"),
    ("15","EQUIPES",    "Treinamento NR-18 (construção civil)",           "SESMT",     "7 dias",   "Certificados NR-18"),
    ("16","EQUIPES",    "Treinamento NR-35 (trabalho em altura, >1,8m)",  "SESMT",     "5 dias",  "Certificados NR-35"),
    ("17","CANTEIRO",   "Pátio montado: almoxarifado + vestiário + WC",  "Mestre",    "5 dias",   "Foto pátio + endereço"),
    ("18","CANTEIRO",   "Placa de obra instalada (padrão SABESP)",       "FCN",       "3 dias",   "Foto placa"),
    ("19","TOPOGRAFIA", "Locação dos eixos das valas (piquetes)",         "Topógrafo", "1 dia",   "Caderneta topografia"),
    ("20","OPERAÇÃO",   "Sinalização de obra instalada antes da 1ª vala", "Sinaliz.",  "0 dias",  "Foto sinalização"),
    ("21","LAVAGEM",    "Lavagem SABESP AGENDADA (mín 2 dias antes)",     "FCN+SABESP","Durante", "Protocolo agendamento"),
    ("22","MEDIÇÃO",    "Planilha de medição diária configurada",         "FCN",       "1 dia",   "Execução_Geral.xlsx núcleo"),
    ("23","FISCAL",     "Fiscal SABESP comunicado sobre início",          "FCN",       "3 dias",  "E-mail fiscal"),
    ("24","FINANCEIRO", "BM (Boletim Medição) anterior quitado",          "Financeiro","Antes",   "Comprovante pagamento BM"),
]

for (num,fase,atv,resp,prazo,doc) in CHECKLIST:
    fase_fills = {"PROJETO":pf(AZUL_CLA),"LICENÇAS":pf(VERM_CLA),"AMBIENTAL":pf(AMAR_CLA),
                  "MATERIAL":pf(VERDE_CLA),"SOCIAL":pf("FFD966"),"COMERCIAL":pf("FFD966"),
                  "EQUIPES":pf(AZUL_CLA),"CANTEIRO":pf(CINZA),"TOPOGRAFIA":pf(CINZA),
                  "OPERAÇÃO":pf(VERDE_CLA),"LAVAGEM":pf(VERM_CLA),"MEDIÇÃO":pf(CINZA),
                  "FISCAL":pf(AZUL_CLA),"FINANCEIRO":pf(AMAR_CLA)}
    fill_ = fase_fills.get(fase, None)
    for ci,v in enumerate([num,fase,atv,resp,prazo,"☐",doc],1):
        f = fn(True,"000000",9) if ci in(1,2) else fn(False,"000000",8)
        wr(ws6,rn,ci,v,fill_,f,al("left" if ci>2 else "center",wrap=True),THIN)
    rh(ws6,rn,18); rn+=1
print("  06_CHECKLIST_NUCLEO ok")

# ╔══════════════════════════════════════╗
# ║  ABA 07 — LAVAGEM DE REDE            ║
# ╚══════════════════════════════════════╝
ws7 = wb.create_sheet("07_LAVAGEM_REDE")
rn = titulo_aba(ws7,"07 — LAVAGEM DE REDE: Processo Completo",
                "⚠ CRÍTICO: sem lavagem aprovada pela SABESP não se pode fazer ligações de água",14)

mc(ws7,rn,1,rn,14)
wr(ws7,rn,1,"POR QUE É CRÍTICO: A rede nova tem sujeira, sedimentos e contaminação do processo construtivo."
   " Sem lavagem e desinfecção aprovada, a água seria entregue imprópria para consumo. SABESP exige"
   " teste bacteriológico. Todo o processo leva 3-5 dias úteis.",
   pf(VERM_CLA),fn(True,VERM_ESC,9),al("left",wrap=True))
rh(ws7,rn,36); rn+=1; rn+=1

mc(ws7,rn,1,rn,14)
wr(ws7,rn,1,"PASSO A PASSO — LAVAGEM E DESINFECÇÃO",pf(AZUL_ESC),fn(True,"FFFFFF",11),al("left"))
rn+=1

hdr(ws7,rn,["Passo","Atividade","Quando","Duração","Quem","Antecipação\nmín.","Formulário/Doc"],
    [6,40,20,12,18,14,36]); rn+=1

LAVAGEM_STEPS = [
    ("1","Concluir 100% assentamento da rede no trecho","Após último tubo","—","FCN","—","Diário de obra assinado"),
    ("2","Registrar comprimento e diâmetro do trecho","Após conclusão","2h","FCN + Topógrafo","—","Caderneta de campo"),
    ("3","Calcular volume de lavagem (L = π×(D/2)²×L×2)","Antes agendamento","30min","Eng. Responsável","—","Planilha de cálculo"),
    ("4","⚠ LIGAR para SABESP e agendar lavagem","MÍNIMO 2 DIAS ANTES","—","FCN (responsável)","48h mín.","Protocolo número: ________"),
    ("5","Preencher formulário de agendamento SABESP","Ao agendar","30min","FCN","48h mín.","Form. SABESP LA-001"),
    ("6","Confirmar agendamento 1 dia antes (e-mail/tel)","1 dia antes","—","FCN","24h mín.","E-mail de confirmação"),
    ("7","Instalar pontos de descarga (plugs + mangote)","Manhã do dia agendado","2h","Encanador","—","—"),
    ("8","Abertura válvulas e lavagem com velocidade ≥1 m/s","Hora agendada","1-2h/trecho","Técnico SABESP + FCN","—","Laudo de lavagem SABESP"),
    ("9","Adicionar cloro (solução hipoclorito) — concentração SABESP","Após lavagem","30min","Técnico SABESP","—","Dosagem registrada em laudo"),
    ("10","Período de contato cloro (mínimo 24h de espera)","Após cloração","24h","—","—","—"),
    ("11","Coleta de amostras bacteriológicas","Após 24h contato","1h","Técnico SABESP","—","Formulário coleta SABESP"),
    ("12","Aguardar resultado laboratório","Após coleta","24-48h","—","—","—"),
    ("13","Resultado OK → SABESP libera trecho p/ ligações","Após resultado","—","SABESP","—","Laudo bacteriológico OK"),
    ("14","⚠ Resultado NOK → repetir processo do Passo 8","Se falhar","—","FCN + SABESP","—","Reinício do processo"),
    ("15","Emitir 'Autorização de Ligações' para o trecho","Após liberação","—","SABESP","—","Autorização por rua/trecho"),
    ("16","Iniciar ligações de água no trecho liberado","Após autorização","—","FCN equipe lig.","—","—"),
]

for s in LAVAGEM_STEPS:
    is_crit = "⚠" in s[0]+s[1] or "SABESP" in s[6]
    fill_ = pf(VERM_CLA) if is_crit else (pf(AMAR_CLA) if s[0] in("11","12","13","14","15") else None)
    linha(ws7,rn,list(s),fill_=fill_,aln_=al("left",wrap=True))
    rh(ws7,rn,22); rn+=1

rn+=1
mc(ws7,rn,1,rn,14)
wr(ws7,rn,1,"CRONOGRAMA TÍPICO — DO FIM DA REDE ATÉ 1ª LIGAÇÃO:",
   pf(AZUL_MED),fn(True,"FFFFFF",10),al("left")); rn+=1
crono_lav = [
    ("Dia 0","Fim do assentamento da rede","Marco 0"),
    ("Dia 1","Topografia as-built + cálculo volume + AGENDAMENTO SABESP","≥48h antes da lavagem"),
    ("Dia 2","Preparação pontos descarga + confirmação agendamento","—"),
    ("Dia 3","LAVAGEM + CLORAÇÃO (SABESP presente)","Dia agendado"),
    ("Dia 4","Período contato cloro (24h) — sem atividade na rede","Aguardar"),
    ("Dia 5","Coleta amostras bacteriológicas","SABESP"),
    ("Dia 6-7","Aguardar resultado laboratório","48h laboratório"),
    ("Dia 7-8","RESULTADO OK → Início das LIGAÇÕES DE ÁGUA","LIBERA"),
]
hdr(ws7,rn,["Dia","Atividade","Observação"],[8,60,40]); rn+=1
for (dia,atv,obs) in crono_lav:
    fill_ = pf(VERDE_CLA) if "OK" in obs or "LIBERA" in obs else (pf(VERM_CLA) if "SABESP" in atv else None)
    linha(ws7,rn,[dia,atv,obs],fill_=fill_)
    rh(ws7,rn,16); rn+=1

rn+=1
mc(ws7,rn,1,rn,14)
wr(ws7,rn,1,"ESTRATÉGIA: Para não perder tempo — enquanto aguarda liberação da Rua A, inicie Ligações na Rua B (já liberada)",
   pf(VERDE_CLA),fn(True,VERDE_ESC,9),al("left",wrap=True))
rh(ws7,rn,20)
print("  07_LAVAGEM_REDE ok")

# ╔══════════════════════════════════════╗
# ║  ABA 08 — CARTOGRAFIA                ║
# ╚══════════════════════════════════════╝
ws8 = wb.create_sheet("08_CARTOGRAFIA")
rn = titulo_aba(ws8,"08 — CARTOGRAFIA E AS-BUILT",
                "O que fazer, quando, quem e como | Pode ser simultâneo com a obra",12)

mc(ws8,rn,1,rn,12)
wr(ws8,rn,1,"CARTOGRAFIA = Registro preciso de ONDE estão as redes instaladas. Sem isso não há GRD (encerramento).",
   pf(AMAR_CLA),fn(True,AMAR_ESC,9),al("left",wrap=True)); rn+=1

CARTO_ITEMS = [
    ("ANTES DA OBRA","Planta base (existente)","FCN/SABESP","Download SABESP ou compra PMSantos","DWG/PDF",
     "Verificar escala 1:500 ou 1:1000","Gratuito via SABESP"),
    ("ANTES DA OBRA","Levantamento topográfico inicial","Topógrafo","1 dia/quadra","DWG + relatório",
     "Cotas de terreno e sarjeta para calcular profundidades","~R$800/km"),
    ("DURANTE OBRA","GPS cada ponto de PV","Topógrafo/FCN","Contínuo — cada PV escavado","Planilha + GPS",
     "Registrar: Northing, Easting, Altitude, DN, Material, Profundidade","GPS ±30cm precisão"),
    ("DURANTE OBRA","Foto georreferenciada de cada PV","FCN/fiscal","Contínuo","JPG + coordenadas",
     "Celular com GPS ligado — foto tirada NA vala aberta","Grátis (celular)"),
    ("DURANTE OBRA","Croqui de ligação (por imóvel)","FCN","Após cada lig.","Croqui A4",
     "Posição cavalete, comprimento ramal, DN, data","Formulário FCN-LIG"),
    ("APÓS REDE","As-built rede (levantamento GPS)","Topógrafo","1-2 dias/km","DWG georreferenciado",
     "Polilinhas com atributos: DN, mat., Prof, Data","~R$1200/km"),
    ("APÓS LIGAÇÕES","Cadastro de ligações (coordenadas cavalete)","FCN","Diário","Planilha Excel",
     "Colunas: Rua, Número, Hidrômetro, Lat, Lon, Data LA, Data LE","Formato SABESP"),
    ("ENCERRAMENTO","GRD — Guia de Recebimento Definitivo","FCN + SABESP","5-10 dias","Processo formal",
     "Exige: as-built aprovado + fotos + fichas + laudos","Obrigatório p/ pagamento final"),
]

hdr(ws8,rn,["Momento","Atividade","Responsável","Prazo","Entregável","Detalhe Crítico","Custo/Obs"],
    [16,36,16,16,20,44,24]); rn+=1
for item in CARTO_ITEMS:
    mom = item[0]
    fill_ = (pf(AMAR_CLA) if mom=="ANTES DA OBRA" else
             pf(VERDE_CLA) if mom=="DURANTE OBRA" else
             pf(AZUL_CLA) if mom=="APÓS REDE" else
             pf(CINZA))
    linha(ws8,rn,list(item),fill_=fill_,aln_=al("left",wrap=True))
    rh(ws8,rn,24); rn+=1

rn+=1
mc(ws8,rn,1,rn,12)
wr(ws8,rn,1,"💡 DICA: O topógrafo pode trabalhar 1 quadra À FRENTE da escavação — cadastra o terreno enquanto a equipe ainda está escavando a quadra anterior. CUSTO ZERO adicional se incluído no contrato de topografia.",
   pf(VERDE_CLA),fn(True,VERDE_ESC,9),al("left",wrap=True))
rh(ws8,rn,28)
print("  08_CARTOGRAFIA ok")

# ╔══════════════════════════════════════╗
# ║  ABA 09 — CADASTROS                  ║
# ╚══════════════════════════════════════╝
ws9 = wb.create_sheet("09_CADASTROS")
rn = titulo_aba(ws9,"09 — CADASTRO SOCIAL + COMERCIAL",
                "Ambos podem ser feitos ANTES e DURANTE a obra | Não bloqueiam o avanço físico",14)

for titulo_secao, cor_sec, itens in [
    ("CADASTRO SOCIAL — O que é e como fazer", AZUL_MED, [
        ("OBJETIVO","Identificar moradores, resistências, casos especiais (deficientes, idosos, inadimplentes crônicos)",),
        ("QUEM FAZ","Assistente Social designado pela FCN + estagiários de serviço social",),
        ("QUANDO","Iniciar 15 dias antes da obra — continuar durante toda a execução",),
        ("COMO","Visita domiciliar porta-a-porta com tablet/formulário | App SABESP ou Excel",),
        ("DADOS COLETADOS","Nome morador, RG, composição familiar, renda, aceite da ligação, restrições",),
        ("CASOS ESPECIAIS","Isenção tarifa social: renda ≤ R$1.412 | Deficiência: adaptar cavalete | Resistência: acionar SABESP"),
        ("RESULTADO","Planilha: % aceitação (meta ≥85%), lista de pendências, casos para SABESP",),
        ("IMPACTO NA META","Alta aceitação = mais ligações executadas = mais LA/mês contabilizados",),
    ]),
    ("CADASTRO COMERCIAL — Sincronizar com SABESP", VERDE_ESC, [
        ("OBJETIVO","Confirmar dados cadastrais dos imóveis no sistema SABESP (RGI, endereço, situação)",),
        ("QUEM FAZ","FCN Construções e Saneamento + técnico SABESP designado",),
        ("QUANDO","Iniciar 10 dias antes da obra — atualizar continuamente",),
        ("SISTEMA","GSAN (Sistema de Gestão Comercial SABESP) — acesso com login FCN",),
        ("O QUE VERIFICAR","Situação: Normal/Supresso/Cortado | RGI correto | Hidrômetro instalado (água) | Ramal existente",),
        ("IMÓVEL SEM RGI","Abrir processo de cadastramento — prazo 5-10 dias — aguardar antes da ligação",),
        ("IMÓVEL SUPRESSO","Verificar motivo: débito? Irregularidade? Acionar equipe SABESP para regularizar",),
        ("RESULTADO","Planilha: imóveis aptos (sem pendência), imóveis com pendência, previsão regularização",),
        ("IMPACTO NA META","Imóvel com pendência ≠ LA contabilizada → resolver antecipadamente = mais LA/mês",),
    ]),
]:
    mc(ws9,rn,1,rn,14)
    wr(ws9,rn,1,titulo_secao,pf(cor_sec),fn(True,"FFFFFF",11),al("left"))
    rn+=1
    hdr(ws9,rn,["Aspecto","Descrição"],[20,100]); rn+=1
    for item in itens:
        fill_ = pf(VERDE_CLA) if "IMPACTO" in item[0] else (pf(VERM_CLA) if "PENDÊNCIA" in item[0] or "SEM RGI" in item[0] else None)
        wr(ws9,rn,1,item[0],fill_,fn(True,"000000",9),al("left"),THIN)
        mc(ws9,rn,2,rn,14)
        wr(ws9,rn,2,item[1],fill_,fn(False,"000000",9),al("left",wrap=True),THIN)
        rh(ws9,rn,20); rn+=1
    rn+=1

# cronograma paralelo
mc(ws9,rn,1,rn,14)
wr(ws9,rn,1,"LINHA DO TEMPO: Cadastros vs Obra (exemplo núcleo de 8 semanas)",
   pf(AZUL_ESC),fn(True,"FFFFFF",10),al("left")); rn+=1
hdr(ws9,rn,["Atividade","Sem1","Sem2","Sem3","Sem4","Sem5","Sem6","Sem7","Sem8"],
    [30,10,10,10,10,10,10,10,10],fill_col=AZUL_MED); rn+=1
paralelo_rows = [
    ("Cadastro Social",    "████","████","████","████","████","░░░░","░░░░","░░░░"),
    ("Cadastro Comercial", "████","████","████","░░░░","░░░░","░░░░","░░░░","░░░░"),
    ("Escavação/Rede",     "░░░░","████","████","████","████","░░░░","░░░░","░░░░"),
    ("Lavagem + Lab.",     "░░░░","░░░░","░░░░","░░░░","████","█▲░░","░░░░","░░░░"),
    ("Ligações Água",      "░░░░","░░░░","░░░░","░░░░","░░░░","████","████","████"),
    ("Cartografia",        "████","████","████","████","████","████","████","████"),
    ("Boletim Medição",    "░░░░","░░░░","░░░░","████","░░░░","░░░░","░░░░","████"),
]
for row in paralelo_rows:
    fills = [pf(VERDE_CLA) if "█" in v else pf(CINZA) for v in row[1:]]
    wr(ws9,rn,1,row[0],pf(AZUL_CLA),fn(True,"1F4E79",9),al("left"),THIN)
    for ci, (v, f) in enumerate(zip(row[1:], fills), 2):
        wr(ws9,rn,ci,v,f,fn(False,"000000",8),al("center"),THIN)
    rh(ws9,rn,16); rn+=1
print("  09_CADASTROS ok")

# ╔══════════════════════════════════════╗
# ║  ABA 10 — LIGAÇÕES PASSO A PASSO     ║
# ╚══════════════════════════════════════╝
ws10 = wb.create_sheet("10_LIGACOES")
rn = titulo_aba(ws10,"10 — LIGAÇÕES: Passo a Passo Completo (Água e Esgoto)",
                "Guia para leigos | Cada passo com duração, materiais e como conferir",14)

for tipo_lig, passos in [
    ("LIGAÇÃO DE ÁGUA — Sequência Completa", [
        ("ANTES","Conferir liberação da rua (laudo lavagem SABESP)","FCN Encarregado","1h","Laudo OK no sistema","Sem laudo = NÃO INICIAR"),
        ("ANTES","Confirmar dados cadastrais no GSAN (RGI, endereço)","FCN","30min","GSAN atualizado","RGI errado = problema na medição"),
        ("ANTES","Marcar posição do cavalete (a 30cm do alinhamento)","Topógrafo/FCN","15min/lig","Gabarito marcado","Padrão SABESP: cavalete no passeio"),
        ("EXECUÇÃO","Escavação ramal: prof. 60cm, larg. 40cm","Equipe (2 pessoas)","1-2h/lig","Vala aberta","Cuidado com redes de energia elétrica"),
        ("EXECUÇÃO","Instalar colar de tomada no tubo da rede","Encanador","30min","Colar fixado + vedado","Apertar porcas em cruz — não vazar"),
        ("EXECUÇÃO","Ramal PVC 25mm (ou conforme projeto)","Encanador","1h","Ramal instalado","Inclinação 1% mínimo até cavalete"),
        ("EXECUÇÃO","Instalar registro de passeio (bronze ½\")","Encanador","20min","Registro + caixa PVC","Profundidade 30-40cm"),
        ("EXECUÇÃO","Instalar cavalete DN 15 (½\") padrão SABESP","Encanador","30min","Cavalete posicionado","Verificar cota no projeto"),
        ("EXECUÇÃO","Instalar hidrômetro (fornecido pela SABESP)","Encanador","20min","Hidrômetro instalado","Nº de série registrado na ficha"),
        ("EXECUÇÃO","Testar pressão do ramal (abrir registro passeio)","Encanador+FCN","10min","Vazão OK, sem vazamento","Pressão mínima 5mca no cavalete"),
        ("EXECUÇÃO","Reaterro compactado manual (camadas 20cm)","Equipe","30min","Vala fechada","Não usar compactador em cima do tubo"),
        ("EXECUÇÃO","Recompor calçada (piso = padrão original)","Pedreiro","1-2h","Calçada restaurada","Foto antes e depois obrigatória"),
        ("REGISTRO","Preencher Ficha de Ligação de Água (FCN-LA)","FCN","15min","Ficha assinada","Nº hidrômetro + foto do cavalete"),
        ("REGISTRO","Foto: cavalete instalado + nº hidrômetro","FCN","5min","Foto no sistema","Upload no app/planilha FCN"),
        ("REGISTRO","Atualizar Execução_Geral.xlsx — lançar +1 LA","FCN/encarregado","5min","LA + 1 na planilha","Confirmar com Fiscal SABESP"),
        ("MEDIÇÃO","Agrupar lote de 20-30 lig. para medição fiscal","FCN","—","Lista para SABESP","Não medir menos de 20 por visita"),
    ]),
    ("LIGAÇÃO DE ESGOTO — Sequência Completa", [
        ("ANTES","Confirmar que a rede coletora está instalada e testada","FCN","—","Projeto as-built","Sem rede = NÃO CONECTAR"),
        ("ANTES","Verificar cota de saída do imóvel (ralo/privada)","FCN/Topógrafo","30min","Cota registrada","Cota saída > cota rede + 10cm mín."),
        ("ANTES","Confirmar RGI e dados no GSAN","FCN","15min","GSAN OK","—"),
        ("EXECUÇÃO","Escavar vala para ramal 100mm (prof. variável)","Equipe (2 pess.)","2-4h","Vala aberta","Respeitar declividade mínima 2%"),
        ("EXECUÇÃO","Instalar caixa de inspeção (CI) padrão SABESP","Pedreiro","1-2h","CI instalada","Altura CI: nível do passeio"),
        ("EXECUÇÃO","Ramal PVC DN 100mm da CI ao PV (rede)","Encanador","1-2h","Ramal instalado","Decl. ≥2% — usar nível de bolha"),
        ("EXECUÇÃO","Conectar no PV (furo + vedação)","Encanador","30min","Conexão vedada","Argamassa expansiva no furo"),
        ("EXECUÇÃO","Ligar tubulação interna do imóvel à CI","Encanador","30min","Ligação interna OK","Verificar que TODOS os drenos vão p/ CI"),
        ("EXECUÇÃO","Tampar CI com tampa PVC padrão","Encanador","10min","CI tampada","Tampa nivelada com passeio"),
        ("EXECUÇÃO","Reaterro e recomposição calçada","Equipe","1h","Vala fechada","Igual padrão água"),
        ("REGISTRO","Ficha de Ligação de Esgoto (FCN-LE)","FCN","15min","Ficha assinada","—"),
        ("REGISTRO","Foto: CI instalada + conexão ao PV","FCN","5min","Foto no sistema","—"),
        ("REGISTRO","Lançar +1 LE em Execução_Geral.xlsx","FCN/encarregado","5min","LE + 1","—"),
    ]),
]:
    mc(ws10,rn,1,rn,14)
    wr(ws10,rn,1,tipo_lig,pf(AZUL_ESC),fn(True,"FFFFFF",11),al("left")); rn+=1
    hdr(ws10,rn,["Fase","Atividade","Quem","Duração","Verificar","Atenção"],
        [12,50,18,12,28,40]); rn+=1
    for p in passos:
        fase_fill = {"ANTES":pf(AMAR_CLA),"EXECUÇÃO":pf(VERDE_CLA),"REGISTRO":pf(AZUL_CLA),"MEDIÇÃO":pf(CINZA)}.get(p[0],None)
        linha(ws10,rn,list(p),fill_=fase_fill,aln_=al("left",wrap=True))
        rh(ws10,rn,22); rn+=1
    rn+=1
print("  10_LIGACOES ok")

# ╔══════════════════════════════════════╗
# ║  ABA 11 — HISTOGRAMA MO              ║
# ╚══════════════════════════════════════╝
ws11 = wb.create_sheet("11_HISTOGRAMA_MO")
rn = titulo_aba(ws11,"11 — HISTOGRAMA DE MÃO DE OBRA",
                "Pico de equipes por mês Abr-Dez/2026 | Planejamento de RH e folha",18)

MESES_FULL = ["Abr/26","Mai/26","Jun/26","Jul/26","Ago/26","Set/26","Out/26","Nov/26","Dez/26"]

MO_PERFIS = [
    # (perfil, qtd por equipe)
    ("Encanador (líder de equipe)",1),
    ("Auxiliar escavação",2),
    ("Pedreiro (pavimentação/CI)",0.5),
    ("Sinaleiro",0.5),
]

GRUPOS_MO = [
    # (grupo, equipes por mês [abr..dez])
    ("6 núcleos ativos (existente)", [20,10,0,0,0,0,0,0,0]),
    ("R3 Cruzeiro/Prog II",          [6,6,6,6,6,4,4,0,0]),
    ("R8 Noroeste/Gilda",            [6,6,6,6,6,4,4,0,0]),
    ("R7 Piratininga/NW",            [0,5,5,5,5,5,0,0,0]),
    ("R4 Cruzeiro/Penha",            [0,5,5,5,5,5,0,0,0]),
    ("R2 Cruzeiro/Prog I",           [0,8,8,8,8,8,8,4,0]),
    ("R5 Cruzeiro/Bento",            [0,8,8,8,8,8,8,4,0]),
    ("R1 Zona Leste",                [0,0,6,6,6,6,6,6,4]),
    ("R6 Monte Serrat",              [0,0,4,4,4,4,4,4,4]),
]

hdr(ws11,rn,["Grupo"]+MESES_FULL,
    [32]+[11]*9, fill_col=AZUL_MED); rn+=1

row_data_start = rn
for grupo, equipes in GRUPOS_MO:
    linha(ws11,rn,[grupo]+equipes,fill_=pf(AZUL_CLA) if "existente" in grupo else None)
    rn+=1

# totais por mês
f_tots = []
for mi in range(9):
    col_letter = get_column_letter(2+mi)
    f_tots.append(f"=SUM({col_letter}{row_data_start}:{col_letter}{rn-1})")

mc(ws11,rn,1,rn,1)
wr(ws11,rn,1,"TOTAL EQUIPES",pf(VERDE_ESC),fn(True,"FFFFFF",10),al("left"))
for mi, f in enumerate(f_tots, 2):
    wr(ws11,rn,mi,f,pf(VERDE_ESC),fn(True,"FFFFFF",11),al("center"))
rh(ws11,rn,20); rn+=2

# MO por perfil
mc(ws11,rn,1,rn,10)
wr(ws11,rn,1,"COMPOSIÇÃO MÃO DE OBRA POR PERFIL (por equipe × total equipes)",
   pf(AZUL_MED),fn(True,"FFFFFF",10),al("left")); rn+=1

hdr(ws11,rn,["Perfil","Qt/Equipe"]+MESES_FULL,
    [30,12]+[11]*9); rn+=1

for perfil, qtd_eq in MO_PERFIS:
    row_mo = [perfil, qtd_eq]
    for mi in range(9):
        col_eq = get_column_letter(2+mi)
        # referencia ao total de equipes calculado acima
        row_tot = row_data_start + len(GRUPOS_MO)
        row_mo.append(f"=ROUND({col_eq}{row_tot}*{qtd_eq},0)")
    linha(ws11,rn,row_mo)
    rn+=1

# gráfico de barras
rn+=2
mc(ws11,rn,1,rn,18)
wr(ws11,rn,1,"GRÁFICO: Total equipes por mês",pf(AZUL_MED),fn(True,"FFFFFF",10),al("center"))
rn+=1

row_tot_equipes = row_data_start + len(GRUPOS_MO)
chart = BarChart()
chart.type = "col"
chart.title = "Histograma de Equipes por Mês"
chart.y_axis.title = "Total de Equipes"
chart.x_axis.title = "Mês"
chart.height = 12; chart.width = 22
chart.grouping = "clustered"

ref_eq = Reference(ws11, min_col=2, max_col=10,
                   min_row=row_tot_equipes, max_row=row_tot_equipes)
chart.add_data(ref_eq)
chart.series[0].title = SeriesLabel(v="Total Equipes")

cats = Reference(ws11, min_col=2, max_col=10, min_row=1)
ws11.add_chart(chart, f"A{rn}")
print("  11_HISTOGRAMA_MO ok")

# ╔══════════════════════════════════════╗
# ║  ABA 12 — HISTOGRAMA EQUIPAMENTOS    ║
# ╚══════════════════════════════════════╝
ws12 = wb.create_sheet("12_HISTOGRAMA_EQ")
rn = titulo_aba(ws12,"12 — HISTOGRAMA DE EQUIPAMENTOS",
                "Planejamento de máquinas, veículos e ferramentas por mês",18)

EQ_TIPOS = [
    ("Retroescavadeira (escavação vala)","un","mês"),
    ("Compactador de placa (reaterro)","un","mês"),
    ("Caminhão basculante (bota-fora)","un","mês"),
    ("Caminhão pipa (controle poeira/limpeza)","un","mês"),
    ("Grupo gerador (iluminação noturna se necessário)","un","mês"),
    ("Bomba d'água (esgotamento)","un","mês"),
    ("Veículo pickup (supervisão/material urgente)","un","mês"),
    ("Escoramento (jogos = 6 pranchas)","jogos","mês"),
    ("Vibrador de concreto (PVs)","un","mês"),
    ("GPS topografia (cadastro as-built)","un","mês"),
    ("Detector de redes (evitar danos)","un","mês"),
]
EQ_MES = [
    # retroesc, compact, bascu, pipa, gerador, bomba, pickup, escor, vibr, gps, detec
    ([2,2,2,2,0,0,0,0,0],  # abr - 6 núcleos existentes
     [2,2,2,2,0,0,0,0,0],  # mai
     [1,1,1,1,0,0,0,0,0],  # jun
     [0,0,0,0,0,0,0,0,0],  # jul
     [0,0,0,0,0,0,0,0,0],  # ago
     [0,0,0,0,0,0,0,0,0],  # set
     [0,0,0,0,0,0,0,0,0],  # out
     [0,0,0,0,0,0,0,0,0],  # nov
     [0,0,0,0,0,0,0,0,0],) # dez
]

# Simplificado: tabela por equipamento
EQ_ROWS = [
    ("Retroescavadeira",    [4,8,8,6,6,4,4,2,2]),
    ("Compactador placa",   [4,8,8,6,6,4,4,2,2]),
    ("Caminhão basculante", [2,4,4,4,4,2,2,2,2]),
    ("Caminhão pipa",       [1,2,2,2,2,1,1,1,1]),
    ("Bomba d'água",        [2,4,4,4,4,2,2,2,2]),
    ("Pickup supervisão",   [3,5,5,5,5,4,4,4,4]),
    ("Escoramento (jogos)", [4,8,8,8,8,6,6,4,4]),
    ("Vibrador concreto",   [2,4,4,4,4,2,2,2,2]),
    ("GPS topografia",      [1,2,2,2,2,2,2,2,2]),
    ("Detector redes",      [1,2,2,2,2,1,1,1,1]),
]

hdr(ws12,rn,["Equipamento"]+MESES_FULL+["Unid."],
    [32]+[10]*9+[8]); rn+=1
eq_start = rn
for eq, qtds in EQ_ROWS:
    linha(ws12,rn,[eq]+qtds+["un"])
    rn+=1

# Custo por mês (estimativa)
CUSTO_EQ = {"Retroescavadeira":8500,"Compactador placa":800,"Caminhão basculante":4500,
            "Caminhão pipa":2500,"Bomba d'água":600,"Pickup supervisão":3200,
            "Escoramento (jogos)":400,"Vibrador concreto":350,"GPS topografia":1800,"Detector redes":1200}
rn+=1
mc(ws12,rn,1,rn,11)
wr(ws12,rn,1,"CUSTO LOCAÇÃO ESTIMADO (R$/mês por unidade)",pf(AZUL_MED),fn(True,"FFFFFF",10),al("left")); rn+=1
hdr(ws12,rn,["Equipamento","R$/mês/un"]+MESES_FULL,
    [32,12]+[12]*9); rn+=1

for i,(eq, qtds) in enumerate(EQ_ROWS):
    custo_un = CUSTO_EQ.get(eq, 1000)
    row_eq = [eq, custo_un]
    for q in qtds:
        row_eq.append(q * custo_un)
    linha(ws12,rn,row_eq)
    rn+=1

# total
wr(ws12,rn,1,"TOTAL EQUIPAMENTOS/MÊS",pf(VERDE_ESC),fn(True,"FFFFFF",10),al("left"))
for mi in range(9):
    col = get_column_letter(3+mi)
    wr(ws12,rn,3+mi,f"=SUM({col}{eq_start+len(EQ_ROWS)+3}:{col}{rn-1})",
       pf(VERDE_ESC),fn(True,"FFFFFF",10),al("center"))
rh(ws12,rn,20)
print("  12_HISTOGRAMA_EQ ok")

# ╔══════════════════════════════════════╗
# ║  ABA 13 — PMBOK + LEAN               ║
# ╚══════════════════════════════════════╝
ws13 = wb.create_sheet("13_PMBOK_LEAN")
rn = titulo_aba(ws13,"13 — PMBOK 7ª Edição + LEAN Last Planner aplicados ao Contrato",
                "Metodologias internacionais adaptadas ao SLNR 11481051 | Santos/SP",14)

pmbok_content = [
    ("PMBOK 7ª EDIÇÃO — 12 PRINCÍPIOS aplicados","AZUL"),
    ("1. Administração","Ser diligente, respeitoso e cuidadoso nas responsabilidades. → Reunião semanal FCN+FCN+SABESP todo a 2ª feira 08h.",),
    ("2. Equipe","Construir um ambiente colaborativo de alto desempenho. → Metas diárias afixadas no canteiro | Prêmio por meta atingida.",),
    ("3. Partes Interessadas","Engajar partes interessadas proativamente. → Reunião mensal com moradores | Canal WhatsApp SABESP-FCN.",),
    ("4. Valor","Focar continuamente no valor. → KPI principal: LA/mês. Tudo medido contra essa meta.",),
    ("5. Pensamento Sistêmico","Ver o todo. → Mapa de processos integrado (Aba 04). Lavagem ≠ opcional.",),
    ("6. Liderança","Demonstrar comportamentos de liderança. → Encarregado com metas pessoais | Painel visual no canteiro.",),
    ("7. Adaptação","Adaptar baseado no contexto. → XGBoost recomenda ajustar equipes conforme DIFIC do trecho.",),
    ("8. Qualidade","Incorporar qualidade nos processos. → Teste estanqueidade + laudo bacteriológico antes de qualquer ligação.",),
    ("9. Complexidade","Navegar a complexidade. → Clusters geográficos reduzem deslocamento e aumentam eficiência.",),
    ("10. Riscos","Otimizar respostas a riscos. → Aba 09_RISCOS com plano de contingência (Cronograma SLNR).",),
    ("11. Adaptabilidade","Ser ágil e adaptável. → Last Planner semanal + lookahead 6 semanas.",),
    ("12. Mudança","Habilitar a mudança para alcançar o estado futuro desejado. → Abertura de novos núcleos = mudança planejada.",),
]

mc(ws13,rn,1,rn,14)
wr(ws13,rn,1,pmbok_content[0][0],pf(AZUL_ESC),fn(True,"FFFFFF",11),al("left")); rn+=1
for item in pmbok_content[1:]:
    princip, desc = item[0].split(".",1)[0]+". "+item[0].split(".",1)[1].split("→")[0].strip(), ""
    texto = item[0]
    mc(ws13,rn,1,rn,2)
    wr(ws13,rn,1,texto.split("→")[0].strip(),pf(AZUL_CLA),fn(True,"1F4E79",9),al("left"))
    mc(ws13,rn,3,rn,14)
    wr(ws13,rn,3,"→ "+texto.split("→")[1].strip() if "→" in texto else "",
       None,fn(False,"000000",9),al("left",wrap=True))
    rh(ws13,rn,18); rn+=1

rn+=1

# LEAN
mc(ws13,rn,1,rn,14)
wr(ws13,rn,1,"LEAN CONSTRUCTION — LAST PLANNER SYSTEM (LPS) aplicado",
   pf(VERDE_ESC),fn(True,"FFFFFF",11),al("left")); rn+=1

lean_items = [
    ("O QUE É LEAN?","Eliminar desperdícios. No canteiro: espera por material, retrabalho, deslocamento desnecessário, over-processing."),
    ("LAST PLANNER","Sistema de planejamento colaborativo em 3 horizontes: Mestre (projeto) → Lookahead 6 sem. → Semanal comprometido."),
    ("LOOKAHEAD 6 SEM.","Identificar restrições antes que bloqueiem a obra. Cartografia, lavagem, materiais, licenças → remover 6 semanas antes."),
    ("REUNIÃO SEMANAL LPS","Toda 2ª-feira: O que foi feito? O que está planejado? O que bloqueia? → PPC (% plano cumprido). Meta: PPC ≥ 80%."),
    ("PPC — % Plano Cumprido","PPC = tarefas concluídas / tarefas comprometidas × 100. PPC < 70%: analisar causa raiz. Meta: 85%."),
    ("PULL PLANNING","Planejar DE TRÁS PARA A FRENTE. Meta Maio = 1000 LA. Quantas equipes? Quando abrir grupos? → Cronograma puxa a realidade."),
    ("KANBAN","Quadro visual no canteiro: Pendente | Em execução | Concluído. Por rua/trecho. Todo encarregado atualiza diariamente."),
    ("ELIMINAÇÃO ESPERAS","Lavagem SABESP: agendar antes de precisar. Material: estoque 2 semanas. Licenças: solicitar 3 semanas antes."),
    ("5S NO CANTEIRO","Seiri (classificar), Seiton (organizar), Seiso (limpar), Seiketsu (padronizar), Shitsuke (disciplinar). Canteiro padrão FCN."),
    ("TAKT TIME","Ritmo de produção necessário. Meta 1000 LA/mês ÷ 22 dias úteis = 45,5 LA/dia. Cada grupo deve contribuir para esse takt."),
    ("OBEYA (WAR ROOM)","Sala de gestão visual na obra: painel com Gantt, KPIs, fotos, pendências. Reunião de pé (standing meeting) 15min/dia."),
    ("VALUE STREAM MAP","Mapeamento do fluxo de valor: da escavação da vala até a ligação contabilizada. Identificar onde o processo para."),
]

hdr(ws13,rn,["Ferramenta Lean","Aplicação Prática no SLNR 11481051"],[22,90]); rn+=1
for (tool, aplicacao) in lean_items:
    wr(ws13,rn,1,tool,pf(VERDE_CLA),fn(True,VERDE_ESC,9),al("left"),THIN)
    mc(ws13,rn,2,rn,14)
    wr(ws13,rn,2,aplicacao,None,fn(False,"000000",9),al("left",wrap=True),THIN)
    rh(ws13,rn,20); rn+=1

rn+=1
# PMBOK Grupos de Processos
mc(ws13,rn,1,rn,14)
wr(ws13,rn,1,"GRUPOS DE PROCESSOS PMBOK × FASE DO CONTRATO SLNR",
   pf(AZUL_ESC),fn(True,"FFFFFF",11),al("left")); rn+=1
hdr(ws13,rn,["Grupo PMBOK","O que fazer","Quando","Saídas/Entregáveis"],
    [18,36,20,40]); rn+=1
pmbok_grupos = [
    ("INICIAÇÃO","Formalizar abertura de cada núcleo com TAP","Antes de iniciar núcleo","TAP (Term. Abertura Proj.), designação Eng. Resp."),
    ("PLANEJAMENTO","Planejar escopo, prazo, custo, RH, riscos","2 sem antes","Cronograma + Histograma MO + Plano Riscos"),
    ("EXECUÇÃO","Executar conforme planejado + LPS semanal","Durante obra","Registros diários, fotos, fichas LA/LE"),
    ("MONIT. E CONTROLE","Medir KPIs vs metas | Ação corretiva imediata","Diário/semanal","Relatório semanal, PPC, desvios"),
    ("ENCERRAMENTO","GRD, as-built, documentação, desligamento","Após conclusão","GRD assinado, documentação entregue"),
]
for g in pmbok_grupos:
    fills = [pf(AMAR_CLA),None,None,pf(VERDE_CLA)]
    for ci,v in enumerate(g,1):
        wr(ws13,rn,ci,v,fills[ci-1] if ci-1<len(fills) else None,
           fn(True if ci==1 else False,"000000",9),al("left",wrap=True),THIN)
    rh(ws13,rn,22); rn+=1
print("  13_PMBOK_LEAN ok")

# ╔══════════════════════════════════════╗
# ║  ABA 14 — META 1000 PASSO A PASSO   ║
# ╚══════════════════════════════════════╝
ws14 = wb.create_sheet("14_META_1000")
rn = titulo_aba(ws14,"14 — PLANO TÁTICO: 1.000 LIGAÇÕES DE ÁGUA EM MAIO 2026",
                "Semana a semana | Quem faz o quê | Resultado esperado por ação",16)

mc(ws14,rn,1,rn,16)
wr(ws14,rn,1,
   "SITUAÇÃO HOJE (25/Mar): ~320 LA/mês | META MAI: 1.000 LA/mês | GAP: +680 LA/mês | PRAZO: 5 semanas",
   pf(VERM_CLA),fn(True,VERM_ESC,11),al("center")); rh(ws14,rn,22); rn+=1

SEMANAS_PLANO = [
    ("SEM 01\n25-31/Mar", "Pré-mobilização",
     "• Emitir OF (Ordem de Fornecimento) R3 e R8\n"
     "• Solicitar licença sinalização Pref. Santos para R3+R8\n"
     "• Contratar/designar 12 novas equipes (R3: 6eq + R8: 6eq)\n"
     "• Iniciar cadastro social R3+R8\n"
     "• Pedido de compra materiais R3+R8 (tubos, conexões)\n"
     "• Manter ritmo 6 núcleos ativos → meta 70+ LA/sem",
     "6 núcleos ativos\n0 novos grupos",
     "~70 LA",
     "OF emitida\nLicença protocolada\nMaterial pedido"),

    ("SEM 02\n01-05/Abr", "Mobilização R3+R8",
     "• Montar canteiros R3 (Cruzeiro/Prog II) e R8 (Noroeste/Gilda)\n"
     "• Treinamento NR-18 novas equipes R3+R8\n"
     "• Topógrafo: locação eixos R3+R8\n"
     "• Início escavação R3 (trecho 1) e R8 (trecho 1)\n"
     "• Agendar lavagem R3+R8 (data prevista: sem 4)\n"
     "• Boletim medição Mar/26 enviado até dia 20",
     "6 ativos + R3+R8 iniciando",
     "~80 LA",
     "Canteiros montados\nEscavação iniciada"),

    ("SEM 03\n06-12/Abr", "Execução R3+R8 acelerada",
     "• R3+R8: concluir 60% da rede (escavação + tubos)\n"
     "• Emitir OF R7 e R4 (mobilização para 22/Abr)\n"
     "• Solicitar licença sinalização R7+R4\n"
     "• Contratar 10 equipes R7 (5eq) + R4 (5eq)\n"
     "• Cadastro Social R3+R8: 80% das casas visitadas\n"
     "• Cadastro Comercial R3+R8: confirmar GSAN",
     "6 ativos + R3+R8 em execução",
     "~95 LA",
     "Rede R3+R8: 60% ok\nOF R7+R4 emitida"),

    ("SEM 04\n13-19/Abr", "Lavagem R3+R8 + início ligações",
     "• LAVAGEM DE REDE R3+R8 (SABESP — confirmado)\n"
     "• Resultado lab. bacteriológico R3+R8 esperado\n"
     "• Início ligações de água nos trechos liberados R3+R8\n"
     "• R3+R8: concluir 100% da rede\n"
     "• Mobilização R7+R4: montar canteiros (22/Abr)\n"
     "• 6 núcleos ativos: manter pressão máxima",
     "6 ativos + R3+R8 lig. + R7+R4 mobiliz.",
     "~130 LA",
     "Lavagem realizada\n1as ligações R3+R8"),

    ("SEM 05\n20-26/Abr", "R7+R4 início + R3+R8 pleno",
     "• R7 e R4 iniciam escavação (22/Abr)\n"
     "• R3+R8: pleno ritmo de ligações (6+6 equipes lig.)\n"
     "• Agendar lavagem R7+R4 (semana 7)\n"
     "• Emitir OF R2+R5 (mobilização 12/Mai)\n"
     "• Medição coletiva: agrupar 60+ LA para fiscal SABESP\n"
     "• Reunião LPS: avaliar PPC semana anterior",
     "6 ativos + R3+R8 pleno + R7+R4 inicia",
     "~170 LA",
     "R7+R4 escavando\nR3+R8 em ligações"),

    ("SEM 06\n27/Abr-03/Mai", "Aceleração rumo à meta",
     "• R3+R8: 50%+ das ligações concluídas\n"
     "• R7+R4: 50% da rede concluída\n"
     "• R2+R5: mobilização (canteiros, equipes, material)\n"
     "• Fiscal SABESP: medição lote LA R3+R8\n"
     "• Boletim Medição Abr/26 preparar\n"
     "• XGBoost dashboard: verificar projeção vs real",
     "6 ativos + R3+R8 lig. + R7+R4 rede + R2+R5 mob.",
     "~210 LA",
     "R7+R4 rede 50%\nMedição R3+R8"),

    ("SEM 07\n04-10/Mai", "Lavagem R7+R4 + R2+R5 início",
     "• LAVAGEM DE REDE R7+R4 (SABESP)\n"
     "• Resultado laboratorial R7+R4\n"
     "• Início ligações R7+R4 (trechos liberados)\n"
     "• R2+R5: início escavação (DIFIC 2 — mais rápido)\n"
     "• R3+R8: 80%+ ligações concluídas\n"
     "• Boletim Medição Abr/26 enviado até dia 20",
     "6 ativos + R3+R8+R7+R4 lig. + R2+R5 início",
     "~280 LA",
     "Lavagem R7+R4\nLigações 4 grupos"),

    ("SEM 08\n11-17/Mai", "Pico produtivo — meta à vista",
     "• TODOS os grupos (R3,R8,R7,R4,R2,R5) em ligações\n"
     "• 6 núcleos ativos em ritmo máximo\n"
     "• R2+R5: concluir rede (DIFIC 2 = mais veloz)\n"
     "• Agendar lavagem R2+R5 (semana 9)\n"
     "• Medição fiscal: lote unificado R3+R8+R7+R4\n"
     "• Contagem parcial: projetar fechamento mês",
     "6 ativos + 4 grupos ligações + R2+R5 rede",
     "~320 LA",
     "Todos grupos ativos\nProjeção: 1000+"),

    ("SEM 09-10\n18-31/Mai", "META 1.000 LA/mês ATINGIDA",
     "• Lavagem e liberação R2+R5\n"
     "• R2+R5 início ligações (DIFIC 2 = alta eficiência)\n"
     "• Consolidar medição total Mai/26\n"
     "• Boletim Mai/26 preparar com total LA\n"
     "• Reunião de resultados com gerente geral SABESP\n"
     "• Planejar Jun/26: abrir R1+R6",
     "6 ativos + 6 grupos ligações",
     "≥1.000 LA/mês ✓",
     "META ATINGIDA\nBM Mai enviado"),
]

hdr(ws14,rn,["Semana","Fase","Ações Táticas (COMPLETO)","Grupos Ativos",
              "Meta LA\nAcum. Sem.","Marcos/Verificar"],
    [12,16,70,32,12,32]); rn+=1

for (sem, fase, acoes, grupos, meta, marcos) in SEMANAS_PLANO:
    is_meta = "META" in fase or "1.000" in meta
    fill_ = pf(VERDE_CLA) if is_meta else (pf(AMAR_CLA) if int(sem.split("\n")[0].replace("SEM ","").strip()[:2])>5 else None)
    for ci, v in enumerate([sem,fase,acoes,grupos,meta,marcos],1):
        f = fn(True,VERDE_ESC,9) if is_meta else fn(False,"000000",8 if ci==3 else 9)
        wr(ws14,rn,ci,v,fill_,f,al("left",wrap=True),THIN)
    rh(ws14,rn,80); rn+=1

print("  14_META_1000 ok")

# ╔══════════════════════════════════════╗
# ║  ABA 15 — APRESENTAÇÃO FINAL         ║
# ╚══════════════════════════════════════╝
ws15 = wb.create_sheet("15_APRESENTACAO")
rn = titulo_aba(ws15,"15 — APRESENTAÇÃO GERENCIAL: Resumo Executivo",
                "Para usar na reunião com Gerente Geral SABESP | Versão condensada | 25/Mar/2026",14)

slides = [
    ("SLIDE 1 — ONDE ESTAMOS HOJE", AZUL_ESC, [
        ("Contrato", "11481051 — SE LIGA NA REDE — Santos/SP"),
        ("Executado LA", "894 ligações de água (206% da meta mensal proporcional)"),
        ("Executado LE", "815 ligações de esgoto (133% da meta)"),
        ("Rede Água", "5.235m executados (193% da meta)"),
        ("Rede Esgoto", "4.316m executados (107% da meta)"),
        ("Produtividade", "3,0 m/eq/dia REAL vs 6,0 m/eq/dia contratado → GAP -50%"),
        ("Diagnóstico ML", "XGBoost: fator #1 = n° de equipes | DIFIC impacta -18% | Aprendizado +12%"),
    ]),
    ("SLIDE 2 — O DESAFIO", VERM_ESC, [
        ("Meta Maio/26", "1.000 ligações de água/mês"),
        ("Situação atual", "~320 LA/mês (6 núcleos ativos)"),
        ("GAP a fechar", "+680 LA/mês em 5 semanas"),
        ("Solução", "Abrir 4 novos grupos de proximidade em Abril → 2 mais em Maio"),
        ("Processos críticos", "Lavagem SABESP (2 dias antecedência) | Cadastro Social | Licenças"),
        ("Premissa", "Todos os processos de apoio em paralelo com a obra"),
    ]),
    ("SLIDE 3 — O PLANO (XGBoost recomenda)", VERDE_ESC, [
        ("07/Abril", "Abrir R3 Cruzeiro/Prog II (6 eq, DIFIC 3) + R8 Noroeste/Gilda (6 eq, DIFIC 3)"),
        ("22/Abril", "Abrir R7 Piratininga/NW (5 eq, DIFIC 3) + R4 Cruzeiro/Penha (5 eq, DIFIC 3)"),
        ("12/Maio",  "Abrir R2 Cruzeiro/Prog I (8 eq, DIFIC 2) + R5 Cruzeiro/Bento (8 eq, DIFIC 2)"),
        ("Resultado", "38 equipes ativas em 10 grupos = 1.050+ LA/mês"),
        ("Custo adicional", "~R$18M/mês em medição vs R$241M contrato total"),
        ("ROI", "Cada LA adicional = R$7.500 de medição"),
    ]),
    ("SLIDE 4 — PROCESSOS CRÍTICOS E TEMPOS", AMAR_ESC, [
        ("Lavagem SABESP", "Agendar 48h antes | Processo completo: 8 dias do fim da rede à 1ª ligação"),
        ("Cartografia", "Iniciar 4 semanas antes | Pode ser 100% paralelo com a obra"),
        ("Cadastro Social", "Iniciar 2 semanas antes | 85% aceitação esperada"),
        ("Cadastro Comercial", "Verificar GSAN | Imóveis sem RGI bloqueiam ligação"),
        ("Licença sinalização", "15 dias úteis Prefeitura Santos | Solicitar AGORA para grupos de Abril"),
        ("ART/RRT", "20 dias antes | CREA confirmado"),
    ]),
    ("SLIDE 5 — RECURSOS NECESSÁRIOS", AZUL_MED, [
        ("Equipes adicionais (Abr)", "+22 equipes (R3+R8: 12, R7+R4: 10)"),
        ("Equipes adicionais (Mai)", "+16 equipes (R2+R5: 16)"),
        ("Retroescavadeiras (Abr)", "+4 máquinas (2/grupo)"),
        ("Caminhões basculante", "+4 (2/grupo abertura)"),
        ("Material Tubos DN100", "Pedido urgente: ~4.500m para R3+R8"),
        ("Total folha adicional (Abr)", "~R$132k/mês (22 equipes × R$6k)"),
        ("Aprovação necessária", "Liberação gerencial para contratação imediata"),
    ]),
    ("SLIDE 6 — MONITORAMENTO (LEAN + PMBOK)", AZUL_ESC, [
        ("KPI diário", "LA do dia vs meta (45 LA/dia = 1.000/mês)"),
        ("Reunião LPS", "2ª-feira 08h | 15min | PPC meta ≥80%"),
        ("Dashboard", "Execução_Geral.xlsx atualizado diariamente por encarregado"),
        ("Lookahead", "6 semanas à frente | Restrições identificadas e removidas"),
        ("Alerta vermelho", "LA semanal <80% da meta → acionar plano contingência"),
        ("Boletim Medição", "Enviado até dia 20 de cada mês para SABESP"),
    ]),
]

for slide_title, cor, items in slides:
    mc(ws15,rn,1,rn,14)
    wr(ws15,rn,1,slide_title,pf(cor),fn(True,"FFFFFF",12),al("left"))
    rh(ws15,rn,22); rn+=1
    for (label, valor) in items:
        wr(ws15,rn,1,label,pf(AZUL_CLA),fn(True,"1F4E79",9),al("left"),THIN)
        mc(ws15,rn,2,rn,14)
        wr(ws15,rn,2,valor,None,fn(False,"000000",9),al("left",wrap=True),THIN)
        rh(ws15,rn,18); rn+=1
    rn+=1

# colunags largura
for c in range(1,15): cw(ws15,c,16 if c==1 else 14)
cw(ws15,1,22); cw(ws15,2,90)

print("  15_APRESENTACAO ok")

# ── salvar ──────────────────────────────────────────────────────────────────
wb.save(OUT_FILE)
print(f"\nSalvo: {OUT_FILE}")
abas = [ws.title for ws in wb.worksheets]
print(f"Total abas: {len(abas)}")
for a in abas: print(f"  • {a}")
