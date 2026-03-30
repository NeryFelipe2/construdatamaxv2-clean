#!/usr/bin/env python3
"""
GERAR_XLSX.PY — Gerador de Planilhas Profissionais
ConstruData - HydroNetwork · FCN Construções e Saneamento

Cada módulo que gerava JSON/CSV fraco agora gera XLSX profissional.
JSON continua existindo (pra máquina). XLSX é pra gente.

PLANILHAS GERADAS:
  1. LEAN_LPS_{nucleo}.xlsx      — 5 abas: Resumo, Takt, LPS, Lookahead, 6D
  2. CURVA_S_{nucleo}.xlsx       — Previsto × Realizado com gráfico
  3. MICROPLAN_{nucleo}.xlsx     — Morfologia + Material JIT + Gantt
  4. CUSTOS_{nucleo}.xlsx        — Detalhamento + BM + Composição
  5. PERDAS_{nucleo}.xlsx        — Balanço IWA + UARL + ILI + Risco
  6. ML_{nucleo}.xlsx            — Cenários + Pipeline + Predição
  7. HIDRAULICA_{nucleo}.xlsx    — Manning + Alertas
"""

import json, math, os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════
# ESTILOS PADRÃO CONSTRUDATA
# ══════════════════════════════════════════════════════════

H_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
H_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
D_FONT = Font(name="Arial", size=10)
D_BOLD = Font(name="Arial", size=10, bold=True)
TITLE = Font(name="Arial", size=14, bold=True, color="003366")
SUB = Font(name="Arial", size=10, italic=True, color="666666")
OK_FONT = Font(name="Arial", size=10, color="008844", bold=True)
WARN_FONT = Font(name="Arial", size=10, color="CC8800", bold=True)
ERR_FONT = Font(name="Arial", size=10, color="CC0033", bold=True)
MONEY = '#,##0'
PCT = '0.0%'
DEC1 = '0.0'
DEC2 = '0.00'
BRD = Border(left=Side('thin', 'CCCCCC'), right=Side('thin', 'CCCCCC'),
             top=Side('thin', 'CCCCCC'), bottom=Side('thin', 'CCCCCC'))
BG_G = PatternFill(start_color="E8FFE8", fill_type="solid")
BG_R = PatternFill(start_color="FFE8E8", fill_type="solid")
BG_Y = PatternFill(start_color="FFFFF0", fill_type="solid")
BG_B = PatternFill(start_color="E8F4FF", fill_type="solid")
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)


def _hdr(ws, row, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = H_FONT; cell.fill = H_FILL; cell.alignment = CENTER; cell.border = BRD

def _dat(ws, r1, r2, nc):
    for r in range(r1, r2 + 1):
        for c in range(1, nc + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = D_FONT; cell.border = BRD; cell.alignment = CENTER

def _aw(ws, nc):
    for c in range(1, nc + 1):
        mx = max((len(str(ws.cell(r, c).value or '')) for r in range(1, ws.max_row + 1)), default=8)
        ws.column_dimensions[get_column_letter(c)].width = min(mx + 4, 35)

def _title(ws, text, row=1, cols=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    ws.cell(row=row, column=1, value=text).font = TITLE

def _sub(ws, text, row=2, cols=6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = SUB

def _footer():
    return f"FCN Construções e Saneamento · ConstruData HydroNetwork · {datetime.now().strftime('%d/%m/%Y')}"


# ══════════════════════════════════════════════════════════
# 1. LEAN + LPS + 6D
# ══════════════════════════════════════════════════════════

def gerar_xlsx_lean(relatorio, pvs, trechos, nucleo, out_path):
    """XLSX 5 abas: Resumo, Takt, LPS, Lookahead, BIM 6D."""
    wb = Workbook()
    lean = relatorio.get("lean", {})
    lps_data = relatorio.get("lps", {})
    d6 = relatorio.get("bim_6d", {})
    takt = lean.get("takt_time", {})
    ext = sum(t.get("ext_m", 0) for t in trechos)
    n_tr = len(trechos)

    # ─── RESUMO ───
    ws = wb.active; ws.title = "RESUMO"; ws.sheet_properties.tabColor = "003366"
    _title(ws, f"LEAN + LPS + BIM 6D — {nucleo}", 1, 6)
    _sub(ws, _footer(), 2, 6)

    _hdr(ws, 4, ["Indicador", "Valor", "Unidade", "Meta", "Status", "Observação"])
    kpis = [
        ("Extensão total", f"{ext:.0f}", "m", "", "", f"{n_tr} trechos"),
        ("Takt Time", f"{takt.get('takt_m_dia_equipe', 0):.1f}", "m/dia/eq", "30", "", ""),
        ("Cycle Time", f"{takt.get('cycle_time_dias', 0):.1f}", "dias/NS", "5", "", ""),
        ("Throughput", f"{takt.get('throughput_ns_semana', 0):.0f}", "NS/semana", "10", "", ""),
        ("PPC Médio", f"{lps_data.get('ppc_atual', {}).get('ppc_pct', 0):.0f}", "%", "80", "", ""),
        ("Eficiência Fluxo", f"{lean.get('vsm', {}).get('flow_efficiency_pct', 0):.0f}", "%", "50", "", "VA / (VA+NVA)"),
        ("CO2 Total", f"{d6.get('co2_total_ton', 0):.1f}", "ton", "", "", ""),
        ("Custo Ciclo Vida", f"{d6.get('custo_ciclo_vida_total', 0):,.0f}", "R$", "", "", "50 anos"),
        ("Manutenção/ano", f"{d6.get('manutencao_anual', 0):,.0f}", "R$/ano", "", "", ""),
    ]
    for r, row_data in enumerate(kpis, 5):
        for c, v in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=v)
        # Status automático
        try:
            val = float(str(row_data[1]).replace(",", ""))
            meta = float(row_data[3]) if row_data[3] else 0
            if meta > 0:
                ok = val >= meta
                ws.cell(row=r, column=5, value="✅ OK" if ok else "⚠️ ABAIXO")
                ws.cell(row=r, column=5).font = OK_FONT if ok else WARN_FONT
                ws.cell(row=r, column=1).fill = BG_G if ok else BG_Y
        except:
            pass
    _dat(ws, 5, 5 + len(kpis) - 1, 6); _aw(ws, 6)

    # Gráfico barras
    chart = BarChart(); chart.title = "Indicadores Lean"; chart.style = 10; chart.height = 12; chart.width = 18
    data_ref = Reference(ws, min_col=2, min_row=4, max_row=8, max_col=2)
    cats_ref = Reference(ws, min_col=1, min_row=5, max_row=8)
    chart.add_data(data_ref, titles_from_data=True); chart.set_categories(cats_ref)
    ws.add_chart(chart, "H4")

    # ─── TAKT TIME ───
    ws2 = wb.create_sheet("TAKT TIME"); ws2.sheet_properties.tabColor = "008844"
    _title(ws2, "CÁLCULO DE TAKT TIME", 1, 5)
    _hdr(ws2, 3, ["Parâmetro", "Valor", "Unidade", "Fórmula", "Obs"])
    equipes = takt.get("equipes", 6)
    dados_t = [
        ("Extensão total", ext, "m", "", ""),
        ("Nº trechos (NS)", n_tr, "un", "", ""),
        ("Nº equipes", equipes, "un", "", ""),
        ("Dias úteis/mês", 22, "dias", "", ""),
        ("Takt (m/dia/equipe)", round(ext / max(equipes, 1) / 22, 1), "m/dia/eq", "Ext/(Eq×22)", f"Meta: 30"),
        ("Takt (m/dia total)", round(ext / 22, 1), "m/dia", "Ext/22", ""),
        ("Cycle Time (dias/NS)", round(n_tr / max(equipes, 1) / 22 * 30, 1), "dias", "Tr/(Eq×22)×30", "Meta: 5"),
        ("Throughput (NS/sem)", round(equipes * 5, 0), "NS/sem", "Eq×5", ""),
        ("Lead Time estimado", round(n_tr / max(equipes * 5, 1), 0), "semanas", "Tr/Throughput", ""),
        ("Duração estimada", round(n_tr / max(equipes * 5, 1) / 4, 1), "meses", "LeadTime/4", ""),
    ]
    for r, row_data in enumerate(dados_t, 4):
        for c, v in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=v)
    _dat(ws2, 4, 4 + len(dados_t) - 1, 5); _aw(ws2, 5)

    # ─── LPS - PLANO SEMANAL ───
    ws3 = wb.create_sheet("LPS - PLANO SEMANAL"); ws3.sheet_properties.tabColor = "CC8800"
    _title(ws3, "WEEKLY WORK PLAN — Last Planner System", 1, 9)
    _hdr(ws3, 3, ["NS", "PV Ini", "PV Fim", "DN", "Ext (m)", "Material", "Equipe", "Restrições", "Status"])
    wwp = lps_data.get("weekly_work_plan", {}).get("ns_planejadas", [])
    if not wwp:
        # Gerar WWP a partir dos trechos
        for i, t in enumerate(trechos[:30], 4):
            ws3.cell(row=i, column=1, value=f"NS_{i-3:03d}")
            ws3.cell(row=i, column=2, value=t.get("pv_ini", ""))
            ws3.cell(row=i, column=3, value=t.get("pv_fim", ""))
            ws3.cell(row=i, column=4, value=t.get("dn_mm", 200))
            ws3.cell(row=i, column=5, value=round(t.get("ext_m", 0), 1))
            ws3.cell(row=i, column=6, value=t.get("material", "PVC"))
            ws3.cell(row=i, column=7, value=f"EQ_{((i-4) % equipes)+1:02d}")
            ws3.cell(row=i, column=8, value="Nenhuma")
            ws3.cell(row=i, column=9, value="PLANEJADA")
            ws3.cell(row=i, column=8).fill = BG_G
            ws3.cell(row=i, column=9).fill = BG_G
        _dat(ws3, 4, min(33, 3 + n_tr), 9)
    _aw(ws3, 9)

    # ─── LOOKAHEAD 6 SEM ───
    ws4 = wb.create_sheet("LOOKAHEAD 6 SEM"); ws4.sheet_properties.tabColor = "0066CC"
    _title(ws4, "LOOKAHEAD 6 SEMANAS", 1, 8)
    _hdr(ws4, 3, ["Semana", "NS Plan.", "Restrições", "Make-Ready", "NS Livres", "NS Exec.", "PPC %", "Ação"])
    ns_por_sem = max(1, n_tr // 12)
    for i in range(6):
        r = i + 4
        restr = max(0, 3 - i)  # restrições diminuem no futuro
        livres = ns_por_sem - restr
        exec_est = round(livres * 0.8, 0)
        ppc = round(exec_est / max(ns_por_sem, 1) * 100, 0)
        ws4.cell(row=r, column=1, value=f"S+{i+1}")
        ws4.cell(row=r, column=2, value=ns_por_sem)
        ws4.cell(row=r, column=3, value=restr)
        ws4.cell(row=r, column=4, value="Material" if restr > 0 else "—")
        ws4.cell(row=r, column=5, value=livres)
        ws4.cell(row=r, column=6, value=int(exec_est))
        ws4.cell(row=r, column=7, value=f"{ppc:.0f}%")
        ws4.cell(row=r, column=7).fill = BG_G if ppc >= 80 else (BG_Y if ppc >= 60 else BG_R)
        ws4.cell(row=r, column=8, value="Solicitar material" if restr > 0 else "OK")
    _dat(ws4, 4, 9, 8); _aw(ws4, 8)

    # ─── BIM 6D ───
    ws5 = wb.create_sheet("BIM 6D - CICLO VIDA"); ws5.sheet_properties.tabColor = "CC0033"
    _title(ws5, "BIM 6D — CICLO DE VIDA 50 ANOS", 1, 8)
    _hdr(ws5, 3, ["Material", "Ext (m)", "Vida Útil", "CO2 (kg)", "Custo Impl.", "Manut 50a", "Total 50a", "Reciclável"])
    vida = {"PVC": 50, "PEAD": 100, "CONCRETO": 80, "PE 80": 100, "PE 100": 100}
    co2k = {"PVC": 3.2, "PEAD": 2.8, "CONCRETO": 12.5, "PE 80": 2.8, "PE 100": 2.8}
    manut_p = {"PVC": 0.005, "PEAD": 0.003, "CONCRETO": 0.01}
    recicl = {"PVC": "Sim", "PEAD": "Sim", "CONCRETO": "Não", "PE 80": "Sim"}

    por_mat = {}
    for t in trechos:
        m = t.get("material", "PVC")
        por_mat.setdefault(m, 0)
        por_mat[m] += t.get("ext_m", 0)

    r = 4
    for m, ext_m in por_mat.items():
        custo_i = ext_m * 910
        manut = custo_i * manut_p.get(m, 0.005) * 50
        ws5.cell(row=r, column=1, value=m)
        ws5.cell(row=r, column=2, value=round(ext_m, 1))
        ws5.cell(row=r, column=3, value=f"{vida.get(m, 50)} anos")
        ws5.cell(row=r, column=4, value=round(ext_m * co2k.get(m, 3.2), 0))
        ws5.cell(row=r, column=5, value=round(custo_i, 0)); ws5.cell(row=r, column=5).number_format = MONEY
        ws5.cell(row=r, column=6, value=round(manut, 0)); ws5.cell(row=r, column=6).number_format = MONEY
        ws5.cell(row=r, column=7, value=round(custo_i + manut, 0)); ws5.cell(row=r, column=7).number_format = MONEY
        ws5.cell(row=r, column=8, value=recicl.get(m, "?"))
        r += 1

    # Total
    ws5.cell(row=r, column=1, value="TOTAL").font = D_BOLD
    for c in [2, 4, 5, 6, 7]:
        ws5.cell(row=r, column=c, value=f"=SUM({get_column_letter(c)}4:{get_column_letter(c)}{r-1})")
        ws5.cell(row=r, column=c).font = D_BOLD
        if c >= 5: ws5.cell(row=r, column=c).number_format = MONEY
    _dat(ws5, 4, r, 8); _aw(ws5, 8)

    chart = PieChart(); chart.title = "CO2 por Material (kg)"; chart.height = 12; chart.width = 14
    data_ref = Reference(ws5, min_col=4, min_row=3, max_row=r-1)
    cats_ref = Reference(ws5, min_col=1, min_row=4, max_row=r-1)
    chart.add_data(data_ref, titles_from_data=True); chart.set_categories(cats_ref)
    ws5.add_chart(chart, "J3")

    wb.save(out_path)
    return out_path


# ══════════════════════════════════════════════════════════
# 2. CURVA S
# ══════════════════════════════════════════════════════════

def gerar_xlsx_curva_s(trechos, nucleo, out_path, custo_metro=910):
    """XLSX Curva S com gráfico previsto × realizado."""
    wb = Workbook()
    ws = wb.active; ws.title = "CURVA S"; ws.sheet_properties.tabColor = "008844"
    _title(ws, f"CURVA S — {nucleo}", 1, 8)
    _sub(ws, _footer(), 2, 8)

    ext = sum(t.get("ext_m", 0) for t in trechos)
    n_meses = max(2, int(ext / (30 * 22)) + 1)
    mensal = ext / n_meses

    _hdr(ws, 4, ["Mês", "Prev. (m)", "Prev. Acum (m)", "Prev. Acum %",
                  "Real. (m)", "Real. Acum (m)", "Real. Acum %", "Custo Acum (R$)"])

    for i in range(n_meses):
        r = i + 5
        acum_prev = mensal * (i + 1)
        pct_prev = acum_prev / ext * 100
        ws.cell(row=r, column=1, value=f"Mês {i+1}")
        ws.cell(row=r, column=2, value=round(mensal, 0))
        ws.cell(row=r, column=3, value=round(acum_prev, 0))
        ws.cell(row=r, column=4, value=round(pct_prev, 1)); ws.cell(row=r, column=4).number_format = '0.0'
        ws.cell(row=r, column=5, value=0).fill = BG_Y  # engenheiro preenche
        ws.cell(row=r, column=6, value=f"=SUM(E5:E{r})")
        ws.cell(row=r, column=7, value=f"=F{r}/{ext}*100"); ws.cell(row=r, column=7).number_format = '0.0'
        ws.cell(row=r, column=8, value=f"=F{r}*{custo_metro}"); ws.cell(row=r, column=8).number_format = MONEY

    end = 4 + n_meses
    _dat(ws, 5, end, 8); _aw(ws, 8)

    # Gráfico Curva S
    chart = LineChart(); chart.title = "Curva S — Previsto × Realizado"; chart.style = 10
    chart.height = 15; chart.width = 25; chart.y_axis.title = "% Acumulado"; chart.x_axis.title = "Mês"
    prev = Reference(ws, min_col=4, min_row=4, max_row=end)
    real = Reference(ws, min_col=7, min_row=4, max_row=end)
    cats = Reference(ws, min_col=1, min_row=5, max_row=end)
    chart.add_data(prev, titles_from_data=True); chart.add_data(real, titles_from_data=True)
    chart.set_categories(cats)
    s0 = chart.series[0]; s0.graphicalProperties.line.solidFill = "008844"; s0.graphicalProperties.line.width = 28000
    if len(chart.series) > 1:
        s1 = chart.series[1]; s1.graphicalProperties.line.solidFill = "FFD000"; s1.graphicalProperties.line.width = 28000
    ws.add_chart(chart, f"A{end + 3}")

    # Resumo
    ws.cell(row=end + 2, column=1, value="Extensão Total:").font = D_BOLD
    ws.cell(row=end + 2, column=2, value=f"{ext:.0f} m")
    ws.cell(row=end + 2, column=3, value="Custo Previsto:").font = D_BOLD
    ws.cell(row=end + 2, column=4, value=f"R$ {ext * custo_metro:,.0f}")

    wb.save(out_path)
    return out_path


# ══════════════════════════════════════════════════════════
# 3. MICROPLANEJAMENTO
# ══════════════════════════════════════════════════════════

def gerar_xlsx_microplan(resultado, pvs, trechos, nucleo, out_path):
    """XLSX 3 abas: Morfologia + Trechos + Material JIT."""
    wb = Workbook()

    # ─── MORFOLOGIA ───
    ws = wb.active; ws.title = "MORFOLOGIA"; ws.sheet_properties.tabColor = "CC8800"
    _title(ws, f"MICRO-PLANEJAMENTO — {nucleo}", 1, 9)
    _sub(ws, _footer(), 2, 9)
    _hdr(ws, 4, ["Morfologia", "Trechos", "Ext (m)", "% Total", "Prod m/dia", "Fator Custo",
                  "Equipes", "Dias", "Custo Frente (R$)"])

    cores = {"planicie": BG_G, "encosta": BG_Y, "morro": BG_R, "mangue": BG_B, "viela": BG_Y}
    morf = resultado.get("por_morfologia", {})
    r = 5
    for tipo, d in morf.items():
        ws.cell(row=r, column=1, value=tipo.upper())
        ws.cell(row=r, column=2, value=d.get("n_trechos", 0))
        ws.cell(row=r, column=3, value=round(d.get("extensao_m", 0), 0))
        ws.cell(row=r, column=4, value=f"{d.get('pct_extensao', 0):.1f}%")
        ws.cell(row=r, column=5, value=d.get("prod_media_m_dia", 0))
        ws.cell(row=r, column=6, value=d.get("fator_custo", 1.0))
        ws.cell(row=r, column=7, value=d.get("equipes_recomendadas", 1))
        ws.cell(row=r, column=8, value=d.get("dias_estimados", 0))
        custo_f = d.get("extensao_m", 0) * 910 * d.get("fator_custo", 1.0)
        ws.cell(row=r, column=9, value=round(custo_f, 0)); ws.cell(row=r, column=9).number_format = MONEY
        fill = cores.get(tipo.lower())
        if fill:
            for c in range(1, 10): ws.cell(row=r, column=c).fill = fill
        r += 1

    # Total
    ws.cell(row=r, column=1, value="TOTAL").font = D_BOLD
    for c in [2, 3, 9]:
        ws.cell(row=r, column=c, value=f"=SUM({get_column_letter(c)}5:{get_column_letter(c)}{r-1})")
        ws.cell(row=r, column=c).font = D_BOLD
    ws.cell(row=r, column=9).number_format = MONEY
    _dat(ws, 5, r, 9); _aw(ws, 9)

    chart = PieChart(); chart.title = "Extensão por Morfologia"; chart.height = 12; chart.width = 14
    data_ref = Reference(ws, min_col=3, min_row=4, max_row=r-1)
    cats_ref = Reference(ws, min_col=1, min_row=5, max_row=r-1)
    chart.add_data(data_ref, titles_from_data=True); chart.set_categories(cats_ref)
    ws.add_chart(chart, "K4")

    # ─── TRECHOS DETALHADO ───
    ws2 = wb.create_sheet("TRECHOS DETALHADO"); ws2.sheet_properties.tabColor = "003366"
    _title(ws2, "TRECHOS COM CLASSIFICAÇÃO", 1, 10)
    _hdr(ws2, 3, ["NS", "PV Ini", "PV Fim", "DN", "Ext (m)", "Material", "Morfologia",
                   "Prod m/dia", "Dias", "Custo (R$)"])
    trechos_plan = resultado.get("trechos", [])
    for i, t in enumerate(trechos_plan[:200], 4):
        ws2.cell(row=i, column=1, value=f"NS_{i-3:03d}")
        ws2.cell(row=i, column=2, value=t.get("pv_ini", ""))
        ws2.cell(row=i, column=3, value=t.get("pv_fim", ""))
        ws2.cell(row=i, column=4, value=t.get("dn_mm", 200))
        ws2.cell(row=i, column=5, value=round(t.get("ext_m", 0), 1))
        ws2.cell(row=i, column=6, value=t.get("material", "PVC"))
        ws2.cell(row=i, column=7, value=t.get("morfologia", ""))
        ws2.cell(row=i, column=8, value=t.get("prod_m_dia", 0))
        ws2.cell(row=i, column=9, value=t.get("duracao_dias", 0))
        ws2.cell(row=i, column=10, value=round(t.get("custo_frente", 0), 0))
        ws2.cell(row=i, column=10).number_format = MONEY
        fill = cores.get(t.get("morfologia", "").lower())
        if fill: ws2.cell(row=i, column=7).fill = fill
    _dat(ws2, 4, min(203, 3 + len(trechos_plan)), 10); _aw(ws2, 10)

    # ─── MATERIAL JIT ───
    ws3 = wb.create_sheet("MATERIAL JIT"); ws3.sheet_properties.tabColor = "0066CC"
    _title(ws3, "MATERIAL JUST-IN-TIME — Consolidado", 1, 5)
    _hdr(ws3, 3, ["Material", "Unidade", "Quantidade", "Preço Unit.", "Total (R$)"])
    materiais = resultado.get("material_consolidado", {})
    r = 4
    for mat, d in materiais.items():
        ws3.cell(row=r, column=1, value=mat)
        ws3.cell(row=r, column=2, value=d.get("un", ""))
        ws3.cell(row=r, column=3, value=round(d.get("qtd", 0), 2))
        pu = d.get("preco_unit", 0)
        ws3.cell(row=r, column=4, value=round(pu, 2)); ws3.cell(row=r, column=4).number_format = MONEY
        ws3.cell(row=r, column=5, value=round(d.get("qtd", 0) * pu, 0)); ws3.cell(row=r, column=5).number_format = MONEY
        r += 1
    _dat(ws3, 4, r - 1, 5); _aw(ws3, 5)

    wb.save(out_path)
    return out_path


# ══════════════════════════════════════════════════════════
# 4. CUSTOS DETALHADO
# ══════════════════════════════════════════════════════════

def gerar_xlsx_custos(pvs, trechos, nucleo, out_path, custo_metro=910):
    """XLSX 2 abas: Detalhamento por trecho + Composição."""
    wb = Workbook()

    ws = wb.active; ws.title = "CUSTOS POR TRECHO"; ws.sheet_properties.tabColor = "003366"
    _title(ws, f"DETALHAMENTO DE CUSTOS — {nucleo}", 1, 12)
    _sub(ws, _footer(), 2, 12)
    _hdr(ws, 4, ["NS", "PV Ini", "PV Fim", "DN", "Ext (m)", "Escavação", "Tubo", "PV/Cx",
                  "Reaterro", "Pavim.", "Ramal", "TOTAL"])

    for i, t in enumerate(trechos[:300], 5):
        ext = t.get("ext_m", 0)
        ws.cell(row=i, column=1, value=f"NS_{i-4:03d}")
        ws.cell(row=i, column=2, value=t.get("pv_ini", ""))
        ws.cell(row=i, column=3, value=t.get("pv_fim", ""))
        ws.cell(row=i, column=4, value=t.get("dn_mm", 200))
        ws.cell(row=i, column=5, value=round(ext, 1))
        ws.cell(row=i, column=6, value=round(ext * 145, 0)); ws.cell(row=i, column=6).number_format = MONEY
        ws.cell(row=i, column=7, value=round(ext * 240, 0)); ws.cell(row=i, column=7).number_format = MONEY
        ws.cell(row=i, column=8, value=round(ext * 120, 0)); ws.cell(row=i, column=8).number_format = MONEY
        ws.cell(row=i, column=9, value=round(ext * 80, 0)); ws.cell(row=i, column=9).number_format = MONEY
        ws.cell(row=i, column=10, value=round(ext * 45, 0)); ws.cell(row=i, column=10).number_format = MONEY
        ws.cell(row=i, column=11, value=round(ext * 65, 0)); ws.cell(row=i, column=11).number_format = MONEY
        ws.cell(row=i, column=12, value=f"=SUM(F{i}:K{i})*1.25"); ws.cell(row=i, column=12).number_format = MONEY

    end = 4 + len(trechos[:300])
    # Totais
    r = end + 1
    ws.cell(row=r, column=1, value="TOTAL").font = D_BOLD
    for c in range(5, 13):
        ws.cell(row=r, column=c, value=f"=SUM({get_column_letter(c)}5:{get_column_letter(c)}{end})")
        ws.cell(row=r, column=c).font = D_BOLD; ws.cell(row=r, column=c).number_format = MONEY

    _dat(ws, 5, end, 12); _aw(ws, 12)

    # ─── COMPOSIÇÃO ───
    ws2 = wb.create_sheet("COMPOSICAO R$ metro"); ws2.sheet_properties.tabColor = "CC8800"
    _title(ws2, "COMPOSIÇÃO DE CUSTO POR METRO", 1, 4)
    _hdr(ws2, 4, ["Serviço", "R$/metro", "% do Total", "Obs"])
    comp = [("Escavação mecânica", 145), ("Tubo ESG (DN200/300)", 240), ("Tubo AG (PEAD)", 95),
            ("PV/Caixas/PIs", 120), ("Reaterro e compactação", 80), ("Ramal predial", 65),
            ("Pavimentação CBUQ", 45), ("Sinalização", 15)]
    sub = sum(v for _, v in comp)
    for i, (srv, val) in enumerate(comp, 5):
        ws2.cell(row=i, column=1, value=srv)
        ws2.cell(row=i, column=2, value=val); ws2.cell(row=i, column=2).number_format = MONEY
        ws2.cell(row=i, column=3, value=f"{val/sub*100:.1f}%")
    r = 5 + len(comp)
    ws2.cell(row=r, column=1, value="Subtotal").font = D_BOLD
    ws2.cell(row=r, column=2, value=sub).font = D_BOLD; ws2.cell(row=r, column=2).number_format = MONEY
    ws2.cell(row=r+1, column=1, value="BDI 25%").font = D_BOLD
    ws2.cell(row=r+1, column=2, value=round(sub*0.25)).font = D_BOLD
    ws2.cell(row=r+2, column=1, value="TOTAL c/ BDI").font = Font(name="Arial", size=12, bold=True, color="CC0033")
    ws2.cell(row=r+2, column=2, value=round(sub*1.25)).font = Font(name="Arial", size=12, bold=True, color="CC0033")
    ws2.cell(row=r+2, column=2).number_format = MONEY
    _dat(ws2, 5, r+2, 4); _aw(ws2, 4)

    chart = PieChart(); chart.title = "Composição R$/m"; chart.height = 12; chart.width = 14
    data_ref = Reference(ws2, min_col=2, min_row=4, max_row=r-1)
    cats_ref = Reference(ws2, min_col=1, min_row=5, max_row=r-1)
    chart.add_data(data_ref, titles_from_data=True); chart.set_categories(cats_ref)
    ws2.add_chart(chart, "F4")

    wb.save(out_path)
    return out_path


# ══════════════════════════════════════════════════════════
# 5. HIDRÁULICA
# ══════════════════════════════════════════════════════════

def gerar_xlsx_hidraulica(trechos_enr, pvs, nucleo, out_path):
    """XLSX Manning completo com alertas coloridos."""
    wb = Workbook()
    ws = wb.active; ws.title = "HIDRÁULICA"; ws.sheet_properties.tabColor = "0066CC"
    _title(ws, f"VERIFICAÇÃO HIDRÁULICA — {nucleo}", 1, 12)
    _sub(ws, _footer(), 2, 12)
    _hdr(ws, 4, ["NS", "PV Ini", "PV Fim", "DN (mm)", "Ext (m)", "Decl (‰)",
                  "n Manning", "V (m/s)", "Q (L/s)", "τ (Pa)", "Status", "Ação"])

    ok = warn = err = 0
    for i, t in enumerate(trechos_enr[:500], 5):
        v = t.get("v_ms", 0)
        tau = t.get("tau_pa", 0)
        q = t.get("q_ls", 0)

        if v > 0:
            problemas = []
            if v > 5: problemas.append("V>5 erosão")
            if 0 < v < 0.6: problemas.append("V<0.6 sedimentação")
            if 0 < tau < 1: problemas.append("τ<1Pa fora NBR")
            if t.get("decl_mm", 0) < 0: problemas.append("Decl negativa")

            if not problemas:
                status = "✅ OK"; ok += 1
                font = OK_FONT; fill = BG_G; acao = "Nenhuma"
            else:
                status = "⚠️ VERIFICAR"; warn += 1
                font = WARN_FONT; fill = BG_Y; acao = "; ".join(problemas)
                if any("erosão" in p or "NBR" in p for p in problemas):
                    status = "❌ CRÍTICO"; err += 1; warn -= 1
                    font = ERR_FONT; fill = BG_R
        else:
            status = "— SEM DADOS"; font = D_FONT; fill = None; acao = "Calcular"

        ws.cell(row=i, column=1, value=f"NS_{i-4:03d}")
        ws.cell(row=i, column=2, value=t.get("pv_ini", ""))
        ws.cell(row=i, column=3, value=t.get("pv_fim", ""))
        ws.cell(row=i, column=4, value=t.get("dn_mm", 0))
        ws.cell(row=i, column=5, value=round(t.get("ext_m", 0), 1))
        ws.cell(row=i, column=6, value=round(t.get("decl_mm", 0), 2))
        ws.cell(row=i, column=7, value=0.013)
        ws.cell(row=i, column=8, value=round(v, 3))
        ws.cell(row=i, column=9, value=round(q, 3))
        ws.cell(row=i, column=10, value=round(tau, 2))
        ws.cell(row=i, column=11, value=status); ws.cell(row=i, column=11).font = font
        ws.cell(row=i, column=12, value=acao)
        if fill:
            ws.cell(row=i, column=11).fill = fill

    end = 4 + len(trechos_enr[:500])
    _dat(ws, 5, end, 12); _aw(ws, 12)

    # Resumo
    r = end + 2
    ws.cell(row=r, column=1, value="RESUMO").font = D_BOLD
    ws.cell(row=r+1, column=1, value="OK"); ws.cell(row=r+1, column=2, value=ok); ws.cell(row=r+1, column=1).fill = BG_G
    ws.cell(row=r+2, column=1, value="Verificar"); ws.cell(row=r+2, column=2, value=warn); ws.cell(row=r+2, column=1).fill = BG_Y
    ws.cell(row=r+3, column=1, value="Crítico"); ws.cell(row=r+3, column=2, value=err); ws.cell(row=r+3, column=1).fill = BG_R

    wb.save(out_path)
    return out_path


# ══════════════════════════════════════════════════════════
# 6. PERDAS
# ══════════════════════════════════════════════════════════

def gerar_xlsx_perdas(relatorio, nucleo, out_path):
    """XLSX Gestão de Perdas: Balanço + UARL + ILI + Risco + DMAs."""
    wb = Workbook()
    uarl = relatorio.get("uarl", {})
    bal = relatorio.get("balanco_hidrico")
    ili = relatorio.get("ili")
    risco = relatorio.get("risco", {})
    dmas = relatorio.get("dmas", {})
    proj = relatorio.get("projecao_economica", {})

    ws = wb.active; ws.title = "INDICADORES"; ws.sheet_properties.tabColor = "0066CC"
    _title(ws, f"GESTÃO DE PERDAS — {nucleo}", 1, 6)
    _sub(ws, _footer(), 2, 6)
    _hdr(ws, 4, ["Indicador", "Valor", "Unidade", "Classificação", "Meta", "Obs"])

    dados = [
        ("UARL", f"{uarl.get('uarl_litros_dia', 0):,.0f}", "L/dia", "", "", "Mínimo técnico"),
        ("UARL Anual", f"{uarl.get('uarl_m3_ano', 0):,.0f}", "m³/ano", "", "", ""),
        ("Comp. Rede", f"{uarl.get('componentes', {}).get('rede', 0):,.0f}", "L/dia", "", "", "18×km×P"),
        ("Comp. Conexões", f"{uarl.get('componentes', {}).get('conexoes', 0):,.0f}", "L/dia", "", "", "0.8×N×P"),
        ("Comp. Ramais", f"{uarl.get('componentes', {}).get('ramais', 0):,.0f}", "L/dia", "", "", "25×km×P"),
    ]
    if ili:
        dados.append(("ILI", f"{ili.get('ili', 0):.2f}", "", ili.get("classificacao", ""), "< 4", ""))
        dados.append(("Potencial Redução", f"{ili.get('potencial_reducao_pct', 0)}%", "", "", "", ""))
    if bal:
        dados.append(("NRW", f"{bal.get('nrw_pct', 0):.1f}", "%", "", "< 25%", ""))
        dados.append(("Perdas Reais", f"{bal.get('perdas_reais_pct', 0):.1f}", "%", "", "", ""))
        dados.append(("Receita Perdida", f"{bal.get('receita_perdida_rs', 0):,.0f}", "R$/mês", "", "", ""))

    dados.append(("Custo Perdas/Ano", f"{proj.get('custo_perdas_ano_estimado', 0):,.0f}", "R$/ano", "", "", ""))
    dados.append(("Energia Perdida", f"{proj.get('energia_perdida_kwh_ano', 0):,.0f}", "kWh/ano", "", "", ""))
    dados.append(("N° DMAs", dmas.get("n_dmas", 0), "setores", "", "", ""))
    dados.append(("Risco Total/Ano", f"{risco.get('resumo', {}).get('custo_risco_total_ano', 0):,.0f}", "R$/ano", "", "", ""))

    for r, row_data in enumerate(dados, 5):
        for c, v in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=v)
        # ILI color
        if row_data[0] == "ILI":
            cls = ili.get("classificacao", "") if ili else ""
            if "EXCELENTE" in cls or "BOM" in cls: ws.cell(row=r, column=4).font = OK_FONT; ws.cell(row=r, column=4).fill = BG_G
            elif "REGULAR" in cls: ws.cell(row=r, column=4).font = WARN_FONT; ws.cell(row=r, column=4).fill = BG_Y
            else: ws.cell(row=r, column=4).font = ERR_FONT; ws.cell(row=r, column=4).fill = BG_R

    _dat(ws, 5, 5 + len(dados) - 1, 6); _aw(ws, 6)

    # ─── ABA RISCO ───
    ws2 = wb.create_sheet("RISCO RUPTURA"); ws2.sheet_properties.tabColor = "CC0033"
    _title(ws2, "TOP 10 TRECHOS MAIS CRÍTICOS", 1, 8)
    _hdr(ws2, 3, ["NS", "Material", "DN", "Ext (m)", "Idade", "Taxa Falha", "Prob. Ruptura", "Nível"])
    top10 = risco.get("top10", [])
    for i, t in enumerate(top10[:10], 4):
        ws2.cell(row=i, column=1, value=t.get("ns", ""))
        ws2.cell(row=i, column=2, value=t.get("material", ""))
        ws2.cell(row=i, column=3, value=t.get("dn_mm", 0))
        ws2.cell(row=i, column=4, value=round(t.get("ext_m", 0), 1))
        ws2.cell(row=i, column=5, value=t.get("idade_anos", 0))
        ws2.cell(row=i, column=6, value=t.get("taxa_falha_km_ano", 0))
        ws2.cell(row=i, column=7, value=f"{t.get('prob_ruptura_ano', 0)*100:.2f}%")
        nivel = t.get("nivel_risco", "")
        ws2.cell(row=i, column=8, value=nivel)
        if "CRÍTICO" in nivel or "ALTO" in nivel: ws2.cell(row=i, column=8).fill = BG_R; ws2.cell(row=i, column=8).font = ERR_FONT
        elif "MÉDIO" in nivel: ws2.cell(row=i, column=8).fill = BG_Y
        else: ws2.cell(row=i, column=8).fill = BG_G
    _dat(ws2, 4, min(13, 3 + len(top10)), 8); _aw(ws2, 8)

    wb.save(out_path)
    return out_path


# ══════════════════════════════════════════════════════════
# MAIN — testa todos
# ══════════════════════════════════════════════════════════

if __name__ == "__main__":
    import sys
    from pathlib import Path
    sys.path.insert(0, str(Path(__file__).parent))

    print("Testando geração de XLSX...")

    pvs = {"PV01": {"x": 0, "y": 0, "ct": 5, "cf": 3, "tipo": "esgoto"},
           "PV02": {"x": 20, "y": 0, "ct": 4.8, "cf": 2.8, "tipo": "esgoto"}}
    trechos = [{"pv_ini": "PV01", "pv_fim": "PV02", "dn_mm": 200, "ext_m": 20,
                "decl_mm": 10, "material": "PVC", "tipo": "esgoto", "v_ms": 0.89, "q_ls": 28, "tau_pa": 4.9}]

    # Curva S
    gerar_xlsx_curva_s(trechos, "Teste", "/tmp/CURVA_S_TESTE.xlsx")
    print("  ✅ Curva S")

    # Custos
    gerar_xlsx_custos(pvs, trechos, "Teste", "/tmp/CUSTOS_TESTE.xlsx")
    print("  ✅ Custos")

    # Hidráulica
    gerar_xlsx_hidraulica(trechos, pvs, "Teste", "/tmp/HIDRAULICA_TESTE.xlsx")
    print("  ✅ Hidráulica")

    print("\n✅ Todos os XLSX gerados com sucesso!")
