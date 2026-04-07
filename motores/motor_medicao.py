#!/usr/bin/env python3
"""
MOTOR_MEDICAO.PY — Motor de Medição e Acompanhamento
ConstruData - HydroNetwork · CT 11481051 · FCN Construções e Saneamento

Fluxo: NS executada → RDO → BM mensal → Curva S → Pagamento
Condição pagamento: cadastro NTS 292 aprovado (contrato pág. 64)

Integra com:
- Execução_Geral.xlsx (dados reais diários por núcleo)
- motor_custo.py (preços do contrato)
- gerar_ns.py (notas de serviço)
"""

import json, math, os
from datetime import datetime, timedelta
from collections import defaultdict
from pathlib import Path


def carregar_execucao_xlsx(path):
    """Carrega dados de execução diária do Excel (Execução_Geral.xlsx)."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    
    dados = {}
    for sname in wb.sheetnames:
        if sname in ("APOIO", "Planilha1"):
            continue
        ws = wb[sname]
        registros = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            r = list(row)
            if i >= 4 and r[1]:
                try:
                    dia = str(r[1])[:10]
                    equipe = str(r[5] or "").strip()
                    rua = str(r[6] or "").strip()
                    la = float(r[7]) if r[7] else 0    # ligações água
                    le = float(r[8]) if r[8] else 0    # ligações esgoto
                    pra = float(r[9]) if r[9] else 0
                    if equipe or la or le or rua:
                        registros.append({
                            "dia": dia, "equipe": equipe, "rua": rua,
                            "lig_agua": la, "lig_esg": le, "pra": pra,
                        })
                except:
                    pass
        dados[sname] = registros
    wb.close()
    return dados


def carregar_execucao_json(path):
    """Carrega dados de execução de JSON exportado."""
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def gerar_resumo_execucao(dados_exec):
    """Gera resumo de execução por núcleo, mês e total."""
    resumo = {
        "por_nucleo": {},
        "por_mes": {},
        "totais": {"dias_trabalhados": 0, "lig_agua": 0, "lig_esg": 0, "lig_total": 0},
    }
    
    for nucleo, registros in dados_exec.items():
        dias = len(registros)
        la = sum(r.get("lig_agua", 0) for r in registros)
        le = sum(r.get("lig_esg", 0) for r in registros)
        
        resumo["por_nucleo"][nucleo] = {
            "dias_trabalhados": dias,
            "lig_agua": la, "lig_esg": le, "lig_total": la + le,
            "prod_dia_agua": round(la / dias, 1) if dias else 0,
            "prod_dia_esg": round(le / dias, 1) if dias else 0,
        }
        
        resumo["totais"]["dias_trabalhados"] += dias
        resumo["totais"]["lig_agua"] += la
        resumo["totais"]["lig_esg"] += le
        resumo["totais"]["lig_total"] += la + le
        
        # Por mês
        for r in registros:
            mes = r.get("dia", "")[:7]  # YYYY-MM
            if mes not in resumo["por_mes"]:
                resumo["por_mes"][mes] = {"lig_agua": 0, "lig_esg": 0, "dias": 0}
            resumo["por_mes"][mes]["lig_agua"] += r.get("lig_agua", 0)
            resumo["por_mes"][mes]["lig_esg"] += r.get("lig_esg", 0)
            resumo["por_mes"][mes]["dias"] += 1
    
    return resumo


def gerar_curva_s(trechos, dados_exec=None, custo_metro=910.0):
    """
    Gera dados para Curva S: previsto × realizado.
    
    Se dados_exec fornecidos, usa execução real.
    Senão, distribui linearmente pelo cronograma.
    """
    ext_total = sum(t.get("ext_m", 0) for t in trechos)
    custo_total = ext_total * custo_metro
    
    # Previsto: distribuição linear em N meses
    n_meses = max(1, int(math.ceil(ext_total / 2000)))  # ~2km/mês por frente
    previsto = []
    for m in range(n_meses):
        pct = (m + 1) / n_meses
        previsto.append({
            "mes": m + 1,
            "pct_acum": round(pct * 100, 1),
            "custo_acum": round(pct * custo_total, 0),
            "ext_acum": round(pct * ext_total, 0),
        })
    
    # Realizado: se tem dados de execução
    realizado = []
    if dados_exec:
        acum_lig = 0
        meses_exec = defaultdict(lambda: {"lig": 0, "dias": 0})
        
        for nucleo, registros in dados_exec.items():
            for r in registros:
                mes = r.get("dia", "")[:7]
                lig = r.get("lig_agua", 0) + r.get("lig_esg", 0)
                meses_exec[mes]["lig"] += lig
                meses_exec[mes]["dias"] += 1
        
        for mes in sorted(meses_exec.keys()):
            acum_lig += meses_exec[mes]["lig"]
            # Estimativa: 6.1m por ligação (dado ML)
            ext_acum = acum_lig * 6.1
            pct = (ext_acum / ext_total * 100) if ext_total else 0
            realizado.append({
                "mes": mes,
                "lig_mes": meses_exec[mes]["lig"],
                "lig_acum": acum_lig,
                "ext_acum_est": round(ext_acum, 0),
                "pct_acum": round(min(pct, 100), 1),
                "custo_acum_est": round(ext_acum * custo_metro, 0),
            })
    
    return {
        "ext_total": ext_total,
        "custo_total": custo_total,
        "custo_metro": custo_metro,
        "previsto": previsto,
        "realizado": realizado,
    }


def vincular_ns_execucao(trechos, dados_exec):
    """
    Tenta vincular trechos (NS) com dados de execução por rua.
    Retorna trechos com status atualizado.
    """
    # Extrair todas as ruas executadas
    ruas_exec = set()
    for nucleo, registros in dados_exec.items():
        for r in registros:
            rua = r.get("rua", "").strip().upper()
            if rua:
                ruas_exec.add(rua)
    
    for tr in trechos:
        # Se o trecho tem rua e ela aparece nos dados de execução
        rua_tr = str(tr.get("rua", "")).strip().upper()
        if rua_tr and rua_tr in ruas_exec:
            tr["status"] = "executado"
            tr["origem_status"] = "Execução_Geral.xlsx"
    
    return trechos


def gerar_boletim_medicao(trechos_exec, pvs, periodo, bm_num, custo_func=None):
    """
    Gera BM formal a partir de trechos executados.
    
    Args:
        trechos_exec: trechos com status='executado'
        pvs: dict de PVs
        periodo: "Mar/2026"
        bm_num: número do BM
        custo_func: função de custo (default: motor_custo.custo_trecho)
    """
    try:
        from motor_custo import custo_trecho
        calc = custo_func or custo_trecho
    except ImportError:
        calc = lambda tr, pvs, **kw: {"total": tr.get("ext_m", 0) * 910}
    
    itens = []
    total = 0
    
    for i, tr in enumerate(trechos_exec):
        c = calc(tr, pvs)
        has_cadastro = tr.get("ct_ini_real") or tr.get("cadastro_ok")
        
        itens.append({
            "ns_id": i + 1,
            "ns": f"{tr.get('pv_ini','?')}→{tr.get('pv_fim','?')}",
            "dn_mm": tr.get("dn_mm"),
            "ext_m": tr.get("ext_m"),
            "tipo": tr.get("tipo", "esgoto"),
            "custo": c.get("total", 0) if isinstance(c, dict) else c,
            "cadastro": "OK" if has_cadastro else "PENDENTE",
            "liberado_pgto": "SIM" if has_cadastro else "NÃO (s/ cadastro)",
        })
        if has_cadastro:
            total += c.get("total", 0) if isinstance(c, dict) else c
    
    return {
        "bm_numero": bm_num,
        "periodo": periodo,
        "empresa": "FCN Construções e Saneamento",
        "contrato": "CT 11481051 — SLNR Santos",
        "data_emissao": datetime.now().strftime("%d/%m/%Y"),
        "n_trechos_exec": len(trechos_exec),
        "n_trechos_liberados": sum(1 for it in itens if it["cadastro"] == "OK"),
        "extensao_exec": round(sum(t.get("ext_m", 0) for t in trechos_exec), 1),
        "itens": itens,
        "total_liberado": round(total, 2),
        "total_pendente_cadastro": round(sum(it["custo"] for it in itens if it["cadastro"] != "OK"), 2),
        "obs": "Pagamento condicionado ao cadastro NTS 292 aprovado (contrato pág. 64)",
    }


def gerar_acompanhamento_semanal(dados_exec, semana_inicio, semana_fim):
    """Gera resumo semanal de produção."""
    resumo = {}
    for nucleo, registros in dados_exec.items():
        total_la = 0
        total_le = 0
        dias = 0
        for r in registros:
            dia = r.get("dia", "")[:10]
            if semana_inicio <= dia <= semana_fim:
                total_la += r.get("lig_agua", 0)
                total_le += r.get("lig_esg", 0)
                dias += 1
        if dias > 0:
            resumo[nucleo] = {
                "dias": dias,
                "lig_agua": total_la,
                "lig_esg": total_le,
                "lig_total": total_la + total_le,
                "prod_dia": round((total_la + total_le) / dias, 1),
            }
    return {
        "periodo": f"{semana_inicio} a {semana_fim}",
        "nucleos": resumo,
        "total_lig": sum(r["lig_total"] for r in resumo.values()),
    }


def gerar_xlsx_curva_s(trechos, nucleo, out_path, custo_metro=910):
    """XLSX Curva S com grafico previsto x realizado."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, Reference
    from openpyxl.utils import get_column_letter

    HF = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    HN = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    TF = Font(name="Arial", size=14, bold=True, color="003366")
    BD = Border(*(Side(style='thin', color='CCCCCC'),)*4)

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "CURVA S"
    ws.merge_cells('A1:H1'); ws['A1'] = f"CURVA S — {nucleo}"; ws['A1'].font = TF
    ws.merge_cells('A2:H2'); ws['A2'] = "FCN Construcoes e Saneamento"; ws['A2'].font = Font(size=10, color="666666")

    headers = ["Mes","Previsto(m)","Prev Acum(m)","Prev Acum(%)","Realizado(m)","Real Acum(m)","Real Acum(%)","Custo Acum(R$)"]
    for c, h in enumerate(headers, 1):
        cl = ws.cell(row=4, column=c, value=h); cl.font=HN; cl.fill=HF; cl.border=BD
        cl.alignment=Alignment(horizontal='center', wrap_text=True)

    ext_total = sum(t.get("ext_m", 0) for t in trechos)
    meses = max(1, int(ext_total / (30 * 22)))
    prev_m = ext_total / meses

    for r in range(meses):
        row = r + 5
        ws.cell(row=row, column=1, value=f"Mes {r+1}")
        ws.cell(row=row, column=2, value=round(prev_m))
        ws.cell(row=row, column=3, value=f"=SUM(B5:B{row})")
        ws.cell(row=row, column=4, value=f"=C{row}/{ext_total}*100")
        ws.cell(row=row, column=5, value=0)  # engenheiro preenche
        ws.cell(row=row, column=6, value=f"=SUM(E5:E{row})")
        ws.cell(row=row, column=7, value=f"=F{row}/{ext_total}*100")
        ws.cell(row=row, column=8, value=f"=F{row}*{custo_metro}")
        ws.cell(row=row, column=8).number_format = '#,##0'
        for c in range(1, 9):
            ws.cell(row=row, column=c).border = BD

    end_row = 4 + meses
    chart = LineChart(); chart.title = "Curva S — Previsto x Realizado"
    chart.y_axis.title = "% Acumulado"; chart.x_axis.title = "Mes"
    chart.height = 15; chart.width = 25
    prev = Reference(ws, min_col=4, min_row=4, max_row=end_row)
    real = Reference(ws, min_col=7, min_row=4, max_row=end_row)
    cats = Reference(ws, min_col=1, min_row=5, max_row=end_row)
    chart.add_data(prev, titles_from_data=True)
    chart.add_data(real, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.line.solidFill = "008844"
    chart.series[0].graphicalProperties.line.width = 25000
    if len(chart.series) > 1:
        chart.series[1].graphicalProperties.line.solidFill = "FFD000"
        chart.series[1].graphicalProperties.line.width = 25000
    ws.add_chart(chart, f"A{end_row + 3}")

    for c in range(1, 9):
        mx = max((len(str(cl.value or "")) for cl in ws[get_column_letter(c)]), default=8)
        ws.column_dimensions[get_column_letter(c)].width = min(mx+4, 30)
    wb.save(str(out_path))
    return str(out_path)


if __name__ == "__main__":
    # Teste com dados reais
    if os.path.exists("/home/claude/EXECUCAO_DIARIA.json"):
        dados = carregar_execucao_json("/home/claude/EXECUCAO_DIARIA.json")
        resumo = gerar_resumo_execucao(dados)
        
        print("RESUMO EXECUÇÃO:")
        for nuc, r in resumo["por_nucleo"].items():
            print(f"  {nuc:15s} | {r['dias_trabalhados']:3d}d | AG={r['lig_agua']:.0f} ESG={r['lig_esg']:.0f} | prod={r['prod_dia_agua']:.1f}+{r['prod_dia_esg']:.1f}/d")
        
        print(f"\n  TOTAL: {resumo['totais']['lig_total']:.0f} ligações")
