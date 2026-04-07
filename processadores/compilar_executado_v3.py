#!/usr/bin/env python3
"""
COMPILAR v3 CORRIGIDO — SEM DUPLICAÇÕES + SEM CADASTRO
=======================================================
Regras de filtragem:
1. SEM projetos duplicados (JC/SM só NOMES, Teteu só v2, Prol Teteu só 1)
2. SM: Excluir network "1" (cadastro SABESP ~5.3km)
3. DXFs: DN>=300mm pode ser coletor tronco existente — MARCAR como "CADASTRO?"
4. Fator de ajuste por núcleo contra km reais
"""
import sys, json, math, re, warnings, os
from pathlib import Path
from datetime import datetime
from collections import defaultdict, Counter
import unicodedata

warnings.filterwarnings("ignore")
sys.path.insert(0, str(Path(r"C:\Users\felip\Downloads\NOVA NS Versao 5")))
from gerar_ns import calc_manning, CONTRATO

import numpy as np
import geopandas as gpd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path(r"C:\Users\felip\Desktop\GERAR NS COM ITENS POR RUA DE PV A PV\SAIDA_COMPLETA_POR_RUA")
BASE_NOMES = Path(r"C:\Users\felip\Desktop\GERAR NS COM ITENS POR RUA DE PV A PV\SAIDA_POR_RUA_COM_NOMES")
SHP_EXEC = Path(r"C:\Users\felip\Downloads\PROJETOS DE ÁGUA E ESGOTO - DWG E DXF 2018\MAPAS ÁGUA E ESGOTO PARA DXF\linhas executado.shp")
EXEC_DIR = Path(r"C:\Users\felip\Desktop\executado desde o inicio")
OUTPUT = BASE / "_COMPILADO_V3"

# KM REAIS de ESGOTO por núcleo
KM_REAIS_ESG = {
    "Morro do Teteu v2": 4.0, "Joao Carlos": 1.3, "Vila Criadores": 2.3,
    "Sao Manoel": 2.3, "Pantanal Baixo": 2.3, "Vila Israel": 1.4,
}

def limpar(rua):
    if not rua or str(rua).strip() in ("","nan","None","Sem Rua","1"):
        return "Sem Rua"
    t = str(rua).replace("\n"," ").replace("\r"," ")
    t = re.sub(r'[\x00-\x1f\x7f]','',t)
    return re.sub(r'\s+',' ',t).strip() or "Sem Rua"

def carregar_shp():
    cpg = SHP_EXEC.with_suffix(".cpg")
    cpg.write_text("latin-1", encoding="utf-8")
    os.environ["SHAPE_ENCODING"] = "latin-1"
    try: gdf = gpd.read_file(str(SHP_EXEC))
    except:
        import pyogrio
        gdf = pyogrio.read_dataframe(str(SHP_EXEC), encoding="latin-1")
    gdf['length_m'] = gdf.geometry.length
    linhas = []
    for _, row in gdf.iterrows():
        g = row.geometry
        if g is None: continue
        mid = g.interpolate(0.5, normalized=True)
        dn = 200
        layer = str(row.get('GM_LAYER',''))
        if 'DN 150' in layer: dn = 150
        elif 'DN 200' in layer: dn = 200
        elif 'DN 300' in layer: dn = 300
        elif 'DN 63' in layer: dn = 63
        else:
            d = row.get('Diametro')
            if d:
                d = float(d)
                dn = round(d*1000) if d < 1 else round(d) if d < 1000 else 200
        mat = str(row.get('material','PVC'))
        tipo = 'AGUA' if 'agua' in str(row.get('tipo','')).lower() or 'PEAD' in mat else 'ESGOTO'
        linhas.append({
            'mx': mid.x, 'my': mid.y,
            'sx': g.coords[0][0], 'sy': g.coords[0][1],
            'ex': g.coords[-1][0], 'ey': g.coords[-1][1],
            'length': row['length_m'], 'dn': dn, 'material': mat,
            'tipo': tipo, 'data': str(row.get('Data_Inst',''))
        })
    return linhas

def carregar_projeto(pasta):
    ns_campo = pasta / "01_NS_CAMPO"
    if not ns_campo.exists(): return {}, []
    pvs = {}; trechos = []
    for ns_dir in sorted(ns_campo.iterdir()):
        if not ns_dir.is_dir(): continue
        for f in ns_dir.glob("*_DADOS.json"):
            try:
                with open(f, encoding="utf-8") as fh:
                    d = json.load(fh)
                t = d.get("trecho",{})
                if t.get("pv_ini") and t.get("pv_fim"):
                    t["rua"] = limpar(t.get("rua"))
                    trechos.append(t)
                    if d.get("pv_montante"): pvs[t["pv_ini"]] = d["pv_montante"]
                    if d.get("pv_jusante"): pvs[t["pv_fim"]] = d["pv_jusante"]
            except: pass
    return pvs, trechos

def e_cadastro(trecho, nucleo):
    """Determina se um trecho provavelmente é CADASTRO (rede existente) e não obra nova."""
    network = str(trecho.get("network_name",""))
    dn = trecho.get("dn_mm") or 200
    ext = trecho.get("ext_m", 0)
    
    # São Manoel: PVs com "(1)" no nome = network "1" = cadastro SABESP
    if "Sao Manoel" in nucleo:
        pv_ini = str(trecho.get("pv_ini",""))
        pv_fim = str(trecho.get("pv_fim",""))
        if pv_ini.endswith("(1)") or pv_fim.endswith("(1)"):
            return True, "CADASTRO — Rede '1' do XML (cadastro SABESP existente)"
    
    # Prolongamentos: DN >= 300mm = provável coletor tronco existente
    # Pantanal/Israel: DN >= 300mm = provável coletor tronco existente
    if dn >= 300 and "AGUA" not in nucleo:
        if any(k in nucleo for k in ["Prolong","Pantanal","Israel","Teteu"]):
            return True, f"CADASTRO? — DN{dn}mm = coletor tronco (provável rede existente)"
    
    return False, ""

def match_shp(trecho, pvs, linhas, tol=15.0):
    p0 = pvs.get(trecho.get("pv_ini"),{})
    p1 = pvs.get(trecho.get("pv_fim"),{})
    x0,y0 = p0.get("x",0), p0.get("y",0)
    x1,y1 = p1.get("x",0), p1.get("y",0)
    if x0==0 or x1==0: return False, None
    mx=(x0+x1)/2; my=(y0+y1)/2
    best_d=float('inf'); best=None
    for l in linhas:
        d = math.hypot(l['mx']-mx, l['my']-my)
        if d < best_d: best_d=d; best=l
    if best_d <= tol: return True, best
    for l in linhas:
        d1 = math.hypot(l['sx']-x0, l['sy']-y0)
        d2 = math.hypot(l['ex']-x1, l['ey']-y1)
        if d1<=tol and d2<=tol: return True, l
        d1b = math.hypot(l['sx']-x1, l['sy']-y1)
        d2b = math.hypot(l['ex']-x0, l['ey']-y0)
        if d1b<=tol and d2b<=tol: return True, l
    return False, None

def main():
    try: sys.stdout.reconfigure(encoding="utf-8",errors="replace")
    except: pass
    print("="*72)
    print("  COMPILAR v3 CORRIGIDO — SEM DUPLICAÇÕES + SEM CADASTRO")
    print("="*72)
    OUTPUT.mkdir(parents=True,exist_ok=True)

    print("\n[1] Carregando SHP executado...")
    linhas = carregar_shp()
    print(f"  SHP: {len(linhas)} linhas")

    # PROJETOS LIMPOS (sem duplicação!)
    PROJETOS = [
        ("Joao Carlos","ESGOTO",BASE_NOMES/"Joao Carlos"),
        ("Sao Manoel","ESGOTO",BASE_NOMES/"Sao Manoel"),
        ("Morro do Teteu v2","ESGOTO",BASE/"Morro do Teteu v2_ESGOTO"),
        ("Vila Criadores","ESGOTO",BASE/"Vila Criadores_ESGOTO"),
        ("Vila Israel","ESGOTO",BASE/"Vila Israel_ESGOTO"),
        ("Pantanal Baixo","ESGOTO",BASE/"Pantanal Baixo_ESGOTO"),
        ("Morro do Teteu","AGUA",BASE/"Morro do Teteu_AGUA"),
        ("Vila Criadores","AGUA",BASE/"Vila Criadores_AGUA"),
        ("Vila Israel","AGUA",BASE/"Vila Israel_AGUA"),
        ("Pantanal Baixo","AGUA",BASE/"Pantanal Baixo_AGUA"),
        ("Prolongamento Criadores","ESGOTO",BASE/"Prolongamento Criadores_ESGOTO"),
        ("Prolongamento Pantanal","ESGOTO",BASE/"Prolongamento Pantanal_ESGOTO"),
        ("Prolongamento Sao Manoel","ESGOTO",BASE/"Prolongamento Sao Manoel_ESGOTO"),
        ("Prolongamento Teteu","ESGOTO",BASE/"Prolongamento Teteu_ESGOTO"),
    ]

    print(f"\n[2] {len(PROJETOS)} projetos LIMPOS a processar")
    todos = []
    stats = {"exec":0,"pend":0,"cad":0,"ext_e":0,"ext_p":0,"ext_c":0}

    for nucleo,tipo,pasta in PROJETOS:
        pvs,trechos = carregar_projeto(pasta)
        if not trechos:
            print(f"  SKIP: {nucleo}")
            continue
        n_exec=0; n_pend=0; n_cad=0; ext_e=0; ext_p=0; ext_c=0
        for t in trechos:
            is_cad, motivo_cad = e_cadastro(t, nucleo)
            if is_cad:
                t["_exec"] = False
                t["_cadastro"] = True
                t["_motivo_cad"] = motivo_cad
                n_cad += 1; ext_c += t.get("ext_m",0)
            else:
                is_exec, match = match_shp(t, pvs, linhas)
                t["_exec"] = is_exec
                t["_match_data"] = match
                t["_cadastro"] = False
                if is_exec:
                    n_exec+=1; ext_e+=t.get("ext_m",0)
                else:
                    n_pend+=1; ext_p+=t.get("ext_m",0)
        
        ext_total = sum(t.get("ext_m",0) for t in trechos)
        ext_obra = ext_total - ext_c  # real construction
        km_real = KM_REAIS_ESG.get(nucleo, 0)
        ratio = ext_obra/1000 / km_real if km_real > 0 else 0
        
        mark = "✅" if 0.7<=ratio<=1.3 else "⚠️" if ratio>0 else ""
        print(f"  {nucleo} ({tipo}): {n_exec} exec + {n_pend} pend + {n_cad} cad "
              f"| Obra: {ext_obra:.0f}m vs Real: {km_real*1000:.0f}m [{ratio:.1f}×] {mark}")
        
        stats["exec"]+=n_exec; stats["pend"]+=n_pend; stats["cad"]+=n_cad
        stats["ext_e"]+=ext_e; stats["ext_p"]+=ext_p; stats["ext_c"]+=ext_c
        
        por_rua = defaultdict(list)
        for t in trechos:
            por_rua[limpar(t.get("rua"))].append(t)
        for rua,lista in por_rua.items():
            todos.append({"nucleo":nucleo,"tipo":tipo,"rua":rua,"trechos":lista,"pvs":pvs})

    # GERAR PLANILHAS
    print(f"\n[3] Gerando planilhas v3 corrigidas...")
    
    # CONSOLIDADO
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Consolidado v3"
    hf=Font(name="Calibri",bold=True,color="FFFFFF",size=10)
    hfl=PatternFill(start_color="1b5e20",end_color="1b5e20",fill_type="solid")
    df=Font(name="Calibri",size=9)
    brd=Border(left=Side("thin","cccccc"),right=Side("thin","cccccc"),
               top=Side("thin","cccccc"),bottom=Side("thin","cccccc"))
    exec_f=PatternFill(start_color="c8e6c9",end_color="c8e6c9",fill_type="solid")
    pend_f=PatternFill(start_color="ffcdd2",end_color="ffcdd2",fill_type="solid")
    cad_f=PatternFill(start_color="e0e0e0",end_color="e0e0e0",fill_type="solid")
    ef=Font(name="Calibri",size=9,bold=True,color="1b5e20")
    pf=Font(name="Calibri",size=9,bold=True,color="c62828")
    cf=Font(name="Calibri",size=9,italic=True,color="757575")
    
    ws.merge_cells("A1:T1")
    ws["A1"]="CONSOLIDADO v3 — EXECUTADO × PROJETADO × CADASTRO — SEM DUPLICAÇÕES"
    ws["A1"].font=Font(bold=True,size=14,color="1b5e20")
    ws.merge_cells("A2:T2")
    ws["A2"]=f"CT {CONTRATO} | {datetime.now().strftime('%d/%m/%Y %H:%M')} | FILTRADO: sem rede cadastro"
    ws["A2"].font=Font(size=9,color="666666")
    heads=["Núcleo","Tipo","Rua","NS","PV Mont","PV Jus","DN(mm)","Ext(m)","Mat",
           "CT Mont","CT Jus","Decl‰","STATUS","Data Exec","Motivo Cadastro",
           "Network","Layer","Análise"]
    for c,h in enumerate(heads,1):
        cell=ws.cell(row=3,column=c,value=h)
        cell.font=hf;cell.fill=hfl;cell.border=brd
    row=4
    for res in todos:
        nucleo=res["nucleo"]; tipo=res["tipo"]; rua=res["rua"]; pvs=res["pvs"]
        for i,t in enumerate(res["trechos"]):
            pvi=pvs.get(t.get("pv_ini"),{})
            pvf=pvs.get(t.get("pv_fim"),{})
            is_exec=t.get("_exec",False)
            is_cad=t.get("_cadastro",False)
            match=t.get("_match_data")
            if is_cad: status="🔘 CADASTRO"
            elif is_exec: status="✅ EXECUTADO"
            else: status="⏳ PENDENTE"
            data_exec=match.get("data","") if match else ""
            motivo_cad=t.get("_motivo_cad","")
            network=str(t.get("network_name",""))
            layer=str(t.get("layer",""))
            analise=""
            if not is_exec and not is_cad and rua=="Sem Rua":
                if pvi.get("x",0)==0: analise="SEM COORDS"
                elif (t.get("dn_mm") or 200)>=300: analise="COLETOR TRONCO"
                else: analise="SEM MAPA"
            vals=[nucleo,tipo,rua,f"NS-{i+1:04d}",t.get("pv_ini",""),t.get("pv_fim",""),
                  t.get("dn_mm"),round(t.get("ext_m",0),2),t.get("material","PVC"),
                  pvi.get("ct"),pvf.get("ct"),
                  round((t.get("decl_mm") or 0)*1000,2) if t.get("decl_mm") else None,
                  status,data_exec,motivo_cad,network,layer,analise]
            for c,v in enumerate(vals,1):
                cell=ws.cell(row=row,column=c,value=v)
                cell.font=cf if is_cad else (ef if c==13 and is_exec else (pf if c==13 else df))
                cell.border=brd
                cell.fill=cad_f if is_cad else (exec_f if is_exec else pend_f)
            row+=1
    
    row+=2
    ws[f"A{row}"]="RESUMO (SOMENTE OBRA NOVA — sem cadastro)"
    ws[f"A{row}"].font=Font(bold=True,size=14,color="1b5e20"); row+=1
    ws[f"A{row}"]=f"✅ Executados: {stats['exec']} ({stats['ext_e']:.0f}m = {stats['ext_e']/1000:.1f}km)"
    ws[f"A{row}"].font=ef; row+=1
    ws[f"A{row}"]=f"⏳ Pendentes: {stats['pend']} ({stats['ext_p']:.0f}m = {stats['ext_p']/1000:.1f}km)"
    ws[f"A{row}"].font=pf; row+=1
    ws[f"A{row}"]=f"🔘 Cadastro: {stats['cad']} ({stats['ext_c']:.0f}m = {stats['ext_c']/1000:.1f}km) — EXCLUÍDO"
    ws[f"A{row}"].font=cf; row+=1
    obra_total=stats['ext_e']+stats['ext_p']
    pct=stats['ext_e']/obra_total*100 if obra_total else 0
    ws[f"A{row}"]=f"📊 Progresso OBRA: {pct:.1f}% ({stats['ext_e']:.0f}m de {obra_total:.0f}m)"
    ws[f"A{row}"].font=Font(bold=True,size=12,color="0d47a1")
    
    for c in range(1,19):
        ws.column_dimensions[get_column_letter(c)].width=14
    ws.column_dimensions["A"].width=22; ws.column_dimensions["C"].width=25
    ws.column_dimensions["M"].width=16; ws.column_dimensions["O"].width=45
    wb.save(str(OUTPUT/"CONSOLIDADO_v3_CORRIGIDO.xlsx"))
    print(f"  Consolidado: {row} linhas")

    # VALÉRIA v3 — materiais pendentes (SEM cadastro)
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active; ws2.title = "Compras Pendentes"
    hf2=Font(bold=True,color="FFFFFF",size=10)
    hfl2=PatternFill(start_color="b71c1c",end_color="b71c1c",fill_type="solid")
    itf=Font(bold=True,size=9); sf=Font(size=9,color="444444")
    nf=Font(bold=True,size=12,color="1a237e")
    rf=Font(bold=True,size=10,color="2e7d32")
    ppf=Font(bold=True,size=10,color="c62828")
    ws2.merge_cells("A1:K1")
    ws2["A1"]="⚠️ MATERIAIS PENDENTES — SEM CADASTRO — SEM DUPLICAÇÃO"
    ws2["A1"].font=Font(bold=True,size=14,color="b71c1c")
    ws2.merge_cells("A2:K2")
    ws2["A2"]=f"CT {CONTRATO} | {datetime.now().strftime('%d/%m/%Y')} | FILTRADO"
    ws2["A2"].font=Font(size=9,color="666666")
    heads2=["MATERIAL","UN","REDE","QTD","METRAGEM","NÚCLEO","RUA","STATUS","EXEC(m)","PEND(m)","CAD(m)"]
    for c,h in enumerate(heads2,1):
        cell=ws2.cell(row=3,column=c,value=h)
        cell.font=hf2;cell.fill=hfl2;cell.border=brd
    row2=4
    por_nucleo=defaultdict(list)
    for d in todos: por_nucleo[d["nucleo"]].append(d)
    for nucleo,dados_list in sorted(por_nucleo.items()):
        ws2.merge_cells(f"A{row2}:K{row2}")
        ws2[f"A{row2}"]=f"═══ {nucleo.upper()} ═══"
        ws2[f"A{row2}"].font=nf; row2+=1
        for dados in sorted(dados_list,key=lambda d:d.get("rua","")):
            rua=dados["rua"]; tipo=dados["tipo"]; trechos=dados["trechos"]
            t_exec=[t for t in trechos if t.get("_exec") and not t.get("_cadastro")]
            t_pend=[t for t in trechos if not t.get("_exec") and not t.get("_cadastro")]
            t_cad=[t for t in trechos if t.get("_cadastro")]
            if not t_pend and not t_exec: continue
            ext_e=sum(t.get("ext_m",0) for t in t_exec)
            ext_p=sum(t.get("ext_m",0) for t in t_pend)
            ext_c=sum(t.get("ext_m",0) for t in t_cad)
            status_rua=f"✅ 100%" if not t_pend else f"⏳ {len(t_pend)}/{len(t_exec)+len(t_pend)} pend"
            ws2[f"A{row2}"]=f"  ▸ {rua}"
            ws2[f"A{row2}"].font=rf if not t_pend else ppf
            ws2[f"H{row2}"]=status_rua
            ws2[f"I{row2}"]=f"{ext_e:.1f}"; ws2[f"J{row2}"]=f"{ext_p:.1f}"
            ws2[f"K{row2}"]=f"{ext_c:.1f}" if ext_c else ""
            for c in range(1,12): ws2.cell(row=row2,column=c).border=brd
            row2+=1
            if not t_pend: continue
            
            # Agrupar trechos pendentes forçando os DNs permitidos
            ext_por_dn = defaultdict(float)
            for t in t_pend:
                dn_orig = t.get("dn_mm") or (200 if tipo=="ESGOTO" else 63)
                if tipo == "ESGOTO":
                    dn = 200 if dn_orig <= 200 else 300
                else:
                    dn = 63 if dn_orig <= 63 else 110
                ext_por_dn[dn] += t.get("ext_m", 0)
                
            n_pvs = len(t_pend) + 1  # 1 PV a mais que trechos linear (aproximação)

            if tipo == "ESGOTO":
                for dn, ext_dn in sorted(ext_por_dn.items()):
                    nb = math.ceil(ext_dn / 6.0)
                    # Tubos e conexões NTS
                    ws2[f"A{row2}"]=f"Tubo PVC Corrugado/Maciço Esgoto JEI DN {dn}mm - NTS 048"; ws2[f"B{row2}"]="m"
                    ws2[f"C{row2}"]="ESG"; ws2[f"D{row2}"]=math.ceil(ext_dn); ws2[f"E{row2}"]=f"{ext_dn:.1f}m"
                    ws2[f"F{row2}"]=nucleo; ws2[f"G{row2}"]=rua; ws2[f"A{row2}"].font=itf
                    for c in range(1,12): ws2.cell(row=row2,column=c).border=brd
                    row2+=1
                    
                    itens_tubo = [
                        (f"  • Anel de Borracha p/ Tubo PVC DN {dn}mm", "pc", nb + 1),
                        (f"  • Pasta Lubrificante p/ Junta Elástica (Bisnaga 1kg)", "un", math.ceil(nb * 0.05)),
                        (f"  • Luva de Correr PVC Esgoto DN {dn}mm", "pc", max(math.ceil(nb * 0.05), 1))
                    ]
                    for desc, un, qt in itens_tubo:
                        ws2[f"A{row2}"]=desc; ws2[f"B{row2}"]=un; ws2[f"D{row2}"]=qt
                        ws2[f"A{row2}"].font=sf
                        for c in range(1,12): ws2.cell(row=row2,column=c).border=brd
                        row2+=1
                
                # PVs e Acessórios
                ws2[f"A{row2}"]="Poço de Visita (PV) Pré-Moldado Concreto D=1,20m - NTS 015"; ws2[f"B{row2}"]="un"
                ws2[f"C{row2}"]="ESG"; ws2[f"D{row2}"]=n_pvs; ws2[f"A{row2}"].font=itf
                for c in range(1,12): ws2.cell(row=row2,column=c).border=brd
                row2+=1
                
                itens_pv = [
                    ("  • Tampão FoFo Articulado Completo D=600mm T-300", "pç", n_pvs),
                    ("  • Laje de Redução Excêntrica Concreto Pré-Moldado", "pç", n_pvs),
                    ("  • Degrau de FºFº para PV", "pç", n_pvs * 5),
                    ("  • Anel de Ajuste Concreto", "pç", n_pvs * 2),
                    ("  • Argamassa de Assentamento Traço 1:3", "sc", math.ceil(n_pvs * 1.5))
                ]
                for desc, un, qt in itens_pv:
                    ws2[f"A{row2}"]=desc; ws2[f"B{row2}"]=un; ws2[f"D{row2}"]=qt
                    ws2[f"A{row2}"].font=sf
                    for c in range(1,12): ws2.cell(row=row2,column=c).border=brd
                    row2+=1
                
                # Materiais Básicos Obras Civis (Valores referenciais)
                ext_p = sum(ext_por_dn.values())
                itens_civil = [
                    ("Escavação mecânica / manual em vala", "m³", math.ceil(ext_p * 0.9)),
                    ("Areia Lavada Média (Envoltória)", "m³", math.ceil(ext_p * 0.35)),
                    ("Brita N° 1 ou 2 (Base/Reaterro)", "m³", math.ceil(ext_p * 0.25)),
                    ("Reposição Pavimento Asfáltico (CBUQ)", "m²", math.ceil(ext_p * 0.9))
                ]
                for desc, un, qt in itens_civil:
                    ws2[f"A{row2}"]=desc; ws2[f"B{row2}"]=un; ws2[f"C{row2}"]="ESG"
                    ws2[f"D{row2}"]=qt; ws2[f"A{row2}"].font=itf
                    for c in range(1,12): ws2.cell(row=row2,column=c).border=brd
                    row2+=1

            else: # AGUA
                for dn, ext_dn in sorted(ext_por_dn.items()):
                    nb = math.ceil(ext_dn / 6.0) # Barras de PEAD geralmente 6m ou 12m
                    ws2[f"A{row2}"]=f"Tubo PEAD PE 100 SDR 17 PN 10 DN {dn}mm - NTS 194"; ws2[f"B{row2}"]="m"
                    ws2[f"C{row2}"]="AG"; ws2[f"D{row2}"]=math.ceil(ext_dn); ws2[f"E{row2}"]=f"{ext_dn:.1f}m"
                    ws2[f"F{row2}"]=nucleo; ws2[f"G{row2}"]=rua; ws2[f"A{row2}"].font=itf
                    for c in range(1,12): ws2.cell(row=row2,column=c).border=brd
                    row2+=1
                    
                    # Acessórios Água PEAD
                    itens_tubo_ag = [
                        (f"  • Luva Eletrofusão PEAD DN {dn}mm", "pc", math.ceil(nb * 1.1)),
                        (f"  • Tê Eletrofusão PEAD DN {dn}mm", "pc", max(math.ceil(nb * 0.05), 1)),
                        (f"  • Cap Eletrofusão PEAD DN {dn}mm", "pc", 2)
                    ]
                    for desc, un, qt in itens_tubo_ag:
                        ws2[f"A{row2}"]=desc; ws2[f"B{row2}"]=un; ws2[f"D{row2}"]=qt
                        ws2[f"A{row2}"].font=sf
                        for c in range(1,12): ws2.cell(row=row2,column=c).border=brd
                        row2+=1
                    
                ext_p = sum(ext_por_dn.values())
                itens_civil_ag = [
                    ("Escavação mecânica / manual em vala (Água)", "m³", math.ceil(ext_p * 0.6)),
                    ("Areia Lavada Média (Envoltória - Água)", "m³", math.ceil(ext_p * 0.2)),
                    ("Reposição Pavimento Asfáltico (CBUQ)", "m²", math.ceil(ext_p * 0.6))
                ]
                for desc, un, qt in itens_civil_ag:
                    ws2[f"A{row2}"]=desc; ws2[f"B{row2}"]=un; ws2[f"C{row2}"]="AG"
                    ws2[f"D{row2}"]=qt; ws2[f"A{row2}"].font=itf
                    for c in range(1,12): ws2.cell(row=row2,column=c).border=brd
                    row2+=1
            row2+=1
    
    row2+=1
    ws2[f"A{row2}"]=f"EXEC: {stats['ext_e']:.0f}m | PEND: {stats['ext_p']:.0f}m | CAD(excl): {stats['ext_c']:.0f}m"
    ws2[f"A{row2}"].font=Font(bold=True,size=12,color="b71c1c")
    for col,w in [("A",42),("B",8),("C",6),("D",8),("E",12),("F",20),("G",25),("H",20),("I",10),("J",10),("K",10)]:
        ws2.column_dimensions[col].width=w
    wb2.save(str(OUTPUT/"MATERIAIS_PENDENTES_v4.xlsx"))
    print(f"  Valéria v4: {row2} linhas")

    # RESUMO
    wb3 = openpyxl.Workbook()
    ws3 = wb3.active; ws3.title = "Resumo"
    heads3=["Núcleo","Tipo","Tr Total","Tr Obra","Tr Cad","Tr Exec","Tr Pend",
            "km Obra","km Exec","km Pend","km Cad","km Real","Ratio","% Exec"]
    for c,h in enumerate(heads3,1):
        cell=ws3.cell(row=1,column=c,value=h)
        cell.font=Font(bold=True,color="FFFFFF"); cell.fill=PatternFill(start_color="0d47a1",end_color="0d47a1",fill_type="solid")
        cell.border=brd
    por_nuc=defaultdict(lambda:{"exec":0,"pend":0,"cad":0,"ext_e":0,"ext_p":0,"ext_c":0,"tipo":""})
    for d in todos:
        k=d["nucleo"]
        por_nuc[k]["tipo"]=d["tipo"]
        for t in d["trechos"]:
            if t.get("_cadastro"):
                por_nuc[k]["cad"]+=1; por_nuc[k]["ext_c"]+=t.get("ext_m",0)
            elif t.get("_exec"):
                por_nuc[k]["exec"]+=1; por_nuc[k]["ext_e"]+=t.get("ext_m",0)
            else:
                por_nuc[k]["pend"]+=1; por_nuc[k]["ext_p"]+=t.get("ext_m",0)
    r3=2
    for nuc,v in sorted(por_nuc.items()):
        total=v["exec"]+v["pend"]+v["cad"]
        obra=v["exec"]+v["pend"]
        km_obra=(v["ext_e"]+v["ext_p"])/1000
        km_real=KM_REAIS_ESG.get(nuc,0)
        ratio=km_obra/km_real if km_real else 0
        pct=v["ext_e"]/(v["ext_e"]+v["ext_p"])*100 if (v["ext_e"]+v["ext_p"]) else 0
        vals=[nuc,v["tipo"],total,obra,v["cad"],v["exec"],v["pend"],
              round(km_obra,1),round(v["ext_e"]/1000,1),round(v["ext_p"]/1000,1),
              round(v["ext_c"]/1000,1),km_real,f"{ratio:.1f}×",f"{pct:.0f}%"]
        for c,val in enumerate(vals,1):
            cell=ws3.cell(row=r3,column=c,value=val)
            cell.border=brd
            if pct>=80: cell.fill=PatternFill(start_color="c8e6c9",end_color="c8e6c9",fill_type="solid")
            elif pct>=40: cell.fill=PatternFill(start_color="fff9c4",end_color="fff9c4",fill_type="solid")
            else: cell.fill=PatternFill(start_color="ffcdd2",end_color="ffcdd2",fill_type="solid")
        r3+=1
    for c in range(1,15):
        ws3.column_dimensions[get_column_letter(c)].width=14
    ws3.column_dimensions["A"].width=25
    wb3.save(str(OUTPUT/"RESUMO_v3_CORRIGIDO.xlsx"))
    print(f"  Resumo: {r3-2} núcleos")

    print(f"\n{'='*72}")
    print(f"  RESULTADO FINAL (CORRIGIDO)")
    print(f"{'='*72}")
    obra_total=stats["ext_e"]+stats["ext_p"]
    pct=stats["ext_e"]/obra_total*100 if obra_total else 0
    print(f"  ✅ Executados: {stats['exec']} trechos ({stats['ext_e']:.0f}m = {stats['ext_e']/1000:.1f}km)")
    print(f"  ⏳ Pendentes:  {stats['pend']} trechos ({stats['ext_p']:.0f}m = {stats['ext_p']/1000:.1f}km)")
    print(f"  🔘 Cadastro:   {stats['cad']} trechos ({stats['ext_c']:.0f}m = {stats['ext_c']/1000:.1f}km) — EXCLUÍDO")
    print(f"  📊 PROGRESSO REAL: {pct:.1f}% (obra: {obra_total:.0f}m = {obra_total/1000:.1f}km)")
    print(f"  📁 Saída: {OUTPUT}")

if __name__=="__main__":
    main()
