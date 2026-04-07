#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PROCESSAR TUDO POR RUA — Pipeline Completo
ConstruData HydroNetwork v9

Processa TODOS os XML e DXF da pasta de projetos:
  - Gera NS completas (A4, Desenho, Satélite, HTML)
  - Planilhas de trechos POR RUA (água e esgoto separados)
  - Imagem satélite por trecho na planilha por rua
  - Planilha de materiais modelo Valéria
  - Consolidado final de todos os núcleos
"""

import sys, os, json, math, traceback, re, warnings, io
from pathlib import Path
from datetime import datetime
from collections import defaultdict, Counter

warnings.filterwarnings("ignore")

# ── Setup: adicionar motor ao path ───────────────────────────────────────────
MOTOR_DIR = Path(r"C:\Users\felip\Downloads\NOVA NS Versao 5")
sys.path.insert(0, str(MOTOR_DIR))

# ── Imports da plataforma ────────────────────────────────────────────────────
from ler_landxml import ler_landxml
from ler_dxf_gdal import ler_dxf_gdal
from gerar_ns import (
    gerar_ns_a4, gerar_ns_desenho, gerar_ns_sat, gerar_html,
    gerar_geojson, enriquecer_trechos, calcular_materiais,
    calc_manning, CONTRATO, NS_VERSION, _ns_folder_name,
    processar_nucleo_from_data
)

try:
    from motor_custo import custo_trecho, FATORES, BDI
    HAS_CUSTO = True
except ImportError:
    HAS_CUSTO = False
    FATORES = {"CBUQ provisório": 0.88, "Areia": 0.25, "Berço": 0.24, "Brita base": 0.22}

try:
    from exportar_completo import (
        _gerar_planilha_trechos_completa, _slug, _carregar_ruas_gpkg
    )
    HAS_EXPORT = True
except ImportError:
    HAS_EXPORT = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XlImage
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import numpy as np
    import geopandas as gpd
    HAS_GEO = True
except ImportError:
    HAS_GEO = False

try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    HAS_MPL = True
except ImportError:
    HAS_MPL = False

try:
    import contextily as cx
    HAS_CTX = True
except ImportError:
    HAS_CTX = False

try:
    from pyproj import Transformer
    _TF = Transformer.from_crs("EPSG:31983", "EPSG:4326", always_xy=True)
    HAS_PROJ = True
except ImportError:
    HAS_PROJ = False

import unicodedata

def slug(texto):
    if not texto:
        return "DESCONHECIDO"
    # Limpar quebras de linha e caracteres de controle
    txt = str(texto).replace("\n", " ").replace("\r", " ").replace("\t", " ")
    txt = re.sub(r'[\x00-\x1f\x7f]', '', txt)  # remove control chars
    normalizado = unicodedata.normalize("NFKD", txt)
    ascii_txt = normalizado.encode("ascii", "ignore").decode("ascii")
    safe = re.sub(r'[^\w\s-]', '_', ascii_txt).strip()
    safe = re.sub(r'\s+', ' ', safe)  # colapsar espaços múltiplos
    return safe[:80] or "DESCONHECIDO"


def limpar_rua(rua):
    """Limpa nome de rua de caracteres inválidos."""
    if not rua or str(rua).strip() in ("", "nan", "None", "Sem Rua"):
        return "Sem Rua"
    txt = str(rua).replace("\n", " ").replace("\r", " ").replace("\t", " ")
    txt = re.sub(r'[\x00-\x1f\x7f]', '', txt)
    txt = re.sub(r'\s+', ' ', txt).strip()
    return txt if txt else "Sem Rua"


# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURAÇÃO: PROJETOS A PROCESSAR
# ══════════════════════════════════════════════════════════════════════════════

INPUT_DIR = Path(r"C:\Users\felip\Desktop\GERAR NS COM ITENS POR RUA DE PV A PV")
OUTPUT_DIR = INPUT_DIR / "SAIDA_COMPLETA_POR_RUA"

PROJETOS = [
    # ── XMLs (LandXML) ────────────────────────────────────
    {"arquivo": "ESTUDO - CT JOÃO CARLOS DA SILVA.xml",     "nucleo": "Joao Carlos",               "tipo": "ESGOTO", "gpkg": "MAPA JOÃO CARLOS DA SILVA_RV07 (QGIS).gpkg"},
    {"arquivo": "ESTUDO - SÃO MANOEL.xml",                  "nucleo": "Sao Manoel",                "tipo": "ESGOTO", "gpkg": "MAPA SÃO MANOEL_RV05 (QGIS).gpkg"},
    {"arquivo": "PROLONGAMENTO CRIADORES.xml",               "nucleo": "Prolongamento Criadores",   "tipo": "ESGOTO", "gpkg": "CARTOGRAFIA VILA CRIADORES_R06 (QGIS).gpkg"},
    {"arquivo": "PROLONGAMENTO PANTANAL BAIXO.xml",          "nucleo": "Prolongamento Pantanal",    "tipo": "ESGOTO", "gpkg": "MAPA PANTANAL BAIXO_R04 (QGIS).gpkg"},
    {"arquivo": "PROLONGAMENTO SÃO MANOEL.xml",              "nucleo": "Prolongamento Sao Manoel",  "tipo": "ESGOTO", "gpkg": "MAPA SÃO MANOEL_RV05 (QGIS).gpkg"},
    {"arquivo": "PROLONGAMENTO TETEU ALT-01.xml",            "nucleo": "Prolongamento Teteu Alt",   "tipo": "ESGOTO", "gpkg": "MAPA TETEU-VALE VERDE_R04 (QGIS).gpkg"},
    {"arquivo": "PROLONGAMENTO TETEU.xml",                   "nucleo": "Prolongamento Teteu",       "tipo": "ESGOTO", "gpkg": "MAPA TETEU-VALE VERDE_R04 (QGIS).gpkg"},

    # ── DXFs ESGOTO ───────────────────────────────────────
    {"arquivo": "CRIADORES_ESGOTO_REV.02.dxf",              "nucleo": "Vila Criadores",            "tipo": "ESGOTO", "gpkg": "CARTOGRAFIA VILA CRIADORES_R06 (QGIS).gpkg"},
    {"arquivo": "ISRAEL_ESGOTO.dxf",                         "nucleo": "Vila Israel",               "tipo": "ESGOTO", "gpkg": "MAPA PANTANAL ALTO (VILA ISRAEL)_RV02 (QGIS).gpkg"},
    {"arquivo": "PANTANAL_ESGOTO.dxf",                       "nucleo": "Pantanal Baixo",            "tipo": "ESGOTO", "gpkg": "MAPA PANTANAL BAIXO_R04 (QGIS).gpkg"},
    {"arquivo": "TETÉU_ESGOTOrevisao.dxf",                  "nucleo": "Morro do Teteu",            "tipo": "ESGOTO", "gpkg": "MAPA TETEU-VALE VERDE_R04 (QGIS).gpkg"},
    {"arquivo": "TETÉU_ESGOTO2.dxf",                        "nucleo": "Morro do Teteu v2",         "tipo": "ESGOTO", "gpkg": "MAPA TETEU-VALE VERDE_R04 (QGIS).gpkg"},

    # ── DXFs ÁGUA ─────────────────────────────────────────
    {"arquivo": "CRIADORES_ÁGUA.dxf",                        "nucleo": "Vila Criadores",            "tipo": "AGUA",   "gpkg": "CARTOGRAFIA VILA CRIADORES_R06 (QGIS).gpkg"},
    {"arquivo": "ISRAEL_ÁGUA.dxf",                           "nucleo": "Vila Israel",               "tipo": "AGUA",   "gpkg": "MAPA PANTANAL ALTO (VILA ISRAEL)_RV02 (QGIS).gpkg"},
    {"arquivo": "PANTANAL ÁGUA.dxf",                         "nucleo": "Pantanal Baixo",            "tipo": "AGUA",   "gpkg": "MAPA PANTANAL BAIXO_R04 (QGIS).gpkg"},
    {"arquivo": "TETEU_ÁGUA.dxf",                            "nucleo": "Morro do Teteu",            "tipo": "AGUA",   "gpkg": "MAPA TETEU-VALE VERDE_R04 (QGIS).gpkg"},
]


def log(msg, nivel="INFO"):
    ts = datetime.now().strftime("%H:%M:%S")
    prefix = {"OK": "[OK]  ", "WARN": "[!]   ", "STEP": ">>> ", "ERR": "[X]  "}.get(nivel, "      ")
    line = f"[{ts}] {prefix}{msg}"
    try:
        print(line)
    except UnicodeEncodeError:
        print(line.encode("ascii", "replace").decode())


# ══════════════════════════════════════════════════════════════════════════════
# GERAR IMAGEM SATELITE DE UM TRECHO
# ══════════════════════════════════════════════════════════════════════════════

def gerar_img_trecho(pvs, trecho, out_path, width_px=600, height_px=400):
    """Gera imagem PNG do trecho com mapa satélite (ou UTM fallback)."""
    if not HAS_MPL:
        return False

    p0 = pvs.get(trecho["pv_ini"], {})
    p1 = pvs.get(trecho["pv_fim"], {})
    if not p0.get("x") or not p1.get("x"):
        return False

    fig, ax = plt.subplots(figsize=(width_px/100, height_px/100), dpi=100)

    x0, y0 = p0["x"], p0["y"]
    x1, y1 = p1["x"], p1["y"]
    dx = abs(x1 - x0)
    dy = abs(y1 - y0)
    margin = max(dx, dy, 30) * 0.6

    ax.set_xlim(min(x0, x1) - margin, max(x0, x1) + margin)
    ax.set_ylim(min(y0, y1) - margin, max(y0, y1) + margin)

    # Tentar satélite
    sat_ok = False
    if HAS_CTX:
        try:
            cx.add_basemap(ax, crs="EPSG:31983",
                           source=cx.providers.Esri.WorldImagery, zoom=18)
            sat_ok = True
        except Exception:
            try:
                cx.add_basemap(ax, crs="EPSG:31983",
                               source=cx.providers.Esri.WorldImagery, zoom=17)
                sat_ok = True
            except Exception:
                pass

    if not sat_ok:
        ax.set_facecolor("#e8f0e8")
        ax.grid(True, color="#ccc", alpha=0.5)

    # Tubo
    ax.plot([x0, x1], [y0, y1], 'r-', linewidth=3, zorder=5)

    # PVs
    for nome, pv, color in [(trecho["pv_ini"], p0, "red"), (trecho["pv_fim"], p1, "blue")]:
        ax.plot(pv["x"], pv["y"], 'o', color=color, markersize=8, zorder=6)
        ct = pv.get("ct", "")
        ct_txt = f"\nCT={ct:.3f}" if isinstance(ct, (int, float)) else ""
        ax.annotate(f"{nome}{ct_txt}", (pv["x"], pv["y"]),
                    fontsize=7, color="white" if sat_ok else "black",
                    fontweight="bold", ha="center", va="bottom",
                    xytext=(0, 8), textcoords="offset points",
                    bbox=dict(boxstyle="round,pad=0.2", fc="black" if sat_ok else "white",
                              alpha=0.7), zorder=7)

    dn = trecho.get("dn_mm", "?")
    ext = trecho.get("ext_m", 0)
    ax.set_title(f"DN{dn}mm | {ext:.1f}m | {trecho.get('rua', '')}", fontsize=8, fontweight="bold")
    ax.tick_params(labelsize=5)

    fig.tight_layout()
    fig.savefig(str(out_path), dpi=100, bbox_inches="tight", pad_inches=0.1)
    plt.close(fig)
    return True


# ══════════════════════════════════════════════════════════════════════════════
# PLANILHA DE TRECHOS POR RUA COM IMAGEM
# ══════════════════════════════════════════════════════════════════════════════

def gerar_planilha_trechos_rua(trechos, pvs, nucleo, tipo_rede, rua, out_path, img_dir=None):
    """Gera planilha de trechos de uma rua com imagem satélite opcional."""
    if not HAS_OPENPYXL:
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trechos"

    # Estilos
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="Calibri", size=9)
    data_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin", color="999999"),
        right=Side(style="thin", color="999999"),
        top=Side(style="thin", color="999999"),
        bottom=Side(style="thin", color="999999"),
    )
    fill_alt = PatternFill(start_color="f0f4ff", end_color="f0f4ff", fill_type="solid")

    # Título
    ws.merge_cells("A1:N1")
    ws["A1"] = f"TRECHOS POR RUA — {rua} — {nucleo} — {tipo_rede}"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="1a237e")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:N2")
    ws["A2"] = f"Contrato {CONTRATO} | SE LIGA NA REDE | Gerado {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(name="Calibri", size=9, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Cabeçalhos
    headers = [
        "NS", "PV Montante", "PV Jusante", "DN (mm)", "Extensão (m)",
        "Material", "CT Mont.", "CF Mont.", "CT Jus.", "CF Jus.",
        "Decl (‰)", "V (m/s)", "Q (l/s)", "Imagem"
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # Dados
    for i, t in enumerate(trechos):
        r = i + 4
        pvi = pvs.get(t["pv_ini"], {})
        pvf = pvs.get(t["pv_fim"], {})
        hidr = calc_manning(t.get("dn_mm"), t.get("decl_mm"))

        values = [
            f"NS-{i+1:04d}",
            t.get("pv_ini", ""),
            t.get("pv_fim", ""),
            t.get("dn_mm"),
            round(t.get("ext_m", 0), 2),
            t.get("material", "PVC"),
            pvi.get("ct"),
            pvi.get("cf"),
            pvf.get("ct"),
            pvf.get("cf"),
            round(t.get("decl_mm", 0) * 1000, 2) if t.get("decl_mm") else None,
            hidr.get("v_ms"),
            hidr.get("q_ls"),
            "",  # Imagem placeholder
        ]
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = border
            if i % 2 == 1:
                cell.fill = fill_alt

        # Gerar e inserir imagem do trecho
        if img_dir and HAS_MPL:
            img_path = Path(img_dir) / f"trecho_{i+1:03d}.png"
            try:
                if gerar_img_trecho(pvs, t, img_path):
                    img = XlImage(str(img_path))
                    img.width = 300
                    img.height = 180
                    ws.add_image(img, f"N{r}")
                    ws.row_dimensions[r].height = 140
            except Exception:
                pass

    # Ajustar larguras
    widths = [8, 14, 14, 8, 10, 8, 10, 10, 10, 10, 8, 8, 8, 45]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # Rodapé resumo
    r_final = len(trechos) + 5
    ws.merge_cells(f"A{r_final}:N{r_final}")
    ext_total = sum(t.get("ext_m", 0) for t in trechos)
    ws[f"A{r_final}"] = f"TOTAL: {len(trechos)} trechos | {ext_total:.1f}m | {tipo_rede}"
    ws[f"A{r_final}"].font = Font(name="Calibri", bold=True, size=11, color="1a237e")

    wb.save(str(out_path))
    return len(trechos)


# ══════════════════════════════════════════════════════════════════════════════
# PLANILHA DE MATERIAIS ESTILO VALÉRIA
# ══════════════════════════════════════════════════════════════════════════════

def gerar_planilha_valeria(todos_dados, out_path):
    """
    Gera planilha de materiais no formato Valéria: material principal + sub-itens.
    todos_dados = [{"nucleo": ..., "tipo": ..., "rua": ..., "trechos": [...], "pvs": {...}}, ...]
    """
    if not HAS_OPENPYXL:
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Materiais por Rua"

    # Estilos
    h_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    h_fill = PatternFill(start_color="0d47a1", end_color="0d47a1", fill_type="solid")
    sub_font = Font(name="Calibri", size=9, color="444444")
    item_font = Font(name="Calibri", bold=True, size=9)
    nucleo_font = Font(name="Calibri", bold=True, size=12, color="1a237e")
    rua_font = Font(name="Calibri", bold=True, size=10, color="2e7d32")
    border = Border(
        left=Side(style="thin", color="cccccc"),
        right=Side(style="thin", color="cccccc"),
        top=Side(style="thin", color="cccccc"),
        bottom=Side(style="thin", color="cccccc"),
    )

    # Título
    ws.merge_cells("A1:H1")
    ws["A1"] = "MATERIAIS POR RUA — BASEADO NOS PROJETOS DXF/XML"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="0d47a1")

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Contrato {CONTRATO} | SE LIGA NA REDE | {datetime.now().strftime('%d/%m/%Y')}"
    ws["A2"].font = Font(name="Calibri", size=9, color="666666")

    # Cabeçalhos
    headers = ["MATERIAL", "UN", "REDE", "QTD TOTAL", "POR RUA", "NÚCLEO", "RUA", "OBS"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = h_font
        cell.fill = h_fill
        cell.border = border

    row = 4

    # Agrupar por núcleo → rua
    por_nucleo = defaultdict(list)
    for d in todos_dados:
        por_nucleo[d["nucleo"]].append(d)

    for nucleo, dados_nucleo in sorted(por_nucleo.items()):
        # Header do núcleo
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"] = f"═══ {nucleo.upper()} ═══"
        ws[f"A{row}"].font = nucleo_font
        row += 1

        for dados in sorted(dados_nucleo, key=lambda d: d.get("rua", "")):
            rua = dados.get("rua", "Sem Rua")
            tipo = dados.get("tipo", "ESGOTO")
            trechos = dados.get("trechos", [])
            pvs = dados.get("pvs", {})

            if not trechos:
                continue

            # Header da rua
            ws[f"A{row}"] = f"  ▸ {rua}"
            ws[f"A{row}"].font = rua_font
            ws[f"C{row}"] = tipo
            ws[f"F{row}"] = nucleo
            ws[f"G{row}"] = rua
            row += 1

            # Calcular materiais
            ext_total = sum(t.get("ext_m", 0) for t in trechos)
            n_pvs_trecho = len(trechos) + 1  # aprox
            dn_default = 200 if tipo == "ESGOTO" else 63
            dns = Counter((t.get("dn_mm") or dn_default) for t in trechos)
            mats = Counter(t.get("material", "PVC") for t in trechos)

            if tipo == "ESGOTO":
                # Tubos PVC por DN
                for dn, qtd_trechos in sorted(dns.items(), key=lambda x: x[0] or 0):
                    ext_dn = sum(t.get("ext_m", 0) for t in trechos if (t.get("dn_mm") or dn_default) == dn)
                    n_barras = math.ceil(ext_dn / 6)

                    ws[f"A{row}"] = f"Tubo PVC DN{dn}mm"
                    ws[f"B{row}"] = "barra"
                    ws[f"C{row}"] = "ESG"
                    ws[f"D{row}"] = n_barras
                    ws[f"E{row}"] = f"{ext_dn:.1f}m"
                    ws[f"A{row}"].font = item_font
                    for c in range(1, 9):
                        ws.cell(row=row, column=c).border = border
                    row += 1

                    # Sub-itens
                    sub_items = [
                        (f"   • Anel de Borracha DN{dn}", "pc", n_barras + 1),
                        (f"   • Pasta Lubrificante", "kg", round(n_barras * 0.04, 2)),
                        (f"   • Luva correr PVC DN{dn}", "pc", max(n_barras - 1, 0)),
                    ]
                    for desc, un, qtd in sub_items:
                        ws[f"A{row}"] = desc
                        ws[f"B{row}"] = un
                        ws[f"D{row}"] = qtd
                        ws[f"A{row}"].font = sub_font
                        for c in range(1, 9):
                            ws.cell(row=row, column=c).border = border
                        row += 1

                # PVs concreto
                ws[f"A{row}"] = "PV concreto DN1200"
                ws[f"B{row}"] = "un"
                ws[f"C{row}"] = "ESG"
                ws[f"D{row}"] = n_pvs_trecho
                ws[f"A{row}"].font = item_font
                for c in range(1, 9):
                    ws.cell(row=row, column=c).border = border
                row += 1

                pv_subs = [
                    ("   • Tampão Ferro Fundido DN600/900", "pc", n_pvs_trecho),
                    ("   • Laje de Cobertura / Cone Redução", "pc", n_pvs_trecho),
                    ("   • Anel Borracha (Junta Elástica)", "pc", n_pvs_trecho * 3),
                    ("   • Degrau Ferro Polipropileno", "pc", n_pvs_trecho * 6),
                ]
                for desc, un, qtd in pv_subs:
                    ws[f"A{row}"] = desc
                    ws[f"B{row}"] = un
                    ws[f"D{row}"] = qtd
                    ws[f"A{row}"].font = sub_font
                    for c in range(1, 9):
                        ws.cell(row=row, column=c).border = border
                    row += 1

                # Escavação e aterro
                vol_esc = round(ext_total * 0.60 * 1.5, 2)
                vol_areia = round(ext_total * 0.32, 2)
                vol_brita = round(ext_total * 0.22, 2)
                area_pav = round(ext_total * 0.88, 2)

                extras = [
                    ("Escavação mecanizada", "m³", vol_esc),
                    ("Areia (lastro + envoltória)", "m³", vol_areia),
                    ("Brita (dreno)", "m³", vol_brita),
                    ("Pavimentação CBUQ", "m²", area_pav),
                ]
                for desc, un, qtd in extras:
                    ws[f"A{row}"] = desc
                    ws[f"B{row}"] = un
                    ws[f"C{row}"] = "ESG"
                    ws[f"D{row}"] = qtd
                    ws[f"A{row}"].font = item_font
                    for c in range(1, 9):
                        ws.cell(row=row, column=c).border = border
                    row += 1

            elif tipo == "AGUA":
                # Tubos PEAD
                for dn, qtd_trechos in sorted(dns.items(), key=lambda x: x[0] or 0):
                    ext_dn = sum(t.get("ext_m", 0) for t in trechos if (t.get("dn_mm") or dn_default) == dn)
                    n_barras = math.ceil(ext_dn / 6)

                    ws[f"A{row}"] = f"Tubo PEAD DN{dn}mm"
                    ws[f"B{row}"] = "barra"
                    ws[f"C{row}"] = "AG"
                    ws[f"D{row}"] = n_barras
                    ws[f"E{row}"] = f"{ext_dn:.1f}m"
                    ws[f"A{row}"].font = item_font
                    for c in range(1, 9):
                        ws.cell(row=row, column=c).border = border
                    row += 1

                    sub_items = [
                        (f"   • Sela de Eletrofusão DN{dn}", "pc", max(round(n_barras * 0.1), 1)),
                        (f"   • Luva de Eletrofusão DN{dn}", "pc", max(round(n_barras * 0.05), 1)),
                    ]
                    for desc, un, qtd in sub_items:
                        ws[f"A{row}"] = desc
                        ws[f"B{row}"] = un
                        ws[f"D{row}"] = qtd
                        ws[f"A{row}"].font = sub_font
                        for c in range(1, 9):
                            ws.cell(row=row, column=c).border = border
                        row += 1

            row += 1  # espaço entre ruas

    # Ajustar larguras
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 6
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 25
    ws.column_dimensions["H"].width = 15

    wb.save(str(out_path))
    log(f"  Planilha Valéria: {row} linhas → {out_path}", "OK")


# ══════════════════════════════════════════════════════════════════════════════
# PLANILHA CONSOLIDADA FINAL
# ══════════════════════════════════════════════════════════════════════════════

def gerar_consolidado_final(todos_resultados, out_path):
    """Gera planilha consolidada com TODOS os trechos de TODOS os núcleos."""
    if not HAS_OPENPYXL:
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Consolidado Geral"

    h_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    h_fill = PatternFill(start_color="1b5e20", end_color="1b5e20", fill_type="solid")
    data_font = Font(name="Calibri", size=9)
    border = Border(
        left=Side(style="thin", color="cccccc"),
        right=Side(style="thin", color="cccccc"),
        top=Side(style="thin", color="cccccc"),
        bottom=Side(style="thin", color="cccccc"),
    )
    fill_alt = PatternFill(start_color="e8f5e9", end_color="e8f5e9", fill_type="solid")

    # Título
    ws.merge_cells("A1:Q1")
    ws["A1"] = "CONSOLIDADO GERAL — TRECHOS POR RUA — TODOS OS NÚCLEOS"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="1b5e20")

    headers = [
        "Núcleo", "Tipo", "Rua", "NS", "PV Mont.", "PV Jus.",
        "DN (mm)", "Ext. (m)", "Material",
        "CT Mont.", "CF Mont.", "CT Jus.", "CF Jus.",
        "Decl (‰)", "V (m/s)", "Q (l/s)", "Obs"
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = h_font
        cell.fill = h_fill
        cell.border = border

    row = 4
    for res in todos_resultados:
        nucleo = res.get("nucleo", "")
        tipo = res.get("tipo", "")
        rua = res.get("rua", "Sem Rua")
        pvs = res.get("pvs", {})

        for i, t in enumerate(res.get("trechos", [])):
            pvi = pvs.get(t["pv_ini"], {})
            pvf = pvs.get(t["pv_fim"], {})
            hidr = calc_manning(t.get("dn_mm"), t.get("decl_mm"))

            values = [
                nucleo, tipo, rua, f"NS-{i+1:04d}",
                t.get("pv_ini", ""), t.get("pv_fim", ""),
                t.get("dn_mm"), round(t.get("ext_m", 0), 2),
                t.get("material", "PVC"),
                pvi.get("ct"), pvi.get("cf"),
                pvf.get("ct"), pvf.get("cf"),
                round(t.get("decl_mm", 0) * 1000, 2) if t.get("decl_mm") else None,
                hidr.get("v_ms"), hidr.get("q_ls"), ""
            ]
            for c, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.font = data_font
                cell.border = border
                if row % 2 == 0:
                    cell.fill = fill_alt
            row += 1

    # Totais
    row += 1
    ws[f"A{row}"] = "TOTAIS"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True, size=12, color="1b5e20")

    total_trechos = row - 5
    total_ext = sum(
        sum(t.get("ext_m", 0) for t in res.get("trechos", []))
        for res in todos_resultados
    )
    ws[f"A{row+1}"] = f"Total de trechos: {total_trechos}"
    ws[f"A{row+2}"] = f"Extensão total: {total_ext:.1f}m"

    # Ajustar larguras
    for c in range(1, 18):
        ws.column_dimensions[get_column_letter(c)].width = 12
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["C"].width = 30

    wb.save(str(out_path))
    log(f"  Consolidado final: {total_trechos} trechos → {out_path}", "OK")


# ══════════════════════════════════════════════════════════════════════════════
# PIPELINE PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════

def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

    print("=" * 72)
    print("  PROCESSAMENTO COMPLETO POR RUA — ConstruData HydroNetwork v9")
    print(f"  {len(PROJETOS)} projetos | {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print("=" * 72)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    todos_dados_valeria = []
    todos_dados_consolidado = []
    stats = {"total_pvs": 0, "total_trechos": 0, "total_ext": 0, "projetos_ok": 0, "projetos_err": 0}

    for pi, proj in enumerate(PROJETOS):
        arquivo = proj["arquivo"]
        nucleo = proj["nucleo"]
        tipo = proj["tipo"]
        gpkg_name = proj.get("gpkg")

        arq_path = INPUT_DIR / arquivo
        if not arq_path.exists():
            log(f"[{pi+1}/{len(PROJETOS)}] NAO ENCONTRADO: {arquivo}", "WARN")
            stats["projetos_err"] += 1
            continue

        log(f"[{pi+1}/{len(PROJETOS)}] {arquivo} → {nucleo} ({tipo})", "STEP")
        tamanho_mb = arq_path.stat().st_size / 1024 / 1024
        log(f"  Tamanho: {tamanho_mb:.1f} MB", "INFO")

        # ── 1. LER ARQUIVO ───────────────────────────────────────────────────
        try:
            ext = arq_path.suffix.lower()
            if ext == ".xml":
                pvs, trechos, ruas_xml, meta = ler_landxml(str(arq_path))
            elif ext == ".dxf":
                pvs, trechos, ruas_raw, meta = ler_dxf_gdal(str(arq_path))
            else:
                log(f"  Formato não suportado: {ext}", "ERR")
                stats["projetos_err"] += 1
                continue
        except Exception as e:
            log(f"  ERRO LEITURA: {e}", "ERR")
            traceback.print_exc()
            stats["projetos_err"] += 1
            continue

        if not trechos:
            log(f"  Sem trechos! Pulando.", "WARN")
            stats["projetos_err"] += 1
            continue

        # ── 2. ENRIQUECER ────────────────────────────────────────────────────
        trechos = enriquecer_trechos(trechos, pvs)
        ext_total = sum(t.get("ext_m", 0) for t in trechos)
        log(f"  Rede: {len(pvs)} PVs, {len(trechos)} trechos, {ext_total:.0f}m", "OK")

        # ── 3. CARREGAR RUAS DO GPKG ─────────────────────────────────────────
        if gpkg_name and HAS_GEO:
            gpkg_path = INPUT_DIR / gpkg_name
            if gpkg_path.exists():
                try:
                    if HAS_EXPORT:
                        n_ruas = _carregar_ruas_gpkg(str(gpkg_path), trechos, pvs)
                    else:
                        # Fallback: carregar manualmente
                        gdf = gpd.read_file(str(gpkg_path))
                        rua_col = None
                        for col in gdf.columns:
                            c = col.upper()
                            if any(k in c for k in ["RUA", "LOGRADOURO", "NOME", "NAME", "TEXT", "LABEL"]):
                                rua_col = col
                                break
                        n_ruas = 0
                        if rua_col:
                            ruas_gdf = gdf[gdf[rua_col].notna() & (gdf[rua_col] != "")]
                            centroids = ruas_gdf.geometry.centroid
                            rua_xy = np.array([[c.x, c.y] for c in centroids])
                            rua_names = ruas_gdf[rua_col].values
                            for t in trechos:
                                if t.get("rua") and str(t["rua"]).strip() not in ("", "Sem Rua"):
                                    continue
                                p0 = pvs.get(t.get("pv_ini"), {})
                                p1 = pvs.get(t.get("pv_fim"), {})
                                if not p0.get("x") or not p1.get("x"):
                                    continue
                                mx = (p0["x"] + p1["x"]) / 2
                                my = (p0["y"] + p1["y"]) / 2
                                dists = np.hypot(rua_xy[:, 0] - mx, rua_xy[:, 1] - my)
                                idx = int(np.argmin(dists))
                                if dists[idx] <= 100.0:
                                    t["rua"] = str(rua_names[idx]).strip()
                                    n_ruas += 1
                    log(f"  Ruas GPKG: {n_ruas} trechos com rua", "OK")
                except Exception as e:
                    log(f"  Erro GPKG: {e}", "WARN")

        # ── 4. CRIAR PASTAS DE SAÍDA ─────────────────────────────────────────
        nucleo_slug = slug(nucleo)
        tipo_slug = slug(tipo)
        projeto_slug = f"{nucleo_slug}_{tipo_slug}"
        pasta_projeto = OUTPUT_DIR / projeto_slug

        pastas = {}
        for nome in ["01_NS_CAMPO", "02_DESENHOS", "03_HTML", "04_GIS",
                      "05_PLANILHAS", "11_POR_RUA", "12_LOG", "IMG_TRECHOS"]:
            p = pasta_projeto / nome
            p.mkdir(parents=True, exist_ok=True)
            pastas[nome] = p

        # ── 5. GERAR NS (A4 + DESENHO + SAT + HTML) ─────────────────────────
        log(f"  Gerando NS...", "INFO")
        n_ok = n_err = 0
        for i, t in enumerate(trechos):
            ns_id = i + 1
            pv_i = t.get("pv_ini", "PVX")
            pv_f = t.get("pv_fim", "PVY")
            ns_name = _ns_folder_name(ns_id, pv_i, pv_f)
            ns_dir = pastas["01_NS_CAMPO"] / ns_name
            ns_dir.mkdir(parents=True, exist_ok=True)
            try:
                gerar_ns_a4(ns_id, t, pvs, nucleo, str(ns_dir / f"NS{ns_id:03d}_A4.pdf"))
                try:
                    gerar_ns_desenho(ns_id, t, pvs, trechos, nucleo,
                                     str(pastas["02_DESENHOS"] / f"NS{ns_id:03d}_DESENHO.pdf"))
                except Exception:
                    pass
                try:
                    gerar_ns_sat(ns_id, t, pvs, nucleo,
                                 str(pastas["02_DESENHOS"] / f"NS{ns_id:03d}_SAT.pdf"))
                except Exception:
                    pass
                try:
                    gerar_html(ns_id, t, pvs, trechos, nucleo,
                               str(pastas["03_HTML"] / f"NS{ns_id:03d}.html"))
                except Exception:
                    pass

                # DADOS JSON
                materiais = calcular_materiais(t, pvs)
                dados = {
                    "ns_id": ns_id, "nucleo": nucleo, "contrato": CONTRATO,
                    "tipo_rede": tipo,
                    "trecho": {k: v for k, v in t.items()},
                    "pv_montante": pvs.get(pv_i, {}),
                    "pv_jusante": pvs.get(pv_f, {}),
                    "hidraulica": calc_manning(t.get("dn_mm"), t.get("decl_mm")),
                    "materiais": materiais,
                    "gerado_em": datetime.now().isoformat(),
                }
                with open(ns_dir / f"NS{ns_id:03d}_DADOS.json", "w", encoding="utf-8") as f:
                    json.dump(dados, f, indent=2, ensure_ascii=False)

                n_ok += 1
                if ns_id <= 3 or ns_id % 50 == 0:
                    log(f"    NS{ns_id:03d}: {pv_i}→{pv_f} DN{t.get('dn_mm')} {t.get('ext_m',0):.1f}m", "OK")
            except Exception as e:
                n_err += 1
                if ns_id <= 3:
                    log(f"    NS{ns_id:03d}: ERRO - {e}", "ERR")

        log(f"  NS: {n_ok} ok, {n_err} erros", "OK")

        # ── 6. HTML REDE GERAL ───────────────────────────────────────────────
        try:
            gerar_html(0, trechos[0], pvs, trechos, nucleo,
                       str(pastas["03_HTML"] / "REDE_GERAL.html"))
        except Exception:
            pass

        # ── 7. GeoJSON ───────────────────────────────────────────────────────
        try:
            n_feat = gerar_geojson(trechos, pvs, str(pastas["04_GIS"] / "rede.geojson"))
            log(f"  GeoJSON: {n_feat} features", "OK")
        except Exception:
            pass

        # ── 8. PLANILHA MESTRE ───────────────────────────────────────────────
        if HAS_EXPORT:
            try:
                _gerar_planilha_trechos_completa(
                    trechos, pvs, nucleo, tipo,
                    pastas["05_PLANILHAS"] / f"TODOS_TRECHOS_{projeto_slug}.xlsx"
                )
                log(f"  Planilha MESTRE: {len(trechos)} trechos", "OK")
            except Exception as e:
                log(f"  Erro planilha mestre: {e}", "WARN")

        # ── 9. SEPARAR POR RUA E GERAR PLANILHAS ────────────────────────────
        log(f"  Separando por rua...", "INFO")
        trechos_por_rua = defaultdict(list)
        for t in trechos:
            rua = limpar_rua(t.get("rua"))
            t["rua"] = rua  # atualizar no trecho
            trechos_por_rua[rua].append(t)

        for rua, lista in trechos_por_rua.items():
            rua_slug = slug(rua)
            pasta_rua = pastas["11_POR_RUA"] / rua_slug
            pasta_rua.mkdir(parents=True, exist_ok=True)
            img_dir = pasta_rua / "imgs"
            img_dir.mkdir(exist_ok=True)

            try:
                gerar_planilha_trechos_rua(
                    lista, pvs, nucleo, tipo, rua,
                    pasta_rua / f"Trechos_{rua_slug}.xlsx",
                    img_dir=str(img_dir)
                )
            except Exception as e:
                log(f"    Erro planilha rua '{rua}': {e}", "WARN")

            # Dados para Valéria e consolidado
            dados_rua = {
                "nucleo": nucleo, "tipo": tipo, "rua": rua,
                "trechos": lista, "pvs": pvs
            }
            todos_dados_valeria.append(dados_rua)
            todos_dados_consolidado.append(dados_rua)

        log(f"  Ruas: {len(trechos_por_rua)} separadas", "OK")

        # ── 10. LOG ──────────────────────────────────────────────────────────
        log_data = {
            "projeto": arquivo, "nucleo": nucleo, "tipo": tipo,
            "n_pvs": len(pvs), "n_trechos": len(trechos),
            "ext_total_m": round(ext_total, 1),
            "n_ruas": len(trechos_por_rua),
            "n_ns_ok": n_ok, "n_ns_err": n_err,
            "gerado_em": datetime.now().isoformat(),
        }
        with open(pastas["12_LOG"] / "log.json", "w", encoding="utf-8") as f:
            json.dump(log_data, f, indent=2, ensure_ascii=False)

        stats["total_pvs"] += len(pvs)
        stats["total_trechos"] += len(trechos)
        stats["total_ext"] += ext_total
        stats["projetos_ok"] += 1

        log(f"  ✓ PROJETO CONCLUÍDO: {nucleo} ({tipo})", "OK")

    # ══════════════════════════════════════════════════════════════════════════
    # GERAR PLANILHA VALÉRIA (MATERIAIS POR RUA)
    # ══════════════════════════════════════════════════════════════════════════
    log("Gerando planilha VALÉRIA (materiais por rua)...", "STEP")
    try:
        gerar_planilha_valeria(
            todos_dados_valeria,
            OUTPUT_DIR / "MATERIAIS_POR_RUA_VALERIA.xlsx"
        )
    except Exception as e:
        log(f"  Erro Valéria: {e}", "ERR")
        traceback.print_exc()

    # ══════════════════════════════════════════════════════════════════════════
    # GERAR CONSOLIDADO FINAL
    # ══════════════════════════════════════════════════════════════════════════
    log("Gerando CONSOLIDADO FINAL (todos núcleos por rua)...", "STEP")
    try:
        gerar_consolidado_final(
            todos_dados_consolidado,
            OUTPUT_DIR / "TRECHOS_CONSOLIDADO_TODOS_NUCLEOS.xlsx"
        )
    except Exception as e:
        log(f"  Erro consolidado: {e}", "ERR")
        traceback.print_exc()

    # ══════════════════════════════════════════════════════════════════════════
    # RESUMO FINAL
    # ══════════════════════════════════════════════════════════════════════════
    print("\n" + "=" * 72)
    print("  PROCESSAMENTO CONCLUÍDO!")
    print("=" * 72)
    print(f"  Projetos OK:    {stats['projetos_ok']}")
    print(f"  Projetos ERRO:  {stats['projetos_err']}")
    print(f"  PVs totais:     {stats['total_pvs']}")
    print(f"  Trechos totais: {stats['total_trechos']}")
    print(f"  Extensão total: {stats['total_ext']:.0f}m")
    print(f"  Pasta saída:    {OUTPUT_DIR}")
    print(f"  Valéria:        MATERIAIS_POR_RUA_VALERIA.xlsx")
    print(f"  Consolidado:    TRECHOS_CONSOLIDADO_TODOS_NUCLEOS.xlsx")
    print("=" * 72)

    # Salvar resumo JSON
    with open(OUTPUT_DIR / "RESUMO_PROCESSAMENTO.json", "w", encoding="utf-8") as f:
        json.dump({**stats, "gerado_em": datetime.now().isoformat(),
                   "saida": str(OUTPUT_DIR)}, f, indent=2, ensure_ascii=False)


if __name__ == "__main__":
    main()
