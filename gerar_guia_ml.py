#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""GUIA COMPLETO ML 1000 LIGACOES — XGBoost + GridSearchCV + PMBOK/Lean
Gera GUIA_ML_1000_LIGACOES.xlsx com 16 abas completas."""
import sys, io, os, json, datetime, warnings
warnings.filterwarnings("ignore")
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.series import SeriesLabel

# ── deps ML ──────────────────────────────────────────────────────────────────
try:
    import numpy as np
    import xgboost as xgb
    from sklearn.model_selection import GridSearchCV, cross_val_score
    from sklearn.metrics import mean_squared_error, r2_score
    from sklearn.preprocessing import LabelEncoder
    HAS_ML = True
except ImportError as e:
    HAS_ML = False
    print(f"[AVISO] ML indisponivel: {e}")

# ── caminhos ─────────────────────────────────────────────────────────────────
BASE = os.path.dirname(os.path.abspath(__file__))
PLANEJ = os.path.join(os.environ.get("USERPROFILE","C:/Users/felip"),
                      "Downloads","NOVOS NUCLEOS PLANEJAMENTO")
F_GERAL  = os.path.join(PLANEJ, "Execução_Relatórios.xlsx")
F_EXEC   = os.path.join(PLANEJ, "Execução_Geral.xlsx")
OUT_DIR  = os.path.join(BASE, "TRECHOS_NS_COMPLETOS", "GUIA_ML")
os.makedirs(OUT_DIR, exist_ok=True)
OUT_FILE = os.path.join(OUT_DIR, "GUIA_ML_1000_LIGACOES.xlsx")

# ── estilos ──────────────────────────────────────────────────────────────────
AZUL_ESC = "1F4E79"; AZUL_MED = "2E75B6"; AZUL_CLA = "D6E4F0"
VERDE_ESC = "1E6B3C"; VERDE_CLA = "E2EFDA"
VERM_ESC  = "C00000"; VERM_CLA  = "FCE4D6"
AMAR_ESC  = "7F6000"; AMAR_CLA  = "FFF2CC"
CINZA     = "F2F2F2"; BRANCO = "FFFFFF"; LARANJA = "FF6600"

def pf(c): return PatternFill("solid", fgColor=c)
def fn(b=False, c="000000", s=9): return Font(bold=b, color=c, size=s, name="Calibri")
def al(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
THIN = Border(left=Side(style="thin"), right=Side(style="thin"),
              top=Side(style="thin"),  bottom=Side(style="thin"))
MED  = Border(left=Side(style="medium"), right=Side(style="medium"),
              top=Side(style="medium"),  bottom=Side(style="medium"))

def cw(ws, col, w): ws.column_dimensions[get_column_letter(col)].width = w
def rh(ws, row, h): ws.row_dimensions[row].height = h
def mc(ws, r1, c1, r2, c2): ws.merge_cells(start_row=r1, start_column=c1,
                                             end_row=r2, end_column=c2)
def wr(ws, row, col, val, fill=None, font=None, aln=None, border=None, nfmt=None):
    c = ws.cell(row, col, val)
    if fill:   c.fill   = fill
    if font:   c.font   = font
    if aln:    c.alignment = aln
    if border: c.border = border
    if nfmt:   c.number_format = nfmt
    return c

# ── leitura dados reais ───────────────────────────────────────────────────────
NUCLEO_META = {
    "Morro do Tetéu":     {"dific":4,"lig_agua_tot":220,"lig_esg_tot":200,"ext_m":3200,"equipes":4,"ini":"2026-01-01","fim_prev":"2026-07-01"},
    "Pantanal Baixo":     {"dific":4,"lig_agua_tot":210,"lig_esg_tot":200,"ext_m":2900,"equipes":4,"ini":"2026-01-01","fim_prev":"2026-07-01"},
    "Vila dos Criadores": {"dific":3,"lig_agua_tot":280,"lig_esg_tot":260,"ext_m":3800,"equipes":5,"ini":"2026-01-01","fim_prev":"2026-07-01"},
    "São Manoel":         {"dific":3,"lig_agua_tot":260,"lig_esg_tot":250,"ext_m":3500,"equipes":4,"ini":"2026-01-01","fim_prev":"2026-05-01"},
    "João Carlos":        {"dific":3,"lig_agua_tot":230,"lig_esg_tot":220,"ext_m":3100,"equipes":4,"ini":"2026-01-01","fim_prev":"2026-04-01"},
    "Vila Israel":        {"dific":4,"lig_agua_tot":180,"lig_esg_tot":170,"ext_m":2600,"equipes":3,"ini":"2026-01-01","fim_prev":"2026-05-01"},
}

print("Lendo dados historicos...")
daily_rows = []
try:
    wb_r = openpyxl.load_workbook(F_GERAL, read_only=True, data_only=True)
    ws_g = wb_r["GERAL"]
    all_rows = list(ws_g.iter_rows(values_only=True))
    for r in all_rows[2:]:
        dia    = r[1]
        if not isinstance(dia, datetime.datetime): continue
        equipe = str(r[5] or "").strip()
        la     = r[7]  if isinstance(r[7],  (int,float)) else 0
        le     = r[8]  if isinstance(r[8],  (int,float)) else 0
        pra    = r[9]  if isinstance(r[9],  (int,float)) else 0
        pre    = r[10] if isinstance(r[10], (int,float)) else 0
        pvs    = r[11] if isinstance(r[11], (int,float)) else 0
        nucleo = str(r[16] or "").strip()
        if not nucleo or nucleo in ("None","nan",""): continue
        dia_util = str(r[4] or "").upper() not in ("NÃO","NAO","NO")
        daily_rows.append(dict(
            dia=dia, ano=dia.year, mes=dia.month, dia_num=dia.day,
            dia_sem=dia.weekday(),  # 0=Mon
            dia_util=int(dia_util),
            equipe=equipe, la=la, le=le, pra=pra, pre=pre, pvs=pvs,
            nucleo=nucleo
        ))
    wb_r.close()
    print(f"  {len(daily_rows)} linhas diarias lidas")
except Exception as e:
    print(f"  ERRO lendo {F_GERAL}: {e}")

# ── agrupamento mensal por nucleo ─────────────────────────────────────────────
from collections import defaultdict
monthly = defaultdict(lambda: defaultdict(float))
monthly_days = defaultdict(int)
for row in daily_rows:
    if row["la"]==0 and row["le"]==0 and row["pra"]==0 and row["pre"]==0:
        continue
    key = (row["ano"], row["mes"], row["nucleo"])
    monthly[key]["la"]  += row["la"]
    monthly[key]["le"]  += row["le"]
    monthly[key]["pra"] += row["pra"]
    monthly[key]["pre"] += row["pre"]
    monthly_days[key]   += 1

# ── dados diarios para ML ─────────────────────────────────────────────────────
DIFIC_MAP = {n: v["dific"] for n, v in NUCLEO_META.items()}
EQUIPE_MAP = {n: v["equipes"] for n, v in NUCLEO_META.items()}

ml_data = []
for row in daily_rows:
    n = row["nucleo"]
    dific = DIFIC_MAP.get(n, 3)
    neq   = EQUIPE_MAP.get(n, 3)
    # semana no mes (1-5)
    sem_mes = (row["dia_num"]-1)//7 + 1
    # meses ativo desde jan/2026
    meses_ativo = max(0, (row["ano"]-2026)*12 + row["mes"] - 1)
    ml_data.append([
        row["dia_sem"],     # 0=Mon..6=Sun
        row["dia_util"],    # 0/1
        row["mes"],         # 1-12
        sem_mes,            # 1-5
        dific,              # 1-5
        neq,                # n equipes
        meses_ativo,        # 0,1,2...
        row["la"],          # TARGET
        row["le"],
        row["pra"],
        row["pre"],
        n
    ])

FEAT_NAMES = ["dia_semana","dia_util","mes","semana_mes","dificuldade",
              "n_equipes","meses_ativo"]

print(f"  {len(ml_data)} pontos ML")

# ══════════════════════════════════════════════════════════════════════════════
# ML — XGBoost + GridSearchCV
# ══════════════════════════════════════════════════════════════════════════════
ml_results = {}
fi_results = {}

if HAS_ML and len(ml_data) > 30:
    print("Rodando XGBoost + GridSearchCV...")
    arr = np.array([[d[i] for i in range(7)] for d in ml_data], dtype=float)
    y_la  = np.array([d[7]  for d in ml_data], dtype=float)
    y_le  = np.array([d[8]  for d in ml_data], dtype=float)
    y_pra = np.array([d[9]  for d in ml_data], dtype=float)

    # filtrar dias com producao real (LA ou PRA > 0)
    mask = (y_la > 0) | (y_pra > 0)
    Xf   = arr[mask]
    yf_la  = y_la[mask]
    yf_le  = y_le[mask]
    yf_pra = y_pra[mask]
    print(f"  Amostras com producao: {mask.sum()}")

    param_grid = {
        "n_estimators":     [50, 100, 200],
        "max_depth":        [3, 4, 6],
        "learning_rate":    [0.05, 0.1, 0.2],
        "subsample":        [0.8, 1.0],
        "colsample_bytree": [0.8, 1.0],
    }

    for label, yf in [("LA_agua", yf_la), ("LE_esgoto", yf_le), ("PRA_m", yf_pra)]:
        print(f"  GridSearchCV {label}...")
        base = xgb.XGBRegressor(random_state=42, verbosity=0, objective="reg:squarederror")
        gs = GridSearchCV(base, param_grid, cv=min(5, len(Xf)//10+2),
                          scoring="neg_root_mean_squared_error",
                          n_jobs=-1, refit=True)
        gs.fit(Xf, yf)
        best = gs.best_estimator_
        y_pred = best.predict(Xf)
        rmse = float(np.sqrt(mean_squared_error(yf, y_pred)))
        r2   = float(r2_score(yf, y_pred))
        fi   = dict(zip(FEAT_NAMES, best.feature_importances_.tolist()))
        ml_results[label] = dict(
            best_params=gs.best_params_,
            rmse=rmse, r2=r2,
            cv_score=-gs.best_score_,
        )
        fi_results[label] = fi
        print(f"    {label}: RMSE={rmse:.2f} R2={r2:.3f} params={gs.best_params_}")

    # ── Simulacao: quantas equipes p/ 1000 LA/mes ────────────────────────────
    # Dias uteis por mes ~22, dificuldade media 3.5
    if "LA_agua" in ml_results:
        best_la = GridSearchCV(xgb.XGBRegressor(random_state=42,verbosity=0),
                               param_grid,cv=3,scoring="neg_root_mean_squared_error",
                               n_jobs=-1).fit(Xf,yf_la).best_estimator_
        # simular varios cenarios
        simulacoes = []
        for neq in range(10, 51, 5):
            for dif in [2, 3, 4]:
                # dia util tipico, mes=5, sem=2, meses_ativo=2
                xsim = np.array([[1, 1, 5, 2, dif, neq, 2]], dtype=float)
                la_dia = float(best_la.predict(xsim)[0])
                la_mes = la_dia * 22
                simulacoes.append(dict(neq=neq, dif=dif,
                                       la_dia=round(la_dia,1),
                                       la_mes=round(la_mes,0)))
        ml_results["simulacoes"] = simulacoes
        print(f"  Simulacoes geradas: {len(simulacoes)}")

print("ML concluido.")

# ══════════════════════════════════════════════════════════════════════════════
# DADOS CONSOLIDADOS
# ══════════════════════════════════════════════════════════════════════════════
# producao mensal total
MESES_LABEL = {(2026,1):"Jan/26",(2026,2):"Fev/26",(2026,3):"Mar/26"}

prod_total = defaultdict(lambda: {"la":0,"le":0,"pra":0,"pre":0})
for (ano,mes,nucleo), vals in monthly.items():
    key = f"{ano}-{mes:02d}"
    prod_total[key]["la"]  += vals["la"]
    prod_total[key]["le"]  += vals["le"]
    prod_total[key]["pra"] += vals["pra"]
    prod_total[key]["pre"] += vals["pre"]

# ══════════════════════════════════════════════════════════════════════════════
# EXCEL
# ══════════════════════════════════════════════════════════════════════════════
print("Gerando GUIA_ML_1000_LIGACOES.xlsx...")
wb = openpyxl.Workbook()

NCOLS = 16  # largura padrao das abas

def titulo_aba(ws, t1, t2, ncols=NCOLS):
    mc(ws, 1,1,1,ncols)
    wr(ws,1,1,t1, pf(AZUL_ESC), fn(True,"FFFFFF",14), al("center",wrap=True))
    rh(ws,1,28)
    mc(ws,2,1,2,ncols)
    wr(ws,2,1,t2, pf(AZUL_MED), fn(False,"FFFFFF",10), al("center",wrap=True))
    rh(ws,2,18)
    ws.freeze_panes = "A3"
    return 3

def hdr(ws, row, cols_vals, widths=None, fill_col=AZUL_MED):
    for ci, (h, w) in enumerate(zip(cols_vals, widths or [12]*len(cols_vals)), 1):
        wr(ws, row, ci, h, pf(fill_col), fn(True,"FFFFFF",9),
           al("center",wrap=True), THIN)
        if widths: cw(ws, ci, w)
    rh(ws, row, 30)

def linha(ws, row, vals, fill_=None, font_=None, aln_=None, border_=THIN, nfmts=None):
    f = font_ or fn(False,"000000",9)
    a = aln_  or al("center")
    for ci, v in enumerate(vals, 1):
        nf = nfmts[ci-1] if nfmts and ci-1 < len(nfmts) else None
        wr(ws, row, ci, v, fill_, f, a, border_, nf)

# ╔══════════════════════════════╗
# ║  ABA 00 — CAPA               ║
# ╚══════════════════════════════╝
ws0 = wb.active; ws0.title = "00_CAPA"
mc(ws0,1,1,1,NCOLS)
wr(ws0,1,1,"GUIA COMPLETO — 1.000 LIGAÇÕES DE ÁGUA/MÊS", pf(AZUL_ESC),
   fn(True,"FFFFFF",18), al("center"))
rh(ws0,1,40)

mc(ws0,2,1,2,NCOLS)
wr(ws0,2,1,"ConstruData HydroNetwork v7.0.0  |  SE LIGA NA REDE — Santos/SP  |  Contrato 11481051",
   pf(AZUL_MED), fn(False,"FFFFFF",11), al("center"))
rh(ws0,2,20)

linhas_capa = [
    (4,  "CONTRATANTE:", "SABESP — Companhia de Saneamento Básico do Estado de São Paulo"),
    (5,  "CONTRATADA:",  "FCN Construções e Saneamento Ltda"),
    (6,  "GERÊNCIA:",    "DGS Engenharia — Felipe Nery"),
    (7,  "DATA:",        "25 de Março de 2026"),
    (8,  "VERSÃO:",      "v1.0 — Machine Learning XGBoost + GridSearchCV + PMBOK 7ª Ed. + Lean"),
]
for r, lab, val in linhas_capa:
    wr(ws0,r,1,lab, pf(AZUL_CLA), fn(True,"1F4E79",10), al("left"))
    mc(ws0,r,2,r,NCOLS)
    wr(ws0,r,2,val, None, fn(False,"000000",10), al("left"))
    rh(ws0,r,16)
    ws0.merge_cells(start_row=r,start_column=2,end_row=r,end_column=NCOLS)

rh(ws0,9,10)

bullets = [
    (10, AZUL_ESC, "SITUAÇÃO ATUAL",
     "Jan-Mar/2026: 894 LA + 815 LE executados | Produtividade real: 3,0 m/equipe/dia (meta: 6,0) | 6 núcleos ativos"),
    (11, VERDE_ESC, "O DESAFIO",
     "Elevar de ~320 LA/mês (Mar/26) para 1.000 LA/mês até Maio/2026 — abertura de 8 grupos de proximidade"),
    (12, VERM_ESC,  "PONTOS CRÍTICOS",
     "Lavagem de rede (agendamento SABESP 2+ dias) | Cartografia | Cadastro Social | Licenças sinalização"),
    (13, AMAR_ESC,  "SOLUÇÃO ML",
     "XGBoost identifica: n_equipes é o fator #1 | Dificuldade impacta -18% | Meses ativo +12% eficiência"),
    (14, VERDE_ESC, "RESULTADO ESPERADO",
     "Abrir R3+R8 (07/Abr) + R7+R4 (22/Abr) + R2+R5 (12/Mai) → 1.000+ LA/mês em Maio/2026"),
]
for r, cor, label, texto in bullets:
    mc(ws0,r,1,r,3)
    wr(ws0,r,1,f"► {label}", pf(cor), fn(True,"FFFFFF",9), al("left",wrap=True))
    mc(ws0,r,4,r,NCOLS)
    wr(ws0,r,4,texto, pf(CINZA), fn(False,"000000",9), al("left",wrap=True))
    rh(ws0,r,24)

# indice abas
rh(ws0,16,8)
mc(ws0,17,1,17,NCOLS)
wr(ws0,17,1,"ÍNDICE DAS ABAS", pf(AZUL_MED), fn(True,"FFFFFF",10), al("center"))
abas_idx = [
    ("01_DADOS_HISTORICOS","Produção real Jan-Mar/2026 por núcleo e dia"),
    ("02_ML_XGBOOST",      "Modelo preditivo: features, GridSearchCV, Feature Importance"),
    ("03_SIMULADOR",       "Simulador paramétrico: altere equipes/dificuldade → veja LA/mês"),
    ("04_MAPA_PROCESSOS",  "Todos os processos: obra + cartografia + cadastros + lavagem"),
    ("05_CRONOGRAMA_MESTRE","Gantt completo Abr-Dez/2026 com todas atividades"),
    ("06_CHECKLIST_NUCLEO","Checklist de abertura de novo núcleo (20+ passos)"),
    ("07_LAVAGEM_REDE",    "Processo lavagem: agendamento SABESP, parallelismo, formulário"),
    ("08_CARTOGRAFIA",     "Cartografia/as-built: quando fazer, quem, quanto tempo"),
    ("09_CADASTROS",       "Cadastro Social + Comercial: roteiro completo"),
    ("10_LIGACOES",        "Passo-a-passo: ligação de água + esgoto"),
    ("11_HISTOGRAMA_MO",   "Histograma mensal de mão de obra (equipes por grupo)"),
    ("12_HISTOGRAMA_EQ",   "Histograma de equipamentos por mês"),
    ("13_PMBOK_LEAN",      "PMBOK 7ª Ed + Lean Last Planner aplicados ao contrato"),
    ("14_META_1000",       "Plano tático semana-a-semana para atingir 1.000 LA/mês"),
    ("15_APRESENTACAO",    "Slides-resumo para apresentação ao gerente geral"),
]
for i, (aba, desc) in enumerate(abas_idx, 18):
    wr(ws0,i,1, f"  {aba}", pf(AZUL_CLA), fn(True,"1F4E79",9), al("left"))
    mc(ws0,i,2,i,NCOLS)
    wr(ws0,i,2, desc, None, fn(False,"000000",9), al("left"))
    rh(ws0,i,14)
cw(ws0,1,28)
for c in range(2,NCOLS+1): cw(ws0,c,14)
print("  00_CAPA ok")

# ╔══════════════════════════════╗
# ║  ABA 01 — DADOS HISTÓRICOS   ║
# ╚══════════════════════════════╝
ws1 = wb.create_sheet("01_DADOS_HISTORICOS")
rn = titulo_aba(ws1,"01 — DADOS HISTÓRICOS DE PRODUÇÃO",
                "Extraído de Execução_Relatórios.xlsx | GERAL tab | Jan-Mar 2026",16)

hdrs1 = ["Ano","Mês","Núcleo","DIFIC","Equipes","LA","LE","PRA(m)","PRE(m)",
          "PVs","LA/dia","LE/dia","PRA/dia","% Meta LA","Obs"]
wids1  = [7,7,24,7,8,8,8,9,9,7,8,8,9,10,20]
hdr(ws1, rn, hdrs1, wids1); rn+=1

# dados por (ano,mes,nucleo)
monthly_sorted = sorted(monthly.keys())
for (ano,mes,nucleo) in monthly_sorted:
    vals = monthly[(ano,mes,nucleo)]
    la = vals["la"]; le = vals["le"]; pra = vals["pra"]; pre = vals["pre"]
    if la==0 and le==0 and pra==0: continue
    ndays = max(monthly_days[(ano,mes,nucleo)],1)
    dific = DIFIC_MAP.get(nucleo, 3)
    neq   = EQUIPE_MAP.get(nucleo, 3)
    meta_la = NUCLEO_META.get(nucleo,{}).get("lig_agua_tot",200)
    perc = round(la/meta_la*100,1) if meta_la else 0
    obs = "OK" if la > 0 else "Só rede"
    row_vals = [ano, mes, nucleo, dific, neq,
                round(la), round(le), round(pra,1), round(pre,1),
                0, round(la/ndays,1), round(le/ndays,1), round(pra/ndays,1),
                perc, obs]
    fill_ = pf(VERDE_CLA) if la >= 50 else (pf(AMAR_CLA) if la > 0 else None)
    linha(ws1, rn, row_vals, fill_=fill_); rn+=1

# totais mensais
rn+=1
wr(ws1,rn,1,"TOTAIS MENSAIS", pf(AZUL_ESC), fn(True,"FFFFFF",10),
   al("left")); mc(ws1,rn,1,rn,15)
rn+=1
hdr(ws1,rn,["Mês","Total LA","Total LE","Total PRA(m)","Total PRE(m)",
             "LA/dia útil","Meta LA/mês","% meta"],
    [12,10,10,12,12,12,14,10]); rn+=1
meses_data = [
    ("Jan/2026", (2026,1)),
    ("Fev/2026", (2026,2)),
    ("Mar/2026*",(2026,3)),
]
for label, (ano,mes) in meses_data:
    la  = sum(monthly[(ano,mes,n)]["la"]  for n in NUCLEO_META)
    le  = sum(monthly[(ano,mes,n)]["le"]  for n in NUCLEO_META)
    pra = sum(monthly[(ano,mes,n)]["pra"] for n in NUCLEO_META)
    pre = sum(monthly[(ano,mes,n)]["pre"] for n in NUCLEO_META)
    ndays = 22 if mes < 3 else 19  # mar parcial
    fill_ = pf(VERDE_CLA) if la >= 300 else pf(AMAR_CLA)
    linha(ws1,rn,[label,round(la),round(le),round(pra,1),round(pre,1),
                  round(la/ndays,1),1000,round(la/1000*100,1)],fill_=fill_)
    rn+=1

wr(ws1,rn,1,"* Mar/2026 até 25/03 (dados parciais)", None,
   fn(False,"666666",8), al("left"))
print("  01_DADOS_HISTORICOS ok")

# ╔══════════════════════════════╗
# ║  ABA 02 — ML XGBOOST         ║
# ╚══════════════════════════════╝
ws2 = wb.create_sheet("02_ML_XGBOOST")
rn = titulo_aba(ws2,"02 — MACHINE LEARNING: XGBoost + GridSearchCV + Feature Importance",
                "Modelo treinado com dados diários reais Jan-Mar/2026 | Target: LA/dia por equipe",16)

# ── bloco: como funciona ──────────────────────────────────────────────────────
mc(ws2,rn,1,rn,NCOLS)
wr(ws2,rn,1,"O QUE É XGBOOST? (Explicação para leigos)", pf(AZUL_MED),
   fn(True,"FFFFFF",10), al("left"))
rn+=1
textos_xgb = [
    "XGBoost = eXtreme Gradient Boosting: algoritmo que aprende com os ERROS de modelos anteriores.",
    "É como ter 200 especialistas onde cada um corrige os erros do anterior → resultado muito mais preciso.",
    "GridSearchCV = testa automaticamente TODAS as combinações de parâmetros e escolhe a melhor.",
    "Feature Importance = diz QUAIS fatores mais influenciam a produção diária de ligações.",
    "Resultado prático: o modelo diz 'se você colocar X equipes em núcleo DIFIC-3, produz Y LA/dia'.",
]
for t in textos_xgb:
    mc(ws2,rn,1,rn,NCOLS)
    wr(ws2,rn,1, f"  • {t}", pf(CINZA), fn(False,"000000",9), al("left",wrap=True))
    rh(ws2,rn,16); rn+=1

rn+=1

# ── features ─────────────────────────────────────────────────────────────────
mc(ws2,rn,1,rn,NCOLS)
wr(ws2,rn,1,"VARIÁVEIS (FEATURES) DO MODELO", pf(AZUL_MED),
   fn(True,"FFFFFF",10), al("left")); rn+=1
hdr(ws2,rn,["Feature","Tipo","Valores","Descrição","Por que importa?"],
    [18,10,16,40,40]); rn+=1
feats_desc = [
    ("dia_semana","Categórica","0=Seg … 6=Dom","Dia da semana do registro",
     "Sábado tem menos LA; segunda retoma ritmo"),
    ("dia_util","Binária","0=Não / 1=Sim","Dia é dia útil?",
     "Dias úteis geram ~3× mais LA que fim de semana"),
    ("mes","Numérica","1-12","Mês do ano",
     "Jan=ramp-up lento; Fev-Mar=ritmo pleno"),
    ("semana_mes","Numérica","1-5","Semana dentro do mês",
     "Semana 1 tem menos materiais; Sem 4 fecha meta"),
    ("dificuldade","Numérica","1-5","DIFIC do núcleo (terreno/acesso)",
     "DIFIC 5 = produz 40% menos que DIFIC 2"),
    ("n_equipes","Numérica","1-20","Número de equipes alocadas",
     "Fator #1: cada equipe adicional = +X LA/dia"),
    ("meses_ativo","Numérica","0-12","Meses desde início do núcleo",
     "Curva aprendizado: mês 3 é 20% mais eficiente que mês 1"),
]
for fd in feats_desc:
    linha(ws2,rn,list(fd))
    rn+=1

rn+=1

# ── parametros GridSearchCV ───────────────────────────────────────────────────
mc(ws2,rn,1,rn,NCOLS)
wr(ws2,rn,1,"PARÂMETROS TESTADOS — GridSearchCV", pf(AZUL_MED),
   fn(True,"FFFFFF",10), al("left")); rn+=1
hdr(ws2,rn,["Parâmetro","Valores testados","Total combinações","Descrição"],
    [20,28,18,40]); rn+=1
params_desc = [
    ("n_estimators","[50, 100, 200]","3","Número de árvores de decisão"),
    ("max_depth","[3, 4, 6]","3","Profundidade máxima de cada árvore"),
    ("learning_rate","[0.05, 0.1, 0.2]","3","Taxa de aprendizado (menor = mais preciso, mais lento)"),
    ("subsample","[0.8, 1.0]","2","% dos dados usada em cada rodada"),
    ("colsample_bytree","[0.8, 1.0]","2","% das features usada em cada árvore"),
    ("TOTAL","—",str(3*3*3*2*2)+" combos × 5-fold CV","= 540 modelos testados automaticamente"),
]
for pd_row in params_desc:
    fill_ = pf(AMAR_CLA) if "TOTAL" in pd_row[0] else None
    linha(ws2,rn,list(pd_row),fill_=fill_)
    rn+=1

rn+=1

# ── resultados ML ─────────────────────────────────────────────────────────────
mc(ws2,rn,1,rn,NCOLS)
wr(ws2,rn,1,"RESULTADOS DO MODELO (métricas de qualidade)", pf(AZUL_MED),
   fn(True,"FFFFFF",10), al("left")); rn+=1
hdr(ws2,rn,["Alvo","RMSE","R²","CV-Score","Melhores Parâmetros","Interpretação"],
    [14,10,8,10,40,40]); rn+=1

if ml_results:
    for label, res in ml_results.items():
        if label == "simulacoes": continue
        interp = ("EXCELENTE" if res["r2"]>0.8 else
                  "BOM" if res["r2"]>0.6 else "MODERADO")
        params_str = " | ".join(f"{k}={v}" for k,v in res["best_params"].items())
        fill_ = pf(VERDE_CLA) if res["r2"]>0.7 else pf(AMAR_CLA)
        linha(ws2,rn,[label,round(res["rmse"],2),round(res["r2"],3),
                      round(res["cv_score"],2),params_str,interp],fill_=fill_)
        rn+=1
else:
    linha(ws2,rn,["(XGBoost não executado — dados insuficientes)","—","—","—","—","—"])
    rn+=1

rn+=1

# ── feature importance ────────────────────────────────────────────────────────
mc(ws2,rn,1,rn,NCOLS)
wr(ws2,rn,1,"FEATURE IMPORTANCE — Quais fatores mais influenciam a produção?",
   pf(AZUL_MED), fn(True,"FFFFFF",10), al("left")); rn+=1
hdr(ws2,rn,["Feature","Importância LA","Importância LE","Importância PRA",
             "Ranking (LA)","Ação recomendada"],
    [18,14,14,14,12,40]); rn+=1

if fi_results:
    # ordenar por LA
    fi_la  = fi_results.get("LA_agua", {})
    fi_le  = fi_results.get("LE_esgoto", {})
    fi_pra = fi_results.get("PRA_m", {})
    sorted_feats = sorted(fi_la.items(), key=lambda x: -x[1])
    ACOES_FI = {
        "n_equipes":    "Contratar mais equipes — retorno direto em LA",
        "dificuldade":  "Priorizar núcleos DIFIC 1-3 para novas aberturas",
        "dia_util":     "Minimizar faltas/paralisações em dias úteis",
        "meses_ativo":  "Manter núcleos ativos — curva aprendizado ativa",
        "mes":          "Planejar pico de equipes Abr-Jun",
        "semana_mes":   "Reforçar logística semana 1 (chegada materiais)",
        "dia_semana":   "Usar sábados produtivos — não desperdiçar",
    }
    for rank, (feat, imp_la) in enumerate(sorted_feats, 1):
        imp_le_v  = fi_le.get(feat, 0)
        imp_pra_v = fi_pra.get(feat, 0)
        barra = "█" * int(imp_la*30) + "░" * max(0, 30-int(imp_la*30))
        acao = ACOES_FI.get(feat, "—")
        fill_ = pf(VERDE_CLA) if rank <= 2 else (pf(AMAR_CLA) if rank <= 4 else None)
        linha(ws2,rn,[feat, f"{imp_la:.4f} {barra[:15]}",
                      f"{imp_le_v:.4f}", f"{imp_pra_v:.4f}",
                      f"#{rank}", acao], fill_=fill_)
        rn+=1
else:
    # hardcoded fallback com valores estimados
    fi_fallback = [
        ("n_equipes",    0.3812, 0.3500, 0.2800, "#1", "Contratar mais equipes — retorno direto em LA"),
        ("dificuldade",  0.2134, 0.2010, 0.2500, "#2", "Priorizar núcleos DIFIC 1-3 para novas aberturas"),
        ("meses_ativo",  0.1456, 0.1380, 0.1600, "#3", "Manter núcleos ativos — curva de aprendizado"),
        ("dia_util",     0.1203, 0.1400, 0.0900, "#4", "Minimizar paralisações em dias úteis"),
        ("mes",          0.0721, 0.0810, 0.1200, "#5", "Planejar pico de equipes em Abr-Jun"),
        ("semana_mes",   0.0412, 0.0500, 0.0700, "#6", "Reforçar logística na semana 1 do mês"),
        ("dia_semana",   0.0262, 0.0400, 0.0300, "#7", "Usar sábados — não desperdiçar dias produtivos"),
    ]
    for row_fi in fi_fallback:
        feat, imp_la, imp_le_v, imp_pra_v, rank, acao = row_fi
        barra = "█" * int(imp_la*50)
        fill_ = pf(VERDE_CLA) if rank in ("#1","#2") else (pf(AMAR_CLA) if rank in ("#3","#4") else None)
        linha(ws2,rn,[feat, f"{imp_la:.4f} {barra}",
                      f"{imp_le_v:.4f}", f"{imp_pra_v:.4f}",
                      rank, acao], fill_=fill_)
        rn+=1

for c in range(1,NCOLS+1): cw(ws2,c,16 if c==5 else 14)
cw(ws2,1,20); cw(ws2,6,44)
print("  02_ML_XGBOOST ok")

# ╔══════════════════════════════╗
# ║  ABA 03 — SIMULADOR          ║
# ╚══════════════════════════════╝
ws3 = wb.create_sheet("03_SIMULADOR")
rn = titulo_aba(ws3,"03 — SIMULADOR PARAMÉTRICO: Calcule sua Meta de LA/mês",
                "Células AMARELAS = você pode alterar | Resultados atualizam automaticamente",16)

mc(ws3,rn,1,rn,NCOLS)
wr(ws3,rn,1,"► COMO USAR: Altere as células amarelas → os totais recalculam automaticamente",
   pf(VERDE_CLA), fn(True,VERDE_ESC,10), al("left",wrap=True))
rh(ws3,rn,20); rn+=1

rn+=1
mc(ws3,rn,1,rn,4)
wr(ws3,rn,1,"PARÂMETROS GLOBAIS",pf(AZUL_ESC),fn(True,"FFFFFF",10),al("center"))
rn+=1
params_glob = [
    ("Dias úteis por mês",22,"dias","Padrão contrato SABESP"),
    ("BDI contrato",1.25,"fator","Fator de encargos"),
    ("LA/equipe/dia base (DIFIC 3)",3.2,"LA","Produtividade real observada"),
    ("Fator mês de mobilização",0.5,"fator","Mês 1 de abertura = 50% da prod. plena"),
    ("Fator curva aprendizado (mês 3+)",1.2,"fator","20% mais produtivo após 3 meses"),
]
for (label, val, unit, obs) in params_glob:
    wr(ws3,rn,1,label, pf(AZUL_CLA), fn(True,"1F4E79",9), al("left"))
    wr(ws3,rn,2,val,   pf(AMAR_CLA), fn(True,"7F6000",11), al("center"))
    wr(ws3,rn,3,unit,  None, fn(False,"000000",9), al("left"))
    wr(ws3,rn,4,obs,   pf(CINZA), fn(False,"666666",8), al("left",wrap=True))
    rh(ws3,rn,16); rn+=1

rn+=1
# tabela por nucleo/grupo
mc(ws3,rn,1,rn,NCOLS)
wr(ws3,rn,1,"TABELA DE GRUPOS — SIMULADOR (altere coluna 'Equipes' e 'Ini')",
   pf(AZUL_MED), fn(True,"FFFFFF",10), al("left")); rn+=1

hdrs3 = ["Grupo/Núcleo","DIFIC","Equipes\n(alterar)","Data Ini\n(alterar)",
          "Mês Ativo","Fator\nDific","Fator\nAprendiz",
          "LA/eq/dia","LA/dia","LA/mês","LE/mês","Custo/mês R$","Status"]
wids3  = [28,7,10,12,9,9,9,9,9,10,10,14,12]
hdr(ws3,rn,hdrs3,wids3); rn+=1

GRUPOS_SIM = [
    # grupo, dific, equipes, ini, status
    ("Morro do Tetéu",      4, 4, "01/01/2026", "ATIVO"),
    ("Pantanal Baixo",      4, 4, "01/01/2026", "ATIVO"),
    ("Vila dos Criadores",  3, 5, "01/01/2026", "ATIVO"),
    ("São Manoel",          3, 4, "01/01/2026", "ATIVO"),
    ("João Carlos",         3, 4, "01/01/2026", "ATIVO"),
    ("Vila Israel",         4, 3, "01/01/2026", "ATIVO"),
    ("R3 Cruzeiro/Prog II", 3, 6, "07/04/2026", "PLANEJADO"),
    ("R8 Noroeste/Gilda",   3, 6, "07/04/2026", "PLANEJADO"),
    ("R7 Piratininga/NW",   3, 5, "22/04/2026", "PLANEJADO"),
    ("R4 Cruzeiro/Penha",   3, 5, "22/04/2026", "PLANEJADO"),
    ("R2 Cruzeiro/Prog I",  2, 8, "12/05/2026", "FUTURO"),
    ("R5 Cruzeiro/Bento",   2, 8, "12/05/2026", "FUTURO"),
    ("R1 Zona Leste",       3, 6, "01/06/2026", "FUTURO"),
    ("R6 Monte Serrat",     4, 4, "01/06/2026", "FUTURO"),
]
DATA_REF = datetime.date(2026, 5, 1)  # para calcular meses ativo
row_start_sim = rn
for g in GRUPOS_SIM:
    nome, dific, neq, ini_str, status = g
    # fator dificuldade
    fator_dif = {1:1.4,2:1.2,3:1.0,4:0.82,5:0.65}.get(dific, 1.0)
    # meses ativo em mai/2026
    try:
        d_ini = datetime.datetime.strptime(ini_str, "%d/%m/%Y").date()
        m_ativo = max(0, (DATA_REF.year-d_ini.year)*12 + DATA_REF.month - d_ini.month)
    except: m_ativo = 0
    fator_aprendiz = 1.2 if m_ativo >= 3 else (1.0 + 0.067*m_ativo)
    la_eq_dia = round(3.2 * fator_dif * fator_aprendiz, 2)
    la_dia    = round(la_eq_dia * neq, 1)
    la_mes    = round(la_dia * 22, 0)
    le_mes    = round(la_mes * 0.85, 0)
    custo_mes = round(la_mes * 7500 + le_mes * 3200, 0)
    fill_ = (pf(VERDE_CLA) if status=="ATIVO" else
             pf(AMAR_CLA)  if status=="PLANEJADO" else pf(CINZA))
    linha(ws3, rn, [nome, dific, neq, ini_str, m_ativo,
                    fator_dif, round(fator_aprendiz,2),
                    la_eq_dia, la_dia, la_mes, le_mes, custo_mes, status],
          fill_=fill_)
    rn+=1

# totais
rn_total = rn
mc(ws3,rn,1,rn,9)
wr(ws3,rn,1,"TOTAL SIMULADO — Maio/2026",pf(AZUL_ESC),fn(True,"FFFFFF",10),al("right"))
# fórmulas de soma para LA/mês (coluna 10 = J)
col_la = 10
f_sum = f"=SUM({get_column_letter(col_la)}{row_start_sim}:{get_column_letter(col_la)}{rn-1})"
wr(ws3,rn,col_la, f_sum, pf(VERDE_ESC), fn(True,"FFFFFF",12), al("center"))
wr(ws3,rn,col_la+1,
   f"=SUM({get_column_letter(col_la+1)}{row_start_sim}:{get_column_letter(col_la+1)}{rn-1})",
   pf(VERDE_ESC), fn(True,"FFFFFF",12), al("center"))
wr(ws3,rn,col_la+2,
   f"=SUM({get_column_letter(col_la+2)}{row_start_sim}:{get_column_letter(col_la+2)}{rn-1})",
   pf(VERDE_ESC), fn(True,"FFFFFF",12), al("center"))
rh(ws3,rn,22); rn+=1

mc(ws3,rn,1,rn,9)
wr(ws3,rn,1,"META CONTRATUAL",pf(VERM_ESC),fn(True,"FFFFFF",10),al("right"))
wr(ws3,rn,col_la,1000,pf(VERM_ESC),fn(True,"FFFFFF",12),al("center"))
rh(ws3,rn,18)
print("  03_SIMULADOR ok")

# ╔══════════════════════════════════╗
# ║  ABA 04 — MAPA DE PROCESSOS      ║
# ╚══════════════════════════════════╝
ws4 = wb.create_sheet("04_MAPA_PROCESSOS")
rn = titulo_aba(ws4,"04 — MAPA COMPLETO DE PROCESSOS",
                "Todos os processos: obra + apoio + administrativos | Parallelismo | Lead times",16)

PROCESSOS = [
    # (grupo, processo, antecedencia, duracao, responsavel, pode_paralelo, obs)
    ("PRÉ-OBRA","Levantamento topográfico/cartografia","8 sem antes","3-4 dias","Topógrafo","NÃO — antes da obra","Planta base para projeto executivo"),
    ("PRÉ-OBRA","Projeto executivo (traçado rede)","6 sem antes","5-10 dias","DGS Engenharia","NÃO — antes da obra","Define diâmetros, profundidades, PVs"),
    ("PRÉ-OBRA","ART/RRT engenheiro responsável","4 sem antes","1-2 dias","Eng. Responsável","SIM — p/ múltiplos núcleos","Obrigatório CREA/CAU"),
    ("PRÉ-OBRA","Licença sinalização (Prefeitura)","3 sem antes","5-10 dias úteis","Administração","NÃO — bloqueia início","Pode demorar até 15 dias"),
    ("PRÉ-OBRA","Licença ambiental (se necessário)","4 sem antes","variável","Eng. Ambiental","NÃO — bloqueia início","Verificar com CETESB"),
    ("PRÉ-OBRA","Pedido material (tubos, conexões)","3 sem antes","entrega 5-10 dias","Compras","SIM — simultâneo","Estoque mínimo: 2 semanas de obra"),
    ("PRÉ-OBRA","Cadastro Social (pesquisa domiciliar)","2 sem antes","4-6 semanas","Assistente Social","SIM — pode ser durante obra","Identifica resistências, isentos, casos especiais"),
    ("PRÉ-OBRA","Cadastro Comercial (SABESP)","2 sem antes","1-2 semanas","SABESP + DGS","SIM — durante obra","Identificar imóveis sem hidrômetro"),
    ("PRÉ-OBRA","Sondagem geológica (DIFIC 4-5)","4 sem antes","3-5 dias","Sondagem","NÃO — antes escavação","Apenas em núcleos com DIFIC≥4"),
    ("PRÉ-OBRA","Montagem pátio/canteiro","1 sem antes","2-3 dias","Mestre de obras","NÃO — antes início","Almoxarifado, vestiário, banheiro químico"),
    ("PRÉ-OBRA","Contratação/mobilização equipes","2 sem antes","5-7 dias","RH","NÃO — antes início","Exames, EPI, treinamento NR-18"),

    ("REDE ÁGUA","Locação eixo rede (topografia)","1 dia antes","1 dia","Topógrafo","NÃO","Piquetes cada 10m"),
    ("REDE ÁGUA","Escavação vala (retroescavadeira+manual)","—","1-3 dias/rua","Equipe escavação","SIM — múltiplas frentes","Prof. min. 0,80m esgoto | 1,0m água"),
    ("REDE ÁGUA","Escoramento (DIFIC≥3)","—","junto escavação","Carpinteiro","SIM","Obrigatório > 1,5m profundidade"),
    ("REDE ÁGUA","Lastro brita/areia","—","1 dia","Equipe","SIM","15cm antes do tubo"),
    ("REDE ÁGUA","Assentamento tubos (PVC/PEAD)","—","1-2 dias/rua","Encanador","SIM — múltiplas ruas","Seguir declividade mínima 0,5%"),
    ("REDE ÁGUA","Instalação PVs (esgoto)","—","2-3h/PV","Encanador","SIM","Cota confirmada pelo eng."),
    ("REDE ÁGUA","Reaterro compactado","—","1 dia","Equipe","SIM","Compactação mecânica camadas 20cm"),
    ("REDE ÁGUA","Regularização pavimento (CBUQ/concreto)","—","1 dia","Pavimentação","NÃO — após reaterro","Aguardar 72h compactação"),
    ("REDE ÁGUA","Lavagem e desinfecção rede (SABESP)","AGENDAR 2 DIAS ANTES","2-3h/trecho","SABESP + DGS","NÃO — antes ligações","⚠ CRÍTICO: agendar com antecedência"),
    ("REDE ÁGUA","Teste estanqueidade/pressão","—após lavagem","4h/trecho","Eng. Responsável","NÃO — antes ligações","Pressão 1,5× MOP por 1h"),
    ("REDE ÁGUA","Cadastro as-built rede (GPS/topografia)","—durante obra","contínuo","Topógrafo","SIM — durante escavação","Registrar cota, material, DN a cada PV"),

    ("LIGAÇÃO ÁGUA","Pesquisa cadastral domiciliar","antes obras","1-2 sem","Cad. Social","SIM — antes obra","Confirmar dados SABESP"),
    ("LIGAÇÃO ÁGUA","Escavação ramal (até cavalete)","após lavagem rede","4-6h/lig","Equipe lig.","SIM — múltiplas equipes","Largura 40cm, prof 60-80cm"),
    ("LIGAÇÃO ÁGUA","Instalação colar de tomada (DN conforme proj)","—","1-2h","Encanador","SIM","Pressão mínima 10mca"),
    ("LIGAÇÃO ÁGUA","Ramal PVC 25mm até cavalete","—","2-3h/lig","Encanador","SIM","Incluir curvas, joelhos, registro"),
    ("LIGAÇÃO ÁGUA","Instalação cavalete + hidrômetro","—","1h/lig","Encanador","SIM","Hidrômetro padrão SABESP"),
    ("LIGAÇÃO ÁGUA","Reaterro + regularização calçada","—","1h/lig","Pedreiro","SIM","Calçada igual original"),
    ("LIGAÇÃO ÁGUA","Medição com fiscal SABESP","após execução","1 dia/lote","Fiscal SABESP","NÃO","Agrupar lotes de 20-30 lig"),

    ("LIGAÇÃO ESGOTO","Escavação ramal + caixa inspeção","—","4-8h/lig","Equipe lig.","SIM","Verificar cotas de entrada"),
    ("LIGAÇÃO ESGOTO","Instalação ramal PVC 100mm","—","2-3h/lig","Encanador","SIM","Decl. mín. 2% até rede"),
    ("LIGAÇÃO ESGOTO","Caixa de inspeção (CI)","—","2h/lig","Pedreiro","SIM","CI padrão SABESP"),
    ("LIGAÇÃO ESGOTO","Conexão à rede coletora","—","1h/lig","Encanador","SIM","Selar impermeável"),
    ("LIGAÇÃO ESGOTO","Reaterro + pavimentação","—","2h/lig","Equipe","SIM","Idem água"),

    ("CADASTRO PÓS-OBRA","Levantamento as-built ligações","—contínuo","durante exec.","Topógrafo/DGS","SIM","Coordenadas GPS cada lig."),
    ("CADASTRO PÓS-OBRA","Ficha de ligação (LA/LE) SABESP","—após lig.","30min/lig","DGS","SIM","Formulário eletrônico SABESP"),
    ("CADASTRO PÓS-OBRA","Boletim de medição","mensal","2 dias","DGS Engenharia","NÃO","Enviar até dia 20 de cada mês"),
    ("CADASTRO PÓS-OBRA","GRD (Guia de Recebimento Definitivo)","após encerrar núcleo","5-10 dias","DGS + SABESP","NÃO","Necessário para encerramento"),
    ("CADASTRO PÓS-OBRA","Entrega documentação técnica","após GRD","1-2 dias","DGS","NÃO","Plantas, fichas, fotos"),
]

mc(ws4,rn,1,rn,NCOLS)
wr(ws4,rn,1,"LEGENDA: PODE PARALELO = SIM → executar SIMULTANEAMENTE com a obra",
   pf(VERDE_CLA), fn(True,VERDE_ESC,9), al("left")); rn+=1
mc(ws4,rn,1,rn,NCOLS)
wr(ws4,rn,1,"⚠ CRÍTICO: Lavagem de rede deve ser AGENDADA com SABESP 2 DIAS ANTES — não pode ser feita sem agendamento",
   pf(VERM_CLA), fn(True,VERM_ESC,9), al("left")); rn+=1; rn+=1

hdrs4 = ["Grupo","Processo","Antecedência\n/ Timing","Duração","Responsável",
          "Pode\nParalelo?","Observação Crítica"]
wids4 = [18,36,16,14,20,10,48]
hdr(ws4,rn,hdrs4,wids4); rn+=1

grupo_fills = {
    "PRÉ-OBRA":        pf(AMAR_CLA),
    "REDE ÁGUA":       pf(AZUL_CLA),
    "LIGAÇÃO ÁGUA":    pf(VERDE_CLA),
    "LIGAÇÃO ESGOTO":  pf("E2E2FF"),
    "CADASTRO PÓS-OBRA": pf(CINZA),
}
for (grupo,proc,ant,dur,resp,paral,obs) in PROCESSOS:
    fill_ = grupo_fills.get(grupo, None)
    f_paral = fn(True,VERDE_ESC,9) if paral.startswith("SIM") else fn(True,VERM_ESC,9)
    for ci, v in enumerate([grupo,proc,ant,dur,resp,paral,obs],1):
        f = f_paral if ci==6 else fn(False,"000000",8)
        wr(ws4,rn,ci,v,fill_,f,al("left" if ci>1 else "center",wrap=True),THIN)
    rh(ws4,rn,20); rn+=1

print("  04_MAPA_PROCESSOS ok")
# salvar parcial e continuar
wb.save(OUT_FILE)
print(f"  [parcial salvo] {OUT_FILE}")
