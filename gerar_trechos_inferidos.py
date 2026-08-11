#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Gera TRECHOS_NUCLEOS_INFERIDOS.xlsx
Lê 'todos os nucleos.xlsx' (68 núcleos) e gera trechos PV-a-PV sintéticos
por similaridade com os 6 núcleos reais calibrados do projeto SLNR Santos.

Parâmetros calibrados dos núcleos reais:
  Morro Teteu      avg_L=16.7m DN≈200 prof=0.67m CT=[2..125]  → Morro
  Pantanal Baixo   avg_L=17.0m DN≈220 prof=0.78m CT=[-3..2]   → Dique/Baixo
  Sao Manoel       avg_L=17.2m DN≈200 prof=0.70m CT=[1..2]    → Baixo urbano
  PROL Teteu       avg_L=43.7m DN≈190 prof=2.39m CT=[0..11]   → Prolongamento
  Vila Criadores   avg_L=22.3m DN≈250 prof=0.97m CT=[-3..1]   → Urbano baixo
  Vila Israel      avg_L=12.0m DN≈210 prof=0.68m CT=[60..108] → Alto morro
"""
import sys, io, os, json, math, random
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

IN_FILE  = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        "TRECHOS_NS_COMPLETOS", "todos os nucleos.xlsx")
OUT_DIR  = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        "TRECHOS_NS_COMPLETOS")
OUT_FILE = os.path.join(OUT_DIR, "TRECHOS_NUCLEOS_INFERIDOS.xlsx")

# ────────── PERFIS DE TERRENO ──────────
# Each profile: avg_L_m, dn_mm, prof_ini, prof_fim_delta, ct_base, ct_slope_pct, pav
PERFIS = {
    # terrain_key → (avg_L, dn, prof_base, prof_var, ct_base_min, ct_base_max, slope_pct_min, slope_pct_max, pavimento)
    "dique":    (18, 200, 1.20, 0.30, -2.0,  2.0,  0.3,  0.8,  "CIMENTADO"),
    "morro":    (17, 200, 1.80, 0.80, 15.0, 80.0,  3.0,  8.0,  "CIMENTADO"),
    "monte":    (17, 200, 2.20, 1.00, 20.0,100.0,  4.0, 10.0,  "CIMENTADO"),
    "encosta":  (20, 200, 1.60, 0.60, 10.0, 50.0,  2.0,  5.0,  "CIMENTADO"),
    "urbano":   (20, 200, 1.50, 0.40,  4.0, 25.0,  0.8,  2.5,  "CIMENTADO"),
    "urbano_a": (22, 200, 1.40, 0.35,  2.0, 15.0,  0.5,  1.5,  "CBUQ"),
    "prolongamento": (44, 200, 2.40, 0.50,  0.0, 10.0,  0.5,  2.0,  "CIMENTADO"),
}

CUSTO_TUBO = {100:120, 150:160, 200:200, 250:270, 300:340, 350:420, 400:500, 500:700}

# ────────── CLASSIFICAÇÃO ──────────
def classificar(nome, tipo_ter, ext_m, ligs):
    n = nome.upper()
    if any(x in n for x in ["DIQUE","ALAGADO","PALAFITA","REGULARIZADO_AG"]):
        return "dique"
    if "MONTE SERRAT" in n:
        return "monte"
    if any(x in n for x in ["MORRO","MORRO DO"]):
        return "morro"
    if any(x in n for x in ["PENHA","LOMBA","ENCOSTA","MORRO"]):
        return "encosta"
    if "Alagado" in tipo_ter or "Dique" in tipo_ter:
        return "dique"
    if ext_m and ext_m > 3000 and ligs and ligs > 400:
        return "prolongamento"
    if any(x in n for x in ["AV.","AV ","RUA","TRAVESSIA","DIVISA"]):
        return "urbano_a"
    return "urbano"


def dn_por_ligs(ligs):
    if not ligs or ligs < 60:    return 150
    if ligs < 220:               return 200
    if ligs < 450:               return 200
    if ligs < 700:               return 250
    return 300


def pavimento_por_nome(nome):
    n = nome.upper()
    if any(x in n for x in ["AV.","AVENIDA","AV ","BANDEIRANTES","MARTINS FONTES"]): return "CBUQ"
    if any(x in n for x in ["MORRO","MONTE","ENCOSTA","DIQUE","PALAFITA"]):           return "CIMENTADO"
    if any(x in n for x in ["TRAVESSIA","RUA","R."]): return "PARALELEPIPEDO"
    return "CIMENTADO"


# ────────── GERADOR DE TRECHOS ──────────
def gerar_trechos(tag, nome, ext_m, ligs, tipo_ter):
    """Gera lista de trechos sintéticos realistas para o núcleo."""
    if not ext_m or ext_m <= 0:
        return []

    perf_key = classificar(nome, tipo_ter, ext_m, ligs)
    avg_L, dn_def, prof_b, prof_v, ct_min, ct_max, sl_min, sl_max, pav_def = PERFIS[perf_key]

    dn   = dn_por_ligs(ligs)
    pav  = pavimento_por_nome(nome)

    # Semente determinística por TAG para reprodutibilidade
    rng = random.Random(int(tag) if str(tag).isdigit() else hash(tag) % 10000)

    # Número de trechos
    n_trechos = max(1, round(ext_m / avg_L))

    # Elevação inicial (sorteada dentro do range do perfil)
    ct_ini = rng.uniform(ct_min, ct_max)

    # Declive geral do trecho (descida ou subida)
    slope_pct = rng.uniform(sl_min, sl_max) / 100.0
    slope_sign = rng.choice([-1, 1])  # descida ou subida

    trechos = []
    ct_cur = ct_ini

    # Distribuição da extensão em n_trechos (variação ±30%)
    comprimentos = []
    for _ in range(n_trechos):
        l = avg_L * rng.uniform(0.7, 1.3)
        comprimentos.append(l)
    # Normalizar para somar ext_m
    fator = ext_m / sum(comprimentos)
    comprimentos = [c * fator for c in comprimentos]

    for i, L in enumerate(comprimentos):
        pv_ini = f"PV-{tag}-{i+1:03d}"
        pv_fim = f"PV-{tag}-{i+2:03d}"

        # Profundidade com variação
        prof_i = max(0.60, prof_b + rng.uniform(-prof_v, prof_v))
        prof_f = max(0.60, prof_b + rng.uniform(-prof_v, prof_v))

        # Cota terreno (CT)
        delta_ct = L * slope_pct * slope_sign
        # Adiciona micro-variação de terreno
        delta_ct += rng.uniform(-0.3, 0.3)
        ct_fim = ct_cur + delta_ct

        cf_ini = ct_cur - prof_i
        cf_fim = ct_fim - prof_f

        # Declividade do coletor (m/m)
        decl = (cf_ini - cf_fim) / L if L > 0 else 0.005
        decl = max(0.002, abs(decl))  # mínimo ProSaneamento 0.002

        # Velocidade Manning n=0.013, raio hidráulico estimado
        r_h = (dn / 1000) / 4.0
        v_ms = (1.0 / 0.013) * (r_h ** (2/3)) * (decl ** 0.5)
        area = math.pi * (dn/1000)**2 / 4
        q_ls = v_ms * area * 1000  # L/s

        trechos.append({
            "pv_ini": pv_ini, "pv_fim": pv_fim,
            "ext_m": round(L, 2), "dn_mm": dn,
            "material": "PVC",
            "ct_ini": round(ct_cur, 3), "cf_ini": round(cf_ini, 3), "prof_ini": round(prof_i, 2),
            "ct_fim": round(ct_fim, 3),  "cf_fim": round(cf_fim, 3),  "prof_fim": round(prof_f, 2),
            "decl_mm": round(decl * 1000, 2),
            "v_ms": round(v_ms, 3), "q_ls": round(q_ls, 3),
            "pavimento": pav,
            "rua": nome.split("_", 1)[-1] if "_" in nome else nome,
            "perf_key": perf_key,
        })
        ct_cur = ct_fim

    return trechos


def preco_tubo(dn):
    return CUSTO_TUBO.get(dn, dn * 1.1)


def calcular_quant(t):
    L    = t["ext_m"]
    prof_m = (t["prof_ini"] + t["prof_fim"]) / 2
    escav    = round(L * prof_m, 3)
    reaterro = round(escav * 0.90, 3)
    bota     = round(escav * 0.10, 3)
    escoram  = round(L * prof_m * 2, 2)
    pavim    = round(L * 0.80, 2)
    c_tubo   = round(L * preco_tubo(t["dn_mm"]), 2)
    c_escav  = round(escav * 30, 2)
    c_pv     = 3078.0
    t_sbdi   = round(c_tubo + c_escav + c_pv, 2)
    t_cbdi   = round(t_sbdi * 1.25, 2)
    return dict(escav=escav, reaterro=reaterro, bota=bota,
                escoram=escoram, pavim=pavim,
                c_tubo=c_tubo, c_escav=c_escav, c_pv=c_pv,
                t_sbdi=t_sbdi, t_cbdi=t_cbdi)


# ────────── ESTILOS ──────────
fill_h1   = PatternFill("solid", fgColor="1F4E79")
fill_h2   = PatternFill("solid", fgColor="2E75B6")
fill_nucl = PatternFill("solid", fgColor="D6E4F0")
fill_sub  = PatternFill("solid", fgColor="FFF2CC")
fill_zb   = PatternFill("solid", fgColor="EBF3FB")
fill_inf  = PatternFill("solid", fgColor="E2EFDA")  # verde suave = inferido
font_h1   = Font(bold=True, color="FFFFFF", size=11)
font_h2   = Font(bold=True, color="FFFFFF", size=9)
font_nucl = Font(bold=True, size=9)
font_sub  = Font(bold=True, size=9)
font_d    = Font(size=9)
font_di   = Font(size=9, italic=True, color="2E4057")  # itálico = inferido
ac = Alignment(horizontal="center", vertical="center", wrap_text=True)
al = Alignment(horizontal="left",   vertical="center")
thin = Border(left=Side(style="thin"), right=Side(style="thin"),
              top=Side(style="thin"),  bottom=Side(style="thin"))

COLS = ["ID","Nucleo","TAG","PV Ini","PV Fim","L (m)","DN (mm)",
        "Mat","CT Ini","CF Ini","Prof Ini","CT Fim","CF Fim","Prof Fim",
        "Decl (m/m)","V (m/s)","Q (l/s)","Pavimento",
        "Escav (m3)","Reaterro (m3)","Bota-fora (m3)","Escoram (m2)","Pavim (m2)",
        "Custo Tubo","Custo Escav","Custo PV","Total s/BDI","Total c/BDI",
        "Perfil","Rua / Obs"]
LARG = [8,25,8,14,14,7,8,5,9,9,10,9,9,10,9,7,7,14,10,10,10,10,10,12,12,12,14,14,12,30]


def escrever_cabecalho(ws, titulo):
    ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
    c = ws.cell(1, 1, titulo)
    c.fill = fill_h1; c.font = font_h1; c.alignment = ac
    ws.row_dimensions[1].height = 22
    # Nota inferido
    ws.merge_cells(f"A2:{get_column_letter(len(COLS))}2")
    c = ws.cell(2, 1,
        "* DADOS INFERIDOS por similaridade com núcleos reais SLNR Santos "
        "(Teteu / Pantanal Baixo / São Manoel / Vila Criadores / Vila Israel / Prolongamentos) — "
        "ConstruData HydroNetwork v7.0.0  |  FCN Construções e Saneamento  |  Contrato 11481051")
    c.fill = fill_inf; c.font = Font(italic=True, size=8, color="2E4057"); c.alignment = al
    ws.row_dimensions[2].height = 18
    for i, col in enumerate(COLS, 1):
        c = ws.cell(3, i, col)
        c.fill = fill_h2; c.font = font_h2; c.alignment = ac; c.border = thin
    ws.row_dimensions[3].height = 32
    for i, w in enumerate(LARG, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


def escrever_linha(ws, row, tid, nome_nucleo, tag, t, q, zebra):
    vals = [
        f"T-{tid:04d}", nome_nucleo, tag,
        t["pv_ini"], t["pv_fim"],
        t["ext_m"], t["dn_mm"], t["material"],
        t["ct_ini"], t["cf_ini"], t["prof_ini"],
        t["ct_fim"], t["cf_fim"], t["prof_fim"],
        round(t["decl_mm"]/1000, 4), t["v_ms"], t["q_ls"],
        t["pavimento"],
        q["escav"], q["reaterro"], q["bota"],
        q["escoram"], q["pavim"],
        q["c_tubo"], q["c_escav"], q["c_pv"],
        q["t_sbdi"], q["t_cbdi"],
        t["perf_key"], t["rua"],
    ]
    for i, v in enumerate(vals, 1):
        c = ws.cell(row, i, v)
        c.font = font_di  # itálico = inferido
        c.alignment = al if i == len(COLS) else ac
        c.border = thin
        if zebra: c.fill = fill_zb


def escrever_subtotal(ws, row, nome, s, n_t):
    vals = [f"SUBTOTAL {nome.upper()} ({n_t} trechos)", "", "", "", "",
            round(s["l"],2),"","","","","","","","",
            "","","","",
            round(s["escav"],3), round(s["reaterro"],3), round(s["bota"],3),
            "", round(s["pavim"],2),
            round(s["c_tubo"],2), round(s["c_escav"],2), round(s["c_pv"],2),
            round(s["t_sbdi"],2), round(s["t_cbdi"],2), "", ""]
    for col, v in enumerate(vals, 1):
        c = ws.cell(row, col, v)
        c.fill = fill_sub; c.font = font_sub; c.alignment = ac; c.border = thin


# ────────── LER DADOS ──────────
print(f"Lendo: {IN_FILE}")
wb_in = openpyxl.load_workbook(IN_FILE)
ws_in = wb_in.active

nucleos = []
for row in ws_in.iter_rows(min_row=3, max_row=ws_in.max_row, values_only=True):
    tag = row[0]
    if not tag: continue
    tag      = str(tag)
    nome     = str(row[1] or tag)
    tipo_ter = str(row[3] or "Urbano")
    ligs_esg = int(row[7]) if row[7] else 0
    ext_m    = float(row[10]) if row[10] else 0
    nucleos.append((tag, nome, tipo_ter, ligs_esg, ext_m))

print(f"  {len(nucleos)} núcleos encontrados")

# ────────── GERAR EXCEL ──────────
wb = openpyxl.Workbook()
ws_all = wb.active
ws_all.title = "TODOS_68_NUCLEOS"
escrever_cabecalho(ws_all, "TRECHOS INFERIDOS — 68 NÚCLEOS SLNR SANTOS — ConstruData HydroNetwork v7.0.0")

row_all = 4
tid_global = 1
grand = dict(l=0, escav=0, reaterro=0, bota=0, pavim=0,
             c_tubo=0, c_escav=0, c_pv=0, t_sbdi=0, t_cbdi=0)

total_skipped = 0

for tag, nome, tipo_ter, ligs, ext_m in nucleos:
    trechos = gerar_trechos(tag, nome, ext_m, ligs, tipo_ter)
    if not trechos:
        print(f"  [SKIP] {tag} {nome[:40]} ext={ext_m}")
        total_skipped += 1
        continue

    # Nome curto para aba (max 28 chars, sem chars especiais)
    nome_curto = (nome.split("_",1)[-1] if "_" in nome else nome)[:28]
    nome_curto = "".join(c if c.isalnum() or c in " -_" else "_" for c in nome_curto).strip()
    aba_nome   = f"{tag}_{nome_curto}"[:31]

    print(f"  {tag}  {nome[:45]:<45s}  {len(trechos):3d} trechos  {ext_m:.0f}m  ({classificar(nome,tipo_ter,ext_m,ligs)})")

    # Separador aba geral
    ws_all.merge_cells(f"A{row_all}:{get_column_letter(len(COLS))}{row_all}")
    label = f"  {tag} — {nome_curto.upper()}  |  {len(trechos)} trechos  |  {ext_m:.0f}m  |  {ligs} lig."
    c = ws_all.cell(row_all, 1, label)
    c.fill = fill_nucl; c.font = font_nucl; c.alignment = al
    ws_all.row_dimensions[row_all].height = 16
    row_all += 1

    # Aba individual
    ws_n = wb.create_sheet(title=aba_nome)
    escrever_cabecalho(ws_n, f"TRECHOS INFERIDOS — {tag} {nome_curto.upper()}")
    row_n = 4
    tid_nucleo = 1

    sub = dict(l=0, escav=0, reaterro=0, bota=0, pavim=0,
               c_tubo=0, c_escav=0, c_pv=0, t_sbdi=0, t_cbdi=0)

    for t in trechos:
        q = calcular_quant(t)
        zebra = (tid_nucleo % 2 == 0)
        escrever_linha(ws_all, row_all, tid_global, nome_curto, tag, t, q, zebra)
        escrever_linha(ws_n,   row_n,   tid_nucleo, nome_curto, tag, t, q, zebra)
        sub["l"]       += t["ext_m"]
        sub["escav"]   += q["escav"];   sub["reaterro"] += q["reaterro"]
        sub["bota"]    += q["bota"];    sub["pavim"]    += q["pavim"]
        sub["c_tubo"]  += q["c_tubo"];  sub["c_escav"]  += q["c_escav"]
        sub["c_pv"]    += q["c_pv"];    sub["t_sbdi"]   += q["t_sbdi"]
        sub["t_cbdi"]  += q["t_cbdi"]
        row_all += 1; row_n += 1
        tid_global += 1; tid_nucleo += 1

    escrever_subtotal(ws_all, row_all, nome_curto, sub, len(trechos))
    escrever_subtotal(ws_n,   row_n,   nome_curto, sub, len(trechos))
    row_all += 2; row_n += 1

    for k in grand: grand[k] += sub.get(k, 0)

# Grand Total
gt_vals = ["GRAND TOTAL — 68 NUCLEOS SLNR SANTOS", "", "", "", "",
           round(grand["l"],2),"","","","","","","","",
           "","","","",
           round(grand["escav"],3), round(grand["reaterro"],3), round(grand["bota"],3),
           "", round(grand["pavim"],2),
           round(grand["c_tubo"],2), round(grand["c_escav"],2), round(grand["c_pv"],2),
           round(grand["t_sbdi"],2), round(grand["t_cbdi"],2), "", ""]
for col, v in enumerate(gt_vals, 1):
    c = ws_all.cell(row_all, col, v)
    c.fill = fill_h1; c.font = font_h1; c.alignment = ac; c.border = thin

wb.save(OUT_FILE)
print(f"\nSalvo: {OUT_FILE}")
print(f"Total nucleos: {len(nucleos) - total_skipped}  |  Skipped: {total_skipped}")
print(f"Total trechos: {tid_global - 1}")
print(f"Extensao total: {round(grand['l'],0):.0f} m  ({round(grand['l']/1000,1)} km)")
print(f"Total c/BDI: R$ {grand['t_cbdi']:,.2f}")
print(f"Abas: {len(wb.sheetnames)}")
