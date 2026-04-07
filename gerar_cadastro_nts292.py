#!/usr/bin/env python3
"""
GERAR_CADASTRO_NTS292.PY — Gerador de Cadastro Técnico As-Built
Conforme NTS 292 Rev.3 (2017) — SABESP
Contrato 11481051 — SLNR Santos

Gera DXF georreferenciado com:
- Planta (Model Space) em UTM SIRGAS 2000
- Perfil longitudinal
- Layout A4 com carimbo SABESP (NTS 116)
- Simbologia SIGNOS (Tabela 4 NTS 292)
- Layers padronizados

Entrada: pvs, trechos (de ler_dxf_gdal ou ler_landxml)
Saída:   DXF por cruzamento/rua + PDF

Autor: ConstruData - HydroNetwork — FCN Construções e Saneamento
"""

import ezdxf
from ezdxf.enums import TextEntityAlignment
import math, json, os
from pathlib import Path
from datetime import datetime
from collections import defaultdict

# ══════════════════════════════════════════════════════════════
# CONSTANTES NTS 292
# ══════════════════════════════════════════════════════════════
ESCALA_H = 500   # Planta/Perfil horizontal 1:500
ESCALA_V = 100   # Perfil vertical 1:100

# Layers padronizados NTS 292 / SIGNOS
LAYERS = {
    "REDE_ESGOTO":       {"color": 3,  "ltype": "CONTINUOUS", "lw": 50},   # verde
    "REDE_AGUA":         {"color": 5,  "ltype": "CONTINUOUS", "lw": 50},   # azul
    "PV_ESGOTO":         {"color": 3,  "ltype": "CONTINUOUS", "lw": 35},
    "PI_ESGOTO":         {"color": 3,  "ltype": "CONTINUOUS", "lw": 25},
    "TXT_REDE":          {"color": 7,  "ltype": "CONTINUOUS", "lw": 18},
    "TXT_PV":            {"color": 7,  "ltype": "CONTINUOUS", "lw": 18},
    "TXT_PERFIL":        {"color": 7,  "ltype": "CONTINUOUS", "lw": 13},
    "CARTOGRAFIA":       {"color": 8,  "ltype": "CONTINUOUS", "lw": 13},
    "PERFIL_TERRENO":    {"color": 8,  "ltype": "DASHED",     "lw": 18},
    "PERFIL_TUBO":       {"color": 3,  "ltype": "CONTINUOUS", "lw": 35},
    "PERFIL_PV":         {"color": 3,  "ltype": "CONTINUOUS", "lw": 25},
    "PERFIL_GRADE":      {"color": 9,  "ltype": "CONTINUOUS", "lw": 9},
    "CARIMBO":           {"color": 7,  "ltype": "CONTINUOUS", "lw": 25},
    "COORD_GRID":        {"color": 9,  "ltype": "CONTINUOUS", "lw": 9},
    "RAMAL":             {"color": 6,  "ltype": "CONTINUOUS", "lw": 18},
    "LIGACAO_PREDIAL":   {"color": 6,  "ltype": "DASHED",     "lw": 13},
}

# Simbologia PV / PI (diâmetros em mm no DXF)
PV_RADIUS = 0.5    # raio PV em metros (Model Space)
PI_RADIUS = 0.3    # raio PI


def _setup_doc():
    """Cria documento DXF com layers NTS 292."""
    doc = ezdxf.new("R2010")
    doc.header["$INSUNITS"] = 6  # metros
    
    for name, props in LAYERS.items():
        doc.layers.add(name, color=props["color"], lineweight=props["lw"])
    
    # Estilos de texto
    doc.styles.add("SABESP_NORMAL", font="Arial")
    doc.styles.add("SABESP_TITULO", font="Arial")
    doc.styles.add("SABESP_PERFIL", font="Arial Narrow")
    
    return doc


def _draw_pv_symbol(msp, x, y, nome, ct, cf, prof, layer_pv="PV_ESGOTO", layer_txt="TXT_PV"):
    """Desenha símbolo de PV/PI conforme simbologia SIGNOS."""
    is_pi = nome.upper().startswith("PI") or nome.upper().startswith("P.I.")
    radius = PI_RADIUS if is_pi else PV_RADIUS
    
    # Círculo do PV
    msp.add_circle((x, y), radius, dxfattribs={"layer": layer_pv})
    
    # Cruz interna (para PV)
    if not is_pi:
        msp.add_line((x - radius*0.7, y), (x + radius*0.7, y), dxfattribs={"layer": layer_pv})
        msp.add_line((x, y - radius*0.7), (x, y + radius*0.7), dxfattribs={"layer": layer_pv})
    
    # Linha de chamada + texto
    lc_dx, lc_dy = 2.0, 2.0
    msp.add_line((x, y), (x + lc_dx, y + lc_dy), dxfattribs={"layer": layer_txt})
    msp.add_line((x + lc_dx, y + lc_dy), (x + lc_dx + 8, y + lc_dy), dxfattribs={"layer": layer_txt})
    
    # Texto do PV
    h = 1.2  # altura texto
    txt_lines = [
        nome,
        f"CT={ct:.3f}" if ct else "CT=—",
        f"CF={cf:.3f}" if cf else "CF=—",
    ]
    if prof:
        txt_lines.append(f"Prof={prof:.2f}m")
    
    for i, line in enumerate(txt_lines):
        msp.add_text(line, height=h, dxfattribs={
            "layer": layer_txt, "style": "SABESP_NORMAL"
        }).set_placement((x + lc_dx + 0.5, y + lc_dy + 1.5 - i * 1.8))


def _draw_tubo(msp, x0, y0, x1, y1, dn_mm, material, decl, ext, layer="REDE_ESGOTO", layer_txt="TXT_REDE"):
    """Desenha tubulação entre dois PVs com anotações."""
    msp.add_line((x0, y0), (x1, y1), dxfattribs={"layer": layer})
    
    # Seta de fluxo no meio
    mx, my = (x0 + x1) / 2, (y0 + y1) / 2
    ang = math.atan2(y1 - y0, x1 - x0)
    
    # Texto no meio do tubo
    txt = f"DN{dn_mm}" if dn_mm else "DN?"
    if material:
        txt += f" {material}"
    if decl:
        txt += f" i={decl*1000:.2f}‰"
    txt += f" L={ext:.2f}m"
    
    # Rotacionar texto para acompanhar tubo
    ang_deg = math.degrees(ang)
    if ang_deg > 90 or ang_deg < -90:
        ang_deg += 180
    
    msp.add_text(txt, height=1.0, rotation=ang_deg, dxfattribs={
        "layer": layer_txt, "style": "SABESP_NORMAL"
    }).set_placement((mx + 0.5 * math.sin(ang), my - 0.5 * math.cos(ang)))


def _draw_perfil(msp, trechos_seq, pvs, origin_x, origin_y, escala_h=ESCALA_H, escala_v=ESCALA_V):
    """
    Desenha perfil longitudinal abaixo da planta.
    origin_x, origin_y = canto inferior esquerdo do perfil no Model Space.
    """
    if not trechos_seq:
        return
    
    # Calcular fator de escala para perfil
    # No perfil: X = distância acumulada / escala_h * 1000
    #            Y = cota / escala_v * 1000
    fx = 1000.0 / escala_h   # mm por metro horizontal
    fy = 1000.0 / escala_v   # mm por metro vertical
    
    # Coletar dados sequenciais
    dist_acum = 0
    pontos_terreno = []
    pontos_tubo = []
    pvs_perfil = []
    
    for i, tr in enumerate(trechos_seq):
        pv_ini = pvs.get(tr["pv_ini"], {})
        pv_fim = pvs.get(tr["pv_fim"], {})
        
        ct_ini = pv_ini.get("ct", 0) or 0
        cf_ini = pv_ini.get("cf", 0) or 0
        ct_fim = pv_fim.get("ct", 0) or 0
        cf_fim = pv_fim.get("cf", 0) or 0
        ext = tr.get("ext_m", 0)
        
        if i == 0:
            pontos_terreno.append((dist_acum, ct_ini))
            pontos_tubo.append((dist_acum, cf_ini))
            pvs_perfil.append((dist_acum, tr["pv_ini"], ct_ini, cf_ini))
        
        dist_acum += ext
        pontos_terreno.append((dist_acum, ct_fim))
        pontos_tubo.append((dist_acum, cf_fim))
        pvs_perfil.append((dist_acum, tr["pv_fim"], ct_fim, cf_fim))
    
    if not pontos_terreno:
        return
    
    # Calcular cota mínima para referência
    all_cotas = [p[1] for p in pontos_terreno + pontos_tubo if p[1] > 0]
    if not all_cotas:
        return
    cota_min = min(all_cotas) - 2
    cota_max = max(all_cotas) + 2
    dist_max = max(p[0] for p in pontos_terreno)
    
    # Converter para coordenadas do perfil
    def to_perfil(d, c):
        px = origin_x + d * fx
        py = origin_y + (c - cota_min) * fy
        return px, py
    
    # Grade do perfil
    layer_g = "PERFIL_GRADE"
    # Linhas horizontais de cota
    cota_step = 1.0 if (cota_max - cota_min) < 20 else 2.0
    c = math.floor(cota_min)
    while c <= math.ceil(cota_max):
        px0 = origin_x
        px1 = origin_x + dist_max * fx
        py = origin_y + (c - cota_min) * fy
        msp.add_line((px0, py), (px1, py), dxfattribs={"layer": layer_g})
        msp.add_text(f"{c:.2f}", height=0.8, dxfattribs={
            "layer": "TXT_PERFIL", "style": "SABESP_PERFIL"
        }).set_placement((px0 - 5, py - 0.3))
        c += cota_step
    
    # Linhas verticais nos PVs
    for d, nome, ct, cf in pvs_perfil:
        px = origin_x + d * fx
        py_bot = origin_y
        py_top = origin_y + (cota_max - cota_min) * fy + 2
        msp.add_line((px, py_bot), (px, py_top), dxfattribs={"layer": layer_g})
        
        # Nome do PV na base
        msp.add_text(nome, height=0.8, rotation=90, dxfattribs={
            "layer": "TXT_PERFIL", "style": "SABESP_PERFIL"
        }).set_placement((px + 0.3, py_bot - 1))
    
    # Linha do terreno (tracejada)
    pts_terr = [to_perfil(d, c) for d, c in pontos_terreno if c > 0]
    if len(pts_terr) >= 2:
        msp.add_lwpolyline(pts_terr, dxfattribs={"layer": "PERFIL_TERRENO"})
    
    # Linha do tubo (contínua)
    pts_tubo = [to_perfil(d, c) for d, c in pontos_tubo if c > 0]
    if len(pts_tubo) >= 2:
        msp.add_lwpolyline(pts_tubo, dxfattribs={"layer": "PERFIL_TUBO"})
    
    # PVs no perfil
    for d, nome, ct, cf in pvs_perfil:
        if ct > 0 and cf > 0:
            px_ct = origin_x + d * fx
            py_ct = origin_y + (ct - cota_min) * fy
            py_cf = origin_y + (cf - cota_min) * fy
            # Linha vertical PV (tampão ao fundo)
            msp.add_line((px_ct, py_ct), (px_ct, py_cf), dxfattribs={"layer": "PERFIL_PV"})
    
    # Tabela de dados abaixo do perfil
    tab_y = origin_y - 3
    headers = ["PV", "Cota Tampão", "Cota Fundo", "Prof.", "Distância"]
    col_w = dist_max * fx / max(len(pvs_perfil), 1)
    
    for i, (d, nome, ct, cf) in enumerate(pvs_perfil):
        px = origin_x + d * fx
        prof = (ct - cf) if ct and cf else 0
        
        data = [nome, f"{ct:.3f}" if ct else "—", f"{cf:.3f}" if cf else "—",
                f"{prof:.2f}" if prof else "—", f"{d:.2f}"]
        
        for j, val in enumerate(data):
            msp.add_text(val, height=0.6, dxfattribs={
                "layer": "TXT_PERFIL", "style": "SABESP_PERFIL"
            }).set_placement((px - 1, tab_y - j * 1.5))


def _draw_carimbo_sabesp(msp, x, y, info):
    """Desenha carimbo padrão SABESP conforme NTS 116."""
    w, h = 185, 40  # mm convertido para unidades do layout
    layer = "CARIMBO"
    
    # Retângulo externo
    msp.add_lwpolyline(
        [(x, y), (x + w, y), (x + w, y + h), (x, y + h), (x, y)],
        close=True, dxfattribs={"layer": layer}
    )
    
    # Divisões internas
    div_x = x + 90
    msp.add_line((div_x, y), (div_x, y + h), dxfattribs={"layer": layer})
    
    # Logo area (esquerda)
    msp.add_text("SABESP", height=5, dxfattribs={
        "layer": layer, "style": "SABESP_TITULO"
    }).set_placement((x + 5, y + h - 10))
    
    msp.add_text("Companhia de Saneamento Básico", height=2, dxfattribs={
        "layer": layer, "style": "SABESP_NORMAL"
    }).set_placement((x + 5, y + h - 18))
    
    msp.add_text("do Estado de São Paulo", height=2, dxfattribs={
        "layer": layer, "style": "SABESP_NORMAL"
    }).set_placement((x + 5, y + h - 22))
    
    # Dados do projeto (direita)
    campos = [
        ("CONTRATO:", info.get("contrato", "11481051")),
        ("LOCAL:", info.get("local", "Santos/SP")),
        ("OBRA:", info.get("obra", "SE LIGA NA REDE")),
        ("CONTEÚDO:", info.get("conteudo", "CADASTRO AS-BUILT")),
        ("ESCALA:", info.get("escala", f"H 1:{ESCALA_H} V 1:{ESCALA_V}")),
        ("FOLHA:", info.get("folha", "01/01")),
        ("DATA:", info.get("data", datetime.now().strftime("%m/%Y"))),
        ("Nº PLANTA:", info.get("n_planta", "")),
    ]
    
    for i, (label, val) in enumerate(campos):
        row_y = y + h - 5 - i * 4.5
        msp.add_text(label, height=1.5, dxfattribs={
            "layer": layer, "style": "SABESP_NORMAL"
        }).set_placement((div_x + 3, row_y))
        msp.add_text(val, height=1.8, dxfattribs={
            "layer": layer, "style": "SABESP_TITULO"
        }).set_placement((div_x + 30, row_y))


def _draw_coord_grid(msp, xmin, ymin, xmax, ymax, step=100):
    """Malha de coordenadas UTM conforme NTS 292."""
    layer = "COORD_GRID"
    
    x0 = math.floor(xmin / step) * step
    y0 = math.floor(ymin / step) * step
    
    x = x0
    while x <= xmax + step:
        msp.add_line((x, ymin - 10), (x, ymax + 10), dxfattribs={"layer": layer})
        msp.add_text(f"E {x:.0f}", height=1.5, rotation=90, dxfattribs={
            "layer": layer, "style": "SABESP_NORMAL"
        }).set_placement((x + 0.5, ymin - 15))
        x += step
    
    y = y0
    while y <= ymax + step:
        msp.add_line((xmin - 10, y), (xmax + 10, y), dxfattribs={"layer": layer})
        msp.add_text(f"N {y:.0f}", height=1.5, dxfattribs={
            "layer": layer, "style": "SABESP_NORMAL"
        }).set_placement((xmin - 40, y + 0.5))
        y += step


def _ensure_layer(doc, name: str, color: int = 7):
    """Cria layer no documento ezdxf se ainda não existir."""
    try:
        layers = doc.layers
        if name not in layers:
            layer = layers.add(name)
            layer.color = color
    except Exception:
        pass  # ezdxf API pode variar entre versões


def _gerar_divergencias_xlsx(divergencias: list, out_dir, nucleo: str) -> str:
    """
    Gera DIVERGENCIAS_NTS292.xlsx com tabela projeto vs campo por PV/NS.
    Retorna path ou '' se openpyxl não disponível.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return ""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Divergencias"

    # Cabeçalho
    headers = ["NS", "PV", "Campo", "Projeto (m)", "Campo Real (m)", "Diferença (m)", "Obs"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="CC0000")
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 30

    # Dados
    fill_warn = PatternFill("solid", fgColor="FFF2CC")
    fill_ok   = PatternFill("solid", fgColor="E2EFDA")
    for row_idx, d in enumerate(divergencias, 2):
        diff = d.get("diferenca_m", 0)
        obs = "DIVERGÊNCIA SIGNIFICATIVA" if abs(diff) > 0.10 else "Dentro da tolerância"
        row_data = [
            d.get("ns_key", ""),
            d.get("pv", ""),
            d.get("campo", ""),
            d.get("projeto", ""),
            d.get("campo_real", ""),
            diff,
            obs,
        ]
        fill = fill_warn if abs(diff) > 0.10 else fill_ok
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = fill
            if col in (4, 5, 6):
                cell.number_format = "0.000"

    # Resumo
    n_sig = sum(1 for d in divergencias if abs(d.get("diferenca_m", 0)) > 0.10)
    ws_r = wb.create_sheet("Resumo")
    ws_r["A1"] = f"Nucleo: {nucleo}"
    ws_r["A2"] = f"Total divergencias: {len(divergencias)}"
    ws_r["A3"] = f"Significativas (>10cm): {n_sig}"
    ws_r["A4"] = f"Dentro da tolerância: {len(divergencias) - n_sig}"

    fname = f"DIVERGENCIAS_NTS292_{nucleo.upper().replace(' ', '_')}.xlsx"
    fpath = str(out_dir / fname)
    wb.save(fpath)
    return fpath


def gerar_cadastro_nts292(pvs, trechos, nucleo, out_dir, info=None, status_ns=None,
                          topo_path=None, cartografia_path=None):
    """
    Gera cadastro técnico as-built conforme NTS 292.

    Args:
        pvs: dict nome→{x,y,ct,cf,prof}
        trechos: list of {pv_ini, pv_fim, dn_mm, ext_m, decl_mm, material}
        nucleo: nome do núcleo
        out_dir: diretório de saída
        info: dict com dados do carimbo
        status_ns: dict do motor_status_ns — se fornecido, usa campo_real{CT/CF}
                   para dados de campo reais e anota divergências do projeto.

    Returns:
        list of paths gerados
    """
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)

    if info is None:
        info = {}
    info.setdefault("contrato", "11481051")
    info.setdefault("local", f"{nucleo} — Santos/SP")
    info.setdefault("obra", "SE LIGA NA REDE — FCN Construções e Saneamento")
    info.setdefault("conteudo", f"CADASTRO AS-BUILT — REDE COLETORA DE ESGOTO — {nucleo}")

    # Enriquecer PVs com dados reais de campo se status_ns fornecido
    pvs_campo: dict = {}          # {pv_nome: {ct_real, cf_real, prof_real}}
    divergencias: list = []       # [{ns_key, pv, campo, projeto, diff}]

    # Pipeline do PROJETO (ficticio) sai sempre em out_dir.
    # Se topo_path foi fornecido, dispara um SEGUNDO pipeline depois com o
    # cadastro REAL COMPILADO em out_dir/REAL_COMPILADO_<NUCLEO>/.
    pvs_ajustados_campo = 0
    _topo_diferido = topo_path  # guarda para 2a passada
    topo_path = None            # zera nesta passada (geramos o ficticio)

    # ── Cartografia de fundo (DWG/DXF) ──
    cartografia_polys = None
    if cartografia_path:
        try:
            from cadastro.ler_cartografia import ler_cartografia
            cartografia_polys = ler_cartografia(cartografia_path)
            print(f"  [NTS292] cartografia: {len(cartografia_polys or [])} polilinhas")
        except Exception as e:
            print(f"  [NTS292] falha ao ler cartografia: {e}")
    data_levant = ""
    resp_levant = ""

    if status_ns and status_ns.get("notas"):
        from datetime import datetime as _dt
        for ns_key, nota in status_ns["notas"].items():
            cr = nota.get("campo_real", {})
            pv_i = nota.get("pv_ini", "")
            pv_f = nota.get("pv_fim", "")

            for pv_nome, ct_r, cf_r in [
                (pv_i, cr.get("ct_ini_real"), cr.get("cf_ini_real")),
                (pv_f, cr.get("ct_fim_real"), cr.get("cf_fim_real")),
            ]:
                if not pv_nome:
                    continue
                entry = pvs_campo.setdefault(pv_nome, {})
                if ct_r is not None:
                    entry["ct_real"] = ct_r
                if cf_r is not None:
                    entry["cf_real"] = cf_r
                if ct_r is not None and cf_r is not None:
                    entry["prof_real"] = round(ct_r - cf_r, 3)

                # Registrar divergências
                pv_proj = pvs.get(pv_nome, {})
                ct_proj = pv_proj.get("ct", 0) or 0
                cf_proj = pv_proj.get("cf", 0) or 0
                if ct_r is not None and ct_proj:
                    diff = round(ct_r - ct_proj, 3)
                    if abs(diff) > 0.05:  # divergência > 5cm
                        divergencias.append({
                            "ns_key": ns_key, "pv": pv_nome, "campo": "CT",
                            "projeto": ct_proj, "campo_real": ct_r, "diferenca_m": diff,
                        })
                if cf_r is not None and cf_proj:
                    diff = round(cf_r - cf_proj, 3)
                    if abs(diff) > 0.05:
                        divergencias.append({
                            "ns_key": ns_key, "pv": pv_nome, "campo": "CF",
                            "projeto": cf_proj, "campo_real": cf_r, "diferenca_m": diff,
                        })

            # Detectar data e responsável de levantamento mais recente
            hist = nota.get("status_historico", [])
            for h in reversed(hist):
                if h.get("para") in ("CADASTRADO", "EXECUTADO"):
                    data_levant = h.get("data", "")[:10]
                    resp_levant = h.get("resp", "")
                    break

        if data_levant:
            info["levantamento"] = f"LEVANTAMENTO: {data_levant} | RESP: {resp_levant}"
            info["conteudo"] = (
                f"CADASTRO AS-BUILT (campo real) — REDE COLETORA — {nucleo}"
            )
    
    # Agrupar trechos por rua (se disponível) ou por proximidade
    # Para rede coletora: formato A4, uma planta por arquivo

    doc = _setup_doc()
    msp = doc.modelspace()

    # Garantir layers CAMPO_REAL (vermelho) e PROJETO (cinza) no documento
    _ensure_layer(doc, "CAMPO_REAL", color=1)   # ACI 1 = vermelho
    _ensure_layer(doc, "PROJETO",    color=8)   # ACI 8 = cinza

    # 1. PLANTA (Model Space) — georreferenciada
    # Malha de coordenadas UTM
    xs = [p["x"] for p in pvs.values() if p.get("x")]
    ys = [p["y"] for p in pvs.values() if p.get("x")]

    if not xs:
        print(f"  [WARN] Sem coordenadas para {nucleo}")
        return []

    xmin, xmax = min(xs) - 20, max(xs) + 20
    ymin, ymax = min(ys) - 20, max(ys) + 20

    _draw_coord_grid(msp, xmin, ymin, xmax, ymax, step=50)

    # Cartografia de fundo no DXF consolidado (georreferenciada, sem reprojeto)
    if cartografia_polys:
        _ensure_layer(doc, "CARTOGRAFIA_FUNDO", color=8)
        n_poly = 0
        for elem in cartografia_polys:
            pts = elem.get("pontos") or []
            if len(pts) < 2:
                continue
            # filtra polilinhas que tocam a area da rede
            if not any(xmin <= p[0] <= xmax and ymin <= p[1] <= ymax for p in pts):
                continue
            try:
                msp.add_lwpolyline(pts, dxfattribs={"layer": "CARTOGRAFIA_FUNDO"})
                n_poly += 1
            except Exception:
                pass
        print(f"  [NTS292] cartografia desenhada no DXF consolidado: {n_poly} polilinhas")

    # Desenhar tubulações
    for tr in trechos:
        pv0 = pvs.get(tr["pv_ini"], {})
        pv1 = pvs.get(tr["pv_fim"], {})
        if pv0.get("x") and pv1.get("x"):
            tipo = "REDE_AGUA" if tr.get("tipo") == "AGUA" else "REDE_ESGOTO"
            _draw_tubo(msp,
                       pv0["x"], pv0["y"], pv1["x"], pv1["y"],
                       tr.get("dn_mm"), tr.get("material"),
                       tr.get("decl_mm"), tr.get("ext_m", 0),
                       layer=tipo)

    # Desenhar PVs
    for nome, pv in pvs.items():
        if not pv.get("x"):
            continue
        # Usar dados reais de campo se disponíveis (status_ns)
        cr = pvs_campo.get(nome, {})
        ct_uso = cr.get("ct_real") or pv.get("ct", 0) or 0
        cf_uso = cr.get("cf_real") or pv.get("cf", 0) or 0
        prof_uso = cr.get("prof_real") or pv.get("prof", 0) or 0
        tem_campo = bool(cr.get("ct_real") or cr.get("cf_real"))
        # Layer: vermelho se dado de campo real, cinza se só projeto (quando há status_ns)
        layer_pv = "CAMPO_REAL" if tem_campo else ("PROJETO" if pvs_campo else "PV_ESGOTO")
        _draw_pv_symbol(msp, pv["x"], pv["y"], nome,
                        ct_uso, cf_uso, prof_uso,
                        layer_pv=layer_pv)

    # 2. PERFIL LONGITUDINAL (abaixo da planta)
    # Sequenciar trechos para perfil — usar CT/CF reais se disponíveis
    pvs_perfil = dict(pvs)  # cópia
    for pv_nome, cr in pvs_campo.items():
        if pv_nome in pvs_perfil:
            pv_c = dict(pvs_perfil[pv_nome])
            if cr.get("ct_real") is not None:
                pv_c["ct"] = cr["ct_real"]
            if cr.get("cf_real") is not None:
                pv_c["cf"] = cr["cf_real"]
            pvs_perfil[pv_nome] = pv_c

    perfil_origin_x = xmin
    perfil_origin_y = ymin - 80  # abaixo da planta
    _draw_perfil(msp, trechos[:50], pvs_perfil, perfil_origin_x, perfil_origin_y)

    # 3. CARIMBO no Model Space (para referência)
    carimbo_x = xmax + 10
    carimbo_y = ymin
    _draw_carimbo_sabesp(msp, carimbo_x, carimbo_y, info)

    # 4. Norte magnético
    norte_x, norte_y = xmax + 5, ymax
    msp.add_line((norte_x, norte_y), (norte_x, norte_y + 15), dxfattribs={"layer": "CARIMBO"})
    msp.add_text("N", height=3, dxfattribs={
        "layer": "CARIMBO", "style": "SABESP_TITULO"
    }).set_placement((norte_x - 1, norte_y + 16))

    # 5. Legenda AS-BUILT se houver dados de campo
    if pvs_campo:
        leg_x, leg_y = xmin, ymax + 5
        msp.add_text(
            f"AS-BUILT  |  {data_levant}  |  Resp: {resp_levant}",
            height=2.5,
            dxfattribs={"layer": "CAMPO_REAL", "style": "SABESP_TITULO"},
        ).set_placement((leg_x, leg_y))
        msp.add_text(
            "CAMPO_REAL (vermelho) = dados levantados em campo  |  "
            "PROJETO (cinza) = conforme projeto",
            height=1.8,
            dxfattribs={"layer": "CAMPO_REAL"},
        ).set_placement((leg_x, leg_y - 4))

    # Salvar DXF
    fname = f"CADASTRO_ASBUILT_{nucleo.upper().replace(' ','_')}.dxf"
    fpath = out / fname
    doc.saveas(str(fpath))

    # 6. Gerar DIVERGENCIAS_NTS292.xlsx se houver divergências
    paths_gerados = [str(fpath)]
    if divergencias:
        div_path = _gerar_divergencias_xlsx(divergencias, out, nucleo)
        if div_path:
            paths_gerados.append(div_path)
            print(f"  [NTS292] {len(divergencias)} divergências projeto/campo → {Path(div_path).name}")

    # Gerar JSON com metadados NTS 292
    meta = {
        "norma": "NTS 292 Rev.3 (2017)",
        "contrato": info["contrato"],
        "nucleo": nucleo,
        "datum_h": "SIRGAS 2000 / UTM Zone 23S (EPSG:31983)",
        "datum_v": "Imbituba-SC",
        "escala_h": f"1:{ESCALA_H}",
        "escala_v": f"1:{ESCALA_V}",
        "n_pvs": len(pvs),
        "n_trechos": len(trechos),
        "extensao_m": sum(t.get("ext_m", 0) for t in trechos),
        "data": datetime.now().isoformat(),
        "arquivo_dxf": fname,
        "formato": "A4 (rede coletora) / A1 (planta-perfil)",
        "dados_campo": {
            "pvs_com_campo_real": len(pvs_campo),
            "pvs_ajustados_por_campo": pvs_ajustados_campo,
            "divergencias": len(divergencias),
            "data_levantamento": data_levant,
            "responsavel_levantamento": resp_levant,
            "topo_path": topo_path or "",
            "cartografia_path": cartografia_path or "",
        },
        "layers": list(LAYERS.keys()) + ["CAMPO_REAL", "PROJETO"],
        "requisitos_entrega": {
            "dwg_georref": True,
            "pdf_assinado": "PENDENTE — assinar digitalmente",
            "art_cau": "PENDENTE — emitir ART",
            "ctb_plotstyle": "PENDENTE — gerar .ctb",
            "signos_lancamento": "PENDENTE — via VisualBIM/1DOC",
        }
    }
    
    meta_path = out / f"META_{nucleo.upper().replace(' ','_')}.json"
    with open(meta_path, "w", encoding="utf-8") as f:
        json.dump(meta, f, indent=2, ensure_ascii=False)

    paths_gerados.append(str(meta_path))

    # 7. Folha A4 SABESP (DXF + PDF) — uma folha por trecho contiguo
    try:
        from cadastro import folha_a4
        base = f"CADASTRO A4 - {nucleo.upper().replace(' ','_')}"
        # converte PVs para o formato esperado pela folha
        def _conv(nm):
            p = pvs_perfil.get(nm, {}) or pvs.get(nm, {})
            return {
                "nome": nm,
                "este": p.get("x", 0),
                "norte": p.get("y", 0),
                "cota_tampa": p.get("ct", 0) or 0,
                "cota_fundo": p.get("cf", 0) or 0,
                "prof": p.get("prof", 0) or 0,
                "tipo": "PV",
            }
        # cada trecho vira uma folha (par PV_ini -> PV_fim)
        folhas = []
        for tr in trechos:
            a = _conv(tr.get("pv_ini", ""))
            b = _conv(tr.get("pv_fim", ""))
            if a["este"] and b["este"]:
                folhas.append([a, b])
        if folhas:
            redes_info = {
                "diametro": (trechos[0].get("dn_mm") if trechos else "") or "",
                "material": (trechos[0].get("material") if trechos else "") or "",
                "met_construtivo": "VALA ABERTA",
            }
            paths_a4 = folha_a4.generate_batch(folhas, str(out), base, info=info, redes_info=redes_info, cartografia=cartografia_polys)
            for dxf_p, pdf_p in paths_a4:
                paths_gerados.extend([dxf_p, pdf_p])
            print(f"  [NTS292] {len(folhas)} folhas A4 SABESP geradas (DXF+PDF)")
    except Exception as e:
        print(f"  [NTS292] folha A4 nao gerada: {e}")

    # ── 2a passada: CADASTRO REAL COMPILADO (topo de campo aplicado) ──
    if _topo_diferido:
        try:
            from cadastro.ler_topo import ler_topo
            from cadastro.compilar_campo import compilar_campo
            topo_pts = ler_topo(_topo_diferido)
            pvs_real, trechos_real, divs_topo = compilar_campo(pvs, trechos, topo_pts)
            n_aj = sum(1 for p in pvs_real.values() if p.get("fonte") == "CAMPO")
            print(f"  [NTS292] topo lido: {len(topo_pts)} pts | {n_aj} PVs ajustados -> REAL_COMPILADO")
            sub = out / f"REAL_COMPILADO_{nucleo.upper().replace(' ','_')}"
            info_real = dict(info)
            info_real["conteudo"] = f"CADASTRO REAL COMPILADO (topo) — {nucleo}"
            extras = gerar_cadastro_nts292(
                pvs_real, trechos_real, nucleo + "_REAL", str(sub),
                info=info_real, status_ns=status_ns,
                topo_path=None, cartografia_path=cartografia_path,
            )
            paths_gerados.extend(extras)
        except Exception as e:
            print(f"  [NTS292] falha REAL_COMPILADO: {e}")

    print(f"[NTS292] Cadastro gerado: {fpath}")
    print(f"         {len(pvs)} PVs | {len(trechos)} trechos | {meta['extensao_m']:.0f}m")
    print(f"         Meta: {meta_path}")
    if pvs_campo:
        print(f"         Campo real: {len(pvs_campo)} PVs levantados | {len(divergencias)} divergencias")

    return paths_gerados


# ══════════════════════════════════════════════════════════════
# CLI
# ══════════════════════════════════════════════════════════════
if __name__ == "__main__":
    import sys
    print("USO: from gerar_cadastro_nts292 import gerar_cadastro_nts292")
    print("     gerar_cadastro_nts292(pvs, trechos, 'NUCLEO', './saida')")
