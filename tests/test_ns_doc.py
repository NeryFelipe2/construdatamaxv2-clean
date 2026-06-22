import pytest
from datetime import date


def _ns_exemplo(geom):
    from planejamento.ns_pdf import materiais_trecho
    return {
        "ns": "NS001", "sistema": "ESGOTO", "nucleo": "BOI MALHADO",
        "equipe": "Wellington", "lider": "Rodrigo",
        "composicao": [("Rodrigo", "Líder"), ("Juan", "Operador")],
        "servico": "Assentamento de rede de esgoto", "rua": None, "dn": 200,
        "comp": 50.0, "produt": "25 m/dia", "inicio": date(2026, 6, 22),
        "fim": date(2026, 6, 24), "status": "PLANEJADO", "geom": geom,
        "materiais": materiais_trecho(50.0, "ESGOTO", 200),
    }


def test_dias_uteis_pula_fim_de_semana():
    from planejamento.ns_doc import dias_uteis
    # seg 22/06 a qua 24/06 = 3 dias úteis
    assert dias_uteis(date(2026, 6, 22), date(2026, 6, 24)) == 3
    # sex 19/06 a seg 22/06 = 2 dias úteis (pula sáb/dom)
    assert dias_uteis(date(2026, 6, 19), date(2026, 6, 22)) == 2


def test_gerar_pdf_ns_cria_arquivo(tmp_path):
    pytest.importorskip("matplotlib")
    from shapely.geometry import LineString
    from planejamento.ns_doc import gerar_pdf_ns
    g = LineString([(0, 0), (50, 0)])
    out = tmp_path / "ns.pdf"
    gerar_pdf_ns([_ns_exemplo(g)], {"ESGOTO": [g]}, str(out))
    assert out.exists() and out.stat().st_size > 1000


def test_gerar_excel_ns_abas(tmp_path):
    pytest.importorskip("openpyxl")
    import openpyxl
    from shapely.geometry import LineString
    from planejamento.ns_doc import gerar_excel_ns
    g = LineString([(0, 0), (50, 0)])
    out = tmp_path / "ns.xlsx"
    info = gerar_excel_ns([_ns_exemplo(g)], {"Wellington": [("Rodrigo", "Líder")]},
                          {"Wellington": "Rodrigo"}, str(out))
    assert set(["NOTAS DE SERVIÇO", "MATERIAIS", "EQUIPES", "CRONOGRAMA", "RESUMO"]) <= set(info["abas"])
    wb = openpyxl.load_workbook(out)
    assert "MATERIAIS" in wb.sheetnames
    assert wb["NOTAS DE SERVIÇO"].cell(3, 1).value == "NS001"
