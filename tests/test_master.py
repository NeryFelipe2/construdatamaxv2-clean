import pytest
from datetime import date


def test_escrever_workbook_multi_aba(tmp_path):
    pytest.importorskip("openpyxl")
    from planejamento.planilha import escrever_workbook
    abas = [{"nome": "ABA1", "colunas": [{"campo": "a", "titulo": "A"}],
             "linhas": [{"a": 7}], "titulo": "T"}]
    out = tmp_path / "wb.xlsx"
    assert escrever_workbook(abas, str(out)) == 1
    from openpyxl import load_workbook
    wb = load_workbook(str(out))
    assert wb.sheetnames == ["ABA1"]
    ws = wb["ABA1"]
    assert ws.cell(1, 1).value == "T" and ws.cell(2, 1).value == "A" and ws.cell(3, 1).value == 7


def test_gerar_master_abas_e_cabecalho(tmp_path):
    pytest.importorskip("openpyxl")
    from planejamento.master import gerar_master
    trechos = [{"ext_m": 92.0, "dn_mm": 150}, {"ext_m": 46.0, "dn_mm": 150}]
    out = tmp_path / "master.xlsx"
    info = gerar_master(str(out), trechos_esgoto=trechos,
                        pvs_esgoto={"PV1": {}, "PV2": {}}, data_inicio=date(2026, 6, 20))
    esperadas = {"TRECHO A TRECHO ESGOTO", "CRONOGRAMA", "MATERIAIS", "POR RUA", "ACOMPANHAMENTO"}
    assert esperadas.issubset(set(info["abas"]))

    from openpyxl import load_workbook
    wb = load_workbook(str(out))
    assert esperadas.issubset(set(wb.sheetnames))
    ws = wb["TRECHO A TRECHO ESGOTO"]
    assert ws.cell(2, 1).value == "Trecho"          # row1=título, row2=cabeçalho
    assert ws.cell(2, 5).value == "Comp (m)"
    assert ws.cell(3, 1).value == 1                 # primeiro trecho
