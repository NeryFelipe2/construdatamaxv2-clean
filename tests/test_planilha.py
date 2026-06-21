import pytest


def test_escreve_material_xlsx(tmp_path):
    pytest.importorskip("openpyxl")
    from planejamento.planilha import escrever_material_xlsx

    material = [{"num": 1, "quant": 70.0, "und": "m", "dimensao": "DN150",
                "codigo": "", "descricao": "Tubo PVC DN150"}]
    out = tmp_path / "material.xlsx"
    n = escrever_material_xlsx(material, str(out))
    assert n == 1 and out.exists()

    from openpyxl import load_workbook
    ws = load_workbook(str(out)).active
    assert ws.cell(row=1, column=1).value == "LISTA DE MATERIAL"   # título
    assert ws.cell(row=2, column=1).value == "Num."                # cabeçalho (mapa)
    assert ws.cell(row=2, column=6).value == "Descrição"
    assert ws.cell(row=3, column=2).value == 70.0                  # dado
    assert ws.cell(row=3, column=6).value == "Tubo PVC DN150"
