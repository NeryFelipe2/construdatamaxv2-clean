import pytest

ITENS = [
    {"descricao": "Tubo PEAD 63mm água", "qtd": 4000, "und": "metros"},
    {"descricao": "União 20mm água", "qtd": 200, "und": "unidade"},
]
CAB = {"nucleo": "BOI MALHADO", "sistema": "ÁGUA"}


def test_cab_padrao_preenche_defaults():
    from planejamento.pedido_material import _cab_padrao
    c = _cab_padrao({"nucleo": "X"})
    assert c["nucleo"] == "X" and c["pedido_n"] and c["destino"]


def test_gerar_pedido_xlsx(tmp_path):
    pytest.importorskip("openpyxl")
    import openpyxl
    from planejamento.pedido_material import gerar_pedido_xlsx
    out = tmp_path / "pedido.xlsx"
    info = gerar_pedido_xlsx(ITENS, CAB, str(out))
    assert info["itens"] == 2
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    hr = next(r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == "Item")
    assert ws.cell(hr + 1, 2).value == "Tubo PEAD 63mm água"
    assert ws.cell(hr + 1, 3).value == 4000


def test_gerar_pedido_pdf(tmp_path):
    pytest.importorskip("matplotlib")
    from planejamento.pedido_material import gerar_pedido_pdf
    out = tmp_path / "pedido.pdf"
    gerar_pedido_pdf(ITENS, CAB, str(out))
    assert out.exists() and out.stat().st_size > 1000


def test_paginar():
    from planejamento.pedido_material import _paginar
    assert len(_paginar(list(range(60)), por_pag=26)) == 3
    assert _paginar([]) == [[]]
