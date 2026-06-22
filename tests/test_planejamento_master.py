import pytest
from datetime import date


DADOS = {
    "contrato": "Sabesp 13.546/25-00", "nucleo": "BOI MALHADO",
    "data_inicio": date(2026, 6, 16), "dias_uteis_semana": 6, "cenario": "VCA",
    "prazo_alvo_dias": 60,
    "quantitativos": [
        ("Rede de água (×2 dois terços)", 7261.2, "m"),
        ("Rede de esgoto", 2638.4, "m"),
    ],
    "produtividade": {
        "Rede de água (×2 dois terços)": (100, 160),
        "Rede de esgoto": (42, 80),
    },
    "frentes": [
        ("Jesse / Ediel", "Jesse", "Ediel", "Rede", 10, "Rede água", "Caixa U.M.A"),
        ("Wellington / Rodrigo", "Wellington", "Rodrigo", "Rede", 4, "Rede esgoto", "Caixa inspeção + ramal"),
        ("Robert (rede)", "Robert", "Robert", "Rede", 4, "Rede esgoto", "Caixa inspeção + ramal"),
    ],
}


def test_gera_abas_master(tmp_path):
    pytest.importorskip("openpyxl")
    import openpyxl
    from planejamento.planejamento_master import gerar_master_xlsx
    out = tmp_path / "master.xlsx"
    info = gerar_master_xlsx(DADOS, str(out))
    esperadas = {"PESSOAL", "PREMISSAS", "QUANTITATIVOS", "ALOCACAO",
                 "DIMENSIONAMENTO", "CENÁRIOS", "CRONOGRAMA"}
    assert esperadas <= set(info["abas"])
    wb = openpyxl.load_workbook(out)
    # CENÁRIOS tem fórmula de dias (ROUNDUP/IFERROR) que cruza ALOCACAO
    cen = wb["CENÁRIOS"]
    f = cen.cell(5, 5).value
    assert isinstance(f, str) and f.startswith("=IFERROR(ROUNDUP")
    # ALOCACAO escreveu as 3 frentes
    al = wb["ALOCACAO"]
    assert al.cell(4, 1).value == "Jesse / Ediel"
    # PREMISSAS tem o cenário editável
    assert wb["PREMISSAS"].cell(3, 2).value == "VCA"


def test_cenario_mnd_usa_frentes_manuais(tmp_path):
    from planejamento.planejamento_master import gerar_master_xlsx
    import openpyxl
    out = tmp_path / "m2.xlsx"
    gerar_master_xlsx(DADOS, str(out))
    cen = openpyxl.load_workbook(out)["CENÁRIOS"]
    # coluna Frentes MND da rede esgoto referencia a premissa manual
    assert "PREMISSAS!$B$" in str(cen.cell(6, 6).value)
