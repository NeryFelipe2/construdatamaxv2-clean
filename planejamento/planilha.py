"""Escritor de planilha por TEMPLATE + MAPA DE COLUNAS (padrão DATOSE.DEF/LST_MAT).

`colunas` é o mapa: lista de {campo, titulo}. As linhas são dicts; cada
linha[campo] é escrito na coluna correspondente. Trocar de layout = trocar o
mapa, não o código.
"""

# Mapa de colunas da lista de material (LST_MAT.DEF: Num/Quant/Und/Dimensão/Código/Descrição)
COLUNAS_MATERIAL = [
    {"campo": "num", "titulo": "Num."},
    {"campo": "quant", "titulo": "Quant."},
    {"campo": "und", "titulo": "Und."},
    {"campo": "dimensao", "titulo": "Dimensão"},
    {"campo": "codigo", "titulo": "Código"},
    {"campo": "descricao", "titulo": "Descrição"},
]


def escrever_planilha(linhas, colunas, caminho, titulo=None):
    """Escreve um .xlsx com cabeçalho do mapa e os dados. Retorna nº de linhas de dados."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    r = 1
    if titulo:
        ws.cell(row=1, column=1, value=titulo)
        r = 2
    for ci, c in enumerate(colunas, 1):
        ws.cell(row=r, column=ci, value=c["titulo"])
    for i, linha in enumerate(linhas):
        for ci, c in enumerate(colunas, 1):
            ws.cell(row=r + 1 + i, column=ci, value=linha.get(c["campo"]))
    wb.save(str(caminho))
    return len(linhas)


def escrever_material_xlsx(material, caminho, titulo="LISTA DE MATERIAL"):
    return escrever_planilha(material, COLUNAS_MATERIAL, caminho, titulo)
