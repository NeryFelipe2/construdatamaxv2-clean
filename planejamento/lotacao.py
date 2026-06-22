# -*- coding: utf-8 -*-
"""Lotação da obra — FONTE ÚNICA (1 JSON) que os 3 fronts leem (planilha, GUI, construdata).
Assim a divisão das equipes nunca diverge entre eles.

JSON: {obra, contrato, data_ref, equipes:[{nome, lider, sistema, tipo, pessoas:[{nome,funcao}]}]}
"""
import json


def carregar(caminho):
    """Lê a lotação de .xlsx (planilha editável — preferida) OU .json."""
    if str(caminho).lower().endswith((".xlsx", ".xlsm")):
        return ler_xlsx(caminho)
    with open(caminho, encoding="utf-8") as f:
        return json.load(f)


def ler_xlsx(caminho):
    """Planilha = FONTE EDITÁVEL. Você edita o Excel (1 linha por pessoa: Nome,
    Função, Equipe, Líder, Sistema, Tipo) e o motor lê daqui. Sistema/Tipo/Líder
    são herdados da equipe (basta preencher numa linha)."""
    import collections
    from openpyxl import load_workbook
    wb = load_workbook(caminho, data_only=True)
    ws = wb["LOTAÇÃO"] if "LOTAÇÃO" in wb.sheetnames else wb.active
    hdr = next((r for r in range(1, 6) if str(ws.cell(r, 1).value).strip().lower() == "nome"), 3)
    eqs = collections.OrderedDict()
    for r in range(hdr + 1, ws.max_row + 1):
        nome = ws.cell(r, 1).value
        if not nome or str(nome).startswith(("Efetivo", "Pessoas")):
            continue
        func, equipe, lider, sis, tipo = (ws.cell(r, c).value for c in range(2, 7))
        if not equipe:
            continue
        e = eqs.setdefault(str(equipe), {"nome": str(equipe), "lider": "", "sistema": "",
                                         "tipo": "", "pessoas": []})
        e["pessoas"].append({"nome": str(nome), "funcao": str(func or "")})
        if lider and not e["lider"]:
            e["lider"] = str(lider)
        if sis and not e["sistema"]:
            e["sistema"] = str(sis)
        if tipo and not e["tipo"]:
            e["tipo"] = str(tipo)
    return {"obra": "Boi Malhado", "equipes": list(eqs.values())}


def pessoas_flat(lot):
    """[(nome, função, equipe, líder, sistema, tipo)] — uma linha por pessoa."""
    out = []
    for eq in lot["equipes"]:
        for p in eq["pessoas"]:
            out.append((p["nome"], p["funcao"], eq["nome"], eq["lider"], eq["sistema"], eq["tipo"]))
    return out


def resumo(lot):
    """Totais: campo, por sistema, por tipo."""
    flat = pessoas_flat(lot)
    return {
        "campo": len(flat),
        "rede": sum(1 for *_, tipo in flat if tipo == "Rede"),
        "caixa_ramal": sum(1 for *_, tipo in flat if tipo != "Rede"),
        "equipes": len(lot["equipes"]),
    }


def composicao(lot, equipe):
    """[(nome, função)] de uma equipe pelo nome."""
    for eq in lot["equipes"]:
        if eq["nome"] == equipe:
            return [(p["nome"], p["funcao"]) for p in eq["pessoas"]]
    return []


def frentes_supabase(lot):
    """Frentes prontas p/ inserir no construdata (tabela frentes)."""
    return [{"nome": eq["nome"], "tipo_rede": eq["sistema"].lower(), "encarregado": eq["lider"],
             "lider": eq["lider"], "tipo": eq["tipo"], "pessoas": len(eq["pessoas"]),
             "status": "ativa"} for eq in lot["equipes"]]


def gerar_xlsx(lot, caminho):
    """Planilha PESSOAL (mesma do MASTER): Nome/Função/Equipe/Líder/Sistema/Tipo + resumo."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    H = PatternFill("solid", fgColor="1F4E79"); HF = Font(bold=True, color="FFFFFF")
    TIT = Font(bold=True, size=13, color="1F4E79"); B = Font(bold=True)
    thin = Side(style="thin", color="BBBBBB"); BORD = Border(thin, thin, thin, thin)
    wb = Workbook(); ws = wb.active; ws.title = "LOTAÇÃO"
    ws.cell(1, 1, "LOTAÇÃO — %s  (ref. %s)" % (lot.get("obra", ""), lot.get("data_ref", ""))).font = TIT
    cols = ["Nome", "Função", "Equipe", "Líder", "Sistema", "Tipo"]
    for ci, h in enumerate(cols, 1):
        c = ws.cell(3, ci, h); c.font = HF; c.fill = H; c.alignment = Alignment(horizontal="center"); c.border = BORD
    r = 4
    for linha in pessoas_flat(lot):
        for ci, v in enumerate(linha, 1):
            ws.cell(r, ci, v).border = BORD
        r += 1
    res = resumo(lot)
    ws.cell(r + 1, 1, "Efetivo de campo").font = B; ws.cell(r + 1, 2, res["campo"])
    ws.cell(r + 2, 1, "Pessoas REDE").font = B; ws.cell(r + 2, 2, res["rede"])
    ws.cell(r + 3, 1, "Pessoas CAIXA E RAMAL").font = B; ws.cell(r + 3, 2, res["caixa_ramal"])
    for col, w in zip("ABCDEF", [30, 20, 22, 22, 10, 16]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"
    wb.save(str(caminho))
    return {"saida": str(caminho), **res}
