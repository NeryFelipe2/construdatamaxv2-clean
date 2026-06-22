# -*- coding: utf-8 -*-
"""Planejamento MASTER — gera a planilha VIVA (estilo PLANEJAMENTO_BOI_MALHADO_MASTER)
a partir dos dados da obra. Cenários VCA × MND com FÓRMULAS que recalculam sozinhas
(troca o cenário/produtividade/alocação → dimensionamento e cronograma mudam).

É o motor único do replanejamento: o construdata (online), o GUI e a planilha
chamam isto. `gerar_master_xlsx(dados, caminho)` escreve o .xlsx.

dados = {
  contrato, nucleo, data_inicio (date), dias_uteis_semana, cenario ('VCA'|'MND'),
  quantitativos: [(servico, qtd, und)],                 # 5 linhas
  produtividade: {servico: (vca, mnd)},                 # por serviço
  frentes: [(frente, encarregado, lideranca, tipo, pessoas, func_vca, func_mnd)],
  prazo_alvo_dias,
}
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

AZUL = "1F4E79"
H_FILL = PatternFill("solid", fgColor=AZUL)
H_FONT = Font(bold=True, color="FFFFFF")
TIT = Font(bold=True, size=13, color=AZUL)
SUB = Font(italic=True, size=9, color="555555")
B = Font(bold=True)
AMARELO = PatternFill("solid", fgColor="FFF2CC")   # células editáveis
thin = Side(style="thin", color="BBBBBB")
BORD = Border(thin, thin, thin, thin)
CEN = Alignment(horizontal="center", vertical="center")


def _hdr(ws, row, cols, start=1):
    for i, h in enumerate(cols):
        c = ws.cell(row, start + i, h); c.font = H_FONT; c.fill = H_FILL
        c.alignment = CEN; c.border = BORD


def gerar_master_xlsx(dados, caminho):
    wb = Workbook()

    # ---------- PESSOAL (referência) ----------
    ws = wb.active; ws.title = "PESSOAL"
    ws.cell(1, 1, "LOTAÇÃO REAL — %s" % dados.get("nucleo", "")).font = TIT
    _hdr(ws, 3, ["Frente", "Encarregado", "Liderança", "Tipo", "Pessoas"])
    for i, fr in enumerate(dados["frentes"]):
        for ci, v in enumerate(fr[:5]):
            ws.cell(4 + i, 1 + ci, v).border = BORD
    n = len(dados["frentes"]); r = 4 + n + 1
    ws.cell(r, 1, "Efetivo de campo").font = B
    ws.cell(r, 5, "=SUM(E4:E%d)" % (3 + n))
    for col, w in zip("ABCDE", [26, 16, 16, 16, 10]):
        ws.column_dimensions[col].width = w

    # ---------- PREMISSAS (edite o amarelo) ----------
    wp = wb.create_sheet("PREMISSAS")
    wp.cell(1, 1, "PREMISSAS (edite o amarelo)").font = TIT
    wp.cell(3, 1, "CENÁRIO").font = B
    cen = wp.cell(3, 2, dados.get("cenario", "VCA")); cen.fill = AMARELO; cen.font = B
    wp.cell(3, 3, "← VCA (vala aberta) ou MND (não destrutivo). O cronograma muda sozinho.").font = SUB
    wp.cell(5, 1, "Dias úteis/semana").font = B
    du = wp.cell(5, 2, dados.get("dias_uteis_semana", 6)); du.fill = AMARELO
    wp.cell(6, 1, "Data início da obra").font = B
    di = wp.cell(6, 2, dados["data_inicio"]); di.fill = AMARELO; di.number_format = "dd/mm/yyyy"
    wp.cell(8, 1, "PRODUTIVIDADE por frente/dia").font = B
    _hdr(wp, 9, ["Serviço", "VCA", "MND"])
    serv_rows = {}   # serviço -> linha na PREMISSAS (p/ fórmulas)
    pr = 10
    for serv, _qtd, _und in dados["quantitativos"]:
        vca, mnd = dados["produtividade"].get(serv, (1, 1))
        wp.cell(pr, 1, serv).border = BORD
        a = wp.cell(pr, 2, vca); a.fill = AMARELO; a.border = BORD
        bnd = wp.cell(pr, 3, mnd); bnd.fill = AMARELO; bnd.border = BORD
        serv_rows[serv] = pr; pr += 1
    # nº de frentes MND manuais (rede) — usados no CENÁRIOS p/ água/esgoto
    wp.cell(pr + 1, 1, "Frentes MND rede água (manual)").font = B
    fa = wp.cell(pr + 1, 2, 1); fa.fill = AMARELO
    wp.cell(pr + 2, 1, "Frentes MND rede esgoto (manual)").font = B
    fe = wp.cell(pr + 2, 2, 1); fe.fill = AMARELO
    _FA, _FE = pr + 1, pr + 2
    for col, w in zip("ABC", [34, 10, 10]):
        wp.column_dimensions[col].width = w

    # ---------- QUANTITATIVOS ----------
    wq = wb.create_sheet("QUANTITATIVOS")
    wq.cell(1, 1, "QUANTITATIVOS — %s" % dados.get("nucleo", "")).font = TIT
    _hdr(wq, 3, ["Serviço", "Qtd", "Unid"])
    q_rows = {}
    for i, (serv, qtd, und) in enumerate(dados["quantitativos"]):
        rr = 4 + i
        wq.cell(rr, 1, serv).border = BORD
        wq.cell(rr, 2, qtd).border = BORD
        wq.cell(rr, 3, und).border = BORD
        q_rows[serv] = rr
    for col, w in zip("ABC", [32, 12, 8]):
        wq.column_dimensions[col].width = w

    # ---------- ALOCACAO ----------
    wa = wb.create_sheet("ALOCACAO")
    wa.cell(1, 1, "ALOCAÇÃO DE EQUIPES — você decide o que cada frente faz").font = TIT
    wa.cell(2, 1, "Escolha a Função de cada frente em cada cenário (amarelo). Dimensionamento e cronograma recalculam.").font = SUB
    _hdr(wa, 3, ["Frente", "Encarregado", "Liderança", "Tipo", "Pessoas", "Função VCA", "Função MND"])
    for i, fr in enumerate(dados["frentes"]):
        rr = 4 + i
        for ci, v in enumerate(fr):
            c = wa.cell(rr, 1 + ci, v); c.border = BORD
            if ci in (5, 6):   # funções = editáveis
                c.fill = AMARELO
    aloc_ini, aloc_fim = 4, 3 + len(dados["frentes"])
    for col, w in zip("ABCDEFG", [22, 14, 16, 14, 9, 22, 22]):
        wa.column_dimensions[col].width = w

    # refs absolutas p/ COUNTIF na ALOCACAO
    FVCA = "ALOCACAO!$F$%d:$F$%d" % (aloc_ini, aloc_fim)
    FMND = "ALOCACAO!$G$%d:$G$%d" % (aloc_ini, aloc_fim)

    # ---------- DIMENSIONAMENTO ----------
    wd = wb.create_sheet("DIMENSIONAMENTO")
    wd.cell(1, 1, "DIMENSIONAMENTO — equipes necessárias (VCA × MND)").font = TIT
    wd.cell(2, 1, "Quantas frentes preciso pra cada prazo. Muda CENÁRIO/produtividade → recalcula.").font = SUB
    wd.cell(3, 1, "Prazo-alvo (dias úteis)").font = B
    pa = wd.cell(3, 2, dados.get("prazo_alvo_dias", 60)); pa.fill = AMARELO
    _hdr(wd, 5, ["Serviço", "Qtd", "Prod VCA", "Frentes nec. VCA", "Prod MND",
                 "Frentes nec. MND", "Frentes hoje", "+ Frentes VCA", "+ Frentes MND"])
    # mapa serviço -> rótulo de função usado na ALOCACAO (p/ COUNTIF "frentes hoje")
    func_label = {
        "Rede de água (×2 dois terços)": "Rede água", "Rede de água": "Rede água",
        "Caixa U.M.A": "Caixa U.M.A", "Ligação água + HM": "Ligação água",
        "Rede de esgoto": "Rede esgoto", "Caixa inspeção + ramal": "Caixa inspeção + ramal",
    }
    for i, (serv, _qtd, _und) in enumerate(dados["quantitativos"]):
        rr = 6 + i
        qref = "QUANTITATIVOS!B%d" % q_rows[serv]
        pvca = "PREMISSAS!B%d" % serv_rows[serv]
        pmnd = "PREMISSAS!C%d" % serv_rows[serv]
        lbl = func_label.get(serv, serv)
        wd.cell(rr, 1, "=QUANTITATIVOS!A%d" % q_rows[serv]).border = BORD
        wd.cell(rr, 2, "=" + qref).border = BORD
        wd.cell(rr, 3, "=" + pvca).border = BORD
        wd.cell(rr, 4, "=ROUNDUP(B%d/(C%d*$B$3),0)" % (rr, rr)).border = BORD
        wd.cell(rr, 5, "=" + pmnd).border = BORD
        wd.cell(rr, 6, "=ROUNDUP(B%d/(E%d*$B$3),0)" % (rr, rr)).border = BORD
        wd.cell(rr, 7, '=COUNTIF(%s,"%s")' % (FVCA, lbl)).border = BORD
        wd.cell(rr, 8, "=MAX(0,D%d-G%d)" % (rr, rr)).border = BORD
        wd.cell(rr, 9, "=MAX(0,F%d-G%d)" % (rr, rr)).border = BORD
    rt = 6 + len(dados["quantitativos"]) + 1
    wd.cell(rt, 1, "TOTAL").font = B
    wd.cell(rt, 4, "=SUM(D6:D%d)" % (rt - 2))
    wd.cell(rt, 6, "=SUM(F6:F%d)" % (rt - 2))
    wd.cell(rt, 8, "=SUM(H6:H%d)" % (rt - 2))
    wd.cell(rt, 9, "=SUM(I6:I%d)" % (rt - 2))
    for col, w in zip("ABCDEFGHI", [28, 10, 9, 15, 9, 15, 12, 13, 13]):
        wd.column_dimensions[col].width = w

    # ---------- CENÁRIOS (prazo VCA × MND) ----------
    wc = wb.create_sheet("CENÁRIOS")
    wc.cell(1, 1, "CENÁRIOS — VCA × MND (prazo e término)").font = TIT
    wc.cell(2, 1, "Com as frentes alocadas, em quanto tempo cada cenário entrega. Recalcula sozinho.").font = SUB
    _hdr(wc, 4, ["Serviço", "Qtd", "Frentes VCA", "Prod VCA", "Dias VCA",
                 "Frentes MND", "Prod MND", "Dias MND"])
    for i, (serv, _qtd, _und) in enumerate(dados["quantitativos"]):
        rr = 5 + i
        lbl = func_label.get(serv, serv)
        wc.cell(rr, 1, serv).border = BORD
        wc.cell(rr, 2, "=QUANTITATIVOS!B%d" % q_rows[serv]).border = BORD
        wc.cell(rr, 3, '=MAX(1,COUNTIF(%s,"%s"))' % (FVCA, lbl)).border = BORD
        wc.cell(rr, 4, "=PREMISSAS!B%d" % serv_rows[serv]).border = BORD
        wc.cell(rr, 5, "=IFERROR(ROUNDUP(B%d/(D%d*C%d),0),0)" % (rr, rr, rr)).border = BORD
        # MND: rede usa frentes manuais; demais COUNTIF MND
        if "Rede de água" in serv:
            fmnd = "MAX(1,PREMISSAS!$B$%d)" % _FA
        elif serv == "Rede de esgoto":
            fmnd = "MAX(1,PREMISSAS!$B$%d)" % _FE
        else:
            fmnd = '=MAX(1,COUNTIF(%s,"%s"))' % (FMND, lbl); fmnd = fmnd[1:]
        wc.cell(rr, 6, "=" + fmnd).border = BORD
        wc.cell(rr, 7, "=PREMISSAS!C%d" % serv_rows[serv]).border = BORD
        wc.cell(rr, 8, "=IFERROR(ROUNDUP(B%d/(G%d*F%d),0),0)" % (rr, rr, rr)).border = BORD
    rp = 5 + len(dados["quantitativos"]) + 1
    wc.cell(rp, 1, "Prazo (dias úteis, caminho crítico)").font = B
    wc.cell(rp, 5, "=MAX(E5:E%d)" % (rp - 2)).font = B
    wc.cell(rp, 8, "=MAX(H5:H%d)" % (rp - 2)).font = B
    wc.cell(rp + 1, 1, "Término previsto").font = B
    wc.cell(rp + 1, 5, "=PREMISSAS!B6+E%d/PREMISSAS!B5*7" % rp).number_format = "dd/mm/yyyy"
    wc.cell(rp + 1, 8, "=PREMISSAS!B6+H%d/PREMISSAS!B5*7" % rp).number_format = "dd/mm/yyyy"
    for col, w in zip("ABCDEFGH", [30, 10, 11, 9, 10, 11, 9, 10]):
        wc.column_dimensions[col].width = w

    # ---------- CRONOGRAMA (por frente, muda com o cenário) ----------
    wk = wb.create_sheet("CRONOGRAMA")
    wk.cell(1, 1, "CRONOGRAMA POR EQUIPE — muda com o CENÁRIO (VCA/MND)").font = TIT
    wk.cell(2, 1, '=("Cenário atual: "&PREMISSAS!B3&"  ·  prazo ≈ "&MAX(G6:G%d)&" semanas")'
            % (5 + len(dados["quantitativos"]))).font = SUB
    n_sem = 26
    cab = ["Frente / serviço", "Sistema", "Qtd", "Prod/dia", "Frentes", "Dias",
           "Semanas", "Ini(sem)", "Fim(sem)"] + list(range(1, n_sem + 1))
    _hdr(wk, 4, cab)
    sistema_de = {"Rede de água (×2 dois terços)": "AGUA", "Caixa U.M.A": "AGUA",
                  "Ligação água + HM": "AGUA", "Rede de esgoto": "ESGOTO",
                  "Caixa inspeção + ramal": "ESGOTO"}
    for i, (serv, _qtd, _und) in enumerate(dados["quantitativos"]):
        rr = 6 + i
        lbl = func_label.get(serv, serv)
        wk.cell(rr, 1, serv).border = BORD
        wk.cell(rr, 2, sistema_de.get(serv, "")).border = BORD
        wk.cell(rr, 3, "=QUANTITATIVOS!B%d" % q_rows[serv]).border = BORD
        wk.cell(rr, 4, '=IF(PREMISSAS!$B$3="MND",PREMISSAS!C%d,PREMISSAS!B%d)'
                % (serv_rows[serv], serv_rows[serv])).border = BORD
        if "Rede de água" in serv:
            frentes = 'MAX(1,IF(PREMISSAS!$B$3="MND",PREMISSAS!$B$%d,COUNTIF(%s,"%s")))' % (_FA, FVCA, lbl)
        elif serv == "Rede de esgoto":
            frentes = 'MAX(1,IF(PREMISSAS!$B$3="MND",PREMISSAS!$B$%d,COUNTIF(%s,"%s")))' % (_FE, FVCA, lbl)
        else:
            frentes = 'MAX(1,IF(PREMISSAS!$B$3="MND",COUNTIF(%s,"%s"),COUNTIF(%s,"%s")))' % (FMND, lbl, FVCA, lbl)
        wk.cell(rr, 5, "=" + frentes).border = BORD
        wk.cell(rr, 6, "=ROUNDUP(C%d/(D%d*E%d),0)" % (rr, rr, rr)).border = BORD
        wk.cell(rr, 7, "=ROUNDUP(F%d/PREMISSAS!$B$5,0)" % rr).border = BORD
        wk.cell(rr, 8, 1).border = BORD
        wk.cell(rr, 9, "=H%d+G%d-1" % (rr, rr)).border = BORD
        # barras (█) por semana
        for s in range(n_sem):
            col = 10 + s
            cl = wk.cell(rr, col, '=IF(AND(%s$4>=$H%d,%s$4<=$I%d),"█","")'
                         % (_col_letter(col), rr, _col_letter(col), rr))
            cl.alignment = CEN
    for col, w in zip("ABCDEFGHI", [26, 9, 9, 9, 9, 7, 9, 9, 9]):
        wk.column_dimensions[col].width = w

    wb.save(str(caminho))
    return {"saida": str(caminho), "abas": wb.sheetnames, "frentes": len(dados["frentes"])}


def _col_letter(idx):
    from openpyxl.utils import get_column_letter
    return get_column_letter(idx)
