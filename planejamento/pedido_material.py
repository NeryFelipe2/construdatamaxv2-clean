# -*- coding: utf-8 -*-
"""Pedido de Material — padrão WCR Saneamento (Contrato Sabesp 13.546/25-00).

Gera o pedido em Excel editável + PDF para envio/impressão. `itens` =
[{descricao, qtd, und, obs?}] (a numeração é automática). `cab` = cabeçalho
{nucleo, sistema, pedido_n, data, solicitante, destino}.
"""
EMPRESA = "WCR Saneamento"
CONTRATO = "Sabesp 13.546/25-00"
AZUL = "#1a237e"


def _cab_padrao(cab):
    base = {"nucleo": "", "sistema": "", "pedido_n": "______",
            "data": "____/____/______", "solicitante": "________________________",
            "destino": "Almoxarifado central", "encarregado": "________________________"}
    base.update(cab or {})
    return base


# ----------------------------------------------------------------- Excel
def gerar_pedido_xlsx(itens, cab, caminho):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    cab = _cab_padrao(cab)
    H_FILL = PatternFill("solid", fgColor="1F4E79"); H_FONT = Font(bold=True, color="FFFFFF")
    TIT = Font(bold=True, size=14, color="1F4E79"); B = Font(bold=True)
    thin = Side(style="thin", color="9AA0B4"); BORD = Border(thin, thin, thin, thin)
    cen = Alignment(horizontal="center", vertical="center")

    wb = Workbook(); ws = wb.active; ws.title = "PEDIDO DE MATERIAL"
    ws.merge_cells("A1:E1")
    ws.cell(1, 1, "PEDIDO DE MATERIAL — %s" % EMPRESA).font = TIT
    ws.cell(2, 1, "Contrato %s" % CONTRATO).font = B
    # bloco de identificação
    ident = [("Núcleo:", cab["nucleo"], "Sistema:", cab["sistema"]),
             ("Pedido nº:", cab["pedido_n"], "Data:", cab["data"]),
             ("Solicitante:", cab["solicitante"], "Destino:", cab["destino"]),
             ("Encarregado geral:", cab["encarregado"], "", "")]
    r = 4
    for a, b, c, d in ident:
        ws.cell(r, 1, a).font = B; ws.cell(r, 2, b)
        ws.cell(r, 4, c).font = B; ws.cell(r, 5, d)
        r += 1
    # tabela
    r += 1
    headers = ["Item", "Descrição do material", "Quantidade", "Unidade", "Conferência"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(r, ci, h); cell.font = H_FONT; cell.fill = H_FILL; cell.alignment = cen; cell.border = BORD
    r0 = r + 1
    for i, it in enumerate(itens, 1):
        ws.cell(r0, 1, i).alignment = cen
        ws.cell(r0, 2, it["descricao"])
        ws.cell(r0, 3, it["qtd"]).alignment = cen
        ws.cell(r0, 4, it["und"]).alignment = cen
        ws.cell(r0, 5, it.get("obs", ""))
        for ci in range(1, 6):
            ws.cell(r0, ci).border = BORD
        r0 += 1
    # total + assinaturas
    ws.cell(r0 + 1, 2, "Total de itens: %d" % len(itens)).font = B
    ws.cell(r0 + 3, 1, "Solicitante: ____________________")
    ws.cell(r0 + 3, 3, "Almoxarife: ____________________")
    ws.cell(r0 + 5, 1, "Aprovação (Eng.): ____________________")
    for col, w in zip("ABCDE", [6, 52, 12, 12, 20]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A%d" % (r + 1)        # congela o cabeçalho da tabela
    wb.save(str(caminho))
    return {"saida": str(caminho), "itens": len(itens)}


# ----------------------------------------------------------------- PDF
def _desenha_pagina(fig, plt, FancyBboxPatch, bloco, cab, total_itens, pi, npags, n0_start):
    # banner
    axb = fig.add_axes([0.06, 0.93, 0.88, 0.055]); axb.axis("off")
    axb.add_patch(FancyBboxPatch((0, 0), 1, 1, boxstyle="round,pad=0.01",
                                 transform=axb.transAxes, fc=AZUL, ec="none"))
    axb.text(0.5, 0.62, "PEDIDO DE MATERIAL", ha="center", va="center",
             fontsize=16, fontweight="bold", color="white", transform=axb.transAxes)
    axb.text(0.5, 0.22, "%s · Contrato %s" % (EMPRESA, CONTRATO), ha="center",
             va="center", fontsize=8.5, color="#bbdefb", transform=axb.transAxes)
    # identificação
    axi = fig.add_axes([0.06, 0.805, 0.88, 0.115]); axi.axis("off")
    linhas = [("Núcleo: %s" % cab["nucleo"], "Sistema: %s" % cab["sistema"]),
              ("Pedido nº: %s" % cab["pedido_n"], "Data: %s" % cab["data"]),
              ("Solicitante: %s" % cab["solicitante"], "Destino: %s" % cab["destino"]),
              ("Encarregado geral: %s" % cab["encarregado"], "")]
    yy = 0.88
    for esq, dir_ in linhas:
        axi.text(0.0, yy, esq, fontsize=9.5, transform=axi.transAxes)
        if dir_:
            axi.text(0.52, yy, dir_, fontsize=9.5, transform=axi.transAxes)
        yy -= 0.26
    # tabela
    axt = fig.add_axes([0.05, 0.06, 0.90, 0.73]); axt.axis("off")
    col_x = [0.0, 0.08, 0.60, 0.73, 0.85, 1.0]   # bordas: Item|Descrição|Qtd|Und|Conferência
    headers = ["Item", "Descrição do material", "Qtd", "Und", "Conferência"]
    ytop = 1.0; rowh = min(0.034, 0.92 / (len(bloco) + 2))
    axt.add_patch(plt.Rectangle((0, ytop - rowh), 1, rowh, transform=axt.transAxes, fc=AZUL, ec="white"))
    for j, h in enumerate(headers):
        axt.text((col_x[j] + col_x[j + 1]) / 2 if j != 1 else col_x[1] + 0.01,
                 ytop - rowh / 2, h, transform=axt.transAxes, ha="center" if j != 1 else "left",
                 va="center", fontsize=8, color="white", fontweight="bold")
    y = ytop - rowh; n0 = n0_start
    for k, it in enumerate(bloco):
        n0 += 1
        if k % 2 == 0:
            axt.add_patch(plt.Rectangle((0, y - rowh), 1, rowh, transform=axt.transAxes, fc="#eef1f8", ec="none"))
        cells = [str(n0), it["descricao"], str(it["qtd"]), str(it["und"]), it.get("obs", "")]
        for j, c in enumerate(cells):
            ha = "left" if j == 1 else "center"
            x = col_x[1] + 0.01 if j == 1 else (col_x[j] + col_x[j + 1]) / 2
            axt.text(x, y - rowh / 2, c, transform=axt.transAxes, ha=ha, va="center", fontsize=8)
        y -= rowh
    for xb in col_x:
        axt.plot([xb, xb], [y, ytop], transform=axt.transAxes, color="#9aa0b4", lw=0.6)
    yy = ytop
    for _ in range(len(bloco) + 1):
        axt.plot([0, 1], [yy, yy], transform=axt.transAxes, color="#9aa0b4", lw=0.6); yy -= rowh
    axt.plot([0, 1], [y, y], transform=axt.transAxes, color="#9aa0b4", lw=0.6)
    if pi == npags - 1:
        axt.text(0.0, y - 0.04, "Total de itens: %d" % total_itens, transform=axt.transAxes,
                 fontsize=9.5, fontweight="bold")
        axt.text(0.0, y - 0.10, "Solicitante: ___________________     "
                 "Almoxarife: ___________________     Aprovação (Eng.): ___________________",
                 transform=axt.transAxes, fontsize=8.5, color="#333")
    else:
        axt.text(0.98, 0.0, "continua…", transform=axt.transAxes, ha="right", fontsize=8, color="#666")


def _paginar(itens, por_pag=26):
    return [itens[i:i + por_pag] for i in range(0, len(itens), por_pag)] or [[]]


def gerar_pedido_pdf(itens, cab, caminho):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.patches import FancyBboxPatch
    from matplotlib.backends.backend_pdf import PdfPages
    cab = _cab_padrao(cab); paginas = _paginar(itens)
    with PdfPages(str(caminho)) as pdf:
        n0 = 0
        for pi, bloco in enumerate(paginas):
            fig = plt.figure(figsize=(8.27, 11.69)); fig.patch.set_facecolor("white")
            _desenha_pagina(fig, plt, FancyBboxPatch, bloco, cab, len(itens), pi, len(paginas), n0)
            n0 += len(bloco)
            pdf.savefig(fig); plt.close(fig)
    return str(caminho)


def gerar_pedido_png(itens, cab, caminho):
    """Renderiza a 1ª página em PNG (para conferência rápida)."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.patches import FancyBboxPatch
    cab = _cab_padrao(cab); bloco = _paginar(itens)[0]
    fig = plt.figure(figsize=(8.27, 11.69)); fig.patch.set_facecolor("white")
    _desenha_pagina(fig, plt, FancyBboxPatch, bloco, cab, len(itens), 0, len(_paginar(itens)), 0)
    fig.savefig(str(caminho), dpi=110, bbox_inches="tight"); plt.close(fig)
    return str(caminho)
