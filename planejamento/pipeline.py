# -*- coding: utf-8 -*-
"""MOTOR ÚNICO de planejamento por núcleo — consolida os scripts descartáveis do
C:\\tmp (build_final_*, build_project_*, build_pedido_*, fix_gpkg_*) num só lugar,
parametrizável por config. UM comando por núcleo → Project + planilha + NS PDF +
pedido + gpkg atualizado + nota Obsidian. Todos os canais (WhatsApp/online/GUI)
chamam isto. Agenda por HORAS úteis (serve água MND 300 m/dia e esgoto ~48 m/dia).

cfg (dict) — ver planejamento/nucleos.py para exemplos.
"""
import os, math, unicodedata
from collections import OrderedDict
from datetime import date, datetime, timedelta
import fiona
import geopandas as gpd
from shapely.geometry import shape, box
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.patches import FancyBboxPatch, Rectangle
from matplotlib.backends.backend_pdf import PdfPages
from . import msproject


# ---------------- utils ----------------
def nrm(s):
    return "".join(c for c in unicodedata.normalize("NFD", str(s or "")) if not unicodedata.combining(c)).upper()

def G(pr, *ks):
    for k in ks:
        for kk in pr:
            if nrm(kk) == nrm(k):
                return pr[kk]
    return None


# ---------------- calendário (horas úteis, 6 dias/sem) ----------------
def _normaliza(dt):
    cur = dt
    while True:
        if cur.weekday() == 6:
            cur = (cur + timedelta(days=1)).replace(hour=8, minute=0, second=0, microsecond=0); continue
        h = cur.hour + cur.minute / 60.0
        if h < 8: cur = cur.replace(hour=8, minute=0, second=0, microsecond=0); continue
        if 12 <= h < 13: cur = cur.replace(hour=13, minute=0, second=0, microsecond=0); continue
        if h >= 17:
            nd = cur + timedelta(days=1)
            while nd.weekday() == 6: nd += timedelta(days=1)
            cur = nd.replace(hour=8, minute=0, second=0, microsecond=0); continue
        return cur

def _avanca(dt, horas):
    cur = _normaliza(dt); rem = float(horas)
    while rem > 1e-9:
        cur = _normaliza(cur); h = cur.hour + cur.minute / 60.0
        seg = 12.0 if h < 12 else 17.0; avail = seg - h
        if rem <= avail + 1e-9: cur = cur + timedelta(hours=rem); rem = 0
        else: rem -= avail; cur = cur + timedelta(hours=avail)
    return cur


# ---------------- 1) leitura + sequência + datas ----------------
def ler_trechos(cfg):
    pa = cfg["gpkg"]
    feats = list(fiona.open(pa, layer=fiona.listlayers(pa)[0]))
    tr = []
    for f in feats:
        p = dict(f["properties"]); g = shape(f["geometry"])
        comp = G(p, "comp_label", "comprimento", "Tubo_m")
        comp = float(comp) if comp not in (None, "") else round(g.length, 1)
        tr.append({"p": p, "g": g, "comp": comp,
                   "dn": str(G(p, "DN") or cfg.get("dn_padrao", "")),
                   "metodo": "MND" if "MND" in nrm(G(p, "TIpo de vala", "Tipo de vala", "Mec-Manual", "metodo")) or cfg.get("metodo") == "MND" else "VCA"})
    # NS estável = ordem nativa do gpkg
    for i, t in enumerate(tr, 1):
        t["ns"] = i
    return tr

def sequenciar(tr, cfg):
    """Define equipe (t['team'], t['lider']) e ordem (t['pos']) por equipe."""
    eqmap = cfg["equipes"]
    campo = cfg.get("equipe_campo")
    for t in tr:
        if campo:
            v = str(G(t["p"], campo) or "").strip()
            v = v[:-2] if v.endswith(".0") else v
            t["team"] = v if v in eqmap else next(iter(eqmap))
        else:
            t["team"] = next(iter(eqmap))
        t["lider"] = eqmap[t["team"]]["lider"]
    grupos = OrderedDict((k, [t for t in tr if t["team"] == k]) for k in eqmap if any(t["team"] == k for t in tr))
    ocampo = cfg.get("ordem_campo")
    for k, lst in grupos.items():
        if ocampo and any(G(t["p"], ocampo) not in (None, "") for t in lst):
            lst.sort(key=lambda t: (float(G(t["p"], ocampo) or 999), t["ns"]))
        else:  # vizinho-mais-próximo do canto SO
            rest = sorted(lst, key=lambda t: t["g"].centroid.x + t["g"].centroid.y); seq = [rest.pop(0)]
            while rest:
                c = seq[-1]["g"].centroid
                rest.sort(key=lambda t: (t["g"].centroid.x - c.x) ** 2 + (t["g"].centroid.y - c.y) ** 2)
                seq.append(rest.pop(0))
            lst[:] = seq
        for i, t in enumerate(lst, 1):
            t["pos"] = i
    return grupos

def agendar(grupos, cfg):
    prod = float(cfg["produtividade"]); y, m, d = cfg["data_inicio"]
    for k, lst in grupos.items():
        cur = _normaliza(datetime(y, m, d, 8))
        for t in lst:
            t["dur_h"] = max(0.25, t["comp"] / prod * 8.0)
            t["s"] = _normaliza(cur); t["f"] = _avanca(t["s"], t["dur_h"])
            t["ini"] = t["s"].strftime("%d/%m/%Y"); t["fim"] = t["f"].strftime("%d/%m/%Y")
            cur = t["f"]
    todos = [t for lst in grupos.values() for t in lst]
    return min(t["s"] for t in todos), max(t["f"] for t in todos)


# ---------------- materiais ----------------
def materiais(t, cfg):
    comp = t["comp"]; bar = math.ceil(comp / 6) if comp else 0
    if cfg["sistema"] == "agua":
        its = [("Tubo PEAD %s água" % (t["dn"] or "DN63"), round(comp, 1), "m"), ("  └ barras 6 m", bar, "barra")]
        if t["metodo"] != "MND":
            its += [("Escavação", round(comp * 0.6 * 0.8, 1), "m³"), ("Areia", round(comp * 0.6 * 0.15, 1), "m³"),
                    ("Brita", round(comp * 0.6 * 0.15, 1), "m³")]
        its += [("Caixa U.M.A", "conf. casas", "un"), ("Ligação de água", "conf. casas", "un")]
        return its
    # esgoto
    prof = float(G(t["p"], "Prof_med") or 1.5)
    its = [("Tubo PVC %s esgoto" % (t["dn"] or "DN200"), round(comp, 1), "m"), ("  └ barras 6 m", bar, "barra")]
    if t["metodo"] != "MND":
        its += [("Escavação", round(comp * 0.7 * prof, 1), "m³"), ("Areia", round(comp * 0.7 * 0.15, 1), "m³"),
                ("Brita", round(comp * 0.7 * 0.15, 1), "m³")]
    for fld, lab in [("PV", "Poço de visita"), ("Cx_Insp", "Caixa de inspeção"), ("Ramais", "Ramal de esgoto")]:
        q = int(G(t["p"], fld) or 0)
        if q: its.append((lab, q, "un"))
    return its


# ---------------- contexto + mapas ----------------
def carregar_fundo(cfg):
    guia, rede = [], []
    for fb in cfg.get("fundo", []):
        try:
            if fb.get("tipo") == "dxf":
                import ezdxf
                doc = ezdxf.readfile(fb["path"])
                for e in doc.modelspace().query('LINE[layer=="%s"]' % fb["layer"]):
                    a, b = e.dxf.start, e.dxf.end
                    guia.append(((a.x, a.y), (b.x, b.y)))
            else:
                with fiona.open(fb["path"], layer=fb.get("layer") or fiona.listlayers(fb["path"])[0]) as src:
                    for f in src:
                        if not fb.get("filtro_layer") or nrm(f["properties"].get("layer")) == nrm(fb["filtro_layer"]):
                            (rede if fb.get("rede") else guia).append(shape(f["geometry"]))
        except Exception as e:
            print("(fundo %s: %s)" % (fb.get("path"), e))
    return guia, rede

def _desenha(ax, t, guia, rede, redetr, cor, buf=110):
    b = t["g"].buffer(buf).bounds; clip = box(*b)
    for gg in guia:
        if isinstance(gg, tuple):
            ax.plot([gg[0][0], gg[1][0]], [gg[0][1], gg[1][1]], color="0.85", lw=0.4)
        elif gg.intersects(clip):
            for ln in ([gg] if gg.geom_type == "LineString" else list(getattr(gg, "geoms", [gg]))):
                try: xs, ys = ln.xy; ax.plot(xs, ys, color="0.84", lw=0.5)
                except Exception: pass
    for gg in list(rede) + list(redetr):
        if gg.intersects(clip):
            for ln in ([gg] if gg.geom_type == "LineString" else list(getattr(gg, "geoms", [gg]))):
                try: xs, ys = ln.xy; ax.plot(xs, ys, color="#b0bec5", lw=0.8)
                except Exception: pass
    for ln in ([t["g"]] if t["g"].geom_type == "LineString" else list(t["g"].geoms)):
        xs, ys = ln.xy; ax.plot(xs, ys, color=cor, lw=3.2)
        ax.plot(xs[0], ys[0], "o", color="#2e7d32", ms=7); ax.plot(xs[-1], ys[-1], "s", color="#1565c0", ms=7)
    ax.set_xlim(b[0], b[2]); ax.set_ylim(b[1], b[3]); ax.set_aspect("equal"); ax.set_xticks([]); ax.set_yticks([])


# ---------------- saídas ----------------
def gerar_project(grupos, cfg, ini, fim):
    g2 = OrderedDict()
    for k, lst in grupos.items():
        eq = cfg["equipes"][k]
        trs = [{"rotulo": "%dº · %s %02d · %.0f m · %dd" % (t["pos"], cfg["prefixo"], t["ns"], t["comp"], max(1, round(t["dur_h"] / 8))),
                "s": t["s"], "f": t["f"], "dur_h": t["dur_h"],
                "notas": "%s — %s" % (eq["nome"], ", ".join("%s (%s)" % (n, fn) for n, fn in eq.get("crew", [])))} for t in lst]
        g2[eq["nome"]] = {"recurso": eq["nome"], "trechos": trs}
    out = os.path.join(cfg["saida_dir"], cfg["arq_project"])
    return msproject.gerar(out, "Cronograma %s - %s (WCR)" % (cfg["sistema"].upper(), cfg["nucleo"]), g2, ini, fim)

def _mini(t, cfg, guia, rede, redetr, mdir):
    path = os.path.join(mdir, "t_%02d.png" % t["ns"])
    fig = plt.figure(figsize=(3.4, 2.3), dpi=100); ax = fig.add_axes([0, 0, 1, 1]); ax.axis("off")
    _desenha(ax, t, guia, rede, redetr, cfg["cor"], 90); fig.savefig(path, dpi=100); plt.close(fig)
    return path

def gerar_planilha(grupos, cfg, guia, rede, mdir):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.worksheet.datavalidation import DataValidation
    redetr = [t["g"] for lst in grupos.values() for t in lst]
    cor = cfg["cor"].lstrip("#")
    HF = PatternFill("solid", fgColor=cor); HFONT = Font(bold=True, color="FFFFFF"); TIT = Font(bold=True, size=13, color=cor)
    thin = Side(style="thin", color="CCCCCC"); BORD = Border(thin, thin, thin, thin); cen = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "TRECHOS"
    ws.cell(1, 1, "%s — %s · ordem + equipe + datas + mapa" % (cfg["prefixo"], cfg["nucleo"])).font = TIT
    heads = ["Equipe", "Ordem", "NS", "Método", "Comp (m)", "LÍDER", "Início", "Fim", "Mapa do local"]
    for ci, h in enumerate(heads, 1):
        c = ws.cell(2, ci, h); c.font = HFONT; c.fill = HF; c.alignment = cen; c.border = BORD
    lideres = sorted({e["lider"] for e in cfg["equipes"].values()})
    dvl = DataValidation(type="list", formula1='"%s"' % ",".join(lideres), allow_blank=True); ws.add_data_validation(dvl)
    r = 2
    for k, lst in grupos.items():
        eqnome = cfg["equipes"][k]["nome"]
        for t in lst:
            r += 1
            for ci, v in enumerate([eqnome, "%dº" % t["pos"], "%s %02d" % (cfg["prefixo"], t["ns"]), t["metodo"],
                                    round(t["comp"], 1), t["lider"], t["ini"], t["fim"]], 1):
                c = ws.cell(r, ci, v); c.border = BORD; c.alignment = cen
            img = XLImage(_mini(t, cfg, guia, rede, redetr, mdir)); img.width = 235; img.height = 158
            ws.add_image(img, ws.cell(r, 9).coordinate); ws.row_dimensions[r].height = 120
    dvl.add("F3:F%d" % r)
    for ci, w in zip(range(1, 10), [20, 8, 14, 9, 10, 14, 11, 11, 36]):
        ws.column_dimensions[ws.cell(2, ci).column_letter].width = w
    ws.freeze_panes = "A3"
    xls = os.path.join(cfg["saida_dir"], cfg["arq_planilha"])
    try: wb.save(xls)
    except PermissionError:
        xls = xls.replace(".xlsx", "_NOVA.xlsx"); wb.save(xls)
    return xls

def gerar_ns_pdf(grupos, cfg, guia, rede):
    redetr = [t["g"] for lst in grupos.values() for t in lst]
    AZ = cfg["cor"]
    out = os.path.join(cfg["saida_dir"], cfg["arq_ns"])
    try: pdf = PdfPages(out)
    except PermissionError:
        out = out.replace(".pdf", "_NOVA.pdf"); pdf = PdfPages(out)
    for k, lst in grupos.items():
        eq = cfg["equipes"][k]
        for t in lst:
            fig = plt.figure(figsize=(8.27, 11.69)); fig.patch.set_facecolor("white")
            ab = fig.add_axes([0.05, 0.905, 0.90, 0.075]); ab.axis("off")
            ab.add_patch(FancyBboxPatch((0, 0), 1, 1, boxstyle="round,pad=0.01", transform=ab.transAxes, fc=AZ, ec="none"))
            ab.text(0.5, 0.62, "NOTA DE SERVIÇO  Nº %s %02d" % (cfg["prefixo"], t["ns"]), ha="center", va="center", fontsize=15, fontweight="bold", color="white", transform=ab.transAxes)
            ab.text(0.5, 0.22, "WCR Saneamento · Contrato Sabesp 13.546/25-00 · %s" % cfg["nucleo"], ha="center", va="center", fontsize=8.5, color="#eeeeee", transform=ab.transAxes)
            ad = fig.add_axes([0.05, 0.735, 0.90, 0.155]); ad.axis("off")
            ad.add_patch(FancyBboxPatch((0, 0), 1, 1, boxstyle="round,pad=0.005", transform=ad.transAxes, fc="#eef2f6", ec="#b0bec5", lw=1))
            cps = [("EQUIPE", eq["nome"]), ("ORDEM DE EXECUÇÃO", "%dº" % t["pos"]),
                   ("MÉTODO", "MND (não destrutivo)" if t["metodo"] == "MND" else "VCA (vala aberta)"),
                   ("COMPRIMENTO", "%.1f m" % t["comp"]), ("TUBO", t["dn"] or "—"),
                   ("INÍCIO", t["ini"]), ("FIM", t["fim"]), ("STATUS", "PLANEJADA")]
            for j, (lab, val) in enumerate(cps):
                col = j % 2; row = j // 2; x = 0.02 + col * 0.5; y = 0.90 - row * 0.205
                ad.text(x, y, lab, fontsize=6.5, color="#666", fontweight="bold", transform=ad.transAxes)
                ad.text(x, y - 0.090, str(val), fontsize=9.5, color="#111", transform=ad.transAxes)
            am = fig.add_axes([0.08, 0.40, 0.84, 0.31]); _desenha(am, t, guia, rede, redetr, AZ, 120)
            am.set_title("Local — trecho sobre a rede e os lotes", fontsize=9)
            at = fig.add_axes([0.05, 0.02, 0.90, 0.29]); at.axis("off")
            at.text(0, 0.97, "EQUIPE — %s" % eq["nome"], fontsize=10, fontweight="bold", color=AZ); yy = 0.88
            for nome, func in eq.get("crew", [])[:10]:
                at.text(0.02, yy, "•  %s — %s" % (nome, func), fontsize=8.3); yy -= 0.052
            at.text(0.52, 0.97, "MATERIAIS / SERVIÇOS", fontsize=10, fontweight="bold", color="#bf360c"); y2 = 0.88
            for it, q, u in materiais(t, cfg):
                at.text(0.54, y2, "•  %s: %s %s" % (it, q, u), fontsize=8); y2 -= 0.050
            at.text(0, 0.02, "Encarregado: __________   Fiscal Sabesp: __________   Data: ___/___/____", fontsize=8, color="#333")
            pdf.savefig(fig); plt.close(fig)
    pdf.close()
    return out

def atualizar_gpkg(grupos, cfg):
    pa = cfg["gpkg"]; byns = {t["ns"]: t for lst in grupos.values() for t in lst}
    g = gpd.read_file(pa).reset_index(drop=True)
    for col in ("NS_numero", "Equipe_NS", "Data_planejado", "Ordem_exec", "Realizado"):
        if col not in g.columns: g[col] = None
    upd = 0
    for k in range(len(g)):
        t = byns.get(k + 1)
        if not t: continue
        g.loc[k, "NS_numero"] = "%s %02d" % (cfg["prefixo"], t["ns"]); g.loc[k, "Equipe_NS"] = cfg["equipes"][t["team"]]["nome"]
        g.loc[k, "Data_planejado"] = t["ini"]; g.loc[k, "Ordem_exec"] = t["pos"]; upd += 1
    try:
        os.remove(pa); g.to_file(pa, driver="GPKG"); saida = pa
    except Exception:
        saida = pa.replace(".gpkg", "_ATUALIZADO.gpkg"); g.to_file(saida, driver="GPKG")
    return {"saida": saida, "atualizados": upd}

def nota_obsidian(grupos, cfg, ini, fim, resumo):
    if not cfg.get("obsidian_dir"): return None
    os.makedirs(cfg["obsidian_dir"], exist_ok=True)
    nuc = cfg["nucleo"]; sis = cfg["sistema"]
    linhas = [f"---", f"nucleo: {nuc}", f"sistema: {sis}", f"status: planejado",
              f"inicio: {ini.strftime('%Y-%m-%d')}", f"fim: {fim.strftime('%Y-%m-%d')}", f"---", "",
              f"# {nuc} — {sis.upper()} (planejamento)", "",
              f"- Início: **{ini.strftime('%d/%m/%Y')}** · Fim: **{fim.strftime('%d/%m/%Y')}**",
              f"- Trechos: **{sum(len(v) for v in grupos.values())}** · Produtividade: {round(cfg['produtividade'],1)} m/dia", "",
              "## Equipes", ""]
    for k, lst in grupos.items():
        eq = cfg["equipes"][k]
        linhas.append(f"- **{eq['nome']}** — {len(lst)} trechos (NS " + ", ".join("%02d" % t["ns"] for t in lst) + ")")
    linhas += ["", "## Saídas",
               f"- Project: `{cfg['arq_project']}`", f"- Planilha: `{cfg['arq_planilha']}`",
               f"- NS PDF: `{cfg['arq_ns']}`", "", "| Ordem | NS | Equipe | Início | Fim | Realizado |",
               "|---|---|---|---|---|---|"]
    for k, lst in grupos.items():
        for t in lst:
            linhas.append(f"| {t['pos']}º | {cfg['prefixo']} {t['ns']:02d} | {cfg['equipes'][k]['nome']} | {t['ini']} | {t['fim']} |  |")
    p = os.path.join(cfg["obsidian_dir"], "%s_%s.md" % (nrm(nuc).replace(" ", "_"), sis.upper()))
    with open(p, "w", encoding="utf-8") as fh:
        fh.write("\n".join(linhas))
    return p


# ---------------- orquestrador ----------------
def pipeline_nucleo(cfg):
    """Roda o ciclo completo de 1 núcleo a partir do config. Retorna o resumo."""
    mdir = cfg.get("maps_dir") or os.path.join(os.environ.get("TEMP", "."), "maps_pipe_%s" % nrm(cfg["nucleo"]).replace(" ", "_"))
    os.makedirs(mdir, exist_ok=True)
    tr = ler_trechos(cfg)
    grupos = sequenciar(tr, cfg)
    ini, fim = agendar(grupos, cfg)
    guia, rede = carregar_fundo(cfg)
    res = {"nucleo": cfg["nucleo"], "sistema": cfg["sistema"], "trechos": len(tr),
           "inicio": ini.strftime("%d/%m/%Y"), "fim": fim.strftime("%d/%m/%Y"),
           "equipes": {cfg["equipes"][k]["nome"]: len(v) for k, v in grupos.items()}}
    res["project"] = gerar_project(grupos, cfg, ini, fim)["saida"]
    res["planilha"] = gerar_planilha(grupos, cfg, guia, rede, mdir)
    res["ns_pdf"] = gerar_ns_pdf(grupos, cfg, guia, rede)
    res["gpkg"] = atualizar_gpkg(grupos, cfg)["saida"]
    res["nota"] = nota_obsidian(grupos, cfg, ini, fim, res)
    return res
