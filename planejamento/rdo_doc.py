# -*- coding: utf-8 -*-
"""Diário de Obra (RDO) — Excel editável + PDF no formato oficial do construdata
(WCR Saneamento / Contrato Sabesp 13.546/25-00).

Entrada: `dias` = [{data, apontador?, efetivo?, equipes:[], equipamentos:[],
ocorrencias:[], servicos:[{servico, sistema, local, quantidade, equipe, evidencia}]}].
Gera 1 página por dia + 1 página TEMPLATE em branco. O Excel traz DIÁRIO + RESUMO
+ TEMPLATE para preenchimento diário.
"""
from collections import defaultdict, OrderedDict

CONTRATO = "Sabesp 13.546/25-00"
EMPRESA = "WCR Saneamento"
NUCLEO = "BOI MALHADO"
AZUL = "#1a237e"


def _data_br(s):
    s = str(s or "")
    if len(s) == 10 and s[4] == "-":          # 2026-06-12 -> 12/06/2026
        return "%s/%s/%s" % (s[8:10], s[5:7], s[0:4])
    return s


def _wrap(txt, n=95):
    txt = str(txt or ""); out = []
    while len(txt) > n:
        corte = txt.rfind(" ", 0, n)
        if corte <= 0:
            corte = n
        out.append(txt[:corte]); txt = txt[corte:].lstrip()
    out.append(txt); return out


def _por_equipe(dia):
    """Agrupa serviços do dia por equipe (preserva ordem)."""
    grupos = OrderedDict()
    for s in dia.get("servicos", []):
        eq = (s.get("equipe") or "Geral").strip() or "Geral"
        grupos.setdefault(eq, []).append(s)
    return grupos


# ----------------------------------------------------------------- PDF
def gerar_rdo_pdf(dias, caminho, composicao_fn=None, mapa=None):
    """1 página por dia + TEMPLATE. composicao_fn(equipe)->[(nome,func)].
    mapa = {'executado':[geoms], 'a_fazer':[geoms]} → desenha o mapa da rede no topo."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.patches import FancyBboxPatch
    from matplotlib.backends.backend_pdf import PdfPages

    with PdfPages(str(caminho)) as pdf:
        for dia in dias:
            _pagina_dia(plt, FancyBboxPatch, dia, composicao_fn, pdf, mapa)
        _pagina_template(plt, FancyBboxPatch, pdf)
    return str(caminho)


def _coords(geom):
    if geom is None:
        return []
    if geom.geom_type == "LineString":
        return [list(geom.coords)]
    return [list(p.coords) for p in geom.geoms]


def _desenha_mapa(ax, mapa):
    ax.set_title("Mapa — rede executada (verde) × a-fazer (cinza)", fontsize=9, color=AZUL)
    for g in mapa.get("a_fazer", []):
        for cc in _coords(g):
            ax.plot([p[0] for p in cc], [p[1] for p in cc], color="0.78", lw=0.8)
    for g in mapa.get("executado", []):
        for cc in _coords(g):
            ax.plot([p[0] for p in cc], [p[1] for p in cc], color="#2e7d32", lw=1.7)
    ax.set_aspect("equal"); ax.set_xticks([]); ax.set_yticks([])


def _cabecalho(plt, FancyBboxPatch, fig, data_txt, apontador):
    ax = fig.add_axes([0.06, 0.93, 0.88, 0.06]); ax.axis("off")
    ax.add_patch(FancyBboxPatch((0, 0), 1, 1, boxstyle="round,pad=0.01",
                                transform=ax.transAxes, fc=AZUL, ec="none"))
    ax.text(0.5, 0.66, "DIÁRIO DE OBRA (RDO)", ha="center", va="center",
            fontsize=15, fontweight="bold", color="white", transform=ax.transAxes)
    ax.text(0.5, 0.24, "Núcleo %s   ·   Data %s   ·   Apontador: %s   ·   %s — %s"
            % (NUCLEO, data_txt, apontador or "—", CONTRATO, EMPRESA),
            ha="center", va="center", fontsize=8, color="#bbdefb", transform=ax.transAxes)


def _pagina_dia(plt, FancyBboxPatch, dia, composicao_fn, pdf, mapa=None):
    fig = plt.figure(figsize=(8.27, 11.69)); fig.patch.set_facecolor("white")
    _cabecalho(plt, FancyBboxPatch, fig, _data_br(dia.get("data")), dia.get("apontador", "Apontamento WCR"))
    if mapa:
        axm = fig.add_axes([0.06, 0.585, 0.88, 0.315])
        _desenha_mapa(axm, mapa)
        ax = fig.add_axes([0.06, 0.05, 0.88, 0.50]); ax.axis("off")
    else:
        ax = fig.add_axes([0.06, 0.05, 0.88, 0.86]); ax.axis("off")
    y = 1.0
    grupos = _por_equipe(dia)
    for eq, servs in grupos.items():
        comp = composicao_fn(eq) if composicao_fn else []
        ax.text(0.0, y, "Equipe %s" % eq, fontsize=10.5, fontweight="bold", color=AZUL); y -= 0.024
        if comp:
            comp_txt = ", ".join("%s (%s)" % (p, f) for p, f in comp)
            for ln in _wrap("Composição: " + comp_txt, 100):
                ax.text(0.02, y, ln, fontsize=7.8, color="#333"); y -= 0.0185
        ax.text(0.02, y, "Serviços executados:", fontsize=8.5, fontweight="bold"); y -= 0.020
        for s in servs:
            q = (" — " + str(s["quantidade"])) if s.get("quantidade") else ""
            loc = (" [" + s["local"] + "]") if s.get("local") else ""
            for j, ln in enumerate(_wrap(s.get("servico", "") + q + loc, 95)):
                ax.text(0.04, y, ("• " if j == 0 else "  ") + ln, fontsize=8, color="#111"); y -= 0.0185
        y -= 0.012
        if y < 0.12:
            break
    # ocorrências
    occ = dia.get("ocorrencias") or []
    if occ:
        ax.text(0.0, y, "Ocorrências:", fontsize=9, fontweight="bold", color="#b71c1c"); y -= 0.020
        for o in occ:
            for ln in _wrap("— " + str(o), 100):
                ax.text(0.02, y, ln, fontsize=8, color="#b71c1c"); y -= 0.0185
    # rodapé assinaturas
    ax.text(0.0, 0.02, "Responsável técnico: ___________________________     "
            "Encarregado: ___________________________", fontsize=8, color="#333")
    pdf.savefig(fig); plt.close(fig)


def _pagina_template(plt, FancyBboxPatch, pdf):
    fig = plt.figure(figsize=(8.27, 11.69)); fig.patch.set_facecolor("white")
    _cabecalho(plt, FancyBboxPatch, fig, "____/____/______", "____________________")
    ax = fig.add_axes([0.06, 0.05, 0.88, 0.86]); ax.axis("off")
    y = 1.0
    ax.text(0.0, y, "TEMPLATE — preencher em campo (1 por dia)", fontsize=10,
            fontweight="bold", color="#666"); y -= 0.04
    for _ in range(4):
        ax.text(0.0, y, "Equipe: ____________________   Líder: ____________________", fontsize=9); y -= 0.028
        ax.text(0.02, y, "Composição: ________________________________________________", fontsize=8.5); y -= 0.026
        ax.text(0.02, y, "Serviços executados:", fontsize=8.5, fontweight="bold"); y -= 0.022
        for _ in range(3):
            ax.text(0.04, y, "•  ____________________________________________________________", fontsize=8); y -= 0.022
        ax.text(0.02, y, "Frota/Equipamentos: ________________________________________", fontsize=8.5); y -= 0.034
    ax.text(0.0, y, "Ocorrências: ___________________________________________________", fontsize=9); y -= 0.030
    ax.text(0.0, 0.02, "Responsável técnico: ___________________________     "
            "Encarregado: ___________________________", fontsize=8, color="#333")
    pdf.savefig(fig); plt.close(fig)


# ----------------------------------------------------------------- Excel
def gerar_rdo_xlsx(dias, caminho):
    """Workbook: DIÁRIO (serviços/dia/equipe) + RESUMO (quantidades) + TEMPLATE."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from .diario import resumo_quantidades

    H_FILL = PatternFill("solid", fgColor="1F4E79"); H_FONT = Font(bold=True, color="FFFFFF")
    TIT = Font(bold=True, size=13, color="1F4E79")
    thin = Side(style="thin", color="BBBBBB"); BORD = Border(thin, thin, thin, thin)

    def cab(ws, titulo, cols):
        ws.cell(1, 1, titulo).font = TIT
        for ci, h in enumerate(cols, 1):
            c = ws.cell(2, ci, h); c.font = H_FONT; c.fill = H_FILL
            c.alignment = Alignment(horizontal="center"); c.border = BORD

    wb = Workbook()
    # DIÁRIO
    ws = wb.active; ws.title = "DIÁRIO"
    cab(ws, "DIÁRIO DE OBRA — %s (%s)" % (NUCLEO, CONTRATO),
        ["Data", "Equipe", "Sistema", "Serviço", "Local", "Quantidade"])
    r = 3
    for dia in dias:
        for s in dia.get("servicos", []):
            for ci, v in enumerate([_data_br(dia.get("data")), s.get("equipe", ""), s.get("sistema", ""),
                                    s.get("servico", ""), s.get("local", ""), s.get("quantidade", "")], 1):
                cell = ws.cell(r, ci, v); cell.border = BORD
            r += 1
    for ci, w in enumerate([12, 16, 10, 60, 22, 16], 1):
        ws.column_dimensions[ws.cell(2, ci).column_letter].width = w
    ws.freeze_panes = "A3"
    # RESUMO
    ws2 = wb.create_sheet("RESUMO")
    cab(ws2, "RESUMO DE QUANTIDADES EXECUTADAS", ["Item executado", "Quantidade", "Und."])
    for i, row in enumerate(resumo_quantidades(dias)):
        for ci, v in enumerate([row["item"], row["qtd"], row["und"]], 1):
            ws2.cell(3 + i, ci, v).border = BORD
    for ci, w in enumerate([34, 14, 8], 1):
        ws2.column_dimensions[ws2.cell(2, ci).column_letter].width = w
    # TEMPLATE
    ws3 = wb.create_sheet("TEMPLATE")
    cab(ws3, "TEMPLATE RDO — preencher em campo", ["Data", "Equipe", "Líder", "Serviço", "Local", "Quantidade"])
    for ci, w in enumerate([12, 16, 14, 60, 22, 16], 1):
        ws3.column_dimensions[ws3.cell(2, ci).column_letter].width = w
    wb.save(str(caminho))
    return {"saida": str(caminho), "dias": len(dias),
            "servicos": sum(len(d.get("servicos", [])) for d in dias)}
