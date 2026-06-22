# -*- coding: utf-8 -*-
"""Documento de Nota de Serviço CONSOLIDADO (junção dos motores V5 + planejamento).

Cada NS reúne, numa folha A4: cabeçalho SABESP/WCR (carimbo), dados do trecho,
desenho do trecho, cronograma (datas/dias úteis), composição da equipe (RDO) e a
BOM rica (materiais/serviços). Saída em PDF (campo) e Excel master (editável).

`ns_list`: [{ns, sistema, nucleo, equipe, lider, servico, rua, dn, comp, produt,
inicio, fim, dias, geom, materiais}]  — materiais = [(item, qtd, und)].
`contexto`: {sistema: [geometrias de fundo]} p/ situar o trecho no desenho.
"""
from datetime import date, timedelta

CONTRATO = "Sabesp 13.546/25-00"
EMPRESA = "WCR Saneamento"
AZUL = "#1a237e"
AZUL_CLARO = "#90caf9"


def _coords(geom):
    if geom is None:
        return []
    if geom.geom_type == "LineString":
        return [list(geom.coords)]
    return [list(g.coords) for g in geom.geoms]


def _fmt(d):
    return d.strftime("%d/%m/%Y") if hasattr(d, "strftime") else (str(d) if d else "—")


def dias_uteis(inicio, fim, feriados=frozenset()):
    """Conta dias úteis (seg-sex, fora feriados) de início a fim, inclusive."""
    if not (hasattr(inicio, "weekday") and hasattr(fim, "weekday")):
        return None
    n, d = 0, inicio
    while d <= fim:
        if d.weekday() < 5 and d not in feriados:
            n += 1
        d += timedelta(days=1)
    return n


# ----------------------------------------------------------------- PDF (campo)
def gerar_pdf_ns(ns_list, contexto, caminho, feriados=frozenset()):
    """Escreve o PDF multipágina (1 NS por página). Retorna o caminho."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.patches import FancyBboxPatch, Rectangle
    from matplotlib.backends.backend_pdf import PdfPages

    with PdfPages(str(caminho)) as pdf:
        for ns in ns_list:
            _pagina_ns(plt, FancyBboxPatch, Rectangle, ns, contexto, feriados, pdf.savefig)
    return str(caminho)


def _pagina_ns(plt, FancyBboxPatch, Rectangle, ns, contexto, feriados, sink):
    sis = ns.get("sistema", "")
    fig = plt.figure(figsize=(8.27, 11.69))            # A4 retrato
    fig.patch.set_facecolor("white")

    # --- faixa de título (carimbo SABESP/WCR) ---
    axb = fig.add_axes([0.05, 0.905, 0.90, 0.075]); axb.axis("off")
    axb.add_patch(FancyBboxPatch((0, 0), 1, 1, boxstyle="round,pad=0.01",
                                 transform=axb.transAxes, fc=AZUL, ec="none"))
    axb.text(0.5, 0.62, "NOTA DE SERVIÇO  Nº %s  —  %s" % (ns["ns"], sis),
             ha="center", va="center", fontsize=16, fontweight="bold", color="white",
             transform=axb.transAxes)
    axb.text(0.5, 0.22, "%s · Contrato %s · Núcleo %s" % (EMPRESA, CONTRATO, ns.get("nucleo", "")),
             ha="center", va="center", fontsize=8.5, color=AZUL_CLARO, transform=axb.transAxes)

    # --- bloco de dados ---
    axd = fig.add_axes([0.05, 0.735, 0.90, 0.155]); axd.axis("off")
    d = dias_uteis(ns.get("inicio"), ns.get("fim"), feriados)
    campos = [
        ("FRENTE/EQUIPE", "%s%s" % (ns.get("equipe", ""),
                                    ("  (líder: %s)" % ns["lider"]) if ns.get("lider") else "")),
        ("SERVIÇO", ns.get("servico", "")),
        ("RUA", ns.get("rua") or "—"),
        ("DN", "%s mm" % ns.get("dn", "—")),
        ("EXTENSÃO", "%.1f m" % ns.get("comp", 0)),
        ("PRODUTIVIDADE MÍN.", ns.get("produt", "—")),
        ("INÍCIO", _fmt(ns.get("inicio"))),
        ("FIM", _fmt(ns.get("fim"))),
        ("DIAS ÚTEIS", "%s dia(s)" % d if d else "—"),
        ("STATUS", ns.get("status", "PLANEJADO")),
    ]
    axd.add_patch(FancyBboxPatch((0, 0), 1, 1, boxstyle="round,pad=0.005",
                                 transform=axd.transAxes, fc="#f5f7fb", ec="#c5cae9", lw=1.0))
    ncol = 2
    for i, (lab, val) in enumerate(campos):
        col = i % ncol; row = i // ncol
        x = 0.02 + col * 0.5
        y = 0.90 - row * 0.205
        axd.text(x, y, lab, fontsize=6.5, color="#666", fontweight="bold", transform=axd.transAxes)
        axd.text(x, y - 0.090, str(val), fontsize=9.5, color="#111", transform=axd.transAxes)

    # --- desenho do trecho ---
    axm = fig.add_axes([0.08, 0.40, 0.84, 0.31])
    ctx = contexto.get(sis, []) if isinstance(contexto, dict) else (contexto or [])
    for g in ctx:
        for cc in _coords(g):
            axm.plot([p[0] for p in cc], [p[1] for p in cc], color="0.82", lw=0.7, zorder=1)
    allx, ally = [], []
    for cc in _coords(ns.get("geom")):
        xs = [p[0] for p in cc]; ys = [p[1] for p in cc]
        axm.plot(xs, ys, color="#d32f2f", lw=3.2, solid_capstyle="round", zorder=3)
        axm.plot(xs[0], ys[0], "o", color="#2e7d32", ms=8, zorder=4)
        axm.plot(xs[-1], ys[-1], "s", color="#1565c0", ms=8, zorder=4)
        allx += xs; ally += ys
    if allx:
        cx = (max(allx) + min(allx)) / 2; cy = (max(ally) + min(ally)) / 2
        mg = max(max(allx) - min(allx), max(ally) - min(ally), 40) * 0.65 + 25
        axm.set_xlim(cx - mg, cx + mg); axm.set_ylim(cy - mg, cy + mg)
    axm.set_aspect("equal"); axm.set_xticks([]); axm.set_yticks([])
    axm.set_title("Desenho do trecho   (vermelho = a executar · ● início · ■ fim)", fontsize=9.5)
    axm.annotate("N", xy=(0.96, 0.92), xytext=(0.96, 0.78), xycoords="axes fraction",
                 ha="center", fontsize=11, fontweight="bold",
                 arrowprops=dict(arrowstyle="->", lw=1.4))

    # --- cronograma (mini timeline início->fim) ---
    axc = fig.add_axes([0.08, 0.325, 0.84, 0.045]); axc.axis("off")
    axc.text(0, 0.6, "CRONOGRAMA:", fontsize=8.5, fontweight="bold", color=AZUL, transform=axc.transAxes)
    axc.add_patch(Rectangle((0.16, 0.30), 0.80, 0.34, transform=axc.transAxes,
                            fc="#e8eaf6", ec="#9fa8da", lw=0.8))
    axc.add_patch(Rectangle((0.16, 0.30), 0.80, 0.34, transform=axc.transAxes,
                            fc="#3949ab", ec="none", alpha=0.85))
    axc.text(0.17, 0.47, _fmt(ns.get("inicio")), fontsize=7.5, color="white",
             va="center", transform=axc.transAxes)
    axc.text(0.95, 0.47, _fmt(ns.get("fim")), fontsize=7.5, color="white",
             va="center", ha="right", transform=axc.transAxes)

    # --- equipe + materiais ---
    axt = fig.add_axes([0.05, 0.02, 0.90, 0.29]); axt.axis("off")
    y = 0.97
    axt.text(0.0, y, "EQUIPE  —  %s%s" % (ns.get("equipe", ""),
             ("  (líder: %s)" % ns["lider"]) if ns.get("lider") else ""),
             fontsize=10, fontweight="bold", color=AZUL)
    y -= 0.085
    for pessoa, func in ns.get("composicao", []):
        axt.text(0.02, y, "•  %s — %s" % (pessoa, func), fontsize=8.5); y -= 0.058
    y2 = 0.97
    axt.text(0.52, y2, "MATERIAIS / SERVIÇOS", fontsize=10, fontweight="bold", color="#ef6c00")
    y2 -= 0.085
    for item, q, u in ns.get("materiais", []):
        axt.text(0.54, y2, "•  %s:  %s %s" % (item, q, u), fontsize=8); y2 -= 0.050
    axt.text(0.0, 0.02, "Encarregado: ____________________      Fiscal Sabesp: ____________________      "
             "Data: ___/___/______", fontsize=8, color="#333")

    sink(fig); plt.close(fig)


# ----------------------------------------------------------------- Excel master
COLS_NOTAS = [
    {"campo": "ns", "titulo": "NS"}, {"campo": "sistema", "titulo": "Sistema"},
    {"campo": "nucleo", "titulo": "Núcleo"}, {"campo": "equipe", "titulo": "Frente/Equipe"},
    {"campo": "lider", "titulo": "Líder"}, {"campo": "servico", "titulo": "Serviço"},
    {"campo": "rua", "titulo": "Rua"}, {"campo": "dn", "titulo": "DN (mm)"},
    {"campo": "comp", "titulo": "Comp (m)"}, {"campo": "produt", "titulo": "Produt. mín."},
    {"campo": "inicio", "titulo": "Início"}, {"campo": "fim", "titulo": "Fim"},
    {"campo": "dias", "titulo": "Dias úteis"}, {"campo": "status", "titulo": "Status"},
]
COLS_MAT = [{"campo": "ns", "titulo": "NS"}, {"campo": "sistema", "titulo": "Sistema"},
            {"campo": "item", "titulo": "Material/Serviço"},
            {"campo": "quant", "titulo": "Quant."}, {"campo": "und", "titulo": "Und."}]
COLS_EQ = [{"campo": "equipe", "titulo": "Frente/Equipe"}, {"campo": "lider", "titulo": "Líder"},
           {"campo": "pessoa", "titulo": "Integrante"}, {"campo": "funcao", "titulo": "Função"}]
COLS_CRON = [{"campo": "equipe", "titulo": "Frente/Equipe"}, {"campo": "ns", "titulo": "NS"},
             {"campo": "rua", "titulo": "Rua"}, {"campo": "comp", "titulo": "Comp (m)"},
             {"campo": "inicio", "titulo": "Início"}, {"campo": "fim", "titulo": "Fim"}]
COLS_RES = [{"campo": "item", "titulo": "Indicador"}, {"campo": "valor", "titulo": "Valor"}]


def _fmtd(d):
    return d.strftime("%d/%m/%Y") if hasattr(d, "strftime") else (d or "")


def gerar_excel_ns(ns_list, equipes_dict, lider_dict, caminho, feriados=frozenset()):
    """Workbook master editável: NOTAS, MATERIAIS, EQUIPES, CRONOGRAMA, RESUMO.
    Aplica formatação (cabeçalho, larguras, freeze)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    H_FILL = PatternFill("solid", fgColor="1F4E79"); H_FONT = Font(bold=True, color="FFFFFF")
    TIT = Font(bold=True, size=13, color="1F4E79")
    thin = Side(style="thin", color="BBBBBB"); BORD = Border(thin, thin, thin, thin)

    notas, mats, crono = [], [], []
    for ns in ns_list:
        du = dias_uteis(ns.get("inicio"), ns.get("fim"), feriados)
        notas.append({**ns, "inicio": _fmtd(ns.get("inicio")), "fim": _fmtd(ns.get("fim")),
                      "dias": du, "status": ns.get("status", "PLANEJADO")})
        for item, q, u in ns.get("materiais", []):
            mats.append({"ns": ns["ns"], "sistema": ns["sistema"], "item": item, "quant": q, "und": u})
        crono.append({"equipe": ns.get("equipe"), "ns": ns["ns"], "rua": ns.get("rua") or "—",
                      "comp": ns.get("comp"), "inicio": _fmtd(ns.get("inicio")), "fim": _fmtd(ns.get("fim"))})
    crono.sort(key=lambda r: (str(r["equipe"]), str(r["inicio"]), r["ns"]))

    eq_rows = []
    for rot, comp in equipes_dict.items():
        for pessoa, func in comp:
            eq_rows.append({"equipe": rot, "lider": lider_dict.get(rot, ""), "pessoa": pessoa, "funcao": func})

    # resumo
    esg = [n for n in ns_list if n["sistema"] == "ESGOTO"]
    agu = [n for n in ns_list if n["sistema"] != "ESGOTO"]
    fim_max = max((n["fim"] for n in ns_list if hasattr(n.get("fim"), "strftime")), default=None)
    resumo = [
        {"item": "Total de NS", "valor": len(ns_list)},
        {"item": "NS de esgoto", "valor": len(esg)},
        {"item": "NS de água", "valor": len(agu)},
        {"item": "Extensão esgoto (m)", "valor": round(sum(n.get("comp", 0) for n in esg), 1)},
        {"item": "Extensão água (m)", "valor": round(sum(n.get("comp", 0) for n in agu), 1)},
        {"item": "Início da obra", "valor": _fmtd(min((n["inicio"] for n in ns_list if hasattr(n.get("inicio"), "strftime")), default=""))},
        {"item": "Término previsto", "valor": _fmtd(fim_max)},
        {"item": "Meta de término", "valor": "30/09/2026"},
        {"item": "Contrato", "valor": CONTRATO},
    ]

    abas = [
        ("NOTAS DE SERVIÇO", COLS_NOTAS, notas, "NOTAS DE SERVIÇO — BOI MALHADO (continuidade · meta 30/09/2026)"),
        ("MATERIAIS", COLS_MAT, mats, "MATERIAIS / SERVIÇOS POR NS"),
        ("EQUIPES", COLS_EQ, eq_rows, "COMPOSIÇÃO DAS EQUIPES (RDO oficial)"),
        ("CRONOGRAMA", COLS_CRON, crono, "CRONOGRAMA POR FRENTE"),
        ("RESUMO", COLS_RES, resumo, "RESUMO DA OBRA"),
    ]
    wb = Workbook(); first = True
    for nome, cols, linhas, titulo in abas:
        ws = wb.active if first else wb.create_sheet(); first = False
        ws.title = nome[:31]
        ws.cell(1, 1, titulo).font = TIT
        for ci, c in enumerate(cols, 1):
            cell = ws.cell(2, ci, c["titulo"]); cell.font = H_FONT; cell.fill = H_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = BORD
        for i, linha in enumerate(linhas):
            for ci, c in enumerate(cols, 1):
                cell = ws.cell(3 + i, ci, linha.get(c["campo"])); cell.border = BORD
        for ci, c in enumerate(cols, 1):
            w = max(len(str(c["titulo"])), *(len(str(l.get(c["campo"], ""))) for l in linhas)) if linhas else len(str(c["titulo"]))
            ws.column_dimensions[ws.cell(2, ci).column_letter].width = min(max(w + 2, 9), 48)
        ws.freeze_panes = "A3"
    wb.save(str(caminho))
    return {"abas": [a[0] for a in abas], "ns": len(ns_list), "saida": str(caminho)}
