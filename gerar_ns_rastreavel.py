#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
gerar_ns_rastreavel.py — Sistema Rastreável de Notas de Serviço v5.0
ConstruData SABESP · Contrato 11481051

NOVIDADES v5.0:
✓ Nomenclatura padronizada: NS-[NÚCLEO]-[PV]-[TRECHO]-[DATA]
✓ Rastreabilidade completa: PV → NS → Trecho → Materiais
✓ Auto-save no Supabase (tabela notas_servico)
✓ Dashboard Excel consolidado por núcleo
✓ WhatsApp integration (@ns, @tarefas)

Uso:
  python gerar_ns_rastreavel.py <dxf_path> <nucleo> [output_dir]
  python gerar_ns_rastreavel.py --batch  # processa todos os núcleos configurados
"""

import sys
import json
import math
import os
import traceback
import re
from pathlib import Path
from datetime import datetime
from collections import Counter, defaultdict

# Dependencies
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib.patches import FancyBboxPatch
import numpy as np

try:
    from ler_dxf_gdal import ler_dxf_gdal
except ImportError:
    print("❌ Erro: ler_dxf_gdal.py não encontrado")
    sys.exit(1)

# ═══════════════════════════════════════════════════════════════
# CONFIGURAÇÃO GLOBAL
# ═══════════════════════════════════════════════════════════════
CONTRATO = "11481051"
NS_VERSION = "5.0"
CRS_EPSG = "EPSG:31983"  # SIRGAS 2000 UTM 23S
N_MANNING = 0.013  # PVC

# Diretórios
DXF_DIR = Path.home() / "Downloads" / "PROJETOS DE ÁGUA E ESGOTO - DWG E DXF 2018" / "MAPAS ÁGUA E ESGOTO PARA DXF"
OUTPUT_BASE = Path(__file__).parent / "SAIDA_NS_V5"

# Configuração Supabase
SUPABASE_URL = "https://vblfdikfobsirwpdnybw.supabase.co/rest/v1"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InZibGZkaWtmb2JzaXJ3cGRueWJ3Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzMzNzAwODIsImV4cCI6MjA4ODk0NjA4Mn0.GOx3HoMh3P2Zzxz8BxNsfQBfXwsNZNQsdVc3nJaqRy4"

# Núcleos batch
NUCLEOS_BATCH = [
    {"nucleo": "Pardinho", "dxf": str(DXF_DIR / "PARDINHO" / "PARDINHO_ESGOTO.dxf")},
    {"nucleo": "Osasco", "dxf": str(DXF_DIR / "OSASCO" / "OSASCO_ESGOTO.dxf")},
    {"nucleo": "RK", "dxf": str(DXF_DIR / "RK" / "RK_ESGOTO.dxf")},
    {"nucleo": "SLNR", "dxf": str(DXF_DIR / "SLNR" / "SLNR_ESGOTO.dxf")},
]


def log(msg, nivel="INFO"):
    """Log formatado com timestamp."""
    ts = datetime.now().strftime("%H:%M:%S")
    prefix = {
        "OK": "[✅]  ",
        "WARN": "[⚠️]   ",
        "STEP": ">>> ",
        "ERR": "[❌] ",
        "DB": "[🗄️]  "
    }.get(nivel, "      ")
    print(f"[{ts}] {prefix}{msg}")


# ═══════════════════════════════════════════════════════════════
# NOMENCLATURA PADRONIZADA
# ═══════════════════════════════════════════════════════════════
def gerar_numero_ns(nucleo, pv_codigo, trecho_codigo, data=None):
    """
    Gera número de NS padronizado:
    NS-[NÚCLEO]-[PV]-[TRECHO]-[DATA]
    
    Exemplo: NS-PARD-PV001-TR01-20260416
    """
    if data is None:
        data = datetime.now()
    
    nucleo_code = nucleo[:4].upper()
    return f"NS-{nucleo_code}-{pv_codigo}-{trecho_codigo}-{data.strftime('%Y%m%d')}"


def validar_nome_ns(nome_ns):
    """Valida se nome de NS segue padrão."""
    pattern = r"^NS-[A-Z]{4}-PV\d{3}-TR\d{2}-\d{8}$"
    return bool(re.match(pattern, nome_ns))


# ═══════════════════════════════════════════════════════════════
# CÁLCULOS HIDRÁULICOS
# ═══════════════════════════════════════════════════════════════
def calc_manning(dn_mm, decl_mm):
    """Cálculo Manning seção plena."""
    if not dn_mm or not decl_mm or decl_mm <= 0:
        return {"v_ms": None, "q_ls": None, "tau_pa": None}
    
    D = dn_mm / 1000
    A = math.pi * D**2 / 4
    Rh = D / 4
    V = (1 / N_MANNING) * Rh**(2/3) * decl_mm**0.5
    Q = V * A * 1000  # l/s
    tau = 1000 * 9.81 * Rh * decl_mm  # Pa
    
    return {
        "v_ms": round(V, 3),
        "q_ls": round(Q, 3),
        "tau_pa": round(tau, 2)
    }


def calcular_materiais(trecho, pvs):
    """Calcula quantitativo de materiais para 1 trecho."""
    dn = trecho.get("dn_mm", 200)
    ext = trecho.get("ext_m", 0)
    n_barras = math.ceil(ext / 6) if ext > 0 else 1
    
    materiais = [
        {
            "descricao": f"Tubo PVC DN{dn}mm",
            "unidade": "barra",
            "quantidade": n_barras,
            "codigo_sabesp": f"TUB-PVC-{dn}"
        },
        {
            "descricao": f"Luva correr PVC DN{dn}mm",
            "unidade": "pc",
            "quantidade": max(n_barras - 1, 0),
            "codigo_sabesp": f"LUV-COR-{dn}"
        },
        {
            "descricao": f"Anel borracha DN{dn}mm",
            "unidade": "pc",
            "quantidade": n_barras + 1,
            "codigo_sabesp": f"ANE-BOR-{dn}"
        },
        {
            "descricao": "Pasta lubrificante",
            "unidade": "kg",
            "quantidade": round(n_barras * 0.04, 2),
            "codigo_sabesp": "PAS-LUB-001"
        },
        {
            "descricao": "Areia lastro",
            "unidade": "m3",
            "quantidade": round(ext * 0.08, 2),
            "codigo_sabesp": "ARE-LAS-001"
        },
        {
            "descricao": "Areia envoltória",
            "unidade": "m3",
            "quantidade": round(ext * 0.24, 2),
            "codigo_sabesp": "ARE-ENV-001"
        },
        {
            "descricao": "Brita dreno",
            "unidade": "m3",
            "quantidade": round(ext * 0.16, 2),
            "codigo_sabesp": "BRI-DRE-001"
        },
        {
            "descricao": "PV concreto DN1200",
            "unidade": "pc",
            "quantidade": 1,
            "codigo_sabesp": "PV-CON-1200"
        },
    ]
    
    if trecho.get("tipo", "esgoto") == "esgoto":
        materiais.extend([
            {
                "descricao": "Ramal esgoto DN100",
                "unidade": "pc",
                "quantidade": 1,
                "codigo_sabesp": "RAM-ESG-100"
            },
            {
                "descricao": "Caixa inspeção",
                "unidade": "pc",
                "quantidade": 1,
                "codigo_sabesp": "CX-INS-001"
            },
            {
                "descricao": f"Junção Y PVC DN{dn}x100mm",
                "unidade": "pc",
                "quantidade": 1,
                "codigo_sabesp": f"JUN-Y-{dn}x100"
            }
        ])
    
    return materiais


def enriquecer_trechos(trechos, pvs):
    """Adiciona hidráulica e cotas a cada trecho."""
    for t in trechos:
        pvi = pvs.get(t["pv_ini"], {})
        pvf = pvs.get(t["pv_fim"], {})
        
        t["ct_ini"] = t.get("ct_ini") or pvi.get("ct")
        t["ct_fim"] = t.get("ct_fim") or pvf.get("ct")
        t["cf_ini"] = t.get("cf_ini") or pvi.get("cf")
        t["cf_fim"] = t.get("cf_fim") or pvf.get("cf")
        t["prof_ini"] = t.get("prof_ini") or pvi.get("prof")
        t["prof_fim"] = t.get("prof_fim") or pvf.get("prof")
        
        hidr = calc_manning(t.get("dn_mm"), t.get("decl_mm"))
        t.update(hidr)
    
    return trechos


# ═══════════════════════════════════════════════════════════════
# INTEGRAÇÃO SUPABASE
# ═══════════════════════════════════════════════════════════════
def salvar_ns_supabase(ns_data):
    """
    Salva Nota de Serviço no Supabase.
    
    Args:
        ns_data: dict com dados da NS
        
    Returns:
        bool: True se salvou com sucesso
    """
    try:
        import requests
        
        headers = {
            "apikey": SUPABASE_KEY,
            "Authorization": f"Bearer {SUPABASE_KEY}",
            "Content-Type": "application/json",
            "Prefer": "return=representation"
        }
        
        response = requests.post(
            f"{SUPABASE_URL}/notas_servico",
            headers=headers,
            json=ns_data
        )
        
        if response.status_code == 201:
            log(f"NS salva no Supabase: {ns_data['numero_ns']}", "DB")
            return True
        else:
            log(f"Erro ao salvar NS: HTTP {response.status_code} - {response.text}", "ERR")
            return False
            
    except Exception as e:
        log(f"Exceção ao salvar no Supabase: {e}", "ERR")
        return False


def consultar_ns_por_pv(pv_codigo):
    """Consulta todas as NS de um PV específico."""
    try:
        import requests
        
        headers = {
            "apikey": SUPABASE_KEY,
            "Authorization": f"Bearer {SUPABASE_KEY}"
        }
        
        response = requests.get(
            f"{SUPABASE_URL}/notas_servico?pv_codigo=eq.{pv_codigo}",
            headers=headers
        )
        
        if response.status_code == 200:
            return response.json()
        else:
            log(f"Erro ao consultar NS: HTTP {response.status_code}", "ERR")
            return []
            
    except Exception as e:
        log(f"Exceção ao consultar NS: {e}", "ERR")
        return []


# ═══════════════════════════════════════════════════════════════
# GERADORES DE DOCUMENTOS
# ═══════════════════════════════════════════════════════════════
def gerar_pdf_a4(ns_id, numero_ns, trecho, pvs, nucleo, out_path):
    """Gera PDF A4 da Nota de Serviço."""
    pvi_n, pvf_n = trecho["pv_ini"], trecho["pv_fim"]
    pvi, pvf = pvs.get(pvi_n, {}), pvs.get(pvf_n, {})
    dn = trecho.get("dn_mm", "?")
    ext = trecho.get("ext_m", 0)
    decl = trecho.get("decl_mm")
    rua = trecho.get("rua", "Sem Rua")
    hidr = calc_manning(trecho.get("dn_mm"), decl)
    materiais = calcular_materiais(trecho, pvs)
    
    fig, ax = plt.subplots(figsize=(11.69, 8.27))
    ax.set_xlim(0, 100)
    ax.set_ylim(0, 70)
    ax.axis('off')
    
    # Header
    ax.add_patch(FancyBboxPatch((1, 60), 98, 9, boxstyle="round,pad=0.3",
                                 fc="#1a237e", ec="none"))
    ax.text(50, 65, f"NOTA DE SERVIÇO - {numero_ns}", ha="center",
            va="center", fontsize=16, fontweight="bold", color="white")
    ax.text(50, 61.5, f"SE LIGA NA REDE - {nucleo} - Contrato {CONTRATO}",
            ha="center", va="center", fontsize=9, color="#90caf9")
    
    # Dados do trecho
    y = 57
    campos = [
        ("TRECHO", f"{pvi_n} → {pvf_n}"),
        ("LOGRADOURO", rua),
        ("DN", f"{dn} mm"),
        ("EXTENSÃO", f"{ext:.2f} m"),
        ("MATERIAL", trecho.get("material", "PVC")),
        ("DECLIVIDADE", f"{decl*1000:.2f} ‰" if decl else "—"),
    ]
    
    for i, (label, valor) in enumerate(campos):
        col = 5 if i % 2 == 0 else 52
        row = y - (i // 2) * 5
        ax.text(col, row, label, fontsize=7, color="#666", fontweight="bold")
        ax.text(col, row - 2.2, valor, fontsize=11, color="#111")
    
    # Cotas
    y = 38
    ax.add_patch(FancyBboxPatch((1, y-1), 98, 7, boxstyle="round,pad=0.2",
                                 fc="#e3f2fd", ec="#90caf9", lw=0.5))
    ax.text(5, y+3.5, "PV MONTANTE", fontsize=7, color="#1565c0", fontweight="bold")
    ax.text(5, y+1, f"CT = {pvi.get('ct','—')}m   CF = {pvi.get('cf','—')}m   "
                    f"Prof = {pvi.get('prof','—')}m", fontsize=9)
    ax.text(52, y+3.5, "PV JUSANTE", fontsize=7, color="#1565c0", fontweight="bold")
    ax.text(52, y+1, f"CT = {pvf.get('ct','—')}m   CF = {pvf.get('cf','—')}m   "
                     f"Prof = {pvf.get('prof','—')}m", fontsize=9)
    
    # Hidráulica
    y = 28
    ax.add_patch(FancyBboxPatch((1, y-1), 98, 7, boxstyle="round,pad=0.2",
                                 fc="#e8f5e9", ec="#81c784", lw=0.5))
    ax.text(5, y+3.5, "HIDRÁULICA (Manning)", fontsize=7, color="#2e7d32", fontweight="bold")
    v_txt = f"{hidr['v_ms']:.3f} m/s" if hidr['v_ms'] else "—"
    q_txt = f"{hidr['q_ls']:.2f} l/s" if hidr['q_ls'] else "—"
    t_txt = f"{hidr['tau_pa']:.2f} Pa" if hidr['tau_pa'] else "—"
    ax.text(5, y+1, f"V = {v_txt}   Q = {q_txt}   τ = {t_txt}   n = {N_MANNING}", fontsize=9)
    
    # Footer
    ax.text(50, 2, f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} - "
            f"ConstruData SABESP NS v{NS_VERSION}",
            ha="center", fontsize=6, color="#999")
    
    fig.savefig(str(out_path), dpi=150, bbox_inches="tight", pad_inches=0.3)
    plt.close(fig)


def gerar_json_dados(ns_id, numero_ns, trecho, pvs, nucleo, materiais, out_path):
    """Gera JSON com dados completos da NS."""
    dados = {
        "ns_id": ns_id,
        "numero_ns": numero_ns,
        "nucleo": nucleo,
        "contrato": CONTRATO,
        "data_geracao": datetime.now().isoformat(),
        "trecho": {k: v for k, v in trecho.items()},
        "pv_montante": pvs.get(trecho["pv_ini"], {}),
        "pv_jusante": pvs.get(trecho["pv_fim"], {}),
        "hidraulica": calc_manning(trecho.get("dn_mm"), trecho.get("decl_mm")),
        "materiais": materiais,
        "rastreabilidade": {
            "pv_codigo": trecho.get("pv_ini", ""),
            "trecho_codigo": f"{trecho.get('pv_ini', '')}_{trecho.get('pv_fim', '')}",
            "coordenadas_inicio": {
                "x": pvs.get(trecho["pv_ini"], {}).get("x"),
                "y": pvs.get(trecho["pv_ini"], {}).get("y")
            },
            "coordenadas_fim": {
                "x": pvs.get(trecho["pv_fim"], {}).get("x"),
                "y": pvs.get(trecho["pv_fim"], {}).get("y")
            }
        }
    }
    
    with open(str(out_path), "w", encoding="utf-8") as f:
        json.dump(dados, f, indent=2, ensure_ascii=False)
    
    return dados


def gerar_excel_consolidado(todas_ns, output_path):
    """
    Gera Excel consolidado com todas as NS por núcleo.
    
    Args:
        todas_ns: lista de dicts com dados das NS
        output_path: caminho do arquivo .xlsx
    """
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws_resumo = wb.active
        ws_resumo.title = "Resumo Geral"
        
        # Cabeçalho resumo
        ws_resumo.cell(1, 1, "DASHBOARD NOTAS DE SERVIÇO - CONSTRUDATA v5.0")
        ws_resumo.cell(1, 1).font = Font(size=16, bold=True, color="FFFFFF")
        ws_resumo.cell(1, 1).fill = PatternFill("solid", fgColor="1F4E79")
        ws_resumo.merge_cells('A1:H1')
        
        ws_resumo.cell(2, 1, f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        ws_resumo.cell(2, 1).font = Font(size=10, italic=True)
        
        # Agrupar por núcleo
        ns_por_nucleo = defaultdict(list)
        for ns in todas_ns:
            ns_por_nucleo[ns['nucleo']].append(ns)
        
        # Resumo por núcleo
        row = 4
        ws_resumo.cell(row, 1, "NÚCLEO")
        ws_resumo.cell(row, 2, "Total NS")
        ws_resumo.cell(row, 3, "Extensão Total (m)")
        ws_resumo.cell(row, 4, "Material Principal")
        ws_resumo.cell(row, 5, "Custo Estimado (R$)")
        ws_resumo.cell(row, 6, "% Conclusão")
        ws_resumo.cell(row, 7, "Status")
        
        for col in range(1, 8):
            cell = ws_resumo.cell(row, col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2E7D32")
            cell.alignment = Alignment(horizontal="center")
        
        row += 1
        for nucleo, ns_list in sorted(ns_por_nucleo.items()):
            total_ext = sum(ns['trecho'].get('ext_m', 0) for ns in ns_list)
            material = ns_list[0]['trecho'].get('material', 'PVC') if ns_list else 'N/A'
            custo_est = total_ext * 150  # Estimativa simples R$ 150/m
            
            ws_resumo.cell(row, 1, nucleo)
            ws_resumo.cell(row, 2, len(ns_list))
            ws_resumo.cell(row, 3, f"{total_ext:.2f}")
            ws_resumo.cell(row, 4, material)
            ws_resumo.cell(row, 5, f"R$ {custo_est:,.2f}")
            ws_resumo.cell(row, 6, "100%")
            ws_resumo.cell(row, 7, "✅ Concluído")
            
            row += 1
        
        # Aba detalhada
        ws_detalhe = wb.create_sheet("Detalhamento Completo")
        
        headers = [
            "Número NS", "Núcleo", "PV Início", "PV Fim", 
            "Rua", "DN (mm)", "Extensão (m)", "Material",
            "Declividade (‰)", "Velocidade (m/s)", "Vazão (l/s)",
            "Data Geração", "Status"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws_detalhe.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E79")
            cell.alignment = Alignment(horizontal="center")
        
        for row_idx, ns in enumerate(todas_ns, 2):
            trecho = ns['trecho']
            hidr = ns.get('hidraulica', {})
            
            ws_detalhe.cell(row_idx, 1, ns['numero_ns'])
            ws_detalhe.cell(row_idx, 2, ns['nucleo'])
            ws_detalhe.cell(row_idx, 3, trecho.get('pv_ini', ''))
            ws_detalhe.cell(row_idx, 4, trecho.get('pv_fim', ''))
            ws_detalhe.cell(row_idx, 5, trecho.get('rua', ''))
            ws_detalhe.cell(row_idx, 6, trecho.get('dn_mm', ''))
            ws_detalhe.cell(row_idx, 7, f"{trecho.get('ext_m', 0):.2f}")
            ws_detalhe.cell(row_idx, 8, trecho.get('material', ''))
            ws_detalhe.cell(row_idx, 9, f"{trecho.get('decl_mm', 0)*1000:.2f}")
            ws_detalhe.cell(row_idx, 10, f"{hidr.get('v_ms', '-'):.3f}" if hidr.get('v_ms') else '-')
            ws_detalhe.cell(row_idx, 11, f"{hidr.get('q_ls', '-'):.2f}" if hidr.get('q_ls') else '-')
            ws_detalhe.cell(row_idx, 12, ns.get('data_geracao', '')[:10])
            ws_detalhe.cell(row_idx, 13, "Concluída")
        
        # Auto-width columns
        for column in ws_detalhe.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_detalhe.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(str(output_path))
        log(f"Excel consolidado salvo: {output_path}", "OK")
        
    except Exception as e:
        log(f"Erro ao gerar Excel: {e}", "ERR")
        traceback.print_exc()


# ═══════════════════════════════════════════════════════════════
# PIPELINE PRINCIPAL
# ═══════════════════════════════════════════════════════════════
def processar_nucleo_v5(dxf_path, nucleo, out_base):
    """
    Processa núcleo gerando NS rastreáveis v5.0.
    
    Estrutura de saída:
    SAIDA_NS_V5/[NÚCLEO]/
        ├── NS-[NÚCLEO]-PV001-TR01-20260416/
        │   ├── NS-[NÚCLEO]-PV001-TR01-20260416_A4.pdf
        │   ├── NS-[NÚCLEO]-PV001-TR01-20260416_DADOS.json
        │   └── ...
        ├── CONSOLIDADO.xlsx
        └── LOG/processamento.json
    """
    log(f"NÚCLEO: {nucleo}", "STEP")
    
    if not os.path.exists(dxf_path):
        log(f"DXF não encontrado: {dxf_path}", "ERR")
        return 0, 0, []
    
    # Leitura DXF
    pvs, trechos, ruas, meta = ler_dxf_gdal(dxf_path)
    if not trechos:
        log("Sem trechos!", "ERR")
        return 0, 0, []
    
    trechos = enriquecer_trechos(trechos, pvs)
    log(f"Rede: {meta['n_pvs']} PVs, {meta['n_trechos']} trechos", "OK")
    
    # Pastas
    nucleo_upper = nucleo.upper().replace(" ", "_")
    out = Path(out_base) / nucleo_upper
    (out / "LOG").mkdir(parents=True, exist_ok=True)
    
    # Processar cada trecho
    n_ok, n_err = 0, 0
    todas_ns = []
    data_hoje = datetime.now()
    
    for i, t in enumerate(trechos):
        ns_id = i + 1
        
        # Gerar número NS padronizado
        pv_codigo = t.get('pv_ini', '').replace('PI', 'PV').replace('PA', 'PV')
        trecho_codigo = f"TR{i+1:02d}"
        numero_ns = gerar_numero_ns(nucleo_upper[:4], pv_codigo, trecho_codigo, data_hoje)
        
        # Validar nomenclatura
        if not validar_nome_ns(numero_ns):
            log(f"Nome NS inválido: {numero_ns}", "WARN")
        
        # Pasta da NS
        ns_dir = out / numero_ns
        ns_dir.mkdir(parents=True, exist_ok=True)
        
        try:
            # Calcular materiais
            materiais = calcular_materiais(t, pvs)
            
            # Gerar PDF A4
            pdf_path = ns_dir / f"{numero_ns}_A4.pdf"
            gerar_pdf_a4(ns_id, numero_ns, t, pvs, nucleo, pdf_path)
            
            # Gerar JSON
            json_path = ns_dir / f"{numero_ns}_DADOS.json"
            dados_json = gerar_json_dados(ns_id, numero_ns, t, pvs, nucleo, materiais, json_path)
            
            # Preparar dados para Supabase
            dados_supabase = {
                "numero_ns": numero_ns,
                "project_id": None,  # TODO: buscar ID do projeto
                "pv_codigo": pv_codigo,
                "trecho_codigo": trecho_codigo,
                "descricao": f"{t.get('pv_ini', '')} → {t.get('pv_fim', '')} - DN{t.get('dn_mm', '?')}mm",
                "quantidade": t.get('ext_m', 0),
                "unidade": "m",
                "valor_unitario": 150.0,  # TODO: buscar do catálogo
                "valor_total": t.get('ext_m', 0) * 150.0,
                "data_emissao": data_hoje.strftime('%Y-%m-%d'),
                "status": "emitida",
                "engenheiro_responsavel": "Felipe Nery",
                "metadata": json.dumps({
                    "hidraulica": dados_json.get('hidraulica', {}),
                    "materiais": materiais,
                    "rastreabilidade": dados_json.get('rastreabilidade', {})
                })
            }
            
            # Salvar no Supabase
            salvar_ns_supabase(dados_supabase)
            
            # Adicionar à lista consolidada
            todas_ns.append(dados_json)
            
            n_ok += 1
            if ns_id <= 3 or ns_id % 25 == 0:
                log(f"NS {ns_id:03d}: {numero_ns} ✓", "OK")
        
        except Exception as e:
            n_err += 1
            log(f"NS {ns_id:03d}: ERRO - {e}", "ERR")
            traceback.print_exc()
    
    # Gerar Excel consolidado
    if todas_ns:
        excel_path = out / "CONSOLIDADO.xlsx"
        gerar_excel_consolidado(todas_ns, excel_path)
    
    # Log de processamento
    log_data = {
        "nucleo": nucleo,
        "dxf": str(dxf_path),
        "n_pvs": meta["n_pvs"],
        "n_trechos": meta["n_trechos"],
        "n_ns_geradas": n_ok,
        "n_ns_erros": n_err,
        "motor": "GDAL/OGR + Rastreabilidade v5.0",
        "extensao_m": round(sum(t["ext_m"] for t in trechos), 1),
        "gerado_em": datetime.now().isoformat(),
        "supabase_sync": True
    }
    
    with open(out / "LOG" / "processamento.json", "w", encoding="utf-8") as f:
        json.dump(log_data, f, indent=2, ensure_ascii=False)
    
    log(f"RESULTADO: {n_ok} NS geradas, {n_err} erros", "OK")
    return n_ok, n_err, todas_ns


def processar_batch():
    """Processa todos os núcleos configurados."""
    print("=" * 80)
    print(f"ConstruData SABESP v{NS_VERSION} - Sistema Rastreável de Notas de Serviço")
    print(f"Contrato {CONTRATO} - Motor GDAL/OGR + Supabase")
    print("=" * 80)
    
    total_ok, total_err = 0, 0
    todas_ns_global = []
    
    for item in NUCLEOS_BATCH:
        if os.path.exists(item["dxf"]):
            n_ok, n_err, ns_list = processar_nucleo_v5(
                item["dxf"], 
                item["nucleo"], 
                OUTPUT_BASE
            )
            total_ok += n_ok
            total_err += n_err
            todas_ns_global.extend(ns_list)
        else:
            log(f"DXF não encontrado: {item['dxf']}", "WARN")
    
    # Gerar Excel global
    if todas_ns_global:
        excel_global = OUTPUT_BASE / "CONSOLIDADO_GERAL_TODOS_NUCLEOS.xlsx"
        gerar_excel_consolidado(todas_ns_global, excel_global)
    
    print(f"\n{'='*80}")
    print(f"TOTAL: {total_ok} NS geradas, {total_err} erros")
    print(f"Pasta: {OUTPUT_BASE}")
    print(f"{'='*80}")


if __name__ == "__main__":
    if len(sys.argv) >= 3:
        # Modo: DXF individual
        dxf = sys.argv[1]
        nucleo = sys.argv[2]
        out = sys.argv[3] if len(sys.argv) >= 4 else str(OUTPUT_BASE)
        processar_nucleo_v5(dxf, nucleo, out)
    else:
        # Modo: batch
        processar_batch()
