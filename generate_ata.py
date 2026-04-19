import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.dimensions import ColumnDimension, RowDimension

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Ata de Reunião"

# Colors
HEADER_COLOR = "002060" # Dark blue
SUBHEADER_COLOR = "4F81BD" # Medium blue
TEXT_COLOR = "000000"
WHITE = "FFFFFF"
WARNING_COLOR = "FFC000" # Yellow
URGENT_COLOR = "FF0000" # Red
LIGHT_BLUE = "DCE6F1"

# Fonts
font_title = Font(name='Segoe UI', size=16, bold=True, color=WHITE)
font_subtitle = Font(name='Segoe UI', size=12, bold=True, color=WHITE)
font_bold = Font(name='Segoe UI', size=11, bold=True)
font_normal = Font(name='Segoe UI', size=11)
font_warning = Font(name='Segoe UI', size=11, bold=True, color="9C6500")

# Alignments
align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Borders
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def set_range_bg(ws, range_str, color):
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for row in ws[range_str]:
        for cell in row:
            cell.fill = fill

def merge_and_format(ws, range_str, value, font, alignment, fill_color=None, border=None):
    ws.merge_cells(range_str)
    cell = ws[range_str.split(':')[0]]
    cell.value = value
    cell.font = font
    cell.alignment = alignment
    if fill_color:
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        cell.fill = fill
    if border:
        set_border(ws, range_str, border)

def set_border(ws, cell_range, border):
    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            cell.border = border

# Column widths
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 45
ws.column_dimensions['D'].width = 25
ws.column_dimensions['E'].width = 25
ws.column_dimensions['F'].width = 3

row = 2

# Main Title
merge_and_format(ws, f'B{row}:E{row}', "ATA DE REUNIÃO - CONSTRUDATAMAX", font_title, align_center, HEADER_COLOR, thin_border)
ws.row_dimensions[row].height = 30
row += 2

# INFORMAÇÕES GERAIS
merge_and_format(ws, f'B{row}:E{row}', "INFORMAÇÕES GERAIS", font_subtitle, align_center, SUBHEADER_COLOR, thin_border)
row += 1

info_data = [
    ("Data:", "07/04/2026", "Horário:", "13:00h"),
    ("Local:", "Escritório de obra / Remoto", "Projeto:", "Contrato CT 11481051"),
    ("Elaborado por:", "Engenharia — ConstruDataMax", "", "")
]

for d1, d2, d3, d4 in info_data:
    ws[f'B{row}'] = d1; ws[f'B{row}'].font = font_bold; ws[f'B{row}'].border = thin_border; ws[f'B{row}'].fill = PatternFill(fill_type="solid", start_color=LIGHT_BLUE)
    ws[f'C{row}'] = d2; ws[f'C{row}'].font = font_normal; ws[f'C{row}'].border = thin_border
    ws[f'D{row}'] = d3; ws[f'D{row}'].font = font_bold; ws[f'D{row}'].border = thin_border; ws[f'D{row}'].fill = PatternFill(fill_type="solid", start_color=LIGHT_BLUE)
    ws[f'E{row}'] = d4; ws[f'E{row}'].font = font_normal; ws[f'E{row}'].border = thin_border
    row += 1

row += 1

# PARTICIPANTES
merge_and_format(ws, f'B{row}:E{row}', "PARTICIPANTES", font_subtitle, align_center, SUBHEADER_COLOR, thin_border)
row += 1

participants = [
    ("Cosme", "AVANT — Representante / Coordenação"),
    ("(preencher)", "Consórcio — Engenheiro Responsável"),
    ("(preencher)", "Consórcio — Coordenador de Topografia"),
    ("(preencher)", "Consórcio — Encarregado de Campo")
]

ws[f'B{row}'] = "NOME"; ws[f'B{row}'].font = font_bold; ws[f'B{row}'].alignment = align_center; ws[f'B{row}'].border = thin_border; ws[f'B{row}'].fill = PatternFill(fill_type="solid", start_color=LIGHT_BLUE)
ws.merge_cells(f'C{row}:E{row}'); cell = ws[f'C{row}']; cell.value = "EMPRESA / FUNÇÃO"; cell.font = font_bold; cell.alignment = align_center; cell.fill = PatternFill(fill_type="solid", start_color=LIGHT_BLUE)
set_border(ws, f'C{row}:E{row}', thin_border)
row += 1

for name, role in participants:
    ws[f'B{row}'] = name; ws[f'B{row}'].font = font_normal; ws[f'B{row}'].border = thin_border; ws[f'B{row}'].alignment = align_center
    ws.merge_cells(f'C{row}:E{row}'); cell = ws[f'C{row}']; cell.value = role; cell.font = font_normal; cell.alignment = align_left
    set_border(ws, f'C{row}:E{row}', thin_border)
    row += 1

row += 1

# PONTOS DISCUTIDOS
merge_and_format(ws, f'B{row}:E{row}', "PONTOS DISCUTIDOS E DELIBERAÇÕES", font_subtitle, align_center, SUBHEADER_COLOR, thin_border)
row += 1

points = [
    ("1. Deficiência no cadastro de redes", "Constatado que o cadastro das redes executadas apresenta deficiência significativa. O Consórcio deve realizar diagnóstico completo e priorizar regularização."),
    ("2. Comunicação Topografia × Produção", "Topógrafo não conversa com o Engenheiro de campo. Fica OBRIGATÓRIA a comunicação diária a partir de hoje. A falta de comunicação é considerada falha crítica."),
    ("3. Programação da Topografia", "Atrasos recorrentes nas demandas. O Consórcio deve apurar razões do não cumprimento e apresentar justificativa formal."),
    ("4. Diário de Obras (Topógrafo × Produção)", "Será implementado um Diário de Obras compartilhado a partir de 08/04/2026 para registrar serviços, pendências e solicitações diárias com assinaturas."),
    ("5. Cadastros Atrasados (Fevereiro/Março)", "Resolver passivo antigo + fev/mar com PRAZO FIRME para 14/04/2026. Ação distribuída por núcleo. Inclui revisão de cadastro A4 sem pedido do gerente geral e levantamento do não medido."),
    ("6. Prazos de Entrega SABESP", "Extensão de ramais com números sem cadastro oficial (Atualização do Survey) para 16/04/2026. Segunda entrega para 02/05/2026. Prazos inegociáveis.")
]

for title, desc in points:
    ws.merge_cells(f'B{row}:E{row}')
    cell = ws[f'B{row}']
    cell.value = title
    cell.font = font_bold
    cell.fill = PatternFill(fill_type="solid", start_color=LIGHT_BLUE)
    set_border(ws, f'B{row}:E{row}', thin_border)
    row += 1
    
    ws.merge_cells(f'B{row}:E{row}')
    cell = ws[f'B{row}']
    cell.value = desc
    cell.font = font_normal
    cell.alignment = align_left_top
    ws.row_dimensions[row].height = 40 
    set_border(ws, f'B{row}:E{row}', thin_border)
    row += 1

row += 1

# QUADRO RESUMO
merge_and_format(ws, f'B{row}:E{row}', "QUADRO RESUMO DE AÇÕES E PRAZOS", font_subtitle, align_center, HEADER_COLOR, thin_border)
row += 1

headers = ["#", "Ação", "Responsável", "Prazo"]
ws[f'B{row}'] = headers[0]; ws[f'B{row}'].font = font_bold; ws[f'B{row}'].fill = PatternFill(fill_type="solid", start_color=LIGHT_BLUE); ws[f'B{row}'].border = thin_border; ws[f'B{row}'].alignment = align_center
ws[f'C{row}'] = headers[1]; ws[f'C{row}'].font = font_bold; ws[f'C{row}'].fill = PatternFill(fill_type="solid", start_color=LIGHT_BLUE); ws[f'C{row}'].border = thin_border; ws[f'C{row}'].alignment = align_center
ws[f'D{row}'] = headers[2]; ws[f'D{row}'].font = font_bold; ws[f'D{row}'].fill = PatternFill(fill_type="solid", start_color=LIGHT_BLUE); ws[f'D{row}'].border = thin_border; ws[f'D{row}'].alignment = align_center
ws[f'E{row}'] = headers[3]; ws[f'E{row}'].font = font_bold; ws[f'E{row}'].fill = PatternFill(fill_type="solid", start_color=LIGHT_BLUE); ws[f'E{row}'].border = thin_border; ws[f'E{row}'].alignment = align_center
row += 1

actions = [
    (1, "Diagnóstico completo das pendências de cadastro de redes", "Consórcio — Topografia", "09/04/2026"),
    (2, "Apurar razões do não cumprimento da programação da Topografia", "Consórcio — Coord.", "09/04/2026"),
    (3, "Implementar comunicação diária Topógrafo ↔ Engenheiro", "Consórcio", "Imediato"),
    (4, "Iniciar Diário de Obras (Topógrafo × Encarregado)", "Consórcio", "08/04/2026"),
    (5, "Regularizar cadastros atrasados (fev/mar e passivo antigo)", "Consórcio — Topografia", "14/04/2026"),
    (6, "Revisão de cadastro A4 + levantamento do que não foi medido", "Consórcio — Topografia", "14/04/2026"),
    (7, "Entrega extensão ramais para SABESP — Survey", "Consórcio", "16/04/2026"),
    (8, "Entrega extensão ramais para SABESP — 2ª remessa", "Consórcio", "02/05/2026")
]

for a, b, c, d in actions:
    ws[f'B{row}'] = a; ws[f'B{row}'].font = font_normal; ws[f'B{row}'].border = thin_border; ws[f'B{row}'].alignment = align_center
    ws[f'C{row}'] = b; ws[f'C{row}'].font = font_normal; ws[f'C{row}'].border = thin_border; ws[f'C{row}'].alignment = align_left; ws[f'C{row}'].alignment = Alignment(wrap_text=True)
    ws[f'D{row}'] = c; ws[f'D{row}'].font = font_normal; ws[f'D{row}'].border = thin_border; ws[f'D{row}'].alignment = align_center; ws[f'D{row}'].alignment = Alignment(wrap_text=True)
    ws[f'E{row}'] = d; ws[f'E{row}'].font = Font(name='Segoe UI', size=11, bold=True, color="9C6500"); ws[f'E{row}'].border = thin_border; ws[f'E{row}'].alignment = align_center 
    ws.row_dimensions[row].height = 30
    row += 1

row += 2

# ASSINATURAS
ws.merge_cells(f'B{row}:C{row}')
ws[f'B{row}'] = "________________________________________"
ws[f'B{row}'].alignment = align_center

ws.merge_cells(f'D{row}:E{row}')
ws[f'D{row}'] = "________________________________________"
ws[f'D{row}'].alignment = align_center
row += 1

ws.merge_cells(f'B{row}:C{row}')
ws[f'B{row}'] = "Cosme / AVANT"
ws[f'B{row}'].alignment = align_center
ws[f'B{row}'].font = font_bold

ws.merge_cells(f'D{row}:E{row}')
ws[f'D{row}'] = "Consórcio — Eng. Responsável"
ws[f'D{row}'].alignment = align_center
ws[f'D{row}'].font = font_bold
row += 3

ws.merge_cells(f'B{row}:C{row}')
ws[f'B{row}'] = "________________________________________"
ws[f'B{row}'].alignment = align_center

ws.merge_cells(f'D{row}:E{row}')
ws[f'D{row}'] = "________________________________________"
ws[f'D{row}'].alignment = align_center
row += 1

ws.merge_cells(f'B{row}:C{row}')
ws[f'B{row}'] = "Consórcio — Coord. de Topografia"
ws[f'B{row}'].alignment = align_center
ws[f'B{row}'].font = font_bold

ws.merge_cells(f'D{row}:E{row}')
ws[f'D{row}'] = "Consórcio — Encarregado de Campo"
ws[f'D{row}'].alignment = align_center
ws[f'D{row}'].font = font_bold

wb.save("ATA_REUNIAO_AVANT_CONSORCIO_07_04_2026.xlsx")
print("Excel generated successfully.")
