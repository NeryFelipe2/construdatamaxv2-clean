#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MOTOR_MAESTRO_LOTE.PY   —   Maestro Central RK Engenharia
ConstruDataMax Gestão 360 · Versão 1.0 · Abril 2026

Modo de operação: LOTE MULTI-OBRA
  - Processa TODAS as obras cadastradas em OBRAS_ATIVAS em um único ciclo
  - Roda como script direto (CLI) ou como endpoint FastAPI
  - Chamado pelo n8n Schedule Trigger (todo dia útil às 18h)
  - Salva snapshot JSON no Supabase + envia resumo WhatsApp via Evolution API

Arquitetura:
  n8n Schedule Trigger 18h
    → POST /api/maestro/rodar
    → FastAPI (Render / local)
    → motor_maestro_lote.py (lote multi-obra)
    → Supabase (snapshot diário)
    → Evolution API (mensagem resumo diretores)
"""

import json
import os
import sys
import requests
from datetime import datetime
from pathlib import Path
from collections import defaultdict

# ── Dependências ─────────────────────────────────────────────────────────────
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter
    XLSX_OK = True
except ImportError:
    XLSX_OK = False

# ── Caminhos ──────────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).parent
OUTPUT_DIR = SCRIPT_DIR / "output"
OUTPUT_DIR.mkdir(exist_ok=True)


# ══════════════════════════════════════════════════════════════════════════════
# 1. CONFIGURAÇÃO CENTRAL DE OBRAS
#    → Para adicionar/remover uma obra, mexa APENAS aqui.
# ══════════════════════════════════════════════════════════════════════════════

OBRAS_ATIVAS = {
    # --- Obras SLNR (Sabesp / FCN) ---
    "SLNR_SANTOS": {
        "nome": "SLNR Santos",
        "contrato": "CT 11481051",
        "cliente": "Sabesp",
        "responsavel": "João",
        "whatsapp_grupo": "5561981846325",   # número do responsável (Evolution API)
        "custo_lig": 1820.0,
        "custo_metro": 910.0,
        "supabase_project_id": "SLNR",
    },
    # --- Obras RK Engenharia ---
    "TATUI": {
        "nome": "Obra Tatuí",
        "contrato": "RK-TATUI-2025",
        "cliente": "RK Engenharia",
        "responsavel": "Ícaro",
        "whatsapp_grupo": "5561981846325",
        "custo_lig": 0,
        "custo_metro": 0,
        "supabase_project_id": "TATUI",
    },
    "OSASCO": {
        "nome": "Obra Osasco CLU",
        "contrato": "RK-OSASCO-2025",
        "cliente": "RK Engenharia",
        "responsavel": "Mateus",
        "whatsapp_grupo": "5561981846325",
        "custo_lig": 0,
        "custo_metro": 0,
        "supabase_project_id": "OSASCO",
    },
    "PARDINHO": {
        "nome": "Obra Pardinho",
        "contrato": "RK-PARDINHO-2025",
        "cliente": "RK Engenharia",
        "responsavel": "Ícaro",
        "whatsapp_grupo": "5561981846325",
        "custo_lig": 0,
        "custo_metro": 0,
        "supabase_project_id": "PARDINHO",
    },
    "CACHOEIRO": {
        "nome": "Obra Cachoeiro",
        "contrato": "RK-CACHOEIRO-2025",
        "cliente": "RK Engenharia",
        "responsavel": "Thalita",
        "whatsapp_grupo": "5561981846325",
        "custo_lig": 0,
        "custo_metro": 0,
        "supabase_project_id": "CACHOEIRO",
    },
    "TEOFILO": {
        "nome": "Obra Teófilo",
        "contrato": "RK-TEOFILO-2025",
        "cliente": "RK Engenharia",
        "responsavel": "Gabriel",
        "whatsapp_grupo": "5561981846325",
        "custo_lig": 0,
        "custo_metro": 0,
        "supabase_project_id": "TEOFILO",
    },
}

# Grupo dos Diretores para receber o resumo consolidado
GRUPO_DIRETORES_WHATSAPP = os.environ.get("GRUPO_DIRETORES_WPP", "5561981846325")


# ══════════════════════════════════════════════════════════════════════════════
# 2. CARREGAMENTO DE CREDENCIAIS (Fallback inteligente multi-env)
# ══════════════════════════════════════════════════════════════════════════════

def load_env():
    """Carrega .env com fallback em cascata (local → mestre → Railway)."""
    candidatos = [
        SCRIPT_DIR / ".env",
        Path(r"C:\Users\felip\Desktop\_ORGANIZADO\01-OBRAS-RK\.env"),
        Path(r"C:\Users\felip\.env"),
    ]
    for p in candidatos:
        if p.exists():
            with open(p, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith("#") and "=" in line:
                        k, v = line.split("=", 1)
                        os.environ.setdefault(k.strip(), v.strip())
            print(f"[ENV] Credenciais carregadas de: {p}")
            return str(p)
    print("[ENV] Sem .env encontrado. Usando variáveis de ambiente do sistema.")
    return None


load_env()

SUPABASE_URL  = os.environ.get("SUPABASE_URL", "")
SUPABASE_KEY  = (
    os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    or os.environ.get("SUPABASE_KEY")
    or os.environ.get("SUPABASE_ANON_KEY", "")
)
EVOLUTION_URL = os.environ.get("EVOLUTION_API_URL",
                               "https://evolution-api-production-b130.up.railway.app")
EVOLUTION_KEY = os.environ.get("EVOLUTION_API_KEY", "")
EVOLUTION_INST = os.environ.get("EVOLUTION_INSTANCE", "construdata-felipe")
GEMINI_KEY    = os.environ.get("GEMINI_API_KEY", "")

SUPABASE_HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "resolution=merge-duplicates",
}


# ══════════════════════════════════════════════════════════════════════════════
# 3. MOTOR DE ANÁLISE (por obra)
# ══════════════════════════════════════════════════════════════════════════════

class MotorObra:
    """Analisa GAP financeiro de UMA obra específica."""

    def __init__(self, obra_id: str, config: dict):
        self.obra_id = obra_id
        self.cfg     = config
        self.dados   = []          # registros brutos do Supabase
        self.resumo  = {}
        self.status  = "OK"        # OK | ATENCAO | CRITICO

    def buscar_supabase(self) -> bool:
        """Busca lançamentos da obra no Supabase."""
        if not SUPABASE_URL or not SUPABASE_KEY:
            return False

        pid = self.cfg["supabase_project_id"]
        url = (
            f"{SUPABASE_URL}/rest/v1/custos"
            f"?project_id=eq.{pid}"
            f"&order=data.desc"
            f"&limit=2000"
        )
        try:
            r = requests.get(url, headers=SUPABASE_HEADERS, timeout=15)
            if r.status_code == 200:
                self.dados = r.json()
                return True
            else:
                print(f"  [WRN] {self.obra_id} Supabase {r.status_code}: {r.text[:100]}")
        except Exception as e:
            print(f"  [ERR] {self.obra_id} conexão Supabase: {e}")
        return False

    def calcular(self) -> dict:
        """Calcula resumo financeiro da obra."""
        receitas = sum(float(d.get("valor", 0)) for d in self.dados
                       if d.get("tipo") in ("RECEITA", "CONTRATO"))
        gastos   = sum(float(d.get("valor", 0)) for d in self.dados
                       if d.get("tipo") in ("GASTO", "CUSTO", None, ""))

        # Para SLNR: usa lógica de GAP específica
        if self.obra_id == "SLNR_SANTOS":
            gap_registros = [d for d in self.dados
                             if d.get("categoria") == "GAP_MEDICAO"]
            gap_total = sum(float(d.get("valor", 0)) for d in gap_registros)
            n_meses_criticos = sum(
                1 for d in gap_registros
                if json.loads(d.get("metadata") or "{}").get("status") == "CRITICO"
            )
        else:
            gap_total = max(0, gastos - receitas)
            n_meses_criticos = 0

        saldo = receitas - gastos
        n_lancamentos = len(self.dados)

        if saldo < -50_000:
            self.status = "CRITICO"
        elif saldo < 0:
            self.status = "ATENCAO"
        else:
            self.status = "OK"

        self.resumo = {
            "obra_id":          self.obra_id,
            "nome":             self.cfg["nome"],
            "responsavel":      self.cfg["responsavel"],
            "status":           self.status,
            "receitas":         round(receitas, 2),
            "gastos":           round(gastos, 2),
            "saldo":            round(saldo, 2),
            "gap_medicao":      round(gap_total, 2),
            "n_lancamentos":    n_lancamentos,
            "meses_criticos":   n_meses_criticos,
            "rodado_em":        datetime.now().isoformat(),
        }
        return self.resumo


# ══════════════════════════════════════════════════════════════════════════════
# 4. ORQUESTRADOR LOTE (O Maestro)
# ══════════════════════════════════════════════════════════════════════════════

class MaestroLote:
    """Orquestra o ciclo completo para TODAS as obras ativas."""

    def __init__(self):
        self.ts       = datetime.now().strftime("%Y%m%d_%H%M")
        self.resultados = {}      # {obra_id: resumo}
        self.supabase_ok = bool(SUPABASE_URL and SUPABASE_KEY)

    def rodar(self) -> dict:
        """Executa o ciclo completo: busca → calcula → salva → notifica."""
        print("\n" + "═"*60)
        print(f"  MAESTRO LOTE — {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        print(f"  Obras ativas: {len(OBRAS_ATIVAS)}")
        print(f"  Supabase: {'✓ conectado' if self.supabase_ok else '✗ sem credenciais'}")
        print("═"*60 + "\n")

        for obra_id, cfg in OBRAS_ATIVAS.items():
            print(f"  ▶ {obra_id} — {cfg['nome']}")
            motor = MotorObra(obra_id, cfg)

            if self.supabase_ok:
                motor.buscar_supabase()
            else:
                print("    [AVISO] Sem Supabase — usando dados zerados")

            resumo = motor.calcular()
            self.resultados[obra_id] = resumo

            emoji_status = {"OK": "✅", "ATENCAO": "⚠️", "CRITICO": "🔴"}.get(resumo["status"], "—")
            print(f"    {emoji_status} Saldo: R$ {resumo['saldo']:,.2f} | "
                  f"GAP: R$ {resumo['gap_medicao']:,.2f} | "
                  f"Lançamentos: {resumo['n_lancamentos']}")
            print()

        # Consolidado geral
        self.consolidado = self._consolidar()

        # Ações
        self._salvar_snapshot()
        if self.supabase_ok:
            self._push_supabase()
        self._gerar_excel()
        self._notificar_whatsapp()

        return self.consolidado

    def _consolidar(self) -> dict:
        """Consolida resultados de todas as obras."""
        total_receitas = sum(r["receitas"] for r in self.resultados.values())
        total_gastos   = sum(r["gastos"]   for r in self.resultados.values())
        total_gap      = sum(r["gap_medicao"] for r in self.resultados.values())
        n_criticos     = sum(1 for r in self.resultados.values() if r["status"] == "CRITICO")
        n_atencao      = sum(1 for r in self.resultados.values() if r["status"] == "ATENCAO")

        return {
            "rodado_em":     datetime.now().isoformat(),
            "obras_total":   len(self.resultados),
            "obras_criticas": n_criticos,
            "obras_atencao": n_atencao,
            "total_receitas": round(total_receitas, 2),
            "total_gastos":   round(total_gastos, 2),
            "saldo_geral":    round(total_receitas - total_gastos, 2),
            "gap_financeiro": round(total_gap, 2),
            "obras":          self.resultados,
        }

    def _salvar_snapshot(self):
        """Salva snapshot JSON local."""
        out = OUTPUT_DIR / f"maestro_snapshot_{self.ts}.json"
        with open(out, "w", encoding="utf-8") as f:
            json.dump(self.consolidado, f, indent=2, ensure_ascii=False)
        print(f"  💾 Snapshot salvo: {out.name}")
        return str(out)

    def _push_supabase(self):
        """Envia snapshot consolidado para o Supabase (tabela maestro_snapshots)."""
        if not self.supabase_ok:
            return

        url = f"{SUPABASE_URL}/rest/v1/maestro_snapshots"
        payload = {
            "ts":            self.ts,
            "rodado_em":     self.consolidado["rodado_em"],
            "obras_total":   self.consolidado["obras_total"],
            "obras_criticas": self.consolidado["obras_criticas"],
            "saldo_geral":   self.consolidado["saldo_geral"],
            "gap_financeiro": self.consolidado["gap_financeiro"],
            "payload_json":  json.dumps(self.consolidado),
        }
        try:
            r = requests.post(url, headers=SUPABASE_HEADERS, json=payload, timeout=15)
            if r.status_code in (200, 201):
                print("  ☁️  Snapshot enviado ao Supabase (maestro_snapshots)")
            else:
                print(f"  [WRN] Supabase push: {r.status_code} — {r.text[:200]}")
        except Exception as e:
            print(f"  [ERR] Supabase push: {e}")

    def _gerar_excel(self):
        """Gera Excel consolidado com uma aba por obra + dashboard."""
        if not XLSX_OK:
            return

        out = OUTPUT_DIR / f"MAESTRO_CONSOLIDADO_{self.ts}.xlsx"
        wb = Workbook()

        # ── Estilos ──────────────────────────────────────────────
        AZUL   = PatternFill("solid", fgColor="1B2A4A")
        VERDE  = PatternFill("solid", fgColor="27AE60")
        VERM   = PatternFill("solid", fgColor="E74C3C")
        AMAR   = PatternFill("solid", fgColor="F39C12")
        CINZA  = PatternFill("solid", fgColor="F5F5F5")
        BOLD_W = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        BOLD_B = Font(bold=True, size=10, name="Calibri")
        BD     = Border(*(Side(style="thin", color="CCCCCC"),)*4)
        MONEY  = 'R$ #,##0.00'

        def hdr(ws, row, cols, data):
            for j, h in enumerate(data, 1):
                c = ws.cell(row, j, h)
                c.font = BOLD_W; c.fill = AZUL
                c.alignment = Alignment(horizontal="center")
                c.border = BD

        def tit(ws, row, c1, c2, texto, cor="1B2A4A"):
            ws.merge_cells(start_row=row, start_column=c1,
                           end_row=row, end_column=c2)
            cell = ws.cell(row, c1, texto)
            cell.font = Font(bold=True, color="FFFFFF", size=13, name="Calibri")
            cell.fill = PatternFill("solid", fgColor=cor)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # ── ABA: DASHBOARD ────────────────────────────────────────
        ws = wb.active
        ws.title = "DASHBOARD"
        ws.row_dimensions[1].height = 28

        tit(ws, 1, 1, 7, f"RK ENGENHARIA · CONSOLIDADO {datetime.now().strftime('%d/%m/%Y')}")

        # KPIs
        kpis = [
            ("OBRAS ATIVAS",  str(self.consolidado["obras_total"]),   AZUL),
            ("CRITICAS 🔴",   str(self.consolidado["obras_criticas"]), VERM),
            ("ATENÇÃO ⚠️",    str(self.consolidado["obras_atencao"]),  AMAR),
            ("SALDO GERAL",   f"R$ {self.consolidado['saldo_geral']:,.2f}",
             VERDE if self.consolidado["saldo_geral"] >= 0 else VERM),
            ("GAP FINANCEIRO", f"R$ {self.consolidado['gap_financeiro']:,.2f}", AMAR),
        ]
        for i, (titulo, valor, cor) in enumerate(kpis):
            row = 3 + i
            ws.cell(row, 1, titulo).font = BOLD_B
            ws.cell(row, 2, valor).font = Font(bold=True, size=14)
            ws.cell(row, 3).fill = cor

        # Tabela resumo por obra
        hdr(ws, 10, 7,
            ["Obra", "Responsável", "Status", "Receitas", "Gastos", "Saldo", "GAP Medição"])
        for i, (oid, r) in enumerate(self.resultados.items(), 11):
            status_emoji = {"OK": "✅ OK", "ATENCAO": "⚠️ ATENÇÃO",
                            "CRITICO": "🔴 CRÍTICO"}.get(r["status"], r["status"])
            row_data = [
                r["nome"], r["responsavel"], status_emoji,
                r["receitas"], r["gastos"], r["saldo"], r["gap_medicao"]
            ]
            for j, v in enumerate(row_data, 1):
                cell = ws.cell(i, j, v)
                if j >= 4:
                    cell.number_format = MONEY
                cell.border = BD
                if i % 2 == 0:
                    cell.fill = CINZA

            # Cor do status
            status_cell = ws.cell(i, 3)
            if r["status"] == "CRITICO":
                status_cell.fill = VERM
                status_cell.font = Font(bold=True, color="FFFFFF")
            elif r["status"] == "ATENCAO":
                status_cell.fill = AMAR

        for col, w in zip("ABCDEFG", [22, 16, 16, 18, 18, 18, 18]):
            ws.column_dimensions[col].width = w

        # ── ABA por OBRA ──────────────────────────────────────────
        for oid, r in self.resultados.items():
            ws_o = wb.create_sheet(oid[:12])
            tit(ws_o, 1, 1, 4, f"{r['nome']} · {r['responsavel']}")
            campos = [
                ("Obra ID",       oid),
                ("Contrato",      OBRAS_ATIVAS[oid]["contrato"]),
                ("Status",        r["status"]),
                ("Receitas",      r["receitas"]),
                ("Gastos",        r["gastos"]),
                ("Saldo",         r["saldo"]),
                ("GAP Medição",   r["gap_medicao"]),
                ("Lançamentos",   r["n_lancamentos"]),
                ("Rodado em",     r["rodado_em"]),
            ]
            for i, (k, v) in enumerate(campos, 3):
                ws_o.cell(i, 1, k).font = BOLD_B
                cell = ws_o.cell(i, 2, v)
                if isinstance(v, float):
                    cell.number_format = MONEY
                cell.border = BD
            ws_o.column_dimensions["A"].width = 20
            ws_o.column_dimensions["B"].width = 28

        wb.save(str(out))
        wb.close()
        print(f"  📊 Excel salvo: {out.name}")
        return str(out)

    def _notificar_whatsapp(self):
        """Envia resumo consolidado via Evolution API para o grupo dos diretores."""
        if not EVOLUTION_KEY:
            print("  [WRN] Evolution API key não configurada. Pulando WhatsApp.")
            return

        # Monta mensagem
        linhas = [
            f"📊 *RELATÓRIO DIÁRIO RK ENGENHARIA*",
            f"_{datetime.now().strftime('%d/%m/%Y %H:%M')}_",
            "",
            f"📌 *{self.consolidado['obras_total']} obras analisadas*",
            f"🔴 Críticas: {self.consolidado['obras_criticas']}",
            f"⚠️ Atenção: {self.consolidado['obras_atencao']}",
            "",
        ]

        for oid, r in self.resultados.items():
            emoji = {"OK": "✅", "ATENCAO": "⚠️", "CRITICO": "🔴"}.get(r["status"], "—")
            saldo_fmt = f"R$ {r['saldo']:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            linhas.append(f"{emoji} *{r['nome']}*: Saldo {saldo_fmt}")

        saldo_g = f"R$ {self.consolidado['saldo_geral']:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        gap_g   = f"R$ {self.consolidado['gap_financeiro']:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

        linhas.extend([
            "",
            f"💰 *Saldo Consolidado:* {saldo_g}",
            f"⚡ *GAP Financeiro:* {gap_g}",
            "",
            "_ConstruDataMax Gestão 360 · Motor Maestro_",
        ])

        mensagem = "\n".join(linhas)

        url = f"{EVOLUTION_URL}/message/sendText/{EVOLUTION_INST}"
        payload = {
            "number": GRUPO_DIRETORES_WHATSAPP,
            "options": {"delay": 1000, "presence": "composing"},
            "textMessage": {"text": mensagem},
        }
        headers = {
            "Content-Type": "application/json",
            "apikey": EVOLUTION_KEY,
        }
        try:
            r = requests.post(url, json=payload, headers=headers, timeout=15)
            if r.status_code in (200, 201):
                print("  📱 Resumo enviado ao WhatsApp (diretores)")
            else:
                print(f"  [WRN] WhatsApp: {r.status_code} — {r.text[:200]}")
        except Exception as e:
            print(f"  [ERR] WhatsApp: {e}")


# ══════════════════════════════════════════════════════════════════════════════
# 5. ENDPOINT FASTAPI (para n8n chamar via HTTP)
# ══════════════════════════════════════════════════════════════════════════════

try:
    from fastapi import FastAPI, HTTPException, Header
    from fastapi.responses import JSONResponse

    app_fastapi = FastAPI(
        title="Motor Maestro — ConstruDataMax",
        description="Endpoint para n8n acionar o ciclo lote de todas as obras",
        version="1.0.0",
    )

    MAESTRO_SECRET = os.environ.get("MAESTRO_SECRET", "construdata-rk-2026")

    @app_fastapi.post("/api/maestro/rodar")
    async def rodar_maestro(x_maestro_key: str = Header(None)):
        """
        Endpoint chamado pelo n8n Schedule Trigger (18h dias úteis).
        Header: x-maestro-key: <MAESTRO_SECRET>
        """
        if x_maestro_key != MAESTRO_SECRET:
            raise HTTPException(status_code=401, detail="Chave inválida")

        maestro = MaestroLote()
        resultado = maestro.rodar()
        return JSONResponse(content=resultado)

    @app_fastapi.get("/api/maestro/status")
    async def status_maestro():
        """Health check rápido."""
        return {
            "status": "online",
            "supabase": bool(SUPABASE_URL and SUPABASE_KEY),
            "evolution": bool(EVOLUTION_KEY),
            "obras_cadastradas": list(OBRAS_ATIVAS.keys()),
            "ts": datetime.now().isoformat(),
        }

    FASTAPI_AVAILABLE = True

except ImportError:
    FASTAPI_AVAILABLE = False


# ══════════════════════════════════════════════════════════════════════════════
# 6. MAIN (CLI direto)
# ══════════════════════════════════════════════════════════════════════════════

def main():
    maestro = MaestroLote()
    resultado = maestro.rodar()

    print("\n" + "═"*60)
    print("  CONSOLIDADO FINAL")
    print("═"*60)
    print(f"  Obras analisadas:  {resultado['obras_total']}")
    print(f"  Obras CRÍTICAS:    {resultado['obras_criticas']}")
    print(f"  Saldo Geral:       R$ {resultado['saldo_geral']:,.2f}")
    print(f"  GAP Financeiro:    R$ {resultado['gap_financeiro']:,.2f}")
    print("═"*60 + "\n")

    return resultado


if __name__ == "__main__":
    main()
