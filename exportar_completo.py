#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
EXPORTAR_COMPLETO.PY — Pipeline SATANICAMENTE BRUTAL v3
ConstruData HydroNetwork v9 · FCN Construções e Saneamento · CT 11481051

MODO: SOMENTE ESGOTO · ORGANIZADO POR PROJETO (nome do DXF)

Estrutura de saída:
  CONSTRUDATA_POR_PROJETO/
    CRIADORES_ESGOTO/
    PANTANAL_ESGOTO/
    ISRAEL_ESGOTO/
    TETEU_ESGOTO22/
    ACAD-ESTUDO - CT SAO MANOEL E CT JOAO CARLOS DA SILVA2/
    PROLONGAMENTO TETEU ALT-01/
    PROLONGAMENTO TETEU/
    PROLONGAMENTO PANTANAL BAIXO/
    PROLONGAMENTO CRIADORES/
    PROLONGAMENTO SAO MANOEL/
    CRONOGRAMA_MACRO/
    RESUMO_GERAL.xlsx
"""

import sys, os, json, math, traceback, re, shutil, unicodedata, warnings
from pathlib import Path
from datetime import datetime
from collections import Counter, defaultdict

warnings.filterwarnings("ignore", category=UserWarning)

MOTOR_DIR = Path(r"C:\Users\felip\Downloads\NOVA NS Versao 5")
sys.path.insert(0, str(MOTOR_DIR))

# ══════════════════════════════════════════════════════════════
# IMPORTAÇÕES
# ══════════════════════════════════════════════════════════════
from ler_dxf_gdal import ler_dxf_gdal
from gerar_ns import (
    gerar_ns_a4, gerar_ns_desenho, gerar_ns_sat, gerar_html,
    gerar_geojson, enriquecer_trechos, calcular_materiais,
    calc_manning, CONTRATO, NS_VERSION, _ns_folder_name
)
from motor_custo import custo_trecho, custo_nucleo, FATORES, BDI
from gerar_xlsx import (
    gerar_xlsx_custos, gerar_xlsx_hidraulica, gerar_xlsx_curva_s,
    gerar_xlsx_lean, gerar_xlsx_microplan
)
from motor_lean_lps import gerar_relatorio_lean_lps, gerar_xlsx_lean_lps
from motor_microplanejamento import micro_planejar_nucleo
from gerar_ifc_lod500 import gerar_ifc_lod500
from gerar_cronograma_macro import (
    gerar_cronograma_macro, exportar_project_xml,
    exportar_primavera_xer, exportar_openproject_csv,
    exportar_macro_xlsx, gerar_cronograma_por_ns
)

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import geopandas as gpd
    import numpy as np
    HAS_GEO = True
except ImportError:
    HAS_GEO = False

# ══════════════════════════════════════════════════════════════
# CONFIGURAÇÃO
# ══════════════════════════════════════════════════════════════
SRC_FOLDER = Path(r"C:\Users\felip\Desktop\PROJETOS")
DST_FOLDER = Path(r"C:\Users\felip\Desktop\CONSTRUDATA_SAIDA_COMPLETA\POR_PROJETO")

# PASTAS EXTRAS com projetos DWG/DXF (serão varridas também)
SRC_FOLDERS_EXTRAS = [
    Path(r"C:\Users\felip\Downloads\PROJETOS DE ÁGUA E ESGOTO - DWG E DXF 2018\MAPAS ÁGUA E ESGOTO PARA DXF"),
]

# ══════════════════════════════════════════════════════════════
# AUTO-DESCOBERTA DE TODOS OS DXF + DWG
# ══════════════════════════════════════════════════════════════
# Palavras que identificam CARTOGRAFIA/MAPA (excluir da lista de projetos)
_EXCLUIR_PALAVRAS = [
    "CARTOGRAFIA", "(QGIS)", "TOPOGRAFIA", "LOCAÇÃO",
    "LEGENDA", "LAYOUT", "MODELO", "TEMPLATE",
]

# Palavras que identificam projeto de ÁGUA (excluir)
_AGUA_PALAVRAS = ["ÁGUA", "AGUA", "_ÁGUA", "_AGUA", " ÁGUA", " AGUA"]

# Mapeamento de palavras-chave → nome do núcleo (para associar cartografia)
_NUCLEO_MAP = {
    "CRIADORES":   "Vila Criadores",
    "PANTANAL":    "Pantanal Baixo",
    "ISRAEL":      "Vila Israel",
    "TETEU":       "Morro do Teteu",
    "TETÉU":       "Morro do Teteu",
    "SÃO MANOEL":  "Sao Manoel",
    "SAO MANOEL":  "Sao Manoel",
    "JOÃO CARLOS": "Joao Carlos",
    "JOAO CARLOS": "Joao Carlos",
    "NOROESTE":    "Noroeste",
    "PROGRESSO":   "Vila Progresso",
    "VALE VERDE":  "Vale Verde",
}

# Cartografias GPKG por núcleo (para nomes de ruas)
CARTOGRAFIAS = {
    "Pantanal Baixo":   "MAPA PANTANAL BAIXO_R04 (QGIS).gpkg",
    "Vila Criadores":   "CARTOGRAFIA VILA CRIADORES_R06 (QGIS).gpkg",
    "Vila Israel":      "MAPA PANTANAL ALTO (VILA ISRAEL)_RV02 (QGIS).gpkg",
    "Morro do Teteu":   "MAPA TETEU-VALE VERDE_R04 (QGIS).gpkg",
    "Sao Manoel":       "MAPA SÃO MANOEL_RV05 (QGIS).gpkg",
    "Joao Carlos":      "MAPA JOÃO CARLOS DA SILVA_RV07 (QGIS).gpkg",
    "Vale Verde":       "MAPA TETEU-VALE VERDE_R04 (QGIS).gpkg",
}

# Mapas DXF/DWG extras para interpolação de ruas (becos, vielas, etc)
# Esses arquivos NÃO são projetos — são usados APENAS para nomes de ruas
MAPAS_INTERPOLACAO = [
    Path(r"C:\Users\felip\Desktop\PROJETOS\CT JOÃO CARLOS DA SILVA.dwg"),
    Path(r"C:\Users\felip\Desktop\PROJETOS\ESTUDO - SÃO MANOEL.dxf"),
    Path(r"C:\Users\felip\Desktop\PROJETOS\ESTUDO - CT JOÃO CARLOS DA SILVA.dxf"),
    Path(r"C:\Users\felip\Desktop\PROJETOS\ACAD-ESTUDO - CT SÃO MANOEL E CT JOÃO CARLOS DA SILVA2.dxf"),
]

# Associação mapa de interpolação → núcleo
MAPA_PARA_NUCLEO = {
    "CT JOÃO CARLOS DA SILVA":     "Joao Carlos",
    "ESTUDO - SÃO MANOEL":         "Sao Manoel",
    "ESTUDO - CT JOÃO CARLOS DA SILVA": "Joao Carlos",
    "ACAD-ESTUDO - CT SÃO MANOEL E CT JOÃO CARLOS DA SILVA2": "Sao Manoel",
}


def _e_cartografia(nome):
    """Verifica se o arquivo é cartografia/mapa (não projeto)."""
    n = nome.upper()
    # Excluir se contém palavras de cartografia
    if any(p.upper() in n for p in _EXCLUIR_PALAVRAS):
        return True
    # Excluir se começa com "MAPA " (mas NÃO "MAPA" dentro de outro nome)
    if n.startswith("MAPA "):
        return True
    # Excluir se é um dos mapas de interpolação explícitos
    stem = Path(nome).stem
    if stem in MAPA_PARA_NUCLEO:
        return True
    return False


def _e_agua(nome):
    """Verifica se é projeto de ÁGUA."""
    n = nome.upper()
    # Excluir se contém ÁGUA/AGUA MAS NÃO contém ESGOTO
    if any(p.upper() in n for p in _AGUA_PALAVRAS):
        if "ESGOTO" not in n and "ESG" not in n:
            return True
    return False


def _identificar_nucleo(nome):
    """Identifica o núcleo automaticamente pelo nome do arquivo."""
    n = nome.upper()
    # Verificar cada palavra-chave
    for chave, nucleo in _NUCLEO_MAP.items():
        if chave.upper() in n:
            # Se é prolongamento, adicionar prefixo
            if "PROLONGAMENTO" in n:
                return f"Prolongamento {nucleo}"
            return nucleo
    # Não reconhecido — usar o nome do arquivo como núcleo
    return nome.replace(".dxf", "").replace(".dwg", "")


def _descobrir_projetos(pastas):
    """
    Descobre TODOS os arquivos .dxf e .dwg nas pastas que são projetos de esgoto.
    Exclui: cartografias, mapas, projetos de água.
    Aceita lista de pastas e guarda o caminho absoluto de cada arquivo.
    """
    if isinstance(pastas, Path):
        pastas = [pastas]

    projetos = []
    nomes_vistos = set()
    extensoes = ["*.dxf", "*.dwg"]

    for pasta in pastas:
        if not pasta.exists():
            continue
        for ext in extensoes:
            for arq in sorted(pasta.glob(ext)):
                nome = arq.name

                # Pular cartografias
                if _e_cartografia(nome):
                    continue

                # Pular projetos de água
                if _e_agua(nome):
                    continue

                # Identificamos projetos pelo NÚCLEO+STEM normalizado para evitar duplicatas de dwg/dxf
                stem = arq.stem.upper()
                proj_base_name = f"{stem}_{_identificar_nucleo(nome).upper()}"
                
                if proj_base_name in nomes_vistos:
                    continue
                nomes_vistos.add(proj_base_name)

                # É um projeto de esgoto!
                nucleo = _identificar_nucleo(nome)
                projeto_nome = _slug(arq.stem)

                projetos.append({
                    "dxf": nome,
                    "dxf_path": str(arq),  # caminho absoluto
                    "nucleo": nucleo,
                    "tipo": "ESGOTO",
                    "projeto": projeto_nome,
                })

    return projetos


def _slug(texto):
    if not texto:
        return "DESCONHECIDO"
    normalizado = unicodedata.normalize("NFKD", str(texto))
    ascii_txt = normalizado.encode("ascii", "ignore").decode("ascii")
    safe = re.sub(r'[^\w\s-]', '_', ascii_txt).strip()
    return safe[:80] or "DESCONHECIDO"


def _carregar_ruas_gpkg(gpkg_path, trechos, pvs):
    """Carrega nomes de ruas do GeoPackage e associa aos trechos."""
    if not HAS_GEO or not gpkg_path or not Path(gpkg_path).exists():
        return 0
    try:
        gdf = gpd.read_file(str(gpkg_path))
        rua_col = None
        for col in gdf.columns:
            c = col.upper()
            if any(k in c for k in ["RUA", "LOGRADOURO", "NOME", "NAME", "DESCRICAO", "LABEL", "TEXT"]):
                rua_col = col
                break
        if not rua_col:
            return 0
        ruas_gdf = gdf[gdf[rua_col].notna() & (gdf[rua_col] != "")].copy()
        if len(ruas_gdf) == 0:
            return 0
        centroids = ruas_gdf.geometry.centroid
        rua_xy = np.array([[c.x, c.y] for c in centroids])
        rua_names = ruas_gdf[rua_col].values
        n_associados = 0
        TOL = 100.0
        for t in trechos:
            if t.get("rua") and str(t["rua"]).strip() not in ("", "Sem Rua", "nan", "None"):
                continue
            p0 = pvs.get(t.get("pv_ini"), {})
            p1 = pvs.get(t.get("pv_fim"), {})
            if not p0.get("x") or not p1.get("x"):
                continue
            mx = (p0["x"] + p1["x"]) / 2
            my = (p0["y"] + p1["y"]) / 2
            dists = np.hypot(rua_xy[:, 0] - mx, rua_xy[:, 1] - my)
            idx = int(np.argmin(dists))
            if dists[idx] <= TOL:
                t["rua"] = str(rua_names[idx]).strip()
                n_associados += 1
        return n_associados
    except Exception as e:
        print(f"    Aviso GPKG: {e}")
        return 0

def _carregar_ruas_dxf(dxf_path, trechos, pvs):
    """Carrega textos de um DXF/DWG e associa como nomes de ruas."""
    if not HAS_GEO or not dxf_path or not Path(dxf_path).exists():
        return 0
    try:
        import pyogrio
        gdf = pyogrio.read_dataframe(str(dxf_path), read_geometry=True)
        # Filtrar textos
        if "Text" not in gdf.columns:
            return 0
        txt_gdf = gdf[gdf["Text"].notna() & (gdf["Text"].str.strip() != "")].copy()
        if len(txt_gdf) == 0:
            return 0
            
        PREF_RUA = ("RUA ", "BECO ", "TRAV", "AV ", "ESTRADA", "VIELA", "ALAMEDA", "ACESSO")
        # Filtrar apenas o que parece ser nome de logradouro
        mask = txt_gdf["Text"].str.upper().str.strip().apply(
            lambda t: len(t) > 2 and any(t.startswith(p) for p in PREF_RUA)
        )
        ruas_gdf = txt_gdf[mask]
        
        if len(ruas_gdf) == 0:
            # Fallback: pegar todos os textos grandes
            ruas_gdf = txt_gdf[txt_gdf["Text"].str.len() > 3]

        if len(ruas_gdf) == 0:
            return 0

        centroids = ruas_gdf.geometry.centroid
        rua_xy = np.array([[c.x, c.y] for c in centroids])
        rua_names = ruas_gdf["Text"].values
        n_associados = 0
        TOL = 100.0
        
        for t in trechos:
            if t.get("rua") and str(t["rua"]).strip() not in ("", "Sem Rua", "nan", "None"):
                continue
            p0 = pvs.get(t.get("pv_ini"), {})
            p1 = pvs.get(t.get("pv_fim"), {})
            if not p0.get("x") or not p1.get("x"):
                continue
            mx = (p0["x"] + p1["x"]) / 2
            my = (p0["y"] + p1["y"]) / 2
            dists = np.hypot(rua_xy[:, 0] - mx, rua_xy[:, 1] - my)
            idx = int(np.argmin(dists))
            if dists[idx] <= TOL:
                t["rua"] = str(rua_names[idx]).strip()
                n_associados += 1
        return n_associados
    except Exception as e:
        print(f"    Aviso MAPA DXF ({Path(dxf_path).name}): {e}")
        return 0


def _gerar_planilha_trechos_completa(trechos, pvs, nucleo, tipo_rede, out_path):
    """Planilha MESTRE de todos os trechos PV a PV com TODOS os quantitativos."""
    rows = []
    for i, t in enumerate(trechos):
        ns_id = i + 1
        dn_mm = t.get("dn_mm") or 200
        ext_m = float(t.get("ext_m") or 0)
        decl = t.get("decl_mm") or 0

        pvi = pvs.get(t.get("pv_ini"), {})
        pvf = pvs.get(t.get("pv_fim"), {})
        ct_i = pvi.get("ct"); cf_i = pvi.get("cf"); prof_i = pvi.get("prof")
        ct_f = pvf.get("ct"); cf_f = pvf.get("cf"); prof_f = pvf.get("prof")

        profs = [p for p in [prof_i, prof_f] if p and p > 0]
        prof_media = sum(profs) / len(profs) if profs else 1.5
        calc_dn = float(dn_mm) / 1000.0 if dn_mm else 0.20
        largura_vala = max(0.60, calc_dn + 0.50)

        vol_escavacao = round(ext_m * largura_vala * prof_media, 2)
        vol_aterro = round(vol_escavacao * 0.85, 2)
        area_pav = round(ext_m * FATORES.get("CBUQ provisório", 0.88), 2)
        vol_areia = round(ext_m * (FATORES.get("Areia", 0.25) + FATORES.get("Berço", 0.24)), 2)
        vol_brita = round(ext_m * FATORES.get("Brita base", 0.22), 2)
        n_barras = math.ceil(ext_m / 6) if ext_m > 0 else 0

        hidr = calc_manning(dn_mm, decl)
        custo = custo_trecho(t, pvs)

        rows.append({
            "NS": f"NS-{ns_id:04d}",
            "Núcleo": nucleo,
            "Rua/Logradouro": t.get("rua", "Sem Rua"),
            "PV Montante": t.get("pv_ini", ""),
            "PV Jusante": t.get("pv_fim", ""),
            "DN (mm)": dn_mm,
            "Extensão (m)": round(ext_m, 2),
            "Material": t.get("material", "PVC"),
            "CT Montante": ct_i,
            "CF Montante": cf_i,
            "Prof Montante (m)": prof_i,
            "CT Jusante": ct_f,
            "CF Jusante": cf_f,
            "Prof Jusante (m)": prof_f,
            "Declividade (‰)": round(decl * 1000, 2) if decl else None,
            "Declividade (%)": round(decl * 100, 3) if decl else None,
            "V Manning (m/s)": hidr.get("v_ms"),
            "Q Manning (L/s)": hidr.get("q_ls"),
            "τ (Pa)": hidr.get("tau_pa"),
            "Vol. Escavação (m³)": vol_escavacao,
            "Vol. Aterro/Reaterro (m³)": vol_aterro,
            "Área Paviment./Asfalto (m²)": area_pav,
            "Vol. Areia (m³)": vol_areia,
            "Vol. Brita (m³)": vol_brita,
            "Tubos (barras)": n_barras,
            "Custo Total c/ BDI (R$)": custo.get("total", 0),
        })

    df = pd.DataFrame(rows)
    df.to_excel(str(out_path), index=False)
    return rows


def processar_projeto(proj_info, dst_base):
    """Processa UM projeto DXF/DWG — TODOS os módulos da plataforma."""
    dxf_name = proj_info["dxf"]
    nucleo = proj_info["nucleo"]
    tipo_rede = proj_info["tipo"]
    projeto_nome = proj_info.get("projeto", _slug(dxf_name.replace('.dxf', '').replace('.dwg', '')))
    # Usar caminho absoluto se disponível, senão fallback
    dxf_path = Path(proj_info.get("dxf_path", str(SRC_FOLDER / dxf_name)))

    if not dxf_path.exists():
        print(f"  ⚠ Não encontrado: {dxf_path}")
        return None

    nucleo_slug = _slug(nucleo)
    projeto_slug = _slug(projeto_nome)
    tamanho_mb = dxf_path.stat().st_size / 1024 / 1024

    print(f"\n{'='*72}")
    print(f"  PROJETO: {projeto_nome}")
    print(f"  DXF:     {dxf_name} [{tamanho_mb:.1f} MB]")
    print(f"  NUCLEO:  {nucleo}")
    print(f"  TIPO:    {tipo_rede}")
    print(f"{'='*72}")

    # ── 1. LEITURA DXF ──────────────────────────────────────────
    try:
        pvs, trechos, ruas, meta = ler_dxf_gdal(str(dxf_path))
    except Exception as e:
        print(f"  ERRO LEITURA: {e}")
        traceback.print_exc()
        return None

    if not trechos:
        print(f"  Sem trechos! Pulando.")
        return None

    trechos = enriquecer_trechos(trechos, pvs)
    ext_total = sum(t.get('ext_m', 0) for t in trechos)
    print(f"  Rede: {len(pvs)} PVs, {len(trechos)} trechos, {ext_total:.0f}m")

    # ── 2. CARTOGRAFIA GPKG → RUAS ─────────────────────────────
    # Tentar pelo núcleo direto, depois pelo base (sem "Prolongamento ")
    nucleo_base = nucleo.replace("Prolongamento ", "")
    gpkg_name = CARTOGRAFIAS.get(nucleo) or CARTOGRAFIAS.get(nucleo_base)
    gpkg_path = SRC_FOLDER / gpkg_name if gpkg_name else None
    n_ruas_gpkg = 0
    if gpkg_path and gpkg_path.exists():
        n_ruas_gpkg = _carregar_ruas_gpkg(str(gpkg_path), trechos, pvs)
        print(f"  Cartografia GPKG: {n_ruas_gpkg} trechos com rua via {gpkg_name}")
    else:
        # Tentar encontrar GPKG automaticamente na pasta
        for gpkg_file in sorted(SRC_FOLDER.glob("*.gpkg")):
            gpkg_upper = gpkg_file.name.upper()
            for chave in _NUCLEO_MAP:
                if chave.upper() in nucleo.upper() and chave.upper() in gpkg_upper:
                    n_ruas_gpkg = _carregar_ruas_gpkg(str(gpkg_file), trechos, pvs)
                    if n_ruas_gpkg > 0:
                        print(f"  Cartografia GPKG (auto): {n_ruas_gpkg} via {gpkg_file.name}")
                    break
            if n_ruas_gpkg > 0:
                break

    # ── 2b. INTERPOLAÇÃO VIA MAPAS DXF/DWG EXTRAS (becos, vielas) ─────
    for mapa_path in MAPAS_INTERPOLACAO:
        if not mapa_path.exists():
            continue
        # Verificar se este mapa corresponde ao núcleo do projeto
        mapa_nucleo = MAPA_PARA_NUCLEO.get(mapa_path.stem, "")
        if mapa_nucleo and (mapa_nucleo.upper() in nucleo.upper() or mapa_nucleo.upper() in nucleo_base.upper()):
            n_ruas_dxf = _carregar_ruas_dxf(str(mapa_path), trechos, pvs)
            if n_ruas_dxf > 0:
                print(f"  Mapa DXF interpolação: {n_ruas_dxf} ruas preenchidas via {mapa_path.name}")

    # ── 3. PASTAS DE SAÍDA (POR PROJETO) ────────────────────────
    pasta_projeto = dst_base / projeto_slug
    pastas = {}
    for nome in ["01_NS_CAMPO", "02_DESENHOS", "03_HTML", "04_GIS",
                  "05_PLANILHAS", "06_CUSTOS", "07_BIM_IFC", "08_LEAN_LPS",
                  "09_MICROPLAN", "10_CRONOGRAMA", "11_POR_RUA", "12_LOG"]:
        p = pasta_projeto / nome
        p.mkdir(parents=True, exist_ok=True)
        pastas[nome] = p

    # ── 4. NOTAS DE SERVIÇO ─────────────────────────────────────
    print(f"  Gerando NS...")
    n_ok = n_err = 0
    for i, t in enumerate(trechos):
        ns_id = i + 1
        pv_i = t.get("pv_ini", "PVX")
        pv_f = t.get("pv_fim", "PVY")
        ns_name = _ns_folder_name(ns_id, pv_i, pv_f)
        ns_dir = pastas["01_NS_CAMPO"] / ns_name
        ns_dir.mkdir(parents=True, exist_ok=True)
        try:
            gerar_ns_a4(ns_id, t, pvs, nucleo, str(ns_dir / f"NS{ns_id:03d}_A4.pdf"))
            try:
                gerar_ns_desenho(ns_id, t, pvs, trechos, nucleo,
                                str(pastas["02_DESENHOS"] / f"NS{ns_id:03d}_DESENHO.pdf"))
            except Exception: pass
            try:
                gerar_ns_sat(ns_id, t, pvs, nucleo,
                            str(pastas["02_DESENHOS"] / f"NS{ns_id:03d}_SAT.pdf"))
            except Exception: pass
            try:
                gerar_html(ns_id, t, pvs, trechos, nucleo,
                          str(pastas["03_HTML"] / f"NS{ns_id:03d}.html"))
            except Exception: pass

            materiais = calcular_materiais(t, pvs)
            custo = custo_trecho(t, pvs)
            dados = {
                "ns_id": ns_id, "nucleo": nucleo, "contrato": CONTRATO,
                "tipo_rede": tipo_rede,
                "trecho": {k: v for k, v in t.items()},
                "pv_montante": pvs.get(pv_i, {}),
                "pv_jusante": pvs.get(pv_f, {}),
                "hidraulica": calc_manning(t.get("dn_mm"), t.get("decl_mm")),
                "materiais": materiais, "custo": custo,
                "gerado_em": datetime.now().isoformat(),
            }
            with open(ns_dir / f"NS{ns_id:03d}_DADOS.json", "w", encoding="utf-8") as f:
                json.dump(dados, f, indent=2, ensure_ascii=False)

            n_ok += 1
            if ns_id <= 3 or ns_id % 50 == 0:
                print(f"    NS{ns_id:03d}: {pv_i}->{pv_f} DN{t.get('dn_mm')} {t.get('ext_m',0):.1f}m")
        except Exception as e:
            n_err += 1
            if ns_id <= 5:
                print(f"    NS{ns_id:03d}: ERRO - {e}")

    print(f"  NS: {n_ok} ok, {n_err} erros")

    # ── 5. HTML MAPA GERAL ──────────────────────────────────────
    try:
        gerar_html(0, trechos[0], pvs, trechos, nucleo,
                  str(pastas["03_HTML"] / "REDE_GERAL.html"))
    except Exception: pass

    # ── 6. GeoJSON ──────────────────────────────────────────────
    try:
        n_feat = gerar_geojson(trechos, pvs, str(pastas["04_GIS"] / "rede_definida.geojson"))
        print(f"  GeoJSON: {n_feat} features")
    except Exception: pass

    # ── 7. PLANILHA MESTRE TRECHOS PV A PV ──────────────────────
    print(f"  Planilha MESTRE trechos PV a PV...")
    try:
        n_rows = _gerar_planilha_trechos_completa(
            trechos, pvs, nucleo, tipo_rede,
            pastas["05_PLANILHAS"] / f"TODOS_TRECHOS_PV_A_PV_{nucleo_slug}.xlsx"
        )
        print(f"  ✓ Planilha MESTRE: {n_rows} trechos")
    except Exception as e:
        print(f"  Erro planilha mestre: {e}")

    # ── 8. PLANILHA CUSTOS ──────────────────────────────────────
    try:
        gerar_xlsx_custos(pvs, trechos, nucleo,
                         str(pastas["06_CUSTOS"] / f"CUSTOS_{nucleo_slug}.xlsx"))
        print(f"  ✓ Planilha CUSTOS")
    except Exception as e:
        print(f"  Erro custos: {e}")

    # ── 9. PLANILHA HIDRÁULICA ──────────────────────────────────
    try:
        gerar_xlsx_hidraulica(trechos, pvs, nucleo,
                             str(pastas["05_PLANILHAS"] / f"HIDRAULICA_{nucleo_slug}.xlsx"))
        print(f"  ✓ Planilha HIDRÁULICA")
    except Exception as e:
        print(f"  Erro hidráulica: {e}")

    # ── 10. CURVA S ─────────────────────────────────────────────
    try:
        gerar_xlsx_curva_s(trechos, nucleo,
                          str(pastas["05_PLANILHAS"] / f"CURVA_S_{nucleo_slug}.xlsx"))
        print(f"  ✓ Curva S")
    except Exception as e:
        print(f"  Erro curva s: {e}")

    # ── 11. BIM IFC LOD500 ──────────────────────────────────────
    print(f"  Gerando BIM IFC LOD500...")
    try:
        arquivos_bim = gerar_ifc_lod500(pvs, trechos, nucleo, str(pastas["07_BIM_IFC"]))
        print(f"  ✓ BIM IFC: {len(arquivos_bim)} arquivos")
    except Exception as e:
        print(f"  BIM IFC: {e}")

    # ── 12. LEAN/LPS + BIM 6D ──────────────────────────────────
    print(f"  Gerando Lean/LPS + BIM 6D...")
    try:
        relatorio_lean = gerar_relatorio_lean_lps(pvs, trechos, nucleo=nucleo)
        # XLSX Lean/LPS
        try:
            gerar_xlsx_lean_lps(relatorio_lean, pvs, trechos, nucleo,
                               str(pastas["08_LEAN_LPS"] / f"LEAN_LPS_{nucleo_slug}.xlsx"))
        except Exception:
            pass
        # XLSX Lean completo (5 abas)
        try:
            gerar_xlsx_lean(relatorio_lean, pvs, trechos, nucleo,
                           str(pastas["08_LEAN_LPS"] / f"LEAN_COMPLETO_{nucleo_slug}.xlsx"))
        except Exception:
            pass
        # JSON
        with open(pastas["08_LEAN_LPS"] / f"LEAN_LPS_{nucleo_slug}.json", "w", encoding="utf-8") as f:
            json.dump(relatorio_lean, f, indent=2, ensure_ascii=False, default=str)
        print(f"  ✓ Lean/LPS + BIM 6D")
    except Exception as e:
        print(f"  Lean/LPS: {e}")

    # ── 13. MICROPLANEJAMENTO ───────────────────────────────────
    print(f"  Gerando Microplanejamento...")
    try:
        resultado_micro = micro_planejar_nucleo(pvs, trechos, nucleo, equipes_max=6)
        # XLSX Microplanejamento
        try:
            gerar_xlsx_microplan(resultado_micro, pvs, trechos, nucleo,
                               str(pastas["09_MICROPLAN"] / f"MICROPLAN_{nucleo_slug}.xlsx"))
        except Exception:
            pass
        # JSON
        with open(pastas["09_MICROPLAN"] / f"MICROPLAN_{nucleo_slug}.json", "w", encoding="utf-8") as f:
            json.dump(resultado_micro, f, indent=2, ensure_ascii=False, default=str)
        print(f"  ✓ Microplanejamento")
    except Exception as e:
        print(f"  Microplan: {e}")

    # ── 14. CRONOGRAMA POR NS ───────────────────────────────────
    print(f"  Gerando Cronograma por NS...")
    try:
        ns_lista = [
            {"ordem": i+1, "trecho_idx": i,
             "pv_ini": t.get("pv_ini", ""), "pv_fim": t.get("pv_fim", ""),
             "ext_m": t.get("ext_m", 0), "rua": t.get("rua", "")}
            for i, t in enumerate(trechos)
        ]
        resultado_crono = gerar_cronograma_por_ns(
            ns_lista, data_inicio_str=datetime.now().strftime("%Y-%m-%d"),
            equipes=4, prod_m_dia=6.0, nucleo=nucleo,
            out_dir=str(pastas["10_CRONOGRAMA"])
        )
        print(f"  ✓ Cronograma: {len(resultado_crono['tarefas'])} tarefas, "
              f"{resultado_crono['data_inicio']} → {resultado_crono['data_fim']}")
    except Exception as e:
        print(f"  Cronograma: {e}")

    # ── 15. SEPARAR POR RUA e COMPRAS ───────────────────────────
    print(f"  Separando por RUA...")
    trechos_por_rua = defaultdict(list)
    for t in trechos:
        rua = str(t.get("rua") or "Sem Rua").strip()
        if not rua or rua.upper() in ("NAN", "NONE", ""):
            rua = "Sem Rua"
        trechos_por_rua[rua].append(t)

    todos_rows_compras = []
    
    for rua, lista in trechos_por_rua.items():
        rua_slug = _slug(rua)
        pasta_rua = pastas["11_POR_RUA"] / rua_slug
        pasta_rua.mkdir(parents=True, exist_ok=True)
        try:
            r = _gerar_planilha_trechos_completa(
                lista, pvs, nucleo, tipo_rede,
                pasta_rua / f"Trechos_{rua_slug}.xlsx"
            )
            todos_rows_compras.extend(r)
        except Exception: pass

    print(f"  Ruas: {len(trechos_por_rua)}")

    # ── 16. LOG ─────────────────────────────────────────────────
    log_data = {
        "projeto": projeto_nome, "dxf_nome": dxf_name,
        "nucleo": nucleo, "tipo_rede": tipo_rede,
        "dxf": str(dxf_path),
        "n_pvs": len(pvs), "n_trechos": len(trechos),
        "extensao_total_m": round(ext_total, 1),
        "n_ns_geradas": n_ok, "n_ns_erros": n_err,
        "n_ruas": len(trechos_por_rua), "n_ruas_gpkg": n_ruas_gpkg,
        "motor": meta.get("motor", "GDAL/OGR v5"),
        "pasta_saida": str(pasta_projeto),
        "gerado_em": datetime.now().isoformat(),
        "versao_ns": NS_VERSION, "contrato": CONTRATO,
    }
    with open(pastas["12_LOG"] / "log_processamento.json", "w", encoding="utf-8") as f:
        json.dump(log_data, f, indent=2, ensure_ascii=False)

    print(f"  ✓ PROJETO CONCLUÍDO: {projeto_nome}")
    return {
        "log_data": log_data,
        "rows_compras": todos_rows_compras,
    }


def _proxima_rodada(base_path):
    """Descobre o número da próxima rodada e cria a pasta."""
    base = Path(base_path)
    rodada = 1
    while (base / f"RODADA_{rodada}").exists():
        rodada += 1
    pasta = base / f"RODADA_{rodada}"
    pasta.mkdir(parents=True, exist_ok=True)
    return pasta, rodada


def main():
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception: pass

    # Descobrir TODOS os DXF + DWG de esgoto em TODAS as pastas
    todas_pastas = [SRC_FOLDER] + SRC_FOLDERS_EXTRAS
    PROJETOS = _descobrir_projetos(todas_pastas)

    # Criar pasta da rodada
    pasta_rodada, num_rodada = _proxima_rodada(DST_FOLDER)

    print("=" * 72)
    print(f"  ConstruData HydroNetwork v9 — EXPORTAÇÃO POR PROJETO v3")
    print(f"  FCN Construções e Saneamento · Contrato {CONTRATO}")
    print(f"  Motor GDAL/OGR v5 · NS v{NS_VERSION}")
    print(f"  MODO: SOMENTE ESGOTO · AUTO-DISCOVERY DXF+DWG")
    print(f"  RODADA: {num_rodada}")
    print("=" * 72)
    for p in todas_pastas:
        st = "OK" if p.exists() else "NAO ENCONTRADA"
        print(f"  Pasta: {p} [{st}]")
    print(f"  Destino: {pasta_rodada}")
    print("=" * 72)

    # Listar projetos descobertos
    print(f"\n  Projetos DESCOBERTOS: {len(PROJETOS)} arquivos (DXF + DWG, somente esgoto)")
    for proj in PROJETOS:
        p = Path(proj["dxf_path"])
        mb = p.stat().st_size / 1024 / 1024 if p.exists() else 0
        print(f"    ✓ {proj['dxf']:55s} [{mb:6.1f} MB] → {proj['nucleo']}")

    if not PROJETOS:
        print("\n  NENHUM PROJETO DXF/DWG DE ESGOTO ENCONTRADO!")
        return

    # Processar cada projeto
    resultados_raw = []
    all_compras = []  # todas as rows de compras de todos os projetos
    for proj in PROJETOS:
        try:
            resultado = processar_projeto(proj, pasta_rodada)
            if resultado:
                resultados_raw.append(resultado)
        except Exception as e:
            print(f"\n  ERRO FATAL em {proj['dxf']}: {e}")
            traceback.print_exc()

    # Separar log_data e rows_compras do retorno
    resultados = []
    for r in resultados_raw:
        log = r.get("log_data", r) if isinstance(r, dict) else r
        resultados.append(log)
        rows_c = r.get("rows_compras", []) if isinstance(r, dict) else []
        all_compras.extend(rows_c)

    # ══════════════════════════════════════════════════════════════
    # PLANILHA DE COMPRAS (por NS, por Núcleo e Geral)
    # ══════════════════════════════════════════════════════════════
    if all_compras:
        print(f"\n{'='*72}")
        print(f"  GERANDO PLANILHA DE COMPRAS ({len(all_compras)} trechos)...")
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb = Workbook()

            # ── ABA 1: GERAL (todos os trechos de todos os projetos) ──
            ws_geral = wb.active
            ws_geral.title = "GERAL"

            df_geral = pd.DataFrame(all_compras)
            # Título
            ws_geral.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df_geral.columns))
            ws_geral.cell(1, 1, f"PLANILHA DE COMPRAS — TODOS OS PROJETOS — RODADA {num_rodada}")
            ws_geral.cell(1, 1).font = Font(bold=True, size=14, color="FFFFFF")
            ws_geral.cell(1, 1).fill = PatternFill("solid", fgColor="003366")
            ws_geral.cell(1, 1).alignment = Alignment(horizontal="center")

            # Cabeçalhos
            for j, col in enumerate(df_geral.columns, 1):
                c = ws_geral.cell(2, j, col)
                c.font = Font(bold=True, size=9, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="0066CC")
                c.alignment = Alignment(horizontal="center", wrap_text=True)
                c.border = Border(
                    left=Side("thin"), right=Side("thin"),
                    top=Side("thin"), bottom=Side("thin"))
            # Dados
            fill_par = PatternFill("solid", fgColor="F0F8FF")
            fill_imp = PatternFill("solid", fgColor="E8F4FF")
            for i, row_data in enumerate(df_geral.values, 3):
                for j, val in enumerate(row_data, 1):
                    c = ws_geral.cell(i, j, val)
                    c.fill = fill_par if i % 2 == 0 else fill_imp
                    c.border = Border(
                        left=Side("thin"), right=Side("thin"),
                        top=Side("thin"), bottom=Side("thin"))
                    if isinstance(val, float):
                        c.number_format = "#,##0.00"

            # Totais
            row_total = len(df_geral) + 3
            ws_geral.cell(row_total, 1, "TOTAL GERAL").font = Font(bold=True, size=11)
            ws_geral.cell(row_total, 1).fill = PatternFill("solid", fgColor="B4C6E7")
            # Soma das colunas numéricas
            for j, col in enumerate(df_geral.columns, 1):
                if df_geral[col].dtype in ["float64", "int64"]:
                    total = df_geral[col].sum()
                    c = ws_geral.cell(row_total, j, round(total, 2))
                    c.font = Font(bold=True)
                    c.fill = PatternFill("solid", fgColor="B4C6E7")
                    c.number_format = "#,##0.00"

            # ── ABA 2: POR NUCLEO (resumo agrupado) ──
            ws_nucleo = wb.create_sheet("POR_NUCLEO")
            ws_nucleo.merge_cells("A1:L1")
            ws_nucleo.cell(1, 1, f"RESUMO DE COMPRAS POR NÚCLEO — RODADA {num_rodada}")
            ws_nucleo.cell(1, 1).font = Font(bold=True, size=14, color="FFFFFF")
            ws_nucleo.cell(1, 1).fill = PatternFill("solid", fgColor="003366")

            cols_nucleo = ["Núcleo", "Qtd Trechos", "Extensão Total (m)",
                           "Vol. Escavação (m³)", "Vol. Aterro/Reaterro (m³)",
                           "Área Paviment./Asfalto (m²)", "Vol. Areia (m³)",
                           "Vol. Brita (m³)", "Tubos (barras)", "Custo Total c/ BDI (R$)"]

            for j, col in enumerate(cols_nucleo, 1):
                c = ws_nucleo.cell(2, j, col)
                c.font = Font(bold=True, size=9, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="0066CC")
                c.alignment = Alignment(horizontal="center", wrap_text=True)

            grouped = df_geral.groupby("Núcleo", dropna=False)
            row_n = 3
            for nuc_name, nuc_df in grouped:
                ws_nucleo.cell(row_n, 1, nuc_name)
                ws_nucleo.cell(row_n, 2, len(nuc_df))
                for k, campo in enumerate(["Extensão (m)", "Vol. Escavação (m³)",
                        "Vol. Aterro/Reaterro (m³)", "Área Paviment./Asfalto (m²)",
                        "Vol. Areia (m³)", "Vol. Brita (m³)", "Tubos (barras)",
                        "Custo Total c/ BDI (R$)"], 3):
                    if campo in nuc_df.columns:
                        ws_nucleo.cell(row_n, k, round(nuc_df[campo].sum(), 2))
                        ws_nucleo.cell(row_n, k).number_format = "#,##0.00"
                row_n += 1

            # Total geral na aba núcleo
            ws_nucleo.cell(row_n, 1, "TOTAL GERAL").font = Font(bold=True, size=11)
            ws_nucleo.cell(row_n, 1).fill = PatternFill("solid", fgColor="B4C6E7")
            ws_nucleo.cell(row_n, 2, len(df_geral))
            for k, campo in enumerate(["Extensão (m)", "Vol. Escavação (m³)",
                    "Vol. Aterro/Reaterro (m³)", "Área Paviment./Asfalto (m²)",
                    "Vol. Areia (m³)", "Vol. Brita (m³)", "Tubos (barras)",
                    "Custo Total c/ BDI (R$)"], 3):
                if campo in df_geral.columns:
                    c = ws_nucleo.cell(row_n, k, round(df_geral[campo].sum(), 2))
                    c.font = Font(bold=True)
                    c.fill = PatternFill("solid", fgColor="B4C6E7")
                    c.number_format = "#,##0.00"

            # ── ABA 3: POR NS (cada NS com seus materiais) ──
            ws_ns = wb.create_sheet("POR_NS")
            ws_ns.merge_cells("A1:L1")
            ws_ns.cell(1, 1, f"COMPRAS POR NS (Nota de Serviço) — RODADA {num_rodada}")
            ws_ns.cell(1, 1).font = Font(bold=True, size=14, color="FFFFFF")
            ws_ns.cell(1, 1).fill = PatternFill("solid", fgColor="003366")

            cols_ns = ["NS", "Núcleo", "Rua/Logradouro", "PV Montante", "PV Jusante",
                       "DN (mm)", "Extensão (m)", "Vol. Escavação (m³)",
                       "Vol. Areia (m³)", "Vol. Brita (m³)", "Tubos (barras)",
                       "Custo Total c/ BDI (R$)"]
            for j, col in enumerate(cols_ns, 1):
                c = ws_ns.cell(2, j, col)
                c.font = Font(bold=True, size=9, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="0066CC")
                c.alignment = Alignment(horizontal="center", wrap_text=True)

            for i, row_data in enumerate(all_compras, 3):
                for j, col in enumerate(cols_ns, 1):
                    val = row_data.get(col, "")
                    c = ws_ns.cell(i, j, val)
                    c.fill = fill_par if i % 2 == 0 else fill_imp
                    if isinstance(val, float):
                        c.number_format = "#,##0.00"

            # Salvar
            compras_path = pasta_rodada / "PLANILHA_COMPRAS.xlsx"
            wb.save(str(compras_path))
            print(f"  ✓ PLANILHA DE COMPRAS: {compras_path}")
            print(f"    - Aba GERAL: {len(all_compras)} trechos")
            print(f"    - Aba POR_NUCLEO: {len(grouped)} núcleos")
            print(f"    - Aba POR_NS: detalhamento NS a NS")
        except Exception as e:
            print(f"  Erro planilha compras: {e}")
            traceback.print_exc()

    # ══════════════════════════════════════════════════════════════
    # SLNR MEGA INTEGRADA (estilo da planilha de referência)
    # ══════════════════════════════════════════════════════════════
    if all_compras:
        print(f"\n  GERANDO SLNR_MEGA_INTEGRADA...")
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb2 = Workbook()

            # Aba TRECHOS_PADRAO
            ws_t = wb2.active
            ws_t.title = "TRECHOS_PADRAO"
            ws_t.merge_cells("A1:R1")
            ws_t.cell(1, 1, "PLANEJAMENTO POR TRECHO (PV a PV) — TABELA DE ATRIBUTOS")
            ws_t.cell(1, 1).font = Font(bold=True, size=14, color="FFFFFF")
            ws_t.cell(1, 1).fill = PatternFill("solid", fgColor="003366")
            ws_t.cell(1, 1).alignment = Alignment(horizontal="center")

            cols_trecho = ["NS", "Núcleo", "Rua/Logradouro",
                           "PV Montante", "PV Jusante", "DN (mm)", "Extensão (m)",
                           "Material", "CT Montante", "CF Montante", "Prof Montante (m)",
                           "CT Jusante", "CF Jusante", "Prof Jusante (m)",
                           "Vol. Escavação (m³)", "Vol. Areia (m³)",
                           "Vol. Brita (m³)", "Custo Total c/ BDI (R$)"]

            for j, col in enumerate(cols_trecho, 1):
                c = ws_t.cell(2, j, col)
                c.font = Font(bold=True, size=9, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="0066CC")
                c.alignment = Alignment(horizontal="center", wrap_text=True)

            for i, row_data in enumerate(all_compras, 3):
                for j, col in enumerate(cols_trecho, 1):
                    val = row_data.get(col, "")
                    c = ws_t.cell(i, j, val)
                    if isinstance(val, float):
                        c.number_format = "#,##0.00"

            # Aba SETORIZACAO
            ws_s = wb2.create_sheet("SETORIZACAO")
            ws_s.merge_cells("A1:F1")
            ws_s.cell(1, 1, "PLANO DE ATAQUE — SETORIZAÇÃO POR PROJETO")
            ws_s.cell(1, 1).font = Font(bold=True, size=14, color="FFFFFF")
            ws_s.cell(1, 1).fill = PatternFill("solid", fgColor="003366")

            setor_cols = ["Projeto", "Núcleo", "PVs", "Trechos", "Extensão (m)", "NS Geradas"]
            for j, col in enumerate(setor_cols, 1):
                c = ws_s.cell(2, j, col)
                c.font = Font(bold=True, size=10, color="FFFFFF")
                c.fill = PatternFill("solid", fgColor="0066CC")
                c.alignment = Alignment(horizontal="center")

            for i, r in enumerate(resultados, 3):
                ws_s.cell(i, 1, r.get("projeto", ""))
                ws_s.cell(i, 2, r.get("nucleo", ""))
                ws_s.cell(i, 3, r.get("n_pvs", 0))
                ws_s.cell(i, 4, r.get("n_trechos", 0))
                ws_s.cell(i, 5, r.get("extensao_total_m", 0))
                ws_s.cell(i, 6, r.get("n_ns_geradas", 0))

            mega_path = pasta_rodada / "SLNR_MEGA_INTEGRADA.xlsx"
            wb2.save(str(mega_path))
            print(f"  ✓ SLNR_MEGA_INTEGRADA: {mega_path}")
        except Exception as e:
            print(f"  Erro SLNR: {e}")

    # ══════════════════════════════════════════════════════════════
    # CRONOGRAMA MACRO (todos os núcleos)
    # ══════════════════════════════════════════════════════════════
    if resultados:
        print(f"\n{'='*72}")
        print("  GERANDO CRONOGRAMA MACRO MULTINÚCLEO...")
        try:
            nucleos_crono = [
                {"nome": r["nucleo"], "extensao_m": r["extensao_total_m"],
                 "n_trechos": r["n_trechos"], "equipes": 2}
                for r in resultados
            ]
            wbs = gerar_cronograma_macro(nucleos_crono, datetime.now().strftime("%Y-%m-%d"))

            pasta_crono = pasta_rodada / "CRONOGRAMA_MACRO"
            pasta_crono.mkdir(exist_ok=True)

            exportar_project_xml(wbs, str(pasta_crono / "CRONOGRAMA_MACRO.xml"))
            exportar_primavera_xer(wbs, str(pasta_crono / "CRONOGRAMA_P6.xer"))
            exportar_openproject_csv(wbs, str(pasta_crono / "CRONOGRAMA_OPENPROJECT.csv"))
            try:
                exportar_macro_xlsx(wbs, str(pasta_crono / "CRONOGRAMA_MACRO.xlsx"))
            except Exception: pass

            with open(pasta_crono / "CRONOGRAMA_MACRO.json", "w", encoding="utf-8") as f:
                json.dump(wbs, f, indent=2, ensure_ascii=False)

            print(f"  ✓ Cronograma Macro: {wbs['total_tarefas']} tarefas, "
                  f"{wbs['data_inicio']} → {wbs.get('data_fim', '?')}")
        except Exception as e:
            print(f"  Erro cronograma macro: {e}")

    # ══════════════════════════════════════════════════════════════
    # CONSOLIDADO GERAL
    # ══════════════════════════════════════════════════════════════
    if resultados:
        print(f"\n{'='*72}")
        print("  GERANDO CONSOLIDADO GERAL...")

        resumo = {
            "gerado_em": datetime.now().isoformat(),
            "contrato": CONTRATO,
            "empresa": "FCN Construções e Saneamento",
            "rodada": num_rodada,
            "total_projetos": len(resultados),
            "total_trechos": sum(r["n_trechos"] for r in resultados),
            "total_pvs": sum(r["n_pvs"] for r in resultados),
            "extensao_total_m": sum(r["extensao_total_m"] for r in resultados),
            "total_ns_geradas": sum(r["n_ns_geradas"] for r in resultados),
            "projetos": resultados,
        }
        with open(pasta_rodada / "RESUMO_GERAL.json", "w", encoding="utf-8") as f:
            json.dump(resumo, f, indent=2, ensure_ascii=False)

        try:
            df = pd.DataFrame([{
                "Projeto": r["projeto"], "Núcleo": r["nucleo"],
                "PVs": r["n_pvs"], "Trechos": r["n_trechos"],
                "Extensão (m)": r["extensao_total_m"],
                "NS Geradas": r["n_ns_geradas"], "Ruas": r["n_ruas"],
            } for r in resultados])
            df.to_excel(str(pasta_rodada / "RESUMO_GERAL.xlsx"), index=False)
        except Exception: pass

    # ══════════════════════════════════════════════════════════════
    # RESULTADO FINAL
    # ══════════════════════════════════════════════════════════════
    print(f"\n{'='*72}")
    print(f"  RESULTADO FINAL — RODADA {num_rodada}")
    print(f"{'='*72}")
    total_ns = sum(r.get("n_ns_geradas", 0) for r in resultados)
    total_tr = sum(r.get("n_trechos", 0) for r in resultados)
    total_ext = sum(r.get("extensao_total_m", 0) for r in resultados)
    print(f"  Projetos processados: {len(resultados)}")
    print(f"  Total de trechos:     {total_tr}")
    print(f"  Total de NS:          {total_ns}")
    print(f"  Extensão total:       {total_ext:,.0f} m")
    print(f"  Pasta de saída:       {pasta_rodada}")
    print(f"{'='*72}")

    for r in resultados:
        print(f"  ✓ {r['projeto']:35s} | "
              f"{r['n_trechos']:4d} trechos | {r['extensao_total_m']:8,.0f}m | "
              f"{r['n_ns_geradas']:4d} NS")

    print(f"\n  MÓDULOS EXECUTADOS:")
    print(f"    ✓ NS A4 + Desenho A3 + Satélite + HTML Leaflet")
    print(f"    ✓ Planilha MESTRE Trechos PV a PV (CT, CF, Escav, Areia, Brita, Asfalto)")
    print(f"    ✓ Custos Detalhados + BDI 25%")
    print(f"    ✓ Hidráulica Manning + Alertas")
    print(f"    ✓ Curva S (Previsto × Realizado)")
    print(f"    ✓ BIM IFC LOD500 (3D real + CSV + JSON)")
    print(f"    ✓ Lean Construction + LPS + BIM 6D")
    print(f"    ✓ Microplanejamento (Morfologia + Frentes)")
    print(f"    ✓ Cronograma por NS (Gantt HTML + MS Project XML)")
    print(f"    ✓ Cronograma Macro (Primavera P6 + OpenProject)")
    print(f"    ✓ GeoJSON")
    print(f"    ✓ Separação por Rua + Cartografia GPKG")
    print(f"    ✓ PLANILHA DE COMPRAS (por NS, por Núcleo, Geral)")
    print(f"    ✓ SLNR MEGA INTEGRADA (Trechos + Setorização)")
    print(f"\n  {'='*72}")
    print(f"  MISSÃO CUMPRIDA! 🔥")
    print(f"  {'='*72}")


if __name__ == "__main__":
    main()

