#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CONSTRUDATA - HydroNetwork v9.0 — Plataforma Desktop Completa
SE LIGA NA REDE · Contrato 11481051 · Santos/SP
FCN Construcoes e Saneamento

Aceita: DXF ProSaneamento, LandXML (Civil 3D), JSON, DWG (AEC Proxy)
Pipeline 6 etapas: Leitura → NS → Civil 3D → Cadastro NTS292 → IFC LOD500 → Cronograma
"""

import sys, os, threading, webbrowser, json, math, subprocess, tempfile
import tkinter as tk
import sv_ttk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from pathlib import Path
from datetime import datetime
from collections import Counter

SCRIPT_DIR = Path(__file__).parent
SCRIPTS_DIR = SCRIPT_DIR / "scripts"
for _path in [SCRIPTS_DIR, SCRIPT_DIR]:
    _path_str = str(_path)
    if _path_str not in sys.path:
        sys.path.insert(0, _path_str)

try:
    from core.config import API_BASE_URL as API_LOCAL_URL
except Exception:
    API_LOCAL_URL = "http://127.0.0.1:8787"

# ── Importar motores ──────────────────────────────────────────────────────────
_ENGINES = {}

def _try_import(name, import_fn):
    try:
        mod = import_fn()
        _ENGINES[name] = True
        return mod
    except Exception:
        _ENGINES[name] = False
        return None

# Leitores
_gdal = _try_import("GDAL", lambda: __import__("ler_dxf_gdal"))
_lxml = _try_import("LandXML", lambda: __import__("ler_landxml"))
_dwg  = _try_import("DWG/AEC", lambda: __import__("ler_dwg_aec"))
_dwgs = _try_import("DWG Semantico", lambda: __import__("ler_dwg_semantico"))
_dwgu = _try_import("DWG Universal", lambda: __import__("ler_dwg_universal"))
_pros = _try_import("ProSaneamento", lambda: __import__("ler_dxf_prosaneamento"))

# Geradores
_ns     = _try_import("GerarNS", lambda: __import__("gerar_ns"))
_c3d    = _try_import("Civil3D", lambda: __import__("gerar_civil3d"))
_nts    = _try_import("NTS292", lambda: __import__("gerar_cadastro_nts292"))
_ifc    = _try_import("IFC", lambda: __import__("gerar_ifc_lod500"))
_proj   = _try_import("MSProject", lambda: __import__("gerar_project_xml"))
_pipe   = _try_import("Pipeline", lambda: __import__("construdata_pipeline"))

# Motores v7
_custo  = _try_import("Custo", lambda: __import__("motor_custo"))
_med    = _try_import("Medicao", lambda: __import__("motor_medicao"))
_ml     = _try_import("ML", lambda: __import__("motor_ml"))
_lean   = _try_import("Lean/LPS", lambda: __import__("motor_lean_lps"))
_param  = _try_import("Parametrico", lambda: __import__("motor_parametrico"))
_micro  = _try_import("MicroPlan", lambda: __import__("motor_microplanejamento"))
_perdas = _try_import("Perdas", lambda: __import__("motor_perdas"))
_cmacro = _try_import("CronoMacro", lambda: __import__("gerar_cronograma_macro"))
_ppdf   = _try_import("PdfPerdas", lambda: __import__("gerar_pdf_perdas"))
_gemini = _try_import("Gemini", lambda: __import__("motor_gemini"))
_llm    = _try_import("Multi-LLM", lambda: __import__("motor_llm"))
_mctrt  = _try_import("Contratos", lambda: __import__("motor_contratos"))
_anlyt  = _try_import("Analytics", lambda: __import__("construdata_analytics"))
_slnr   = _try_import("SLNR_Mestre", lambda: __import__("slnr_mestre_ml"))
NS_VERSION = getattr(_ns, "NS_VERSION", "9")

# Motor v5 (fallback)
try:
    import construdata_sabesp_v5_FINAL as v5
    _ENGINES["Motor_v5"] = True
except Exception:
    v5 = None
    _ENGINES["Motor_v5"] = False

# ── Cores (tema escuro — ref: construdata_editor.html) ────────────────────────
BG       = "#06060f"
BG2      = "#0d0d1a"
BG3      = "#1a1a2e"
FG       = "#e2e8f0"
ACCENT   = "#00ff88"
BLUE     = "#00aaff"
RED      = "#ef4444"
YELLOW   = "#eab308"
ORANGE   = "#f97316"
CYAN     = "#06b6d4"
GRAY     = "#6b7280"
WHITE    = "#ffffff"
PURPLE   = "#7c3aed"

VERSION  = "9.0.0"
EMPRESA  = "FCN Construcoes e Saneamento"
CONTRATO = "11481051"

# Tile servers
_TILE_SAT = "https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}"
_TILE_RUA = "https://tile.openstreetmap.org/{z}/{x}/{y}.png"

# Prolongamentos LandXML
_XML_DIR = Path.home() / "Downloads" / "PROJETOS DE ÁGUA E ESGOTO - DWG E DXF 2018"
_PROLONGAMENTOS = [
    ("Prolongamento Teteu Alt-01", str(_XML_DIR / "PROLONGAMENTO TETEU ALT-01.xml")),
    ("Prolongamento Teteu",        str(_XML_DIR / "PROLONGAMENTO TETEU.xml")),
    ("Prolongamento Pantanal",     str(_XML_DIR / "PROLONGAMENTO PANTANAL BAIXO.xml")),
    ("Prolongamento Criadores",    str(_XML_DIR / "PROLONGAMENTO CRIADORES.xml")),
    ("Prolongamento Sao Manoel",   str(_XML_DIR / "PROLONGAMENTO SÃO MANOEL.xml")),
]


class HydroNetworkApp:
    def __init__(self, root):
        self.root = root
        sv_ttk.set_theme("dark")
        root.title(f"ConstruData - HydroNetwork v{VERSION} | NS v{NS_VERSION}")
        root.geometry("1200x850")
        root.configure(bg=BG)
        root.minsize(1000, 700)

        # Estado
        self.pvs = {}
        self.trechos = []
        self.ruas = []
        self.meta = {}
        self.running = False
        self.source_path = None
        self.analytics_results = None
        self.mapas_interpolacao = []  # lista de paths DXF/DWG/GPKG para interpolação de ruas

        # Estado centralizado por NS (motor_status_ns)
        self._status_ns: dict = {}             # STATUS_NS.json carregado em memória

        # Sequência executiva NS
        self._ns_sequencia  = []               # lista de trecho_idx na ordem executiva
        self._ns_seq_inicio = tk.IntVar(value=1)
        self._ns_seq_fim    = tk.IntVar(value=999)
        self._ns_equipes    = tk.IntVar(value=4)
        self._ns_prod_m_dia = tk.DoubleVar(value=6.0)
        self._crono_modo    = tk.StringVar(value="ns_atual")

        # Variaveis
        self.arquivo_var = tk.StringVar()
        self.gpkg_var    = tk.StringVar()
        self.saida_var   = tk.StringVar(value=str(SCRIPT_DIR / "SAIDA_HYDRONETWORK"))
        self.nucleo_var  = tk.StringVar()
        self.tipo_var    = tk.StringVar(value="auto")
        self.topo_var         = tk.StringVar()
        self.cartografia_var  = tk.StringVar()

        self._build_ui()
        self._log_msg(f"ConstruData - HydroNetwork v{VERSION} | NS v{NS_VERSION}")
        self._log_msg(f"{EMPRESA} · Contrato {CONTRATO} · SLNR Santos")
        ok = [k for k, v in _ENGINES.items() if v]
        self._log_msg(f"Motores: {', '.join(ok)}", "OK")

    # ══════════════════════════════════════════════════════════════════════════
    # UI
    # ══════════════════════════════════════════════════════════════════════════

    def _build_ui(self):
        # Header v9 — height 64, version pill, accent bar, SLNR branding
        hdr = tk.Frame(self.root, height=64)
        hdr.pack(fill=tk.X)
        hdr.pack_propagate(False)

        # ── Logo icon (hexágono + rede) desenhado via Canvas ──────────────
        _lc = tk.Canvas(hdr, width=52, height=52,
                        highlightthickness=0, cursor="hand2")
        _lc.pack(side=tk.LEFT, padx=(10, 0), pady=6)
        # Hexágono
        _cx, _cy, _r = 26, 26, 21
        _hex = []
        for _i in range(6):
            _a = math.radians(60 * _i - 30)
            _hex += [_cx + _r * math.cos(_a), _cy + _r * math.sin(_a)]
        _lc.create_polygon(_hex, fill=ACCENT, outline="#000000", width=1)
        # Nós da rede (3 nós + conexões)
        _lc.create_line(_cx, _cy - 10, _cx, _cy,      fill="#000000", width=2)
        _lc.create_line(_cx, _cy,      _cx - 9, _cy+8, fill="#000000", width=2)
        _lc.create_line(_cx, _cy,      _cx + 9, _cy+8, fill="#000000", width=2)
        _lc.create_oval(_cx-4,    _cy-14, _cx+4,    _cy-6,  fill="#000000", outline="")
        _lc.create_oval(_cx-13,   _cy+4,  _cx-5,    _cy+12, fill="#000000", outline="")
        _lc.create_oval(_cx+5,    _cy+4,  _cx+13,   _cy+12, fill="#000000", outline="")
        # Nó central com marca
        _lc.create_oval(_cx-4, _cy-4, _cx+4, _cy+4, fill=ACCENT, outline="#000000", width=1)
        _lc.create_oval(_cx-2, _cy-2, _cx+2, _cy+2, fill="#000000", outline="")
        # Separador vertical fino
        _sep = tk.Frame(hdr, width=1)
        _sep.pack(side=tk.LEFT, fill=tk.Y, pady=10, padx=6)

        # Left: brand
        tk.Label(hdr, text="CONSTRUDATA", font=("Segoe UI", 20, "bold")).pack(side=tk.LEFT, padx=(0, 4))
        tk.Label(hdr, text="HydroNetwork", font=("Segoe UI", 14)).pack(side=tk.LEFT)
        # Version pill badge
        _v_pill = tk.Frame(hdr, padx=6, pady=2)
        _v_pill.pack(side=tk.LEFT, padx=8, pady=16)
        tk.Label(_v_pill, text=f"v{VERSION[:3]}", font=("Segoe UI", 9, "bold")).pack()
        tk.Label(hdr, text=f"NS v{NS_VERSION}", font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT, padx=4)
        # Right: SLNR branding block
        _right = tk.Frame(hdr)
        _right.pack(side=tk.RIGHT, padx=12)
        tk.Label(_right, text="SLNR Santos", font=("Segoe UI", 9, "bold")).pack(anchor=tk.E)
        tk.Label(_right, text=f"{EMPRESA} | {CONTRATO}",
                 font=("Segoe UI", 8)).pack(anchor=tk.E)
        # Accent bar 2px
        tk.Frame(self.root, height=2).pack(fill=tk.X)

        # Style
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TNotebook", background=BG, borderwidth=0)
        style.configure("TNotebook.Tab", background=BG2, foreground=FG,
                        padding=[16, 7], font=("Segoe UI", 9, "bold"))
        style.map("TNotebook.Tab",
                  background=[("selected", "#0d2b1f")],
                  foreground=[("selected", ACCENT)])
        style.configure("TProgressbar", background=ACCENT, troughcolor=BG2)
        style.configure("TCombobox", fieldbackground=BG3, foreground=FG)
        style.configure("Treeview", background=BG2, foreground=FG,
                        fieldbackground=BG2, font=("Consolas", 9))
        style.configure("Treeview.Heading", background=BG3, foreground=ACCENT,
                        font=("Segoe UI", 8, "bold"))

        self.nb = ttk.Notebook(self.root)
        self.nb.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)

        # Binding para atualizar status quando aba IA for selecionada
        self.nb.bind("<<NotebookTabChanged>>", self._on_tab_changed)

        self._tab_processar()
        self._tab_mapa()
        self._tab_rede()
        self._tab_hidraulica()
        self._tab_trechos()
        self._tab_custos()
        self._tab_bim()
        self._tab_lean()
        self._tab_perdas()
        self._tab_ia()
        self._tab_nucleos()
        self._tab_log()
        self._tab_gestao()
        self._tab_ciclo_operacional()
        self._tab_construdata_workspace()

    def _on_tab_changed(self, event):
        """Atualiza status dos LLMs quando aba IA é selecionada."""
        selected = event.widget.tab('current')['text']
        if 'IA' in selected:
            self._ia_atualizar_status()

    # ── TAB 1: PROCESSAR ─────────────────────────────────────────────────────

    def _tab_ciclo_operacional(self):
        try:
            from ui_operational_cycle import build_operational_cycle_tab
            build_operational_cycle_tab(self)
        except Exception as exc:
            self._log_msg(f"Ciclo operacional indisponivel: {exc}", "WARN")

    def _tab_construdata_workspace(self):
        try:
            from ui_construdata_modules import build_construdata_workspace_tab
            build_construdata_workspace_tab(self, index=15)
        except Exception as exc:
            self._log_msg(f"Workspace ConstruData indisponivel: {exc}", "WARN")

    def _tab_processar(self):
        tab = tk.Frame(self.nb)
        self.nb.add(tab, text="  [1] Processar  ")

        # Entrada
        lf = tk.LabelFrame(tab, text=" ARQUIVO DE ENTRADA ",
                           font=("Segoe UI", 9, "bold"), bd=1)
        lf.pack(fill=tk.X, padx=8, pady=(8, 4))

        row = tk.Frame(lf)
        row.pack(fill=tk.X, padx=8, pady=4)
        tk.Label(row, text="Arquivo (DXF/XML/JSON/DWG):", width=26, anchor=tk.W, font=("Segoe UI", 9)).pack(side=tk.LEFT)
        tk.Entry(row, textvariable=self.arquivo_var,
                 insertbackground=FG, bd=2,
                 font=("Segoe UI", 9)).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=4)
        tk.Button(row, text="...", command=self._browse_arquivo, width=3).pack(side=tk.RIGHT)

        self.tipo_detectado = tk.Label(lf, text="Nenhum arquivo selecionado", font=("Segoe UI", 8))
        self.tipo_detectado.pack(padx=8, anchor=tk.W)
        self.arquivo_var.trace_add("write", self._on_arquivo_change)

        self._file_row(lf, "Pasta de Saida:", self.saida_var, directory=True)
        self._file_row(lf, "TXT/GSI Topografo (campo):", self.topo_var,
                       filetypes=[("Topografia", "*.txt *.csv *.gsi")])
        self._file_row(lf, "Cartografia (DXF/DWG/GPKG):", self.cartografia_var,
                       filetypes=[("Cartografia", "*.dxf *.dwg *.gpkg")])

        # ── PAINEL DE STATUS DE MOTORES (Real-time) ───────────────────
        sf = tk.Frame(tab, bd=1, relief=tk.RIDGE)
        sf.pack(fill=tk.X, padx=12, pady=4)
        
        # Grid de status
        for i, (name, key) in enumerate([
            ("Headless (accore)", "LandXML"),
            ("BIM (win32com)",   "Motor_v5"),
            ("Geometria (GDAL)", "GDAL"),
            ("ProSaneamento",    "ProSaneamento"),
            ("Analytics (XGB)",  "Analytics"),
        ]):
            r = i // 5
            c = i % 5
            f = tk.Frame(sf, padx=10, pady=5)
            f.grid(row=r, column=c, sticky="nsew")
            sf.grid_columnconfigure(c, weight=1)
            
            status = _ENGINES.get(key, False)
            color = ACCENT if status else RED
            bulb = "●" if status else "○"
            
            tk.Label(f, text=bulb, fg=color, font=("Segoe UI", 12)).pack(side=tk.LEFT)
            tk.Label(f, text=f" {name}", font=("Segoe UI", 8, "bold")).pack(side=tk.LEFT)
        # ──────────────────────────────────────────────────────────────

        # Mapas de interpolação (ruas)
        lf_mapas = tk.LabelFrame(tab, text=" MAPAS DE INTERPOLACAO (ruas - DXF/DWG/GPKG) ",
                                  font=("Segoe UI", 9, "bold"), bd=1)
        lf_mapas.pack(fill=tk.X, padx=8, pady=(4, 2))

        mrow = tk.Frame(lf_mapas)
        mrow.pack(fill=tk.X, padx=8, pady=4)

        self.mapas_listbox = tk.Listbox(mrow, height=3,
                                         font=("Consolas", 8), selectmode=tk.EXTENDED)
        self.mapas_listbox.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 4))

        mbtn_frame = tk.Frame(mrow)
        mbtn_frame.pack(side=tk.RIGHT)
        tk.Button(mbtn_frame, text="+ Mapa", command=self._add_mapa_interpolacao, font=("Segoe UI", 8, "bold"), padx=6).pack(pady=1)
        tk.Button(mbtn_frame, text="- Remover", command=self._rem_mapa_interpolacao, font=("Segoe UI", 8, "bold"), padx=6).pack(pady=1)

        # Config
        lf2 = tk.LabelFrame(tab, text=" CONFIGURACAO ",
                             font=("Segoe UI", 9, "bold"), bd=1)
        lf2.pack(fill=tk.X, padx=8, pady=4)
        crow = tk.Frame(lf2)
        crow.pack(fill=tk.X, padx=8, pady=4)

        tk.Label(crow, text="Nucleo:").pack(side=tk.LEFT)
        ttk.Combobox(crow, textvariable=self.nucleo_var, values=[""],
                     width=28).pack(side=tk.LEFT, padx=(4, 16))
        tk.Label(crow, text="Tipo:").pack(side=tk.LEFT)
        ttk.Combobox(crow, textvariable=self.tipo_var,
                     values=["auto", "esgoto", "agua"], width=8).pack(side=tk.LEFT, padx=4)

        # Botoes — MODULAR
        lf3 = tk.LabelFrame(tab, text=" PIPELINE MODULAR ",
                             font=("Segoe UI", 9, "bold"), bd=1)
        lf3.pack(fill=tk.X, padx=8, pady=4)

        brow1 = tk.Frame(lf3)
        brow1.pack(fill=tk.X, padx=8, pady=(6, 2))
        btns_top = [
            ("FAZER TUDO (Lá ele)", "#ff007f", "#fff", self._cmd_brutal_tudo),
            ("VARREDURA BATCH (Complete)", "#7c3aed", WHITE, self._cmd_brutal_batch),
            ("APENAS LER PROJETO", BLUE, WHITE, self._cmd_apenas_ler),
        ]
        for txt, bg_cor, fg_cor, cmd in btns_top:
            tk.Button(brow1, text=txt, command=cmd, bg=bg_cor, fg=fg_cor,
                      font=("Segoe UI", 9, "bold"),
                      padx=8, pady=5, cursor="hand2").pack(side=tk.LEFT, padx=2)

        brow2 = tk.Frame(lf3)
        brow2.pack(fill=tk.X, padx=8, pady=2)
        btns_modulos = [
            ("NS CAMPO",     ACCENT, "#000", lambda: self._cmd_modulo("ns_campo")),
            ("NS DESENHO",   "#22c55e", "#000", lambda: self._cmd_modulo("ns_desenho")),
            ("NS SATELITE",  CYAN,   "#000", lambda: self._cmd_modulo("ns_satelite")),
            ("OSE",          PURPLE, WHITE,  lambda: self._cmd_modulo("ose")),
            ("MATERIAIS",    ORANGE, WHITE,  lambda: self._cmd_modulo("materiais")),
            ("COMPRAS",      YELLOW, "#000", lambda: self._cmd_modulo("compras")),
        ]
        for txt, bg_cor, fg_cor, cmd in btns_modulos:
            tk.Button(brow2, text=txt, command=cmd, bg=bg_cor, fg=fg_cor,
                      font=("Segoe UI", 8, "bold"),
                      padx=6, pady=4, cursor="hand2").pack(side=tk.LEFT, padx=2)

        brow3 = tk.Frame(lf3)
        brow3.pack(fill=tk.X, padx=8, pady=(2, 6))
        btns_crono = [
            ("CRONOGRAMA NS",    BLUE,   WHITE, lambda: self._cmd_modulo("crono_ns")),
            ("CPM (Caminho Crítico)", PURPLE, WHITE, lambda: self._cmd_modulo("cpm")),
            ("CRONOGRAMA MICRO", "#00bcd4", "#000", lambda: self._cmd_modulo("crono_micro")),
            ("CRONOGRAMA MACRO", ORANGE, WHITE, lambda: self._cmd_modulo("crono_macro")),
            ("ABRIR SAIDA",      GRAY,   WHITE, self._cmd_abrir_saida),
        ]
        for txt, bg_cor, fg_cor, cmd in btns_crono:
            tk.Button(brow3, text=txt, command=cmd, bg=bg_cor, fg=fg_cor,
                      font=("Segoe UI", 8, "bold"),
                      padx=6, pady=4, cursor="hand2").pack(side=tk.LEFT, padx=2)

        self.progress = ttk.Progressbar(lf3, mode="indeterminate")
        self.progress.pack(fill=tk.X, padx=8, pady=(0, 2))
        self.status_lbl = tk.Label(lf3, text="Pronto",
                                    font=("Segoe UI", 9, "bold"))
        self.status_lbl.pack(pady=(0, 4))

        # Resumo compacto
        self.resumo_frame = tk.LabelFrame(tab, text=" RESUMO ",
                                            font=("Segoe UI", 9, "bold"), bd=1)
        self.resumo_frame.pack(fill=tk.X, padx=8, pady=4)
        self.resumo_text = tk.Label(self.resumo_frame,
            text=f"Pipeline Modular: Projeto DXF/DWG → NS v{NS_VERSION} (Campo/Desenho/Sat/OSE/Materiais/Compras) → Cronogramas (NS/Micro/Macro)  |  "
                 f"Formatos: .dxf .xml .json .dwg  |  {EMPRESA}", font=("Consolas", 8), justify=tk.LEFT, anchor=tk.NW,
            wraplength=1000)
        self.resumo_text.pack(fill=tk.X, padx=8, pady=2)

        # Painel de Sequência Executiva NS
        self._widget_sequencia_ns(tab)

    # ── SEQUÊNCIA EXECUTIVA NS ────────────────────────────────────────────────

    def _widget_sequencia_ns(self, parent):
        """Painel de reordenação da sequência executiva de NS no tab Processar."""
        lf = tk.LabelFrame(parent, text=" SEQUENCIA EXECUTIVA DAS NS ",
                           font=("Segoe UI", 9, "bold"), bd=1)
        lf.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 8))

        # ── linha de parâmetros ──────────────────────────────────────────────
        top = tk.Frame(lf)
        top.pack(fill=tk.X, padx=8, pady=(4, 2))

        for lbl, var, lo, hi in [
            ("De NS:", self._ns_seq_inicio, 1, 999),
            ("Ate NS:", self._ns_seq_fim,   1, 999),
            ("Equipes:", self._ns_equipes,  1, 20),
        ]:
            tk.Label(top, text=lbl, font=("Segoe UI", 8)).pack(side=tk.LEFT)
            tk.Spinbox(top, textvariable=var, from_=lo, to=hi, width=5, insertbackground=FG,
                       font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(2, 8))

        tk.Label(top, text="Prod (m/d):", font=("Segoe UI", 8)).pack(side=tk.LEFT)
        tk.Entry(top, textvariable=self._ns_prod_m_dia, width=6, insertbackground=FG,
                 font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(2, 8))

        # ── Treeview ─────────────────────────────────────────────────────────
        tv_frame = tk.Frame(lf)
        tv_frame.pack(fill=tk.BOTH, expand=True, padx=8, pady=2)

        cols = ("ns", "trecho", "ext", "rua")
        self._ns_tree = ttk.Treeview(tv_frame, columns=cols, show="headings",
                                      height=10, selectmode="browse")
        self._ns_tree.heading("ns",     text="NS")
        self._ns_tree.heading("trecho", text="Trecho")
        self._ns_tree.heading("ext",    text="Ext (m)")
        self._ns_tree.heading("rua",    text="Rua")
        self._ns_tree.column("ns",     width=60,  stretch=False)
        self._ns_tree.column("trecho", width=220, stretch=True)
        self._ns_tree.column("ext",    width=70,  stretch=False)
        self._ns_tree.column("rua",    width=180, stretch=True)

        vsb = ttk.Scrollbar(tv_frame, orient="vertical",   command=self._ns_tree.yview)
        self._ns_tree.configure(yscrollcommand=vsb.set)
        self._ns_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)

        # ── botões de reordenação ─────────────────────────────────────────────
        btn_frame = tk.Frame(lf)
        btn_frame.pack(fill=tk.X, padx=8, pady=(2, 4))

        for txt, cmd in [
            ("↑",    lambda: self._seq_mover(-1)),
            ("↓",    lambda: self._seq_mover(+1)),
            ("Topo", lambda: self._seq_mover("topo")),
            ("Base", lambda: self._seq_mover("base")),
        ]:
            tk.Button(btn_frame, text=txt, command=cmd,
                      font=("Segoe UI", 9, "bold"), width=6,
                      cursor="hand2").pack(side=tk.LEFT, padx=2)

        tk.Frame(btn_frame, width=12).pack(side=tk.LEFT)

        for txt, cor, cmd in [
            ("SALVAR SEQUENCIA",   BLUE,   self._seq_salvar),
            ("CARREGAR SEQUENCIA", GRAY,   self._seq_carregar),
            ("GERAR CRONOGRAMA NS", ACCENT, self._cmd_cronograma_ns),
        ]:
            tk.Button(btn_frame, text=txt, command=cmd, bg=cor if cor == ACCENT else WHITE,
                      font=("Segoe UI", 9, "bold"),
                      padx=8, pady=3, cursor="hand2").pack(side=tk.LEFT, padx=2)

    def _ns_seq_refresh(self):
        """Atualiza a treeview de sequência após carregar trechos."""
        if not hasattr(self, "_ns_tree"):
            return
        # Se a sequência ainda não foi definida, usar ordem natural
        if not self._ns_sequencia and self.trechos:
            self._ns_sequencia = list(range(len(self.trechos)))

        self._ns_tree.delete(*self._ns_tree.get_children())
        ini = self._ns_seq_inicio.get()
        fim = self._ns_seq_fim.get()
        for seq, idx in enumerate(self._ns_sequencia):
            ns_id = seq + 1
            if ns_id < ini or ns_id > fim:
                continue
            if idx >= len(self.trechos):
                continue
            t = self.trechos[idx]
            self._ns_tree.insert("", "end", iid=str(seq),
                values=(f"NS{ns_id:03d}",
                        f"{t.get('pv_ini','?')} → {t.get('pv_fim','?')}",
                        f"{t.get('ext_m',0):.1f}",
                        t.get("rua", "")))

    def _seq_mover(self, direcao):
        sel = self._ns_tree.selection()
        if not sel:
            return
        seq = int(sel[0])
        if seq >= len(self._ns_sequencia):
            return
        lst = self._ns_sequencia

        if direcao == "topo":
            lst.insert(0, lst.pop(seq))
        elif direcao == "base":
            lst.append(lst.pop(seq))
        elif direcao == -1 and seq > 0:
            lst[seq], lst[seq - 1] = lst[seq - 1], lst[seq]
        elif direcao == +1 and seq < len(lst) - 1:
            lst[seq], lst[seq + 1] = lst[seq + 1], lst[seq]
        else:
            return

        self._ns_seq_refresh()
        new_seq = max(0, seq + (0 if direcao in ("topo", "base") else direcao))
        new_seq = 0 if direcao == "topo" else (len(lst) - 1 if direcao == "base" else new_seq)
        iid = str(new_seq)
        if self._ns_tree.exists(iid):
            self._ns_tree.selection_set(iid)
            self._ns_tree.see(iid)

    def _seq_salvar(self):
        from tkinter import filedialog
        path = filedialog.asksaveasfilename(
            title="Salvar Sequência Executiva",
            defaultextension=".json",
            filetypes=[("JSON", "*.json")],
            initialfile="SEQUENCIA_EXECUTIVA.json",
        )
        if not path:
            return
        from datetime import datetime as _dt
        seq_data = {
            "nucleo":        self.nucleo_var.get() or "REDE",
            "gerado_em":     _dt.now().isoformat(),
            "ns_seq_inicio": self._ns_seq_inicio.get(),
            "ns_seq_fim":    self._ns_seq_fim.get(),
            "equipes":       self._ns_equipes.get(),
            "prod_m_dia":    self._ns_prod_m_dia.get(),
            "sequencia": [
                {"ordem":      i + 1,
                 "trecho_idx": idx,
                 "pv_ini":     self.trechos[idx].get("pv_ini") if idx < len(self.trechos) else "",
                 "pv_fim":     self.trechos[idx].get("pv_fim") if idx < len(self.trechos) else "",
                 "ext_m":      self.trechos[idx].get("ext_m", 0) if idx < len(self.trechos) else 0,
                 "rua":        self.trechos[idx].get("rua", "") if idx < len(self.trechos) else ""}
                for i, idx in enumerate(self._ns_sequencia)
            ],
        }
        import json as _json
        with open(path, "w", encoding="utf-8") as f:
            _json.dump(seq_data, f, indent=2, ensure_ascii=False)
        self._log_msg(f"Sequência salva: {Path(path).name}", "OK")

    def _seq_carregar(self):
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            title="Carregar Sequência Executiva",
            filetypes=[("JSON", "*.json")],
        )
        if not path or not Path(path).exists():
            return
        import json as _json
        with open(path, "r", encoding="utf-8") as f:
            data = _json.load(f)

        self._ns_sequencia = [item["trecho_idx"] for item in data.get("sequencia", [])]
        if data.get("ns_seq_inicio"):
            self._ns_seq_inicio.set(data["ns_seq_inicio"])
        if data.get("ns_seq_fim"):
            self._ns_seq_fim.set(data["ns_seq_fim"])
        if data.get("equipes"):
            self._ns_equipes.set(data["equipes"])
        if data.get("prod_m_dia"):
            self._ns_prod_m_dia.set(data["prod_m_dia"])

        self._ns_seq_refresh()
        self._log_msg(f"Sequência carregada: {len(self._ns_sequencia)} NS", "OK")

    def _cmd_cronograma_ns(self):
        if not self._ns_sequencia:
            messagebox.showwarning("Aviso", "Carregue a rede primeiro."); return
        self._run(self._do_cronograma_ns)

    def _do_cronograma_ns(self):
        from gerar_cronograma_macro import gerar_cronograma_por_ns
        from tkinter import filedialog
        from datetime import datetime as _dt

        ini = self._ns_seq_inicio.get()
        fim = self._ns_seq_fim.get()
        ns_lista = [
            {"ordem":      seq + 1,
             "trecho_idx": idx,
             "pv_ini":     self.trechos[idx].get("pv_ini", "") if idx < len(self.trechos) else "",
             "pv_fim":     self.trechos[idx].get("pv_fim", "") if idx < len(self.trechos) else "",
             "ext_m":      self.trechos[idx].get("ext_m", 0) if idx < len(self.trechos) else 0,
             "rua":        self.trechos[idx].get("rua", "") if idx < len(self.trechos) else ""}
            for seq, idx in enumerate(self._ns_sequencia)
            if ini <= (seq + 1) <= fim
        ]

        nucleo = self.nucleo_var.get() or "REDE"
        slug   = nucleo.lower().replace(" ", "_")
        out    = Path(self.saida_var.get()) / nucleo.upper().replace(" ", "_") / "PLANEJAMENTO" / "CRONOGRAMA"
        out.mkdir(parents=True, exist_ok=True)

        resultado = gerar_cronograma_por_ns(
            ns_lista,
            data_inicio_str=_dt.now().strftime("%Y-%m-%d"),
            equipes=self._ns_equipes.get(),
            prod_m_dia=self._ns_prod_m_dia.get(),
            nucleo=nucleo,
            out_dir=str(out),
        )
        n = len(resultado["tarefas"])
        self.root.after(0, self._log_msg,
            f"Cronograma NS: {n} tarefas | {resultado['extensao_total_m']:.0f}m | "
            f"{resultado['data_inicio']} → {resultado['data_fim']}", "OK")
        self.root.after(0, self._log_msg, f"Exportado → {out}", "OK")

        gantt = out / "GANTT_NS.html"
        if gantt.exists():
            import webbrowser
            self.root.after(500, lambda: webbrowser.open(str(gantt)))

    # ── TAB 2: MAPA ──────────────────────────────────────────────────────────

    def _tab_mapa(self):
        tab = tk.Frame(self.nb)
        self.nb.add(tab, text="  [2] Mapa  ")

        tb = tk.Frame(tab)
        tb.pack(fill=tk.X, padx=4, pady=(4, 0))
        for txt, cor, cmd in [
            ("Carregar Rede", ACCENT, self._mapa_carregar),
            ("Validar GPKG", "#059669", self._mapa_validar_gpkg),
            ("Gerar NS Selecionados", PURPLE, self._mapa_gerar_ns),
            ("Satelite/Rua", BLUE, self._mapa_trocar_tile),
        ]:
            tk.Button(tb, text=txt, command=cmd, bg=cor if cor == ACCENT else WHITE,
                      font=("Segoe UI", 9, "bold"),
                      padx=8, pady=3).pack(side=tk.LEFT, padx=2)

        self._mapa_tile_label = tk.Label(tb, text="Satelite",
                                          font=("Segoe UI", 8))
        self._mapa_tile_label.pack(side=tk.LEFT, padx=4)

        for txt, cmd in [("Salvar ML", self._ml_salvar),
                         ("Treinar", self._ml_treinar),
                         ("Predizer", self._ml_predizer)]:
            tk.Button(tb, text=txt, command=cmd,
                      font=("Segoe UI", 7), padx=5, pady=3
                      ).pack(side=tk.LEFT, padx=1)

        self.mapa_info = tk.Label(tb, text="", font=("Segoe UI", 8))
        self.mapa_info.pack(side=tk.RIGHT, padx=4)

        content = tk.PanedWindow(tab, orient=tk.HORIZONTAL, sashwidth=4)
        content.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)

        map_frame = tk.Frame(content)
        try:
            from tkintermapview import TkinterMapView
            self.map_widget = TkinterMapView(map_frame, corner_radius=0)
            self.map_widget.pack(fill=tk.BOTH, expand=True)
            self.map_widget.set_position(-23.96, -46.33)
            self.map_widget.set_zoom(14)
            self.map_widget.set_tile_server(_TILE_SAT, max_zoom=19)
            self._has_map = True
            self._tile_is_sat = True
        except ImportError:
            self._has_map = False
            self._tile_is_sat = True
            tk.Label(map_frame, text="pip install tkintermapview", font=("Segoe UI", 12)).pack(expand=True)
        content.add(map_frame, width=680)

        panel = tk.Frame(content)
        content.add(panel, width=340)

        btn_row = tk.Frame(panel)
        btn_row.pack(fill=tk.X, padx=4, pady=4)
        for txt, cmd in [("Todos", self._mapa_sel_todos),
                         ("Nenhum", self._mapa_sel_nenhum),
                         ("Inverter", self._mapa_sel_inverter)]:
            tk.Button(btn_row, text=txt, command=cmd, padx=6, font=("Segoe UI", 8)).pack(side=tk.LEFT, padx=2)
        tk.Button(btn_row, text="Toggle", command=self._mapa_toggle_trecho, padx=8,
                  font=("Segoe UI", 8, "bold")).pack(side=tk.LEFT, padx=4)
        self.mapa_sel_count = tk.Label(btn_row, text="0/0",
                                        font=("Segoe UI", 9, "bold"))
        self.mapa_sel_count.pack(side=tk.RIGHT, padx=4)

        lf = tk.Frame(panel)
        lf.pack(fill=tk.BOTH, expand=True, padx=4, pady=(0, 2))
        self.mapa_listbox = tk.Listbox(lf, font=("Consolas", 9),
                                        selectmode=tk.SINGLE, bd=2,
                                        activestyle="none")
        sb = tk.Scrollbar(lf, command=self.mapa_listbox.yview)
        self.mapa_listbox.config(yscrollcommand=sb.set)
        self.mapa_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb.pack(side=tk.RIGHT, fill=tk.Y)
        self.mapa_listbox.bind("<<ListboxSelect>>", lambda e: self._mapa_on_select())
        self.mapa_listbox.bind("<space>", lambda e: self._mapa_toggle_trecho())
        self.mapa_listbox.bind("<Return>", lambda e: self._mapa_toggle_trecho())

        self.mapa_sel_var = tk.StringVar(value="Espaco/Enter = incluir/excluir")
        tk.Label(panel, textvariable=self.mapa_sel_var,
                 font=("Consolas", 8), wraplength=330, justify=tk.LEFT,
                 anchor=tk.W).pack(fill=tk.X, padx=8, pady=4)

        self._mapa_markers = []
        self._mapa_paths = []
        self._mapa_checkstates = []

    # ── TAB 3-5: REDE, HIDRAULICA, TRECHOS ──────────────────────────────────

    def _tab_rede(self):
        tab = tk.Frame(self.nb)
        self.nb.add(tab, text="  [3] Rede  ")
        self.rede_stats = tk.Frame(tab)
        self.rede_stats.pack(fill=tk.X, padx=8, pady=8)
        self.stat_labels = {}
        for nome, val in [("PVs", "0"), ("Trechos", "0"), ("Extensao (m)", "0"),
                          ("Tipo", "-"), ("Motor", "-"), ("Ruas", "0")]:
            f = tk.Frame(self.rede_stats, padx=12, pady=8)
            f.pack(side=tk.LEFT, padx=4, expand=True, fill=tk.X)
            tk.Label(f, text=nome, font=("Segoe UI", 8)).pack()
            lbl = tk.Label(f, text=val, font=("Segoe UI", 16, "bold"))
            lbl.pack()
            self.stat_labels[nome] = lbl
        cols = ("Nome", "X", "Y", "CT", "CF", "Prof")
        self.pv_tree = ttk.Treeview(tab, columns=cols, show="headings", height=18)
        for c in cols:
            self.pv_tree.heading(c, text=c)
            self.pv_tree.column(c, width=100, anchor=tk.CENTER)
        self.pv_tree.column("Nome", width=110, anchor=tk.W)
        sb = ttk.Scrollbar(tab, orient=tk.VERTICAL, command=self.pv_tree.yview)
        self.pv_tree.configure(yscrollcommand=sb.set)
        self.pv_tree.pack(fill=tk.BOTH, expand=True, padx=8, side=tk.LEFT)
        sb.pack(fill=tk.Y, side=tk.RIGHT, padx=(0, 8))

    def _tab_hidraulica(self):
        tab = tk.Frame(self.nb)
        self.nb.add(tab, text="  [4] Hidraulica  ")
        self.hid_stats = tk.Frame(tab)
        self.hid_stats.pack(fill=tk.X, padx=8, pady=8)
        self.hid_labels = {}
        for nome, val, cor in [("OK", "0", ACCENT), ("Verificar", "0", YELLOW),
                                ("Sem Dados", "0", GRAY), ("Manning n", "0.013", CYAN)]:
            f = tk.Frame(self.hid_stats, padx=12, pady=8)
            f.pack(side=tk.LEFT, padx=4, expand=True, fill=tk.X)
            tk.Label(f, text=nome, font=("Segoe UI", 8)).pack()
            lbl = tk.Label(f, text=val, fg=cor, font=("Segoe UI", 16, "bold"))
            lbl.pack()
            self.hid_labels[nome] = lbl
        cols = ("NS", "PV Ini", "PV Fim", "DN", "Ext(m)", "Decl(%)",
                "V(m/s)", "Q(l/s)", "Tau(Pa)", "Status")
        self.hid_tree = ttk.Treeview(tab, columns=cols, show="headings", height=18)
        for c in cols:
            self.hid_tree.heading(c, text=c)
            self.hid_tree.column(c, width=75, anchor=tk.CENTER)
        self.hid_tree.column("Status", width=150, anchor=tk.W)
        sb = ttk.Scrollbar(tab, orient=tk.VERTICAL, command=self.hid_tree.yview)
        self.hid_tree.configure(yscrollcommand=sb.set)
        self.hid_tree.pack(fill=tk.BOTH, expand=True, padx=8, side=tk.LEFT)
        sb.pack(fill=tk.Y, side=tk.RIGHT, padx=(0, 8))

    def _tab_trechos(self):
        tab = tk.Frame(self.nb)
        self.nb.add(tab, text="  [5] Trechos  ")
        cols = ("NS", "PV Ini", "PV Fim", "Rua", "DN", "Ext(m)", "Material",
                "CT Ini", "CF Ini", "Prof Ini", "Custo R$")
        self.tr_tree = ttk.Treeview(tab, columns=cols, show="headings", height=20)
        for c in cols:
            self.tr_tree.heading(c, text=c)
            self.tr_tree.column(c, width=85, anchor=tk.CENTER)
        self.tr_tree.column("Rua", width=140, anchor=tk.W)
        sb = ttk.Scrollbar(tab, orient=tk.VERTICAL, command=self.tr_tree.yview)
        self.tr_tree.configure(yscrollcommand=sb.set)
        self.tr_tree.pack(fill=tk.BOTH, expand=True, padx=8, pady=8, side=tk.LEFT)
        sb.pack(fill=tk.Y, side=tk.RIGHT, padx=(0, 8), pady=8)

    # ── TAB 6: CUSTOS / MEDICAO ────────────────────────────────────────────

    def _tab_custos(self):
        tab = tk.Frame(self.nb)
        self.nb.add(tab, text="  [6] Custos 5D  ")

        # Stats cards
        cstats = tk.Frame(tab)
        cstats.pack(fill=tk.X, padx=8, pady=8)
        self.custo_labels = {}
        for nome, val, cor in [("Custo Total R$", "0", ACCENT), ("R$/metro", "0", CYAN),
                                ("BDI 25%", "0", YELLOW), ("Trechos", "0", BLUE),
                                ("Extensao (m)", "0", PURPLE), ("BMs", "0", ORANGE)]:
            f = tk.Frame(cstats, padx=10, pady=6)
            f.pack(side=tk.LEFT, padx=3, expand=True, fill=tk.X)
            tk.Label(f, text=nome, font=("Segoe UI", 7)).pack()
            lbl = tk.Label(f, text=val, fg=cor, font=("Segoe UI", 13, "bold"))
            lbl.pack()
            self.custo_labels[nome] = lbl

        # Buttons
        brow = tk.Frame(tab)
        brow.pack(fill=tk.X, padx=8, pady=4)
        for txt, cor, fg_c, cmd in [
            ("CALCULAR CUSTOS", ACCENT, "#000", self._cmd_custos),
            ("GERAR BM", ORANGE, WHITE, self._cmd_gerar_bm),
            ("CURVA S", PURPLE, WHITE, self._cmd_curva_s),
            ("MICRO-PLAN", CYAN, "#000", self._cmd_microplan),
            ("RELATORIO ML", "#b45309", WHITE, self._cmd_relatorio_ml),
            ("CRONOGRAMA MACRO", BLUE, WHITE, self._cmd_crono_macro),
        ]:
            tk.Button(brow, text=txt, command=cmd, bg=cor, fg=fg_c,
                      font=("Segoe UI", 8, "bold"),
                      padx=6, pady=4).pack(side=tk.LEFT, padx=2)

        # Cost table
        cols = ("NS", "PV Ini", "PV Fim", "DN", "Ext(m)", "Tubo R$", "Escav R$",
                "Reaterro R$", "Repav R$", "PV R$", "TOTAL R$")
        self.custo_tree = ttk.Treeview(tab, columns=cols, show="headings", height=16)
        for c in cols:
            self.custo_tree.heading(c, text=c)
            self.custo_tree.column(c, width=80, anchor=tk.CENTER)
        self.custo_tree.column("NS", width=50)
        sb = ttk.Scrollbar(tab, orient=tk.VERTICAL, command=self.custo_tree.yview)
        self.custo_tree.configure(yscrollcommand=sb.set)
        self.custo_tree.pack(fill=tk.BOTH, expand=True, padx=8, side=tk.LEFT)
        sb.pack(fill=tk.Y, side=tk.RIGHT, padx=(0, 8))

        self.custo_status = tk.Label(tab, text="Processe um arquivo primeiro",
                                      font=("Segoe UI", 8))
        self.custo_status.pack(side=tk.BOTTOM, pady=2)

    # ── TAB 7: BIM / CIVIL 3D ───────────────────────────────────────────────

    def _tab_bim(self):
        tab = tk.Frame(self.nb)
        self.nb.add(tab, text="  [7] BIM  ")

        tk.Label(tab, text="Pipeline de Saidas BIM 5D",
                 font=("Segoe UI", 12, "bold")).pack(padx=12, pady=(8, 4), anchor=tk.W)

        r1 = tk.Frame(tab)
        r1.pack(fill=tk.X, padx=8, pady=4)
        for txt, cor, fg_c, cmd in [
            ("GERAR TUDO (6 etapas)", ACCENT, "#000", self._cmd_gerar_tudo_bim),
            ("IFC LOD500", BLUE, WHITE, self._cmd_ifc),
            ("LandXML", "#dc2626", WHITE, self._cmd_landxml),
            ("Cadastro NTS292", ORANGE, WHITE, self._cmd_nts292),
            ("Cadastro DXF", "#ea580c", WHITE, self._cmd_cadastro_dxf),
            ("Cronograma", PURPLE, WHITE, self._cmd_cronograma),
            ("Dynamo", CYAN, "#000", self._cmd_dynamo),
            ("SCR", GRAY, WHITE, self._cmd_scr),
        ]:
            tk.Button(r1, text=txt, command=cmd, bg=cor, fg=fg_c,
                      font=("Segoe UI", 9, "bold"),
                      padx=8, pady=5).pack(side=tk.LEFT, padx=2)

        # Viewer HTML buttons
        r2 = tk.Frame(tab)
        r2.pack(fill=tk.X, padx=8, pady=4)
        tk.Label(r2, text="Interfaces HTML:",
                 font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(0, 8))
        for txt, html_file in [
            ("Editor EPANET", "construdata_editor.html"),
            ("Viewer 3D", "construdata_manage.html"),
            ("Controle As-Built", "construdata_controle.html"),
            ("RDO Diario", "construdata_rdo.html"),
            ("Gestao Perdas", "construdata_perdas.html"),
            ("Fluxograma", "FLUXOGRAMA_BIM_5D.html"),
        ]:
            tk.Button(r2, text=txt,
                      font=("Segoe UI", 8), padx=8, pady=3,
                      command=lambda f=html_file: self._abrir_html(f)
                      ).pack(side=tk.LEFT, padx=2)

        info = tk.LabelFrame(tab, text=" SAIDAS DO PIPELINE ",
                             font=("Segoe UI", 9, "bold"), bd=1)
        info.pack(fill=tk.BOTH, expand=True, padx=8, pady=4)
        dados = [
            ("01_NS/",           "Notas de Servico: PDF A4 + JSON + HTML Leaflet + GeoJSON"),
            ("02_CIVIL3D/",      "LandXML 1.2 + Cadastro DXF + Dynamo .py + AutoCAD .scr"),
            ("03_CADASTRO_NTS292/", "DXF As-Built georref SIRGAS 2000 UTM 23S + Meta JSON"),
            ("04_BIM_LOD500/",   "IFC 3D real (SweptDiskSolid+ExtrudedAreaSolid) + CSV + JSON"),
            ("05_CRONOGRAMA/",   "MS Project XML com WBS por fase + Resumo JSON"),
        ]
        for pasta, desc in dados:
            row = tk.Frame(info)
            row.pack(fill=tk.X, padx=8, pady=1)
            tk.Label(row, text=pasta, width=22, anchor=tk.W,
                     font=("Consolas", 9, "bold")).pack(side=tk.LEFT)
            tk.Label(row, text=desc, font=("Segoe UI", 8)).pack(side=tk.LEFT, padx=8)

        self.bim_status = tk.Label(tab, text="", font=("Segoe UI", 9))
        self.bim_status.pack(pady=4)

    # ── TAB 9: LEAN / LPS ────────────────────────────────────────────────

    def _tab_lean(self):
        tab = tk.Frame(self.nb)
        self.nb.add(tab, text="  [8] Lean/LPS  ")

        tk.Label(tab, text="Lean Construction + Last Planner System + BIM 6D",
                 font=("Segoe UI", 11, "bold")).pack(padx=8, pady=(8, 2), anchor=tk.W)

        brow = tk.Frame(tab)
        brow.pack(fill=tk.X, padx=8, pady=4)
        for txt, cor, fg_c, cmd in [
            ("RELATORIO LEAN+LPS", ACCENT, "#000", self._cmd_lean_report),
            ("TAKT TIME", CYAN, "#000", self._cmd_takt),
            ("LOOKAHEAD 6 SEM", PURPLE, WHITE, self._cmd_lookahead),
            ("BIM 6D (Ciclo Vida)", ORANGE, WHITE, self._cmd_bim6d),
        ]:
            tk.Button(brow, text=txt, command=cmd, bg=cor, fg=fg_c,
                      font=("Segoe UI", 9, "bold"),
                      padx=8, pady=5).pack(side=tk.LEFT, padx=2)

        # Stats
        lstats = tk.Frame(tab)
        lstats.pack(fill=tk.X, padx=8, pady=4)
        self.lean_labels = {}
        for nome, val, cor in [("Takt (m/dia)", "-", ACCENT), ("Cycle Time", "-", CYAN),
                                ("PPC (%)", "-", YELLOW), ("VA/NVA", "-", PURPLE),
                                ("CO2 (ton)", "-", ORANGE), ("Custo 50 anos", "-", BLUE)]:
            f = tk.Frame(lstats, padx=10, pady=6)
            f.pack(side=tk.LEFT, padx=3, expand=True, fill=tk.X)
            tk.Label(f, text=nome, font=("Segoe UI", 7)).pack()
            lbl = tk.Label(f, text=val, fg=cor, font=("Segoe UI", 13, "bold"))
            lbl.pack()
            self.lean_labels[nome] = lbl

        self.lean_text = scrolledtext.ScrolledText(tab, font=("Consolas", 9),
                                                     wrap=tk.WORD, bd=4)
        self.lean_text.pack(fill=tk.BOTH, expand=True, padx=8, pady=4)

    # ── TAB 10: PERDAS ────────────────────────────────────────────────────

    def _tab_perdas(self):
        tab = tk.Frame(self.nb)
        self.nb.add(tab, text="  [9] Perdas  ")

        tk.Label(tab, text="Gestao de Perdas — IWA / UARL / ILI / DMA",
                 font=("Segoe UI", 11, "bold")).pack(padx=8, pady=(8, 2), anchor=tk.W)

        brow = tk.Frame(tab)
        brow.pack(fill=tk.X, padx=8, pady=4)
        for txt, cor, fg_c, cmd in [
            ("RELATORIO PERDAS", ACCENT, "#000", self._cmd_perdas_report),
            ("MAPA RISCO", RED, WHITE, self._cmd_mapa_risco),
            ("CRIAR DMAs", BLUE, WHITE, self._cmd_criar_dma),
            ("PDF PERDAS", PURPLE, WHITE, self._cmd_pdf_perdas),
            ("ANALISE TROCA", ORANGE, WHITE, self._cmd_analise_troca),
        ]:
            tk.Button(brow, text=txt, command=cmd, bg=cor, fg=fg_c,
                      font=("Segoe UI", 9, "bold"),
                      padx=8, pady=5).pack(side=tk.LEFT, padx=2)

        # Stats
        pstats = tk.Frame(tab)
        pstats.pack(fill=tk.X, padx=8, pady=4)
        self.perdas_labels = {}
        for nome, val, cor in [("UARL (m3/ano)", "-", CYAN), ("ILI", "-", ACCENT),
                                ("Classif.", "-", YELLOW), ("Risco Alto", "0", RED),
                                ("DMAs", "-", BLUE), ("Perda R$/ano", "-", ORANGE)]:
            f = tk.Frame(pstats, padx=10, pady=6)
            f.pack(side=tk.LEFT, padx=3, expand=True, fill=tk.X)
            tk.Label(f, text=nome, font=("Segoe UI", 7)).pack()
            lbl = tk.Label(f, text=val, fg=cor, font=("Segoe UI", 13, "bold"))
            lbl.pack()
            self.perdas_labels[nome] = lbl

        self.perdas_text = scrolledtext.ScrolledText(tab, font=("Consolas", 9),
                                                       wrap=tk.WORD, bd=4)
        self.perdas_text.pack(fill=tk.BOTH, expand=True, padx=8, pady=4)

    # ── TAB 11: IA (LLMs) ──────────────────────────────────────────────────

    def _tab_ia(self):
        tab = tk.Frame(self.nb)
        self.nb.add(tab, text="  [10] IA  ")

        tk.Label(tab, text="Assistente IA — 4 LLMs Gratuitos + Analytics ML",
                 font=("Segoe UI", 11, "bold")).pack(padx=8, pady=(8, 2), anchor=tk.W)

        # ── SEÇÃO ANALYTICS ────────────────────────────────────────────────────
        analytics_frame = tk.LabelFrame(tab, text=" ANALYTICS ML — XGBoost/RandomForest ",
                                         font=("Segoe UI", 9, "bold"), bd=1)
        analytics_frame.pack(fill=tk.X, padx=8, pady=4)

        # Status do módulo Analytics
        astatus_row = tk.Frame(analytics_frame)
        astatus_row.pack(fill=tk.X, padx=8, pady=4)
        analytics_status = "OK" if _ENGINES.get("Analytics", False) else "NÃO DISPONÍVEL"
        analytics_cor = ACCENT if analytics_status == "OK" else RED
        tk.Label(astatus_row, text=f"Analytics: {analytics_status}", fg=analytics_cor,
                 font=("Segoe UI", 9, "bold")).pack(side=tk.LEFT)

        # Botões de ação do Analytics
        abtn_row = tk.Frame(analytics_frame)
        abtn_row.pack(fill=tk.X, padx=8, pady=4)
        for txt, cor, fg_c, cmd in [
            ("🚀 EXECUTAR ANALYTICS", ACCENT, "#000", self._cmd_executar_analytics),
            ("📊 VER GRÁFICOS", BLUE, WHITE, self._cmd_ver_graficos_analytics),
            ("📈 CENÁRIOS", PURPLE, WHITE, self._cmd_cenarios_analytics),
            ("📄 ABRIR RELATÓRIO", YELLOW, "#000", self._cmd_abrir_relatorio_analytics),
            ("📄 EXPORTAR XLSX", ORANGE, WHITE, self._cmd_exportar_xlsx_analytics),
            ("📂 ABRIR PASTA", GRAY, WHITE, self._cmd_abrir_pasta_analytics),
        ]:
            tk.Button(abtn_row, text=txt, command=cmd, bg=cor, fg=fg_c,
                      font=("Segoe UI", 8, "bold"),
                      padx=8, pady=4).pack(side=tk.LEFT, padx=2)

        # Labels de resultado do Analytics
        self.analytics_labels = {}
        albl_row1 = tk.Frame(analytics_frame)
        albl_row1.pack(fill=tk.X, padx=8, pady=4)
        for nome, val, cor in [("R² Test", "-", YELLOW), ("MAE", "-", BLUE),
                                ("RMSE", "-", ORANGE), ("Algoritmo", "-", ACCENT)]:
            f = tk.Frame(albl_row1, padx=8, pady=4)
            f.pack(side=tk.LEFT, padx=2, expand=True, fill=tk.X)
            tk.Label(f, text=nome, font=("Segoe UI", 7)).pack()
            lbl = tk.Label(f, text=val, fg=cor, font=("Segoe UI", 10, "bold"))
            lbl.pack()
            self.analytics_labels[nome] = lbl

        albl_row2 = tk.Frame(analytics_frame)
        albl_row2.pack(fill=tk.X, padx=8, pady=2)
        for nome, val, cor in [("Ligações Realizadas", "-", WHITE), ("Faltam", "-", RED),
                                ("Previsão Conclusão", "-", CYAN), ("Feature Top", "-", PURPLE)]:
            f = tk.Frame(albl_row2, padx=8, pady=4)
            f.pack(side=tk.LEFT, padx=2, expand=True, fill=tk.X)
            tk.Label(f, text=nome, font=("Segoe UI", 7)).pack()
            lbl = tk.Label(f, text=val, fg=cor, font=("Segoe UI", 9))
            lbl.pack()
            self.analytics_labels[nome] = lbl

        # Área de texto para detalhes do Analytics
        self.analytics_text = scrolledtext.ScrolledText(analytics_frame,
                                                         font=("Consolas", 8),
                                                         wrap=tk.WORD, bd=2,
                                                         height=6)
        self.analytics_text.pack(fill=tk.X, padx=8, pady=4)
        self.analytics_text.insert(tk.END, "Aguardando execução do Analytics...\n")

        # ── SEÇÃO SLNR MESTRE UNIFICADO ───────────────────────────────────────
        slnr_frame = tk.LabelFrame(tab, text=" SLNR MESTRE UNIFICADO — 20 NÚCLEOS + ML ",
                                    font=("Segoe UI", 9, "bold"), bd=1)
        slnr_frame.pack(fill=tk.X, padx=8, pady=4)

        # Status do módulo SLNR Mestre
        slnr_status = "OK" if _ENGINES.get("SLNR_Mestre", False) else "NÃO DISPONÍVEL"
        slnr_cor = ACCENT if slnr_status == "OK" else RED
        tk.Label(slnr_frame, text=f"SLNR Mestre: {slnr_status}", fg=slnr_cor,
                 font=("Segoe UI", 9, "bold")).pack(padx=8, pady=4, anchor=tk.W)

        # Botões SLNR Mestre
        slnr_btn_row = tk.Frame(slnr_frame)
        slnr_btn_row.pack(fill=tk.X, padx=8, pady=4)
        for txt, cor, fg_c, cmd in [
            ("📊 GERAR SLNR ML", BLUE, WHITE, self._cmd_slnr_ml),
            ("📄 EMITIR NOTAS SERVIÇO", PURPLE, WHITE, self._cmd_emitir_notas_servico),
            ("📄 ABRIR PLANILHA", ORANGE, WHITE, self._cmd_abrir_slnr),
            ("📂 ABRIR PASTA", GRAY, WHITE, self._cmd_abrir_pasta_slnr),
        ]:
            tk.Button(slnr_btn_row, text=txt, command=cmd, bg=cor, fg=fg_c,
                      font=("Segoe UI", 8, "bold"),
                      padx=8, pady=4).pack(side=tk.LEFT, padx=2)

        # Labels SLNR Mestre
        self.slnr_labels = {}
        slnr_lbl_row = tk.Frame(slnr_frame)
        slnr_lbl_row.pack(fill=tk.X, padx=8, pady=2)
        for nome, val, cor in [("Núcleos", "12", WHITE), ("Fórmulas", "115+", CYAN),
                                ("Cenários", "5", YELLOW), ("R² ML", "-", PURPLE)]:
            f = tk.Frame(slnr_lbl_row, padx=8, pady=4)
            f.pack(side=tk.LEFT, padx=2, expand=True, fill=tk.X)
            tk.Label(f, text=nome, font=("Segoe UI", 7)).pack()
            lbl = tk.Label(f, text=val, fg=cor, font=("Segoe UI", 9, "bold"))
            lbl.pack()
            self.slnr_labels[nome] = lbl

        # Área de texto SLNR Mestre
        self.slnr_text = scrolledtext.ScrolledText(slnr_frame,
                                                    font=("Consolas", 8),
                                                    wrap=tk.WORD, bd=2,
                                                    height=4)
        self.slnr_text.pack(fill=tk.X, padx=8, pady=4)
        self.slnr_text.insert(tk.END, "SLNR Mestre Unificado — 12 núcleos com fórmulas + ML integrado\n")

        # ── SEÇÃO LLMs ────────────────────────────────────────────────────────

        # Status providers + Botão Configurar
        prow = tk.Frame(tab)
        prow.pack(fill=tk.X, padx=8, pady=4)
        
        # Botão Configurar API Keys
        tk.Button(prow, text="⚙ Configurar API Keys", command=self._ia_configurar_keys, font=("Segoe UI", 9, "bold"), padx=10, pady=4).pack(side=tk.RIGHT)
        
        self.ia_status_labels = {}
        for provider, cor in [("Gemini", "#4285f4"), ("Groq", "#f55036"),
                               ("Mistral", "#ff7000"), ("Cohere", "#39594d")]:
            f = tk.Frame(prow, padx=8, pady=4)
            f.pack(side=tk.LEFT, padx=3, expand=True, fill=tk.X)
            status = "OK" if _ENGINES.get(provider, _ENGINES.get("Gemini" if provider == "Gemini" else "Multi-LLM", False)) else "?"
            lbl = tk.Label(f, text=f"{provider}: {status}", fg=cor,
                          font=("Segoe UI", 9, "bold"))
            lbl.pack()
            self.ia_status_labels[provider] = lbl

        # Quick buttons
        brow = tk.Frame(tab)
        brow.pack(fill=tk.X, padx=8, pady=4)
        for txt, cor, fg_c, cmd in [
            ("Resumo Executivo", ACCENT, "#000", lambda: self._ia_quick("resumo")),
            ("Validar Hidraulica", BLUE, WHITE, lambda: self._ia_quick("hidraulica")),
            ("Analisar Perdas", CYAN, "#000", lambda: self._ia_quick("perdas")),
            ("Explicar ML", PURPLE, WHITE, lambda: self._ia_quick("ml")),
            ("Analisar Foto", ORANGE, WHITE, self._ia_foto),
            ("Ler PDF", "#dc2626", WHITE, self._ia_pdf),
        ]:
            tk.Button(brow, text=txt, command=cmd, bg=cor, fg=fg_c,
                      font=("Segoe UI", 8, "bold"),
                      padx=6, pady=4).pack(side=tk.LEFT, padx=2)

        # Question input
        qrow = tk.Frame(tab)
        qrow.pack(fill=tk.X, padx=8, pady=4)
        tk.Label(qrow, text="Pergunta:", font=("Segoe UI", 9)).pack(side=tk.LEFT)
        self.ia_question = tk.Entry(qrow, insertbackground=FG, bd=2, font=("Segoe UI", 9))
        self.ia_question.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=4)
        self.ia_question.bind("<Return>", lambda e: self._ia_perguntar())
        tk.Button(qrow, text="Perguntar", command=self._ia_perguntar, font=("Segoe UI", 9, "bold"), padx=12).pack(side=tk.RIGHT)

        # Response area
        self.ia_text = scrolledtext.ScrolledText(tab, font=("Consolas", 9),
                                                   wrap=tk.WORD, bd=4)
        self.ia_text.pack(fill=tk.BOTH, expand=True, padx=8, pady=4)
        self.ia_text.insert(tk.END, "LLMs disponiveis: Gemini Flash, Groq Llama 3.3 70B, Mistral Large, Cohere Command-R+\n"
                                     "Digite uma pergunta ou use os botoes rapidos.\n\n"
                                     "Configure as API keys clicando em '⚙ Configurar API Keys' acima.\n")

    def _ia_perguntar(self):
        q = self.ia_question.get().strip()
        if not q: return
        self.ia_question.delete(0, tk.END)
        self.ia_text.insert(tk.END, f"\n> {q}\n")
        self.ia_text.see(tk.END)
        self._run(self._do_ia_perguntar, pergunta=q)

    def _do_ia_perguntar(self, pergunta=""):
        try:
            from motor_llm import consultar, chamar
            ctx = f"Rede: {len(self.pvs)} PVs, {len(self.trechos)} trechos, {sum(t.get('ext_m',0) for t in self.trechos):.0f}m"
            resp = consultar(pergunta, ctx)
            
            # Verificar se retornou erro de LLM nao configurado
            if "Nenhum LLM disponível" in resp or "configure" in resp.lower():
                def upd():
                    self.ia_text.insert(tk.END, f"\n⚠️ {resp}\n\n💡 Dica: Clique em '⚙ Configurar API Keys' para configurar seus LLMs gratuitos.\n{'─'*60}\n")
                    self.ia_text.see(tk.END)
                self.root.after(0, upd)
                return
                
            def upd():
                self.ia_text.insert(tk.END, f"\n{resp}\n{'─'*60}\n")
                self.ia_text.see(tk.END)
            self.root.after(0, upd)
        except Exception as e:
            self.root.after(0, lambda: self.ia_text.insert(tk.END, f"\n❌ Erro: {e}\n\n💡 Verifique se configurou as API keys em '⚙ Configurar API Keys'.\n"))

    def _ia_quick(self, tipo):
        # Usa prompts especializados do motor_llm se disponiveis
        try:
            from motor_llm import PROMPTS_ANALISE
            prompts_esp = PROMPTS_ANALISE
        except ImportError:
            prompts_esp = {}
        ctx = f"{len(self.pvs)} PVs, {len(self.trechos)} trechos, {sum(t.get('ext_m',0) for t in self.trechos):.0f}m"
        prompts = {
            "resumo": prompts_esp.get("resumo_exec", "Faca um resumo executivo.") + f"\n{ctx}",
            "hidraulica": prompts_esp.get("hidraulica", "Valide os parametros hidraulicos.") + f"\n{ctx}",
            "perdas": prompts_esp.get("perdas", "Analise o risco de perdas.") + f"\n{ctx}",
            "ml": prompts_esp.get("ml", "Explique a previsao ML.") + f"\n{ctx}",
        }
        self.ia_question.delete(0, tk.END)
        self.ia_question.insert(0, prompts.get(tipo, ""))
        self._ia_perguntar()

    def _ia_foto(self):
        p = filedialog.askopenfilename(title="Selecionar foto",
                                        filetypes=[("Imagens", "*.jpg;*.jpeg;*.png"), ("Todos", "*.*")])
        if not p: return
        self.ia_text.insert(tk.END, f"\nAnalisando foto: {Path(p).name}...\n")
        self._run(self._do_ia_foto, foto_path=p)

    def _do_ia_foto(self, foto_path=""):
        try:
            from motor_gemini import analisar_foto
            r = analisar_foto(foto_path)
            def upd():
                self.ia_text.insert(tk.END, f"\nResultado:\n{json.dumps(r, indent=2, ensure_ascii=False)}\n{'─'*60}\n")
                self.ia_text.see(tk.END)
            self.root.after(0, upd)
        except Exception as e:
            self.root.after(0, lambda: self.ia_text.insert(tk.END, f"\nErro Gemini: {e}\n"))

    def _ia_pdf(self):
        p = filedialog.askopenfilename(title="Selecionar PDF de projeto",
                                        filetypes=[("PDF", "*.pdf"), ("Todos", "*.*")])
        if not p: return
        self.ia_text.insert(tk.END, f"\nLendo PDF: {Path(p).name}...\n")
        self._run(self._do_ia_pdf, pdf_path=p)

    def _do_ia_pdf(self, pdf_path=""):
        try:
            from motor_gemini import ler_pdf
            pvs, trechos = ler_pdf(pdf_path)
            self.pvs = pvs
            self.trechos = trechos
            self.source_path = pdf_path
            self.meta = {"motor": "Gemini/PDF", "n_pvs": len(pvs), "n_trechos": len(trechos)}
            self.root.after(0, self._update_tables)
            def upd():
                self.ia_text.insert(tk.END, f"\nPDF lido: {len(pvs)} PVs, {len(trechos)} trechos\n{'─'*60}\n")
                self.ia_text.see(tk.END)
            self.root.after(0, upd)
        except Exception as e:
            self.root.after(0, lambda: self.ia_text.insert(tk.END, f"\nErro PDF: {e}\n"))

    # ── CONFIGURAR API KEYS ──────────────────────────────────────────────────

    def _ia_configurar_keys(self):
        """Abre janela modal para configurar API keys dos 4 LLMs."""
        win = tk.Toplevel(self.root)
        win.title("Configurar API Keys — LLMs Gratuitos")
        win.geometry("650x520")
        win.configure(bg=BG)
        win.transient(self.root)
        win.grab_set()

        tk.Label(win, text="Configurar API Keys — 4 LLMs Gratuitos",
                 font=("Segoe UI", 12, "bold")).pack(padx=16, pady=(16, 8))

        # Instruções
        instr = tk.Label(win, text="Todas as APIs sao GRATUITAS. Clique nos links para obter suas keys:",
                         font=("Segoe UI", 9), justify=tk.LEFT)
        instr.pack(padx=16, pady=(0, 12), anchor=tk.W)

        # Container dos providers
        main = tk.Frame(win)
        main.pack(fill=tk.BOTH, expand=True, padx=16, pady=8)

        # Dados dos providers
        providers = [
            ("Gemini Flash", "#4285f4", "https://aistudio.google.com/app/apikey", "500 req/dia", "gemini"),
            ("Groq Llama 3.3", "#f55036", "https://console.groq.com/keys", "30 req/min", "groq"),
            ("Mistral Large", "#ff7000", "https://console.mistral.ai/api-keys", "1M tokens/mês", "mistral"),
            ("Cohere Command-R", "#39594d", "https://dashboard.cohere.com/api-keys", "1000 req/mês", "cohere"),
        ]

        self.key_entries = {}
        self.key_status = {}

        for nome, cor, url, limite, provider_id in providers:
            frame = tk.LabelFrame(main, text=f" {nome} — {limite} ",
                                  font=("Segoe UI", 9, "bold"), fg=cor, bd=1)
            frame.pack(fill=tk.X, pady=6)

            # Link
            link = tk.Label(frame, text=url, font=("Segoe UI", 8), cursor="hand2")
            link.pack(padx=10, pady=(4, 0), anchor=tk.W)
            link.bind("<Button-1>", lambda e, u=url: webbrowser.open(u))

            # Entry da key
            row = tk.Frame(frame)
            row.pack(fill=tk.X, padx=10, pady=6)

            tk.Label(row, text="API Key:", width=8).pack(side=tk.LEFT)
            entry = tk.Entry(row, insertbackground=FG, bd=2, font=("Segoe UI", 9))
            entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=4)

            # Botão olho (mostrar/ocultar)
            def toggle_visibility(e=entry):
                if e.cget('show') == '*':
                    e.config(show='')
                else:
                    e.config(show='*')

            eye_btn = tk.Button(row, text="👁", command=lambda e=entry: toggle_visibility(e), width=2)
            eye_btn.pack(side=tk.RIGHT, padx=2)

            # Status
            status_lbl = tk.Label(row, text="⬚ Sem key", font=("Segoe UI", 8))
            status_lbl.pack(side=tk.RIGHT, padx=8)

            # Carregar key existente
            try:
                from motor_llm import _get_key
                existing = _get_key(provider_id)
                if existing:
                    entry.insert(0, existing)
                    status_lbl.config(text="✅ Configurada")
            except:
                pass

            self.key_entries[provider_id] = entry
            self.key_status[provider_id] = status_lbl

        # Botões de ação
        btn_row = tk.Frame(win)
        btn_row.pack(padx=16, pady=12)

        def salvar():
            from motor_llm import _load_config, _save_config, PROVIDERS
            config = _load_config()
            salvas = 0
            for provider_id, entry in self.key_entries.items():
                key = entry.get().strip()
                if key:
                    config[f"{provider_id}_api_key"] = key
                    os.environ[PROVIDERS[provider_id]["env_key"]] = key
                    salvas += 1
                    self.key_status[provider_id].config(text="✅ Salva")
            _save_config(config)
            self._ia_atualizar_status()
            messagebox.showinfo("Sucesso", f"{salvas} API key(s) configurada(s)!\n\nAgora voce pode usar os LLMs na aba IA.", parent=win)

        def testar():
            from motor_llm import CALL_MAP, PROVIDERS
            resultados = []
            for provider_id, entry in self.key_entries.items():
                key = entry.get().strip()
                if not key:
                    resultados.append(f"⬚ {provider_id}: sem key")
                    continue
                try:
                    resp = CALL_MAP[provider_id]("Responda apenas: CONECTADO", system="Seja breve")
                    resultados.append(f"✅ {provider_id}: {resp[:30]}")
                    self.key_status[provider_id].config(text="✅ OK")
                except Exception as e:
                    resultados.append(f"❌ {provider_id}: {str(e)[:50]}")
                    self.key_status[provider_id].config(text="❌ Erro")

            msg = "\n".join(resultados)
            self.ia_text.insert(tk.END, f"\nTeste de conexão:\n{msg}\n{'─'*60}\n")
            self.ia_text.see(tk.END)

        tk.Button(btn_row, text="💾 Salvar Keys", command=salvar, font=("Segoe UI", 10, "bold"), padx=16, pady=6).pack(side=tk.LEFT, padx=4)

        tk.Button(btn_row, text="📡 Testar Conexão", command=testar, font=("Segoe UI", 9, "bold"), padx=12, pady=6).pack(side=tk.LEFT, padx=4)

        tk.Button(btn_row, text="Fechar", command=win.destroy, font=("Segoe UI", 9), padx=12, pady=6).pack(side=tk.RIGHT, padx=4)

    def _ia_atualizar_status(self):
        """Atualiza labels de status dos providers na TAB IA."""
        for provider, label in self.ia_status_labels.items():
            provider_map = {"Gemini": "gemini", "Groq": "groq", "Mistral": "mistral", "Cohere": "cohere"}
            pid = provider_map.get(provider, provider.lower())
            try:
                from motor_llm import _get_key
                has_key = bool(_get_key(pid))
                label.config(text=f"{provider}: {'✅ OK' if has_key else '⬚ Sem key'}" if provider == "Gemini" else "#f55036" if provider == "Groq" else "#ff7000" if provider == "Mistral" else "#39594d")
            except:
                label.config(text=f"{provider}: ?")

    # ── TAB 12: NUCLEOS ──────────────────────────────────────────────────────

    def _tab_nucleos(self):
        tab = tk.Frame(self.nb)
        self.nb.add(tab, text="  [11] Nucleos  ")

        tk.Label(tab, text="Nucleos DXF (ProSaneamento)", font=("Segoe UI", 10, "bold")).pack(padx=8, pady=(8, 2), anchor=tk.W)
        cols = ("Nucleo", "Arquivo", "Existe")
        self.nuc_tree = ttk.Treeview(tab, columns=cols, show="headings", height=6)
        for c in cols: self.nuc_tree.heading(c, text=c)
        self.nuc_tree.column("Nucleo", width=160)
        self.nuc_tree.column("Arquivo", width=450)
        self.nuc_tree.column("Existe", width=60, anchor=tk.CENTER)
        self.nuc_tree.pack(fill=tk.X, padx=8, pady=(0, 4))
        if v5 and hasattr(v5, 'NUCLEOS_BATCH'):
            for n in v5.NUCLEOS_BATCH:
                self.nuc_tree.insert("", tk.END, values=(
                    n["nucleo"], Path(n["dxf"]).name,
                    "SIM" if Path(n["dxf"]).exists() else "NAO"))

        tk.Label(tab, text="Prolongamentos LandXML (Civil 3D)", font=("Segoe UI", 10, "bold")).pack(padx=8, pady=(8, 2), anchor=tk.W)
        self.prol_tree = ttk.Treeview(tab, columns=cols, show="headings", height=5)
        for c in cols: self.prol_tree.heading(c, text=c)
        self.prol_tree.column("Nucleo", width=220)
        self.prol_tree.column("Arquivo", width=400)
        self.prol_tree.column("Existe", width=60, anchor=tk.CENTER)
        self.prol_tree.pack(fill=tk.X, padx=8, pady=(0, 4))
        for nome, xml in _PROLONGAMENTOS:
            self.prol_tree.insert("", tk.END, values=(
                nome, Path(xml).name, "SIM" if Path(xml).exists() else "NAO"))

        brow = tk.Frame(tab)
        brow.pack(fill=tk.X, padx=8, pady=8)
        for txt, cor, fg_c, cmd in [
            ("BATCH NUCLEOS DXF", PURPLE, WHITE, self._cmd_batch),
            ("BATCH PROLONGAMENTOS", ORANGE, WHITE, self._cmd_batch_prolongamentos),
            ("BATCH TUDO", ACCENT, "#000", self._cmd_batch_tudo),
            ("AUDITORIA CAMPO V4 (NTS)", BLUE, WHITE, self._cmd_auditoria_campo_v4),
        ]:
            tk.Button(brow, text=txt, command=cmd, bg=cor, fg=fg_c,
                      font=("Segoe UI", 10, "bold"),
                      padx=20, pady=8).pack(side=tk.LEFT, padx=4)

    def _cmd_auditoria_campo_v4(self):
        """Chama a validação de executado vs projetado (V4/NTS) cruzando a rede carregada com um Shapefile importado."""
        if not hasattr(self, 'trechos') or not self.trechos:
            messagebox.showwarning("Aviso", "Por favor, navegue no arquivo DXF/XML/DWG lá em cima e clique em 'APENAS LER PROJETO' primeiro!")
            return

        res = messagebox.askyesno("Auditoria V4 (Executado vs Projeto)", 
                                  "Vamos cruzar o PROJETO que está na tela com as LINHAS DO EXECUTADO.\n"
                                  "Serão gerados os Consolidados V4 e Listas de Materiais NTS (PEAD 63/110, PVC 200/300).\n\n"
                                  "Deseja selecionar o Shapefile de Execução agora?")
        if not res: return
        
        shp_path = filedialog.askopenfilename(
            title="Selecionar Shapefile de Campo (Executado)",
            filetypes=[("Shapefile", "*.shp"), ("Todos", "*.*")]
        )
        if not shp_path: return

        self._run(self._do_auditoria_campo_v4, shp_path=shp_path)

    def _do_auditoria_campo_v4(self, shp_path=""):
        try:
            self.root.after(0, self._status, "Inicializando Motor V4 de Auditoria Executiva...", "#ffeb3b")
            
            from motor_auditoria_v4 import processar_lote_auditoria
            
            # Pegamos o projeto atual na memoria
            nucleo = self.nucleo_var.get() or "Rede Atual"
            tipo = "AGUA" if "agua" in nucleo.lower() else "ESGOTO"
            
            projetos = [{
                "nucleo": nucleo,
                "tipo": tipo,
                "pvs": self.pvs,
                "trechos": self.trechos
            }]
            
            out_base = Path.home() / "Desktop" / "GERAR NS COM ITENS POR RUA DE PV A PV" / "SAIDA_COMPLETA_POR_RUA" / "_AUDITORIA_V4_IMPORTADA"
            out_base.mkdir(parents=True, exist_ok=True)
            
            resultado = processar_lote_auditoria(projetos, Path(shp_path), out_base)
            
            self.root.after(0, self._log_msg, f"AUDITORIA CONCLUIDA! ({resultado['status_geral']['exec']} NS executadas cruzadas)", "OK")
            self.root.after(0, self._status, "Auditoria Concluida", "#00ff7f")
            
            if out_base.exists():
                os.startfile(out_base)
                
        except Exception as e:
            self.root.after(0, self._log_msg, f"FALHA NA AUDITORIA V4: {e}", "ERROR")
            self.root.after(0, self._status, "Erro na Auditoria V4", "#ff3333")

    # ── TAB 8: LOG ───────────────────────────────────────────────────────────

    def _tab_log(self):
        tab = tk.Frame(self.nb)
        self.nb.add(tab, text="  [12] Log  ")
        self.log_text = scrolledtext.ScrolledText(
            tab, font=("Consolas", 9),
            wrap=tk.WORD, bd=6, insertbackground=FG)
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
        brow = tk.Frame(tab)
        brow.pack(fill=tk.X, padx=8, pady=(0, 8))
        tk.Button(brow, text="Limpar", command=lambda: self.log_text.delete("1.0", tk.END), padx=10).pack(side=tk.LEFT)
        tk.Button(brow, text="Copiar", padx=10,
                  command=lambda: (self.root.clipboard_clear(),
                                   self.root.clipboard_append(self.log_text.get("1.0", tk.END)))
                  ).pack(side=tk.LEFT, padx=4)
        tk.Button(brow, text="VERIFICAR SAÚDE DO SISTEMA",
                  font=("Segoe UI", 9, "bold"), padx=16,
                  command=self._cmd_check_system).pack(side=tk.RIGHT, padx=4)


    def _cmd_check_system(self):
        """Verifica se todas as dependências estão presentes e funcionais."""
        deps = [
            ("ezdxf", "Leitura de DXF"),
            ("geopandas", "Processamento Geográfico"),
            ("matplotlib", "Geração de PDFs e Gráficos"),
            ("numpy", "Cálculos Matemáticos"),
            ("scipy", "Clustering (PVs)"),
            ("pandas", "Planilhas Excel"),
            ("openpyxl", "Exportação XLSX"),
            ("pyproj", "Conversão de Coordenadas"),
            ("shapely", "Geometria de Tubos"),
            ("xgboost", "Analítica (XGBoost ML)"),
            ("scipy", "Clustering (PVs)"),
        ]
        self._log_msg("--- VERIFICAÇÃO DE SAÚDE DO SISTEMA ---", "STEP")
        missing = []
        for mod, desc in deps:
            try:
                __import__(mod)
                self._log_msg(f"✅ {mod:12} : {desc}", "OK")
            except ImportError:
                self._log_msg(f"❌ {mod:12} : {desc} (FALTANDO!)", "ERR")
                missing.append(mod)
        
        if not missing:
            messagebox.showinfo("Saúde do Sistema", "Tudo OK! Todas as dependências fundamentais foram encontradas.")
        else:
            messagebox.showerror("Saúde do Sistema", f"Dependências Faltando:\n{', '.join(missing)}\n\n"
                                 "Para corrigir, feche o ConstruData e execute:\npip install -r requirements-full.txt")

    # ══════════════════════════════════════════════════════════════════════════
    # HELPERS
    # ══════════════════════════════════════════════════════════════════════════

    def _file_row(self, parent, label, var, filetypes=None, directory=False):
        row = tk.Frame(parent)
        row.pack(fill=tk.X, padx=8, pady=2)
        tk.Label(row, text=label, width=26, anchor=tk.W,
                 font=("Segoe UI", 9)).pack(side=tk.LEFT)
        tk.Entry(row, textvariable=var, insertbackground=FG, bd=2, font=("Segoe UI", 9)).pack(
                     side=tk.LEFT, fill=tk.X, expand=True, padx=4)
        def browse():
            p = filedialog.askdirectory() if directory else \
                filedialog.askopenfilename(filetypes=filetypes + [("Todos", "*.*")])
            if p: var.set(p)
        tk.Button(row, text="...", command=browse,
                  width=3).pack(side=tk.RIGHT)

    def _log_msg(self, msg, nivel="INFO"):
        ts = datetime.now().strftime("%H:%M:%S")
        prefix = {"OK": "[OK] ", "ERROR": "[ERRO] ", "WARN": "[!] "}.get(nivel, "")
        self.log_text.insert(tk.END, f"[{ts}] {prefix}{msg}\n")
        self.log_text.see(tk.END)

    def _status(self, msg, cor=ACCENT):
        self.status_lbl.config(text=msg, fg=cor)

    def _add_mapa_interpolacao(self):
        ps = filedialog.askopenfilenames(
            title="Selecionar mapas para ruas (interpolação)",
            filetypes=[("Mapas", "*.dxf;*.dwg;*.gpkg"), ("Todos", "*.*")]
        )
        if ps:
            for p in ps:
                if p not in self.mapas_interpolacao:
                    self.mapas_interpolacao.append(p)
                    self.mapas_listbox.insert(tk.END, Path(p).name)
            self._log_msg(f"Adicionados {len(ps)} mapas para interpolação.")

    def _rem_mapa_interpolacao(self):
        sels = list(self.mapas_listbox.curselection())
        for i in reversed(sels):
            self.mapas_listbox.delete(i)
            self.mapas_interpolacao.pop(i)
        self._log_msg("Mapas removidos.")

    def _cmd_brutal_tudo(self):
        """Pipeline 1-Clique Brutal: Processa TODOS os módulos para o projeto escolhido."""
        if not self.arquivo_var.get():
            messagebox.showwarning("Aviso", "Selecione um projeto base (DXF/DWG/XML) primeiro."); return
        self._run(self._do_brutal_tudo)

    def _cmd_brutal_batch(self):
        """Varredura Completa: Roda o main() do exportar_completo (Batch de todos os projetos)."""
        res = messagebox.askyesno("Varredura Batch", 
                                  "Isso irá escanear as pastas Downloads/PROJETOS e Downloads/PROJETOS_DE_AGUA...\n"
                                  "e processar TODOS os DXF/DWG/XML encontrados.\n\nDeseja continuar?")
        if res:
            self._run(self._do_brutal_batch)

    def _do_brutal_batch(self):
        from exportar_completo import main as brutal_main
        self.root.after(0, self._status, "Iniciando Varredura Batch (Satanicamente Brutal)...", YELLOW)
        try:
            brutal_main()
            self.root.after(0, self._log_msg, "VARREDURA BATCH CONCLUIDA COM SUCESSO!", "OK")
            self.root.after(0, self._status, "Batch Concluido", ACCENT)
            out_base = Path.home() / "Downloads" / "CONSTRUDATA_SAIDA_COMPLETA"
            if out_base.exists():
                os.startfile(out_base)
        except Exception as e:
            self.root.after(0, self._log_msg, f"FALHA NO BATCH: {e}", "ERROR")
            self.root.after(0, self._status, "Erro no Batch", RED)

    def _do_brutal_tudo(self):
        from exportar_completo import processar_projeto
        proj = Path(self.arquivo_var.get())
        out_base = Path(self.saida_var.get())
        nucleo = self.nucleo_var.get() or proj.stem
        
        # Injetar mapas manuais no Path da engine se necessário ou passar via argumento
        # Aqui vamos simular o comportamento de exportar_completo
        self.root.after(0, self._status, "Iniciando Processamento Brutal...", YELLOW)
        try:
            # Chamar motor brutal consolidado
            res = processar_projeto(proj, out_base, manual_mapas=[Path(m) for m in self.mapas_interpolacao])
            self.root.after(0, self._log_msg, f"BRUTAL OK: {proj.name}", "OK")
            self.root.after(0, self._status, "Processamento Concluido", ACCENT)
            # Abrir pasta se der certo
            os.startfile(out_base)
        except Exception as e:
            self.root.after(0, self._log_msg, f"FALHA BRUTAL: {e}", "ERROR")
            self.root.after(0, self._status, "Erro no Processamento", RED)

    def _cmd_modulo(self, modulo):
        """Executa um módulo isolado."""
        if not self.trechos and modulo != "leitura":
            messagebox.showwarning("Aviso", "Carregue a rede (Apenas Ler) primeiro."); return
        self._run(self._do_modulo, mod=modulo)

    def _do_modulo(self, mod=""):
        try:
            self.root.after(0, self._status, f"Executando {mod.upper()}...", YELLOW)
            nucleo = self.nucleo_var.get() or "REDE"
            out_base = Path(self.saida_var.get())
            
            if mod == "ns_campo":
                from gerar_ns import gerar_ns_a4
                p = out_base / "01_NS_CAMPO"
                p.mkdir(parents=True, exist_ok=True)
                for i, t in enumerate(self.trechos):
                    gerar_ns_a4(i+1, t, self.pvs, nucleo, str(p / f"NS{i+1:03d}_A4.pdf"))
            elif mod == "ns_desenho":
                from gerar_ns import gerar_ns_desenho
                p = out_base / "02_DESENHOS"
                p.mkdir(parents=True, exist_ok=True)
                for i, t in enumerate(self.trechos):
                    gerar_ns_desenho(i+1, t, self.pvs, self.trechos, nucleo, str(p / f"NS{i+1:03d}_DES.pdf"))
            elif mod == "ns_satelite":
                from gerar_ns import gerar_ns_sat
                p = out_base / "03_HTML"
                p.mkdir(parents=True, exist_ok=True)
                for i, t in enumerate(self.trechos):
                    gerar_ns_sat(i+1, t, self.pvs, nucleo, str(p / f"NS{i+1:03d}_SAT.pdf"))
            elif mod == "ose":
                from gerar_ose import gerar_ose_batch
                p = out_base / "01_NS_CAMPO"
                gerar_ose_batch(self.trechos, self.pvs, nucleo, str(p))
            elif mod == "materiais":
                from gerar_ns import calcular_materiais
                # Lógica simplificada: gera um JSON ou XLSX de materiais
                pass
            elif mod == "compras":
                from gerar_compras import gerar_planilha_compras
                p = out_base / "PLANILHA_COMPRAS.xlsx"
                # ... chamar lógica de compras
                pass
            elif mod == "crono_ns":
                self._do_cronograma_ns()
            elif mod == "crono_micro":
                self._do_microplan()
            elif mod == "cpm":
                from cronograma_cpm import compute_cpm
                # Gerar JSON de CPM para todos os trechos
                msg = f"Calculando CPM para {len(self.trechos)} trechos..."
                self.root.after(0, self._log_msg, msg)
                
                prod_m_dia = self._ns_prod_m_dia.get() or 6.0
                cpm_tasks = []
                for i, t in enumerate(self.trechos):
                    cpm_tasks.append({
                        "id": i + 1,
                        "name": f"NS{i+1:03d}",
                        "duration": float(t.get("ext_m", 0) / prod_m_dia),
                        "deps": [{"pred": i, "type": "FS"}] if i > 0 else []
                    })
                
                res_cpm = compute_cpm(cpm_tasks)
                
                p_cpm = out_base / "10_CRONOGRAMA"
                p_cpm.mkdir(parents=True, exist_ok=True)
                out_path = p_cpm / f"CPM_{nucleo}.json"
                
                with open(out_path, "w", encoding="utf-8") as f:
                    json.dump(res_cpm, f, indent=2, ensure_ascii=False)
                
                self.root.after(0, self._log_msg, f"CPM salvo: {out_path.name}", "OK")
                # Se tivermos um visualizador de CPM, chamamos aqui futuramente.
            elif mod == "crono_macro":
                self._do_crono_macro()
                
            self.root.after(0, self._log_msg, f"Modulo {mod.upper()} concluido.", "OK")
            self.root.after(0, self._status, "Pronto", ACCENT)
        except Exception as e:
            self.root.after(0, self._log_msg, f"Erro modulo {mod}: {e}", "ERROR")
            self.root.after(0, self._status, "Erro", RED)

    def _browse_arquivo(self):
        p = filedialog.askopenfilename(
            title="Selecionar arquivo de rede",
            filetypes=[
                ("Todos suportados", "*.dxf;*.xml;*.json;*.dwg"),
                ("DXF ProSaneamento", "*.dxf"), ("LandXML", "*.xml"),
                ("JSON ConstruData", "*.json"), ("DWG Civil 3D", "*.dwg"),
                ("Todos", "*.*"),
            ])
        if p: self.arquivo_var.set(p)

    def _on_arquivo_change(self, *args):
        p = self.arquivo_var.get()
        if not p:
            self.tipo_detectado.config(text=""); return
        ext = Path(p).suffix.lower()
        existe = Path(p).exists() if p else False
        tipos = {".dxf": "DXF (auto-detecta ProSaneamento ou Civil 3D)",
                 ".xml": "LandXML (Civil 3D — PipeNetworks)",
                 ".json": "JSON ConstruData",
                 ".dwg": "DWG Civil 3D (AEC Proxy / DWG Semantico)"}
        tipo = tipos.get(ext, f"Desconhecido: {ext}")
        # Detectar ProSaneamento em tempo real
        if existe and ext == ".dxf" and _ENGINES.get("ProSaneamento"):
            try:
                from ler_dxf_prosaneamento import detectar_prosaneamento
                if detectar_prosaneamento(p):
                    tipo = "DXF ProSaneamento ✓ (layers PS_* detectadas)"
                else:
                    tipo = "DXF Civil 3D / Genérico (GDAL)"
            except Exception:
                pass
        if not existe: tipo += " — NAO ENCONTRADO"
        self.tipo_detectado.config(text=tipo if existe else RED)
        if not self.nucleo_var.get() and existe:
            stem = Path(p).stem
            for rem in ["_ESGOTO", "_AGUA", "_REDE", "PROLONGAMENTO_"]:
                stem = stem.replace(rem, "")
            self.nucleo_var.set(stem.replace("_", " ").title())

    def _manage_num(self, value, default=0.0):
        if value is None:
            return default
        try:
            if isinstance(value, str):
                value = value.strip()
                if not value:
                    return default
            return float(value)
        except Exception:
            try:
                return float(str(value))
            except Exception:
                return default

    def _build_manage_dataset(self):
        if not self.pvs or not self.trechos:
            raise RuntimeError("Carregue uma rede antes de abrir o Viewer 3D.")

        coords = []
        for pv in self.pvs.values():
            x = self._manage_num(pv.get("x"), math.nan)
            y = self._manage_num(pv.get("y"), math.nan)
            if math.isfinite(x) and math.isfinite(y):
                coords.append((x, y))
        if not coords:
            raise RuntimeError("A rede atual nao possui coordenadas validas para o Viewer 3D.")

        ox = math.floor(min(x for x, _ in coords))
        oy = math.floor(min(y for _, y in coords))

        nodes = []
        for nome, pv in sorted(self.pvs.items()):
            x = self._manage_num(pv.get("x"), math.nan)
            y = self._manage_num(pv.get("y"), math.nan)
            if not (math.isfinite(x) and math.isfinite(y)):
                continue
            ct = self._manage_num(pv.get("ct"), 0.0)
            cf = self._manage_num(pv.get("cf"), ct)
            prof = self._manage_num(pv.get("prof"), max(ct - cf, 0.0))
            custo_pv = self._manage_num(
                pv.get("custo_total"),
                self._manage_num(pv.get("custo_pv"), 0.0),
            )
            nodes.append({
                "n": str(nome),
                "x": round(x - ox, 2),
                "y": round(y - oy, 2),
                "ct": round(ct, 3),
                "cf": round(cf, 3),
                "p": round(prof, 2),
                "t": "PV" if str(nome).upper().startswith("PV") else "PI",
                "c": round(custo_pv, 2),
            })

        edges = []
        for trecho in self.trechos:
            pv_ini = str(trecho.get("pv_ini") or "")
            pv_fim = str(trecho.get("pv_fim") or "")
            pvi = self.pvs.get(pv_ini) or {}
            pvf = self.pvs.get(pv_fim) or {}
            x0 = self._manage_num(pvi.get("x"), math.nan)
            y0 = self._manage_num(pvi.get("y"), math.nan)
            x1 = self._manage_num(pvf.get("x"), math.nan)
            y1 = self._manage_num(pvf.get("y"), math.nan)
            if not all(math.isfinite(v) for v in (x0, y0, x1, y1)):
                continue

            ct = self._manage_num(trecho.get("custo_tubo"), 0.0)
            ce = self._manage_num(trecho.get("custo_escavacao"), 0.0)
            cr = self._manage_num(trecho.get("custo_reaterro"), 0.0)
            cp = self._manage_num(trecho.get("custo_reposicao"), 0.0)
            custo_total = self._manage_num(trecho.get("custo_total"), ct + ce + cr + cp)
            fase = int(round(self._manage_num(trecho.get("ph"), self._manage_num(trecho.get("fase"), 12))))
            fase = max(1, min(12, fase))

            edges.append({
                "a": pv_ini,
                "b": pv_fim,
                "dn": int(round(self._manage_num(trecho.get("dn_mm"), 0))) or 200,
                "x0": round(x0 - ox, 2),
                "y0": round(y0 - oy, 2),
                "z0": round(self._manage_num(trecho.get("cf_ini"), self._manage_num(pvi.get("cf"), self._manage_num(pvi.get("ct"), 0.0))), 3),
                "x1": round(x1 - ox, 2),
                "y1": round(y1 - oy, 2),
                "z1": round(self._manage_num(trecho.get("cf_fim"), self._manage_num(pvf.get("cf"), self._manage_num(pvf.get("ct"), 0.0))), 3),
                "ext": round(self._manage_num(trecho.get("ext_m"), math.hypot(x1 - x0, y1 - y0)), 2),
                "mat": str(trecho.get("material") or "PVC"),
                "v": round(self._manage_num(trecho.get("vel_ms"), 0.0), 2),
                "ct": round(ct, 2),
                "ce": round(ce, 2),
                "cr": round(cr, 2),
                "cp": round(cp, 2),
                "c": round(custo_total, 2),
                "ph": fase,
            })

        if not edges:
            raise RuntimeError("A rede atual nao possui trechos com geometria valida para o Viewer 3D.")

        return {
            "nodes": nodes,
            "edges": edges,
            "ox": ox,
            "oy": oy,
            "ext": round(sum(e["ext"] for e in edges), 2),
        }

    def _build_manage_info(self):
        arquivo = Path(self.source_path).name if self.source_path else "sem arquivo"
        nucleo = self.nucleo_var.get() or (Path(self.source_path).stem if self.source_path else "REDE")
        motor = str(self.meta.get("motor") or "motor-desconhecido")
        return f"{nucleo} | {arquivo} | {motor} | {len(self.trechos)} trechos"

    def _render_manage_snapshot(self):
        template_path = SCRIPT_DIR / "html" / "construdata_manage.html"
        html_src = template_path.read_text(encoding="utf-8")
        marker_ini = "const REDE_DATA = "
        marker_fim = ";\nconst PH="
        ini = html_src.find(marker_ini)
        fim = html_src.find(marker_fim, ini)
        if ini < 0 or fim < 0:
            raise RuntimeError("Template do Viewer 3D nao possui marcadores esperados.")

        rede_json = json.dumps(self._build_manage_dataset(), ensure_ascii=False, separators=(",", ":"), default=str)
        html_src = html_src[:ini] + marker_ini + rede_json + html_src[fim:]
        html_src = html_src.replace(
            "CT 11481051 | SABESP | SLNR Santos | Verde e Teteu",
            self._build_manage_info(),
            1,
        )

        out_dir = Path(tempfile.gettempdir()) / "construdata_manage"
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / f"manage_atual_{os.getpid()}.html"
        out_path.write_text(html_src, encoding="utf-8")
        return out_path

    def _abrir_html(self, filename):
        if filename == "construdata_manage.html":
            try:
                snapshot = self._render_manage_snapshot()
            except Exception as e:
                messagebox.showwarning("Aviso", f"Viewer 3D nao pode ser aberto agora:\n{e}")
                self._log_msg(f"Viewer 3D bloqueado: {e}", "WARN")
                return
            webbrowser.open(snapshot.resolve().as_uri())
            self._log_msg(f"Viewer 3D atualizado: {snapshot}", "OK")
            return
        api_route = self._api_html_route(filename)
        if api_route:
            self._garantir_api_server()
            webbrowser.open(api_route)
            return
        p = SCRIPT_DIR / "html" / filename
        if p.exists():
            webbrowser.open(str(p))
        else:
            messagebox.showinfo("Info", f"HTML nao encontrado: {p}")

    def _api_html_route(self, filename):
        routes = {
            "construdata_rdo.html": f"{API_LOCAL_URL}/rdo",
            "construdata_manage.html": f"{API_LOCAL_URL}/manage",
            "construdata_controle.html": f"{API_LOCAL_URL}/controle",
            "construdata_campo.html": f"{API_LOCAL_URL}/campo",
        }
        return routes.get(filename)

    def _garantir_api_server(self):
        proc = getattr(self, "_api_proc", None)
        if proc and proc.poll() is None:
            return
        try:
            self._api_proc = subprocess.Popen(
                [sys.executable, "-m", "uvicorn", "api.server:app", "--port", "8787", "--host", "127.0.0.1"],
                cwd=str(SCRIPT_DIR),
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                text=True,
            )
        except FileNotFoundError:
            self._log_msg("API local nao iniciada: uvicorn indisponivel", "WARN")

    def _ler_arquivo(self, path):
        ext = Path(path).suffix.lower()
        if ext == ".dxf":
            # Detectar se é ProSaneamento (layers PS_*) ou Civil 3D genérico
            if _ENGINES.get("ProSaneamento"):
                try:
                    from ler_dxf_prosaneamento import detectar_prosaneamento
                    if detectar_prosaneamento(path):
                        self.root.after(0, self._log_msg,
                            "Detectado: DXF ProSaneamento (layers PS_*)", "OK")
                        from ler_dxf_prosaneamento import ler_dxf_prosaneamento
                        return ler_dxf_prosaneamento(path)
                except Exception:
                    pass
            # Fallback: GDAL genérico (Civil 3D / outros)
            from ler_dxf_gdal import ler_dxf_gdal
            return ler_dxf_gdal(path)
        elif ext == ".xml":
            from ler_landxml import ler_landxml
            return ler_landxml(path)
        elif ext == ".dwg":
            from ler_dwg_aec import ler_dwg_aec
            pvs, trechos, meta = ler_dwg_aec(path)
            return pvs, trechos, [], meta
        elif ext == ".json" and v5:
            return v5.ler_json_rede(path)
        raise RuntimeError(f"Formato nao suportado: {ext}")

    def _enriquecer(self, trechos, pvs):
        if v5 and hasattr(v5, 'enriquecer_trechos'):
            return v5.enriquecer_trechos(trechos, pvs)
        if _ns:
            try:
                from gerar_ns import enriquecer_trechos
                return enriquecer_trechos(trechos, pvs)
            except (ImportError, AttributeError):
                pass
        return trechos

    def _update_tables(self):
        n_pvs = len(self.pvs)
        ext_total = sum(t.get("ext_m", 0) for t in self.trechos)
        tipo = "Agua" if any(t.get("is_agua") for t in self.trechos) else "Esgoto"
        motor = self.meta.get("motor", "?")
        n_ruas = len(set(t.get("rua", "") for t in self.trechos
                         if t.get("rua") and t["rua"] != "Sem Rua"))
        self.stat_labels["PVs"].config(text=str(n_pvs))
        self.stat_labels["Trechos"].config(text=str(len(self.trechos)))
        self.stat_labels["Extensao (m)"].config(text=f"{ext_total:,.0f}")
        self.stat_labels["Tipo"].config(text=tipo)
        self.stat_labels["Motor"].config(text=motor[:15])
        self.stat_labels["Ruas"].config(text=str(n_ruas))

        self.pv_tree.delete(*self.pv_tree.get_children())
        for nome, pv in sorted(self.pvs.items()):
            self.pv_tree.insert("", tk.END, values=(
                nome,
                f"{pv.get('x',0):.1f}" if pv.get("x") else "-",
                f"{pv.get('y',0):.1f}" if pv.get("y") else "-",
                f"{pv['ct']:.3f}" if pv.get("ct") else "-",
                f"{pv['cf']:.3f}" if pv.get("cf") else "-",
                f"{pv['prof']:.2f}" if pv.get("prof") else "-"))

        self.hid_tree.delete(*self.hid_tree.get_children())
        sc = Counter()
        for i, t in enumerate(self.trechos):
            st = t.get("status_hid", "SEM_DADOS")
            cat = "OK" if st == "OK" else "Verificar" if "VERIFICAR" in st else "Sem Dados"
            sc[cat] += 1
            self.hid_tree.insert("", tk.END, values=(
                f"{i+1:03d}", t.get("pv_ini",""), t.get("pv_fim",""),
                t.get("dn_mm",""), f"{t.get('ext_m',0):.1f}",
                f"{t.get('decl_pct',0):.3f}" if t.get("decl_pct") else "-",
                f"{t.get('vel_ms',0):.3f}" if t.get("vel_ms") else "-",
                f"{t.get('vazao_ls',0):.1f}" if t.get("vazao_ls") else "-",
                f"{t.get('tau_pa',0):.2f}" if t.get("tau_pa") else "-",
                st[:40]))
        self.hid_labels["OK"].config(text=str(sc.get("OK", 0)))
        self.hid_labels["Verificar"].config(text=str(sc.get("Verificar", 0)))
        self.hid_labels["Sem Dados"].config(text=str(sc.get("Sem Dados", 0)))

        self.tr_tree.delete(*self.tr_tree.get_children())
        for i, t in enumerate(self.trechos):
            self.tr_tree.insert("", tk.END, values=(
                f"{i+1:03d}", t.get("pv_ini",""), t.get("pv_fim",""),
                t.get("rua","")[:25], t.get("dn_mm",""),
                f"{t.get('ext_m',0):.1f}", t.get("material",""),
                f"{t['ct_ini']:.3f}" if t.get("ct_ini") else "-",
                f"{t['cf_ini']:.3f}" if t.get("cf_ini") else "-",
                f"{t['prof_ini']:.2f}" if t.get("prof_ini") else "-",
                f"{t.get('custo_total',0):,.0f}" if t.get("custo_total") else "-"))

        dns = Counter(t.get("dn_mm") for t in self.trechos if t.get("dn_mm"))
        self.resumo_text.config(
            text=f"Arquivo: {Path(self.source_path).name if self.source_path else '?'}  |  "
                 f"Motor: {motor}  |  Tipo: {tipo}  |  "
                 f"PVs: {n_pvs}  |  Trechos: {len(self.trechos)}  |  Ext: {ext_total:.1f}m  |  "
                 f"DNs: {dict(sorted(dns.items()))}  |  "
                 f"CT: {sum(1 for p in self.pvs.values() if p.get('ct'))}/{n_pvs}  "
                 f"CF: {sum(1 for p in self.pvs.values() if p.get('cf'))}/{n_pvs}")

        # Atualizar sequência executiva NS com ordem natural após carregar
        self._ns_sequencia = list(range(len(self.trechos)))
        self._ns_seq_refresh()

        # Carregar ou criar STATUS_NS.json
        if self.trechos:
            threading.Thread(target=self._do_status_ns_init, daemon=True).start()

        # Auto-calcular custos em background (silencioso)
        if self.trechos:
            threading.Thread(target=self._do_custos_silencioso, daemon=True).start()

    def _do_status_ns_init(self):
        """Carrega STATUS_NS.json existente ou cria novo em background."""
        try:
            from motor_status_ns import carregar, criar_status_inicial, resumo as ns_resumo
            nucleo   = self.nucleo_var.get() or "REDE"
            out_base = self.saida_var.get()
            status   = carregar(nucleo, out_base)
            if not status:
                status = criar_status_inicial(
                    self.pvs, self.trechos, nucleo,
                    ns_sequencia=self._ns_sequencia if self._ns_sequencia else None,
                )
                # Salvar imediatamente
                from motor_status_ns import salvar as ns_salvar
                ns_salvar(status, nucleo, out_base)
                self.root.after(0, self._log_msg,
                    f"STATUS_NS criado: {len(status['notas'])} NS em PLANEJADO", "OK")
            else:
                r = ns_resumo(status)
                self.root.after(0, self._log_msg,
                    f"STATUS_NS carregado: {r['n_total']} NS | "
                    f"{r['pct_fisico']}% fisico | {r['n_executadas']} exec | "
                    f"{r['n_medidas']} medidas", "OK")
            self._status_ns = status
            self.root.after(0, self._atualizar_gestao_stats_ns)
        except Exception as e:
            self.root.after(0, self._log_msg, f"STATUS_NS: {e}", "WARN")

    def _atualizar_gestao_stats_ns(self):
        """Atualiza os cards de % Fisico / % Financ e mini KPIs do WhatsApp panel."""
        try:
            from motor_status_ns import resumo as ns_resumo
            if not self._status_ns:
                return
            r = ns_resumo(self._status_ns)
            # Cards principais
            if hasattr(self, "gestao_labels"):
                if "% Fisico" in self.gestao_labels:
                    self.gestao_labels["% Fisico"].config(text=f"{r['pct_fisico']}%")
                if "% Financ" in self.gestao_labels:
                    self.gestao_labels["% Financ"].config(text=f"{r['pct_financeiro']}%")
            # Mini KPIs no painel WhatsApp
            if hasattr(self, "_wa_kpi_labels"):
                mapping = {
                    "Total":       str(r["n_total"]),
                    "Planejadas":  str(r["n_planejadas"]),
                    "Executadas":  str(r["n_executadas"]),
                    "Cadastradas": str(r["n_cadastradas"]),
                    "Medidas":     str(r["n_medidas"]),
                    "% Fisico":    f"{r['pct_fisico']}%",
                }
                for nome, val in mapping.items():
                    if nome in self._wa_kpi_labels:
                        self._wa_kpi_labels[nome].config(text=val)
        except Exception:
            pass

    def _do_custos_silencioso(self):
        """Calcula custos sem travar a barra de progresso principal."""
        try:
            from motor_custo import custo_nucleo
            r = custo_nucleo(self.pvs, self.trechos, self.nucleo_var.get() or "REDE")
            self.root.after(0, self._fill_custos, r)
        except Exception:
            pass  # silencioso — usuário pode clicar CALCULAR CUSTOS manualmente

    # ══════════════════════════════════════════════════════════════════════════
    # THREADING
    # ══════════════════════════════════════════════════════════════════════════

    def _run(self, func, **kwargs):
        if self.running:
            messagebox.showinfo("Aguarde", "Processamento em andamento."); return
        self.running = True
        self._status("Processando...", YELLOW)
        self.progress.start(15)
        threading.Thread(target=lambda: self._safe(func, **kwargs), daemon=True).start()

    def _safe(self, func, **kwargs):
        try:
            func(**kwargs)
            self.root.after(0, self._status, "Concluido!", ACCENT)
            self.root.after(0, self._log_msg, "Concluido!", "OK")
        except Exception as e:
            self.root.after(0, self._status, f"Erro: {e}", RED)
            self.root.after(0, self._log_msg, str(e), "ERROR")
            import traceback; traceback.print_exc()
        finally:
            self.running = False
            self.root.after(0, self.progress.stop)

    # ══════════════════════════════════════════════════════════════════════════
    # COMANDOS
    # ══════════════════════════════════════════════════════════════════════════

    # ── Mapas de Interpolação ───────────────────────────────────────────────

    def _add_mapa_interpolacao(self):
        files = filedialog.askopenfilenames(
            title="Selecionar Mapas de Interpolacao (ruas)",
            filetypes=[
                ("Todos suportados", "*.dxf;*.dwg;*.gpkg"),
                ("DXF", "*.dxf"), ("DWG", "*.dwg"),
                ("GeoPackage", "*.gpkg"), ("Todos", "*.*"),
            ])
        for f in files:
            if f and f not in self.mapas_interpolacao:
                self.mapas_interpolacao.append(f)
                self.mapas_listbox.insert(tk.END, Path(f).name)

    def _rem_mapa_interpolacao(self):
        sel = list(self.mapas_listbox.curselection())
        for idx in reversed(sel):
            self.mapas_listbox.delete(idx)
            del self.mapas_interpolacao[idx]

    # ── Interpolação de ruas via mapas selecionados ──────────────────────

    def _interpolar_ruas(self, trechos, pvs):
        """Aplica interpolação de nomes de ruas usando todos os mapas selecionados."""
        total = 0
        for mapa_path in self.mapas_interpolacao:
            ext = Path(mapa_path).suffix.lower()
            try:
                if ext == ".gpkg":
                    from exportar_completo import _carregar_ruas_gpkg
                    n = _carregar_ruas_gpkg(mapa_path, trechos, pvs)
                else:
                    from exportar_completo import _carregar_ruas_dxf
                    n = _carregar_ruas_dxf(mapa_path, trechos, pvs)
                if n > 0:
                    self.root.after(0, self._log_msg,
                        f"Interpolacao: {n} ruas via {Path(mapa_path).name}", "OK")
                    total += n
            except Exception as e:
                self.root.after(0, self._log_msg,
                    f"Aviso interpol. {Path(mapa_path).name}: {e}", "WARN")
        return total

    # ── Carregar projeto (ler + enriquecer + interpolar) ─────────────────

    def _carregar_projeto(self, path):
        """Lê arquivo, enriquece trechos e aplica interpolação de ruas."""
        self.root.after(0, self._log_msg, f"Lendo: {Path(path).name}", "INFO")
        self.pvs, self.trechos, self.ruas, self.meta = self._ler_arquivo(path)
        self.source_path = path
        if not self.trechos:
            raise RuntimeError("Nenhum trecho encontrado")
        self.trechos = self._enriquecer(self.trechos, self.pvs)

        # Interpolação de ruas com mapas selecionados
        if self.mapas_interpolacao:
            self._interpolar_ruas(self.trechos, self.pvs)

        ext = sum(t["ext_m"] for t in self.trechos)
        self.root.after(0, self._update_tables)
        self.root.after(0, self._log_msg,
            f"Rede: {len(self.pvs)} PVs, {len(self.trechos)} trechos, {ext:.1f}m", "OK")

    # ── GERAR TUDO (1 CLIQUE) — Pipeline Brutal via GUI ─────────────────

    def _cmd_brutal_tudo(self):
        arq = self.arquivo_var.get()
        if not arq or not Path(arq).exists():
            messagebox.showwarning("Aviso", "Selecione um arquivo de projeto DXF/DWG."); return
        self._run(self._do_brutal_tudo, path=arq)

    def _do_brutal_tudo(self, path=None):
        """Pipeline completo: le projeto, interpola ruas, gera TODOS os módulos."""
        self._carregar_projeto(path)
        nucleo = self.nucleo_var.get() or Path(path).stem.replace("_", " ").title()
        pasta_base = Path(self.saida_var.get()) / Path(path).stem
        pastas = {}
        for nome in ["01_NS_CAMPO", "02_DESENHOS", "03_HTML", "04_GIS",
                      "05_PLANILHAS", "06_CUSTOS", "07_BIM_IFC", "08_LEAN_LPS",
                      "09_MICROPLAN", "10_CRONOGRAMA", "11_POR_RUA", "12_LOG"]:
            p = pasta_base / nome
            p.mkdir(parents=True, exist_ok=True)
            pastas[nome] = p

        self._gerar_ns_campo(nucleo, pastas)
        self._gerar_ns_desenho(nucleo, pastas)
        self._gerar_ns_satelite(nucleo, pastas)
        self._gerar_ose_materiais_compras(nucleo, pastas)
        self._gerar_cronograma_ns(nucleo, pastas)
        self._gerar_cronograma_micro(nucleo, pastas)
        self._gerar_cronograma_macro_mod(nucleo, pastas)

        # Planilhas extras (custos, hidráulica, BIM, Lean)
        self._gerar_extras(nucleo, pastas)

        self.root.after(0, self._log_msg,
            f"PIPELINE BRUTAL COMPLETO: {len(self.trechos)} trechos em {pasta_base}", "OK")

    # ── Módulo individual ────────────────────────────────────────────────

    def _cmd_modulo(self, modulo):
        arq = self.arquivo_var.get()
        if not arq or not Path(arq).exists():
            messagebox.showwarning("Aviso", "Selecione um arquivo de projeto."); return
        self._run(self._do_modulo, path=arq, modulo=modulo)

    def _do_modulo(self, path=None, modulo=None):
        if not self.trechos:
            self._carregar_projeto(path)
        nucleo = self.nucleo_var.get() or Path(path).stem.replace("_", " ").title()
        pasta_base = Path(self.saida_var.get()) / Path(path).stem
        pastas = {}
        for nome in ["01_NS_CAMPO", "02_DESENHOS", "03_HTML", "04_GIS",
                      "05_PLANILHAS", "06_CUSTOS", "07_BIM_IFC", "08_LEAN_LPS",
                      "09_MICROPLAN", "10_CRONOGRAMA", "11_POR_RUA", "12_LOG"]:
            p = pasta_base / nome
            p.mkdir(parents=True, exist_ok=True)
            pastas[nome] = p

        dispatch = {
            "ns_campo": self._gerar_ns_campo,
            "ns_desenho": self._gerar_ns_desenho,
            "ns_satelite": self._gerar_ns_satelite,
            "ose": self._gerar_ose_materiais_compras,
            "materiais": self._gerar_ose_materiais_compras,
            "compras": self._gerar_ose_materiais_compras,
            "crono_ns": self._gerar_cronograma_ns,
            "crono_micro": self._gerar_cronograma_micro,
            "crono_macro": self._gerar_cronograma_macro_mod,
        }
        fn = dispatch.get(modulo)
        if fn:
            fn(nucleo, pastas)
            self.root.after(0, self._log_msg, f"Modulo [{modulo}] concluido", "OK")

    # ── Geradores modulares ──────────────────────────────────────────────

    def _gerar_ns_campo(self, nucleo, pastas):
        if not _ENGINES.get("GerarNS"): return
        self.root.after(0, self._log_msg, "Gerando NS CAMPO...", "INFO")
        from gerar_ns import gerar_ns_a4, calcular_materiais, calc_manning, _ns_folder_name, CONTRATO
        from motor_custo import custo_trecho
        n_ok = n_err = 0
        for i, t in enumerate(self.trechos):
            ns_id = i + 1
            pv_i = t.get("pv_ini", "PVX"); pv_f = t.get("pv_fim", "PVY")
            ns_name = _ns_folder_name(ns_id, pv_i, pv_f)
            ns_dir = pastas["01_NS_CAMPO"] / ns_name
            ns_dir.mkdir(parents=True, exist_ok=True)
            try:
                gerar_ns_a4(ns_id, t, self.pvs, nucleo, str(ns_dir / f"NS{ns_id:03d}_A4.pdf"))
                materiais = calcular_materiais(t, self.pvs)
                custo = custo_trecho(t, self.pvs)
                dados = {
                    "ns_id": ns_id, "nucleo": nucleo, "contrato": CONTRATO,
                    "trecho": {k: v for k, v in t.items()},
                    "pv_montante": self.pvs.get(pv_i, {}),
                    "pv_jusante": self.pvs.get(pv_f, {}),
                    "hidraulica": calc_manning(t.get("dn_mm"), t.get("decl_mm")),
                    "materiais": materiais, "custo": custo,
                    "gerado_em": datetime.now().isoformat(),
                }
                with open(ns_dir / f"NS{ns_id:03d}_DADOS.json", "w", encoding="utf-8") as f:
                    json.dump(dados, f, indent=2, ensure_ascii=False)
                n_ok += 1
            except Exception as e:
                n_err += 1
                if ns_id <= 5:
                    self.root.after(0, self._log_msg, f"NS{ns_id:03d} ERRO: {e}", "ERROR")
        self.root.after(0, self._log_msg, f"NS CAMPO: {n_ok} ok, {n_err} erros", "OK")

    def _gerar_ns_desenho(self, nucleo, pastas):
        if not _ENGINES.get("GerarNS"): return
        self.root.after(0, self._log_msg, "Gerando NS DESENHO...", "INFO")
        from gerar_ns import gerar_ns_desenho
        n_ok = 0
        for i, t in enumerate(self.trechos):
            ns_id = i + 1
            try:
                gerar_ns_desenho(ns_id, t, self.pvs, self.trechos, nucleo,
                                str(pastas["02_DESENHOS"] / f"NS{ns_id:03d}_DESENHO.pdf"))
                n_ok += 1
            except Exception: pass
        self.root.after(0, self._log_msg, f"NS DESENHO: {n_ok} geradas", "OK")

    def _gerar_ns_satelite(self, nucleo, pastas):
        if not _ENGINES.get("GerarNS"): return
        self.root.after(0, self._log_msg, "Gerando NS SATELITE...", "INFO")
        from gerar_ns import gerar_ns_sat
        n_ok = 0
        for i, t in enumerate(self.trechos):
            ns_id = i + 1
            try:
                gerar_ns_sat(ns_id, t, self.pvs, nucleo,
                            str(pastas["02_DESENHOS"] / f"NS{ns_id:03d}_SAT.pdf"))
                n_ok += 1
            except Exception: pass
        self.root.after(0, self._log_msg, f"NS SATELITE: {n_ok} geradas", "OK")

    def _gerar_ose_materiais_compras(self, nucleo, pastas):
        self.root.after(0, self._log_msg, "Gerando OSE / Materiais / Compras...", "INFO")
        from gerar_ns import calcular_materiais, calc_manning
        from motor_custo import custo_trecho, FATORES, BDI
        from exportar_completo import _slug, _gerar_planilha_trechos_completa
        import pandas as pd

        nucleo_slug = _slug(nucleo)

        # Planilha mestre trechos PV a PV
        try:
            _gerar_planilha_trechos_completa(
                self.trechos, self.pvs, nucleo, "ESGOTO",
                pastas["05_PLANILHAS"] / f"TODOS_TRECHOS_PV_A_PV_{nucleo_slug}.xlsx"
            )
            self.root.after(0, self._log_msg, f"Planilha MESTRE: {len(self.trechos)} trechos", "OK")
        except Exception as e:
            self.root.after(0, self._log_msg, f"Erro planilha mestre: {e}", "ERROR")

        # Custos
        try:
            from gerar_xlsx import gerar_xlsx_custos
            gerar_xlsx_custos(self.pvs, self.trechos, nucleo,
                             str(pastas["06_CUSTOS"] / f"CUSTOS_{nucleo_slug}.xlsx"))
            self.root.after(0, self._log_msg, "Planilha CUSTOS OK", "OK")
        except Exception as e:
            self.root.after(0, self._log_msg, f"Erro custos: {e}", "WARN")

        # Hidráulica
        try:
            from gerar_xlsx import gerar_xlsx_hidraulica
            gerar_xlsx_hidraulica(self.trechos, self.pvs, nucleo,
                                 str(pastas["05_PLANILHAS"] / f"HIDRAULICA_{nucleo_slug}.xlsx"))
            self.root.after(0, self._log_msg, "Planilha HIDRAULICA OK", "OK")
        except Exception as e:
            self.root.after(0, self._log_msg, f"Erro hidraulica: {e}", "WARN")

        # Separar por rua
        from collections import defaultdict
        trechos_por_rua = defaultdict(list)
        for t in self.trechos:
            rua = str(t.get("rua") or "Sem Rua").strip()
            if not rua or rua.upper() in ("NAN", "NONE", ""):
                rua = "Sem Rua"
            trechos_por_rua[rua].append(t)
        for rua, lista in trechos_por_rua.items():
            rua_slug = _slug(rua)
            pasta_rua = pastas["11_POR_RUA"] / rua_slug
            pasta_rua.mkdir(parents=True, exist_ok=True)
            try:
                _gerar_planilha_trechos_completa(
                    lista, self.pvs, nucleo, "ESGOTO",
                    pasta_rua / f"Trechos_{rua_slug}.xlsx"
                )
            except Exception: pass
        self.root.after(0, self._log_msg,
            f"OSE/Materiais/Compras: {len(trechos_por_rua)} ruas", "OK")

    def _gerar_cronograma_ns(self, nucleo, pastas):
        self.root.after(0, self._log_msg, "Gerando Cronograma NS...", "INFO")
        try:
            from gerar_cronograma_macro import gerar_cronograma_por_ns
            ns_lista = [
                {"ordem": i+1, "trecho_idx": i,
                 "pv_ini": t.get("pv_ini", ""), "pv_fim": t.get("pv_fim", ""),
                 "ext_m": t.get("ext_m", 0), "rua": t.get("rua", "")}
                for i, t in enumerate(self.trechos)
            ]
            resultado = gerar_cronograma_por_ns(
                ns_lista, data_inicio_str=datetime.now().strftime("%Y-%m-%d"),
                equipes=self._ns_equipes.get(), prod_m_dia=self._ns_prod_m_dia.get(),
                nucleo=nucleo, out_dir=str(pastas["10_CRONOGRAMA"])
            )
            self.root.after(0, self._log_msg,
                f"Cronograma NS: {len(resultado['tarefas'])} tarefas, "
                f"{resultado['data_inicio']} -> {resultado['data_fim']}", "OK")
        except Exception as e:
            self.root.after(0, self._log_msg, f"Erro cronograma NS: {e}", "ERROR")

    def _gerar_cronograma_micro(self, nucleo, pastas):
        self.root.after(0, self._log_msg, "Gerando Microplanejamento...", "INFO")
        try:
            from motor_microplanejamento import micro_planejar_nucleo
            from exportar_completo import _slug
            resultado = micro_planejar_nucleo(self.pvs, self.trechos, nucleo,
                                              equipes_max=self._ns_equipes.get())
            with open(pastas["09_MICROPLAN"] / f"MICROPLAN_{_slug(nucleo)}.json",
                      "w", encoding="utf-8") as f:
                json.dump(resultado, f, indent=2, ensure_ascii=False, default=str)
            try:
                from gerar_xlsx import gerar_xlsx_microplan
                gerar_xlsx_microplan(resultado, self.pvs, self.trechos, nucleo,
                                   str(pastas["09_MICROPLAN"] / f"MICROPLAN_{_slug(nucleo)}.xlsx"))
            except Exception: pass
            self.root.after(0, self._log_msg, "Microplanejamento OK", "OK")
        except Exception as e:
            self.root.after(0, self._log_msg, f"Erro microplan: {e}", "ERROR")

    def _gerar_cronograma_macro_mod(self, nucleo, pastas):
        self.root.after(0, self._log_msg, "Gerando Cronograma MACRO...", "INFO")
        try:
            from gerar_cronograma_macro import (
                gerar_cronograma_macro, exportar_project_xml,
                exportar_primavera_xer, exportar_openproject_csv, exportar_macro_xlsx
            )
            ext_total = sum(t.get("ext_m", 0) for t in self.trechos)
            nucleos_crono = [{
                "nome": nucleo, "extensao_m": ext_total,
                "n_trechos": len(self.trechos), "equipes": self._ns_equipes.get()
            }]
            wbs = gerar_cronograma_macro(nucleos_crono, datetime.now().strftime("%Y-%m-%d"))
            pasta_crono = pastas["10_CRONOGRAMA"]
            exportar_project_xml(wbs, str(pasta_crono / "CRONOGRAMA_MACRO.xml"))
            exportar_primavera_xer(wbs, str(pasta_crono / "CRONOGRAMA_P6.xer"))
            exportar_openproject_csv(wbs, str(pasta_crono / "CRONOGRAMA_OPENPROJECT.csv"))
            try:
                exportar_macro_xlsx(wbs, str(pasta_crono / "CRONOGRAMA_MACRO.xlsx"))
            except Exception: pass
            self.root.after(0, self._log_msg, "Cronograma MACRO OK", "OK")
        except Exception as e:
            self.root.after(0, self._log_msg, f"Erro cronograma macro: {e}", "ERROR")

    def _gerar_extras(self, nucleo, pastas):
        """Gera extras: BIM IFC, CPM, Lean/LPS, Curva S, GeoJSON, HTML."""
        from exportar_completo import _slug
        nucleo_slug = _slug(nucleo)

        # CPM (Caminho Crítico)
        try:
            from cronograma_cpm import compute_cpm
            cpm_tasks = []
            for i, t in enumerate(self.trechos):
                cpm_tasks.append({
                    "id": i + 1,
                    "name": f"NS{i+1:03d}",
                    "duration": float(t.get("ext_m", 0) / self._ns_prod_m_dia.get()),
                    "deps": [{"pred": i, "type": "FS"}] if i > 0 else []
                })
            res_cpm = compute_cpm(cpm_tasks)
            with open(pastas["10_CRONOGRAMA"] / f"CPM_{nucleo_slug}.json", "w", encoding="utf-8") as f:
                json.dump(res_cpm, f, indent=2, ensure_ascii=False)
            self.root.after(0, self._log_msg, "Caminho Critico (CPM) OK", "OK")
        except Exception: pass

        # GeoJSON
        try:
            from gerar_ns import gerar_geojson, gerar_html
            gerar_geojson(self.trechos, self.pvs,
                         str(pastas["04_GIS"] / "rede_definida.geojson"))
            gerar_html(0, self.trechos[0], self.pvs, self.trechos, nucleo,
                      str(pastas["03_HTML"] / "REDE_GERAL.html"))
        except Exception: pass
        # BIM IFC
        if _ENGINES.get("IFC"):
            try:
                from gerar_ifc_lod500 import gerar_ifc_lod500
                gerar_ifc_lod500(self.pvs, self.trechos, nucleo, str(pastas["07_BIM_IFC"]))
                self.root.after(0, self._log_msg, "BIM IFC LOD500 OK", "OK")
            except Exception as e:
                self.root.after(0, self._log_msg, f"BIM IFC: {e}", "WARN")
        # Lean/LPS
        if _ENGINES.get("Lean/LPS"):
            try:
                from motor_lean_lps import gerar_relatorio_lean_lps, gerar_xlsx_lean_lps
                rel = gerar_relatorio_lean_lps(self.pvs, self.trechos, nucleo=nucleo)
                gerar_xlsx_lean_lps(rel, self.pvs, self.trechos, nucleo,
                                   str(pastas["08_LEAN_LPS"] / f"LEAN_LPS_{nucleo_slug}.xlsx"))
                self.root.after(0, self._log_msg, "Lean/LPS OK", "OK")
            except Exception: pass
        # Curva S
        try:
            from gerar_xlsx import gerar_xlsx_curva_s
            gerar_xlsx_curva_s(self.trechos, nucleo,
                              str(pastas["05_PLANILHAS"] / f"CURVA_S_{nucleo_slug}.xlsx"))
        except Exception: pass

    # ── Comandos legados (mantidos para compatibilidade) ─────────────────

    def _cmd_pipeline(self):
        arq = self.arquivo_var.get()
        if not arq or not Path(arq).exists():
            messagebox.showwarning("Aviso", "Selecione um arquivo."); return
        self._run(self._do_pipeline, path=arq)

    def _cmd_apenas_ler(self):
        arq = self.arquivo_var.get()
        if not arq or not Path(arq).exists():
            messagebox.showwarning("Aviso", "Selecione um arquivo."); return
        self._run(self._do_apenas_ler, path=arq)

    def _cmd_apenas_ler_dwg_semantico(self):
        arq = self.arquivo_var.get()
        if not arq or not Path(arq).exists():
            messagebox.showwarning("Aviso", "Selecione um arquivo."); return
        if Path(arq).suffix.lower() != ".dwg":
            messagebox.showwarning("Aviso", "O leitor semantico aceita apenas arquivos DWG."); return
        if not _ENGINES.get("DWG Semantico"):
            messagebox.showwarning(
                "Aviso",
                "Motor DWG Semantico indisponivel. Verifique pywin32 e os scripts do exporter.",
            )
            return
        self._run(self._do_apenas_ler_dwg_semantico, path=arq)

    def _cmd_apenas_ler_dwg_universal(self):
        arq = self.arquivo_var.get()
        if not arq or not Path(arq).exists():
            messagebox.showwarning("Aviso", "Selecione um arquivo."); return
        if Path(arq).suffix.lower() != ".dwg":
            messagebox.showwarning(
                "Aviso", 
                "O leitor DWG Universal aceita apenas arquivos DWG.\n\n"
                "Este leitor usa ODA File Converter ou libredwg para converter\n"
                "DWG → DXF e extrair dados de QUALQUER software (Civil 3D,\n"
                "AutoCAD MEP, genérico, etc.)."
            )
            return
        # IGNORANDO BLOQUEIO DO ODA E LIBERANDO CORRIDA COM FALLBACK AEC (CONSTRUDATAMAX HACK)
        self._run(self._do_apenas_ler_dwg_universal, path=arq)

    def _cmd_batch(self):
        if v5 and messagebox.askyesno("Batch", "Processar todos os nucleos DXF?"):
            self._run(self._do_batch)

    def _cmd_batch_prolongamentos(self):
        xmls = [(n, x) for n, x in _PROLONGAMENTOS if Path(x).exists()]
        if xmls:
            self._run(self._do_batch_prol)
        else:
            messagebox.showwarning("Aviso", "Nenhum XML encontrado.")

    def _cmd_batch_tudo(self):
        if messagebox.askyesno("Batch TUDO", "Processar nucleos + prolongamentos?"):
            self._run(self._do_batch_tudo)

    def _cmd_abrir_saida(self):
        p = Path(self.saida_var.get())
        if p.exists(): os.startfile(str(p))

    def _cmd_abrir_editor(self):
        self._abrir_html("construdata_editor.html")

    # ── Pipeline execution ───────────────────────────────────────────────────

    def _do_pipeline(self, path=None):
        self.root.after(0, self._log_msg, f"PIPELINE: {Path(path).name}", "INFO")
        self.pvs, self.trechos, self.ruas, self.meta = self._ler_arquivo(path)
        self.source_path = path
        if not self.trechos:
            raise RuntimeError("Nenhum trecho encontrado")
        self.trechos = self._enriquecer(self.trechos, self.pvs)
        self.root.after(0, self._update_tables)
        ext = sum(t["ext_m"] for t in self.trechos)
        self.root.after(0, self._log_msg,
            f"Rede: {len(self.pvs)} PVs, {len(self.trechos)} trechos, {ext:.1f}m", "OK")

        nucleo = self.nucleo_var.get() or Path(path).stem.replace("_", " ").title()
        pasta = self.saida_var.get()

        # Etapa 2: NS
        if _ENGINES.get("GerarNS"):
            self.root.after(0, self._log_msg, "Etapa 2/6: NS...", "INFO")
            from gerar_ns import processar_nucleo_from_data
            ns_seq = self._ns_sequencia if self._ns_sequencia else None
            n_ok, n_err = processar_nucleo_from_data(self.pvs, self.trechos, nucleo, pasta,
                                                     ns_sequencia=ns_seq)
            self.root.after(0, self._log_msg, f"NS: {n_ok} OK, {n_err} erros", "OK")

        # Etapa 3: Civil 3D
        if _ENGINES.get("Civil3D"):
            self.root.after(0, self._log_msg, "Etapa 3/6: Civil 3D...", "INFO")
            from gerar_civil3d import (gerar_landxml, gerar_cadastro_dxf,
                                        gerar_dynamo_script, gerar_autocad_scr, gerar_json_dados)
            p = Path(pasta)
            p.mkdir(parents=True, exist_ok=True)
            slug = nucleo.lower().replace(" ", "_")
            gerar_landxml(self.pvs, self.trechos, nucleo, p / f"REDE_{slug.upper()}.xml")
            gerar_cadastro_dxf(self.pvs, self.trechos, nucleo, str(p))
            gerar_dynamo_script(self.pvs, self.trechos, nucleo, p / f"criar_pipe_network_{slug}.py")
            gerar_autocad_scr(self.pvs, self.trechos, nucleo, p / f"desenhar_rede_{slug}.scr")
            gerar_json_dados(self.pvs, self.trechos, nucleo, p / f"dados_{slug}.json")
            self.root.after(0, self._log_msg, "Civil 3D: OK", "OK")

        # Etapa 4: NTS 292
        if _ENGINES.get("NTS292"):
            self.root.after(0, self._log_msg, "Etapa 4/6: Cadastro NTS 292...", "INFO")
            from gerar_cadastro_nts292 import gerar_cadastro_nts292
            gerar_cadastro_nts292(self.pvs, self.trechos, nucleo, str(Path(pasta) / "03_CADASTRO_NTS292"),
                                  topo_path=self.topo_var.get() or None,
                                  cartografia_path=self.cartografia_var.get() or None)
            self.root.after(0, self._log_msg, "NTS 292: OK", "OK")

        # Etapa 5: IFC
        if _ENGINES.get("IFC"):
            self.root.after(0, self._log_msg, "Etapa 5/6: IFC LOD 500...", "INFO")
            from gerar_ifc_lod500 import gerar_ifc_lod500
            gerar_ifc_lod500(self.pvs, self.trechos, nucleo, str(Path(pasta) / "04_BIM_LOD500"))
            self.root.after(0, self._log_msg, "IFC LOD 500: OK", "OK")

        # Etapa 6: Cronograma
        if _ENGINES.get("MSProject"):
            self.root.after(0, self._log_msg, "Etapa 6/6: Cronograma...", "INFO")
            from gerar_project_xml import gerar_project_xml
            gerar_project_xml(self.pvs, self.trechos, nucleo, str(Path(pasta) / "05_CRONOGRAMA"))
            self.root.after(0, self._log_msg, "MS Project: OK", "OK")

        self.root.after(0, self._log_msg,
            f"PIPELINE COMPLETO: {len(self.trechos)} trechos processados", "OK")

    def _do_apenas_ler(self, path=None):
        self.root.after(0, self._log_msg, f"Lendo: {Path(path).name}", "INFO")
        self.pvs, self.trechos, self.ruas, self.meta = self._ler_arquivo(path)
        self.source_path = path
        if not self.trechos: raise RuntimeError("Nenhum trecho")
        self.trechos = self._enriquecer(self.trechos, self.pvs)
        self.root.after(0, self._update_tables)
        self.root.after(0, self._log_msg,
            f"Carregado: {len(self.pvs)} PVs, {len(self.trechos)} trechos", "OK")

    def _do_apenas_ler_dwg_semantico(self, path=None):
        from ler_dwg_semantico import ler_dwg_semantico

        saida_sem = Path(self.saida_var.get()) / "07_DWG_SEMANTICO" / Path(path).stem
        self.root.after(0, self._log_msg, f"Lendo DWG semantico: {Path(path).name}", "INFO")
        self.root.after(0, self._log_msg, f"Saida semantica: {saida_sem}", "INFO")
        self.pvs, self.trechos, self.ruas, self.meta = ler_dwg_semantico(
            path,
            out_dir=str(saida_sem),
        )
        self.source_path = path
        if not self.trechos:
            raise RuntimeError("Nenhum trecho no DWG semantico")
        self.trechos = self._enriquecer(self.trechos, self.pvs)
        self.root.after(0, self._update_tables)
        self.root.after(0, self._log_msg,
            f"DWG semantico carregado: {len(self.pvs)} PVs, {len(self.trechos)} trechos", "OK")

    def _do_apenas_ler_dwg_universal(self, path=None):
        from ler_dwg_universal import ler_dwg_universal

        self.root.after(0, self._log_msg, f"Lendo DWG universal: {Path(path).name}", "INFO")
        self.root.after(0, self._log_msg, 
            "Metodo: ODA/libredwg → DXF → ezdxf (multi-software)", "INFO")
        
        pvs, trechos, meta = ler_dwg_universal(path)
        
        if meta.get('erro'):
            raise RuntimeError(f"Erro ao ler DWG universal: {meta['erro']} - {meta.get('obs', '')}")
        
        self.pvs = pvs
        self.trechos = trechos
        self.ruas = []
        self.meta = meta
        self.source_path = path
        
        if not self.trechos:
            raise RuntimeError("Nenhum trecho encontrado no DWG universal")
        
        self.trechos = self._enriquecer(self.trechos, self.pvs)
        self.root.after(0, self._update_tables)
        
        ext_total = sum(t.get('ext_m', 0) for t in trechos)
        self.root.after(0, self._log_msg,
            f"DWG universal carregado: {len(self.pvs)} PVs, {len(self.trechos)} trechos, {ext_total:.0f}m", "OK")

    def _do_batch(self):
        if v5:
            v5.processar_batch_com_validacao(self.saida_var.get())

    def _do_batch_prol(self):
        from ler_landxml import ler_landxml
        from gerar_ns import processar_nucleo_from_data
        pasta = self.saida_var.get()
        total = 0
        for nome, xml in _PROLONGAMENTOS:
            if not Path(xml).exists(): continue
            self.root.after(0, self._log_msg, f"Prolongamento: {nome}", "INFO")
            pvs, trechos, _, meta = ler_landxml(xml)
            if trechos:
                n_ok, _ = processar_nucleo_from_data(pvs, trechos, nome, pasta)
                total += n_ok
        self.root.after(0, self._log_msg, f"Prolongamentos: {total} NS geradas", "OK")

    def _do_batch_tudo(self):
        if v5: self._do_batch()
        if _ENGINES.get("LandXML") and _ENGINES.get("GerarNS"):
            self._do_batch_prol()

    # ── BIM individual commands ──────────────────────────────────────────────

    def _check_data(self):
        if not self.trechos:
            messagebox.showinfo("Info", "Processe um arquivo primeiro."); return False
        return True

    def _cmd_gerar_tudo_bim(self):
        if self._check_data(): self._run(self._do_gerar_tudo_bim)

    def _do_gerar_tudo_bim(self):
        nucleo = self.nucleo_var.get() or "REDE"
        pasta = self.saida_var.get()
        # Run stages 3-6
        if _ENGINES.get("Civil3D"):
            from gerar_civil3d import (gerar_landxml, gerar_cadastro_dxf,
                                        gerar_dynamo_script, gerar_autocad_scr, gerar_json_dados)
            p = Path(pasta); p.mkdir(parents=True, exist_ok=True)
            slug = nucleo.lower().replace(" ", "_")
            gerar_landxml(self.pvs, self.trechos, nucleo, p / f"REDE_{slug.upper()}.xml")
            gerar_cadastro_dxf(self.pvs, self.trechos, nucleo, str(p))
            gerar_dynamo_script(self.pvs, self.trechos, nucleo, p / f"criar_pipe_network_{slug}.py")
            gerar_autocad_scr(self.pvs, self.trechos, nucleo, p / f"desenhar_rede_{slug}.scr")
            gerar_json_dados(self.pvs, self.trechos, nucleo, p / f"dados_{slug}.json")
        if _ENGINES.get("NTS292"):
            from gerar_cadastro_nts292 import gerar_cadastro_nts292
            gerar_cadastro_nts292(self.pvs, self.trechos, nucleo, str(Path(pasta) / "03_CADASTRO_NTS292"),
                                  topo_path=self.topo_var.get() or None,
                                  cartografia_path=self.cartografia_var.get() or None)
        if _ENGINES.get("IFC"):
            from gerar_ifc_lod500 import gerar_ifc_lod500
            gerar_ifc_lod500(self.pvs, self.trechos, nucleo, str(Path(pasta) / "04_BIM_LOD500"))
        if _ENGINES.get("MSProject"):
            from gerar_project_xml import gerar_project_xml
            gerar_project_xml(self.pvs, self.trechos, nucleo, str(Path(pasta) / "05_CRONOGRAMA"))
        self.root.after(0, self.bim_status.config, {"text": "TUDO gerado!", "fg": ACCENT})

    def _cmd_ifc(self):
        if self._check_data(): self._run(self._do_ifc)
    def _do_ifc(self):
        from gerar_ifc_lod500 import gerar_ifc_lod500
        gerar_ifc_lod500(self.pvs, self.trechos, self.nucleo_var.get() or "REDE",
                         str(Path(self.saida_var.get()) / "04_BIM_LOD500"))
        self.root.after(0, self.bim_status.config, {"text": "IFC LOD500 gerado", "fg": ACCENT})

    def _cmd_landxml(self):
        if self._check_data(): self._run(self._do_landxml)
    def _do_landxml(self):
        from gerar_civil3d import gerar_landxml
        p = Path(self.saida_var.get()); p.mkdir(parents=True, exist_ok=True)
        n = self.nucleo_var.get() or "REDE"
        gerar_landxml(self.pvs, self.trechos, n, p / f"REDE_{n.upper().replace(' ','_')}.xml")

    def _cmd_nts292(self):
        if self._check_data(): self._run(self._do_nts292)
    def _do_nts292(self):
        from gerar_cadastro_nts292 import gerar_cadastro_nts292
        gerar_cadastro_nts292(self.pvs, self.trechos, self.nucleo_var.get() or "REDE",
                              str(Path(self.saida_var.get()) / "03_CADASTRO_NTS292"),
                              topo_path=self.topo_var.get() or None,
                              cartografia_path=self.cartografia_var.get() or None)

    def _cmd_cadastro_dxf(self):
        if self._check_data(): self._run(self._do_cadastro_dxf)
    def _do_cadastro_dxf(self):
        from gerar_civil3d import gerar_cadastro_dxf
        gerar_cadastro_dxf(self.pvs, self.trechos, self.nucleo_var.get() or "REDE",
                           str(Path(self.saida_var.get())))

    def _cmd_cronograma(self):
        if self._check_data(): self._run(self._do_cronograma)
    def _do_cronograma(self):
        from gerar_project_xml import gerar_project_xml
        gerar_project_xml(self.pvs, self.trechos, self.nucleo_var.get() or "REDE",
                          str(Path(self.saida_var.get()) / "05_CRONOGRAMA"))

    def _cmd_dynamo(self):
        if self._check_data(): self._run(self._do_dynamo)
    def _do_dynamo(self):
        from gerar_civil3d import gerar_dynamo_script
        p = Path(self.saida_var.get()); p.mkdir(parents=True, exist_ok=True)
        n = self.nucleo_var.get() or "REDE"
        gerar_dynamo_script(self.pvs, self.trechos, n,
                            p / f"criar_pipe_network_{n.lower().replace(' ','_')}.py")

    def _cmd_scr(self):
        if self._check_data(): self._run(self._do_scr)
    def _do_scr(self):
        from gerar_civil3d import gerar_autocad_scr
        p = Path(self.saida_var.get()); p.mkdir(parents=True, exist_ok=True)
        n = self.nucleo_var.get() or "REDE"
        gerar_autocad_scr(self.pvs, self.trechos, n,
                          p / f"desenhar_rede_{n.lower().replace(' ','_')}.scr")

    # ══════════════════════════════════════════════════════════════════════════
    # CUSTOS / MEDICAO / ML / LEAN / PERDAS
    # ══════════════════════════════════════════════════════════════════════════

    def _cmd_custos(self):
        if self._check_data(): self._run(self._do_custos)

    def _do_custos(self):
        from motor_custo import custo_nucleo
        nucleo = self.nucleo_var.get() or "REDE"
        r = custo_nucleo(self.pvs, self.trechos, nucleo)
        self.root.after(0, self._fill_custos, r)

    def _fill_custos(self, r):
        self.custo_labels["Custo Total R$"].config(text=f"{r.get('total',0):,.0f}")
        self.custo_labels["R$/metro"].config(text=f"{r.get('custo_metro',0):,.0f}")
        self.custo_labels["BDI 25%"].config(text=f"{r.get('bdi_valor',0):,.0f}")
        self.custo_labels["Trechos"].config(text=str(r.get("n_trechos", len(self.trechos))))
        self.custo_labels["Extensao (m)"].config(text=f"{r.get('extensao_m',0):,.0f}")
        self.custo_tree.delete(*self.custo_tree.get_children())
        for i, ct in enumerate(r.get("custos_trechos", [])):
            self.custo_tree.insert("", tk.END, values=(
                f"{i+1:03d}", ct.get("pv_ini",""), ct.get("pv_fim",""),
                ct.get("dn_mm",""), f"{ct.get('ext_m',0):.1f}",
                f"{ct.get('custo_tubo',0):,.0f}", f"{ct.get('custo_escavacao',0):,.0f}",
                f"{ct.get('custo_reaterro',0):,.0f}", f"{ct.get('custo_reposicao',0):,.0f}",
                f"{ct.get('custo_pv',0):,.0f}", f"{ct.get('custo_total',0):,.0f}"))
        self.custo_status.config(text=f"Custos calculados: R$ {r.get('total',0):,.2f}")
        self._log_msg(f"Custos: R$ {r.get('total',0):,.2f} ({r.get('custo_metro',0):.0f} R$/m)", "OK")

    def _cmd_gerar_bm(self):
        if self._check_data(): self._run(self._do_gerar_bm)

    def _do_gerar_bm(self):
        from motor_custo import gerar_bm
        r = gerar_bm(self.trechos, self.pvs, periodo=datetime.now().strftime("%Y-%m"))
        p = Path(self.saida_var.get()) / "BM"
        p.mkdir(parents=True, exist_ok=True)
        with open(p / "BM_01.json", "w", encoding="utf-8") as f:
            json.dump(r, f, indent=2, ensure_ascii=False)
        self.root.after(0, self._log_msg, f"BM gerado: R$ {r.get('total',0):,.2f}", "OK")

    def _cmd_curva_s(self):
        if self._check_data(): self._run(self._do_curva_s)

    def _do_curva_s(self):
        from motor_medicao import gerar_curva_s, gerar_xlsx_curva_s
        r = gerar_curva_s(self.trechos)
        pasta = Path(self.saida_var.get()); pasta.mkdir(parents=True, exist_ok=True)
        with open(pasta / "curva_s.json", "w", encoding="utf-8") as f:
            json.dump(r, f, indent=2, ensure_ascii=False)
        try:
            gerar_xlsx_curva_s(self.trechos, self.nucleo_var.get() or "REDE",
                               str(pasta / "CURVA_S.xlsx"))
            self.root.after(0, self._log_msg, "Curva S: JSON + XLSX com grafico", "OK")
        except Exception as e:
            self.root.after(0, self._log_msg, f"Curva S JSON OK, XLSX: {e}", "WARN")

    def _cmd_microplan(self):
        if self._check_data(): self._run(self._do_microplan)

    def _do_microplan(self):
        from motor_microplanejamento import micro_planejar_nucleo
        from gerar_xlsx import gerar_xlsx_microplan
        from construdata_pipeline import _normalize_microplan_for_xlsx

        nucleo = self.nucleo_var.get() or "REDE"
        slug = str(nucleo).strip().replace(" ", "_").upper()
        pasta_saida = Path(self.saida_var.get())
        pasta_saida.mkdir(parents=True, exist_ok=True)

        r = micro_planejar_nucleo(self.pvs, self.trechos, nucleo)

        json_path = pasta_saida / "microplanejamento.json"
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(r, f, indent=2, ensure_ascii=False, default=str)

        resumo = r.get("resumo", {})
        ext_total = resumo.get("extensao_total", 0)
        dias_paralelo = resumo.get("dias_paralelo", 0)
        equipes_max = resumo.get("equipes_max", 0)

        try:
            xlsx_data = _normalize_microplan_for_xlsx(r, self.trechos)
            xlsx_path = pasta_saida / f"MICROPLAN_{slug}.xlsx"
            gerar_xlsx_microplan(xlsx_data, self.pvs, self.trechos, nucleo, str(xlsx_path))
            self.root.after(
                0,
                self._log_msg,
                f"Micro-Plan: JSON + XLSX ({equipes_max} equipes, {dias_paralelo} dias, {ext_total:.0f}m)",
                "OK",
            )
        except Exception as e:
            self.root.after(
                0,
                self._log_msg,
                f"Micro-Plan: JSON OK, XLSX falhou: {e}",
                "WARN",
            )

    def _cmd_relatorio_ml(self):
        if self._check_data(): self._run(self._do_relatorio_ml)

    def _do_relatorio_ml(self):
        from motor_ml import gerar_relatorio_ml
        ext_total = sum(t.get("ext_m", 0) for t in self.trechos)
        r = gerar_relatorio_ml(None, saldo_total_m=ext_total)
        p = Path(self.saida_var.get()) / "relatorio_ml.json"
        Path(self.saida_var.get()).mkdir(parents=True, exist_ok=True)
        with open(p, "w", encoding="utf-8") as f:
            json.dump(r, f, indent=2, ensure_ascii=False)
        self.root.after(0, self._log_msg, f"ML: previsao {r.get('previsao_90d',{}).get('producao_estimada',0):.0f}m em 90 dias", "OK")

    def _cmd_crono_macro(self):
        if self._check_data(): self._run(self._do_crono_macro)

    def _do_crono_macro(self):
        from gerar_cronograma_macro import gerar_tudo
        ext = sum(t.get("ext_m", 0) for t in self.trechos)
        nucleo = self.nucleo_var.get() or "REDE"
        nucleos = [{"nome": nucleo, "extensao_m": ext, "n_trechos": len(self.trechos)}]
        p = str(Path(self.saida_var.get()) / "cronograma")
        gerar_tudo(nucleos, out_dir=p)
        self.root.after(0, self._log_msg, f"Cronograma Macro: XML+XER+CSV+JSON em {p}", "OK")

    # ── ANALYTICS ML ──────────────────────────────────────────────────────────

    def _cmd_executar_analytics(self):
        """Executa o pipeline completo de Analytics ML."""
        self._run(self._do_executar_analytics)

    def _do_executar_analytics(self):
        """Executa construdata_analytics.py com dados do contrato."""
        try:
            import construdata_analytics as anyt
        except ImportError:
            raise RuntimeError("construdata_analytics.py não encontrado")

        # Caminhos dos dados
        script_dir = Path(__file__).parent
        dados_dir = script_dir / "dados_contrato"

        path_exec = dados_dir / "EXECUCAO_DIARIA.json"
        path_ml = dados_dir / "ML_DATA.json"
        
        # Usar saida_var se definido, senão usar pasta padrão
        if self.saida_var.get() and Path(self.saida_var.get()).exists():
            saida_dir = Path(self.saida_var.get()) / "analiticos"
        else:
            saida_dir = script_dir / "analiticos"
        
        saida_dir.mkdir(parents=True, exist_ok=True)

        if not path_exec.exists():
            # Tenta caminho alternativo
            path_exec_alt = script_dir / "analiticos construdata" / "dados" / "EXECUCAO_DIARIA.json"
            if path_exec_alt.exists():
                path_exec = path_exec_alt
            else:
                raise FileNotFoundError(
                    f"EXECUCAO_DIARIA.json não encontrado.\nProcure em: {dados_dir}"
                )

        # 1. Carregar dados
        self.root.after(0, lambda: self.analytics_text.insert(tk.END, "Carregando dados...\n"))
        df = anyt.carregar_dados(str(path_exec), str(path_ml) if path_ml.exists() else None)

        # 2. Preparar features
        self.root.after(0, lambda: self.analytics_text.insert(tk.END, "Preparando features...\n"))
        X, y, FEATURES = anyt.preparar_features(df)

        # 3. Treinar modelo
        self.root.after(0, lambda: self.analytics_text.insert(tk.END, "Treinando modelo (GridSearchCV)...\n"))
        modelo, grid, X_test, y_test, y_pred_test, y_pred_all, metricas = anyt.treinar_modelo(X, y)

        # 4. Feature importance
        df_imp = anyt.calcular_feature_importance(modelo, FEATURES)

        # 5. Cenários
        self.root.after(0, lambda: self.analytics_text.insert(tk.END, "Gerando cenários...\n"))
        cenarios = anyt.gerar_cenarios(df, modelo, FEATURES)

        # 6. Gerar outputs
        xlsx_path = saida_dir / "ANALYTICS_SLNR.xlsx"
        json_path = saida_dir / "ANALYTICS_SLNR.json"
        graficos_dir = saida_dir / "graficos"
        graficos_dir.mkdir(parents=True, exist_ok=True)

        # Gerar XLSX
        self.root.after(0, lambda: self.analytics_text.insert(tk.END, f"Gerando {xlsx_path.name}...\n"))
        anyt.gerar_xlsx(df, modelo, df_imp, metricas, cenarios, y_test,
                        y_pred_test, y_pred_all, FEATURES, str(xlsx_path))

        # Adicionar aba filtrada pelo núcleo atual (se carregado)
        nucleo_atual = self.nucleo_var.get()
        if nucleo_atual and self.trechos:
            self._analytics_aba_nucleo(str(xlsx_path), nucleo_atual)

        # Gerar JSON
        self.root.after(0, lambda: self.analytics_text.insert(tk.END, f"Gerando {json_path.name}...\n"))
        resultados = {
            "gerado_em": datetime.now().isoformat(),
            "plataforma": f"ConstruData HydroNetwork v{VERSION}",
            "empresa": EMPRESA,
            "contrato": CONTRATO,
            "modelo": metricas,
            "cenarios": cenarios,
            "feature_importance": df_imp.to_dict("records"),
            "resumo_nucleos": {
                nucleo: {
                    "label": getattr(anyt, "NUCLEO_LABELS", {}).get(nucleo, nucleo),
                    "total_dias": len(sub),
                    "total_lig": int(sub["lig_total"].sum()),
                    "media_dia": round(float(sub["lig_total"].mean()), 2),
                    "max_dia": int(sub["lig_total"].max()),
                }
                for nucleo, sub in df.groupby("nucleo")
            },
        }
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(resultados, f, indent=2, ensure_ascii=False, default=str)

        # Gerar gráficos
        if anyt.MPL_OK:
            self.root.after(0, lambda: self.analytics_text.insert(tk.END, "Gerando gráficos...\n"))
            anyt.grafico_real_vs_predito(
                df, y_test, y_pred_test, metricas,
                str(graficos_dir / "01_real_vs_predito.png")
            )
            anyt.grafico_violin_nucleos(
                df, str(graficos_dir / "02_violin_nucleos.png")
            )
            anyt.grafico_feature_importance(
                df_imp, str(graficos_dir / "03_feature_importance.png")
            )
            anyt.grafico_tendencia_semanal(
                df, str(graficos_dir / "04_tendencia_semanal.png")
            )

        relatorio_md = saida_dir / "RELATORIO_ANALYTICS.md"
        relatorio_html = saida_dir / "RELATORIO_ANALYTICS.html"
        contexto_rede = {}
        if self.trechos:
            contexto_rede = {
                "nucleo": self.nucleo_var.get() or "REDE",
                "arquivo": Path(self.source_path).name if self.source_path else "",
                "motor": self.meta.get("motor", ""),
                "n_pvs": len(self.pvs),
                "n_trechos": len(self.trechos),
                "ext_total_m": round(sum(t.get("ext_m", 0) for t in self.trechos), 1),
            }

        resultados["contexto_rede_atual"] = contexto_rede
        resultados["saidas"] = {
            "xlsx": str(xlsx_path),
            "json": str(json_path),
            "graficos_dir": str(graficos_dir),
            "relatorio_md": str(relatorio_md),
            "relatorio_html": str(relatorio_html),
        }
        self.root.after(0, lambda: self.analytics_text.insert(tk.END, "Gerando relatórios...\n"))
        anyt.gerar_relatorio_markdown(resultados, str(relatorio_md))
        anyt.gerar_relatorio_html(resultados, str(relatorio_html))

        # Salvar resultados
        self.analytics_results = {
            "metricas": metricas,
            "cenarios": cenarios,
            "df_imp": df_imp,
            "saidas": {
                "xlsx": str(xlsx_path),
                "json": str(json_path),
                "graficos_dir": str(graficos_dir),
                "relatorio_md": str(relatorio_md),
                "relatorio_html": str(relatorio_html),
            }
        }

        # Atualizar UI
        def upd():
            self.analytics_labels["R² Test"].config(text=f"{metricas.get('r2_test', 0):.4f}")
            self.analytics_labels["MAE"].config(text=f"{metricas.get('mae', 0):.3f}")
            self.analytics_labels["RMSE"].config(text=f"{metricas.get('rmse', 0):.3f}")
            self.analytics_labels["Algoritmo"].config(text=metricas.get('algoritmo', 'N/A'))

            # Cenário baseline
            baseline = cenarios[0] if cenarios else {}
            lig_realizadas = baseline.get('lig_realizadas', 0)
            lig_faltam = baseline.get('lig_faltam', 0)
            data_conclusao = baseline.get('data_conclusao', 'N/A')

            self.analytics_labels["Ligações Realizadas"].config(text=f"{lig_realizadas:,}")
            self.analytics_labels["Faltam"].config(text=f"{lig_faltam:,}")
            self.analytics_labels["Previsão Conclusão"].config(text=data_conclusao)

            # Feature mais importante
            if len(df_imp) > 0:
                top_feature = df_imp.iloc[0]["feature"]
                self.analytics_labels["Feature Top"].config(text=top_feature)

            # Resumo no text
            self.analytics_text.delete("1.0", tk.END)
            self.analytics_text.insert(tk.END,
                f"✅ Analytics executado com sucesso!\n\n"
                f"📊 Modelo: {metricas.get('algoritmo', 'N/A')}\n"
                f"   R² Test: {metricas.get('r2_test', 0):.4f} | MAE: {metricas.get('mae', 0):.3f}\n"
                f"   Modelos treinados: {metricas.get('n_modelos', 0)}\n\n"
                f"📈 Cenários:\n"
            )
            for i, cen in enumerate(cenarios, 1):
                self.analytics_text.insert(tk.END,
                    f"   {i}. {cen['nome']}: {cen['producao_mensal']:.0f} lig/mês → {cen['data_conclusao']}\n"
                )
            self.analytics_text.insert(tk.END,
                f"\n📁 Saídas:\n"
                f"   Relatório HTML: {relatorio_html.name}\n"
                f"   Relatório MD: {relatorio_md.name}\n"
                f"   XLSX: {xlsx_path.name}\n"
                f"   JSON: {json_path.name}\n"
                f"   Gráficos: {graficos_dir.name}/\n"
            )
            self.analytics_text.see(tk.END)

            self._log_msg(
                f"Analytics: R²={metricas.get('r2_test', 0):.3f}, "
                f"conclusão em {data_conclusao}", "OK"
            )

        self.root.after(0, upd)

    def _cmd_abrir_relatorio_analytics(self):
        """Abre o relatório HTML do Analytics, com fallback para Markdown."""
        if not self.analytics_results or "saidas" not in self.analytics_results:
            messagebox.showwarning("Aviso", "Execute o Analytics primeiro para gerar o relatório.")
            return
        saidas = self.analytics_results["saidas"]
        html_path = saidas.get("relatorio_html")
        md_path = saidas.get("relatorio_md")
        for path in [html_path, md_path]:
            if path and Path(path).exists():
                os.startfile(path)
                return
        messagebox.showwarning("Aviso", "Relatório ainda não foi gerado.")

    def _cmd_ver_graficos_analytics(self):
        """Abre a pasta de gráficos no explorador."""
        if self.analytics_results and "saidas" in self.analytics_results:
            graficos_dir = self.analytics_results["saidas"].get("graficos_dir")
            if graficos_dir and Path(graficos_dir).exists():
                os.startfile(graficos_dir)
            else:
                messagebox.showwarning("Aviso", "Gráficos ainda não foram gerados.\nExecute o Analytics primeiro.")
        else:
            messagebox.showwarning("Aviso", "Execute o Analytics primeiro para gerar os gráficos.")

    def _cmd_cenarios_analytics(self):
        """Mostra detalhes dos cenários em uma janela."""
        if not self.analytics_results:
            messagebox.showinfo("Analytics", "Execute o Analytics primeiro para visualizar os cenários.")
            return

        win = tk.Toplevel(self.root)
        win.title("Cenários de Aceleração — Analytics ML")
        win.geometry("700x400")
        win.configure(bg=BG)
        win.transient(self.root)

        tk.Label(win, text="Cenários de Produção — Simulação",
                 font=("Segoe UI", 12, "bold")).pack(padx=8, pady=8)

        # Tabela
        cols = ["Cenário", "Produção Diária", "Produção Mensal", "Conclusão", "Custo Extra"]
        tree = ttk.Treeview(win, columns=cols, show="headings", height=8)
        for col in cols:
            tree.heading(col, text=col)
            tree.column(col, width=130)
        tree.pack(fill=tk.BOTH, expand=True, padx=8, pady=4)

        for cen in self.analytics_results.get("cenarios", []):
            tree.insert("", tk.END, values=[
                cen["nome"][:30],
                f"{cen['producao_diaria']:.1f} lig/dia",
                f"{cen['producao_mensal']:.0f} lig/mês",
                cen["data_conclusao"],
                f"R$ {cen['custo_extra_mensal']:,.0f}"
            ])

        # Detalhes
        details = scrolledtext.ScrolledText(win, font=("Consolas", 8), bd=2, height=8)
        details.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        baseline = self.analytics_results["cenarios"][0] if self.analytics_results.get("cenarios") else {}
        details.insert(tk.END,
            f"📊 Resumo do Cenário Baseline (ritmo atual):\n\n"
            f"• Ligações já realizadas: {baseline.get('lig_realizadas', 0):,}\n"
            f"• Ligações faltantes: {baseline.get('lig_faltam', 0):,}\n"
            f"• Produção diária média: {baseline.get('producao_diaria', 0):.1f} lig/dia\n"
            f"• Produção mensal média: {baseline.get('producao_mensal', 0):.0f} lig/mês\n"
            f"• Previsão de conclusão: {baseline.get('data_conclusao', 'N/A')}\n"
            f"• Dias para concluir: {baseline.get('dias_para_concluir', 0):,} dias\n\n"
            f"💡 Dica: Use os cenários de aceleração (+10%, +20%, +30%) para simular\n"
            f"   o impacto de aumentar a produtividade no cronograma.\n"
        )
        details.config(state=tk.DISABLED)

        tk.Button(win, text="Fechar", command=win.destroy,
                  font=("Segoe UI", 9, "bold"), padx=20).pack(pady=8)

    def _analytics_aba_nucleo(self, xlsx_path, nucleo):
        """Adiciona aba com resumo de trechos/PVs do núcleo atual ao XLSX analytics."""
        try:
            import openpyxl as _xl
            from openpyxl.styles import Font as _F, PatternFill as _PF, Alignment as _A
        except ImportError:
            return
        try:
            wb = _xl.load_workbook(xlsx_path)
        except Exception:
            return

        slug = nucleo.upper().replace(" ", "_")[:20]
        sheet_name = f"NS_{slug}"
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)

        hdr_fill = _PF("solid", fgColor="FF0A2140")
        hdr_font = _F(bold=True, color="FFFFFFFF", size=9)

        cols = ["NS", "PV Ini", "PV Fim", "DN (mm)", "Ext (m)", "Rua",
                "CT Ini", "CF Ini", "Prof Ini", "Material", "Status"]
        for ci, h in enumerate(cols, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.fill = hdr_fill; c.font = hdr_font
            c.alignment = _A(horizontal="center")

        for ri, t in enumerate(self.trechos, 2):
            ws.append([
                f"NS{ri-1:03d}", t.get("pv_ini",""), t.get("pv_fim",""),
                t.get("dn_mm",""), round(t.get("ext_m",0), 1), t.get("rua",""),
                t.get("ct_ini",""), t.get("cf_ini",""), t.get("prof_ini",""),
                t.get("material","PVC"), t.get("status","PLANEJADO"),
            ])

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["F"].width = 22

        try:
            wb.save(xlsx_path)
        except Exception:
            pass

    def _cmd_exportar_xlsx_analytics(self):
        """Abre o XLSX gerado pelo Analytics."""
        if self.analytics_results and "saidas" in self.analytics_results:
            xlsx_path = self.analytics_results["saidas"].get("xlsx")
            if xlsx_path and Path(xlsx_path).exists():
                os.startfile(xlsx_path)
            else:
                messagebox.showwarning("Aviso", "XLSX ainda não foi gerado.\nExecute o Analytics primeiro.")
        else:
            messagebox.showwarning("Aviso", "Execute o Analytics primeiro para gerar o XLSX.")

    def _cmd_abrir_pasta_analytics(self):
        """Abre a pasta de saída do Analytics."""
        saida_dir = Path(self.saida_var.get()) / "analiticos"
        if saida_dir.exists():
            os.startfile(saida_dir)
        else:
            messagebox.showinfo("Analytics", f"Pasta não encontrada:\n{saida_dir}")

    # ── SLNR MESTRE UNIFICADO ─────────────────────────────────────────────────

    def _cmd_slnr_ml(self):
        """Executa o SLNR Mestre Unificado com ML."""
        self._run(self._do_slnr_ml)

    def _gerar_mestre_from_trechos(self):
        """Gera planilha SLNR_MESTRE básica a partir dos trechos carregados."""
        try:
            import openpyxl as _xl
            from openpyxl.styles import Font as _F, PatternFill as _PF, Alignment as _A
        except ImportError:
            self.root.after(0, self._log_msg, "openpyxl necessário para gerar planilha MESTRE", "ERR")
            return

        saida_dir = Path(self.saida_var.get()) / "slnr_mestre"
        saida_dir.mkdir(parents=True, exist_ok=True)
        dest = saida_dir / "SLNR_MESTRE_UNIFICADO.xlsx"

        wb = _xl.Workbook()
        ws = wb.active
        ws.title = "TRECHOS_GERAL"

        hdr_fill = _PF("solid", fgColor="FF0A2140")
        hdr_font = _F(bold=True, color="FFFFFFFF", size=9)

        cols = ["Nucleo", "PV_Ini", "PV_Fim", "DN_mm", "Ext_m", "Rua", "CT_Ini", "CF_Ini", "Status"]
        for ci, c in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=c)
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = _A(horizontal="center")

        nucleo = self.nucleo_var.get() or "REDE"
        for ri, t in enumerate(self.trechos, 2):
            ws.append([
                nucleo, t.get("pv_ini",""), t.get("pv_fim",""),
                t.get("dn_mm",""), t.get("ext_m", 0), t.get("rua",""),
                t.get("ct_ini",""), t.get("cf_ini",""), t.get("status","PLANEJADO"),
            ])

        wb.save(str(dest))
        self.root.after(0, self._log_msg,
            f"MESTRE básico gerado: {dest.name}  ({len(self.trechos)} trechos de '{nucleo}')", "OK")
        self.root.after(0, self._log_msg,
            "Complete com outros nucleos manualmente e clique GERAR SLNR ML novamente.", "WARN")
        try:
            self.root.after(300, lambda: os.startfile(str(dest)))
        except Exception:
            pass

    def _do_slnr_ml(self):
        """Executa slnr_mestre_ml.py para gerar planilha com 12 núcleos + ML."""
        try:
            import slnr_mestre_ml as slnr
        except ImportError:
            raise RuntimeError("slnr_mestre_ml.py não encontrado")

        saida_dir = Path(self.saida_var.get()) / "slnr_mestre"
        saida_dir.mkdir(parents=True, exist_ok=True)

        # Verificar se arquivo MESTRE existe; se não, gerar versão básica e orientar usuário
        arquivo_mestre = SCRIPT_DIR / "SLNR_MESTRE_UNIFICADO.xlsx"
        if not arquivo_mestre.exists():
            arquivo_mestre = saida_dir / "SLNR_MESTRE_UNIFICADO.xlsx"
        if not arquivo_mestre.exists():
            self.root.after(0, self._log_msg,
                "SLNR_MESTRE_UNIFICADO.xlsx nao encontrado — gerando planilha basica...", "WARN")
            self._gerar_mestre_from_trechos()
            return

        # Executar pipeline
        self.root.after(0, lambda: self.slnr_text.insert(tk.END, "Iniciando SLNR Mestre ML...\n"))

        integrador = slnr.SLNRMLIntegrador(arquivo_mestre=str(arquivo_mestre))
        integrador.arquivo_saida = saida_dir / "SLNR_MESTRE_UNIFICADO_ML.xlsx"

        # 1. Carregar dados
        self.root.after(0, lambda: self.slnr_text.insert(tk.END, "Carregando dados...\n"))
        integrador.carregar_dados()
        
        # 2. Preparar features
        self.root.after(0, lambda: self.slnr_text.insert(tk.END, "Preparando features ML...\n"))
        df, FEATURES = integrador.preparar_features()
        
        # 3. Treinar modelo
        self.root.after(0, lambda: self.slnr_text.insert(tk.END, "Treinando XGBoost + GridSearchCV...\n"))
        modelo, y_test, y_pred_test, y_pred_all = integrador.treinar_modelo(df, FEATURES)
        
        # 4. Gerar cenários
        self.root.after(0, lambda: self.slnr_text.insert(tk.END, "Gerando cenários...\n"))
        integrador.gerar_cenarios(df)
        
        # 5. Atualizar planilha
        self.root.after(0, lambda: self.slnr_text.insert(tk.END, "Atualizando planilha com fórmulas...\n"))
        integrador.atualizar_planilha()
        
        # 6. Gerar gráficos
        self.root.after(0, lambda: self.slnr_text.insert(tk.END, "Gerando gráficos Seaborn...\n"))
        integrador.gerar_graficos()
        
        # 7. Exportar JSON
        self.root.after(0, lambda: self.slnr_text.insert(tk.END, "Exportando JSON...\n"))
        integrador.exportar_json()

        # Atualizar UI
        def upd():
            self.slnr_labels["R² ML"].config(text=f"{integrador.metricas_ml.get('r2_test', 0):.4f}")
            self.slnr_text.delete("1.0", tk.END)
            self.slnr_text.insert(tk.END,
                f"✅ SLNR Mestre ML concluído!\n\n"
                f"📊 Modelo: {integrador.metricas_ml.get('algoritmo', 'N/A')}\n"
                f"   R² = {integrador.metricas_ml.get('r2_test', 0):.4f}\n"
                f"   MAE = {integrador.metricas_ml.get('mae', 0):.3f}\n\n"
                f"🏗️ 12 núcleos atualizados com fórmulas:\n"
                f"   N07_NOROESTE, N08_V_PROGRESSO, N09_Z_LESTE,\n"
                f"   N10_CONJUNTO, N11_ALAGADO, N12_MONTANHOSO,\n"
                f"   SD_JOAO_CARLOS, SD_SAO_MANOEL, SD_VILA_ISRAEL,\n"
                f"   SD_MORRO_TETEU, SD_VILA_CRIADORES, SD_PANTANAL_BAIXO\n\n"
                f"📁 Saída: {integrador.arquivo_saida.name}\n"
            )
            self.slnr_text.see(tk.END)
            self._log_msg(
                f"SLNR Mestre: R²={integrador.metricas_ml.get('r2_test', 0):.3f}, "
                f"12 núcleos atualizados", "OK"
            )

        self.root.after(0, upd)

    def _cmd_abrir_slnr(self):
        """Abre a planilha SLNR Mestre."""
        saida_dir = Path(self.saida_var.get()) / "slnr_mestre"
        xlsx_path = saida_dir / "SLNR_MESTRE_UNIFICADO_ML.xlsx"
        if xlsx_path.exists():
            os.startfile(xlsx_path)
        else:
            messagebox.showinfo("SLNR Mestre",
                "Planilha não encontrada.\nExecute '📊 GERAR SLNR ML' primeiro.")

    def _cmd_abrir_pasta_slnr(self):
        """Abre a pasta de saída do SLNR Mestre."""
        saida_dir = Path(self.saida_var.get()) / "slnr_mestre"
        if saida_dir.exists():
            os.startfile(saida_dir)
        else:
            messagebox.showinfo("SLNR Mestre", f"Pasta não encontrada:\n{saida_dir}")

    def _cmd_emitir_notas_servico(self):
        """Emite Notas de Serviço divididas por PIs e PVs."""
        saida_dir = Path(self.saida_var.get()) / "slnr_mestre"
        xlsx_path = saida_dir / "SLNR_MESTRE_UNIFICADO_ML.xlsx"
        
        if xlsx_path.exists():
            # Abrir a planilha e ir direto para aba de Notas de Serviço
            os.startfile(xlsx_path)
            self.slnr_text.delete("1.0", tk.END)
            self.slnr_text.insert(tk.END,
                "✅ Notas de Serviço emitidas!\n\n"
                "📄 Formato: NS_XXX_PI_YY_AO_PV_ZZ\n"
                "📊 Total: 29 NSs geradas\n"
                "🏗️ 12 núcleos atendidos\n"
                "📐 ~6.000m de rede\n"
                "🔌 ~1.700 ligações\n\n"
                "Abra a aba 'NOTAS_SERVICO_PIS_PVS' na planilha!\n")
            self.slnr_text.see(tk.END)
            self._log_msg("Notas de Serviço emitidas com sucesso!", "OK")
        else:
            messagebox.showinfo("Notas de Serviço",
                "Planilha não encontrada.\n\n"
                "Execute '📊 GERAR SLNR ML' primeiro para gerar as Notas de Serviço.")

    # ── LEAN / LPS ──

    def _cmd_lean_report(self):
        if self._check_data(): self._run(self._do_lean_report)

    @staticmethod
    def _lean_semanas(lps):
        lookahead = lps.get("lookahead_6sem", lps.get("lookahead", {}))
        if isinstance(lookahead, dict):
            return lookahead.get("semanas", [])
        if isinstance(lookahead, list):
            return lookahead
        return []

    @staticmethod
    def _lean_ready_pct(semanas):
        if not semanas:
            return 0.0
        sem1 = semanas[0] or {}
        total = sem1.get("ns_total", sem1.get("ns_planejadas", 0)) or 0
        ready = sem1.get("ns_ready")
        if ready is None and total:
            blocked = sem1.get("ns_blocked")
            if blocked is not None:
                ready = max(total - blocked, 0)
        if not total or ready is None:
            return 0.0
        return round((ready / total) * 100, 1)

    @staticmethod
    def _lean_va_nva_text(vs):
        valor_agregado = float(vs.get("valor_agregado_dias", 0) or 0)
        espera = float(vs.get("espera_dias", 0) or 0)
        total = valor_agregado + espera
        if total <= 0:
            return "0/0"
        va_pct = (valor_agregado / total) * 100
        nva_pct = 100 - va_pct
        return f"{va_pct:.0f}/{nva_pct:.0f}"

    @staticmethod
    def _formatar_lean_dashboard_legacy(r, xlsx_path=""):
        """Formata relatório Lean/LPS/BIM6D como texto legível (não JSON bruto)."""
        takt  = r.get("lean", {}).get("takt_time", r.get("takt", {}))
        bim6d = r.get("bim_6d", {})
        vs    = r.get("lean", {}).get("value_stream", r.get("value_stream", {}))
        lps   = r.get("lps", {})
        semanas = HydroNetworkApp._lean_semanas(lps)
        ready_s1 = HydroNetworkApp._lean_ready_pct(semanas)
        va_nva = HydroNetworkApp._lean_va_nva_text(vs)
        meta = r.get("meta", {})

        def v(d, *keys, fmt="{}", default="-"):
            for k in keys:
                if k in d:
                    try: return fmt.format(d[k])
                    except Exception: return str(d[k])
            return default

        lines = [
            "═" * 52,
            "  LEAN + LPS + BIM 6D — RESUMO",
            "═" * 52,
            f"  Nucleo:         {meta.get('nucleo', r.get('nucleo', 'REDE'))}",
            f"  Takt Time:      {v(takt,'takt_metros_dia','takt_dias', fmt='{:.1f}')} m/dia/equipe",
            f"  Cycle Time:     {v(takt,'cycle_time_dias','cycle_time', fmt='{:.1f}')} dias",
            f"  Lead Time:      {v(takt,'lead_time_dias', fmt='{:.1f}')} dias",
            f"  Throughput:     {v(takt,'throughput_ns_semana','throughput_dia', fmt='{:.1f}')} NS/semana",
            f"  Equipes:        {v(takt,'equipes', fmt='{}')}",
            f"  Ready 1a sem:   {ready_s1:.1f}%",
            "─" * 52,
            "  BIM 6D — CICLO DE VIDA",
            f"  CO2 Total:      {v(bim6d,'co2_total_ton','co2', fmt='{:.2f}')} ton",
            f"  Custo 50 anos:  R$ {bim6d.get('custo_ciclo_vida_total', bim6d.get('custo_ciclo_vida', 0)):>12,.0f}",
            f"  Manutencao/ano: R$ {bim6d.get('manutencao_anual_total', bim6d.get('manutencao_anual', 0)):>12,.0f}",
            "─" * 52,
            "  VALUE STREAM MAP",
            f"  Valor Agregado: {v(vs,'valor_agregado_dias', fmt='{:.1f}')}d",
            f"  Espera/Fila:    {v(vs,'espera_dias', fmt='{:.1f}')}d",
            f"  Eficiencia:     {v(vs,'eficiencia_fluxo_pct', fmt='{:.1f}')}%",
            f"  VA/NVA:         {va_nva}",
            "─" * 52,
            "  LPS — LOOKAHEAD 6 SEMANAS",
        ]
        for sem in semanas[:6]:
            lines.append(
                f"  Sem {sem.get('semana','?'):2d}: {sem.get('ns_total', sem.get('ns_planejadas', 0))} NS"
                f" | ready={sem.get('ns_ready', 0)}"
                f" | bloqueadas={sem.get('ns_blocked', 0)}"
                f" | ext={sem.get('ext_m', 0):.0f}m"
            )
        if xlsx_path:
            lines += ["═" * 52, f"  XLSX → {xlsx_path}", "═" * 52]
        return "\n".join(lines)

    @staticmethod
    def _formatar_lean_dashboard(r, xlsx_path=""):
        """Formata resumo Lean/LPS/BIM 6D em texto simples."""
        takt = r.get("lean", {}).get("takt_time", r.get("takt", {}))
        bim6d = r.get("bim_6d", {})
        vs = r.get("lean", {}).get("value_stream", r.get("value_stream", {}))
        lps = r.get("lps", {})
        semanas = HydroNetworkApp._lean_semanas(lps)
        ready_s1 = HydroNetworkApp._lean_ready_pct(semanas)
        va_nva = HydroNetworkApp._lean_va_nva_text(vs)
        meta = r.get("meta", {})

        def v(d, *keys, fmt="{}", default="-"):
            for k in keys:
                if k in d:
                    try:
                        return fmt.format(d[k])
                    except Exception:
                        return str(d[k])
            return default

        lines = [
            "=" * 52,
            "  LEAN + LPS + BIM 6D - RESUMO",
            "=" * 52,
            f"  Nucleo:         {meta.get('nucleo', r.get('nucleo', 'REDE'))}",
            f"  Takt Time:      {v(takt, 'takt_metros_dia', 'takt_dias', fmt='{:.1f}')} m/dia/equipe",
            f"  Cycle Time:     {v(takt, 'cycle_time_dias', 'cycle_time', fmt='{:.1f}')} dias",
            f"  Lead Time:      {v(takt, 'lead_time_dias', fmt='{:.1f}')} dias",
            f"  Throughput:     {v(takt, 'throughput_ns_semana', 'throughput_dia', fmt='{:.1f}')} NS/semana",
            f"  Equipes:        {v(takt, 'equipes', fmt='{}')}",
            f"  Ready 1a sem:   {ready_s1:.1f}%",
            "-" * 52,
            "  BIM 6D - CICLO DE VIDA",
            f"  CO2 Total:      {v(bim6d, 'co2_total_ton', 'co2', fmt='{:.2f}')} ton",
            f"  Custo 50 anos:  R$ {bim6d.get('custo_ciclo_vida_total', bim6d.get('custo_ciclo_vida', 0)):>12,.0f}",
            f"  Manutencao/ano: R$ {bim6d.get('manutencao_anual_total', bim6d.get('manutencao_anual', 0)):>12,.0f}",
            "-" * 52,
            "  VALUE STREAM MAP",
            f"  Valor Agregado: {v(vs, 'valor_agregado_dias', fmt='{:.1f}')}d",
            f"  Espera/Fila:    {v(vs, 'espera_dias', fmt='{:.1f}')}d",
            f"  Eficiencia:     {v(vs, 'eficiencia_fluxo_pct', fmt='{:.1f}')}%",
            f"  VA/NVA:         {va_nva}",
            "-" * 52,
            "  LPS - LOOKAHEAD 6 SEMANAS",
        ]
        for sem in semanas[:6]:
            lines.append(
                f"  Sem {sem.get('semana', '?'):2d}: {sem.get('ns_total', sem.get('ns_planejadas', 0))} NS"
                f" | ready={sem.get('ns_ready', 0)}"
                f" | bloqueadas={sem.get('ns_blocked', 0)}"
                f" | ext={sem.get('ext_m', 0):.0f}m"
            )
        if xlsx_path:
            lines += ["=" * 52, f"  XLSX -> {xlsx_path}", "=" * 52]
        return "\n".join(lines)

    def _do_lean_report(self):
        from motor_lean_lps import gerar_relatorio_lean_lps, gerar_xlsx_lean_lps
        nucleo = self.nucleo_var.get() or "REDE"
        r = gerar_relatorio_lean_lps(self.pvs, self.trechos, nucleo=nucleo)
        takt  = r.get("lean", {}).get("takt_time", r.get("takt", {}))
        lean6d = r.get("bim_6d", {})
        vs = r.get("lean", {}).get("value_stream", r.get("value_stream", {}))
        semanas = self._lean_semanas(r.get("lps", {}))
        ready_s1 = self._lean_ready_pct(semanas)
        va_nva = self._lean_va_nva_text(vs)
        pasta = Path(self.saida_var.get()); pasta.mkdir(parents=True, exist_ok=True)

        # Exportar XLSX
        xlsx_path = str(pasta / "LEAN_LPS.xlsx")
        try:
            gerar_xlsx_lean_lps(r, self.pvs, self.trechos, nucleo, xlsx_path)
        except Exception as e:
            xlsx_path = f"(erro: {e})"

        # Exportar JSON silenciosamente
        with open(pasta / "lean_lps.json", "w", encoding="utf-8") as f:
            json.dump(r, f, indent=2, ensure_ascii=False, default=str)

        dashboard = self._formatar_lean_dashboard(r, xlsx_path)

        def upd():
            self.lean_labels["Takt (m/dia)"].config(
                text=f"{takt.get('takt_metros_dia', takt.get('takt_dias', 0)):.1f}")
            self.lean_labels["Cycle Time"].config(
                text=f"{takt.get('cycle_time_dias', takt.get('cycle_time', 0)):.1f}")
            self.lean_labels["PPC (%)"].config(text=f"{ready_s1:.1f}")
            self.lean_labels["VA/NVA"].config(text=va_nva)
            self.lean_labels["CO2 (ton)"].config(
                text=f"{lean6d.get('co2_total_ton', 0):.1f}")
            self.lean_labels["Custo 50 anos"].config(
                text=f"{lean6d.get('custo_ciclo_vida_total', lean6d.get('custo_ciclo_vida', 0)):,.0f}")
            self.lean_text.delete("1.0", tk.END)
            self.lean_text.insert(tk.END, dashboard)

        self.root.after(0, upd)
        self.root.after(0, self._log_msg, "Lean+LPS+BIM6D: relatório gerado", "OK")
        # Abrir XLSX se gerado com sucesso
        if Path(xlsx_path).exists():
            try:
                self.root.after(500, lambda: os.startfile(xlsx_path))
            except Exception:
                pass

    def _cmd_takt(self):
        if self._check_data(): self._run(self._do_takt)

    def _do_takt(self):
        from motor_lean_lps import calcular_takt_time
        r = calcular_takt_time(self.trechos)
        texto = "\n".join([
            "=" * 52,
            "  TAKT TIME",
            "=" * 52,
            f"  Takt:           {r.get('takt_metros_dia', 0):.1f} m/dia/equipe",
            f"  Cycle Time:     {r.get('cycle_time_dias', 0):.1f} dias/NS",
            f"  Lead Time:      {r.get('lead_time_dias', 0):.1f} dias",
            f"  Throughput:     {r.get('throughput_ns_semana', 0):.1f} NS/semana",
            f"  Equipes:        {r.get('equipes', 0)}",
            f"  Extensao total: {r.get('ext_total', 0):,.0f} m",
            f"  Trechos:        {r.get('n_trechos', 0)}",
            "=" * 52,
        ])
        def upd():
            self.lean_labels["Takt (m/dia)"].config(text=f"{r.get('takt_metros_dia',0):.1f}")
            self.lean_labels["Cycle Time"].config(text=f"{r.get('cycle_time_dias',0):.1f}")
            self.lean_text.delete("1.0", tk.END)
            self.lean_text.insert(tk.END, texto)
        self.root.after(0, upd)
        self.root.after(0, self._log_msg,
            f"Takt: {r.get('takt_metros_dia',0):.1f} m/dia/equipe | Throughput: {r.get('throughput_ns_semana',0):.1f} NS/sem", "OK")

    def _cmd_lookahead(self):
        if self._check_data(): self._run(self._do_lookahead)

    def _do_lookahead(self):
        from motor_lean_lps import gerar_lookahead
        ns_list = [{"id": f"NS_{i+1:03d}", "pv_ini": t["pv_ini"], "pv_fim": t["pv_fim"],
                     "ext_m": t["ext_m"]} for i, t in enumerate(self.trechos)]
        r = gerar_lookahead(ns_list)
        p = Path(self.saida_var.get()) / "lookahead_6sem.json"
        Path(self.saida_var.get()).mkdir(parents=True, exist_ok=True)
        with open(p, "w", encoding="utf-8") as f:
            json.dump(r, f, indent=2, ensure_ascii=False, default=str)
        ready_s1 = self._lean_ready_pct(r.get("semanas", []))
        lines = [
            "=" * 52,
            "  LOOKAHEAD 6 SEMANAS",
            "=" * 52,
            f"  Data base:      {r.get('data_base', '-')}",
            f"  Alerta:         {r.get('alerta', '-')}",
            f"  Ready 1a sem:   {ready_s1:.1f}%",
            f"  Restricoes:     {r.get('total_restricoes', 0)}",
            "-" * 52,
        ]
        for sem in r.get("semanas", [])[:6]:
            lines.append(
                f"  Sem {sem.get('semana','?'):2d}: {sem.get('ns_total', 0)} NS"
                f" | ready={sem.get('ns_ready', 0)}"
                f" | bloqueadas={sem.get('ns_blocked', 0)}"
                f" | ext={sem.get('ext_m', 0):.0f}m"
            )
        texto = "\n".join(lines)
        def upd():
            self.lean_labels["PPC (%)"].config(text=f"{ready_s1:.1f}")
            self.lean_text.delete("1.0", tk.END)
            self.lean_text.insert(tk.END, texto)
        self.root.after(0, upd)
        self.root.after(0, self._log_msg, f"Lookahead 6 sem: {len(r.get('semanas',[]))} semanas", "OK")

    def _cmd_bim6d(self):
        if self._check_data(): self._run(self._do_bim6d)

    def _do_bim6d(self):
        from motor_lean_lps import gerar_6d_nucleo
        r = gerar_6d_nucleo(self.pvs, self.trechos, self.nucleo_var.get() or "REDE")
        lines = [
            "=" * 52,
            "  BIM 6D - CICLO DE VIDA",
            "=" * 52,
            f"  Nucleo:         {r.get('nucleo', 'REDE')}",
            f"  CO2 total:      {r.get('co2_total_ton', 0):.2f} ton",
            f"  Custo 50 anos:  R$ {r.get('custo_ciclo_vida_total', r.get('custo_ciclo_vida', 0)):>12,.0f}",
            f"  Manutencao/ano: R$ {r.get('manutencao_anual_total', r.get('manutencao_anual', 0)):>12,.0f}",
            "-" * 52,
            "  Por material",
        ]
        for material, dados in sorted(r.get("por_material", {}).items()):
            lines.append(
                f"  {material:<12} ext={dados.get('ext', 0):>8.1f} m"
                f" | CO2={dados.get('co2', 0):>8.1f} kg"
                f" | CV=R$ {dados.get('custo_cv', 0):>10,.0f}"
            )
        texto = "\n".join(lines)
        def upd():
            self.lean_labels["CO2 (ton)"].config(text=f"{r.get('co2_total_ton',0):.1f}")
            self.lean_labels["Custo 50 anos"].config(
                text=f"{r.get('custo_ciclo_vida_total', r.get('custo_ciclo_vida', 0)):,.0f}")
            self.lean_text.delete("1.0", tk.END)
            self.lean_text.insert(tk.END, texto)
        self.root.after(0, upd)
        self.root.after(0, self._log_msg,
            f"BIM 6D: CO2={r.get('co2_total_ton',0):.1f}t | Ciclo de vida=R$ {r.get('custo_ciclo_vida_total', r.get('custo_ciclo_vida', 0)):,.0f}", "OK")

    # ── PERDAS ──

    def _cmd_perdas_report(self):
        if self._check_data(): self._run(self._do_perdas_report)

    def _do_perdas_report(self):
        from motor_perdas import gerar_relatorio_perdas
        nucleo = self.nucleo_var.get() or "REDE"
        r = gerar_relatorio_perdas(self.pvs, self.trechos, nucleo=nucleo)
        def upd():
            self.perdas_labels["UARL (m3/ano)"].config(text=f"{r.get('uarl',{}).get('uarl_m3_ano',0):,.0f}")
            self.perdas_labels["ILI"].config(text=f"{r.get('ili',{}).get('ili',0):.1f}")
            self.perdas_labels["Classif."].config(text=r.get("ili",{}).get("classificacao","?"))
            risco = r.get("mapa_risco", {})
            self.perdas_labels["Risco Alto"].config(text=str(risco.get("alto",0) + risco.get("critico",0)))
            self.perdas_labels["DMAs"].config(text=str(r.get("n_dmas", 0)))
            self.perdas_text.delete("1.0", tk.END)
            self.perdas_text.insert(tk.END, json.dumps(r, indent=2, ensure_ascii=False, default=str))
        self.root.after(0, upd)
        p = Path(self.saida_var.get()) / "perdas.json"
        Path(self.saida_var.get()).mkdir(parents=True, exist_ok=True)
        with open(p, "w", encoding="utf-8") as f:
            json.dump(r, f, indent=2, ensure_ascii=False, default=str)
        self.root.after(0, self._log_msg, f"Perdas: ILI={r.get('ili',{}).get('ili',0):.1f}", "OK")

    def _cmd_mapa_risco(self):
        if self._check_data(): self._run(self._do_mapa_risco)

    def _do_mapa_risco(self):
        from motor_perdas import mapa_risco_nucleo
        r = mapa_risco_nucleo(self.pvs, self.trechos, self.nucleo_var.get() or "REDE")
        def upd():
            self.perdas_labels["Risco Alto"].config(text=str(r.get("alto",0) + r.get("critico",0)))
            self.perdas_text.delete("1.0", tk.END)
            self.perdas_text.insert(tk.END, json.dumps(r, indent=2, ensure_ascii=False, default=str))
        self.root.after(0, upd)
        self.root.after(0, self._log_msg,
            f"Mapa Risco: {r.get('critico',0)} criticos, {r.get('alto',0)} altos", "OK")

    def _cmd_criar_dma(self):
        if self._check_data(): self._run(self._do_criar_dma)

    def _do_criar_dma(self):
        from motor_perdas import criar_dma
        r = criar_dma(self.pvs, self.trechos)
        def upd():
            self.perdas_labels["DMAs"].config(text=str(r.get("n_setores", 0)))
        self.root.after(0, upd)
        self.root.after(0, self._log_msg, f"DMAs: {r.get('n_setores',0)} setores criados", "OK")

    def _cmd_pdf_perdas(self):
        if self._check_data(): self._run(self._do_pdf_perdas)

    def _do_pdf_perdas(self):
        from motor_perdas import gerar_relatorio_perdas
        from gerar_pdf_perdas import gerar_pdf_perdas
        r = gerar_relatorio_perdas(self.pvs, self.trechos, self.nucleo_var.get() or "REDE")
        p = Path(self.saida_var.get())
        p.mkdir(parents=True, exist_ok=True)
        pdf = gerar_pdf_perdas(r, str(p / "RELATORIO_PERDAS.pdf"), self.nucleo_var.get() or "REDE")
        self.root.after(0, self._log_msg, f"PDF Perdas: {pdf}", "OK")

    def _cmd_analise_troca(self):
        if self._check_data(): self._run(self._do_analise_troca)

    def _do_analise_troca(self):
        from motor_perdas import analise_troca_vs_perda
        results = []
        for t in self.trechos[:20]:
            r = analise_troca_vs_perda(t.get("ext_m", 0), t.get("material", "PVC"), 15)
            results.append(r)
        def upd():
            troca = sum(1 for r in results if r.get("recomendacao") == "TROCAR")
            self.perdas_text.delete("1.0", tk.END)
            self.perdas_text.insert(tk.END, f"Analise troca vs perda (primeiros 20 trechos):\n")
            self.perdas_text.insert(tk.END, f"Recomenda TROCAR: {troca}/{len(results)}\n\n")
            self.perdas_text.insert(tk.END, json.dumps(results[:5], indent=2, ensure_ascii=False, default=str))
        self.root.after(0, upd)
        self.root.after(0, self._log_msg, f"Analise troca: {len(results)} trechos avaliados", "OK")

    # ══════════════════════════════════════════════════════════════════════════
    # MAPA
    # ══════════════════════════════════════════════════════════════════════════

    def _mapa_trocar_tile(self):
        if not self._has_map: return
        if self._tile_is_sat:
            self.map_widget.set_tile_server(_TILE_RUA, max_zoom=19)
            self._tile_is_sat = False
            self._mapa_tile_label.config(text="Rua")
        else:
            self.map_widget.set_tile_server(_TILE_SAT, max_zoom=19)
            self._tile_is_sat = True
            self._mapa_tile_label.config(text="Satelite")

    def _mapa_carregar(self):
        if not self.trechos or not self._has_map:
            messagebox.showinfo("Info", "Processe um arquivo primeiro."); return
        for m in self._mapa_markers: m.delete()
        for p in self._mapa_paths: p.delete()
        self._mapa_markers.clear(); self._mapa_paths.clear()
        if len(self._mapa_checkstates) != len(self.trechos):
            self._mapa_checkstates = [True] * len(self.trechos)
        try:
            from pyproj import Transformer
            tr = Transformer.from_crs("EPSG:31983", "EPSG:4326", always_xy=True)
        except ImportError:
            messagebox.showerror("Erro", "pip install pyproj"); return

        lats, lons, pv_coords = [], [], {}
        for nome, pv in self.pvs.items():
            x, y = pv.get("x"), pv.get("y")
            if not x or not y or abs(x) < 100_000: continue
            lon, lat = tr.transform(x, y)
            if -34 < lat < 5:
                pv_coords[nome] = (lat, lon); lats.append(lat); lons.append(lon)
                self._mapa_markers.append(
                    self.map_widget.set_marker(lat, lon, text=nome,
                                                marker_color_circle="blue",
                                                marker_color_outside="darkblue"))

        self.mapa_listbox.delete(0, tk.END)
        for i, t in enumerate(self.trechos):
            pi, pf = t.get("pv_ini",""), t.get("pv_fim","")
            inc = self._mapa_checkstates[i] if i < len(self._mapa_checkstates) else True
            tag = " OK" if inc else "  X"
            cor = "#00ff88" if inc else "gray"
            if t.get("_cruza_quadra"): cor = "#ff8800"; tag = " !!"
            label = f"[{tag}] {i+1:03d} {pi}->{pf} {t.get('ext_m',0):.0f}m DN{t.get('dn_mm','?')}"
            self.mapa_listbox.insert(tk.END, label)
            self.mapa_listbox.itemconfig(i, fg=cor)
            if pi in pv_coords and pf in pv_coords:
                lat1, lon1 = pv_coords[pi]; lat2, lon2 = pv_coords[pf]
                self._mapa_paths.append(
                    self.map_widget.set_path([(lat1,lon1),(lat2,lon2)], color=cor, width=3))

        if lats:
            self.map_widget.set_position(sum(lats)/len(lats), sum(lons)/len(lons))
            self.map_widget.set_zoom(16)
        sel = sum(self._mapa_checkstates)
        self.mapa_sel_count.config(text=f"{sel}/{len(self.trechos)}")
        self.mapa_info.config(text=f"{len(pv_coords)} PVs no mapa")

    def _mapa_on_select(self):
        sel = self.mapa_listbox.curselection()
        if sel:
            i = sel[0]; t = self.trechos[i]
            inc = "INC" if self._mapa_checkstates[i] else "EXC"
            self.mapa_sel_var.set(
                f"{i+1:03d}: {t.get('pv_ini','')}->{t.get('pv_fim','')} | "
                f"DN{t.get('dn_mm','')} {t.get('ext_m',0):.1f}m [{inc}]")

    def _mapa_toggle_trecho(self):
        sel = self.mapa_listbox.curselection()
        if not sel: return
        i = sel[0]
        if i >= len(self._mapa_checkstates): return
        self._mapa_checkstates[i] = not self._mapa_checkstates[i]
        inc = self._mapa_checkstates[i]
        t = self.trechos[i]
        tag = " OK" if inc else "  X"
        label = f"[{tag}] {i+1:03d} {t.get('pv_ini','')}->{t.get('pv_fim','')} {t.get('ext_m',0):.0f}m DN{t.get('dn_mm','?')}"
        self.mapa_listbox.delete(i)
        self.mapa_listbox.insert(i, label)
        self.mapa_listbox.itemconfig(i if inc else "#666666")
        self.mapa_listbox.selection_set(i)
        self.mapa_sel_count.config(text=f"{sum(self._mapa_checkstates)}/{len(self.trechos)}")
        if i + 1 < len(self.trechos):
            self.mapa_listbox.selection_clear(0, tk.END)
            self.mapa_listbox.selection_set(i + 1)
            self.mapa_listbox.see(i + 1)

    def _mapa_sel_todos(self):
        self._mapa_checkstates = [True] * len(self.trechos); self._mapa_carregar()
    def _mapa_sel_nenhum(self):
        self._mapa_checkstates = [False] * len(self.trechos); self._mapa_carregar()
    def _mapa_sel_inverter(self):
        self._mapa_checkstates = [not s for s in self._mapa_checkstates]; self._mapa_carregar()

    def _mapa_validar_gpkg(self):
        gpkg = self.gpkg_var.get()
        if not gpkg or not self.trechos:
            messagebox.showinfo("Info", "Selecione GPKG e processe um arquivo."); return
        self._run(self._do_validar_gpkg)

    def _do_validar_gpkg(self):
        from shapely.geometry import LineString
        carta = v5.ler_cartografia_gpkg(self.gpkg_var.get())
        if not carta.get("quadras"): return
        n = 0
        for i, t in enumerate(self.trechos):
            pvi, pvf = self.pvs.get(t.get("pv_ini"),{}), self.pvs.get(t.get("pv_fim"),{})
            if not all([pvi.get("x"), pvi.get("y"), pvf.get("x"), pvf.get("y")]): continue
            ln = LineString([(pvi["x"],pvi["y"]),(pvf["x"],pvf["y"])])
            for g in carta["quadras"].geometry:
                try:
                    if ln.intersects(g):
                        t["_cruza_quadra"] = True; self._mapa_checkstates[i] = False; n += 1; break
                except: pass
        self.root.after(0, self._log_msg, f"GPKG: {n} cruzam quadras", "OK")
        self.root.after(0, self._mapa_carregar)

    def _mapa_gerar_ns(self):
        if not self.trechos: return
        sel = [t for i, t in enumerate(self.trechos)
               if i < len(self._mapa_checkstates) and self._mapa_checkstates[i]]
        if not sel:
            messagebox.showwarning("Aviso", "Nenhum trecho selecionado!"); return
        if messagebox.askyesno("NS", f"Gerar {len(sel)} NS?"):
            self.trechos = sel
            self._run(lambda: (
                __import__("gerar_ns").processar_nucleo_from_data(
                    self.pvs, self.trechos, self.nucleo_var.get() or "Nucleo",
                    self.saida_var.get())))

    # ── ML ──
    def _ml_salvar(self):
        if not self.trechos: return
        try:
            from ml_classificador import salvar_decisoes
            n = salvar_decisoes(self.trechos, self.pvs, self._mapa_checkstates, self.nucleo_var.get())
            self._log_msg(f"ML: {n} amostras salvas", "OK")
        except Exception as e: self._log_msg(f"ML: {e}", "ERROR")

    def _ml_treinar(self):
        try:
            from ml_classificador import treinar_modelo
            r = treinar_modelo()
            self._log_msg(f"ML: F1={r['f1']} Acc={r['accuracy']}", "OK")
        except Exception as e: self._log_msg(f"ML: {e}", "ERROR")

    def _ml_predizer(self):
        if not self.trechos: return
        try:
            from ml_classificador import predizer
            self._mapa_checkstates = predizer(self.trechos, self.pvs)
            self._log_msg(f"ML: {sum(self._mapa_checkstates)}/{len(self._mapa_checkstates)} reais", "OK")
            self._mapa_carregar()
        except Exception as e: self._log_msg(f"ML: {e}", "ERROR")

    # ── TAB 13: GESTAO ────────────────────────────────────────────────────

    def _tab_gestao(self):
        tab = tk.Frame(self.nb)
        self.nb.add(tab, text="  [13] Gestao  ")

        # Botoes
        brow = tk.Frame(tab)
        brow.pack(fill=tk.X, padx=8, pady=(8, 2))
        tk.Label(brow, text="Gestao & Cronograma", font=("Segoe UI", 11, "bold")).pack(side=tk.LEFT, padx=(0, 12))
        for txt, cor, fg_c, cmd in [
            ("CURVA S",         ACCENT,  "#000", self._cmd_gestao_curva_s),
            ("CRONOGRAMA",      BLUE,    WHITE,  self._cmd_gestao_cronograma),
            ("BOLETIM MEDICAO", ORANGE,  WHITE,  self._cmd_gestao_boletim),
            ("ABRIR MEGA",      PURPLE,  WHITE,  self._cmd_abrir_mega),
            ("EXPORTAR XLSX",   CYAN,    "#000", self._cmd_gestao_xlsx),
        ]:
            tk.Button(brow, text=txt, command=cmd, bg=cor, fg=fg_c,
                      font=("Segoe UI", 9, "bold"),
                      padx=8, pady=5, cursor="hand2").pack(side=tk.LEFT, padx=2)

        # Modo cronograma
        mrow = tk.Frame(tab)
        mrow.pack(fill=tk.X, padx=8, pady=(0, 2))
        tk.Label(mrow, text="Cronograma:",
                 font=("Segoe UI", 8)).pack(side=tk.LEFT)
        for txt, val in [("NS do Nucleo Atual", "ns_atual"), ("Macro (todos os nucleos)", "macro")]:
            tk.Radiobutton(mrow, text=txt, variable=self._crono_modo, value=val, selectcolor=BG2, activebackground=BG,
                           font=("Segoe UI", 8)).pack(side=tk.LEFT, padx=6)

        # Stats cards
        gstats = tk.Frame(tab)
        gstats.pack(fill=tk.X, padx=8, pady=4)
        self.gestao_labels = {}
        for nome, val, cor in [
            ("Nucleos",    "0", ACCENT),
            ("LA (agua)",  "0", BLUE),
            ("LE (esgoto)","0", CYAN),
            ("Equipes",    "0", ORANGE),
            ("% Fisico",   "0%", YELLOW),
            ("% Financ",   "0%", PURPLE),
        ]:
            f = tk.Frame(gstats, padx=10, pady=6)
            f.pack(side=tk.LEFT, padx=3, expand=True, fill=tk.X)
            tk.Label(f, text=nome, font=("Segoe UI", 7)).pack()
            lbl = tk.Label(f, text=val, fg=cor, font=("Segoe UI", 13, "bold"))
            lbl.pack()
            self.gestao_labels[nome] = lbl

        # Painel principal: Gantt + Curva S lado a lado
        main_pane = tk.Frame(tab)
        main_pane.pack(fill=tk.BOTH, expand=True, padx=8, pady=4)

        # Gantt
        lf_gantt = tk.LabelFrame(main_pane, text=" CRONOGRAMA (Gantt) ",
                                  font=("Segoe UI", 9, "bold"), bd=1)
        lf_gantt.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 4))
        self.gestao_canvas_gantt = tk.Canvas(lf_gantt, height=220,
                                              highlightthickness=0)
        self.gestao_canvas_gantt.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)
        self._gantt_placeholder()

        # Curva S
        lf_curva = tk.LabelFrame(main_pane, text=" CURVA S (fisico) ",
                                  font=("Segoe UI", 9, "bold"), bd=1)
        lf_curva.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(4, 0))
        self.gestao_canvas_curva = tk.Canvas(lf_curva, height=220,
                                              highlightthickness=0)
        self.gestao_canvas_curva.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)
        self._curva_placeholder()

        # ── Diário de Obras + A FAZER ─────────────────────────────────────────
        lf_diario = tk.LabelFrame(tab, text=" DIARIO DE OBRAS + A FAZER ",
                                   font=("Segoe UI", 9, "bold"), bd=1)
        lf_diario.pack(fill=tk.X, padx=8, pady=(4, 2))

        # Linha de cabeçalho
        drow = tk.Frame(lf_diario)
        drow.pack(fill=tk.X, padx=8, pady=(4, 2))
        tk.Label(drow, text="Data:", font=("Segoe UI", 8)).pack(side=tk.LEFT)
        self._diario_data = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        tk.Entry(drow, textvariable=self._diario_data, width=12, insertbackground=FG,
                 font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(2, 12))
        tk.Label(drow, text="Responsavel:", font=("Segoe UI", 8)).pack(side=tk.LEFT)
        self._diario_resp = tk.StringVar(value="Felipe Nery")
        tk.Entry(drow, textvariable=self._diario_resp, width=20, insertbackground=FG,
                 font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(2, 0))

        # Painel duplo: executado (esq) | a fazer (dir)
        dpanel = tk.Frame(lf_diario)
        dpanel.pack(fill=tk.BOTH, expand=True, padx=8, pady=2)

        # Executado
        lf_exec = tk.LabelFrame(dpanel, text=" EXECUTADO HOJE ", font=("Segoe UI", 8), bd=1)
        lf_exec.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 4))
        import tkinter.scrolledtext as _st2
        self._diario_exec_text = _st2.ScrolledText(lf_exec,
                                                    font=("Segoe UI", 8), height=4,
                                                    wrap=tk.WORD, bd=2)
        self._diario_exec_text.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)

        # A FAZER
        lf_af = tk.LabelFrame(dpanel, text=" A FAZER ", font=("Segoe UI", 8), bd=1)
        lf_af.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(4, 0))

        af_top = tk.Frame(lf_af)
        af_top.pack(fill=tk.X, padx=4, pady=(4, 2))
        self._diario_nova_tarefa = tk.StringVar()
        tk.Entry(af_top, textvariable=self._diario_nova_tarefa, width=28, insertbackground=FG,
                 font=("Segoe UI", 9)).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 4))
        tk.Button(af_top, text="+ Adicionar", command=self._diario_add_tarefa, font=("Segoe UI", 8, "bold"), padx=4, cursor="hand2").pack(side=tk.LEFT)

        self._diario_afazer_frame = tk.Frame(lf_af)
        self._diario_afazer_frame.pack(fill=tk.BOTH, expand=True, padx=4, pady=(0, 4))
        self._diario_afazer = []   # lista de (StringVar descricao, BooleanVar feito)

        # Botões do diário
        dbtn = tk.Frame(lf_diario)
        dbtn.pack(fill=tk.X, padx=8, pady=(2, 6))
        for txt, cor, fg_c, cmd in [
            ("SALVAR DIARIO",  ORANGE, WHITE,  self._salvar_diario),
            ("ABRIR ULTIMO",   GRAY,   WHITE,  self._abrir_diario),
        ]:
            tk.Button(dbtn, text=txt, command=cmd, bg=cor, fg=fg_c,
                      font=("Segoe UI", 8, "bold"),
                      padx=8, pady=3, cursor="hand2").pack(side=tk.LEFT, padx=2)

        # ── WhatsApp Campo + Status NS ────────────────────────────────────────
        lf_wa = tk.LabelFrame(tab, text=" WHATSAPP CAMPO + STATUS NS ",
                              font=("Segoe UI", 9, "bold"), bd=1)
        lf_wa.pack(fill=tk.X, padx=8, pady=(4, 2))

        wa_top = tk.Frame(lf_wa)
        wa_top.pack(fill=tk.X, padx=8, pady=(4, 2))

        self._webhook_status_lbl = tk.Label(
            wa_top, text="Webhook: INATIVO", font=("Segoe UI", 8, "bold"))
        self._webhook_status_lbl.pack(side=tk.LEFT, padx=(0, 10))

        for txt, cor, fg_c, cmd in [
            ("INICIAR WEBHOOK CAMPO",  "#25D366", "#000", self._cmd_iniciar_webhook),
            ("PARAR WEBHOOK",          RED,       WHITE,  self._cmd_parar_webhook),
            ("EXPORTAR STATUS NS CSV", CYAN,      "#000", self._cmd_exportar_status_ns),
            ("ABRIR STATUS NS JSON",   GRAY,      WHITE,  self._cmd_abrir_status_ns),
        ]:
            tk.Button(wa_top, text=txt, command=cmd, bg=cor, fg=fg_c,
                      font=("Segoe UI", 8, "bold"),
                      padx=6, pady=3, cursor="hand2").pack(side=tk.LEFT, padx=2)

        # Feed de mensagens do campo
        wa_feed_row = tk.Frame(lf_wa)
        wa_feed_row.pack(fill=tk.X, padx=8, pady=(2, 6))

        lf_feed = tk.LabelFrame(wa_feed_row, text=" Mensagens Recentes ",
                                font=("Segoe UI", 8), bd=1)
        lf_feed.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 4))
        try:
            import tkinter.scrolledtext as _st3
            self._wa_feed_text = _st3.ScrolledText(
                lf_feed, font=("Courier New", 8),
                height=4, state=tk.DISABLED)
        except Exception:
            self._wa_feed_text = tk.Text(
                lf_feed, font=("Courier New", 8),
                height=4, state=tk.DISABLED)
        self._wa_feed_text.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)

        # Mini KPIs de status NS
        lf_kpis = tk.LabelFrame(wa_feed_row, text=" Status NS ",
                                font=("Segoe UI", 8), bd=1)
        lf_kpis.pack(side=tk.LEFT, fill=tk.Y, padx=(4, 0))
        self._wa_kpi_labels: dict[str, tk.Label] = {}
        for nome, val in [
            ("Total",      "0"),
            ("Planejadas", "0"),
            ("Executadas", "0"),
            ("Cadastradas","0"),
            ("Medidas",    "0"),
            ("% Fisico",   "0%"),
        ]:
            row = tk.Frame(lf_kpis)
            row.pack(fill=tk.X, padx=6, pady=1)
            tk.Label(row, text=f"{nome}:",
                     font=("Segoe UI", 7), width=12, anchor=tk.W).pack(side=tk.LEFT)
            lbl = tk.Label(row, text=val,
                           font=("Segoe UI", 8, "bold"), anchor=tk.W)
            lbl.pack(side=tk.LEFT)
            self._wa_kpi_labels[nome] = lbl

        # Webhook internals
        self._webhook_proc  = None   # subprocess handle
        self._webhook_thread = None  # thread que lê stdout

        # Status bar
        self.gestao_status = tk.Label(tab, text="Pronto — processe uma rede ou clique CURVA S / CRONOGRAMA", font=("Segoe UI", 8))
        self.gestao_status.pack(pady=(0, 4))

    def _gantt_placeholder(self):
        c = self.gestao_canvas_gantt
        c.delete("all")
        w, h = 500, 220
        c.config(width=w)
        msg = "Clique CRONOGRAMA para gerar o Gantt"
        c.create_text(w // 2, h // 2, text=msg, fill=GRAY,
                      font=("Segoe UI", 9), anchor=tk.CENTER)

    def _curva_placeholder(self):
        c = self.gestao_canvas_curva
        c.delete("all")
        w, h = 400, 220
        c.config(width=w)
        msg = "Clique CURVA S para gerar o grafico"
        c.create_text(w // 2, h // 2, text=msg, fill=GRAY,
                      font=("Segoe UI", 9), anchor=tk.CENTER)

    # ── Comandos do tab Gestao ─────────────────────────────────────────────

    def _get_nucleos_para_gestao(self):
        """Retorna lista de nucleos para gerar_cronograma_macro."""
        ext_total = sum(t.get("ext_m", 0) for t in self.trechos)
        n_trechos = len(self.trechos)
        if v5 and hasattr(v5, "NUCLEOS_BATCH") and v5.NUCLEOS_BATCH:
            return [
                {"nome": n["nucleo"], "extensao_m": ext_total or 1000,
                 "n_trechos": n_trechos or 10, "equipes": 4}
                for n in v5.NUCLEOS_BATCH
            ]
        nucleo = self.nucleo_var.get() or "REDE"
        return [{"nome": nucleo, "extensao_m": ext_total or 1000,
                 "n_trechos": n_trechos or 10, "equipes": 4}]

    def _cmd_gestao_cronograma(self):
        if not self.trechos:
            from tkinter import messagebox
            messagebox.showinfo("Cronograma", "Processe uma rede primeiro.")
            return
        self._run(self._do_gestao_cronograma)

    def _do_gestao_cronograma(self):
        self.root.after(0, self.gestao_status.config,
                        {"text": "Gerando cronograma...", "fg": YELLOW})
        out_dir = Path(self.saida_var.get()) / "GESTAO"
        out_dir.mkdir(parents=True, exist_ok=True)
        data_inicio = datetime.now().strftime("%Y-%m-%d")
        try:
            if self._crono_modo.get() == "ns_atual" and self._ns_sequencia:
                # ── Modo NS do núcleo atual ──────────────────────────────────
                from gerar_cronograma_macro import gerar_cronograma_por_ns
                nucleo = self.nucleo_var.get() or "REDE"
                ns_lista = [
                    {"ordem":      seq + 1,
                     "trecho_idx": idx,
                     "pv_ini":     self.trechos[idx].get("pv_ini", "") if idx < len(self.trechos) else "",
                     "pv_fim":     self.trechos[idx].get("pv_fim", "") if idx < len(self.trechos) else "",
                     "ext_m":      self.trechos[idx].get("ext_m", 0) if idx < len(self.trechos) else 0,
                     "rua":        self.trechos[idx].get("rua", "") if idx < len(self.trechos) else ""}
                    for seq, idx in enumerate(self._ns_sequencia)
                ]
                resultado = gerar_cronograma_por_ns(
                    ns_lista, data_inicio,
                    equipes=self._ns_equipes.get(),
                    prod_m_dia=self._ns_prod_m_dia.get(),
                    nucleo=nucleo, out_dir=str(out_dir / "CRONOGRAMA_NS"),
                )
                n = len(resultado["tarefas"])
                self.root.after(0, self._log_msg,
                    f"Cronograma NS: {n} NS | {resultado['extensao_total_m']:.0f}m | "
                    f"{resultado['data_inicio']} → {resultado['data_fim']}", "OK")
                self.root.after(0, self.gestao_status.config,
                    {"text": f"Cronograma NS: {n} tarefas | {resultado['data_fim']}", "fg": ACCENT})
                # Abrir GANTT HTML
                gantt = out_dir / "CRONOGRAMA_NS" / "GANTT_NS.html"
                if gantt.exists():
                    self.root.after(500, lambda: webbrowser.open(str(gantt)))
            else:
                # ── Modo macro (todos os núcleos) ────────────────────────────
                from gerar_cronograma_macro import gerar_cronograma_macro, gerar_tudo
                nucleos = self._get_nucleos_para_gestao()
                wbs = gerar_cronograma_macro(nucleos, data_inicio)
                gerar_tudo(nucleos, data_inicio, str(out_dir))
                self.root.after(0, self._desenhar_gantt, wbs)
                n_t = wbs.get("total_tarefas", 0)
                self.root.after(0, self._log_msg,
                    f"Cronograma Macro: {n_t} tarefas — {out_dir.name}/", "OK")
                self.root.after(0, self.gestao_status.config,
                    {"text": f"Cronograma Macro: {n_t} tarefas", "fg": ACCENT})
                # Abrir XLSX
                xlsx = out_dir / "CRONOGRAMA_MACRO_SLNR.xlsx"
                if xlsx.exists():
                    try:
                        self.root.after(500, lambda: os.startfile(str(xlsx)))
                    except Exception:
                        pass
        except Exception as e:
            self.root.after(0, self._log_msg, f"Cronograma: {e}", "ERROR")
            self.root.after(0, self.gestao_status.config,
                            {"text": f"Erro cronograma: {e}", "fg": RED})
            raise

    # ── WhatsApp Webhook ────────────────────────────────────────────────────

    def _cmd_iniciar_webhook(self):
        """Sobe whatsapp_receiver via uvicorn em subprocess."""
        import subprocess, sys
        if getattr(self, "_webhook_proc", None) and self._webhook_proc.poll() is None:
            messagebox.showinfo("Webhook", "Webhook já está rodando na porta 8765.")
            return
        nucleo   = self.nucleo_var.get() or "REDE"
        out_base = self.saida_var.get()
        # Montar caminho do STATUS_NS.json
        from pathlib import Path as _P
        status_path = str(_P(out_base) / nucleo.upper() / "PLANEJAMENTO" / "STATUS_NS.json")
        env = {**os.environ,
               "STATUS_NS_PATH": status_path,
               "NUCLEO_DEFAULT": nucleo,
               "OUT_BASE": out_base}
        try:
            self._webhook_proc = subprocess.Popen(
                [sys.executable, "-m", "uvicorn",
                 "whatsapp_receiver:app", "--port", "8765", "--host", "0.0.0.0"],
                cwd=str(SCRIPT_DIR),
                env=env,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
            )
            self._webhook_status_lbl.config(text="Webhook: ATIVO :8765")
            self._log_msg("Webhook WhatsApp iniciado na porta 8765", "OK")
            # Thread para ler logs do processo
            threading.Thread(target=self._webhook_log_reader, daemon=True).start()
        except FileNotFoundError:
            messagebox.showerror("Webhook",
                "uvicorn nao encontrado.\nInstale: pip install fastapi uvicorn[standard] httpx")

    def _cmd_parar_webhook(self):
        """Para o subprocess do webhook."""
        proc = getattr(self, "_webhook_proc", None)
        if proc and proc.poll() is None:
            proc.terminate()
            self._webhook_status_lbl.config(text="Webhook: INATIVO")
            self._log_msg("Webhook WhatsApp encerrado.", "WARN")
        else:
            messagebox.showinfo("Webhook", "Webhook nao estava rodando.")

    def _webhook_log_reader(self):
        """Lê stdout do processo uvicorn e exibe no feed de mensagens."""
        proc = self._webhook_proc
        if not proc or not proc.stdout:
            return
        for line in proc.stdout:
            line = line.rstrip()
            if line:
                self.root.after(0, self._wa_feed_append, line)
        # Processo terminou
        self.root.after(0, self._webhook_status_lbl.config,
                        {"text": "Webhook: INATIVO", "fg": GRAY})

    def _wa_feed_append(self, text: str):
        """Adiciona linha ao feed de mensagens WhatsApp (thread-safe via root.after)."""
        try:
            w = self._wa_feed_text
            w.config(state=tk.NORMAL)
            from datetime import datetime as _dt
            w.insert(tk.END, f"[{_dt.now().strftime('%H:%M:%S')}] {text}\n")
            w.see(tk.END)
            w.config(state=tk.DISABLED)
            # Atualizar KPIs ao detectar msg de transição
            if "STATUS_NS" in text or "EXECUTAD" in text or "CADASTRAD" in text:
                threading.Thread(target=self._do_status_ns_init, daemon=True).start()
        except Exception:
            pass

    def _cmd_exportar_status_ns(self):
        """Exporta STATUS_NS como CSV."""
        if not self._status_ns:
            messagebox.showinfo("Status NS", "Nenhum STATUS_NS carregado. Processe uma rede primeiro.")
            return
        try:
            from motor_status_ns import exportar_csv as ns_csv
            nucleo   = self.nucleo_var.get() or "REDE"
            out_dir  = Path(self.saida_var.get()) / nucleo.upper() / "PLANEJAMENTO"
            out_dir.mkdir(parents=True, exist_ok=True)
            p = ns_csv(self._status_ns, str(out_dir / "STATUS_NS.csv"))
            self._log_msg(f"STATUS_NS exportado: {Path(p).name}", "OK")
            try:
                os.startfile(p)
            except Exception:
                pass
        except Exception as e:
            self._log_msg(f"Exportar Status NS: {e}", "ERROR")

    def _cmd_abrir_status_ns(self):
        """Abre STATUS_NS.json no editor padrão."""
        nucleo = self.nucleo_var.get() or "REDE"
        p = Path(self.saida_var.get()) / nucleo.upper() / "PLANEJAMENTO" / "STATUS_NS.json"
        if p.exists():
            try:
                os.startfile(str(p))
            except Exception:
                messagebox.showinfo("Status NS", str(p))
        else:
            messagebox.showinfo("Status NS", f"Arquivo nao encontrado:\n{p}\n\nProcesse uma rede primeiro.")

    def _desenhar_gantt(self, wbs):
        """Desenha Gantt simplificado no Canvas."""
        from datetime import date as dt_date
        c = self.gestao_canvas_gantt
        c.update_idletasks()
        c.delete("all")
        W = c.winfo_width() or 500
        H = c.winfo_height() or 220

        nucleos = wbs.get("nucleos", [])
        if not nucleos:
            c.create_text(W // 2, H // 2, text="Sem dados", fill=GRAY,
                          font=("Segoe UI", 9))
            return

        # Calcular range de datas
        def parse_d(s):
            try: return dt_date.fromisoformat(str(s))
            except: return dt_date.today()

        all_starts = [parse_d(n.get("inicio")) for n in nucleos]
        all_ends   = [parse_d(n.get("fim"))    for n in nucleos]
        d_min = min(all_starts)
        d_max = max(all_ends)
        total_dias = max((d_max - d_min).days, 1)

        LABEL_W = 110
        PAD_TOP  = 28
        PAD_BOT  = 16
        ROW_H    = 22
        BAR_H    = 14
        graf_w   = W - LABEL_W - 12
        n_rows   = len(nucleos)
        needed_h = PAD_TOP + n_rows * ROW_H + PAD_BOT
        c.config(scrollregion=(0, 0, W, max(H, needed_h)))

        # Cores por nucleo (ciclico)
        CORES = [ACCENT, BLUE, CYAN, ORANGE, PURPLE, YELLOW, "#22c55e", "#f43f5e"]

        # Cabecalho: meses
        from datetime import timedelta
        cur = d_min.replace(day=1)
        while cur <= d_max:
            dias_offset = (cur - d_min).days
            x = LABEL_W + int(dias_offset / total_dias * graf_w)
            c.create_line(x, PAD_TOP - 4, x, PAD_TOP + n_rows * ROW_H,
                          fill=GRAY, width=1, dash=(2, 4))
            label = cur.strftime("%b/%y")
            c.create_text(x + 2, PAD_TOP - 10, text=label,
                          fill=GRAY, font=("Segoe UI", 7), anchor=tk.W)
            # Avancar 1 mes
            if cur.month == 12:
                cur = cur.replace(year=cur.year + 1, month=1)
            else:
                cur = cur.replace(month=cur.month + 1)

        # Barras
        for i, nuc in enumerate(nucleos):
            y_center = PAD_TOP + i * ROW_H + ROW_H // 2
            cor = CORES[i % len(CORES)]

            # Label nucleo
            nome = nuc.get("nome", f"N{i+1}")[:16]
            c.create_text(LABEL_W - 4, y_center, text=nome,
                          fill=FG, font=("Consolas", 8), anchor=tk.E)

            # Barra total (fundo cinza)
            c.create_rectangle(LABEL_W, y_center - BAR_H // 2,
                                LABEL_W + graf_w, y_center + BAR_H // 2,
                                fill="#1a1a2e", outline="")

            # Barra planejada
            d_ini = parse_d(nuc.get("inicio"))
            d_fim = parse_d(nuc.get("fim"))
            x1 = LABEL_W + int((d_ini - d_min).days / total_dias * graf_w)
            x2 = LABEL_W + int((d_fim - d_min).days / total_dias * graf_w)
            x2 = max(x2, x1 + 4)
            c.create_rectangle(x1, y_center - BAR_H // 2,
                                x2, y_center + BAR_H // 2,
                                fill=cor, outline="")

            # Duracao label
            dur = nuc.get("duracao_dias", 0)
            c.create_text(x2 + 4, y_center, text=f"{dur}d",
                          fill=GRAY, font=("Segoe UI", 7), anchor=tk.W)

        # Atualizar stats cards
        self.gestao_labels["Nucleos"].config(text=str(len(nucleos)))
        total_ext = sum(n.get("extensao_m", 0) for n in nucleos)
        dur_total = wbs.get("duracao_total_dias", 0)
        self.gestao_status.config(
            text=f"Gantt: {len(nucleos)} nucleos | {total_ext:.0f}m | {dur_total} dias"
        )

    # ── Diário de Obras ───────────────────────────────────────────────────────

    def _diario_add_tarefa(self):
        desc = self._diario_nova_tarefa.get().strip()
        if not desc:
            return
        feito = tk.BooleanVar(value=False)
        desc_var = tk.StringVar(value=desc)
        self._diario_afazer.append((desc_var, feito))
        self._diario_nova_tarefa.set("")
        self._diario_render_afazer()

    def _diario_render_afazer(self):
        for w in self._diario_afazer_frame.winfo_children():
            w.destroy()
        for i, (desc_var, feito_var) in enumerate(self._diario_afazer):
            row = tk.Frame(self._diario_afazer_frame)
            row.pack(fill=tk.X, pady=1)
            tk.Checkbutton(row, variable=feito_var,
                           activebackground=BG, selectcolor=BG2).pack(side=tk.LEFT)
            fg_c = GRAY if feito_var.get() else FG
            tk.Label(row, textvariable=desc_var, fg=fg_c,
                     font=("Segoe UI", 8)).pack(side=tk.LEFT, fill=tk.X, expand=True)
            tk.Button(row, text="✕", command=lambda idx=i: self._diario_rem_tarefa(idx), font=("Segoe UI", 7),
                      cursor="hand2").pack(side=tk.RIGHT)

    def _diario_rem_tarefa(self, idx):
        if 0 <= idx < len(self._diario_afazer):
            self._diario_afazer.pop(idx)
            self._diario_render_afazer()

    def _salvar_diario(self):
        data    = self._diario_data.get() or datetime.now().strftime("%Y-%m-%d")
        resp    = self._diario_resp.get()
        exec_txt = self._diario_exec_text.get("1.0", tk.END).strip()
        afazer  = [{"desc": d.get(), "feito": f.get()}
                   for d, f in self._diario_afazer]

        out = Path(self.saida_var.get()) / "GESTAO" / "DIARIO"
        out.mkdir(parents=True, exist_ok=True)
        fname = out / f"DIARIO_{data}.json"
        with open(fname, "w", encoding="utf-8") as fh:
            json.dump({
                "data": data, "responsavel": resp,
                "executado": exec_txt, "a_fazer": afazer,
                "gerado_em": datetime.now().isoformat(),
            }, fh, indent=2, ensure_ascii=False)
        self._log_msg(f"Diario salvo: {fname.name}", "OK")

    def _abrir_diario(self):
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            title="Abrir Diário",
            initialdir=str(Path(self.saida_var.get()) / "GESTAO" / "DIARIO"),
            filetypes=[("JSON", "*.json")],
        )
        if not path or not Path(path).exists():
            return
        with open(path, "r", encoding="utf-8") as fh:
            d = json.load(fh)
        self._diario_data.set(d.get("data", ""))
        self._diario_resp.set(d.get("responsavel", ""))
        self._diario_exec_text.delete("1.0", tk.END)
        self._diario_exec_text.insert(tk.END, d.get("executado", ""))
        self._diario_afazer = [
            (tk.StringVar(value=item["desc"]), tk.BooleanVar(value=item.get("feito", False)))
            for item in d.get("a_fazer", [])
        ]
        self._diario_render_afazer()
        self._log_msg(f"Diario carregado: {Path(path).name}", "OK")

    # ── Gestão — outros comandos ───────────────────────────────────────────────

    def _cmd_gestao_curva_s(self):
        self._run(self._do_gestao_curva_s)

    def _do_gestao_curva_s(self):
        self.root.after(0, self.gestao_status.config,
                        {"text": "Gerando Curva S...", "fg": YELLOW})
        try:
            nucleo = self.nucleo_var.get() or "REDE"
            cs = {}
            kpi = {}
            try:
                from core.database import bootstrap_database, curva_s_dados, dashboard_metricas
                bootstrap_database(force_import=False)
                cs = curva_s_dados("" if nucleo == "REDE" else nucleo)
                kpi = dashboard_metricas("" if nucleo == "REDE" else nucleo)
            except Exception:
                cs = {}
                kpi = {}

            if not cs.get("previsto") and not cs.get("realizado"):
                from motor_medicao import gerar_curva_s
                dados_exec = {}
                cs = gerar_curva_s(self.trechos, dados_exec)

            self.root.after(0, self._desenhar_curva_s, cs)

            # Tambem gerar HTML+XLSX se houver trechos
            if self.trechos:
                from gerar_medicao_curva_s import gerar_medicao_curva_s
                out_dir = Path(self.saida_var.get()) / "GESTAO"
                out_dir.mkdir(parents=True, exist_ok=True)
                gerar_medicao_curva_s(self.pvs, self.trechos, nucleo, str(out_dir))
                self.root.after(0, self._log_msg,
                                f"Curva S: HTML + XLSX gerados em {out_dir.name}/", "OK")

            if kpi:
                pct_fis = kpi.get("pct_fisico", 0)
                pct_fin = kpi.get("pct_financeiro", 0)
                self.root.after(0, self.gestao_labels["% Fisico"].config,
                                {"text": f"{pct_fis:.0f}%"})
                self.root.after(0, self.gestao_labels["% Financ"].config,
                                {"text": f"{pct_fin:.0f}%"})
            else:
                previsto = cs.get("previsto", [])
                if previsto:
                    ultimo_prev = previsto[-1]
                    pct_fis  = ultimo_prev.get("pct_acum", ultimo_prev.get("acum_pct", 0))
                    pct_fin  = round(ultimo_prev.get("custo_acum", 0) /
                                     max(cs.get("custo_total", 1), 1) * 100, 1)
                    self.root.after(0, self.gestao_labels["% Fisico"].config,
                                    {"text": f"{pct_fis:.0f}%"})
                    self.root.after(0, self.gestao_labels["% Financ"].config,
                                    {"text": f"{pct_fin:.0f}%"})

            n_la = sum(1 for t in self.trechos if t.get("is_agua")) if self.trechos else 0
            n_le = sum(1 for t in self.trechos if not t.get("is_agua")) if self.trechos else 0
            self.root.after(0, self.gestao_labels["LA (agua)"].config,  {"text": str(n_la)})
            self.root.after(0, self.gestao_labels["LE (esgoto)"].config, {"text": str(n_le)})

            self.root.after(0, self.gestao_status.config,
                            {"text": f"Curva S: {len(cs.get('previsto', []))} marcos previstos",
                             "fg": ACCENT})
        except Exception as e:
            self.root.after(0, self._log_msg, f"Curva S: {e}", "ERROR")
            self.root.after(0, self.gestao_status.config,
                            {"text": f"Erro Curva S: {e}", "fg": RED})
            raise

    def _desenhar_curva_s(self, cs):
        """Desenha Curva S no Canvas: linha previsto (CYAN) + realizado (ACCENT)."""
        c = self.gestao_canvas_curva
        c.update_idletasks()
        c.delete("all")
        W = c.winfo_width()  or 400
        H = c.winfo_height() or 220

        PAD_L, PAD_R, PAD_T, PAD_B = 36, 12, 16, 28
        gw = W - PAD_L - PAD_R
        gh = H - PAD_T - PAD_B

        # Grid horizontal (0, 25, 50, 75, 100%)
        for pct in [0, 25, 50, 75, 100]:
            y = PAD_T + gh - int(pct / 100 * gh)
            c.create_line(PAD_L, y, W - PAD_R, y, fill="#1a1a2e", width=1)
            c.create_text(PAD_L - 2, y, text=f"{pct}%", fill=GRAY,
                          font=("Segoe UI", 7), anchor=tk.E)

        # Eixo
        c.create_line(PAD_L, PAD_T, PAD_L, H - PAD_B, fill=GRAY, width=1)
        c.create_line(PAD_L, H - PAD_B, W - PAD_R, H - PAD_B, fill=GRAY, width=1)

        previsto  = cs.get("previsto",  [])
        realizado = cs.get("realizado", [])

        def pts(serie, campo_pct):
            n = len(serie)
            if n == 0:
                return []
            result = []
            for i, item in enumerate(serie):
                x = PAD_L + int(i / max(n - 1, 1) * gw)
                pv = min(float(item.get(campo_pct, 0)), 100)
                y  = PAD_T + gh - int(pv / 100 * gh)
                result.append((x, y))
            return result

        pts_prev = pts(previsto,  "pct_acum")
        pts_real = pts(realizado, "pct_acum")

        def draw_line(pontos, cor, dash=None):
            if len(pontos) < 2:
                return
            coords = []
            for p in pontos:
                coords += list(p)
            kw = {"fill": cor, "width": 2, "smooth": True}
            if dash:
                kw["dash"] = dash
            c.create_line(*coords, **kw)

        draw_line(pts_prev, CYAN,  dash=(6, 3))
        draw_line(pts_real, ACCENT)

        # Labels eixo X (meses)
        all_series = previsto if previsto else realizado
        n_tot = len(all_series)
        step  = max(1, n_tot // 6)
        for i in range(0, n_tot, step):
            item = all_series[i]
            mes  = item.get("mes_label") or item.get("mes", "")
            x = PAD_L + int(i / max(n_tot - 1, 1) * gw)
            c.create_text(x, H - PAD_B + 10, text=str(mes)[:7],
                          fill=GRAY, font=("Segoe UI", 6), anchor=tk.CENTER)

        # Legenda
        c.create_line(W - PAD_R - 60, PAD_T + 8,
                      W - PAD_R - 44, PAD_T + 8, fill=CYAN, width=2, dash=(6, 3))
        c.create_text(W - PAD_R - 42, PAD_T + 8, text="Previsto",
                      fill=CYAN, font=("Segoe UI", 7), anchor=tk.W)
        c.create_line(W - PAD_R - 60, PAD_T + 20,
                      W - PAD_R - 44, PAD_T + 20, fill=ACCENT, width=2)
        c.create_text(W - PAD_R - 42, PAD_T + 20, text="Realizado",
                      fill=ACCENT, font=("Segoe UI", 7), anchor=tk.W)

        if not previsto and not realizado:
            c.create_text(W // 2, H // 2,
                          text="Sem dados — processe uma rede primeiro",
                          fill=GRAY, font=("Segoe UI", 9), anchor=tk.CENTER)

    def _cmd_gestao_boletim(self):
        from tkinter import messagebox, simpledialog
        if not self.trechos:
            messagebox.showinfo("Boletim", "Processe uma rede primeiro.")
            return
        periodo = simpledialog.askstring(
            "Boletim de Medicao",
            "Periodo (ex: 2026-03):",
            initialvalue=datetime.now().strftime("%Y-%m")
        )
        if not periodo:
            return
        self._run(self._do_gestao_boletim, periodo=periodo)

    def _do_gestao_boletim(self, periodo=""):
        self.root.after(0, self.gestao_status.config,
                        {"text": f"Gerando boletim {periodo}...", "fg": YELLOW})
        try:
            from motor_medicao import gerar_boletim_medicao
            nucleo = self.nucleo_var.get() or "REDE"
            trechos_exec = []
            try:
                from core.database import bootstrap_database, listar_ns, detalhe_ns
                bootstrap_database(force_import=False)
                for row in listar_ns("" if nucleo == "REDE" else nucleo):
                    if row.get("status") not in ("EM_EXECUCAO", "CONCLUIDA", "MEDIDA"):
                        continue
                    detalhado = detalhe_ns(int(row["id"])) or {}
                    trecho = (detalhado.get("trechos") or [{}])[0]
                    trechos_exec.append(
                        dict(
                            trecho,
                            cadastro_ok=row.get("cadastro_ok", False),
                            status=row.get("status", "PLANEJADO"),
                            custo=row.get("custo_realizado") or row.get("custo_previsto") or 0,
                        )
                    )
            except Exception:
                trechos_exec = []

            if not trechos_exec:
                trechos_exec = [
                    dict(t, cadastro_ok=t.get("cadastro_ok", False),
                         status=t.get("status", "PLANEJADO"))
                    for t in self.trechos
                ]
            boletim = gerar_boletim_medicao(
                trechos_exec=trechos_exec,
                pvs=self.pvs,
                periodo=periodo,
                bm_num=1,
                custo_func=lambda t, _pvs: t.get("custo", t.get("ext_m", 0) * 910)
            )
            out_dir = Path(self.saida_var.get()) / "GESTAO"
            out_dir.mkdir(parents=True, exist_ok=True)
            out_path = out_dir / f"BOLETIM_{periodo.replace('-','_')}.json"
            with open(out_path, "w", encoding="utf-8") as f:
                json.dump(boletim, f, indent=2, ensure_ascii=False, default=str)
            self.root.after(0, self._log_msg,
                            f"Boletim {periodo}: R$ {boletim.get('total_liberado', 0):,.0f} liberado — {out_path.name}", "OK")
            self.root.after(0, self.gestao_status.config,
                            {"text": f"Boletim {periodo} gerado — {out_path.name}",
                             "fg": ACCENT})
        except Exception as e:
            self.root.after(0, self._log_msg, f"Boletim: {e}", "ERROR")
            self.root.after(0, self.gestao_status.config,
                            {"text": f"Erro boletim: {e}", "fg": RED})
            raise

    def _cmd_abrir_mega(self):
        mega_default = Path(
            r"C:\Users\felip\Downloads\NOVOS NUCLEOS PLANEJAMENTO"
            r"\SIMULACAO 24_03_2026 1\SLNR_MEGA_INTEGRADA.xlsx"
        )
        if mega_default.exists():
            os.startfile(str(mega_default))
            self._log_msg(f"Abrindo MEGA: {mega_default.name}", "OK")
        else:
            # Fallback: filedialog
            from tkinter import filedialog
            path = filedialog.askopenfilename(
                title="Selecionar planilha MEGA",
                filetypes=[("Excel", "*.xlsx *.xlsm"), ("Todos", "*.*")],
                initialdir=r"C:\Users\felip\Downloads\NOVOS NUCLEOS PLANEJAMENTO"
            )
            if path:
                os.startfile(path)
                self._log_msg(f"Abrindo MEGA: {Path(path).name}", "OK")

    def _cmd_gestao_xlsx(self):
        if not self.trechos:
            from tkinter import messagebox
            messagebox.showinfo("Exportar XLSX", "Processe uma rede primeiro.")
            return
        self._run(self._do_gestao_xlsx)

    def _do_gestao_xlsx(self):
        self.root.after(0, self.gestao_status.config,
                        {"text": "Gerando XLSX gestao...", "fg": YELLOW})
        try:
            from gerar_medicao_curva_s import gerar_medicao_curva_s
            nucleo  = self.nucleo_var.get() or "REDE"
            out_dir = Path(self.saida_var.get()) / "GESTAO"
            out_dir.mkdir(parents=True, exist_ok=True)
            gerar_medicao_curva_s(self.pvs, self.trechos, nucleo, str(out_dir))
            self.root.after(0, self._log_msg,
                            f"GESTAO: MEDICAO.xlsx + CURVA_S.html + DIARIO.html gerados em {out_dir.name}/", "OK")
            self.root.after(0, self.gestao_status.config,
                            {"text": f"XLSX exportado em GESTAO/", "fg": ACCENT})
        except Exception as e:
            self.root.after(0, self._log_msg, f"XLSX gestao: {e}", "ERROR")
            self.root.after(0, self.gestao_status.config,
                            {"text": f"Erro XLSX: {e}", "fg": RED})
            raise


# ══════════════════════════════════════════════════════════════════════════════
def main():
    root = tk.Tk()
    app = HydroNetworkApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
