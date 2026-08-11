#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gera TRECHOS_MEGA_CLUSTERS.xlsx
Trechos PV-a-PV para os 83 nucleos SLNR MEGA INTEGRADA
Organizados por clusters de proximidade geografica (Union-Find, threshold=500m)
ConstruData HydroNetwork v7.0.0 | FCN Construcoes e Saneamento | Contrato 11481051
"""
import sys, io, os, math, random, csv
from collections import defaultdict, Counter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------- #
#  Caminhos
# --------------------------------------------------------------------------- #
BASE = os.path.join(
    os.environ.get("USERPROFILE", "C:\\Users\\felip"),
    "Downloads", "NOVOS NUCLEOS PLANEJAMENTO", "CONSTRUDATA_v5")

CSV_NUCLEOS = os.path.join(BASE, "nucleos_enriquecido.csv")
CSV_PROX    = os.path.join(BASE, "SLNR_MEGA_INTEGRADA_NUCLEOS_PROX_SETOR.csv")

OUT_DIR  = os.path.join(os.path.dirname(os.path.abspath(__file__)), "TRECHOS_NS_COMPLETOS")
os.makedirs(OUT_DIR, exist_ok=True)
OUT_FILE = os.path.join(OUT_DIR, "TRECHOS_MEGA_CLUSTERS.xlsx")

DIST_THRESHOLD = 500  # metros — limiar de vizinhanca para agrupar em cluster

# --------------------------------------------------------------------------- #
#  Profundidade base por DIFIC (1-5)
# --------------------------------------------------------------------------- #
PROF_DIFIC = {1: 1.0, 2: 1.3, 3: 1.7, 4: 2.1, 5: 2.5}

# --------------------------------------------------------------------------- #
#  Estilos
# --------------------------------------------------------------------------- #
fill_h1    = PatternFill("solid", fgColor="1F4E79")
fill_h2    = PatternFill("solid", fgColor="2E75B6")
fill_clust = PatternFill("solid", fgColor="D6E4F0")
fill_nucl  = PatternFill("solid", fgColor="E2EFDA")
fill_sub   = PatternFill("solid", fgColor="FFF2CC")
fill_grand = PatternFill("solid", fgColor="FCE4D6")
fill_zebra = PatternFill("solid", fgColor="EBF3FB")
font_h1    = Font(bold=True, color="FFFFFF", size=11)
font_h2    = Font(bold=True, color="FFFFFF", size=9)
font_clust = Font(bold=True, size=10, color="1F4E79")
font_nucl  = Font(bold=True, size=9, color="375623")
font_sub   = Font(bold=True, size=9)
font_grand = Font(bold=True, color="FFFFFF", size=10)
font_d     = Font(size=9)
ac  = Alignment(horizontal="center", vertical="center", wrap_text=True)
al  = Alignment(horizontal="left",   vertical="center")
ar  = Alignment(horizontal="right",  vertical="center")
thin = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"))

# --------------------------------------------------------------------------- #
#  Colunas trechos (34)
# --------------------------------------------------------------------------- #
COLS = [
    "ID_Trecho", "Cluster", "Nucleo", "TAG", "Setor",
    "PV Ini", "PV Fim", "L(m)", "DN", "Mat",
    "CT Ini", "CF Ini", "Prof Ini(m)", "CT Fim", "CF Fim", "Prof Fim(m)",
    "Decl(m/m)", "V(m/s)", "Q(l/s)", "Pavimento",
    "Escav(m3)", "Reaterro(m3)", "Bota-fora(m3)", "Escoram(m2)", "Pavim(m2)",
    "Custo Tubo", "Custo Escav", "Custo PV", "Total s/BDI", "Total c/BDI",
    "DIFIC", "Urgencia", "Prox_Vizinho", "Dist_Viz(m)",
]
LARGURAS = [
    10, 18, 26, 8, 14,
    10, 10, 7, 6, 5,
    9, 9, 10, 9, 9, 10,
    9, 7, 7, 13,
    9, 10, 11, 9, 8,
    12, 12, 12, 14, 14,
    6, 10, 22, 10,
]

# Colunas RESUMO (11)
COLS_RES  = ["Cluster", "Ancora (maior nucleo)", "N Nucleos", "N Trechos",
             "Ligacoes", "Ext Total(m)", "Custo Total c/BDI",
             "N_PVS", "VOL_ESCAV(m3)", "Urgencia dom.", "Nucleos (TAGs)"]
LARG_RES  = [22, 32, 10, 10, 10, 14, 20, 8, 14, 12, 70]


# --------------------------------------------------------------------------- #
#  Helpers
# --------------------------------------------------------------------------- #
def _f(v):
    try:
        return float(str(v).replace(",", ".").strip()) if str(v).strip() else 0.0
    except Exception:
        return 0.0

def _i(v):
    try:
        return int(float(str(v).replace(",", ".").strip())) if str(v).strip() else 0
    except Exception:
        return 0

def safe_aba(name, max_len=31):
    """Nome seguro para aba Excel."""
    for ch in r'\/?*[]:|':
        name = name.replace(ch, "_")
    return name[:max_len]


# --------------------------------------------------------------------------- #
#  Leitura CSV
# --------------------------------------------------------------------------- #
def ler_nucleos(path):
    with open(path, encoding="utf-8-sig", newline="") as f:
        return {r["TAG"]: r for r in csv.DictReader(f)}

def ler_prox(path):
    with open(path, encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


# --------------------------------------------------------------------------- #
#  Union-Find
# --------------------------------------------------------------------------- #
def connected_components(graph, all_nodes):
    parent = {n: n for n in all_nodes}

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[rb] = ra

    for node in all_nodes:
        for nbr in graph.get(node, []):
            union(node, nbr)

    comps = defaultdict(list)
    for n in all_nodes:
        comps[find(n)].append(n)
    return list(comps.values())


# --------------------------------------------------------------------------- #
#  Inferencia trecho-a-trecho
# --------------------------------------------------------------------------- #
# Parametros por tipo de terreno: (pavimento, dn_base, decl_base)
TERR = {
    "dique":         ("CBUQ",           200, 0.010),
    "morro":         ("PARALELEPIPEDO", 150, 0.015),
    "monte":         ("PARALELEPIPEDO", 150, 0.015),
    "encosta":       ("PARALELEPIPEDO", 150, 0.012),
    "urbano":        ("CBUQ",           150, 0.010),
    "prolongamento": ("CIMENTADO",      150, 0.008),
    "planicie":      ("CBUQ",           150, 0.008),
}

def inferir_trechos(nr, prox1_nome, prox1_dist):
    """Gera lista de trechos inferidos para um nucleo."""
    seed = _i(nr.get("TAG", 0)) * 7 + 13
    rng  = random.Random(seed)

    tag       = str(nr["TAG"]).strip()
    nome      = str(nr["NOME"]).strip()
    setor     = str(nr.get("SETOR", "")).strip()
    dific     = max(1, min(5, _i(nr.get("DIFIC", 3)) or 3))
    n_pvs     = _i(nr.get("N_PVS", 0))
    ext_m     = _f(nr.get("EXT_ES_M", 0))
    vol_escav = _f(nr.get("VOL_ESCAV", 0))
    c_red_es  = _f(nr.get("C_RED_ES", 0))
    c_pvs_tot = _f(nr.get("C_PVS", 0))
    lig_total = _i(nr.get("LIG_TOTAL", 0))
    urgencia  = str(nr.get("URGENCIA", "")).strip()
    tipo_ter  = str(nr.get("TIPO_TER", "urbano")).strip().lower()

    # N trechos
    if n_pvs >= 2:
        n_trechos = n_pvs - 1
    elif ext_m > 0:
        n_trechos = max(1, round(ext_m / 60))
    else:
        n_trechos = max(1, round(lig_total / 30))

    # Parametros de terreno
    pav_params = None
    for k, v in TERR.items():
        if k in tipo_ter:
            pav_params = v
            break
    if pav_params is None:
        pav_params = TERR["urbano"]
    pav_str, dn_base, decl_base = pav_params

    # Comprimento medio por trecho
    L_medio = ext_m / n_trechos if ext_m > 0 else 60.0

    # Prof base por dific
    prof_base = PROF_DIFIC.get(dific, 1.5)

    # Custo unitario de tubo (R$/m)
    custo_tubo_m = c_red_es / ext_m if ext_m > 0 else 200.0
    # Custo unitario de PV
    custo_pv_un  = c_pvs_tot / n_pvs if n_pvs > 0 else 3078.0

    # CT inicial estimada (Santos: cotas baixas, podem ser negativas)
    ct_ini = rng.uniform(3.0, 7.0)

    trechos = []
    for i in range(n_trechos):
        L     = round(L_medio * rng.uniform(0.80, 1.20), 1)
        pr0   = round(prof_base * rng.uniform(0.85, 1.15), 2)
        pr1   = round(prof_base * rng.uniform(0.85, 1.15), 2)
        decl  = round(decl_base * rng.uniform(0.75, 1.35), 5)

        # DN crescente ao longo da rede
        if dific >= 4:
            dn = 200
        elif dific >= 3:
            dn = 150 if i < n_trechos * 0.6 else 200
        else:
            dn = 100 if i < n_trechos * 0.5 else 150

        # Cotas
        ct0 = round(ct_ini - i * L_medio * decl_base * 0.25, 3)
        cf0 = round(ct0 - pr0, 3)
        ct1 = round(ct0 - L * decl, 3)
        cf1 = round(ct1 - pr1, 3)

        # Hidraulica (Manning, secao circular cheia, n=0.013)
        r_hid = (dn / 1000) / 4
        v_ms  = round((1.0 / 0.013) * (r_hid ** (2.0 / 3.0)) * (decl ** 0.5), 3)
        area  = math.pi * ((dn / 1000) / 2) ** 2
        q_ls  = round(v_ms * area * 1000.0, 3)

        # Quantitativos
        prof_med = (pr0 + pr1) / 2
        if vol_escav > 0:
            escav = round(vol_escav / n_trechos, 3)
        else:
            escav = round(L * prof_med, 3)
        reaterro = round(escav * 0.90, 3)
        bota     = round(escav * 0.10, 3)
        escoram  = round(L * prof_med * 2.0, 2)
        pavim    = round(L * 0.80, 2)

        # Custos
        c_tubo = round(L * custo_tubo_m, 2)
        c_esc  = round(escav * 30.0, 2)
        c_pv   = round(custo_pv_un, 2)
        t_sbdi = round(c_tubo + c_esc + c_pv, 2)
        t_cbdi = round(t_sbdi * 1.25, 2)

        trechos.append(dict(
            tag=tag, nome=nome, setor=setor,
            pv_ini=f"PV-{tag}-{i+1:03d}",
            pv_fim=f"PV-{tag}-{i+2:03d}",
            L=L, dn=dn, mat="PVC",
            ct0=ct0, cf0=cf0, pr0=pr0,
            ct1=ct1, cf1=cf1, pr1=pr1,
            decl=decl, v_ms=v_ms, q_ls=q_ls,
            pav=pav_str,
            escav=escav, reaterro=reaterro, bota=bota,
            escoram=escoram, pavim=pavim,
            c_tubo=c_tubo, c_escav=c_esc, c_pv=c_pv,
            t_sbdi=t_sbdi, t_cbdi=t_cbdi,
            dific=dific, urgencia=urgencia,
            prox1_nome=prox1_nome or "",
            prox1_dist=prox1_dist,
        ))
    return trechos


# --------------------------------------------------------------------------- #
#  Escrita XLSX
# --------------------------------------------------------------------------- #
def escrever_header_trechos(ws, titulo):
    nc = len(COLS)
    ws.merge_cells(f"A1:{get_column_letter(nc)}1")
    c = ws.cell(1, 1, titulo)
    c.fill = fill_h1; c.font = font_h1; c.alignment = ac
    ws.row_dimensions[1].height = 22
    for i, col in enumerate(COLS, 1):
        c = ws.cell(2, i, col)
        c.fill = fill_h2; c.font = font_h2; c.alignment = ac; c.border = thin
    ws.row_dimensions[2].height = 32
    for i, w in enumerate(LARGURAS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"


def escrever_trecho(ws, row_idx, tid, cluster_nome, t, zebra=False):
    vals = [
        f"T-{tid:05d}", cluster_nome, t["nome"], t["tag"], t["setor"],
        t["pv_ini"], t["pv_fim"],
        round(t["L"], 1), t["dn"], t["mat"],
        round(t["ct0"], 3), round(t["cf0"], 3), round(t["pr0"], 2),
        round(t["ct1"], 3), round(t["cf1"], 3), round(t["pr1"], 2),
        round(t["decl"], 5), round(t["v_ms"], 3), round(t["q_ls"], 3),
        t["pav"],
        t["escav"], t["reaterro"], t["bota"],
        t["escoram"], t["pavim"],
        t["c_tubo"], t["c_escav"], t["c_pv"],
        t["t_sbdi"], t["t_cbdi"],
        t["dific"], t["urgencia"],
        t["prox1_nome"],
        round(t["prox1_dist"], 1) if t["prox1_dist"] else "",
    ]
    LEFT_COLS = {3, 5, 33}
    for col, v in enumerate(vals, 1):
        c = ws.cell(row_idx, col, v)
        c.font = font_d
        c.alignment = al if col in LEFT_COLS else ac
        c.border = thin
        if zebra:
            c.fill = fill_zebra


def escrever_sep_cluster(ws, row_idx, cluster_nome, n_nuc, n_tr):
    nc = len(COLS)
    ws.merge_cells(f"A{row_idx}:{get_column_letter(nc)}{row_idx}")
    c = ws.cell(row_idx, 1,
        f"  CLUSTER: {cluster_nome}  —  {n_nuc} nucleo(s)  |  {n_tr} trechos")
    c.fill = fill_clust; c.font = font_clust; c.alignment = al
    ws.row_dimensions[row_idx].height = 18


def escrever_sep_nucleo(ws, row_idx, nome, n_tr):
    nc = len(COLS)
    ws.merge_cells(f"A{row_idx}:{get_column_letter(nc)}{row_idx}")
    c = ws.cell(row_idx, 1, f"    Nucleo: {nome}  —  {n_tr} trechos")
    c.fill = fill_nucl; c.font = font_nucl; c.alignment = al
    ws.row_dimensions[row_idx].height = 15


def escrever_subtotal(ws, row_idx, label, tot):
    vals = [label] + [""] * 6 + [round(tot["L"], 1)] + [""] * 11 + [
        round(tot["escav"], 2), round(tot["reaterro"], 2), round(tot["bota"], 2),
        "", round(tot["pavim"], 2),
        round(tot["c_tubo"], 2), round(tot["c_escav"], 2), round(tot["c_pv"], 2),
        round(tot["t_sbdi"], 2), round(tot["t_cbdi"], 2),
        "", "", "", "",
    ]
    # pad to len(COLS)
    while len(vals) < len(COLS):
        vals.append("")
    for col, v in enumerate(vals[:len(COLS)], 1):
        c = ws.cell(row_idx, col, v)
        c.fill = fill_sub; c.font = font_sub; c.alignment = ac; c.border = thin


def escrever_grand_total(ws, row_idx, n_tr, tot):
    label = f"GRAND TOTAL — 83 NUCLEOS — {n_tr} TRECHOS"
    vals = [label] + [""] * 6 + [round(tot["L"], 1)] + [""] * 11 + [
        round(tot["escav"], 2), round(tot["reaterro"], 2), round(tot["bota"], 2),
        "", round(tot["pavim"], 2),
        round(tot["c_tubo"], 2), round(tot["c_escav"], 2), round(tot["c_pv"], 2),
        round(tot["t_sbdi"], 2), round(tot["t_cbdi"], 2),
        "", "", "", "",
    ]
    while len(vals) < len(COLS):
        vals.append("")
    for col, v in enumerate(vals[:len(COLS)], 1):
        c = ws.cell(row_idx, col, v)
        c.fill = fill_h1; c.font = font_grand; c.alignment = ac; c.border = thin


def zero_tot():
    return dict(L=0.0, escav=0.0, reaterro=0.0, bota=0.0, pavim=0.0,
                c_tubo=0.0, c_escav=0.0, c_pv=0.0, t_sbdi=0.0, t_cbdi=0.0)

def add_tot(a, t):
    a["L"]       += t["L"]
    a["escav"]   += t["escav"]
    a["reaterro"]+= t["reaterro"]
    a["bota"]    += t["bota"]
    a["pavim"]   += t["pavim"]
    a["c_tubo"]  += t["c_tubo"]
    a["c_escav"] += t["c_escav"]
    a["c_pv"]    += t["c_pv"]
    a["t_sbdi"]  += t["t_sbdi"]
    a["t_cbdi"]  += t["t_cbdi"]


def escrever_resumo(ws, clusters_info):
    nc = len(COLS_RES)
    titulo = ("RESUMO CLUSTERS DE PROXIMIDADE GEOGRAFICA — "
              "ConstruData HydroNetwork v7.0.0  |  FCN Construcoes  |  Contrato 11481051")
    ws.merge_cells(f"A1:{get_column_letter(nc)}1")
    c = ws.cell(1, 1, titulo)
    c.fill = fill_h1; c.font = font_h1; c.alignment = ac
    ws.row_dimensions[1].height = 22
    for i, col in enumerate(COLS_RES, 1):
        c = ws.cell(2, i, col)
        c.fill = fill_h2; c.font = font_h2; c.alignment = ac; c.border = thin
    ws.row_dimensions[2].height = 28
    for i, w in enumerate(LARG_RES, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"

    for row_idx, ci in enumerate(clusters_info, 3):
        zebra = (row_idx % 2 == 0)
        vals = [
            ci["nome"], ci["ancora"], ci["n_nucleos"], ci["n_trechos"],
            ci["lig_total"], round(ci["ext_m"], 1), round(ci["custo_cbdi"], 2),
            ci["n_pvs"], round(ci["vol_escav"], 1), ci["urgencia_dom"],
            ci["tags_str"],
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row_idx, col, v)
            c.font = font_d
            c.alignment = al if col in (2, 10, 11) else ac
            c.border = thin
            if zebra:
                c.fill = fill_zebra


# --------------------------------------------------------------------------- #
#  MAIN
# --------------------------------------------------------------------------- #
print("Lendo CSVs...")
if not os.path.isfile(CSV_NUCLEOS):
    print(f"ERRO: nao encontrado: {CSV_NUCLEOS}")
    sys.exit(1)
if not os.path.isfile(CSV_PROX):
    print(f"ERRO: nao encontrado: {CSV_PROX}")
    sys.exit(1)

nucleos_dict = ler_nucleos(CSV_NUCLEOS)
prox_rows    = ler_prox(CSV_PROX)
print(f"  {len(nucleos_dict)} nucleos em nucleos_enriquecido.csv")
print(f"  {len(prox_rows)} linhas em SLNR_MEGA_INTEGRADA_NUCLEOS_PROX_SETOR.csv")

# Proximidade mais proxima para display
prox_info = {}
for row in prox_rows:
    tag = str(row.get("TAG", "")).strip()
    p1n = str(row.get("PROX1_NOME", "") or "").strip()
    p1d = _f(row.get("PROX1_DIST_M", 0))
    prox_info[tag] = (p1n, p1d)

# Grafo de proximidade
print(f"Montando grafo (threshold={DIST_THRESHOLD}m)...")
all_tags = list(nucleos_dict.keys())
graph = defaultdict(set)
for row in prox_rows:
    tag = str(row.get("TAG", "")).strip()
    if tag not in nucleos_dict:
        continue
    for i in [1, 2, 3]:
        prox_tag  = str(row.get(f"PROX{i}_TAG", "") or "").strip()
        prox_dist = _f(row.get(f"PROX{i}_DIST_M", "") or 9999)
        if prox_tag and prox_tag in nucleos_dict and prox_dist <= DIST_THRESHOLD:
            graph[tag].add(prox_tag)
            graph[prox_tag].add(tag)

# Union-Find
clusters_raw = connected_components(graph, all_tags)

# Ordena clusters: maior numero de ligacoes primeiro
def cluster_lig_sum(tags_list):
    return sum(_i(nucleos_dict[t].get("LIG_TOTAL", 0)) for t in tags_list if t in nucleos_dict)

clusters_raw.sort(key=lambda tl: -cluster_lig_sum(tl))
for tl in clusters_raw:
    tl.sort(key=lambda t: -_i(nucleos_dict[t].get("LIG_TOTAL", 0)) if t in nucleos_dict else 0)

# Nomeia clusters: setor dominante + letra
setor_cnt = defaultdict(int)
cluster_names = []
for tl in clusters_raw:
    setor_lig = defaultdict(int)
    for t in tl:
        if t in nucleos_dict:
            raw = nucleos_dict[t].get("SETOR", "GERAL")
            s = raw.replace("ST-", "").replace("ZONA ", "").strip()
            setor_lig[s] += _i(nucleos_dict[t].get("LIG_TOTAL", 0))
    dom = max(setor_lig, key=setor_lig.get) if setor_lig else "GERAL"
    short = dom[:18].replace(" ", "_").replace("/", "_").upper()
    setor_cnt[short] += 1
    letra = chr(64 + setor_cnt[short])
    cluster_names.append(f"{short}_{letra}")

print(f"  {len(clusters_raw)} clusters formados")
for cn, tl in zip(cluster_names, clusters_raw):
    print(f"    {cn}: {len(tl)} nucleo(s) — {', '.join(tl)}")

# --------------------------------------------------------------------------- #
#  Gerar trechos
# --------------------------------------------------------------------------- #
print("Gerando trechos inferidos...")

# Estrutura: list of (cluster_nome, [(tag, nome, [trechos])], tot_cluster)
all_clusters = []
grand_tot = zero_tot()
grand_n_tr = 0
clusters_info = []

for cname, tags in zip(cluster_names, clusters_raw):
    cl_tot = zero_tot()
    cl_lig = cl_pvs = cl_n_tr = 0
    cl_ext = cl_vol = cl_cbdi = 0.0
    urgencias = []
    ancora_nome = ""
    nucleo_list = []

    for idx, tag in enumerate(tags):
        if tag not in nucleos_dict:
            continue
        nr = nucleos_dict[tag]
        pn, pd = prox_info.get(tag, ("", 0.0))
        trechos = inferir_trechos(nr, pn, pd)
        nucleo_list.append((tag, str(nr["NOME"]).strip(), trechos))

        for t in trechos:
            add_tot(cl_tot, t)
            add_tot(grand_tot, t)
        cl_n_tr   += len(trechos)
        grand_n_tr += len(trechos)
        cl_lig    += _i(nr.get("LIG_TOTAL", 0))
        cl_ext    += _f(nr.get("EXT_ES_M", 0))
        cl_pvs    += _i(nr.get("N_PVS", 0))
        cl_vol    += _f(nr.get("VOL_ESCAV", 0))
        cl_cbdi   += sum(t["t_cbdi"] for t in trechos)
        urgencias.append(str(nr.get("URGENCIA", "")).strip())
        if idx == 0:
            ancora_nome = str(nr["NOME"]).strip()

    all_clusters.append((cname, nucleo_list, cl_tot, cl_n_tr))

    urg_dom = Counter(urgencias).most_common(1)[0][0] if urgencias else ""
    clusters_info.append(dict(
        nome=cname, ancora=ancora_nome,
        n_nucleos=len(nucleo_list), n_trechos=cl_n_tr,
        lig_total=cl_lig, ext_m=cl_ext, custo_cbdi=cl_cbdi,
        n_pvs=cl_pvs, vol_escav=cl_vol, urgencia_dom=urg_dom,
        tags_str=", ".join(tags),
    ))

print(f"  Total: {grand_n_tr} trechos em {len(all_clusters)} clusters")

# --------------------------------------------------------------------------- #
#  Montar XLSX
# --------------------------------------------------------------------------- #
print("Criando XLSX...")
wb = openpyxl.Workbook()

TITULO_GERAL = ("TRECHOS MEGA INTEGRADA — CLUSTERS DE PROXIMIDADE GEOGRAFICA — "
                "ConstruData HydroNetwork v7.0.0  |  FCN Construcoes  |  Contrato 11481051")

# Aba RESUMO_CLUSTERS (primeira)
ws_res = wb.active
ws_res.title = "RESUMO_CLUSTERS"
escrever_resumo(ws_res, clusters_info)

# Aba TODOS_83_NUCLEOS
ws_all = wb.create_sheet("TODOS_83_NUCLEOS")
escrever_header_trechos(ws_all, TITULO_GERAL)
row_all = 3
tid_all = 1

for cname, nucleo_list, cl_tot, cl_n_tr in all_clusters:
    escrever_sep_cluster(ws_all, row_all, cname,
                         len(nucleo_list), cl_n_tr)
    row_all += 1
    cl_cont_tot = zero_tot()

    for tag, nome, trechos in nucleo_list:
        escrever_sep_nucleo(ws_all, row_all, nome, len(trechos))
        row_all += 1
        nuc_tot = zero_tot()
        for zi, t in enumerate(trechos):
            zebra = (zi % 2 == 0)
            escrever_trecho(ws_all, row_all, tid_all, cname, t, zebra)
            add_tot(nuc_tot, t)
            row_all += 1
            tid_all += 1
        escrever_subtotal(ws_all, row_all, f"  Sub {nome[:30]}", nuc_tot)
        add_tot(cl_cont_tot, nuc_tot)
        row_all += 1

    escrever_subtotal(ws_all, row_all, f"SUBTOTAL {cname}", cl_cont_tot)
    row_all += 2

escrever_grand_total(ws_all, row_all, grand_n_tr, grand_tot)

# Abas por cluster
used_aba = set(["RESUMO_CLUSTERS", "TODOS_83_NUCLEOS"])
for cname, nucleo_list, cl_tot, cl_n_tr in all_clusters:
    aba = safe_aba(cname)
    # garante unicidade
    base = aba
    suf  = 2
    while aba in used_aba:
        aba = safe_aba(f"{base}_{suf}", 31)
        suf += 1
    used_aba.add(aba)

    ws_c = wb.create_sheet(title=aba)
    titulo_cl = (f"CLUSTER: {cname} — {len(nucleo_list)} nucleo(s) — {cl_n_tr} trechos  |  "
                 "ConstruData HydroNetwork v7.0.0  |  FCN Construcoes  |  Contrato 11481051")
    escrever_header_trechos(ws_c, titulo_cl)
    row_c = 3
    tid_c = 1
    cl_tot2 = zero_tot()

    for tag, nome, trechos in nucleo_list:
        escrever_sep_nucleo(ws_c, row_c, nome, len(trechos))
        row_c += 1
        nuc_tot = zero_tot()
        for zi, t in enumerate(trechos):
            zebra = (zi % 2 == 0)
            escrever_trecho(ws_c, row_c, tid_c, cname, t, zebra)
            add_tot(nuc_tot, t)
            row_c += 1
            tid_c += 1
        escrever_subtotal(ws_c, row_c, f"  Sub {nome[:30]}", nuc_tot)
        add_tot(cl_tot2, nuc_tot)
        row_c += 1

    escrever_subtotal(ws_c, row_c, f"TOTAL {cname}", cl_tot2)

wb.save(OUT_FILE)
print(f"\nSalvo: {OUT_FILE}")
print(f"Abas ({len(wb.worksheets)}): {[ws.title for ws in wb.worksheets]}")
print(f"Total clusters: {len(all_clusters)}")
print(f"Total trechos : {grand_n_tr}")
print(f"Total ext.(m) : {round(grand_tot['L'], 1)}")
print(f"Total c/BDI   : R$ {grand_tot['t_cbdi']:,.2f}")
