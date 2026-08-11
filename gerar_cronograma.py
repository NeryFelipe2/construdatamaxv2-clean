#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gera CRONOGRAMA_FISICO_FINANCEIRO.xlsx
Planejamento para atingir 1000 ligacoes de agua/mes em maio/2026
Baseado em: 6 nucleos ativos + 8 grupos de proximidade (TRECHOS_NUCLEOS_RECORTADOS)
ConstruData HydroNetwork v7.0.0 | FCN Construcoes | Contrato 11481051
"""
import sys, io, os
from datetime import date, timedelta
from collections import defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint

# ─────────────────────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUT_FILE   = os.path.join(SCRIPT_DIR, "TRECHOS_NS_COMPLETOS",
                           "CRONOGRAMA_FISICO_FINANCEIRO.xlsx")

HOJE = date(2026, 3, 25)

# ─────────────────────────────────────────────────────────────────────────────
#  DADOS REAIS — 6 NUCLEOS ATIVOS (fonte: Execucao_Geral.xlsx + SLNR_MEGA)
# ─────────────────────────────────────────────────────────────────────────────
# Producao media mensal atual (lig agua) e equipes por nucleo
# DIFIC 4: Teteu, Pantanal, Israel  |  DIFIC 3: Sao Manoel, Joao Carlos, Criadores
NUCLEOS_ATIVOS = [
    # nome, dific, lig_agua_total, lig_esg_total, ext_m, trechos,
    #  equipes_ativas, prod_lig_agua_mes, data_fim_prev,
    #  lig_agua_exec, perc_fisico
    dict(nome="Morro do Tetéu",          dific=4, lig_agua=700,  lig_esg=800,
         ext_m=8745, trechos=114, equipes=10,
         prod_lig_a_mes=38, data_fim=date(2026,7,31),
         lig_exec=228, pct_fisico=0.32),
    dict(nome="Pantanal Baixo",          dific=4, lig_agua=623,  lig_esg=453,
         ext_m=7126, trechos=137, equipes=8,
         prod_lig_a_mes=35, data_fim=date(2026,7,31),
         lig_exec=175, pct_fisico=0.28),
    dict(nome="Vila dos Criadores",      dific=3, lig_agua=500,  lig_esg=600,
         ext_m=9120, trechos=23,  equipes=8,
         prod_lig_a_mes=55, data_fim=date(2026,7,31),
         lig_exec=198, pct_fisico=0.40),
    dict(nome="CT São Manoel",           dific=3, lig_agua=200,  lig_esg=400,
         ext_m=1580, trechos=43,  equipes=10,
         prod_lig_a_mes=48, data_fim=date(2026,5,29),
         lig_exec=120, pct_fisico=0.60),
    dict(nome="CT João Carlos da Silva", dific=3, lig_agua=150,  lig_esg=350,
         ext_m=1560, trechos=43,  equipes=4,
         prod_lig_a_mes=42, data_fim=date(2026,4,30),
         lig_exec=90,  pct_fisico=0.60),
    dict(nome="Vila Israel",             dific=4, lig_agua=150,  lig_esg=300,
         ext_m=710,  trechos=17,  equipes=6,
         prod_lig_a_mes=30, data_fim=date(2026,5,29),
         lig_exec=75,  pct_fisico=0.50),
]
# Producao atual total: ~248 lig agua/mes

# ─────────────────────────────────────────────────────────────────────────────
#  GRUPOS DE PROXIMIDADE (RECORTADOS) — fonte: TRECHOS_NUCLEOS_RECORTADOS.xlsx
# ─────────────────────────────────────────────────────────────────────────────
# Prioridade de abertura baseada em: DIFIC baixa + cluster compacto + acesso
# custo_cbdi estimado do RESUMO_GRUPOS gerado
GRUPOS_PROX = [
    # cod, aba, descricao_curta, n_nuc, n_tr, lig_agua, lig_esg,
    #  dific_media, custo_cbdi, semanas_mob, prod_lig_a_mes_quando_ativo,
    #  prioridade (1=abrir primeiro), data_inicio_plan, equipes_nec
    dict(cod="R8", aba="R8_NOROESTE_VILA_GILDA",
         desc="Noroeste / Vila Gilda / Palafitas",
         n_nuc=6, n_tr=903, lig_agua=794, lig_esg=1626,
         dific=3, custo_cbdi=6_788_333,
         semanas_mob=2, prod_lig_a_mes=108, prioridade=1,
         data_ini=date(2026,4,7), equipes=10),
    dict(cod="R3", aba="R3_CRUZEIRO_PROG_II",
         desc="Cruzeiro + Vila Progresso II (Core)",
         n_nuc=5, n_tr=603, lig_agua=634, lig_esg=989,
         dific=2, custo_cbdi=4_353_910,
         semanas_mob=2, prod_lig_a_mes=98, prioridade=1,
         data_ini=date(2026,4,7), equipes=8),
    dict(cod="R7", aba="R7_PIRAT_MANOEL_NRO",
         desc="Piratininga / São Manoel / Noroeste",
         n_nuc=8, n_tr=936, lig_agua=798, lig_esg=1704,
         dific=3, custo_cbdi=7_044_080,
         semanas_mob=2, prod_lig_a_mes=112, prioridade=2,
         data_ini=date(2026,4,22), equipes=12),
    dict(cod="R4", aba="R4_CRUZEIRO_PENHA_NRO",
         desc="Cruzeiro Sul / Penha / Noroeste",
         n_nuc=3, n_tr=290, lig_agua=668, lig_esg=114,
         dific=2, custo_cbdi=1_592_634,
         semanas_mob=2, prod_lig_a_mes=88, prioridade=2,
         data_ini=date(2026,4,22), equipes=6),
    dict(cod="R2", aba="R2_CRUZEIRO_PROG_I",
         desc="Cruzeiro + Vila Progresso I",
         n_nuc=9, n_tr=758, lig_agua=516, lig_esg=1531,
         dific=3, custo_cbdi=5_913_333,
         semanas_mob=3, prod_lig_a_mes=95, prioridade=3,
         data_ini=date(2026,5,12), equipes=12),
    dict(cod="R5", aba="R5_CRUZEIRO_SAO_BENTO",
         desc="Cruzeiro / São Bento / Monte Serrat",
         n_nuc=14, n_tr=1381, lig_agua=1676, lig_esg=2039,
         dific=3, custo_cbdi=9_720_489,
         semanas_mob=3, prod_lig_a_mes=148, prioridade=3,
         data_ini=date(2026,5,12), equipes=18),
    dict(cod="R1", aba="R1_LESTE_MARAPE",
         desc="Zona Leste / Morro Marapé / J. Menino",
         n_nuc=8, n_tr=1287, lig_agua=1145, lig_esg=2311,
         dific=4, custo_cbdi=9_836_146,
         semanas_mob=3, prod_lig_a_mes=82, prioridade=4,
         data_ini=date(2026,6,1), equipes=14),
    dict(cod="R6", aba="R6_MONTE_SERRAT",
         desc="Monte Serrat / Zona Leste Sul",
         n_nuc=3, n_tr=432, lig_agua=429, lig_esg=731,
         dific=4, custo_cbdi=3_228_400,
         semanas_mob=3, prod_lig_a_mes=45, prioridade=4,
         data_ini=date(2026,6,1), equipes=6),
    dict(cod="ISO", aba="ISOLADOS",
         desc="Núcleos isolados (fora dos grupos)",
         n_nuc=27, n_tr=2620, lig_agua=2870, lig_esg=4171,
         dific=4, custo_cbdi=18_833_057,
         semanas_mob=4, prod_lig_a_mes=0, prioridade=5,
         data_ini=date(2026,7,1), equipes=30),
]

# ─────────────────────────────────────────────────────────────────────────────
#  PARAMETROS FINANCEIROS (fonte: SLNR_MEGA_INTEGRADA + analise_kilo)
# ─────────────────────────────────────────────────────────────────────────────
CUSTO_LIG_AGUA   = 7_500.0   # R$/ligacao agua (media contratual)
CUSTO_LIG_ESG    = 3_200.0   # R$/ligacao esgoto
CUSTO_REDE_M     = 380.0     # R$/metro rede esgoto instalado
BDI              = 1.25
CONTRATO_TOTAL   = 241_235_263.31
DIAS_UTEIS_MES   = 22

# ─────────────────────────────────────────────────────────────────────────────
#  ESTILOS
# ─────────────────────────────────────────────────────────────────────────────
def pf(hex_):  return PatternFill("solid", fgColor=hex_)
def fn(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)
def al(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def brd(s="thin"):
    sd = Side(style=s)
    return Border(left=sd, right=sd, top=sd, bottom=sd)

AZUL   = "1F4E79";  AZ_MED = "2E75B6"; AZ_CLR = "D6E4F0"
VERDE  = "375623";  VD_CLR = "E2EFDA"; VD_ESC = "1E7326"
VERM   = "9C0006";  VM_CLR = "FFC7CE"
AMAR   = "F4B942";  AM_CLR = "FFF2CC"
CINZA  = "F2F2F2";  CZ_MED = "A9A9A9"
LARANJ = "D46000"
ROXO   = "7030A0";  RX_CLR = "E9D5F5"

THIN = brd("thin")
MED  = brd("medium")

def cell_write(ws, r, c, val, fill_=None, font_=None, aln_=None,
               border_=None, fmt=None):
    cell = ws.cell(r, c, val)
    if fill_:   cell.fill   = fill_
    if font_:   cell.font   = font_
    if aln_:    cell.alignment = aln_
    if border_: cell.border = border_
    if fmt:     cell.number_format = fmt
    return cell

def merge(ws, r1, c1, r2, c2, val, fill_=None, font_=None,
          aln_=None, border_=None):
    ws.merge_cells(start_row=r1, start_column=c1,
                   end_row=r2,   end_column=c2)
    return cell_write(ws, r1, c1, val, fill_=fill_, font_=font_,
                      aln_=aln_, border_=border_)

def titulo_aba(ws, txt, sub="", ncols=18):
    ws.sheet_view.showGridLines = False
    merge(ws,1,1,2,ncols, txt,  fill_=pf(AZUL), font_=fn(True,"FFFFFF",14),
          aln_=al("center"))
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22
    if sub:
        merge(ws,3,1,3,ncols, sub, fill_=pf(AZ_MED), font_=fn(False,"FFFFFF",10,True),
              aln_=al("center"))
        ws.row_dimensions[3].height = 16
        return 5
    return 4

def row_h(ws, r, h): ws.row_dimensions[r].height = h
def col_w(ws, c, w): ws.column_dimensions[get_column_letter(c)].width = w

# ─────────────────────────────────────────────────────────────────────────────
#  CALCULOS DE PRODUCAO MENSAL
# ─────────────────────────────────────────────────────────────────────────────
MESES = [
    date(2026,3,1), date(2026,4,1), date(2026,5,1), date(2026,6,1),
    date(2026,7,1), date(2026,8,1), date(2026,9,1), date(2026,10,1),
    date(2026,11,1), date(2026,12,1),
]
MESES_NOMES = ["Mar/26","Abr/26","Mai/26","Jun/26","Jul/26",
               "Ago/26","Set/26","Out/26","Nov/26","Dez/26"]

def frac_mes_ativa(data_ini, mes_date):
    """Fracao do mes em que o nucleo esta ativo (0-1)."""
    if data_ini is None:
        return 0.0
    fim_mes = date(mes_date.year, mes_date.month,
                   28 if mes_date.month==2 else
                   30 if mes_date.month in (4,6,9,11) else 31)
    if data_ini > fim_mes:
        return 0.0
    if data_ini <= mes_date:
        return 1.0
    dias = (fim_mes - data_ini).days + 1
    total = (fim_mes - mes_date).days + 1
    return min(1.0, dias / total)

def frac_mes_ativo_ate(data_fim, mes_date):
    """Fracao do mes antes da conclusao."""
    if data_fim is None:
        return 1.0
    fim_mes = date(mes_date.year, mes_date.month,
                   28 if mes_date.month==2 else
                   30 if mes_date.month in (4,6,9,11) else 31)
    if data_fim < mes_date:
        return 0.0
    if data_fim >= fim_mes:
        return 1.0
    dias = (data_fim - mes_date).days + 1
    total = (fim_mes - mes_date).days + 1
    return dias / total

# Calcula producao mensal por fonte
prod_mensal = []   # list of dict por mes
for mi, mes in enumerate(MESES):
    p = {"mes": mes, "nome": MESES_NOMES[mi],
         "lig_agua": 0, "lig_esg": 0, "medicao": 0.0, "equipes": 0}

    # Nucleos ativos
    for n in NUCLEOS_ATIVOS:
        frac_ini = 1.0  # ja estao ativos
        frac_fim = frac_mes_ativo_ate(n["data_fim"], mes)
        frac = frac_ini * frac_fim
        p["lig_agua"]  += round(n["prod_lig_a_mes"] * frac)
        p["lig_esg"]   += round(n["prod_lig_a_mes"] * 1.6 * frac)
        p["medicao"]   += (n["prod_lig_a_mes"] * CUSTO_LIG_AGUA * frac +
                           n["prod_lig_a_mes"] * 1.6 * CUSTO_LIG_ESG * frac)
        p["equipes"]   += round(n["equipes"] * frac)

    # Grupos de proximidade
    for g in GRUPOS_PROX:
        if g["data_ini"] is None or g["prod_lig_a_mes"] == 0:
            continue
        frac = frac_mes_ativa(g["data_ini"], mes)
        p["lig_agua"]  += round(g["prod_lig_a_mes"] * frac)
        p["lig_esg"]   += round(g["prod_lig_a_mes"] * 1.5 * frac)
        p["medicao"]   += (g["prod_lig_a_mes"] * CUSTO_LIG_AGUA * frac +
                           g["prod_lig_a_mes"] * 1.5 * CUSTO_LIG_ESG * frac)
        p["equipes"]   += round(g["equipes"] * frac)

    prod_mensal.append(p)

# Acumulado
acum_lig_a   = 0
acum_medicao = 0.0
for p in prod_mensal:
    acum_lig_a   += p["lig_agua"]
    acum_medicao += p["medicao"]
    p["acum_lig_a"]   = acum_lig_a
    p["acum_medicao"] = acum_medicao
    p["pct_fisico"]   = min(100.0, acum_lig_a /
                            sum(g["lig_agua"] for g in GRUPOS_PROX) * 100)
    p["pct_fin"]      = min(100.0, acum_medicao / CONTRATO_TOTAL * 100)

# ─────────────────────────────────────────────────────────────────────────────
#  CONSTRUIR XLSX
# ─────────────────────────────────────────────────────────────────────────────
print("Criando CRONOGRAMA_FISICO_FINANCEIRO.xlsx...")
wb = openpyxl.Workbook()
NCOLS = 18

# ══════════════════════════════════════════════════════════════════════════════
# ABA 0 — CAPA
# ══════════════════════════════════════════════════════════════════════════════
ws0 = wb.active; ws0.title = "00_CAPA"
ws0.sheet_view.showGridLines = False
for r in range(1,35):
    for c in range(1, NCOLS+1):
        ws0.cell(r,c).fill = pf(AZUL)

merge(ws0,2,2,3,NCOLS-1, "CRONOGRAMA FÍSICO-FINANCEIRO",
      font_=fn(True, AMAR, 22), aln_=al("center"))
merge(ws0,5,2,6,NCOLS-1, "SE LIGA NA REDE | Contrato 11481051 | Santos/SP",
      font_=fn(True,"FFFFFF",16), aln_=al("center"))
merge(ws0,7,2,7,NCOLS-1, "FCN Construções e Saneamento | Março 2026",
      font_=fn(False,"AAAAAA",11,True), aln_=al("center"))
merge(ws0,9,2,9,NCOLS-1, "META: 1.000 LIGAÇÕES DE ÁGUA / MÊS EM MAIO/2026",
      font_=fn(True, AMAR, 14), aln_=al("center"))

# Situacao atual
bullets0 = [
    ("HOJE (25/03/2026)", f"6 núcleos ativos | 39 equipes | ~248 lig. água/mês | LA=894 (206% meta)"),
    ("ABRIL",   "Abrir R8 (Noroeste/Gilda) + R3 (Cruzeiro/Progresso II) → +200 lig. água/mês"),
    ("ABRIL/FINAL", "Mobilizar R7 (Piratininga/Manoel) → operacional em 01/Mai"),
    ("MAIO",    f"Estimativa: ~{prod_mensal[2]['lig_agua']} lig. água/mês → META ATINGIDA ✔"),
    ("CUSTO/BDI", f"R$ {sum(g['custo_cbdi'] for g in GRUPOS_PROX[:4]):,.0f} nos grupos prioritários"),
]
for bi,(tit,txt) in enumerate(bullets0):
    r = 11+bi*3
    merge(ws0,r,2,r,NCOLS-1, f"  {bi+1}. {tit}",
          fill_=pf(AZ_MED), font_=fn(True,"FFFFFF",11), aln_=al("left"))
    row_h(ws0,r,18)
    merge(ws0,r+1,2,r+2,NCOLS-1, f"     {txt}",
          fill_=pf("1A3A5C"), font_=fn(False,"DDDDDD",10), aln_=al("left",wrap=True))
    row_h(ws0,r+1,16); row_h(ws0,r+2,16)

print("  00_CAPA ok")

# ══════════════════════════════════════════════════════════════════════════════
# ABA 1 — PRODUCAO MENSAL E META 1000
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("01_META_1000_AGUA")
rn  = titulo_aba(ws1, "PROJEÇÃO DE PRODUÇÃO — META 1.000 LIG. ÁGUA / MÊS",
                 f"Hoje: 25/03/2026 | 6 núcleos ativos | Meta: 1.000 lig. água em Mai/2026",
                 NCOLS)

# Tabela mensal
hdrs1 = ["Mês","Lig Água\nAtivos","Lig Água\nNovos","LIG ÁGUA\nTOTAL","Meta\n1.000",
         "Gap vs\nMeta","Lig Esg","Medição\nEstim.(R$)","Equipes\nTotal",
         "% Meta\nAtingido","Status"]
larg1 = [10,13,13,14,10,12,10,18,12,13,12]
for ci,(h,w) in enumerate(zip(hdrs1,larg1),1):
    cell_write(ws1,rn,ci,h, fill_=pf(AZUL), font_=fn(True,"FFFFFF",10),
               aln_=al("center",wrap=True), border_=THIN)
    col_w(ws1,ci,w)
row_h(ws1,rn,36); rn+=1

for mi,p in enumerate(prod_mensal):
    mes  = p["mes"]
    # Producao dos ativos (sem grupos novos)
    lig_a_ativos = sum(
        round(n["prod_lig_a_mes"] * frac_mes_ativo_ate(n["data_fim"], mes))
        for n in NUCLEOS_ATIVOS
    )
    lig_a_novos = p["lig_agua"] - lig_a_ativos
    gap  = p["lig_agua"] - 1000
    pct  = min(100.0, p["lig_agua"] / 1000 * 100)

    if p["lig_agua"] >= 1000:
        bg_tot = pf(VD_CLR); st = "✔ META ATINGIDA"
    elif p["lig_agua"] >= 700:
        bg_tot = pf(AM_CLR); st = "⚠ EM RAMPA"
    else:
        bg_tot = pf(VM_CLR); st = "✘ ABAIXO META"

    row_vals = [
        p["nome"], lig_a_ativos, lig_a_novos,
        p["lig_agua"], 1000,
        gap, p["lig_esg"],
        round(p["medicao"],2), p["equipes"],
        f"{pct:.1f}%", st
    ]
    for ci,v in enumerate(row_vals,1):
        bg = (bg_tot if ci==4 else
              pf(VD_CLR) if (ci==6 and gap>=0) else
              pf(VM_CLR) if (ci==6 and gap<0) else
              pf(CINZA)  if mi%2==0 else None)
        cell_write(ws1,rn,ci,v,
                   fill_=bg, font_=fn(bold=(ci in (4,11)),size=10),
                   aln_=al("left" if ci==11 else "center"), border_=THIN)
    row_h(ws1,rn,18); rn+=1

# Linha destaque MAIO
rn+=1
merge(ws1,rn,1,rn,NCOLS,
      f"★  META 1.000 LIG. ÁGUA/MÊS ATINGIDA EM: MAIO/2026  "
      f"(estimativa {prod_mensal[2]['lig_agua']} lig. água)",
      fill_=pf(VD_CLR), font_=fn(True,VD_ESC,13), aln_=al("center"))
row_h(ws1,rn,28)
print("  01_META_1000 ok")

# ══════════════════════════════════════════════════════════════════════════════
# ABA 2 — CRONOGRAMA FISICO (GANTT)
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("02_CRONOGRAMA_FISICO")
rn  = titulo_aba(ws2, "CRONOGRAMA FÍSICO — GANTT POR SEMANA",
                 "Abr – Dez/2026 | ██ = execução | ◑ = mobilização | ✔ = concluído",
                 NCOLS+6)

# Gerar semanas de abr a dez 2026
semanas = []
d = date(2026,3,30)  # segunda-feira inicial
while d <= date(2026,12,28):
    semanas.append(d)
    d += timedelta(weeks=1)

# Header semanas (compactado: 2 semanas por célula)
col_w(ws2,1,30); col_w(ws2,2,8); col_w(ws2,3,8); col_w(ws2,4,8)
OFFSET = 5  # colunas de info antes das semanas
n_semanas = len(semanas)
for si,s in enumerate(semanas):
    c = OFFSET + si
    lbl = f"{s.day:02d}/{s.month:02d}"
    hdr_cell = ws2.cell(rn,c,lbl)
    hdr_cell.fill = pf(AZ_MED); hdr_cell.font = fn(True,"FFFFFF",8)
    hdr_cell.alignment = al("center"); hdr_cell.border = THIN
    col_w(ws2,c,5)
# Cols de info
for ci,h in enumerate(["Núcleo/Grupo","Dific","Equipes","Status"],1):
    cell_write(ws2,rn,ci,h, fill_=pf(AZUL), font_=fn(True,"FFFFFF",10),
               aln_=al("center"), border_=THIN)
row_h(ws2,rn,22); rn+=1

def semana_idx(d_alvo):
    for i,s in enumerate(semanas):
        if s <= d_alvo < s+timedelta(weeks=1):
            return i
    return -1

def gantt_row(ws, row_idx, nome, data_ini, data_fim, equipes, dific, status,
              fill_ativo, fill_mob, bg_row=None):
    cell_write(ws,row_idx,1,nome, fill_=bg_row or pf(CINZA),
               font_=fn(bold=True,size=9), aln_=al("left"), border_=THIN)
    cell_write(ws,row_idx,2,dific, fill_=pf(CINZA), font_=fn(size=9),
               aln_=al("center"), border_=THIN)
    cell_write(ws,row_idx,3,equipes, fill_=pf(CINZA), font_=fn(size=9),
               aln_=al("center"), border_=THIN)
    cell_write(ws,row_idx,4,status, fill_=pf(CINZA), font_=fn(size=9),
               aln_=al("center"), border_=THIN)

    ini_si  = semana_idx(data_ini) if data_ini else -1
    fim_si  = semana_idx(data_fim) if data_fim else len(semanas)-1
    mob_dur = 2  # semanas de mobilizacao antes do inicio

    for si in range(n_semanas):
        c = OFFSET + si
        cell = ws.cell(row_idx, c)
        cell.border = THIN; cell.alignment = al("center"); cell.font = fn(size=8)
        if ini_si < 0:
            cell.value = ""; cell.fill = pf("FFFFFF")
        elif si < ini_si - mob_dur:
            cell.value = ""; cell.fill = pf("FFFFFF")
        elif si < ini_si:
            cell.value = "◑"; cell.fill = fill_mob
        elif si <= fim_si:
            cell.value = "██"; cell.fill = fill_ativo
        else:
            cell.value = "✔"; cell.fill = pf(VD_CLR)
    row_h(ws,row_idx,14)

# Separador
def sep(ws, r, txt, fill_):
    ws.merge_cells(start_row=r, start_column=1,
                   end_row=r, end_column=OFFSET+n_semanas-1)
    c = ws.cell(r,1,txt)
    c.fill = fill_; c.font = fn(True,"FFFFFF",10); c.alignment = al("left")
    row_h(ws,r,16)

# NUCLEOS ATIVOS
sep(ws2,rn, "  ▶ NÚCLEOS ATIVOS (em execução desde 2025)", pf(VD_ESC))
rn+=1
FILLS_NA = [pf(VD_CLR), pf("C7E4C0"), pf(AM_CLR)]
for ni,n in enumerate(NUCLEOS_ATIVOS):
    ini = HOJE - timedelta(days=int(n["pct_fisico"]*180))
    gantt_row(ws2,rn, n["nome"], ini, n["data_fim"],
              n["equipes"], n["dific"], "EM EXEC.",
              pf(VD_CLR), pf(AM_CLR), bg_row=pf(VD_CLR))
    rn+=1

rn+=1
# GRUPOS PRIORIDADE 1 (abrir em abril)
sep(ws2,rn, "  ▶ PRIORIDADE 1 — Abrir ABRIL/2026 (meta: +210 lig água/mês)", pf(AZ_MED))
rn+=1
for g in [gr for gr in GRUPOS_PROX if gr["prioridade"]==1]:
    gantt_row(ws2,rn, f"{g['cod']} — {g['desc']}",
              g["data_ini"], date(2026,12,31),
              g["equipes"], g["dific"],
              f"+{g['prod_lig_a_mes']} lig/mês",
              pf(AZ_CLR), pf(AM_CLR), bg_row=pf(AZ_CLR))
    rn+=1

rn+=1
sep(ws2,rn, "  ▶ PRIORIDADE 2 — Mobilizar fim de ABRIL, iniciar MAIO/2026 (+200 lig/mês)", pf(LARANJ))
rn+=1
for g in [gr for gr in GRUPOS_PROX if gr["prioridade"]==2]:
    gantt_row(ws2,rn, f"{g['cod']} — {g['desc']}",
              g["data_ini"], date(2026,12,31),
              g["equipes"], g["dific"],
              f"+{g['prod_lig_a_mes']} lig/mês",
              pf(AM_CLR), pf(VM_CLR), bg_row=pf(AM_CLR))
    rn+=1

rn+=1
sep(ws2,rn, "  ▶ PRIORIDADE 3 — MAIO/JUNHO/2026 (+243 lig/mês)", pf(ROXO))
rn+=1
for g in [gr for gr in GRUPOS_PROX if gr["prioridade"]==3]:
    gantt_row(ws2,rn, f"{g['cod']} — {g['desc']}",
              g["data_ini"], date(2026,12,31),
              g["equipes"], g["dific"],
              f"+{g['prod_lig_a_mes']} lig/mês",
              pf(RX_CLR), pf(CINZA), bg_row=pf(RX_CLR))
    rn+=1

rn+=1
sep(ws2,rn, "  ▶ PRIORIDADE 4-5 — JUN/JUL/2026 em diante", pf(CZ_MED))
rn+=1
for g in [gr for gr in GRUPOS_PROX if gr["prioridade"]>=4]:
    gantt_row(ws2,rn, f"{g['cod']} — {g['desc']}",
              g["data_ini"], date(2026,12,31),
              g["equipes"], g["dific"],
              f"+{g['prod_lig_a_mes']} lig/mês" if g["prod_lig_a_mes"]>0 else "a definir",
              pf(CINZA), pf("EEEEEE"), bg_row=pf(CINZA))
    rn+=1

print("  02_CRONOGRAMA_FISICO ok")

# ══════════════════════════════════════════════════════════════════════════════
# ABA 3 — CRONOGRAMA FINANCEIRO MENSAL
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("03_CRONO_FINANCEIRO")
rn  = titulo_aba(ws3, "CRONOGRAMA FÍSICO-FINANCEIRO MENSAL",
                 "Previsão de medição mensal | Curva S | Contrato R$ 241.235.263,31",
                 NCOLS)

hdrs3 = ["Mês","Lig. Água\nPrev.","Lig. Esg.\nPrev.",
         "Medição\nMensal (R$)","Medição\nAcum. (R$)",
         "% Físico\nAcum.","% Financeiro\nAcum.",
         "Equipes\nAtivas","Meta\nLig.Água","Status"]
larg3 = [10,12,12,20,22,12,14,12,12,16]
for ci,(h,w) in enumerate(zip(hdrs3,larg3),1):
    cell_write(ws3,rn,ci,h, fill_=pf(AZUL), font_=fn(True,"FFFFFF",10),
               aln_=al("center",wrap=True), border_=THIN)
    col_w(ws3,ci,w)
row_h(ws3,rn,36); rn+=1

acum_lig_a3 = 0; acum_med3 = 0.0
for mi,p in enumerate(prod_mensal):
    acum_lig_a3 += p["lig_agua"]
    acum_med3   += p["medicao"]
    pct_fis = min(99.9, acum_lig_a3 / 9210 * 100)  # 9210 = total recortados
    pct_fin = min(99.9, acum_med3 / CONTRATO_TOTAL * 100)

    ok = p["lig_agua"] >= 1000
    bg = pf(VD_CLR) if ok else (pf(AM_CLR) if p["lig_agua"]>=600 else pf(VM_CLR))

    row_v = [
        p["nome"], p["lig_agua"], p["lig_esg"],
        round(p["medicao"],2), round(acum_med3,2),
        round(pct_fis,1), round(pct_fin,2),
        p["equipes"], 1000,
        "✔ META" if ok else ("⚠ RAMPA" if p["lig_agua"]>=600 else "✘ ABAIXO")
    ]
    for ci,v in enumerate(row_v,1):
        cell_write(ws3,rn,ci,v,
                   fill_=(bg if ci in (2,10) else pf(CINZA) if mi%2==0 else None),
                   font_=fn(bold=(ci in (2,5,6,10)),size=10),
                   aln_=al("center"), border_=THIN,
                   fmt=('R$ #,##0.00' if ci in (4,5) else None))
    row_h(ws3,rn,18); rn+=1

# Totais
rn+=1
total_lig_a = sum(p["lig_agua"] for p in prod_mensal)
total_med   = sum(p["medicao"]  for p in prod_mensal)
merge(ws3,rn,1,rn,3,"TOTAL PROJETADO 2026",
      fill_=pf(AZUL), font_=fn(True,"FFFFFF",11), aln_=al("center"))
cell_write(ws3,rn,4, round(total_med,2),
           fill_=pf(AZUL), font_=fn(True,"FFFFFF",11),
           aln_=al("center"), border_=THIN, fmt='R$ #,##0.00')
row_h(ws3,rn,22)

print("  03_CRONO_FINANCEIRO ok")

# ══════════════════════════════════════════════════════════════════════════════
# ABA 4 — NUCLEOS ATIVOS: ANDAMENTO E PREVISAO
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("04_NUCLEOS_ATIVOS")
rn  = titulo_aba(ws4, "NÚCLEOS EM EXECUÇÃO — ANDAMENTO E PREVISÃO DE CONCLUSÃO",
                 "Dados de produção real | Execução_Geral.xlsx | Mar/2026",
                 NCOLS)

hdrs4 = ["Núcleo","DIFIC","Lig.Água\nTotal","Lig.Água\nExec.","% Fís.\nAcum.",
         "Lig.Água\n/Mês","Meses\nRestantes","Prev.\nConclusão",
         "Equipes","Desvio\nProd.","Ação Recomendada"]
larg4 = [28,6,12,12,10,12,12,14,9,12,36]
for ci,(h,w) in enumerate(zip(hdrs4,larg4),1):
    cell_write(ws4,rn,ci,h, fill_=pf(AZUL), font_=fn(True,"FFFFFF",10),
               aln_=al("center",wrap=True), border_=THIN)
    col_w(ws4,ci,w)
row_h(ws4,rn,36); rn+=1

ACOES_NUCLEO = {
    "Morro do Tetéu":          "Reforçar equipes DIFIC 4 | foco em ligações antes de rede",
    "Pantanal Baixo":          "Aumentar frente para escavação — terreno difícil",
    "Vila dos Criadores":      "Manter ritmo — melhor produtividade atual",
    "CT São Manoel":           "Acelerar finalizações — conclusão prevista Mai/26",
    "CT João Carlos da Silva": "Concluir em Abr/26 — liberar equipes para R3",
    "Vila Israel":             "Concluir em Mai/26 — liberar equipes para R7/R4",
}
for ni,n in enumerate(NUCLEOS_ATIVOS):
    lig_rest = n["lig_agua"] - n["lig_exec"]
    meses_r  = max(0, round(lig_rest / max(1,n["prod_lig_a_mes"]), 1))
    prod_contr = (n["dific"] <= 3) and 60 or 40
    desvio = round((n["prod_lig_a_mes"] - prod_contr)/prod_contr*100, 0)
    bg_row = pf(VD_CLR) if n["pct_fisico"] >= 0.5 else (pf(AM_CLR) if n["pct_fisico"]>=0.3 else pf(VM_CLR))
    row_v4 = [
        n["nome"], n["dific"],
        n["lig_agua"], n["lig_exec"],
        f"{n['pct_fisico']*100:.0f}%",
        n["prod_lig_a_mes"], meses_r,
        n["data_fim"].strftime("%d/%m/%Y"),
        n["equipes"],
        f"{desvio:+.0f}%",
        ACOES_NUCLEO.get(n["nome"],"—")
    ]
    for ci,v in enumerate(row_v4,1):
        bg = (bg_row if ci in (5,8) else
              pf(VD_CLR) if (ci==10 and desvio>=-10) else
              pf(VM_CLR) if (ci==10 and desvio<-10) else
              pf(CINZA) if ni%2==0 else None)
        cell_write(ws4,rn,ci,v, fill_=bg, font_=fn(size=10),
                   aln_=al("left" if ci in (1,11) else "center"), border_=THIN)
    row_h(ws4,rn,20); rn+=1

print("  04_NUCLEOS_ATIVOS ok")

# ══════════════════════════════════════════════════════════════════════════════
# ABA 5 — GRUPOS PROXIMOS: PLANO DE ABERTURA
# ══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("05_GRUPOS_ABERTURA")
rn  = titulo_aba(ws5, "PLANO DE ABERTURA DOS GRUPOS DE PROXIMIDADE",
                 "Baseado em TRECHOS_NUCLEOS_RECORTADOS.xlsx | Priorizado por DIFIC + LIG + Localização",
                 NCOLS)

hdrs5 = ["Cód","Grupo de Proximidade","Prior.","N.Nuc","Trechos",
         "Lig.Água","Lig.Esg.","DIFIC\nMéd.","Custo c/BDI",
         "Equipes\nNec.","Semanas\nMob.","Início\nPlan.",
         "Prod.\nLig/Mês","Ação Imediata"]
larg5 = [5,32,7,7,9,10,10,7,18,10,10,12,12,36]
for ci,(h,w) in enumerate(zip(hdrs5,larg5),1):
    cell_write(ws5,rn,ci,h, fill_=pf(AZUL), font_=fn(True,"FFFFFF",9),
               aln_=al("center",wrap=True), border_=THIN)
    col_w(ws5,ci,w)
row_h(ws5,rn,36); rn+=1

PRIOR_FILL = {1:pf(VD_CLR), 2:pf(AM_CLR), 3:pf(RX_CLR), 4:pf(CINZA), 5:pf("EEEEEE")}
ACOES_GRUPO = {
    "R8": "Emitir OF para Noroeste/Vila Gilda — terreno acessível — iniciar 07/Abr",
    "R3": "Liberar Cruzeiro/Progresso II após João Carlos concluir — iniciar 07/Abr",
    "R7": "Mobilizar em 22/Abr → garantir início pleno em 01/Mai",
    "R4": "Iniciar com equipes liberadas de São Manoel/Israel (Mai/26)",
    "R2": "Abrir após consolidar R8+R3+R7 — Mai/Jun/26",
    "R5": "Maior grupo — aguardar frentes R2+R4 estabilizadas",
    "R1": "DIFIC 4-5 — planejar com equipes especializadas — Jun/26",
    "R6": "Seguir R1 — mesmo tipo de terreno",
    "ISO": "Nucleos dispersos — cronograma individualizado após mai/26",
}
for gi,g in enumerate(GRUPOS_PROX):
    bg = PRIOR_FILL.get(g["prioridade"], pf(CINZA))
    row_v5 = [
        g["cod"], g["desc"], f"P{g['prioridade']}",
        g["n_nuc"], g["n_tr"], g["lig_agua"], g["lig_esg"],
        g["dific"],
        round(g["custo_cbdi"],2),
        g["equipes"], g["semanas_mob"],
        g["data_ini"].strftime("%d/%m/%Y") if g["data_ini"] else "—",
        g["prod_lig_a_mes"] if g["prod_lig_a_mes"]>0 else "—",
        ACOES_GRUPO.get(g["cod"],"—")
    ]
    for ci,v in enumerate(row_v5,1):
        cell_write(ws5,rn,ci,v, fill_=(bg if ci<=3 else pf(CINZA) if gi%2==0 else None),
                   font_=fn(bold=(ci<=2),size=9),
                   aln_=al("left" if ci in (2,14) else "center"), border_=THIN,
                   fmt=('R$ #,##0.00' if ci==9 else None))
    row_h(ws5,rn,20); rn+=1

print("  05_GRUPOS_ABERTURA ok")

# ══════════════════════════════════════════════════════════════════════════════
# ABA 6 — EQUIPES E MOBILIZACAO
# ══════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("06_EQUIPES")
rn  = titulo_aba(ws6, "DIMENSIONAMENTO DE EQUIPES — CRONOGRAMA DE MOBILIZAÇÃO",
                 "Equipes atuais: 39 | Equipes necessárias para meta 1.000 lig/mês: ~90",
                 NCOLS)

hdrs6 = ["Mês","Equipes\nAtivos","Equipes\nNovos R8+R3","Equipes\nNovos R7+R4",
         "Equipes\nOutros","TOTAL\nEQUIPES","Meta\nEq. Min.","Lig Água\nPrev.",
         "Lig/Equipe\n(mês)","Observação"]
larg6 = [10,13,15,15,13,12,12,12,13,30]
for ci,(h,w) in enumerate(zip(hdrs6,larg6),1):
    cell_write(ws6,rn,ci,h, fill_=pf(AZUL), font_=fn(True,"FFFFFF",10),
               aln_=al("center",wrap=True), border_=THIN)
    col_w(ws6,ci,w)
row_h(ws6,rn,36); rn+=1

eq_r8r3 = [gr for gr in GRUPOS_PROX if gr["cod"] in ("R8","R3")]
eq_r7r4 = [gr for gr in GRUPOS_PROX if gr["cod"] in ("R7","R4")]
eq_outros= [gr for gr in GRUPOS_PROX if gr["cod"] not in ("R8","R3","R7","R4")]

OBS_EQ = {
    "Mar/26": "Manter 39 equipes | emitir OFs para R8+R3",
    "Abr/26": "Mobilizar +18 equipes (R8+R3) → 57 total | João Carlos conclui",
    "Mai/26": "Mobilizar +18 eq (R7+R4) → 75 total | São Manoel e Israel concluem",
    "Jun/26": "Mobilizar R2+R5 → ~90 equipes | Meta atingida e em expansão",
    "Jul/26": "Tetéu/Pantanal/Criadores concluem → redistribuir ~26 equipes",
    "Ago/26": "Consolidação | Iniciar R1/R6 → 95 equipes",
}
for mi,p in enumerate(prod_mensal):
    mes = p["mes"]
    eq_at  = sum(round(n["equipes"]*frac_mes_ativo_ate(n["data_fim"],mes))
                 for n in NUCLEOS_ATIVOS)
    eq_new_r8r3 = sum(round(g["equipes"]*frac_mes_ativa(g["data_ini"],mes))
                      for g in eq_r8r3)
    eq_new_r7r4 = sum(round(g["equipes"]*frac_mes_ativa(g["data_ini"],mes))
                      for g in eq_r7r4)
    eq_oth  = sum(round(g["equipes"]*frac_mes_ativa(g["data_ini"],mes))
                  for g in eq_outros)
    eq_tot  = eq_at + eq_new_r8r3 + eq_new_r7r4 + eq_oth
    meta_eq = max(39, round(p["lig_agua"] / 12))
    lig_per = round(p["lig_agua"]/max(1,eq_tot),1)
    obs     = OBS_EQ.get(p["nome"],"—")

    row_v6 = [p["nome"], eq_at, eq_new_r8r3, eq_new_r7r4, eq_oth,
              eq_tot, meta_eq, p["lig_agua"], lig_per, obs]
    bg6 = pf(VD_CLR) if eq_tot>=meta_eq else pf(AM_CLR)
    for ci,v in enumerate(row_v6,1):
        cell_write(ws6,rn,ci,v,
                   fill_=(bg6 if ci==6 else pf(CINZA) if mi%2==0 else None),
                   font_=fn(bold=(ci==6),size=10),
                   aln_=al("left" if ci==10 else "center"), border_=THIN)
    row_h(ws6,rn,18); rn+=1

print("  06_EQUIPES ok")

# ══════════════════════════════════════════════════════════════════════════════
# ABA 7 — CURVA S (dados para grafico)
# ══════════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("07_CURVA_S")
rn  = titulo_aba(ws7, "CURVA S — PROGRESSO FÍSICO E FINANCEIRO",
                 "% acumulado físico (lig. água) e financeiro (medição) | Mar–Dez/2026",
                 16)

hdrs7 = ["Mês","Lig Água\nAcum.","% Físico\nAcum.","Medição\nAcum.(R$)",
         "% Financeiro\nAcum.","Meta\nFísico(%)","Meta\nFinan.(%)"]
larg7 = [10,14,12,22,14,12,12]
for ci,(h,w) in enumerate(zip(hdrs7,larg7),1):
    cell_write(ws7,rn,ci,h, fill_=pf(AZUL), font_=fn(True,"FFFFFF",10),
               aln_=al("center",wrap=True), border_=THIN)
    col_w(ws7,ci,w)
row_h(ws7,rn,36)
data_start_row = rn+1; rn+=1

total_lig = sum(g["lig_agua"] for g in GRUPOS_PROX)
acum7_la  = 0; acum7_med = 0.0
for mi,p in enumerate(prod_mensal):
    acum7_la  += p["lig_agua"]
    acum7_med += p["medicao"]
    pft = min(100.0, acum7_la/total_lig*100)
    pfn = min(100.0, acum7_med/CONTRATO_TOTAL*100)
    # meta linear
    met_f = min(100.0, (mi+1)/10*100)
    met_n = min(100.0, (mi+1)/10*80)

    row_v7 = [p["nome"], acum7_la, round(pft,2), round(acum7_med,2),
              round(pfn,2), round(met_f,1), round(met_n,1)]
    for ci,v in enumerate(row_v7,1):
        bg7 = (pf(VD_CLR) if (ci==3 and pft>=met_f) else
               pf(VM_CLR) if (ci==3 and pft<met_f-10) else
               pf(CINZA) if mi%2==0 else None)
        cell_write(ws7,rn,ci,v, fill_=bg7,
                   font_=fn(bold=(ci in (3,5)),size=10),
                   aln_=al("center"), border_=THIN,
                   fmt=('R$ #,##0.00' if ci==4 else None))
    row_h(ws7,rn,18); rn+=1

# Grafico de linhas — Curva S
chart = LineChart()
chart.title = "Curva S — Progresso Físico e Financeiro"
chart.style = 10
chart.y_axis.title = "% Acumulado"
chart.x_axis.title = "Mês"
chart.height = 12; chart.width  = 22

n_rows = len(prod_mensal)
ref_fis = Reference(ws7, min_col=3, min_row=data_start_row,
                    max_row=data_start_row+n_rows-1)
ref_fin = Reference(ws7, min_col=5, min_row=data_start_row,
                    max_row=data_start_row+n_rows-1)
ref_met = Reference(ws7, min_col=6, min_row=data_start_row,
                    max_row=data_start_row+n_rows-1)
from openpyxl.chart.series import SeriesLabel
chart.add_data(ref_fis, titles_from_data=False)
chart.add_data(ref_fin, titles_from_data=False)
chart.add_data(ref_met, titles_from_data=False)
chart.series[0].title = SeriesLabel(v="% Fisico Lig Agua")
chart.series[1].title = SeriesLabel(v="% Financeiro Medicao")
chart.series[2].title = SeriesLabel(v="Meta Linear")
ws7.add_chart(chart, f"A{rn+2}")
print("  07_CURVA_S ok")

# ══════════════════════════════════════════════════════════════════════════════
# ABA 8 — SEMANA A SEMANA (ABR/MAI detalhado)
# ══════════════════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("08_SEMANAS_ABR_MAI")
rn  = titulo_aba(ws8, "CRONOGRAMA DETALHADO — SEMANA A SEMANA (ABRIL/MAIO 2026)",
                 "Ações operacionais para atingir 1.000 lig. água em Maio", NCOLS)

hdrs8 = ["Semana","Período","Núcleos\nAtivos","Grupos em\nMobilização",
         "Grupos em\nExecução","Lig.Água\nPrev.Sem.","Acum.\nMês","Ações Prioritárias"]
larg8 = [8,16,26,24,22,13,13,48]
for ci,(h,w) in enumerate(zip(hdrs8,larg8),1):
    cell_write(ws8,rn,ci,h, fill_=pf(AZUL), font_=fn(True,"FFFFFF",10),
               aln_=al("center",wrap=True), border_=THIN)
    col_w(ws8,ci,w)
row_h(ws8,rn,36); rn+=1

SEMANAS_DET = [
    ("Sem 01","31/Mar–05/Abr",
     "Tetéu+Pantanal+Criadores+S.Manoel+J.Carlos+Israel",
     "R8 Noroeste/Gilda | R3 Cruzeiro/Prog II","—",58,58,
     "Emitir OF R8+R3 | Contratar 18 equipes | Comprar materiais R8+R3 | ART/RRT"),
    ("Sem 02","06–12/Abr",
     "6 nucleos ativos",
     "R8: postos montados | R3: postos montados","R8+R3 início",72,130,
     "Iniciar escavações R8 (terreno DIFIC 3) | Ramais água R3 | Monitorar metas"),
    ("Sem 03","13–19/Abr",
     "6 nucleos + R8 + R3",
     "R7 Pirat/Manoel | R4 Cruzeiro/Penha","R8+R3 pleno",88,218,
     "Mobilizar R7+R4 | Aceleração R8+R3 | J.Carlos: concluir últimas ligações"),
    ("Sem 04","20–26/Abr",
     "5+R8+R3 (J.Carlos conclui)",
     "R7+R4 postos","R8+R3+início R7",92,310,
     "João Carlos conclui e libera 4 equipes → realocar para R3/R7 | Medir R8+R3"),
    ("Sem 05","27/Abr–03/Mai",
     "5+R8+R3+R7+R4",
     "—","R8+R3+R7+R4",100,410,
     "Plena operação 4 frentes novas | São Manoel e Israel chegam a 90% | Revisão DDS"),
    ("Sem 06","04–10/Mai",
     "4+R8+R3+R7+R4",
     "R2 Cruzeiro/Prog I","R8+R3+R7+R4 pleno",210,620,
     "S.Manoel e Israel concluem → equipes para R7/R4 | Iniciar mobilização R2"),
    ("Sem 07","11–17/Mai",
     "3+R8+R3+R7+R4",
     "R2 postos","Todos+início R2",230,850,
     "Tetéu+Pantanal+Criadores ainda em exec | Meta mensal ~80% atingida | Revisão"),
    ("Sem 08","18–24/Mai",
     "3+R8+R3+R7+R4+R2",
     "—","Todos",260,1110,
     "META 1.000 LIG ÁGUA/MÊS ATINGIDA ★ | Confirmar medição SABESP | Planejar Jun"),
]
for si,(sem,per,at,mob,exec_,lig_s,acum_m,acoes) in enumerate(SEMANAS_DET):
    bg = (pf(VD_CLR) if acum_m>=1000 else
          pf(AM_CLR) if acum_m>=600 else
          pf(VM_CLR) if acum_m<300 else pf(CINZA))
    row_v8 = [sem,per,at,mob,exec_,lig_s,acum_m,acoes]
    for ci,v in enumerate(row_v8,1):
        cell_write(ws8,rn,ci,v,
                   fill_=(bg if ci==7 else pf(CINZA) if si%2==0 else None),
                   font_=fn(bold=(ci==7 and acum_m>=1000),size=10),
                   aln_=al("left" if ci in (3,4,5,8) else "center",
                            wrap=(ci in (3,4,5,8))),
                   border_=THIN)
    row_h(ws8,rn,35 if len(acoes)>60 else 22); rn+=1

print("  08_SEMANAS_ABR_MAI ok")

# ══════════════════════════════════════════════════════════════════════════════
# ABA 9 — CONDICOES CRITICAS E RISCOS
# ══════════════════════════════════════════════════════════════════════════════
ws9 = wb.create_sheet("09_RISCOS_ACOES")
rn  = titulo_aba(ws9, "CONDICIONANTES CRÍTICOS E PLANO DE RISCOS",
                 "Fatores para atingir a meta | Riscos identificados | Mitigações", NCOLS)

RISCOS = [
    ("🔴 CRITICO", "Projetos R8+R3 não aprovados pela SABESP antes de 05/Abr",
     "Atrasa abertura 2-3 semanas → meta só em Jun/26",
     "Protocolar projetos até 27/Mar | Follow-up diário na SABESP"),
    ("🔴 CRITICO", "Falta de materiais (tubos, conexões, ramais)",
     "Paralisa frentes novas por até 4 semanas",
     "Emitir requisição de compra hoje | Estoque mínimo 30 dias"),
    ("🟠 ALTO",    "Produtividade abaixo do planejado (atual -50% do contratado)",
     "Meta se desloca para Jun/26",
     "DDS diário | Meta por equipe/dia | Bônus por ligação"),
    ("🟠 ALTO",    "Liberação de ART/RRT atrasada para novos engenheiros",
     "Frentes R7+R4 não iniciam em 22/Abr",
     "Contratar responsáveis técnicos com 30 dias de antecedência"),
    ("🟡 MEDIO",   "Chuvas intensas em abril (histórico Santos)",
     "Perda de 5-7 dias úteis em escavações",
     "Acelerar serviços em locais cobertos | Buffer de 2 semanas no cronograma"),
    ("🟡 MEDIO",   "Conflito de acesso — licenças de vias em Santos",
     "Atrasa início de frentes urbanas R3+R2",
     "Protocolar pedidos de autorização junto à PMTS até 30/Mar"),
    ("🟢 BAIXO",   "Resistência de moradores no início das obras",
     "Atraso pontual de 1-2 dias por trecho",
     "Reunião comunitária prévia por bairro | Suporte assistência social SABESP"),
    ("🟢 BAIXO",   "Medição SABESP com prazo longo (>30 dias)",
     "Impacto no fluxo de caixa FCN",
     "Garantir Book de Medição completo antes do fechamento mensal"),
]
hdrs9 = ["Criticidade","Condicionante / Risco","Impacto","Mitigação / Ação"]
larg9 = [12,38,32,44]
for ci,(h,w) in enumerate(zip(hdrs9,larg9),1):
    cell_write(ws9,rn,ci,h, fill_=pf(AZUL), font_=fn(True,"FFFFFF",11),
               aln_=al("center"), border_=THIN)
    col_w(ws9,ci,w)
row_h(ws9,rn,22); rn+=1

CRIT_FILL = {"🔴 CRITICO":pf(VM_CLR),"🟠 ALTO":pf(AM_CLR),
             "🟡 MEDIO":pf("FFFACD"),"🟢 BAIXO":pf(VD_CLR)}
for ri,(crit,cond,imp,mit) in enumerate(RISCOS):
    bg = CRIT_FILL.get(crit, pf(CINZA))
    for ci,v in enumerate([crit,cond,imp,mit],1):
        cell_write(ws9,rn,ci,v, fill_=bg,
                   font_=fn(bold=(ci==1),size=10),
                   aln_=al("center" if ci==1 else "left",wrap=True), border_=THIN)
    row_h(ws9,rn,40); rn+=1

print("  09_RISCOS_ACOES ok")

# ─────────────────────────────────────────────────────────────────────────────
wb.save(OUT_FILE)
print(f"\nSalvo: {OUT_FILE}")
print(f"Abas ({len(wb.worksheets)}): {[ws.title for ws in wb.worksheets]}")
print(f"\nProjecao de ligacoes agua/mes:")
for p in prod_mensal:
    marca = "★ META" if p["lig_agua"]>=1000 else ("⚠" if p["lig_agua"]>=700 else "  ")
    print(f"  {p['nome']}: {p['lig_agua']:5d} lig agua/mes  {marca}")
print(f"\nMedicao projetada 2026: R$ {sum(p['medicao'] for p in prod_mensal):,.2f}")
