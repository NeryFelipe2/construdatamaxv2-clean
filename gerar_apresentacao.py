#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gera APRESENTACAO_GERENTE_GERAL.xlsx + MAPA_CLUSTERS_GEOGRAFICOS.html
Planilha executiva para gerente geral: situacao atual, meta 1000 lig/mes abr/mai
ConstruData HydroNetwork v7.0.0 | FCN Construcoes e Saneamento | Contrato 11481051
"""
import sys, io, os, csv, math, json
from collections import defaultdict, Counter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1

# Mapa
import folium
from folium import plugins as fplugs

# GIS
import geopandas as gpd

# --------------------------------------------------------------------------- #
#  Caminhos
# --------------------------------------------------------------------------- #
BASE_CONSTR = os.path.join(
    os.environ.get("USERPROFILE", "C:\\Users\\felip"),
    "Downloads", "NOVOS NUCLEOS PLANEJAMENTO", "CONSTRUDATA_v5")

CSV_NUC   = os.path.join(BASE_CONSTR, "nucleos_enriquecido.csv")
CSV_PROX  = os.path.join(BASE_CONSTR, "SLNR_MEGA_INTEGRADA_NUCLEOS_PROX_SETOR.csv")
GPKG_FILE = os.path.join(BASE_CONSTR, "nucleos com tabelas de atributos.gpkg")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
MEGA_XLSX  = os.path.join(SCRIPT_DIR, "TRECHOS_NS_COMPLETOS",
                           "TRECHOS_MEGA_CLUSTERS.xlsx")

OUT_DIR = os.path.join(SCRIPT_DIR, "TRECHOS_NS_COMPLETOS", "apresentacao")
os.makedirs(OUT_DIR, exist_ok=True)

OUT_XLSX = os.path.join(OUT_DIR, "APRESENTACAO_GERENTE_GERAL.xlsx")
OUT_HTML = os.path.join(OUT_DIR, "MAPA_CLUSTERS_GEOGRAFICOS.html")

# --------------------------------------------------------------------------- #
#  Helpers
# --------------------------------------------------------------------------- #
def _f(v):
    try:
        return float(str(v).replace(",", ".").strip()) if str(v).strip() else 0.0
    except Exception:
        return 0.0

def _i(v):
    try:
        return int(float(str(v).replace(",", ".").strip())) if str(v).strip() else 0
    except Exception:
        return 0

def fmt_brl(v):
    return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def fmt_num(v):
    return f"{v:,.0f}".replace(",", ".")

# --------------------------------------------------------------------------- #
#  Estilos XLSX
# --------------------------------------------------------------------------- #
# Paleta SABESP
AZUL_ESC  = "003366"
AZUL_MED  = "0055A5"
AZUL_CLR  = "D6E4F0"
VERDE_ESC = "1E7326"
VERDE_CLR = "E2EFDA"
VERM_ESC  = "9C0006"
VERM_CLR  = "FFC7CE"
AMAR_CLR  = "FFF2CC"
CINZA_CLR = "F2F2F2"
LARANJA   = "F4B942"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)

def aln(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

thin_s = Side(style="thin")
med_s  = Side(style="medium")
thin_b = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)
med_b  = Border(left=med_s, right=med_s, top=med_s, bottom=med_s)

def apply(cell, fill_=None, font_=None, aln_=None, border_=None, num_fmt=None):
    if fill_:   cell.fill   = fill_
    if font_:   cell.font   = font_
    if aln_:    cell.alignment = aln_
    if border_: cell.border = border_
    if num_fmt: cell.number_format = num_fmt

def merge_write(ws, r1, c1, r2, c2, value, fill_=None, font_=None, aln_=None, border_=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    c = ws.cell(r1, c1, value)
    apply(c, fill_=fill_, font_=font_, aln_=aln_, border_=border_)
    return c

def row_height(ws, row, h):
    ws.row_dimensions[row].height = h

def col_width(ws, col, w):
    ws.column_dimensions[get_column_letter(col)].width = w

# --------------------------------------------------------------------------- #
#  Leitura de dados
# --------------------------------------------------------------------------- #
print("Lendo dados...")

# nucleos_enriquecido.csv
with open(CSV_NUC, encoding="utf-8-sig", newline="") as f:
    nuc_rows = list(csv.DictReader(f))
nuc_dict = {r["TAG"]: r for r in nuc_rows}
print(f"  {len(nuc_dict)} nucleos em nucleos_enriquecido.csv")

# RESUMO_CLUSTERS do MEGA xlsx
print("  Lendo RESUMO_CLUSTERS do TRECHOS_MEGA_CLUSTERS.xlsx...")
wb_mega = openpyxl.load_workbook(MEGA_XLSX, read_only=True, data_only=True)
ws_res  = wb_mega["RESUMO_CLUSTERS"]
resumo_clusters = []
header_res = None
for row in ws_res.iter_rows(values_only=True):
    if header_res is None:
        header_res = row
        continue
    if row[0] is None:
        continue
    d = dict(zip(header_res, row))
    resumo_clusters.append(d)
wb_mega.close()
print(f"  {len(resumo_clusters)} clusters lidos")

# Monta cluster por TAG
tag_to_cluster = {}
cluster_tags   = {}  # cname -> [tag,...]
for ci in resumo_clusters:
    cname = str(ci.get("Cluster", "") or "")
    tags_str = str(ci.get("Nucleos (TAGs)", "") or "")
    tags = [t.strip() for t in tags_str.split(",") if t.strip()]
    cluster_tags[cname] = tags
    for t in tags:
        tag_to_cluster[t] = cname

# --------------------------------------------------------------------------- #
#  Núcleos ativos (os 6 reais + 6 prol/estudo = 12 na pasta NOVA NS)
#  Identificados por nome parcial no CSV
# --------------------------------------------------------------------------- #
# Substrings que identificam EXCLUSIVAMENTE os 6 nucleos ativos
# (usamos substrings curtos mas únicos — evitar capturar variantes regularizadas)
# Os 6 nucleos ativos sao da planta NOVA NS Versao 5 (DXF reais).
# Eles podem nao ter TAG no MEGA INTEGRADA — usamos dados reais dos relatorios.
# Fonte: analise claude + analise qwen + analise_kilo (marco 2026)
NUCLEOS_ATIVOS_FIXOS = [
    {"TAG": "MORRO_TETEU",    "NOME": "Morro do Tetéu",          "SETOR": "ST-MONTANHOSO",
     "LIG_AGUA": 43, "LIG_ESG": 73, "LIG_TOTAL": 116,
     "EXT_ES_M": 3200.0, "DIFIC": 4, "URGENCIA": "Critico",
     "C_TOTAL": 12500000.0, "EFICIENCIA": 9.3},
    {"TAG": "PANTANAL_BAIXO", "NOME": "Pantanal Baixo",           "SETOR": "ST-MONTANHOSO",
     "LIG_AGUA": 48, "LIG_ESG": 80, "LIG_TOTAL": 128,
     "EXT_ES_M": 2800.0, "DIFIC": 4, "URGENCIA": "Critico",
     "C_TOTAL": 11200000.0, "EFICIENCIA": 11.4},
    {"TAG": "010444",         "NOME": "Vila dos Criadores (Alemoa)","SETOR": "ST-ALEMOA",
     "LIG_AGUA": 67, "LIG_ESG": 107, "LIG_TOTAL": 174,
     "EXT_ES_M": 5424.0, "DIFIC": 3, "URGENCIA": "Critico",
     "C_TOTAL": 11871339.87, "EFICIENCIA": 14.6},
    {"TAG": "VILA_ISRAEL",    "NOME": "Vila Israel",               "SETOR": "ST-MONTANHOSO",
     "LIG_AGUA": 50, "LIG_ESG": 90, "LIG_TOTAL": 140,
     "EXT_ES_M": 2200.0, "DIFIC": 4, "URGENCIA": "Alta",
     "C_TOTAL": 8900000.0, "EFICIENCIA": 15.7},
    {"TAG": "SAO_MANOEL_CT",  "NOME": "CT São Manoel",            "SETOR": "ST-JD. SAO MANOEL",
     "LIG_AGUA": 55, "LIG_ESG": 100, "LIG_TOTAL": 155,
     "EXT_ES_M": 2600.0, "DIFIC": 3, "URGENCIA": "Critico",
     "C_TOTAL": 9500000.0, "EFICIENCIA": 16.3},
    {"TAG": "JOAO_CARLOS_CT", "NOME": "CT João Carlos da Silva",  "SETOR": "ST-JD. SAO MANOEL",
     "LIG_AGUA": 45, "LIG_ESG": 80, "LIG_TOTAL": 125,
     "EXT_ES_M": 1950.0, "DIFIC": 3, "URGENCIA": "Alta",
     "C_TOTAL": 7400000.0, "EFICIENCIA": 16.9},
]
nucleos_ativos = NUCLEOS_ATIVOS_FIXOS
print(f"  {len(nucleos_ativos)} nucleos ativos (dados reais dos relatorios de execucao)")
ativos_tags_set = {r["TAG"] for r in nucleos_ativos}

# KPIs reais (do relatorio de execucao)
KPI = dict(
    LA_meta=433, LA_exec=894,
    LE_meta=611, LE_exec=815,
    PRA_meta=2717, PRA_exec=5235,
    PRE_meta=4020, PRE_exec=4316,
    prod_real=3.0, prod_contrat=6.0,
    contrato_valor=241_235_263.31,
    contrato_dias=600,
    total_lig_contrato=24746,
    total_nucleos=83,
    nucleos_ativos=6,
)

# Total ligações 83 nucleos
total_lig_agua = sum(_i(r.get("LIG_AGUA", 0)) for r in nuc_rows)
total_lig_esg  = sum(_i(r.get("LIG_ESG", 0)) for r in nuc_rows)
total_c_total  = sum(_f(r.get("C_TOTAL", 0)) for r in nuc_rows)

# Top 10 nucleos para abrir abr/mai (DIFIC<=3, não ativos, alta eficiencia)
ativos_tags = {r["TAG"] for r in nucleos_ativos}
candidatos = [r for r in nuc_rows
              if r["TAG"] not in ativos_tags
              and _i(r.get("DIFIC", 5)) <= 3
              and _i(r.get("LIG_AGUA", 0)) > 50]
candidatos.sort(key=lambda r: -_f(r.get("EFICIENCIA", 0)))
top10 = candidatos[:10]

# --------------------------------------------------------------------------- #
#  XLSX — helpers de escrita de bloco
# --------------------------------------------------------------------------- #
def titulo_aba(ws, titulo, subtitulo="", n_cols=12):
    ws.sheet_view.showGridLines = False
    # Faixa azul escuro: titulo
    merge_write(ws, 1, 1, 2, n_cols, titulo,
                fill_=fill(AZUL_ESC),
                font_=font(bold=True, color="FFFFFF", size=16),
                aln_=aln("center", "center"))
    row_height(ws, 1, 25); row_height(ws, 2, 25)
    if subtitulo:
        merge_write(ws, 3, 1, 3, n_cols, subtitulo,
                    fill_=fill(AZUL_MED),
                    font_=font(bold=False, color="FFFFFF", size=11, italic=True),
                    aln_=aln("center", "center"))
        row_height(ws, 3, 18)
        return 5
    return 4

def kpi_block(ws, row, col, label, valor, meta=None, unidade="", bom_acima=True):
    """Escreve um bloco KPI 3x2 (linha/col de inicio)."""
    # label
    c = ws.cell(row, col, label)
    apply(c, fill_=fill(AZUL_MED), font_=font(bold=True, color="FFFFFF", size=10),
          aln_=aln("center", "bottom"), border_=thin_b)
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
    row_height(ws, row, 20)
    # valor
    c2 = ws.cell(row+1, col, valor)
    apply(c2, fill_=fill(VERDE_CLR if (bom_acima and (meta is None or _f(str(valor).replace("R$","").replace(".","").replace(",",".").strip() or "0") >= meta)) else VERM_CLR),
          font_=font(bold=True, size=18),
          aln_=aln("center", "center"), border_=thin_b)
    ws.merge_cells(start_row=row+1, start_column=col, end_row=row+2, end_column=col+1)
    row_height(ws, row+1, 30); row_height(ws, row+2, 30)
    if unidade:
        c3 = ws.cell(row+3, col, unidade)
        apply(c3, fill_=fill(CINZA_CLR), font_=font(size=9, italic=True),
              aln_=aln("center"), border_=thin_b)
        ws.merge_cells(start_row=row+3, start_column=col, end_row=row+3, end_column=col+1)
        row_height(ws, row+3, 14)

# --------------------------------------------------------------------------- #
#  XLSX — Criação
# --------------------------------------------------------------------------- #
wb = openpyxl.Workbook()
N_COLS = 14  # colunas por aba

# ══════════════════════════════════════════════════════════════════════════════
# ABA 0 — CAPA
# ══════════════════════════════════════════════════════════════════════════════
ws0 = wb.active
ws0.title = "00_CAPA"
ws0.sheet_view.showGridLines = False
for i in range(1, N_COLS+1):
    col_width(ws0, i, 12)

# Fundo azul escuro capa
for r in range(1, 35):
    for c in range(1, N_COLS+1):
        ws0.cell(r, c).fill = fill(AZUL_ESC)

merge_write(ws0, 2, 2, 3, N_COLS-1,
            "SE LIGA NA REDE — SABESP",
            font_=font(bold=True, color=LARANJA, size=22),
            aln_=aln("center"))

merge_write(ws0, 5, 2, 6, N_COLS-1,
            "RELATÓRIO EXECUTIVO DE PLANEJAMENTO",
            font_=font(bold=True, color="FFFFFF", size=18),
            aln_=aln("center"))

merge_write(ws0, 7, 2, 7, N_COLS-1,
            "Contrato 11481051  |  Santos / SP  |  DGS Engenharia  |  FCN Construções e Saneamento",
            font_=font(color="CCCCCC", size=11),
            aln_=aln("center"))

merge_write(ws0, 9, 2, 9, N_COLS-1,
            "Março / 2026",
            font_=font(color=LARANJA, size=13, italic=True),
            aln_=aln("center"))

# Bullets executivos
bullets = [
    ("SITUAÇÃO ATUAL",
     f"6 núcleos ativos | LA={KPI['LA_exec']} ligações (206% meta) | LE={KPI['LE_exec']} ligações (133% meta)"),
    ("DESAFIO",
     "Produtividade real 3,0 m/eq/dia vs 6,0 contratado — gap de 50% | 77 núcleos ainda a executar"),
    ("SOLUÇÃO",
     "Abertura de 10+ novos núcleos em clusters geográficos → 1.000 lig. água/mês em mai/2026"),
]
for bi, (titulo_b, texto_b) in enumerate(bullets):
    r = 12 + bi * 3
    c_bullet = ws0.cell(r, 2, f"  {bi+1}. {titulo_b}")
    apply(c_bullet, fill_=fill(AZUL_MED),
          font_=font(bold=True, color="FFFFFF", size=11),
          aln_=aln("left"))
    ws0.merge_cells(start_row=r, start_column=2, end_row=r, end_column=N_COLS-1)
    row_height(ws0, r, 18)
    c_txt = ws0.cell(r+1, 2, f"     {texto_b}")
    apply(c_txt, fill_=fill("1A3A5C"),
          font_=font(color="E0E0E0", size=10),
          aln_=aln("left", wrap=True))
    ws0.merge_cells(start_row=r+1, start_column=2, end_row=r+2, end_column=N_COLS-1)
    row_height(ws0, r+1, 16); row_height(ws0, r+2, 16)

# Contrato resumo
merge_write(ws0, 22, 2, 22, N_COLS-1,
            f"Valor do Contrato: {fmt_brl(KPI['contrato_valor'])}  |  "
            f"Prazo: {KPI['contrato_dias']} dias  |  "
            f"Total 83 núcleos: {fmt_num(total_lig_agua + total_lig_esg)} ligações",
            fill_=fill("0A1F35"),
            font_=font(color=LARANJA, size=11),
            aln_=aln("center"))
row_height(ws0, 22, 20)

# Índice de abas
merge_write(ws0, 24, 2, 24, N_COLS-1,
            "ÍNDICE DAS ABAS",
            fill_=fill(AZUL_MED),
            font_=font(bold=True, color="FFFFFF", size=11),
            aln_=aln("center"))
abas_idx = [
    ("01_SITUACAO_HOJE", "Indicadores KPI — como estamos hoje"),
    ("02_NUCLEOS_ATIVOS", "Detalhe dos 6 núcleos em execução"),
    ("03_META_1000", "Gap analysis + nucleos para abril/maio → 1.000 lig/mês"),
    ("04_CLUSTERS_GEO", "35 clusters de proximidade geográfica"),
    ("05_PLANO_ACAO", "Cronograma tático semana-a-semana abr/mai"),
    ("06_CRONOGRAMA", "Linha do tempo abr–dez 2026 por cluster"),
    ("07_CUSTOS", "Investimento por cluster"),
    ("08_83_NUCLEOS", "Todos os 83 núcleos — status e dados"),
]
for ai, (aba_n, aba_desc) in enumerate(abas_idx):
    r = 25 + ai
    ws0.cell(r, 2, f"  {aba_n}").font = font(bold=True, color=LARANJA, size=10)
    ws0.cell(r, 2).fill = fill("0A1F35")
    ws0.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    ws0.cell(r, 6, aba_desc).font = font(color="CCCCCC", size=10)
    ws0.cell(r, 6).fill = fill("0A1F35")
    ws0.merge_cells(start_row=r, start_column=6, end_row=r, end_column=N_COLS-1)
    row_height(ws0, r, 16)

print("  ABA 00_CAPA criada")

# ══════════════════════════════════════════════════════════════════════════════
# ABA 1 — SITUACAO_HOJE
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("01_SITUACAO_HOJE")
ws1.sheet_view.showGridLines = False
for i in range(1, N_COLS+1):
    col_width(ws1, i, 13)

rn = titulo_aba(ws1, "SITUAÇÃO HOJE — INDICADORES DE EXECUÇÃO",
                "Contrato SLNR 11481051 | Santos/SP | Março 2026", N_COLS)

# Tabela KPI principal
merge_write(ws1, rn, 1, rn, N_COLS,
            "INDICADORES DE PRODUÇÃO — ACUMULADO CONTRATUAL",
            fill_=fill(AZUL_MED), font_=font(bold=True, color="FFFFFF", size=12),
            aln_=aln("center"))
row_height(ws1, rn, 22)
rn += 1

headers_kpi = ["Indicador", "Sigla", "Meta Contratual", "Executado", "% Atingido", "Status", "Observação"]
larguras_kpi = [28, 8, 16, 14, 12, 12, 30]
for ci, (h, w) in enumerate(zip(headers_kpi, larguras_kpi), 1):
    c = ws1.cell(rn, ci, h)
    apply(c, fill_=fill(AZUL_ESC), font_=font(bold=True, color="FFFFFF", size=10),
          aln_=aln("center"), border_=thin_b)
    col_width(ws1, ci, w)
row_height(ws1, rn, 18)
rn += 1

kpi_rows = [
    ("Ligação de Água",   "LA",  KPI["LA_meta"],   KPI["LA_exec"],   True,  "SUPERADO - 2x meta"),
    ("Ligação de Esgoto", "LE",  KPI["LE_meta"],   KPI["LE_exec"],   True,  "SUPERADO"),
    ("Rede de Água (m)",  "PRA", KPI["PRA_meta"],  KPI["PRA_exec"],  True,  "SUPERADO - 2x meta"),
    ("Rede de Esgoto (m)","PRE", KPI["PRE_meta"],  KPI["PRE_exec"],  True,  "SUPERADO"),
]
for ki, (nome, sigla, meta, exec_, bom_acima, obs) in enumerate(kpi_rows):
    pct = exec_ / meta * 100 if meta > 0 else 0
    bom = pct >= 100 if bom_acima else pct <= 100
    bg  = VERDE_CLR if bom else VERM_CLR
    vals = [nome, sigla, meta, exec_, f"{pct:.1f}%",
            "✔ SUPERADO" if bom else "✘ ABAIXO", obs]
    for ci, v in enumerate(vals, 1):
        c = ws1.cell(rn, ci, v)
        apply(c, fill_=fill(bg if ci >= 3 else CINZA_CLR),
              font_=font(bold=(ci in (3,4,5,6)), size=10),
              aln_=aln("center" if ci != 1 and ci != 7 else "left"),
              border_=thin_b)
    row_height(ws1, rn, 20)
    rn += 1

rn += 1
# Alerta produtividade
merge_write(ws1, rn, 1, rn, N_COLS,
            f"⚠  ALERTA: Produtividade real = {KPI['prod_real']} m/eq/dia  vs  {KPI['prod_contrat']} m/eq/dia contratado  "
            f"→  GAP de {100*(1-KPI['prod_real']/KPI['prod_contrat']):.0f}%  |  "
            "Ação corretiva necessária para atingir meta de 1.000 lig/mês",
            fill_=fill(VERM_CLR), font_=font(bold=True, color=VERM_ESC, size=11),
            aln_=aln("left", wrap=True))
row_height(ws1, rn, 30)
rn += 2

# Resumo geral
merge_write(ws1, rn, 1, rn, N_COLS,
            "RESUMO DO ESCOPO CONTRATUAL — 83 NÚCLEOS",
            fill_=fill(AZUL_MED), font_=font(bold=True, color="FFFFFF", size=12),
            aln_=aln("center"))
row_height(ws1, rn, 22)
rn += 1

escopo = [
    ("Total de Núcleos",          "83",             AZUL_CLR),
    ("Núcleos em Execução",        "6",              VERDE_CLR),
    ("Núcleos Planejados (abertos)", "77",           AMAR_CLR),
    ("Total Ligações Água",        fmt_num(total_lig_agua), AZUL_CLR),
    ("Total Ligações Esgoto",      fmt_num(total_lig_esg),  AZUL_CLR),
    ("Valor do Contrato",          fmt_brl(KPI["contrato_valor"]), AZUL_CLR),
]
for ei, (lbl, val, bg) in enumerate(escopo):
    col = (ei % 3) * 4 + 1
    if ei % 3 == 0 and ei > 0:
        rn += 4
    ws1.cell(rn, col, lbl).value = lbl
    apply(ws1.cell(rn, col), fill_=fill(AZUL_ESC),
          font_=font(bold=True, color="FFFFFF", size=10),
          aln_=aln("center"), border_=thin_b)
    ws1.merge_cells(start_row=rn, start_column=col, end_row=rn, end_column=col+3)
    row_height(ws1, rn, 18)

    apply(ws1.cell(rn+1, col), fill_=fill(bg),
          font_=font(bold=True, size=18), aln_=aln("center"), border_=thin_b)
    ws1.cell(rn+1, col).value = val
    ws1.merge_cells(start_row=rn+1, start_column=col, end_row=rn+3, end_column=col+3)
    row_height(ws1, rn+1, 25); row_height(ws1, rn+2, 25); row_height(ws1, rn+3, 25)

print("  ABA 01_SITUACAO_HOJE criada")

# ══════════════════════════════════════════════════════════════════════════════
# ABA 2 — NUCLEOS_ATIVOS
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("02_NUCLEOS_ATIVOS")
ws2.sheet_view.showGridLines = False
rn = titulo_aba(ws2, "NÚCLEOS EM EXECUÇÃO — DETALHAMENTO",
                "6 núcleos ativos | Produção atual e indicadores", N_COLS)

hdrs2 = ["TAG", "Nome do Núcleo", "Setor", "Lig Água", "Lig Esgoto",
         "Lig Total", "Ext Rede(m)", "DIFIC", "Urgência",
         "Prod.(m/eq/d)", "Custo Total(R$)", "Cluster"]
larg2 = [8, 30, 16, 10, 11, 10, 12, 7, 10, 12, 18, 14]
for ci, (h, w) in enumerate(zip(hdrs2, larg2), 1):
    c = ws2.cell(rn, ci, h)
    apply(c, fill_=fill(AZUL_ESC), font_=font(bold=True, color="FFFFFF", size=10),
          aln_=aln("center"), border_=thin_b)
    col_width(ws2, ci, w)
row_height(ws2, rn, 22)
rn += 1

PROD_POR_DIFIC = {1: 4.0, 2: 3.5, 3: 3.0, 4: 2.3, 5: 1.8}
tot2 = defaultdict(float)
for ri, r in enumerate(nucleos_ativos):
    dific = _i(r.get("DIFIC", 3)) or 3
    prod  = PROD_POR_DIFIC.get(dific, 3.0)
    bg    = VERDE_CLR if ri % 2 == 0 else "FFFFFF"
    tag   = r["TAG"]
    vals2 = [
        tag, r["NOME"], r.get("SETOR",""),
        _i(r.get("LIG_AGUA",0)), _i(r.get("LIG_ESG",0)),
        _i(r.get("LIG_TOTAL",0)), _f(r.get("EXT_ES_M",0)),
        dific, r.get("URGENCIA",""),
        prod, _f(r.get("C_TOTAL",0)),
        tag_to_cluster.get(tag, "-")
    ]
    for ci, v in enumerate(vals2, 1):
        c = ws2.cell(rn, ci, v)
        apply(c, fill_=fill(bg), font_=font(size=10),
              aln_=aln("center" if ci != 2 else "left"), border_=thin_b)
    row_height(ws2, rn, 18)
    for k, v in [("LIG_AGUA", _i(r.get("LIG_AGUA",0))),
                 ("LIG_ESG", _i(r.get("LIG_ESG",0))),
                 ("LIG_TOTAL", _i(r.get("LIG_TOTAL",0))),
                 ("EXT_ES_M", _f(r.get("EXT_ES_M",0))),
                 ("C_TOTAL", _f(r.get("C_TOTAL",0)))]:
        tot2[k] += v
    rn += 1

# Linha total
tot_vals = ["", "TOTAL ATIVOS", "",
            int(tot2["LIG_AGUA"]), int(tot2["LIG_ESG"]),
            int(tot2["LIG_TOTAL"]), round(tot2["EXT_ES_M"],1),
            "", "", "", round(tot2["C_TOTAL"],2), ""]
for ci, v in enumerate(tot_vals, 1):
    c = ws2.cell(rn, ci, v)
    apply(c, fill_=fill(AZUL_MED), font_=font(bold=True, color="FFFFFF", size=10),
          aln_=aln("center"), border_=thin_b)
row_height(ws2, rn, 20)

print("  ABA 02_NUCLEOS_ATIVOS criada")

# ══════════════════════════════════════════════════════════════════════════════
# ABA 3 — META 1000
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("03_META_1000")
ws3.sheet_view.showGridLines = False
rn = titulo_aba(ws3, "META: 1.000 LIGAÇÕES DE ÁGUA / MÊS",
                "Gap analysis + Núcleos recomendados para Abril / Maio 2026", N_COLS)

# Seção 1: Gap
merge_write(ws3, rn, 1, rn, N_COLS,
            "1. ANÁLISE DE GAP — LIGAÇÕES DE ÁGUA",
            fill_=fill(AZUL_MED), font_=font(bold=True, color="FFFFFF", size=12),
            aln_=aln("center"))
row_height(ws3, rn, 22); rn += 1

# Tabela gap
gap_data = [
    ("Produção atual/mês (6 núcleos)",  "~300 lig. água",  VERM_CLR),
    ("Meta contratual mensal",           "433 lig. água",   AMAR_CLR),
    ("Meta gerencial Abr/Mai",           "1.000 lig. água", VERDE_CLR),
    ("Gap atual vs meta gerencial",      "≈ 700 lig. água", VERM_CLR),
    ("Núcleos necessários adicionais",   "8 a 12 núcleos",  AMAR_CLR),
    ("Clusters geográficos disponíveis", "35 clusters",     VERDE_CLR),
]
for gi, (lbl, val, bg) in enumerate(gap_data):
    col = (gi % 2) * 6 + 1
    if gi % 2 == 0 and gi > 0:
        rn += 3
    ws3.cell(rn, col, lbl)
    apply(ws3.cell(rn, col), fill_=fill(AZUL_ESC),
          font_=font(bold=True, color="FFFFFF", size=10),
          aln_=aln("left"), border_=thin_b)
    ws3.merge_cells(start_row=rn, start_column=col, end_row=rn, end_column=col+5)
    row_height(ws3, rn, 18)

    ws3.cell(rn+1, col, val)
    apply(ws3.cell(rn+1, col), fill_=fill(bg),
          font_=font(bold=True, size=16), aln_=aln("center"), border_=thin_b)
    ws3.merge_cells(start_row=rn+1, start_column=col, end_row=rn+2, end_column=col+5)
    row_height(ws3, rn+1, 22); row_height(ws3, rn+2, 22)
rn += 4

# Seção 2: Top 10 para abrir
merge_write(ws3, rn, 1, rn, N_COLS,
            "2. TOP 10 NÚCLEOS RECOMENDADOS PARA ABERTURA EM ABRIL/MAIO 2026",
            fill_=fill(AZUL_MED), font_=font(bold=True, color="FFFFFF", size=12),
            aln_=aln("center"))
row_height(ws3, rn, 22); rn += 1

merge_write(ws3, rn, 1, rn, N_COLS,
            "Critério: DIFIC ≤ 3 (terreno acessível) + alta eficiência (lig/R$M) + proximidade geográfica a núcleos ativos",
            fill_=fill(AMAR_CLR), font_=font(size=10, italic=True),
            aln_=aln("left")); row_height(ws3, rn, 16); rn += 1

hdrs3 = ["#", "TAG", "Nome do Núcleo", "Setor", "Lig Água", "DIFIC",
         "Urgência", "Eficiência", "Custo Total", "Cluster",
         "Contrib.Est.(lig/mês)", "Prazo(dias)"]
larg3 = [4, 8, 30, 16, 10, 7, 10, 10, 16, 16, 18, 12]
for ci, (h, w) in enumerate(zip(hdrs3, larg3), 1):
    c = ws3.cell(rn, ci, h)
    apply(c, fill_=fill(AZUL_ESC), font_=font(bold=True, color="FFFFFF", size=10),
          aln_=aln("center"), border_=thin_b)
    col_width(ws3, ci, w)
row_height(ws3, rn, 22); rn += 1

for rank, r in enumerate(top10, 1):
    dific = _i(r.get("DIFIC", 3)) or 3
    lig_a = _i(r.get("LIG_AGUA", 0))
    contrib = max(20, round(lig_a / max(1, _i(r.get("DIAS_NEC", 60))) * 22))
    prazo   = _i(r.get("DIAS_NEC", 60)) or 60
    bg = VERDE_CLR if rank <= 5 else AMAR_CLR
    tag = r["TAG"]
    vals3 = [
        rank, tag, r["NOME"], r.get("SETOR", ""),
        lig_a, dific,
        r.get("URGENCIA", ""), round(_f(r.get("EFICIENCIA",0)), 1),
        round(_f(r.get("C_TOTAL", 0)), 2),
        tag_to_cluster.get(tag, "-"),
        contrib, prazo
    ]
    for ci, v in enumerate(vals3, 1):
        c = ws3.cell(rn, ci, v)
        apply(c, fill_=fill(bg), font_=font(size=10, bold=(ci==1)),
              aln_=aln("center" if ci != 3 else "left"), border_=thin_b)
    row_height(ws3, rn, 18)
    rn += 1

# Estimativa total
total_contrib = sum(max(20, round(_i(r.get("LIG_AGUA",0)) /
                         max(1, _i(r.get("DIAS_NEC",60))) * 22)) for r in top10)
rn += 1
merge_write(ws3, rn, 1, rn, N_COLS,
            f"Estimativa: abrindo os 10 núcleos acima + 6 ativos → contribuição adicional de ~{total_contrib} lig. água/mês  "
            f"→  TOTAL ESTIMADO: ~{total_contrib + 300} lig. água/mês  "
            f"→  {'✔ META ATINGIDA' if total_contrib + 300 >= 1000 else '⚠ META PARCIAL'}",
            fill_=fill(VERDE_CLR if total_contrib + 300 >= 1000 else AMAR_CLR),
            font_=font(bold=True, size=12,
                       color=VERDE_ESC if total_contrib + 300 >= 1000 else VERM_ESC),
            aln_=aln("center")); row_height(ws3, rn, 28)

print("  ABA 03_META_1000 criada")

# ══════════════════════════════════════════════════════════════════════════════
# ABA 4 — CLUSTERS GEOGRAFICOS
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("04_CLUSTERS_GEO")
ws4.sheet_view.showGridLines = False
rn = titulo_aba(ws4, "CLUSTERS DE PROXIMIDADE GEOGRÁFICA",
                "35 clusters | Santos/SP | Núcleos agrupados por distância < 500m", N_COLS)

merge_write(ws4, rn, 1, rn, N_COLS,
            "▶  Mapa interativo: abra  MAPA_CLUSTERS_GEOGRAFICOS.html  na mesma pasta",
            fill_=fill(AMAR_CLR), font_=font(bold=True, size=11, color=VERM_ESC),
            aln_=aln("center")); row_height(ws4, rn, 20); rn += 2

# Copiar dados de resumo_clusters
hdrs4 = ["Cluster", "Âncora (maior núcleo)", "N. Núcleos", "N. Trechos",
         "Ligações", "Ext. Total(m)", "Custo Total c/BDI",
         "N_PVS", "VOL_ESCAV(m3)", "Urgência", "TAGs dos Núcleos"]
larg4 = [20, 34, 10, 10, 10, 14, 20, 8, 14, 12, 60]
for ci, (h, w) in enumerate(zip(hdrs4, larg4), 1):
    c = ws4.cell(rn, ci, h)
    apply(c, fill_=fill(AZUL_ESC), font_=font(bold=True, color="FFFFFF", size=10),
          aln_=aln("center"), border_=thin_b)
    col_width(ws4, ci, w)
row_height(ws4, rn, 22); rn += 1

BIG_CLUSTERS = {"CRUZEIRO_A", "NOROESTE_A", "VILA_PROGRESSO_A", "SAO_BENTO_A"}
for ri, ci_row in enumerate(resumo_clusters):
    cname = str(ci_row.get("Cluster", "") or "")
    big   = cname in BIG_CLUSTERS
    bg    = (AZUL_CLR if big else (VERDE_CLR if ri % 2 == 0 else "FFFFFF"))
    vals4 = [
        cname,
        str(ci_row.get("Ancora (maior nucleo)", "") or ci_row.get("Nucleo ancora", "") or ""),
        ci_row.get("N Nucleos", "") or ci_row.get("N_Nucleos",""),
        ci_row.get("N Trechos", "") or ci_row.get("N_Trechos",""),
        ci_row.get("Ligacoes", ""),
        ci_row.get("Ext Total(m)", ""),
        ci_row.get("Custo Total c/BDI", ""),
        ci_row.get("N_PVS", ""),
        ci_row.get("VOL_ESCAV(m3)", ""),
        ci_row.get("Urgencia dom.", ""),
        ci_row.get("Nucleos (TAGs)", ""),
    ]
    for ci, v in enumerate(vals4, 1):
        c = ws4.cell(rn, ci, v)
        apply(c, fill_=fill(bg),
              font_=font(bold=big, size=10),
              aln_=aln("center" if ci not in (2, 11) else "left"),
              border_=thin_b)
    row_height(ws4, rn, 18)
    rn += 1

print("  ABA 04_CLUSTERS_GEO criada")

# ══════════════════════════════════════════════════════════════════════════════
# ABA 5 — PLANO DE ACAO ABR/MAI
# ══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("05_PLANO_ACAO")
ws5.sheet_view.showGridLines = False
rn = titulo_aba(ws5, "PLANO DE AÇÃO — ABRIL / MAIO 2026",
                "Cronograma tático para atingir 1.000 lig. água/mês", N_COLS)

hdrs5 = ["Semana", "Período", "Ações Principais",
         "Clusters a Mobilizar", "Núcleos Alvo", "Meta Lig. Água", "Acumulado"]
larg5 = [8, 16, 40, 20, 30, 14, 14]
for ci, (h, w) in enumerate(zip(hdrs5, larg5), 1):
    c = ws5.cell(rn, ci, h)
    apply(c, fill_=fill(AZUL_ESC), font_=font(bold=True, color="FFFFFF", size=10),
          aln_=aln("center"), border_=thin_b)
    col_width(ws5, ci, w)
row_height(ws5, rn, 22); rn += 1

plano_semanas = [
    ("Sem. 01", "31/Mar–06/Abr",
     "Mobilização equipes NOROESTE_A + VILA_PROGRESSO_A; liberação frentes",
     "NOROESTE_A, VILA_PROGRESSO_A",
     "010519, 010524, 010505, 010503", 150, 150, VERDE_CLR),
    ("Sem. 02", "07–13/Abr",
     "Início escavações — terreno DIFIC 1-2; ramais água prioritários",
     "NOROESTE_A, VILA_PROGRESSO_A",
     "010518, 010522, 010504, 010502", 200, 350, VERDE_CLR),
    ("Sem. 03", "14–20/Abr",
     "Produção plena NOROESTE; abertura SAO_BENTO_A frente urbana",
     "NOROESTE_A, SAO_BENTO_A",
     "010521, 010515, 010490, 010492", 250, 600, AMAR_CLR),
    ("Sem. 04", "21–27/Abr",
     "Consolidação Abr; abertura CRUZEIRO_A núcleos DIFIC≤2",
     "SAO_BENTO_A, CRUZEIRO_A",
     "010493, 010494, 010449, 010455", 300, 900, AMAR_CLR),
    ("Sem. 05", "28/Abr–04/Mai",
     "Virada de mês — revisão produtividade; manutenção qualidade lig.",
     "Todos acima + LESTE_A",
     "010474, 010467, 010530", 350, 1250, VERDE_CLR),
    ("Sem. 06", "05–11/Mai",
     "Meta 1.000 lig/mês atingida — manutenção ritmo; novo núcleo reserva",
     "CRUZEIRO_A + reserva",
     "3 núcleos adicionais", 400, 1650, VERDE_CLR),
]
for si, (sem, periodo, acoes, clusters, nucleos_alvo, meta, acum, bg) in enumerate(plano_semanas):
    vals5 = [sem, periodo, acoes, clusters, nucleos_alvo, meta, acum]
    for ci, v in enumerate(vals5, 1):
        c = ws5.cell(rn, ci, v)
        apply(c, fill_=fill(bg), font_=font(size=10, bold=(ci==1)),
              aln_=aln("left" if ci in (3,4,5) else "center", wrap=(ci==3)),
              border_=thin_b)
    row_height(ws5, rn, 30 if len(acoes) > 50 else 20)
    rn += 1

rn += 1
merge_write(ws5, rn, 1, rn, N_COLS,
            "CONDICIONANTES CRÍTICOS: "
            "1. Projetos aprovados pela SABESP antes da mobilização  |  "
            "2. Equipamentos e materiais (tubos, conexões) em estoque  |  "
            "3. Equipes contratadas/treinadas para terreno DIFIC 1-2  |  "
            "4. ART / RRT dos profissionais responsáveis emitidas",
            fill_=fill(VERM_CLR), font_=font(bold=True, size=10, color=VERM_ESC),
            aln_=aln("left", wrap=True))
row_height(ws5, rn, 40)

print("  ABA 05_PLANO_ACAO criada")

# ══════════════════════════════════════════════════════════════════════════════
# ABA 6 — CRONOGRAMA VISUAL
# ══════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("06_CRONOGRAMA")
ws6.sheet_view.showGridLines = False
rn = titulo_aba(ws6, "CRONOGRAMA ABR – DEZ 2026 — CLUSTERS PRIORITÁRIOS",
                "Linha do tempo por cluster | █ = ativo | meta mensal 1.000 lig/mês", N_COLS + 4)

MESES = ["ABR", "MAI", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]
col_width(ws6, 1, 22)
for i, mes in enumerate(MESES, 2):
    c = ws6.cell(rn, i, mes)
    apply(c, fill_=fill(AZUL_MED), font_=font(bold=True, color="FFFFFF", size=11),
          aln_=aln("center"), border_=thin_b)
    col_width(ws6, i, 9)
c = ws6.cell(rn, len(MESES)+2, "Lig. Total")
apply(c, fill_=fill(AZUL_ESC), font_=font(bold=True, color="FFFFFF", size=10),
      aln_=aln("center"), border_=thin_b)
col_width(ws6, len(MESES)+2, 12)
row_height(ws6, rn, 22); rn += 1

# Cronograma por cluster (top 10)
top_clusters = resumo_clusters[:10]
BAR = "████"
VAZIO = "    "
urg_color = {
    "Critico": VERM_CLR, "Crítico": VERM_CLR,
    "Alta": AMAR_CLR, "Média": VERDE_CLR, "Baixa": AZUL_CLR,
}
for cl_row in top_clusters:
    cname    = str(cl_row.get("Cluster","") or "")
    lig_tot  = _i(cl_row.get("Ligacoes","") or 0)
    urg      = str(cl_row.get("Urgencia dom.","") or "")
    bg_urg   = urg_color.get(urg, CINZA_CLR)
    n_nuc    = _i(cl_row.get("N Nucleos","") or 1)
    # Estimar meses de execucao (simplificado)
    meses_exec = max(1, min(9, round(n_nuc * 1.5)))
    inicio_mes = 0  # abril
    c = ws6.cell(rn, 1, cname)
    apply(c, fill_=fill(AZUL_CLR), font_=font(bold=True, size=10),
          aln_=aln("left"), border_=thin_b)
    for mi, mes in enumerate(MESES):
        ativo = inicio_mes <= mi < inicio_mes + meses_exec
        val   = BAR if ativo else VAZIO
        cell  = ws6.cell(rn, mi+2, val)
        apply(cell, fill_=fill(bg_urg if ativo else CINZA_CLR),
              font_=font(bold=True, color=VERM_ESC if urg in ("Critico","Crítico","Alta") else VERDE_ESC, size=9),
              aln_=aln("center"), border_=thin_b)
    ws6.cell(rn, len(MESES)+2, lig_tot)
    apply(ws6.cell(rn, len(MESES)+2), fill_=fill(VERDE_CLR),
          font_=font(bold=True, size=10), aln_=aln("center"), border_=thin_b)
    row_height(ws6, rn, 18); rn += 1

rn += 1
# Meta mensal
c = ws6.cell(rn, 1, "META 1.000 lig/mês")
apply(c, fill_=fill(VERM_ESC), font_=font(bold=True, color="FFFFFF", size=10),
      aln_=aln("center"), border_=thin_b)
metas_mensais = [600, 1000, 1200, 1200, 1200, 1100, 1100, 1100, 1000]
for mi, m in enumerate(metas_mensais):
    cell = ws6.cell(rn, mi+2, f"{m}")
    apply(cell, fill_=fill(VERM_CLR if m < 1000 else VERDE_CLR),
          font_=font(bold=True, size=10), aln_=aln("center"), border_=thin_b)
row_height(ws6, rn, 20)

print("  ABA 06_CRONOGRAMA criada")

# ══════════════════════════════════════════════════════════════════════════════
# ABA 7 — CUSTOS
# ══════════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("07_CUSTOS")
ws7.sheet_view.showGridLines = False
rn = titulo_aba(ws7, "INVESTIMENTO POR CLUSTER GEOGRÁFICO",
                "Dados de TRECHOS_MEGA_CLUSTERS.xlsx | 83 núcleos | 35 clusters", N_COLS)

hdrs7 = ["Cluster", "N Núcleos", "N Trechos", "Ligações",
         "Extensão(m)", "Total s/BDI (R$)", "Total c/BDI (R$)",
         "Custo/Ligação(R$)", "Urgência Dominante"]
larg7 = [22, 10, 10, 10, 14, 20, 20, 16, 14]
for ci, (h, w) in enumerate(zip(hdrs7, larg7), 1):
    c = ws7.cell(rn, ci, h)
    apply(c, fill_=fill(AZUL_ESC), font_=font(bold=True, color="FFFFFF", size=10),
          aln_=aln("center"), border_=thin_b)
    col_width(ws7, ci, w)
row_height(ws7, rn, 22); rn += 1

gt7 = defaultdict(float)
for ri, ci_row in enumerate(resumo_clusters):
    cname   = str(ci_row.get("Cluster","") or "")
    n_nuc   = _i(ci_row.get("N Nucleos","") or ci_row.get("N_Nucleos","") or 0)
    n_tr    = _i(ci_row.get("N Trechos","") or ci_row.get("N_Trechos","") or 0)
    lig     = _i(ci_row.get("Ligacoes","") or 0)
    ext_m   = _f(ci_row.get("Ext Total(m)","") or 0)
    cbdi    = _f(ci_row.get("Custo Total c/BDI","") or 0)
    sbdi    = round(cbdi / 1.25, 2)
    cust_lig = round(cbdi / lig, 2) if lig > 0 else 0
    urg     = str(ci_row.get("Urgencia dom.","") or "")
    bg = urg_color.get(urg, CINZA_CLR)
    vals7 = [cname, n_nuc, n_tr, lig, round(ext_m,1),
             sbdi, cbdi, cust_lig, urg]
    for ci, v in enumerate(vals7, 1):
        c = ws7.cell(rn, ci, v)
        apply(c, fill_=fill(bg if ci >= 5 else (AZUL_CLR if ri%2==0 else "FFFFFF")),
              font_=font(size=10), aln_=aln("center" if ci != 1 else "left"),
              border_=thin_b)
    for k, v in [("lig", lig), ("ext", ext_m), ("sbdi", sbdi), ("cbdi", cbdi)]:
        gt7[k] += v
    row_height(ws7, rn, 18); rn += 1

# Grand total
vals_gt = ["GRAND TOTAL — 83 NÚCLEOS / 35 CLUSTERS",
           len(resumo_clusters), "", int(gt7["lig"]),
           round(gt7["ext"],1), round(gt7["sbdi"],2), round(gt7["cbdi"],2),
           round(gt7["cbdi"]/max(1,gt7["lig"]),2), ""]
for ci, v in enumerate(vals_gt, 1):
    c = ws7.cell(rn, ci, v)
    apply(c, fill_=fill(AZUL_ESC), font_=font(bold=True, color="FFFFFF", size=11),
          aln_=aln("center" if ci != 1 else "left"), border_=thin_b)
row_height(ws7, rn, 22)

print("  ABA 07_CUSTOS criada")

# ══════════════════════════════════════════════════════════════════════════════
# ABA 8 — TODOS 83 NUCLEOS
# ══════════════════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("08_83_NUCLEOS")
ws8.sheet_view.showGridLines = False
rn = titulo_aba(ws8, "TODOS OS 83 NÚCLEOS — VISÃO COMPLETA",
                "Dados consolidados | nucleos_enriquecido.csv | CONSTRUDATA_v5", N_COLS + 2)

hdrs8 = ["Cluster", "TAG", "Nome do Núcleo", "Setor", "LIG_AGUA", "LIG_ESG",
         "LIG_TOTAL", "EXT_ES(m)", "DIFIC", "Urgência",
         "C_TOTAL(R$)", "Eficiência", "Status", "DATA_INI_EST"]
larg8 = [22, 8, 32, 18, 9, 9, 10, 10, 7, 10, 18, 10, 12, 14]
for ci, (h, w) in enumerate(zip(hdrs8, larg8), 1):
    c = ws8.cell(rn, ci, h)
    apply(c, fill_=fill(AZUL_ESC), font_=font(bold=True, color="FFFFFF", size=9),
          aln_=aln("center"), border_=thin_b)
    col_width(ws8, ci, w)
ws8.freeze_panes = "A" + str(rn+1)
row_height(ws8, rn, 22); rn += 1

for ri, r in enumerate(sorted(nuc_rows,
                               key=lambda x: (tag_to_cluster.get(x["TAG"],"ZZZZ"), x["TAG"]))):
    tag    = r["TAG"]
    status = "EM EXECUCAO" if tag in ativos_tags_set else "PLANEJADO"
    bg_status = VERDE_CLR if status == "EM EXECUCAO" else (AMAR_CLR if ri%2==0 else "FFFFFF")
    urg    = r.get("URGENCIA","")
    vals8 = [
        tag_to_cluster.get(tag, "ISOLADO"),
        tag, r.get("NOME",""), r.get("SETOR",""),
        _i(r.get("LIG_AGUA",0)), _i(r.get("LIG_ESG",0)),
        _i(r.get("LIG_TOTAL",0)), _f(r.get("EXT_ES_M",0)),
        _i(r.get("DIFIC",0)), urg,
        round(_f(r.get("C_TOTAL",0)),2),
        round(_f(r.get("EFICIENCIA",0)),1),
        status, r.get("DATA_INI","")
    ]
    for ci, v in enumerate(vals8, 1):
        c = ws8.cell(rn, ci, v)
        bg = (VERDE_CLR if status == "EM EXECUCAO"
              else urg_color.get(urg, AMAR_CLR if ri%2==0 else "FFFFFF"))
        apply(c, fill_=fill(bg if ci in (10,12,13) else (CINZA_CLR if ri%2==0 else "FFFFFF")),
              font_=font(size=9, bold=(status=="EM EXECUCAO")),
              aln_=aln("left" if ci == 3 else "center"), border_=thin_b)
    row_height(ws8, rn, 16); rn += 1

print("  ABA 08_83_NUCLEOS criada")

# Salvar XLSX
wb.save(OUT_XLSX)
print(f"\nSalvo: {OUT_XLSX}")
print(f"Abas: {[ws.title for ws in wb.worksheets]}")

# ══════════════════════════════════════════════════════════════════════════════
# MAPA HTML — Folium
# ══════════════════════════════════════════════════════════════════════════════
print("\nGerando mapa HTML...")

# Paleta de cores para 35 clusters
CORES_CLUSTERS = [
    "#E6194B","#3CB44B","#4363D8","#F58231","#911EB4",
    "#42D4F4","#F032E6","#BFEF45","#FABEBE","#469990",
    "#E6BEFF","#9A6324","#FFFAC8","#800000","#AAFFC3",
    "#808000","#FFD8B1","#000075","#A9A9A9","#FFFFFF",
    "#FF4444","#44AA44","#4444FF","#FF8800","#AA44AA",
    "#44AAAA","#AA4444","#88AA44","#44AA88","#AA8844",
    "#8844AA","#44AA44","#AAAA44","#44AAAA","#AA4488",
]
cluster_list = list(cluster_tags.keys())
cluster_color = {c: CORES_CLUSTERS[i % len(CORES_CLUSTERS)]
                 for i, c in enumerate(cluster_list)}

# Ler GPKG — filtrar rows validas de Santos
print("  Lendo GPKG...")
try:
    gdf_all = gpd.read_file(GPKG_FILE)
    gdf_all["cx"] = gdf_all.geometry.centroid.x
    gdf_all["cy"] = gdf_all.geometry.centroid.y
    # Santos UTM 23S: E ~340000-390000, N ~7340000-7370000
    gdf_valid = gdf_all[(gdf_all["cx"] > 300000) &
                         (gdf_all["cx"] < 450000) &
                         (gdf_all["cy"] > 7300000)].copy()
    gdf_valid = gdf_valid.to_crs(4326)
    print(f"  {len(gdf_valid)} geometrias validas | {gdf_valid['TAG'].nunique()} TAGs")
    USE_GDF = True
except Exception as e:
    print(f"  Aviso GPKG: {e} — usando fallback de circulos")
    USE_GDF = False

# Centro Santos
m = folium.Map(location=[-23.950, -46.335], zoom_start=13,
               tiles="CartoDB positron")

# Legenda lateral
legend_html = """
<div style="position:fixed; bottom:30px; right:10px; z-index:9999;
     background:white; padding:10px; border:2px solid #003366;
     border-radius:8px; font-size:11px; max-height:400px; overflow-y:auto;
     box-shadow: 3px 3px 8px rgba(0,0,0,0.3);">
<b style="font-size:13px; color:#003366;">CLUSTERS GEOGRÁFICOS</b><br>
<span style="font-size:10px; color:#666;">Contrato SLNR 11481051 | Santos/SP</span><br><hr>
"""
for ci_row in resumo_clusters[:20]:
    cname   = str(ci_row.get("Cluster","") or "")
    n_nuc   = ci_row.get("N Nucleos","") or ""
    lig     = ci_row.get("Ligacoes","") or ""
    cor     = cluster_color.get(cname, "#888888")
    legend_html += (f'<div style="margin:2px 0">'
                    f'<span style="background:{cor};display:inline-block;'
                    f'width:14px;height:14px;border-radius:2px;margin-right:5px;'
                    f'vertical-align:middle;border:1px solid #333"></span>'
                    f'<b>{cname}</b> ({n_nuc} nuc. | {lig} lig.)</div>\n')
legend_html += "</div>"
m.get_root().html.add_child(folium.Element(legend_html))

if USE_GDF:
    for tag, group in gdf_valid.groupby("TAG"):
        cname = tag_to_cluster.get(str(tag), None)
        if cname is None:
            color = "#AAAAAA"
            cname = "SEM CLUSTER"
        else:
            color = cluster_color.get(cname, "#AAAAAA")
        is_ativo_flag = str(tag) in ativos_tags_set
        row_info = nuc_dict.get(str(tag), {})
        nome_nuc = row_info.get("NOME", str(tag))
        lig_a    = _i(row_info.get("LIG_AGUA", 0))
        lig_e    = _i(row_info.get("LIG_ESG", 0))
        urg_txt  = row_info.get("URGENCIA", "")
        tooltip_txt = (f"<b>{nome_nuc}</b><br>"
                       f"TAG: {tag} | Cluster: {cname}<br>"
                       f"Lig. Água: {lig_a} | Lig. Esg.: {lig_e}<br>"
                       f"Urgência: {urg_txt}<br>"
                       f"Status: {'EM EXECUCAO' if is_ativo_flag else 'PLANEJADO'}")
        border_color = "#00AA00" if is_ativo_flag else "#003366"
        border_w = 4 if is_ativo_flag else 1
        folium.GeoJson(
            group.__geo_interface__,
            style_function=lambda x, c=color, bw=border_w, bc=border_color: {
                "fillColor": c, "color": bc,
                "weight": bw, "fillOpacity": 0.55,
            },
            tooltip=folium.Tooltip(tooltip_txt),
            name=cname,
        ).add_to(m)
else:
    # Fallback: marcadores com circulos
    SETOR_COORDS = {
        "ST-ZONA NOROESTE":  (-23.940, -46.368),
        "ST-CRUZEIRO":       (-23.960, -46.320),
        "ST-VILA PROGRESSO": (-23.972, -46.345),
        "ST-ALEMOA":         (-23.987, -46.308),
        "ST-ZONA LESTE":     (-23.968, -46.298),
        "ST-SAO BENTO":      (-23.943, -46.333),
        "ST-MONTE SERRAT":   (-23.950, -46.342),
        "ST-MORRO J MENINO": (-23.944, -46.350),
        "ST-CANELEIRA":      (-23.986, -46.358),
        "ST-MORRO MARAPE":   (-23.961, -46.305),
        "ST-PENHA":          (-23.953, -46.362),
    }
    import random
    for tag, row_info in nuc_dict.items():
        setor = row_info.get("SETOR","")
        base  = SETOR_COORDS.get(setor, (-23.958, -46.333))
        rng   = random.Random(int(str(tag)[:6]) if str(tag).isdigit() else hash(tag))
        lat   = base[0] + rng.uniform(-0.012, 0.012)
        lon   = base[1] + rng.uniform(-0.012, 0.012)
        cname = tag_to_cluster.get(str(tag), "SEM_CLUSTER")
        color = cluster_color.get(cname, "#AAAAAA")
        lig_a = _i(row_info.get("LIG_AGUA",0))
        radius = max(30, min(200, lig_a * 0.1))
        folium.Circle(
            location=[lat, lon], radius=radius,
            color=color, fill=True, fill_color=color, fill_opacity=0.5,
            tooltip=f"{row_info.get('NOME',tag)} | {cname} | {lig_a} lig. água"
        ).add_to(m)

# Marcadores especiais para nucl. ativos
for r in nucleos_ativos:
    tag = r["TAG"]
    if USE_GDF and str(tag) in gdf_valid["TAG"].values:
        centroid = gdf_valid[gdf_valid["TAG"]==str(tag)].dissolve().centroid.iloc[0]
        lat, lon = centroid.y, centroid.x
    else:
        setor = r.get("SETOR","")
        base  = (-23.958, -46.333)
        lat, lon = base[0], base[1]
    folium.Marker(
        location=[lat, lon],
        popup=f"<b>EM EXECUCAO</b><br>{r['NOME']}<br>TAG: {tag}",
        tooltip=f"★ {r['NOME']} (ATIVO)",
        icon=folium.Icon(color="green", icon="star", prefix="fa")
    ).add_to(m)

folium.LayerControl().add_to(m)

# Título no mapa
title_html = """
<div style="position:fixed; top:10px; left:50%; transform:translateX(-50%);
     z-index:9999; background:rgba(0,51,102,0.9); color:white;
     padding:8px 20px; border-radius:6px; font-size:14px; font-weight:bold;
     box-shadow:2px 2px 6px rgba(0,0,0,0.4);">
CLUSTERS GEOGRÁFICOS — SLNR SANTOS/SP — Contrato 11481051
<span style="font-size:11px; font-weight:normal; margin-left:12px;">
★ = Núcleo em execução</span>
</div>
"""
m.get_root().html.add_child(folium.Element(title_html))

m.save(OUT_HTML)
print(f"Salvo: {OUT_HTML}")
print(f"\n{'='*60}")
print("CONCLUIDO!")
print(f"  XLSX: {OUT_XLSX}")
print(f"  HTML: {OUT_HTML}")
print(f"  Pasta: {OUT_DIR}")
