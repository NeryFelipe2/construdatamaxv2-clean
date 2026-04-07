#!/usr/bin/env python3
"""
gerar_compras.py — Sistema de Compras por NS
ConstruData SABESP v5.0 — Contrato 11481051

Calcula materiais por trecho (NS) e consolida para compras.
Gera: Excel com 4 abas + HTML interativo.

Uso:
    from gerar_compras import gerar_compras
    gerar_compras(pvs, trechos, nucleo, pasta_saida)
"""

import json
import math
import os
from collections import defaultdict
from datetime import datetime
from pathlib import Path

# ═══════════════════════════════════════════════════════════════════
# CATÁLOGO DE MATERIAIS E PREÇOS (contrato SLNR Santos)
# ═══════════════════════════════════════════════════════════════════

CATALOGO = {
    # Tubos esgoto
    "Tubo PVC DN150":   {"un": "m", "preco": 155.00, "lead_time": "2 sem", "tipo": "ESG"},
    "Tubo PVC DN200":   {"un": "m", "preco": 200.12, "lead_time": "2 sem", "tipo": "ESG"},
    "Tubo PVC DN250":   {"un": "m", "preco": 255.00, "lead_time": "2 sem", "tipo": "ESG"},
    "Tubo PVC DN300":   {"un": "m", "preco": 310.00, "lead_time": "3 sem", "tipo": "ESG"},
    "Tubo PVC DN400":   {"un": "m", "preco": 450.00, "lead_time": "4 sem", "tipo": "ESG"},
    # Tubos água
    "PEAD PE80 DN63":   {"un": "m", "preco": 85.00,  "lead_time": "4 sem", "tipo": "AG"},
    "PEAD PE80 DN110":  {"un": "m", "preco": 101.80, "lead_time": "4 sem", "tipo": "AG"},
    "PEAD PE80 DN160":  {"un": "m", "preco": 145.00, "lead_time": "4 sem", "tipo": "AG"},
    # Estruturas
    "PV concreto DN1200":{"un": "un", "preco": 3686.00,"lead_time": "2 sem", "tipo": "ESG"},
    "PI plástico":       {"un": "un", "preco": 1412.00,"lead_time": "1 sem", "tipo": "ESG"},
    "Caixa inspeção":    {"un": "un", "preco": 890.00, "lead_time": "1 sem", "tipo": "ESG"},
    # Conexões
    "Luva correr PVC":   {"un": "un", "preco": 25.00,  "lead_time": "1 sem", "tipo": "ESG"},
    "Anel borracha":     {"un": "un", "preco": 8.50,   "lead_time": "1 sem", "tipo": "ESG"},
    "Junção Y PVC":      {"un": "un", "preco": 45.00,  "lead_time": "1 sem", "tipo": "ESG"},
    # Agregados
    "Areia berço":       {"un": "m³", "preco": 160.00, "lead_time": "1 sem", "tipo": "GERAL"},
    "Brita drenagem":    {"un": "m³", "preco": 180.00, "lead_time": "1 sem", "tipo": "GERAL"},
    "Concreto":          {"un": "m³", "preco": 720.00, "lead_time": "1 sem", "tipo": "GERAL"},
    # Pavimentação
    "CBUQ repavim.":     {"un": "t",  "preco": 820.00, "lead_time": "2 sem", "tipo": "GERAL"},
    # Ramal
    "Ramal esgoto DN100":{"un": "un", "preco": 350.00, "lead_time": "1 sem", "tipo": "ESG"},
    "Sela/Cavalete AG":  {"un": "un", "preco": 280.00, "lead_time": "2 sem", "tipo": "AG"},
}


def calcular_materiais_ns(trecho, pvs):
    """Calcula lista de materiais para 1 NS (trecho)."""
    ext = trecho.get("ext_m", 0) or 0
    dn = trecho.get("dn_mm", 200)
    material = trecho.get("material", "PVC").upper()
    tipo = trecho.get("tipo", "esgoto")

    # Profundidade média
    pv_a = pvs.get(trecho.get("pv_ini", ""), {})
    pv_b = pvs.get(trecho.get("pv_fim", ""), {})
    prof_a = abs((pv_a.get("ct") or 0) - (pv_a.get("cf") or 0)) or 1.5
    prof_b = abs((pv_b.get("ct") or 0) - (pv_b.get("cf") or 0)) or 1.5
    prof_med = (prof_a + prof_b) / 2
    largura = 0.60 if dn <= 200 else 0.80

    materiais = []

    if tipo == "esgoto" or "PVC" in material:
        # Tubo PVC
        n_barras = math.ceil(ext / 6) if ext > 0 else 0
        tubo_nome = f"Tubo PVC DN{dn}"
        materiais.append({"material": tubo_nome, "un": "m", "qtd": round(ext, 1)})
        materiais.append({"material": "Luva correr PVC", "un": "un", "qtd": max(0, n_barras - 1)})
        materiais.append({"material": "Anel borracha", "un": "un", "qtd": n_barras + 1})
        materiais.append({"material": "Junção Y PVC", "un": "un", "qtd": 1})
        materiais.append({"material": "Ramal esgoto DN100", "un": "un", "qtd": 1})

        # PV ou PI
        pv_ini_nome = trecho.get("pv_ini", "")
        if "PV" in pv_ini_nome.upper():
            materiais.append({"material": "PV concreto DN1200", "un": "un", "qtd": 1})
        else:
            materiais.append({"material": "PI plástico", "un": "un", "qtd": 1})

    elif "PEAD" in material or "PE" in material:
        # Tubo PEAD
        if dn <= 75:
            tubo_nome = "PEAD PE80 DN63"
        elif dn <= 125:
            tubo_nome = "PEAD PE80 DN110"
        else:
            tubo_nome = "PEAD PE80 DN160"
        materiais.append({"material": tubo_nome, "un": "m", "qtd": round(ext, 1)})
        materiais.append({"material": "Sela/Cavalete AG", "un": "un", "qtd": 1})

    else:
        materiais.append({"material": f"Tubo PVC DN{dn}", "un": "m", "qtd": round(ext, 1)})

    # Agregados (baseado em volume de vala)
    vol_areia = round(ext * largura * 0.15, 2)  # berço 15cm
    vol_brita = round(ext * largura * 0.10, 2)  # drenagem 10cm
    vol_concreto = round(0.5, 2)  # PV base
    peso_cbuq = round(ext * largura * 0.05 * 2.4, 2)  # 5cm × 2.4 t/m³

    materiais.append({"material": "Areia berço", "un": "m³", "qtd": vol_areia})
    materiais.append({"material": "Brita drenagem", "un": "m³", "qtd": vol_brita})
    materiais.append({"material": "Concreto", "un": "m³", "qtd": vol_concreto})
    materiais.append({"material": "CBUQ repavim.", "un": "t", "qtd": peso_cbuq})

    return materiais


def consolidar_compras(pvs, trechos):
    """Consolida materiais de todos os trechos."""
    consolidado = defaultdict(lambda: {"un": "", "qtd": 0, "ns_list": []})

    for i, tr in enumerate(trechos):
        ns_id = i + 1
        materiais = calcular_materiais_ns(tr, pvs)
        for m in materiais:
            nome = m["material"]
            consolidado[nome]["un"] = m["un"]
            consolidado[nome]["qtd"] += m["qtd"]
            consolidado[nome]["ns_list"].append(ns_id)

    # Adicionar preços e lead time do catálogo
    resultado = []
    for nome, dados in sorted(consolidado.items(), key=lambda x: -x[1]["qtd"]):
        cat = CATALOGO.get(nome, {})
        # Match parcial para tubos com DN específico
        if not cat:
            for cat_nome, cat_dados in CATALOGO.items():
                if cat_nome in nome or nome in cat_nome:
                    cat = cat_dados
                    break
        preco = cat.get("preco", 0)
        resultado.append({
            "material": nome,
            "un": dados["un"],
            "qtd": round(dados["qtd"], 1),
            "preco_unit": preco,
            "custo": round(dados["qtd"] * preco, 2),
            "lead_time": cat.get("lead_time", "?"),
            "tipo": cat.get("tipo", "GERAL"),
            "n_ns": len(set(dados["ns_list"])),
        })

    return sorted(resultado, key=lambda x: -x["custo"])


def gerar_compras(pvs, trechos, nucleo, pasta_saida):
    """
    Gera sistema de compras completo.
    Retorna dict com dados consolidados.
    """
    pasta = Path(pasta_saida)
    pasta.mkdir(parents=True, exist_ok=True)

    compras = consolidar_compras(pvs, trechos)
    custo_total = sum(c["custo"] for c in compras)
    n_itens = len(compras)

    # ── Gerar HTML ────────────────────────────────────────────────
    _gerar_html_compras(compras, nucleo, custo_total, len(trechos), pasta / "COMPRAS.html")

    # ── Gerar Excel ───────────────────────────────────────────────
    _gerar_xlsx_compras(compras, pvs, trechos, nucleo, custo_total, pasta / "COMPRAS.xlsx")

    # ── Gerar JSON ────────────────────────────────────────────────
    dados = {
        "nucleo": nucleo,
        "gerado_em": datetime.now().isoformat(),
        "total_ns": len(trechos),
        "total_itens": n_itens,
        "custo_total": custo_total,
        "compras": compras,
    }
    with open(pasta / "COMPRAS.json", "w", encoding="utf-8") as f:
        json.dump(dados, f, indent=2, ensure_ascii=False)

    print(f"[COMPRAS] {n_itens} itens | {len(trechos)} NS | R$ {custo_total:,.0f}")

    return dados


def _gerar_html_compras(compras, nucleo, custo_total, n_ns, out_path):
    """Gera HTML interativo de compras."""
    rows = ""
    for c in compras:
        prio = "🔴 URGENTE" if c["custo"] > 100000 else ("🟡 ALTO" if c["custo"] > 10000 else "🟢 NORMAL")
        rows += f"""<tr>
          <td style="font-weight:600">{c['material']}</td>
          <td class="c">{c['tipo']}</td>
          <td class="c">{c['un']}</td>
          <td class="r mono">{c['qtd']:,.1f}</td>
          <td class="r mono">R$ {c['preco_unit']:,.2f}</td>
          <td class="r mono" style="font-weight:700">R$ {c['custo']:,.0f}</td>
          <td class="c">{c['lead_time']}</td>
          <td class="c">{c['n_ns']}</td>
          <td class="c">{prio}</td>
        </tr>"""

    # Top 5 por custo para gráfico
    top5 = compras[:5]
    labels = [c['material'][:20] for c in top5]
    values = [c['custo'] for c in top5]

    html = f"""<!DOCTYPE html>
<html lang="pt-BR"><head><meta charset="UTF-8">
<title>Compras — {nucleo}</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.2/dist/chart.umd.min.js"></script>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:Arial,sans-serif;background:#f5f5f5;padding:16px}}
.hdr{{background:linear-gradient(135deg,#C62828,#E53935);color:#fff;padding:20px;border-radius:12px;margin-bottom:16px}}
.hdr h1{{font-size:20px}}.hdr p{{opacity:.85;font-size:13px;margin-top:4px}}
.kpis{{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:16px}}
.kpi{{background:#fff;border-radius:10px;padding:14px;text-align:center}}
.kpi .l{{font-size:11px;color:#666;font-weight:600}}.kpi .v{{font-size:20px;font-weight:700;color:#C62828}}
.kpi .s{{font-size:10px;color:#999}}
.grid{{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:16px}}
.card{{background:#fff;border-radius:10px;padding:16px}}
table{{width:100%;border-collapse:collapse;font-size:12px;background:#fff;border-radius:10px;overflow:hidden}}
thead tr{{background:#C62828;color:#fff}}
th,td{{padding:8px;border-bottom:1px solid #f0f0f0}}
th{{text-align:left;font-size:11px}}.r{{text-align:right}}.c{{text-align:center}}
.mono{{font-family:monospace}}
tr:hover{{background:#fff3f3}}
</style></head><body>
<div class="hdr">
<h1>Sistema de Compras — {nucleo}</h1>
<p>FCN Construções e Saneamento · Contrato 11481051 · {n_ns} Notas de Serviço</p>
</div>
<div class="kpis">
<div class="kpi"><div class="l">Custo Total</div><div class="v">R$ {custo_total:,.0f}</div><div class="s">materiais + agregados</div></div>
<div class="kpi"><div class="l">Itens</div><div class="v">{len(compras)}</div><div class="s">tipos de material</div></div>
<div class="kpi"><div class="l">NS Atendidas</div><div class="v">{n_ns}</div><div class="s">trechos</div></div>
<div class="kpi"><div class="l">Urgentes</div><div class="v" style="color:#C62828">{sum(1 for c in compras if c['custo']>100000)}</div><div class="s">&gt; R$ 100k</div></div>
</div>
<div class="grid">
<div class="card"><h3 style="margin-bottom:12px;font-size:14px">Top 5 — Maior Custo</h3><canvas id="chart1" height="200"></canvas></div>
<div class="card"><h3 style="margin-bottom:12px;font-size:14px">Por Tipo (ESG / AG / GERAL)</h3><canvas id="chart2" height="200"></canvas></div>
</div>
<table>
<thead><tr><th>Material</th><th class="c">Tipo</th><th class="c">Un</th><th class="r">Qtd</th><th class="r">P.Unit</th><th class="r">Custo</th><th class="c">Lead Time</th><th class="c">NS</th><th class="c">Prioridade</th></tr></thead>
<tbody>{rows}</tbody>
<tfoot><tr style="background:#f5f5f5;font-weight:700"><td colspan="5">TOTAL</td><td class="r mono">R$ {custo_total:,.0f}</td><td colspan="3"></td></tr></tfoot>
</table>
<div style="margin-top:16px;font-size:11px;color:#999;text-align:center">
Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} · ConstruData SABESP v5.0
</div>
<script>
new Chart(document.getElementById('chart1'),{{type:'bar',data:{{labels:{json.dumps(labels)},datasets:[{{label:'Custo (R$)',data:{json.dumps(values)},backgroundColor:['#C62828','#E53935','#EF5350','#E57373','#EF9A9A']}}]}},options:{{indexAxis:'y',plugins:{{legend:{{display:false}}}}}}}});
const tipoData={json.dumps({t: sum(c['custo'] for c in compras if c['tipo']==t) for t in ['ESG','AG','GERAL']})};
new Chart(document.getElementById('chart2'),{{type:'doughnut',data:{{labels:Object.keys(tipoData),datasets:[{{data:Object.values(tipoData),backgroundColor:['#C62828','#1565C0','#2E7D32']}}]}}}});
</script>
</body></html>"""

    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)


def _gerar_xlsx_compras(compras, pvs, trechos, nucleo, custo_total, out_path):
    """Gera Excel de compras com 2 abas."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return

    wb = openpyxl.Workbook()

    # ── Aba 1: CONSOLIDADO ────────────────────────────────────────
    ws = wb.active
    ws.title = "COMPRAS_CONSOLIDADO"

    hdr_fill = PatternFill("solid", fgColor="C62828")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)

    headers = ["Material", "Tipo", "Un", "Qtd", "P.Unit (R$)", "Custo (R$)", "Lead Time", "NS", "Prioridade"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hdr_fill
        c.font = hdr_font

    for i, comp in enumerate(compras, 2):
        prio = "URGENTE" if comp["custo"] > 100000 else ("ALTO" if comp["custo"] > 10000 else "NORMAL")
        ws.cell(row=i, column=1, value=comp["material"])
        ws.cell(row=i, column=2, value=comp["tipo"])
        ws.cell(row=i, column=3, value=comp["un"])
        ws.cell(row=i, column=4, value=comp["qtd"])
        ws.cell(row=i, column=5, value=comp["preco_unit"])
        ws.cell(row=i, column=6, value=comp["custo"])
        ws.cell(row=i, column=7, value=comp["lead_time"])
        ws.cell(row=i, column=8, value=comp["n_ns"])
        ws.cell(row=i, column=9, value=prio)

    # Totais
    row_tot = len(compras) + 2
    ws.cell(row=row_tot, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row_tot, column=6, value=custo_total).font = Font(bold=True)

    # Ajustar larguras
    for col in range(1, 10):
        ws.column_dimensions[chr(64 + col)].width = 18

    # ── Aba 2: POR NS ────────────────────────────────────────────
    ws2 = wb.create_sheet("MATERIAIS_POR_NS")
    headers2 = ["NS", "PV Ini", "PV Fim", "Ext (m)", "DN", "Material", "Un", "Qtd", "Custo (R$)"]
    for col, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.fill = PatternFill("solid", fgColor="1F4E79")
        c.font = Font(bold=True, color="FFFFFF", size=10)

    row = 2
    for i, tr in enumerate(trechos):
        mats = calcular_materiais_ns(tr, pvs)
        for m in mats:
            cat = CATALOGO.get(m["material"], {})
            preco = cat.get("preco", 0)
            ws2.cell(row=row, column=1, value=f"NS_{i+1:03d}")
            ws2.cell(row=row, column=2, value=tr.get("pv_ini", ""))
            ws2.cell(row=row, column=3, value=tr.get("pv_fim", ""))
            ws2.cell(row=row, column=4, value=tr.get("ext_m", 0))
            ws2.cell(row=row, column=5, value=tr.get("dn_mm", 0))
            ws2.cell(row=row, column=6, value=m["material"])
            ws2.cell(row=row, column=7, value=m["un"])
            ws2.cell(row=row, column=8, value=m["qtd"])
            ws2.cell(row=row, column=9, value=round(m["qtd"] * preco, 2))
            row += 1

    for col in range(1, 10):
        ws2.column_dimensions[chr(64 + col)].width = 16

    wb.save(out_path)


# ═══════════════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import sys
    json_path = sys.argv[1] if len(sys.argv) > 1 else "_tmp_dwg/REDE_EXTRAIDA_BIM.json"

    if not Path(json_path).exists():
        print(f"Uso: python gerar_compras.py <rede.json>")
        sys.exit(1)

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    resultado = gerar_compras(
        data.get("pvs", {}),
        data.get("trechos", []),
        data.get("arquivo", "Teste"),
        "SAIDA_BIM_SABESP/COMPRAS",
    )
