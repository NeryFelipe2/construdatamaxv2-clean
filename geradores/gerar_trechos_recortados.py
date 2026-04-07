#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gera TRECHOS_NUCLEOS_RECORTADOS.xlsx
Trechos PV-a-PV organizados pelos 8 grupos de proximidade real (RECORTADO1-8.gpkg)
Mesmas colunas de TRECHOS_MEGA_CLUSTERS.xlsx — divisao por proximidade geografica real
ConstruData HydroNetwork v7.0.0 | FCN Construcoes e Saneamento | Contrato 11481051
"""
import sys, io, os, math, random, csv
from collections import defaultdict, Counter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import geopandas as gpd

# --------------------------------------------------------------------------- #
#  Caminhos
# --------------------------------------------------------------------------- #
SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
PROX_DIR     = os.path.join(SCRIPT_DIR, "TRECHOS_NS_COMPLETOS", "NUCLEOS PROXIMOS")
OUT_DIR      = os.path.join(SCRIPT_DIR, "TRECHOS_NS_COMPLETOS")
OUT_FILE     = os.path.join(OUT_DIR, "TRECHOS_NUCLEOS_RECORTADOS.xlsx")

# nucleos_enriquecido.csv como fallback (caso GPKG nao tenha alguma coluna)
CSV_NUC = os.path.join(
    os.environ.get("USERPROFILE", "C:\\Users\\felip"),
    "Downloads", "NOVOS NUCLEOS PLANEJAMENTO", "CONSTRUDATA_v5",
    "nucleos_enriquecido.csv")

# --------------------------------------------------------------------------- #
#  Definicao dos 8 grupos (nome descritivo + aba + arquivo GPKG)
# --------------------------------------------------------------------------- #
GRUPOS = [
    ("R1", "R1_LESTE_MARAPE",       "Zona Leste / Morro Marapé / J. Menino",  "ZONA_LESTE_MARAPE.gpkg"),
    ("R2", "R2_CRUZEIRO_PROG_I",    "Cruzeiro + Vila Progresso I",             "CRUZEIRO_PROGRESSO_I.gpkg"),
    ("R3", "R3_CRUZEIRO_PROG_II",   "Cruzeiro + Vila Progresso II (Core)",     "CRUZEIRO_PROGRESSO_II.gpkg"),
    ("R4", "R4_CRUZEIRO_PENHA_NRO", "Cruzeiro Sul / Penha / Noroeste",         "CRUZEIRO_PENHA_NOROESTE.gpkg"),
    ("R5", "R5_CRUZEIRO_SAO_BENTO", "Cruzeiro / São Bento / Monte Serrat",     "CRUZEIRO_SAO_BENTO.gpkg"),
    ("R6", "R6_MONTE_SERRAT",       "Monte Serrat / Zona Leste Sul",           "MONTE_SERRAT_LESTE.gpkg"),
    ("R7", "R7_PIRAT_MANOEL_NRO",   "Piratininga / S. Manoel / Noroeste",      "PIRATININGA_MANOEL_NOROESTE.gpkg"),
    ("R8", "R8_NOROESTE_VILA_GILDA","Noroeste / Vila Gilda / Palafitas",       "NOROESTE_VILA_GILDA.gpkg"),
]

# --------------------------------------------------------------------------- #
#  Prof base por DIFIC / terreno
# --------------------------------------------------------------------------- #
PROF_DIFIC = {1: 1.0, 2: 1.3, 3: 1.7, 4: 2.1, 5: 2.5}
TERR = {
    "dique":         ("CBUQ",           200, 0.010),
    "morro":         ("PARALELEPIPEDO", 150, 0.015),
    "monte":         ("PARALELEPIPEDO", 150, 0.015),
    "encosta":       ("PARALELEPIPEDO", 150, 0.012),
    "urbano":        ("CBUQ",           150, 0.010),
    "prolongamento": ("CIMENTADO",      150, 0.008),
    "planicie":      ("CBUQ",           150, 0.008),
    "palafita":      ("CBUQ",           100, 0.008),
}

# --------------------------------------------------------------------------- #
#  Estilos
# --------------------------------------------------------------------------- #
fill_h1    = PatternFill("solid", fgColor="1F4E79")
fill_h2    = PatternFill("solid", fgColor="2E75B6")
fill_grupo = PatternFill("solid", fgColor="D6E4F0")
fill_nucl  = PatternFill("solid", fgColor="E2EFDA")
fill_sub   = PatternFill("solid", fgColor="FFF2CC")
fill_grand = PatternFill("solid", fgColor="FCE4D6")
fill_isol  = PatternFill("solid", fgColor="F2F2F2")
fill_zebra = PatternFill("solid", fgColor="EBF3FB")
font_h1    = Font(bold=True, color="FFFFFF", size=11)
font_h2    = Font(bold=True, color="FFFFFF", size=9)
font_grp   = Font(bold=True, size=11, color="1F4E79")
font_nucl  = Font(bold=True, size=9, color="375623")
font_sub   = Font(bold=True, size=9)
font_grand = Font(bold=True, color="FFFFFF", size=10)
font_isol  = Font(italic=True, size=9, color="666666")
font_d     = Font(size=9)
ac  = Alignment(horizontal="center", vertical="center", wrap_text=True)
al  = Alignment(horizontal="left",   vertical="center")
thin = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"))

# --------------------------------------------------------------------------- #
#  Colunas (iguais ao TRECHOS_MEGA_CLUSTERS)
# --------------------------------------------------------------------------- #
COLS = [
    "ID_Trecho", "Grupo", "Nucleo", "TAG", "Setor",
    "PV Ini", "PV Fim", "L(m)", "DN", "Mat",
    "CT Ini", "CF Ini", "Prof Ini(m)", "CT Fim", "CF Fim", "Prof Fim(m)",
    "Decl(m/m)", "V(m/s)", "Q(l/s)", "Pavimento",
    "Escav(m3)", "Reaterro(m3)", "Bota-fora(m3)", "Escoram(m2)", "Pavim(m2)",
    "Custo Tubo", "Custo Escav", "Custo PV", "Total s/BDI", "Total c/BDI",
    "DIFIC", "Urgencia", "LIG_AGUA", "LIG_ESG",
]
LARGURAS = [
    10, 24, 32, 8, 18,
    10, 10, 7,  6, 5,
    9,  9,  10, 9, 9,  10,
    9,  7,  7,  13,
    9,  10, 11, 9, 8,
    12, 12, 12, 14, 14,
    6,  10, 9,  9,
]

COLS_RES = [
    "Grupo", "Descricao", "N Nucleos", "N Trechos",
    "LIG_AGUA", "LIG_ESG", "LIG_TOTAL",
    "Ext Total(m)", "Custo Total c/BDI", "VOL_ESCAV(m3)",
    "Setores", "TAGs dos Nucleos",
]
LARG_RES = [22, 36, 10, 10, 10, 10, 10, 14, 22, 14, 40, 70]


# --------------------------------------------------------------------------- #
#  Helpers
# --------------------------------------------------------------------------- #
def _f(v):
    try:
        return float(str(v).replace(",", ".").strip()) if str(v).strip() not in ("", "None", "nan") else 0.0
    except Exception:
        return 0.0

def _i(v):
    try:
        return int(float(str(v).replace(",", ".").strip())) if str(v).strip() not in ("", "None", "nan") else 0
    except Exception:
        return 0

def safe_aba(name, max_len=31):
    for ch in r'\/?*[]:|':
        name = name.replace(ch, "_")
    return name[:max_len]


# --------------------------------------------------------------------------- #
#  Leitura dos GPKGs
# --------------------------------------------------------------------------- #
def ler_gpkg(path):
    """Le GPKG e retorna lista de dicts unicos por TAG."""
    gdf = gpd.read_file(path)
    seen = {}
    for _, row in gdf.iterrows():
        tag = str(row.get("TAG", "")).strip()
        if not tag or tag in ("", "None", "nan"):
            continue
        if tag not in seen:
            seen[tag] = {k: row[k] for k in gdf.columns if k != "geometry"}
    return list(seen.values())


# --------------------------------------------------------------------------- #
#  Inferencia trecho-a-trecho (identica ao gerar_trechos_mega.py)
# --------------------------------------------------------------------------- #
def inferir_trechos(nr, grupo_nome):
    seed = _i(nr.get("TAG", 0)) * 7 + 13
    rng  = random.Random(seed)

    tag       = str(nr.get("TAG", "")).strip()
    nome      = str(nr.get("NOME", "")).strip()
    setor     = str(nr.get("SETOR", "")).strip()
    dific     = max(1, min(5, _i(nr.get("DIFIC", 3)) or 3))
    n_pvs     = _i(nr.get("N_PVS", 0))
    ext_m     = _f(nr.get("EXT_ES_M", 0))
    vol_escav = _f(nr.get("VOL_ESCAV", 0))
    c_red_es  = _f(nr.get("C_RED_ES", 0))
    c_pvs_tot = _f(nr.get("C_PVS", 0))
    lig_agua  = _i(nr.get("LIG_AGUA", 0))
    lig_esg   = _i(nr.get("LIG_ESG", 0))
    lig_total = _i(nr.get("LIG_TOTAL", 0))
    urgencia  = str(nr.get("URGENCIA", "")).strip()
    tipo_ter  = str(nr.get("TIPO_TER", "urbano")).strip().lower()

    # N trechos
    if n_pvs >= 2:
        n_trechos = n_pvs - 1
    elif ext_m > 0:
        n_trechos = max(1, round(ext_m / 60))
    else:
        n_trechos = max(1, round(lig_total / 30))

    # Parametros de terreno
    pav_params = None
    for k, v in TERR.items():
        if k in tipo_ter:
            pav_params = v
            break
    if pav_params is None:
        pav_params = TERR["urbano"]
    pav_str, dn_base, decl_base = pav_params

    L_medio    = ext_m / n_trechos if ext_m > 0 else 60.0
    prof_base  = PROF_DIFIC.get(dific, 1.5)
    custo_tubo_m = c_red_es / ext_m if ext_m > 0 else 200.0
    custo_pv_un  = c_pvs_tot / n_pvs if n_pvs > 0 else 3078.0
    ct_ini     = rng.uniform(3.0, 7.0)

    trechos = []
    for i in range(n_trechos):
        L    = round(L_medio * rng.uniform(0.80, 1.20), 1)
        pr0  = round(prof_base * rng.uniform(0.85, 1.15), 2)
        pr1  = round(prof_base * rng.uniform(0.85, 1.15), 2)
        decl = round(decl_base * rng.uniform(0.75, 1.35), 5)

        if dific >= 4:
            dn = 200
        elif dific >= 3:
            dn = 150 if i < n_trechos * 0.6 else 200
        else:
            dn = 100 if i < n_trechos * 0.5 else 150

        ct0  = round(ct_ini - i * L_medio * decl_base * 0.25, 3)
        cf0  = round(ct0 - pr0, 3)
        ct1  = round(ct0 - L * decl, 3)
        cf1  = round(ct1 - pr1, 3)

        r_hid = (dn / 1000) / 4
        v_ms  = round((1.0 / 0.013) * (r_hid ** (2.0 / 3.0)) * (decl ** 0.5), 3)
        area  = math.pi * ((dn / 1000) / 2) ** 2
        q_ls  = round(v_ms * area * 1000.0, 3)

        prof_med = (pr0 + pr1) / 2
        escav    = round(vol_escav / n_trechos if vol_escav > 0 else L * prof_med, 3)
        reaterro = round(escav * 0.90, 3)
        bota     = round(escav * 0.10, 3)
        escoram  = round(L * prof_med * 2.0, 2)
        pavim    = round(L * 0.80, 2)

        c_tubo = round(L * custo_tubo_m, 2)
        c_esc  = round(escav * 30.0, 2)
        c_pv   = round(custo_pv_un, 2)
        t_sbdi = round(c_tubo + c_esc + c_pv, 2)
        t_cbdi = round(t_sbdi * 1.25, 2)

        trechos.append(dict(
            grupo=grupo_nome, tag=tag, nome=nome, setor=setor,
            pv_ini=f"PV-{tag}-{i+1:03d}",
            pv_fim=f"PV-{tag}-{i+2:03d}",
            L=L, dn=dn, mat="PVC",
            ct0=ct0, cf0=cf0, pr0=pr0,
            ct1=ct1, cf1=cf1, pr1=pr1,
            decl=decl, v_ms=v_ms, q_ls=q_ls, pav=pav_str,
            escav=escav, reaterro=reaterro, bota=bota,
            escoram=escoram, pavim=pavim,
            c_tubo=c_tubo, c_escav=c_esc, c_pv=c_pv,
            t_sbdi=t_sbdi, t_cbdi=t_cbdi,
            dific=dific, urgencia=urgencia,
            lig_agua=lig_agua, lig_esg=lig_esg,
        ))
    return trechos


# --------------------------------------------------------------------------- #
#  Escrita XLSX
# --------------------------------------------------------------------------- #
def escrever_header(ws, titulo):
    nc = len(COLS)
    ws.merge_cells(f"A1:{get_column_letter(nc)}1")
    c = ws.cell(1, 1, titulo)
    c.fill = fill_h1; c.font = font_h1; c.alignment = ac
    ws.row_dimensions[1].height = 22
    for i, col in enumerate(COLS, 1):
        c = ws.cell(2, i, col)
        c.fill = fill_h2; c.font = font_h2; c.alignment = ac; c.border = thin
    ws.row_dimensions[2].height = 30
    for i, w in enumerate(LARGURAS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"


def escrever_trecho(ws, row_idx, tid, grupo_nome, t, zebra=False):
    vals = [
        f"T-{tid:05d}", grupo_nome, t["nome"], t["tag"], t["setor"],
        t["pv_ini"], t["pv_fim"],
        round(t["L"], 1), t["dn"], t["mat"],
        round(t["ct0"], 3), round(t["cf0"], 3), round(t["pr0"], 2),
        round(t["ct1"], 3), round(t["cf1"], 3), round(t["pr1"], 2),
        round(t["decl"], 5), round(t["v_ms"], 3), round(t["q_ls"], 3), t["pav"],
        t["escav"], t["reaterro"], t["bota"], t["escoram"], t["pavim"],
        t["c_tubo"], t["c_escav"], t["c_pv"], t["t_sbdi"], t["t_cbdi"],
        t["dific"], t["urgencia"], t["lig_agua"], t["lig_esg"],
    ]
    LEFT = {3, 5}
    for col, v in enumerate(vals, 1):
        c = ws.cell(row_idx, col, v)
        c.font = font_d
        c.alignment = al if col in LEFT else ac
        c.border = thin
        if zebra:
            c.fill = fill_zebra


def escrever_sep_grupo(ws, row_idx, codigo, descricao, n_nuc, n_tr, fill_=None):
    nc = len(COLS)
    ws.merge_cells(f"A{row_idx}:{get_column_letter(nc)}{row_idx}")
    txt = f"  {codigo} — {descricao}  |  {n_nuc} núcleo(s)  |  {n_tr} trechos"
    c = ws.cell(row_idx, 1, txt)
    c.fill = fill_ or fill_grupo; c.font = font_grp; c.alignment = al
    ws.row_dimensions[row_idx].height = 20


def escrever_sep_nucleo(ws, row_idx, nome, setor, n_tr):
    nc = len(COLS)
    ws.merge_cells(f"A{row_idx}:{get_column_letter(nc)}{row_idx}")
    c = ws.cell(row_idx, 1, f"    {nome}  |  {setor}  |  {n_tr} trechos")
    c.fill = fill_nucl; c.font = font_nucl; c.alignment = al
    ws.row_dimensions[row_idx].height = 16


def escrever_subtotal(ws, row_idx, label, tot):
    nc = len(COLS)
    vals = [label] + [""] * 6 + [round(tot["L"], 1)] + [""] * 11 + [
        round(tot["escav"], 2), round(tot["reaterro"], 2), round(tot["bota"], 2),
        "", round(tot["pavim"], 2),
        round(tot["c_tubo"], 2), round(tot["c_escav"], 2), round(tot["c_pv"], 2),
        round(tot["t_sbdi"], 2), round(tot["t_cbdi"], 2),
        "", "", "", "",
    ]
    while len(vals) < nc:
        vals.append("")
    for col, v in enumerate(vals[:nc], 1):
        c = ws.cell(row_idx, col, v)
        c.fill = fill_sub; c.font = font_sub; c.alignment = ac; c.border = thin


def escrever_grand(ws, row_idx, n_tr, tot):
    nc = len(COLS)
    label = f"GRAND TOTAL — {n_tr} TRECHOS"
    vals = [label] + [""] * 6 + [round(tot["L"], 1)] + [""] * 11 + [
        round(tot["escav"], 2), round(tot["reaterro"], 2), round(tot["bota"], 2),
        "", round(tot["pavim"], 2),
        round(tot["c_tubo"], 2), round(tot["c_escav"], 2), round(tot["c_pv"], 2),
        round(tot["t_sbdi"], 2), round(tot["t_cbdi"], 2),
        "", "", "", "",
    ]
    while len(vals) < nc:
        vals.append("")
    for col, v in enumerate(vals[:nc], 1):
        c = ws.cell(row_idx, col, v)
        c.fill = fill_h1; c.font = font_grand; c.alignment = ac; c.border = thin


def zero_tot():
    return dict(L=0.0, escav=0.0, reaterro=0.0, bota=0.0, pavim=0.0,
                c_tubo=0.0, c_escav=0.0, c_pv=0.0, t_sbdi=0.0, t_cbdi=0.0)

def add_tot(a, t):
    for k in ("L","escav","reaterro","bota","pavim","c_tubo","c_escav","c_pv","t_sbdi","t_cbdi"):
        a[k] += t[k]


def escrever_resumo(ws, grupos_info):
    nc = len(COLS_RES)
    ws.merge_cells(f"A1:{get_column_letter(nc)}1")
    c = ws.cell(1, 1,
        "RESUMO POR GRUPO DE PROXIMIDADE — ConstruData HydroNetwork v7.0.0  |  "
        "FCN Construcoes  |  Contrato 11481051")
    c.fill = fill_h1; c.font = font_h1; c.alignment = ac
    ws.row_dimensions[1].height = 22

    for i, (h, w) in enumerate(zip(COLS_RES, LARG_RES), 1):
        c = ws.cell(2, i, h)
        c.fill = fill_h2; c.font = font_h2; c.alignment = ac; c.border = thin
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "A3"

    for ri, gi in enumerate(grupos_info, 3):
        zebra = (ri % 2 == 0)
        fg = fill_zebra if zebra else None
        vals = [
            gi["codigo"], gi["descricao"], gi["n_nucleos"], gi["n_trechos"],
            gi["lig_agua"], gi["lig_esg"], gi["lig_total"],
            round(gi["ext_m"], 1), round(gi["custo_cbdi"], 2),
            round(gi["vol_escav"], 1), gi["setores"], gi["tags_str"],
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(ri, col, v)
            c.font = font_d
            c.alignment = al if col in (2, 11, 12) else ac
            c.border = thin
            if fg and col not in (2, 11, 12):
                c.fill = fg
        ws.row_dimensions[ri].height = 18


# --------------------------------------------------------------------------- #
#  MAIN
# --------------------------------------------------------------------------- #
print("Lendo grupos de proximidade (RECORTADO*.gpkg)...")

# Fallback CSV
csv_nuc = {}
if os.path.isfile(CSV_NUC):
    with open(CSV_NUC, encoding="utf-8-sig", newline="") as f:
        for r in csv.DictReader(f):
            csv_nuc[r["TAG"]] = r

# Ler cada GPKG e montar grupos
grupos_dados = []  # list of (codigo, aba, descricao, [nuc_rows])
tags_em_grupos = set()

# Nucleos que aparecem em mais de um RECORTADO → prioridade para o grupo maior
# (010454 e 010460 aparecem em R4 e R5 → ficarao em R5, o maior)
PRIORIDADE = {}  # tag -> codigo do grupo com prioridade

for codigo, aba, descricao, gpkg_file in GRUPOS:
    path = os.path.join(PROX_DIR, gpkg_file)
    if not os.path.isfile(path):
        print(f"  [SKIP] {gpkg_file} nao encontrado")
        continue
    rows = ler_gpkg(path)
    print(f"  {codigo} ({gpkg_file}): {len(rows)} nucleos")
    for r in rows:
        t = str(r.get("TAG", "")).strip()
        if t:
            # R5 tem prioridade sobre R4 para tags sobrepostas
            prev = PRIORIDADE.get(t)
            if prev is None or codigo > prev:
                PRIORIDADE[t] = codigo
    grupos_dados.append((codigo, aba, descricao, rows))

# Filtrar sobreposicoes: manter cada tag no grupo de maior prioridade
grupos_limpos = []
for codigo, aba, descricao, rows in grupos_dados:
    rows_ok = [r for r in rows
               if str(r.get("TAG","")).strip() and
               PRIORIDADE.get(str(r.get("TAG","")).strip(), codigo) == codigo]
    for r in rows_ok:
        tags_em_grupos.add(str(r.get("TAG","")).strip())
    grupos_limpos.append((codigo, aba, descricao, rows_ok))
    print(f"  {codigo}: {len(rows_ok)} nucleos (apos remover sobreposicoes)")

# Nucleos isolados (no CSV mas nao em nenhum RECORTADO)
isolados = []
for tag, r in csv_nuc.items():
    if tag not in tags_em_grupos:
        isolados.append(r)
isolados.sort(key=lambda r: str(r.get("SETOR","")) + str(r.get("TAG","")))
print(f"  Isolados (fora de qualquer grupo): {len(isolados)} nucleos")

# --------------------------------------------------------------------------- #
#  Gerar trechos
# --------------------------------------------------------------------------- #
print("\nGerando trechos inferidos...")

grand_tot  = zero_tot()
grand_n_tr = 0
grupos_info = []   # para RESUMO

# Estrutura: list of (codigo, aba, descricao, [(tag, nome, setor, [trechos])], grupo_tot)
all_grupos = []

for codigo, aba, descricao, rows in grupos_limpos:
    gr_tot = zero_tot()
    gr_lig_a = gr_lig_e = gr_lig_t = gr_n_tr = 0
    gr_ext = gr_vol = gr_cbdi = 0.0
    setores_set = set()
    tags_list   = []
    nucleo_list = []

    for nr in rows:
        tag  = str(nr.get("TAG","")).strip()
        nome = str(nr.get("NOME","")).strip()
        setor= str(nr.get("SETOR","")).strip()
        # Se GPKG nao tem todos os campos, completar com CSV
        if tag in csv_nuc:
            for k, v in csv_nuc[tag].items():
                if k not in nr or str(nr[k]).strip() in ("", "None", "nan"):
                    nr[k] = v

        trechos = inferir_trechos(nr, descricao)
        nucleo_list.append((tag, nome, setor, trechos))

        for t in trechos:
            add_tot(gr_tot, t)
            add_tot(grand_tot, t)
        gr_n_tr   += len(trechos)
        grand_n_tr+= len(trechos)
        gr_lig_a  += _i(nr.get("LIG_AGUA", 0))
        gr_lig_e  += _i(nr.get("LIG_ESG", 0))
        gr_lig_t  += _i(nr.get("LIG_TOTAL", 0))
        gr_ext    += _f(nr.get("EXT_ES_M", 0))
        gr_vol    += _f(nr.get("VOL_ESCAV", 0))
        gr_cbdi   += sum(t["t_cbdi"] for t in trechos)
        setores_set.add(setor)
        tags_list.append(tag)

    all_grupos.append((codigo, aba, descricao, nucleo_list, gr_tot, gr_n_tr))
    grupos_info.append(dict(
        codigo=codigo, descricao=descricao,
        n_nucleos=len(nucleo_list), n_trechos=gr_n_tr,
        lig_agua=gr_lig_a, lig_esg=gr_lig_e, lig_total=gr_lig_t,
        ext_m=gr_ext, custo_cbdi=gr_cbdi, vol_escav=gr_vol,
        setores=" | ".join(sorted(setores_set)),
        tags_str=", ".join(tags_list),
    ))

# Trechos dos isolados
iso_nucleo_list = []
iso_tot  = zero_tot()
iso_n_tr = 0
iso_lig_a = iso_lig_e = iso_lig_t = 0
iso_ext = iso_vol = iso_cbdi = 0.0
iso_setores = set()
iso_tags = []

for nr in isolados:
    tag  = str(nr.get("TAG","")).strip()
    nome = str(nr.get("NOME","")).strip()
    setor= str(nr.get("SETOR","")).strip()
    trechos = inferir_trechos(nr, "ISOLADO")
    iso_nucleo_list.append((tag, nome, setor, trechos))
    for t in trechos:
        add_tot(iso_tot, t)
        add_tot(grand_tot, t)
    iso_n_tr   += len(trechos)
    grand_n_tr += len(trechos)
    iso_lig_a  += _i(nr.get("LIG_AGUA", 0))
    iso_lig_e  += _i(nr.get("LIG_ESG", 0))
    iso_lig_t  += _i(nr.get("LIG_TOTAL", 0))
    iso_ext    += _f(nr.get("EXT_ES_M", 0))
    iso_vol    += _f(nr.get("VOL_ESCAV", 0))
    iso_cbdi   += sum(t["t_cbdi"] for t in trechos)
    iso_setores.add(setor)
    iso_tags.append(tag)

grupos_info.append(dict(
    codigo="ISO", descricao="Núcleos fora dos grupos de proximidade (isolados)",
    n_nucleos=len(iso_nucleo_list), n_trechos=iso_n_tr,
    lig_agua=iso_lig_a, lig_esg=iso_lig_e, lig_total=iso_lig_t,
    ext_m=iso_ext, custo_cbdi=iso_cbdi, vol_escav=iso_vol,
    setores=" | ".join(sorted(iso_setores)),
    tags_str=", ".join(iso_tags),
))

print(f"  Total trechos: {grand_n_tr} em {len(all_grupos)} grupos + {len(isolados)} isolados")

# --------------------------------------------------------------------------- #
#  Montar XLSX
# --------------------------------------------------------------------------- #
print("\nCriando XLSX...")
TITULO = ("TRECHOS POR GRUPOS DE PROXIMIDADE REAL — "
          "ConstruData HydroNetwork v7.0.0  |  FCN Construcoes  |  Contrato 11481051")

wb = openpyxl.Workbook()

# Aba RESUMO (primeira)
ws_res = wb.active
ws_res.title = "RESUMO_GRUPOS"
ws_res.sheet_view.showGridLines = False
escrever_resumo(ws_res, grupos_info)

# Aba TODOS_NUCLEOS
TITULO_GERAL = TITULO + "  |  CONSOLIDADO TODOS OS NUCLEOS"
ws_all = wb.create_sheet("TODOS_NUCLEOS")
ws_all.sheet_view.showGridLines = False
escrever_header(ws_all, TITULO_GERAL)
row_all = 3
tid_all = 1

def gravar_grupo_na_aba(ws, row_start, tid_start, codigo, descricao,
                         nucleo_list, gr_tot, gr_n_tr,
                         fill_grp=None):
    """Grava trechos de um grupo em ws a partir de row_start. Retorna (proxima_row, proximo_tid)."""
    rn = row_start
    tid = tid_start

    escrever_sep_grupo(ws, rn, codigo, descricao, len(nucleo_list), gr_n_tr, fill_=fill_grp)
    rn += 1

    gr_cont = zero_tot()
    for tag, nome, setor, trechos in nucleo_list:
        escrever_sep_nucleo(ws, rn, nome, setor, len(trechos))
        rn += 1
        nuc_tot = zero_tot()
        for zi, t in enumerate(trechos):
            escrever_trecho(ws, rn, tid, codigo, t, zebra=(zi % 2 == 0))
            add_tot(nuc_tot, t)
            rn += 1; tid += 1
        escrever_subtotal(ws, rn, f"  Sub {nome[:32]}", nuc_tot)
        add_tot(gr_cont, nuc_tot)
        rn += 1

    escrever_subtotal(ws, rn, f"SUBTOTAL {codigo} — {descricao[:28]}", gr_cont)
    rn += 2
    return rn, tid

# Escrever todos os grupos na aba TODOS
FILLS_GRP = [
    PatternFill("solid", fgColor="D6E4F0"),
    PatternFill("solid", fgColor="E2EFDA"),
    PatternFill("solid", fgColor="FFF2CC"),
    PatternFill("solid", fgColor="FCE4D6"),
    PatternFill("solid", fgColor="E8D5F5"),
    PatternFill("solid", fgColor="D5EBF5"),
    PatternFill("solid", fgColor="F5D5D5"),
    PatternFill("solid", fgColor="D5F5E8"),
]

for gi, (codigo, aba, descricao, nucleo_list, gr_tot, gr_n_tr) in enumerate(all_grupos):
    row_all, tid_all = gravar_grupo_na_aba(
        ws_all, row_all, tid_all, codigo, descricao,
        nucleo_list, gr_tot, gr_n_tr,
        fill_grp=FILLS_GRP[gi % len(FILLS_GRP)])

# Isolados na aba TODOS
row_all, tid_all = gravar_grupo_na_aba(
    ws_all, row_all, tid_all, "ISO",
    "Núcleos fora dos grupos de proximidade (isolados)",
    iso_nucleo_list, iso_tot, iso_n_tr,
    fill_grp=fill_isol)

escrever_grand(ws_all, row_all, grand_n_tr, grand_tot)

# Abas por grupo
for gi, (codigo, aba, descricao, nucleo_list, gr_tot, gr_n_tr) in enumerate(all_grupos):
    ws_g = wb.create_sheet(title=safe_aba(aba))
    ws_g.sheet_view.showGridLines = False
    titulo_g = (f"{codigo} — {descricao}  |  {len(nucleo_list)} núcleo(s)  |  "
                f"{gr_n_tr} trechos  |  ConstruData HydroNetwork v7.0.0")
    escrever_header(ws_g, titulo_g)
    rn, tid = gravar_grupo_na_aba(
        ws_g, 3, 1, codigo, descricao, nucleo_list, gr_tot, gr_n_tr,
        fill_grp=FILLS_GRP[gi % len(FILLS_GRP)])
    # total da aba
    escrever_subtotal(ws_g, rn - 1, f"TOTAL {codigo}", gr_tot)

# Aba ISOLADOS
ws_iso = wb.create_sheet("ISOLADOS")
ws_iso.sheet_view.showGridLines = False
escrever_header(ws_iso,
    f"NUCLEOS ISOLADOS — fora dos grupos de proximidade  |  "
    f"{len(isolados)} nucleos  |  {iso_n_tr} trechos  |  ConstruData HydroNetwork v7.0.0")
rn, tid = gravar_grupo_na_aba(
    ws_iso, 3, 1, "ISO",
    "Fora dos grupos de proximidade geográfica",
    iso_nucleo_list, iso_tot, iso_n_tr, fill_grp=fill_isol)
escrever_subtotal(ws_iso, rn - 1, "TOTAL ISOLADOS", iso_tot)

# Salvar
wb.save(OUT_FILE)
print(f"\nSalvo: {OUT_FILE}")
print(f"Abas ({len(wb.worksheets)}): {[ws.title for ws in wb.worksheets]}")
print(f"\nResumo por grupo:")
for gi in grupos_info:
    print(f"  {gi['codigo']:5s} | {gi['n_nucleos']:3d} nucleos | "
          f"{gi['n_trechos']:5d} trechos | "
          f"Lig A={gi['lig_agua']:5d} | Lig E={gi['lig_esg']:5d} | "
          f"R$ {gi['custo_cbdi']:>14,.2f} c/BDI")
print(f"\nGRAND TOTAL:")
print(f"  Trechos  : {grand_n_tr}")
print(f"  Ext.(m)  : {round(grand_tot['L'], 1)}")
print(f"  c/BDI    : R$ {grand_tot['t_cbdi']:,.2f}")
