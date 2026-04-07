#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ESTATISTICA_ML.PY — XGBoost + GridSearchCV + Feature Importance + Seaborn Violin
ConstruData HydroNetwork v7.0.0 | FCN Construções e Saneamento | Contrato 11481051

Pipeline:
  1. Lê trechos REAIS dos JSONs das NS (813 amostras de treino)
  2. Lê trechos INFERIDOS das 68 áreas novas (5166 amostras para predição)
  3. XGBoost com GridSearchCV para 4 targets: prof_ini, prof_fim, decl_mm, escav_m3
  4. Feature Importance plot (barplot seaborn)
  5. Violin plots das distribuições reais vs preditas
  6. Gera ESTATISTICA.xlsx com predições melhoradas e métricas
"""
import sys, io, os, json, glob, warnings
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
import matplotlib; matplotlib.use("Agg")
import matplotlib.pyplot as plt
import seaborn as sns
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

try:
    from xgboost import XGBRegressor
    HAS_XGB = True
except ImportError:
    from sklearn.ensemble import GradientBoostingRegressor as XGBRegressor
    HAS_XGB = False
    print("[ML] XGBoost não instalado — usando GradientBoostingRegressor")

from sklearn.model_selection import GridSearchCV, cross_val_score
from sklearn.preprocessing import LabelEncoder
from sklearn.metrics import r2_score, mean_absolute_error

BASE_NS = os.path.join(os.environ["USERPROFILE"], "Downloads",
    "PROJETOS DE \u00c1GUA E ESGOTO - DWG E DXF 2018",
    "MAPAS \u00c1GUA E ESGOTO PARA DXF",
    "NOTAS DE SERVICO COMPLETAS 24-03-2026")

INF_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
    "TRECHOS_NS_COMPLETOS", "TRECHOS_NUCLEOS_INFERIDOS.xlsx")
OUT_DIR  = os.path.join(os.path.dirname(os.path.abspath(__file__)),
    "TRECHOS_NS_COMPLETOS")
OUT_XLSX = os.path.join(OUT_DIR, "ESTATISTICA.xlsx")
OUT_FIGS = os.path.join(OUT_DIR, "figs_ml")
os.makedirs(OUT_FIGS, exist_ok=True)

TERRAIN_MAP = {
    "MORRO_DO_TETEU": "morro", "PANTANAL_BAIXO": "dique",
    "SAO_MANOEL": "urbano",   "JOAO_CARLOS": "urbano",
    "VILA_CRIADORES": "urbano","VILA_ISRAEL": "morro",
    "PROL_TETEU": "prolongamento", "PROL_TETEU_ALT-01": "prolongamento",
    "PROL_CRIADORES": "prolongamento", "PROL_PANTANAL_BAIXO": "dique",
    "PROL_SAO_MANOEL": "urbano", "ESTUDO_SM_JC": "urbano",
}

# ════════════════════════════════════════════════════════════
# 1. COLETAR DADOS REAIS (treino)
# ════════════════════════════════════════════════════════════
print("\n[1/5] Coletando trechos REAIS (treino)...")
rows_real = []
for pasta in os.listdir(BASE_NS):
    p = os.path.join(BASE_NS, pasta)
    if not os.path.isdir(p): continue
    terrain = TERRAIN_MAP.get(pasta, "urbano")
    jsons = glob.glob(os.path.join(p, "**", "*.json"), recursive=True)
    for j in jsons:
        try:
            d = json.load(open(j, encoding="utf-8"))
            t = d.get("trecho", {})
            pvm = d.get("pv_montante", {}); pvj = d.get("pv_jusante", {})
            ext   = float(t.get("ext_m") or 0)
            dn    = int(t.get("dn_mm") or 200)
            pr_i  = float(t.get("prof_ini") or pvm.get("prof") or 0)
            pr_f  = float(t.get("prof_fim") or pvj.get("prof") or 0)
            ct0   = float(t.get("ct_ini") or pvm.get("ct") or 0)
            cf0   = float(t.get("cf_ini") or pvm.get("cf") or 0)
            decl  = float(t.get("decl_mm") or 0)
            v_ms  = float(t.get("v_ms") or 0)
            q_ls  = float(t.get("q_ls") or 0)
            if ext <= 0: continue
            prof_m = (pr_i + pr_f) / 2 if pr_i and pr_f else max(pr_i, pr_f)
            escav = ext * max(prof_m, 0.5)
            rows_real.append({
                "ext_m": ext, "dn_mm": dn, "terrain": terrain,
                "ct_ini": ct0, "cf_ini": cf0,
                "prof_ini": pr_i, "prof_fim": pr_f,
                "decl_mm": decl, "v_ms": v_ms, "q_ls": q_ls,
                "escav_m3": escav, "nucleo": pasta, "fonte": "real"
            })
        except: pass

df_real = pd.DataFrame(rows_real)
# Filtrar outliers (prof <= 0 ou >8)
df_real = df_real[(df_real["prof_ini"] > 0.3) & (df_real["prof_ini"] < 8)]
df_real = df_real[(df_real["prof_fim"] > 0.3) & (df_real["prof_fim"] < 8)]
print(f"   {len(df_real)} trechos reais carregados (após filtro)")

# ════════════════════════════════════════════════════════════
# 2. COLETAR TRECHOS INFERIDOS (predição)
# ════════════════════════════════════════════════════════════
print("\n[2/5] Lendo trechos INFERIDOS (5166)...")
wb_inf = openpyxl.load_workbook(INF_FILE, read_only=True)

rows_inf = []
for sh_name in wb_inf.sheetnames:
    if sh_name == "TODOS_68_NUCLEOS": continue
    ws = wb_inf[sh_name]
    tag = sh_name.split("_")[0] if "_" in sh_name else sh_name
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[0] or not str(row[0]).startswith("T-"): continue
        try:
            ext  = float(row[5] or 0)
            dn   = int(row[6] or 200)
            ct0  = float(row[8] or 0)
            cf0  = float(row[9] or 0)
            pr_i = float(row[10] or 0)
            ct1  = float(row[11] or 0)
            pr_f = float(row[13] or 0)
            decl = float(row[14] or 0)
            v_ms = float(row[15] or 0)
            q_ls = float(row[16] or 0)
            perf = str(row[28] or "urbano")
            escav = float(row[18] or 0)
            if ext <= 0: continue
            rows_inf.append({
                "id": row[0], "nucleo_nome": str(row[1] or ""),
                "tag": tag, "pv_ini": str(row[3] or ""),
                "pv_fim": str(row[4] or ""),
                "ext_m": ext, "dn_mm": dn, "terrain": perf,
                "ct_ini": ct0, "cf_ini": cf0,
                "prof_ini": pr_i, "prof_fim": pr_f,
                "decl_mm": decl * 1000, "v_ms": v_ms, "q_ls": q_ls,
                "escav_m3": escav, "fonte": "inferido"
            })
        except: pass

df_inf = pd.DataFrame(rows_inf)
print(f"   {len(df_inf)} trechos inferidos carregados")

wb_inf.close()

# ════════════════════════════════════════════════════════════
# 3. XGBOOST + GRIDSEARCHCV
# ════════════════════════════════════════════════════════════
print("\n[3/5] XGBoost + GridSearchCV...")

le_terrain = LabelEncoder()
all_terrains = list(df_real["terrain"].unique()) + list(df_inf["terrain"].unique())
le_terrain.fit(all_terrains)

df_real["terrain_enc"] = le_terrain.transform(df_real["terrain"])
df_inf["terrain_enc"]  = le_terrain.transform(
    df_inf["terrain"].apply(lambda x: x if x in le_terrain.classes_ else "urbano"))

FEATURES = ["ext_m", "dn_mm", "terrain_enc", "ct_ini"]
TARGETS   = ["prof_ini", "prof_fim", "decl_mm", "escav_m3"]

# Remover linhas com NaN
df_train = df_real[FEATURES + TARGETS].dropna()
X_train = df_train[FEATURES].values
metrics = {}
models  = {}

param_grid = {
    "n_estimators": [100, 200],
    "max_depth":    [3, 5],
    "learning_rate":[0.05, 0.1],
    "subsample":    [0.8, 1.0],
}
if not HAS_XGB:
    param_grid = {"n_estimators": [100, 200], "max_depth": [3, 5], "learning_rate": [0.05, 0.1]}

for target in TARGETS:
    y = df_train[target].values
    if HAS_XGB:
        base = XGBRegressor(random_state=42, tree_method="hist", verbosity=0)
    else:
        base = XGBRegressor(random_state=42)

    gs = GridSearchCV(base, param_grid, cv=3, scoring="r2",
                      n_jobs=-1, verbose=0)
    gs.fit(X_train, y)
    model = gs.best_estimator_

    # Cross-val
    cv_r2 = cross_val_score(model, X_train, y, cv=5, scoring="r2")
    y_pred_train = model.predict(X_train)
    mae  = mean_absolute_error(y, y_pred_train)
    r2   = r2_score(y, y_pred_train)

    models[target]  = model
    metrics[target] = {
        "best_params": gs.best_params_,
        "r2_train": round(r2, 4),
        "mae_train": round(mae, 4),
        "cv_r2_mean": round(cv_r2.mean(), 4),
        "cv_r2_std": round(cv_r2.std(), 4),
    }
    print(f"   {target:12s}  R²={r2:.3f}  MAE={mae:.3f}  CV_R²={cv_r2.mean():.3f}±{cv_r2.std():.3f}  best={gs.best_params_}")

# Predição nos dados inferidos
X_inf = df_inf[FEATURES].fillna(0).values
for target in TARGETS:
    df_inf[f"{target}_ml"] = models[target].predict(X_inf)

# Aplicar clamps físicos
df_inf["prof_ini_ml"]  = df_inf["prof_ini_ml"].clip(0.6, 6.0)
df_inf["prof_fim_ml"]  = df_inf["prof_fim_ml"].clip(0.6, 6.0)
df_inf["decl_mm_ml"]   = df_inf["decl_mm_ml"].clip(2.0, 100.0)
df_inf["escav_m3_ml"]  = df_inf["escav_m3_ml"].clip(0.5, 500.0)

# ════════════════════════════════════════════════════════════
# 4. FEATURE IMPORTANCE
# ════════════════════════════════════════════════════════════
print("\n[4/5] Gerando gráficos...")

fig, axes = plt.subplots(2, 2, figsize=(14, 10), facecolor="#0D0D0D")
fig.suptitle("Feature Importance — XGBoost | ConstruData SLNR Santos",
             color="white", fontsize=14, fontweight="bold")

feat_labels = ["Extensão (m)", "DN (mm)", "Terreno", "CT Ini (m)"]
colors_fi   = ["#00BFFF","#00FF87","#FFD700","#FF6B6B"]

for ax, target in zip(axes.flatten(), TARGETS):
    model = models[target]
    if hasattr(model, "feature_importances_"):
        imp = model.feature_importances_
    else:
        imp = np.ones(len(FEATURES)) / len(FEATURES)
    idx = np.argsort(imp)[::-1]
    ax.set_facecolor("#1A1A1A")
    bars = ax.bar([feat_labels[i] for i in idx],
                  [imp[i] for i in idx],
                  color=[colors_fi[i] for i in idx], edgecolor="white", linewidth=0.5)
    for bar, val in zip(bars, [imp[i] for i in idx]):
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.01,
                f"{val:.3f}", ha="center", color="white", fontsize=8)
    m = metrics[target]
    ax.set_title(f"{target}  |  R²={m['r2_train']}  CV={m['cv_r2_mean']}±{m['cv_r2_std']}",
                 color="white", fontsize=9)
    ax.tick_params(colors="white"); ax.spines[:].set_color("#444")
    ax.set_ylabel("Importance", color="white", fontsize=8)
    ax.set_xlabel("Feature", color="white", fontsize=8)

plt.tight_layout(rect=[0,0,1,0.94])
fi_path = os.path.join(OUT_FIGS, "feature_importance.png")
fig.savefig(fi_path, dpi=130, bbox_inches="tight", facecolor="#0D0D0D")
plt.close(fig)
print(f"   Feature importance salvo: {fi_path}")

# ════════════════════════════════════════════════════════════
# 5. VIOLIN PLOTS — Real vs Inferido vs ML
# ════════════════════════════════════════════════════════════
sns.set_theme(style="darkgrid", palette="muted")
fig2, axes2 = plt.subplots(2, 4, figsize=(20, 10), facecolor="#0D0D0D")
fig2.suptitle("Distribuições: Real vs Inferido vs ML — ConstruData SLNR Santos",
              color="white", fontsize=13, fontweight="bold")

violins = [
    ("prof_ini",    "Prof Ini (m)",    "prof_ini",     "prof_ini_ml"),
    ("prof_fim",    "Prof Fim (m)",    "prof_fim",     "prof_fim_ml"),
    ("decl_mm",     "Decl (mm/m)",     "decl_mm",      "decl_mm_ml"),
    ("escav_m3",    "Escav (m³)",      "escav_m3",     "escav_m3_ml"),
    ("ext_m",       "Extensão (m)",    "ext_m",        None),
    ("dn_mm",       "DN (mm)",         "dn_mm",        None),
    ("ct_ini",      "CT Ini (m)",      "ct_ini",       None),
    ("cf_ini",      "CF Ini (m)",      "cf_ini",       None),
]

for ax, (col, label, col_real, col_ml) in zip(axes2.flatten(), violins):
    ax.set_facecolor("#1A1A1A")
    dfs = []
    if col in df_real.columns:
        tmp = df_real[[col]].dropna().copy(); tmp["Fonte"] = "Real (DXF)"
        dfs.append(tmp.rename(columns={col: "valor"}))
    if col in df_inf.columns:
        tmp = df_inf[[col]].dropna().copy(); tmp["Fonte"] = "Inferido"
        dfs.append(tmp.rename(columns={col: "valor"}))
    if col_ml and col_ml in df_inf.columns:
        tmp = df_inf[[col_ml]].dropna().copy(); tmp["Fonte"] = "ML (XGBoost)"
        dfs.append(tmp.rename(columns={col_ml: "valor"}))

    if dfs:
        df_v = pd.concat(dfs, ignore_index=True)
        # Remover outliers extremos para visualização
        p1, p99 = df_v["valor"].quantile(0.01), df_v["valor"].quantile(0.99)
        df_v = df_v[(df_v["valor"] >= p1) & (df_v["valor"] <= p99)]
        palette = {"Real (DXF)": "#00BFFF", "Inferido": "#FFD700", "ML (XGBoost)": "#00FF87"}
        sns.violinplot(data=df_v, x="Fonte", y="valor", ax=ax,
                       palette=palette, inner="quartile", cut=0, linewidth=0.8)
        sns.stripplot(data=df_v.sample(min(200, len(df_v)), random_state=42),
                      x="Fonte", y="valor", ax=ax,
                      color="white", size=1.5, alpha=0.3, jitter=True)

    ax.set_title(label, color="white", fontsize=9, fontweight="bold")
    ax.set_xlabel(""); ax.set_ylabel("", fontsize=8)
    ax.tick_params(colors="white", labelsize=7)
    ax.spines[:].set_color("#444")
    for label_ in ax.get_xticklabels(): label_.set_color("white")

plt.tight_layout(rect=[0,0,1,0.94])
vio_path = os.path.join(OUT_FIGS, "violin_plots.png")
fig2.savefig(vio_path, dpi=130, bbox_inches="tight", facecolor="#0D0D0D")
plt.close(fig2)
print(f"   Violin plots salvos: {vio_path}")

# ════════════════════════════════════════════════════════════
# 6. GERAR ESTATISTICA.XLSX
# ════════════════════════════════════════════════════════════
print("\n[5/5] Gerando ESTATISTICA.xlsx...")

fill_h1   = PatternFill("solid", fgColor="1F4E79")
fill_h2   = PatternFill("solid", fgColor="2E75B6")
fill_ok   = PatternFill("solid", fgColor="E2EFDA")  # verde = ML
fill_w    = PatternFill("solid", fgColor="FFF2CC")  # amarelo = inferido original
fill_info = PatternFill("solid", fgColor="D6E4F0")
font_h1   = Font(bold=True, color="FFFFFF", size=11)
font_h2   = Font(bold=True, color="FFFFFF", size=9)
font_ok   = Font(size=9, color="1A4D1A")
font_w    = Font(size=9, italic=True, color="7D5A00")
font_d    = Font(size=9)
ac = Alignment(horizontal="center", vertical="center", wrap_text=True)
al = Alignment(horizontal="left",   vertical="center")
thin = Border(left=Side(style="thin"), right=Side(style="thin"),
              top=Side(style="thin"),  bottom=Side(style="thin"))

wb_out = openpyxl.Workbook()

# ─── ABA 1: METRICAS ───
ws_m = wb_out.active
ws_m.title = "METRICAS_ML"
ws_m.column_dimensions["A"].width = 20
ws_m.column_dimensions["B"].width = 14
ws_m.column_dimensions["C"].width = 14
ws_m.column_dimensions["D"].width = 14
ws_m.column_dimensions["E"].width = 16
ws_m.column_dimensions["F"].width = 30

ws_m.merge_cells("A1:F1")
c = ws_m.cell(1,1,"MÉTRICAS XGBoost + GridSearchCV — ConstruData ML v7.0.0")
c.fill=fill_h1; c.font=font_h1; c.alignment=ac

heads = ["Target","R² Treino","MAE Treino","CV R² Médio","CV R² Std","Melhores Parâmetros"]
for i,h in enumerate(heads,1):
    c=ws_m.cell(2,i,h); c.fill=fill_h2; c.font=font_h2; c.alignment=ac; c.border=thin

for r, (target, m) in enumerate(metrics.items(), 3):
    vals = [target, m["r2_train"], m["mae_train"],
            m["cv_r2_mean"], m["cv_r2_std"],
            str(m["best_params"])]
    for i, v in enumerate(vals, 1):
        c=ws_m.cell(r,i,v)
        c.fill=fill_ok; c.font=font_ok; c.alignment=ac; c.border=thin

# Info modelo
ws_m.cell(8,1,"Algoritmo").font=font_h2
ws_m.cell(8,2,"XGBoost" if HAS_XGB else "GradientBoosting").font=font_d
ws_m.cell(9,1,"Features").font=font_h2
ws_m.cell(9,2," | ".join(feat_labels)).font=font_d
ws_m.cell(10,1,"Amostras Treino").font=font_h2
ws_m.cell(10,2,len(df_train)).font=font_d
ws_m.cell(11,1,"Amostras Pred.").font=font_h2
ws_m.cell(11,2,len(df_inf)).font=font_d

# ─── ABA 2: PREDIÇÕES ML ───
ws_p = wb_out.create_sheet("PREDICOES_ML")
COLS_P = ["ID","Nucleo","TAG","PV Ini","PV Fim","L (m)","DN",
          "CT Ini","Terrain",
          "Prof Ini Orig","Prof Ini ML","Prof Fim Orig","Prof Fim ML",
          "Decl Orig (mm)","Decl ML (mm)","Escav Orig","Escav ML",
          "Delta Prof Ini","Delta Escav","Custo c/BDI (ML)"]
LARG_P = [9,22,7,12,12,7,6,9,12,11,10,11,10,12,11,11,11,12,12,14]

ws_p.merge_cells(f"A1:{get_column_letter(len(COLS_P))}1")
c = ws_p.cell(1,1,"PREDIÇÕES ML — XGBoost vs Inferência Original")
c.fill=fill_h1; c.font=font_h1; c.alignment=ac
for i,col in enumerate(COLS_P,1):
    c=ws_p.cell(2,i,col); c.fill=fill_h2; c.font=font_h2; c.alignment=ac; c.border=thin
for i,w in enumerate(LARG_P,1):
    ws_p.column_dimensions[get_column_letter(i)].width=w
ws_p.freeze_panes="A3"

for r_idx, (_, row) in enumerate(df_inf.iterrows(), 3):
    delta_prof = round(float(row.get("prof_ini_ml",0)) - float(row.get("prof_ini",0)), 3)
    delta_escav= round(float(row.get("escav_m3_ml",0)) - float(row.get("escav_m3",0)), 2)
    prof_med_ml= (float(row.get("prof_ini_ml",0)) + float(row.get("prof_fim_ml",0)))/2
    escav_ml   = round(float(row.get("ext_m",0)) * prof_med_ml, 3)
    c_tubo = float(row.get("ext_m",0)) * {150:160,200:200,250:270,300:340}.get(int(row.get("dn_mm",200)), 200)
    c_escav_ml = escav_ml * 30
    total_cbdi = round((c_tubo + c_escav_ml + 3078) * 1.25, 2)

    vals = [
        str(row.get("id","")), str(row.get("nucleo_nome","")), str(row.get("tag","")),
        str(row.get("pv_ini","")), str(row.get("pv_fim","")),
        round(float(row.get("ext_m",0)),2), int(row.get("dn_mm",200)),
        round(float(row.get("ct_ini",0)),3), str(row.get("terrain","")),
        round(float(row.get("prof_ini",0)),2), round(float(row.get("prof_ini_ml",0)),3),
        round(float(row.get("prof_fim",0)),2), round(float(row.get("prof_fim_ml",0)),3),
        round(float(row.get("decl_mm",0)),2), round(float(row.get("decl_mm_ml",0)),2),
        round(float(row.get("escav_m3",0)),3), round(escav_ml,3),
        delta_prof, delta_escav, total_cbdi
    ]
    zebra = (r_idx % 2 == 0)
    for i, v in enumerate(vals, 1):
        c=ws_p.cell(r_idx, i, v)
        # Colorir col ML vs orig
        if i in [11,13,15,17]:
            c.fill=fill_ok; c.font=font_ok
        elif i in [10,12,14,16]:
            c.fill=fill_w; c.font=font_w
        else:
            c.font=font_d
            if zebra: c.fill=PatternFill("solid",fgColor="EBF3FB")
        c.alignment=ac; c.border=thin

# ─── ABA 3: ESTATÍSTICAS DESCRITIVAS ───
ws_e = wb_out.create_sheet("ESTATISTICAS")
ws_e.merge_cells("A1:J1")
c = ws_e.cell(1,1,"ESTATÍSTICAS DESCRITIVAS — Real vs Inferido vs ML")
c.fill=fill_h1; c.font=font_h1; c.alignment=ac

STATS_VARS = ["ext_m","dn_mm","prof_ini","prof_fim","decl_mm","escav_m3","ct_ini"]
STATS_LABELS= ["Extensão (m)","DN (mm)","Prof Ini (m)","Prof Fim (m)",
               "Declive (mm/m)","Escavação (m³)","CT Ini (m)"]
stat_fns = [("count",lambda x: len(x)), ("mean",np.mean), ("std",np.std),
            ("min",np.min), ("p25",lambda x: np.percentile(x,25)),
            ("median",np.median), ("p75",lambda x: np.percentile(x,75)),
            ("max",np.max), ("CV%",lambda x: np.std(x)/np.mean(x)*100 if np.mean(x)!=0 else 0)]

heads_e = ["Variável","Fonte"] + [s[0] for s in stat_fns]
for i,h in enumerate(heads_e,1):
    c=ws_e.cell(2,i,h); c.fill=fill_h2; c.font=font_h2; c.alignment=ac; c.border=thin
    ws_e.column_dimensions[get_column_letter(i)].width = 12 if i>2 else 22

row_e = 3
for var, lbl in zip(STATS_VARS, STATS_LABELS):
    for fonte, df_src, ml_suffix in [("Real (DXF)", df_real, None),
                                      ("Inferido",  df_inf, None),
                                      ("ML (XGBoost)", df_inf, f"{var}_ml")]:
        col = ml_suffix if ml_suffix and ml_suffix in df_src.columns else var
        if col not in df_src.columns: continue
        data = df_src[col].dropna().values
        if len(data) == 0: continue
        data = data[(data > -999) & (data < 9999)]
        vals_row = [lbl if fonte=="Real (DXF)" else "", fonte]
        for _, fn in stat_fns:
            try: vals_row.append(round(fn(data), 3))
            except: vals_row.append("")
        for i, v in enumerate(vals_row, 1):
            c=ws_e.cell(row_e,i,v)
            if fonte=="Real (DXF)": c.fill=PatternFill("solid",fgColor="DDEEFF")
            elif fonte=="Inferido": c.fill=fill_w
            else: c.fill=fill_ok
            c.font=font_d; c.alignment=ac; c.border=thin
        row_e += 1
    row_e += 1  # separador

# ─── ABA 4: IMAGENS ───
ws_img = wb_out.create_sheet("GRAFICOS_ML")
ws_img.merge_cells("A1:M1")
c = ws_img.cell(1,1,"GRÁFICOS — Feature Importance & Violin Plots")
c.fill=fill_h1; c.font=font_h1; c.alignment=ac

try:
    img1 = XLImage(fi_path)
    img1.width = 800; img1.height = 520
    ws_img.add_image(img1, "A2")
    img2 = XLImage(vio_path)
    img2.width = 900; img2.height = 500
    ws_img.add_image(img2, "A32")
except Exception as e:
    ws_img.cell(2,1,f"Erro ao inserir imagens: {e}")

wb_out.save(OUT_XLSX)
print(f"\nSalvo: {OUT_XLSX}")
print(f"Abas: {[ws.title for ws in wb_out.worksheets]}")
print(f"\nMétricas finais:")
for t, m in metrics.items():
    print(f"  {t:12s}  R²={m['r2_train']}  CV={m['cv_r2_mean']}±{m['cv_r2_std']}")
print(f"\nFiguras em: {OUT_FIGS}")
print("Concluido!")
