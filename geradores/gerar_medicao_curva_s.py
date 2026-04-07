#!/usr/bin/env python3
"""
gerar_medicao_curva_s.py — Medição Mensal + Curva S + Diário de Obras
ConstruData SABESP v5.0 — Contrato 11481051

Gera:
  1. MEDICAO_MENSAL.html — Medição por NS com status
  2. CURVA_S.html — Curva S físico-financeiro (previsto × realizado)
  3. DIARIO_OBRAS.html — Diário de obras integrado ao GIS (Leaflet)
  4. MEDICAO.xlsx — Excel com 3 abas
  5. MEDICAO.json — Dados estruturados

Uso:
    from gerar_medicao_curva_s import gerar_medicao_curva_s
    gerar_medicao_curva_s(pvs, trechos, nucleo, pasta_saida)
"""

import json
import math
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path


def _dias_uteis_entre(d1, d2):
    """Conta dias úteis entre duas datas."""
    count = 0
    d = d1
    while d <= d2:
        if d.weekday() < 5:
            count += 1
        d += timedelta(days=1)
    return count


def gerar_medicao_curva_s(pvs, trechos, nucleo, pasta_saida,
                           data_inicio="2026-04-01", equipes=4, prod_m_dia=6.0):
    """
    Gera medição mensal, curva S e diário de obras.
    """
    pasta = Path(pasta_saida)
    pasta.mkdir(parents=True, exist_ok=True)

    dt_ini = datetime.strptime(data_inicio, "%Y-%m-%d")
    custo_metro = 910  # R$/m contrato

    # ── Calcular cronograma por NS ────────────────────────────────
    equipe_livre = [dt_ini] * equipes
    ns_plan = []

    for i, tr in enumerate(trechos):
        ext = tr.get("ext_m", 0) or 0
        dias = max(1, math.ceil(ext / prod_m_dia))

        eq_idx = equipe_livre.index(min(equipe_livre))
        inicio = equipe_livre[eq_idx]
        fim = inicio
        count = 0
        while count < dias:
            fim += timedelta(days=1)
            if fim.weekday() < 5:
                count += 1
        equipe_livre[eq_idx] = fim

        custo = round(ext * custo_metro, 2)
        ns_plan.append({
            "ns_id": i + 1,
            "nome": f"NS_{i+1:03d}",
            "pv_ini": tr.get("pv_ini", "?"),
            "pv_fim": tr.get("pv_fim", "?"),
            "ext_m": round(ext, 1),
            "dn_mm": tr.get("dn_mm", 200),
            "rua": tr.get("rua", "Sem Rua"),
            "custo": custo,
            "inicio": inicio,
            "fim": fim,
            "equipe": eq_idx + 1,
            "dias": dias,
            "status": "PLANEJADO",  # PLANEJADO / EM_EXEC / CONCLUIDO / CADASTRADO
            "pct_exec": 0,
        })

    # ── Distribuir por mês (Curva S prevista) ─────────────────────
    if ns_plan:
        dt_min = min(n["inicio"] for n in ns_plan)
        dt_max = max(n["fim"] for n in ns_plan)
    else:
        dt_min = dt_max = dt_ini

    meses = []
    d = datetime(dt_min.year, dt_min.month, 1)
    while d <= dt_max + timedelta(days=31):
        mes_fim = datetime(d.year + (d.month // 12), (d.month % 12) + 1, 1) - timedelta(days=1)

        ext_mes = 0
        custo_mes = 0
        ns_mes = 0
        for n in ns_plan:
            # Proporção do trecho neste mês
            overlap_ini = max(n["inicio"], d)
            overlap_fim = min(n["fim"], mes_fim)
            if overlap_ini <= overlap_fim:
                dias_overlap = _dias_uteis_entre(overlap_ini, overlap_fim)
                dias_total = max(1, n["dias"])
                frac = dias_overlap / dias_total
                ext_mes += n["ext_m"] * frac
                custo_mes += n["custo"] * frac
                if frac > 0:
                    ns_mes += 1

        meses.append({
            "mes": d.strftime("%Y-%m"),
            "mes_label": d.strftime("%b/%Y"),
            "ext_previsto": round(ext_mes, 1),
            "custo_previsto": round(custo_mes, 2),
            "ns_previstas": ns_mes,
            "ext_realizado": 0,  # Será preenchido com dados reais
            "custo_realizado": 0,
        })

        d = datetime(d.year + (d.month // 12), (d.month % 12) + 1, 1)
        if d > dt_max + timedelta(days=60):
            break

    # Acumulados
    ext_total = sum(n["ext_m"] for n in ns_plan)
    custo_total = sum(n["custo"] for n in ns_plan)
    ext_acum = 0
    custo_acum = 0
    for m in meses:
        ext_acum += m["ext_previsto"]
        custo_acum += m["custo_previsto"]
        m["ext_acum_prev"] = round(ext_acum, 1)
        m["custo_acum_prev"] = round(custo_acum, 2)
        m["pct_fisico_prev"] = round(ext_acum / ext_total * 100, 1) if ext_total else 0
        m["pct_financ_prev"] = round(custo_acum / custo_total * 100, 1) if custo_total else 0

    stats = {
        "nucleo": nucleo,
        "total_ns": len(ns_plan),
        "ext_total": round(ext_total, 1),
        "custo_total": round(custo_total, 2),
        "equipes": equipes,
        "data_inicio": data_inicio,
        "data_fim": dt_max.strftime("%Y-%m-%d") if ns_plan else data_inicio,
        "meses_total": len(meses),
    }

    # ── Gerar HTMLs ───────────────────────────────────────────────
    _gerar_medicao_html(ns_plan, stats, pasta / "MEDICAO_MENSAL.html")
    _gerar_curva_s_html(meses, stats, pasta / "CURVA_S.html")
    _gerar_diario_html(ns_plan, pvs, stats, pasta / "DIARIO_OBRAS.html")

    # ── Gerar Excel ───────────────────────────────────────────────
    _gerar_xlsx(ns_plan, meses, stats, pasta / "MEDICAO.xlsx")

    # ── Gerar JSON ────────────────────────────────────────────────
    dados = {
        "stats": stats,
        "ns_plan": [{**n, "inicio": n["inicio"].strftime("%Y-%m-%d"),
                     "fim": n["fim"].strftime("%Y-%m-%d")} for n in ns_plan],
        "curva_s": meses,
    }
    with open(pasta / "MEDICAO.json", "w", encoding="utf-8") as f:
        json.dump(dados, f, indent=2, ensure_ascii=False)

    print(f"[MEDIÇÃO] {len(ns_plan)} NS | {len(meses)} meses | "
          f"R$ {custo_total:,.0f} | {ext_total:,.0f}m")

    return dados


def _gerar_medicao_html(ns_plan, stats, out_path):
    """HTML de medição mensal por NS."""
    rows = ""
    for n in ns_plan:
        status_cor = {"PLANEJADO": "#999", "EM_EXEC": "#F57F17",
                      "CONCLUIDO": "#2E7D32", "CADASTRADO": "#1565C0"}
        cor = status_cor.get(n["status"], "#999")
        rows += f"""<tr>
          <td style="font-weight:600">{n['nome']}</td>
          <td>{n['pv_ini']} → {n['pv_fim']}</td>
          <td class="r">{n['ext_m']:.0f}m</td>
          <td class="r">DN{n['dn_mm']}</td>
          <td class="r">Eq.{n['equipe']}</td>
          <td class="r">{n['inicio'].strftime('%d/%m')}</td>
          <td class="r">{n['fim'].strftime('%d/%m')}</td>
          <td class="r mono">R$ {n['custo']:,.0f}</td>
          <td class="c"><span style="color:{cor};font-weight:700">{n['status']}</span></td>
          <td class="c">{n['pct_exec']}%</td>
        </tr>"""

    html = f"""<!DOCTYPE html>
<html lang="pt-BR"><head><meta charset="UTF-8">
<title>Medição Mensal — {stats['nucleo']}</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:Arial,sans-serif;background:#f5f5f5;padding:16px}}
.hdr{{background:linear-gradient(135deg,#1565C0,#42A5F5);color:#fff;padding:20px;border-radius:12px;margin-bottom:16px}}
.hdr h1{{font-size:20px}}.hdr p{{opacity:.85;font-size:13px;margin-top:4px}}
.kpis{{display:grid;grid-template-columns:repeat(5,1fr);gap:10px;margin-bottom:16px}}
.kpi{{background:#fff;border-radius:10px;padding:12px;text-align:center}}
.kpi .l{{font-size:11px;color:#666;font-weight:600}}.kpi .v{{font-size:18px;font-weight:700;color:#1565C0}}
table{{width:100%;border-collapse:collapse;font-size:12px;background:#fff;border-radius:10px;overflow:hidden}}
thead tr{{background:#1565C0;color:#fff}}
th,td{{padding:7px 8px;border-bottom:1px solid #f0f0f0}}
th{{text-align:left;font-size:11px}}.r{{text-align:right}}.c{{text-align:center}}.mono{{font-family:monospace}}
tr:hover{{background:#e3f2fd}}
.footer{{margin-top:12px;font-size:11px;color:#999;text-align:center}}
</style></head><body>
<div class="hdr"><h1>Medição por Nota de Serviço — {stats['nucleo']}</h1>
<p>FCN Construções e Saneamento · Contrato 11481051 · {stats['total_ns']} NS</p></div>
<div class="kpis">
<div class="kpi"><div class="l">Total NS</div><div class="v">{stats['total_ns']}</div></div>
<div class="kpi"><div class="l">Extensão</div><div class="v">{stats['ext_total']:,.0f}m</div></div>
<div class="kpi"><div class="l">Custo Total</div><div class="v">R$ {stats['custo_total']:,.0f}</div></div>
<div class="kpi"><div class="l">Equipes</div><div class="v">{stats['equipes']}</div></div>
<div class="kpi"><div class="l">Período</div><div class="v">{stats['meses_total']} meses</div></div>
</div>
<table>
<thead><tr><th>NS</th><th>Trecho</th><th class="r">Ext</th><th class="r">DN</th><th class="r">Eq</th>
<th class="r">Início</th><th class="r">Fim</th><th class="r">Custo</th><th class="c">Status</th><th class="c">% Exec</th></tr></thead>
<tbody>{rows}</tbody>
</table>
<div class="footer">Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} · ConstruData SABESP v5.0</div>
</body></html>"""

    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)


def _gerar_curva_s_html(meses, stats, out_path):
    """HTML Curva S com Chart.js."""
    labels = [m["mes_label"] for m in meses]
    pct_fisico = [m["pct_fisico_prev"] for m in meses]
    pct_financ = [m["pct_financ_prev"] for m in meses]
    custo_mensal = [m["custo_previsto"] for m in meses]

    html = f"""<!DOCTYPE html>
<html lang="pt-BR"><head><meta charset="UTF-8">
<title>Curva S — {stats['nucleo']}</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.2/dist/chart.umd.min.js"></script>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:Arial,sans-serif;background:#f5f5f5;padding:16px}}
.hdr{{background:linear-gradient(135deg,#2E7D32,#66BB6A);color:#fff;padding:20px;border-radius:12px;margin-bottom:16px}}
.hdr h1{{font-size:20px}}.hdr p{{opacity:.85;font-size:13px;margin-top:4px}}
.grid{{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:16px}}
.card{{background:#fff;border-radius:10px;padding:16px}}
.card h3{{font-size:14px;margin-bottom:12px;color:#333}}
.kpis{{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:16px}}
.kpi{{background:#fff;border-radius:10px;padding:12px;text-align:center}}
.kpi .l{{font-size:11px;color:#666;font-weight:600}}.kpi .v{{font-size:18px;font-weight:700;color:#2E7D32}}
table{{width:100%;border-collapse:collapse;font-size:12px;background:#fff;border-radius:10px;overflow:hidden}}
thead tr{{background:#2E7D32;color:#fff}}
th,td{{padding:7px 8px;border-bottom:1px solid #f0f0f0}}
th{{text-align:left;font-size:11px}}.r{{text-align:right}}.mono{{font-family:monospace}}
.footer{{margin-top:12px;font-size:11px;color:#999;text-align:center}}
</style></head><body>
<div class="hdr"><h1>Curva S — Físico-Financeiro — {stats['nucleo']}</h1>
<p>FCN Construções e Saneamento · Contrato 11481051 · Previsto × Realizado</p></div>
<div class="kpis">
<div class="kpi"><div class="l">Extensão Total</div><div class="v">{stats['ext_total']:,.0f}m</div></div>
<div class="kpi"><div class="l">Custo Total</div><div class="v">R$ {stats['custo_total']:,.0f}</div></div>
<div class="kpi"><div class="l">Período</div><div class="v">{stats['meses_total']} meses</div></div>
<div class="kpi"><div class="l">R$/mês médio</div><div class="v">R$ {stats['custo_total']/max(1,stats['meses_total']):,.0f}</div></div>
</div>
<div class="grid">
<div class="card"><h3>Curva S — % Acumulado</h3><canvas id="curvaS" height="250"></canvas></div>
<div class="card"><h3>Desembolso Mensal (R$)</h3><canvas id="mensal" height="250"></canvas></div>
</div>
<table>
<thead><tr><th>Mês</th><th class="r">Ext Previsto (m)</th><th class="r">Ext Acum (m)</th>
<th class="r">Custo Previsto (R$)</th><th class="r">Custo Acum (R$)</th>
<th class="r">% Físico</th><th class="r">% Financeiro</th><th class="r">NS Previstas</th></tr></thead>
<tbody>{''.join(f"""<tr><td>{m['mes_label']}</td>
<td class="r mono">{m['ext_previsto']:,.0f}</td><td class="r mono">{m['ext_acum_prev']:,.0f}</td>
<td class="r mono">R$ {m['custo_previsto']:,.0f}</td><td class="r mono">R$ {m['custo_acum_prev']:,.0f}</td>
<td class="r">{m['pct_fisico_prev']:.1f}%</td><td class="r">{m['pct_financ_prev']:.1f}%</td>
<td class="r">{m['ns_previstas']}</td></tr>""" for m in meses)}</tbody>
</table>
<div class="footer">Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} · ConstruData SABESP v5.0</div>
<script>
new Chart(document.getElementById('curvaS'),{{type:'line',data:{{labels:{json.dumps(labels)},datasets:[
{{label:'% Físico (previsto)',data:{json.dumps(pct_fisico)},borderColor:'#2E7D32',backgroundColor:'rgba(46,125,50,0.1)',fill:true,tension:0.3}},
{{label:'% Financeiro (previsto)',data:{json.dumps(pct_financ)},borderColor:'#1565C0',borderDash:[5,5],tension:0.3}}
]}},options:{{scales:{{y:{{max:100,title:{{display:true,text:'%'}}}}}}}}}});
new Chart(document.getElementById('mensal'),{{type:'bar',data:{{labels:{json.dumps(labels)},datasets:[
{{label:'Custo Mensal (R$)',data:{json.dumps(custo_mensal)},backgroundColor:'#66BB6A'}}
]}},options:{{plugins:{{legend:{{display:false}}}}}}}});
</script>
</body></html>"""

    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)


def _gerar_diario_html(ns_plan, pvs, stats, out_path):
    """HTML Diário de Obras com mapa Leaflet."""
    # Preparar markers para PVs com coords
    markers_js = ""
    lines_js = ""
    valid_count = 0

    for n in ns_plan:
        pv_a = pvs.get(n["pv_ini"], {})
        pv_b = pvs.get(n["pv_fim"], {})
        xa, ya = pv_a.get("x", 0), pv_a.get("y", 0)
        xb, yb = pv_b.get("x", 0), pv_b.get("y", 0)

        if xa > 100000 and ya > 1000000:
            # Converter UTM → lat/lon (aproximação Santos)
            lat_a = -23.96 + (ya - 7353000) / 111000
            lon_a = -46.33 + (xa - 359000) / 95000
            lat_b = -23.96 + (yb - 7353000) / 111000
            lon_b = -46.33 + (xb - 359000) / 95000

            if -25 < lat_a < -22 and -47 < lon_a < -45:
                status_cor = {"PLANEJADO": "gray", "EM_EXEC": "orange",
                              "CONCLUIDO": "green", "CADASTRADO": "blue"}
                cor = status_cor.get(n["status"], "gray")
                lines_js += f"L.polyline([[{lat_a:.6f},{lon_a:.6f}],[{lat_b:.6f},{lon_b:.6f}]],{{color:'{cor}',weight:4,opacity:0.8}}).addTo(map).bindPopup('<b>{n['nome']}</b><br>{n['pv_ini']}→{n['pv_fim']}<br>DN{n['dn_mm']} | {n['ext_m']:.0f}m<br>Status: {n['status']}<br>Eq.{n['equipe']} | {n['inicio'].strftime('%d/%m')}—{n['fim'].strftime('%d/%m')}<br>R$ {n['custo']:,.0f}');\n"
                valid_count += 1

    html = f"""<!DOCTYPE html>
<html lang="pt-BR"><head><meta charset="UTF-8">
<title>Diário de Obras — {stats['nucleo']}</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:Arial,sans-serif}}
.hdr{{background:linear-gradient(135deg,#E65100,#FF9800);color:#fff;padding:16px 20px;display:flex;justify-content:space-between;align-items:center}}
.hdr h1{{font-size:18px}}.hdr p{{font-size:12px;opacity:.85}}
#map{{height:calc(100vh - 130px);width:100%}}
.info{{display:flex;gap:16px;padding:10px 20px;background:#fff;border-bottom:1px solid #e0e0e0;font-size:12px}}
.info b{{color:#E65100}}
.legend{{position:absolute;bottom:30px;left:10px;background:#fff;padding:10px;border-radius:8px;z-index:1000;font-size:11px;box-shadow:0 2px 6px rgba(0,0,0,.2)}}
.legend div{{margin:3px 0;display:flex;align-items:center;gap:6px}}
.legend span{{width:20px;height:4px;border-radius:2px;display:inline-block}}
</style></head><body>
<div class="hdr">
<div><h1>Diário de Obras — {stats['nucleo']}</h1>
<p>FCN Construções e Saneamento · Contrato 11481051</p></div>
<div style="text-align:right"><p><b>{stats['total_ns']}</b> NS | <b>{stats['ext_total']:,.0f}m</b> | <b>R$ {stats['custo_total']:,.0f}</b></p>
<p>{stats['data_inicio']} a {stats['data_fim']}</p></div>
</div>
<div class="info">
<span><b>{valid_count}</b> trechos no mapa</span>
<span><b>{stats['equipes']}</b> equipes</span>
<span><b>{stats['meses_total']}</b> meses</span>
<span>Gerado: {datetime.now().strftime('%d/%m/%Y %H:%M')}</span>
</div>
<div id="map"></div>
<div class="legend">
<b>STATUS</b>
<div><span style="background:gray"></span> Planejado</div>
<div><span style="background:orange"></span> Em Execução</div>
<div><span style="background:green"></span> Concluído</div>
<div><span style="background:blue"></span> Cadastrado</div>
</div>
<script>
var map=L.map('map').setView([-23.96,-46.33],14);
L.tileLayer('https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{{z}}/{{y}}/{{x}}',{{maxZoom:19}}).addTo(map);
L.tileLayer('https://{{s}}.tile.openstreetmap.org/{{z}}/{{x}}/{{y}}.png',{{maxZoom:19,opacity:0.4}}).addTo(map);
{lines_js}
</script>
</body></html>"""

    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)


def _gerar_xlsx(ns_plan, meses, stats, out_path):
    """Excel com 3 abas: Medição + Curva S + Resumo."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return

    wb = openpyxl.Workbook()

    # ── Aba 1: MEDIÇÃO POR NS ────────────────────────────────────
    ws = wb.active
    ws.title = "MEDICAO_POR_NS"
    hdr = PatternFill("solid", fgColor="1565C0")
    hf = Font(bold=True, color="FFFFFF", size=10)

    headers = ["NS", "PV Ini", "PV Fim", "Ext (m)", "DN", "Rua",
               "Equipe", "Início", "Fim", "Dias", "Custo (R$)", "Status", "% Exec"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hdr
        cell.font = hf

    for i, n in enumerate(ns_plan, 2):
        ws.cell(row=i, column=1, value=n["nome"])
        ws.cell(row=i, column=2, value=n["pv_ini"])
        ws.cell(row=i, column=3, value=n["pv_fim"])
        ws.cell(row=i, column=4, value=n["ext_m"])
        ws.cell(row=i, column=5, value=n["dn_mm"])
        ws.cell(row=i, column=6, value=n["rua"])
        ws.cell(row=i, column=7, value=n["equipe"])
        ws.cell(row=i, column=8, value=n["inicio"].strftime("%d/%m/%Y"))
        ws.cell(row=i, column=9, value=n["fim"].strftime("%d/%m/%Y"))
        ws.cell(row=i, column=10, value=n["dias"])
        ws.cell(row=i, column=11, value=n["custo"])
        ws.cell(row=i, column=12, value=n["status"])
        ws.cell(row=i, column=13, value=n["pct_exec"])

    for c in range(1, 14):
        ws.column_dimensions[chr(64 + c)].width = 14

    # ── Aba 2: CURVA S ───────────────────────────────────────────
    ws2 = wb.create_sheet("CURVA_S")
    hdr2 = PatternFill("solid", fgColor="2E7D32")
    headers2 = ["Mês", "Ext Prev (m)", "Ext Acum (m)", "Custo Prev (R$)",
                "Custo Acum (R$)", "% Físico", "% Financeiro", "NS Prev"]
    for c, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.fill = hdr2
        cell.font = hf

    for i, m in enumerate(meses, 2):
        ws2.cell(row=i, column=1, value=m["mes_label"])
        ws2.cell(row=i, column=2, value=m["ext_previsto"])
        ws2.cell(row=i, column=3, value=m["ext_acum_prev"])
        ws2.cell(row=i, column=4, value=m["custo_previsto"])
        ws2.cell(row=i, column=5, value=m["custo_acum_prev"])
        ws2.cell(row=i, column=6, value=m["pct_fisico_prev"])
        ws2.cell(row=i, column=7, value=m["pct_financ_prev"])
        ws2.cell(row=i, column=8, value=m["ns_previstas"])

    for c in range(1, 9):
        ws2.column_dimensions[chr(64 + c)].width = 16

    # ── Aba 3: RESUMO ────────────────────────────────────────────
    ws3 = wb.create_sheet("RESUMO")
    ws3.cell(row=1, column=1, value="RESUMO DO PLANEJAMENTO").font = Font(bold=True, size=14)
    dados_resumo = [
        ("Núcleo", stats["nucleo"]),
        ("Total NS", stats["total_ns"]),
        ("Extensão Total (m)", stats["ext_total"]),
        ("Custo Total (R$)", stats["custo_total"]),
        ("Equipes", stats["equipes"]),
        ("Data Início", stats["data_inicio"]),
        ("Data Fim", stats["data_fim"]),
        ("Meses", stats["meses_total"]),
        ("R$/mês médio", round(stats["custo_total"] / max(1, stats["meses_total"]), 0)),
        ("Produtividade", "6.0 m/dia/equipe"),
        ("Custo/metro", "R$ 910,00"),
    ]
    for i, (k, v) in enumerate(dados_resumo, 3):
        ws3.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws3.cell(row=i, column=2, value=v)

    wb.save(out_path)


# ═══════════════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import sys
    json_path = sys.argv[1] if len(sys.argv) > 1 else "_tmp_dwg/REDE_EXTRAIDA_BIM.json"

    if not Path(json_path).exists():
        print(f"Uso: python gerar_medicao_curva_s.py <rede.json>")
        sys.exit(1)

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    gerar_medicao_curva_s(
        data.get("pvs", {}),
        data.get("trechos", []),
        data.get("arquivo", "Teste"),
        "SAIDA_BIM_SABESP/MEDICAO",
    )
