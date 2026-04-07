#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
gerar_ose.py — Ordem de Servico para Gabarito (OSE) formato SABESP
ConstruData - HydroNetwork · FCN Construcoes e Saneamento

Layout identico ao template ProSaneamento OSE-Modelo_1 (NS_017rev1).
Colunas intercaladas (par=dado, impar=espacador):
  B=TRECHO, D=ESTACA INT, F=ESTACA FRAC, H=DIST PARCIAL, J=DIST ACUM,
  L=CT, N=I, P=CP(CF), R=CR, T=DN, V=G, X=H, Z=P,
  AB=PV NOME, AD=PV TIPO, AF=PV PROF, AH=OBSERVACOES
"""

import os, math
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter


def _fmtv(val, fmt=".3f"):
    """Formata valor numerico ou retorna '---'."""
    if val is None:
        return "---"
    try:
        return format(float(val), fmt)
    except (TypeError, ValueError):
        return str(val)


def gerar_ose(ns_id, trecho, pvs, nucleo, out_path, cfg=None):
    """
    Gera OSE padrao SABESP — layout identico ao template ProSaneamento.

    Args:
        ns_id: numero da NS (int)
        trecho: dict com pv_ini, pv_fim, dn_mm, ext_m, decl_mm, etc.
        pvs: dict {nome: {x, y, ct, cf, ...}}
        nucleo: nome do nucleo (str)
        out_path: caminho do arquivo .xlsx
        cfg: dict opcional com empresa, contrato, cidade, engenheiro
    """
    if cfg is None:
        # Tenta pegar do motor_contratos
        try:
            from motor_contratos import get_contrato
            ct = get_contrato()
            if ct:
                cfg = {
                    "empresa": ct.get("empresa", "FCN Construcoes e Saneamento"),
                    "contrato": ct.get("numero", "11481051"),
                    "cidade": ct.get("cidade", "Santos"),
                    "nucleo": nucleo,
                    "engenheiro": "Felipe Nery",
                }
            else:
                cfg = {}
        except Exception:
            cfg = {}

    cfg.setdefault("empresa", "FCN Construcoes e Saneamento")
    cfg.setdefault("contrato", "11481051")
    cfg.setdefault("cidade", "Santos")
    cfg.setdefault("nucleo", nucleo)
    cfg.setdefault("engenheiro", "Felipe Nery")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OS"
    ws.sheet_view.showGridLines = False

    # ── estilos ──
    thin  = Side(style="thin",   color="888888")
    thick = Side(style="medium", color="1F4E79")
    brd   = Border(left=thin, right=thin, top=thin, bottom=thin)
    brd_h = Border(left=thick, right=thick, top=thick, bottom=thick)

    is_agua = trecho.get("is_agua", False) or trecho.get("tipo") == "agua"
    ext = trecho.get("ext_m") or 0
    pvi_d = pvs.get(trecho.get("pv_ini", ""), {})
    pvf_d = pvs.get(trecho.get("pv_fim", ""), {})

    # ── helpers ──
    def _hc(row, col, val, fc="1F4E79", fg="FFFFFF", bold=True, sz=9):
        c = ws.cell(row=row, column=col, value=val)
        c.fill      = PatternFill("solid", fgColor=fc)
        c.font      = Font(name="Calibri", bold=bold, color=fg, size=sz)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = brd
        return c

    def _dc(row, col, val, fmt=None, bold=False, fc="FFFFFF", halign="center"):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(name="Calibri", size=9, bold=bold)
        c.alignment = Alignment(horizontal=halign, vertical="center")
        c.fill      = PatternFill("solid", fgColor=fc)
        c.border    = brd
        if fmt:
            c.number_format = fmt
        return c

    # ── mapeamento colunas (template ProSaneamento) ──
    CB  = 2;  CD  = 4;  CF  = 6;  CH  = 8;  CJ  = 10
    CL  = 12; CN  = 14; CP_ = 16; CR  = 18; CT_ = 20
    CV  = 22; CX  = 24; CZ  = 26
    CAB = 28; CAD = 30; CAF = 32; CAH = 34; CAX = 50

    # ── larguras ──
    COL_WIDTHS = {
        1: 1.5, 2: 14, 3: 1.5, 4: 7, 5: 1.5, 6: 7, 7: 1.5,
        8: 8, 9: 1.5, 10: 8, 11: 1.5, 12: 8, 13: 1.5, 14: 8,
        15: 1.5, 16: 8, 17: 1.5, 18: 8, 19: 1.5, 20: 6, 21: 1.5,
        22: 6, 23: 1.5, 24: 7, 25: 1.5, 26: 7, 27: 1.5, 28: 9,
        29: 1.5, 30: 6, 31: 1.5, 32: 6, 33: 1.5, 34: 18,
    }
    for ci, cw in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(ci)].width = cw
    for ci in range(35, 51):
        ws.column_dimensions[get_column_letter(ci)].width = 2.2

    # ── alturas ──
    for r, h in {3: 18, 4: 14, 5: 14, 6: 22, 7: 18, 8: 14, 9: 14,
                 10: 14, 11: 14, 12: 14, 13: 14, 14: 6, 15: 22, 16: 14, 17: 16}.items():
        ws.row_dimensions[r].height = h

    # ══════════════════════════════════════════════════════════
    # BLOCO 1 — CABECALHO INSTITUCIONAL (linhas 3-13)
    # ══════════════════════════════════════════════════════════
    titulo = ("SISTEMA DE ABASTECIMENTO DE AGUA FRIA"
              if is_agua else "SISTEMA DE ESGOTAMENTO SANITARIO")
    ws["L3"].value     = titulo
    ws["L3"].font      = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ws["L3"].fill      = PatternFill("solid", fgColor="1F4E79")
    ws["L3"].alignment = Alignment(horizontal="center", vertical="center")
    ws["L3"].border    = brd_h
    ws.merge_cells("L3:Y3")

    ws["L4"].value     = f"CIDADE: {cfg.get('cidade', '')}"
    ws["L4"].font      = Font(name="Calibri", bold=True, size=10)
    ws["L4"].alignment = Alignment(horizontal="center", vertical="center")
    ws["L4"].border    = brd
    ws.merge_cells("L4:Y4")

    ws["L6"].value     = "ORDEM DE SERVICO"
    ws["L6"].font      = Font(name="Calibri", bold=True, size=13, color="1F4E79")
    ws["L6"].alignment = Alignment(horizontal="center", vertical="center")
    ws["L6"].border    = brd
    ws.merge_cells("L6:Y6")

    subtitulo = "PARA EXECUCAO" if is_agua else "PARA GABARITO"
    ws["L7"].value     = subtitulo
    ws["L7"].font      = Font(name="Calibri", bold=True, size=11, color="1F4E79")
    ws["L7"].alignment = Alignment(horizontal="center", vertical="center")
    ws["L7"].border    = brd
    ws.merge_cells("L7:Y7")

    # Info institucional
    for ref, merge, val in [
        ("B9",  "B9:K9",   f"EMPRESA: {cfg.get('empresa', '')}"),
        ("L9",  "L9:Y9",   f"CALCULISTA: {cfg.get('engenheiro', 'Felipe Nery')}"),
        ("B10", "B10:K10",  f"CONTRATO No: {cfg.get('contrato', '')}"),
        ("B11", "B11:K11",  f"NUCLEO: {cfg.get('nucleo', '')}"),
        ("L11", "L11:Y11",  f"DATA: {datetime.now().strftime('%d/%m/%Y')}"),
        ("B12", "B12:K12",  f"LOGRADOURO: {trecho.get('rua', 'Sem Rua')}"),
    ]:
        ws[ref].value     = val
        ws[ref].font      = Font(name="Calibri", size=9)
        ws[ref].alignment = Alignment(horizontal="left", vertical="center")
        ws[ref].border    = brd
        ws.merge_cells(merge)

    # OS numero + nome dos PVs (OBRIGATORIO)
    pv_ini = trecho.get("pv_ini", "?")
    pv_fim = trecho.get("pv_fim", "?")
    ws["B13"].value     = f"O.S. No: NS{ns_id:03d} -- {pv_ini} ao {pv_fim}"
    ws["B13"].font      = Font(name="Calibri", size=9, bold=True)
    ws["B13"].alignment = Alignment(horizontal="left", vertical="center")
    ws["B13"].border    = brd
    ws.merge_cells("B13:K13")

    # Convencoes (Z10:AD13)
    conv_fc  = "FFF2CC"
    conv_hdr = "FFD966"
    ws["Z10"].value     = "CONVENCOES"
    ws["Z10"].font      = Font(name="Calibri", bold=True, size=8)
    ws["Z10"].fill      = PatternFill("solid", fgColor=conv_hdr)
    ws["Z10"].alignment = Alignment(horizontal="center", vertical="center")
    ws["Z10"].border    = brd
    ws.merge_cells("Z10:AD10")

    for sig_ref, sig_val, desc_ref, desc_val in [
        ("Z11",  "CT",  "AA11", "Cota do Terreno (m)"),
        ("AB11", "CP",  "AC11", "Cota de Projeto (m)"),
        ("Z12",  "I",   "AA12", "Declividade (m/m)"),
        ("AB12", "DN",  "AC12", "Diametro Nominal (mm)"),
        ("Z13",  "G",   "AA13", "Largura da Vala (m)"),
        ("AB13", "H",   "AC13", "Altura Regua ao Greide (m)"),
    ]:
        ws[sig_ref].value     = sig_val
        ws[sig_ref].font      = Font(name="Calibri", bold=True, size=8)
        ws[sig_ref].fill      = PatternFill("solid", fgColor=conv_fc)
        ws[sig_ref].alignment = Alignment(horizontal="center", vertical="center")
        ws[sig_ref].border    = brd
        ws[desc_ref].value     = desc_val
        ws[desc_ref].font      = Font(name="Calibri", size=7.5)
        ws[desc_ref].fill      = PatternFill("solid", fgColor=conv_fc)
        ws[desc_ref].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws[desc_ref].border    = brd

    for sig_ref, sig_val in [("AD11", "CR"), ("AD12", "P")]:
        ws[sig_ref].value = sig_val
        ws[sig_ref].font  = Font(name="Calibri", bold=True, size=8)
        ws[sig_ref].fill  = PatternFill("solid", fgColor=conv_fc)
        ws[sig_ref].alignment = Alignment(horizontal="center", vertical="center")
        ws[sig_ref].border = brd

    # ══════════════════════════════════════════════════════════
    # BLOCO 2 — CABECALHOS TABELA (linhas 15-17)
    # ══════════════════════════════════════════════════════════
    FC_H1 = "1F4E79"; FC_H2 = "2E6DAA"

    for col, lbl, mr in [
        (CB, "TRECHO", "B15:B17"), (CL, "CT", "L15:L17"), (CN, "I", "N15:N17"),
        (CP_, "CP", "P15:P17"), (CR, "CR", "R15:R17"), (CT_, "DN", "T15:T17"),
        (CV, "G", "V15:V17"), (CX, "H", "X15:X17"), (CZ, "P", "Z15:Z17"),
    ]:
        _hc(15, col, lbl, fc=FC_H1, sz=8)
        ws.merge_cells(mr)

    _hc(15, CD, "ESTACA", fc=FC_H1, sz=8); ws.merge_cells("D15:F15")
    _hc(15, CH, "DISTANCIA (m)", fc=FC_H1, sz=8); ws.merge_cells("H15:J15")
    _hc(15, CAB, "POCO DE VISITA", fc=FC_H1, sz=8); ws.merge_cells("AB15:AF15")
    _hc(15, CAH, "OBSERVACOES", fc=FC_H1, sz=8); ws.merge_cells("AH15:AX17")

    for col, lbl, mr in [
        (CD, "INTEIRO", "D16:D17"), (CF, "FRACAO (m)", "F16:F17"),
        (CH, "PARCIAL", "H16:H17"), (CJ, "ACUMUL.", "J16:J17"),
        (CAB, "NOME", "AB16:AB17"), (CAD, "TIPO", "AD16:AD17"),
        (CAF, "PROF.", "AF16:AF17"),
    ]:
        _hc(16, col, lbl, fc=FC_H2, sz=7.5)
        ws.merge_cells(mr)

    # ══════════════════════════════════════════════════════════
    # BLOCO 3 — DADOS (linha 19+)
    # ══════════════════════════════════════════════════════════
    DATA_START = 19
    # Lê largura de vala do ProSaneamento (LST_VALA.DEF), fallback 0.60 m
    try:
        _vala_vals = []
        with open(r"C:\pro_sane\LST_VALA.DEF", encoding="cp1252", errors="ignore") as _f:
            for _ln in _f:
                _ln = _ln.strip()
                if _ln and not _ln.startswith(("#", "*", '"')):
                    try: _vala_vals.append(float(_ln.replace(",", ".")))
                    except ValueError: pass
        G_VALA = (_vala_vals[0] / 100.0) if _vala_vals else 0.60
    except OSError:
        G_VALA = 0.60

    trecho_nome = f"{pv_ini} --> {pv_fim}"
    decl_val = None
    try:
        decl_raw = trecho.get("decl_mm")
        decl_val = float(decl_raw) if decl_raw is not None else None
    except (TypeError, ValueError):
        pass

    ct_ini = trecho.get("ct_ini") or pvi_d.get("ct")
    cf_ini = trecho.get("cf_ini") or pvi_d.get("cf")
    ct_fim = trecho.get("ct_fim") or pvf_d.get("ct")
    cf_fim = trecho.get("cf_fim") or pvf_d.get("cf")
    prof_ini = trecho.get("prof_ini") or pvi_d.get("prof")
    prof_fim = trecho.get("prof_fim") or pvf_d.get("prof")
    if prof_ini is None and ct_ini and cf_ini:
        try: prof_ini = round(abs(float(ct_ini) - float(cf_ini)), 3)
        except: pass
    if prof_fim is None and ct_fim and cf_fim:
        try: prof_fim = round(abs(float(ct_fim) - float(cf_fim)), 3)
        except: pass

    dn = trecho.get("dn_mm") or 200
    ext_f = round(float(ext), 4) if ext else 0.0

    rows_data = [
        (trecho_nome, 0, 0.0, 0.0, 0.0,
         ct_ini, decl_val, cf_ini, prof_ini,
         dn, G_VALA if not is_agua else None, prof_ini, prof_ini,
         pv_ini, "PV" if "PV" in pv_ini.upper() else "PI", prof_ini),
        (trecho_nome, 0, ext_f, ext_f, ext_f,
         ct_fim, decl_val, cf_fim, prof_fim,
         dn, G_VALA if not is_agua else None, prof_fim, prof_fim,
         pv_fim, "PV" if "PV" in pv_fim.upper() else "PI", prof_fim),
    ]

    for ri, row_vals in enumerate(rows_data):
        (trn, est_i, est_f, dist_p, dist_a,
         ct, decl, cp, cr, dn_v, g_val, h_val, p_val,
         pv_nome, pv_tipo, pv_prof) = row_vals
        row = DATA_START + ri
        ws.row_dimensions[row].height = 15
        fc_c = "F0F4FA" if ri % 2 == 0 else "FFFFFF"

        _dc(row, CB, trn, bold=True, fc=fc_c, halign="left")
        _dc(row, CD, est_i, fmt="0", fc=fc_c)
        _dc(row, CF, est_f, fmt="0.0000", fc=fc_c)
        _dc(row, CH, dist_p, fmt="0.0000", fc=fc_c)
        _dc(row, CJ, dist_a, fmt="0.0000", fc=fc_c)
        _dc(row, CL, ct, fmt="0.0000", fc=fc_c)
        _dc(row, CN, decl, fmt="0.00000", fc=fc_c)
        _dc(row, CP_, cp, fmt="0.0000", fc=fc_c)
        _dc(row, CR, cr, fmt="0.0000", fc=fc_c)
        _dc(row, CT_, dn_v, fmt="0", fc=fc_c)
        _dc(row, CV, g_val, fmt="0.00", fc=fc_c)
        _dc(row, CX, h_val, fmt="0.0000", fc=fc_c)
        _dc(row, CZ, p_val, fmt="0.0000", fc=fc_c)
        _dc(row, CAB, pv_nome, fc=fc_c)
        _dc(row, CAD, pv_tipo, fc=fc_c)
        _dc(row, CAF,
            _fmtv(pv_prof, ".2f") if pv_prof is not None else "---", fc=fc_c)
        _dc(row, CAH, None, fc=fc_c)

    # ══════════════════════════════════════════════════════════
    # BLOCO 4 — TOTAIS
    # ══════════════════════════════════════════════════════════
    row_tot = DATA_START + len(rows_data)
    ws.row_dimensions[row_tot].height = 15
    c_tot = ws.cell(row=row_tot, column=CB, value="TOTAIS")
    c_tot.font = Font(name="Calibri", bold=True, size=9)
    c_tot.fill = PatternFill("solid", fgColor="DCE6F1")
    c_tot.alignment = Alignment(horizontal="center", vertical="center")
    c_tot.border = brd
    _dc(row_tot, CH, 0.0, fmt="0.0000", fc="DCE6F1")
    _dc(row_tot, CJ, ext_f, fmt="0.0000", fc="DCE6F1")

    # ══════════════════════════════════════════════════════════
    # BLOCO 5 — DADOS HIDRAULICOS
    # ══════════════════════════════════════════════════════════
    row_hid = row_tot + 2
    ws.row_dimensions[row_hid].height = 14

    v = trecho.get("v_ms") or 0
    q = trecho.get("q_ls") or 0
    tau = trecho.get("tau_pa") or 0

    if is_agua:
        hid_txt = (f"DN: {dn}mm  |  Material: {trecho.get('material', 'PE80')}  |  "
                   f"Prof. ini: {_fmtv(prof_ini, '.2f')}m  |  Prof. fim: {_fmtv(prof_fim, '.2f')}m")
    else:
        status = "APROVADO" if (0.6 <= v <= 5 and tau >= 1) else "VERIFICAR"
        if v > 5: status = "VERIFICAR: V>5m/s"
        if v < 0.6 and v > 0: status = f"VERIFICAR: V={v:.3f}<0.6m/s"
        if tau < 1 and tau > 0: status += f", tau={tau:.2f}<1.0Pa"
        hid_txt = (f"V={v:.3f}m/s  |  Tau={tau:.2f}Pa  |  "
                   f"Qp={q:.3f}L/s  |  {status}  |  R$ {ext * 910:,.2f}")

    ws[f"B{row_hid}"].value     = hid_txt
    ws[f"B{row_hid}"].font      = Font(name="Calibri", size=8.5, italic=True,
                                        color="1F4E79" if "APROVADO" in str(hid_txt) else "CC0000")
    ws[f"B{row_hid}"].fill      = PatternFill("solid", fgColor="E8F4FD")
    ws[f"B{row_hid}"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"B{row_hid}"].border    = brd
    ws.merge_cells(f"B{row_hid}:AX{row_hid}")

    # ══════════════════════════════════════════════════════════
    # BLOCO 6 — ASSINATURAS
    # ══════════════════════════════════════════════════════════
    row_ass = row_hid + 3
    ws.row_dimensions[row_ass].height = 14
    ws.row_dimensions[row_ass + 1].height = 18

    ws[f"B{row_ass}"].value = "ASSINATURAS"
    ws[f"B{row_ass}"].font  = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
    ws[f"B{row_ass}"].fill  = PatternFill("solid", fgColor="1F4E79")
    ws[f"B{row_ass}"].alignment = Alignment(horizontal="center")
    ws.merge_cells(f"B{row_ass}:AX{row_ass}")

    sigs = ["ENG. CAMPO", "EXECUTOR", "COORD.", "GERENTE", "C. PROJ.", "G. ENG."]
    sig_cols = [CB, CJ, CR, CZ, CAH, 42]
    for k, (s, sc) in enumerate(zip(sigs, sig_cols)):
        ec = sig_cols[k + 1] - 2 if k < len(sig_cols) - 1 else 50
        ws.merge_cells(start_row=row_ass + 1, start_column=sc,
                       end_row=row_ass + 1, end_column=ec)
        cell = ws.cell(row=row_ass + 1, column=sc, value=s)
        cell.font = Font(name="Calibri", bold=True, size=8.5, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2E6DAA")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = brd

    # ══════════════════════════════════════════════════════════
    # BLOCO 7 — RODAPE
    # ══════════════════════════════════════════════════════════
    row_rod = row_ass + 3
    ws.row_dimensions[row_rod].height = 12

    ws[f"B{row_rod}"].value = (
        f"SABESP {cfg.get('cidade', 'Santos').upper()}  --  {nucleo}  --  "
        f"{'AGUA FRIA' if is_agua else 'ESGOTO'}  --  "
        f"NS{ns_id:03d} {pv_ini} ao {pv_fim}  --  Folha 1 de 1  --  Rev. 0")
    ws[f"B{row_rod}"].font = Font(name="Calibri", size=8, color="555555", italic=True)
    ws[f"B{row_rod}"].alignment = Alignment(horizontal="left")
    ws.merge_cells(f"B{row_rod}:X{row_rod}")

    ws[f"Z{row_rod}"].value = "ConstruData - HydroNetwork"
    ws[f"Z{row_rod}"].font  = Font(name="Calibri", size=8, color="555555", italic=True)
    ws[f"Z{row_rod}"].alignment = Alignment(horizontal="right")
    ws.merge_cells(f"Z{row_rod}:AX{row_rod}")

    # ── salvar ──
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return str(out_path)


if __name__ == "__main__":
    print("gerar_ose.py - Usar: gerar_ose(ns_id, trecho, pvs, nucleo, out_path)")
