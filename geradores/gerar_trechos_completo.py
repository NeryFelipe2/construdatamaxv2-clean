#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Gera planilha TRECHOS_POR_NUCLEO_NS_COMPLETO.xlsx
com todos os trechos PV-a-PV de todos os nucleos, baseado nos JSONs das NS."""
import sys, io, os, json, glob, math
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_NS = os.path.join(
    os.environ.get("USERPROFILE", "C:\\Users\\felip"),
    "Downloads", "PROJETOS DE ÁGUA E ESGOTO - DWG E DXF 2018",
    "MAPAS ÁGUA E ESGOTO PARA DXF", "NOTAS DE SERVICO COMPLETAS 24-03-2026")

OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "TRECHOS_NS_COMPLETOS")
os.makedirs(OUT_DIR, exist_ok=True)

NUCLEOS = [
    ("MORRO_DO_TETEU",      "Morro do Teteu"),
    ("PANTANAL_BAIXO",      "Pantanal Baixo"),
    ("SAO_MANOEL",          "Sao Manoel"),
    ("JOAO_CARLOS",         "Joao Carlos"),
    ("VILA_CRIADORES",      "Vila Criadores"),
    ("VILA_ISRAEL",         "Vila Israel"),
    ("PROL_TETEU_ALT-01",   "Prol Teteu Alt-01"),
    ("PROL_TETEU",          "Prol Teteu"),
    ("PROL_PANTANAL_BAIXO", "Prol Pantanal Baixo"),
    ("PROL_CRIADORES",      "Prol Criadores"),
    ("PROL_SAO_MANOEL",     "Prol Sao Manoel"),
    ("ESTUDO_SM_JC",        "Estudo CT SM JC"),
]

CUSTO_TUBO = {100:120, 150:160, 200:200, 250:270, 300:340, 350:420, 400:500, 500:700}

COLS = ["ID", "Nucleo", "PV Ini", "PV Fim", "Trecho", "L (m)", "DN (mm)",
        "Mat", "CT Ini", "CF Ini", "Prof Ini (m)", "CT Fim", "CF Fim", "Prof Fim (m)",
        "Decl (m/m)", "V (m/s)", "Q (l/s)", "Pavimento",
        "Escav (m3)", "Reaterro (m3)", "Bota-fora (m3)", "Escoram (m2)", "Pavim (m2)",
        "Custo Tubo", "Custo Escav", "Custo PV", "Total s/BDI", "Total c/BDI", "Rua / Obs"]

LARGURAS = [8,18,12,12,26,8,8,6,9,9,10,9,9,10,9,7,7,14,10,10,10,10,10,12,12,12,14,14,30]

fill_h1    = PatternFill("solid", fgColor="1F4E79")
fill_h2    = PatternFill("solid", fgColor="2E75B6")
fill_nucl  = PatternFill("solid", fgColor="D6E4F0")
fill_sub   = PatternFill("solid", fgColor="FFF2CC")
fill_zebra = PatternFill("solid", fgColor="EBF3FB")
font_h1    = Font(bold=True, color="FFFFFF", size=11)
font_h2    = Font(bold=True, color="FFFFFF", size=9)
font_nucl  = Font(bold=True, size=9)
font_sub   = Font(bold=True, size=9)
font_d     = Font(size=9)
ac         = Alignment(horizontal="center", vertical="center", wrap_text=True)
al         = Alignment(horizontal="left",   vertical="center")
thin = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"))


def preco_tubo(dn):
    return CUSTO_TUBO.get(dn, dn * 1.1)


def pav_tipo(t):
    rua = (t.get("rua") or "").upper()
    if "ASFALTO" in rua or "CBUQ" in rua:            return "CBUQ"
    if "PARALEPIPEDO" in rua or "PEDRA" in rua:       return "PARALELEPIPEDO"
    return "CIMENTADO"


def _num(v):
    return v if v not in (None, 0, "") else ""


def calcular(t, pv_m, pv_j):
    ext  = t.get("ext_m") or 0
    dn   = t.get("dn_mm") or 200
    mat  = t.get("material", "PVC")
    rua  = (t.get("rua") or "").replace("\n", " ")
    ct0  = t.get("ct_ini") or pv_m.get("ct") or 0
    cf0  = t.get("cf_ini") or pv_m.get("cf") or 0
    pr0  = t.get("prof_ini") or pv_m.get("prof") or 0
    ct1  = t.get("ct_fim") or pv_j.get("ct") or 0
    cf1  = t.get("cf_fim") or pv_j.get("cf") or 0
    pr1  = t.get("prof_fim") or pv_j.get("prof") or 0
    decl = t.get("decl_mm") or 0
    v_ms = t.get("v_ms") or 0
    q_ls = t.get("q_ls") or 0
    pav  = pav_tipo(t)

    prof_med = (pr0 + pr1) / 2 if (pr0 and pr1) else max(pr0, pr1, 1.0)
    escav    = round(ext * prof_med, 3)
    reaterro = round(escav * 0.90, 3)
    bota     = round(escav * 0.10, 3)
    escoram  = round(ext * prof_med * 2, 2)
    pavim    = round(ext * 0.80, 2)
    c_tubo   = round(ext * preco_tubo(dn), 2)
    c_escav  = round(escav * 30, 2)
    c_pv     = 3078.0
    t_sbdi   = round(c_tubo + c_escav + c_pv, 2)
    t_cbdi   = round(t_sbdi * 1.25, 2)

    return dict(ext=ext, dn=dn, mat=mat, rua=rua,
                ct0=ct0, cf0=cf0, pr0=pr0, ct1=ct1, cf1=cf1, pr1=pr1,
                decl=decl, v_ms=v_ms, q_ls=q_ls, pav=pav,
                escav=escav, reaterro=reaterro, bota=bota,
                escoram=escoram, pavim=pavim,
                c_tubo=c_tubo, c_escav=c_escav, c_pv=c_pv,
                t_sbdi=t_sbdi, t_cbdi=t_cbdi)


def escrever_cabecalho(ws):
    ws.merge_cells(f"A1:{get_column_letter(len(COLS))}1")
    c = ws.cell(1, 1, "TRECHOS POR NUCLEO — TODOS PVs — ConstruData HydroNetwork v7.0.0  |  FCN Construcoes e Saneamento  |  Contrato 11481051")
    c.fill = fill_h1; c.font = font_h1; c.alignment = ac
    ws.row_dimensions[1].height = 22
    for i, col in enumerate(COLS, 1):
        c = ws.cell(2, i, col)
        c.fill = fill_h2; c.font = font_h2; c.alignment = ac; c.border = thin
    ws.row_dimensions[2].height = 32
    for i, w in enumerate(LARGURAS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"


def escrever_linha(ws, row, tid, nome, pvi, pvf, r, zebra=False):
    vals = [f"T-{tid:04d}", nome, pvi, pvf, f"{pvi}->{pvf}",
            round(r["ext"],2), r["dn"], r["mat"],
            round(r["ct0"],3) if r["ct0"] else "",
            round(r["cf0"],3) if r["cf0"] else "",
            round(r["pr0"],2) if r["pr0"] else "",
            round(r["ct1"],3) if r["ct1"] else "",
            round(r["cf1"],3) if r["cf1"] else "",
            round(r["pr1"],2) if r["pr1"] else "",
            round(r["decl"]/1000,4) if r["decl"] else "",
            round(r["v_ms"],2) if r["v_ms"] else "",
            round(r["q_ls"],2) if r["q_ls"] else "",
            r["pav"], r["escav"], r["reaterro"], r["bota"],
            r["escoram"], r["pavim"],
            r["c_tubo"], r["c_escav"], r["c_pv"],
            r["t_sbdi"], r["t_cbdi"], r["rua"]]
    for i, v in enumerate(vals, 1):
        c = ws.cell(row, i, v)
        c.font = font_d
        c.alignment = al if i == len(COLS) else ac
        c.border = thin
        if zebra:
            c.fill = fill_zebra


def escrever_subtotal(ws, row, nome, s):
    # Escreve sem merge para evitar MergedCell conflict
    vals = ["SUBTOTAL " + nome.upper(), "", "", "", "",
            round(s["l"],2), "", "", "", "", "", "", "", "",
            "", "", "", "",
            round(s["escav"],3), round(s["reaterro"],3), round(s["bota"],3),
            "", round(s["pavim"],2),
            round(s["c_tubo"],2), round(s["c_escav"],2), round(s["c_pv"],2),
            round(s["t_sbdi"],2), round(s["t_cbdi"],2), ""]
    for col, v in enumerate(vals, 1):
        c = ws.cell(row, col, v)
        c.fill = fill_sub; c.font = font_sub; c.alignment = ac; c.border = thin


# -------- MAIN --------
wb = openpyxl.Workbook()
ws_all = wb.active
ws_all.title = "TRECHOS_COMPLETOS"
escrever_cabecalho(ws_all)

row_all = 3
tid_global = 1
grand = dict(l=0, escav=0, reaterro=0, bota=0, pavim=0,
             c_tubo=0, c_escav=0, c_pv=0, t_sbdi=0, t_cbdi=0)

for pasta, nome in NUCLEOS:
    caminho = os.path.join(BASE_NS, pasta)
    if not os.path.isdir(caminho):
        print(f"  [SKIP] {pasta} nao encontrado")
        continue
    jsons = sorted(glob.glob(os.path.join(caminho, "**", "*.json"), recursive=True))
    if not jsons:
        print(f"  [SKIP] {nome}: sem JSONs")
        continue

    print(f"  {nome}: {len(jsons)} trechos")

    # Separador no ALL
    ws_all.merge_cells(f"A{row_all}:{get_column_letter(len(COLS))}{row_all}")
    c = ws_all.cell(row_all, 1, f"  NUCLEO: {nome.upper()}  —  {len(jsons)} trechos")
    c.fill = fill_nucl; c.font = font_nucl; c.alignment = al
    ws_all.row_dimensions[row_all].height = 16
    row_all += 1

    # Aba por nucleo
    ws_n = wb.create_sheet(title=pasta[:25])
    escrever_cabecalho(ws_n)
    row_n = 3

    sub = dict(l=0, escav=0, reaterro=0, bota=0, pavim=0,
               c_tubo=0, c_escav=0, c_pv=0, t_sbdi=0, t_cbdi=0)
    tid_nucleo = 1

    for jpath in jsons:
        try:
            d = json.load(open(jpath, encoding="utf-8"))
        except Exception as e:
            print(f"    ERRO lendo {jpath}: {e}")
            continue
        t   = d.get("trecho", {})
        pv_m = d.get("pv_montante", {})
        pv_j = d.get("pv_jusante", {})
        pvi  = t.get("pv_ini", "?")
        pvf  = t.get("pv_fim", "?")
        r    = calcular(t, pv_m, pv_j)

        zebra = (tid_nucleo % 2 == 0)
        escrever_linha(ws_all, row_all, tid_global, nome, pvi, pvf, r, zebra)
        escrever_linha(ws_n,   row_n,   tid_nucleo, nome, pvi, pvf, r, zebra)

        for k in sub:
            sub[k] += r.get({"l":"ext","c_tubo":"c_tubo","c_escav":"c_escav",
                              "c_pv":"c_pv","t_sbdi":"t_sbdi","t_cbdi":"t_cbdi"}.get(k,k), 0)

        row_all += 1; row_n += 1
        tid_global += 1; tid_nucleo += 1

    escrever_subtotal(ws_all, row_all, nome, sub)
    escrever_subtotal(ws_n,   row_n,   nome, sub)
    row_all += 2; row_n  += 1

    for k in grand: grand[k] += sub.get(k, 0)

# Linha GRAND TOTAL (sem merge para evitar MergedCell)
gt_vals = ["GRAND TOTAL — TODOS OS NUCLEOS", "", "", "", "",
           round(grand["l"],2), "", "", "", "", "", "", "", "",
           "", "", "", "",
           round(grand["escav"],3), round(grand["reaterro"],3), round(grand["bota"],3),
           "", round(grand["pavim"],2),
           round(grand["c_tubo"],2), round(grand["c_escav"],2), round(grand["c_pv"],2),
           round(grand["t_sbdi"],2), round(grand["t_cbdi"],2), ""]
for col, v in enumerate(gt_vals, 1):
    c = ws_all.cell(row_all, col, v)
    c.fill = fill_h1; c.font = font_h1; c.alignment = ac; c.border = thin

out = os.path.join(OUT_DIR, "TRECHOS_POR_NUCLEO_NS_COMPLETO.xlsx")
wb.save(out)
print(f"\nSalvo: {out}")
print(f"Abas: {[ws.title for ws in wb.worksheets]}")
print(f"Total trechos: {tid_global - 1}")
print(f"Total extensao: {round(grand['l'],1)} m")
print(f"Total c/BDI: R$ {grand['t_cbdi']:,.2f}")
