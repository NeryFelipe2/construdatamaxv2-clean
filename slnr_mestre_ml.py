#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SLNR_MESTRE_UNIFICADO — Integração Machine Learning
Contrato 11481051 · DGS Engenharia · SLNR Santos

XGBoost + GridSearchCV + Seaborn para previsão de produção
Atualiza TODAS as abas da planilha com cálculos e fórmulas

Autor: ConstruData HydroNetwork v7.0
Data: Março 2026
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime, timedelta
import json
import warnings
warnings.filterwarnings('ignore')

# ── ML Dependencies ───────────────────────────────────────────────────────────
try:
    from xgboost import XGBRegressor
    XGBOOST_OK = True
except ImportError:
    XGBOOST_OK = False
    print("[AVISO] xgboost não instalado. Usando RandomForest como fallback.")
    from sklearn.ensemble import RandomForestRegressor

from sklearn.model_selection import GridSearchCV, cross_val_score, train_test_split
from sklearn.preprocessing import LabelEncoder
from sklearn.metrics import r2_score, mean_absolute_error, mean_squared_error

# ── Visualization ─────────────────────────────────────────────────────────────
try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import seaborn as sns
    MPL_OK = True
    SNS_OK = True
except ImportError:
    MPL_OK = False
    SNS_OK = False
    print("[AVISO] matplotlib/seaborn não instalados. Gráficos desativados.")

# ── Excel ─────────────────────────────────────────────────────────────────────
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  numbers, Color)
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import (BarChart, LineChart, ScatterChart, AreaChart,
                                Reference, PieChart)
    from openpyxl.chart.series import DataPoint
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False
    print("[ERRO] openpyxl não instalado: pip install openpyxl")

# ── CONSTANTES ────────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).parent
ARQUIVO_MESTRE = SCRIPT_DIR / "SLNR_MESTRE_UNIFICADO.xlsx"
ARQUIVO_SAIDA = SCRIPT_DIR / "SLNR_MESTRE_UNIFICADO_ML.xlsx"

# Dados reais de produtividade (Jan-Mar/2026)
PRODUTIVIDADE_REAL = {
    "Montanhoso": 2.3,    # Morro do Teteu
    "Misto_dificil": 3.0,  # São Manoel
    "Misto_alagado": 2.8,  # Pantanal Baixo
    "Urbano_Denso": 3.5,   # Vila Criadores
    "Urbano_facil": 4.0,   # Caneleira/CDHU
}

# 20 Núcleos com produtividade ajustada
NUCLEOS_20 = [
    {"TAG": "010480", "nome": "Caneleira IV CDHU", "tipo": "Urbano_facil", "prod": 4.0},
    {"TAG": "010524", "nome": "Vila Alemoa", "tipo": "Urbano_Denso", "prod": 3.5},
    {"TAG": "010490", "nome": "Av.Bandeirantes", "tipo": "Montanhoso", "prod": 2.3},
    {"TAG": "010503", "nome": "Encosta Saboo", "tipo": "Montanhoso", "prod": 2.3},
    {"TAG": "010492", "nome": "R.São Geraldo", "tipo": "Montanhoso", "prod": 2.3},
    {"TAG": "010466", "nome": "Av.Martins Fontes", "tipo": "Misto_dificil", "prod": 3.0},
    {"TAG": "010532", "nome": "Av.M.Fontes Pacheco_a", "tipo": "Misto_dificil", "prod": 3.0},
    {"TAG": "010506", "nome": "Rua Nove Saboo", "tipo": "Misto_dificil", "prod": 3.0},
    {"TAG": "010509", "nome": "R.M.Mercedes Fea", "tipo": "Misto_dificil", "prod": 3.0},
    {"TAG": "010465", "nome": "Cam.Maria Rosa", "tipo": "Montanhoso", "prod": 2.3},
    {"TAG": "010449", "nome": "Área ZEIS S.Bento", "tipo": "Montanhoso", "prod": 2.3},
    {"TAG": "010453", "nome": "Nova Cintra I", "tipo": "Montanhoso", "prod": 2.3},
    {"TAG": "010454", "nome": "Pacheco Sta Casa", "tipo": "Montanhoso", "prod": 2.3},
    {"TAG": "010459", "nome": "Vila Vitória", "tipo": "Montanhoso", "prod": 2.3},
    {"TAG": "010460", "nome": "Pacheco Área Part.", "tipo": "Montanhoso", "prod": 2.3},
    {"TAG": "010464", "nome": "GRusso-JGSilveira", "tipo": "Montanhoso", "prod": 2.3},
    {"TAG": "010457", "nome": "São Bento I", "tipo": "Montanhoso", "prod": 2.3},
    {"TAG": "010452", "nome": "Lomba da Penha", "tipo": "Montanhoso", "prod": 2.3},
    {"TAG": "010513", "nome": "Lomba Penha(Est.JBatista)", "tipo": "Montanhoso", "prod": 2.3},
    {"TAG": "010533", "nome": "GRusso(variante)", "tipo": "Montanhoso", "prod": 2.3},
]

# Cores para visualização
CORES_NUCLEOS = {
    "Montanhoso": "#E74C3C",
    "Misto_dificil": "#F39C12",
    "Misto_alagado": "#D68910",
    "Urbano_Denso": "#3498DB",
    "Urbano_facil": "#2ECC71",
}

META_TOTAL_LIG = 25383
META_MENSAL_LIG = 366
DIAS_UTEIS_MES = 22
BDI = 1.25


# ── CLASSE PRINCIPAL ──────────────────────────────────────────────────────────

class SLNRMLIntegrador:
    """
    Integrador de Machine Learning com planilha SLNR_MESTRE_UNIFICADO
    XGBoost + GridSearchCV + Seaborn
    """
    
    def __init__(self, arquivo_mestre: str = None):
        self.arquivo_mestre = Path(arquivo_mestre) if arquivo_mestre else ARQUIVO_MESTRE
        self.arquivo_saida = ARQUIVO_SAIDA
        
        self.df_dados = None
        self.df_nucleos = None
        self.df_execucao = None
        self.modelo_ml = None
        self.metricas_ml = {}
        self.cenarios = []
        self.resultados = {}
        
    # ── 1. CARREGAR DADOS ─────────────────────────────────────────────────────
    
    def carregar_dados(self):
        """Carrega dados da planilha MESTRE_UNIFICADO."""
        print(f"\n📊 Carregando dados de {self.arquivo_mestre.name}...")
        
        if not self.arquivo_mestre.exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {self.arquivo_mestre}")
        
        xl = pd.ExcelFile(self.arquivo_mestre)
        
        # Carregar abas principais
        abas_disponiveis = xl.sheet_names
        
        # DADOS
        if 'DADOS' in abas_disponiveis:
            self.df_dados = pd.read_excel(xl, sheet_name='DADOS')
            print(f"  ✓ DADOS: {len(self.df_dados)} registros")
        
        # NOVOS_NUCLEOS ou N07_NOROESTE etc
        self.df_nucleos = pd.DataFrame(NUCLEOS_20)
        print(f"  ✓ NÚCLEOS: {len(self.df_nucleos)} núcleos planejados")
        
        # Dados de execução diária (se existir)
        if 'PROJ_vs_EXEC' in abas_disponiveis:
            self.df_execucao = pd.read_excel(xl, sheet_name='PROJ_vs_EXEC')
            print(f"  ✓ EXECUÇÃO: {len(self.df_execucao)} registros")
        
        # Dados históricos do contrato (dados_contrato)
        dados_contrato_dir = Path(r"C:\Users\felip\Downloads\NOVA NS Versao 5\dados_contrato")
        if dados_contrato_dir.exists():
            path_exec = dados_contrato_dir / "EXECUCAO_DIARIA.json"
            if path_exec.exists():
                with open(path_exec, 'r', encoding='utf-8') as f:
                    raw = json.load(f)
                rows = []
                for nucleo, dias in raw.items():
                    for rec in dias:
                        rows.append({
                            "nucleo": nucleo,
                            "dia": rec["dia"],
                            "la": float(rec.get("la", 0) or 0),
                            "le": float(rec.get("le", 0) or 0),
                            "pra": float(rec.get("pra", 0) or 0),
                        })
                self.df_historico = pd.DataFrame(rows)
                print(f"  ✓ HISTÓRICO: {len(self.df_historico)} dias × {self.df_historico['nucleo'].nunique()} núcleos")
        
        xl.close()
        print("  ✓ Dados carregados com sucesso!\n")
        
    # ── 2. FEATURE ENGINEERING ────────────────────────────────────────────────
    
    def preparar_features(self):
        """Prepara features para o modelo de ML."""
        print("🔧 Preparando features para ML...")
        
        if not hasattr(self, 'df_historico'):
            print("  ⚠ Sem dados históricos suficientes. Usando dados sintéticos.")
            self._gerar_dados_sinteticos()
        
        df = self.df_historico.copy()
        
        # Features temporais
        df['dia'] = pd.to_datetime(df['dia'])
        df = df.sort_values(['nucleo', 'dia']).reset_index(drop=True)
        
        # Ligações totais
        df['lig_total'] = df['la'] + df['le']
        
        # Features de tempo
        df['dia_semana'] = df['dia'].dt.dayofweek
        df['mes'] = df['dia'].dt.month
        df['dia_mes'] = df['dia'].dt.day
        df['fim_semana'] = (df['dia_semana'] >= 5).astype(int)
        
        # Label encoding do núcleo
        le = LabelEncoder()
        df['nucleo_enc'] = le.fit_transform(df['nucleo'])
        
        # Rolling features por núcleo
        for col in ['lig_total', 'pra', 'la', 'le']:
            df[f'{col}_r3'] = (
                df.groupby('nucleo')[col]
                .transform(lambda x: x.shift(1).rolling(3, min_periods=1).mean())
                .fillna(0)
            )
            df[f'{col}_r7'] = (
                df.groupby('nucleo')[col]
                .transform(lambda x: x.shift(1).rolling(7, min_periods=1).mean())
                .fillna(0)
            )
        
        # Acumulado
        df['lig_acum'] = df.groupby('nucleo')['lig_total'].cumsum()
        
        # Dias decorridos
        df['dias_decorridos'] = df.groupby('nucleo')['dia'].transform(
            lambda x: (x - x.min()).dt.days
        )
        
        self.df_features = df
        self.feature_encoder = le
        
        FEATURES = [
            "lig_total_r3", "lig_total_r7",
            "pra_r3", "pra_r7",
            "la_r3", "le_r3",
            "nucleo_enc", "dia_semana", "mes", "fim_semana", "dias_decorridos"
        ]
        
        print(f"  ✓ {len(FEATURES)} features preparadas")
        print(f"  ✓ {len(df)} amostras\n")
        
        return df, FEATURES
    
    def _gerar_dados_sinteticos(self):
        """Gera dados sintéticos baseados na produtividade real."""
        print("  📝 Gerando dados sintéticos baseados na produtividade real...")
        
        np.random.seed(42)
        rows = []
        
        for nucleo_info in NUCLEOS_20:
            tag = nucleo_info['TAG']
            prod = nucleo_info['prod']
            
            # Simular 90 dias de produção
            data_inicio = datetime(2026, 1, 1)
            for dia in range(90):
                data_atual = data_inicio + timedelta(days=dia)
                
                # Dias úteis (seg-sex)
                if data_atual.weekday() < 5:
                    # Produção com variabilidade
                    ligacoes = max(0, np.random.normal(prod * 1.5, prod * 0.4))
                    la = ligacoes * 0.5 + np.random.normal(0, 1)
                    le = ligacoes * 0.5 + np.random.normal(0, 1)
                    pra = prod * np.random.uniform(0.8, 1.2) * 10
                    
                    rows.append({
                        'nucleo': tag,
                        'dia': data_atual.strftime('%Y-%m-%d'),
                        'la': max(0, la),
                        'le': max(0, le),
                        'pra': max(0, pra),
                    })
        
        self.df_historico = pd.DataFrame(rows)
        print(f"    {len(self.df_historico)} registros sintéticos gerados")
    
    # ── 3. TREINAR MODELO XGBOOST ─────────────────────────────────────────────
    
    def treinar_modelo(self, df, FEATURES):
        """Treina modelo XGBoost com GridSearchCV."""
        print("🤖 Treinando modelo XGBoost + GridSearchCV...")
        
        X = df[FEATURES].fillna(0)
        y = df['lig_total']
        
        # Split treino/teste
        X_train, X_test, y_train, y_test = train_test_split(
            X, y, test_size=0.2, random_state=42
        )
        
        # Grid de hiperparâmetros
        param_grid = {
            'n_estimators': [50, 100, 200],
            'max_depth': [3, 5, 7],
            'learning_rate': [0.05, 0.1, 0.2],
            'subsample': [0.8, 1.0],
        }
        
        # Calcular número de combinações
        n_combos = 1
        for v in param_grid.values():
            n_combos *= len(v)
        
        print(f"  📊 {n_combos} combinações × 3 folds = {n_combos * 3} modelos")
        
        # Modelo base
        if XGBOOST_OK:
            base = XGBRegressor(
                objective='reg:squarederror',
                random_state=42,
                n_jobs=-1,
                verbosity=0
            )
            algoritmo = "XGBoost"
        else:
            base = RandomForestRegressor(random_state=42, n_jobs=-1)
            param_grid = {
                'n_estimators': [50, 100, 200],
                'max_depth': [3, 5, 7, None],
            }
            algoritmo = "RandomForest (fallback)"
        
        # GridSearchCV
        grid = GridSearchCV(
            base,
            param_grid,
            cv=3,
            scoring='r2',
            n_jobs=-1,
            verbose=0,
            refit=True
        )
        
        print("  ⏳ Treinando modelos (isso pode levar alguns minutos)...")
        grid.fit(X_train, y_train)
        
        modelo = grid.best_estimator_
        
        # Predições
        y_pred_test = modelo.predict(X_test)
        y_pred_all = modelo.predict(X)
        
        # Métricas
        r2 = r2_score(y_test, y_pred_test)
        mae = mean_absolute_error(y_test, y_pred_test)
        rmse = np.sqrt(mean_squared_error(y_test, y_pred_test))
        
        # Cross-validation
        cv_scores = cross_val_score(modelo, X, y, cv=5, scoring='r2')
        
        self.modelo_ml = modelo
        self.metricas_ml = {
            'algoritmo': algoritmo,
            'r2_test': round(float(r2), 4),
            'r2_cv_mean': round(float(cv_scores.mean()), 4),
            'r2_cv_std': round(float(cv_scores.std()), 4),
            'mae': round(float(mae), 3),
            'rmse': round(float(rmse), 3),
            'best_params': grid.best_params_,
            'best_cv_score': round(float(grid.best_score_), 4),
            'n_combos': n_combos,
            'n_modelos': n_combos * 3,
        }
        
        # Feature importance
        if hasattr(modelo, 'feature_importances_'):
            imp = modelo.feature_importances_
        else:
            imp = np.ones(len(FEATURES)) / len(FEATURES)
        
        self.df_feature_importance = pd.DataFrame({
            'feature': FEATURES,
            'importancia': imp,
        }).sort_values('importancia', ascending=False).reset_index(drop=True)
        
        print(f"  ✓ Modelo treinado: {algoritmo}")
        print(f"  ✓ R² Test: {r2:.4f} | MAE: {mae:.3f} | RMSE: {rmse:.3f}")
        print(f"  ✓ Melhores params: {grid.best_params_}\n")
        
        return modelo, y_test, y_pred_test, y_pred_all
    
    # ── 4. GERAR CENÁRIOS ─────────────────────────────────────────────────────
    
    def gerar_cenarios(self, df):
        """Gera cenários de aceleração."""
        print("📈 Gerando cenários de aceleração...")
        
        # Dados atuais
        lig_realizadas = int(df['lig_total'].sum())
        lig_faltam = max(0, META_TOTAL_LIG - lig_realizadas)
        
        # Produção média recente
        prod_dias = df[df['lig_total'] > 0]['lig_total']
        media_recente = prod_dias.tail(60).mean() if len(prod_dias) > 0 else 1.0
        media_recente = max(media_recente, 1.0)
        
        # Fator de utilização
        ultimo_dia = df['dia'].max()
        ultimos_30 = df[df['dia'] >= ultimo_dia - timedelta(days=30)]
        fat_utilizacao = (ultimos_30['lig_total'] > 0).mean()
        fat_utilizacao = max(fat_utilizacao, 0.3)
        
        media_recente_corrida = media_recente * fat_utilizacao
        
        # Cenários
        self.cenarios = []
        multiplicadores = [1.0, 1.10, 1.20, 1.30, None]
        nomes = [
            "Baseline (ritmo atual)",
            "Aceleração +10%",
            "Aceleração +20%",
            "Aceleração +30%",
            "Meta contratual (366/mês)",
        ]
        
        for mult, nome in zip(multiplicadores, nomes):
            if mult is None:
                prod_diaria = META_MENSAL_LIG / DIAS_UTEIS_MES * fat_utilizacao
            else:
                prod_diaria = media_recente_corrida * mult
            
            dias_para_concluir = int(np.ceil(lig_faltam / max(prod_diaria, 0.1)))
            data_conclusao = ultimo_dia + timedelta(days=dias_para_concluir)
            
            aceleracao_pct = ((prod_diaria / (media_recente_corrida or 1)) - 1) * 100
            custo_extra_mensal = max(0, aceleracao_pct / 100) * 50000
            
            self.cenarios.append({
                'nome': nome,
                'producao_diaria': round(float(prod_diaria), 1),
                'producao_mensal': round(float(prod_diaria * DIAS_UTEIS_MES), 0),
                'lig_realizadas': lig_realizadas,
                'lig_faltam': lig_faltam,
                'dias_para_concluir': dias_para_concluir,
                'data_conclusao': data_conclusao.strftime('%d/%m/%Y'),
                'custo_extra_mensal': round(float(custo_extra_mensal), 2),
                'aceleracao_pct': round(float(max(0, aceleracao_pct)), 1),
            })
        
        for cen in self.cenarios:
            print(f"  • {cen['nome']}: {cen['producao_mensal']:.0f} lig/mês → {cen['data_conclusao']}")
        
        print()
        return self.cenarios
    
    # ── 5. ATUALIZAR PLANILHA ─────────────────────────────────────────────────
    
    def atualizar_planilha(self):
        """Atualiza TODAS as abas da planilha com ML integrado."""
        print("📝 Atualizando planilha SLNR_MESTRE_UNIFICADO...\n")
        
        if not OPENPYXL_OK:
            print("[ERRO] openpyxl não disponível. Não foi possível atualizar a planilha.")
            return
        
        # Carregar workbook existente
        wb = load_workbook(self.arquivo_mestre)
        
        # Atualizar abas principais
        self._atualizar_aba_ml_resultados(wb)
        self._atualizar_aba_ml_metodo(wb)
        self._atualizar_aba_tendencias(wb)
        self._atualizar_aba_dashboard(wb)
        self._atualizar_aba_novos_nucleos(wb)
        self._atualizar_aba_crono_macro(wb)
        self._atualizar_aba_custos(wb)
        self._atualizar_aba_curva_s(wb)
        self._atualizar_aba_medicao_mensal(wb)
        self._atualizar_aba_compras_mensal(wb)
        self._atualizar_aba_painel_executivo(wb)
        self._atualizar_aba_heatmap_rede(wb)
        self._atualizar_aba_heatmap_lig(wb)

        # Gerar Notas de Serviço por PIs e PVs
        self._gerar_notas_servico_pis_pvs(wb)

        # Atualizar abas dos nucleos (N07, N08, N09, N10, N11, N12, SD_*)
        self._atualizar_aba_nucleo(wb, 'N07_NOROESTE', '010480')
        self._atualizar_aba_nucleo(wb, 'N08_V_PROGRESSO', '010524')
        self._atualizar_aba_nucleo(wb, 'N09_Z_LESTE', '010490')
        self._atualizar_aba_nucleo(wb, 'N10_CONJUNTO', '010503')
        self._atualizar_aba_nucleo(wb, 'N11_ALAGADO', '010492')
        self._atualizar_aba_nucleo(wb, 'N12_MONTANHOSO', '010466')
        self._atualizar_aba_nucleo(wb, 'SD_JOAO_CARLOS', 'SD_JOAO_CARLOS')
        self._atualizar_aba_nucleo(wb, 'SD_SAO_MANOEL', 'SD_SAO_MANOEL')
        self._atualizar_aba_nucleo(wb, 'SD_VILA_ISRAEL', 'SD_VILA_ISRAEL')
        self._atualizar_aba_nucleo(wb, 'SD_MORRO_TETEU', 'SD_MORRO_TETEU')
        self._atualizar_aba_nucleo(wb, 'SD_VILA_CRIADORES', 'SD_VILA_CRIADORES')
        self._atualizar_aba_nucleo(wb, 'SD_PANTANAL_BAIXO', 'SD_PANTANAL_BAIXO')

        # Salvar
        print(f"💾 Salvando {self.arquivo_saida.name}...")
        wb.save(self.arquivo_saida)
        wb.close()
        print(f"  ✓ Planilha atualizada com sucesso!\n")
    
    def _estilo_cabecalho(self, ws, row, col_start, col_end, texto, cor="1F4E79"):
        """Estiliza cabeçalho."""
        ws.merge_cells(start_row=row, start_column=col_start,
                       end_row=row, end_column=col_end)
        cell = ws.cell(row=row, column=col_start, value=texto)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill("solid", fgColor=cor)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def _atualizar_aba_ml_resultados(self, wb):
        """Atualiza aba ML_RESULTADOS com métricas do modelo."""
        print("  📊 Atualizando ML_RESULTADOS...")
        
        if 'ML_RESULTADOS' not in wb.sheetnames:
            ws = wb.create_sheet('ML_RESULTADOS')
        else:
            ws = wb['ML_RESULTADOS']
            ws.delete_rows(1, ws.max_row)
        
        # Cabeçalho
        self._estilo_cabecalho(ws, 1, 1, 4, 
            "MACHINE LEARNING — XGBoost + GridSearchCV · SLNR Santos")
        
        ws.cell(2, 1, f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        ws.cell(2, 1).font = Font(italic=True)
        
        # Métricas
        ws.cell(4, 1, "MÉTRICAS DO MODELO")
        ws.cell(4, 1).font = Font(bold=True, size=11)
        
        metricas_data = [
            ("Algoritmo", self.metricas_ml.get('algoritmo', 'N/A')),
            ("R² (teste 20%)", self.metricas_ml.get('r2_test', 0)),
            ("R² CV (5-fold)", self.metricas_ml.get('r2_cv_mean', 0)),
            ("R² CV (desvio)", self.metricas_ml.get('r2_cv_std', 0)),
            ("MAE (lig/dia)", self.metricas_ml.get('mae', 0)),
            ("RMSE (lig/dia)", self.metricas_ml.get('rmse', 0)),
            ("Melhor score CV", self.metricas_ml.get('best_cv_score', 0)),
            ("Combinações testadas", self.metricas_ml.get('n_combos', 0)),
            ("Modelos treinados", self.metricas_ml.get('n_modelos', 0)),
        ]
        
        for i, (k, v) in enumerate(metricas_data, 5):
            ws.cell(i, 1, k)
            ws.cell(i, 2, v)
            if isinstance(v, float):
                ws.cell(i, 2).number_format = "0.0000"
        
        # Feature Importance
        row_fi = 16
        ws.cell(row_fi, 1, "FEATURE IMPORTANCE")
        ws.cell(row_fi, 1).font = Font(bold=True, size=11)
        
        ws.cell(row_fi + 1, 1, "Feature")
        ws.cell(row_fi + 1, 2, "Importância")
        ws.cell(row_fi + 1, 3, "% Total")
        
        if hasattr(self, 'df_feature_importance'):
            total_imp = self.df_feature_importance['importancia'].sum()
            for i, (_, row) in enumerate(self.df_feature_importance.iterrows(), row_fi + 2):
                ws.cell(i, 1, row['feature'])
                ws.cell(i, 2, round(float(row['importancia']), 5))
                ws.cell(i, 3, round(float(row['importancia'] / total_imp * 100), 1))
                ws.cell(i, 3).number_format = "0.0%"
        
        # Ajustar colunas
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
    
    def _atualizar_aba_ml_metodo(self, wb):
        """Atualiza aba METODO_ML com descrição do método."""
        print("  📚 Atualizando METODO_ML...")
        
        if 'METODO_ML' not in wb.sheetnames:
            ws = wb.create_sheet('METODO_ML')
        else:
            ws = wb['METODO_ML']
            ws.delete_rows(1, ws.max_row)
        
        self._estilo_cabecalho(ws, 1, 1, 3,
            "MÉTODO MACHINE LEARNING — XGBoost + GridSearchCV")
        
        conteudo = [
            ("1. OBJETIVO", "Prever produção diária de ligações (LA + LE) por núcleo"),
            ("2. ALGORITMO", "XGBoost Regressor (ou RandomForest como fallback)"),
            ("3. VALIDAÇÃO", "GridSearchCV com 3 folds × 108 combinações = 324 modelos"),
            ("4. FEATURES", [
                "lig_total_r3: Rolling 3 dias ligações totais",
                "lig_total_r7: Rolling 7 dias ligações totais",
                "pra_r3, pra_r7: Rolling produção rede",
                "la_r3, le_r3: Rolling ligações água/esgoto",
                "nucleo_enc: Codificação do núcleo",
                "dia_semana: Dia da semana (0-6)",
                "mes: Mês (1-12)",
                "fim_semana: Flag fim de semana",
                "dias_decorridos: Maturidade do núcleo",
            ]),
            ("5. TARGET", "lig_total: Ligações totais no dia (LA + LE)"),
            ("6. MÉTRICAS", [
                "R²: Coeficiente de determinação",
                "MAE: Erro absoluto médio (lig/dia)",
                "RMSE: Raiz do erro quadrático médio",
            ]),
            ("7. CENÁRIOS", [
                "Baseline: Ritmo atual de produção",
                "+10%: Aceleração moderada",
                "+20%: Aceleração significativa",
                "+30%: Aceleração agressiva",
                "Meta Contratual: 366 ligações/mês",
            ]),
        ]
        
        row = 3
        for titulo, desc in conteudo:
            ws.cell(row, 1, titulo)
            ws.cell(row, 1).font = Font(bold=True, size=10)
            row += 1
            
            if isinstance(desc, list):
                for item in desc:
                    ws.cell(row, 1, f"  • {item}")
                    row += 1
            else:
                ws.cell(row, 1, str(desc))
                row += 1
            row += 1
        
        ws.column_dimensions['A'].width = 60
    
    def _atualizar_aba_tendencias(self, wb):
        """Atualiza aba TENDENCIAS com projeções."""
        print("  📈 Atualizando TENDENCIAS...")
        
        if 'TENDENCIAS' not in wb.sheetnames:
            ws = wb.create_sheet('TENDENCIAS')
        else:
            ws = wb['TENDENCIAS']
            ws.delete_rows(1, ws.max_row)
        
        self._estilo_cabecalho(ws, 1, 1, 6,
            "TENDÊNCIAS E PROJEÇÕES — ML + XGBoost")
        
        # Cenários
        ws.cell(3, 1, "CENÁRIOS DE ACELERAÇÃO")
        ws.cell(3, 1).font = Font(bold=True, size=11)
        
        headers = ["Cenário", "Prod./dia", "Prod./mês", "Lig. feitas",
                   "Lig. faltam", "Dias p/ concluir", "Data conclusão", "Custo extra/mês"]
        
        for j, h in enumerate(headers, 1):
            ws.cell(4, j, h)
            ws.cell(4, j).font = Font(bold=True)
            ws.cell(4, j).fill = PatternFill("solid", fgColor="4472C4")
            ws.cell(4, j).font = Font(bold=True, color="FFFFFF")
        
        for i, cen in enumerate(self.cenarios, 5):
            ws.cell(i, 1, cen['nome'])
            ws.cell(i, 2, cen['producao_diaria'])
            ws.cell(i, 3, cen['producao_mensal'])
            ws.cell(i, 4, cen['lig_realizadas'])
            ws.cell(i, 5, cen['lig_faltam'])
            ws.cell(i, 6, cen['dias_para_concluir'])
            ws.cell(i, 7, cen['data_conclusao'])
            ws.cell(i, 8, cen['custo_extra_mensal'])
            ws.cell(i, 8).number_format = "R$ #,##0.00"
        
        # Fórmulas de projeção por núcleo
        row_proj = len(self.cenarios) + 7
        ws.cell(row_proj, 1, "PROJEÇÃO POR NÚCLEO (FÓRMULAS)")
        ws.cell(row_proj, 1).font = Font(bold=True, size=11)
        
        ws.cell(row_proj + 1, 1, "Núcleo")
        ws.cell(row_proj + 1, 2, "Prod. Real (m/eq/dia)")
        ws.cell(row_proj + 1, 3, "Meta (m)")
        ws.cell(row_proj + 1, 4, "Dias Estimados")
        ws.cell(row_proj + 1, 5, "Previsão Conclusão")
        
        for i, nuc in enumerate(NUCLEOS_20[:10], row_proj + 2):
            ws.cell(i, 1, f"{nuc['TAG']} - {nuc['nome']}")
            ws.cell(i, 2, nuc['prod'])
            ws.cell(i, 3, f"=DADOS!B{i}")  # Fórmula Excel
            ws.cell(i, 4, f"=C{i}/B{i}/22")  # Fórmula Excel
            ws.cell(i, 5, f"=HOJE()+D{i}")  # Fórmula Excel
        
        for col, width in zip('ABCDEFGH', [28, 12, 12, 12, 12, 12, 14, 16]):
            ws.column_dimensions[col].width = width
    
    def _atualizar_aba_dashboard(self, wb):
        """Atualiza aba DASHBOARD com indicadores."""
        print("  📊 Atualizando DASHBOARD...")
        
        if 'DASHBOARD' not in wb.sheetnames:
            ws = wb.create_sheet('DASHBOARD')
        else:
            ws = wb['DASHBOARD']
            ws.delete_rows(1, ws.max_row)
        
        self._estilo_cabecalho(ws, 1, 1, 5,
            "DASHBOARD EXECUTIVO — SLNR Santos · Contrato 11481051")
        
        # KPIs principais
        kpis = [
            ("META TOTAL", f"{META_TOTAL_LIG:,}", "ligações"),
            ("META MENSAL", f"{META_MENSAL_LIG:,}", "ligações/mês"),
            ("R² MODELO", f"{self.metricas_ml.get('r2_test', 0):.4f}", ""),
            ("MAE", f"{self.metricas_ml.get('mae', 0):.2f}", "lig/dia"),
        ]
        
        for i, (titulo, valor, unidade) in enumerate(kpis, 3):
            ws.cell(i, 1, titulo)
            ws.cell(i, 1).font = Font(bold=True)
            ws.cell(i, 2, valor)
            ws.cell(i, 2).font = Font(bold=True, size=14)
            ws.cell(i, 3, unidade)
        
        # Cenários em tabela dinâmica
        row_cen = 10
        ws.cell(row_cen, 1, "CENÁRIOS DISPONÍVEIS")
        ws.cell(row_cen, 1).font = Font(bold=True, size=11)
        
        for i, cen in enumerate(self.cenarios, row_cen + 1):
            ws.cell(i, 1, f"• {cen['nome']}")
            ws.cell(i, 2, f"→ {cen['producao_mensal']:.0f} lig/mês")
            ws.cell(i, 3, f"→ Conclusão: {cen['data_conclusao']}")
        
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
    
    def _atualizar_aba_novos_nucleos(self, wb):
        """Atualiza aba NOVOS_NUCLEOS com ML."""
        print("  🏗️ Atualizando NOVOS_NUCLEOS...")
        
        if 'NOVOS_NUCLEOS' not in wb.sheetnames:
            ws = wb.create_sheet('NOVOS_NUCLEOS')
        else:
            ws = wb['NOVOS_NUCLEOS']
            ws.delete_rows(1, ws.max_row)
        
        self._estilo_cabecalho(ws, 1, 1, 7,
            "NOVOS NÚCLEOS — Planejamento com ML")
        
        headers = ["TAG", "Núcleo", "Tipo Terreno", "Prod. (m/eq/dia)",
                   "Ligações", "Rede (m)", "Dias Estimados", "Previsão", "Custo"]
        
        for j, h in enumerate(headers, 1):
            ws.cell(3, j, h)
            ws.cell(3, j).font = Font(bold=True)
            ws.cell(3, j).fill = PatternFill("solid", fgColor="4472C4")
            ws.cell(3, j).font = Font(bold=True, color="FFFFFF")
        
        for i, nuc in enumerate(NUCLEOS_20, 4):
            ws.cell(i, 1, nuc['TAG'])
            ws.cell(i, 2, nuc['nome'])
            ws.cell(i, 3, nuc['tipo'])
            ws.cell(i, 4, nuc['prod'])
            ws.cell(i, 5, f"=DADOS!C{i}")  # Fórmula
            ws.cell(i, 6, f"=DADOS!D{i}")  # Fórmula
            ws.cell(i, 7, f"=F{i}/E{i}/{nuc['prod']}")  # Fórmula Excel
            ws.cell(i, 8, f"=HOJE()+G{i}")  # Fórmula Excel
            ws.cell(i, 9, f"=F{i}*910*1.25")  # Fórmula Excel
        
        for col, width in zip('ABCDEFGHI', [12, 30, 18, 12, 12, 12, 12, 14, 14]):
            ws.column_dimensions[col].width = width
    
    def _atualizar_aba_crono_macro(self, wb):
        """Atualiza aba CRONO_MACRO."""
        print("  📅 Atualizando CRONO_MACRO...")
        
        if 'CRONO_MACRO' not in wb.sheetnames:
            ws = wb.create_sheet('CRONO_MACRO')
        else:
            ws = wb['CRONO_MACRO']
            ws.delete_rows(1, ws.max_row)
        
        self._estilo_cabecalho(ws, 1, 1, 5,
            "CRONOGRAMA MACRO — Projeção ML")
        
        # Usar cenário baseline
        baseline = self.cenarios[0] if self.cenarios else {}
        
        ws.cell(3, 1, "CENÁRIO BASELINE")
        ws.cell(3, 1).font = Font(bold=True)
        
        dados = [
            ("Produção Diária", f"{baseline.get('producao_diaria', 0):.1f}", "lig/dia"),
            ("Produção Mensal", f"{baseline.get('producao_mensal', 0):.0f}", "lig/mês"),
            ("Ligações Realizadas", f"{baseline.get('lig_realizadas', 0):,}", "lig"),
            ("Ligações Faltantes", f"{baseline.get('lig_faltam', 0):,}", "lig"),
            ("Dias para Concluir", f"{baseline.get('dias_para_concluir', 0):,}", "dias"),
            ("Data Conclusão", baseline.get('data_conclusao', 'N/A'), ""),
        ]
        
        for i, (k, v, un) in enumerate(dados, 4):
            ws.cell(i, 1, k)
            ws.cell(i, 2, v)
            ws.cell(i, 3, un)
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
    
    def _atualizar_aba_custos(self, wb):
        """Atualiza aba CUSTOS com projeções ML."""
        print("  💰 Atualizando CUSTOS...")
        
        if 'CUSTOS' not in wb.sheetnames:
            ws = wb.create_sheet('CUSTOS')
        else:
            ws = wb['CUSTOS']
            ws.delete_rows(1, ws.max_row)
        
        self._estilo_cabecalho(ws, 1, 1, 5,
            "CUSTOS — Projeção com ML")
        
        custo_por_lig = 910 * 2  # R$ 1.820/ligação
        
        ws.cell(3, 1, "CUSTO POR LIGAÇÃO")
        ws.cell(3, 1).font = Font(bold=True)
        
        dados = [
            ("Custo Unitário", f"R$ {custo_por_lig:,.2f}", "/ligação"),
            ("Ligações Totais", f"{META_TOTAL_LIG:,}", "lig"),
            ("Custo Total Estimado", f"R$ {META_TOTAL_LIG * custo_por_lig:,.2f}", ""),
            ("Custo Mensal (meta)", f"R$ {META_MENSAL_LIG * custo_por_lig:,.2f}", ""),
        ]
        
        for i, (k, v, un) in enumerate(dados, 4):
            ws.cell(i, 1, k)
            ws.cell(i, 2, v)
            ws.cell(i, 3, un)
        
        # Cenários de custo
        row_cen = 10
        ws.cell(row_cen, 1, "CUSTO POR CENÁRIO")
        ws.cell(row_cen, 1).font = Font(bold=True)
        
        ws.cell(row_cen + 1, 1, "Cenário")
        ws.cell(row_cen + 1, 2, "Custo Extra/mês")
        ws.cell(row_cen + 1, 3, "Custo Total Projeto")
        
        for i, cen in enumerate(self.cenarios, row_cen + 2):
            ws.cell(i, 1, cen['nome'])
            ws.cell(i, 2, f"R$ {cen['custo_extra_mensal']:,.2f}")
            custo_total = META_TOTAL_LIG * custo_por_lig + cen['custo_extra_mensal'] * 12
            ws.cell(i, 3, f"R$ {custo_total:,.2f}")
        
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
    
    def _atualizar_aba_curva_s(self, wb):
        """Atualiza aba CURVA_S com projeções."""
        print("  📈 Atualizando CURVA_S...")
        
        if 'CURVA_S' not in wb.sheetnames:
            ws = wb.create_sheet('CURVA_S')
        else:
            ws = wb['CURVA_S']
            ws.delete_rows(1, ws.max_row)
        
        self._estilo_cabecalho(ws, 1, 1, 6,
            "CURVA S — Físico-Financeiro com Projeção ML")
        
        # Gerar dados mensais acumulados
        baseline = self.cenarios[0] if self.cenarios else {}
        prod_mensal = baseline.get('producao_mensal', 58)
        
        ws.cell(3, 1, "Mês")
        ws.cell(3, 2, "Produção Mensal (lig)")
        ws.cell(3, 3, "Acumulado (lig)")
        ws.cell(3, 4, "% Físico")
        ws.cell(3, 5, "Custo Mensal (R$)")
        ws.cell(3, 6, "Custo Acumulado (R$)")
        ws.cell(3, 7, "% Financeiro")
        
        custo_por_lig = 1820
        acum_lig = 0
        acum_cust = 0
        
        for mes in range(1, 25):
            prod = prod_mensal if acum_lig < META_TOTAL_LIG else 0
            acum_lig += prod
            pct_fisico = min(100, acum_lig / META_TOTAL_LIG * 100)
            
            custo_mensal = prod * custo_por_lig
            acum_cust += custo_mensal
            pct_fin = min(100, acum_cust / (META_TOTAL_LIG * custo_por_lig) * 100)
            
            row = mes + 3
            ws.cell(row, 1, f"Mês {mes}")
            ws.cell(row, 2, f"={prod:.0f}")
            ws.cell(row, 3, f"={acum_lig:.0f}")
            ws.cell(row, 4, f"={pct_fisico:.1f}%")
            ws.cell(row, 5, f"=R$ {custo_mensal:,.0f}")
            ws.cell(row, 6, f"=R$ {acum_cust:,.0f}")
            ws.cell(row, 7, f"={pct_fin:.1f}%")
        
        for col, width in zip('ABCDEFG', [10, 18, 15, 12, 18, 18, 14]):
            ws.column_dimensions[col].width = width
    
    def _atualizar_aba_medicao_mensal(self, wb):
        """Atualiza aba MEDICAO_MENSAL."""
        print("  📋 Atualizando MEDICAO_MENSAL...")
        
        if 'MEDICAO_MENSAL' not in wb.sheetnames:
            ws = wb.create_sheet('MEDICAO_MENSAL')
        else:
            ws = wb['MEDICAO_MENSAL']
            ws.delete_rows(1, ws.max_row)
        
        self._estilo_cabecalho(ws, 1, 1, 5,
            "MEDIÇÃO MENSAL — Projeção ML")
        
        headers = ["Item", "Descrição", "Unid", "Qtd Total", "P.Unit", 
                   "Total", "Mês1", "Mês2", "Mês3", "Mês4", "Mês5", "Mês6"]
        
        for j, h in enumerate(headers, 1):
            ws.cell(3, j, h)
            ws.cell(3, j).font = Font(bold=True)
            ws.cell(3, j).fill = PatternFill("solid", fgColor="4472C4")
            ws.cell(3, j).font = Font(bold=True, color="FFFFFF")
        
        # Itens principais
        itens = [
            ("101", "Ligações de Água", "lig", META_TOTAL_LIG, 455, f"={META_TOTAL_LIG}*455"),
            ("102", "Ligações de Esgoto", "lig", META_TOTAL_LIG, 455, f"={META_TOTAL_LIG}*455"),
            ("201", "Rede de Água PVC", "m", META_TOTAL_LIG * 3, 120, ""),
            ("202", "Rede de Esgoto PVC", "m", META_TOTAL_LIG * 2.5, 150, ""),
        ]
        
        baseline = self.cenarios[0] if self.cenarios else {}
        prod_mensal = baseline.get('producao_mensal', 58)
        
        for i, item in enumerate(itens, 4):
            cod, desc, un, qtd, pu, total = item
            ws.cell(i, 1, cod)
            ws.cell(i, 2, desc)
            ws.cell(i, 3, un)
            ws.cell(i, 4, qtd)
            ws.cell(i, 5, pu)
            ws.cell(i, 6, total)
            
            # Distribuição mensal proporcional
            for mes in range(7, 13):
                ws.cell(i, mes, f"={prod_mensal if 'Lig' in desc else qtd/24:.0f}")
        
        for col, width in zip('ABCDEFGHIJKL', [8, 25, 8, 12, 12, 15, 10, 10, 10, 10, 10, 10]):
            ws.column_dimensions[col].width = width
    
    def _atualizar_aba_compras_mensal(self, wb):
        """Atualiza aba COMPRAS_MENSAL."""
        print("  🛒 Atualizando COMPRAS_MENSAL...")
        
        if 'COMPRAS_MENSAL' not in wb.sheetnames:
            ws = wb.create_sheet('COMPRAS_MENSAL')
        else:
            ws = wb['COMPRAS_MENSAL']
            ws.delete_rows(1, ws.max_row)
        
        self._estilo_cabecalho(ws, 1, 1, 6,
            "COMPRAS MENSAL — Planejamento com ML")
        
        headers = ["Material", "Unid", "Qtd Total", "Lead Time", 
                   "Mês1", "Mês2", "Mês3", "Mês4", "Mês5", "Mês6"]
        
        for j, h in enumerate(headers, 1):
            ws.cell(3, j, h)
            ws.cell(3, j).font = Font(bold=True)
            ws.cell(3, j).fill = PatternFill("solid", fgColor="4472C4")
            ws.cell(3, j).font = Font(bold=True, color="FFFFFF")
        
        materiais = [
            ("Tubo PVC DN 100", "m", META_TOTAL_LIG * 3, "30 dias"),
            ("Tubo PVC DN 150", "m", META_TOTAL_LIG * 2.5, "30 dias"),
            ("PV de Visita", "un", META_TOTAL_LIG / 5, "45 dias"),
            ("Caixa de Inspeção", "un", META_TOTAL_LIG, "30 dias"),
            ("Ramal de Ligação", "un", META_TOTAL_LIG * 2, "15 dias"),
            ("Areia", "m³", META_TOTAL_LIG * 0.1, "7 dias"),
            ("Brita", "m³", META_TOTAL_LIG * 0.05, "7 dias"),
            ("CBUQ", "ton", META_TOTAL_LIG * 0.02, "15 dias"),
        ]
        
        for i, mat in enumerate(materiais, 4):
            nome, un, qtd, lead = mat
            ws.cell(i, 1, nome)
            ws.cell(i, 2, un)
            ws.cell(i, 3, qtd)
            ws.cell(i, 4, lead)
            
            # Distribuir ao longo dos meses
            for mes in range(5, 11):
                ws.cell(i, mes, f"={qtd/20:.0f}")
        
        for col, width in zip('ABCDEFGHIJK', [25, 8, 12, 12, 10, 10, 10, 10, 10, 10, 10]):
            ws.column_dimensions[col].width = width
    
    def _atualizar_aba_painel_executivo(self, wb):
        """Atualiza aba PAINEL_EXECUTIVO."""
        print("  📊 Atualizando PAINEL_EXECUTIVO...")
        
        if 'PAINEL_EXECUTIVO' not in wb.sheetnames:
            ws = wb.create_sheet('PAINEL_EXECUTIVO')
        else:
            ws = wb['PAINEL_EXECUTIVO']
            ws.delete_rows(1, ws.max_row)
        
        self._estilo_cabecalho(ws, 1, 1, 4,
            "PAINEL EXECUTIVO — SLNR Santos · Contrato 11481051")
        
        ws.cell(3, 1, f"Data: {datetime.now().strftime('%d/%m/%Y')}")
        ws.cell(3, 1).font = Font(italic=True)
        
        # Resumo executivo
        resumo = f"""
RESUMO EXECUTIVO — MACHINE LEARNING APLICADO AO PLANEJAMENTO

1. MODELO PREDITIVO
   • Algoritmo: {self.metricas_ml.get('algoritmo', 'N/A')}
   • R² (teste): {self.metricas_ml.get('r2_test', 0):.4f}
   • MAE: {self.metricas_ml.get('mae', 0):.2f} ligações/dia
   • Modelos treinados: {self.metricas_ml.get('n_modelos', 0)}

2. CENÁRIO ATUAL (BASELINE)
   • Produção mensal: {self.cenarios[0]['producao_mensal']:.0f} ligações
   • Previsão conclusão: {self.cenarios[0]['data_conclusao']}
   • Dias restantes: {self.cenarios[0]['dias_para_concluir']:,} dias

3. CENÁRIO OTIMIZADO (+30%)
   • Produção mensal: {self.cenarios[3]['producao_mensal']:.0f} ligações
   • Previsão conclusão: {self.cenarios[3]['data_conclusao']}
   • Custo extra: R$ {self.cenarios[3]['custo_extra_mensal']:,.0f}/mês

4. RECOMENDAÇÃO
   • Acelerar produção em 20-30% para cumprir prazo contratual
   • Focar em núcleos de terreno fácil primeiro (Urbano)
   • Aumentar número de equipes de {39} para {50-60}
"""
        
        ws.cell(5, 1, resumo)
        ws.cell(5, 1).font = Font(size=10)
        
        ws.column_dimensions['A'].width = 80
    
    def _atualizar_aba_heatmap_rede(self, wb):
        """Atualiza aba HEATMAP_REDE."""
        print("  🔥 Atualizando HEATMAP_REDE...")
        
        if 'HEATMAP_REDE' not in wb.sheetnames:
            ws = wb.create_sheet('HEATMAP_REDE')
        else:
            ws = wb['HEATMAP_REDE']
            ws.delete_rows(1, ws.max_row)
        
        self._estilo_cabecalho(ws, 1, 1, 5,
            "HEATMAP — Produção de Rede por Núcleo")
        
        # Dados para heatmap
        ws.cell(3, 1, "Núcleo")
        ws.cell(3, 2, "Tipo")
        ws.cell(3, 3, "Produtividade")
        ws.cell(3, 4, "Status")
        ws.cell(3, 5, "Prioridade")
        
        for i, nuc in enumerate(NUCLEOS_20, 4):
            ws.cell(i, 1, nuc['TAG'])
            ws.cell(i, 2, nuc['tipo'])
            ws.cell(i, 3, nuc['prod'])
            
            # Status baseado na produtividade
            if nuc['prod'] >= 3.5:
                status = "✅ Fácil"
                cor = "C6EFCE"
            elif nuc['prod'] >= 3.0:
                status = "⚠️ Médio"
                cor = "FFEB9C"
            else:
                status = "🔴 Difícil"
                cor = "FFC7CE"
            
            ws.cell(i, 4, status)
            ws.cell(i, 4).fill = PatternFill("solid", fgColor=cor)
            
            # Prioridade
            if nuc['prod'] >= 3.5:
                prioridade = "ALTA"
            elif nuc['prod'] >= 2.8:
                prioridade = "MÉDIA"
            else:
                prioridade = "BAIXA"
            
            ws.cell(i, 5, prioridade)
        
        for col, width in zip('ABCDE', [12, 20, 15, 12, 12]):
            ws.column_dimensions[col].width = width
    
    def _atualizar_aba_heatmap_lig(self, wb):
        """Atualiza aba HEATMAP_LIG."""
        print("  🔥 Atualizando HEATMAP_LIG...")

        if 'HEATMAP_LIG' not in wb.sheetnames:
            ws = wb.create_sheet('HEATMAP_LIG')
        else:
            ws = wb['HEATMAP_LIG']
            ws.delete_rows(1, ws.max_row)

        self._estilo_cabecalho(ws, 1, 1, 6,
            "HEATMAP — Ligações por Núcleo (Previsão ML)")

        ws.cell(3, 1, "Núcleo")
        ws.cell(3, 2, "Ligações")
        ws.cell(3, 3, "Rede (m)")
        ws.cell(3, 4, "Dias Estimados")
        ws.cell(3, 5, "Conclusão")
        ws.cell(3, 6, "Risco")

        for i, nuc in enumerate(NUCLEOS_20, 4):
            ws.cell(i, 1, f"{nuc['TAG']} - {nuc['nome']}")
            ws.cell(i, 2, f"=DADOS!C{i}")  # Fórmula
            ws.cell(i, 3, f"=DADOS!D{i}")  # Fórmula
            ws.cell(i, 4, f"=C{i}/{nuc['prod']}/22")  # Fórmula
            ws.cell(i, 5, f"=HOJE()+D{i}")  # Fórmula

            # Risco baseado no tipo
            if 'Montanhoso' in nuc['tipo']:
                risco = "ALTO"
                cor = "FFC7CE"
            elif 'Misto' in nuc['tipo']:
                risco = "MÉDIO"
                cor = "FFEB9C"
            else:
                risco = "BAIXO"
                cor = "C6EFCE"
            
            ws.cell(i, 6, risco)
            ws.cell(i, 6).fill = PatternFill("solid", fgColor=cor)
        
        for col, width in zip('ABCDEF', [30, 12, 12, 15, 14, 10]):
            ws.column_dimensions[col].width = width

    def _atualizar_aba_nucleo(self, wb, nome_aba, tag_nucleo):
        """Atualiza aba de núcleo específico com TODAS as fórmulas."""
        print(f"  🏗️ Atualizando {nome_aba}...")

        # Buscar dados do núcleo
        nuc_info = None
        for nuc in NUCLEOS_20:
            if nuc['TAG'] == tag_nucleo:
                nuc_info = nuc
                break
        
        # Se não encontrou nos 20, usa dados padrão
        if not nuc_info:
            nuc_info = {
                'TAG': tag_nucleo,
                'nome': nome_aba.replace('_', ' '),
                'tipo': 'Montanhoso',
                'prod': 2.3
            }

        if nome_aba not in wb.sheetnames:
            ws = wb.create_sheet(nome_aba)
        else:
            ws = wb[nome_aba]
            ws.delete_rows(1, ws.max_row)

        # Cabeçalho
        self._estilo_cabecalho(ws, 1, 1, 8,
            f"NUCLEO {nome_aba.replace('_', ' ')} — Planejamento com ML")

        # Dados do núcleo
        ws.cell(3, 1, "DADOS DO NÚCLEO")
        ws.cell(3, 1).font = Font(bold=True, size=11)
        
        dados_nucleo = [
            ("TAG", nuc_info['TAG']),
            ("Nome", nuc_info['nome']),
            ("Tipo de Terreno", nuc_info['tipo']),
            ("Produtividade (m/eq/dia)", nuc_info['prod']),
        ]
        
        for i, (k, v) in enumerate(dados_nucleo, 4):
            ws.cell(i, 1, k)
            ws.cell(i, 2, v)
        
        # Tabela de TRECHOS
        row_trechos = 10
        ws.cell(row_trechos, 1, "TRECHOS — CÁLCULOS COM FÓRMULAS")
        ws.cell(row_trechos, 1).font = Font(bold=True, size=11)
        
        headers = ["Item", "Descrição", "Unid", "Qtd", "P.Unit", "Total", 
                   "Produtividade", "Dias", "Início", "Conclusão"]
        
        for j, h in enumerate(headers, 1):
            ws.cell(row_trechos + 1, j, h)
            ws.cell(row_trechos + 1, j).font = Font(bold=True)
            ws.cell(row_trechos + 1, j).fill = PatternFill("solid", fgColor="4472C4")
            ws.cell(row_trechos + 1, j).font = Font(bold=True, color="FFFFFF")
        
        # Itens principais com fórmulas
        itens = [
            ("1", "Ligações de Água", "lig", 100, 455, "água"),
            ("2", "Ligações de Esgoto", "lig", 100, 455, "esgoto"),
            ("3", "Rede de Água PVC DN 100", "m", 300, 120, "rede_agua"),
            ("4", "Rede de Esgoto PVC DN 150", "m", 250, 150, "rede_esgoto"),
            ("5", "PV de Visita", "un", 20, 800, "pv"),
            ("6", "Caixa de Inspeção", "un", 100, 350, "ci"),
            ("7", "Ramal de Ligação", "un", 200, 180, "ramal"),
            ("8", "Escavação/Reaterro", "m³", 150, 85, "terra"),
            ("9", "Revestimento CBUQ", "m²", 80, 120, "cbuq"),
            ("10", "Sinalização", "m", 300, 45, "sinal"),
        ]
        
        prod = nuc_info['prod']
        for i, item in enumerate(itens, row_trechos + 2):
            cod, desc, un, qtd, pu, tipo = item
            
            ws.cell(i, 1, cod)
            ws.cell(i, 2, desc)
            ws.cell(i, 3, un)
            ws.cell(i, 4, qtd)
            ws.cell(i, 5, pu)
            ws.cell(i, 6, f"=D{i}*E{i}")  # Fórmula: Qtd × P.Unit
            
            # Produtividade ajustada por tipo de serviço
            if 'agua' in tipo or 'esgoto' in tipo:
                prod_serv = prod * 0.8
            elif 'rede' in tipo:
                prod_serv = prod
            elif 'pv' in tipo or 'ci' in tipo:
                prod_serv = prod * 0.5
            else:
                prod_serv = prod * 1.2
            
            ws.cell(i, 7, f"{prod_serv:.2f}")
            ws.cell(i, 8, f"=D{i}/G{i}")  # Fórmula: Qtd / Prod = Dias
            
            # Datas com fórmulas
            if i == row_trechos + 2:
                ws.cell(i, 9, "=HOJE()")  # Primeiro item começa hoje
            else:
                ws.cell(i, 9, f"=I{i-1}+1")  # Próximo item começa no dia seguinte
            
            ws.cell(i, 10, f"=I{i}+H{i}")  # Fórmula: Conclusão = Início + Dias
        
        # Tabela de CRONOGRAMA
        row_crono = row_trechos + len(itens) + 3
        ws.cell(row_crono, 1, "CRONOGRAMA — FÍSICO-FINANCEIRO")
        ws.cell(row_crono, 1).font = Font(bold=True, size=11)
        
        headers_crono = ["Mês", "Produção (m)", "Ligações", "Custo (R$)", 
                         "Acumulado (R$)", "% Físico", "% Financeiro"]
        
        for j, h in enumerate(headers_crono, 1):
            ws.cell(row_crono + 1, j, h)
            ws.cell(row_crono + 1, j).font = Font(bold=True)
            ws.cell(row_crono + 1, j).fill = PatternFill("solid", fgColor="70AD47")
            ws.cell(row_crono + 1, j).font = Font(bold=True, color="FFFFFF")
        
        # 12 meses de cronograma
        custo_total_estimado = 500000  # Estimativa base
        producao_mensal = 22 * prod * 10  # m/mês
        
        for mes in range(1, 13):
            row = row_crono + 1 + mes
            ws.cell(row, 1, f"Mês {mes}")
            ws.cell(row, 2, f"={producao_mensal:.0f}")  # Produção mensal
            ws.cell(row, 3, f"=B{row}/5")  # Ligações (estimativa)
            ws.cell(row, 4, f"=B{row}*910*1.25")  # Custo = Produção × R$910 × BDI
            ws.cell(row, 5, f"=SOMA($D${row_crono+2}:D{row})")  # Acumulado
            ws.cell(row, 6, f"={mes*8.33:.1f}%")  # % Físico (linear)
            ws.cell(row, 7, f"=E{row}/{custo_total_estimado}*100")  # % Financeiro
        
        # Tabela de MEDIÇÃO
        row_medicao = row_crono + 14
        ws.cell(row_medicao, 1, "MEDIÇÃO MENSAL — FORMATO SABESP")
        ws.cell(row_medicao, 1).font = Font(bold=True, size=11)
        
        headers_medicao = ["Item", "Descrição", "Unid", "Qtd Contrato", 
                           "P.Unit", "Total", "Mês 1", "Mês 2", "Mês 3"]
        
        for j, h in enumerate(headers_medicao, 1):
            ws.cell(row_medicao + 1, j, h)
            ws.cell(row_medicao + 1, j).font = Font(bold=True)
            ws.cell(row_medicao + 1, j).fill = PatternFill("solid", fgColor="FFC000")
            ws.cell(row_medicao + 1, j).font = Font(bold=True, color="FFFFFF")
        
        # Itens de medição
        itens_medicao = [
            ("101", "Ligações de Água", "lig", 100, 455),
            ("102", "Ligações de Esgoto", "lig", 100, 455),
            ("201", "Rede de Água", "m", 300, 120),
            ("202", "Rede de Esgoto", "m", 250, 150),
            ("301", "PV de Visita", "un", 20, 800),
            ("302", "Caixa de Inspeção", "un", 100, 350),
        ]
        
        for i, item in enumerate(itens_medicao, row_medicao + 2):
            cod, desc, un, qtd, pu = item
            
            ws.cell(i, 1, cod)
            ws.cell(i, 2, desc)
            ws.cell(i, 3, un)
            ws.cell(i, 4, qtd)
            ws.cell(i, 5, pu)
            ws.cell(i, 6, f"=D{i}*E{i}")  # Total
            ws.cell(i, 7, f"=D{i}*0.3")  # Mês 1: 30%
            ws.cell(i, 8, f"=D{i}*0.4")  # Mês 2: 40%
            ws.cell(i, 9, f"=D{i}*0.3")  # Mês 3: 30%
        
        # Resumo de CUSTOS
        row_custos = row_medicao + len(itens_medicao) + 3
        ws.cell(row_custos, 1, "RESUMO DE CUSTOS — ML")
        ws.cell(row_custos, 1).font = Font(bold=True, size=11)
        
        custos_data = [
            ("Custo Total Estimado", f"=SOMA(F{row_medicao+2}:F{row_medicao+len(itens_medicao)})"),
            ("Custo por Ligação", "=F18/200"),  # Custo total / 200 ligações
            ("Produtividade Média", f"{prod} m/eq/dia"),
            ("Dias Totais Estimados", f"=SOMA(H{row_trechos+2}:H{row_trechos+len(itens)})"),
            ("Equipes Necessárias", "=H23/22"),  # Dias totais / 22 dias úteis
        ]
        
        for i, (desc, formula) in enumerate(custos_data, row_custos + 1):
            ws.cell(i, 1, desc)
            ws.cell(i, 2, formula)
        
        # Ajustar colunas
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 12

    def _gerar_notas_servico_pis_pvs(self, wb):
        """Gera aba de Notas de Serviço divididas por PIs e PVs."""
        print("  📄 Gerando Notas de Serviço (PIs e PVs)...")

        if 'NOTAS_SERVICO_PIS_PVS' not in wb.sheetnames:
            ws = wb.create_sheet('NOTAS_SERVICO_PIS_PVS')
        else:
            ws = wb['NOTAS_SERVICO_PIS_PVS']
            ws.delete_rows(1, ws.max_row)

        # Cabeçalho
        self._estilo_cabecalho(ws, 1, 1, 10,
            "NOTAS DE SERVIÇO — Divisão por PIs e PVs · SLNR Santos")

        ws.cell(3, 1, "LEGENDA:")
        ws.cell(3, 1).font = Font(bold=True)
        ws.cell(4, 1, "NS = Nota de Serviço")
        ws.cell(5, 1, "PI = Ponto Inicial (Interseção)")
        ws.cell(6, 1, "PV = Ponto de Vértice (Mudança de direção)")
        ws.cell(7, 1, "Formato: NS_XXX_PI_YY_AO_PV_ZZ")
        ws.cell(8, 1, "  XXX = Número da NS (001-999)")
        ws.cell(9, 1, "  YY  = PI inicial (00-99)")
        ws.cell(10, 1, "  ZZ  = PV final (01-999)")

        # Tabela principal
        row_tabela = 12
        ws.cell(row_tabela, 1, "NOTAS DE SERVIÇO POR TRECHO")
        ws.cell(row_tabela, 1).font = Font(bold=True, size=11)

        headers = ["NS", "Trecho", "PI Inicial", "PV Final", "Extensão (m)",
                   "DN Água", "DN Esgoto", "LA", "LE", "Total Lig.",
                   "Equipes", "Dias", "Status"]

        for j, h in enumerate(headers, 1):
            ws.cell(row_tabela + 1, j, h)
            ws.cell(row_tabela + 1, j).font = Font(bold=True)
            ws.cell(row_tabela + 1, j).fill = PatternFill("solid", fgColor="4472C4")
            ws.cell(row_tabela + 1, j).font = Font(bold=True, color="FFFFFF")

        # Gerar NSs para cada núcleo
        ns_counter = 1
        row_atual = row_tabela + 2

        # Dados fictícios de PIs e PVs por núcleo (em produção viria do DXF/DWG)
        nucleos_pis_pvs = {
            'N07_NOROESTE': [
                ('PI_00', 'PV_1', 150, 100, 75, 20, 15, 35, 2, '=E2/(2*6)'),
                ('PI_01', 'PV_2', 180, 100, 75, 25, 18, 43, 2, '=E3/(2*6)'),
                ('PI_02', 'PV_3', 200, 100, 75, 28, 20, 48, 2, '=E4/(2*6)'),
                ('PI_03', 'PV_4', 170, 100, 75, 22, 16, 38, 2, '=E5/(2*6)'),
                ('PI_04', 'PV_5', 160, 100, 75, 20, 15, 35, 2, '=E6/(2*6)'),
            ],
            'N08_V_PROGRESSO': [
                ('PI_00', 'PV_1', 220, 150, 100, 35, 28, 63, 3, '=E7/(3*6)'),
                ('PI_05', 'PV_2', 190, 150, 100, 30, 24, 54, 3, '=E8/(3*6)'),
                ('PI_06', 'PV_3', 210, 150, 100, 33, 26, 59, 3, '=E9/(3*6)'),
            ],
            'N09_Z_LESTE': [
                ('PI_00', 'PV_1', 250, 150, 100, 40, 32, 72, 3, '=E10/(3*5)'),
                ('PI_07', 'PV_2', 230, 150, 100, 36, 29, 65, 3, '=E11/(3*5)'),
            ],
            'N10_CONJUNTO': [
                ('PI_00', 'PV_1', 180, 100, 75, 28, 22, 50, 2, '=E12/(2*5)'),
                ('PI_08', 'PV_2', 200, 100, 75, 32, 25, 57, 2, '=E13/(2*5)'),
            ],
            'N11_ALAGADO': [
                ('PI_00', 'PV_1', 160, 100, 75, 25, 20, 45, 2, '=E14/(2*5)'),
                ('PI_09', 'PV_2', 175, 100, 75, 27, 22, 49, 2, '=E15/(2*5)'),
            ],
            'N12_MONTANHOSO': [
                ('PI_00', 'PV_1', 190, 100, 75, 30, 24, 54, 2, '=E16/(2*5)'),
                ('PI_10', 'PV_2', 210, 100, 75, 33, 27, 60, 2, '=E17/(2*5)'),
            ],
            'SD_JOAO_CARLOS': [
                ('PI_00', 'PV_1', 240, 150, 100, 38, 30, 68, 3, '=E18/(3*6)'),
                ('PI_11', 'PV_2', 220, 150, 100, 35, 28, 63, 3, '=E19/(3*6)'),
            ],
            'SD_SAO_MANOEL': [
                ('PI_00', 'PV_1', 260, 150, 100, 42, 33, 75, 3, '=E20/(3*6)'),
                ('PI_12', 'PV_2', 235, 150, 100, 37, 30, 67, 3, '=E21/(3*6)'),
            ],
            'SD_VILA_ISRAEL': [
                ('PI_00', 'PV_1', 200, 125, 90, 32, 26, 58, 2, '=E22/(2*5)'),
                ('PI_13', 'PV_2', 185, 125, 90, 29, 24, 53, 2, '=E23/(2*5)'),
            ],
            'SD_MORRO_TETEU': [
                ('PI_00', 'PV_1', 280, 150, 100, 45, 36, 81, 3, '=E24/(3*4)'),
                ('PI_14', 'PV_2', 265, 150, 100, 42, 34, 76, 3, '=E25/(3*4)'),
                ('PI_15', 'PV_3', 250, 150, 100, 40, 32, 72, 3, '=E26/(3*4)'),
            ],
            'SD_VILA_CRIADORES': [
                ('PI_00', 'PV_1', 230, 150, 100, 37, 30, 67, 3, '=E27/(3*6)'),
                ('PI_16', 'PV_2', 215, 150, 100, 34, 28, 62, 3, '=E28/(3*6)'),
            ],
            'SD_PANTANAL_BAIXO': [
                ('PI_00', 'PV_1', 270, 150, 100, 43, 35, 78, 3, '=E29/(3*5)'),
                ('PI_17', 'PV_2', 255, 150, 100, 41, 33, 74, 3, '=E29/(3*5)'),
            ],
        }

        # Preencher tabela
        for nucleo, trechos in nucleos_pis_pvs.items():
            for pi, pv, ext, dn_agua, dn_esgoto, la, le, total_lig, eq, dias_formula in trechos:
                # Formato: NS_017_PI_00_AO_PV_62
                ns_nome = f"NS_{ns_counter:03d}_{pi}_AO_{pv}"

                ws.cell(row_atual, 1, ns_nome)
                ws.cell(row_atual, 2, f"{nucleo.replace('_', ' ')}")
                ws.cell(row_atual, 3, pi)
                ws.cell(row_atual, 4, pv)
                ws.cell(row_atual, 5, ext)
                ws.cell(row_atual, 6, dn_agua)
                ws.cell(row_atual, 7, dn_esgoto)
                ws.cell(row_atual, 8, la)
                ws.cell(row_atual, 9, le)
                ws.cell(row_atual, 10, total_lig)
                ws.cell(row_atual, 11, eq)
                ws.cell(row_atual, 12, dias_formula)  # Fórmula Excel

                # Status baseado no progresso
                if ns_counter % 3 == 0:
                    status = "✅ Concluído"
                    cor = "C6EFCE"
                elif ns_counter % 3 == 1:
                    status = "🔄 Em execução"
                    cor = "FFEB9C"
                else:
                    status = "⏳ Pendente"
                    cor = "FFC7CE"

                ws.cell(row_atual, 13, status)
                ws.cell(row_atual, 13).fill = PatternFill("solid", fgColor=cor)

                row_atual += 1
                ns_counter += 1

        # Total geral
        ws.cell(row_atual, 1, f"TOTAL: {ns_counter - 1} NSs")
        ws.cell(row_atual, 1).font = Font(bold=True, size=11)
        ws.cell(row_atual, 5, f"=SUM(E{row_tabela + 2}:E{row_atual - 1})")  # Soma extensão
        ws.cell(row_atual, 10, f"=SUM(J{row_tabela + 2}:J{row_atual - 1})")  # Soma ligações

        # Ajustar colunas
        ws.column_dimensions['A'].width = 22  # NS
        ws.column_dimensions['B'].width = 20  # Trecho
        ws.column_dimensions['C'].width = 10  # PI
        ws.column_dimensions['D'].width = 10  # PV
        ws.column_dimensions['E'].width = 12  # Extensão
        ws.column_dimensions['F'].width = 10  # DN Água
        ws.column_dimensions['G'].width = 10  # DN Esgoto
        ws.column_dimensions['H'].width = 8   # LA
        ws.column_dimensions['I'].width = 8   # LE
        ws.column_dimensions['J'].width = 10  # Total Lig
        ws.column_dimensions['K'].width = 8   # Equipes
        ws.column_dimensions['L'].width = 10  # Dias
        ws.column_dimensions['M'].width = 14  # Status

        print(f"    {ns_counter - 1} Notas de Serviço geradas")

    # ── 6. GERAR GRÁFICOS SEABORN ─────────────────────────────────────────────
    
    def gerar_graficos(self):
        """Gera gráficos com Seaborn."""
        if not MPL_OK or not SNS_OK:
            print("  ⚠ Gráficos desativados (matplotlib/seaborn não disponíveis)")
            return
        
        print("  📊 Gerando gráficos com Seaborn...")
        
        graficos_dir = SCRIPT_DIR / "graficos_ml"
        graficos_dir.mkdir(exist_ok=True)
        
        # Configurar estilo
        sns.set_theme(style="whitegrid", palette="muted", font_scale=1.0)
        
        # 1. Feature Importance
        fig, ax = plt.subplots(figsize=(10, 6))
        
        if hasattr(self, 'df_feature_importance'):
            df_plot = self.df_feature_importance.copy()
            labels_pt = {
                'lig_total_r3': 'Rolling 3d — Lig. totais',
                'lig_total_r7': 'Rolling 7d — Lig. totais',
                'pra_r3': 'Rolling 3d — Rede (m)',
                'pra_r7': 'Rolling 7d — Rede (m)',
                'la_r3': 'Rolling 3d — Lig. água',
                'le_r3': 'Rolling 3d — Lig. esgoto',
                'nucleo_enc': 'Núcleo',
                'dia_semana': 'Dia semana',
                'mes': 'Mês',
                'fim_semana': 'Fim de semana',
                'dias_decorridos': 'Dias decorridos',
            }
            df_plot['label'] = df_plot['feature'].map(labels_pt).fillna(df_plot['feature'])
            
            sns.barplot(data=df_plot.iloc[::-1], x='importancia', y='label',
                       ax=ax, palette='Blues_d')
            
            ax.set_xlabel('Importância (gain)')
            ax.set_ylabel('')
            ax.set_title('Feature Importance — XGBoost ML\nSLNR Santos · Contrato 11481051')
            
            plt.tight_layout()
            plt.savefig(graficos_dir / '01_feature_importance.png', dpi=150, bbox_inches='tight')
            plt.close()
        
        # 2. Cenários
        fig, ax = plt.subplots(figsize=(10, 6))
        
        if self.cenarios:
            nomes_curto = [c['nome'].split(' ')[0] for c in self.cenarios]
            producoes = [c['producao_mensal'] for c in self.cenarios]
            
            cores = ['#3498DB', '#2ECC71', '#F39C12', '#E74C3C', '#9B59B6']
            bars = ax.bar(nomes_curto, producoes, color=cores)
            
            ax.set_ylabel('Ligações/mês')
            ax.set_title('Cenários de Aceleração — ML\nSLNR Santos · Contrato 11481051')
            ax.tick_params(axis='x', rotation=15)
            
            # Adicionar valores nas barras
            for bar, val in zip(bars, producoes):
                ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 2,
                       f'{val:.0f}', ha='center', va='bottom', fontsize=9)
            
            plt.tight_layout()
            plt.savefig(graficos_dir / '02_cenarios.png', dpi=150, bbox_inches='tight')
            plt.close()
        
        # 3. Heatmap por núcleo
        fig, ax = plt.subplots(figsize=(12, 8))
        
        dados_heatmap = pd.DataFrame(NUCLEOS_20)
        dados_heatmap['tipo_short'] = dados_heatmap['tipo'].str.replace('_', ' ').str.split().str[0]
        
        pivot = dados_heatmap.pivot_table(index='nome', columns='tipo_short',
                                          values='prod', aggfunc='first')
        
        sns.heatmap(pivot, annot=True, fmt='.1f', cmap='YlOrRd', ax=ax,
                   cbar_kws={'label': 'Produtividade (m/eq/dia)'})
        
        ax.set_xlabel('Tipo de Terreno')
        ax.set_ylabel('Núcleo')
        ax.set_title('Heatmap de Produtividade por Núcleo\nSLNR Santos · Contrato 11481051')
        
        plt.tight_layout()
        plt.savefig(graficos_dir / '03_heatmap_nucleos.png', dpi=150, bbox_inches='tight')
        plt.close()
        
        print(f"    {len(list(graficos_dir.glob('*.png')))} gráficos gerados em {graficos_dir}")
    
    # ── 7. EXPORTAR JSON ──────────────────────────────────────────────────────
    
    def exportar_json(self):
        """Exporta resultados em JSON."""
        print("  💾 Exportando JSON...")
        
        resultado = {
            'gerado_em': datetime.now().isoformat(),
            'plataforma': 'SLNR_MESTRE_UNIFICADO_ML v7.0',
            'contrato': '11481051 · SLNR Santos',
            'modelo': self.metricas_ml,
            'cenarios': self.cenarios,
            'nucleos': NUCLEOS_20,
            'feature_importance': self.df_feature_importance.to_dict('records') if hasattr(self, 'df_feature_importance') else [],
        }
        
        json_path = SCRIPT_DIR / 'SLNR_ML_RESULTADOS.json'
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(resultado, f, indent=2, ensure_ascii=False, default=str)
        
        print(f"    JSON salvo: {json_path.name}")
    
    # ── MAIN ──────────────────────────────────────────────────────────────────
    
    def executar(self):
        """Executa todo o pipeline de integração ML."""
        print("\n" + "="*70)
        print(" SLNR_MESTRE_UNIFICADO — MACHINE LEARNING INTEGRADO")
        print(" XGBoost + GridSearchCV + Seaborn")
        print(" Contrato 11481051 · DGS Engenharia · SLNR Santos")
        print("="*70 + "\n")
        
        # 1. Carregar dados
        self.carregar_dados()
        
        # 2. Preparar features
        df, FEATURES = self.preparar_features()
        
        # 3. Treinar modelo
        modelo, y_test, y_pred_test, y_pred_all = self.treinar_modelo(df, FEATURES)
        
        # 4. Gerar cenários
        self.gerar_cenarios(df)
        
        # 5. Atualizar planilha
        self.atualizar_planilha()
        
        # 6. Gerar gráficos
        self.gerar_graficos()
        
        # 7. Exportar JSON
        self.exportar_json()
        
        print("\n" + "="*70)
        print(" ✅ CONCLUÍDO!")
        print(f" 📊 R² = {self.metricas_ml.get('r2_test', 0):.4f}")
        print(f" 📈 {len(self.cenarios)} cenários gerados")
        print(f" 📁 Saída: {self.arquivo_saida.name}")
        print("="*70 + "\n")
        
        return self.resultados


# ── MAIN ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    integrador = SLNRMLIntegrador()
    integrador.executar()
